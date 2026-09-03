# ズレの警告が、事実を言っているか ── 2026-08-31。
# Namakoo「この指す中身が変わるとはどういうこと？」（画面の ⚠ を見て）
#
# ★★ 実測（「金額と単価を入れ替えて」）── ⚠ の中身が **3 つとも事実に反していた**:
#     ⚠ の主張                    実際
#     「指す先の中身が変わる」  →  変わらない（税込金額は金額を指し続けた）
#     「**行**が入れ替わる」    →  入れ替えたのは**列**
#     「直していません」        →  **直している**（=E2*1.1 → =D2*1.1・値も同じ）
#   そして嘘の ⚠ のせいで、**正しく動いた操作の ✓ が △ に落ちていた**。
#
# ★★ 原因は片配線の**逆**:
#   この警告は並べ替え（式を直さない）用に作り、入れ替えにも同じものを配線した。
#   そのあと**入れ替えだけ式を直すようになった**のに、警告は昔の前提のまま残った。
#   ★ 直しを両方に配ったとき、**その直しが前提を変えていないか**を見ていなかった。
#
# ★ 黙らせすぎない: **別シートから指している式**は書き直していないので、名指しは残す。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import cellmap as cm  # noqa: E402
from _product_source import window_around  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む


# --- ① 書き直した式は「ずれる式」ではない ------------------------------------------------

def test_a_rewritten_formula_is_not_called_drifting():
    hits = [("売上", "F2", "=E2*1.1", 2, 5)]
    assert cm.drop_rewritten(hits, {(2, 6)}, "売上") == [], "書き直した式が名指しに残っている"


def test_a_formula_on_another_sheet_still_counts():
    """★ 別シートからの参照は書き直していない ── 黙らせない。"""
    hits = [("まとめ", "B1", "=売上!C2", 2, 3)]
    assert cm.drop_rewritten(hits, {(1, 2)}, "売上") == hits


def test_nothing_rewritten_means_nothing_dropped():
    """★ 並べ替え（式を直さない op）では、今までどおり全部名指しする。"""
    hits = [("売上", "F2", "=E2*1.1", 2, 5)]
    assert cm.drop_rewritten(hits, set(), "売上") == hits
    assert cm.drop_rewritten(hits, None, "売上") == hits


# --- ② 軸を取り違えない -------------------------------------------------------------------

@pytest.mark.parametrize("unit", ["行", "列"])
def test_the_note_names_the_axis_it_was_given(unit):
    """★ 列の入れ替えなのに「行が入れ替わります」と言っていた
       （同じ形の言い間違いを 08-30 に別の場所で直したばかり）。"""
    note = cm.reference_drift_note([("S", "B1", "=A2", 2, 1)], unit=unit)
    assert f"指している{unit}が入れ替わります" in note, note


def test_the_swap_passes_the_axis_through():
    """★ 変異試験: 入れ替えが軸を渡していること（既定の「行」に戻ったら赤くする）。"""
    seg = window_around("reference_drift_warning(book_meta, _sw_sheet", after=300)
    assert 'unit=("列" if as_col else "行")' in seg, seg[:200]
    assert "rewritten=set(_rw)" in seg, "書き直した式を渡していない"


# --- ③ 実物で（LibreOffice を通す）--------------------------------------------------------

def _run(book: Path, task: str, sheet: str):
    import subprocess
    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}
    return subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(book), task, "--copy", "--sheet", sheet],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=900, cwd=str(REPO), env=env)


@pytest.mark.local
def test_swapping_columns_no_longer_warns_about_its_own_formulas(tmp_path):
    """★★ Namakoo の実例そのもの ── 正しく動いた操作が ✓ になること。"""
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "単価", "金額", "税込金額"])
    for i, (n, t, k) in enumerate([("りんご", 100, 1200), ("みかん", 200, 800)], start=2):
        ws.append([n, t, k, f"=C{i}*1.1"])
    wb.save(p)
    r = _run(p, "単価と金額を入れ替えて", "売上")
    assert "✓" in r.stdout, r.stdout[-1200:]
    assert "指す先の中身が変わる式" not in r.stdout, r.stdout[-1200:]
    v = openpyxl.load_workbook(tmp_path / "b.out.xlsx", data_only=True)["売上"]
    assert [v.cell(1, c).value for c in (2, 3)] == ["金額", "単価"]
    assert v.cell(2, 4).value == pytest.approx(1320)   # 税込は金額に付いていく


@pytest.mark.local
def test_a_cross_sheet_reference_is_still_named(tmp_path):
    """★ 黙らせすぎていないこと ── 書き直さない参照は今までどおり名指しする。"""
    p = tmp_path / "c.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "単価", "金額"])
    for r_ in (["りんご", 100, 1200], ["みかん", 200, 800]):
        ws.append(r_)
    s2 = wb.create_sheet("まとめ")
    s2["A1"] = "りんごの金額"
    s2["B1"] = "=売上!C2"
    wb.save(p)
    r = _run(p, "単価と金額を入れ替えて", "売上")
    assert "まとめ!B1" in r.stdout, r.stdout[-1200:]
    assert "指している列が入れ替わります" in r.stdout, r.stdout[-1200:]
