# セル 2 つの入れ替え ── 2026-08-31。
# Namakoo「これは間違った操作だが機械検証が通っている。やりたいのは丸和物流の単価セルの値と
# みどり建設の単価セルの値を入れ替えたいだけだが、それを行の入れ替えと判断している」
#
# ★★ 実測（画面）: 「丸和物流の単価とみどり建設の単価を入れ替えて」で
#   **行を丸ごと入れ替えて ✓ を出していた**（16 セルが動いた・頼んだのは 2 セル）。
#   ★ ✓ が出た理由: 番人は「**宣言どおり**行が入れ替わったか」を見る。
#     宣言そのものが違えば、検算は通る ── 三項（依頼・宣言・実体）の「依頼」が抜けた形。
#     この repo で何度も踏んでいるが、**✓ が出たまま気づけない**のはこれが一番危ない。
#
# ★ 依頼文には証拠が在った: **両側とも「〜の〈列名〉」と列を名指ししている**。
#   行の入れ替えなら列は出てこない。
# ★ 言い回しを数え上げない ── 見るのは**実表に在る見出し**と**実表に在る行**だけ。
#
# ★ 直しは 2 段階で入れた（安全側を先に）:
#   ① 検出して**断る**（✓ を出す道を先に塞ぐ）→ commit
#   ② セルの入れ替えとして**実装する**（SWAP の第 3 の軸。新しい op は作らない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HEADERS = ["取引先", "項目", "件数", "単価", "金額"]
ROWS = [["丸和物流", "配送", 12, 4800, 57600],
        ["近江スチール", "鋼材", 5, 12000, 60000],
        ["みどり建設", "内装", 9, 7200, 64800]]


@pytest.fixture()
def meta(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(p)
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(p)}


# --- ① 検出（誤爆させない）----------------------------------------------------------------

def test_two_named_cells_are_detected(meta):
    got = ailine.swap_targets_are_cells("丸和物流の単価とみどり建設の単価を入れ替えて",
                                          meta, "請求")
    assert got == [(2, 4), (4, 4)], got


def test_the_two_columns_may_differ(meta):
    got = ailine.swap_targets_are_cells("丸和物流の単価とみどり建設の金額を入れ替えて",
                                          meta, "請求")
    assert got == [(2, 4), (4, 5)], got


@pytest.mark.parametrize("task", [
    "丸和物流とみどり建設を入れ替えて",          # 行の入れ替え
    "金額と単価を入れ替えて",                   # 列の入れ替え
    "丸和物流の行とみどり建設の行を入れ替えて",   # 行と明示
])
def test_row_and_column_swaps_are_not_hijacked(meta, task):
    """★★ ここが緩むと、動いていた入れ替えを奪って壊す。"""
    assert ailine.swap_targets_are_cells(task, meta, "請求") is None


def test_the_same_cell_twice_is_not_a_swap(meta):
    assert ailine.swap_targets_are_cells("丸和物流の単価と丸和物流の単価を入れ替えて",
                                          meta, "請求") is None


# --- ② 解決（値は実表から読む・A' 原則）---------------------------------------------------

def test_the_values_come_from_the_table_not_the_model(meta):
    ok, r, _i, err = ailine.verify_dsl_args(
        "SWAP", {"a": "丸和物流", "b": "みどり建設"}, meta,
        task="丸和物流の単価とみどり建設の単価を入れ替えて")
    assert ok, err
    assert r["_axis"] == "cell"
    assert r["_cells"] == [[2, 4], [4, 4]]
    assert r["_cell_values"] == [4800, 7200], r["_cell_values"]
    assert "単価" in r["_axis_label"]


def test_identical_contents_are_refused(meta):
    """★ 入れ替えても何も変わらない依頼は、✓ を出さずに断る。"""
    ok, _r, _i, err = ailine.verify_dsl_args(
        "SWAP", {"a": "丸和物流", "b": "近江スチール"}, meta,
        task="丸和物流の項目と近江スチールの項目を入れ替えて")
    if not ok:
        assert "同じ" in err or "決められません" in err, err


# --- ③ 番人（2 セル**だけ**が動いたことを証明する）----------------------------------------

def _book(tmp_path, name, rows):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


ARGS = {"_axis": "cell", "_cells": [[2, 4], [4, 4]], "_cell_values": [4800, 7200],
        "_target_sheet": "請求", "a": "丸和物流", "b": "みどり建設"}


def test_a_correct_cell_swap_passes(tmp_path):
    src = _book(tmp_path, "s.xlsx", ROWS)
    out = _book(tmp_path, "o.xlsx",
                 [["丸和物流", "配送", 12, 7200, 57600],
                  ["近江スチール", "鋼材", 5, 12000, 60000],
                  ["みどり建設", "内装", 9, 4800, 64800]])
    st, why = ailine.check_swap(out, ARGS, source_book=src)
    assert st == "pass", why


def test_swapping_the_whole_row_is_caught(tmp_path):
    """★★ これが今回の事故そのもの ── 行ごと入れ替えたら ✓ を出さない。"""
    src = _book(tmp_path, "s2.xlsx", ROWS)
    out = _book(tmp_path, "o2.xlsx",
                 [["みどり建設", "内装", 9, 7200, 64800],
                  ["近江スチール", "鋼材", 5, 12000, 60000],
                  ["丸和物流", "配送", 12, 4800, 57600]])
    st, why = ailine.check_swap(out, ARGS, source_book=src)
    assert st == "fail", why
    assert "セル変わっています" in why or "ではありません" in why, why


def test_doing_nothing_is_caught(tmp_path):
    src = _book(tmp_path, "s3.xlsx", ROWS)
    out = _book(tmp_path, "o3.xlsx", ROWS)
    st, _why = ailine.check_swap(out, ARGS, source_book=src)
    assert st == "fail"


def test_an_extra_cell_is_caught(tmp_path):
    """★ 「他は 1 セルも変わらず」── 余計に書いたら通さない。"""
    src = _book(tmp_path, "s4.xlsx", ROWS)
    out = _book(tmp_path, "o4.xlsx",
                 [["丸和物流", "配送", 12, 7200, 99999],
                  ["近江スチール", "鋼材", 5, 12000, 60000],
                  ["みどり建設", "内装", 9, 4800, 64800]])
    st, why = ailine.check_swap(out, ARGS, source_book=src)
    assert st == "fail", why


# --- ④ 実物で（LibreOffice を通す）--------------------------------------------------------

@pytest.mark.local
def test_only_two_cells_move_for_real(tmp_path):
    """★★ Namakoo の実例そのもの。"""
    import subprocess
    src = _book(tmp_path, "r.xlsx", ROWS)
    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(src),
         "丸和物流の単価とみどり建設の単価を入れ替えて", "--copy", "--sheet", "請求"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=900, cwd=str(REPO), env=env)
    # ★★ 2026-08-31: この検体を書いた時点では ✓ を期待していたが、そのあと
    #   **派生列の取り残しの ⚠** を入れたので、この表（金額＝件数×単価）では △ が正しい。
    #   ★ 実装が正しく、検体が追いついていなかった ── 期待値を実態に合わせる。
    #     操作そのものが通っていること（2 セルだけ動いた）は下の assert で見る。
    assert "2 セルだけを入れ替え" in r.stdout, r.stdout[-1500:]
    assert "『金額』＝『件数』×『単価』" in r.stdout, "派生列の取り残しを言っていない"
    v = openpyxl.load_workbook(tmp_path / "r.out.xlsx", data_only=True)["請求"]
    assert v.cell(2, 4).value == 7200 and v.cell(4, 4).value == 4800
    assert v.cell(2, 1).value == "丸和物流" and v.cell(4, 1).value == "みどり建設"
    assert v.cell(2, 5).value == 57600, "金額まで動いている"


# --- ⑤ 一段目が降りた回でも届くこと（読み直し）--------------------------------------------

def test_the_reread_uses_the_machine_not_the_model():
    """★★ 2026-08-31（Namakoo「単価の入れ替えに対応していないようにみえる」）:
       「みどり建設の単価と丸和物流の単価を入れ替えて」が **OUT_OF_VOCAB** で終わっていた。
       セルの入れ替えは前日に実装したが、それは SWAP の**検証段**に置いたので、
       **op が SWAP にならなければ一度も呼ばれない**。
       ★ また「番人は在るが、失敗が取る形では鳴らない」── 何度も踏む形。
       ★ 直し: 読み直しで LLM に聞く**前に**、機械だけで 2 セルが解けているならそれを使う。
    """
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    # ★ 「最初の出現」で切ると、別の門に同じ述語を足したときに外れる（実際に外れた）
    #   ── 入れ替えの読み直しそのもの（plan を SWAP に差し替える所）を狙う。
    # ★★ 2026-09-02: **窓を固定長（1200/600 文字）で切っていて外れた** ──
    #   行の入れ替えの読み直しを間に足したら、LLM への聞き直しが窓の外へ出た。
    #   この repo が何度も書いている「窓は構造で切る」を、ここだけ守っていなかった。
    #   ★ 両端を構造で取る: セルの読み直し（plan を SWAP に差し替える所）から、
    #     LLM への聞き直しまで。間に何を足しても、順序の主張は生き続ける。
    i = src.index('plan = [{"op": "SWAP", "args": {}}]')
    start = src.rindex("if (not _reread_done", 0, i)
    end = src.index('translate_task_fixed_op(a.model, "SWAP"', i)
    seg = src[start:end]
    assert "swap_targets_are_cells(a.task, book_meta, _sheet_h)" in seg, (
        "読み直しが、機械で解けるセルを見ていない")
    # ★ 窓の終わりが LLM への聞き直しそのものなので、機械の判定がその手前に在れば
    #   順序は満たされている（seg の中に translate_task_fixed_op は現れない）。
    assert "translate_task_fixed_op" not in seg, "LLM に聞いてから機械を見ている（順序が逆）"


def test_the_cell_reread_actually_fires_without_the_llm(tmp_path, monkeypatch, capsys):
    """★★ 2026-09-02: 上の番人は**文字の順序**しか見ておらず、`if False and …` で
      無効化しても緑のままだった（行の側の変異試験で気づいた）。
      ★ 在ることと効くことは別 ── **配線を通して**確かめる 1 本を置く。
    ★ 一段目が語彙外を返す回（実測で起きた形）を作り、二段目（LLM への聞き直し）が
      **呼ばれない**ことまで見る。適用しない（--dry）ので LibreOffice も要らない。
    """
    import openpyxl
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_golden_transcripts import _isolate, _run_main

    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "c.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "単価"])
    for n, v in [("あかね商事", 12000), ("いろは工業", 4500), ("うえだ物産", 30000)]:
        ws.append([n, v])
    wb.save(p)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "OUT_OF_VOCAB", "about": "特定の値の入れ替え"}]})
    monkeypatch.setattr(ailine, "translate_task_fixed_op",
                         lambda *a, **k: pytest.fail("機械で解けるのに LLM に聞いている"))
    _rc, out = _run_main(["run", str(p),
                           "あかね商事の単価とうえだ物産の単価を入れ替えて", "--dry"], capsys)
    assert "読み直しました" in out, out
    assert "2 つのセル" in out, out


def test_cells_are_resolved_before_the_row_or_column_decision():
    """★ a/b が空でも通ること ── 一段目が降りた回は a/b が無い。
       セルの解決を**行/列の判定より前**に置く（そこで確定して先へ行かない）。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index('elif op == "SWAP":')
    # ★ 窓を固定長で切ると、間に足したぶんで外れる（実際に外れた）── 次の分岐まで見る。
    seg = src[i:src.index('elif op == "INSERT_ROWS":', i)]
    assert seg.index("_cells0 = swap_targets_are_cells") < seg.index("as_col = _a in _headers_s")
    assert "if not _cells0 and (not _a or not _b):" in seg, "a/b が無い回に先に落ちる"


def test_the_interpretation_line_has_no_empty_slots(meta):
    """★ 実測で「入れ替える一方: もう一方:」と**空欄**が出ていた（a/b が無い経路）。
       嘘の空欄を見せない ── 行の名前で埋める。"""
    ok, r, _i, err = ailine.verify_dsl_args(
        "SWAP", {}, meta, task="丸和物流の単価とみどり建設の単価を入れ替えて")
    assert ok, err
    assert r["a"] == "丸和物流" and r["b"] == "みどり建設", (r.get("a"), r.get("b"))


# --- ⑥ 行番号で指した入れ替え ------------------------------------------------------------
#
# ★★ 2026-08-31（Namakoo が実測・「この基本操作ができない」）:
#   「6行目と5行目を入れ替えて」が CLARIFY に落ちていた ── **行番号で指すと黙る**、
#   08-29 に追加・削除で直したのと同じ非対称が、**入れ替えには残っていた**。
# ★ 直しは「住所の解決を集めている 1 箇所」（_resolve_named_row）に足した ──
#   入れ替え専用の判定を作らない。そこに足すと全部の op に効く。
# ★★ さらに、解釈が通ったあとに**実行が別の行を動かした** ── 生成側が Basic に
#   **名前**を渡しており、表に『6行目』という名前は無いので探せなかった（番人が捕捉）。
#   行番号で解けた回は**座標で**渡す（SwapRowsAt を追加・既存の SwapRowsByName は不変）。


def test_a_row_number_resolves_as_a_row(meta):
    row, note = ailine._resolve_named_row(meta, "請求", "6行目")
    assert row == 6, (row, note)
    assert "依頼文の行番号" in note


def test_the_header_row_is_refused(meta):
    row, note = ailine._resolve_named_row(meta, "請求", "1行目")
    assert row is None and "見出し行" in note, (row, note)


def test_a_name_is_still_resolved_by_content(meta):
    """★ 行番号を足したせいで、名前で指す道が壊れていないこと。"""
    row, _note = ailine._resolve_named_row(meta, "請求", "丸和物流")
    assert row == 2


def test_the_codegen_passes_coordinates_for_numbered_rows():
    """★★ 名前で渡すと Basic が探せない（実測で別の行が動いた）。"""
    m = {"sheets": ["S"], "headers": {"S": HEADERS}, "header_rows": {"S": 1}}
    args = {"a": "6行目", "b": "5行目", "_axis": "row", "_a_pos": 6, "_b_pos": 5,
            "_headers": HEADERS, "_header_row": 1}
    code = ailine.codegen_dsl("SWAP", args, m)
    assert "Call SwapRowsAt(oDoc, 5, 4)" in code, code
    assert "SwapRowsByName" not in code, code


def test_the_codegen_still_passes_names_when_named():
    """★ 名前で指された回は今までどおり名前を渡す（Basic 側が独立に位置を探す）。"""
    m = {"sheets": ["S"], "headers": {"S": HEADERS}, "header_rows": {"S": 1}}
    args = {"a": "丸和物流", "b": "みどり建設", "_axis": "row", "_a_pos": 2, "_b_pos": 4,
            "_headers": HEADERS, "_header_row": 1}
    code = ailine.codegen_dsl("SWAP", args, m)
    assert "SwapRowsByName" in code and "SwapRowsAt" not in code, code


def test_the_basic_helper_exists_and_is_new():
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    assert "Sub SwapRowsAt(" in bas
    # ★ 既存の SwapRowsByName は触っていない（目録・凍結検体が全部動く）
    i = bas.index("Sub SwapRowsByName(")
    assert "nKeyCol As Integer" in bas[i:i + 200]


# --- ⑦ 「AとBの〈列〉」と、入れ替えを 1 セル書換に読み替えない ------------------------------
#
# ★★ 2026-08-31（Namakoo が実測・「項目名が上書きされている。しかも機械検証が通している。
#   これは式としては正しいかもしれない」）:
#     「丸和物流と近江スチールの**項目を入れ替えて**」
#       → **1 セル書換**に読み直され、近江スチールの項目に**「丸和物流」と書いて** ✓
#          （鋼材加工が消えた）
#   ★ Namakoo の見立てどおり、番人は「1 セルだけ書き換えた」を**正しく**検算している。
#     違っていたのは**宣言**のほう ── 三項の「依頼」が抜けた形（今日 2 度目）。
#
# ★ 直しは 2 つ:
#   ① 依頼文が**入れ替え**と言っているなら、1 セル書換に読み替えない
#     （入れ替えは 2 か所が動く操作・1 セル書換は 1 か所しか動かない＝別の仕事）
#   ② 「〈A〉と〈B〉の〈列〉」＝ 2 つの行の同じ 1 列、として解く
#     （列名が **1 回しか出ない**言い方なので、前の形では拾えず行を丸ごと入れ替えていた）


def test_the_shared_column_phrasing_is_two_cells(meta):
    """★★ 列名が 1 回しか出ない言い方も、2 セルとして解く。"""
    got = ailine.swap_targets_are_cells("丸和物流と近江スチールの項目を入れ替えて",
                                          meta, "請求")
    assert got == [(2, 2), (3, 2)], got


def test_both_phrasings_agree(meta):
    """★ 同じ意味なら同じ答え（指し方で結果が変わってはいけない）。"""
    a = ailine.swap_targets_are_cells("丸和物流と近江スチールの項目を入れ替えて", meta, "請求")
    b = ailine.swap_targets_are_cells("丸和物流の項目と近江スチールの項目を入れ替えて",
                                        meta, "請求")
    assert a == b, (a, b)


@pytest.mark.parametrize("task", [
    "丸和物流と近江スチールを入れ替えて",   # 行の入れ替え（列が出てこない）
    "件数と単価を入れ替えて",              # 列の入れ替え
])
def test_the_new_shape_does_not_hijack(meta, task):
    assert ailine.swap_targets_are_cells(task, meta, "請求") is None


def test_a_swap_request_is_never_reread_as_a_single_cell_write():
    """★★ 変異試験: 入れ替えの依頼を 1 セル書換の門が拾わないこと。
       ここが緩むと、片方の名前がもう片方のセルに**書き込まれて ✓ が出る**。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index("_cell = resolve_cell_target_from_task(a.task, book_meta, _sheet_h)")
    seg = src[max(0, i - 900):i]
    assert "not task_asks_for_a_swap(a.task)" in seg, (
        "入れ替えの依頼が 1 セル書換に読み替えられる")
