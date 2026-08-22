# operator 盲検 9 回目（$25・PRED-20260823-operator9.md）の BROKEN 2 件 ── 修正より先に
# 凍結した赤い検体。Namakoo「かかってくれ」(2026-08-23 08:01)。
#
# ①比較語の等号側丸め: 「より大きい」→gte・「未満」→lte と LLM が写し間違えても素通し
#   （cmp だけ A' 原則の外・value は機械確定するのに比較種は LLM を信じていた）。
#   実害: 「4万円未満だけ抜き出して」に 4 万ちょうどが混入 ── 行数と Σ が黙って狂う。
#   直し: 依頼文から比較語を機械抽出し、LLM と食い違えば機械が勝つ+開示（factor と同じ作法）
# ②捏造段: 複合計画に依頼文へ根拠を持たない op の段（けい線）が湧いても ✓ が出る。
#   単位E は対象スロットの食い違いしか見ない ── 「段の存在自体の根拠」が無検査だった。
#   直し: 各段の op について 根拠 =（照合プール句が依頼文に語として在る）∨（解決済み args が
#   依頼文の語に接地している）を機械照合し、根拠なしの段は ★ で名指し+決裁③の △ 降格に乗せる

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "extract_cmp_from_task"),
    reason="operator9 修正 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


# --- ① 比較語の機械抽出 ---------------------------------------------------------

@needs_impl
@pytest.mark.parametrize("task,expected", [
    ("金額が5000より大きい行を抜き出して", "gt"),
    ("金額が5000を超える行を抜き出して", "gt"),
    ("金額が5000未満の行を抜き出して", "lt"),
    ("金額が5000より小さい行を抜き出して", "lt"),
    ("金額が5000以上の行を抜き出して", "gte"),
    ("金額が5000以下の行を抜き出して", "lte"),
    ("備考に東京を含む行を抜き出して", "contains"),
    ("状態が完了と等しい行を抜き出して", "eq"),
    ("東京の行を抜き出して", None),          # 比較語なし → 機械は断定しない
])
def test_extract_cmp_from_task(task, expected):
    assert ailine.extract_cmp_from_task(task) == expected


@needs_impl
def test_llm_cmp_loses_to_task_words_with_disclosure():
    """LLM が gte を返しても、依頼文の「より大きい」が gt に上書きし、食い違いを開示する。"""
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]},
             "header_rows": {"Sheet": 1}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "金額", "cmp": "gte", "value": 5000}, meta,
        task="金額が5000より大きい行を抜き出して")
    assert ok, err
    assert resolved["cmp"] == "gt", f"機械抽出が LLM に負けた: {resolved}"
    assert resolved.get("_warnings"), "食い違いの開示が無い"
    assert "gte" in str(resolved["_warnings"]) or "以上" in str(resolved["_warnings"])


def test_llm_cmp_stands_when_task_silent_and_agrees_without_noise():
    # xfail 対象外: 現状でも通る退行防止（実装後も静かさが保たれることの番人）
    """依頼文に比較語が無ければ LLM の cmp を従来どおり使う（開示なし）。
       一致していれば開示も出さない（オオカミ少年防止）。"""
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]},
             "header_rows": {"Sheet": 1}}
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "金額", "cmp": "gt", "value": 5000}, meta,
        task="金額が5000を超える行を抜き出して")
    assert ok and resolved["cmp"] == "gt"
    assert not resolved.get("_warnings"), f"一致なのに開示が出た: {resolved.get('_warnings')}"


# --- ② 捏造段の検出 -------------------------------------------------------------

def _star_book(tmp_path):
    return _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])


@needs_impl
def test_fabricated_step_is_named_and_checkmark_demoted(tmp_path, monkeypatch, capsys):
    """operator の実再現: 依頼に一度も出てこない『けい線』の段が湧いたら、
       その段を ★ で名指しし、最後の ✓ は △ に落ちる（決裁③の機械に乗せる）。"""
    _isolate(monkeypatch, tmp_path)
    book = _star_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                   {"op": "DRAW_BORDERS", "args": {}}]})
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if calls["n"] == 1:
            ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        else:
            from openpyxl.styles import Border, Side
            thin = Side(style="thin")
            for row in ws.iter_rows():
                for c in row:
                    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(
        ["run", str(book), "金額を降順に並べ替えて、一番高い行に印をつけて", "--copy"], capsys)
    assert "けい線" in out and "根拠" in out, f"捏造段の名指しが無い: {out}"
    assert "✓" not in out, f"根拠なしの段があるのに ✓ が出ている: {out}"
    assert "△" in out, f"降格の顔が無い: {out}"


@needs_impl
def test_legitimate_steps_pool_word_and_args_grounding_not_flagged(tmp_path, monkeypatch, capsys):
    """誤爆防止の二本立て:
       - 段1 SORT: 「並べ替え」が依頼文に語として在る（照合プールの根拠）
       - 段2 COMPUTE_COLUMN: プール語は無いが operands の列名（数量・単価）が依頼文に在る
         （args 接地の根拠）
       どちらも ★ を受けず、✓ が出る。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "数量", "単価"], ["a", 2, 100], ["b", 1, 50]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "SORT", "args": {"col": "単価", "order": "desc"}},
                   {"op": "COMPUTE_COLUMN",
                    "args": {"operands": ["数量", "単価"], "operator": "*",
                              "new_col": "金額"}}]})
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if calls["n"] == 1:
            ws["A2"], ws["B2"], ws["C2"] = "a", 2, 100
            ws["A3"], ws["B3"], ws["C3"] = "b", 1, 50
        else:
            ws["D1"] = "金額"
            ws["D2"] = 200
            ws["D3"] = 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(
        ["run", str(book), "単価で並べ替えして、数量と単価を掛けた列を足して", "--copy"], capsys)
    assert rc == 0, out
    assert "根拠が見つかりません" not in out, f"正当な段に誤爆: {out}"
    assert "✓" in out, f"誤爆で ✓ が消えた: {out}"
