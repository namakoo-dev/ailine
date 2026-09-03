# 雛形が 2 枚あると × ── 2026-08-28。Namakoo「請求書のレイアウトが異なる場合には？」
#
# ★★ 実測: まったく違う配置の雛形（別様式）を 2 枚目として置き、そちらで作らせたら
#   **5 枚とも正しく埋まったのに ×** が出た。
#     × 報告シートの枚数が合いません（欠落 []・余剰(孤児シートの疑い) ['雛形']）
#
# ★ 真因は**分母を出力側から作っていた**こと:
#     「データ/雛形/検分 以外は全部この操作が作ったはず」
#   という決めつけ。ブックに前から在るシート（2 つ目の雛形・メモ・参照表）が
#   丸ごと孤児扱いになる。この repo で何度も出ている形（分母は入力側から取る）。
#   ★ 正しい分母: **前に無くて後に在る**シートだけが、この操作の産物。
#
# 契約:
#   ① 印は座標を持たない ── 配置が違っても、印さえ在れば埋まる
#   ② 使う雛形は依頼文で選べる（2 枚以上あってよい）
#   ③ 使わなかった雛形・前から在るシートは孤児ではない
#   ④ それでも**この操作が作った**宣言外のシートは今も捕まる（黙りすぎない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from _product_source import window_around  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む

DATA = [["取引先", "金額", "担当"], ["丸和物流", 57600, "田中"], ["近江スチール", 60000, "佐藤"]]


def _book(tmp_path, name="b.xlsx", extra_sheets=()):
    """データ + 雛形（縦）+ 任意の追加シート。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    for r in DATA:
        ws.append(r)
    tp = wb.create_sheet("雛形")
    tp["A1"] = "請求書"
    tp["A3"] = "{{取引先}}"
    tp["B7"] = "{{金額}}"
    tp["B14"] = "{{担当}}"
    for s in extra_sheets:
        wb.create_sheet(s)["A1"] = "メモ"
    wb.save(p)
    return p


def _finish(wb, path, rows):
    ins = wb.create_sheet("検分")
    ins.append(["シート名", "元の行", "埋めた印の数"])
    for name, r in rows:
        ins.append([name, r, 3])
    wb.save(path)
    return path


_MARKS = {"A3": 1, "B7": 2, "B14": 3}


def _args(rows, template="雛形"):
    return {"_report_rows": [{"sheet": n, "row": r} for n, r in rows],
            "_placeholders": [{"cell": "A3", "col_idx": 1, "whole": True,
                                "raw": "{{取引先}}", "column_name": "取引先"},
                               {"cell": "B7", "col_idx": 2, "whole": True,
                                "raw": "{{金額}}", "column_name": "金額"},
                               {"cell": "B14", "col_idx": 3, "whole": True,
                                "raw": "{{担当}}", "column_name": "担当"}],
            "_inspection_sheet": "検分", "template_sheet": template,
            "_target_sheet": "請求"}


def _run(tmp_path, extra_sheets=(), stray=None):
    src = _book(tmp_path, "src.xlsx", extra_sheets)
    out = _book(tmp_path, "out.xlsx", extra_sheets)
    rows = [("丸和物流", 2), ("近江スチール", 3)]
    # 雛形を 1 枚ずつコピーして印を埋めた体を作る（実機を通さず事後条件だけ試す）
    wb = openpyxl.load_workbook(out)
    srcws = wb["請求"]
    for name, r in rows:
        o = wb.copy_worksheet(wb["雛形"])
        o.title = name
        for cell, col in _MARKS.items():
            o[cell] = srcws.cell(r, col).value
    if stray:
        wb.create_sheet(stray)["A1"] = "?"
    _finish(wb, out, rows)
    return ailine.check_report_per_row(out, _args(rows), source_book=src)


# --- ③ 前から在るシートは孤児ではない -------------------------------------------------

def test_a_plain_book_still_passes(tmp_path):
    status, reason = _run(tmp_path)
    assert status == "pass", reason


def test_a_second_template_is_not_an_orphan(tmp_path):
    """★★ Namakoo の実測そのもの。2 枚目の雛形を置いただけで × になっていた。"""
    status, reason = _run(tmp_path, extra_sheets=("別様式",))
    assert status == "pass", reason


def test_unrelated_pre_existing_sheets_are_not_orphans(tmp_path):
    """★ メモ・参照表が在るだけの、ごく普通のブックで × にしない。"""
    status, reason = _run(tmp_path, extra_sheets=("メモ", "去年の分", "単価表"))
    assert status == "pass", reason


# --- ④ 黙りすぎていないこと -----------------------------------------------------------

def test_a_sheet_born_in_this_run_is_still_caught(tmp_path):
    """★ 恒久の芯: LibreOffice の copyByName は失敗時に孤児を残す。
       **この操作で生まれた**宣言外のシートは、今も捕まること。"""
    status, reason = _run(tmp_path, stray="丸和物流_1")
    assert status == "fail" and "孤児" in reason, reason


def test_a_missing_declared_sheet_is_still_caught(tmp_path):
    src = _book(tmp_path, "s2.xlsx")
    out = _book(tmp_path, "o2.xlsx")
    wb = openpyxl.load_workbook(out)
    srcws = wb["請求"]
    o = wb.copy_worksheet(wb["雛形"])
    o.title = "丸和物流"
    for cell, col in _MARKS.items():
        o[cell] = srcws.cell(2, col).value
    _finish(wb, out, [("丸和物流", 2)])
    rows = [("丸和物流", 2), ("近江スチール", 3)]
    status, reason = ailine.check_report_per_row(out, _args(rows), source_book=src)
    assert status == "fail", reason


def test_the_orphan_check_reads_the_before_book():
    """★ 分母が入力側から来ていること（出力側から作り直すと恒真に戻る）。"""
    seg = window_around("declared_sheet_names = {rr[", after=1800)
    assert "before_sheets" in seg and "BookView(source_book)" in seg, seg[:400]
    assert "born = set(bv.sheetnames) - before_sheets" in seg, seg[:400]
