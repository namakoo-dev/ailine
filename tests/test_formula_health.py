"""ailine_core/formula_health.py の単体テスト（宣言つき挙動変更 #1: 型破壊の安全網）。

★ ブラインド査定の実測: 数値の『原価』列を SET_COLUMN_VALUE で文字列『0円』に一括書換
すると、それを参照する数式（例: `=D2-C2`）が `#VALUE!` に壊れるのに、事後条件チェッカー
（対象列が指定文字列になったか、だけを見る）は「✓ 達成を機械検証済み」を出していた。

(a) formula_error_advisory: 適用の前後でエラー値セルを比較し、新たにエラーになった
    セルがあれば助言を返す一般の網。
(b) detect_write_target_type_change: OP_WRITE_TARGET が宣言する書き込み先列で、数値が
    数値に見えない文字列に変わった場合の予防的な確認行（関所にはしない・理由は
    ailine_core/formula_health.py のモジュール docstring 参照）。

★ 純ロジックのみ・LibreOffice/basrun 不要（CI で走る）。実機の basrun 通しで
「本当に #VALUE! が起きて拾えるか」を見る回帰試験は tests/test_formula_health_local.py
（@pytest.mark.local）が担う。CLI 全体（ailine.py 側の配線）を fake_apply 経由で通す
統合試験は tests/test_ailine.py（宣言つき挙動変更#1 のセクション）に置く。
"""
import re
import sys
import zipfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from ailine_core.formula_health import (   # noqa: E402
    _error_cells, formula_error_advisory, _parses_as_number, detect_write_target_type_change,
)


def _cell_ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _book(tmp_path, name, rows) -> Path:
    p = tmp_path / name
    wb = openpyxl.Workbook(); ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(p)
    return p


def _inject_error_cache(path, sheet_filename: str, addr_to_err: dict) -> None:
    """テスト専用: 数式セルへエラー値のキャッシュ(t="e")を直接注入する（openpyxl は式を
       計算しないため、LO を使わずに『保存時に LibreOffice/Excel がエラー値を計算・保存
       した』状態を再現する小道具・tests/test_ailine.py の _inject_formula_string_cache と
       同型・OOXML のエラー型は t="str" ではなく t="e"）。"""
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_filename:
                text = data.decode("utf-8")
                for addr, err in addr_to_err.items():
                    pattern = re.compile(rf'<c r="{addr}"([^>]*)>(.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?</c>')

                    def _sub(m, err=err):
                        attrs = re.sub(r'\s*t="[^"]*"', '', m.group(1))
                        return f'<c r="{addr}"{attrs} t="e">{m.group(2)}<v>{err}</v></c>'

                    text, n = pattern.subn(_sub, text, count=1)
                    assert n == 1, (
                        f"_inject_error_cache: {addr} に注入できなかった（xlsx の直列化の形が想定外の可能性）")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


# --- _error_cells / formula_error_advisory (a) -----------------------------

def test_error_cells_reads_cached_error_values(tmp_path):
    p = _book(tmp_path, "e.xlsx", [["x", "y"], [1, "=A2/0"]])
    _inject_error_cache(p, "xl/worksheets/sheet1.xml", {"B2": "#DIV/0!"})
    assert _error_cells(p) == {("Sheet", 2, 2): "#DIV/0!"}

def test_error_cells_empty_when_no_errors(tmp_path):
    p = _book(tmp_path, "e.xlsx", [["x", "y"], [1, 2]])
    assert _error_cells(p) == {}

def test_error_cells_empty_for_missing_file(tmp_path):
    assert _error_cells(tmp_path / "nope.xlsx") == {}

def test_formula_error_advisory_fires_on_new_break(tmp_path):
    before = _book(tmp_path, "before.xlsx", [["x", "y"], [1, "=A2/0"]])
    after = _book(tmp_path, "after.xlsx", [["x", "y"], [1, "=A2/0"]])
    _inject_error_cache(after, "xl/worksheets/sheet1.xml", {"B2": "#DIV/0!"})
    msg = formula_error_advisory(before, after, cell_ref=_cell_ref)
    assert len(msg) == 1
    assert msg[0].startswith("★ 疑わしい: 適用後にエラー値のセルが増えました")
    assert "Sheet!B2=#DIV/0!" in msg[0]

def test_formula_error_advisory_silent_when_error_preexisting(tmp_path):
    # 前から壊れていたセルは「新たに」壊れたのではないので何も言わない(保守的)。
    before = _book(tmp_path, "before.xlsx", [["x", "y"], [1, "=A2/0"]])
    _inject_error_cache(before, "xl/worksheets/sheet1.xml", {"B2": "#DIV/0!"})
    after = _book(tmp_path, "after.xlsx", [["x", "y"], [1, "=A2/0"]])
    _inject_error_cache(after, "xl/worksheets/sheet1.xml", {"B2": "#DIV/0!"})
    assert formula_error_advisory(before, after, cell_ref=_cell_ref) == []

def test_formula_error_advisory_silent_when_no_errors(tmp_path):
    before = _book(tmp_path, "before.xlsx", [["x", "y"], [1, 2]])
    after = _book(tmp_path, "after.xlsx", [["x", "y"], [1, 2]])
    assert formula_error_advisory(before, after, cell_ref=_cell_ref) == []

def test_formula_error_advisory_counts_beyond_five_shown(tmp_path):
    # ★ 総数が5件を超えたら先頭5件だけ列挙し「ほかN件」を添える（表示が肥大しない）。
    header = ["k"] + [f"c{i}" for i in range(7)]
    row = [1] + ["=A2/0"] * 7
    before = _book(tmp_path, "before.xlsx", [header, row])
    after = _book(tmp_path, "after.xlsx", [header, row])
    addrs = {f"{get_column_letter(c)}2": "#DIV/0!" for c in range(2, 9)}   # B2..H2 (7件)
    _inject_error_cache(after, "xl/worksheets/sheet1.xml", addrs)
    msg = formula_error_advisory(before, after, cell_ref=_cell_ref)
    assert len(msg) == 1
    assert "計7件" in msg[0]
    assert "ほか2件" in msg[0]


# --- _parses_as_number -------------------------------------------------------

def test_parses_as_number_true_for_digit_strings():
    assert _parses_as_number("500") is True
    assert _parses_as_number("3.5") is True
    assert _parses_as_number("1,200") is True   # 桁区切りは許容

def test_parses_as_number_false_for_non_numeric_text():
    assert _parses_as_number("0円") is False
    assert _parses_as_number("N/A") is False
    assert _parses_as_number("") is False
    assert _parses_as_number(None) is False


# --- detect_write_target_type_change (b) ------------------------------------

def _snap(cells: dict, sheets=("Sheet",)) -> dict:
    return {"sheets": list(sheets), "charts": 0, "cells": dict(cells),
            "merges": {s: [] for s in sheets}, "colw": {s: {} for s in sheets},
            "rowh": {s: {} for s in sheets}, "truncated": False}

def _v(val):
    return (val, "General", None, False, None, None)

def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

# ★ 単位C: 宣言の形が「列」から「領域」へ広がったので、ここの模擬も `.col_key`/`.sheet_key`
#   を持つオブジェクトにする。ailine_core は ailine を import できない（移植可能性の番人）ので、
#   本物の ailine.WriteTarget ではなく同じ属性を持つ最小の模擬を置く
#   （detect_write_target_type_change が読むのはこの2属性だけ、という契約の明示でもある）。
_OP_WRITE_TARGET = {
    "SET_COLUMN_VALUE": SimpleNamespace(col_key="col", sheet_key=None),
    "COMPUTE_COLUMN": SimpleNamespace(col_key="target", sheet_key=None),
    "SORT": SimpleNamespace(col_key=None, sheet_key=None),   # 書き込み先列は無いと確認した宣言
}

def test_detect_write_target_type_change_fires_numeric_to_nonnumeric_text():
    before = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v(300), "Sheet!3,1": _v(200)})
    after = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v("0円"), "Sheet!3,1": _v("0円")})
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["原価"]}}
    msg = detect_write_target_type_change(
        before, after, op="SET_COLUMN_VALUE", resolved={"col": "原価", "value": "0円"}, meta=meta,
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number)
    assert msg is not None
    assert "『原価』" in msg and "2 件" in msg

def test_detect_write_target_type_change_silent_for_numeric_looking_string():
    # ★ DoD 過剰検出でないこと: SET_COLUMN_VALUE は常に setString で書くため文字列型には
    #   なるが、書いた中身が数字そのもの("500")なら型変化の実害の芽ではない＝黙る。
    before = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v(300)})
    after = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v("500")})
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["原価"]}}
    msg = detect_write_target_type_change(
        before, after, op="SET_COLUMN_VALUE", resolved={"col": "原価", "value": "500"}, meta=meta,
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number)
    assert msg is None

def test_detect_write_target_type_change_silent_for_new_column():
    before = _snap({"Sheet!1,1": _v("既存")})
    after = _snap({"Sheet!1,1": _v("既存"), "Sheet!1,2": _v("新規"), "Sheet!2,2": _v("x")})
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["既存"]}}
    msg = detect_write_target_type_change(
        before, after, op="COMPUTE_COLUMN", resolved={"target": "新規"}, meta=meta,
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number)
    assert msg is None

def test_detect_write_target_type_change_silent_for_numeric_to_numeric():
    before = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v(300)})
    after = _snap({"Sheet!1,1": _v("原価"), "Sheet!2,1": _v(330)})
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["原価"]}}
    msg = detect_write_target_type_change(
        before, after, op="COMPUTE_COLUMN", resolved={"target": "原価"}, meta=meta,
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number)
    assert msg is None

def test_detect_write_target_type_change_none_for_op_without_write_target():
    msg = detect_write_target_type_change(
        _snap({}), _snap({}), op="SORT", resolved={"col": "原価"}, meta={"sheets": [], "headers": {}},
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number)
    assert msg is None

def test_detect_write_target_type_change_none_when_args_missing():
    assert detect_write_target_type_change(
        {}, {}, op=None, resolved=None, meta=None,
        op_write_target=_OP_WRITE_TARGET, is_number=_is_number) is None
