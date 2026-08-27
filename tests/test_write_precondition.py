"""単位F: 宣言した書き込み先の『前提』を、反映前に1箇所で確かめる関所の検体。

★ この検体群は**実装より先に**書いた（先に赤を見る）。後から検体を作ると、実装が通る
検体を作ってしまう ―― この repo は「凍結コーパスが誤爆を1件も含んでいなかった」を
実測している。

## 設計（ブリーフで凍結・ここでは変えない）

| writes | 前提 | 破れたら |
|---|---|---|
| new_column | その列は空 | 既存の関所 |
| new_row_at_end | 書き込んだ行は before で空だった | 関所へ |
| new_sheet | その名前のシートは before に存在しない | 関所へ |
| existing_column | （前提なし＝上書き前提） | 既存の関所 |
| format_only | 値が1つも変わらない | 関所へ |
| row_shift / reorder | 値の多重集合が保存される | 関所へ |

判定は「適用前の予測」でなく「**適用後の実測**」で行う ―― 全工程は out_book（コピー）
の上で走り、原本への反映は最後の _finish_apply。before/after が両方手元にあって原本は
まだ無傷、という窓がある。

## 検体の作り方（なぜ basrun_apply を差し替えるか）

実機の LibreOffice 往復（@pytest.mark.local）はここでは使わない。代わりに
basrun_apply を「実測した破壊そのものを再現する fake」に差し替える ―― 再現すべき挙動は
ヘルパ本体（helpers/AiLineHelpers.bas）に書いてあるとおりに写す。例:
SummaryTable は `If oDoc.Sheets.hasByName("集計") Then oDoc.Sheets.removeByName("集計")`
＝既存の『集計』シートを**消してから**作り直す（③の実測はこの1行そのもの）。
事後条件チェッカー・助言・関所は全部**本物**が走る。
"""
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

from _run_argv import run_argv  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402  — 数式キャッシュ注入の小道具を二重管理しない


# --- 土台 -------------------------------------------------------------------

def _book(tmp_path, sheets: dict, name="b.xlsx") -> Path:
    """sheets: {シート名: [[行の値...], ...]}（1つ目が既定シート）。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    first = True
    for title, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(title)
        ws.title = title
        first = False
        for row in rows:
            ws.append(row)
    wb.save(p)
    return p


def _eof(prompt=""):
    raise EOFError()   # 非対話（パイプ/CI と同じ）＝関所は exit 7 を返すべき状況


def _run(tmp_path, monkeypatch, book, task, op, args, fake_apply, **over):
    """既定（原本へ直接反映）で1回 run する。翻訳と LibreOffice だけを差し替える。"""
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    # 既定変更の一度きり告知（★ で始まる）は環境の初回性の産物で、検体の所見ではない。
    # marker を先に置いて黙らせる（★ の本数を数える邪魔をしない）。
    (tmp_path / "notice_v2_shown").write_text("shown", encoding="utf-8")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task_, book_meta, temperature=0.1: {"op": op, "args": dict(args)})
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr("builtins.input", _eof)
    return ailine.main(run_argv(book=str(book), task=task, repair=0, **over))


def _stars(out: str) -> list:
    """★ で始まる行（＝疑わしい/警告の側）の一覧。GO/NO-GO の数はこれで数える。"""
    return [ln for ln in out.splitlines() if ln.lstrip().startswith("★")]


def _cell(path: Path, sheet: str, addr: str):
    wb = openpyxl.load_workbook(path)
    try:
        return wb[sheet][addr].value
    finally:
        wb.close()


# ===========================================================================
# 真陽性側（止まること）
# ===========================================================================

def test_t1_handwritten_total_row_is_not_silent(tmp_path, monkeypatch, capsys):
    """T1: 手入力の合計行（A 列にラベル有）+「合計を出して」。

    実測（t_a）: 既存の合計行の**下**に新しい行を足し、合計に既存の合計行まで足し込む。
    ★ これは根1（表の下端の判定 _scan_last_row_basic）であって単位F の対象外 ――
    書き込んだ行そのものは before で空なので、宣言した前提は破れていない。

    ★★ 期待を書き換えた（算術恒等の検算・tests/test_sum_identity.py の T6）:
    単位F の時点での期待は「exit 0 で無言でないこと」だった ―― 追記が起きたことさえ
    画面に見えれば良い、という線。だが独立レビューが指摘したとおり、二重計上した 600 に
    `✓ 機械検証済み` が出て原本が上書きされることは、追記の告知では埋め合わせられない。
    今の期待は **exit 1・✓ を出さない・原本無傷・該当行(B4)を名指し**。
    単位F の関所（宣言した書き込み先の前提）は今もここでは鳴らない（それは正しい）――
    鳴るのは事後条件側の算術の検算。
    """
    book = _book(tmp_path, {"Sheet": [["品名", "金額"], ["あ", 100], ["い", 200], ["合計", 300]]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A5"] = "合計"
        ws["B5"] = "=SUM(B2:INDEX(B:B,ROW()-1))"
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"B5": 600})   # 300+300 の二重計上
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "✓" not in out, out
    assert "B4" in out and "300" in out, out          # 該当行を名指しする
    assert "（表の末尾への追記は意図どおりです）" in out   # 追記の告知も今までどおり出る
    assert _cell(book, "Sheet", "B5") is None        # 原本は無傷


def test_t2_existing_total_row_without_label_in_col_a_is_gated(tmp_path, monkeypatch, capsys):
    """T2: 既存合計行（A 列が空）+「合計を出して」→ exit 7・116600 が保存されている。

    実測（t_b）: D4 の値 116600 が '=SUM(...)' に**上書き**され 106000 になった（差 10,600）。
    画面には `D4: 値 116600→'=SUM(...)'` と出ているのに関所は鳴らなかった。
    """
    book = _book(tmp_path, {"Sheet": [
        ["品名", "数量", "単価", "金額"],
        ["a", 2, 50000, 100000],
        ["b", 1, 6000, 6000],
        [None, None, "合計", 116600],
    ]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        # ★ 2026-08-28: 本物の codegen はラベルを**1 列目**にも書く（旧: 対象列の左隣）。
        #   治具が本物と違うと、関所の手前で事後条件が落ちて**関所そのものを試せない**。
        ws["A4"] = "合計"
        ws["D4"] = "=SUM(D2:INDEX(D:D,ROW()-1))"   # 既存の手入力合計を潰す
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"D4": 106000})
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 7, out
    assert _cell(book, "Sheet", "D4") == 116600   # 原本は無傷


def test_t3_unrelated_existing_sheet_is_gated(tmp_path, monkeypatch, capsys):
    """T3: 無関係な手作りの『集計』シート（年度/予算）+「顧客ごとに集計して」→ exit 7。

    実測（③）: 集計シートの中身が全滅し、y/N の確認も無く exit 0。画面には
    「（既存シート『集計』の更新は意図どおりです）」という**肯定文**まで出ていた。
    """
    book = _book(tmp_path, {
        "工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]],
        "集計": [["年度", "予算"], [2025, 1000], [2026, 2000]],
    })

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        del wb["集計"]                       # SummaryTable の removeByName("集計") そのもの
        out = wb.create_sheet("集計")
        out.append(["取引先", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", 200])
        out.append(["合計", 300])
        wb.save(out_book)
        return True, None, "ok"

    # ★ 依頼文に「集計」の語を入れない: 入れると対象シートの解決が『集計』シート自身へ寄り、
    #   接地検証が先に落ちて（exit 3）適用まで行かない（実測）。③ の本体は「無関係な
    #   『集計』シートが SummaryTable に消される」ことなので、そこだけを残す。
    rc = _run(tmp_path, monkeypatch, book, "取引先ごとに金額を合計して",
              "AGGREGATE", {"group_col": "取引先", "value_col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 7, out
    assert _cell(book, "集計", "A1") == "年度"     # 原本の手作りシートは無傷
    assert _cell(book, "集計", "B2") == 1000


def _fake_pivot_sheet(out_book, code, workdir, helper_files=(), timeout=None):
    """PivotSum の出力先は『ピボット』（『集計』ではない・helpers/AiLineHelpers.bas 参照）。"""
    wb = openpyxl.load_workbook(out_book)
    if "ピボット" in wb.sheetnames:
        del wb["ピボット"]
    out = wb.create_sheet("ピボット")
    out.append(["取引先", "合計 - 金額"])
    out.append(["a", 100])
    out.append(["b", 200])
    wb.save(out_book)
    return True, None, "ok"


def _pivot_structure_check_is_faked(monkeypatch):
    """check_pivot は本物の DataPilot(xl/pivotTables/) の実在を見るが、openpyxl では
       DataPilot を作れない。その**構造確認だけ**を差し替える（他は全部本物）。"""
    monkeypatch.setitem(ailine.POSTCONDITIONS, "PIVOT",
                        lambda path, args, header_row=1: ("pass", "『ピボット』シートと DataPilot を確認"))


def test_t4_pivot_with_unrelated_sheet_does_not_fire(tmp_path, monkeypatch, capsys):
    """T4: 同じ検体を PIVOT で（★ ブリーフは exit 7 を期待・実測は exit 0）。

    ★ 外れた理由は「関所が効かない」ではない: PivotSum の出力先は『ピボット』であって
    無関係な『集計』には最初から触れない ── 宣言した前提（その名前のシートは before に
    存在しない）は破れていないので、鳴らないのが正しい。同じ op が既存シートに書く形
    （= 前提が破れる形）は下の T4b で測る。
    """
    book = _book(tmp_path, {
        "工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]],
        "集計": [["年度", "予算"], [2025, 1000], [2026, 2000]],
    })
    _pivot_structure_check_is_faked(monkeypatch)
    rc = _run(tmp_path, monkeypatch, book, "取引先ごとに金額を合計して",
              "PIVOT", {"group_col": "取引先", "value_col": "金額"}, _fake_pivot_sheet)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert _cell(book, "集計", "A1") == "年度"     # 手作りシートは最初から無関係
    assert "★ 新しいシートを作るはずが" not in out


def test_t4b_pivot_over_an_existing_pivot_sheet_is_gated(tmp_path, monkeypatch, capsys):
    """T4b: 既に『ピボット』シートがあるブックへ PIVOT → exit 7。

    ★ 関所が op ごとの白名簿でなく宣言（writes=new_sheet）駆動であることの実証:
    AGGREGATE で書いた1行も if も足していないのに、PIVOT でも同じ理由で鳴る。
    """
    book = _book(tmp_path, {
        "工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]],
        "ピボット": [["前回のメモ", "手入力"], ["残しておきたい", 999]],
    })
    _pivot_structure_check_is_faked(monkeypatch)
    rc = _run(tmp_path, monkeypatch, book, "取引先ごとに金額を合計して",
              "PIVOT", {"group_col": "取引先", "value_col": "金額"}, _fake_pivot_sheet)
    out = capsys.readouterr().out
    assert rc == 7, out
    assert _cell(book, "ピボット", "A2") == "残しておきたい"   # 原本は無傷


def test_t5_aggregate_onto_the_summary_sheet_itself_stops_at_the_postcondition(
        tmp_path, monkeypatch, capsys):
    """T5: 『集計』シート自身を対象に集計（★ ブリーフは exit 7 or CLARIFY・実測は exit 1）。

    ★ 集計元を出力で潰すので、事後条件（分類列/集計列を集計元で探す）が先に落ちる。
    関所より手前で止まる ―― 原本は無傷で .out に結果が残る、という既存の失敗の形。
    単位F の関所は**事後条件が通った結果**だけを対象にする（通らなかった結果に
    「承知して続行しますか」と聞くのは、壊れた結果の適用を勧めることになる）。
    """
    book = _book(tmp_path, {"集計": [["取引先", "金額"], ["a", 100], ["b", 200]]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        del wb["集計"]                       # 集計元がそのまま消える（helper の removeByName）
        out = wb.create_sheet("集計")
        out.append(["取引先", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", 200])
        out.append(["合計", 300])
        wb.save(out_book)
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "取引先ごとに集計して",
              "AGGREGATE", {"group_col": "取引先", "value_col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "事後条件を満たさない" in out
    assert _cell(book, "集計", "B1") == "金額"   # 原本は無傷


# ===========================================================================
# 偽陽性側（誤爆が復活しないこと）── ここを凍結しないと同じ後退を繰り返す
# ===========================================================================

def test_f1_plain_table_append_total_stays_silent(tmp_path, monkeypatch, capsys):
    """F1: 合計行なしの普通の表 +「合計を出して」→ 中立表示のみ・★ 0 本・exit 0。"""
    book = _book(tmp_path, {"Sheet": [["品名", "金額"], ["あ", 100], ["い", 200]]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A4"] = "合計"
        ws["B4"] = "=SUM(B2:INDEX(B:B,ROW()-1))"
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"B4": 300})
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "（表の末尾への追記は意図どおりです）" in out
    assert _stars(out) == []


def test_f2_aggregate_into_a_brand_new_sheet_stays_silent(tmp_path, monkeypatch, capsys):
    """F2: 単一シート +「取引先ごとに集計して」→ 中立表示のみ・★ 0 本。"""
    book = _book(tmp_path, {"工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        out = wb.create_sheet("集計")
        out.append(["取引先", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", 200])
        out.append(["合計", 300])
        wb.save(out_book)
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "取引先ごとに集計して",
              "AGGREGATE", {"group_col": "取引先", "value_col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "（新規シート『集計』の作成は意図どおりです）" in out
    # ★ この1本は単位F より前から出ている②の範囲注記（依頼文が『金額』と言っていない）。
    #   ここで凍結しておかないと「★ が1本ある」が単位F の誤爆と読み違えられる。
    assert _stars(out) == [
        "★ ただし対象『金額』は依頼文の語と機械照合していません"
        "（ブックの実体・既定から機械決定しました） — 「解釈:」行を確認してください。"]


def test_f3_sheet_name_colliding_with_column_name_stays_silent(tmp_path, monkeypatch, capsys):
    """F3: シート名と列名が衝突するブック（誤爆#3 の検体）→ ★ 0 本。"""
    book = _book(tmp_path, {
        "売上データ": [["商品", "金額"], ["a", 200], ["b", 300]],
        "金額": [["月", "金額"], ["1月", 50]],
    })

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["売上データ"]
        ws["A2"], ws["B2"] = "b", 300      # 降順に並べ替えた後
        ws["A3"], ws["B3"] = "a", 200
        wb.save(out_book)
        return True, None, "ok"

    rc = _run(tmp_path, monkeypatch, book, "金額を降順に並べ替えて",
              "SORT", {"col": "金額", "order": "desc"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert _stars(out) == []


def test_f4_aggregate_twice_in_the_same_book(tmp_path, monkeypatch, capsys):
    """F4: AGGREGATE を同じブックで2回連続実行。

    ★ これは単位F 単体では落ちる見込み（2回目は『集計』が既に存在する＝宣言した前提が
    破れている）。落ちるかどうかを測って報告するための検体であって、直さない。
    """
    book = _book(tmp_path, {"工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        if "集計" in wb.sheetnames:
            del wb["集計"]
        out = wb.create_sheet("集計")
        out.append(["取引先", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", 200])
        out.append(["合計", 300])
        wb.save(out_book)
        return True, None, "ok"

    # ★ 2回目に「集計」の語を含む依頼を出すと、対象シートの解決が1回目の出力シートへ
    #   寄って接地検証で止まる（exit 3・T3 と同じ実測）。それでは「2回目に前提が破れるか」
    #   という問いが測れないので、語を避けて適用まで届かせる。
    args = ("取引先ごとに金額を合計して", "AGGREGATE",
             {"group_col": "取引先", "value_col": "金額"}, fake)
    rc1 = _run(tmp_path, monkeypatch, book, *args)
    capsys.readouterr()
    rc2 = _run(tmp_path, monkeypatch, book, *args)
    out2 = capsys.readouterr().out
    assert rc1 == 0
    assert rc2 == 0, out2
    # ★ 1回目の出力で2シートになるため、2回目は②の範囲注記が1本出る（単位F より前からの挙動）。
    assert _stars(out2) == [
        "★ ただし対象シート『工事台帳』は依頼文の語と機械照合していません"
        "（ブックの実体・既定から機械決定しました） — 「解釈:」行を確認してください。"]


def test_f4b_aggregate_twice_with_edited_source_data(tmp_path, monkeypatch, capsys):
    """F4 の内訳を分ける検体: 2回目の集計結果が1回目と**違う**場合。

    ★ F4 が鳴らないのは「2回目だから」ではなく「書いた値が前回と同じで、実測した変化が
    ゼロだから」。間で元データを直すと同じ2回目が鳴る ── 再集計は日常の使い方なので、
    ここは単位F の誤爆の本命候補として測って残す（この回では直さない）。

    ★★ 単位H（2026-08-19）で **主張を反転させた**。単位F が「この回では直さない」と書いて
    凍結した誤爆が、まさにここで直る対象だった。
    出力先『集計』の見出しが SummaryTable 自身の署名（A1=分類列名 / B1="合計 - 集計列名"・
    helpers/AiLineHelpers.bas:374-375）と一致するので「自分の前回の出力の作り直し」と判定し、
    前提は破れていない ＝ 関所を鳴らさない。
    ★ 人が手で作った『集計』（署名が一致しない）は今までどおり鳴る
    ── その対照は tests/test_write_precondition_unit.py と
    tests/golden/f9_transcripts/unit_g_declared_sheet_premise_broken.txt が持つ。
    """
    book = _book(tmp_path, {"工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]]})
    totals = {"b": 200}

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        if "集計" in wb.sheetnames:
            del wb["集計"]
        out = wb.create_sheet("集計")
        out.append(["取引先", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", totals["b"]])
        out.append(["合計", 100 + totals["b"]])
        wb.save(out_book)
        return True, None, "ok"

    args = ("取引先ごとに金額を合計して", "AGGREGATE",
             {"group_col": "取引先", "value_col": "金額"}, fake)
    assert _run(tmp_path, monkeypatch, book, *args) == 0
    capsys.readouterr()

    wb = openpyxl.load_workbook(book)      # 利用者が元データを直した
    wb["工事台帳"]["B3"] = 500
    wb.save(book)
    totals["b"] = 500

    rc2 = _run(tmp_path, monkeypatch, book, *args)
    out2 = capsys.readouterr().out
    # ★ 単位H 以前は rc2 == 7（誤爆）。以後は正常な再集計として通す。
    assert rc2 == 0, out2
    assert "★ 新しいシートを作るはずが" not in out2, out2


# --- F5: format_only 系 8 op を健全系で1回ずつ ------------------------------------

_F5_ROWS = [["品名", "金額"], ["a", 100], ["b", 200]]


def _f5_bold(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    for c in ("A1", "B1"):
        ws[c].font = Font(bold=True)
    wb.save(out_book)
    return True, None, "ok"


def _f5_fill(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    for c in ("A1", "B1"):
        ws[c].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    wb.save(out_book)
    return True, None, "ok"


def _f5_number_format(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    for c in ("B2", "B3"):
        ws[c].number_format = "#,##0"
    wb.save(out_book)
    return True, None, "ok"


def _f5_merge(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    wb.active.merge_cells("A1:B1")
    wb.save(out_book)
    return True, None, "ok"


def _f5_chart(out_book, code, workdir, helper_files=(), timeout=None):
    # ★ グラフ段: 本物のヘルパ(InsertChart)は項目名列(c:cat)も必ずセットする
    #   （addNewByName に2本の CellRangeAddress を渡す ── 値だけの1本にはならない）。
    #   check_chart_series の恒真殺しがそこまで見るようになったので、fake も同じ形に合わせる。
    #   ★ openpyxl の set_categories() は常に numRef を作る（ChartBase.set_categories
    #   実装）が、実 LO は文字列カテゴリを strRef で書く（fixtures/charts/*.xlsx で実測済み・
    #   ailine_core/chart_check.py が見るのも c:cat/c:strRef）。fake をその形に合わせて
    #   AxDataSource(strRef=...) を手で組む。
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=3)
    for s in ch.series:
        s.cat = AxDataSource(strRef=StrRef(f=str(cat_ref)))
    ws.add_chart(ch, "D2")
    wb.save(out_book)
    return True, None, "ok"


def _f5_center(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    wb.save(out_book)
    return True, None, "ok"


def _f5_borders(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.save(out_book)
    return True, None, "ok"


def _f5_autofit(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ws.column_dimensions["A"].width = 18.5
    ws.column_dimensions["B"].width = 12.25
    wb.save(out_book)
    return True, None, "ok"


_F5_CASES = {
    "BOLD": ("見出しを太字にして", {"target": "row:1"}, _f5_bold, _F5_ROWS),
    "FILL_COLOR": ("見出しを赤で塗って", {"target": "row:1", "color": "red"}, _f5_fill, _F5_ROWS),
    "NUMBER_FORMAT": ("金額に桁区切りを付けて", {"col": "金額", "style": "thousands"},
                       _f5_number_format, _F5_ROWS),
    # ★ 健全系の結合＝左上以外が空（タイトル行を横に伸ばす、という実際の用途）。
    #   左上以外に値がある結合は「値が消える」ので別の検体（下の occupied 版）で測る。
    "MERGE": ("A1:B1 を結合して", {"range": "A1:B1"}, _f5_merge,
               [["タイトル", None], ["品名", "金額"], ["a", 100]]),
    "CHART": ("金額のグラフを作って", {"value_col": "金額"}, _f5_chart, _F5_ROWS),
    "CENTER_ALIGN": ("全体を中央揃えにして", {"target": "all"}, _f5_center, _F5_ROWS),
    "DRAW_BORDERS": ("罫線を引いて", {}, _f5_borders, _F5_ROWS),
    "AUTOFIT": ("列幅を自動調整して", {}, _f5_autofit, _F5_ROWS),
}


@pytest.mark.parametrize("op", sorted(_F5_CASES))
def test_f5_format_only_ops_stay_silent(tmp_path, monkeypatch, capsys, op):
    """F5: format_only 系 8 op を健全系で1回ずつ → ★ 0 本。鳴ったら新しい発見。"""
    task, args, fake, rows = _F5_CASES[op]
    book = _book(tmp_path, {"Sheet": rows})
    rc = _run(tmp_path, monkeypatch, book, task, op, args, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert _stars(out) == [], out


def test_f5b_merge_over_occupied_cells_now_gates(tmp_path, monkeypatch, capsys):
    """★★ F5 の周辺で見つかった**新しい発見**（健全系 8 op は全部沈黙している・上の F5）。

    左上以外に値がある範囲を結合すると、その値は読めなくなる（openpyxl は結合範囲の
    左上以外を必ず None として読む＝観測側の性質でもあり、xlsx 上も表示は消える）。
    MERGE は writes=format_only と宣言しているので、前提「値が1つも変わらない」がここで
    破れ、関所が鳴る（exit 7）。

    ★ この回では直さない（ブリーフの「やらないこと」）。判断の材料として2つ:
      (a) 正しい検出だと見る ── 『金額』という見出しは実際に消えるので、確認を挟むのは妥当。
      (b) 誤爆だと見る ── 結合は宣言どおりの効果であり、MERGE の宣言が format_only で
          あること自体が間違い（値を畳む op なので別の種類が要る）。
    どちらを採るかで直す場所が変わる（(a) は何もしない・(b) は OP_WRITE_TARGET の宣言）。
    """
    book = _book(tmp_path, {"Sheet": [["品名", "金額"], ["a", 100], ["b", 200]]})
    rc = _run(tmp_path, monkeypatch, book, "A1:B1 を結合して", "MERGE", {"range": "A1:B1"}, _f5_merge)
    out = capsys.readouterr().out
    assert rc == 7, out
    assert "★ 書式だけのはずが、セルの値が 1 件変わりました（Sheet!B1: '金額' → (空)）" in out
    assert _cell(book, "Sheet", "B1") == "金額"   # 原本は無傷
