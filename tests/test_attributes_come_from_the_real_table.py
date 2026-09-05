"""語が実表の何を指すかは、**実表の証拠からしか**出ないこと（2026-09-05）。

★★ 出所（Namakoo の設計）:
  「疑わしい単語に対し登録をしてもらおう。登録する属性は有限と考えられるから
    列挙でいいが**無関係な属性は表示させないように工夫する必要がある**」

  ★ 「無関係を出さない」の答えは**選択肢を実表から作る**こと ── 実表に証拠が無い
    属性はそもそも候補に載らない。

★★ Namakoo「属性は 4 つで足りるか？」「LLM が操作を完遂できるだけでいい」への答え:
  道具が語を接地できる先は 12 種で閉じている。うち 5 種（数・単位・番地・日付）は
  残差に出ず、2 種（扱える操作・書式の名前）は**既存の器官が持つ**（別名ストア・
  op の照合語彙）── そこに足すと二重化して片配線を生む。残るのが 4＋3。
  実測 29 語のうち当たらなかった 12 語の内訳は「op の語彙の穴 7 / 用語集 2 /
  別ファイル 1 / 0 が正解 1 / 未対応 1」── **完遂を塞いでいるのは属性より op の語彙**。
  ★ だから「4 で足りるか」より「足りないと分かった時に安く足せるか」が本題で、
    ④ が**名簿に 1 行**であることを機械で縛る。

★ 実装中に 2 つ直した（どちらも実測から）:
  ・部分一致で断定していた ── 「単価の平均値を…」で『単価』を列『平均単価』と読んだ。
    強い証拠（ぴったり）と弱い証拠（部分一致）を分け、**断定は強い証拠だけ**に（②）。
  ・同じ読み方がシート違いで 2 つ並んでいた（『ボルト』）── 畳んだ（③）。
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import ailine  # noqa: E402
from ailine_core import attributes as A  # noqa: E402
from ailine_core.op_axes import AXES  # noqa: E402

NL = chr(10)
SHEETS = ["売上一覧", "在庫"]
HEADERS = {"売上一覧": ["品名", "数量", "平均単価", "金額"], "在庫": ["品名", "在庫数"]}
LACKS = {}
for _ax in AXES.values():
    LACKS.update(getattr(_ax, "lacks", {}) or {})


@pytest.fixture
def book(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上一覧"
    ws.append(HEADERS["売上一覧"])
    for r in [["ボルト", 10, 50, 500], ["ナット", 5, 20, 100], ["ボルト", 3, 50, 150]]:
        ws.append(r)
    wb.create_sheet("在庫")
    wb["在庫"].append(HEADERS["在庫"])
    wb["在庫"].append(["ボルト", 99])
    p = tmp_path / "t.xlsx"
    wb.save(p)
    return p


@pytest.fixture
def material(book):
    return {"sheets": SHEETS, "headers": HEADERS,
            "samples": A.sample_columns(book, HEADERS), "lacks": LACKS}


def _cands(word, material):
    return A.candidates_for(word, **material)


# --- ① 実表に証拠のあるものだけが候補になる --------------------------------

@pytest.mark.parametrize("word, want", [
    ("平均", 2),      # 列『平均単価』＋ まだ無い操作『平均』── ★ 聞くべき語
    ("ボルト", 1),    # 品名 列の値（2 シートに在るが**同じ読み方**なので 1 つ）
    ("金額", 1),      # 列
    ("原価", 0),      # ★ 実表に手がかりが無い ── 候補を捏造しない
    ("ぬるぽ", 0),
])
def test_only_what_the_table_can_witness(word, want, material):
    assert len(_cands(word, material)) == want


def test_no_attribute_appears_without_evidence(material):
    """★ Namakoo の要求そのもの ── どの候補も、実表のどこかを指していること。"""
    for word in ("平均", "ボルト", "金額", "品名", "在庫"):
        for c in _cands(word, material):
            assert c.kind in A.KINDS
            assert c.label, c
            if c.kind == A.KIND_SHEET:
                assert c.label in SHEETS
            elif c.kind == A.KIND_COLUMN:
                assert c.label in HEADERS.get(c.sheet, []) or "・" in c.sheet
            elif c.kind == A.KIND_OP_LACKING:
                assert c.label in set(LACKS.values())


# --- ② 部分一致で断定しない（★ 実測で出た誤読）-----------------------------

def test_a_partial_match_is_weak_evidence(material):
    """『単価』は列『平均単価』に触れるが**ぴったりではない** ── 断定してはいけない。"""
    got = _cands("単価", material)
    assert len(got) == 1 and got[0].label == "平均単価"
    assert got[0].exact is False, "部分一致を強い証拠として扱っている"


def test_an_exact_match_is_strong_evidence(material):
    for word in ("金額", "品名", "ボルト"):
        got = _cands(word, material)
        assert got and any(c.exact for c in got), (word, got)


def test_the_tool_stays_silent_on_weak_evidence_alone(book, capsys):
    """★ 実測の再現形: 「単価の平均値を…」で『単価』を『平均単価』と読んでいた。"""
    a = argparse.Namespace(task="単価の列を消して", json=False, dry=False)
    said = ailine._ask_about_a_suspicious_word(a, book, ailine.build_book_meta(book))
    out = capsys.readouterr().out
    assert not said and "平均単価" not in out, out


# --- ③ 同じ読み方がシート違いで並ばない ------------------------------------

def test_the_same_reading_across_sheets_is_one_choice(material):
    got = _cands("ボルト", material)
    assert len(got) == 1
    assert "在庫" in got[0].sheet and "売上一覧" in got[0].sheet
    assert "計 3 件" in got[0].note, got[0]


# --- ④ 属性を足すのは名簿に 1 行（★ 片配線を作らない）----------------------

def test_adding_an_attribute_is_one_line_in_the_table(material):
    """名簿に 1 行足すだけで、候補作りも表示文も**触らずに**効くこと。"""
    kind = A.Kind("試験用", lambda w, ctx: [(w + "の実体", "", "", True)],
                  lambda c: "試験用『" + c.label + "』")
    A.KIND_TABLE["_test_kind"] = kind
    try:
        got = [c for c in _cands("原価", material) if c.kind == "_test_kind"]
        assert len(got) == 1, "名簿に足しただけでは候補に出ない（配線が散っている）"
        assert got[0].describe() == "試験用『原価の実体』", got[0].describe()
    finally:
        A.KIND_TABLE.pop("_test_kind", None)


def test_the_order_of_kinds_comes_from_the_table():
    assert A.KINDS == tuple(A.KIND_TABLE), "並びの出所が名簿でない（2 箇所になる）"


# --- ⑤ 覚える層（★ 別名ストアの流儀）--------------------------------------

def test_it_remembers_and_reads_back(book, tmp_path):
    store = tmp_path / "attrs.json"
    key = ailine._book_key_now(book, ailine.build_book_meta(book))
    cand = A.Candidate(A.KIND_COLUMN, "平均単価", sheet="売上一覧", exact=False)
    ok, _msg = ailine.attr_add("平均", cand, key, store)
    assert ok
    got = ailine.lookup_attribute("平均", key, store)
    assert got and got["label"] == "平均単価"


def test_a_broken_store_does_not_crash(tmp_path):
    p = tmp_path / "attrs.json"
    p.write_text("{ これは JSON ではない", encoding="utf-8")
    assert ailine.load_attribute_entries(p) == []
    p.write_text(json.dumps({"entries": [{"word": "x"}, 3, None]}), encoding="utf-8")
    assert ailine.load_attribute_entries(p) == []


def test_undo_takes_back_the_last_registration(book, tmp_path):
    store = tmp_path / "attrs.json"
    key = ailine._book_key_now(book, ailine.build_book_meta(book))
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価"), key, store)
    ailine.attr_add("単価", A.Candidate(A.KIND_COLUMN, "平均単価"), key, store)
    ok, _ = ailine.attr_undo(store)
    assert ok
    assert ailine.lookup_attribute("単価", key, store) is None
    assert ailine.lookup_attribute("平均", key, store) is not None


def test_remove_takes_back_a_named_word(book, tmp_path):
    store = tmp_path / "attrs.json"
    key = ailine._book_key_now(book, ailine.build_book_meta(book))
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価"), key, store)
    assert ailine.attr_remove("平均", store)[0]
    assert ailine.attr_remove("平均", store)[0] is False


# --- ⑥ 鍵（パスと署名の両方・どちらか合えば引く）---------------------------

def test_either_half_of_the_key_is_enough():
    a = {"path": "C:/x/t.xlsx", "sig": "1111"}
    assert A.key_matches(a, {"path": "C:/x/t.xlsx", "sig": "9999"})   # 動かした後に列が変わった
    assert A.key_matches(a, {"path": "C:/y/t.xlsx", "sig": "1111"})   # 名前を変えた
    assert not A.key_matches(a, {"path": "C:/y/t.xlsx", "sig": "9999"})
    assert not A.key_matches(a, None)


def test_a_different_book_does_not_borrow_the_registration(book, tmp_path):
    store = tmp_path / "attrs.json"
    key = ailine._book_key_now(book, ailine.build_book_meta(book))
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価"), key, store)
    wb = openpyxl.Workbook()
    wb.active.append(["氏名", "点数"])
    other = tmp_path / "other.xlsx"
    wb.save(other)
    task, notes = ailine.apply_known_attributes("平均を並べ替えて", other,
                                                 ailine.build_book_meta(other), store)
    assert task == "平均を並べ替えて" and not notes


# --- ⑦ 使う時に実表で検算する（★ 論点 E ── 登録は検証の免除ではない）------

def test_a_registration_that_no_longer_holds_is_dropped(book, material):
    entry = {"word": "平均", "kind": A.KIND_COLUMN, "label": "平均単価", "sheet": "売上一覧",
             "book": {"path": str(book), "sig": ""}}
    assert A.entry_still_holds(entry, **material)
    gone = dict(material, headers={"売上一覧": ["品名", "数量", "金額"]}, samples={})
    assert not A.entry_still_holds(entry, **gone), "消えた列を指す登録がまだ成り立っている"


def test_the_rewrite_checks_the_table_before_using_a_registration(book, tmp_path):
    store = tmp_path / "attrs.json"
    bm = ailine.build_book_meta(book)
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価", sheet="売上一覧"),
                    ailine._book_key_now(book, bm), store)
    task, notes = ailine.apply_known_attributes("平均を並べ替えて", book, bm, store)
    assert task == "平均単価を並べ替えて" and notes
    # ★ 列を消したブックでは、同じ登録が使われないこと（パスは同じ＝鍵は当たる）
    wb = openpyxl.load_workbook(book)
    wb["売上一覧"].delete_cols(3)
    wb.save(book)
    bm2 = ailine.build_book_meta(book)
    task2, notes2 = ailine.apply_known_attributes("平均を並べ替えて", book, bm2, store)
    assert task2 == "平均を並べ替えて" and not notes2, (task2, notes2)


# --- ⑧ 言い換えの安全側 ----------------------------------------------------

def test_it_does_not_touch_a_task_that_already_names_the_column(book, tmp_path):
    store = tmp_path / "attrs.json"
    bm = ailine.build_book_meta(book)
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価", sheet="売上一覧"),
                    ailine._book_key_now(book, bm), store)
    task, notes = ailine.apply_known_attributes("平均単価の平均を出して", book, bm, store)
    assert task == "平均単価の平均を出して" and not notes, "実体が既に在るのに書き換えた"


def test_a_word_that_is_only_a_fragment_is_not_rewritten(book, tmp_path):
    """★ 変異試験がすり抜けた穴（2026-09-05）: 断片ガードを外しても番人が全緑だった。

      登録『平均』→ 列『平均単価』のとき、依頼「**平均値**を出して」を素朴に置換すると
      「平均単価値を出して」という**存在しない語**になる。他の語の断片としてしか
      現れない語は書き換えない（別名ストアと同じ判定を使う）。
    """
    store = tmp_path / "attrs.json"
    bm = ailine.build_book_meta(book)
    ailine.attr_add("平均", A.Candidate(A.KIND_COLUMN, "平均単価", sheet="売上一覧"),
                    ailine._book_key_now(book, bm), store)
    task, notes = ailine.apply_known_attributes("平均値を一番下に出して", book, bm, store)
    assert task == "平均値を一番下に出して" and not notes, (task, notes)
    # ★ 対で縛る: 語として立っている回は従来どおり書き換わる
    ok_task, ok_notes = ailine.apply_known_attributes("平均を出して", book, bm, store)
    assert ok_task == "平均単価を出して" and ok_notes


def test_only_column_registrations_change_the_task(book, tmp_path):
    store = tmp_path / "attrs.json"
    bm = ailine.build_book_meta(book)
    ailine.attr_add("売上", A.Candidate(A.KIND_SHEET, "売上一覧", sheet="売上一覧"),
                    ailine._book_key_now(book, bm), store)
    task, notes = ailine.apply_known_attributes("売上を並べ替えて", book, bm, store)
    assert task == "売上を並べ替えて" and not notes


# --- ⑨ 聞く側（★ 注入できないと番人から見えない）--------------------------

def test_choosing_a_reading_registers_it(book, tmp_path, capsys):
    store = tmp_path / "attrs.json"
    a = argparse.Namespace(task="平均を出して", json=False, dry=False)
    said = ailine._ask_about_a_suspicious_word(a, book, ailine.build_book_meta(book),
                                                input_fn=lambda q: "1", attr_path=store)
    out = capsys.readouterr().out
    assert said and "決められません" in out and "登録:" in out, out
    entries = ailine.load_attribute_entries(store)
    assert len(entries) == 1 and entries[0]["word"] == "平均"


def test_it_does_not_ask_when_it_cannot(book, tmp_path, capsys):
    """--json は機械が読む ── プロンプトも選択肢も混ぜない。"""
    a = argparse.Namespace(task="平均を出して", json=True, dry=False)
    said = ailine._ask_about_a_suspicious_word(a, book, ailine.build_book_meta(book),
                                                attr_path=tmp_path / "attrs.json")
    assert not said and capsys.readouterr().out == ""


def test_a_wrong_answer_registers_nothing(book, tmp_path, capsys):
    store = tmp_path / "attrs.json"
    a = argparse.Namespace(task="平均を出して", json=False, dry=False)
    ailine._ask_about_a_suspicious_word(a, book, ailine.build_book_meta(book),
                                         input_fn=lambda q: "9", attr_path=store)
    capsys.readouterr()
    assert ailine.load_attribute_entries(store) == []


# --- ⑩ 配線（★ 純ロジックだけでは守られない）------------------------------

def test_the_refusal_gate_asks_about_the_word():
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    gate = src.split("def cmd_refuse_vocab_miss")[1].split(NL + "def ")[0]
    assert "_ask_about_a_suspicious_word(" in gate, "断りの門に配線されていない"


def test_the_run_path_applies_what_was_learned():
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src.split("def _translate_and_dispatch")[1].split(NL + "def ")[0]
    assert "apply_known_attributes(" in body, "翻訳の前で登録が使われていない"
    head = body.split("apply_known_attributes(")[0]
    assert "translation is None" in head[-400:], "翻訳を使い回す回でも書き換えている"


def test_the_word_extractor_is_not_a_second_implementation():
    """★ 残差検出を 2 つ持たない（持つと片方だけ直る）。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src.split("def suspicious_words")[1].split(NL + "def ")[0]
    assert "suggest_residue.find_unconsumed_words(" in body
