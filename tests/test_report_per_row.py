# 帳票段 REPORT_PER_ROW ── 実装より先に凍結した赤い検体。
# 設計: ~/.nagi/plans/DESIGN-20260823-report-per-row.md（spike の実測で 5 箇所訂正済み）。
# 実需: MARKET-20260823-lancers.md の財務書類 7 件（すべて「表 → 依頼者が既に持っている
# 定型フォーマットへの転写を N 回」）。決裁: Namakoo「B だ」(2026-08-23 17:53)。
#
# ★ 憲法の適用: **雛形は人が作る。機械は埋めるだけ。**
#   LLM はデータに触らないを見た目にも拡張する ── 機械が触ってよいのは印のあるセルだけ。
#
# 契約（✓ の 4 本柱）:
#   ① 枚数の完全会計: データ行 N = 出力 N 枚 + 除外（合計行は total_row の既存機構で除く）
#   ② 値の 3 計数: 各枚の印セル vs 元の行 ── 欠落0・不一致0・余剰0（型込み等値）
#   ③ 出所: 検分シートに シート名/元行 の一覧（★訂正4: stack の出所列は 1行=1枚に合わない）
#   ④ 雛形の不変: 印**以外**のセルが 1 つも変わらない（★訂正3: 型を正しく埋めた場合に限る）
#
# spike の実測による契約（守らないと静かに壊れる）:
#   A. シート名は Excel の規則で機械が自分で切る（禁止文字 \ / ? * [ ] : ・**31 文字**）
#      ★ 31 文字超を LO は無警告で切り詰める（例外も返り値も無い）
#   B. copyByName は失敗時に**孤児シートを残す** → 失敗しうる名前を渡さない（再試行禁止）
#   C. 型は**元データ列**で決める（数値列は setValue）。雛形の数値書式には触らない
#      ── 誤ると値だけでなく number_format まで壊れる実測あり

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "sanitize_sheet_name"),
    reason="帳票段 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


def _invoice_book(tmp_path, name="請求.xlsx"):
    """データ表（取引先 3 件 + 合計行）と、人が作った雛形シート。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in [["取引先", "金額", "日付"],
               ["あかつき商事", 12000, "2026-07-31"],
               ["みどり工業", 8500, "2026-07-31"],
               ["そら建設", 30000, "2026-07-31"],
               ["合計", 50500, ""]]:
        ws.append(r)
    tpl = wb.create_sheet("雛形")
    tpl["A1"] = "請求書"
    tpl["A3"] = "宛先"
    tpl["B3"] = "{{取引先}}"
    tpl["A5"] = "ご請求金額"
    tpl["B5"] = "{{金額}}"
    tpl["B5"].number_format = "#,##0"
    tpl["A7"] = "発行日"
    tpl["B7"] = "{{日付}}"
    tpl["A9"] = "備考"
    tpl["B9"] = "毎度ありがとうございます"      # 印の無いセル（柱④の対象）
    wb.save(p)
    return p


# --- A. シート名の作法（spike 実測・LO は無警告で切る）------------------------------

@needs_impl
@pytest.mark.parametrize("raw,expected", [
    ("あかつき商事", "あかつき商事"),
    ("得意先/AB", "得意先_AB"),                      # Excel の禁止文字
    ("a\\b?c*d[e]f:g", "a_b_c_d_e_f_g"),
    ("あ" * 40, "あ" * 31),                          # ★ 31 文字で機械が自分で切る
    ("", "_"),                                        # 空は使えない
])
def test_sanitize_sheet_name(raw, expected):
    assert ailine.sanitize_sheet_name(raw) == expected


@needs_impl
def test_sheet_name_collision_resolved_before_calling_lo():
    """B: 衝突は**呼ぶ前に**解く（copyByName に失敗しうる名前を渡さない）。
       既存名を渡された時、31 文字を超えない一意名を返すこと。"""
    used = {"あかつき商事", "あかつき商事_2"}
    got = ailine.unique_sheet_name("あかつき商事", used)
    assert got not in used
    assert len(got) <= 31
    long_used = {"あ" * 31}
    got2 = ailine.unique_sheet_name("あ" * 40, long_used)
    assert got2 not in long_used and len(got2) <= 31, f"31 文字を超えた: {got2!r}"


# --- ①〜④ の柱（e2e）----------------------------------------------------------

@needs_impl
def test_report_per_row_end_to_end(tmp_path, monkeypatch, capsys):
    """データ 3 行（+合計行）→ 3 枚。合計行は除外・値は正しい型・雛形の非印セルは不変。"""
    _isolate(monkeypatch, tmp_path)
    book = _invoice_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "REPORT_PER_ROW",
         "args": {"template_sheet": "雛形", "name_col": "取引先"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        # 実 LO の代わりに、雛形を複製して印を埋めた形を作る（型は元列に従う）
        wb = openpyxl.load_workbook(out_book)
        src = wb["売上"]
        tpl = wb["雛形"]
        rows = [(src.cell(r, 1).value, src.cell(r, 2).value, src.cell(r, 3).value)
                 for r in range(2, 5)]           # 合計行(5 行目)は含めない
        for name, amount, day in rows:
            sh = wb.copy_worksheet(tpl)
            sh.title = name
            sh["B3"] = name
            sh["B5"] = amount                    # ★ 数値は数値のまま
            sh["B7"] = day
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "売上表から取引先ごとに請求書を作って", "--copy"],
                         capsys)
    assert rc == 0, out
    got = openpyxl.load_workbook(book.with_suffix("")._str + ".out.xlsx") \
        if False else openpyxl.load_workbook(str(book).replace(".xlsx", ".out.xlsx"))
    # ① 枚数の完全会計（合計行は積まない）
    made = [s for s in got.sheetnames if s not in ("売上", "雛形") and not s.startswith("検分")]
    assert len(made) == 3, f"枚数が合わない（合計行を積んだ疑い）: {got.sheetnames}"
    assert "合計" not in made
    # ② 値と型
    sh = got["あかつき商事"]
    assert sh["B3"].value == "あかつき商事"
    assert sh["B5"].value == 12000, f"金額が数値でない（下流の SUM が壊れる）: {sh['B5'].value!r}"
    # ④ 雛形の非印セルは不変（値も数値書式も）
    assert sh["A1"].value == "請求書" and sh["B9"].value == "毎度ありがとうございます"
    assert sh["B5"].number_format == "#,##0", \
        f"雛形の書式が壊れた（型を誤ると number_format ごと壊れる実測）: {sh['B5'].number_format}"
    # ③ 出所は検分シートに
    assert any(s.startswith("検分") for s in got.sheetnames), \
        f"検分シートが無い（出所の置き場）: {got.sheetnames}"
    assert "✓" in out, out


@needs_impl
def test_placeholder_naming_unknown_column_refuses(tmp_path, monkeypatch, capsys):
    """雛形の印が実在しない列名を指していたら、埋める前に断る（幻覚の封鎖）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "金額"]); ws.append(["a", 100])
    tpl = wb.create_sheet("雛形")
    tpl["B3"] = "{{存在しない列}}"
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "REPORT_PER_ROW", "args": {"template_sheet": "雛形", "name_col": "取引先"}})
    rc, out = _run_main(["run", str(p), "取引先ごとに請求書を作って", "--copy"], capsys)
    assert rc != 0
    assert "存在しない列" in out, f"どの印が悪いか名指ししていない: {out}"


# --- ② の照合関数（印セルだけに絞る・★訂正5: CSV 版の直接転用は不可）----------------

@needs_impl
def test_compare_report_cells_only_ignores_template_cells(tmp_path):
    """雛形由来の非印セル（見出し・固定文）を『余剰』に数えない。"""
    p = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "あかつき商事"
    sh["A1"] = "請求書"                     # 雛形の固定文（比較の対象外）
    sh["B3"] = "あかつき商事"                # 印だった場所
    sh["B5"] = 12000
    wb.save(p)
    declared = {"B3": "あかつき商事", "B5": 12000}
    r = ailine.compare_report_cells(p, sheet_name="あかつき商事", declared=declared)
    assert (r.missing, r.mismatched, r.surplus) == ([], [], []), \
        f"雛形の固定文を余剰に数えた: {r}"


@needs_impl
def test_compare_report_cells_catches_a_wrong_value(tmp_path):
    """恒真殺し: 埋めた値が元の行と違えば不一致として名指しする。"""
    p = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "s"
    sh["B5"] = 999
    wb.save(p)
    r = ailine.compare_report_cells(p, sheet_name="s", declared={"B5": 12000})
    assert r.mismatched, "値が違うのに不一致 0（恒真）"
