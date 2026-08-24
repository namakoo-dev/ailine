# SPLIT_CELL ── 実装より先に凍結した赤い検体（2026-08-24）。
# 出所: 台帳 3203975（1セルの複数 URL を別セルへ）/ 1430969（項目分割）。
#
# 契約（✓ の 4 本柱）:
#   ① 行数不変 ② 元の列を残す（消さない）
#   ③ ★ 割った断片を同じ区切りで繋ぎ直すと元と一致する（恒真にならない検算）
#   ④ 他の列は 1 セルも変わらない

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import split_cell  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    "SPLIT_CELL" not in ailine.OP_SCHEMA,
    reason="SPLIT_CELL 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


# --- 区切りの読み取りと分割（単体・実装前から測れる部分）------------------------------

def test_normalize_separator():
    assert split_cell.normalize_separator("改行") == "\n"
    assert split_cell.normalize_separator("カンマ") == ","
    assert split_cell.normalize_separator("、") == "、"
    assert split_cell.normalize_separator("これは区切りの説明文です") is None


def test_split_keeps_empty_fragments():
    """★ 「a,,b」は 3 つ。空を落とすと繋ぎ直しても元に戻らない。"""
    assert split_cell.split_value("a,,b", ",") == ["a", "", "b"]
    assert split_cell.max_parts(["a,b", "a,b,c"], ",") == 3


def test_verify_rejoin_catches_a_lost_fragment():
    """恒真殺し: 断片が 1 つ落ちたら不一致として名指しする。"""
    r = split_cell.verify_rejoin(["a,b,c"], [["a", "c"]], ",")
    assert r.mismatched, "断片が落ちたのに一致と判定（恒真）"


def test_verify_rejoin_passes_on_a_faithful_split():
    r = split_cell.verify_rejoin(["a,b,c"], [["a", "b", "c"]], ",")
    assert not r.mismatched and r.rows_checked == 1


# --- e2e（op としての配線）-----------------------------------------------------------

@needs_impl
def test_split_cell_end_to_end(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "urls.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["会社", "URL", "備考"])
    ws.append(["A社", "http://a.com\nhttp://b.com", "そのまま"])
    ws.append(["B社", "http://c.com", "そのまま"])
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SPLIT_CELL", "args": {"col": "URL", "sep": "改行"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2["一覧"]
        ws2.cell(1, 4, "URL_1"); ws2.cell(1, 5, "URL_2")
        for r in (2, 3):
            parts = str(ws2.cell(r, 2).value or "").split("\n")
            for k, part in enumerate(parts):
                ws2.cell(r, 4 + k, part)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "URL列を改行ごとに別のセルに分けて", "--copy"], capsys)
    assert rc == 0, out
    got = openpyxl.load_workbook(str(p).replace(".xlsx", ".out.xlsx"))["一覧"]
    assert got.max_row == 3, "行数が変わった（柱①）"
    assert got.cell(2, 2).value == "http://a.com\nhttp://b.com", "元の列が消された（柱②）"
    assert got.cell(2, 4).value == "http://a.com" and got.cell(2, 5).value == "http://b.com"
    assert got.cell(2, 3).value == "そのまま", "他の列が変わった（柱④）"
    assert "✓" in out, out


# ★ needs_impl を付けない ── 実装前は恒真（op が無いので ✓ が出るはずもない）。
#   実装後に初めて意味を持つ「事後条件が在るか」の柵として置く。
def test_split_cell_refuses_when_rejoin_differs(tmp_path, monkeypatch, capsys):
    """柱③の恒真殺し: 断片を落として書いたら ✓ を名乗らない。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "urls2.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["会社", "URL"])
    ws.append(["A社", "http://a.com\nhttp://b.com"])
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SPLIT_CELL", "args": {"col": "URL", "sep": "改行"}})

    def bad_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2["一覧"]
        ws2.cell(1, 3, "URL_1"); ws2.cell(1, 4, "URL_2")
        ws2.cell(2, 3, "http://a.com")       # ★ 2 つ目を書かない（断片の欠落）
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", bad_apply)
    rc, out = _run_main(["run", str(p), "URL列を改行ごとに別のセルに分けて", "--copy"], capsys)
    assert "✓" not in out, f"断片が落ちたのに ✓ を名乗った: {out}"


# --- 英字の見出し名が「列文字」と誤読される（SPLIT_CELL 固有でない・2026-08-24 実測）------
#
# ★ 「URL列を…」の URL を列文字と読んで column_index_from_string("URL")=14676 列目と解釈し、
#   「依頼で言及された『列URL』は存在しません/変更されていません」という偽の ⚠ が出ていた。
#   決裁③でこれが ✓ を △ に降格させる。ID / AB / SUM のような英字見出しを持つブック
#   すべてで起きる（op を問わない）。

def test_letterlike_header_name_is_not_read_as_a_column_letter(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "ids.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["ID", "金額"])
    ws.append(["x", 100])
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2["一覧"].cell(2, 2, 100)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "ID列はそのままで金額を降順に並べ替えて", "--copy"], capsys)
    assert "『列ID』" not in out, f"見出し名 ID を列文字と誤読した: {out}"
