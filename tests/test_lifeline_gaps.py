# 復元の盲検 2 回目・致命②③（2026-08-25）── 命綱の穴。どちらも「今朝入れた物の隣」。
#
# 致命②: 壊れたバックアップを無検査で原本に被せて「✓ 復元した」
#   `restore_backup` は copy2 するだけで、復元元が開ける xlsx かを見なかった。
#   ★ この検査は今朝、**反映側にだけ**入れた（_why_output_is_unusable）。片配線の 4 度目。
#
# 致命③: `--copy` の成果物が、次の原本反映 run に黙って上書き・削除される
#   ★ **今朝入れた関所そのものの穴**。「この道具が過去にそこへ書いたか」で判定するので、
#     利用者がその後どれだけ手を入れても素通りした。
#   ★ 判定には三項が要る（依頼/宣言/実体）── 実体の項＝いま在る物の指紋。
#
# 契約:
#   ① 開けないバックアップは原本に被せない（原本は 1 バイトも変わらない）
#   ② 黙って別の世代へずらさない（どれを使うかは人が決める）
#   ③ 健全なバックアップは従来どおり復元できる（誤爆しない）
#   ④ 人が手を入れた .out は消さない
#   ⑤ 手つかずの .out は従来どおり作り直す（誤爆しない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402


def _xlsx(p: Path, value="ok"):
    wb = openpyxl.Workbook(); wb.active["A1"] = value; wb.save(p)
    return p


# --- ①②③ 壊れたバックアップ ---------------------------------------------------------

def test_broken_backup_is_never_laid_over_the_original(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = _xlsx(tmp_path / "book.xlsx", "current")
    ns = backups / ailine._backup_namespace(book)
    ns.mkdir(parents=True)
    (ns / "book.20260101T000000Z.xlsx").write_bytes(b"this is not a zip at all")
    _xlsx(ns / "book.20250101T000000Z.xlsx", "older-and-fine")   # ② ずらし先の誘惑
    before = book.read_bytes()
    with pytest.raises(ailine.BrokenBackupError) as ei:
        ailine.restore_backup(book)
    assert book.read_bytes() == before, "壊れたものを原本に被せた"
    assert "20260101" in str(ei.value), f"どの世代が壊れているか名指ししていない: {ei.value}"
    # ② 黙って古い方へずらしていないこと（ずらしていたら原本が変わっている）
    assert openpyxl.load_workbook(book).active["A1"].value == "current"


def test_healthy_backup_still_restores(tmp_path, monkeypatch):
    """③ 誤爆しない。"""
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = _xlsx(tmp_path / "book.xlsx", "current")
    ns = backups / ailine._backup_namespace(book)
    ns.mkdir(parents=True)
    _xlsx(ns / "book.20260101T000000Z.xlsx", "previous")
    ailine.restore_backup(book)
    assert openpyxl.load_workbook(book).active["A1"].value == "previous"


def test_unprobeable_format_is_not_called_broken(tmp_path, monkeypatch):
    """★ 調べられないことを「壊れている」と言わない（命綱を丸ごと塞がない）。"""
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.ods"
    book.write_bytes(b"current-ods")
    ns = backups / ailine._backup_namespace(book)
    ns.mkdir(parents=True)
    (ns / "book.20260101T000000Z.ods").write_bytes(b"older-ods")
    ailine.restore_backup(book)
    assert book.read_bytes() == b"older-ods"


# --- ④⑤ 人が育てた .out を消さない -----------------------------------------------------

def _history_with_out(tmp_path, monkeypatch, out: Path, sha):
    """★★ 2026-08-26 に治具を作り直した（復元の盲検 3 回目・致命2）。

    旧版は history の 1 行を**この治具が手で組んで**いた。本番の書き手は
    `build_history_entry` で、そちらは **out_sha を写していなかった** ──
    つまり**検体が、本番の欠けている項を自分で供給していた**。関所は一度も
    発火せず、それでも試験は緑だった。「同じ関数で作った分母は恒真」の親戚で、
    今回は *検体の側が* 中間項を代用した形。

    ★ 直しは「out_sha を足す」だけでは足りない。**治具から手書きの経路を無くす** ──
      history の行は必ず `build_history_entry` を通す。こうすると、本番の書き手が
      キーを落とした瞬間にこの試験が赤くなる（継ぎ目を跨ぐ）。
    """
    import json
    hist = tmp_path / "history.jsonl"
    result = {"ok": True, "out": str(out)}
    if sha is not None:
        result["out_sha"] = sha
    entry = ailine.build_history_entry(result, tmp_path / "book.xlsx", "t", "m", "none")
    assert "out" in entry, "本番の書き手が out を写していない"
    if sha is not None:
        assert entry.get("out_sha") == sha, (
            "本番の書き手が out_sha を写していない ── 関所は永遠に素通りする"
            "（この assert が致命2 の再発を止める）")
    hist.write_text(json.dumps(entry, ensure_ascii=False) + chr(10), encoding="utf-8")
    monkeypatch.setattr(ailine, "HISTORY_FILE", hist)


def test_edited_copy_output_is_refused_not_overwritten(tmp_path, monkeypatch, capsys):
    """④ ★ README が慎重な人に勧める --copy の成果物が、原本反映 1 回で消えていた。"""
    book = _xlsx(tmp_path / "book.xlsx")
    out = _xlsx(ailine.out_book_path(book), "ailine-made")
    _history_with_out(tmp_path, monkeypatch, out, ailine._file_digest(out))
    _xlsx(out, "人が続きを書き足した")            # 人が手を入れた
    rc = ailine.refuse_if_output_is_someone_elses(book)
    assert rc == 7, "人が育てた成果物を黙って上書きしようとした"
    printed = capsys.readouterr().out
    assert "変更されています" in printed, printed


def test_untouched_copy_output_is_still_rebuilt(tmp_path, monkeypatch):
    """⑤ 誤爆しない: 手つかずなら従来どおり作り直す。"""
    book = _xlsx(tmp_path / "book.xlsx")
    out = _xlsx(ailine.out_book_path(book), "ailine-made")
    _history_with_out(tmp_path, monkeypatch, out, ailine._file_digest(out))
    assert ailine.refuse_if_output_is_someone_elses(book) is None


def test_old_history_without_a_fingerprint_behaves_as_before(tmp_path, monkeypatch):
    """指紋を残す前の記録では判定材料が無い ── 従来どおり（新しい嘘を足さない）。"""
    book = _xlsx(tmp_path / "book.xlsx")
    out = _xlsx(ailine.out_book_path(book), "ailine-made")
    _history_with_out(tmp_path, monkeypatch, out, None)
    _xlsx(out, "人が書き足した")
    assert ailine.refuse_if_output_is_someone_elses(book) is None


# --- 重大5: 名前を変えたら「無い」ではなく「別の名前で在る」と言う ----------------------

def test_renaming_points_at_the_generations_instead_of_saying_none(tmp_path, monkeypatch):
    """★ 実測: mv したら `× ... のバックアップが無い` だけ。世代はそのまま在るのに。"""
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = _xlsx(tmp_path / "見積 書.xlsx")
    ns = backups / ailine._backup_namespace(book)
    ns.mkdir(parents=True)
    _xlsx(ns / "見積 書.20260101T000000Z.xlsx", "v0")
    renamed = _xlsx(tmp_path / "見積 書_最終版.xlsx")
    with pytest.raises(FileNotFoundError) as ei:
        ailine.restore_backup(renamed)
    msg = str(ei.value)
    assert "別の名前の世代が 1 件" in msg, f"在り処を言わず行き止まりにした: {msg}"
    assert str(ns) in msg, msg


def test_no_note_when_the_shelf_is_genuinely_empty(tmp_path, monkeypatch):
    """誤爆防止: 本当に 1 件も無い時は余計なことを言わない。"""
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = _xlsx(tmp_path / "b.xlsx")
    with pytest.raises(FileNotFoundError) as ei:
        ailine.restore_backup(book)
    assert "別の名前の世代" not in str(ei.value), str(ei.value)


# --- 致命1: --copy の成果物を、次の原本反映 run が無言で消す ---------------------------

def test_inplace_run_does_not_delete_a_pre_existing_copy_artifact(tmp_path, monkeypatch, capsys):
    """★ 実測: `--copy` で作った `<stem>.out.xlsx` が、次の原本反映 run で
       一言も無く消えた（undo は book しか守らないので戻せない）。

    根: `atomic_replace_inplace` の後始末が **無条件に** out_book を消していた。
    原本反映 run も作業ファイルに同じ名前を使うので、同じ場所を掴む。
    ★ 8/25 に入れた関所は run の**入口**だけを見ており、この**出口**を見ていなかった。
    """
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = _xlsx(tmp_path / "uriage.xlsx", "original")
    out = ailine.out_book_path(book)
    _xlsx(out, "人が育てた --copy の成果物")
    # run の入口（関所）が控える ── 本番と同じ経路を通す（治具で代用しない）
    _history_with_out(tmp_path, monkeypatch, out, ailine._file_digest(out))
    ailine.refuse_if_output_is_someone_elses(book)
    workdir = tmp_path / ".ailine_uriage"
    workdir.mkdir()
    ok, err = ailine.atomic_replace_inplace(book, out, workdir)
    assert ok, err
    assert out.exists(), "人が育てた --copy の成果物を無言で消した"
    assert "消していません" in capsys.readouterr().out


def test_inplace_run_still_cleans_up_its_own_scratch(tmp_path, monkeypatch):
    """誤爆防止: 今回の run が作った作業ファイルは従来どおり片づける。"""
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = _xlsx(tmp_path / "b.xlsx", "original")
    out = ailine.out_book_path(book)
    ailine.refuse_if_output_is_someone_elses(book)   # この時点では .out は無い
    _xlsx(out, "今回の作業結果")
    workdir = tmp_path / ".ailine_b"
    workdir.mkdir()
    ok, err = ailine.atomic_replace_inplace(book, out, workdir)
    assert ok, err
    assert not out.exists(), "自分の作業ファイルを片づけていない（ゴミが残る）"
