"""削除の「行か列か」は、宣言でなく**実表**が決めること（2026-09-05）。

★★ 実測した揺れ: 「ナットを削除して」が **5 回に 1 回**、行削除でなく**列削除**として
  返る。機械は「列『ナット』がこの表にありません」と正しく断る（fail closed）ので
  壊れはしないが、**誰もが最初に打つ形が 5 回に 1 回落ちる**。

★ 構造の原因: 削除の読み直しの門は「**どの段も削除を宣言していないとき**」に開く。
  列削除も削除を宣言しているので門が閉じ、間違った軸のまま断りになっていた。
  ── **宣言による門は「消すのか」は見分けるが「行か列か」は見分けられない。**

★ 直し方: 軸は実表が決める（SWAP が既にそうしている）。
    名指しの語が見出しに在る → 列削除でよい
    どこかの列の値に在る     → 行削除のはず
    両方 / どちらでもない    → 決めない（今までどおり断る）
  ★ 新しい門を増やさない ── **同じ門の鍵に軸を足した**。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402


@pytest.fixture
def stock(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "在庫"
    ws.append(["品名", "棚", "数量", "備考"])
    for r in [["ボルト", "A-1", 120, None], ["ナット", "A-2", 80, None],
              ["ワッシャー", "B-1", 300, None]]:
        ws.append(r)
    p = tmp_path / "b.xlsx"; wb.save(p)
    return ailine.build_book_meta(p)


def test_a_value_named_as_a_column_is_recognised_as_the_wrong_axis(stock):
    """★ 本命 ── 『ナット』は品名列の値であって列名ではない。"""
    why = ailine.removal_axis_is_wrong("DELETE_COLUMN", {"col": "ナット"}, stock, "在庫")
    assert why and "ナット" in why and "品名" in why, why


def test_a_real_column_is_left_alone(stock):
    """★ 見出しに在る名前は列削除で正しい ── 触らない。"""
    for name in ("備考", "数量", "品名"):
        assert ailine.removal_axis_is_wrong(
            "DELETE_COLUMN", {"col": name}, stock, "在庫") is None, name


def test_a_name_that_is_nowhere_is_not_decided(stock):
    """★ どこにも無い名前は決めない（推測しない）── 今までどおり断る側へ。"""
    assert ailine.removal_axis_is_wrong(
        "DELETE_COLUMN", {"col": "担当者"}, stock, "在庫") is None


def test_a_value_that_appears_in_two_columns_is_not_decided(tmp_path):
    """★ 2 つの列に在る値は決めない（どちらの行を消すか機械には分からない）。"""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "T"
    ws.append(["甲", "乙"])
    ws.append(["同じ", "同じ"])
    p = tmp_path / "b.xlsx"; wb.save(p)
    meta = ailine.build_book_meta(p)
    assert ailine.removal_axis_is_wrong("DELETE_COLUMN", {"col": "同じ"}, meta, "T") is None


def test_other_ops_are_untouched(stock):
    """★ 誤爆しない ── 削除以外の op には何も言わない。"""
    for op in ("DELETE_ROWS", "SORT", "SET_WHERE", "EXTRACT", ""):
        assert ailine.removal_axis_is_wrong(op, {"col": "ナット"}, stock, "在庫") is None, op


def test_the_judgement_reads_the_table_not_the_words():
    """★ 語の解釈をしないこと ── 実表に聞くだけ（誤爆の余地を作らない）。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src[src.index("def removal_axis_is_wrong"):src.index("\ndef ", src.index("def removal_axis_is_wrong") + 10)]
    assert "task_names_real_values" in body and "headers" in body
    for word in ("削除", "消して", "行を", "列を"):
        assert f'"{word}"' not in body, f"依頼文の語を見ている（{word}）── 表に聞くこと"


def test_the_gate_was_widened_not_duplicated():
    """★ 新しい門を増やしていないこと ── 既存の門の鍵に足す。

    今日ここまでで片配線を 5 回踏んでいる。門が増えるほど
    「どちらが先に立つか」で事故が起きる（2026-08-27 に 5 つ並んで上書きし合った）。
    """
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    assert src.count("removal_reading(a.task") == 1, "行削除の読み直しが 2 箇所になっている"
    assert src.count("_axis_wrong = removal_axis_is_wrong(") == 1


def test_it_only_speaks_when_it_actually_fixed_something():
    """★ 「軸を直しました」と言ったのに何もしない、を作らない。

    ★ 先に宣言して行が解決しなければ、**鳴ったのに何も起きない**（「在っても鳴らない」の
      反対側）。行が解けた後に言う順序であることを、コードの並びで縛る。
    """
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    i = src.index('_rm = removal_reading(a.task')
    j = src.index('print(f"（軸を直しました', i)
    k = src.index('print(f"（『行削除』として読み直しました', i)
    assert i < j < k, "軸の宣言が、行の解決より前に出ている"


@pytest.mark.local
def test_a_plain_delete_request_reaches_row_deletion(tmp_path):
    """★ 実機 ── 「ナットを削除して」が行削除に届くこと（5 回に 1 回落ちていた形）。"""
    import os, subprocess
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "在庫"
    ws.append(["品名", "棚", "数量", "備考"])
    for r in [["ボルト", "A-1", 120, None], ["ナット", "A-2", 80, None],
              ["ワッシャー", "B-1", 300, None]]:
        ws.append(r)
    src = tmp_path / "in.xlsx"; wb.save(src)
    repo = Path(ailine.__file__).resolve().parents[2]
    p = subprocess.run([sys.executable, "-m", "ailine", "run", str(src),
                        "ナットを削除して", "--copy", "--sheet", "在庫", "--timeout", "120"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       cwd=str(repo), env={**os.environ, "PYTHONPATH": str(repo / "src")})
    assert p.returncode == 0, p.stdout[-500:]
    assert "行削除" in p.stdout, p.stdout[-500:]


# --- ★ 門を通る経路を測る（純ロジックの試験だけでは配線が守られない）------------

def test_the_gate_actually_uses_the_axis(monkeypatch, tmp_path, capsys):
    """★★ 門の条件に軸が入っていること ── **経路で**確かめる。

    ★ 2026-09-05 の変異試験で穴が出た: 門から `_axis_wrong` を外しても、
      純ロジックの試験は 8 本とも緑のままだった。
      **いちばん大事な配線を、1 本も測っていなかった。**
      （「番人が在っても、その事故の形では鳴らない」の実例 ── 今日 2 度目）
    ★ だから一段目が**列削除を返す**状況を作り、行削除へ読み直されることを見る。
    """
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "在庫"
    ws.append(["品名", "棚", "数量", "備考"])
    for r in [["ボルト", "A-1", 120, None], ["ナット", "A-2", 80, None],
              ["ワッシャー", "B-1", 300, None]]:
        ws.append(r)
    book = tmp_path / "b.xlsx"; wb.save(book)

    # ★ 一段目が「列削除（col=ナット）」を返す ── 実測で 5 回に 1 回起きる形。
    monkeypatch.setattr(ailine, "normalize_book", lambda b, w, timeout=None: b)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, meta, temperature=0.1:
                        {"op": "DELETE_COLUMN", "args": {"col": "ナット"}})
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    applied = {}

    def _fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        applied["code"] = code
        return True, None, "ok"

    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply)
    ailine.main(["run", str(book), "ナットを削除して", "--copy", "--sheet", "在庫"])
    out = capsys.readouterr().out
    assert "軸を直しました" in out, out[-700:]
    assert "行削除" in out, out[-700:]
    assert "列削除" not in out.split("軸を直しました")[-1], "列削除のまま進んでいる"
