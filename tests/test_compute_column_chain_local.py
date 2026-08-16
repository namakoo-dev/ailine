"""★ W10f 項目1 実機E2E: 「数量に単価を掛けた列を作って、その列に消費税10%を掛けた
   税込列を作って」相当の複合計画（2段）を、本物の basrun/LibreOffice 経由で通す通し試験。
   ★ target 無指定の新規列は自動命名される（COMPUTE_COLUMN の既存仕様・本タスクの対象外
   ＝target が実在しない列名を指す場合は新規列作成にフォールバックする既存設計のため、
   ここでは「税込」という名前そのものは検証しない。実際の列名は下のアサーション参照）。

   ★ 背景: check_compute_column/check_compute_column_single_factor が use_formula=True の
   とき operand を式ビュー(raw)から読んでいたため、1段目が式で作った計算列(小計)を
   2段目の operand にすると、その値が式文字列のまま『数値でない』扱いになり、
   『検証対象0件』で fail → 計画全体が反映されない（計画ごとロールバック）バグがあった
   （Namakoo が純関数レベルで再現・tests/test_ailine.py の
   test_check_compute_column_operand_from_prior_formula_column_* が回帰カバー）。
   ここではその修正が『本物の basrun 適用が保存のたび式を再計算してキャッシュ値を
   埋める』実運用の形でも効くことを、--values を付けずに（＝既定の use_formula=True で）
   確認する。単体テスト（openpyxl だけで作った・LibreOffice を一度も通していないブック）
   では意図的にキャッシュ値が無いままなので、そちらでは同じシナリオが『検証できない』
   と正直に fail する（区別している設計）。ここは逆に、実際に basrun を通すことで
   キャッシュ値が実在する状態を作り、pass することを見る。

   ★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
   （`pytest -m "not local"` で除外される。手動は `pytest -m local` または
   `pytest tests/test_compute_column_chain_local.py`）。
   ★ tests/test_bold_local.py と同じ後始末方針（taskkill 名前一括はしない・
   ailine._stop_office() に委譲）。"""
import argparse
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
import ailine


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1   # 数えられない環境。local マーカーの手動実行前提なので実害は小さい
    return out.lower().count("soffice.bin")


@pytest.mark.local
def test_compound_plan_compute_column_operand_from_prior_step_passes_without_values(
        tmp_path, monkeypatch, capsys):
    """1段目(数量*単価→数量*単価 列を式のまま作成)、2段目(その列に1.1を掛けた税込列を
       式のまま作成)の2段計画を、本物の basrun/LibreOffice 経由で通す。
       translate_task だけ機械的な計画を直接返すよう差し替える（DSL 経路は ollama を
       呼ばない＝率や列名を LLM に決めさせない A' 原則どおり・ここはその決定論部分と
       本物の basrun 適用・本物の事後条件検証だけを見る）。"""
    before_procs = _soffice_process_count()
    try:
        book = tmp_path / "book.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["商品", "数量", "単価"])
        ws.append(["りんご", 3, 100])
        ws.append(["みかん", 5, 80])
        wb.save(book)

        monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
        monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
        task_text = "数量に単価を掛けた列を作って、その列に消費税10%を掛けた税込列を作って"
        monkeypatch.setattr(
            ailine, "translate_task",
            lambda model, task, book_meta, temperature=0.1: {
                "plan": [
                    {"op": "COMPUTE_COLUMN",
                     "args": {"operands": ["数量", "単価"], "operator": "*"}},
                    {"op": "COMPUTE_COLUMN",
                     "args": {"operands": ["数量*単価"], "operator": "*"}},
                ]
            })

        ns = argparse.Namespace(
            book=str(book), task=task_text, model="qwen2.5-coder:7b",
            refs=None, helpers=None, repair=0, temperature=0.2,
            dry=False, copy=True, json=False, timeout=180.0, ask=False,
            allow_freeform=True)
        rc = ailine.cmd_run(ns)
        captured = capsys.readouterr()

        assert rc == 0, captured.out
        assert "検証対象が0件" not in captured.out, captured.out   # ★ 直った点そのもの
        assert "すべて機械検証済み" in captured.out, captured.out

        out_book = book.with_name(book.stem + ".out" + book.suffix)
        wb2 = openpyxl.load_workbook(out_book, data_only=True)
        ws2 = wb2.active
        headers = [c.value for c in ws2[1]]
        # ★ target 無指定の新規列は自動命名（COMPUTE_COLUMN の既存仕様・本タスクの対象外）。
        #   1段目 "数量*単価"。2段目は依頼文に「税込」とあるため自然な日本語見出し
        #   "税込数量*単価"（_TAX_INCLUSIVE_RE・W10c 中の既存機能、実測で確認）。
        assert "数量*単価" in headers, headers
        assert "税込数量*単価" in headers, headers
        i_mid = headers.index("数量*単価") + 1
        i_tax = headers.index("税込数量*単価") + 1
        # りんご: 3*100=300 → 300*1.1=330 / みかん: 5*80=400 → 400*1.1=440
        assert ws2.cell(row=2, column=i_mid).value == 300
        assert ws2.cell(row=3, column=i_mid).value == 400
        assert abs(ws2.cell(row=2, column=i_tax).value - 330) < 1e-6
        assert abs(ws2.cell(row=3, column=i_tax).value - 440) < 1e-6
        wb2.close()
    finally:
        # ★ 後片付け: basrun.py stop（接続先の LibreOffice だけを終了）。taskkill 一括はしない。
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )
