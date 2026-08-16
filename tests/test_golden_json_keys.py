"""C1-F7: `--json` のキー集合と各値の型を、path=dsl/plan/freeform × dry(2通り) の
6通りについて凍結する。

型は再帰的な型シグネチャ文字列で表す（例: "list[str]"・"list[]"（空リスト）・
"dict{op:str,...}"・"null"）。これにより ★ 既知の不整合 —— `advisories` が経路に
よって dsl/freeform では `list[str]`、plan では `list[{"steps": [...], "text": str}]`
（`list[dict{...}]`）の2形になっている —— がゴールデンの型シグネチャの違いとして
そのまま見える。

★★ これは今回直さない（brief の指示どおり）。網を張るのが先、修理は宣言つきで後。
直すときは、この6ファイルの型シグネチャが変わることを承知の上で
AILINE_REGEN_GOLDEN=1 で再生成し、diff に「advisories の形を統一した」等の理由を
commit メッセージに書くこと。
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402

from golden._harness import GOLDEN_ROOT, assert_golden_json  # noqa: E402
from _run_argv import run_argv  # noqa: E402  — C2: cmd_run 直呼び用 Namespace → main(argv) 変換

F7_DIR = GOLDEN_ROOT / "f7_json_keys"


def _type_sig(v):
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    if isinstance(v, str):
        return "str"
    if isinstance(v, list):
        if not v:
            return "list[]"
        inner = sorted({_type_sig(x) for x in v})
        return f"list[{'|'.join(inner)}]"
    if isinstance(v, dict):
        fields = ",".join(f"{k}:{_type_sig(vv)}" for k, vv in sorted(v.items()))
        return f"dict{{{fields}}}"
    return type(v).__name__


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _base_argv(book, dry, **overrides):
    base = dict(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=dry, copy=True, json=True, timeout=180.0, ask=False)
    base.update(overrides)
    return run_argv(**base)


def _extract_json_line(stdout: str) -> dict:
    for line in reversed(stdout.splitlines()):
        line = line.strip()
        if line.startswith("{"):
            return json.loads(line)
    raise AssertionError(f"--json 出力の JSON 行が見つからない:\n{stdout}")


def _fake_basrun_apply_noop(out_book, code, workdir, helper_files=(), timeout=None):
    """SORT/BOLD 用: ブック側を事前に『適用後の状態』にしておくので basrun_apply 自体は
       ファイルに触れず True を返すだけでよい（test_ailine.py の既存パターンを踏襲）。
       ★ F7: advisories が空のままだと dsl/plan の型差（list[str] vs list[dict]）が
       golden に現れないため、原本の使用範囲外に1セルだけ書き込んで幽霊データ advisory
       を必ず1件は立てる（postcondition の合否には無関係な位置）。"""
    wb = openpyxl.load_workbook(out_book)
    wb.active.cell(row=30, column=10, value="外側")
    wb.save(out_book)
    return True, None, "ok"


def _fake_basrun_apply_touches_cell(out_book, code, workdir, helper_files=(), timeout=None):
    """FREEFORM 用: no-op ガードに引っかからないよう、実際に1セル書き換える。"""
    wb = openpyxl.load_workbook(out_book)
    wb.active.cell(row=1, column=3, value="ok")
    wb.save(out_book)
    return True, None, "ok"


def _run_and_capture(tmp_path, monkeypatch, capsys, translate_result, dry, *,
                      freeform=False, presorted_and_bold=False, plan=False):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1: translate_result)

    if presorted_and_bold:
        book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
        wb = openpyxl.load_workbook(book)
        ws = wb.active
        for c in (1, 2):
            ws.cell(row=1, column=c).font = Font(bold=True)
        wb.save(book)
    else:
        book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])

    if freeform:
        monkeypatch.setattr(ailine, "ollama_generate",
                             lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    if not dry:
        monkeypatch.setattr(
            ailine, "basrun_apply",
            _fake_basrun_apply_touches_cell if freeform else _fake_basrun_apply_noop)

    argv = _base_argv(book, dry, allow_freeform=(freeform and not dry))
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    result = _extract_json_line(captured.out)
    return rc, result


_SORT_STEP = {"op": "SORT", "args": {"col": "金額", "order": "desc"}}
_BOLD_STEP = {"op": "BOLD", "args": {"target": "row:1"}}

CASES = {
    "dsl_dry": dict(translate_result=_SORT_STEP, dry=True),
    "dsl_full": dict(translate_result=_SORT_STEP, dry=False, presorted_and_bold=False),
    "plan_dry": dict(translate_result={"plan": [_SORT_STEP, _BOLD_STEP]}, dry=True),
    "plan_full": dict(translate_result={"plan": [_SORT_STEP, _BOLD_STEP]}, dry=False,
                       presorted_and_bold=True, plan=True),
    "freeform_dry": dict(translate_result={"op": "FREEFORM", "args": {}}, dry=True, freeform=True),
    "freeform_full": dict(translate_result={"op": "FREEFORM", "args": {}}, dry=False, freeform=True),
}


@pytest.mark.parametrize("name", sorted(CASES.keys()))
def test_json_keys_and_types_golden(tmp_path, monkeypatch, capsys, name):
    kwargs = dict(CASES[name])
    rc, result = _run_and_capture(tmp_path, monkeypatch, capsys, **kwargs)
    sig = {k: _type_sig(v) for k, v in sorted(result.items())}
    payload = {"rc": rc, "path": result.get("path"), "dry": bool(kwargs["dry"]),
               "key_type_signature": sig}
    assert_golden_json(F7_DIR / f"{name}.json", payload, label=name)


def test_advisories_shape_inconsistency_is_visible_in_golden():
    """★ 既知の不整合を golden の型シグネチャで直接確認する自己検査
       （このテストが赤くなったら、6ファイルの advisories 型を再確認すること — 直った
       ならこの assert ごと書き換えて『直した』と commit メッセージに書く）。"""
    dsl_full = json.loads((F7_DIR / "dsl_full.json").read_text(encoding="utf-8"))
    freeform_full = json.loads((F7_DIR / "freeform_full.json").read_text(encoding="utf-8"))
    plan_full = json.loads((F7_DIR / "plan_full.json").read_text(encoding="utf-8"))
    dsl_adv = dsl_full["key_type_signature"].get("advisories")
    freeform_adv = freeform_full["key_type_signature"].get("advisories")
    plan_adv = plan_full["key_type_signature"].get("advisories")
    for sig in (dsl_adv, freeform_adv):
        assert sig is None or sig.startswith("list[") and "dict" not in sig, (
            f"dsl/freeform の advisories が dict 形に変わった（既知の不整合が解消された "
            f"可能性 — 直したなら意図的な commit として扱うこと）: {sig}")
    assert plan_adv is None or "dict" in plan_adv or plan_adv == "list[]", (
        f"plan の advisories が list[str] 形に変わった: {plan_adv}")
