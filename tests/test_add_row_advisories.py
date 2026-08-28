# 末尾に 1 行足すと毎回 △ ── 2026-08-28。Namakoo「7行目にヤマノ食品を追加して」
#
# ★★ 実測: 正しく 1 行足したのに、△ に落ちて 2 件の助言が出ていた。
#     ★ 疑わしい: 変更が元データの範囲外です（A7）
#     ★ 疑わしい: 空欄への同一値の一括書き込みです（値 'ヤマノ食品' × 1 セル）
#   どちらも **この op がやると宣言していること そのもの**。
#   「× 1 セル」を「一括」と言う助言まで出ていた ── 助言が仕事をしていない。
#
# ★ 真因は宣言の欠け。ADD_ROW は WRITE_ROW_SHIFT しか宣言しておらず、
#   「末尾に新しい行を足すこともある」と言っていなかった。助言側は宣言しか見ないので
#   知りようがない。APPEND_TOTAL・ADD_COLUMN で **同じ形を 2 度**直していて、これが 3 度目。
#   ★ だから直しは「助言に if を足す」ではなく **宣言を足す**（1 箇所に畳む）。
#
# ★ 助言を黙らせる側の変更なので、番人は**黙りすぎていないこと**を見る:
#   途中への挿入・列の外への書き込み・上の行の破壊は、今も捕まること。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

ROWS = [["取引先", "件数", "金額"], ["丸和物流", 12, 57600],
         ["近江スチール", 5, 60000], ["みどり建設", 9, 64800]]


def _book(tmp_path, rows=None, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    for r in (rows or ROWS):
        ws.append(r)
    wb.save(p)
    return p


# --- ① 宣言 ---------------------------------------------------------------------------

def test_add_row_declares_what_it_actually_does():
    """★ 直しの芯。ここが戻ると、末尾への追記が毎回 △ に落ちる。"""
    wt = ailine.OP_WRITE_TARGET["ADD_ROW"]
    assert ailine.WRITE_NEW_ROW_AT_END in wt.writes, "末尾に足すことを宣言していない"
    assert ailine.WRITE_ROW_SHIFT in wt.writes, "途中に挿すことを宣言していない"
    assert wt.proves_which_cells, "両方向の証明を持つと宣言していない"


def test_the_check_really_proves_both_directions(tmp_path):
    """★ proves_which_cells は「助言を黙らせてよい」という主張 ── 主張の裏を取る。
       上の行を 1 セル壊したら、助言ではなく**事後条件そのもの**が落ちること。"""
    src = _book(tmp_path, name="src.xlsx")
    out = _book(tmp_path, name="out.xlsx")
    wb = openpyxl.load_workbook(out)
    ws = wb["請求"]
    ws.append(["ヤマノ食品", None, None])
    ws.cell(2, 1, "丸和物流(壊)")          # ★ at より上を壊す
    wb.save(out)
    status, reason = ailine.check_add_row(
        out, {"at": 5, "values": {"取引先": "ヤマノ食品"}, "_target_sheet": "請求"},
        source_book=src)
    assert status == "fail", reason


# --- ② 助言（黙るべき所で黙る）--------------------------------------------------------

def _ghost(tmp_path, op, after_mut):
    p = _book(tmp_path, name="g.xlsx")
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    after_mut(wb["請求"])
    wb.save(p)
    return ailine.detect_ghost_data(
        before, ailine.snapshot(p),
        new_row_at_end=ailine._op_writes(op, ailine.WRITE_NEW_ROW_AT_END))


def test_appending_below_the_last_row_is_not_suspicious(tmp_path):
    note = _ghost(tmp_path, "ADD_ROW", lambda ws: ws.cell(5, 1, "ヤマノ食品"))
    assert note and "意図どおり" in note, note
    assert "疑わしい" not in note, note


def test_writing_outside_the_columns_is_still_suspicious(tmp_path):
    """★ 黙りすぎていないこと: 列の外（H 列）へ出たら、宣言があっても助言は出る。"""
    note = _ghost(tmp_path, "ADD_ROW", lambda ws: ws.cell(5, 8, "?"))
    assert note and "疑わしい" in note, note


def test_an_op_without_the_declaration_is_still_warned(tmp_path):
    """★ 恒真殺し: 免除は**宣言した op だけ**に効く（全 op に効いたら助言が死ぬ）。"""
    note = _ghost(tmp_path, "SET_COLUMN_VALUE", lambda ws: ws.cell(5, 1, "ヤマノ食品"))
    assert note and "疑わしい" in note, note


def test_the_bulk_write_advice_is_silent_for_add_row():
    """★ 「× 1 セル」を『一括書き込み』と呼ぶ助言が出ていた ── 証明が届く op では黙る。"""
    assert ailine.OP_WRITE_TARGET["ADD_ROW"].proves_which_cells
    assert not ailine.OP_WRITE_TARGET["SET_COLUMN_VALUE"].proves_which_cells, \
        "一括書換は元が空欄だったかを問わない ── ここの助言は今も仕事がある"


# --- ③ 画面（この操作の分だけを色で示す）----------------------------------------------

def test_the_gui_compares_against_the_previous_state_by_default():
    """★★ Namakoo が見たもの: 「7行目に…」で **8 行目まで**増えたように見えた。
       増やしたのは 1 行で、8 行目は**前の操作**の結果だった。下書きは続きに積む設計
       なのに、色は原本との差で塗っていたので、積んだ分まで今回の結果に見えていた。
       ★ 表示が依頼の範囲を偽る形なので、既定を「直前」にする（原本も選べる）。"""
    html = (REPO / "gui" / "index.html").read_text(encoding="utf-8")
    i = html.index('id="basis"')
    seg = html[i:i + 400]
    assert seg.index('value="prev"') < seg.index('value="orig"'), \
        "くらべる相手の既定が『直前』でない ── 前の操作の結果まで今回の色になる"
    assert 'value="orig"' in seg, "原本とくらべる道も残すこと"


def test_the_gui_says_when_the_draft_carries_earlier_work():
    html = (REPO / "gui" / "index.html").read_text(encoding="utf-8")
    assert "前の操作の結果が入っています" in html, "積んであることを画面で言っていない"
    assert "function countDiff(" in html
