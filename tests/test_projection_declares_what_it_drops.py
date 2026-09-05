"""投影法 ── 表を別の形へ写す op が、**何を保存しないか**を宣言していること（2026-09-05）。

★★ 出所（2026-09-02）: FORMAT_MAP の困難を ailine の語彙だけで書き、盲で立てた
  23 学問分野へ撃って拾った。★ 予測に入れていなかった分野から来た。

    地図の投影法 ── 面積・角度・距離のどれを保存するかを「選ぶ」。
                  全部同時には不可能（**定理**）。だから投影法には名前がつき、
                  地図には必ず明記される。

★ 読み替え: 「依頼文に情報が無い」（依頼者側の欠陥）ではなく
  **「保存する不変量を選ばせていない」**（こちらの設計の穴）。

★★ 実測で効いたこと（2026-09-05）: 式のある表に様式写像を掛けると **必ず × になっていた**。
  元の `金額` が `=B2*C2`、出力は `36000` ── 事後条件は式の文字列と比べて「不一致」と
  判定していた。**正当な変換を事故として扱っていた。**
  ★ 欠けていたのは「この写像は式を保存せず値を保存する」という宣言。
    判定に要る三項（依頼・宣言・実体）のうち、また宣言だった。

★ 変分法の轍を踏まない ── **7 op すべて**に宣言を書く（1 つの検体だけで測らない）。
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import ailine  # noqa: E402
from ailine_core.projection import (  # noqa: E402
    ALL_INVARIANTS, FORMULAS, PROJECTIONS, projection_for, render_projection_notice)

#: 新しい表を作る 7 op（postconditions/derive.py の対象）。
_DERIVE_OPS = ("AGGREGATE", "PIVOT", "EXTRACT", "EXTRACT_COLUMNS",
               "REPORT_PER_ROW", "FORMAT_MAP", "LOOKUP_FILL")


def test_every_deriving_op_declares_its_projection():
    """★ 7 op すべてが宣言していること（FORMAT_MAP だけにしない ── 変分法の轍）。"""
    missing = [op for op in _DERIVE_OPS if projection_for(op) is None]
    assert not missing, f"投影法を宣言していない op: {missing}"


def test_no_op_declares_an_invariant_that_does_not_exist():
    """★ 名前を発明しない ── keeps/drops は決めた語彙の中だけ。"""
    for op, proj in PROJECTIONS.items():
        for name in proj.keeps + proj.drops:
            assert name in ALL_INVARIANTS, f"{op}: 知らない不変量 {name!r}"


def test_keeps_and_drops_never_overlap():
    """★ 保存すると同時に保存しない、は言えない（地図の定理と同じ）。"""
    for op, proj in PROJECTIONS.items():
        both = set(proj.keeps) & set(proj.drops)
        assert not both, f"{op}: 保存すると同時に落とすと言っている {both}"


def test_every_dropped_invariant_that_matters_has_a_reason():
    """★ 「式を落とす」とだけ言って理由を書かない、をさせない。

    ★ 式は**利用者がいちばん驚く**もの（見た目が同じで中身が値になる）。
      理由の無い宣言は、地図に投影法名だけ書いて歪みを説明しないのと同じ。
    """
    for op, proj in PROJECTIONS.items():
        if FORMULAS in proj.drops:
            assert proj.why.get(FORMULAS), f"{op}: 式を落とす理由が書かれていない"


@pytest.mark.parametrize("op", _DERIVE_OPS)
def test_the_notice_leads_with_what_is_lost(op):
    """★ 「保存する」より「**保存しない**」を先に言う ── そちらが人の判断を変える。"""
    lines = render_projection_notice(op)
    assert lines, f"{op}: 何も言っていない"
    assert "保存されない" in lines[0], lines[0]


def test_an_op_that_does_not_transform_says_nothing():
    """★ 誤爆しない ── 写す op でなければ黙る。"""
    for op in ("SORT", "BOLD", "DELETE_ROWS", "SET_WHERE", None):
        assert render_projection_notice(op) == [], op


def test_the_notice_is_printed_at_the_single_confluence():
    """★ 単発も複合計画も通る 1 箇所で出すこと（呼び出し側 2 箇所に配らない）。

    今日ここまでで片配線を 5 回踏んでいる ── 同じ形を新しく作らない。
    """
    src = (Path(ailine.__file__).parent.parent / "ailine_core" / "dsl_step.py"
           ).read_text(encoding="utf-8")
    assert src.count("render_projection_notice(") == 1, \
        "投影法の開示が dsl_step.py の 1 箇所でない"
    main = Path(ailine.__file__).read_text(encoding="utf-8")
    assert "render_projection_notice(" not in main, \
        "呼び出し側にも配っている（合流点 1 箇所に置くこと）"


def test_the_projected_value_read_is_folded_into_one_function():
    """★ 「式はキャッシュ値で比べる」判断が 1 本に畳まれていること。

    ★ BookView はまさにこのために作られた（「事後条件が operand を式ビューから読む」は
      5 件の欠陥の共通の根）。帳票と様式写像が **6・7 件目**だったので、
      2 箇所に書き写さず 1 本を通す。
    """
    derive = (Path(ailine.__file__).parent.parent / "ailine_core" / "postconditions"
              / "derive.py").read_text(encoding="utf-8")
    assert derive.count("source_value_as_projected(") == 3, \
        "定義 1 + 呼び出し 2 になっていない（書き写しが増えている）"
    assert 'src.cell(row=src_row, column=ph["col_idx"]).value' not in derive, \
        "式ビューから直に読む書き方が残っている"


@pytest.mark.local
def test_a_table_with_formulas_can_be_mapped_to_a_template(tmp_path):
    """★ 実機 ── 式のある表に様式写像を掛けて ✓ が出ること（以前は必ず × だった）。"""
    import openpyxl, os, subprocess
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "受注"
    ws.append(["取引先", "数量", "単価", "金額"])
    for j, (t, q, u) in enumerate([("あかね商事", 3, 12000), ("うえだ物産", 7, 8000)]):
        ws.append([t, q, u, f"=B{2 + j}*C{2 + j}"])
    tpl = wb.create_sheet("出荷様式")
    tpl["A1"] = "出荷先"; tpl["B1"] = "個数"; tpl["C1"] = "合計"
    tpl["A2"] = "{{取引先}}"; tpl["B2"] = "{{数量}}"; tpl["C2"] = "{{金額}}"
    src = tmp_path / "in.xlsx"; wb.save(src)
    repo = Path(ailine.__file__).resolve().parents[2]
    p = subprocess.run([sys.executable, "-m", "ailine", "run", str(src),
                        "出荷様式に合わせて写して", "--copy", "--sheet", "受注",
                        "--timeout", "150"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       cwd=str(repo), env={**os.environ, "PYTHONPATH": str(repo / "src")})
    assert p.returncode == 0, p.stdout[-600:]
    assert "✓" in p.stdout, p.stdout[-600:]
    assert "保存されないもの" in p.stdout, "投影法の開示が出ていない"
