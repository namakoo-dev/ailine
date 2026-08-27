"""③ ⚠ による ✓ の降格（Namakoo 決裁 2026-08-22 08:11）の検体。
   ★ 実装前に凍結した赤い検体。根拠: 「結合を解除して」→ セル結合（真逆）の誤訳が
   見出し 2 セルの値を消し、⚠（書式だけのはずが値が変わった）と ✓ が同居した実測 ──
   ⚠ と ✓ の同居は、買い手にどちらを信じろと言っているのか分からない。

   契約: 疑わしい系の ⚠ が 1 件でも出た実行は「✓ 機械検証済み」を名乗らない。
   代わりに「△」で 宣言どおりの変化の確認 と ⚠ の先出し確認依頼を分けて言う。
   ⚠ ゼロの実行は従来どおり ✓（降格の誤爆で ✓ が絶滅しないこと）。"""
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402



def test_suspicious_warning_demotes_the_checkmark(tmp_path, monkeypatch, capsys):
    """★ 実測の再現（2026-08-21 夜・複合計画）: 2 段目のセル結合が見出しの値を消し、
       ⚠『書式だけのはずが値が変わりました』が段の注記として出たのに、最後に ✓ も出た。
       契約: 疑わしい ⚠ が 1 件でも出た実行は ✓ を名乗らない。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                   {"op": "MERGE", "args": {"range": "A1:B1"}}]})
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if calls["n"] == 1:
            ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        else:
            # ★ 結合（openpyxl/LibreOffice とも仕様どおり）で非アンカー側 B1 の値が消える
            #   （実測の形 ── check_merge の事後条件を満たしつつ、単位F の format_only
            #   前提を破る。前提: merge_cells は check_merge が読む merged_cells.ranges に
            #   "A1:B1" を登録し、副作用として B1 の値を None にする）。
            ws.merge_cells("A1:B1")
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて、A1とB1を結合して", "--copy"],
                         capsys)
    assert "値が" in out and "変わりました" in out, f"前提: 値が変わった ⚠ が出ること: {out}"
    assert "✓" not in out, f"⚠ と ✓ が同居している（降格されていない）: {out}"
    assert "△" in out or "確認してください" in out, f"降格後の顔が無い: {out}"


def test_clean_run_still_gets_the_checkmark(tmp_path, monkeypatch, capsys):
    """⚠ ゼロの実行は従来どおり ✓（降格の誤爆で ✓ が絶滅しないこと）。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    assert "✓" in out, f"⚠ ゼロなのに ✓ が出ない（降格の誤爆）:\n{out}"


# ★ 片配線の追補（2026-08-22 検分で発見・実装より先に凍結した赤い検体）:
#   決裁③の初版は advisories（★ 付き）と単位F/G の前提破れだけを数えていたが、
#   確認段で「⚠ 」前置で印字される resolved["_warnings"]（LLM 由来の値と機械抽出の
#   食い違い）が どちらのリストにも入らず、⚠ と ✓ が同居できた。
#   系譜「二重化した経路は片配線が既定で起きる」── 直しは同じ形の経路すべてに・
#   検体も両経路分（単発 cmd_run_dsl / 複合 cmd_run_plan）。

def _tax_book(tmp_path):
    return _book(tmp_path, [["品目", "数量", "単価", "小計"],
                             ["a", 1, 100, 100], ["b", 2, 50, 100]])


def _tax_fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
    # APPEND_TOTAL の事後条件を満たす正しい適用（機械抽出の 1.1 が採用される）
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    # ★ 2026-08-28: 合計行のラベルは**1 列目**に置く（旧: 対象列の左隣）。
    #   旧位置は 1 列目を空のまま残し、道具自身の走査を止めていた。
    ws["A4"] = "税込み合計"
    ws["D4"] = "=SUM(D2:INDEX(D:D,ROW()-1))*1.1"
    wb.save(out_book)
    _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"D4": 200 * 1.1})
    return True, None, "ok"


def test_llm_value_mismatch_warning_demotes_single_op(tmp_path, monkeypatch, capsys):
    """単発経路: LLM の factor(1.08) と機械抽出(1.1) の食い違い ⚠ が出た run は ✓ を名乗らない。"""
    _isolate(monkeypatch, tmp_path)
    book = _tax_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "APPEND_TOTAL", "args": {"col": "小計", "label": "税込み合計", "factor": 1.08}})
    monkeypatch.setattr(ailine, "basrun_apply", _tax_fake_apply)
    rc, out = _run_main(["run", str(book), "税込み合計を一番下に出して（消費税10%）", "--copy"],
                         capsys)
    assert "食い違うため機械抽出" in out, f"前提: 食い違い ⚠ が出ること: {out}"
    assert "✓" not in out, f"⚠ と ✓ が同居している（単発経路が片配線）: {out}"
    assert "△" in out, f"降格後の顔が無い: {out}"


def test_llm_value_mismatch_warning_demotes_plan_step(tmp_path, monkeypatch, capsys):
    """複合計画経路: 段の中で同じ食い違い ⚠ が出た run も ✓ を名乗らない（両経路分）。"""
    _isolate(monkeypatch, tmp_path)
    book = _tax_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "APPEND_TOTAL",
                    "args": {"col": "小計", "label": "税込み合計", "factor": 1.08}}]})
    monkeypatch.setattr(ailine, "basrun_apply", _tax_fake_apply)
    rc, out = _run_main(["run", str(book), "税込み合計を一番下に出して（消費税10%）", "--copy"],
                         capsys)
    assert "食い違うため機械抽出" in out, f"前提: 食い違い ⚠ が出ること: {out}"
    assert "✓" not in out, f"⚠ と ✓ が同居している（複合経路が片配線）: {out}"
    assert "△" in out, f"降格後の顔が無い: {out}"
