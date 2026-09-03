# 同じ取引先を 1 枚にまとめる ── 2026-08-28。Namakoo「同名の取引先から複数の発注が
# あるケースでは請求書を一枚にまとめないといけないだろう？」
#
# ★★ 直す前: 1 データ行 = 1 枚。ヤマノ食品が 2 行あると請求書が『ヤマノ食品』と
#   『ヤマノ食品_2』の 2 枚になり、**それでも ✓ が出た**（宣言どおり 2 枚作ったから）。
#   宣言と実体は合っていて、**依頼（人が本当に欲しいもの）**だけが見られていない
#   ── この repo の三項の型そのもの。
#
# ★ 設計: **雛形が形を決める**（依頼文にも一段目の語彙にも 1 文字も足さない）。
#     {{列名}}      … そのグループで同じはずの値。食い違ったら**埋めずに断る**
#     {{明細:列名}} … 1 件ごと。この行が件数ぶん増える
#     {{合計:列名}} … そのグループの合計（数値列のみ）
#
# ★ 設計査読（fable）が名指しした穴を、そのまま番人にしてある:
#   ① 重複があっても**断らない**（領収書・納品書は取引ごとに 1 枚が正しい）
#   ② 明細行に他種の印が同居すると件数ぶん刷られる ── しかも事後条件は通る
#   ③ 合計の型（int を float にすると型込み等値で偽の × が出る）
#   ④ 縦の結合セルが明細行を横切ると崩れる（値は合うので ✓ が出てしまう）
#   ⑥ 全角コロン『明細：項目』（IME で雛形を書けばこちらが自然）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import report_group as rg  # noqa: E402
from _product_source import window_around  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む


class _Ph:
    def __init__(self, row, cell, name):
        self.row, self.cell, self.column_name = row, cell, name


# --- 印の仕分け -----------------------------------------------------------------------

@pytest.mark.parametrize("mark,want", [
    ("明細:項目", ("detail", "項目")),
    ("明細：項目", ("detail", "項目")),      # ⑥ 全角コロン
    ("明細: 項目", ("detail", "項目")),
    ("合計:金額", ("total", "金額")),
    ("合計：金額", ("total", "金額")),
    ("取引先", ("value", "取引先")),
])
def test_the_three_kinds_of_mark(mark, want):
    assert rg.mark_kind(mark) == want


def test_detail_marks_on_two_rows_are_refused():
    _l, err = rg.classify_placeholders([_Ph(11, "A11", "明細:項目"),
                                         _Ph(12, "A12", "明細:金額")])
    assert err and "散っています" in err, err


def test_a_squatter_on_the_detail_row_is_refused():
    """★★ 査読②: {{担当}} が明細行に居ると、担当名が件数ぶん刷られる。
       しかも埋める側と確かめる側が同じずれ関数を共有するので**事後条件は通る**。"""
    _l, err = rg.classify_placeholders([_Ph(11, "A11", "明細:項目"), _Ph(11, "B11", "担当")])
    assert err and "同居" in err, err
    _l2, err2 = rg.classify_placeholders([_Ph(11, "A11", "明細:項目"), _Ph(11, "D11", "合計:金額")])
    assert err2 and "同居" in err2, err2


def test_a_mark_without_a_column_name_is_refused():
    _l, err = rg.classify_placeholders([_Ph(11, "A11", "明細:")])
    assert err and "列名がありません" in err, err


def test_a_plain_template_has_no_detail_row():
    layout, err = rg.classify_placeholders([_Ph(3, "A3", "取引先"), _Ph(7, "B7", "金額")])
    assert err is None and layout.detail_row is None and len(layout.value) == 2


# --- グループ化・食い違い・合計 -------------------------------------------------------

_ROWS = [(2, ["丸和物流", 100]), (3, ["ヤマノ食品", 200]), (4, ["ヤマノ食品", 300])]


def test_groups_keep_the_order_they_first_appear():
    gs = rg.build_groups(_ROWS, 1)
    assert [(g.name, g.rows) for g in gs] == [("丸和物流", (2,)), ("ヤマノ食品", (3, 4))]
    assert rg.needs_grouping(gs) and not rg.needs_grouping(gs[:1])


def test_a_conflicting_group_value_is_reported():
    g = rg.GroupPlan("ヤマノ食品", (3, 4))
    vals = {3: {"担当": "田中"}, 4: {"担当": "佐藤"}}
    assert rg.value_conflicts(g, vals, "担当") == ["田中", "佐藤"]
    assert rg.value_conflicts(g, {3: {"担当": "田中"}, 4: {"担当": "田中"}}, "担当") == []


def test_an_all_integer_sum_stays_an_integer():
    """★ 査読③: float を返すと、型込み等値の事後条件が 60000 と 60000.0 で偽の × を出す。"""
    g = rg.GroupPlan("x", (3, 4))
    got, err = rg.sum_for(g, {3: {"金額": 42000}, 4: {"金額": 18000}}, "金額")
    assert err is None and got == 60000 and isinstance(got, int), (got, type(got))


def test_a_non_numeric_value_is_refused_not_guessed():
    g = rg.GroupPlan("x", (3,))
    got, err = rg.sum_for(g, {3: {"金額": "1,200"}}, "金額")
    assert got is None and err and "数値ではありません" in err, err


# --- ずれ（1 箇所で決める）------------------------------------------------------------

def test_the_shift_is_decided_in_one_place():
    """★ 埋める側（Basic）と確かめる側が別々にずれを数えると、片方だけ直る。"""
    assert rg.output_rows_for(7, 11, 3) == [7]
    assert rg.output_rows_for(11, 11, 3) == [11, 12, 13]
    assert rg.output_rows_for(13, 11, 3) == [15]
    assert rg.output_rows_for(13, None, 1) == [13]
    assert [rg.detail_index_for(r, 11, 3) for r in (11, 12, 13, 15)] == [0, 1, 2, 0]


# --- ① 断らない（領収書・納品書は取引ごとに 1 枚が正しい）------------------------------

def _plain_book(tmp_path, name="b.xlsx", dup=True, marks=None):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "項目", "金額", "担当"])
    ws.append(["丸和物流", "配送", 100, "田中"])
    ws.append(["ヤマノ食品", "仕入", 200, "田中"])
    if dup:
        ws.append(["ヤマノ食品", "配送", 300, "田中"])
    tp = wb.create_sheet("雛形")
    tp["A1"] = "請求書"
    tp["A3"] = "{{取引先}}"
    for cell, m in (marks or {}).items():
        tp[cell] = m
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["売上", "雛形"],
            "headers": {"売上": ["取引先", "項目", "金額", "担当"], "雛形": []},
            "header_rows": {"売上": 1, "雛形": 1}, "path": str(path)}


def test_duplicates_without_detail_marks_are_warned_not_refused(tmp_path):
    """★★ 査読①（俺が開けかけた穴）: ここで断ると、取引ごとに 1 枚が正しい帳票
       ── 領収書・納品書・締め日違いの月別請求 ── が全部作れなくなる。
       既に在る処方は『断ること』ではなく『✓ を出さないこと』だった。"""
    p = _plain_book(tmp_path)
    ok, r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert ok, err
    assert len(r["_report_rows"]) == 3, r["_report_rows"]
    assert any("同じ値が複数" in w for w in r.get("_warnings", [])), r.get("_warnings")


def test_the_warning_names_the_way_out(tmp_path):
    """★ 行き止まりに出口を置く: 「先に集計してから」はこの道具の中に道が無かった。"""
    p = _plain_book(tmp_path)
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    w = " ".join(r.get("_warnings", []))
    assert "{{明細:列名}}" in w and "領収書" in w, w


# --- 雛形に明細の印があれば、まとめる -------------------------------------------------

_DETAIL = {"A5": "{{明細:項目}}", "B5": "{{明細:金額}}", "A8": "{{合計:金額}}"}


def test_detail_marks_switch_to_grouped_mode(tmp_path):
    p = _plain_book(tmp_path, marks=_DETAIL)
    ok, r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert ok, err
    assert [(g["name"], g["rows"]) for g in r["_groups"]] == \
        [("丸和物流", [2]), ("ヤマノ食品", [3, 4])]
    assert r["_detail_row"] == 5
    assert [rr["sheet"] for rr in r["_report_rows"]] == ["丸和物流", "ヤマノ食品"]


def test_a_conflicting_group_value_refuses_with_the_way_out(tmp_path):
    p = _plain_book(tmp_path, marks=dict(_DETAIL, A10="{{担当}}"))
    wb = openpyxl.load_workbook(p)
    wb["売上"].cell(4, 4).value = "佐藤"        # ヤマノ食品の 2 件目だけ担当が違う
    wb.save(p)
    ok, _r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert not ok and "食い違" in err and "{{明細:担当}}" in err, err


def test_a_vertical_merge_across_the_detail_row_is_refused(tmp_path):
    """★★ 査読④: 値は合うので事後条件は通ってしまう ── 雛形を見る段で断る。"""
    p = _plain_book(tmp_path, marks=_DETAIL)
    wb = openpyxl.load_workbook(p)
    wb["雛形"].merge_cells("D4:D6")
    wb.save(p)
    ok, _r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert not ok and "結合" in err and "D4:D6" in err, err


def test_a_horizontal_merge_is_left_alone(tmp_path):
    """★ 黙りすぎない側の対: 横の結合（見出しの中央寄せ）は崩れないので断らない。"""
    p = _plain_book(tmp_path, marks=_DETAIL)
    wb = openpyxl.load_workbook(p)
    wb["雛形"].merge_cells("A1:D1")
    wb.save(p)
    ok, _r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert ok, err


# --- 事後条件は別の証明（混ぜない）----------------------------------------------------

def test_the_postcondition_is_routed_by_the_declaration_not_the_op_name():
    """★ `_groups` が在れば別の証明を使う ── 1 つの関数に混ぜると、片方の分岐が
       恒真でも全体は緑に見える。"""
    seg = window_around("def _check_report_router(", after=900)
    assert 'args.get("_groups")' in seg and "check_report_per_group" in seg, seg[:300]
    assert ailine.POSTCONDITIONS["REPORT_PER_ROW"] is ailine._check_report_router


# --- 事後条件の中身（敵対検体）--------------------------------------------------------
#
# ★ 実機を通さずに「まとめた紙」を手で組み立てて、壊し方ごとに × が出るかを見る。
#   ★ 変異試験（負の被覆の系譜）: **埋め手を 1 行ずらす**と赤くなること ──
#     ずれ関数を埋める側と確かめる側で共有した代償は、ここで払う。

_TPL = {"A1": "請求書", "A3": "{{取引先}}", "A5": "{{明細:項目}}", "B5": "{{明細:金額}}",
         "A8": "{{合計:金額}}", "A10": "備考", "B10": "毎度どうも"}
_PH = [
    {"cell": "A3", "row": 3, "col": 1, "column_name": "取引先", "kind": "value",
     "mark": "取引先", "whole": True, "raw": "{{取引先}}", "col_idx": 1},
    {"cell": "A5", "row": 5, "col": 1, "column_name": "項目", "kind": "detail",
     "mark": "明細:項目", "whole": True, "raw": "{{明細:項目}}", "col_idx": 2},
    {"cell": "B5", "row": 5, "col": 2, "column_name": "金額", "kind": "detail",
     "mark": "明細:金額", "whole": True, "raw": "{{明細:金額}}", "col_idx": 3},
    {"cell": "A8", "row": 8, "col": 1, "column_name": "金額", "kind": "total",
     "mark": "合計:金額", "whole": True, "raw": "{{合計:金額}}", "col_idx": 3},
]
_GARGS = {"_groups": [{"sheet": "丸和物流", "name": "丸和物流", "rows": [2]},
                       {"sheet": "ヤマノ食品", "name": "ヤマノ食品", "rows": [3, 4]}],
           "_placeholders": _PH, "_inspection_sheet": "検分", "template_sheet": "雛形",
           "_target_sheet": "売上", "_detail_row": 5}


def _grouped_book(tmp_path, name="out.xlsx", shift=0, total_delta=0, damage=None, drop=False):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "項目", "金額", "担当"])
    ws.append(["丸和物流", "配送", 100, "田中"])
    ws.append(["ヤマノ食品", "仕入", 200, "田中"])
    ws.append(["ヤマノ食品", "配送", 300, "田中"])
    tp = wb.create_sheet("雛形")
    for cell, v in _TPL.items():
        tp[cell] = v
    for g in _GARGS["_groups"]:
        rows = g["rows"][:1] if drop and len(g["rows"]) > 1 else g["rows"]
        n = len(rows)
        out = wb.create_sheet(g["sheet"])
        for cell, v in _TPL.items():
            r = int(cell[1:])
            for orow in rg.output_rows_for(r, 5, n):
                out.cell(orow, ord(cell[0]) - 64, v)
        for k, src_row in enumerate(rows):
            out.cell(5 + k + shift, 1, ws.cell(src_row, 2).value)
            out.cell(5 + k + shift, 2, ws.cell(src_row, 3).value)
        out.cell(8 + n - 1, 1, sum(ws.cell(r, 3).value for r in rows) + total_delta)
        out.cell(3, 1, g["name"])
        if damage and g["sheet"] == "ヤマノ食品":
            out[damage[0]] = damage[1]
    ins = wb.create_sheet("検分")
    ins.append(["シート名", "元の行", "まとめた件数"])
    for g in _GARGS["_groups"]:
        ins.append([g["sheet"], ",".join(str(r) for r in g["rows"]), len(g["rows"])])
    wb.save(p)
    return p


def test_a_correct_grouped_output_passes(tmp_path):
    p = _grouped_book(tmp_path)
    status, reason = ailine.check_report_per_group(p, dict(_GARGS))
    assert status == "pass", reason
    assert "取引先2件" in reason and "明細 3 行" in reason, reason


def test_shifting_the_detail_fill_by_one_row_fails(tmp_path):
    """★★ 変異試験: ずれ関数を両側で共有した代償はここで払う。"""
    p = _grouped_book(tmp_path, shift=1)
    status, reason = ailine.check_report_per_group(p, dict(_GARGS))
    assert status == "fail", reason


def test_a_wrong_total_fails(tmp_path):
    """★ 合計は**足し直して**比べる（宣言値と宣言値を比べたら恒真）。"""
    p = _grouped_book(tmp_path, total_delta=1)
    status, reason = ailine.check_report_per_group(p, dict(_GARGS))
    assert status == "fail" and "足すと" in reason, reason


def test_a_damaged_fixed_cell_below_the_detail_block_fails(tmp_path):
    """★★ ⑤の芯: 行の挿入は「増やした所」より**押し下げた所**が静かに壊れる。
       印でないセル（備考）が雛形と違えば落ちること。"""
    p = _grouped_book(tmp_path, damage=("B11", "消えた"))
    status, reason = ailine.check_report_per_group(p, dict(_GARGS))
    assert status == "fail" and "印でないセル" in reason, reason


def test_a_dropped_order_fails(tmp_path):
    """★ グループの完全会計: 2 件のうち 1 件しか刷らなければ落ちること。"""
    p = _grouped_book(tmp_path, drop=True)
    status, reason = ailine.check_report_per_group(p, dict(_GARGS))
    assert status == "fail", reason


def test_a_changed_template_fails(tmp_path):
    """★ 雛形を書き換えると、⑤（印でないセルが雛形のまま）が**先に**落とす。
       ⑥（雛形/データシート無変更）まで行かないが、止まる位置が早いだけで通らない。"""
    src = _grouped_book(tmp_path, name="src.xlsx")
    out = _grouped_book(tmp_path, name="o.xlsx")
    wb = openpyxl.load_workbook(out)
    wb["雛形"]["A1"] = "書き換えた"
    wb.save(out)
    status, reason = ailine.check_report_per_group(out, dict(_GARGS), source_book=src)
    assert status == "fail", reason


def test_a_changed_data_sheet_fails(tmp_path):
    """★ ⑥: データシートは**読むだけ**のはず。1 セルでも変わったら落とす
       （こちらは ⑤ の見ない所なので、⑥ そのものが働いていることが見える）。"""
    src = _grouped_book(tmp_path, name="s2.xlsx")
    out = _grouped_book(tmp_path, name="o2.xlsx")
    wb = openpyxl.load_workbook(out)
    wb["売上"].cell(2, 4).value = "書き換えた"      # 担当列（どの印も指していない）
    wb.save(out)
    status, reason = ailine.check_report_per_group(out, dict(_GARGS), source_book=src)
    assert status == "fail" and "読むだけのはず" in reason, reason


def test_the_way_out_is_only_offered_when_it_exists(tmp_path):
    """★ 「足すなら {{合計:担当}}」は担当のような文字列の列では意味を成さない
       ── 出せる道だけを名指しする（行き止まりに出口を置く、の逆側の作法）。"""
    p = _plain_book(tmp_path, marks=dict(_DETAIL, A10="{{担当}}"))
    wb = openpyxl.load_workbook(p)
    wb["売上"].cell(4, 4).value = "佐藤"
    wb.save(p)
    _ok, _r, _i, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, _meta(p))
    assert "{{明細:担当}}" in err and "{{合計:担当}}" not in err, err
