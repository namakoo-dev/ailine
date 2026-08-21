"""M3（2冊の照合）の追加検体（実装後・ナギが独自に足した分）。
   tests/test_match_e2e.py（実装前に凍結された specimen-first の7検体）とは別ファイル
   ── 依頼どおり「検体ファイルとは別に」自分で足すテスト。

   ここに置くもの:
   - 決定論（同一入力を2回走らせてセル内容が完全一致）
   - 変異試験（差額セルを改竄し、事後条件が非0で止め、出力を本置き場に書かないことを見る）
   - 対称性の追試（凍結specimenとは別データセットで A/B 入替の符号反転を確認）
   - 会計恒等の JSON（--json の a_rows/keys/mismatched が独立に再計算した値と一致すること）
"""
import dataclasses
import datetime
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402
from ailine_core import match as multifile_match  # noqa: E402


def _book(path, headers, rows, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _cells(path):
    """出力の中身（シート名・セル値・型）を決定論比較用に正規化して返す
       （tests/test_stack_e2e.py の _cells と同じ線）。"""
    wb = openpyxl.load_workbook(path)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    out.append((ws.title, c.coordinate, type(c.value).__name__, str(c.value)))
    wb.close()
    return out


NYUKIN_HDRS = ["取引日", "振込人名義", "お預り金額", "摘要"]
SEIKYU_HDRS = ["請求日", "取引先名", "請求金額", "請求番号"]
TASK = "振込人名義と取引先名をキーに、お預り金額と請求金額を突き合わせて"


def _dataset(tmp_path):
    a = tmp_path / "in" / "入金明細_7月.xlsx"
    b = tmp_path / "in" / "請求_7月.xlsx"
    _book(a, NYUKIN_HDRS, [
        (datetime.date(2026, 7, 31), "甲社", 220000, "7月分"),
        (datetime.date(2026, 7, 31), "乙社", 109340, "手数料引き"),
        (datetime.date(2026, 8, 1), "丁社", 55000, "台帳に無い"),
        (datetime.date(2026, 8, 2), "戊社", 330000, "2件合算"),
    ], "入金明細")
    _book(b, SEIKYU_HDRS, [
        (datetime.date(2026, 7, 10), "甲社", 220000, "INV-101"),
        (datetime.date(2026, 7, 12), "乙社", 110000, "INV-102"),
        (datetime.date(2026, 7, 15), "丙社", 88000, "INV-103"),
        (datetime.date(2026, 7, 20), "戊社", 110000, "INV-104"),
        (datetime.date(2026, 7, 25), "戊社", 220000, "INV-105"),
    ], "請求一覧")
    return a, b


def test_deterministic_output_across_two_runs(tmp_path, capsys):
    """決定論: 同一入力で2回走らせて、出力ブックのセル内容（値・型・シート）が完全一致する
       こと ── 2回目は自分の前回出力を『作り直す』経路（rebuilt_own_output）を通る。"""
    a, b = _dataset(tmp_path)
    rc1 = ailine.main(["run", str(a), str(b), TASK])
    capsys.readouterr()
    assert rc1 == 0
    out1 = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    cells1 = _cells(out1)

    rc2 = ailine.main(["run", str(a), str(b), TASK])
    out = capsys.readouterr().out
    assert rc2 == 0
    assert "作り直し" in out, f"前回出力の作り直しが開示されていない:\n{out}"
    out2 = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    assert out2 == out1, "2回目が別ファイルに書かれた（同条件なら同名のはず）"
    cells2 = _cells(out2)
    assert cells1 == cells2, "同一入力なのに出力の中身（値・型・シート）が一致しない"


def test_mutated_diff_cell_is_caught_by_postcondition(tmp_path, monkeypatch, capsys):
    """変異試験: compute_match の戻り値（差額）を改竄して、書いた直後の独立読み検算
       （xml_readback）が非0で止まり、出力を本置き場（A の親）に書かないことを見る。
       ★ Σ(a_sum)/Σ(b_sum) は保存したまま diff だけ壊す（他の事後条件を巻き込まない）。"""
    a, b = _dataset(tmp_path)
    real = multifile_match.compute_match

    def mutant(*args, **kwargs):
        groups = real(*args, **kwargs)
        mutated = []
        touched = False
        for g in groups:
            if not touched and g.key_display != multifile_match.UNKNOWN_KEY_LABEL:
                mutated.append(dataclasses.replace(g, diff=g.diff + 999))
                touched = True
            else:
                mutated.append(g)
        return mutated

    monkeypatch.setattr(multifile_match, "compute_match", mutant)
    before = set((tmp_path / "in").glob("*.xlsx"))
    rc = ailine.main(["run", str(a), str(b), TASK])
    out = capsys.readouterr().out
    assert rc != 0, f"改竄した差額セルで合格した:\n{out}"
    assert "事後条件が破れた" in out, out
    after = set((tmp_path / "in").glob("*.xlsx"))
    assert after == before, "事後条件が破れたのに出力を本置き場に書いた"


def test_symmetry_with_a_second_independent_dataset(tmp_path, capsys):
    """対称性の追試（凍結specimenとは別データセット）: A/B を入替えると差額の符号だけ
       反転し、キー集合・件数・状態の対応は同一であること。"""
    a = tmp_path / "in2" / "入金.xlsx"
    b = tmp_path / "in2" / "請求.xlsx"
    _book(a, NYUKIN_HDRS, [
        (datetime.date(2026, 6, 1), "A商事", 5000, ""),
        (datetime.date(2026, 6, 2), "B商店", 12000, ""),
        (datetime.date(2026, 6, 3), "C工業", 7000, "端数"),
    ], "入金")
    _book(b, SEIKYU_HDRS, [
        (datetime.date(2026, 6, 1), "A商事", 5000, "INV-1"),
        (datetime.date(2026, 6, 2), "B商店", 11500, "INV-2"),
        (datetime.date(2026, 6, 4), "D工務店", 3000, "INV-3"),
    ], "請求")
    task = "振込人名義と取引先名をキーに、お預り金額と請求金額を突き合わせて"
    task_rev = "取引先名と振込人名義をキーに、請求金額とお預り金額を突き合わせて"

    rc1 = ailine.main(["run", str(a), str(b), task])
    capsys.readouterr()
    assert rc1 == 0
    out1 = max((tmp_path / "in2").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    rows1 = {r[0]: r for r in _rows(out1)[1:] if r and r[0]}

    rc2 = ailine.main(["run", str(b), str(a), task_rev])
    capsys.readouterr()
    assert rc2 == 0
    out2 = max((tmp_path / "in2").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    assert out2 != out1
    rows2 = {r[0]: r for r in _rows(out2)[1:] if r and r[0]}

    assert set(rows1) == set(rows2), "キー集合が入替えで変わった"
    for k in rows1:
        d1, d2 = rows1[k][5], rows2[k][5]
        assert d1 == -(d2 or 0) if d1 is not None else d2 is None, \
            f"{k}: 符号だけの反転になっていない: {d1} vs {d2}"


def _rows(path):
    ws = openpyxl.load_workbook(path)["照合"]
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_json_accounting_identity_matches_independent_recount(tmp_path, capsys):
    """会計恒等の JSON: --json の a_rows/b_rows/keys が、こちらで独立に数えた
       行数・キー数と一致すること（②照合出力のΣ ではなく①CLI の報告面を検体化）。"""
    a, b = _dataset(tmp_path)
    rc = ailine.main(["run", str(a), str(b), TASK, "--json"])
    out = capsys.readouterr().out
    assert rc == 0, out
    import json
    data = json.loads(out.strip().splitlines()[-1])
    assert data["a_rows"] == 4, data      # 入金明細のデータ行数（合計行なし・単純に4行）
    assert data["b_rows"] == 5, data      # 請求のデータ行数
    # キー数: 甲社・乙社・丙社・丁社・戊社 の5キー（キー不明なし）。
    assert data["keys"] == 5, data
    # 差額あり: 乙社(-660)・丙社(Bのみ)・丁社(Aのみ) の3件。戊社・甲社は差額0。
    assert data["mismatched"] == 3, data
    assert data["keys"] - data["mismatched"] == 2, "差額0のキー数(甲社・戊社)が合わない"
