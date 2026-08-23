"""DEDUP の実機（LO・basrun）検体。★ 2026-08-21 の実弾で DedupRows が実機で無限ループ
   （Collection.Add の重複キーエラー頼みのイディオムが LO を固める・単離検証で確定）。
   sandbox の pytest は basrun をモックするため構造的に捕まえられない ── 実機検体で凍結する。
   ★ タイムアウトは短く（楔でスイート全体を道連れにしない）。"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent


@pytest.mark.local
def test_dedup_rows_runs_on_real_lo_and_keeps_first_per_key(tmp_path):
    book = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["取引先", "金額"])
    for r in [["甲社", 100], ["甲社", 250], ["乙社", 200]]:
        ws.append(r)
    wb.save(book)
    p = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(book),
         "取引先が同じ行を重複として除いて", "--copy", "--timeout", "60"],
        capture_output=True, text=True, timeout=300, encoding="utf-8")
    assert p.returncode == 0, f"実機 DEDUP が失敗（無限ループの再発?）:\n{p.stdout[-600:]}"
    out = book.with_name(book.stem + ".out" + book.suffix)
    wb2 = openpyxl.load_workbook(out)
    dst = [s for s in wb2.sheetnames if "重複" in s]
    assert dst, wb2.sheetnames
    rows = [[c.value for c in r] for r in wb2[dst[0]].iter_rows(min_row=2)]
    assert rows == [["甲社", 100], ["乙社", 200]], f"最初の 1 行を残す、が破れた: {rows}"
