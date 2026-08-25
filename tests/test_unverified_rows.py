# 塊①（2026-08-25）── 検証できなかった行を、機械の値として残して ✓ を降ろす。
#
# ★ 実測した事故（中核 op の盲検・2026-08-24・致命6）:
#   単価列に文字列の数値が 2 行混ざった 10 行の表に「高い順に並べ替えて」と頼むと、
#     事後条件を確認（操作:並べ替え）: 8 行を検証（降順）（数値でない 2 行は対象外）
#     ✓ w_t19.out.xlsx は機械検証済みの内容です
#   実物の単価列は **250, 1500, 1200, ...** ── 頼んだ「高い順」の先頭が 250。
#   ★ 除外されるのは、**まさに主張を壊す行だけ**。検証対象が自分に都合よく選ばれている。
#
# ★ 根（3 面の盲検が別々に着いた一つの形）:
#   判定に要る項が機械の値として在るのに、判定へ渡していない。
#   `excluded` は int として数えられ、**文章にして捨てられていた**。
#
# 契約:
#   ① 検証できなかった行は **機械の値**として残る（表示文から読み取らせない）
#   ② 1 行でも在れば ✓ を名乗らない（△ に降りる ── 8 行は本当に検証したので ⚠ でもない）
#   ③ 何行が・なぜ検証できなかったかを名指しする
#   ④ 除外ゼロの時は 1 文字も増えない（誤爆しない）
#   ⑤ 除外を数える checker すべてが同じ器官を通る（片配線の禁止・書き写さない）

import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402
from test_ailine import _book  # noqa: E402


# --- ① 機械の値として残る ------------------------------------------------------------

def test_sort_records_unverified_rows_as_data(tmp_path):
    """★ 実測の再現: 文字列の数値が混ざると、その行は検証されないまま ✓ が出ていた。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 250], ["b", "1500"], ["c", 1200],
                          ["d", "250"], ["e", 100]])
    args = {"col": "単価", "order": "desc"}
    status, reason = ailine.check_sort(p, args)
    assert args.get("_unverified"), "検証できなかった行が機械の値として残っていない"
    assert sum(u["rows"] for u in args["_unverified"]) == 2, args["_unverified"]


def test_no_unverified_key_when_everything_was_checked(tmp_path):
    """④ 誤爆しない: 全部数値なら、印は 1 つも付かない。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 300], ["b", 200], ["c", 100]])
    args = {"col": "単価", "order": "desc"}
    status, reason = ailine.check_sort(p, args)
    assert status == "pass"
    assert not args.get("_unverified"), f"除外ゼロなのに印が付いた: {args.get('_unverified')}"


# --- ③ 何行が・なぜ、を名指しする -----------------------------------------------------

def test_unverified_lines_name_the_count_and_the_reason():
    lines = ailine.render_unverified_advisories(
        [{"rows": 2, "why": "数値でないため並び順を確かめられない"}])
    text = "\n".join(lines)
    assert text.lstrip().startswith("⚠"), f"決裁③が数える形（⚠ 始まり）でない: {text}"
    assert "2 行" in text
    assert "数値でない" in text


def test_no_lines_when_nothing_was_excluded():
    assert ailine.render_unverified_advisories([]) == []
    assert ailine.render_unverified_advisories(None) == []


# --- ② ✓ を名乗らない（e2e）----------------------------------------------------------

def test_run_does_not_claim_verified_when_rows_were_excluded(tmp_path, monkeypatch, capsys):
    """★ これが盲検の実測そのもの。8 行を検証したことは事実なので ⚠ ではなく △。"""
    import openpyxl
    from test_golden_transcripts import _isolate, _run_main
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, [["商品", "単価"], ["a", 250], ["b", "1500"], ["c", 1200],
                          ["d", "250"], ["e", 100]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "単価", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        # LibreOffice と同じ落とし方を模す: 数値行だけが降順に並び、文字列はその場に残る
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        rows = [[ws.cell(r, c).value for c in (1, 2)] for r in range(2, ws.max_row + 1)]
        nums = sorted([r for r in rows if isinstance(r[1], (int, float))],
                       key=lambda r: r[1], reverse=True)
        it = iter(nums)
        for i, r in enumerate(rows):
            if isinstance(r[1], (int, float)):
                rows[i] = next(it)
        for i, r in enumerate(rows, start=2):
            ws.cell(i, 1, r[0]); ws.cell(i, 2, r[1])
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    # ★ --header-row 1 を明示する: 対象列に数値と文字列が混ざると見出し行を
    #   判定できない（中核 op の盲検 中#8・**別の所見**）。この検体はそちらでなく
    #   「除外された行が在るのに ✓ が出る」を測りたいので、そこは固定して外す。
    rc, out = _run_main(["run", str(p), "単価が高い順に並べ替えて", "--copy",
                          "--header-row", "1"], capsys)
    assert "対象外" in out or "検証できていません" in out, f"前提: 除外が起きること: {out}"
    assert "✓" not in out, f"検証できていない行が在るのに ✓ を名乗った: {out}"
    assert "△" in out, f"8 行は本当に検証したのに、機械保証なし扱いに落とした: {out}"


# --- ⑤ 全員が同じ器官を通る（片配線の禁止）--------------------------------------------

@pytest.mark.parametrize("name", [
    "check_sort", "check_compute_column", "check_compute_column_single_factor",
    "check_lookup_fill",
])
def test_every_excluding_checker_uses_the_shared_organ(name):
    """★ 除外を数える checker が、自前で握りつぶさず共通の器官へ渡すこと。
       今日までに片配線を 6 回踏んでいる ── 書き写しは構造で禁じる。"""
    import inspect
    src = inspect.getsource(getattr(ailine, name))
    assert "note_unverified(" in src, f"{name} が共通の器官を通っていない"
