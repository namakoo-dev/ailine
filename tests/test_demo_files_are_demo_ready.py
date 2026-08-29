# デモ材料の番人 ── 2026-08-29。Namakoo の通しで 2 度つまずいた所を機械で縛る。
#
# ★★ 実測①: 売上列が**空に見えた** ──「5 個のセルは数式で、計算結果がファイルに
#   入っていません」。openpyxl で作った式にはキャッシュ値が無い。デモで空列は致命的。
#   ★ 直しは LibreOffice に開かせて保存し直すこと（実物に計算させる）。
# ★★ 実測②: シート名『売上』と列名『売上』が同じだと「2 通りに読めます」の 3 択が出る。
#   道具の振る舞いは正しい（曖昧なら聞く）ので、**検体の側**を直す。
#
# ★ どちらも「当日その場で気づく」類なので、機械が先に気づくようにする。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
DEMO = REPO / "demo"


def _books():
    return sorted(p for p in DEMO.glob("*.xlsx")
                   if not any(w in p.stem for w in ("（下書き）", "（捨てた）", "（前回）", ".out")))


def test_the_demo_folder_has_the_material():
    names = {p.stem for p in _books()}
    assert {"1_請求_2026年8月", "4_売上_合計つき", "5_参照つき"} <= names, names


@pytest.mark.parametrize("book", _books(), ids=lambda p: p.stem)
def test_every_formula_has_a_cached_value(book):
    """★ 実測①: 式のキャッシュ値が無いと、画面でその列が**空に見える**。"""
    raw = openpyxl.load_workbook(book)
    val = openpyxl.load_workbook(book, data_only=True)
    missing = []
    for name in raw.sheetnames:
        ws, wv = raw[name], val[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if wv.cell(cell.row, cell.column).value is None:
                        missing.append(f"{name}!{cell.coordinate}")
    assert not missing, (
        f"計算結果がファイルに入っていない式: {missing[:5]}"
        " ── LibreOffice に開かせて保存し直すこと（openpyxl で書いた式には値が付かない）")


@pytest.mark.parametrize("book", _books(), ids=lambda p: p.stem)
def test_no_sheet_name_collides_with_a_column_name(book):
    """★ 実測②: シート名と列名が同じだと「2 通りに読めます」の 3 択が出る
       ── 道具は正しいが、実演では事故になる。"""
    wb = openpyxl.load_workbook(book)
    sheets = set(wb.sheetnames)
    bad = []
    for name in wb.sheetnames:
        ws = wb[name]
        heads = {str(ws.cell(1, c).value).strip()
                  for c in range(1, (ws.max_column or 0) + 1)
                  if ws.cell(1, c).value not in (None, "")}
        for h in heads & sheets:
            bad.append(f"{name}!見出し『{h}』")
    assert not bad, f"シート名と同じ見出しがある: {bad}"
