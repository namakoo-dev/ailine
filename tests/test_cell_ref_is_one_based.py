# `_cell_ref` の起点を凍結する（2026-09-03）。
#
# ★★ なぜ在るか: docstring が「0起点」と書いていたが、実装は 1 起点を要求し、
#   呼び元 8 箇所すべてが 1 起点を渡していた ── **説明だけが取り残されていた**。
#   ★ そして数えたら、この repo には**起点が 2 つある**:
#     0 起点 … LibreOffice Basic 側（getCellByPosition・ヘルパ引数・LLM への説明・used_range）
#     1 起点 … openpyxl 側と検算側（_col_index_by_header・スナップショットの r,c）
#   ★ 2 つの世界が同居する以上、**どちらの側の関数かを機械で固定**しないと、
#     いつか 0 起点の値が 1 起点の関数へ渡る。
#
# 契約:
#   ① 1 起点で正しい A1 を返す
#   ② ★ 0 を渡したら**落ちる**（黙って B0 のような嘘の座標を作らない）
#   ③ 呼び元が 1 起点の側から来ていること（_col_index_by_header が 1 起点を返す）

import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

from ailine_core.table_scan import _cell_ref, _col_index_by_header  # noqa: E402


def test_one_based_coordinates_render_as_a1():
    """① 1 起点で正しい A1 になる。"""
    assert _cell_ref(1, 1) == "A1"
    assert _cell_ref(2, 3) == "C2"
    assert _cell_ref(10, 27) == "AA10"


def test_a_zero_column_raises_instead_of_lying():
    """② 0 を渡したら落ちる ── ★ 黙って嘘の座標を作らない。

    ★ 「B0」のような存在しない座標を人に見せるより、その場で落ちる方が安全。
      0 起点の側（Basic）から来た値が紛れ込んだことに、その場で気づける。
    """
    with pytest.raises(ValueError):
        _cell_ref(1, 0)


def test_the_column_index_source_is_one_based():
    """③ 呼び元の出所（_col_index_by_header）が 1 起点であること。

    ★ ここが 0 起点に変わると、_cell_ref の全出力が 1 列ずれる ──
      しかも**エラーにならず、静かに間違った座標を出し続ける**。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "品名"
    ws.cell(row=1, column=2).value = "金額"
    assert _col_index_by_header(ws, "品名") == 1, "1 起点でない"
    assert _col_index_by_header(ws, "金額") == 2
    assert _col_index_by_header(ws, "無い列") is None
    # ★ 出所と組み合わせて、実際に人が読む座標になること
    assert _cell_ref(1, _col_index_by_header(ws, "金額")) == "B1"
