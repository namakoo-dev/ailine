# operator 盲検 10 回目（$59・PRED-20260823-operator10.md）の BROKEN 修正 ── 実装より先に
# 凍結した赤い検体。★ 今回の所見の性質: データ破壊系はほぼ消え、「見せ方・言い方」へ論点が
# 移った ── ①新機能の意味論の穴 ②誤アラーム（✓ を毀損するオオカミ少年） ③型の黙示変換。
#
# 契約:
#   ① グラフの範囲から合計行を除く（stack/extract は除いているのに CHART だけ未配線＝片配線。
#      実測: 集計!$A$2:$A$5 に「合計 1750」が第 4 の柱として混入し、管理者が総額を部門と誤読）
#   ② シートに**グラフだけ**加わった run で「存在しません/変更されていません」の嘘を出さない
#      （実測: 集計シートにグラフを挿したのに ★ が出て ✓ が △ に落ちた）
#   ③ --sheet で対象シートが確定しているとき、そのシートに無い列名を LLM が返しても、
#      依頼文が名指しする実在列を優先する（実測: 別シートの『値』を聞き返して行き止まり）
#   ④ SET_COLUMN_VALUE の型は列の実体から機械決定して開示する
#      （実測: 数値列が '10'/書式 @ の文字列になり ✓ ── 下流の SUM が黙って壊れる）
#   ⑤ 数式セルはキャッシュ値が数値なら「文字列に変わった」と言わない（偽の破壊アラーム）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402

FIXTURES = Path(__file__).parent / "fixtures" / "charts"

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "chart_data_last_row"),
    reason="operator10 修正 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


def _summary_book(tmp_path, name="集計.xlsx"):
    """部門 3 行 + 合計行（stack/extract が既に除外している形）。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"
    for r in [["部門", "金額"], ["営業", 800], ["開発", 700], ["総務", 250], ["合計", 1750]]:
        ws.append(r)
    wb.save(p)
    return p


# --- ① グラフ範囲から合計行を除く -------------------------------------------------

@needs_impl
def test_chart_range_excludes_total_row(tmp_path):
    """合計行のある表 ── グラフのデータ範囲は 4 行目まで（5 行目の合計を含まない）。"""
    p = _summary_book(tmp_path)
    last = ailine.chart_data_last_row(p, sheet_name="集計", header_row=1)
    assert last == 4, f"合計行がグラフ範囲に入っている（第 4 の柱になる）: last_row={last}"


@needs_impl
def test_chart_range_keeps_all_rows_without_total(tmp_path):
    """誤爆防止: 合計行が無い表では最終行まで使う。"""
    p = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"
    for r in [["部門", "金額"], ["営業", 800], ["開発", 700], ["総務", 250]]:
        ws.append(r)
    wb.save(p)
    assert ailine.chart_data_last_row(p, sheet_name="集計", header_row=1) == 4


# --- ② グラフだけの変更を「変更されていません」と言わない --------------------------

@needs_impl
def test_chart_only_change_is_not_called_unchanged(tmp_path, monkeypatch, capsys):
    """集計シートにグラフを挿しただけの run ── 嘘の ★ を出さず ✓ まで通る。"""
    _isolate(monkeypatch, tmp_path)
    book = _summary_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "CHART", "args": {"value_col": "金額", "kind": "bar", "category_col": "部門"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        import shutil
        # 実 LO 産のグラフ入りブックの chart XML を持つ形に差し替える代わりに、
        # 値は変えずグラフだけが増えた状態を openpyxl の BarChart で作る
        from openpyxl.chart import BarChart, Reference
        wb = openpyxl.load_workbook(out_book)
        ws = wb["集計"]
        ch = BarChart()
        ch.add_data(Reference(ws, min_col=2, min_row=2, max_row=4), titles_from_data=False)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
        ws.add_chart(ch, "D2")
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "集計シートで部門ごとの棒グラフを作って", "--copy"],
                         capsys)
    assert "変更されていません" not in out, f"グラフだけの変更を『変更なし』と嘘をついた: {out}"


# --- ③ --sheet 指定時の列解決（別シートの列名に汚染されない）------------------------

@needs_impl
def test_explicit_sheet_prefers_task_named_real_column(tmp_path, monkeypatch, capsys):
    """--sheet 売上データ + 依頼文が『金額』── LLM が別シートの『値』を返しても、
       依頼文が名指しし対象シートに実在する『金額』を優先する（A' 原則の適用）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "collide.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上データ"
    for r in [["商品", "数量", "金額"], ["a", 2, 100], ["b", 1, 300]]:
        ws.append(r)
    other = wb.create_sheet("金額")
    other.append(["項目", "値"])
    other.append(["税率", 10])
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "値", "order": "desc"}})   # ← 別シートの列に汚染された返答

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2["売上データ"]
        ws2["A2"], ws2["B2"], ws2["C2"] = "b", 1, 300
        ws2["A3"], ws2["B3"], ws2["C3"] = "a", 2, 100
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "金額を降順に並べ替えて", "--sheet", "売上データ",
                          "--copy"], capsys)
    assert "列『値』がありません" not in out, f"別シートの列名で行き止まりになった: {out}"
    assert rc == 0, out
    assert "金額" in out


# --- ④ SET_COLUMN_VALUE の型を列の実体から決めて開示 --------------------------------

@needs_impl
def test_set_column_value_keeps_numeric_column_numeric(tmp_path, monkeypatch, capsys):
    """数値列に「10」を一括書換 ── 数値として書き、型の決定を開示する。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "数量"], ["a", 10], ["b", 2], ["c", 50]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SET_COLUMN_VALUE", "args": {"col": "数量", "value": "10"}})
    seen = {}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        seen["code"] = code
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for row in range(2, 5):
            ws.cell(row=row, column=2, value=10)   # 数値として書く（codegen の意図を再現）
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "数量の列を全部「10」にして", "--copy", "--overwrite"],
                         capsys)
    assert rc == 0, out
    assert "setValue" in seen.get("code", ""), \
        f"数値列なのに文字列で書く codegen（setString）のまま: {seen.get('code','')[:300]}"
    assert "数値" in out, f"型の決定が開示されていない: {out}"


def test_set_column_value_keeps_text_column_text(tmp_path, monkeypatch, capsys):
    # xfail 対象外: 現状も緑（型の実体判定を入れても文字列列は文字列のままであることの番人）
    """誤爆防止: 文字列列への「確認済み」は従来どおり文字列。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "備考"], ["a", "旧1"], ["b", "旧2"]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SET_COLUMN_VALUE", "args": {"col": "備考", "value": "確認済み"}})
    seen = {}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        seen["code"] = code
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for row in range(2, 4):
            ws.cell(row=row, column=2, value="確認済み")
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "備考の列を全部「確認済み」にして", "--copy",
                          "--overwrite"], capsys)
    assert rc == 0, out
    assert "setString" in seen.get("code", ""), "文字列列が数値書きに倒れた（誤爆）"


# --- ⑤ 数式セルへの偽「文字列になった」アラーム ------------------------------------

@needs_impl
def test_formula_cells_with_numeric_cache_are_not_called_strings(tmp_path):
    """数式（=B2*C2・キャッシュ値が数値）は『数値に見えない文字列に変わった』と言わない。"""
    before = {"cells": {"Sheet!C2": 100, "Sheet!C3": 200}}
    p = tmp_path / "after.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b", "c"])
    ws["C2"] = "=A2*B2"
    ws["C3"] = "=A3*B3"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"C2": 100, "C3": 200})
    assert ailine.numeric_cells_became_strings(p, ["C2", "C3"], sheet_name="Sheet") == [], \
        "数式セルを『文字列になった』と誤検出（偽の破壊アラーム）"


# ★ 検分の差し戻し（2026-08-23 16:2x・実機で捕獲）: 上の
#   test_chart_only_change_is_not_called_unchanged は openpyxl 製グラフの治具で緑だったが、
#   実 LibreOffice 産のブックでは charts_by_sheet が 0 を返し、実経路では嘘の
#   「変更されていません」が出続けていた（rels の Target が "../drawings/..." の相対形で、
#   旧実装は ".." を畳まず xl/ を前置していた）。治具が実物と違う形の 5 度目 ──
#   ここは **spike が実 LO で作った fixture** で縛る。

def test_charts_by_sheet_counts_real_libreoffice_output():
    from ailine_core import chart_check
    for name in ("bar.xlsx", "line.xlsx", "pie.xlsx"):
        got = chart_check.charts_by_sheet(FIXTURES / name)
        assert got == {"Sheet": 1}, f"{name}: 実 LO 産のグラフを数えられない: {got}"
