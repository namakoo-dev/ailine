# 様式写像段 FORMAT_MAP ── helpers/AiLineHelpers.bas:FillFormatMapSheet を実際に basrun
# （実 LibreOffice）で走らせ、3行のデータが雛形の見出しのまま1枚のシートへ縦に展開される
# ことを確かめる通し試験。
#
# ★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
#   （`pytest -m "not local"` で除外される。手動は `pytest -m local` または
#   `pytest tests/test_format_map_local.py`）。
# ★ tests/test_format_map.py（凍結済み・basrun_apply をモックした e2e）と役割が違う:
#   あちらは CLI パイプライン全体（②検証〜⑥事後条件）の契約を固定する。こちらは
#   helpers/AiLineHelpers.bas:FillFormatMapSheet 自体が実 LO で本当に動くこと
#   （見出し行の再構成・印の走査・型の出し分け）を見る（test_report_per_row_local.py と
#   同じ役割分担）。
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))
import ailine  # noqa: E402

HELPERS_DIR = REPO_ROOT / "src" / "ailine" / "helpers"


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1
    return out.lower().count("soffice.bin")


@pytest.mark.local
def test_fill_format_map_sheet_makes_one_sheet_from_3_rows_via_basrun(tmp_path):
    """出納帳3行（+合計行）→ 様式シートの見出しのまま1枚のシートへ縦に展開されることを
       実 LO で確かめる。★ この試験の存在意義: FillFormatMapSheet の見出し行の再構成
       （印行の列だけ左詰め）・印の走査（InStr による {{...}} 検出）・型の出し分け
       （getType→setValue/setString）のどれか1つが壊れても、この試験は赤くなる。
       ★ 合計行の除外は verify_dsl_args(total_row.py)の管轄なので、ここは Python 側が
       決め切った srcRowsCsv をそのまま渡す（合計行の 4 行目は含めない）。"""
    before_procs = _soffice_process_count()
    try:
        book = tmp_path / "ledger.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "出納帳"
        for row in [["日付", "摘要", "入金", "出金"],
                     ["2026-07-01", "売上入金", 50000, 0],
                     ["2026-07-03", "仕入", 0, 12000],
                     ["2026-07-05", "備品購入", 0, 3000],
                     ["合計", "", 50000, 15000]]:
            ws.append(row)
        tpl = wb.create_sheet("様式")
        tpl["A1"] = "取引日"
        tpl["B1"] = "内容"
        tpl["C1"] = "金額"
        tpl["A2"] = "{{日付}}"
        tpl["B2"] = "{{摘要}}"
        tpl["C2"] = "{{入金}}"
        wb.save(book)

        workdir = tmp_path / "work"
        workdir.mkdir()
        _catalog, helper_files = ailine.load_helpers(HELPERS_DIR)
        assert helper_files, "helpers/*.bas が見つからない（HELPERS_DIR の場所を確認）"

        # ★ Basic は 0 起点（headerTplRow/phTplRow/srcHeaderRow は 0 起点）。
        #   codegen_dsl と同じ変換規則（Excel の1起点 row → Basic の row-1）。
        #   srcRowsCsv は合計行(物理5行目=0起点4)を含めない 3 データ行(0起点 1,2,3)。
        code = (
            "Option VBASupport 1\nOption Explicit\n\n"
            "Sub Run(oDoc As Object)\n"
            '    Call FillFormatMapSheet(oDoc, "様式", "出納帳", "様式_出力", 0, 1, 0, "1,2,3")\n'
            "End Sub\n"
        )
        ok, err, raw = ailine.basrun_apply(book, code, workdir, helper_files)
        assert ok, f"basrun_apply が失敗した: {err}\n{(raw or '')[-800:]}"

        wb2 = openpyxl.load_workbook(book)
        made = [s for s in wb2.sheetnames if s not in ("出納帳", "様式")]
        assert made == ["様式_出力"], made

        out = wb2["様式_出力"]
        rows = [[c.value for c in r] for r in out.iter_rows()]
        assert rows[0] == ["取引日", "内容", "金額"], f"人の見出しが保たれていない: {rows[0]}"
        assert len(rows) == 4, f"合計行が混入した疑い: {rows}"
        assert rows[1] == ["2026-07-01", "売上入金", 50000]
        assert isinstance(rows[1][2], (int, float)), f"金額が数値でない: {rows[1][2]!r}"
        assert rows[2] == ["2026-07-03", "仕入", 0]
        assert rows[3] == ["2026-07-05", "備品購入", 0]

        # ★ 憲法: 雛形には一切書き込まない（読むだけ）。
        tpl2 = wb2["様式"]
        assert tpl2["A2"].value == "{{日付}}", "雛形が書き換わった（憲法違反）"
        assert tpl2["B2"].value == "{{摘要}}", "雛形が書き換わった（憲法違反）"
        wb2.close()
    finally:
        # ★ 後片付け: basrun.py stop（接続先の LibreOffice だけを終了・taskkill 一括はしない）。
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )
