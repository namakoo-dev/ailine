# 座標の法則 ── 2026-08-29。Namakoo の設計を検算できる形にする。
#
# ★★ 発端（Namakoo）:
#   「原本から先に全てのセルの位置情報と内容を内部的な表にできないか？
#     行の挿入は、今の (x,y) から後ろを (x+1,y) にずらして (x,y) に入れるだけだよな？」
#
# ★ この見立ては**宣言の言語としても、検算の物差しとしても正しい**。
#   基本操作はどれも「座標の写像 π」と「新しく書く座標の集合 W」で書ける。
#   ★ ただし**実行の手順としては足りない**（同日実測）── 行を 1 本挿しただけで
#     数式の中の座標・範囲の終端・結合・行の高さが全部付け替わる。
#     だから実行は LibreOffice に任せ、ここは**宣言と検算**だけを担う。
#     （自分で計算して書き込むと、検算が『自分の計算 vs 自分の書き込み』＝恒真になる）
#
# ★★ この検体が守る芯 ── 法則は **1 本**:
#     適用後の表 ＝ 適用前の表を写像で動かしたもの ＋ 宣言した書き込み
#   ③（それ以外の座標に何も無い）が在るので、**宣言していない変化は全部**捕まる。
#   op が増えても法則は変わらない ＝「消えたものは diff に出ない」への正面からの答え。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core import cellmap as cm  # noqa: E402


# --- ① 写像そのもの -------------------------------------------------------------------

@pytest.mark.parametrize("shift,idx,want", [
    (cm.identity(), 5, 5),
    (cm.insert_rows(3), 2, 2), (cm.insert_rows(3), 3, 4), (cm.insert_rows(3), 9, 10),
    (cm.insert_rows(3, 2), 3, 5),
    (cm.delete_rows(3), 2, 2), (cm.delete_rows(3), 3, None), (cm.delete_rows(3), 4, 3),
    (cm.delete_rows(3, 2), 5, 3), (cm.delete_rows(3, 2), 4, None),
    (cm.swap_rows(2, 5), 2, 5), (cm.swap_rows(2, 5), 5, 2), (cm.swap_rows(2, 5), 3, 3),
])
def test_the_shift_is_a_total_function(shift, idx, want):
    """★ 消える座標だけが None。それ以外は必ず行き先がある（穴を作らない）。"""
    assert shift.moved(idx) == want


def test_rows_and_columns_use_the_same_shape():
    """★ Namakoo の指摘した対称性: 行と列は同じ写像の軸違い。"""
    assert cm.insert_rows(3).map_cell(3, 7) == (4, 7)
    assert cm.insert_cols(3).map_cell(7, 3) == (7, 4)
    assert cm.swap_rows(2, 4).map_cell(2, 9) == (4, 9)
    assert cm.swap_cols(2, 4).map_cell(9, 2) == (9, 4)


def test_a_permutation_covers_sorting():
    """★ 並べ替えも座標の写像（データが決める置換）。"""
    sh = cm.permute_rows([3, 2, 4])      # 適用後 1,2,3 番目に来る「適用前の番号」
    assert (sh.moved(3), sh.moved(2), sh.moved(4)) == (1, 2, 3)


# --- ② 数式の中の座標も同じ写像で動く（実測に一致すること）------------------------------

@pytest.mark.parametrize("formula,shift,want", [
    # ★ 2026-08-29 に実機で観測した通りの結果になること
    ("=B3*C3", cm.insert_rows(3), "=B4*C4"),
    ("=SUM(D2:D4)", cm.insert_rows(3), "=SUM(D2:D5)"),      # 範囲の**終わりだけ**伸びる
    ("=B2-C2", cm.swap_cols(2, 3), "=C2-B2"),               # 08-27 の列入れ替えの実測と一致
    ("=$B$3+1", cm.insert_rows(2), "=$B$4+1"),              # $ は保つ
    ("=A1", cm.identity(), "=A1"),
    ("ただの文字", cm.insert_rows(2), "ただの文字"),
    (42, cm.insert_rows(2), 42),
])
def test_formula_references_move_with_the_same_map(formula, shift, want):
    """★★ 検算を**文字列一致**でやってはいけない理由がここ。
       `=B3*C3` → `=B4*C4` は**正しい変化**で、文字列で比べると毎回『壊れた』と言う。
       写像を通してから比べれば、正しい付け替えと本当の破壊を区別できる。"""
    assert cm.map_formula(formula, shift) == want


def test_a_reference_into_a_deleted_row_is_reported_not_guessed():
    """★ 消える行を指していた数式は**決めない**（#REF! になるのが正しいが、
       それを勝手に書くのは実行系の仕事 ── ここは『分からない』を返す）。"""
    assert cm.map_formula("=B3*2", cm.delete_rows(3)) is None


# --- ③ 法則 1 本で基本操作を検算する ---------------------------------------------------

def _map(rows, merges=(), start=1):
    c = cm.CellMap()
    for i, row in enumerate(rows, start=start):
        for j, v in enumerate(row, start=1):
            if v not in (None, ""):
                c.values[(i, j)] = v
    c.merges = tuple(merges)
    c.max_row = start + len(rows) - 1
    c.max_col = max((len(r) for r in rows), default=0)
    return c


BEFORE = [["品名", "単価", "数量"],
           ["ボルト", 100, 12],
           ["ナット", 200, 5],
           ["ワッシャー", 50, 30]]


def test_a_correct_row_insert_passes():
    before = _map(BEFORE)
    after = _map([BEFORE[0], BEFORE[1], ["新品", None, None], BEFORE[2], BEFORE[3]])
    v = cm.check_shift(before, after, cm.insert_rows(3), {(3, 1): "新品"})
    assert v.ok, v.reason


def test_an_off_by_one_insert_fails():
    """★ 1 つずれた位置に挿したら落ちること（この 2 日で一番多かった壊れ方）。"""
    before = _map(BEFORE)
    after = _map([BEFORE[0], BEFORE[1], ["新品", None, None], BEFORE[2], BEFORE[3]])
    v = cm.check_shift(before, after, cm.insert_rows(4), {(4, 1): "新品"})
    assert not v.ok and "はずです" in v.reason, v.reason


def test_a_correct_row_delete_passes():
    before = _map(BEFORE)
    after = _map([BEFORE[0], BEFORE[1], BEFORE[3]])
    assert cm.check_shift(before, after, cm.delete_rows(3), {}).ok


def test_deleting_the_wrong_row_fails():
    before = _map(BEFORE)
    after = _map([BEFORE[0], BEFORE[2], BEFORE[3]])      # ボルトを消してしまった
    v = cm.check_shift(before, after, cm.delete_rows(3), {})
    assert not v.ok, v.reason


def test_a_correct_swap_passes():
    before = _map(BEFORE)
    after = _map([BEFORE[0], BEFORE[2], BEFORE[1], BEFORE[3]])
    assert cm.check_shift(before, after, cm.swap_rows(2, 3), {}).ok


def test_a_one_cell_write_passes_and_a_column_wide_write_fails():
    """★★ この 2 日で一番悪かった事故（1 セルの依頼で列が全行書き換わって ✓）が、
       この法則 1 本で落ちること。"""
    before = _map(BEFORE)
    ok_after = _map([BEFORE[0], ["ボルト", 999, 12], BEFORE[2], BEFORE[3]])
    assert cm.check_shift(before, ok_after, cm.identity(), {(2, 2): 999}).ok
    bad_after = _map([BEFORE[0], ["ボルト", 999, 12], ["ナット", 999, 5], ["ワッシャー", 999, 30]])
    v = cm.check_shift(before, bad_after, cm.identity(), {(2, 2): 999})
    assert not v.ok, v.reason


# --- ★★ ③ 宣言していない所に書いていないか（負の被覆）---------------------------------

def test_writing_somewhere_nobody_declared_is_caught():
    """★ 法則の芯。op が増えても、この 1 行が全部の『余計な書き込み』を捕まえる。"""
    before = _map(BEFORE)
    after = _map(BEFORE + [[None, None, None, "こっそり"]])
    v = cm.check_shift(before, after, cm.identity(), {})
    assert not v.ok and "宣言していない所" in v.reason, v.reason


def test_a_cell_that_quietly_disappeared_is_caught():
    """★ 消えたものは diff に出ない ── 法則は『動いた先に在るはず』を要求するので出る。"""
    before = _map(BEFORE)
    after = _map([BEFORE[0], ["ボルト", None, 12], BEFORE[2], BEFORE[3]])
    v = cm.check_shift(before, after, cm.identity(), {})
    assert not v.ok, v.reason


# --- ④ 結合セルも同じ写像で動く ---------------------------------------------------------

def test_merges_move_with_the_shift():
    before = _map(BEFORE, merges=[(4, 1, 4, 3)])
    after = _map([BEFORE[0], BEFORE[1], ["新品", None, None], BEFORE[2], BEFORE[3]],
                  merges=[(5, 1, 5, 3)])
    assert cm.check_shift(before, after, cm.insert_rows(3), {(3, 1): "新品"}).ok


def test_a_merge_left_behind_is_caught():
    before = _map(BEFORE, merges=[(4, 1, 4, 3)])
    after = _map([BEFORE[0], BEFORE[1], ["新品", None, None], BEFORE[2], BEFORE[3]],
                  merges=[(4, 1, 4, 3)])          # ずれていない
    v = cm.check_shift(before, after, cm.insert_rows(3), {(3, 1): "新品"})
    assert not v.ok and "結合" in v.reason, v.reason


# --- ⑤ 実機（LibreOffice が本当にこの法則どおりに動くか）--------------------------------

@pytest.mark.local
def test_real_libreoffice_obeys_the_law_for_insert_and_swap(tmp_path):
    """★★ 一番大事な検体: **法則が実物に当たるか**。
       机上で正しい法則を作っても、LibreOffice が違うことをするなら意味が無い。
       数式・範囲参照・結合・行の高さを入れた表で通す。"""
    import subprocess
    from openpyxl.styles import PatternFill

    src = tmp_path / "a.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "在庫"
    ws.append(["品名", "単価", "数量", "金額"])
    for i, (n, u, q) in enumerate([("ボルト", 100, 12), ("ナット", 200, 5),
                                    ("ワッシャー", 50, 30)], start=2):
        ws.cell(i, 1, n), ws.cell(i, 2, u), ws.cell(i, 3, q)
        ws.cell(i, 4, f"=B{i}*C{i}")
    ws.cell(6, 1, "合計"), ws.cell(6, 4, "=SUM(D2:D4)")
    ws.merge_cells("A8:D8"), ws.cell(8, 1, "備考欄")
    ws.cell(3, 1).fill = PatternFill("solid", fgColor="FFF3D6")
    ws.row_dimensions[3].height = 40
    wb.save(src)

    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}
    p = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(src),
         "ボルトとナットの間に新品を作って", "--copy", "--timeout", "90"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=420, cwd=str(REPO), env=env)
    assert p.returncode == 0, p.stdout[-900:]
    out = src.with_name("a.out.xlsx")
    v = cm.check_shift(cm.read_cellmap(src, "在庫"), cm.read_cellmap(out, "在庫"),
                        cm.insert_rows(3), {(3, 1): "新品"})
    assert v.ok, f"法則が実物に当たらなかった: {v.reason}"


# --- 分かっている穴（隠さずに書いておく）-----------------------------------------------

def test_known_gaps_are_written_down():
    """★ まだ扱えていないもの。**黙って通さない**ために、ここに列挙しておく:
       ・別シートを指す参照（Sheet2.B3）と名前付き範囲は写像を通していない
       ・行の高さ・列幅は読んではいるが、まだ検算していない
       ・条件付き書式・入力規則・印刷範囲の座標も未対応
       ★ どれも『検算していない』のであって『大丈夫』ではない ── 使う側で断ること。
    """
    src = (REPO / "src" / "ailine_core" / "cellmap.py").read_text(encoding="utf-8")
    assert "row_heights" in src, "行の高さは読んでいる（検算はまだ）"
    # ★★ この検体が実際に穴を 1 つ暴いた（書いた直後に赤くなった）:
    #   初版は `=Sheet2.B3` を `=Sheet2.B4` に動かしていた。別のシートに行を挿した
    #   わけではないので、これは**静かな破壊**になる。
    #   ★ このシートの写像は、このシートの参照にだけ効く。
    assert cm.map_formula("=Sheet2.B3", cm.insert_rows(2)) == "=Sheet2.B3"
    assert cm.map_formula("=Sheet2!B3", cm.insert_rows(2)) == "=Sheet2!B3"
    assert cm.map_formula("=B3", cm.insert_rows(2)) == "=B4"        # 自分のシートは動く
