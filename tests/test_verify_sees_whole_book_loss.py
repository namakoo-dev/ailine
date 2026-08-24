# `ailine verify` が「冊まるごとの欠落」を見られるようにする（2026-08-24）。
#
# ★ 根（盲検の契約レビュー・俺が verify.py:410-431 で裏取り済み）:
#   `_verify_stack` は元側のファイル一覧を**出力自身の出所列**から作っている。
#   だからフォルダに在るのに積まれなかった冊は、元側にも現れない ──
#   **出力を出力自身と比べている**。しかも `cmd_stack` は `sources` を渡していない。
#   ★ これは「道具を信じる代わりに使う独立チェック」が、
#     一番肝心な「冊が丸ごと落ちた」を原理的に見られない、ということ。
#
#   実測（査定者の再現）: 3 冊のうち 1 冊が見出しの綴り違いで積まれず →
#     `ailine verify out.xlsx <folder>` → 行数 元3/出力3・Σ 元600/出力600・**exit 0**
#     （実際には 2 行 / ¥18,887 が消えている）
#
# 契約:
#   ① フォルダを渡された verify は、**フォルダの実ファイル**を元側の分母にする
#   ② 出所列に現れない冊があれば、それを不一致として名指しする
#   ③ 誤爆しない: 全冊が積まれていれば今までどおり通る

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core import verify as multifile_verify  # noqa: E402

needs_impl = pytest.mark.xfail(
    "missing_source" not in (multifile_verify._verify_stack.__doc__ or ""),
    reason="冊まるごとの欠落の検出 未実装（契約は凍結済み）", strict=True)


def _book(path, rows):
    wb = openpyxl.Workbook(); ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(path)


def _folder_with_three(tmp_path):
    f = tmp_path / "src"; f.mkdir()
    _book(f / "a.xlsx", [["商品", "金額"], ["x", 100], ["y", 200]])
    _book(f / "b.xlsx", [["商品", "金額"], ["z", 300]])
    _book(f / "c.xlsx", [["商品", "金額"], ["w", 400], ["v", 500]])
    return f


def _stacked(tmp_path, folder, names):
    """names の冊だけを積んだ出力（出所列つき）を手で作る。"""
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "金額", "元ファイル", "元行"])
    for name in names:
        src = openpyxl.load_workbook(folder / name).active
        for r in range(2, src.max_row + 1):
            ws.append([src.cell(r, 1).value, src.cell(r, 2).value, name, r])
    wb.save(out)
    return out


@needs_impl
def test_a_whole_missing_book_is_detected(tmp_path):
    """① ② c.xlsx が丸ごと積まれていない ── 見つけて名指しする。"""
    folder = _folder_with_three(tmp_path)
    out = _stacked(tmp_path, folder, ["a.xlsx", "b.xlsx"])   # ★ c.xlsx が無い
    res = multifile_verify._verify_stack(out, folder)
    assert res["mismatch"] is not None, "冊まるごとの欠落を見逃した（出力を出力と比べている）"
    kinds = {m.get("kind") for m in res.get("mismatches", [])}
    assert "missing_source" in kinds, f"欠落の種類が名指しされていない: {res.get('mismatches')}"
    names = " ".join(str(m) for m in res.get("mismatches", []))
    assert "c.xlsx" in names, f"どの冊が落ちたかを言っていない: {names}"


# ★ needs_impl を付けない ── 実装前は恒真（誰も止めないので当然通る）。
#   実装後に初めて意味を持つ「誤爆の柵」として置く。
def test_all_books_stacked_still_passes(tmp_path):
    """③ 誤爆防止: 全冊積まれていれば通る。"""
    folder = _folder_with_three(tmp_path)
    out = _stacked(tmp_path, folder, ["a.xlsx", "b.xlsx", "c.xlsx"])
    res = multifile_verify._verify_stack(out, folder)
    assert res["mismatch"] is None, res.get("mismatches")


@needs_impl
def test_row_count_denominator_comes_from_the_folder(tmp_path):
    """① 分母（元の行数）が**フォルダ側**から作られていること。"""
    folder = _folder_with_three(tmp_path)
    out = _stacked(tmp_path, folder, ["a.xlsx", "b.xlsx"])
    res = multifile_verify._verify_stack(out, folder)
    assert res["row_count"]["source"] == 5, \
        f"元の行数が出力から作られている（5 が正・実測 {res['row_count']}）"
    assert res["row_count"]["output"] == 3
