"""M2（run のフォルダ抽出）の追加検体 ── 凍結検体（tests/test_run_folder.py）が
   覆っていない側面を後から足したもの。

★ ここで測るのは「凍結検体が通ったこと」では足りない性質:
   E2 負の被覆（数だけ合っていても中身が違う偽実装を殺す）・陽性対照（正常な材料で
   ⚠ が 0 件＝測定器の空鳴りが無い）・E7 決定論・E8 原本無変更・E9 ファイル名の sanitize。
★ 7B は使わない（translate_task の monkeypatch）── 測るのは配管であって翻訳の質ではない。
"""
import hashlib
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402
from ailine_core import extract_multi  # noqa: E402

HDRS = ["注文ID", "取引先", "金額"]
_EXTRACT_40000 = [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 40000}}]
TASK = "金額が40000以上の行を抜き出して"


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r) + [None] * (len(headers) - len(r)))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _mock_translation(monkeypatch, plan):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"plan": plan})


def _run(folder, task, *extra, capsys=None):
    rc = ailine.main(["run", str(folder), task, *extra])
    out = capsys.readouterr().out if capsys else ""
    return rc, out


def _data_rows(path):
    ws = openpyxl.load_workbook(path).active
    return [[c.value for c in row] for row in ws.iter_rows(min_row=2)]


def test_e2_matched_row_carries_its_real_cell_values_not_just_a_count(tmp_path, monkeypatch, capsys):
    """★ E2（負の被覆）: 行数だけ合っている偽実装を殺す ── 一致した行の**中身**
       （目印の注文ID と金額）が出力ブックの実セルに在ることを見る。
       件数の照合だけでは『別の行を1行書いた』実装が通ってしまう。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-6", "甲", 100), ("J-7", "乙", 99000), ("J-8", "丙", 39999)])
    rc, out = _run(folder, TASK, capsys=capsys)
    assert rc == 0, out
    rows = _data_rows(next(iter(tmp_path.glob("*.xlsx"))))
    assert len(rows) == 1, rows
    assert rows[0][0] == "J-7" and rows[0][2] == 99000, f"中身が別の行: {rows[0]}"
    assert rows[0][2] != "99000", "数値が文字列化されている（型の保存が壊れている）"
    assert rows[0][3] == "a.xlsx" and rows[0][4] == 3, f"出所列が元の位置を指していない: {rows[0]}"


def test_positive_control_clean_folder_raises_no_warning(tmp_path, monkeypatch, capsys):
    """★ 陽性対照/ノイズ床: 合計行も欠けも無い正常な3冊で ⚠ が 1 件も出ないこと。
       ⚠ が常時鳴る測定器では『⚠ が出た』に情報が無い。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("J-2", "乙", 60000)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 45000)])
    _book(folder / "c.xlsx", HDRS, [("J-4", "丁", 70000)])
    rc, out = _run(folder, TASK, capsys=capsys)
    assert rc == 0, out
    assert out.count("⚠") == 0, f"正常な材料で ⚠ が鳴った:\n{out}"
    assert len(_data_rows(next(iter(tmp_path.glob("*.xlsx"))))) == 4


def test_e7_second_run_reproduces_the_same_book_cell_for_cell(tmp_path, monkeypatch, capsys):
    """★ E7 決定論: 同じフォルダ・同じ依頼を2回走らせたら、出力の中身は完全に同じ
       （行の順序も出所列も）。2回目は自分の前回出力を印で認めて黙って作り直す。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("J-2", "乙", 100)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 45000), ("J-4", "丁", 88000)])
    rc1, out1 = _run(folder, TASK, capsys=capsys)
    assert rc1 == 0, out1
    produced = next(iter(tmp_path.glob("*.xlsx")))
    first = _data_rows(produced)
    header_first = [c.value for c in openpyxl.load_workbook(produced).active[1]]

    rc2, out2 = _run(folder, TASK, capsys=capsys)
    assert rc2 == 0, out2
    assert list(tmp_path.glob("*.xlsx")) == [produced], "2回目が別名のブックを作った"
    assert _data_rows(produced) == first, "2回の実行で中身が違う（決定論でない）"
    assert [c.value for c in openpyxl.load_workbook(produced).active[1]] == header_first
    assert "作り直しました" in out2, f"前回出力を作り直した旨の開示が無い:\n{out2}"


def test_e8_source_books_are_byte_identical_after_a_successful_run(tmp_path, monkeypatch, capsys):
    """★ E8 原本無変更: 成功した実行の後も、元フォルダの全ブックが1バイトも変わらない
       （フォルダ経路は読むだけ ── だから undo も要らない、が製品の約束）。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("合計", None, 50000)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 45000)])
    before = {p.name: (p.stat().st_size, hashlib.sha256(p.read_bytes()).hexdigest())
              for p in sorted(folder.glob("*.xlsx"))}
    rc, out = _run(folder, TASK, capsys=capsys)
    assert rc == 0, out
    after = {p.name: (p.stat().st_size, hashlib.sha256(p.read_bytes()).hexdigest())
             for p in sorted(folder.glob("*.xlsx"))}
    assert before == after, "原本が変わった"


@pytest.mark.parametrize("kind", ["行を捏造", "値を改竄"])
def test_postcondition_actually_fires_when_the_write_side_is_mutated(tmp_path, monkeypatch,
                                                                     capsys, kind):
    """★ 番人への変異試験（負の被覆）: 書く側（extract_multi）に嘘を注入したら、
       独立読み（xml_readback）の事後条件が本当に鳴るか。鳴らない番人は無いのと同じ。
       ★ 破れた時は出力ブックを**書かない**（tmp のまま捨てる）ことも同時に測る。"""
    import dataclasses
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("J-2", "乙", 100)])
    real = extract_multi.evaluate_and_extract

    def mutant(*args, **kwargs):
        r = real(*args, **kwargs)
        if r.status != "取れた" or not r.rows:
            return r
        if kind == "行を捏造":
            return dataclasses.replace(r, rows=r.rows + [(["X-9", "偽", 1], 99)],
                                       rows_matched=r.rows_matched + 1)
        vals, src_row = r.rows[0]
        return dataclasses.replace(r, rows=[([vals[0], vals[1], 12345], src_row)] + r.rows[1:])

    monkeypatch.setattr(extract_multi, "evaluate_and_extract", mutant)
    rc, out = _run(folder, TASK, capsys=capsys)
    assert rc == 1, f"変異を注入したのに合格した:\n{out}"
    assert "事後条件が破れた" in out and "元" in out and "出力" in out, out
    assert list(tmp_path.glob("*.xlsx")) == [], "事後条件が破れたのに出力を書いた"


# --- E9: 出力ファイル名の sanitize（単体） -------------------------------------

@pytest.mark.parametrize("bad", ['a<b', 'a>b', 'a:b', 'a"b', 'a/b', 'a\\b', 'a|b', 'a?b', 'a*b',
                                  'a\x01b', 'a\x1fb'])
def test_e9_forbidden_characters_are_replaced(bad):
    got = extract_multi.sanitize_filename(bad)
    assert not set(got) & set('<>:"/\\|?*'), got
    assert all(ord(ch) >= 32 for ch in got), repr(got)


def test_e9_reserved_device_names_do_not_round_trip():
    for name in ("CON", "con", "NUL", "COM1", "LPT9"):
        assert extract_multi.sanitize_filename(name) != name, name


def test_e9_trailing_dots_and_spaces_are_stripped():
    for raw in ("報告書. ", "報告書...", "報告書   "):
        got = extract_multi.sanitize_filename(raw)
        assert got and not got.endswith(".") and not got.endswith(" "), repr(got)


def test_e9_truncation_collision_is_separated_by_a_hash():
    """★ review3#1/#5 の直し: 切り詰めが起きたら**常に**元の（切り詰め前の）名前 全体 の
       sha256 で分ける ── taken の受け渡しに頼らない（配線されない対策コードを持たない）。
       切り詰め前なら同じ幹に潰れていた2つの長い名前が、別のファイル名になること。"""
    long_a = "あ" * 150 + "_甲"
    long_b = "あ" * 150 + "_乙"
    a = extract_multi.sanitize_filename(long_a)
    b = extract_multi.sanitize_filename(long_b)
    assert a != b, f"切り詰めで別名が同じ幹に潰れた: {a} == {b}"
    for got in (a, b):
        suffix = got.rsplit("_", 1)[-1]
        assert len(suffix) == 6 and all(c in "0123456789abcdef" for c in suffix), got
    short = extract_multi.sanitize_filename("短い名前")
    assert short == "短い名前", "切り詰めが起きていないのにハッシュを付けた"
