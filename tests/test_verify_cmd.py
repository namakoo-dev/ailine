"""`ailine verify <out.xlsx> <srcfolder>`（検算の単独再実行）の検体。
   ★ 実装前に書いた赤い検体。信用の条件⑥「検算側を単独で再実行できること」──
   需要ノートの言葉で「信じる対象が道具から検算に移る」。

   契約の要点:
   - stack の出力ブックと元フォルダだけを引数に、検算（行数照合・数値列 Σ 照合）を再実行
   - 出所列（元ファイル/元行）を使って出力の各行を原本に引き当てる
   - 合格: 両側の数字を並べて exit 0 ／ 不一致: どの列がいくつ対いくつ、まで名指しで exit 5
   - どちらのファイルも変更しない（読むだけ）"""
import hashlib
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
HDRS = ["注文ID", "取引先", "金額"]


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r) + [None] * (len(headers) - len(r)))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _run(cmd, *args):
    return subprocess.run(
        [sys.executable, str(REPO / "ailine.py"), cmd, *map(str, args)],
        capture_output=True, text=True, timeout=180, encoding="utf-8")


def _made(tmp_path):
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100), ("J-2", "乙", 200)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 300)])
    out = tmp_path / "out.xlsx"
    p = _run("stack", folder, "--out", out)
    assert p.returncode == 0, "前提の stack が失敗:\n" + p.stderr[-500:]
    return folder, out


def test_verify_passes_on_untouched_output_with_both_sides(tmp_path):
    folder, out = _made(tmp_path)
    p = _run("verify", out, folder)
    assert p.returncode == 0, p.stdout + p.stderr[-500:]
    assert p.stdout.count("600") >= 2, f"Σ の両側表示が無い:\n{p.stdout}"


def test_verify_names_a_tampered_value_and_exits_5(tmp_path):
    """出力の金額 1 セルを 100→999 に改竄 → 列名と両側の数字つきで不合格・exit 5。"""
    folder, out = _made(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[2].value == 100:
            row[2].value = 999
            break
    wb.save(out)
    p = _run("verify", out, folder)
    assert p.returncode == 5, f"exit={p.returncode}\n{p.stdout}"
    assert "金額" in p.stdout
    assert "600" in p.stdout and "1499" in p.stdout, f"両側の数字が無い:\n{p.stdout}"


def test_verify_names_a_deleted_row_and_exits_5(tmp_path):
    """出力からデータ行を 1 行削除 → 行数の不一致が両側の数字つきで出る・exit 5。"""
    folder, out = _made(tmp_path)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    ws.delete_rows(2)
    wb.save(out)
    p = _run("verify", out, folder)
    assert p.returncode == 5
    assert "3" in p.stdout and "2" in p.stdout, f"行数の両側が無い:\n{p.stdout}"


def test_verify_is_read_only_on_both_sides(tmp_path):
    folder, out = _made(tmp_path)
    before = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
              for p in [*folder.glob("*.xlsx"), out]}
    assert _run("verify", out, folder).returncode == 0
    after = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
             for p in [*folder.glob("*.xlsx"), out]}
    assert before == after


def test_verify_passes_on_output_whose_totals_were_excluded(tmp_path):
    """★ 実機の敵対検分（2026-08-21 06:1x）で踏んだ穴の凍結: 合計行を正しく除外した
       正当な出力に対して、verify が除外の意味論を再現できず 元7/出力4 で誤 FAIL した。
       verify は stack と同じ除外規則（total_row）を XML 直読みの側でも再現すること。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 600), ("J-2", "乙", 400),
           ("小計", None, 1000), ("消費税", None, 100), ("合計", None, 1100)])
    out = tmp_path / "out.xlsx"
    p = _run("stack", folder, "--out", out)
    assert p.returncode == 0, p.stdout + p.stderr[-500:]
    v = _run("verify", out, folder)
    assert v.returncode == 0, f"正当な出力で verify が落ちた:\n{v.stdout}"
    assert v.stdout.count("1100") >= 2, f"Σ の両側（税込 1100）が無い:\n{v.stdout}"


def test_verify_agrees_with_stack_when_dates_and_reorder_present(tmp_path):
    """★ 予測①の的中を凍結（2026-08-21 06:1x 実機 probe の再現）: XML 直読みは日付の
       シリアル値を数値と見るため、数値列の引き当てが stack(openpyxl=datetime は非数値)と
       食い違い、除外規則が別の列で走って 元7/出力4 の誤 FAIL になった。
       両側の読み手は『日付書式のセルは数値列の候補にしない』で一致すること。"""
    import datetime
    folder = tmp_path / "src"
    _book(folder / "inv1.xlsx", ["注文ID", "受注日", "金額"],
          [("J-1", datetime.date(2026, 8, 1), 600), ("J-2", datetime.date(2026, 8, 3), 400),
           ("小計", None, 1000), ("消費税", None, 100), ("合計", None, 1100)])
    _book(folder / "inv2.xlsx", ["金額", "注文ID", "受注日"],
          [(300, "J-3", datetime.date(2026, 8, 5)), (999, "合計", None)])
    out = tmp_path / "out.xlsx"
    p = _run("stack", folder, "--out", out)
    assert p.returncode == 0, p.stdout + p.stderr[-500:]
    assert "Σ受注日" not in p.stdout, "日付列を合計している"
    v = _run("verify", out, folder)
    assert v.returncode == 0, f"正当な出力で verify が落ちた:\n{v.stdout}"
    assert v.stdout.count("1400") >= 2, f"Σ金額の両側（1400）が無い:\n{v.stdout}"
    assert "Σ受注日" not in v.stdout


def test_verify_reads_the_same_named_sheet_as_stack(tmp_path):
    """★ P2（architect 致命5 の後段・出荷済みの食い違い）: 基準名のシートが 2 枚目にある
       ソースでは、stack は同名シートを読むのに verify は常に 1 枚目を読んで偽 ⚠ を出す。
       verify は出力ブックのシート名（= 基準のシート名）で各ソースを引き当てること。"""
    folder = tmp_path / "src"
    base = folder / "a.xlsx"
    _book(base, HDRS, [("J-1", "甲", 100)])
    wb = openpyxl.load_workbook(base)
    wb.active.title = "明細"
    wb.save(base)
    b = folder / "b.xlsx"
    wb = openpyxl.Workbook()
    front = wb.active
    front.title = "表紙"
    front.append(["メモ"]); front.append(["これは表紙"])
    ws = wb.create_sheet("明細")
    ws.append(HDRS); ws.append(["J-2", "乙", 200])
    wb.save(b)
    out = tmp_path / "out.xlsx"
    p = _run("stack", folder, "--out", out)
    assert p.returncode == 0, p.stdout
    ws2 = openpyxl.load_workbook(out).active
    assert ws2.max_row - 1 == 2, f"明細 2 行のはず: {ws2.max_row - 1}"
    v = _run("verify", out, folder)
    assert v.returncode == 0, f"正当な出力に偽 ⚠（verify が別シートを読んでいる）:\n{v.stdout}"


def test_verify_refuses_unmarked_book_instead_of_passing_empty(tmp_path):
    """★ E13/致命3 の半分: ailine の印が無いブックには「検算できません」と言って
       不合格でも合格でもない出口（exit 4）── 0 件照合で合格を名乗らない（空虚な合格の禁止）。"""
    book = tmp_path / "someones.xlsx"
    _book(book, HDRS + ["元ファイル", "元行"], [("X-1", "人の表", 999, "a.xlsx", 2)])
    p = _run("verify", book, tmp_path)
    assert p.returncode == 4, f"exit={p.returncode}\n{p.stdout}"
    assert "検算できません" in p.stdout or "印" in p.stdout


def test_verify_catches_attribution_swap_even_when_sums_match(tmp_path):
    """★ review3#3 major: verify は集計（行数・Σ）しか照合せず、出所の帰属（どの行が
       どのファイルの何行目か）を検算しない ── 値を入れ替えても Σ が同じなら合格していた。
       契約: 各出力行の値は 元ファイル/元行 が指す原本の行と一致すること（exit 5）。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    _book(folder / "b.xlsx", HDRS, [("J-2", "乙", 200)])
    out = tmp_path / "out.xlsx"
    p = _run("stack", folder, "--out", out)
    assert p.returncode == 0
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    v1, v2 = ws.cell(row=2, column=3).value, ws.cell(row=3, column=3).value
    ws.cell(row=2, column=3).value = v2   # 金額だけ入れ替え（Σ 不変・帰属が嘘になる）
    ws.cell(row=3, column=3).value = v1
    wb.save(out)
    p = _run("verify", out, folder)
    assert p.returncode == 5, f"Σ が同じ帰属の嘘を見逃した (exit={p.returncode}):\n{p.stdout}"
    assert "a.xlsx" in p.stdout or "元行" in p.stdout or "2" in p.stdout, "どの行かの名指しが無い"


def test_verify_catches_attribution_swap_in_extract_output_even_when_sums_match(
        tmp_path, monkeypatch):
    """★ jisaku-review 4戦目 F3（minor）: 上の stack 版の抽出（kind=extract）版。
       verify_extract 側にも同型の検体が無かった ── フォルダ抽出（`ailine run`）の
       出力でも、Σ 不変の値入れ替えを帰属検算が拾うこと（exit 5）。
       ★ run の翻訳は 7B を使わず monkeypatch（他の run 検体と同じ作法・
       test_run_folder.py の _mock_translation と同じ線）。"""
    import ailine  # ★ conftest.py が sys.path を通す（他の subprocess ベースの検体と混在可）
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "EXTRACT",
                                   "args": {"column": "金額", "cmp": "gte", "value": 100}}]})
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100), ("J-2", "乙", 200)])
    rc = ailine.main(["run", str(folder), "金額が100以上の行を抜き出して"])
    assert rc == 0
    out = next(tmp_path.glob("*.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    v1, v2 = ws.cell(row=2, column=3).value, ws.cell(row=3, column=3).value
    ws.cell(row=2, column=3).value = v2   # 金額だけ入れ替え（Σ 不変・帰属が嘘になる）
    ws.cell(row=3, column=3).value = v1
    wb.save(out)
    p = _run("verify", out, folder)
    assert p.returncode == 5, f"Σ が同じ帰属の嘘を見逃した (exit={p.returncode}):\n{p.stdout}"
    assert "a.xlsx" in p.stdout or "元行" in p.stdout or "2" in p.stdout, "どの行かの名指しが無い"
