"""ailine_core/csv_quarantine.py の機械の層を検体化する。

★ 分類（classify_column/detect_encoding/control_char_cells/overlong_cells の判定結果）
  の期待値は tests/test_csv_truth_table.py が凍結済みの正本 ── ここでは書き写さない。
  ここが担うのは parser（物理行範囲・重複ヘッダ・列数不一致・閉じない引用）・
  writer（数式化け防止の証明・制御文字隔離）・compare（3計数・型込み等値）・
  sha256 の 4 層。
"""
import datetime

import openpyxl
import pytest

from ailine_core import csv_quarantine as cq

# --- parse_csv: 物理行範囲 -------------------------------------------------


def test_parse_csv_basic_records_without_header():
    result = cq.parse_csv("a,b\nc,d\n", has_header=False)
    assert result.header == []
    assert [r.cells for r in result.records] == [["a", "b"], ["c", "d"]]
    assert (result.records[0].first_line, result.records[0].last_line) == (1, 1)
    assert (result.records[1].first_line, result.records[1].last_line) == (2, 2)


def test_parse_csv_quoted_newline_spans_physical_lines():
    text = 'h1,h2\n"a\nb",c\nd,e\n'
    result = cq.parse_csv(text, has_header=True)
    assert result.header == ["h1", "h2"]
    rec0, rec1 = result.records
    assert rec0.cells == ["a\nb", "c"]
    assert (rec0.first_line, rec0.last_line) == (2, 3)
    assert rec1.cells == ["d", "e"]
    assert (rec1.first_line, rec1.last_line) == (4, 4)


def test_parse_csv_unterminated_quote_is_flagged():
    # 引用が最後まで閉じない ── csv モジュール自体は例外を上げず読み切ってしまう。
    text = 'a,"b,c\nd,e\n'
    result = cq.parse_csv(text, has_header=False)
    assert result.unterminated_quote_records == [0]


def test_parse_csv_closed_quotes_are_not_flagged():
    result = cq.parse_csv('a,"b,c"\nd,e\n', has_header=False)
    assert result.unterminated_quote_records == []


# --- parse_csv: 重複ヘッダの機械リネーム -----------------------------------


def test_duplicate_headers_are_renamed_and_disclosed():
    result = cq.parse_csv("名前,名前,名前\n1,2,3\n", has_header=True)
    assert result.header == ["名前", "名前__dup2", "名前__dup3"]
    assert [(r.index, r.original, r.renamed) for r in result.header_renames] == [
        (1, "名前", "名前__dup2"),
        (2, "名前", "名前__dup3"),
    ]


def test_no_duplicate_headers_means_no_renames():
    result = cq.parse_csv("a,b,c\n1,2,3\n", has_header=True)
    assert result.header == ["a", "b", "c"]
    assert result.header_renames == []


# --- parse_csv: 列数不一致の名指し -----------------------------------------


def test_column_count_mismatch_is_recorded_by_raw_index():
    # raw_index は「ヘッダ込みの物理レコード列」の 0 起点 index（records の index とは
    # 空間が違う ── ParseResult の docstring どおり）。
    result = cq.parse_csv("a,b\n1,2\n3,4,5\n", has_header=True)
    assert result.column_count_mismatches == [(2, 2, 3)]


def test_no_mismatch_when_all_rows_match_header_width():
    result = cq.parse_csv("a,b\n1,2\n3,4\n", has_header=True)
    assert result.column_count_mismatches == []


# --- sha256 は渡した値をそのまま運ぶだけ -----------------------------------


def test_sha256_bytes_matches_hashlib():
    import hashlib

    raw = "テスト".encode("utf-8")
    assert cq.sha256_bytes(raw) == hashlib.sha256(raw).hexdigest()


def test_parse_csv_carries_source_sha256_through():
    raw = b"a,b\n1,2\n"
    digest = cq.sha256_bytes(raw)
    result = cq.parse_csv(raw.decode("utf-8"), has_header=True, source_sha256=digest)
    assert result.source_sha256 == digest


# --- build_columns: classify_column に渡す列の組み立て ---------------------


def test_build_columns_pads_short_rows_with_empty_string():
    result = cq.parse_csv("a,b,c\n1,2\n3,4,5\n", has_header=True)
    columns = cq.build_columns(result.header, result.records)
    assert columns == [["1", "3"], ["2", "4"], ["", "5"]]


# --- write_quarantined_xlsx: 数式化けの防止（実測の罠の核心）---------------


def test_string_column_survives_formula_looking_value(tmp_path):
    """`=SUM(1,2)` を string 分類で書いたら、data_type は 's' のまま・値も変わらない
    ことを openpyxl の読み戻しで自分のテストとして証明する（brief の要求どおり）。"""
    parsed = cq.ParseResult(
        header=["式"],
        header_renames=[],
        records=[cq.ParsedRecord(cells=["=SUM(1,2)"], first_line=2, last_line=2)],
        column_count_mismatches=[],
        unterminated_quote_records=[],
        source_sha256=None,
    )
    classifications = [cq.ColumnClassification(kind="string", reasons=["formula_head"], warn=False)]
    out_path = tmp_path / "formula.xlsx"
    write_result = cq.write_quarantined_xlsx(parsed, classifications, out_path)

    reopened = openpyxl.load_workbook(out_path)
    cell = reopened.active.cell(row=2, column=1)
    assert cell.data_type == "s"
    assert cell.value == "=SUM(1,2)"
    assert write_result.declared[(2, 1)] == "=SUM(1,2)"


# --- write_quarantined_xlsx: 制御文字の隔離 ---------------------------------


def test_control_char_is_stripped_and_removal_is_recorded(tmp_path):
    parsed = cq.ParseResult(
        header=["x"],
        header_renames=[],
        records=[cq.ParsedRecord(cells=["a\x01b"], first_line=2, last_line=2)],
        column_count_mismatches=[],
        unterminated_quote_records=[],
        source_sha256=None,
    )
    classifications = [cq.ColumnClassification(kind="string", reasons=[], warn=False)]
    out_path = tmp_path / "control.xlsx"
    # ★ 実測の罠: 素で書けば IllegalCharacterError で落ちる。落ちないこと自体が
    #   除去が効いている証拠。
    write_result = cq.write_quarantined_xlsx(parsed, classifications, out_path)

    assert write_result.removed_control_chars == [(2, 1, "U+0001")]
    assert write_result.declared[(2, 1)] == "ab"
    reopened = openpyxl.load_workbook(out_path)
    assert reopened.active.cell(row=2, column=1).value == "ab"


# --- write_quarantined_xlsx + compare_against_quarantine: 全セル等値照合 ---


def test_number_and_date_columns_round_trip_cleanly(tmp_path):
    parsed = cq.ParseResult(
        header=["郵便", "金額", "日付"],
        header_renames=[],
        records=[
            cq.ParsedRecord(cells=["1000001", "1,234", "2026/1/2"], first_line=2, last_line=2),
            cq.ParsedRecord(cells=["1000002", "2,345", "2026-01-03"], first_line=3, last_line=3),
        ],
        column_count_mismatches=[],
        unterminated_quote_records=[],
        source_sha256=None,
    )
    columns = cq.build_columns(parsed.header, parsed.records)
    classifications = [cq.classify_column(c) for c in columns]
    assert [c.kind for c in classifications] == ["number", "number", "date"]

    out_path = tmp_path / "clean.xlsx"
    write_result = cq.write_quarantined_xlsx(parsed, classifications, out_path)
    compare = cq.compare_against_quarantine(write_result.declared, out_path)

    assert compare.ok is True
    assert compare.missing == []
    assert compare.mismatched == []
    assert compare.surplus == []
    # 日付は date-style で書かれ、xml_readback から date として読める。
    assert write_result.declared[(2, 3)] == datetime.date(2026, 1, 2)


def test_undecidable_column_is_written_as_preserved_string(tmp_path):
    parsed = cq.ParseResult(
        header=["謎"],
        header_renames=[],
        records=[cq.ParsedRecord(cells=["2026/2/29"], first_line=2, last_line=2)],
        column_count_mismatches=[],
        unterminated_quote_records=[],
        source_sha256=None,
    )
    classification = cq.classify_column(["2026/2/29"])
    assert classification.kind == "undecidable"
    out_path = tmp_path / "undecidable.xlsx"
    write_result = cq.write_quarantined_xlsx(parsed, [classification], out_path)

    # ⚠ は判定に付くのであって、セルの書き方は string と同じ（壊さず持つ）。
    assert write_result.declared[(2, 1)] == "2026/2/29"
    reopened = openpyxl.load_workbook(out_path)
    cell = reopened.active.cell(row=2, column=1)
    assert cell.data_type == "s"
    assert cell.value == "2026/2/29"


# --- compare_against_quarantine: 3 計数の分離 -------------------------------


def _make_actual_grid_xlsx(path, cells: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    for (row, col), value in cells.items():
        ws.cell(row=row, column=col, value=value)
    wb.save(path)


def test_compare_counts_missing_mismatched_and_surplus_separately(tmp_path):
    out_path = tmp_path / "broken.xlsx"
    _make_actual_grid_xlsx(
        out_path,
        {
            (1, 1): "ok",       # declared と一致
            (1, 2): "actual",   # declared と型・値が違う（mismatched）
            (1, 3): "surplus",  # declared に無い（surplus）
            # (2, 1) は declared にあるが実際には書かれていない（missing）
        },
    )
    declared = {
        (1, 1): "ok",
        (1, 2): "declared",
        (2, 1): "gone",
    }
    compare = cq.compare_against_quarantine(declared, out_path)

    assert compare.ok is False
    assert compare.missing == [(2, 1)]
    assert compare.mismatched == [(1, 2, "declared", "actual")]
    assert compare.surplus == [(1, 3, "surplus")]


def test_compare_treats_number_and_numeric_string_as_different_kinds(tmp_path):
    """★ 型込み等値: 宣言側が文字列 "5"・実測側が数値 5 なら一致とみなさない
    （TOLERANCE 不使用の「型込み」の意味 ── 値が同じでも種別が違えば不一致）。"""
    out_path = tmp_path / "typed.xlsx"
    _make_actual_grid_xlsx(out_path, {(1, 1): 5})
    compare = cq.compare_against_quarantine({(1, 1): "5"}, out_path)
    assert compare.mismatched == [(1, 1, "5", 5)]


def test_compare_treats_int_and_float_of_equal_value_as_agreeing(tmp_path):
    # 数値どうしは int/float の表現差を許す（5 と 5.0 は同じ数値）。
    out_path = tmp_path / "numeric.xlsx"
    _make_actual_grid_xlsx(out_path, {(1, 1): 5})
    compare = cq.compare_against_quarantine({(1, 1): 5.0}, out_path)
    assert compare.ok is True


# --- 文字コード検出: 対象外宣言の例外（真理値表がカバーしない失敗系）-------


def test_utf16_bom_is_declared_out_of_scope():
    with pytest.raises(cq.UndecidableEncodingError):
        cq.detect_encoding("a".encode("utf-16"))


def test_undecodable_bytes_raise_explicit_error():
    # utf-8 でも cp932 でも復号できないバイト列（実測: 両方 UnicodeDecodeError）。
    raw = b"\x80\x81"
    with pytest.raises(cq.UndecidableEncodingError):
        cq.detect_encoding(raw)


# --- operator9 CONFUSING の修正（2026-08-23）: 桁数警告のオオカミ少年化 -------------
# 実測: 数量 5/12/3/100・単価 1000/850/15000/50 のような普通の数値列で毎回鳴っていた。
# 絞り: 桁固定の期待を持つのは「コード様の見出し」（番号/コード/品番/No/ID/郵便/TEL 等）
# の列だけ。見出しにその手掛かりが無い数値列では桁数の混在は自然 ── 鳴らさない。

def test_digit_variance_silent_on_ordinary_numeric_columns():
    from ailine_core import csv_quarantine as cq
    cols = [["5", "12", "3", "100"], ["1000", "850", "15000", "50"], ["5000", "10200", "45000", "5000"]]
    cls = [cq.classify_column(c) for c in cols]
    findings = cq.detect_excel_damage(cols, cls, header=["数量", "単価", "金額"])
    assert [f for f in findings if "桁数" in f] == [], f"普通の数値列で鳴った: {findings}"


def test_digit_variance_fires_on_code_like_header_with_varying_lengths():
    from ailine_core import csv_quarantine as cq
    cols = [["1234567", "234567", "1234567"]]   # 郵便 7 桁のうち 1 件だけ 6 桁 = 0 落ちの痕跡
    cls = [cq.classify_column(c) for c in cols]
    findings = cq.detect_excel_damage(cols, cls, header=["郵便番号"])
    assert any("桁数" in f for f in findings), f"コード様の列で鳴らない: {findings}"
