"""★ 単位E の電池: 対象スロットの出所（①照合できた/②無言/③矛盾）の番人。

ここで守るのは、実装より先に**凍結した2本のバー**（ブリーフの事前登録）:

  バー1（真陽性）: 依頼の語と解決値が矛盾する敵対例で **`✓` が出ないこと**
    1. 「見出し」を太字に  → `col:数量*単価`（前段が作った新規列）
    2. 「金額」を降順に    → シート『金額』との衝突で1枚目へ後退（単位D の検体）
    3. 「税込み合計」      → ラベル『合計』・率は未確定（査定B の致命1）
  バー2（偽陽性・★こちらが本番）: LLM が選んだが**正しい**対象で **`✓` が消えないこと**
    1. 「小計の列を太くして」→ 実列名『小計金額』（部分一致で正しい）
    2. 既存のゴールデン/テストの合格線が1本も下がらない（＝この repo の残り全部が証拠）

★ 対照ケース（単位C/D と同じ作法）: **同じブック・同じ op・同じ解決値**で、依頼文だけを
   変えて①②③を撃ち分ける3本を凍結する（下の TestContrast）。どれか1本でも隣の段階へ
   ずれたら赤くなる ―― 「定義をいじれば通る」への歯止め。

★ ollama も LibreOffice も使わない（translate_task / basrun_apply は monkeypatch）。
"""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402
from ailine_core.subject import (  # noqa: E402
    CONTRADICTED, MATCHED, UNSPOKEN, Consumed, Slot, classify_slots, name_matches_task,
)

from _run_argv import run_argv  # noqa: E402

VERIFIED = "は機械検証済みの内容です"


def _isolate(monkeypatch, tmp_path):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", tmp_path / "run.lock")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _bold_book(tmp_path):
    """見出しは既に太字（fake basrun_apply が何もしなくても check_bold の row:1 は通る）。"""
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    for c in (1, 2):
        wb.active.cell(row=1, column=c).font = Font(bold=True)
    wb.save(p)
    return p


def _fixed(monkeypatch, translated, apply_fn=None):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: translated)
    monkeypatch.setattr(ailine, "basrun_apply", apply_fn or (
        lambda out_book, code, workdir, helper_files=(), timeout=None: (True, None, "ok")))


def _inject_formula_cache(path: Path, updates: dict) -> None:
    """保存済み xlsx の sheet1.xml で、<f>式</f> の直後にある空の <v></v> をキャッシュ値で
       埋める（openpyxl は数式を計算しないため、二層検証を模擬するには直接 XML を触るしかない）。
       ★ tests/test_golden_postcondition.py の同名ヘルパと同じ手口（そちらが原典）。"""
    member = "xl/worksheets/sheet1.xml"
    with zipfile.ZipFile(path) as z:
        xml = z.read(member).decode("utf-8")
        others = {n: z.read(n) for n in z.namelist() if n != member}
    for ref, val in updates.items():
        # ★ lxml 有無で openpyxl の直列化が変わる（<v></v> 枠の有無・2026-08-21 CI 実測）。
        #   test_golden_postcondition.py の注入ヘルパと同じ両形対応（無ければ挿す）。
        pattern = re.compile(rf'(<c r="{ref}"[^>]*><f>[^<]*</f>)(<v/>|<v>[^<]*</v>)?')
        m = pattern.search(xml)
        assert m, f"{ref} の <f> セルが見つからない"
        xml = xml[:m.start()] + m.group(1) + f"<v>{val}</v>" + xml[m.end():]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, data in others.items():
            z.writestr(n, data)
        z.writestr(member, xml)


def _append_total_apply(total: int):
    """APPEND_TOTAL の代役: 実 codegen と同じ『挿入耐性 SUM』式＋ラベルを書き、
       キャッシュ値も注入する（事後条件を本物どおり pass させるため）。"""
    def _apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value="合計")
        ws.cell(row=row, column=2, value="=SUM(B2:INDEX(B:B,ROW()-1))")
        wb.save(out_book)
        _inject_formula_cache(out_book, {f"B{row}": str(total)})
        return True, None, "ok"
    return _apply


def _bold_whole_column(out_book, code, workdir, helper_files=(), timeout=None):
    """BOLD col:金額 を実際に効かせる代役（列全体を太字にする）。"""
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value not in (None, ""):
            ws.cell(row=r, column=2).font = Font(bold=True)
    wb.save(out_book)
    return True, None, "ok"


# ===========================================================================
# ★ 凍結した照合の定義（純関数レベル）
# ===========================================================================

class TestFrozenMatchingRule:
    """定義: (i) 実在名が依頼文に現れる or (ii) 長さ2以上の断片が依頼文にあり、かつ
       **他のどの実在名にも現れない**（曖昧な断片は証拠にしない）。"""

    def test_exact_substring_matches(self):
        assert name_matches_task("金額", "金額で降順に並べ替えて", others=["商品"])

    def test_short_hand_matches_when_the_fragment_is_unambiguous(self):
        # ★ バー2の核: 人は実在名を短く言う（『小計金額』を「小計」と呼ぶ）
        assert name_matches_task("小計金額", "小計の列を太くして", others=["商品", "数量"])

    def test_fragment_that_names_another_real_column_is_not_evidence(self):
        # ★ バー1の核: 『数量*単価』の断片『数量』は**他の実在列**も指しうる＝証拠にならない
        assert not name_matches_task(
            "数量*単価", "数量と単価をかけた金額列を作って、見出しを太字にして",
            others=["商品", "数量", "単価"])

    def test_an_occurrence_swallowed_by_another_real_name_is_not_evidence(self):
        # ★ 単位B（実測 2026-08-17 の穴）: 『金額』は依頼文に確かに現れるが、その出現は
        #   『税込金額』の一部としてしか説明できない＝依頼者が言ったのは『税込金額』の方。
        assert not name_matches_task("金額", "税込金額で並べ替えて", others=["商品", "税込金額"])
        # 飲み込まれていない出現が1つでもあれば従来どおり照合できる（(i) を狭めすぎない）。
        assert name_matches_task("金額", "税込金額と金額を比べて", others=["商品", "税込金額"])

    def test_fragments_shorter_than_two_characters_are_never_evidence(self):
        # 『金額』は依頼文に無く、共有する断片は1文字の『金』だけ＝証拠にしない
        # （★ 定義(i) の完全一致はそのまま有効なので、実列名が1文字の場合まで禁じてはいない）。
        assert not name_matches_task("金額", "お金の話", others=[])
        assert name_matches_task("金", "お金の話", others=[])   # (i) 完全一致は従来どおり


# ===========================================================================
# ★ 対照ケース: 同じブック・同じ op・同じ解決値で、依頼文だけが①②③を撃ち分ける
# ===========================================================================

CONTRAST_COLUMNS = ["商品", "金額"]


def _classify_bold_col(task):
    return classify_slots([Slot(key="target", value="col:金額", kind="region")],
                          task=task, columns=CONTRAST_COLUMNS, header_row=1,
                          sheets=["Sheet"], consumed=Consumed())


class TestContrast:
    def test_tier1_matched_when_the_task_names_the_column(self):
        assert [v.tier for v in _classify_bold_col("金額の列を太字にして")] == [MATCHED]

    def test_tier2_unspoken_when_the_task_says_nothing_about_a_target(self):
        assert [v.tier for v in _classify_bold_col("太字にして")] == [UNSPOKEN]

    def test_tier3_contradicted_when_the_task_points_somewhere_else(self):
        verdicts = _classify_bold_col("見出しを太字にして")
        assert [v.tier for v in verdicts] == [CONTRADICTED]
        assert verdicts[0].designators == ("見出し",)   # 誰も拾わなかった語だけが反証

    def test_designator_already_taken_by_another_slot_is_not_evidence(self):
        # ★ 実測で掴んだ誤爆の凍結: 「金額で降順に」+ [SORT 金額 / BOLD row:1] の2段目。
        #   『金額』は1段目が拾っているので、2段目の反証にはならない＝②（✓ は残す）。
        consumed = Consumed()
        first = classify_slots([Slot("col", "金額", "column")], task="金額で降順に並べ替えて",
                               columns=CONTRAST_COLUMNS, consumed=consumed)
        second = classify_slots([Slot("target", "row:1", "region")], task="金額で降順に並べ替えて",
                                columns=CONTRAST_COLUMNS, consumed=consumed)
        assert [v.tier for v in first] == [MATCHED]
        assert [v.tier for v in second] == [UNSPOKEN]


# ===========================================================================
# ★ 単位B: 部分文字列の**片方向**で ✓ が抜けていた穴の対照の対
# ===========================================================================

SUBSTRING_COLUMNS = ["商品", "金額", "税込金額"]


def _classify_sort_col(task, value):
    return classify_slots([Slot(key="col", value=value, kind="column")],
                          task=task, columns=SUBSTRING_COLUMNS, header_row=1,
                          sheets=["Sheet"], consumed=Consumed())


class TestSubstringDirections:
    """★ 実測 2026-08-17 の穴（単位B の残り）: 列が『商品/金額/税込金額』のブックで
       「税込金額で並べ替えて」→ 解決値『金額』が①になり **✓ が出ていた**（税込のつもりで
       頼んで税抜きの列が並べ替えられ、しかも機械検証済みと言う実害のある型）。逆向き
       「金額で」→『税込金額』は元から③で正しかった ―― **片方向だけ抜けていた**。
       同じブック・同じ op で依頼文と解決値を入れ替えた4通りをここで凍結する
       （片方を直して逆を壊す変更が入ったら赤くなる）。"""

    def test_tax_inclusive_request_resolved_to_the_bare_column_is_contradicted(self):
        # ★ これが塞いだ穴そのもの（単位B 以前は matched ＝ ✓ が出ていた）
        verdicts = _classify_sort_col("税込金額で並べ替えて", "金額")
        assert [v.tier for v in verdicts] == [CONTRADICTED]
        assert verdicts[0].designators == ("税込金額",)

    def test_bare_request_resolved_to_the_tax_inclusive_column_is_contradicted(self):
        # ★ 元から捕まえていた側（後退させないことの番人）
        verdicts = _classify_sort_col("金額で並べ替えて", "税込金額")
        assert [v.tier for v in verdicts] == [CONTRADICTED]
        assert verdicts[0].designators == ("金額",)

    def test_both_correct_pairings_stay_matched(self):
        assert [v.tier for v in _classify_sort_col("金額で並べ替えて", "金額")] == [MATCHED]
        assert [v.tier for v in _classify_sort_col("税込金額で並べ替えて", "税込金額")] == [MATCHED]

    def test_silence_about_the_column_is_still_unspoken(self):
        assert [v.tier for v in _classify_sort_col("並べ替えて", "金額")] == [UNSPOKEN]

    def test_a_designator_swallowed_by_another_real_name_is_not_left_over(self):
        """★ 同じ排除を designator 側にも当てた副次効果: 1段目が『税込金額』を拾った後、
           その中に埋まっているだけの『金額』が2段目の反証に化けない（②のまま＝✓ は残る）。"""
        task = "税込金額で並べ替えて、太字にして"
        consumed = Consumed()
        first = classify_slots([Slot("col", "税込金額", "column")], task=task,
                               columns=SUBSTRING_COLUMNS, consumed=consumed)
        second = classify_slots([Slot("target", "row:1", "region")], task=task,
                                columns=SUBSTRING_COLUMNS, consumed=consumed)
        assert [v.tier for v in first] == [MATCHED]
        assert [v.tier for v in second] == [UNSPOKEN]

    def test_end_to_end_the_check_disappears_for_the_wrong_column(self, tmp_path, monkeypatch, capsys):
        """★ 実害の形そのもの: 「税込金額で」と頼んで『金額』列が並べ替えられた run で
           ✓ が出ないこと（適用そのものは成功する ―― 止めるのは主張だけ）。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "金額", "税込金額"], ["a", 300, 330], ["b", 200, 220]])

        def _sort_by_amount(out_book, code, workdir, helper_files=(), timeout=None):
            wb = openpyxl.load_workbook(out_book)
            ws = wb.active
            for i, row in enumerate([["b", 200, 220], ["a", 300, 330]], start=2):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=v)
            wb.save(out_book)
            return True, None, "ok"

        _fixed(monkeypatch, {"op": "SORT", "args": {"col": "金額", "order": "asc"}}, _sort_by_amount)
        rc = ailine.main(run_argv(book=str(p), task="税込金額で並べ替えて", copy=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out
        assert "対象『金額』は依頼文の語と機械照合できません（依頼文が指しているのは: 税込金額）" in out


# ===========================================================================
# ★ バー1（真陽性）: 敵対例で ✓ が出ないこと
# ===========================================================================

class TestBar1TruePositives:
    def test_case1_header_word_vs_new_column_target_suppresses_the_check(self, tmp_path, monkeypatch, capsys):
        """症状そのもの（ブラインド査定2本が独立に致命の筆頭に置いた形）。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "数量", "単価"], ["a", 2, 100], ["b", 3, 200]])

        def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
            wb = openpyxl.load_workbook(out_book)
            ws = wb.active
            if 'setString("数量*単価")' in code:
                ws.cell(row=1, column=4, value="数量*単価")
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=4,
                            value=(ws.cell(row=r, column=2).value or 0) * (ws.cell(row=r, column=3).value or 0))
            if "Call StyleBold(oDoc, 3, " in code:
                for r in range(1, ws.max_row + 1):
                    if ws.cell(row=r, column=4).value not in (None, ""):
                        ws.cell(row=r, column=4).font = Font(bold=True)
            wb.save(out_book)
            return True, None, "ok"

        _fixed(monkeypatch, {"plan": [
            {"op": "COMPUTE_COLUMN", "args": {"operands": ["数量", "単価"], "operator": "*"}},
            {"op": "BOLD", "args": {"target": "col:数量*単価"}}]}, fake_apply)
        rc = ailine.main(run_argv(book=str(p),
                                  task="数量と単価をかけた金額列を作って、見出しを太字にして",
                                  copy=True, values=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out          # ★ バー1: ✓ を出さない
        assert "依頼文が指しているのは: 見出し" in out
        assert "意図した対象か確認してください" in out

    def test_case2_sheet_name_conflict_fallback_suppresses_the_check(self, tmp_path, monkeypatch, capsys):
        """単位D の検体: 依頼文の「金額」はシート名でもあり、曖昧なので1枚目へ後退した run。
           後退そのものは正しい判断だが、**依頼文の語と対象シートは照合できていない**。"""
        _isolate(monkeypatch, tmp_path)
        p = tmp_path / "売上.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "売上データ"
        ws1.append(["商品", "金額"])
        for row in [["a", 200], ["b", 300]]:
            ws1.append(row)
        ws2 = wb.create_sheet("金額")
        ws2.append(["月", "金額"])
        ws2.append(["1月", 50])
        wb.save(p)

        def sort_apply(out_book, code, workdir, helper_files=(), timeout=None):
            w = openpyxl.load_workbook(out_book)
            s = w["売上データ"]
            s.cell(row=2, column=1, value="b"); s.cell(row=2, column=2, value=300)
            s.cell(row=3, column=1, value="a"); s.cell(row=3, column=2, value=200)
            w.save(out_book)
            return True, None, "ok"

        _fixed(monkeypatch, {"op": "SORT", "args": {"col": "金額", "order": "desc"}}, sort_apply)
        monkeypatch.setattr(ailine, "_stdin_isatty", lambda: False)
        rc = ailine.main(run_argv(book=str(p), task="金額を降順に並べ替えて", copy=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out          # ★ バー1
        assert "対象シート『売上データ』は依頼文の語と機械照合できません" in out

    def test_case3_tax_inclusive_request_with_a_plain_total_label(self, tmp_path, monkeypatch, capsys):
        """査定B の致命1: 依頼は「税込み合計」なのに、解釈は『合計』・率は未確定。
           ―― 事後条件は『合計』を正しく検証して通る（機械は嘘をつかない）が、
           それは**依頼どおり**ではない。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
        _fixed(monkeypatch, {"op": "APPEND_TOTAL", "args": {"col": "金額", "label": "合計"}},
               _append_total_apply(500))
        rc = ailine.main(run_argv(book=str(p), task="税込み合計を出して", copy=True, values=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out          # ★ バー1
        assert "金額の性質を限定していますが" in out

    def test_case3_is_silent_when_the_qualifier_is_already_in_the_declaration(self, tmp_path, monkeypatch, capsys):
        """★ 上の裏（誤爆の芽を潰す）: 対象列そのものが『税込金額』なら、ラベル『合計』は正しい。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "税込金額"], ["a", 330], ["b", 220]])
        _fixed(monkeypatch, {"op": "APPEND_TOTAL", "args": {"col": "税込金額", "label": "合計"}},
               _append_total_apply(550))
        rc = ailine.main(run_argv(book=str(p), task="税込金額の合計を出して", copy=True, values=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED in out              # ★ バー2 側（✓ が消えない）
        assert "金額の性質を限定していますが" not in out


# ===========================================================================
# ★ バー2（偽陽性・本番）: 正しい対象で ✓ が消えないこと
# ===========================================================================

class TestBar2FalsePositives:
    def test_short_hand_column_name_keeps_the_check(self, tmp_path, monkeypatch, capsys):
        """「小計の列を太くして」→ 実列名『小計金額』。部分一致で正しい＝① のまま。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "小計金額"], ["a", 300], ["b", 200]])
        _fixed(monkeypatch, {"op": "BOLD", "args": {"target": "col:小計金額"}}, _bold_whole_column)
        rc = ailine.main(run_argv(book=str(p), task="小計の列を太くして", copy=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED in out                                   # ★ バー2
        assert "機械照合していません" not in out                   # ①なので注記も出ない
        assert "機械照合できません" not in out

    def test_correct_chain_to_an_auto_named_new_column_keeps_the_check(self, tmp_path, monkeypatch, capsys):
        """「売上から原価を引いた利益列を作って、利益で降順に」―― 2段目の対象は機械が
           自動命名した『売上-原価』。依頼文の『売上』『原価』は1段目が使っている＝反証は
           残っていないので②（✓ は残り、範囲を狭める1文が付く）。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "売上", "原価"], ["a", 500, 300], ["b", 900, 400]])

        def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
            wb = openpyxl.load_workbook(out_book)
            ws = wb.active
            if "SortByColumn" in code:
                rows = [["b", 900, 400, 500], ["a", 500, 300, 200]]
            else:
                rows = [["a", 500, 300, 200], ["b", 900, 400, 500]]
            ws.cell(row=1, column=4, value="売上-原価")
            for i, row in enumerate(rows, start=2):
                for j, v in enumerate(row, start=1):
                    ws.cell(row=i, column=j, value=v)
            wb.save(out_book)
            return True, None, "ok"

        _fixed(monkeypatch, {"plan": [
            {"op": "COMPUTE_COLUMN", "args": {"operands": ["売上", "原価"], "operator": "-"}},
            {"op": "SORT", "args": {"col": "利益", "order": "desc"}}]}, fake_apply)
        rc = ailine.main(run_argv(book=str(p),
                                  task="売上から原価を引いた利益列を作って、利益で降順に並べ替えて",
                                  copy=True, values=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED in out                                   # ★ バー2
        assert "対象『売上-原価』は依頼文の語と機械照合していません" in out   # ②の1文


# ===========================================================================
# ★ ② の1文（常時注記の置き換え）と、①では何も出ないこと
# ===========================================================================

class TestScopeNote:
    def test_note_is_run_specific_not_always_on(self, tmp_path, monkeypatch, capsys):
        """★ 旧 _VERIFY_SCOPE_NOTE は ✓ が出る全 run で必ず出ていた（発火率100%＝情報量ゼロ）。
           ①だけの run では何も出ない ―― 出たときに意味がある形にした。"""
        _isolate(monkeypatch, tmp_path)
        p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])

        def sort_apply(out_book, code, workdir, helper_files=(), timeout=None):
            wb = openpyxl.load_workbook(out_book)
            ws = wb.active
            ws.cell(row=2, column=1, value="a"); ws.cell(row=2, column=2, value=300)
            ws.cell(row=3, column=1, value="b"); ws.cell(row=3, column=2, value=200)
            wb.save(out_book)
            return True, None, "ok"

        _fixed(monkeypatch, {"op": "SORT", "args": {"col": "金額", "order": "desc"}}, sort_apply)
        rc = ailine.main(run_argv(book=str(p), task="金額で降順に並べ替えて", copy=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED in out
        assert "機械検証済み」は、上の" not in out    # 旧・常時注記の断片
        assert "★ ただし" not in out

    def test_note_names_the_machine_decided_subject(self, tmp_path, monkeypatch, capsys):
        _isolate(monkeypatch, tmp_path)
        p = _bold_book(tmp_path)
        _fixed(monkeypatch, {"op": "BOLD", "args": {"target": "row:1"}})
        rc = ailine.main(run_argv(book=str(p), task="太字にして", copy=True))
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED in out
        assert "★ ただし対象『row:1』は依頼文の語と機械照合していません" in out


# ===========================================================================
# ★ 関所の再利用（新しい exit code は作らない）
# ===========================================================================

class TestGateReuse:
    def _run_contradiction(self, tmp_path, monkeypatch, **overrides):
        p = _bold_book(tmp_path)
        _fixed(monkeypatch, {"op": "BOLD", "args": {"target": "col:金額"}}, _bold_whole_column)
        return ailine.main(run_argv(book=str(p), task="見出しを太字にして", **overrides))

    def test_interactive_no_aborts_with_the_existing_exit_code(self, tmp_path, monkeypatch, capsys):
        _isolate(monkeypatch, tmp_path)
        asked = []

        def _record(prompt=""):
            asked.append(prompt)
            return "n"
        monkeypatch.setattr("builtins.input", _record)
        rc = self._run_contradiction(tmp_path, monkeypatch)   # 既定＝原本へ直接反映
        out = capsys.readouterr().out
        assert rc == 1, out                 # ★ 既存の「対話で拒否」と同じ 1（新設しない）
        assert asked == ["この対象で実行しますか？ [y/N]: "]
        assert "× 中止した" in out

    def test_non_interactive_uses_the_existing_exit_code(self, tmp_path, monkeypatch, capsys):
        _isolate(monkeypatch, tmp_path)

        def _eof(prompt=""):
            raise EOFError()
        monkeypatch.setattr("builtins.input", _eof)
        rc = self._run_contradiction(tmp_path, monkeypatch)
        out = capsys.readouterr().out
        assert rc == 7, out                 # ★ 既存の「非対話で確認できない」と同じ 7
        assert "--copy" in out

    def test_interactive_yes_proceeds_but_still_refuses_to_claim(self, tmp_path, monkeypatch, capsys):
        """★ 人が y と答えても ✓ は出さない ―― ✓ は『機械が確かめた』の意味であって、
           人の承諾は機械の照合の代わりにならない。"""
        _isolate(monkeypatch, tmp_path)
        monkeypatch.setattr("builtins.input", lambda prompt="": "y")
        rc = self._run_contradiction(tmp_path, monkeypatch)
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out

    def test_copy_mode_does_not_ask_but_still_refuses_to_claim(self, tmp_path, monkeypatch, capsys):
        """--copy は原本に触れないので聞かない（既存の関所と同じ条件）。✓ は出さない。"""
        _isolate(monkeypatch, tmp_path)

        def _boom(prompt=""):
            raise AssertionError(f"--copy なのに入力を待った: {prompt!r}")
        monkeypatch.setattr("builtins.input", _boom)
        rc = self._run_contradiction(tmp_path, monkeypatch, copy=True)
        out = capsys.readouterr().out
        assert rc == 0, out
        assert VERIFIED not in out


# ===========================================================================
# ★ 一般則 vs 手書きの if（_maybe_warn_header_col_mismatch を消せるか）の実測
# ===========================================================================

def _classify_sequence(steps, task, columns):
    """steps: [(op, resolved)] を順に流し、③が出た段番号を返す（消費は段をまたいで持ち回る）。"""
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": list(columns)}, "header_rows": {"Sheet": 1}}
    carrier = type("Carrier", (), {})()
    hits = []
    for i, (op, resolved) in enumerate(steps, 1):
        verdicts = ailine.classify_subject_provenance(op, resolved, meta, task, carrier)
        if any(v.tier == CONTRADICTED for v in verdicts):
            hits.append(i)
    return hits


class TestGeneralRuleVsHandWrittenIf:
    """★ ブリーフの指示: 一般則が「同じケースをより広い条件で捕まえる」ことを**電池で確認して
       から**手書きの if を削る（先に削らない）。実測の結論は **削れない** ―― 下の
       `test_the_one_case_only_the_hand_written_if_catches` が反例そのもの。この4本は
       その測定を凍結する（将来この反例が一般則で捕まるようになったら、削る判断ができる）。"""

    def test_general_rule_catches_the_symptom(self):
        assert _classify_sequence(
            [("COMPUTE_COLUMN", {"operands": ["数量", "単価"], "operator": "*", "_target_sheet": "Sheet"}),
             ("BOLD", {"target": "col:数量*単価", "_target_sheet": "Sheet"})],
            "数量と単価をかけた金額列を作って、見出しを太字にして",
            ["商品", "数量", "単価", "数量*単価"]) == [2]

    def test_general_rule_is_wider_for_ops_the_hand_written_if_ignores(self):
        # 旧: BOLD/FILL_COLOR/CENTER_ALIGN だけ・かつ「直前段の新規列」だけ → どちらも沈黙。
        assert ailine._maybe_warn_header_col_mismatch("SORT", {"col": "金額"}, ["金額"],
                                                      "見出しを並べ替えて") is None
        assert ailine._maybe_warn_header_col_mismatch("BOLD", {"target": "col:金額"}, [],
                                                      "見出しを太字にして") is None
        # 新: どちらも③（新規列かどうかにも op の種類にも依らない）。
        assert _classify_sequence([("SORT", {"col": "金額", "order": "desc", "_target_sheet": "Sheet"})],
                                  "見出しを並べ替えて", ["商品", "金額"]) == [1]
        assert _classify_sequence([("BOLD", {"target": "col:金額", "_target_sheet": "Sheet"})],
                                  "見出しを太字にして", ["商品", "金額"]) == [1]

    def test_the_one_case_only_the_hand_written_if_catches(self):
        """★ 反例（実測 2026-08-17）: 前段が作った新規列の名前が**依頼文にそのまま出ている**
           場合（「…税込金額の列を作って、見出しを太字にして」）、一般則は『税込金額』を①と
           判定して沈黙する ―― 依頼文の語と解決値は確かに照合できているから。だが依頼は
           「その列の**見出し**」であって列全体ではない。この食い違いは、依頼文のどの語が
           どの段に属するかを知らないと機械では分けられない（DSL 段は段ごとの依頼文を
           持たない・翻訳が返すのは op と args だけ）。よって手書きの if は残す。"""
        assert ailine._maybe_warn_header_col_mismatch(
            "BOLD", {"target": "col:税込金額"}, ["税込金額"],
            "金額に消費税10%をかけた税込金額の列を作って、見出しを太字にして") is not None
        assert _classify_sequence(
            [("COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "factor": 1.1,
                                 "_target_sheet": "Sheet"}),
             ("BOLD", {"target": "col:税込金額", "_target_sheet": "Sheet"})],
            "金額に消費税10%をかけた税込金額の列を作って、見出しを太字にして",
            ["商品", "金額", "税込金額"]) == []

    def test_neither_fires_on_a_plain_correct_request(self):
        assert ailine._maybe_warn_header_col_mismatch("BOLD", {"target": "col:金額"}, ["金額"],
                                                      "金額の列を太字にして") is None
        assert _classify_sequence([("BOLD", {"target": "col:金額", "_target_sheet": "Sheet"})],
                                  "金額の列を太字にして", ["商品", "金額"]) == []


# ===========================================================================
# ★ 宣言の番人（OP_WRITE_TARGET の番人テストと同じ形）
# ===========================================================================

def test_op_subject_slots_declares_all_ops():
    missing = [op for op in ailine.OP_SCHEMA if op not in ailine.OP_SUBJECT_SLOTS]
    assert missing == [], f"OP_SUBJECT_SLOTS に対象スロットの宣言が無い op: {missing}"


def test_op_subject_slots_are_well_formed():
    from ailine_core.subject import COLUMN, INPUT, LABEL, REGION, ROW, SHEET
    known = {COLUMN, INPUT, LABEL, REGION, ROW, SHEET}
    for op, decl in ailine.OP_SUBJECT_SLOTS.items():
        keys = [k for k, _kind in decl]
        assert len(set(keys)) == len(keys), f"{op}: 同じ slot を二重に宣言している"
        for key, kind in decl:
            assert kind in known, f"{op}.{key}: 未知の種別 {kind!r}"
            assert kind != SHEET, (
                f"{op}.{key}: 対象シートは全 op 共通で足す（_subject_slots）ので表には書かない")


@pytest.mark.parametrize("op", sorted(ailine.OP_SCHEMA))
def test_declared_subject_slots_are_real_slots_of_that_op(op):
    """宣言した slot 名が、その op が実際に持ちうるキーであること（綴り間違いの番人）。
       ★ OP_SCHEMA は必須 slot だけなので、既定値つきの任意 slot（APPEND_TOTAL の label 等）
       は _CONFIRM_FIELDS（確認行に出す＝解釈そのもの）側にあれば正当とみなす。"""
    schema_keys = set(ailine.OP_SCHEMA[op])
    confirm_keys = {key for _label, key, _t in ailine._CONFIRM_FIELDS.get(op, ())}
    for key, _kind in ailine.OP_SUBJECT_SLOTS[op]:
        assert key in (schema_keys | confirm_keys), f"{op}: 存在しない slot を宣言している: {key}"
