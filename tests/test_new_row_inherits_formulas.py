# 追加した行に、既存の式を引き継ぐ番人（2026-09-02）。
#
# ★★ 実測（README の「既知の問題」に自分で書いていた）:
#   「みかんの下に梨を追加して」の後、梨の行の利益列は**空のまま**だった。
#   宣言した値だけを書くので `✓` は正しいが、**人が期待するものとは違う**。
#
# ★ 式は発明ではない ── **隣の行から写す**（A' 原則を破らない）。
#   参照の付け替えは LibreOffice にやらせる（自前で式の文字列を書き換えると、
#   それは 2 つ目の参照解決の実装になる ── SwapRowsByName が moveRange を使うのと同じ線）。
#
# 契約:
#   ① 引き継ぐのは「**全データ行が式を持つ列**」だけ（形で決める・列挙しない）
#   ② 合計列は引き継がない（金額は直値＋合計行だけ式 ＝「全部が式」ではない）
#   ③ 合計行を**写す元にしない**（=SUM を新しい行に配ると壊れる）
#   ④ 人が値を指定した列は触らない（**人の指定が勝つ**）
#   ⑤ **黙ってやらない** ── 解釈行に「式を引き継ぐ列」を出す
#   ⑥ 増えた行を含めて、操作前の等式がまだ成り立つかを見る

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine                                   # noqa: E402
from ailine_core import row_identity            # noqa: E402


def _invoice(tmp_path: Path) -> Path:
    """請求書の形（金額は直値・税込金額は式・合計行つき）。"""
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "件数", "単価", "金額", "税込金額"])
    for i, (n, k, t) in enumerate([("あ", 12, 4800), ("い", 5, 12000), ("う", 9, 7200)], start=2):
        ws.append([n, k, t, k * t, f"=D{i}*1.1"])
    ws.append(["合計", None, None, "=SUM(D2:D4)", "=D5*1.1"])
    wb.save(p)
    return p


def _meta(p: Path) -> dict:
    return ailine.build_book_meta(p)


def test_only_all_formula_columns_are_inherited(tmp_path):
    """①② 税込金額だけ引き継ぐ。金額（直値＋合計行だけ式）は引き継がない。"""
    p = _invoice(tmp_path)
    cols, src = ailine.formula_columns_to_inherit(_meta(p), "請求", 1, 5, set())
    assert cols == [4], f"税込金額(0起点4)だけのはず: {cols}"
    assert src == 4, f"写す元は合計行でない最後のデータ行(4行目)のはず: {src}"


def test_the_total_row_is_never_the_source(tmp_path):
    """③ 合計行から写さない ── =SUM を新しい行に配ったら壊れる。"""
    p = _invoice(tmp_path)
    totals = ailine.total_rows_in(_meta(p), "請求", 1)
    _cols, src = ailine.formula_columns_to_inherit(_meta(p), "請求", 1, 5, set())
    assert totals == [5], totals
    assert src not in totals, f"写す元が合計行になっている: {src}"


def test_a_column_with_only_some_formulas_is_not_inherited(tmp_path):
    """① 「全データ行が式」を単独で縛る検体。

    ★★ 2026-09-02 の変異試験で見つけた穴: `all` を `any` に緩めても、
      既存の検体は**緑のまま**だった ── 合計行を先に除いているので、
      その表では `all` と `any` が同じ答えになっていた（打ち消し合っていた）。
      ★ **データ行の中に式と直値が混ざる列**を作らないと、この性質は縛れない。
    """
    p = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "混在"
    ws.append(["品名", "単価", "数量", "小計"])
    ws.append(["あ", 100, 2, "=B2*C2"])
    ws.append(["い", 200, 3, "=B3*C3"])
    ws.append(["う", 300, 4, 1200])          # ★ ここだけ直値（人が手で入れた形）
    wb.save(p)
    cols, _src = ailine.formula_columns_to_inherit(_meta(p), "混在", 1, 5, set())
    assert cols == [], (
        f"式と直値が混ざる列を引き継ごうとしている: {cols} ── "
        "その列は『式で出している』とは言えない")


def test_the_row_above_may_be_the_total_row(tmp_path):
    """③ 合計行を写す元にしない、を単独で縛る検体。

    ★★ 同じ変異試験で見つけた穴: 合計行を除く処理を消しても緑だった ──
      検体が「合計行より上に挿入する」形しか持っていなかったので、
      **合計行が“すぐ上の行”になる場合**が一度も試されていなかった。
    ★ ここは実害が大きい: =SUM(...) を新しいデータ行に配ると、表が壊れる。
    """
    p = _invoice(tmp_path)                    # 合計行は 5 行目
    cols, src = ailine.formula_columns_to_inherit(_meta(p), "請求", 1, 6, set())
    assert cols == [4], cols
    assert src == 4, f"合計行(5)から写そうとしている: {src}"


def test_a_column_the_person_declared_is_left_alone(tmp_path):
    """④ 人の指定が勝つ（黙って上書きしない）。"""
    p = _invoice(tmp_path)
    cols, _src = ailine.formula_columns_to_inherit(_meta(p), "請求", 1, 5, {"税込金額"})
    assert cols == [], f"人が指定した列を引き継ごうとしている: {cols}"


def test_a_table_without_formulas_inherits_nothing(tmp_path):
    """★ 陰性対照 ── 式が無い表で何かを引き継いだら、それは発明している。"""
    p = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "在庫"
    ws.append(["品名", "数量"])
    for n, q in [("ボルト", 120), ("ナット", 80), ("ワッシャー", 300)]:
        ws.append([n, q])
    wb.save(p)
    assert ailine.formula_columns_to_inherit(_meta(p), "在庫", 1, 4, set()) == ([], 0)


def test_the_interpretation_line_declares_it():
    """⑤ 黙ってやらない ── 解釈行に欄が在る。"""
    labels = [f[0] for f in ailine._CONFIRM_FIELDS["ADD_ROW"]]
    assert "式を引き継ぐ列" in labels, labels


def test_the_generated_basic_copies_instead_of_rewriting():
    """★ 式の文字列を自分で書き換えていないこと（2 つ目の参照解決を作らない）。"""
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    i = bas.index("Sub FillFormulasFromNeighbour")
    j = bas.index("End Sub", i)
    body = bas[i:j]
    assert "copyRange" in body, "copyRange で LibreOffice に付け替えさせる"
    assert "setFormula" not in body, "式を自分で組み立てている（参照解決の二重実装）"


# --- ⑥ 増えた行を含めた等式の検算（純ロジック）------------------------------------

def test_an_inserted_row_that_breaks_the_identity_is_reported():
    """新しい行に数が揃っていて等式を満たさないなら言う。"""
    before = [[12, 4800, 57600], [5, 12000, 60000], [9, 7200, 64800]]
    after = before[:1] + [[2, 100, 999]] + before[1:]      # 2×100=200 のはずが 999
    assert row_identity.broken_after_insert(before, after)


def test_an_inserted_row_without_numbers_is_silent():
    """★ 鳴りすぎない ── 値を入れなかった行では鳴らない（そこは別の番人の担当）。"""
    before = [[12, 4800, 57600], [5, 12000, 60000], [9, 7200, 64800]]
    after = before[:1] + [[None, None, None]] + before[1:]
    assert row_identity.broken_after_insert(before, after) == []


def test_a_correct_inserted_row_is_silent():
    """★ 陰性対照 ── 正しく足した行では鳴らない。"""
    before = [[12, 4800, 57600], [5, 12000, 60000], [9, 7200, 64800]]
    after = before[:1] + [[2, 100, 200]] + before[1:]
    assert row_identity.broken_after_insert(before, after) == []


def test_broken_still_ignores_row_count_changes():
    """★ 元の契約を壊していないこと（broken は行数が変わる回を見ない）。"""
    before = [[12, 4800, 57600], [5, 12000, 60000], [9, 7200, 64800]]
    after = before + [[2, 100, 999]]
    assert row_identity.broken(before, after) == []


@pytest.mark.local
def test_the_formula_really_lands_on_real_libreoffice(tmp_path):
    """★ 実機 ── 新しい行に式が入り、参照が 1 行ぶんずれていること。

    ★ 「生成した」で終わらせない。この repo が何度も踏んだのは
      **走ったのに何も起きていない**形（Basic の予約語衝突など）。
    """
    import subprocess
    p = _invoice(tmp_path)
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p),
         "5行目に えの行を追加して、件数は 2、単価は 100、金額は 200 にして", "--copy"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=900,
        env={**__import__("os").environ, "PYTHONPATH": str(REPO / "src")})
    assert r.returncode == 0, r.stdout[-900:]
    out = p.with_name(p.stem + ".out.xlsx")
    ws = openpyxl.load_workbook(out)["請求"]
    assert ws.cell(row=5, column=5).value == "=D5*1.1", (
        f"新しい行に式が来ていない: {ws.cell(row=5, column=5).value!r}")
