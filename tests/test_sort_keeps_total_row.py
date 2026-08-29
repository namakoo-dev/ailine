# 合計行があると並べ替えできない ── 2026-08-29。Namakoo「並び替えできない」
#
# ★★ 実測: 合計行のある表に「金額の大きい順に行を並び替えて」で
#     E2: 値 57600 → '=SUM(#REF!:INDEX(E:E,ROW()-1))'
#     ★ 疑わしい: 適用後にエラー値のセルが増えました（計2件）
#     × 並び順の検証対象に式はあるがキャッシュ値が無く検証できない
#   合計（一番大きい）が**先頭へ飛び**、その式が壊れていた。番人は止めたが、
#   人は並べ替えられない ── 直すべきは道具のほう。
#
# ★ 合計行は「データ行ではない」── 並べ替えの対象から外し、最下行に残す。
#   判定は既存の凍結規則を借りる（total_rows_in → row_has_total_word）。
#   ★ 見つけたら**必ず画面に出す**（黙って行を外さない）。
#
# ★★ 直す途中で「分母の片側だけを縮める」を 2 回やった（片配線・今日 5 度目）:
#   ① 並べ替えの範囲だけ縮めて、検算は全行を見た → 「指定順に並んでいない」
#   ② 検算の後ろ側だけ縮めて、前側は全行 → 「行数が変わっています」
#   ★ 宣言（_sort_end_row）を**同じ 1 箇所**から全部に配る、が処方。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HEADERS = ["取引先", "金額"]
DATA = [["丸和物流", 57600], ["近江スチール", 60000], ["北斗精機", 114000]]


def _book(tmp_path, with_total=True, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in DATA:
        ws.append(r)
    if with_total:
        ws.cell(5, 1, "合計")
        ws.cell(5, 2, "=SUM(B2:B4)")
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(path)}


# --- ① 合計行は並べ替えない -------------------------------------------------------------

def test_the_total_row_is_excluded_from_the_sort(tmp_path):
    ok, r, _i, err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, _meta(_book(tmp_path)))
    assert ok, err
    assert r["_sort_end_row"] == 4, r.get("_sort_end_row")


def test_the_exclusion_is_disclosed_without_demoting(tmp_path):
    """★ 黙って行を外さない ── 外したことは必ず画面に出す。
    ★★ ただし**警告ではなく解釈行に出す**（SET_WHERE が合計行を外す時と同じ口）。
       警告にすると決裁③で ✓ が △ に落ち、合計行のある表を並べ替えるたびに
       「確かめきれていない」と言うことになる ── 宣言どおりに動いて検算も通って
       いるのだから、それは ✓ でよい。"""
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, _meta(_book(tmp_path)))
    assert r.get("_skip_rows") == [5], r.get("_skip_rows")
    assert "合計行" in r.get("_skip_label", ""), r.get("_skip_label")
    assert not any("合計行" in w for w in r.get("_warnings", [])), r.get("_warnings")


def test_without_a_total_row_nothing_changes(tmp_path):
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, _meta(_book(tmp_path, with_total=False)))
    assert not r.get("_sort_end_row"), r.get("_sort_end_row")
    assert not any("合計行" in w for w in r.get("_warnings", []))


def test_the_codegen_passes_the_end_row(tmp_path):
    p = _book(tmp_path)
    ok, r, _i, err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, _meta(p))
    assert ok, err
    code = ailine.codegen_dsl("SORT", r, _meta(p))
    assert "SortByColumnUpTo(oDoc, 0, 1, 1, False, 3)" in code, code  # Basic は 0 起点


def test_the_basic_uses_the_given_end_row():
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    i = bas.index("Sub SortByColumnUpTo(")
    seg = bas[i:i + 1200]
    assert "endRow As Long" in seg, seg[:200]
    assert "getCellRangeByPosition(0, headerRow + 1, lastCol, endRow)" in seg, seg[:800]
    # ★ 既存の SortByColumn は触っていない（引数を増やすと目録・README・凍結検体が全部動く）
    j = bas.index("Sub SortByColumn(")
    assert "Optional" not in bas[j:j + 200], bas[j:j + 200]


# --- ★★ ② 分母は 1 箇所から配る（片配線の再演を止める）--------------------------------

def _sorted_book(tmp_path, name="out.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in sorted(DATA, key=lambda x: -x[1]):
        ws.append(r)
    ws.cell(5, 1, "合計")
    ws.cell(5, 2, "=SUM(B2:B4)")
    wb.save(p)
    return p


def test_the_check_uses_the_same_range_on_both_sides(tmp_path):
    """★★ 直す途中で 2 回やらかした形:
       ① 並べ替えだけ縮めて検算は全行 → 「指定順に並んでいない」（合計は一番大きい）
       ② 検算の後ろ側だけ縮めて前側は全行 → 「行数が変わっています」
       ★ 宣言（_sort_end_row）を同じ 1 箇所から全部に配ること。"""
    src = _book(tmp_path, name="src.xlsx")
    out = _sorted_book(tmp_path)
    args = {"col": "金額", "order": "desc", "_target_sheet": "請求", "_sort_end_row": 4}
    status, why = ailine.check_sort(out, args, source_book=src)
    assert status == "pass", why
    assert "7 行" not in why


def test_a_wrong_order_still_fails(tmp_path):
    """★ 黙りすぎていないこと: 範囲を縮めても、並びが違えば落ちる。"""
    src = _book(tmp_path, name="src2.xlsx")
    out = _book(tmp_path, name="out2.xlsx")      # 並べ替えていない（昇順のまま）
    args = {"col": "金額", "order": "desc", "_target_sheet": "請求", "_sort_end_row": 4}
    status, why = ailine.check_sort(out, args, source_book=src)
    assert status == "fail", why


def test_the_end_row_reaches_every_denominator():
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    # 検算の後ろ側と前側の両方で使われていること
    assert src.count('args.get("_sort_end_row")') >= 2, "分母の片側にしか配っていない"
