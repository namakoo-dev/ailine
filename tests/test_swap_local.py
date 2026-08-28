"""SWAP の実機（LO・basrun）検体 ── 2026-08-27。

★★ なぜ実機の検体が要るか（今日この形で 3 時間払った）:
  生成した Basic が**コンパイルに失敗しても、basrun は「適用した」と言う**。
  マクロは 1 行も走らず、エラーも 1 行も出ない。sandbox の pytest は basrun をモック
  するので、この壊れ方は構造的に捕まえられない ── 実機でしか分からない。
  ★ 実際の犯人: 変数名 `oR`。Basic は大文字小文字を区別しないので **予約語 Or** と
    衝突してモジュールごと落ちていた。名前を変えるだけで直った。
  ★ 命綱は効いた: 事後条件が「変化なし」で × を出したので、嘘の ✓ にはならなかった。

★ もう 1 つの実測: **ヘルパの module を書き換えた直後の 1 回目は、古い module が走る**。
  切り分け中の「通ったり通らなかったり」はこれだった（同じ入力で 2 回目以降は安定）。

★ タイムアウトは短く（楔でスイート全体を道連れにしない）。
"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent


def _book(tmp_path, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価", "利益"])
    for i, (n, a, b) in enumerate([("りんご", 1200, 700), ("みかん", 800, 300),
                                    ("ぶどう", 1500, 900)], start=2):
        ws.cell(i, 1, n), ws.cell(i, 2, a), ws.cell(i, 3, b)
        ws.cell(i, 4, f"=B{i}-C{i}")
    wb.save(p)
    return p


def _run(book, task):
    return subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(book), task, "--copy", "--timeout", "90"],
        capture_output=True, text=True, timeout=420, encoding="utf-8", errors="replace")


def _out(book):
    return book.with_name(book.stem + ".out" + book.suffix)


@pytest.mark.local
def test_rows_really_swap_on_real_lo_and_formulas_follow(tmp_path):
    book = _book(tmp_path)
    p = _run(book, "みかんとぶどうを入れ替えて")
    assert p.returncode == 0, f"実機の行入れ替えが失敗:\n{p.stdout[-900:]}"
    raw = openpyxl.load_workbook(_out(book))["売上"]
    val = openpyxl.load_workbook(_out(book), data_only=True)["売上"]
    names = [raw.cell(r, 1).value for r in range(2, 5)]
    assert names == ["りんご", "ぶどう", "みかん"], f"入れ替わっていない: {names}"
    # ★ 芯: 式は**自分の行**を指し続ける（値を交換する実装だと、ここが他の行の値になる）
    assert val.cell(3, 4).value == 600, f"ぶどうの利益が自分の値でない: {val.cell(3, 4).value}"
    assert val.cell(4, 4).value == 500, f"みかんの利益が自分の値でない: {val.cell(4, 4).value}"


@pytest.mark.local
def test_columns_really_swap_on_real_lo_and_references_follow(tmp_path):
    book = _book(tmp_path, "c.xlsx")
    p = _run(book, "売上と原価の列を入れ替えて")
    assert p.returncode == 0, f"実機の列入れ替えが失敗:\n{p.stdout[-900:]}"
    raw = openpyxl.load_workbook(_out(book))["売上"]
    val = openpyxl.load_workbook(_out(book), data_only=True)["売上"]
    assert [raw.cell(1, c).value for c in range(1, 5)] == ["商品", "原価", "売上", "利益"]
    # 参照は LibreOffice が付け替える（=B2-C2 → =C2-B2）。**計算結果は変わらない**のが正しい。
    assert str(raw.cell(2, 4).value).replace(" ", "") == "=C2-B2", raw.cell(2, 4).value
    assert val.cell(2, 4).value == 500, val.cell(2, 4).value


@pytest.mark.local
def test_the_generated_basic_actually_runs(tmp_path):
    """★ 恒真殺し: 「適用した」と言われても走っていない、という今日の壊れ方を直接見る。
       入れ替えを頼んで**原本と 1 バイトも違わない**なら、それはマクロが走っていない。"""
    book = _book(tmp_path, "d.xlsx")
    before = book.read_bytes()
    p = _run(book, "みかんとぶどうを入れ替えて")
    assert p.returncode == 0, p.stdout[-600:]
    assert _out(book).read_bytes() != before, (
        "出力が原本と同一 ── マクロが 1 行も走っていない疑い"
        "（Basic のコンパイル失敗は basrun からは『適用した』に見える）")


# --- 列の追加（2026-08-27・Namakoo「列の追加はできないの？」）------------------------
#
# ★ 一段目は**同じ依頼文で回ごとに違う op** を返す（実測）。「原価列の右に列を追加して」で
#   INSERT_ROWS が返る回があり、そのまま走れば**列を頼まれて行を挿す**。
#   軸そのものを間違える形なので、実機で通しておく。

@pytest.mark.local
def test_a_column_really_gets_inserted_at_the_named_position(tmp_path):
    book = _book(tmp_path, "e.xlsx")
    p = _run(book, "原価の右に備考の列を追加して")
    assert p.returncode == 0, f"実機の列追加が失敗:\n{p.stdout[-900:]}"
    raw = openpyxl.load_workbook(_out(book))["売上"]
    val = openpyxl.load_workbook(_out(book), data_only=True)["売上"]
    assert [raw.cell(1, c).value for c in range(1, 6)] == ["商品", "売上", "原価", "備考", "利益"]
    # 利益は D → E へずれるが、参照する 売上/原価 は動いていないので式も値もそのまま
    assert val.cell(2, 5).value == 500, val.cell(2, 5).value
    assert raw.cell(2, 4).value is None, "挿した列に値が入っている"


@pytest.mark.local
def test_asking_for_a_column_never_inserts_a_row(tmp_path):
    """★ 軸の取り違えを実機で縛る ── 行数が増えたら、それは列の依頼に行で答えている。"""
    book = _book(tmp_path, "f.xlsx")
    p = _run(book, "原価列の右に列を追加して")
    assert p.returncode == 0, p.stdout[-600:]
    ws = openpyxl.load_workbook(_out(book))["売上"]
    assert ws.max_row == 4, f"行が増えた（列の依頼に行で答えた）: {ws.max_row}"
    assert ws.max_column == 5, f"列が増えていない: {ws.max_column}"


# --- 条件つき書換（2026-08-27・Namakoo「原価が500以上の項目に◎を付ける」）------------

@pytest.mark.local
def test_only_the_matching_rows_get_the_mark_on_real_lo(tmp_path):
    """★ 芯: **当てはまらない行に付いていない**ことまで実機で見る。
       「付いたか」だけ見る試験は、全行に付ける実装でも通ってしまう。"""
    p = tmp_path / "g.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価", "チェック"])
    for row in [["りんご", 1200, 700], ["みかん", 800, 300], ["ぶどう", 1500, 900]]:
        ws.append(row)
    wb.save(p)
    r = _run(p, "原価が500以上の項目のチェック列に「◎」を付けて")
    assert r.returncode == 0, f"実機の条件つき書換が失敗:\n{r.stdout[-900:]}"
    out = openpyxl.load_workbook(_out(p))["売上"]
    got = [out.cell(i, 4).value for i in range(2, 5)]
    assert got == ["◎", None, "◎"], f"当てはまる行だけ、が破れた: {got}"


@pytest.mark.local
def test_a_replace_touches_only_the_matching_rows_on_real_lo(tmp_path):
    """★ 置き換え「『◎』を『合格』に」── 空欄の行を巻き込まないことを実機で見る
       （一段目は「列を丸ごと『合格』に」を返していた ── そちらだと空欄まで潰れる）。"""
    p = tmp_path / "h.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価", "チェック"])
    for row in [["りんご", 1200, 700, "◎"], ["みかん", 800, 300, None],
                 ["ぶどう", 1500, 900, "◎"]]:
        ws.append(row)
    wb.save(p)
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p), "チェック列の「◎」を全て「合格」に書き換えて",
         "--copy", "--overwrite", "--timeout", "90"],
        capture_output=True, text=True, timeout=420, encoding="utf-8", errors="replace")
    assert r.returncode == 0, f"実機の置き換えが失敗:\n{r.stdout[-900:]}"
    out = openpyxl.load_workbook(_out(p))["売上"]
    got = [out.cell(i, 4).value for i in range(2, 5)]
    assert got == ["合格", None, "合格"], f"空欄の行を巻き込んだ: {got}"


@pytest.mark.local
def test_named_rows_and_columns_extract_on_real_lo(tmp_path):
    """★ 名指しの抽出（行）と列の抽出を実機で。**空でも正しく見える**のが一番こわい形なので、
       中身まで見る（0 行の抽出結果は「成功」に見えてしまう）。"""
    p = tmp_path / "i.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価"])
    for row in [["りんご", 1200, 700], ["みかん", 800, 300], ["ぶどう", 1500, 900]]:
        ws.append(row)
    wb.save(p)

    r = _run(p, "みかんの行とりんごの行だけを抽出して")
    assert r.returncode == 0, f"名指しの抽出が失敗:\n{r.stdout[-900:]}"
    wb2 = openpyxl.load_workbook(_out(p))
    dst = [s for s in wb2.sheetnames if "どれか" in s]
    assert dst, wb2.sheetnames
    got = [[wb2[dst[0]].cell(i, j).value for j in (1, 2)]
            for i in range(2, wb2[dst[0]].max_row + 1)]
    assert got == [["りんご", 1200], ["みかん", 800]], f"抽出の中身が違う: {got}"

    p2 = tmp_path / "j.xlsx"
    wb.save(p2)
    r2 = _run(p2, "商品と原価の列だけ抜き出して")
    assert r2.returncode == 0, f"列の抽出が失敗:\n{r2.stdout[-900:]}"
    wb3 = openpyxl.load_workbook(_out(p2))
    dst2 = [s for s in wb3.sheetnames if "だけ" in s]
    assert dst2, wb3.sheetnames
    out = wb3[dst2[0]]
    assert [out.cell(1, j).value for j in (1, 2)] == ["商品", "原価"]
    assert out.max_row == 4, f"行が減っている: {out.max_row}"


# --- 行番号で 1 セル（2026-08-28・Namakoo「指示文が通らない」）------------------------
#
# ★★ sandbox の試験は「宣言と実体」しか見られない ── 今日の事故は**依頼**を見て
#   いなかったことなので、実機で「頼んだ 1 セル以外が動いていない」ところまで見る。

@pytest.mark.local
def test_a_row_number_writes_exactly_one_cell_on_real_lo(tmp_path):
    p = tmp_path / "k.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "件数", "担当"])
    for row in [["丸和物流", 10, "田中"], ["ヤマノ食品", 20, "鈴木"],
                 ["北斗精機", 30, "田中"], ["ヤマノ食品", 40, None]]:
        ws.append(row)
    wb.save(p)
    before = [[ws.cell(r, c).value for c in range(1, 4)] for r in range(2, 6)]

    r = _run(p, "5行目の担当を「佐藤」にして")
    assert r.returncode == 0, f"行番号の 1 セル書換が失敗:\n{r.stdout[-900:]}"
    out = openpyxl.load_workbook(_out(p))["請求"]
    got = [[out.cell(i, c).value for c in range(1, 4)] for i in range(2, 6)]
    # ★ 芯: 5 行目だけ。「佐藤が入ったか」だけ見る試験は、列を潰す実装でも通る。
    assert got[3][2] == "佐藤", f"5行目に入っていない: {got[3]}"
    assert [g[2] for g in got] == ["田中", "鈴木", "田中", "佐藤"], \
        f"頼んでいない行まで変わった（列を潰した）: {[g[2] for g in got]}"
    assert [g[:2] for g in got] == [b[:2] for b in before], "他の列が変わった"


@pytest.mark.local
def test_the_a1_column_letter_form_reaches_the_same_one_cell(tmp_path):
    """★ 一段目は 3/3 で **ADD_ROW**（行の追加）を返していた形。行が増えないこと。"""
    p = tmp_path / "l.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "件数", "担当"])
    for row in [["丸和物流", 10, "田中"], ["ヤマノ食品", 20, "鈴木"]]:
        ws.append(row)
    wb.save(p)
    r = _run(p, "3 行C列に「佐藤」を追加")
    assert r.returncode == 0, f"A1 の列名が通らない:\n{r.stdout[-900:]}"
    out = openpyxl.load_workbook(_out(p))["請求"]
    assert out.max_row == 3, f"行が増えた（列への書き込みに行で答えた）: {out.max_row}"
    assert [out.cell(i, 3).value for i in (2, 3)] == ["田中", "佐藤"], \
        [out.cell(i, 3).value for i in (2, 3)]


# --- 請求書を 1 枚にまとめる（2026-08-28・Namakoo「同名の取引先から複数の発注」）------
#
# ★ sandbox では Basic が走らない ── 明細行の**増殖**（insertByIndex + copyRange）は
#   実機でしか確かめられない。設計査読が「唯一の未知」と名指しした所なので、
#   書式つき・結合ありの雛形で通す。

@pytest.mark.local
def test_orders_for_the_same_customer_land_on_one_invoice(tmp_path):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "項目", "件数", "単価", "金額", "締め日", "担当"])
    for row in [["丸和物流", "配送業務一式", 12, 4800, 57600, "2026/08/31", "田中"],
                 ["ヤマノ食品", "食品仕入", 28, 1500, 42000, "2026/08/31", "田中"],
                 ["ヤマノ食品", "冷蔵配送", 6, 3000, 18000, "2026/08/31", "田中"]]:
        ws.append(row)
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    tp = wb.create_sheet("雛形")
    tp["A1"] = "請　求　書"
    tp["A1"].font = Font(size=18, bold=True)
    tp.merge_cells("A1:D1")                     # ★ 横の結合（崩れてはいけない）
    tp["A3"] = "{{取引先}}"
    tp["B3"] = "御中"
    tp["A7"] = "ご請求金額"
    tp["B7"] = "{{合計:金額}}"
    tp["B7"].fill = PatternFill("solid", fgColor="FFF3D6")
    for c, t in zip("ABCD", ["項目", "数量", "単価", "金額"]):
        tp[f"{c}10"] = t
        tp[f"{c}10"].font = Font(bold=True)
        tp[f"{c}10"].alignment = Alignment(horizontal="center")
    tp["A11"] = "{{明細:項目}}"
    tp["B11"] = "{{明細:件数}}"
    tp["C11"] = "{{明細:単価}}"
    tp["D11"] = "{{明細:金額}}"
    for c in "ABCD":
        tp[f"{c}11"].border = box            # ★ 罫線が複製先まで乗ること
    tp["C11"].number_format = "#,##0"
    tp["D11"].number_format = "#,##0"
    tp["A13"] = "お支払期限"
    tp["B13"] = "{{締め日}}"
    tp["A16"] = "備考"
    tp["B16"] = "毎度ありがとうございます。"
    wb.save(p)

    r = _run(p, "取引先ごとに請求書を作って")
    assert r.returncode == 0, f"まとめ版の帳票が失敗:\n{r.stdout[-1200:]}"
    out = openpyxl.load_workbook(_out(p))
    assert "ヤマノ食品_2" not in out.sheetnames, f"2 枚に割れた: {out.sheetnames}"
    ya = out["ヤマノ食品"]
    # ★ 明細が 2 行に増え、それぞれ自分の値
    assert [ya.cell(11, c).value for c in range(1, 5)] == ["食品仕入", 28, 1500, 42000]
    assert [ya.cell(12, c).value for c in range(1, 5)] == ["冷蔵配送", 6, 3000, 18000]
    # ★ 合計は足した値（宣言値でなく実際の和）
    assert ya["B7"].value == 60000, ya["B7"].value
    # ★ 増やした分だけ下がずれ、固定文は無傷（ここが一番静かに壊れる）
    assert ya.cell(14, 1).value == "お支払期限" and ya.cell(14, 2).value == "2026/08/31"
    assert ya.cell(17, 2).value == "毎度ありがとうございます。"
    # ★ 書式が複製先まで乗っている／横の結合が残っている
    assert ya.cell(12, 1).border.left.style == "thin", "複製した明細行に罫線が無い"
    assert ya.cell(12, 4).number_format == "#,##0", "複製した明細行の桁区切りが落ちた"
    assert "A1:D1" in [str(m) for m in ya.merged_cells.ranges], "横の結合が崩れた"
    # ★ 1 件だけの取引先は増えない
    assert out["丸和物流"].cell(13, 1).value == "お支払期限"
    # ★ 検分に「どの発注が 1 枚に入ったか」が出る
    ins = out["検分"]
    got = {ins.cell(i, 1).value: (ins.cell(i, 2).value, ins.cell(i, 3).value)
            for i in range(2, ins.max_row + 1)}
    assert got["ヤマノ食品"] == ("3,4", 2), got
