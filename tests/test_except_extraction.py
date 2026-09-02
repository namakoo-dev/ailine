# 「〜以外」の抽出（cmp `nin`）の番人（2026-09-02）。
#
# ★★ この操作は長らく**名指しで断って**いた（README「作らなかったこと」）:
#   「この述語は Python・Basic・凍結した真理値表の 3 箇所が独立に持つので、
#    締切前に触ると 3 つの同期がずれる」── 正しい判断だった。
#   ★ 締切は過ぎ、**ずれたら赤くなる番人**を先に置いてから実装した
#     （tests/test_compare_codes_stay_in_sync.py）。
#
# ★★ 実装中に**片配線を踏んだ**（実測・検体が警告していた事故の再演）:
#   読み直しは正しく `cmp=nin` を立てたのに、検証段の
#     resolved["cmp"] = "in" if len(_named_vals) > 1 else "eq"
#   が**否定を知らないまま上書き**し、「味噌汁以外を抜き出して」で
#   **味噌汁だけを抜き出して △** が出た ── 逆のことをして合格。
#   ★ 読み直しに足して、**決定の場所に足し忘れた**。だからこの番人は
#     読み直しではなく**決定の場所**（verify_dsl_args）を通す。
#
# 契約:
#   ① 「〜以外」は削除に化けない（残したい行を消すのは取り返しがつかない）
#   ② 決定の場所を通っても `nin` が生き残る（片配線を作らない）
#   ③ 述語は `in` の否定（空欄は残す・部分一致にしない）
#   ④ 「以外」と言っていない依頼を否定に読み替えない
#   ⑤ 出力シート名が日本語になっている

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine  # noqa: E402


def _menu(tmp_path: Path) -> Path:
    p = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "献立"
    ws.append(["料理", "主材料", "分量", "備考"])
    for r in [["カレー", "牛肉", 4, ""], ["味噌汁", "豆腐", 2, ""], ["サラダ", "レタス", 3, ""]]:
        ws.append(r)
    wb.save(p)
    return p


def test_it_is_never_flipped_into_a_deletion(tmp_path):
    """① 核心（実装前から変わらない）── 『以外』を削除に化けさせない。"""
    meta = ailine.build_book_meta(_menu(tmp_path))
    assert ailine.removal_reading("味噌汁以外を抜き出して", meta, "献立") is None


def test_the_negation_survives_the_decision_point(tmp_path):
    """★★ ② ここが芯（2026-09-02 に実際に踏んだ片配線）。

      読み直しが `nin` を立てても、検証段が `eq` に上書きしていた
      ── **味噌汁だけを抜き出して △ を出す**、逆のことをして合格する形。
    ★ だから読み直しではなく**決定の場所**を通して見る。
    """
    meta = ailine.build_book_meta(_menu(tmp_path))
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "料理", "cmp": "nin", "value": ["味噌汁"]}, meta,
        task="味噌汁以外を抜き出して")
    assert ok, err
    assert resolved["cmp"] == "nin", f"否定が消えている: {resolved['cmp']!r}"
    assert list(resolved["value"]) == ["味噌汁"], resolved["value"]


def test_a_positive_extraction_still_decides_the_same_way(tmp_path):
    """★ 陰性対照 ── 「以外」でない依頼まで否定にしない。"""
    meta = ailine.build_book_meta(_menu(tmp_path))
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "料理", "cmp": "eq", "value": "味噌汁"}, meta,
        task="味噌汁の行を抜き出して")
    assert ok, err
    assert resolved["cmp"] == "eq", resolved["cmp"]


def test_the_machine_reads_it_without_the_llm(tmp_path):
    """④ 列も値も機械が実表から解く（LLM に「以外」を教えない）。"""
    meta = ailine.build_book_meta(_menu(tmp_path))
    assert ailine.task_says_except("味噌汁以外を抜き出して")
    assert not ailine.task_says_except("味噌汁の行を抜き出して")
    col, vals = ailine.except_extraction_reading(meta, "献立", "味噌汁以外を抜き出して")
    assert col == "料理" and vals == ["味噌汁"], (col, vals)
    assert ailine.except_extraction_reading(
        meta, "献立", "味噌汁の行を抜き出して") == (None, None)


def test_the_predicate_is_the_negation_of_in():
    """③ 述語 ── `in` の否定。空欄は残し、部分一致にしない。"""
    f = ailine._extract_predicate("nin", ["味噌汁"])
    assert f("カレー") is True and f("味噌汁") is False
    assert f("") is True and f(None) is True, "空欄を落とすと残したい行を失う"
    assert f("味噌汁定食") is True, "部分一致の否定にしない"


def test_the_output_sheet_name_reads_as_japanese():
    """⑤ 連結だけだと『料理味噌汁のどれでもない』になる（eq/in と同じ理由）。"""
    assert ailine._extract_output_sheet_name("料理", "nin", ["味噌汁"]) == "料理が味噌汁以外"


def test_the_reread_fires_through_the_real_path(tmp_path, monkeypatch, capsys):
    """★★ **配線そのものを通す**検体（変異試験で開いていた穴）。

      決定の場所（verify_dsl_args）と述語は縛っていたが、**読み直しの経路**を通る
      非実機の試験が無く、「読み直しから否定を外す」変異が緑のままだった。
    ★ 一段目が語彙外を返す回（実測で起きた形）を作り、機械が否定として拾うことを見る。
      適用しない（--dry）ので LibreOffice も要らない。
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_golden_transcripts import _isolate, _run_main

    _isolate(monkeypatch, tmp_path)
    p = _menu(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "OUT_OF_VOCAB", "about": "以外の抽出"}]})
    _rc, out = _run_main(["run", str(p), "味噌汁以外を抜き出して", "--dry"], capsys)
    # ★★ 主張を 1 度きつくした: 最初は「"以外" が出力に在る」で見ていたが、
    #   **依頼文のこだま**とも当たるので、読み直しを外しても緑だった
    #   （決定の場所が結果だけは拾い直すため ── 今日 4 度目の打ち消し合い）。
    #   ★ ここが見たいのは**読み直しが何と言ったか**。文面を名指しで縛る。
    #     結果（cmp=nin）は別の検体（決定の場所）が見ている。
    assert "**以外**の行を抜き出します" in out, (
        "読み直しが否定として説明していない: " + out)
    assert "のどれでもない" in out, ("否定として決まっていない: " + out)


@pytest.mark.local
def test_it_really_extracts_the_others_on_real_libreoffice(tmp_path):
    """★ 実機 ── **味噌汁以外**（カレーとサラダ）が出ること。

    ★ 「行数が減った」だけ見る試験は、逆のことをしても通る（1 行 vs 2 行）。
      中身まで見る。
    """
    import subprocess
    p = _menu(tmp_path)
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p), "味噌汁以外を抜き出して", "--copy"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=900,
        env={**__import__("os").environ, "PYTHONPATH": str(REPO / "src")})
    assert r.returncode == 0, r.stdout[-900:]
    wb = openpyxl.load_workbook(p.with_name(p.stem + ".out.xlsx"))
    dst = [s for s in wb.sheetnames if s != "献立"]
    assert len(dst) == 1, wb.sheetnames
    got = [wb[dst[0]].cell(i, 1).value for i in range(2, wb[dst[0]].max_row + 1)]
    assert got == ["カレー", "サラダ"], f"抜き出したのが違う: {got}"
