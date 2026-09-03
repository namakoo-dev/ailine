# 合計行が既に在る表に、別の列の合計を足す ── 2026-08-29。
# Namakoo「合計を一度入れたあとに単価の合計を入力させれない。9行目が増えたことによる誤動作」
#
# ★★ 実測: 合計行（9行目）が在る表に「単価列の合計行に単価の合計を書いて」と頼むと、
#   **10 行目に『単価合計』という別の行**が増え、事後条件が × を出した。
#   ★ 真因: 合計行を「データ行」と数えて、その**下**に足していた。
#     実務では 1 本の合計行に複数の列の合計が並ぶ ── ごく普通の形。
#
# ★ 直し: 合計行が 1 つに決まり、その列がまだ空なら**その行に書く**（行は増やさない）。
#   判定は既存の凍結規則を借りる（total_rows_in → row_has_total_word）── 新しい規則を
#   書かない。同じことを 2 箇所が決めると必ずずれる。
#
# ★ 「その回だけ宣言が変わる」ものが 3 つ出た。どれも op の宣言でなく**引数から分かる事実**:
#   ① ラベルは**その行に既に在る物**が正（LLM の案『単価合計』で検算しない）
#   ② 1 セルの書き込みなので「空欄への一括書き込み」の助言は出さない
#   ③ 末尾に足していないので「末尾に足すはずが既存行を書き換えた」の前提を外す

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from _product_source import product_text, window_around  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む

HEADERS = ["取引先", "件数", "単価", "金額"]
ROWS = [["丸和物流", 12, 4800, 57600], ["ヤマノ食品", 28, 1500, 42000]]


def _book(tmp_path, with_total=True, filled_col=None, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    if with_total:
        ws.cell(4, 1, "合計")
        ws.cell(4, 4, "=SUM(D2:D3)")
        if filled_col:
            ws.cell(4, filled_col, 999)
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(path)}


def _verify(path, col="単価", task="単価列の合計行に単価の合計を書いて"):
    return ailine.verify_dsl_args("APPEND_TOTAL", {"col": col}, _meta(path), task=task)


# --- ① 既にある合計行に書く（行は増やさない）------------------------------------------

def test_it_writes_into_the_existing_total_row(tmp_path):
    ok, r, _i, err = _verify(_book(tmp_path))
    assert ok, err
    assert r["_at_row"] == 4, r.get("_at_row")
    assert "行は増やしません" in r.get("_at_basis", ""), r.get("_at_basis")


def test_the_label_comes_from_the_row_not_from_the_llm(tmp_path):
    """★ ①: 検算に使うラベルは**その行に既に在る物**。LLM の案で照合すると、
       正しく書けているのに『ラベルが違う』で × になる（実測）。"""
    ok, r, _i, err = _verify(_book(tmp_path))
    assert ok, err
    assert r["label"] == "合計", r["label"]


def test_the_codegen_writes_at_that_row_without_scanning(tmp_path):
    p = _book(tmp_path)
    ok, r, _i, err = _verify(p)
    assert ok, err
    code = ailine.codegen_dsl("APPEND_TOTAL", r, _meta(p))
    assert "totalRow = 3" in code, code          # Basic は 0 起点
    assert "lastRow" not in code, "既にある行に書くのに、末尾を走査している"
    assert 'setString("合計")' not in code, "ラベルを二重に書いている"


# --- ② 合計行が無ければ、今までどおり末尾に足す ----------------------------------------

def test_without_a_total_row_it_still_appends(tmp_path):
    ok, r, _i, err = _verify(_book(tmp_path, with_total=False))
    assert ok, err
    assert not r.get("_at_row"), r.get("_at_row")
    code = ailine.codegen_dsl("APPEND_TOTAL", r, _meta(_book(tmp_path, with_total=False)))
    assert "lastRow" in code and "totalRow = lastRow + 1" in code


# --- ③ 既に値が入っている列は、勝手に上書きしない --------------------------------------

def test_a_filled_cell_is_left_to_the_deeper_guards(tmp_path):
    """★★ ここで「既に値が入っています」と**断るのはやめた**。
       既存の番人（事後条件の算術の検算＝二重計上に ✓ を出さない／単位F の関所）が
       同じ事故を既に止めていて、断りを重ねると**その番人の出番が消える**
       ── 過去の事故を守っている検体が 3 本落ちた（実測）。
    ★ 埋められる時だけ埋め、それ以外は今までどおり深い番人に任せる。"""
    ok, r, _i, err = _verify(_book(tmp_path, filled_col=3))
    assert ok, err
    assert not r.get("_at_row"), "埋められない回に埋めようとしている"


def test_an_existing_sum_formula_may_be_replaced(tmp_path):
    """★ 既にその列の合計が入っているなら、作り直しは自然（=SUM( で始まる時だけ）。"""
    p = _book(tmp_path)
    wb = openpyxl.load_workbook(p)
    wb["請求"].cell(4, 3, "=SUM(C2:C3)")
    wb.save(p)
    ok, r, _i, err = _verify(p)
    assert ok, err
    assert r["_at_row"] == 4


# --- ★ その回だけ宣言が変わる（引数から分かる事実）--------------------------------------

def test_the_run_is_declared_as_a_single_cell_write(tmp_path):
    """★★ ②③: 「既にある合計行に書く」回は末尾に足していないし、1 セルしか書かない。
       宣言のままだと 2 つの助言/前提が誤爆する（実測で両方出た）:
         ・空欄への同一値の一括書き込みです（値 … × 1 セル）
         ・末尾に新しい行を足すはずが、既存の行の値を 1 件書き換えました
       ★ 直しは op の宣言ではなく、**その回の引数**で外す（位置がずれる回に
         位置ベースの前提を外すのと同じ形）。"""
    seg = window_around("_writes = write_target.writes", after=700)
    assert 'get("_at_row")' in seg and "WRITE_NEW_ROW_AT_END" in seg, seg[:300]
    assert "WRITE_SINGLE_CELL" in seg, seg[:300]
    j = product_text().index('{"single_cell": (_op_writes(op, WRITE_SINGLE_CELL)')
    assert '_at_row' in product_text()[j:j + 200], product_text()[j:j + 200]
