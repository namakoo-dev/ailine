# 帳票段 REPORT_PER_ROW ── helpers/AiLineHelpers.bas:FillReportSheet を実際に basrun
# （実 LibreOffice）で走らせ、3 行 → 3 枚が出て雛形の書式が保たれることを確かめる通し試験。
#
# ★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
#   （`pytest -m "not local"` で除外される。手動は `pytest -m local` または
#   `pytest tests/test_report_per_row_local.py`）。
# ★ tests/test_report_per_row.py（凍結済み・basrun_apply をモックした e2e）と役割が違う:
#   あちらは CLI パイプライン全体（②検証〜⑥事後条件）の契約を固定する。こちらは
#   helpers/AiLineHelpers.bas:FillReportSheet 自体が実 LO で本当に動くこと（copyByName・
#   印の走査・型の出し分け・雛形の書式保存）を見る（test_bold_local.py と同じ役割分担）。
import subprocess
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))
import ailine  # noqa: E402

HELPERS_DIR = REPO_ROOT / "src" / "ailine" / "helpers"


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1
    return out.lower().count("soffice.bin")


@pytest.mark.local
def test_fill_report_sheet_makes_3_sheets_from_3_rows_via_basrun(tmp_path):
    """3行のデータ → FillReportSheet を3回 Call → 3枚の報告シートが実 LO で出来ることを
       確かめる。★ この試験の存在意義: FillReportSheet の copyByName・印の走査
       （InStr による {{...}} 検出）・型の出し分け（getType→setValue/setString）の
       どれか1つが壊れても、この試験は赤くなる。"""
    before_procs = _soffice_process_count()
    try:
        book = tmp_path / "invoice.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "売上"
        for row in [["取引先", "金額", "備考"],
                     ["あかつき商事", 12000, "現金"],
                     ["みどり工業", 8500, "振込"],
                     ["そら建設", 30000, "振込"]]:
            ws.append(row)
        tpl = wb.create_sheet("雛形")
        tpl["A1"] = "請求書"
        tpl["A1"].font = Font(bold=True)   # ★ 雛形の書式（太字）が複製で保たれるかも見る
        tpl["A3"] = "宛先"
        tpl["B3"] = "{{取引先}}"
        tpl["A5"] = "ご請求金額"
        tpl["B5"] = "{{金額}}"
        tpl["B5"].number_format = "#,##0"   # ★ 雛形の数値書式（設計文書 訂正3の対象）
        tpl["A7"] = "備考"
        tpl["B7"] = "{{備考}}"
        wb.save(book)

        workdir = tmp_path / "work"
        workdir.mkdir()
        _catalog, helper_files = ailine.load_helpers(HELPERS_DIR)
        assert helper_files, "helpers/*.bas が見つからない（HELPERS_DIR の場所を確認）"

        # ★ Basic は 0 起点（headerRow=0・srcRow は 0 起点の物理行番号）。
        #   codegen_dsl と同じ変換規則（Excel の1起点 row → Basic の row-1）。
        code = (
            "Option VBASupport 1\nOption Explicit\n\n"
            "Sub Run(oDoc As Object)\n"
            '    Call FillReportSheet(oDoc, "雛形", "あかつき商事", "売上", 1, 0)\n'
            '    Call FillReportSheet(oDoc, "雛形", "みどり工業", "売上", 2, 0)\n'
            '    Call FillReportSheet(oDoc, "雛形", "そら建設", "売上", 3, 0)\n'
            "End Sub\n"
        )
        ok, err, raw = ailine.basrun_apply(book, code, workdir, helper_files)
        assert ok, f"basrun_apply が失敗した: {err}\n{(raw or '')[-800:]}"

        wb2 = openpyxl.load_workbook(book)
        made = [s for s in wb2.sheetnames if s not in ("売上", "雛形")]
        assert set(made) == {"あかつき商事", "みどり工業", "そら建設"}, made

        sh = wb2["あかつき商事"]
        # ★ 印セル: 型の出し分け（文字列は setString・数値は setValue）。
        assert sh["B3"].value == "あかつき商事"
        assert sh["B5"].value == 12000, f"金額が数値でない: {sh['B5'].value!r}"
        assert isinstance(sh["B5"].value, (int, float)), f"金額の型: {type(sh['B5'].value)}"
        assert sh["B7"].value == "現金"
        # ★ 雛形の書式が複製で保たれる（訂正3: 印以外は一切変えない・数値書式にも触れない）。
        assert sh["A1"].value == "請求書"
        assert sh["A1"].font.bold is True, "雛形の太字（A1）が複製で失われた"
        assert sh["B5"].number_format == "#,##0", (
            f"雛形の数値書式が壊れた: {sh['B5'].number_format}")

        sh2 = wb2["みどり工業"]
        assert sh2["B3"].value == "みどり工業"
        assert sh2["B5"].value == 8500
        wb2.close()
    finally:
        # ★ 後片付け: basrun.py stop（接続先の LibreOffice だけを終了・taskkill 一括はしない）。
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )
