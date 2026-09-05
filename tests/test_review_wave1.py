# 帳票レビュー第一波 ── 「✓ が恒真」系の致命。修正より先に凍結した赤い検体（2026-08-24）。
# 出典: SEALED-20260824-report-review.md（盲検 4 次元 + 俺の裏取り）
#
# 共通の家系: **分母を作る関数が、計画側と検証側で同じ。** だから「N 行 → N 枚」が
# 常に成立してしまう。恒真は「間違いを見逃す」より悪い ── **正しさを主張する**。
#
# 契約:
#   ① A 列が途中で空いても、データ行の数え方が縮んだなら**その事実を名指しする**
#      （黙って分母を縮めない）
#   ② 1 セルに印が 2 つある雛形は、埋める前に断る（2 つ目が生で顧客の紙に出る）
#   ③ ふりがな（rPh）を本文に混ぜない
#   ④ 雛形の比較対象が 0 セルなら「無変更」を主張しない

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import report_per_row as rpr, xml_readback  # noqa: E402


# --- ③ ふりがな（rPh）の混入 ★確認済み: xml_readback→山田太郎ヤマダタロウ / openpyxl→山田太郎 ---

def _shared_string_book(tmp_path, with_rph=True):
    """本物の Excel の形（sharedStrings + rPh）を作る。openpyxl は inlineStr で書くので手で組む。"""
    import zipfile, re
    src = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"]); ws.append(["山田太郎", 1000])
    wb.save(src)
    dst = tmp_path / "shared.xlsx"
    rph = ('<rPh sb="0" eb="4"><t>ヤマダタロウ</t></rPh><phoneticPr fontId="1"/>'
            if with_rph else "")
    shared = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
              '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="3" uniqueCount="3">'
              '<si><t>取引先</t></si><si><t>金額</t></si>'
              f'<si><t>山田太郎</t>{rph}</si></sst>')
    idx = {"A1": 0, "B1": 1, "A2": 2}
    with zipfile.ZipFile(src) as zi, zipfile.ZipFile(dst, "w") as zo:
        for it in zi.infolist():
            data = zi.read(it.filename).decode("utf-8")
            if it.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(r'<c r="(A1|B1|A2)" t="inlineStr">.*?</c>',
                               lambda m: f'<c r="{m.group(1)}" t="s"><v>{idx[m.group(1)]}</v></c>', data)
            if it.filename == "[Content_Types].xml":
                data = data.replace('</Types>', '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/></Types>')
            if it.filename == "xl/_rels/workbook.xml.rels":
                data = data.replace('</Relationships>', '<Relationship Id="rIdSS" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/></Relationships>')
            zo.writestr(it, data.encode("utf-8"))
        zo.writestr("xl/sharedStrings.xml", shared.encode("utf-8"))
    return dst


def test_furigana_is_not_mixed_into_the_value(tmp_path):
    """★確認済みの事故: 読み仮名が本文に連結され、しかも declared 側も同じ読みなので
       export-csv の照合は恒真で ✓ が出ていた。"""
    book = _shared_string_book(tmp_path)
    grid = xml_readback.read_grid(book, "売上")["grid"]
    assert grid[(2, 1)] == "山田太郎", f"ふりがなが本文に混入した: {grid[(2, 1)]!r}"
    assert grid[(2, 1)] == openpyxl.load_workbook(book)["売上"].cell(2, 1).value, \
        "openpyxl と読みが食い違う（下流の照合が全部ずれる）"


def test_rich_text_runs_are_still_joined(tmp_path):
    """誤爆防止: リッチテキスト（<r><t>）の分割は今までどおり連結する。"""
    book = _shared_string_book(tmp_path, with_rph=False)
    assert xml_readback.read_grid(book, "売上")["grid"][(2, 1)] == "山田太郎"


# --- ② 1 セルに印が 2 つ ★確認済み: 拾えた印は 1 個だけ ---

def test_two_placeholders_in_one_cell_are_detected(tmp_path):
    """`{{取引先}} 様（担当: {{担当者}}）` ── 2 つ目を黙って生のまま残さない。"""
    p = tmp_path / "tpl.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "雛形"
    ws["B5"] = "{{取引先}} 様（担当: {{担当者}}）"
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["雛形"]
    found = rpr.scan_placeholders(ws2, 10, 10)
    names = [f.column_name for f in found]
    assert "担当者" in names, f"2 つ目の印を見落とした（顧客の紙に生で出る）: {names}"


# --- ① 分母が縮む ★確認済み: 5 行の表が 3 行 ---

def test_blank_cell_in_first_column_no_longer_shrinks_the_denominator(tmp_path):
    """★★ 盲検（2026-08-24）の①「5 行の表が 3 行と数えられる」── **根治した**。

    ★ 当時の姿: `_scan_last_row` が A 列を上から見て最初の空で止まるため、
      取引先の名前が 1 行空いているだけで分母が縮み、「データ N 行 → 出力 N 枚」の
      完全会計が**恒真**になっていた。当時は数え方を変える影響が読めなかったので
      「縮んだ事実を必ず言う」（detect_first_column_gap）に留めた。
    ★ 2026-09-05（段B）で数え方そのものを直した ── 走査は「表の幅のどこかに値が
      在れば行」になり、1 列目が空でも届く。**分母はもう縮まない。**
    ★ この試験は「言うこと」から「**縮まないこと**」へ役目が変わる。
      器官が働く形（走査が本当に止まる形）は
      tests/test_review_wave3.py が別に証明している。
    """
    p = tmp_path / "gap.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"])
    for name, amt in [("甲", 1), ("乙", 2), (None, 3), ("丁", 4), ("戊", 5)]:
        ws.append([name, amt])
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["売上"]
    scanned = ailine._scan_last_row(ws2, header_row=1)
    assert scanned == 6, f"5 行の表が {scanned - 1} 行と数えられている（分母が縮んだ）"
    assert hasattr(ailine, "detect_first_column_gap"), "分母の縮みを名指しする器が無い"
    assert ailine.detect_first_column_gap(ws2, header_row=1) is None,         "届いているのに『縮んだ』と言っている（開示が古い前提のまま）"


def test_no_gap_report_when_the_first_column_is_dense(tmp_path):
    """誤爆防止: 隙間の無い表では何も言わない。"""
    p = tmp_path / "dense.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"])
    for n, a in [("甲", 1), ("乙", 2)]:
        ws.append([n, a])
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["売上"]
    assert ailine.detect_first_column_gap(ws2, header_row=1) is None
