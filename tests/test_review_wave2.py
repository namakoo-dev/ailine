# 帳票レビュー第二波 ── 書き出し系の致命。修正より先に凍結（2026-08-24）。
# 出典: SEALED-20260824-report-review.md
#
# 契約:
#   ① `export-pdf --sheet` が効く（指定シートだけが PDF になる）
#      ★ 事故の形: 50 社分の請求書シートを作った後に 1 社を指定すると、**全 50 社＋元データ**
#        が 1 つの PDF に入るのに ✓ が出る ── 他人の売上を同封して顧客に送る
#   ② PDF 照合の分母を開示し、**数式セル（キャッシュ値なし）を黙って落とさない**
#   ③ `export-csv --out` があり、既存ファイルを黙って上書きしない
#   ④ セル分割の検算が列ずれを見逃さない（空断片を両辺から落とさない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import pdf_export, split_cell  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


# --- ④ 分割の列ずれ（純ロジック・実装前から測れる）--------------------------------

def test_rejoin_catches_a_column_shift():
    """`a,,b` の 2 つ目を列2に詰めたら不一致として捕まえる（★ 実測では見逃していた）。"""
    r = split_cell.verify_rejoin(["http://a,,http://b"], [["http://a", "http://b", ""]], ",")
    assert r.mismatched, "空断片を両辺から落として列ずれを見逃した（納品物が列ずれする）"


def test_rejoin_still_passes_a_faithful_split_with_empties():
    """誤爆防止: 位置どおりに割ってあれば（空も空の位置に）一致。"""
    r = split_cell.verify_rejoin(["http://a,,http://b"], [["http://a", "", "http://b"]], ",")
    assert not r.mismatched, f"正しい分割を不一致にした: {r.mismatched}"


# --- ② PDF 照合の分母 -------------------------------------------------------------

def test_pdf_check_reports_uncheckable_cells(tmp_path):
    """数式セル（キャッシュ値なし）は「見なかった」として数え、分母から黙って消さない。"""
    assert hasattr(pdf_export, "PdfCheck")
    r = pdf_export.PdfCheck()
    assert hasattr(r, "uncheckable"), "検証できなかったセルを数える枠が無い（分母の開示）"


needs_sheet = pytest.mark.xfail(
    "sheet_index" not in (ailine._soffice_to_pdf.__doc__ or ""),
    reason="export-pdf のシート指定 未実装（契約は凍結済み）", strict=True)


@needs_sheet
def test_export_pdf_only_the_named_sheet(tmp_path, monkeypatch, capsys):
    """① 指定シートだけが PDF になる（他社の請求書を同封しない）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    a = wb.active; a.title = "甲社"; a.append(["取引先", "金額"]); a.append(["甲社", 100])
    b = wb.create_sheet("乙社"); b.append(["取引先", "金額"]); b.append(["乙社", 999])
    wb.save(p)
    seen = {}

    def fake(book_path, out_path, sheet=None, orientation=None, fit_to_width=False):
        # ★ 治具の訂正（封印者ナギ・2026-08-24）: 初版は渡されたパスを控えて後から開いたが、
        #   一時ディレクトリは呼び出し側が片付けるので消えていた。**中で**中身を控える。
        #   assert の意図（指定しなかったシートを PDF 化しない）は不変。
        seen["sheets"] = openpyxl.load_workbook(book_path).sheetnames
        Path(out_path).write_bytes(b"%PDF-1.4 fake")
        return True, ""
    monkeypatch.setattr(ailine, "_soffice_to_pdf", fake)
    monkeypatch.setattr(pdf_export, "readback_available", lambda: True)
    monkeypatch.setattr(pdf_export, "read_pdf_text", lambda q: "取引先 金額 甲社 100")
    rc, out = _run_main(["export-pdf", str(p), "--sheet", "甲社"], capsys)
    assert rc == 0, out
    got = seen["sheets"]
    assert got == ["甲社"], f"指定しなかったシートまで PDF 化しようとした（他社の情報が同封される）: {got}"


# --- ③ export-csv --out ------------------------------------------------------------

needs_out = pytest.mark.xfail(
    not hasattr(ailine, "_export_csv_out_path"),
    reason="export-csv --out 未実装（契約は凍結済み）", strict=True)


@needs_out
def test_export_csv_refuses_to_overwrite_silently(tmp_path, monkeypatch, capsys):
    """既にある CSV を黙って潰さない（`export-pdf` には --out があるのに非対称だった）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "syuunou.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "出納"
    ws.append(["日付", "金額"]); ws.append(["2026-04-01", 100])
    wb.save(p)
    existing = p.with_suffix(".csv")
    existing.write_text("先に人が置いたファイル", encoding="utf-8")
    rc, out = _run_main(["export-csv", str(p), "--sheet", "出納"], capsys)
    assert rc != 0, f"既存の CSV を黙って上書きした: {out}"
    assert existing.read_text(encoding="utf-8") == "先に人が置いたファイル", "中身が消えた"
    assert "--out" in out or "上書き" in out, f"逃げ道を案内していない: {out}"


# --- ⑤ export-csv は**ディスク上のファイル**を読み戻す（2026-08-24）------------------
#
# ★ 実測（盲検の契約レビュー）: 読み戻していたのはメモリ上の raw_bytes で、
#   `売上.csv` 自体は一度も読んでいなかった。「1 セルも変えずに書いた」は
#   **バイト列についての主張**で、書き込みが途中で切れても ✓ が出る形だった。

def test_export_csv_reads_the_file_back_from_disk(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "出納"
    ws.append(["日付", "金額"]); ws.append(["2026-04-01", 100])
    wb.save(p)
    out = p.with_suffix(".csv")
    real_open = open
    state = {"written": False}

    def truncating_open(file, mode="r", *a, **k):
        f = real_open(file, mode, *a, **k)
        if str(file) == str(out) and "w" in mode:
            state["written"] = True
            class _Half:
                def write(self, data):
                    return f.write(data[:len(data) // 2])   # ★ 途中で切れた書き込みを模す
                def __enter__(self): return self
                def __exit__(self, *e): f.close(); return False
            return _Half()
        return f
    monkeypatch.setattr("builtins.open", truncating_open)
    rc, txt = _run_main(["export-csv", str(p), "--sheet", "出納"], capsys)
    assert state["written"], "前提: 書き込みが起きること"
    assert "✓" not in txt, f"中身が半分しか書けていないのに ✓ を名乗った: {txt}"
