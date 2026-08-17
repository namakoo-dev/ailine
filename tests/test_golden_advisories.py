"""C1-F4: build_advisories(task, before, after, exclude_sheets) / _structural_advisories
(before, after) の list[str] を凍結する。

対象: ①幽霊データ ②一様埋め ③件数の突き合わせ ⑤新規シートの中身
（new_sheet_advisories 経由） ⑥依頼にないシート新設の申告 ⑦既存シートの中身の丸ごと
すり替え ④依頼文言との重なり（mention_overlap_advisory・exclude_sheets 込み）を収載する。
neutralize 系（_neutralize_new_column_ghost_warning 等）は build_advisories 自身の
戻り値には含まれない（呼び出し側 cmd_run_dsl 等が別途後処理として適用する）ため対象外。

before/after は ailine.snapshot() の生 dict をそのままゴールデンに含める
（brief 指定: 「before/after snapshot は JSON で保存できる dict」）。1 ケース1ファイル
（tests/golden/f4_advisories/<name>.json）に {"task", "exclude_sheets", "before",
"after", "advisories"} をまとめる。
"""
import dataclasses
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Alignment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402
from ailine_core.target_sheet import SheetNameConflict  # noqa: E402

from golden._harness import GOLDEN_ROOT, assert_golden_json, sorted_list  # noqa: E402

F4_DIR = GOLDEN_ROOT / "f4_advisories"


def _snap(tmp_path, filename, sheets: dict, extra=None) -> dict:
    """sheets: {シート名: [[行の値...], ...]}。extra(ws_by_name) で個々のセルに
       スタイルを追加で当てられる（呼び出し側が openpyxl の Worksheet を直接いじる）。"""
    p = tmp_path / filename
    wb = openpyxl.Workbook()
    first = True
    ws_by_name = {}
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        first = False
        for row in rows:
            ws.append(row)
        ws_by_name[name] = ws
    if extra:
        extra(ws_by_name)
    wb.save(p)
    return ailine.snapshot(p)


# 名前 -> (task, before_snap_fn(tmp_path), after_snap_fn(tmp_path), exclude_sheets,
#         use_structural_only, op, resolved, meta, sheet_conflict)
# ★ 単位C: op/resolved/meta は「その op が何を書くと宣言しているか」(OP_WRITE_TARGET) を
#   助言側に渡す経路。渡さないケース（既存の全ケース）は従来どおり宣言なしで評価される
#   ＝ゴールデンのバイト列も payload のキー構成も変わらない。
# ★ 誤爆#3: sheet_conflict は「対象シートを決めた側(resolve_target_sheet)が、その語は
#   列名とも一致するので曖昧と判断して既定へ後退した」という判定結果を運ぶ経路（同上・
#   渡さないケースは payload のキー構成もバイト列も変わらない）。
CASES: dict = {}


def _add(name, task, before_fn, after_fn, exclude_sheets=None, structural_only=False,
          op=None, resolved=None, meta=None, sheet_conflict=None):
    assert name not in CASES, f"重複した case 名: {name}"
    CASES[name] = (task, before_fn, after_fn, exclude_sheets, structural_only, op, resolved,
                   meta, sheet_conflict)


# --- 基準: 何も疑わしくない通常の値変更（複数列にまたがる＝③件数突き合わせの対象外にもなる） ---
_add("baseline_no_advisories", "商品名と金額を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額"], ["a2", 150], ["b", 200]]}))

# --- ①幽霊データ ---------------------------------------------------------
_add("ghost_data_outside_used_range", "金額を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]},
                       extra=lambda wb: wb["Sheet"].cell(row=20, column=26, value="幽霊")))
_add("ghost_data_not_triggered_when_inside_range", "金額を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額"], ["a", 999], ["b", 200]]}))

# --- ②一様埋め -----------------------------------------------------------
_add("uniform_fill_zero_into_blank_cells", "備考を埋めて",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "備考"], ["a", None], ["b", None]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "備考"], ["a", 0], ["b", 0]]}))
_add("uniform_fill_not_triggered_when_values_differ", "備考を埋めて",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "備考"], ["a", None], ["b", None]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "備考"], ["a", "x"], ["b", "y"]]}))

# --- ③件数の突き合わせ ------------------------------------------------------
_add("count_reconciliation_partial_column_change", "金額列を更新して",
     lambda tp: _snap(tp, "before.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 150], ["b", 200], ["c", 300]]}))
_add("count_reconciliation_header_also_changed", "金額列を更新して",
     lambda tp: _snap(tp, "before.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "税込金額"], ["a", 150], ["b", 200]]}))

# --- ⑤新規シートの中身（new_sheet_advisories 経由） ---------------------------
# ★ 見出し文字列を含む普通の SummaryTable 形（見出しが値と異なる）は一様埋めとして
#   誤検知しない（見出し込みで全部が「同じ値」でないと発火しない仕様）ことも併せて
#   捉えるため、あえて見出し無し・全セル同一値0の新規シートで発火させる。
_add("new_sheet_uniform_fill_detected", "0で初期化したシートを追加して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]],
                        "新規": [[0, 0], [0, 0]]}))
_add("new_sheet_realistic_summary_table_shape_not_flagged_as_uniform", "部門ごとに集計して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["部門", "金額"], ["A", 100], ["B", 200]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["部門", "金額"], ["A", 100], ["B", 200]],
                        "集計": [["部門", "合計"], ["A", 0], ["B", 0]]}))

# --- ⑥依頼にないシート新設の申告 ------------------------------------------------
_add("unrequested_new_sheet_warns_without_mention", "金額を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100]], "新規": [["x", 1]]}))
_add("new_sheet_silent_when_mentioned", "新しいシートに集計を出して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100]], "新規": [["x", 1]]}))

# --- ⑦既存シートの中身の丸ごとすり替え -----------------------------------------
_add("existing_sheet_fully_replaced_warns", "集計を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["a", 1]], "集計": [["A", 100], ["B", 200]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["a", 1]], "集計": [["X", 999], ["Y", 888]]}))
_add("existing_sheet_partial_update_silent", "集計を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["a", 1]], "集計": [["A", 100], ["B", 200]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["a", 1]], "集計": [["A", 999], ["B", 200]]}))

# --- ④依頼文言との重なり（build_advisories 全体・_structural_only=False） -----------
_add("mention_overlap_column_letter_not_changed_warns", "B列を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "x"]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "y"]]}))
_add("mention_overlap_column_letter_changed_silent", "B列を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "x"]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 150, "x"]]}))
_add("mention_overlap_row_not_changed_warns", "行3を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額"], ["a", 150], ["b", 200]]}))
_add("mention_overlap_sheet_not_changed_warns", "参照シートも更新して",
     lambda tp: _snap(tp, "before.xlsx",
                       {"Sheet": [["a", 1]], "参照シート": [["x", 1]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["a", 2]], "参照シート": [["x", 1]]}))
_add("mention_overlap_sheet_excluded_silent", "参照シートも更新して",
     lambda tp: _snap(tp, "before.xlsx",
                       {"Sheet": [["a", 1]], "参照シート": [["x", 1]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["a", 2]], "参照シート": [["x", 1]]}),
     exclude_sheets={"参照シート"})

# --- ④依頼文言との重なり（数字表記の列・0起点/1起点どちらの解釈でも許す） -----------
_add("mention_overlap_digit_column_not_changed_warns", "列2を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "x"]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a2", 100, "x"]]}))
_add("mention_overlap_digit_column_changed_silent", "列2を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "x"]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 150, "x"]]}))

# --- _structural_advisories 単体（build_advisories との差分= ④が含まれないこと） -----
_add("structural_only_excludes_mention_overlap", "列B を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "x"]]}),
     lambda tp: _snap(tp, "after.xlsx", {"Sheet": [["商品", "金額", "備考"], ["a", 100, "y"]]}),
     structural_only=True)

# --- ★ 単位C: op の宣言（OP_WRITE_TARGET の writes / reads_only）を読む中立化 ---------
# D10: APPEND_TOTAL の合計行は「データ末尾の新規行」＝定義上ずっと原本の使用範囲の外に出る。
#   宣言を渡さない場合（下の _no_declaration 版）は従来どおり
#   「★ 疑わしい: 変更が元データの範囲外です（A4:B4）」が出る ── 同じ before/after で
#   宣言の有無だけが結果を分けることをゴールデンで対にして凍結する。
def _append_total_before(tp):
    return _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]})


def _append_total_after(tp):
    return _snap(tp, "after.xlsx",
                  {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200], ["合計", 300]]})


_add("append_total_new_row_at_end_neutralized_by_declaration", "金額の合計を出して",
     _append_total_before, _append_total_after, structural_only=True,
     op="APPEND_TOTAL", resolved={"col": "金額", "label": "合計", "_target_sheet": "Sheet"},
     meta={"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]},
            "header_rows": {"Sheet": 1}})
_add("append_total_new_row_at_end_no_declaration_still_warns", "金額の合計を出して",
     _append_total_before, _append_total_after, structural_only=True)

# D8: AGGREGATE は新規シートを作るだけで、名指しされた入力シートは読むだけ＝無変更が正常。
#   reads_only の宣言（_target_sheet）を渡すと「★ …は変更されていません」を言わない。
def _aggregate_before(tp):
    return _snap(tp, "before.xlsx", {"工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]]})


def _aggregate_after(tp):
    return _snap(tp, "after.xlsx",
                  {"工事台帳": [["取引先", "金額"], ["a", 100], ["b", 200]],
                   "集計": [["取引先", "合計"], ["a", 100], ["b", 200]]})


_add("aggregate_reads_only_input_sheet_silent_by_declaration", "工事台帳を取引先ごとに集計して",
     _aggregate_before, _aggregate_after,
     op="AGGREGATE", resolved={"group_col": "取引先", "value_col": "金額", "_target_sheet": "工事台帳"},
     meta={"sheets": ["工事台帳"], "headers": {"工事台帳": ["取引先", "金額"]},
            "header_rows": {"工事台帳": 1}})
_add("aggregate_reads_only_input_sheet_no_declaration_warns", "工事台帳を取引先ごとに集計して",
     _aggregate_before, _aggregate_after)


# --- ★ 誤爆#3: シート名と列名が衝突した語は「シート言及」から外す ---------------------
# 実測の型: sheets=['売上データ','金額'] に「金額を降順に並べ替えて」。resolve_target_sheet は
#   『金額』が列名でもあることを見て「曖昧＝既定(1枚目)へ後退」と**既に決めている**のに、
#   助言側はそれを読まず「『金額』シートは変更されていない」と誤爆していた。
#   衝突の記録を渡した時だけ黙る ── 同じ before/after で判定の有無だけが結果を分けることを
#   ゴールデンで対にして凍結する（単位C の append_total/aggregate の対と同じ作法）。
def _sheet_conflict_before(tp):
    return _snap(tp, "before.xlsx",
                  {"売上データ": [["商品", "金額"], ["a", 200], ["b", 300]],
                   "金額": [["月", "金額"], ["1月", 50]]})


def _sheet_conflict_after(tp):
    """1枚目だけを降順に並べ替えた後（2枚目『金額』シートは読んでも触ってもいない）。"""
    return _snap(tp, "after.xlsx",
                  {"売上データ": [["商品", "金額"], ["b", 300], ["a", 200]],
                   "金額": [["月", "金額"], ["1月", 50]]})


_add("sheet_name_conflict_mention_silent_by_conflict", "金額を降順に並べ替えて",
     _sheet_conflict_before, _sheet_conflict_after,
     sheet_conflict=SheetNameConflict(word="金額", alternative="金額", chosen="売上データ"))
_add("sheet_name_conflict_mention_no_conflict_warns", "金額を降順に並べ替えて",
     _sheet_conflict_before, _sheet_conflict_after)


# --- 複合: 幽霊データ + 一様埋め + 件数突き合わせが同時に出るケース -----------------
# ★ 全変更セルが単一列(J)・原本範囲外・空欄→同一値、の3条件を同時に満たすように
#   仕組む（ghost は「全部」範囲外の時だけ発火する設計なので、範囲内の変更は混ぜない）。
_add("multiple_advisories_at_once", "金額を更新して",
     lambda tp: _snap(tp, "before.xlsx", {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]}),
     lambda tp: _snap(tp, "after.xlsx",
                       {"Sheet": [["商品", "金額"], ["a", 100], ["b", 200]]},
                       extra=lambda wb: (wb["Sheet"].cell(row=30, column=10, value="外側"),
                                         wb["Sheet"].cell(row=31, column=10, value="外側"))))


def _case_ids():
    return sorted(CASES.keys())


@pytest.mark.parametrize("name", _case_ids())
def test_build_advisories_golden(tmp_path, name):
    (task, before_fn, after_fn, exclude_sheets, structural_only, op, resolved, meta,
     sheet_conflict) = CASES[name]
    before = before_fn(tmp_path)
    after = after_fn(tmp_path)
    if structural_only:
        advisories = ailine._structural_advisories(before, after, op=op, resolved=resolved, meta=meta)
    else:
        advisories = ailine.build_advisories(task, before, after, exclude_sheets,
                                              op=op, resolved=resolved, meta=meta,
                                              sheet_conflict=sheet_conflict)
    payload = {
        "task": task,
        "exclude_sheets": sorted_list(exclude_sheets) if exclude_sheets else [],
        "before": before,
        "after": after,
        "advisories": advisories,
    }
    if op is not None:   # ★ 単位C: 宣言を渡すケースだけ payload に足す（既存ゴールデンは不変）
        payload["op"] = op
        payload["resolved"] = resolved
        payload["meta"] = meta
    if sheet_conflict is not None:   # ★ 誤爆#3: 衝突を渡すケースだけ足す（同上）
        payload["sheet_conflict"] = dataclasses.asdict(sheet_conflict)
    assert_golden_json(F4_DIR / f"{name}.json", payload, label=name)
