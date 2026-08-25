# 復元の致命2（2026-08-24 の盲検）── 隣にある利用者のファイルを無言で消す。
#
# ★ 実測: `<book>.out.xlsx` は作業ファイル名として**存在確認も警告もなく**上書きされ、
#   原本反映が成功すると unlink される。利用者が自分で作った `売上.out.xlsx` は
#   一言も無く消えた（`--copy` の時は中身だけ上書きされ、しかも
#   「（原本 売上.xlsx は変更していません）」と表示される ── 別の原本は破壊済み）。
# ★ フォルダ経路には同じ危険への関所（_refuse_output_conflict・exit 7・
#   「ailine の印が無い人のファイルです」）が**既に在る**のに、単一ブック経路に
#   配線されていなかった ── 片配線。
#
# 契約:
#   ① 人のファイルが出力先に在れば、**触る前に**止める（exit 7）
#   ② ailine 産（前回の .out）なら従来どおり黙って作り直す
#   ③ 出力先が空いていれば 1 文字も増えない
#   ④ `.out` の場所を決める実装は 1 つ（4 箇所の書き写しを畳む）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _book(tmp_path, name="売上.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"]); ws.append(["a", 100]); ws.append(["b", 250])
    wb.save(p)
    return p


def test_out_path_has_one_implementation():
    """④ 4 箇所が同じ形を書き写していた ── 畳んだことを構造で縛る。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    assert src.count('with_name(book.stem + ".out"') <= 1, \
        ".out の場所を決める式が複数ある（書き写し）"
    assert "def out_book_path(" in src, "共通の実装が無い"


def test_refuses_when_a_human_file_sits_at_the_output_path(tmp_path, monkeypatch, capsys):
    """① 人のファイルは触る前に守る。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    mine = tmp_path / "売上.out.xlsx"
    wb = openpyxl.Workbook(); wb.active["A1"] = "私の大事なメモ"; wb.save(mine)
    before = mine.read_bytes()
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 7, f"人のファイルを守らなかった: exit={rc} / {out}"
    assert mine.read_bytes() == before, "人のファイルが書き換わった"
    assert "人のファイル" in out or "書けません" in out, out


def test_own_previous_output_is_rebuilt_silently(tmp_path, monkeypatch, capsys):
    """② 前回の .out（ailine 産）は従来どおり作り直す ── 誤爆で使えなくしない。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "b", 250, "a", 100
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc1, _ = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc1 == 0
    rc2, out2 = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc2 == 0, f"自分の前回出力で止まった（誤爆）: {out2}"
