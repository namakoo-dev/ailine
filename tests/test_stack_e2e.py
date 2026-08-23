"""M1書き `ailine stack <folder> --out <path>`（縦積み・UNION ALL）の検体。
   ★ 実装前に書いた赤い検体（DESIGN-20260821-multifile v2/v2.1・specimen-first）。

   契約の要点:
   - 基準ファイル方式で照合できたファイルだけを、基準の列順に揃えて新ブックへ縦積み
   - 出所列（『元ファイル』『元行』）を右端に付ける ── 抜き打ち検証の装置 + 自分の出力の署名
   - 単位L: 合計行は積まない（除外は開示・閉じる検査の不一致は両側の数字つき ⚠）
   - 事後条件: Σ行数一致・数値列ごとの Σ 両側表示・原本無変更・分母つき報告
   - 出力先が既存の人のファイルなら止まる（exit 7・--overwrite で通す）。
     自分の前回の縦積み出力なら作り直して開示（単位H の作法）
   - 決定論: 同一入力で 2 回走らせて、セル内容が完全一致（zip バイトではなく中身）"""
import hashlib
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent

HDRS = ["注文ID", "取引先", "金額"]


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        padded = list(r) + [None] * (len(headers) - len(r))
        ws.append(padded[: len(headers)])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _stack(folder, out, *extra):
    return subprocess.run(
        [sys.executable, "-m", "ailine", "stack", str(folder), "--out", str(out), *extra],
        capture_output=True, text=True, timeout=180, encoding="utf-8")


def _cells(path):
    """出力の中身（シート名・セル値・型）を決定論比較用に正規化して返す。"""
    wb = openpyxl.load_workbook(path)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    out.append((ws.title, c.coordinate, type(c.value).__name__, str(c.value)))
    wb.close()
    return out


def _folder3(tmp_path):
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100), ("J-2", "乙", 200)])
    _book(folder / "b.xlsx", ["金額", "注文ID", "取引先"], [(300, "J-3", "丙")])   # 並びだけ違う
    _book(folder / "c.xlsx", HDRS, [("J-4", "丁", 50), ("合計", None, 350)])       # ★ 合計行入り
    return folder


def test_stack_unions_rows_in_base_column_order_with_provenance(tmp_path):
    """本命: 3 冊 → 1 冊。基準（a.xlsx）の列順に揃い、出所列が右端に付く。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stderr[-800:] + p.stdout[-800:]
    ws = openpyxl.load_workbook(out).active
    headers = [c.value for c in ws[1]]
    assert headers == HDRS + ["元ファイル", "元行"], headers
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert ["J-3", "丙", 300, "b.xlsx", 2] in rows, "並べ替え縦積みが崩れている"
    assert len(rows) == 4, f"データ行 4 のはず（合計行は積まない）: {len(rows)}"


def test_total_row_is_not_stacked_and_exclusion_is_disclosed(tmp_path):
    """単位L の配線: c.xlsx の合計行は積まれず、その事実がファイル名つきで開示される。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    assert "合計行" in p.stdout and "c.xlsx" in p.stdout, f"除外の開示が無い:\n{p.stdout}"
    ws = openpyxl.load_workbook(out).active
    labels = [r[0].value for r in ws.iter_rows(min_row=2)]
    assert "合計" not in labels


def test_report_shows_both_sides_of_row_and_sum_reconciliation(tmp_path):
    """事後条件①②: 行数と数値列の合計を 両側の数字 で並べる（「一致」だけの報告は感想）。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    assert "3 ファイル中 3 積んだ" in p.stdout, f"分母つき報告が無い:\n{p.stdout}"
    # 金額の総和 = 100+200+300+50 = 650 が「元」「出力」の両側で出る
    assert p.stdout.count("650") >= 2, f"Σ金額の両側表示が無い:\n{p.stdout}"


def test_unmatched_file_is_skipped_named_and_counted(tmp_path):
    """列が欠けたファイルは積まず、名指し + 分母に現れる。exit は 0（黙る失敗だけが罪）。"""
    folder = _folder3(tmp_path)
    _book(folder / "d.xlsx", ["注文ID", "取引先"], [("J-9", "戊")])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    assert "4 ファイル中 3 積んだ" in p.stdout
    assert "d.xlsx" in p.stdout and "金額" in p.stdout


def test_originals_unchanged_and_output_outside_folder(tmp_path):
    """事後条件④: 原本は 1 バイトも変わらない（sha256）。"""
    folder = _folder3(tmp_path)
    before = {p.name: hashlib.sha256(p.read_bytes()).hexdigest() for p in folder.glob("*.xlsx")}
    out = tmp_path / "out.xlsx"
    r = _stack(folder, out)
    assert r.returncode == 0
    after = {p.name: hashlib.sha256(p.read_bytes()).hexdigest() for p in folder.glob("*.xlsx")}
    assert before == after


def test_deterministic_content_across_two_runs(tmp_path):
    """事後条件⑥: 同一入力で 2 回 → セル内容（シート名・座標・型・値）が完全一致。
       ★ zip バイト一致は openpyxl の保存時刻で原理的に不能（実測）── 中身で比べる。"""
    folder = _folder3(tmp_path)
    out1, out2 = tmp_path / "o1.xlsx", tmp_path / "o2.xlsx"
    assert _stack(folder, out1).returncode == 0
    assert _stack(folder, out2).returncode == 0
    assert _cells(out1) == _cells(out2)


def test_existing_foreign_output_is_a_gate_not_a_casualty(tmp_path):
    """writes=(new_book,) の関所: 出力先に人のファイルが居たら止まる（exit 7・中身無傷）。
       --overwrite で通す（既存の overwrite 作法と同じ）。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "mine.xlsx"
    _book(out, ["大事な列"], [("消えては困る",)])
    sha = hashlib.sha256(out.read_bytes()).hexdigest()
    p = _stack(folder, out)
    assert p.returncode == 7, f"exit={p.returncode}\n{p.stdout}\n{p.stderr}"
    assert hashlib.sha256(out.read_bytes()).hexdigest() == sha, "関所で止まったのに書き換えた"
    p2 = _stack(folder, out, "--overwrite")
    assert p2.returncode == 0


def test_own_previous_output_is_rebuilt_with_disclosure(tmp_path):
    """出力先が自分の前回の縦積み（署名 = 右端の出所列 2 本）なら作り直し + 開示（単位H の作法）。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "out.xlsx"
    assert _stack(folder, out).returncode == 0
    p = _stack(folder, out)   # 2 回目: 自分の出力の上に
    assert p.returncode == 0, f"自分の前回出力で関所が閉まった:\n{p.stdout}\n{p.stderr}"
    assert "前回" in p.stdout or "作り直" in p.stdout, "作り直しの開示が無い"


def test_own_output_inside_input_folder_is_excluded_from_sources(tmp_path):
    """★ 敵対検体（V6・自己参照）: 出力を入力フォルダ内に置いて 2 回走らせても
       行数が増えない（自分の出力を入力から署名で除外 + 開示）。"""
    folder = _folder3(tmp_path)
    out = folder / "まとめ.xlsx"
    assert _stack(folder, out).returncode == 0
    n1 = openpyxl.load_workbook(out).active.max_row
    p = _stack(folder, out)
    assert p.returncode == 0
    n2 = openpyxl.load_workbook(out).active.max_row
    assert n1 == n2, f"2 週目に自分の出力を食って {n1}→{n2} 行"
    assert "除外" in p.stdout or "自分の出力" in p.stdout, "自己参照除外の開示が無い"


def test_mismatched_total_row_warns_with_both_numbers(tmp_path):
    """単位L の閉じる検査: 合計行の値が明細の和と合わないファイルは 両側の数字つき ⚠。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 600), ("J-2", "乙", 400), ("合計", None, 1200)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    assert "1200" in p.stdout and "1000" in p.stdout, f"両側の数字が無い:\n{p.stdout}"


def test_json_carries_denominator_per_file_and_totals(tmp_path):
    """--json: 分母・ファイルごとの積んだ行数・除外合計行数・Σ の両側が機械可読で出る。"""
    folder = _folder3(tmp_path)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    assert data["denominator"] == 3 and data["stacked_files"] == 3
    assert data["rows_written"] == 4
    per = {f["name"]: f for f in data["files"]}
    assert per["c.xlsx"]["rows_stacked"] == 1 and per["c.xlsx"]["total_rows_excluded"] == 1
    assert data["sums"]["金額"]["source"] == 650 and data["sums"]["金額"]["output"] == 650


def test_data_row_with_empty_numeric_cell_is_still_stacked(tmp_path):
    """★ 地雷の先回り（total_row 検分 2026-08-21）: adopted_rows は数値行だけを数える
       （閉じる検査の分母としては正しい）。だが積む対象は『除外行以外の全データ行』──
       金額が空欄のデータ行を adopted_rows 基準で黙って落とさないこと。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 100), ("J-2", "乙", None), ("合計", None, 100)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stderr[-500:]
    ws = openpyxl.load_workbook(out).active
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    assert any(r[0] == "J-2" for r in rows), "金額が空のデータ行が黙って消えた"
    assert len(rows) == 2, f"データ行 2（J-1, J-2）のはず: {len(rows)}"


def test_suffixed_provenance_output_is_still_recognized_as_own(tmp_path):
    """★ 凍結予測③が的中した穴（2026-08-21 06:1x 実機で確認）: 基準が『元ファイル』列を
       持つと出所列は 元ファイル_2 になるが、①その事実の開示が無く ②自分の前回出力の
       署名判定が素の列名しか見ず exit 7 で誤って閉まった。署名はサフィックス形も自分と
       認識し、衝突の開示も 1 行出すこと。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", ["注文ID", "元ファイル", "金額"], [("J-1", "memo1", 100)])
    _book(folder / "b.xlsx", ["注文ID", "元ファイル", "金額"], [("J-2", "memo2", 200)])
    out = tmp_path / "out.xlsx"
    p1 = _stack(folder, out)
    assert p1.returncode == 0
    assert "元ファイル_2" in p1.stdout, f"サフィックスの開示が無い:\n{p1.stdout}"
    p2 = _stack(folder, out)   # 2 周目: 自分のサフィックスつき出力の上に
    assert p2.returncode == 0, f"自分の出力を他人と誤認して閉まった (exit={p2.returncode})"
    assert "前回" in p2.stdout or "作り直" in p2.stdout


# ---- jisaku-review 確定 6 件の凍結（2026-08-21 06:4x・81b9aa9..bdfc4a8 への盲検レビュー）----


def test_foreign_file_with_coincident_provenance_names_is_still_gated(tmp_path):
    """★ review#1 critical: 末尾 2 列がたまたま『元ファイル』『元行』という名前の
       人のファイルを、名前だけの署名で自分の前回出力と誤認して無警告上書きした。
       署名は名前の一致だけで成立してはならない ── 人のデータは exit 7 で守る。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    out = tmp_path / "mine.xlsx"
    _book(out, HDRS + ["元ファイル", "元行"],
          [("X-1", "人の大事なデータ", 999, "note", "n/a")])
    sha = hashlib.sha256(out.read_bytes()).hexdigest()
    p = _stack(folder, out)
    assert p.returncode == 7, f"人のファイルを署名誤認で通した (exit={p.returncode})\n{p.stdout}"
    assert hashlib.sha256(out.read_bytes()).hexdigest() == sha, "人のデータが消えた"


def test_gate_refusal_discloses_reason_and_next_step(tmp_path):
    """★ review#5: exit 7 の関所が無言で閉まっていた（報告が成果物、の自己矛盾）。
       止まる時こそ 理由 + 次の手（--overwrite）を言う。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    out = tmp_path / "mine.xlsx"
    _book(out, ["大事な列"], [("消えては困る",)])
    p = _stack(folder, out)
    assert p.returncode == 7
    combined = p.stdout + p.stderr
    assert "--overwrite" in combined, f"次の手の開示が無い:\n{combined}"
    assert "mine.xlsx" in combined, "何が邪魔なのかの名指しが無い"


def test_multiple_subtotal_groups_do_not_false_alarm(tmp_path):
    """★ review#2 major: 閉じる検査が先頭からの累積和としか比べず、2 個目の小計で
       正しい表に偽 ⚠ を出した。区間和（直前の除外行から）か累積和のどちらかが
       合えば閉じる、が正しい形（総計は累積で・各小計は区間で閉じる）。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("部署A-1", "x", 100), ("部署A-2", "x", 200), ("小計", None, 300),
           ("部署B-1", "y", 50), ("部署B-2", "y", 150), ("小計", None, 200),
           ("総計", None, 500)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    assert "≠" not in p.stdout, f"正しい表に偽の不一致 ⚠:\n{p.stdout}"
    assert p.stdout.count("500") >= 2, "Σ の両側（500）が無い"
    ws = openpyxl.load_workbook(out).active
    assert ws.max_row - 1 == 4, "積む行数が違う（明細 4 行のはず）"


def test_all_numeric_columns_are_reconciled_not_just_the_first(tmp_path):
    """★ review#3/#6 major: stack の Σ 照合・報告が最初の数値列 1 本だけだった。
       契約（このファイル冒頭）は『数値列ごとの Σ 両側表示』── 数量と金額の両方が出ること。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", ["注文ID", "数量", "金額"],
          [("J-1", 2, 100), ("J-2", 3, 200)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    assert "数量" in data["sums"] and "金額" in data["sums"], f"数値列ごとの Σ が無い: {list(data['sums'])}"
    assert data["sums"]["数量"] == {"source": 5, "output": 5}
    assert data["sums"]["金額"] == {"source": 300, "output": 300}
    p2 = _stack(folder, tmp_path / "out2.xlsx")
    assert "数量" in p2.stdout and "金額" in p2.stdout, "人間向けにも両列の Σ が要る"


def test_json_carries_closure_mismatches_for_automation(tmp_path):
    """★ review#4 major (conv=3): --json に閉じる検査の不一致が載らず、自動化経路だけ
       データ不整合が見えなかった。テキストの ⚠ と同じ情報を機械可読でも出す。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 600), ("J-2", "乙", 400), ("合計", None, 1200)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    mm = data.get("mismatches")
    assert mm, f"mismatches が JSON に無い: {list(data)}"
    assert any(m.get("excluded_value") == 1200 and m.get("adopted_sum") == 1000 for m in mm), mm


# ---- P1: 書き手の印の集合化（M2 前の契約改修・architect 致命2 の凍結）----


def _mark_creator(path, mark):
    wb = openpyxl.load_workbook(path)
    wb.properties.creator = mark
    wb.save(path)


def test_other_ailine_kind_output_is_gated_with_named_reason(tmp_path):
    """★ 致命2: 将来の ailine extract 等『別コマンドの出力』に stack --out を向けたら、
       無警告の作り直しではなく exit 7 + 名指し（ailine の別コマンドの出力です）。
       作り直してよいのは印が完全一致（ailine stack 自身の前回出力）の時だけ。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    out = tmp_path / "extracted.xlsx"
    _book(out, HDRS + ["元ファイル", "元行"], [("X-1", "抽出結果", 999, "b.xlsx", 2)])
    _mark_creator(out, "ailine extract")
    sha = hashlib.sha256(out.read_bytes()).hexdigest()
    p = _stack(folder, out)
    assert p.returncode == 7, f"別コマンドの出力を黙って作り直した (exit={p.returncode})\n{p.stdout}"
    assert hashlib.sha256(out.read_bytes()).hexdigest() == sha
    assert "別のコマンド" in p.stdout or "ailine extract" in p.stdout, \
        f"別コマンドの出力である旨の名指しが無い:\n{p.stdout}"


def test_other_ailine_kind_output_inside_folder_is_excluded_from_inputs(tmp_path):
    """★ 致命2 の裏面: 入力フォルダ内の『ailine の別コマンドの出力』は、二重計上を防ぐため
       入力から除外 + 開示（is_own_output の marks 集合判定が V6 除外にも効くこと）。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    other = folder / "extracted.xlsx"
    _book(other, HDRS + ["元ファイル", "元行"], [("X-1", "抽出結果", 999, "b.xlsx", 2)])
    _mark_creator(other, "ailine extract")
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout + p.stderr[-300:]
    ws = openpyxl.load_workbook(out).active
    assert ws.max_row - 1 == 1, "別コマンドの出力を入力に食って二重計上した"
    assert "除外" in p.stdout, "除外の開示が無い"


# ---- P2: シート引き当ての整合と開示（architect 致命5 前段・stack 側の開示）----


def test_sheet_fallback_is_disclosed_in_text_and_json(tmp_path):
    """★ 致命5 前段: 基準名のシートが無いソースは1枚目へ落ちる（従来どおり無警告では
       止まらない）が、その事実は人間向け報告に1行 + --json（files 配列内・sheet_fallbacks
       両方）で開示すること。"""
    folder = tmp_path / "src"
    base = folder / "a.xlsx"
    _book(base, HDRS, [("J-1", "甲", 100)])
    wb = openpyxl.load_workbook(base)
    wb.active.title = "明細"
    wb.save(base)
    # b.xlsx には『明細』シートが無い（1枚目『集計』しかない）── 落ちて照合を試みる。
    b = folder / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"
    ws.append(HDRS)
    ws.append(["J-2", "乙", 200])
    wb.save(b)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    assert "明細" in p.stdout and "集計" in p.stdout and "b.xlsx" in p.stdout, \
        f"シート引き当てのフォールバック開示が無い:\n{p.stdout}"

    out2 = tmp_path / "out2.xlsx"
    p2 = _stack(folder, out2, "--json")
    assert p2.returncode == 0, p2.stdout
    data = json.loads(p2.stdout)
    fb = data.get("sheet_fallbacks")
    assert fb and any(f["name"] == "b.xlsx" and f["wanted"] == "明細" and f["used"] == "集計"
                       for f in fb), f"sheet_fallbacks が JSON に無い/不正: {fb}"


def test_date_cells_keep_date_number_format(tmp_path):
    """★ 実視の磨き残し（2026-08-21）: 受注日が『2026-07-09 0:00:00』と時刻付きで出る ──
       元セルの表示書式（number_format）を運んでいないため。日付セルは元の書式を
       引き継ぎ、時刻の尻尾を見せないこと。"""
    import datetime
    folder = tmp_path / "src"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["注文ID", "受注日", "金額"])
    ws.append(["J-1", datetime.date(2026, 7, 9), 100])
    ws["B2"].number_format = "yyyy/m/d"
    folder.mkdir()
    wb.save(folder / "a.xlsx")
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    ws2 = openpyxl.load_workbook(out).active
    cell = [c for c in ws2[2]][1]
    assert cell.number_format not in ("General", None), \
        f"日付セルの書式が運ばれていない: {cell.number_format!r}"
    assert "h" not in cell.number_format and ":" not in cell.number_format, \
        f"時刻の尻尾が残る書式: {cell.number_format!r}"


def test_postcondition_catches_attribution_swap_via_evaluate_and_stack_mutation(
        tmp_path, monkeypatch, capsys):
    """★ jisaku-review 4戦目 F1（major）の変異試験。cmd_run_folder は verify_extract
       経由で帰属検算（review3#3）を無償継承していたが、cmd_stack の書き込み時経路は
       自前の行数+Σ だけで帰属を見ていなかった ── レビューの実機再現: evaluate_and_stack
       の戻りを Σ 保存的に細工すると cmd_stack が exit 0 で素通りした。
       ここでは行の値を丸ごと2行分入れ替える（行数・Σ とも保存されたまま帰属だけ壊れる）
       変異を注入し、cmd_stack が非0で止まり、出力が本置き場（out）に現れないことを見る。"""
    import dataclasses

    import ailine
    from ailine_core import stack as multifile_stack

    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100), ("J-2", "乙", 200)])
    out = tmp_path / "out.xlsx"

    real = multifile_stack.evaluate_and_stack

    def mutant(*args, **kwargs):
        r = real(*args, **kwargs)
        if r.status != "積んだ" or len(r.rows) < 2:
            return r
        (v0, f0, row0), (v1, f1, row1) = r.rows[0], r.rows[1]
        swapped = [(v1, f0, row0), (v0, f1, row1)] + r.rows[2:]   # ★ Σ 保存・帰属だけ壊す
        return dataclasses.replace(r, rows=swapped)

    monkeypatch.setattr(multifile_stack, "evaluate_and_stack", mutant)
    rc = ailine.main(["stack", str(folder), "--out", str(out)])
    captured = capsys.readouterr()
    assert rc != 0, f"帰属を入れ替えたのに合格した:\n{captured.out}"
    assert "事後条件が破れた" in captured.out and "元" in captured.out and "出力" in captured.out, \
        captured.out
    assert not out.exists(), "事後条件が破れたのに出力を本置き場に書いた"


def test_date_column_width_is_not_inflated_by_time_tail(tmp_path):
    """★ jisaku-review 4戦目 F2（minor）の直し検体。autosize が number_format を無視して
       str(datetime) の生の長さ（時刻付きなら20文字超）を数えると、日付列だけ過大な幅
       （実測: 幅21）になる。日付セルは表示相当の固定幅で数え、常識的な範囲(8〜14)に
       収まること。"""
    import datetime
    folder = tmp_path / "src"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["注文ID", "受注日", "金額"])
    ws.append(["J-1", datetime.date(2026, 7, 9), 100])
    ws["B2"].number_format = "yyyy/m/d"
    folder.mkdir()
    wb.save(folder / "a.xlsx")
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    ws2 = openpyxl.load_workbook(out).active
    width = ws2.column_dimensions["B"].width
    assert width is not None and 8 <= width <= 14, f"日付列の幅が過大/未設定: {width!r}"


# ---- P: 署名の kind 別テーブル（M3 先行・architect 致命2 の凍結）----

_MATCH_HEADERS = ["キー", "A側 件数", "A側 合計", "B側 件数", "B側 合計", "差額", "状態"]


def _fake_match_output(path):
    """M3（照合）出力の形: 1 枚目が『照合』シート・固定見出し・creator 印 ailine match。
       ★ この見出し集合が match kind の列署名（契約）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "照合"
    ws.append(_MATCH_HEADERS)
    ws.append(["甲社", 1, 220000, 1, 220000, 0, "差額 0"])
    wb.properties.creator = "ailine match"
    wb.save(path)


def test_match_kind_output_in_folder_is_excluded_not_misread(tmp_path):
    """★ 致命2: 照合出力は末尾 2 列の出所署名を構造的に持てない。kind 別の列署名
       （照合シートの固定見出し + 印）で ailine 産と認識し、入力列挙から除外 + 開示。
       今は『取れなかった（見出し不一致）』の雑音として数えられてしまう。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    _fake_match_output(folder / "照合結果.xlsx")
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    assert "1 ファイル中 1 積んだ" in p.stdout, f"照合出力が分母に混ざった:\n{p.stdout}"
    assert "除外" in p.stdout and "照合結果.xlsx" in p.stdout, "ailine 産としての除外開示が無い"
    assert "取れなかった" not in p.stdout, "ailine 産を『取れなかった』の雑音にした"


def test_match_kind_output_at_gate_is_named_as_ailine_kind(tmp_path):
    """★ 致命2 の関所側: 出力先が照合出力なら『ailine の別のコマンドの出力（作成: ailine match）』
       と名指し（人のファイル、の誤認をしない）。exit 7・無傷は従来どおり。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100)])
    out = tmp_path / "照合結果.xlsx"
    _fake_match_output(out)
    sha = hashlib.sha256(out.read_bytes()).hexdigest()
    p = _stack(folder, out)
    assert p.returncode == 7
    assert hashlib.sha256(out.read_bytes()).hexdigest() == sha
    assert "ailine match" in p.stdout or "別のコマンド" in p.stdout, \
        f"照合出力を人のファイルと誤認した文言:\n{p.stdout}"


def test_creator_mark_sets_and_signature_table_are_in_sync():
    """★ 番人（片配線の名所の常設監視・2026-08-21）: 印の集合は stack.py と verify.py で
       意図的に二重管理されている。片方だけに kind を足すと『在っても繋がらない』が再演する
       ── 両集合の一致と、署名テーブルのキー一致を機械で見張る。"""
    from ailine_core import stack as st
    from ailine_core import verify as vf
    assert set(st.CREATOR_MARKS) == set(vf._CREATOR_MARKS), \
        f"印の集合が二重管理でずれた: stack={st.CREATOR_MARKS} verify={vf._CREATOR_MARKS}"
    assert set(st.KIND_SIGNATURES.keys()) == set(st.CREATOR_MARKS), \
        f"署名テーブルに無い印がある（署名なしの印は fail closed を破る）: {st.KIND_SIGNATURES.keys()}"


# ---- operator 盲検 7 度目・$0 の主因（2026-08-21 18:4x・発見③の凍結）----

_INV_HDRS = ["日付", "品目", "数量", "単価", "金額"]


def _invoice(path, items, total_label_col, formula=False):
    """operator の検体の再現: 数量・単価つきの実務標準形。合計ラベルの列位置は可変・
       合計の数字は 金額 列だけ（数量・単価は空）── これが実物の合計行の形。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(_INV_HDRS)
    total = 0
    for d, item, qty, price in items:
        ws.append([d, item, qty, price, qty * price])
        total += qty * price
    row = [None] * 5
    row[total_label_col] = "合計"
    row[4] = f"=SUM(E2:E{len(items) + 1})" if formula else total
    ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


_ITEMS = [("2026-07-03", "事務用品", 10, 1200), ("2026-07-10", "コピー用紙", 5, 400)]


@pytest.mark.parametrize("label_col", [0, 1, 3], ids=["日付列", "品目列", "単価列"])
def test_qty_price_invoices_totals_are_excluded_regardless_of_label_column(tmp_path, label_col):
    """★ operator $0 の主因: 最初の数値列（数量）に数字が無い合計行は全トリガが沈黙し、
       Σ が黙って 2 倍になった。合計の数字がどの数値列にあっても・ラベルがどの列でも除外。"""
    folder = tmp_path / "src"
    _invoice(folder / "a.xlsx", _ITEMS, label_col)
    _invoice(folder / "b.xlsx", _ITEMS, label_col)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    ws = openpyxl.load_workbook(out).active
    labels = [str(r[label_col].value) for r in ws.iter_rows(min_row=2)]
    assert "合計" not in labels, f"合計行が積まれた（ラベル列 {label_col}）: {labels}"
    # Σ金額 は明細 4 行分 = (10*1200+5*400)*2 = 28000
    assert "28000" in p.stdout and "56000" not in p.stdout, f"Σ が 2 倍:\n{p.stdout}"


def test_total_word_row_stacked_anyway_trips_loud_wire(tmp_path):
    """★ 恒真切り（第二の検出器）: 除外機構が何かの理由で沈黙しても、『合計/小計/総計』の
       語を持つ行が出力に積まれたら、列解決に依存しない語のトリップワイヤが ⚠ で鳴る
       （黙って倍額、を機械で不可能にする ── 検出できないなら鳴って人に渡す）。
       検体: ラベル『 合 計 』（空白混じり・語トリガの正規化が拾えない変種を想定して
       将来の沈黙も含めて網を張る ── まず現行の沈黙形（数量列に数字なし）で赤を確認）。"""
    folder = tmp_path / "src"
    _invoice(folder / "a.xlsx", _ITEMS, 1)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    ws = openpyxl.load_workbook(out).active
    stacked_texts = " ".join(str(c.value) for row in ws.iter_rows(min_row=2) for c in row if c.value)
    if "合計" in stacked_texts:
        assert "⚠" in p.stdout and "合計" in p.stdout, \
            f"合計語の行が積まれたのにトリップワイヤが黙った:\n{p.stdout}"


def test_verify_also_trips_on_stacked_total_word_rows(tmp_path):
    """★ verify 側の恒真切り: stack と同じ除外規則の再計算だけでは沈黙を共有する ──
       verify にも語のトリップワイヤ（出力データ行に合計語 → ⚠・exit 5）。"""
    folder = tmp_path / "src"
    _invoice(folder / "a.xlsx", _ITEMS, 1)
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    ws = openpyxl.load_workbook(out).active
    stacked_texts = " ".join(str(c.value) for row in ws.iter_rows(min_row=2) for c in row if c.value)
    v = subprocess.run([sys.executable, "-m", "ailine", "verify", str(out), str(folder)],
                       capture_output=True, text=True, timeout=120, encoding="utf-8")
    if "合計" in stacked_texts:
        assert v.returncode == 5, f"合計語の行が積まれた出力を verify が合格させた:\n{v.stdout}"


def test_total_word_in_filename_does_not_trip_the_wire(tmp_path):
    """★ 再演検分（2026-08-21 19:1x）: 『合計』を含むファイル名（A_合計列0.xlsx）が
       出所列の値としてトリップワイヤに引っかかり、正当な出力が verify exit 5 になった。
       ワイヤの走査対象はデータ列のみ ── 出所列（うちが付けた列）は対象外。"""
    folder = tmp_path / "src"
    _book(folder / "月次_合計表.xlsx", HDRS, [("J-1", "甲", 100)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout
    assert "⚠" not in p.stdout.split("出力先:")[-1] or "合計語" not in p.stdout, \
        f"ファイル名の『合計』でワイヤが誤発火:\n{p.stdout}"
    v = subprocess.run([sys.executable, "-m", "ailine", "verify", str(out), str(folder)],
                       capture_output=True, text=True, timeout=120, encoding="utf-8")
    assert v.returncode == 0, f"ファイル名の『合計』で verify が落ちた:\n{v.stdout}"


def test_verify_never_fails_silently(tmp_path):
    """★ 憲法 1: verify が非 0 で終わる時は、必ず理由の ⚠ が 1 行以上出る（黙る不合格の禁止）。
       再演検分の実物: total_word の描画分岐が無く、全数字一致表示のまま exit 5 になった ──
       total_word だけが発火する形（データセルに『合計商事』）で凍結する。
       ★ ワイヤは発火してよい（データ列の語・誤爆でも 30 秒で確認できる ⚠）── 罪は黙ること。"""
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "合計商事", 100)])
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0
    v = subprocess.run([sys.executable, "-m", "ailine", "verify", str(out), str(folder)],
                       capture_output=True, text=True, timeout=120, encoding="utf-8")
    if v.returncode != 0:
        assert "⚠" in v.stdout and ("合計" in v.stdout), \
            f"exit {v.returncode} なのに理由が 1 行も無い（黙る不合格）:\n{v.stdout}"
