"""C1-F2: verify_dsl_args(op, args, book_meta, task, vocab) の (ok, resolved, inferred, err)
を JSON として凍結する（★エラー文言も含めて）。

★★ 本番コード(ailine.py)は一切変更しない。ここは凍結用のテスト追加のみ。

対象: 正常系・未実在列・数字表記・factor 各種（依頼文抽出/用語集/未解決の2パターン/
非正の数/LLM 食い違い警告/税込・税抜ラベル）を全 op にわたって収載する。
CLARIFY/OUT_OF_VOCAB は verify_dsl_args 自身の分岐ではない（_normalize_plan_step 側で
弾かれ、verify_dsl_args には渡らない）ため、代わりに「未対応の op」「シート無し」を
境界ケースとして収載する。

ゴールデンの置き場所: tests/golden/f2_verify/<name>.json。更新の作法は
tests/golden/_harness.py の docstring 参照。
"""
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

from golden._harness import GOLDEN_ROOT, assert_golden_json, sorted_list  # noqa: E402

F2_DIR = GOLDEN_ROOT / "f2_verify"

_HEADERS = {"Sheet": ["商品", "数量", "単価", "金額"], "単価表": ["商品", "単価"]}
BM = {"sheets": ["Sheet", "単価表"], "headers": _HEADERS,
      "header_rows": {"Sheet": 1, "単価表": 1}}
BM_NO_SHEETS = {"sheets": [], "headers": {}, "header_rows": {}}
# LOOKUP_FILL の digit 候補が一意に決まるケース専用（列2つだけの対象シート）。
BM_LOOKUP_2COL = {"sheets": ["Sheet", "単価表"],
                   "headers": {"Sheet": ["商品", "単価"], "単価表": ["商品", "単価"]},
                   "header_rows": {"Sheet": 1, "単価表": 1}}

# 名前 -> (op, args, book_meta, task, vocab)
CASES: dict = {}


def _add(name, op, args, book_meta=BM, task="", vocab=None):
    assert name not in CASES, f"重複した case 名: {name}"
    CASES[name] = (op, args, book_meta, task, vocab or {})


# --- SORT -----------------------------------------------------------------
_add("sort_ok_by_name", "SORT", {"col": "金額", "order": "desc"})
_add("sort_ok_digit_resolve", "SORT", {"col": "3", "order": "asc"})   # 0起点3=金額
_add("sort_unknown_column", "SORT", {"col": "存在しない", "order": "asc"})
_add("sort_bad_order", "SORT", {"col": "金額", "order": "ascending"})

# --- COMPUTE_COLUMN（2列演算） ---------------------------------------------
_add("compute_column_2op_ok", "COMPUTE_COLUMN", {"operands": ["数量", "単価"], "operator": "*"})
_add("compute_column_bad_operand_count", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価", "金額"], "operator": "*"})
_add("compute_column_2op_unknown_operand", "COMPUTE_COLUMN",
     {"operands": ["数量", "存在しない"], "operator": "*"})
_add("compute_column_2op_bad_operator", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "%"})
_add("compute_column_target_ambiguous_digit", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "*", "target": "1"})   # 0起点1=数量/1起点1=商品
_add("compute_column_target_unknown_falls_back_to_new_column", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "*", "target": "存在しない列"})
# --- COMPUTE_COLUMN（単列×倍率） -------------------------------------------
_add("compute_column_single_bad_operator", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "+"})
_add("compute_column_single_factor_from_task_text", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="金額に消費税10%を掛けた列を作って")
_add("compute_column_single_factor_from_vocab", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="消費税込みの金額にして",
     vocab={"消費税": 1.1})
_add("compute_column_single_no_rate_signal_clarify", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="金額の列の値を変えて")
_add("compute_column_single_rate_signal_unresolved_clarify", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="税率をどうにかして")
_add("compute_column_single_factor_le0", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="金額を0倍にした列を作って")
_add("compute_column_single_llm_factor_mismatch_warns", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*", "factor": "1.05"},
     task="金額に消費税10%を掛けた列を作って")
_add("compute_column_single_tax_inclusive_label", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*"}, task="金額に消費税10%を掛けた税込金額列を作って")
_add("compute_column_single_tax_exclusive_label", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "/"}, task="金額を1.1で割った税抜金額列を作って")
_add("compute_column_single_rate_keyword_fallback_label", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "/"}, task="倍率1.1を使った金額列を作って")
_add("compute_column_single_target_present_no_label", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "/", "target": "金額"},
     task="金額を1.1で割った税抜金額にして")

# --- LOOKUP_FILL ------------------------------------------------------------
_add("lookup_fill_ok_exists_and_mentioned", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "数量", "source_sheet": "単価表", "key_col": "商品"},
     task="数量を単価表から転記して")
_add("lookup_fill_ok_exists_matches_value_col_only", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"},
     task="単価表から値を埋めて")
_add("lookup_fill_exists_without_grounds_errors", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "数量", "source_sheet": "単価表", "key_col": "商品"},
     task="単価表から値を埋めて")
_add("lookup_fill_missing_digit_candidate_resolves", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "2", "source_sheet": "単価表", "key_col": "商品"},
     book_meta=BM_LOOKUP_2COL, task="単価表から転記して")
_add("lookup_fill_missing_and_mentioned_creates_new", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "備考", "source_sheet": "単価表", "key_col": "商品"},
     task="備考という列を作って単価表から転記して")
_add("lookup_fill_missing_not_mentioned_errors", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "備考", "source_sheet": "単価表", "key_col": "商品"},
     task="単価表から値を転記して")
_add("lookup_fill_unknown_target_sheet", "LOOKUP_FILL",
     {"target_sheet": "存在しないシート", "target_col": "単価", "source_sheet": "単価表",
      "key_col": "商品"})
_add("lookup_fill_unknown_source_sheet", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "単価", "source_sheet": "存在しないシート",
      "key_col": "商品"})
_add("lookup_fill_target_sheet_not_first_now_ok", "LOOKUP_FILL",   # ★ 挙動変更#2: 旧「対象シートは
     # 1枚目のみ対応」拒否を撤廃（散在した1枚目固定の一箇所）。target_sheet が2枚目でも
     # target_col が実在(単価)かつ値列(source_headers[1])と一致すれば従来どおり通る。
     {"target_sheet": "単価表", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"})
_add("lookup_fill_target_sheet_second_source_first", "LOOKUP_FILL",   # ★ 挙動変更#2: target_sheet
     # が2枚目(単価表)・source_sheet が1枚目(Sheet)という、旧実装なら拒否されていた非対称な
     # 組み合わせが通ることを示す（依頼文に列名の言及あり＝根拠つき）。
     {"target_sheet": "単価表", "target_col": "単価", "source_sheet": "Sheet", "key_col": "商品"},
     task="単価表の単価を更新して")
_add("lookup_fill_unknown_key_col", "LOOKUP_FILL",
     {"target_sheet": "Sheet", "target_col": "単価", "source_sheet": "単価表", "key_col": "不明列"},
     task="単価転記して")

# --- AGGREGATE --------------------------------------------------------------
_add("aggregate_ok", "AGGREGATE", {"group_col": "商品", "value_col": "金額"})
_add("aggregate_unknown_group_col", "AGGREGATE", {"group_col": "不明", "value_col": "金額"})
_add("aggregate_unknown_value_col", "AGGREGATE", {"group_col": "商品", "value_col": "不明"})

# --- BOLD / FILL_COLOR / CENTER_ALIGN ---------------------------------------
_add("bold_target_all_unsupported", "BOLD", {"target": "all"})
_add("center_align_target_all_ok", "CENTER_ALIGN", {"target": "all"})
_add("bold_target_row_ok", "BOLD", {"target": "row:2"})
_add("bold_target_row_bad_number", "BOLD", {"target": "row:abc"})
_add("fill_color_target_col_ok", "FILL_COLOR", {"target": "col:金額", "color": "red"})
_add("fill_color_target_col_unknown", "FILL_COLOR", {"target": "col:存在しない", "color": "red"})
_add("bold_target_unknown_format", "BOLD", {"target": "xyz"})
_add("fill_color_unknown_color", "FILL_COLOR", {"target": "col:金額", "color": "虹色"})

# --- NUMBER_FORMAT -----------------------------------------------------------
_add("number_format_ok", "NUMBER_FORMAT", {"col": "金額", "style": "thousands"})
_add("number_format_unknown_col", "NUMBER_FORMAT", {"col": "不明", "style": "thousands"})
_add("number_format_unsupported_style", "NUMBER_FORMAT", {"col": "金額", "style": "percent"})

# --- MERGE --------------------------------------------------------------
_add("merge_ok", "MERGE", {"range": "A1:C3"})
_add("merge_bad_format", "MERGE", {"range": "A1-C3"})

# --- CHART --------------------------------------------------------------
_add("chart_ok", "CHART", {"value_col": "金額"})
_add("chart_unknown_value_col", "CHART", {"value_col": "不明"})

# --- APPEND_TOTAL ----------------------------------------------------------
_add("append_total_ok_default_factor", "APPEND_TOTAL", {"col": "金額"})
_add("append_total_ok_factor_from_task_text", "APPEND_TOTAL", {"col": "金額"},
     task="金額に消費税10%を掛けた合計を出して")
_add("append_total_ok_factor_from_vocab", "APPEND_TOTAL", {"col": "金額"},
     task="消費税込みの合計を出して", vocab={"消費税": 1.1})
_add("append_total_factor_le0", "APPEND_TOTAL", {"col": "金額"}, task="金額を0倍にした合計を出して")
_add("append_total_tax_label_unresolved_factor_errors", "APPEND_TOTAL",
     {"col": "金額", "label": "税込合計"})
_add("append_total_llm_factor_mismatch_warns", "APPEND_TOTAL",
     {"col": "金額", "factor": "1.05"}, task="金額に消費税10%を掛けた合計を出して")
_add("append_total_unknown_col", "APPEND_TOTAL", {"col": "不明"})

# --- INSERT_ROWS -----------------------------------------------------------
_add("insert_rows_ok_explicit_count", "INSERT_ROWS", {"at": "3", "count": "2"})
_add("insert_rows_ok_default_count_inferred", "INSERT_ROWS", {"at": "3"})
_add("insert_rows_bad_at", "INSERT_ROWS", {"at": "0"})
_add("insert_rows_bad_count", "INSERT_ROWS", {"at": "3", "count": "abc"})

# --- DRAW_BORDERS / AUTOFIT（引数無し） ------------------------------------
_add("draw_borders_ok", "DRAW_BORDERS", {})
_add("autofit_ok", "AUTOFIT", {})

# --- PIVOT --------------------------------------------------------------
_add("pivot_ok", "PIVOT", {"group_col": "商品", "value_col": "金額"})
_add("pivot_unknown_group_col", "PIVOT", {"group_col": "不明", "value_col": "金額"})
_add("pivot_unknown_value_col", "PIVOT", {"group_col": "商品", "value_col": "不明"})

# --- SET_COLUMN_VALUE --------------------------------------------------------
_add("set_column_value_ok_quoted", "SET_COLUMN_VALUE", {"col": "商品"},
     task="商品列を全部『確認済み』にして")
_add("set_column_value_llm_mismatch_warns", "SET_COLUMN_VALUE",
     {"col": "商品", "value": "LLM値"}, task="商品列を全部『確認済み』にして")
_add("set_column_value_no_quoted_literal_errors", "SET_COLUMN_VALUE", {"col": "商品"},
     task="商品列を全部確認済みにして")
_add("set_column_value_unknown_col", "SET_COLUMN_VALUE", {"col": "不明"},
     task="不明列を全部『確認済み』にして")

# --- EXTRACT --------------------------------------------------------------
_add("extract_ok_numeric_gte", "EXTRACT", {"col": "金額", "cmp": "gte", "value": "40000"})
_add("extract_ok_string_contains", "EXTRACT", {"col": "商品", "cmp": "contains", "value": "セット"})
_add("extract_unknown_column", "EXTRACT", {"col": "不明", "cmp": "gte", "value": "100"})
_add("extract_bad_cmp", "EXTRACT", {"col": "金額", "cmp": "between", "value": "100"})
_add("extract_numeric_cmp_needs_number", "EXTRACT", {"col": "金額", "cmp": "gte", "value": "abc"})
_add("extract_eq_falls_back_to_string", "EXTRACT", {"col": "商品", "cmp": "eq", "value": "ノートPC"})
_add("extract_missing_value", "EXTRACT", {"col": "金額", "cmp": "gte", "value": ""})

# --- DEDUP（EXTRACT の兄弟）--------------------------------------------------
_add("dedup_ok_single_key", "DEDUP", {"keys": ["商品"]})
_add("dedup_ok_multi_key", "DEDUP", {"keys": ["商品", "単価"]})
_add("dedup_unknown_key_column", "DEDUP", {"keys": ["不明"]})
_add("dedup_missing_keys_empty_list", "DEDUP", {"keys": []})
_add("dedup_missing_keys_absent", "DEDUP", {})

# --- REPORT_PER_ROW（帳票段）--------------------------------------------------
# ★ 印の実在検証・行の会計は本物のファイルを読む（book_meta["path"]）ため、この golden
#   だけは実ファイルを持つ専用の book_meta を使う（他 op は列名の実在照合だけで足りるため
#   静的な BM で済んでいる）。
_REPORT_DIR = Path(tempfile.mkdtemp(prefix="ailine_golden_f2_report_"))
_REPORT_BOOK_PATH = _REPORT_DIR / "report_src.xlsx"


def _build_report_book() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("雛形")
    tpl["B1"] = "{{商品}}"
    tpl["B2"] = "{{金額}}"
    wb.save(_REPORT_BOOK_PATH)


_build_report_book()
BM_REPORT = {"sheets": ["Sheet", "雛形"],
             "headers": {"Sheet": ["商品", "金額"], "雛形": []},
             "header_rows": {"Sheet": 1, "雛形": 1},
             "path": _REPORT_BOOK_PATH}

_add("report_per_row_ok", "REPORT_PER_ROW",
     {"template_sheet": "雛形", "name_col": "商品"}, book_meta=BM_REPORT)
_add("report_per_row_unknown_template_sheet", "REPORT_PER_ROW",
     {"template_sheet": "存在しない", "name_col": "商品"})
_add("report_per_row_template_same_as_data_sheet", "REPORT_PER_ROW",
     {"template_sheet": "Sheet", "name_col": "商品"})
_add("report_per_row_unknown_name_col", "REPORT_PER_ROW",
     {"template_sheet": "単価表", "name_col": "不明"})
_add("report_per_row_missing_book_path", "REPORT_PER_ROW",
     {"template_sheet": "単価表", "name_col": "商品"})

# --- FORMAT_MAP（様式写像段。REPORT_PER_ROW の兄弟・縦の展開）----------------
# ★ REPORT_PER_ROW と同じ理由でこの golden だけ実ファイルを持つ専用 book_meta を使う。
_FORMAT_MAP_DIR = Path(tempfile.mkdtemp(prefix="ailine_golden_f2_format_map_"))
_FORMAT_MAP_BOOK_PATH = _FORMAT_MAP_DIR / "format_map_src.xlsx"
_FORMAT_MAP_BAD_BOOK_PATH = _FORMAT_MAP_DIR / "format_map_bad.xlsx"


def _build_format_map_book() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "品名"
    tpl["B1"] = "価格"
    tpl["A2"] = "{{商品}}"
    tpl["B2"] = "{{金額}}"
    wb.save(_FORMAT_MAP_BOOK_PATH)


def _build_format_map_bad_book() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "品名"
    tpl["A2"] = "{{存在しない}}"
    wb.save(_FORMAT_MAP_BAD_BOOK_PATH)


_build_format_map_book()
_build_format_map_bad_book()
BM_FORMAT_MAP = {"sheets": ["Sheet", "様式"],
                  "headers": {"Sheet": ["商品", "金額"], "様式": []},
                  "header_rows": {"Sheet": 1, "様式": 1},
                  "path": _FORMAT_MAP_BOOK_PATH}
BM_FORMAT_MAP_BAD = {"sheets": ["Sheet", "様式"],
                      "headers": {"Sheet": ["商品", "金額"], "様式": []},
                      "header_rows": {"Sheet": 1, "様式": 1},
                      "path": _FORMAT_MAP_BAD_BOOK_PATH}

_add("format_map_ok", "FORMAT_MAP", {"template_sheet": "様式"}, book_meta=BM_FORMAT_MAP)
_add("format_map_unknown_template_sheet", "FORMAT_MAP", {"template_sheet": "存在しない"})
_add("format_map_template_same_as_data_sheet", "FORMAT_MAP", {"template_sheet": "Sheet"})
_add("format_map_unknown_placeholder_column", "FORMAT_MAP", {"template_sheet": "様式"},
     book_meta=BM_FORMAT_MAP_BAD)
_add("format_map_missing_book_path", "FORMAT_MAP", {"template_sheet": "単価表"})

# --- SPLIT_CELL（1セルの複数値を右の列へ割る）---------------------------------
# ★ 何列必要かは**実データ**が決めるので、この golden も実ファイルを持つ book_meta を使う。
_SPLIT_DIR = Path(tempfile.mkdtemp(prefix="ailine_golden_f2_split_"))
_SPLIT_BOOK_PATH = _SPLIT_DIR / "split_src.xlsx"


def _build_split_book() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["商品", "金額"])
    ws.append(["a,b", 100])
    ws.append(["c,d,e", 200])
    wb.save(_SPLIT_BOOK_PATH)


_build_split_book()
BM_SPLIT = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]},
            "header_rows": {"Sheet": 1}, "path": _SPLIT_BOOK_PATH}

_add("split_cell_ok", "SPLIT_CELL", {"col": "商品", "sep": ","}, book_meta=BM_SPLIT)
_add("split_cell_unknown_separator", "SPLIT_CELL",
     {"col": "商品", "sep": "これは区切りの説明としか読めない長い文字列"}, book_meta=BM_SPLIT)
_add("split_cell_separator_not_found", "SPLIT_CELL", {"col": "金額", "sep": ","},
     book_meta=BM_SPLIT)
_add("split_cell_unknown_column", "SPLIT_CELL", {"col": "存在しない", "sep": ","},
     book_meta=BM_SPLIT)

# --- 境界: verify_dsl_args 自体の全体ガード ----------------------------------
_add("no_sheets_in_book", "SORT", {"col": "金額", "order": "asc"}, book_meta=BM_NO_SHEETS)
_add("unsupported_op", "FOOBAR", {})


def _case_ids():
    return sorted(CASES.keys())


def test_verify_dsl_args_golden_coverage_matches_declared_ops():
    """★網羅性の自己検査: OP_WRITE_TARGET に載っている全 op が最低1ケース入っているか。"""
    covered_ops = {op for op, *_ in CASES.values()}
    all_ops = set(ailine.OP_WRITE_TARGET.keys())
    missing = all_ops - covered_ops
    assert not missing, f"verify_dsl_args golden が網羅していない op: {sorted(missing)}"


@pytest.mark.parametrize("name", _case_ids())
def test_verify_dsl_args_golden(name):
    op, args, book_meta, task, vocab = CASES[name]
    ok, resolved, inferred, err = ailine.verify_dsl_args(op, dict(args), book_meta, task, vocab)
    payload = {"ok": ok, "resolved": resolved, "inferred": sorted_list(inferred), "err": err}
    assert_golden_json(F2_DIR / f"{name}.json", payload, label=name)
