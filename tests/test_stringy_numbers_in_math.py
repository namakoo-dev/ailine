# 塊③（2026-08-25）── 前提を**別実装で**確かめないと塞がらない 3 件。
#
# ★ 中核 op の盲検・致命3/致命4:
#   営業 = 1000 + **'2000'（文字列）** + 1500 の表を集計すると:
#       事後条件を確認（操作:集計）: 2 グループを検証
#       ✓ 機械検証済みの内容です
#   出力は 営業 **2500**（正 4500）／合計 **6000**（正 8000）。
#   ★ 並べ替え・計算列は「（数値でない N 行は対象外）」と一言添えるのに、
#     **集計は一言も言わない**。
#
# ★ 根（塊①の除外の開示では塞がらない理由）:
#   `check_aggregate` は `v = v if _is_number(v) else 0` を
#   **期待側と観測側の両方**に掛けている。LibreOffice の SUM と同じ落とし方をするので
#   **必ず一致する** ── 検算が恒真。除外を数えても、数える相手が居ない。
#   検分者の言葉:「前提（文字列型の数値）を**別実装で**確かめない限り塞がらない」。
#
# 契約:
#   ① 集計・合計追加の対象列に「数字に見える文字列」があれば、機械の値として残す
#   ② 1 行でもあれば ✓ を名乗らない（金額が黙って小さくなるのを ✓ で覆わない）
#   ③ 何行が・どんな値かを名指しする（人が直せるように）
#   ④ 本物の数値だけなら 1 文字も増えない（誤爆しない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def test_aggregate_records_stringy_numbers(tmp_path):
    """①③ 文字列の数値を機械の値として残す。"""
    p = _book(tmp_path, [["部門", "金額"], ["営業", 1000], ["営業", "2000"],
                          ["営業", 1500], ["開発", 3500]])
    wb = openpyxl.load_workbook(p)
    ws = wb["売上"]
    wb.create_sheet("集計")
    wb["集計"].append(["部門", "金額"])
    wb["集計"].append(["営業", 2500])
    wb["集計"].append(["開発", 3500])
    wb.save(p)
    args = {"group_col": "部門", "value_col": "金額"}
    ailine.check_aggregate(p, args)
    assert args.get("_unverified"), "文字列の数値が黙って 0 になった"
    assert any("2000" in u["why"] or u["rows"] == 1 for u in args["_unverified"]), \
        args["_unverified"]


def _with_total_row(tmp_path, values, total):
    """★ 治具の訂正: 合計行は**本物と同じ形**（挿入耐性 SUM 型 + キャッシュ値）で作らないと、
       検算が値を集める所まで到達しない（初版はただの数値を置いて別の理由で fail した）。"""
    p = tmp_path / "t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["品目", "金額"])
    for i, v in enumerate(values):
        ws.append([f"r{i}", v])
    last = ws.max_row + 1
    ws.cell(last, 1, "合計")
    ws.cell(last, 2, "=SUM(B2:INDEX(B:B,ROW()-1))")
    wb.save(p)
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_readback_claim import _inject_formula_cache
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {f"B{last}": total})
    return p


def test_append_total_records_stringy_numbers(tmp_path):
    """① 合計追加も同じ形（致命4）。画面に 1000・2000・3000 が並んで合計 4000。"""
    p = _with_total_row(tmp_path, [1000, "2000", 3000], total=4000)
    args = {"col": "金額", "label": "合計", "factor": 1}
    ailine.check_append_total(p, args)
    assert args.get("_unverified"), "文字列の 2000 が落ちているのに黙った"


def test_no_noise_on_real_numbers(tmp_path):
    """④ 誤爆しない。"""
    p = _with_total_row(tmp_path, [1000, 2000, 3000], total=6000)
    args = {"col": "金額", "label": "合計", "factor": 1}
    ailine.check_append_total(p, args)
    assert not args.get("_unverified"), args.get("_unverified")


def test_the_detector_is_a_separate_implementation():
    """★ 恒真を切る条件: 期待側と観測側が**同じ落とし方**をしないこと。
       文字列を 0 にする判断（_is_number）とは別の実装で「数字に見えるか」を見る。"""
    from ailine_core import compare_blocked
    assert compare_blocked.looks_numeric("2000") is True
    assert compare_blocked.looks_numeric("1,500") is True
    assert ailine._is_number("2000") is False      # 本体の判断は変えない
