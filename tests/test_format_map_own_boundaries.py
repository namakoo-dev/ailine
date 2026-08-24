# 様式写像段 FORMAT_MAP + CSV_EXPORT ── 自分の境界検体（凍結検体 tests/test_format_map.py
# は事後実測化されているため変更禁止・こちらは実装者が追加する非凍結の追加検体）。
#
# ブリーフの完了条件: 「自分の境界検体（引用が要る値・cp932 で書けない文字・空シート）」。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def test_csv_export_quotes_value_with_delimiter_and_newline(tmp_path, monkeypatch, capsys):
    """値にカンマ・改行・引用符を含む場合、引用して書く（RFC4180 相当・引用符は二重化）。
       読み戻しても欠落/不一致/余剰が出ないことまで確かめる（恒真殺しでなく実際に正しい）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "quote.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["摘要"])
    ws.append(['カンマ,改行\n引用"符'])
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "s"], capsys)
    assert rc == 0, out
    assert "✓" in out and "欠落0" in out.replace(" ", ""), out
    raw = p.with_suffix(".csv").read_bytes()
    text = raw.decode("utf-8-sig")
    # RFC4180: フィールドを引用符で囲み、内部の引用符は二重化する。
    assert '"カンマ,改行\n引用""符"' in text, f"引用の規則が守られていない: {text!r}"


def test_csv_export_cp932_unencodable_char_refuses_cleanly(tmp_path, monkeypatch, capsys):
    """cp932 で書けない文字（絵文字等）があれば、例外を漏らさず名指しで断る（黙って落とさない・
       黙って置換しない）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "emoji.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["メモ"])
    ws.append(["🎉お祝い"])
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "s", "--encoding", "cp932"], capsys)
    assert rc != 0, out
    assert "cp932" in out, f"どの符号化で書けなかったか名指ししていない: {out}"
    assert not p.with_suffix(".csv").exists(), "書けない文字があるのに CSV を残した"


def test_csv_export_empty_sheet_still_claims_correctly(tmp_path, monkeypatch, capsys):
    """空シート（見出しすら無い）を渡しても例外を出さず、0行0列を正直に主張する。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空"
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "空"], capsys)
    assert rc == 0, out
    assert "✓" in out, out
    assert p.with_suffix(".csv").exists()
    assert p.with_suffix(".csv").read_bytes().decode("utf-8-sig") == "", (
        "空シートなのに CSV に中身がある")
