"""条件つき書換の「〜以外」（2026-09-05）。

★★ 出所（2026-09-04 の実測・この repo で一番こわい形）:
  「所属が営業**以外**の行のメモに『○』を付けて」で **営業の行に ○ が付き、
  しかも ✓ が出た** ── ちょうど逆のことをして合格した。

  なぜ通ったか: 依頼は否定・宣言は `eq`・事後条件も**同じ eq** を見た。
  宣言と実体は一致するので検算は正しく通る。**欠けていたのは「依頼 vs 宣言」**で、
  そこは if では塞げない（三項のうち 1 項が渡されていなかった）。

  ★ その日の処置は「否定が在る回は門を開けない」── 正しいが、**扱えるようには
    していなかった**（うるさい失敗に戻しただけ）。ここで扱えるようにする。

★ 新しい述語は 1 つも書いていない。部品は全部在った:
    Basic  `RowMatches` Case 7（区切りごとの丸ごと一致の**否定**）
    Python `_extract_predicate` の "nin"（**別実装** ＝ 事後条件が恒真にならない）
    凍結    tests/test_predicate_truth_table.py が両方を縛る
  ★ 足りなかったのは `_EXTRACT_CMPS`（許可リスト）に "nin" が無かったことだけ ──
    兄弟の EXTRACT は特別扱いの経路で通っていたので、**片側だけ古かった**。
"""
import os
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine  # noqa: E402

HEADERS = ["氏名", "所属", "メモ"]
ROWS = [["田中", "営業", ""], ["鈴木", "経理", ""],
        ["佐藤", "営業", ""], ["山田", "総務", ""]]
NEG = "所属が営業以外の行のメモに「○」を付けて"
POS = "所属が営業の行のメモに「○」を付けて"


@pytest.fixture
def book(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "名簿"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    p = tmp_path / "neg.xlsx"
    wb.save(p)
    return p


def _resolve(book, task, cmp0="eq"):
    bm = ailine.build_book_meta(book)
    return ailine.verify_dsl_args(
        "SET_WHERE", {"col": "メモ", "cond_col": "所属", "cmp": cmp0, "value": "○"},
        bm, task=task, vocab=ailine.load_vocab())


# --- ① 否定が否定として解決されること ---------------------------------------

def test_the_negation_becomes_a_not_in_comparison(book):
    ok, res, _inf, err = _resolve(book, NEG)
    assert ok, err
    assert res["cmp"] == "nin", res["cmp"]
    assert res["cond_value"] == ["営業"], res["cond_value"]
    # ★ 当てはまるのは**営業でない行**（3 行目 経理・5 行目 総務）
    assert res["_match_rows"] == [3, 5], res["_match_rows"]


def test_the_label_is_written_the_way_a_person_writes_it(book):
    """★ 一覧をそのまま出すと画面に ['営業'] と Python のリストが出る（実測）。"""
    _ok, res, _inf, _err = _resolve(book, NEG)
    label = res["_cond_label"]
    assert "[" not in label and "'" not in label, label
    assert "営業" in label and "どれでもない" in label, label


# --- ② 否定でない側が壊れていないこと（★ 対で縛る）--------------------------

def test_the_positive_case_is_untouched(book):
    ok, res, _inf, err = _resolve(book, POS)
    assert ok, err
    assert res["cmp"] == "eq" and res["cond_value"] == "営業"
    assert res["_match_rows"] == [2, 4], res["_match_rows"]


def test_a_numeric_condition_is_untouched(book):
    ok, res, _inf, _err = _resolve(book, "所属が3以上の行のメモに「○」を付けて")
    assert res["cmp"] == "gte"
    del ok


# --- ③ 依頼と宣言が食い違う回は実行しない（★ 三項の欠けていた項）-------------

def test_a_request_that_says_except_never_runs_as_equals(book, monkeypatch):
    """★ 静かな嘘の再演を機械で止める。

    否定の解決を殺して（＝2026-09-04 当時の状態にして）も、`eq` のまま走らないこと。
    """
    monkeypatch.setattr(ailine, "task_names_real_values", lambda *a, **k: [])
    ok, _res, _inf, err = _resolve(book, NEG)
    assert not ok
    assert "以外" in err, err


def test_it_refuses_when_the_named_value_is_not_in_the_table(book):
    """★ 実表に無い値の「以外」は決めない（推測で列を書き換えない）。"""
    ok, _res, _inf, err = _resolve(book, "所属が課長以外の行のメモに「○」を付けて")
    assert not ok and "決められません" in err, err


# --- ④ 門（読み直しの入口）が否定を通すこと ---------------------------------

def test_the_gate_now_lets_the_negation_through(book):
    bm = ailine.build_book_meta(book)
    assert ailine.task_asks_for_a_conditional_write(NEG, bm, "名簿") is True


def test_the_extraction_form_is_not_stolen_by_this_gate(book):
    """★ 「味噌汁以外を抜き出して」（抽出）は引用が無いので、この門の手前で外れる。"""
    bm = ailine.build_book_meta(book)
    assert ailine.task_asks_for_a_conditional_write("営業以外を抜き出して", bm, "名簿") is False


# --- ⑤ Basic に渡る比較コードが「どれでもない」であること --------------------

def test_the_generated_code_uses_the_not_in_case(book):
    _ok, res, _inf, _err = _resolve(book, NEG)
    code = ailine.codegen_dsl("SET_WHERE", res, book_meta=ailine.build_book_meta(book))
    assert "SetColumnValueWhere" in code
    # ★ 7 = RowMatches の Case 7（どれでもない）。ここが 4（等しい）になれば逆になる。
    assert ", 7, " in code, code[code.index("SetColumnValueWhere"):][:200]


# --- ⑥ 実機（★ 事故そのものの再現 ── 逆をやったら赤くなる）------------------

@pytest.mark.local
def test_the_negation_writes_to_the_right_rows_on_the_real_machine(book, tmp_path):
    """★★ 2026-09-04 に ✓ が出てしまった依頼文を、そのまま実機で走らせる。"""
    env = {**os.environ, "PYTHONPATH": str(REPO / "src")}
    got = subprocess.run([sys.executable, "-m", "ailine", "run", str(book), NEG,
                          "--copy", "--timeout", "200"],
                         capture_output=True, text=True, encoding="utf-8",
                         errors="replace", cwd=str(REPO), env=env)
    assert got.returncode == 0, got.stdout[-800:]
    out = book.with_name(book.stem + ".out.xlsx")
    assert out.exists(), got.stdout[-400:]
    wb = openpyxl.load_workbook(out)
    ws = wb["名簿"]
    got_col = [ws.cell(row=r, column=3).value for r in range(2, 6)]
    wb.close()
    # 田中(営業)・佐藤(営業) は空のまま、鈴木(経理)・山田(総務) に ○
    assert got_col == [None, "○", None, "○"], got_col
    assert "✓" in got.stdout, got.stdout[-400:]
