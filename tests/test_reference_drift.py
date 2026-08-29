# 参照のズレ ── 2026-08-29。Namakoo「入れ替えや並び替えによって参照するデータがズレる」
#
# ★★ 実測（同じ表に 4 種類の式を置いて金額の降順に並べ替えた）:
#     =B2*1.1（範囲の中・同じ行）      → ✓ 行と一緒に動いて正しい
#     =SUM(B2:B4)（範囲全体を指す）    → ✓ 600 のまま正しい
#     =売上!B2（別シートから）          → ⚠ 指す先が りんご100 → ぶどう300 に変わった
#     =B3（範囲の外・特定の 1 行を指す）→ ⚠ ラベルは「ぶどうの金額」なのに みかん200 を指す
#
#   ★★ **式は 1 文字も壊れていない**。壊れたのは意味だけ ── 値でも文字列でも捕まらず、
#     **参照を読むしかない**。そして ailine は ✓ を出していた（並べ替え自体は宣言どおり
#     だから）。「静かに壊れて合格が出る」── この製品が一番嫌う形。
#
# ★ 直さない: Excel も LibreOffice も、範囲の外から特定の行を指す式は並べ替えで
#   追従させない（アドレスに留まるのが既定の意味）。
#   「ぶどうの金額 = B3」は追従してほしいが「3行目の値 = B3」は留まってほしい
#   ── 機械には区別できない。**名指しして人に返す**（補正は人が決めてから）。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import cellmap as cm  # noqa: E402


def _book(tmp_path, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "金額", "税込み"])
    for i, (n, v) in enumerate([("りんご", 100), ("ぶどう", 300), ("みかん", 200)], start=2):
        ws.cell(i, 1, n), ws.cell(i, 2, v), ws.cell(i, 3, f"=B{i}*1.1")
    ws.cell(5, 1, "合計"), ws.cell(5, 2, "=SUM(B2:B4)")
    ws.cell(7, 1, "ぶどうの金額"), ws.cell(7, 2, "=B3")
    wb.create_sheet("別紙").cell(1, 2, "=売上!B2")
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["売上", "別紙"],
            "headers": {"売上": ["商品", "金額", "税込み"], "別紙": []},
            "header_rows": {"売上": 1, "別紙": 1}, "path": str(path)}


# --- ① 参照の読み分け（範囲は除く・別シートは拾う）--------------------------------------

@pytest.mark.parametrize("formula,own,want", [
    ("=B3", "売上", [("売上", 3, 2)]),
    ("=SUM(B2:B4)", "売上", []),                    # ★ 範囲は鳴らさない（正しく追従する）
    ("=売上!B2", "別紙", [("売上", 2, 2)]),
    ("=SUM(B2:B4)+C7", "売上", [("売上", 7, 3)]),    # 範囲は除き、単独だけ拾う
    ("='第 1'!B9", "別紙", [("第 1", 9, 2)]),        # 空白入りのシート名
    ("ただの文字", "売上", []),
])
def test_only_single_cell_references_are_collected(formula, own, want):
    assert cm.single_cell_refs(formula, own) == want


# --- ② 区画の外から指しているものだけ拾う ----------------------------------------------

def test_references_from_outside_the_sorted_rows_are_found(tmp_path):
    hits = cm.refs_pointing_into(_book(tmp_path), "売上", row_lo=2, row_hi=4)
    assert [(h[0], h[1]) for h in hits] == [("別紙", "B1"), ("売上", "B7")], hits


def test_references_from_inside_are_not_flagged(tmp_path):
    """★ 区画の中の式（=B2*1.1）は行と一緒に動くので正しく追従する ── 鳴らさない。"""
    hits = cm.refs_pointing_into(_book(tmp_path), "売上", row_lo=2, row_hi=4)
    assert not any(h[1].startswith("C") for h in hits), hits


def test_a_column_swap_uses_the_same_shape(tmp_path):
    """★ 行と列は軸違い ── 列を動かす回は、その列を指す式が対象になる
       （=B2*1.1 は列が動くと指す先が変わるので、こちらでは鳴る）。"""
    hits = cm.refs_pointing_into(_book(tmp_path), "売上", col_lo=2, col_hi=2)
    refs = {h[1] for h in hits}
    assert {"C2", "C3", "C4"} <= refs, refs
    assert "B1" in {h[1] for h in hits if h[0] == "別紙"}


def test_no_formulas_means_no_note(tmp_path):
    p = tmp_path / "plain.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["a", 1])
    wb.save(p)
    assert cm.reference_drift_note(cm.refs_pointing_into(p, "Sheet", 1, 9)) is None


# --- ③ 本体に配線されている（並べ替え・入れ替えの両方）----------------------------------

def test_sort_discloses_the_drift(tmp_path):
    p = _book(tmp_path)
    ok, r, _i, err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, _meta(p), task="金額の大きい順に並べ替えて")
    assert ok, err
    assert any("指す先の中身が変わる式" in w for w in r.get("_warnings", [])), r.get("_warnings")


@pytest.mark.parametrize("args", [
    {"a": "金額", "b": "税込み"},          # 列の入れ替え
    {"a": "りんご", "b": "みかん"},        # 行の入れ替え
])
def test_swap_discloses_the_drift(tmp_path, args):
    p = _book(tmp_path)
    ok, r, _i, err = ailine.verify_dsl_args("SWAP", args, _meta(p), task="入れ替えて")
    assert ok, err
    assert any("指す先の中身が変わる式" in w for w in r.get("_warnings", [])), r.get("_warnings")


def test_the_note_says_it_did_not_fix_anything():
    """★ 直していないことを、はっきり書く（黙って直さない・黙って放置もしない）。"""
    note = cm.reference_drift_note([("別紙", "B1", "=売上!B2", 2, 2)])
    assert "直していません" in note and "人が決める" in note, note


def test_a_clean_book_stays_silent(tmp_path):
    """★ 黙りすぎない側の対でなく、**鳴りすぎない**側の対: 参照が無ければ鳴らさない。"""
    p = tmp_path / "c.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "金額"])
    ws.append(["りんご", 100])
    ws.append(["ぶどう", 300])
    wb.save(p)
    meta = {"sheets": ["売上"], "headers": {"売上": ["商品", "金額"]},
             "header_rows": {"売上": 1}, "path": str(p)}
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "SORT", {"col": "金額", "order": "desc"}, meta, task="金額の大きい順に")
    assert not any("指す先" in w for w in r.get("_warnings", [])), r.get("_warnings")
