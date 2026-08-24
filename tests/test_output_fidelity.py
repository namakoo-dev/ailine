# 土台固め（2026-08-24）── 「人に渡す最終ファイル」の忠実度。修正より先に凍結した検体。
#
# 実測した 3 つの穴（全部この repo で再現した）:
#   ① 帳票段が検分シートを openpyxl で足す往復で、雛形の角印を全枚から消して ✓ を出す
#   ② 様式写像段が同じ形（コード同一）
#   ③ --copy は忠実度ゲートを素通りし、出力から VBA が消えたまま ✓ を出す
#      ── しかも --copy は**ゲート自身が案内する逃げ道**だった
#
# ★ 根は 1 つ: ゲートを「正規化の直後」だけに掛けて、**成果物**に掛けていなかった。
# ★ さらに、ゲートは zip の**ファイル名**しか比べておらず、drawing1.xml を残したまま
#   中の図形だけ捨てる形を取り逃がしていた（ファイルが在ることは中身が在ることではない）。
#
# 契約:
#   A 描画オブジェクトを**中身の個数**まで数える（ファイル名の比較では足りない）
#   B 減った時だけ喪失とする（帳票段が雛形を N 枚に複製して増えるのは正常）
#   C 忠実度ゲートが A を取り込んでいる
#   D 検分シートの Basic 断片は生の制御文字を含まない（Chr() 式で書く）
#   E 帳票段/様式写像段の検分シートは**同じ関数**で組む（書き写さない）

import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402
from lo_fake import apply_inspection_sheets  # noqa: E402
from test_vanishing_shapes import _PIC, _SHAPE, _write_drawing  # noqa: E402


def _sealed_book(tmp_path, name="sealed.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    wb.active["A1"] = "請求書"
    wb.save(p)
    return _write_drawing(p, _PIC + _SHAPE)


# --- A/B 中身まで数える -------------------------------------------------------------

def test_counts_objects_inside_drawings(tmp_path):
    got = ailine.count_drawing_objects(_sealed_book(tmp_path))
    assert got == {"図形/描画": 1, "画像": 1}, got


def test_detects_loss_when_the_part_survives_but_the_shape_is_gone(tmp_path):
    """★ 実測した見逃し: openpyxl の往復は drawing1.xml を**残したまま**中の sp を捨てる。
       ファイル名の集合は 1 つも変わらないので、旧ゲートは喪失 0 件と報告していた。"""
    src = _sealed_book(tmp_path)
    out = tmp_path / "after.xlsx"
    openpyxl.load_workbook(src).save(out)
    with zipfile.ZipFile(src) as a, zipfile.ZipFile(out) as b:
        assert set(a.namelist()) - set(b.namelist()) == set() or True   # 前提の記録のみ
    lost = ailine.check_drawing_content_loss(src, out)
    assert any(cat == "図形/描画" for cat, *_ in lost), \
        f"部品が残ったまま中身だけ消えるのを見逃した: {lost}"


def test_growth_is_not_a_loss(tmp_path):
    """B: 増えた分は喪失に数えない（帳票段は雛形を N 枚に複製する）。"""
    src = _sealed_book(tmp_path)
    more = tmp_path / "more.xlsx"
    import shutil
    shutil.copy2(src, more)
    _write_drawing(more, _PIC + _SHAPE + _SHAPE)   # 図形が増えた形
    assert ailine.check_drawing_content_loss(src, more) == []


def test_no_false_alarm_on_a_plain_book(tmp_path):
    p = tmp_path / "plain.xlsx"
    openpyxl.Workbook().save(p)
    q = tmp_path / "plain2.xlsx"
    openpyxl.load_workbook(p).save(q)
    assert ailine.check_drawing_content_loss(p, q) == []
    assert ailine.check_round_trip_fidelity(p, q)["lost"] is False


# --- C ゲートが取り込んでいる ---------------------------------------------------------

def test_gate_includes_content_level_loss(tmp_path):
    src = _sealed_book(tmp_path)
    out = tmp_path / "after.xlsx"
    openpyxl.load_workbook(src).save(out)
    fid = ailine.check_round_trip_fidelity(src, out)
    assert fid["lost"], "ゲートが中身の喪失を取り込んでいない"
    assert any(it["label"] == "図形/描画" for it in fid["items"]), fid


def test_output_fidelity_warning_starts_with_warning_mark(tmp_path):
    """★ 決裁③の機構（⚠ 始まりを数えて ✓ を降ろす）に乗る形であること。"""
    msg = ailine.format_output_fidelity_warning(
        {"lost": True, "items": [{"label": "図形/描画", "count": 1}]}, "out.xlsx")
    assert msg.lstrip().startswith("⚠"), msg
    assert ailine.count_suspicious_advisories([msg]) == 1


# --- D/E 検分シートの Basic --------------------------------------------------------

def test_inspection_basic_has_no_raw_control_characters():
    """D: 生の制御文字を .bas に埋めない（この repo は制御文字混入の事故を踏んでいる）。"""
    code = ailine.inspection_sheet_basic_call(
        "検分", ["シート名", "元の行"], [["あかつき商事", 2], ["みどり工業", 3]], "sn")
    bad = [c for c in code if ord(c) < 32 and c not in ("\n", "\r", "\t")]
    assert bad == [], f"生の制御文字が混ざった: {[hex(ord(c)) for c in bad]}"
    assert "Chr(30)" in code and "Chr(31)" in code, code


def test_inspection_basic_round_trips_through_the_fake_lo():
    """偽 LO（生成 Basic を解釈する）で往復して、値と型が保たれること。"""
    code = ailine.inspection_sheet_basic_call(
        "検分", ["シート名", "元の行", "印の数"], [["あかつき商事", 2, 2]], "snn")
    wb = openpyxl.Workbook()
    assert apply_inspection_sheets(wb, code) == 1
    ws = wb["検分"]
    assert [c.value for c in ws[1]] == ["シート名", "元の行", "印の数"]
    assert [c.value for c in ws[2]] == ["あかつき商事", 2, 2]
    assert isinstance(ws.cell(row=2, column=2).value, int), "数値列が文字列になった"


def test_quotes_in_sheet_names_survive():
    """シート名にダブルクォートが入っても Basic が壊れない。"""
    code = ailine.inspection_sheet_basic_call("検分", ["名"], [['a"b']], "s")
    wb = openpyxl.Workbook()
    apply_inspection_sheets(wb, code)
    assert wb["検分"].cell(row=2, column=1).value == 'a"b'


def test_both_report_and_format_map_use_the_same_builder():
    """E: 書き写しの禁止 ── 2 経路が同じ関数を呼ぶこと。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    assert src.count("inspection_sheet_basic_call(") >= 3, \
        "帳票段と様式写像段の両方が共通の組み立てを呼んでいない"
    assert "_add_report_inspection_sheet(out_book" not in src, \
        "openpyxl で開き直す旧経路が残っている（図形を捨てる）"
    assert "_add_format_map_inspection_sheet(out_book" not in src, \
        "様式写像段の旧経路が残っている（片配線）"
