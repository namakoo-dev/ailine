"""EXTRACT の端から端（`ailine.main(argv)` を実際に叩く）。tests/test_golden_transcripts.py の
_isolate/_run_main を流用する（EXTRACT op 実装（コミット 2edcb08）の方針どおり、
既存の f9_transcripts golden には一切触れない ── 新規の直接アサーション形式のテスト）。

★ 昨夜の実弾2件を、端から端で再現・退治する:
  - 抽出成功で exit 0・✓・新シートの中身が値と型ごと正しいこと
    → test_extract_success_end_to_end
  - 「できたふり」（空シートを作って exit 0）を事後条件が捕まえること
    → test_extract_fake_empty_sheet_apply_fails_postcondition
  - 単位H の2例目: 同じ抽出を2回実行しても、前回の出力シートの作り直しが
    破壊の関所(単位F/G)で止まらないこと
    → test_extract_second_run_does_not_gate
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402


def _translate_extract(monkeypatch, col="金額", cmp="gte", value=40000):
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "EXTRACT", "args": {"col": col, "cmp": cmp, "value": value}})


def test_extract_success_end_to_end(tmp_path, monkeypatch, capsys):
    """抽出成功: exit 0・✓ が出る・新シート『金額40000以上』の中身が値も型も正しいこと
       （数値セルが int/float のまま。文字列に化けていないこと ── 昨夜の実弾の逆）。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 30000], ["b", 50000], ["c", 45000]])
    _translate_extract(monkeypatch)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        out = wb.create_sheet("金額40000以上")
        out.append(["商品", "金額"])
        out.append(["b", 50000])
        out.append(["c", 45000])
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    rc, out = _run_main(["run", str(book), "金額が40000以上の行だけを別シートに抜き出して",
                          "--copy"], capsys)
    assert rc == 0, out
    assert "✓" in out, out
    assert "3行中2行が一致" in out and "2行を抽出" in out, out

    out_book = book.with_name(book.stem + ".out" + book.suffix)
    ws = openpyxl.load_workbook(out_book)["金額40000以上"]
    assert [c.value for c in ws[1]] == ["商品", "金額"]
    assert ws.cell(row=2, column=1).value == "b"
    assert ws.cell(row=2, column=2).value == 50000
    assert isinstance(ws.cell(row=2, column=2).value, (int, float)), (
        f"金額セルが数値でない（文字列化バグの再現）: {ws.cell(row=2, column=2).value!r}"
    )
    assert ws.cell(row=3, column=1).value == "c"
    assert ws.cell(row=3, column=2).value == 45000


def test_extract_fake_empty_sheet_apply_fails_postcondition(tmp_path, monkeypatch, capsys):
    """★ 昨夜の「できたふり」の再現: 一致する行が実際にはあるのに、basrun_apply が
       見出しだけの空シートを作って exit 0 を返す（列抽出の実弾と同型）。
       事後条件(check_extract)が独立に数え直して捕まえ、exit 0 にならないこと。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 30000], ["b", 50000], ["c", 45000]])
    _translate_extract(monkeypatch)

    def fake_apply_empty(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        out = wb.create_sheet("金額40000以上")
        out.append(["商品", "金額"])   # 見出しだけ・データ行を1件も書かない（できたふり）
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply_empty)

    rc, out = _run_main(["run", str(book), "金額が40000以上の行だけを別シートに抜き出して",
                          "--copy"], capsys)
    assert rc != 0, f"できたふりの空シートが exit 0 で通ってしまった:\n{out}"
    assert "適用されたが事後条件を満たさない" in out, out
    assert "✓" not in out, out   # ★ 機械検証済みを絶対に名乗らないこと


def test_extract_second_run_does_not_gate(tmp_path, monkeypatch, capsys):
    """★★ 単位H の2例目: 前回の EXTRACT 出力（見出し = 元シートの見出しのコピー、という
       自分自身の署名）が既にあるシートへ、同じ条件でもう一度 EXTRACT を実行する。
       元データが増えている（正常な作り直し）ので、破壊の関所(単位F/G)で止まらず exit 0
       になること（AGGREGATE の単位H 実測(F4/test_unit_h_specimens.py)と同型）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工事台帳"
    for r in [["商品", "金額"], ["a", 30000], ["b", 50000], ["c", 45000]]:
        ws.append(r)
    out = wb.create_sheet("金額40000以上")
    # ★ 前回の EXTRACT 自身の出力の署名そのもの（見出し行 = 元シートの見出し行のコピー）。
    for r in [["商品", "金額"], ["b", 50000], ["c", 45000]]:
        out.append(r)
    wb.save(p)

    _translate_extract(monkeypatch)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        if "金額40000以上" in wb.sheetnames:
            del wb["金額40000以上"]
        s = wb.create_sheet("金額40000以上")
        for r in [["商品", "金額"], ["b", 50000], ["c", 45000], ["d", 60000]]:
            s.append(r)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)

    # 元データが増えた状態(d=60000 を先に足す)で2回目を実行 — 依頼文でシートを名指しする
    # （--sheet フラグでは単位Eの②注記が出て検体から無関係な変数が増えるため、F4 と同じ
    #   「語で指す」形にする）。
    wb2 = openpyxl.load_workbook(p)
    wb2["工事台帳"].append(["d", 60000])
    wb2.save(p)

    rc, out = _run_main(
        ["run", str(p), "工事台帳の金額が40000以上の行だけを別シートに抜き出して"], capsys)
    # ★ 本題の主張: 破壊の関所(単位F/G)で止まらないこと（旧実装なら「新しいシートを作る
    #   はずが、既存のシート『金額40000以上』の値を書き換えました」→非対話で exit 7）。
    assert rc == 0, f"rc={rc}\n{out}"
    assert "新しいシートを作るはずが" not in out, out
    # ★ 「変更が元データの範囲外です」は別の助言(幽霊データ検出)で、単位F/G/Hの対象外
    #   （出力シートの行数が前回より増える＝EXTRACTの正常な作り直しでは毎回起きうる）。
    #   ここでは「exit 7 で止まらない」ことだけを主張する。
    assert "事後条件を確認" in out, out


@pytest.mark.local
def test_contains_on_numeric_column_agrees_between_basic_and_checker(tmp_path, monkeypatch, capsys):
    """★ review3#2/#4 critical: contains の意味論を検算側だけ文字列限定に変えたため、
       Basic 書き手（getString の部分一致・数値 140000 も『40』にマッチ）と分裂 ──
       数値列への contains が必ず事後条件 fail になる。契約: 両実装とも
       『contains は文字列セルのみ』で一致し、数値列 contains は 0 行一致で
       事後条件が破れずに通ること。★ 実機（LO・basrun）検体 ── モックしない。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["注文ID", "金額"], ["J-1", 140000], ["J-2", 30000]])
    _translate_extract(monkeypatch, col="金額", cmp="contains", value="40")
    rc, out = _run_main(["run", str(book), "金額に40を含む行を別シートに抜き出して",
                          "--copy"], capsys)
    assert rc == 0, f"書き手と検算が分裂している（事後条件 fail か？）:\n{out}"
    assert "0行" in out or "0 行" in out, f"数値列 contains は 0 行一致のはず:\n{out}"


@pytest.mark.local
def test_eq_on_numeric_column_agrees_between_basic_and_checker(tmp_path, monkeypatch, capsys):
    """★ jisaku-review 4戦目 F4（minor）: eq の許容誤差（TOLERANCE）の実機（LO・basrun）
       検体が無かった ── 上の contains 検体と同型の欠けを埋める回帰番人。単一ブック eq
       （金額が30000の行）で Basic 書き手と検算（事後条件）が一致し、1行抽出で exit 0
       となること。★ 実機（LO・basrun）検体 ── モックしない。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["注文ID", "金額"], ["J-1", 30000], ["J-2", 40000]])
    _translate_extract(monkeypatch, col="金額", cmp="eq", value=30000)
    rc, out = _run_main(["run", str(book), "金額が30000の行を別シートに抜き出して",
                          "--copy"], capsys)
    assert rc == 0, f"書き手と検算が分裂している（事後条件 fail か？）:\n{out}"
    assert "1行" in out or "1 行" in out, f"eq 一致は1行のはず:\n{out}"
