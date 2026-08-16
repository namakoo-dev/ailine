"""C1-F6: `ailine run` 終了コード表（tests/golden/f6_exit_codes.md）の裏取り。

表に書いた {0,1,3,4,5,6,7,8} の各条件を、既存 test_ailine.py が使っている最小限の
monkeypatch パターンを再利用して1関数=1裏取りに集約する（散らばっている既存の
個別テストは残したまま、ここは『表そのものが正しいか』を機械的に確認する専用）。

★ 2 が欠番である理由の調査結果は tests/golden/f6_exit_codes.md 本文を参照
（結論: argparse.ArgumentParser.error() が予約している標準ライブラリの既定動作で、
ailine.py 自身のコードではない）。
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _cf_dv_book(tmp_path, name, add_cf=True, add_dv=True):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["りんご", 100])
    ws.append(["バナナ", 200])
    if add_cf:
        ws.conditional_formatting.add("B2:B3", CellIsRule(operator="greaterThan", formula=["150"]))
    if add_dv:
        dv = DataValidation(type="list", formula1='"a,b,c"')
        ws.add_data_validation(dv)
        dv.add("C2")
    wb.save(p)
    return p


def _patch_lossy_normalize(monkeypatch):
    def fake_normalize(book, workdir, timeout=None):
        norm = workdir / ("normalized" + book.suffix)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["商品", "金額"])
        ws.append(["りんご", 100])
        ws.append(["バナナ", 200])
        wb.save(norm)
        return norm
    monkeypatch.setattr(ailine, "normalize_book", fake_normalize)


def _fidelity_gate_ns(book, **overrides):
    base = dict(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, json=False, timeout=180.0, ask=False,
        accept_loss=False, copy=False, allow_freeform=True)
    base.update(overrides)
    return argparse.Namespace(**base)


def _gate_ns(**overrides):
    base = dict(inplace=True, dry=False, ask=False, overwrite=False)
    base.update(overrides)
    return argparse.Namespace(**base)


def _fgate_ns(**overrides):
    base = dict(allow_freeform=False)
    base.update(overrides)
    return argparse.Namespace(**base)


def test_exit_0_success(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                         lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    ns = argparse.Namespace(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    rc = ailine.cmd_run(ns)
    assert rc == 0


def test_exit_1_generic_failure(monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    rc = ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告")
    captured = capsys.readouterr()
    assert rc == 1
    assert "中止した" in captured.out


def test_exit_2_is_argparse_reserved_not_ailine_own(capsys):
    """★ ailine.py 自身は sys.exit(2)/return 2 を一度も書いていない
       （tests/golden/f6_exit_codes.md 参照）。2 が出るのは argparse 自身の
       ArgumentParser.error() の既定動作であることを実測する。"""
    with pytest.raises(SystemExit) as exc1:
        ailine.build_parser().parse_args(["--bogus-flag"])
    assert exc1.value.code == 2

    with pytest.raises(SystemExit) as exc2:
        ailine.build_parser().parse_args(["badsubcommand"])
    assert exc2.value.code == 2

    import re
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    own_codes = set(int(m) for m in re.findall(r"(?:return|sys\.exit\()\s*(\d+)\b", src))
    assert 2 not in own_codes, "ailine.py が独自に exit code 2 を使い始めた（表を更新すること）"


def test_exit_3_clarify(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "build_struct_dump", lambda book, workdir: ambiguous)
    ns = argparse.Namespace(
        book=str(book), task="いい感じにして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.cmd_run(ns)
    assert rc == 3


def test_exit_4_fidelity_gate(tmp_path, monkeypatch, capsys):
    book = _cf_dv_book(tmp_path, "book.xlsx", add_cf=True, add_dv=False)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    _patch_lossy_normalize(monkeypatch)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda *a, **k: {"op": "FREEFORM", "args": {}})
    rc = ailine.cmd_run(_fidelity_gate_ns(book))
    captured = capsys.readouterr()
    assert rc == 4
    assert "--accept-loss" in captured.out
    assert "--copy" in captured.out


def test_exit_5_excel_lock(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    (tmp_path / f"~${book.name}").write_bytes(b"lock")
    ns = argparse.Namespace(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.cmd_run(ns)
    assert rc == 5


def test_exit_6_run_lock_busy(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    lock_path = tmp_path / "run.lock"
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", lock_path)
    other_pid = 999999
    lock_path.write_text(json.dumps({"pid": other_pid, "ts": ailine.datetime.now(ailine.timezone.utc)
                                     .isoformat(timespec="seconds")}), encoding="utf-8")
    monkeypatch.setattr(ailine, "_pid_alive", lambda pid: pid == other_pid)
    ns = argparse.Namespace(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    rc = ailine.cmd_run(ns)
    assert rc == 6


def test_exit_7_overwrite_gate_noninteractive(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    rc = ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告")
    captured = capsys.readouterr()
    assert rc == 7
    assert "--overwrite" in captured.out
    assert "--copy" in captured.out


def test_exit_8_freeform_gate_noninteractive(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    rc = ailine._confirm_freeform_apply(_fgate_ns())
    captured = capsys.readouterr()
    assert rc == 8
    assert "--allow-freeform" in captured.out
