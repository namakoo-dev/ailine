# 存在しない列を頼まれたとき、黙って別の列でやらない（2026-08-24）。
#
# ★ 盲検の査定者が値段を決めた一点（実測の再現）:
#     ブック: 日付/取引先/商品/数量/単価（**金額列は無い**）
#     依頼:  「金額で降順に並べ替えて」
#     結果:  7B が args を col="数量" に**差し替え**、機械は実在列なので通し、
#            **✓ 機械検証済み** を出して原本相当を書き換えた。
#     査定者の言葉: 「人間なら『金額って列、ありませんけど？』と聞き返す。
#     この道具は聞き返さず、原本を書き換え、緑のチェックを付ける。」
#
# ★ なぜ既存の番人が鳴らなかったか: 対象スロットの 3 段階（①照合できた ②無言 ③矛盾）で
#   これは ② に落ちていた。② は「★ ただし…機械決定しました」の 1 文を出すだけで ✓ を保つ設計。
#   だが**依頼者は名指ししている**（無言ではない）。無言と矛盾を分ける材料は既にあった ──
#   残差（依頼文のうち、どの引数にも使われなかった語）。実測:
#     「金額で降順に並べ替えて」(col=数量) → 残差 ['金額']   ← 捏造
#     「数量で降順に並べ替えて」(col=数量) → 残差 []        ← 正しい
#     「降順に並べ替えて」    (col=数量) → 残差 []        ← 無指定（②のまま）
#   誤爆しない形で切り分けられる。
#
# 契約:
#   ① 対象列が依頼文に無く、かつ依頼文に**使われなかった語**が残っているなら、
#      適用する前に断り、**両方を名指しする**（無い列名と、ある列の一覧）
#   ② 誤爆しない: 対象列が依頼文にある時／依頼文が列を言っていない時は、今までどおり通る

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "fabricated_subject_refusal"),
    reason="捏造列の関所 未実装（契約は凍結済み）",
    strict=True,
)


def _book(tmp_path):
    p = tmp_path / "uriage.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(["日付", "取引先", "商品", "数量", "単価"])
    ws.append(["2026-04-01", "甲", "A", 3, 100])
    ws.append(["2026-04-02", "乙", "B", 9, 200])
    wb.save(p)
    return p


def _sort_to(col):
    return lambda model, task, book_meta, temperature=0.1: {
        "plan": [{"op": "SORT", "args": {"col": col, "order": "desc"}}]}


@needs_impl
def test_fabricated_column_is_refused_and_both_named(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _sort_to("数量"))
    called = []
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda *a, **k: called.append(1) or (True, None, "ok"))
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert "✓" not in out, f"存在しない列を頼まれて別の列で ✓ を出した: {out}"
    assert "金額" in out, f"依頼文の『金額』を名指ししていない: {out}"
    assert "数量" in out, f"ある列の一覧を出していない: {out}"
    assert not called, "断る前に LibreOffice へ適用した（原本に触れる恐れ）"


# ★ needs_impl を付けない ── 実装前から緑（今の挙動を守る誤爆の柵）。
def test_named_existing_column_still_passes(tmp_path, monkeypatch, capsys):
    """誤爆防止①: 実在する列を名指しした依頼は今までどおり通る。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _sort_to("数量"))

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["売上表"]
        ws["A2"], ws["B2"], ws["C2"], ws["D2"], ws["E2"] = "2026-04-02", "乙", "B", 9, 200
        ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"] = "2026-04-01", "甲", "A", 3, 100
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake)
    rc, out = _run_main(["run", str(book), "数量で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    assert "✓" in out, f"正しい依頼が止められた（誤爆）: {out}"


# ★ needs_impl を付けない ── 実装前から緑（今の挙動を守る誤爆の柵）。
def test_unspecified_column_is_not_treated_as_fabrication(tmp_path, monkeypatch, capsys):
    """誤爆防止②: 列を言っていない依頼は、今までどおり ② の注記つきで通る。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _sort_to("数量"))

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["売上表"]
        ws["A2"], ws["B2"], ws["C2"], ws["D2"], ws["E2"] = "2026-04-02", "乙", "B", 9, 200
        ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"] = "2026-04-01", "甲", "A", 3, 100
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake)
    rc, out = _run_main(["run", str(book), "降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    assert "機械決定しました" in out, f"② の注記が消えた: {out}"


# ★ 誤爆③（実装中に実測で見つけた・2026-08-24）: 依存つき連鎖を撃ってはいけない。
#   「売上から原価を引いた利益列を作って、利益で降順に並べ替えて」
#   → 1段目が新しい列を作るが、機械は『売上-原価』と自動命名する。
#   → 2段目の『利益』はブックに無いが、これは**捏造ではなく前段の列の呼び名**。
#   関所の初版はこれを断ってしまい、既存の連鎖テスト 2 本が赤くなった（帰属は git stash で確認）。
#   印（_chained_new_column）を立てて関所を免除する。

def test_chained_new_column_name_is_not_treated_as_fabrication():
    """前段が作った列への言及は関所を通す（印が立っていれば断らない）。"""
    meta = {"sheets": ["Sheet"],
             "headers": {"Sheet": ["商品", "売上", "原価", "売上-原価"]},
             "header_rows": {"Sheet": 1}}
    task = "売上から原価を引いた利益列を作って、利益で降順に並べ替えて"
    # 印が無ければ関所が鳴る（＝この試験が恒真でないことの証明）
    assert ailine.fabricated_subject_refusal(
        "SORT", {"col": "売上-原価", "order": "desc"}, meta, task, "Sheet") is not None
    # 印があれば通る
    assert ailine.fabricated_subject_refusal(
        "SORT", {"col": "売上-原価", "order": "desc", "_chained_new_column": "売上-原価"},
        meta, task, "Sheet") is None
