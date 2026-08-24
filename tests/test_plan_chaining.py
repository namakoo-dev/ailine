# 複合計画の「連鎖しない」を鳴らす番人 ── 修正より先に凍結した赤い検体（2026-08-24）。
#
# ★ 実測した事故（実 7B + 実 LibreOffice・scratchpad/probe3/chain.xlsx）:
#     依頼「売上が60以上の行だけ現場ごとに集計して」
#     1段目 EXTRACT(売上 gte 60) → 新シート『売上60以上』(3行) … 正しい
#     2段目 AGGREGATE(現場,売上) → **元の『売上』を集計** → A=150 / 総計420
#     正解は A=100 / 総計370。**それでも ✓ が出た。**
#   各段の事後条件はどちらも真（抽出は正しい・集計も「そのシートの」集計として正しい）。
#   嘘は段の中でなく**段と段の間**にあり、どの番人も段の中しか見ていなかった。
#
# ★ なぜ ⚠ で止めるのか（自動で連鎖させないのか）: 「抽出して、**元表に**合計を追加して」の
#   ような計画もありうる。どちらの意図かは機械には決まらない。決まらないものを黙って
#   決めるのが今回の事故そのものなので、**名指しして人に返す**（決裁③により ✓→△）。
#   ★ 連鎖そのものは、シート名を依頼文に書けば今でもできる（『売上60以上』を集計して）。
#
# 契約:
#   ① 前段が派生シート（EXTRACT/DEDUP の出力）を作ったのに、後段が元シートを見ていたら
#      ⚠ を出し、両方のシート名を名指しする
#   ② その run は ✓ を名乗らない（決裁③ の降格に乗る）
#   ③ 誤爆しない: 後段が派生シートを見ている計画では ⚠ を出さない

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "PLAN_CHAIN_WARNING_OPS"),
    reason="連鎖の番人 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


def _sales(tmp_path):
    p = tmp_path / "chain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in [["現場", "売上"], ["A", 50], ["A", 100], ["B", 70], ["B", 200]]:
        ws.append(r)
    wb.save(p)
    return p


def _fake_apply_extract_then_agg(out_book, code, workdir, helper_files=(), timeout=None):
    """実 LO の代わり: コードの中身を見て、抽出なら派生シート・集計なら集計シートを作る。
       ★ 窒息点は basrun_apply（実 LO を呼ぶ唯一の門）── ここより上は本物を通す。"""
    wb = openpyxl.load_workbook(out_book)
    if "ExtractRows" in code:
        if "売上60以上" in wb.sheetnames:
            del wb["売上60以上"]
        sh = wb.create_sheet("売上60以上")
        sh.append(["現場", "売上"])
        for row in wb["売上"].iter_rows(min_row=2, values_only=True):
            if row[1] is not None and row[1] >= 60:
                sh.append(list(row))
    else:
        if "集計" in wb.sheetnames:
            del wb["集計"]
        sh = wb.create_sheet("集計")
        sh.append(["現場", "合計 - 売上"])
        sh.append(["A", 150])       # ★ 事故の再現: 元シートを集計した値
        sh.append(["B", 270])
    wb.save(out_book)
    return True, None, "ok"


# ★ 契約の変更（2026-08-24 夕・理由つき）: この検体は初版で「連鎖しない・⚠ を出して降格する」
#   を凍結していた。同日、盲検の 3 人が全員この形で詰まり（台帳の最頻の形＝期間で絞って集計が
#   一度も完遂しない）、Namakoo の指示「問題点があれば修正か**機能追加**で対応」に沿って
#   **連鎖を実装した**。よってこの段は今は成功する ── 検体は「⚠ で止まる」から
#   「連鎖して、連鎖したことを人に見せる」へ、理由を書いた上で差し替える。
#   ★ 連鎖**しない**場合の番人（人がシート名を明示した時）は下に残してある。

@needs_impl
def test_later_step_chains_and_discloses(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book = _sales(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1: {"plan": [
            {"op": "EXTRACT", "args": {"col": "売上", "cmp": "gte", "value": 60}},
            {"op": "AGGREGATE", "args": {"group_col": "現場", "value_col": "売上"}},
        ]})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_extract_then_agg)
    rc, out = _run_main(["run", str(book), "売上が60以上の行だけ現場ごとに集計して", "--copy"], capsys)
    assert "売上60以上" in out, f"連鎖したことを人に見せていない: {out}"
    assert "前段が作った" not in out, f"連鎖したのに「元の表を見ています」の ⚠ が残った: {out}"


# ★ 治具の訂正（封印者ナギ・2026-08-24）: 初版の誤爆検体は AGGREGATE の args に
#   target_sheet を渡していたが、そんな引数は無い（対象シートは**依頼文**から解決される）。
#   しかも派生シートは計画を立てた時点では存在しないので、**同じ run の中では指名できない**
#   ── これは実装の穴でなく設計上の順序。誤爆の柵は現実に起こる形に置き換えた。
#   assert の意図（誤爆しない）は不変。

def test_no_warning_when_no_derived_sheet_was_made(tmp_path, monkeypatch, capsys):
    """誤爆防止①: 派生シートを作らない計画（並べ替え→合計追加）では ⚠ を出さない。"""
    _isolate(monkeypatch, tmp_path)
    book = _sales(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1: {"plan": [
            {"op": "SORT", "args": {"col": "売上", "order": "desc"}},
            {"op": "APPEND_TOTAL", "args": {"col": "売上"}},
        ]})

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["売上"]
        if "合計" not in [c.value for c in ws["A"]]:
            rows = sorted([(r[0], r[1]) for r in ws.iter_rows(min_row=2, values_only=True)
                            if r[0] is not None], key=lambda t: -t[1])
            for j, (n, v) in enumerate(rows, start=2):
                ws.cell(j, 1, n); ws.cell(j, 2, v)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake)
    rc, out = _run_main(["run", str(book), "売上の降順に並べ替えて合計を追加して", "--copy"], capsys)
    assert "前段が作った" not in out, f"派生シートが無いのに連鎖の ⚠ が出た（誤爆）: {out}"


def test_no_warning_when_the_derived_step_is_last(tmp_path, monkeypatch, capsys):
    """誤爆防止②: 抽出が最後の段なら、後段が無いので ⚠ は出ない。"""
    _isolate(monkeypatch, tmp_path)
    book = _sales(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1: {"plan": [
            {"op": "SORT", "args": {"col": "売上", "order": "desc"}},
            {"op": "EXTRACT", "args": {"col": "売上", "cmp": "gte", "value": 60}},
        ]})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_extract_then_agg)
    rc, out = _run_main(["run", str(book), "売上の降順に並べ替えて60以上を抽出して", "--copy"], capsys)
    assert "前段が作った" not in out, f"最終段の抽出で ⚠ が出た（誤爆）: {out}"


# --- 連鎖そのものを実装する（2026-08-24 夕）------------------------------------------
#
# ★ なぜ ⚠ で止めるだけにしなかったか: 台帳の最頻の形（期間で絞って集計）が、
#   ⚠ だけでは**一度も完遂しない**。盲検の 3 人が全員ここで詰まり、うち 1 人は
#   「期間を無視した数字を、ツール自身が作業結果として案内する」と書いた。
#   Namakoo の指示は「問題点があれば修正か**機能追加**で対応」── 連鎖はその機能追加。
#
# ★ 安全のための条件（3 つ全部が揃った時だけ連鎖する）:
#   ① 直前までの段が派生シート（EXTRACT/DEDUP の出力＝絞り込んだ同じ表）を作っている
#   ② その段が表の行を読む op である
#   ③ 依頼文が**シート名を明示していない**（明示があれば人の指定が勝つ）
#   そして連鎖したことは必ず解釈行に出す（黙って対象を変えない）。
#   連鎖**しなかった**場合の ⚠ は今までどおり残す。

def test_extract_then_aggregate_chains_onto_the_derived_sheet(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book = _sales(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1: {"plan": [
            {"op": "EXTRACT", "args": {"col": "売上", "cmp": "gte", "value": 60}},
            {"op": "AGGREGATE", "args": {"group_col": "現場", "value_col": "売上"}},
        ]})
    seen = []

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        seen.append(code)
        wb = openpyxl.load_workbook(out_book)
        if "ExtractRows" in code:
            if "売上60以上" in wb.sheetnames:
                del wb["売上60以上"]
            sh = wb.create_sheet("売上60以上")
            sh.append(["現場", "売上"])
            for row in wb["売上"].iter_rows(min_row=2, values_only=True):
                if row[1] is not None and row[1] >= 60:
                    sh.append(list(row))
        else:
            if "集計" in wb.sheetnames:
                del wb["集計"]
            sh = wb.create_sheet("集計")
            sh.append(["現場", "合計 - 売上"]); sh.append(["A", 100]); sh.append(["B", 270])
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake)
    rc, out = _run_main(["run", str(book), "売上が60以上の行だけ現場ごとに集計して", "--copy"], capsys)
    assert len(seen) == 2, out
    # ★ 2段目の Basic が、派生シートを対象にしていること（wrap_basic_for_sheet が先頭へ動かす）
    assert "売上60以上" in seen[1], f"2段目が派生シートを対象にしていない: {seen[1]}"
    assert "売上60以上" in out, f"連鎖したことを人に見せていない: {out}"
    assert "前段が作った" not in out, f"連鎖したのに ⚠ が残っている: {out}"


def test_explicitly_named_sheet_beats_the_chain(tmp_path, monkeypatch, capsys):
    """③ 人がシート名を書いたら、その指定が連鎖より勝つ（黙って上書きしない）。"""
    _isolate(monkeypatch, tmp_path)
    book = _sales(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1: {"plan": [
            {"op": "EXTRACT", "args": {"col": "売上", "cmp": "gte", "value": 60}},
            {"op": "AGGREGATE", "args": {"group_col": "現場", "value_col": "売上"}},
        ]})
    seen = []

    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        seen.append(code)
        wb = openpyxl.load_workbook(out_book)
        if "ExtractRows" in code:
            sh = wb.create_sheet("売上60以上")
            sh.append(["現場", "売上"]); sh.append(["A", 100])
        else:
            sh = wb.create_sheet("集計")
            sh.append(["現場", "合計 - 売上"]); sh.append(["A", 150])
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake)
    rc, out = _run_main(["run", str(book), "60以上を抜き出してから売上シートを現場ごとに集計して",
                          "--copy"], capsys)
    assert len(seen) == 2, out
    assert "売上60以上" not in seen[1], "人が『売上シート』と指定したのに連鎖で上書きした"
