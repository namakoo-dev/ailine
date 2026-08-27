# 名指しの抽出 ── 2026-08-27。Namakoo「特定行や特定列の抜き出し（抽出）ができない」
#
# ★★ 実測した危なさ: 「みかんの行とりんごの行だけを抽出して」に対し、一段目は
#   `contains "リンゴ"`（片仮名の幻覚）や `eq "みかんとりんご"`（連結）を返していた。
#   どちらも **0 行に当たる** ── 空の抽出結果が ✓ で出るところだった。
#   「動かない」より悪い。**空でも正しく見える**からだ。
#
# 契約:
#   ① 比較語（以上/以下…）が依頼文に**無い**なら、それは条件でなく**名指し**
#   ② 値は LLM に作らせない ── その列に**実在する値**のうち依頼文に現れるものだけ拾う
#   ③ 部分一致にしない（「りんご」が「青りんご」に当たると頼んでいない行が混じる）
#   ④ 列の抽出: 残す列も依頼文の実在列から。全列指定は断る（抜き出す意味がない）
#   ⑤ 列の抽出の事後条件: 見出しの並び・行数・値と型・**元シートが無変更**

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

ROWS = [["商品", "売上", "原価"], ["りんご", 1200, 700],
         ["みかん", 800, 300], ["ぶどう", 1500, 900], ["青りんご", 300, 100]]


def _book(tmp_path, rows=None, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in (rows or ROWS):
        ws.append(r)
    wb.save(p)
    return p


def _meta(path, headers=("商品", "売上", "原価")):
    return {"sheets": ["売上"], "headers": {"売上": list(headers)},
            "header_rows": {"売上": 1}, "path": str(path)}


# --- ①②③ 行を名前で抽出 ---------------------------------------------------------------

def test_named_rows_become_an_in_list(tmp_path):
    ok, r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "商品", "cmp": "contains", "value": "リンゴ"},
        _meta(_book(tmp_path)), task="みかんの行とりんごの行だけを抽出して")
    assert ok, err
    assert r["cmp"] == "in" and r["value"] == ["みかん", "りんご"], (r["cmp"], r["value"])
    assert any("実在の値" in w for w in r.get("_warnings", [])), r.get("_warnings")


def test_a_single_named_row_becomes_eq(tmp_path):
    ok, r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "商品", "cmp": "contains", "value": "ミカン"},
        _meta(_book(tmp_path)), task="みかんの行だけ抜き出して")
    assert ok, err
    assert (r["cmp"], r["value"]) == ("eq", "みかん")


def test_a_longer_name_wins_over_its_substring(tmp_path):
    """③ 「青りんご」と「りんご」が両方在る表で、長い方を先に当てて位置を潰す。"""
    got = ailine.task_names_real_values(
        "青りんごだけ抜き出して", _meta(_book(tmp_path)), "売上", "商品")
    assert got == ["青りんご"], got


def test_a_comparison_request_is_left_alone(tmp_path):
    """★ 恒真殺し: 比較語が在る依頼を名指しと読まない（500 を名前にしない）。"""
    ok, r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "原価", "cmp": "gte", "value": 500},
        _meta(_book(tmp_path)), task="原価が500以上の行を抜き出して")
    assert ok, err
    assert r["cmp"] == "gte" and r["value"] == 500.0


def test_the_in_predicate_is_whole_value_not_substring():
    match = ailine._extract_predicate("in", ["りんご"])
    assert match("りんご") and not match("青りんご") and not match("りん")


def test_codegen_passes_the_list_with_a_separator(tmp_path):
    meta = _meta(_book(tmp_path))
    ok, r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "商品", "cmp": "eq", "value": "x"}, meta,
        task="みかんの行とりんごの行だけを抽出して")
    assert ok, err
    code = ailine.codegen_dsl("EXTRACT", r, meta)
    assert 'Chr(2)' in code and '"みかん"' in code and '"りんご"' in code, code


def test_the_bas_matches_whole_values_for_the_in_case():
    """★ 意味論は 3 箇所が同時に持つ（ここ / Basic / 凍結した真理値表）。"""
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    assert "Case 6" in bas and "Chr(2) & CStr(cmpValue) & Chr(2)" in bas, "丸ごと一致の実装が無い"


# --- ④⑤ 列を抽出 -----------------------------------------------------------------------

def test_columns_are_taken_from_the_request_in_order(tmp_path):
    ok, r, _i, err = ailine.verify_dsl_args(
        "EXTRACT_COLUMNS", {"cols": ["原価"]}, _meta(_book(tmp_path)),
        task="商品と売上の列だけ抜き出して")
    assert ok, err
    assert r["cols"] == ["商品", "売上"], r["cols"]     # ★ LLM の『原価』は採らない
    assert r["_new_sheet"] == "商品・売上だけ"


def test_all_columns_is_refused(tmp_path):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "EXTRACT_COLUMNS", {}, _meta(_book(tmp_path)),
        task="商品と売上と原価の列だけ抜き出して")
    assert not ok and "全部の列" in err, err


def test_no_recognisable_column_is_refused(tmp_path):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "EXTRACT_COLUMNS", {}, _meta(_book(tmp_path)), task="必要な列だけ抜き出して")
    assert not ok and "読み取れません" in err, err


def test_the_request_is_recognised():
    assert ailine.task_asks_to_extract_columns("商品と売上の列だけ抜き出して")
    assert ailine.task_asks_to_extract_columns("必要な列のみ取り出して")
    assert not ailine.task_asks_to_extract_columns("みかんの行だけ抜き出して")
    assert not ailine.task_asks_to_extract_columns("原価の列を削除して")


def _out_book(tmp_path, cols, rows, name="out.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in ROWS:
        ws.append(r)
    out = wb.create_sheet("商品・売上だけ")
    out.append(cols)
    for r in rows:
        out.append(r)
    wb.save(p)
    return p


_XC = {"cols": ["商品", "売上"], "_new_sheet": "商品・売上だけ", "_header_row": 1}


def test_a_correct_column_extract_passes(tmp_path):
    p = _out_book(tmp_path, ["商品", "売上"],
                   [["りんご", 1200], ["みかん", 800], ["ぶどう", 1500], ["青りんご", 300]])
    status, reason = ailine.check_extract_columns(p, dict(_XC))
    assert status == "pass", reason


def test_a_dropped_row_fails(tmp_path):
    """⑤ 列を選ぶ操作で行は減らない ── 減っていたら落とす。"""
    p = _out_book(tmp_path, ["商品", "売上"],
                   [["りんご", 1200], ["みかん", 800], ["ぶどう", 1500]])
    status, reason = ailine.check_extract_columns(p, dict(_XC))
    assert status == "fail" and "行数が合わない" in reason, reason


def test_a_wrong_value_fails(tmp_path):
    p = _out_book(tmp_path, ["商品", "売上"],
                   [["りんご", 9999], ["みかん", 800], ["ぶどう", 1500], ["青りんご", 300]])
    status, reason = ailine.check_extract_columns(p, dict(_XC))
    assert status == "fail" and "元と違います" in reason, reason


def test_wrong_headers_fail(tmp_path):
    p = _out_book(tmp_path, ["商品", "原価"],
                   [["りんご", 700], ["みかん", 300], ["ぶどう", 900], ["青りんご", 100]])
    status, reason = ailine.check_extract_columns(p, dict(_XC))
    assert status == "fail" and "見出しが宣言と違う" in reason, reason


def test_the_column_is_chosen_by_where_the_named_values_live(tmp_path):
    """★ 実測（同じ依頼文で聞かれる回と聞かれない回があった）: 一段目は 2/3 で EXTRACT、
       1/3 で OUT_OF_VOCAB → 「もしかして」の確認に回っていた。
       **聞かれるかどうかが偶然で決まる**のは、道具の性格として悪い。
       ★ 機械が列も値も解けているなら迷う理由が無い（実表を見た側が確かなことを知っている）。"""
    meta = _meta(_book(tmp_path))
    col, vals = ailine.resolve_named_extraction(meta, "売上", "みかんの行とりんごの行だけを抽出して")
    assert (col, vals) == ("商品", ["みかん", "りんご"]), (col, vals)


def test_it_refuses_to_choose_when_two_columns_tie(tmp_path):
    """★ 同数で並んだら**決めない**（推測で別の列を抜き出すのが一番こわい）。"""
    rows = [["商品", "担当"], ["りんご", "みかん"], ["みかん", "りんご"]]
    meta = _meta(_book(tmp_path, rows), headers=("商品", "担当"))
    col, vals = ailine.resolve_named_extraction(meta, "売上", "みかんとりんごを抜き出して")
    assert (col, vals) == (None, None), (col, vals)


def test_a_condition_request_is_never_stolen_by_the_named_extraction_reread():
    """★★ 許可の列挙へ裏返した時に開けかけた穴（今日 3 度目の同じ形）:
       「原価が500以上の行を抜き出して」で 500 が実表にも在ると、条件の抽出を
       **名指しの抽出**として横取りしかねない。比較語が在る依頼には触らない。"""
    assert ailine.extract_cmp_from_task("原価が500以上の行を抜き出して") == "gte"
    assert ailine.extract_cmp_from_task("みかんの行とりんごの行だけを抽出して") is None
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index("_re_extract_ask.search(a.task")
    assert "extract_cmp_from_task(a.task) is None" in src[i:i + 400], \
        "名指しの読み直しが、比較語のある依頼まで拾う形になっている"
