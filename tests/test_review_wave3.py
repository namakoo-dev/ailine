# 帳票レビュー第三波 ── 誤った ⚠ / 誤った × / 黙る、の 4 件（2026-08-24）。
# 出典: SEALED-20260824-report-review.md
#
# 契約:
#   ① PDF 照合は**書式つきの表示**も許す（1000 と 1,000 を別物として × にしない）
#   ② 「取引先ごとに」で同じ名前が複数あるなら、黙って別々の書類にせず名指しする
#      （機械は `_2` を付けた瞬間に重複を知っている）
#   ③ 依頼文に無い年を LLM が入れたら、機械抽出と突き合わせて断る
#   ④ 雛形の A1 が空でも「無変更」の比較セル数が 0 にならない

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import pdf_export  # noqa: E402


# --- ① 書式つきの表示を許す ---------------------------------------------------------

def test_pdf_check_accepts_formatted_numbers():
    """1000 が PDF 上で「1,000」や「¥1,000」でも見つかったと数える。"""
    r = pdf_export.verify_values_in_pdf.__doc__
    assert hasattr(pdf_export, "renderings"), "表示ゆれの候補を作る器が無い"
    cands = pdf_export.renderings(1000)
    assert "1000" in cands and "1,000" in cands, f"桁区切りを候補にしていない: {cands}"


def test_pdf_check_accepts_formatted_dates():
    import datetime as dt
    cands = pdf_export.renderings(dt.datetime(2026, 7, 31))
    assert any("2026/07/31" == c or "2026/7/31" == c for c in cands), \
        f"日付の表示ゆれを候補にしていない: {cands}"
    assert not any("00:00:00" in c for c in cands), \
        f"str(datetime) をそのまま候補にしている（PDF には出ない形）: {cands}"


def test_pdf_check_still_catches_a_truly_missing_value(tmp_path):
    """恒真殺し: 表示ゆれを許しても、本当に無い値は見つからないままであること。"""
    p = tmp_path / "x.pdf"
    p.write_bytes(b"%PDF fake")
    import types
    r = pdf_export.PdfCheck()
    assert r.missing == []
    # renderings が何でも通す作りになっていないこと
    assert "9999" not in pdf_export.renderings(1000)


# --- ② 「ごとに」で重複を黙らない ---------------------------------------------------

needs_dup = pytest.mark.xfail(
    not hasattr(ailine, "duplicate_name_warning"),
    reason="重複の名指し 未実装（契約は凍結済み）", strict=True)


@needs_dup
def test_duplicate_name_column_is_named():
    """★ 実測（盲検）: 3 社なのに請求書 4 枚。同一取引先が別々の書類に分かれ、
       それでも ✓ が出た。機械は `_2` を付けた瞬間に重複を知っている ── 黙らない。"""
    rows = ["あかつき商事", "みどり工業", "あかつき商事", "そら建設"]
    msg = ailine.duplicate_name_warning("取引先", rows)
    assert msg, "重複があるのに黙っている"
    assert "あかつき商事" in msg and "2" in msg, f"どの名前が何枚に分かれるかを言っていない: {msg}"


@needs_dup
def test_no_warning_when_names_are_unique():
    assert ailine.duplicate_name_warning("取引先", ["甲", "乙", "丙"]) is None


# --- ④ 雛形の A1 が空でも比較対象が 0 にならない（分母ゼロの禁止）--------------------
#
# ★ 実測（盲検の契約レビュー・俺の裏取りで確認）: A 列を余白にした典型的な請求書雛形
#   （B2 に「請 求 書」）だと `_scan_last_col` が 0 を返し、「雛形は無変更」を
#   **0 セル比較したまま**宣言していた。雛形が何セル壊れても pass する。

def test_used_range_is_used_for_template_comparison(tmp_path):
    p = tmp_path / "tpl.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "雛形"
    ws["B2"] = "請 求 書"        # ★ A 列も 1 行目も空（実物の請求書雛形の典型）
    ws["B5"] = "{{取引先}}"
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["雛形"]
    assert ailine._scan_last_col(ws2, header_row=1) == 0, "前提: 走査では 0 になること"
    rows, cols = ailine._used_extent(ws2)
    assert rows >= 5 and cols >= 2, f"使用範囲が取れていない: {rows}x{cols}"


def test_used_extent_is_zero_only_for_a_truly_empty_sheet(tmp_path):
    """誤爆防止: 本当に空のシートだけが 0。"""
    p = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook(); wb.active.title = "空"
    wb.save(p)
    ws = openpyxl.load_workbook(p)["空"]
    assert ailine._used_extent(ws) == (0, 0)


# --- ⑤ 依頼文に無い年を LLM が入れたら断る（2026-08-24）------------------------------
#
# ★ 実測（盲検の使い勝手レビュー）: 「3月26日から4月25日まで」（年を言っていない）に対し
#   LLM が **2023 年**を入れてきた。データは全部 2026 年。無言で年を捏造していた。
#   A' 原則 ── 依頼文に無い年は機械が受け取らない。年は人が決めることで、
#   LLM が埋めてよい空白ではない。

def test_year_not_in_the_task_is_refused(tmp_path):
    import datetime as dt
    p = tmp_path / "n.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "日報"
    ws.append(["日付", "売上"])
    ws.append([dt.datetime(2026, 3, 27), 100])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    ok, _res, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "gte", "value": "2023-03-26"}, meta,
        task="3月26日から4月25日までの行だけ抽出して")
    assert not ok, "依頼文に無い 2023 年を黙って使った"
    assert "2023" in (err or "") and "年" in (err or ""), f"何が問題かを言っていない: {err}"


def test_year_written_in_the_task_is_accepted(tmp_path):
    """誤爆防止: 依頼文に年が書いてあれば今までどおり通る。"""
    import datetime as dt
    p = tmp_path / "y.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "日報"
    ws.append(["日付", "売上"])
    ws.append([dt.datetime(2026, 3, 27), 100])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    ok, _res, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "gte", "value": "2026/3/26"}, meta,
        task="2026/3/26以降の行だけ抽出して")
    assert ok, err


# --- ⑥ 合計行の除外の内訳 ── **検体が実装より先に真実を出した**（2026-08-24）-----------
#
# 所見#20 は「run 経路が除外の内訳を開示していない（cmd_stack は開示＝片配線）」だった。
# 開示を書いて検体を回したら**鳴らなかった**ので追いかけたところ:
#   帳票段のデータ行収集は `_scan_last_row`（**1 列目**を上から見て最初の空で打ち切り）を
#   使う。合計行の判定に使うラベルも 1 列目。つまり **run 経路では「ラベルが空欄の行」に
#   そもそも到達しない**（その手前で走査が終わる）。書いた開示は死んだコードだった。
#
# ★ 実際にその事故から守っているのは detect_first_column_gap（第一波で入れた分母の警告）。
#   1 列目が途中で空くと「データ N 行と数えました（M 行目にはまだ中身があります）」と言い、
#   決裁③で ✓→△ に落ちる。**片配線の指摘は正しかったが、run 側の穴は別の器が塞いでいた。**
#   死んだコードを残すのは、番人が在るように見えて鳴らないのと同じなので消した。


def test_a_row_whose_label_is_blank_is_now_reached(tmp_path):
    """★ 2026-09-05（段B）の根治 ── 1 列目が空の行に走査が届くこと。

    ★ 旧版はこう凍結していた:
        assert scanned == 2, "前提: 1 列目の空欄で走査が止まる"
      その「前提」こそが盲検（2026-08-24）で見つかった事故の原因で、当時は
      **数え方を変えると全機構の前提が動く**ため「縮んだ事実を言う」に留めていた。
    ★ 段B で数え方そのものを直したので、いまは届く ── **器官が黙るのが正しい**。
      帳票（取引先ごとに 1 枚）で、名前が空の行が落ちなくなる。
    """
    p = tmp_path / "gap.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"])
    for name, amt in [("甲", 100), (None, 999), ("乙", 200)]:
        ws.append([name, amt])
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["売上"]
    assert ailine._scan_last_row(ws2, header_row=1) == 4, "1 列目が空の行に届いていない"
    assert ailine.detect_first_column_gap(ws2, header_row=1) is None,         "届いているのに『縮んだ』と言っている（開示が古い前提のまま）"


def test_the_gap_organ_still_speaks_when_the_scan_really_stops(tmp_path):
    """★ 器官を消していないこと ── 走査が**本当に**止まる形では、いまも名指しする。

    空行で切れた先にデータが在る形（表の終わりと見分けられない）。
    """
    p = tmp_path / "gap2.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"])
    ws.append(["甲", 100])
    ws.append([None, None])
    ws.append(["乙", 200])
    wb.save(p)
    ws2 = openpyxl.load_workbook(p)["売上"]
    assert ailine._scan_last_row(ws2, header_row=1) == 2, "空行で切れていない"
    gap = ailine.detect_first_column_gap(ws2, header_row=1)
    assert gap and "4" in gap, f"分母が縮んだ事実を言っていない: {gap}"
