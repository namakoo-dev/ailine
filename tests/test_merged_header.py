# 結合セルの見出しで列が丸ごと消えるのを止める（2026-08-24）。
#
# ★ 実測（盲検の実データ耐性レビュー・俺が実物で再現）:
#   `A1:B1` を結合した請求書（日本の請求書でほぼ普遍の形）を stack すると
#     出力: ('品名','規格','元ファイル','元行')     ← **数量と金額が列ごと消える**
#     exit 0 ・ ⚠ なし ・ **Σ金額 の行も出ない**
#   金額が消えたのに、消えたこと自体が画面のどこにも現れない。
#   原因: `read_row_headers` が「先頭列から連続する非空セル」で切るため、
#   結合セルの 2 つ目以降（None）で止まる。docstring に
#   「第一波は単純に ── 列の間に空白を挟む見出しは扱わない」と当時の割り切りが在った。
#
# 契約:
#   ① 見出し行に空セルが挟まっていても、**その先の見出しを落とさない**
#   ② 結合セルは左上の値を引き継ぐ（Excel の見た目と一致させる）
#   ③ 誤爆しない: 見出しの後ろの余白（末尾の空セル）は今までどおり打ち切る

import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core.multifile import read_row_headers  # noqa: E402


def _ws(tmp_path, cells, merges=()):
    wb = openpyxl.Workbook(); ws = wb.active
    for ref, v in cells.items():
        ws[ref] = v
    for m in merges:
        ws.merge_cells(m)
    p = tmp_path / "t.xlsx"; wb.save(p)
    return openpyxl.load_workbook(p).active


def test_merged_header_does_not_truncate_the_rest(tmp_path):
    """① ② A1:B1 結合 → 数量・金額を落とさない。"""
    ws = _ws(tmp_path, {"A1": "商品情報", "C1": "数量", "D1": "金額"}, ["A1:B1"])
    got = read_row_headers(ws, 1)
    assert "数量" in got and "金額" in got, f"結合の先が落ちた: {got}"
    assert len(got) == 4, f"列数が合わない（結合は左上を引き継ぐ）: {got}"


def test_a_single_gap_does_not_truncate(tmp_path):
    """① 結合でない単なる空セルでも、その先を落とさない。"""
    ws = _ws(tmp_path, {"A1": "品名", "C1": "数量", "D1": "金額"})
    got = read_row_headers(ws, 1)
    assert "数量" in got and "金額" in got, f"空セルの先が落ちた: {got}"


def test_trailing_blanks_still_end_the_headers(tmp_path):
    """③ 誤爆防止: 見出しの後ろの余白は今までどおり打ち切る（無限に伸ばさない）。"""
    ws = _ws(tmp_path, {"A1": "品名", "B1": "数量", "C1": "金額"})
    got = read_row_headers(ws, 1)
    assert got == ["品名", "数量", "金額"], got


def test_no_headers_at_all_is_still_empty(tmp_path):
    ws = _ws(tmp_path, {"A5": "何か"})
    assert read_row_headers(ws, 1) == []


# --- ★ 見出し行の推定と、列が消える事故（2026-08-24）--------------------------------
#
# 実測: `商品情報`(A1:B1 結合) / `数量` / `金額` の請求書で、`detect_header_row` は
# **2 行目を confident で返す**（1 行目は結合のせいで非空セルが 1 つに見える）。
# 2 行目は `品名`/`規格` の 2 列しかないので、**数量と金額が列ごと落ちる**。
# しかも exit 0・⚠ なし・Σ金額 の行も出ない ── 消えたこと自体が画面に現れない。
#
# ★ 判定そのものを変えない理由: 2 行目を選ぶのが正しい表も実在する（多段見出し）。
#   ここで直すのは「選ばなかった行に**まだ見出しが在る**なら、それを言う」ことだけ。
#
# 契約: 採用した見出し行より上の行に、採用行より**多くの見出し**が在るなら名指しする

def test_columns_lost_by_header_row_choice_are_named(tmp_path):
    ws = _ws(tmp_path, {"A1": "商品情報", "C1": "数量", "D1": "金額",
                         "A2": "品名", "B2": "規格"}, ["A1:B1"])
    import ailine
    msg = ailine.header_row_drops_columns(ws, chosen_row=2)
    assert msg, "上の行に在る列を落としたのに黙っている"
    assert "数量" in msg and "金額" in msg, f"落ちた列を名指ししていない: {msg}"


def test_no_message_when_the_chosen_row_is_the_widest(tmp_path):
    """誤爆防止: 採用行が一番広ければ何も言わない。"""
    ws = _ws(tmp_path, {"A1": "表題", "A2": "品名", "B2": "数量", "C2": "金額"})
    import ailine
    assert ailine.header_row_drops_columns(ws, chosen_row=2) is None


def test_no_message_for_a_plain_single_header_row(tmp_path):
    ws = _ws(tmp_path, {"A1": "品名", "B1": "数量", "C1": "金額"})
    import ailine
    assert ailine.header_row_drops_columns(ws, chosen_row=1) is None
