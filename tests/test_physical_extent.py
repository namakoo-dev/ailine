# 塊②（2026-08-25）── 分母を「走査」でなく「物理の使用範囲」から作る。
#
# ★★ 2026-09-05（段B）で**原因の側を直した**ので、検体を作り直した。
#   それまでの検体は「末尾に **1 列目だけ空**の行がある表」だった ── 走査が A 列を
#   上から見て最初の空で止まる規則だったので、そこで食い違いが起きた。
#   段B で走査を「**表の幅のどこかに値が在れば行**」に変えたため、この形はもう
#   食い違わない（scanned 4/4）。★ **それは器官が壊れたのではなく、症状が消えた**。
#   だから検体を「今の走査でも届かない形」へ替え、器官が働くことを証明し続ける:
#     ・空行が 3 つ以上あって、その先にデータが在る（表の終わりと見分けられない）
#     ・見出しの右に離れて在る列
#   ★ そして旧検体は「**今は届く**」ことを測る側に回した（根治の証拠として残す）。
# 致命1: 見出しの無い列だけ置き去りになり、全行が入れ替わる
#   D1 が空で D2..D5 にデータがある表を並べ替えると、範囲が見出し由来（A〜C）なので
#   D 列だけ動かず、**全行で備考が別の商品の物に付け替わる**。
#   `check_sort` はキー列の単調性しか見ず、**行の同一性を一度も確かめない**ので ✓ が出る。
#   しかも画面の「変更点」も A〜C しか出さないため、**差分自体が壊れた列を隠す**。
#
# 致命5: 先頭列が空の末尾行が、処理からも分母からも消える
#   `_scan_last_row` は A 列を上から見て最初の空で止まる。末尾に A 列が空の行があると、
#   その行は処理されず、**分母にも数えられない**（「3行中1行が一致」＝真の分母は 5）。
#   ★ 表の**途中**の空きは ⚠ が出て △ に落ちるのに、**末尾だけ鳴らない**
#     （警告条件が「A 列に下方向の中身があるか」なので原理的に発火しない）。
#
# 契約:
#   ① 物理の使用範囲と、走査で得た範囲の食い違いを機械で出せる
#   ② 並べ替えは**行の同一性**を物理の列範囲で確かめる（他列が一緒に動いたか）
#   ③ 行がちぎれていたら ✓ でなく **×**（出力は既に壊れている）
#   ④ 末尾の欠けは分母の食い違いとして名指しされる
#   ⑤ 食い違いが無ければ 1 文字も増えない（誤爆しない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


# --- ① 食い違いを出す器官 -------------------------------------------------------------

def test_extent_gap_sees_columns_beyond_the_headers(tmp_path):
    """D1 が空・D2 以下にデータ ── 見出し走査は 3 列、物理は 4 列。"""
    p = _book(tmp_path, [["商品", "単価", "数量", None],
                          ["a", 120, 3, "特売"], ["b", 200, 5, "通常"]])
    wb = openpyxl.load_workbook(p)
    gap = ailine.extent_gap(wb["売上"], header_row=1)
    assert gap["cols_scanned"] == 3, gap
    assert gap["cols_physical"] == 4, gap
    assert gap["cols_missing"] == 1, gap


def test_extent_gap_sees_rows_beyond_a_run_of_blank_rows(tmp_path):
    """★ 空行が続いた先にデータ ── 走査は「表が終わった」と読むが、物理には続きが在る。

    ★ 2026-09-05（段B）で検体を替えた。旧版は「末尾に **1 列目だけ空**の行」で、
      当時の走査（A 列を上から見て最初の空で止まる）ならそこで食い違った。
      いまの走査は表の幅のどこかに値が在れば行と数えるので、その形はもう届く
      （下の test_a_trailing_row_with_an_empty_first_cell_is_now_reached が証拠）。
    ★ 器官が要らなくなったのではない ── **表の終わりと見分けられない形**では
      いまも食い違う。証明はそちらへ移す。
    """
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200],
                          [None, None], [None, None], [None, None], ["c", 300]])
    wb = openpyxl.load_workbook(p)
    gap = ailine.extent_gap(wb["売上"], header_row=1)
    assert gap["rows_scanned"] == 2, gap
    # ★ 2026-09-05（A′）: 分母は「表と同じ形の行」なので、
    #   物理の使用範囲そのものではない。数でなく**見落としが在ること**を測る。
    assert gap["rows_scanned"] == 2, gap
    assert gap["rows_missing"] > 0, gap


def test_extent_gap_is_silent_on_a_tidy_table(tmp_path):
    """⑤ 誤爆しない。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    gap = ailine.extent_gap(wb["売上"], header_row=1)
    assert gap["rows_missing"] == 0 and gap["cols_missing"] == 0, gap


# --- ②③ 行がちぎれたら × ------------------------------------------------------------

def test_sort_fails_when_rows_were_torn(tmp_path):
    """★ 実測の再現: 見出しの無い列が置き去りになり、備考が別商品の物になった状態。
       今までは ✓ が出ていた。"""
    p = _book(tmp_path, [["商品", "単価", "数量", None],
                          ["ぶどう", 500, 2, "特売"],    # 本来は '高級'
                          ["かき", 200, 5, "通常"],      # 本来は '訳あり'
                          ["りんご", 120, 3, "高級"],    # 本来は '特売'
                          ["みかん", 80, 10, "訳あり"]]) # 本来は '通常'
    before = _book(tmp_path, [["商品", "単価", "数量", None],
                               ["りんご", 120, 3, "特売"],
                               ["みかん", 80, 10, "通常"],
                               ["ぶどう", 500, 2, "高級"],
                               ["かき", 200, 5, "訳あり"]], name="before.xlsx")
    status, reason = ailine.check_sort(p, {"col": "単価", "order": "desc"},
                                        source_book=before)
    assert status == "fail", f"行がちぎれているのに通した: {status} / {reason}"
    assert "行" in reason


def test_sort_passes_when_all_columns_moved_together(tmp_path):
    """⑤ 誤爆防止: 見出しの無い列も一緒に動いていれば通る。"""
    before = _book(tmp_path, [["商品", "単価", None],
                               ["りんご", 120, "特売"], ["ぶどう", 500, "高級"]],
                    name="before.xlsx")
    after = _book(tmp_path, [["商品", "単価", None],
                              ["ぶどう", 500, "高級"], ["りんご", 120, "特売"]])
    status, reason = ailine.check_sort(after, {"col": "単価", "order": "desc"},
                                        source_book=before)
    assert status == "pass", f"正しい並べ替えを落とした: {reason}"


# --- ④ 末尾の欠けを名指しする ---------------------------------------------------------

def test_trailing_rows_are_named_not_swallowed(tmp_path):
    """★ 表の途中の空きは鳴るのに、末尾だけ鳴らなかった。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200],
                          [None, None], [None, None], [None, None], ["c", 300]])
    args = {"col": "単価", "order": "desc"}
    # ★ 2026-08-25: 本番の合流点で測る。以前は check_sort を直に呼んでいたが、
    #   食い違いを数えるのは run_postcondition の入口へ畳んだ（全 op に効かせるため）。
    ailine.run_postcondition("SORT", p, args)
    assert args.get("_unverified"), "空行の先の行が黙って消えた"
    assert any(u["rows"] > 0 for u in args["_unverified"]), \
        args["_unverified"]


# --- ⑥ 見ていない行が在ることは、表の性質であって op の性質でない ------------------------

@pytest.mark.parametrize("op", sorted(ailine.POSTCONDITIONS) + ["CHART"])
def test_every_op_notices_the_trailing_gap(tmp_path, op):
    """★ 盲検 2 回目の致命①②: extent_gap は在ったのに **20 本中 1 本**にしか
       挿さっていなかった。同じ検体で並べ替えだけが △ に落ち、集計・抽出・重複除去・
       一括書換・太字・桁区切り・グラフは ✓ を名乗った。

       ★ 1 本の試験で**全経路**を縛る（各 checker に書き写させない）。この試験は
       note_extent_gap の配線を外すと全件が同時に赤くなる（変異試験で確認済み）。
       事後条件の**判定**は問わない ── 問うのは「見ていない行を数えたか」だけ。
    """
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200],
                          [None, None], [None, None], [None, None], ["c", 300]])
    args = {"col": "単価", "order": "desc", "value_col": "単価"}
    ailine.run_postcondition(op, p, args, source_book=p)
    assert args.get("_unverified"), f"{op}: 空行の先の行を見ていないのに黙っている"
    assert any(u["rows"] > 0 for u in args["_unverified"]), args["_unverified"]


@pytest.mark.parametrize("op", sorted(ailine.POSTCONDITIONS) + ["CHART"])
def test_no_op_cries_wolf_on_a_tidy_table(tmp_path, op):
    """⑤ 誤爆しない: 整った表では 1 op も ⚠ を足さない。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200]])
    args = {"col": "単価", "order": "desc", "value_col": "単価"}
    ailine.run_postcondition(op, p, args, source_book=p)
    assert not args.get("_unverified"), f"{op}: 食い違いが無いのに鳴った: {args['_unverified']}"

def test_a_trailing_row_with_an_empty_first_cell_is_now_reached(tmp_path):
    """★ 2026-09-05（段B）の根治の証拠 ── 旧検体が**もう食い違わない**こと。

    ★ 2026-08-25 の盲検（致命5）はこの形で見つかった:
      「_scan_last_row は A 列を上から見て最初の空で止まる。末尾に A 列が空の行が
        あると、その行は処理されず、**分母にも数えられない**」
    ★ 段B で走査を「表の幅のどこかに値が在れば行」に変えたので、いま A 列が空でも
      届く。**器官が黙るのが正しい** ── 症状が消えたのであって、器官は上の検体で
      まだ働いている。
    """
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200],
                          [None, 1500], [None, 2000]])
    wb = openpyxl.load_workbook(p)
    gap = ailine.extent_gap(wb["売上"], header_row=1)
    assert gap["rows_scanned"] == 4, gap
    assert gap["rows_missing"] == 0, "1 列目が空の末尾行に、走査が届いていない"


def test_the_organ_still_sees_a_column_that_sits_apart(tmp_path):
    """★ 見出しの右に離れて在る列は、いまも食い違いとして名指しされること。"""
    p = _book(tmp_path, [["商品", "単価"], ["a", 120], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    ws = wb["売上"]
    ws.cell(1, 8, "離れた見出し")
    ws.cell(2, 8, "値")
    wb.save(p)
    gap = ailine.extent_gap(openpyxl.load_workbook(p)["売上"], header_row=1)
    assert gap["cols_missing"] > 0, gap

