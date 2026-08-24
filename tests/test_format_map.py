# 様式写像段 FORMAT_MAP + CSV_EXPORT ── 実装より先に凍結した赤い検体。
# 設計: ~/.nagi/plans/DESIGN-20260824-format-map.md
# 実需: MARKET-20260823-lancers.md（FORMAT_MAP 5 件 / CSV_EXPORT 4 件・★ 台帳が
#       「この 2 つは対で、片方だけでは受注 CSV → 出荷 CSV が完結しない」と明言）。
#
# ★ 憲法: **様式は人が作る。機械は埋めるだけ。** 列の対応を LLM に推測させない。
# ★ 形: REPORT_PER_ROW の兄弟 ── 雛形の 1 **行**を、データ行の数だけ**縦**に展開する。
#
# 契約:
#   ① 行の完全会計: データ行 N = 出力 N 行 + 除外（合計行は total_row の既存機構）
#   ② 値の 3 計数: 出力の各セル vs 元の行の対応セル（型込み等値）
#   ③ 出所: 検分シートに 出力行 / 元行
#   ④ 雛形と元データの不変
#   ⑤ CSV_EXPORT: 書いた CSV を**読み戻して**元シートと照合。0 落ちを**作らない**
#      （文字列保持のセルは引用して書く・数値化しない）。文字コードと引用の規則を**開示**する

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "check_format_map"),
    reason="様式写像段 未実装（契約は凍結済み）",
    strict=True,
)


def _ledger_book(tmp_path, name="出納帳.xlsx"):
    """元データ（出納帳・合計行つき）と、人が作った出力様式の雛形。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出納帳"
    for r in [["日付", "摘要", "入金", "出金"],
               ["2026-07-01", "売上入金", 50000, 0],
               ["2026-07-03", "仕入", 0, 12000],
               ["合計", "", 50000, 12000]]:
        ws.append(r)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "取引日"; tpl["B1"] = "内容"; tpl["C1"] = "金額"      # 人が並べた出力の見出し
    tpl["A2"] = "{{日付}}"; tpl["B2"] = "{{摘要}}"; tpl["C2"] = "{{入金}}"
    wb.save(p)
    return p


# --- FORMAT_MAP: 縦の展開 ---------------------------------------------------------

@needs_impl
def test_format_map_end_to_end(tmp_path, monkeypatch, capsys):
    """データ 2 行（+合計行）→ 出力 2 行。見出しは人の様式のまま・型は保たれる。"""
    _isolate(monkeypatch, tmp_path)
    book = _ledger_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "FORMAT_MAP", "args": {"template_sheet": "様式"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        src = wb["出納帳"]
        out = wb.create_sheet("様式_出力")
        out.append(["取引日", "内容", "金額"])
        for r in range(2, 4):                      # 合計行(4 行目)は含めない
            out.append([src.cell(r, 1).value, src.cell(r, 2).value, src.cell(r, 3).value])
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "出納帳を様式シートの形に写して", "--copy"], capsys)
    assert rc == 0, out
    got = openpyxl.load_workbook(str(book).replace(".xlsx", ".out.xlsx"))
    made = [s for s in got.sheetnames if s not in ("出納帳", "様式") and not s.startswith("検分")]
    assert len(made) == 1, f"出力シートが 1 枚でない: {got.sheetnames}"
    sh = got[made[0]]
    rows = [[c.value for c in r] for r in sh.iter_rows()]
    assert rows[0] == ["取引日", "内容", "金額"], f"人の見出しが保たれていない: {rows[0]}"
    assert len(rows) == 3, f"① 行の会計が合わない（合計行を写した疑い）: {rows}"
    assert rows[1][2] == 50000 and isinstance(rows[1][2], int), \
        f"金額が数値でない（下流が壊れる）: {rows[1][2]!r}"
    assert any(s.startswith("検分") for s in got.sheetnames), "③ 出所の検分シートが無い"
    assert got["様式"]["A2"].value == "{{日付}}", "④ 雛形が書き換わった"
    assert "✓" in out, out


@needs_impl
def test_format_map_unknown_placeholder_refuses(tmp_path, monkeypatch, capsys):
    """様式の印が実在しない列を指していたら、写す前に断る（幻覚の封鎖）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "元"
    ws.append(["日付", "金額"]); ws.append(["2026-07-01", 100])
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "取引日"; tpl["A2"] = "{{存在しない列}}"
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "FORMAT_MAP", "args": {"template_sheet": "様式"}})
    rc, out = _run_main(["run", str(p), "様式の形に写して", "--copy"], capsys)
    assert rc != 0
    assert "存在しない列" in out, f"どの印が悪いか名指ししていない: {out}"


# --- CSV_EXPORT: 検疫の逆方向 -----------------------------------------------------

@needs_impl
def test_csv_export_preserves_leading_zero_and_discloses_encoding(tmp_path, monkeypatch, capsys):
    """★ 0 落ちを**作らない**: 文字列で保持されている品番は引用して書く。
       文字コードと引用の規則を必ず開示する。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "出荷"
    ws.append(["品番", "数量"])
    ws.append(["0123", 5])          # 文字列のまま保持されている品番
    ws.append(["00456", 12])
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "出荷"], capsys)
    assert rc == 0, out
    made = p.with_suffix(".csv")
    assert made.exists(), f"CSV が作られていない: {out}"
    raw = made.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    assert "0123" in text and "00456" in text, f"0 落ちを作った: {text!r}"
    assert "文字コード" in out and "引用" in out, f"符号化と引用の規則を開示していない: {out}"
    assert "✓" in out and "欠落0" in out.replace(" ", ""), f"転送段の主張が無い: {out}"


@needs_impl
def test_csv_export_roundtrip_mismatch_is_not_checkmarked(tmp_path, monkeypatch, capsys):
    """恒真殺し: 読み戻して食い違ったら ✓ を名乗らない（CSV 検疫と同じ規律）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "drift.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "s"
    ws.append(["v"]); ws.append([1.1234567890123456789])   # 往復で落ちる桁
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "s"], capsys)
    if "不一致0" not in out.replace(" ", ""):
        assert "✓" not in out, f"食い違っているのに ✓: {out}"


@needs_impl
def test_csv_export_encoding_option_is_honored(tmp_path, monkeypatch, capsys):
    """--encoding cp932 を指定したらその符号化で書き、そう述べる。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "j.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "s"
    ws.append(["名前"]); ws.append(["田中"])
    wb.save(p)
    rc, out = _run_main(["export-csv", str(p), "--sheet", "s", "--encoding", "cp932"], capsys)
    assert rc == 0, out
    raw = p.with_suffix(".csv").read_bytes()
    assert "田中".encode("cp932") in raw, "cp932 で書かれていない"
    assert "cp932" in out, f"選んだ符号化を述べていない: {out}"
