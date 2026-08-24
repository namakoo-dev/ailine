# 帳票段 REPORT_PER_ROW ── 実装者の自分の境界検体（凍結済み tests/test_report_per_row.py には
# 無いが、DESIGN-20260823-report-per-row.md が明示した境界。凍結検体は1文字も変更しない・
# ここは新規追加のみ）。
#
# 対象:
#   ① 部分一致の印（数値列）は原理的に文字列にしかなれないので断る
#      （「使われたら断る・検体には無いが自分の検体で固定すること」の指示どおり）。
#   ② 31文字境界の実配線: sanitize_sheet_name/unique_sheet_name の単体テストは凍結検体に
#      あるが、verify_dsl_args がそれを実際に呼んで衝突を解決するところまでは検体が無い
#      （同じ31文字接頭辞に切り詰まる2つの長い名前を実データで通す）。
#   ③ 型の出し分け: compare_report_cells は「値が違う」だけでなく「型が違う」も
#      恒真殺しとして拾うこと（数値 100 と文字列 "100" は別物）。
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


# --- ① 部分一致の印 × 数値列は断る ---------------------------------------------

def test_partial_placeholder_on_numeric_column_refuses(tmp_path, monkeypatch, capsys):
    """雛形『請求書番号: {{金額}}』のように、印が他の文字と同居する部分一致で、かつ
       指す列が数値列なら、埋める前に断る（設計文書 訂正3: 部分一致の印は原理的に
       文字列にしかなれない ── 数値列には使わせない）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "invoice.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "金額"])
    ws.append(["あかつき商事", 12000])
    tpl = wb.create_sheet("雛形")
    tpl["B3"] = "{{取引先}}"
    tpl["B5"] = "内訳: {{金額}}円"   # ★ 部分一致・指す列(金額)は数値
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "REPORT_PER_ROW", "args": {"template_sheet": "雛形", "name_col": "取引先"}})
    rc, out = _run_main(["run", str(p), "取引先ごとに請求書を作って", "--copy"], capsys)
    assert rc != 0, out
    assert "金額" in out and "数値" in out, f"数値列への部分一致だと名指ししていない: {out}"


def test_whole_cell_placeholder_on_numeric_column_is_fine(tmp_path, monkeypatch, capsys):
    """対照: セル全体が印（部分一致ではない）なら、数値列でも断らない
       （柱④の成立条件は『部分一致 かつ 数値列』の組だけを禁じる・広げすぎない）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "invoice.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "金額"])
    ws.append(["あかつき商事", 12000])
    tpl = wb.create_sheet("雛形")
    tpl["B3"] = "{{取引先}}"
    tpl["B5"] = "{{金額}}"   # ★ 丸ごと一致・数値列でも可
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "REPORT_PER_ROW", "args": {"template_sheet": "雛形", "name_col": "取引先"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        sh = wb2.copy_worksheet(wb2["雛形"])
        sh.title = "あかつき商事"
        sh["B3"] = "あかつき商事"
        sh["B5"] = 12000
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "取引先ごとに請求書を作って", "--copy"], capsys)
    assert rc == 0, out


# --- ② 31文字境界の実配線（verify_dsl_args → unique_sheet_name の衝突解決） -----

def test_verify_dsl_args_resolves_31char_collision_from_real_data():
    """同じ31文字接頭辞に切り詰まる2つの長い取引先名を実データで通し、
       verify_dsl_args が呼ぶ前に一意な31文字以内の名前へ解決していることを確かめる
       （copyByName に失敗しうる名前を渡さない、という B の実配線を検体で固定する）。"""
    import tempfile
    tmp_dir = Path(tempfile.mkdtemp(prefix="ailine_report_31char_"))
    p = tmp_dir / "invoice.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "金額"])
    name1 = "あ" * 35             # 先頭31文字は「あ」*31
    name2 = "あ" * 34 + "B"       # 先頭31文字は同じく「あ」*31（32文字目以降でしか違わない）
    ws.append([name1, 100])
    ws.append([name2, 200])
    tpl = wb.create_sheet("雛形")
    tpl["B3"] = "{{取引先}}"
    tpl["B5"] = "{{金額}}"
    wb.save(p)

    book_meta = {"sheets": ["売上", "雛形"],
                 "headers": {"売上": ["取引先", "金額"], "雛形": []},
                 "header_rows": {"売上": 1, "雛形": 1},
                 "path": p}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "REPORT_PER_ROW", {"template_sheet": "雛形", "name_col": "取引先"}, book_meta)
    assert ok, err
    names = [rr["sheet"] for rr in resolved["_report_rows"]]
    assert len(names) == 2
    assert len(set(names)) == 2, f"31文字で切り詰まった2つの長い名前が衝突したまま: {names}"
    assert all(len(n) <= 31 for n in names), f"31文字を超えた: {names}"


# --- ③ 型の出し分け: compare_report_cells は型違いも拾う ------------------------

def test_compare_report_cells_treats_number_and_string_as_different_types(tmp_path):
    """『値は同じに見えるが型が違う』（数値 100 と文字列 "100"）を同一視しない
       （設計文書 訂正3: 型を誤ると number_format ごと壊れる実測に対応する側の検算）。"""
    p = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = "s"
    sh["B5"] = "100"   # ★ 文字列として書いた（setString 相当の事故を模す）
    wb.save(p)
    r = ailine.compare_report_cells(p, sheet_name="s", declared={"B5": 100})   # ★ 宣言は数値
    assert r.mismatched, f"数値と文字列(見た目は同じ)を型違いとして拾えていない: {r}"
