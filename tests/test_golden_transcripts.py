"""C1-F9（本命）: `ailine.main(argv)` を実際に叩き、標準出力の全文を承認テストで固定する。

★ 現在 ailine.main( を通るテストが test_ailine.py に 0 本だった（CLI の契約=main(argv)
そのものが一度も検証されていなかった）。ここでは translate_task / basrun_apply を
「録画済み」相当の固定戻り値・固定副作用の関数に差し替え、それ以外は本物のパイプライン
（verify_dsl_args → codegen_dsl → build_advisories → run_postcondition → ...）を通す。

シナリオ20+本（brief の目安どおり）:
  dsl: pass / warn / fail / 実行時エラー
  plan: 全ok / 混在warn / 途中fail / CLARIFY含み
  freeform: 関所y / 関所N / 非対話 / 総なめ検出
  破壊の関所: y / N / 非対話
  忠実度ゲート(exit 4) / Excelロック(exit 5) / runロック(exit 6)
  header-row 指定
  --dry × 3経路(dsl/plan/freeform)
  ★ 単位E: 対象スロットの出所（③矛盾で ✓ を出さない / ②無言で範囲を狭める1文）
  ★ 単位B: 部分文字列の片方向（「税込金額で」→『金額』）でも ✓ を出さない
= 4+4+4+3+1+1+1+1+3+2+1 = 25本。

ゴールデンは tests/golden/f9_transcripts/<name>.txt（標準出力そのもの）。
更新の作法は tests/golden/_harness.py 参照。
"""
import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

from golden._harness import GOLDEN_ROOT, assert_golden_text  # noqa: E402

F9_DIR = GOLDEN_ROOT / "f9_transcripts"


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _isolate(monkeypatch, tmp_path):
    """本番のユーザーディレクトリ(~/.ailine 等)に一切触れないようにする（実行するたび
    実ファイルへ書き込む既存の危険な既定を、テストでは全部 tmp_path に寄せる）。"""
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", tmp_path / "run.lock")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)


def _run_main(argv, capsys) -> tuple:
    rc = ailine.main(argv)
    out = capsys.readouterr().out
    return rc, out


# ===========================================================================
# dsl: pass / warn / fail / 実行時エラー
# ===========================================================================

def _t_dsl_pass(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 200], ["b", 300], ["c", 100]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for i, (name, val) in enumerate([("b", 300), ("a", 200), ("c", 100)], start=2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=val)
        wb.save(out_book)
        return True, None, "ok"

    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)


def _t_dsl_warn(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])   # データ1行のみ→順序の意味なし
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)


def _t_dsl_fail(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 300], ["c", 200]])   # 未整列のまま
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    # ★ basrun_apply は「成功はしたが実際には並べ替わっていない」状態を模する
    #   （LibreOffice+LLM がもっともらしく成功報告しつつ何もしないケースの再現）。
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)


def _t_dsl_runtime_error(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 300]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(
        ailine, "basrun_apply",
        lambda out_book, code, workdir, helper_files=(), timeout=None:
        (False, "BasicError: Object variable not set (line 12)", "raw-uno-output"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)


# ===========================================================================
# plan: 全ok / 混在warn / 途中fail / CLARIFY含み
# ===========================================================================

def _plan_book_presorted_bold(tmp_path):
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    wb = openpyxl.load_workbook(book)
    ws = wb.active
    for c in (1, 2):
        ws.cell(row=1, column=c).font = Font(bold=True)
    wb.save(book)
    return book


def _t_plan_all_ok(tmp_path, monkeypatch, capsys):
    book = _plan_book_presorted_bold(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                   {"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて見出しを太字に", "--copy"], capsys)


def _t_plan_mixed_warn(tmp_path, monkeypatch, capsys):
    book = _plan_book_presorted_bold(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                   {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        # ★ SORT 段は既に整列済みなのでファイルを変えなくても postcondition は通る。
        #   語彙外(自由生成)段は「変化した」ことそのものが唯一の機械確認(no-op ガード)
        #   なので、こちらだけ必ず何かしら変更する（code の中身で段を見分ける — 同じ
        #   basrun_apply 差し替えを両方の段が共有するため、SORT 側まで毎回書き換えると
        #   語彙外段の前後比較が『既に書き換え済み』で差分無し=no-op に化ける）。
        if "SortByColumn" not in code:
            wb = openpyxl.load_workbook(out_book)
            wb.active.cell(row=1, column=5, value="cf-applied")
            wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    def fake_ollama(model, msgs, temperature=0.2):
        return "Sub Run(oDoc As Object)\nEnd Sub"
    monkeypatch.setattr(ailine, "ollama_generate", fake_ollama)
    return _run_main(["run", str(book), "金額で降順に並べ替えて条件付き書式もつけて", "--copy",
                       "--allow-freeform"], capsys)


def _t_plan_partial_fail(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])   # 見出しは太字でない
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                   {"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))   # BOLD 側は何もしない→太字にならず fail
    return _run_main(["run", str(book), "金額で降順に並べ替えて見出しを太字に", "--copy"], capsys)


def _t_plan_with_clarify(tmp_path, monkeypatch, capsys):
    book = _plan_book_presorted_bold(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                   {"op": "CLARIFY", "question": "何色にしますか？"}]})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて色も変えて", "--copy"], capsys)


# ===========================================================================
# ★ freeform 最終決定（2026-08-21）: 単発の語彙外は生成に入らず即座に断る（cmd_refuse_
#   vocab_miss）。旧タイトル「関所y / 関所N / 非対話 / 総なめ検出」は cmd_run_freeform の
#   生成→適用ループを前提にしており、その前提ごと無くなった（y/N の分岐・sweep 検出は
#   複合計画側 run_freeform_plan_step にだけ残る ── freeform_plan_step 系の golden は
#   このブリーフの対象外・触っていない）。
#   4本の名前はそのまま維持し、中身は「新契約でも同じ断りになる」ことを示す新しい顔に
#   差し替えた（意図的な再生成 ── ブリーフが明示的に許可）。
# ===========================================================================

def _freeform_setup(monkeypatch, op="FREEFORM", about=""):
    """単発の語彙外を作るだけ。★ 生成は一切呼ばれない契約なので ollama_generate の
       固定戻り値はもう要らない（呼んだら test_freeform_out_only.py 側の禁止アサーションが
       落ちる ── ここでは単に mock しない = 呼ばれたら AttributeError で気づける）。"""
    def _translate(model, task, book_meta, temperature=0.1):
        return {"op": op, "about": about} if op == "OUT_OF_VOCAB" else {"op": op, "args": {}}
    monkeypatch.setattr(ailine, "translate_task", _translate)
    # ★ W10 便C2 検分（2026-08-22 夜）: 判定器は決定論の空に固定する ── mock しないと
    #   本物の ollama の応答が golden に焼き込まれ（非決定・CI には ollama が居ない）、
    #   ローカルでだけ緑に見える。この golden 群の目的は「断りの顔」の凍結であり、
    #   提案の顔は test_suggest_flow.py が担う。
    monkeypatch.setattr(ailine, "judge_ops_via_llm", lambda task, about=None: [])
    # ★ 2026-08-27: 第二段翻訳(translate_task_fixed_op)も止める。
    #   ★ 実測した抜け: 列追加(ADD_COLUMN)を足した日、この治具は translate_task しか
    #     mock しておらず、依頼文が「列を足して」だったため**本物の ollama を叩いていた**
    #     （CI には ollama が居ないので、手元だけ通る形になりかける）。
    #   治具は「外に触らない」ことまで固定する ── 窒息点を 1 つ残すと、そこから漏れる。
    monkeypatch.setattr(ailine, "translate_task_fixed_op",
                         lambda model, op, task, book_meta, temperature=0.1: None)


def _t_freeform_gate_yes(tmp_path, monkeypatch, capsys):
    """★ 旧「関所 y（承知して適用）」に一番近い操作は --allow-freeform（受理はするが
       断りの中身は変えない・廃止告知が足される）。"""
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    _freeform_setup(monkeypatch)
    return _run_main(["run", str(book), "セルに画像を貼って", "--copy", "--allow-freeform"], capsys)


def _t_freeform_gate_no(tmp_path, monkeypatch, capsys):
    """素の断り（--allow-freeform 無し）。"""
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    _freeform_setup(monkeypatch)
    return _run_main(["run", str(book), "セルに画像を貼って", "--copy"], capsys)


def _t_freeform_noninteractive(tmp_path, monkeypatch, capsys):
    """★ 生成/適用が無くなった以上「対話で y/N を聞く」余地自体が無いが、input() に
       一切触れずに断ることを固定する（呼んだら例外で気づける形にする）。"""
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    _freeform_setup(monkeypatch)

    def _boom(prompt=""):
        raise AssertionError("★ 廃止後は input() を呼んではいけない（確認そのものが無い）")
    monkeypatch.setattr("builtins.input", _boom)
    return _run_main(["run", str(book), "セルに画像を貼って", "--copy"], capsys)


def _t_freeform_helper_sweep_detected(tmp_path, monkeypatch, capsys):
    """★ ヘルパ総なめ検出は生成が起きて初めて意味を持つ器官なので、単発の語彙外経路には
       もう乗らない（複合計画側にだけ残る）。名前は互換のため残し、中身は OUT_OF_VOCAB
       （about 付き）でも同じ断りになることの固定に置き換える。"""
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    _freeform_setup(monkeypatch, op="OUT_OF_VOCAB", about="条件付き書式")
    return _run_main(["run", str(book), "条件付き書式で色を付けて", "--copy"], capsys)


# ===========================================================================
# 破壊の関所: y / N / 非対話
# ===========================================================================

def _overwrite_gate_setup(tmp_path):
    return _book(tmp_path, [["商品", "数量", "単価", "金額"],
                             ["a", 2, 100, 999], ["b", 3, 150, 999]])   # 金額に既存値あり


def _t_overwrite_gate_yes(tmp_path, monkeypatch, capsys):
    book = _overwrite_gate_setup(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "COMPUTE_COLUMN",
                          "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws.cell(row=2, column=4, value=200)
        ws.cell(row=3, column=4, value=450)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    # ★ 破壊の関所は --copy 時は走らない（原本に触れないため確認不要・
    #   test_confirm_overwrite_or_gate_none_when_copy_mode の実装どおり）。
    #   関所そのものを見るシナリオなので --copy は付けない（既定=原本直接反映）。
    return _run_main(["run", str(book), "金額を数量×単価で上書きして", "--values"], capsys)


def _t_overwrite_gate_no(tmp_path, monkeypatch, capsys):
    book = _overwrite_gate_setup(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "COMPUTE_COLUMN",
                          "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}})
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    return _run_main(["run", str(book), "金額を数量×単価で上書きして", "--values"], capsys)


def _t_overwrite_gate_noninteractive(tmp_path, monkeypatch, capsys):
    book = _overwrite_gate_setup(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "COMPUTE_COLUMN",
                          "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}})

    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    return _run_main(["run", str(book), "金額を数量×単価で上書きして", "--values"], capsys)


# ===========================================================================
# 忠実度ゲート(exit 4) / Excel ロック(exit 5) / run ロック(exit 6)
# ===========================================================================

def _t_fidelity_gate(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    ws = openpyxl.load_workbook(book).active
    wb = ws.parent
    ws.conditional_formatting.add("B2:B3", CellIsRule(operator="greaterThan", formula=["150"]))
    wb.save(book)

    def fake_normalize(b, workdir, timeout=None):
        norm = workdir / ("normalized" + b.suffix)
        w = openpyxl.Workbook()
        s = w.active
        s.append(["商品", "金額"])
        s.append(["a", 100])
        s.append(["b", 200])   # 正規化後は CF が消えている(往復忠実度ロス)
        w.save(norm)
        return norm
    monkeypatch.setattr(ailine, "normalize_book", fake_normalize)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "asc"}})
    return _run_main(["run", str(book), "金額で並べ替えて"], capsys)   # --copy なし=既定(原本直接)


def _t_excel_lock(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    (tmp_path / f"~${book.name}").write_bytes(b"lock")
    return _run_main(["run", str(book), "何か変更して", "--copy"], capsys)


def _t_run_lock_busy(tmp_path, monkeypatch, capsys):
    # ★ 2026-08-24: 判定を OS の排他ロックに移したので、偽の PID を書いて _pid_alive を
    #   差し替える手はもう通じない。★ 危なかった実測: そのまま golden を再生成したら
    #   「exit 6 のロック拒否」が「exit 9 の ollama 失敗」に化けて保存されかけた
    #   ── 検体が別の場面を撮っていることに、再生成は気づかない。
    #   本物の持ち主を別プロセスで立てる。
    import subprocess
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    lock_path = tmp_path / "run.lock"
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", lock_path)
    parts = [
        "import sys",
        "sys.path.insert(0, r%r)" % str(Path(__file__).resolve().parent.parent / "src"),
        "import ailine, time",
        "from pathlib import Path",
        "ok, _ = ailine.acquire_run_lock(Path(r%r))" % str(lock_path),
        "print(chr(79) + chr(75) if ok else chr(78) + chr(71), flush=True)",
        "time.sleep(120)",
    ]
    child = subprocess.Popen([sys.executable, "-c", chr(10).join(parts)],
                              stdout=subprocess.PIPE, text=True, encoding="utf-8")
    assert child.stdout.readline().strip() == "OK", "子がロックを取れていない"
    try:
        return _run_main(["run", str(book), "何か変更して", "--copy"], capsys)
    finally:
        child.kill(); child.wait(timeout=10)


# ===========================================================================
# header-row 指定
# ===========================================================================

def _t_header_row_explicit(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["x", 1], ["y", 2], ["商品", "金額"], ["a", 300], ["b", 100]])
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0}, 2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0}, 4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "build_struct_dump", lambda b, workdir: ambiguous)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy", "--header-row", "3"],
                      capsys)


# ===========================================================================
# --dry × 3経路
# ===========================================================================

def _t_dry_dsl(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    return _run_main(["run", str(book), "金額で降順に並べ替えて", "--dry"], capsys)


def _t_dry_plan(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                   {"op": "BOLD", "args": {"target": "row:1"}}]})
    return _run_main(["run", str(book), "金額で降順に並べ替えて見出しを太字に", "--dry"], capsys)


def _t_dry_freeform(tmp_path, monkeypatch, capsys):
    """★ freeform 最終決定: 断りにはレビューする生成物が無いので --dry に実質的な意味は
       無いが、クラッシュせず同じ断りになることを固定する。"""
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    _freeform_setup(monkeypatch)
    return _run_main(["run", str(book), "セルに画像を貼って", "--dry"], capsys)


# ===========================================================================
# ★ 単位E: 対象スロットの出所（③矛盾 / ②無言）
# ===========================================================================

def _t_subject_contradiction(tmp_path, monkeypatch, capsys):
    """★ 症状そのもの: 依頼文には「見出し」とあるのに、2段目の対象が前段の新規列
       『数量*単価』に解決された run。✓ を出さず、⚠ と（--copy なので聞かないまま）終わる。"""
    book = _book(tmp_path, [["商品", "数量", "単価"], ["a", 2, 100], ["b", 3, 200]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "COMPUTE_COLUMN",
                                    "args": {"operands": ["数量", "単価"], "operator": "*"}},
                                   {"op": "BOLD", "args": {"target": "col:数量*単価"}}]})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if 'setString("数量*単価")' in code:
            ws.cell(row=1, column=4, value="数量*単価")
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=4,
                        value=(ws.cell(row=r, column=2).value or 0) * (ws.cell(row=r, column=3).value or 0))
        if "Call StyleBold(oDoc, 3, " in code:
            for r in range(1, ws.max_row + 1):
                if ws.cell(row=r, column=4).value not in (None, ""):
                    ws.cell(row=r, column=4).font = Font(bold=True)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    return _run_main(["run", str(book), "数量と単価をかけた金額列を作って、見出しを太字にして",
                       "--copy", "--values"], capsys)


def _t_subject_substring_contradiction(tmp_path, monkeypatch, capsys):
    """★ 単位B: 依頼は「税込金額で」なのに解決値が『金額』（＝税抜きの列）。『金額』は依頼文に
       部分文字列としては現れるが、その出現は『税込金額』の一部としてしか説明できないので
       照合の証拠にしない ―― 適用そのものは成功するが ✓ は出さない。
       ★ 逆向き（依頼「金額で」→ 解決値『税込金額』）は純関数側の対照の対で凍結してある
       （tests/test_subject_provenance.py の TestSubstringDirections）。"""
    book = _book(tmp_path, [["商品", "金額", "税込金額"], ["a", 300, 330], ["b", 200, 220]])
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "asc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for i, row in enumerate([["b", 200, 220], ["a", 300, 330]], start=2):
            for j, v in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=v)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    return _run_main(["run", str(book), "税込金額で並べ替えて", "--copy"], capsys)


def _t_subject_unspoken_note(tmp_path, monkeypatch, capsys):
    """★ ②: 依頼文は対象について無言（「太字にして」だけ）。✓ は出すが、その run 固有の
       1文で範囲を狭める（旧・常時注記の置き換え）。"""
    book = _plan_book_presorted_bold(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "BOLD", "args": {"target": "row:1"}})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))
    return _run_main(["run", str(book), "太字にして", "--copy"], capsys)


def _book_with_handmade_summary(tmp_path):
    """★ 単位G: 人が手で作った『集計』シートがあるブック。盲検査定の致命2件目の実物
       （無関係な手作りの『集計』(年度/予算) が SummaryTable の removeByName で全滅した）。"""
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for r in [["部門", "金額"], ["営業", 100], ["営業", 200], ["開発", 300]]:
        ws.append(r)
    s = wb.create_sheet("集計")
    for r in [["年度", "予算"], [2025, 5000], [2026, 6000]]:
        s.append(r)
    wb.save(p)
    return p


def _t_unit_g_declared_sheet_premise_broken(tmp_path, monkeypatch, capsys):
    """★★ 単位G: 宣言(writes=new_sheet)は「その名前のシートは before に存在しない」を前提に
       警告を黙らせる権利を持つ。手作りの『集計』が既に在る＝前提が破れているので、権利を失う。
       ★ この golden が凍結するのは **「（既存シート『集計』の更新は意図どおりです）」が
       画面に出ないこと**（単位G の完了条件）。関所が止めた 1 行上に肯定文が出ていたのが
       元の欠陥で、純関数のテストではその並びまでは測れない。
       ★ 正常系（前に ailine 自身が作った『集計』の作り直し）を肯定文へ戻すのは 単位H の仕事。"""
    book = _book_with_handmade_summary(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        # SummaryTable ヘルパの removeByName("集計") → 作り直し を模す（中身が全部変わる）
        wb = openpyxl.load_workbook(out_book)
        del wb["集計"]
        s = wb.create_sheet("集計")
        for r in [["部門", "合計"], ["営業", 300], ["開発", 300]]:
            s.append(r)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    # --copy なし = 既定(原本直接) → 前提が破れれば破壊の関所が非対話で exit 7
    return _run_main(["run", str(book), "部門ごとに金額をまとめて"], capsys)


def _t_unit_g_composite_plan_declared_sheet_premise_broken(tmp_path, monkeypatch, capsys):
    """★★ 単位G(複合計画版): 上の _t_unit_g_declared_sheet_premise_broken と同じ致命
       （手作りの『集計』を AGGREGATE が破壊）を、複合計画([AUTOFIT, AGGREGATE])経由で再現する。
       敵対検証で確認された配線漏れ: _run_dsl_plan_step の compose_dsl_step_advisories 呼び出しに
       precondition_broken が渡っていなかったため、単発では出ない
       「（既存シート『集計』の更新は意図どおりです）」が複合計画だけで出ていた。
       ★ --overwrite で関所は素通しになる（承知の上の続行）が、その先の助言は前提が破れた
       ことを正しく伝えなければならない ── 関所を通したことと、破れた宣言を「意図どおり」と
       言うことは別問題（前者は利用者の選択・後者は嘘）。"""
    book = _book_with_handmade_summary(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "AUTOFIT", "args": {}},
                                   {"op": "AGGREGATE",
                                    "args": {"group_col": "部門", "value_col": "金額"}}]})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        if "AutoFitColumns" in code:
            wb["Sheet"].column_dimensions["A"].width = 20   # ★ check_autofit の pass 条件
        elif "SummaryTable" in code:
            # SummaryTable ヘルパの removeByName("集計") → 作り直し を模す（中身が全部変わる）
            del wb["集計"]
            s = wb.create_sheet("集計")
            for r in [["部門", "合計"], ["営業", 300], ["開発", 300]]:
                s.append(r)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    # --copy なし = 既定(原本直接)。--overwrite = 破壊の関所を承知の上で続行（非対話で exit 7 に
    # ならない） ── 単発の EOFError 版と違い、ここは「関所を通した後の助言」の正しさを見る。
    return _run_main(["run", str(book), "列幅を整えて部門ごとに金額をまとめて", "--overwrite"], capsys)


CASES = {
    "subject_contradiction": _t_subject_contradiction,
    "subject_substring_contradiction": _t_subject_substring_contradiction,   # ★ 単位B
    "subject_unspoken_note": _t_subject_unspoken_note,
    "dsl_pass": _t_dsl_pass,
    "dsl_warn": _t_dsl_warn,
    "dsl_fail": _t_dsl_fail,
    "dsl_runtime_error": _t_dsl_runtime_error,
    "plan_all_ok": _t_plan_all_ok,
    "plan_mixed_warn": _t_plan_mixed_warn,
    "plan_partial_fail": _t_plan_partial_fail,
    "plan_with_clarify": _t_plan_with_clarify,
    "freeform_gate_yes": _t_freeform_gate_yes,
    "freeform_gate_no": _t_freeform_gate_no,
    "freeform_noninteractive": _t_freeform_noninteractive,
    "freeform_helper_sweep_detected": _t_freeform_helper_sweep_detected,
    "overwrite_gate_yes": _t_overwrite_gate_yes,
    "overwrite_gate_no": _t_overwrite_gate_no,
    "overwrite_gate_noninteractive": _t_overwrite_gate_noninteractive,
    "fidelity_gate_exit4": _t_fidelity_gate,
    "excel_lock_exit5": _t_excel_lock,
    "run_lock_exit6": _t_run_lock_busy,
    "header_row_explicit": _t_header_row_explicit,
    "dry_dsl": _t_dry_dsl,
    "dry_plan": _t_dry_plan,
    "dry_freeform": _t_dry_freeform,
    "unit_g_declared_sheet_premise_broken": _t_unit_g_declared_sheet_premise_broken,   # ★ 単位G
    # ★ 単位G(複合計画版・敵対検証で発見された配線漏れの再現)
    "unit_g_composite_plan_declared_sheet_premise_broken":
        _t_unit_g_composite_plan_declared_sheet_premise_broken,
}


# ★★ 単位G の番人（ゴールデンとは別に置く）: 承認テストは AILINE_REGEN_GOLDEN=1 で
#   再生成できてしまうので、「肯定文が戻ってきた」を凍結ファイルだけに頼ると、再生成した
#   瞬間に黙って通る。★ 消えたことは diff に出ない（feedback_negative_coverage）ので、
#   禁止文字列の不在をここで直接主張する ── この assert は再生成では消えない。
def test_unit_g_affirmative_line_never_returns(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    rc, out = _t_unit_g_declared_sheet_premise_broken(tmp_path, monkeypatch, capsys)
    assert "意図どおりです" not in out, out
    # 対の主張: 黙らせる代わりに、破れた前提と置き換えの両方をちゃんと述べていること
    assert "既存シート『集計』の中身が置き換わりました" in out
    assert "新しいシートを作るはずが" in out
    assert rc == 7   # 非対話で関所が止めた（原本は無変更）


# ★★ 単位G(複合計画版) の番人: 上の単発版と対になる ── _run_dsl_plan_step が
#   precondition_broken を配線し忘れていた不具合そのものの再発防止（単発だけ塞いで
#   複合計画を塞ぎ忘れる、という同型の事故を機械で止める）。
def test_unit_g_composite_plan_affirmative_line_never_returns(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    rc, out = _t_unit_g_composite_plan_declared_sheet_premise_broken(tmp_path, monkeypatch, capsys)
    assert "意図どおりです" not in out, out
    # 対の主張: 黙らせる代わりに、破れた前提と置き換えの両方をちゃんと述べていること
    # （★ 警告と旧・誤った肯定文が矛盾して同時に出ていたのが敵対検証の再現そのもの）。
    assert "既存シート『集計』の中身が置き換わりました" in out
    assert "新しいシートを作るはずが" in out
    assert rc == 0   # --overwrite で関所は承知の上で通した（原本は書き換わる）


_ISO_TS_RE = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\+\d{2}:\d{2}")


@pytest.mark.parametrize("name", sorted(CASES.keys()))
def test_transcript_golden(tmp_path, monkeypatch, capsys, name):
    _isolate(monkeypatch, tmp_path)
    rc, out = CASES[name](tmp_path, monkeypatch, capsys)
    # ★ run_lock_exit6 だけ「実行中ロックの発行時刻」が実行のたび変わる現在時刻を含む
    #   （_lock_is_stale が30分超で奪取可能と判定するため、固定の過去日時は使えない）。
    #   golden の再現性のため ISO タイムスタンプを一律 <TS> に正規化してから比較する。
    out = _ISO_TS_RE.sub("<TS>", out)
    # ★ 2026-08-24: ロックの持ち主を本物のプロセスにしたので pid が毎回変わる。
    #   時刻と同じく伏せる（撮りたいのは「誰か」でなく「拒否の形」）。
    out = re.sub(r"pid=\d+", "pid=<PID>", out)
    # ★ CI の長期赤の一因（2026-08-21 実測）: freeform バナーの 参照ライブラリ/ヘルパ 行が
    #   repo の絶対パスを含み、golden に生成マシンのパス（C:\Dev\ailine）が焼き込まれて
    #   CI（D:\a\ailine\ailine）で必ず不一致になった。repo root を <REPO> に正規化する。
    out = out.replace(str(Path(__file__).resolve().parent.parent), "<REPO>")
    # rc をトランスクリプトの先頭に埋め込む（終了コードの回帰もこの1ファイルで拾える）。
    text = f"[exit code: {rc}]\n{out}"
    assert_golden_text(F9_DIR / f"{name}.txt", text, label=name)
