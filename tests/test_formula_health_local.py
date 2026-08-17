"""★ 宣言つき挙動変更#1: 型破壊の安全網 — 実機E2E（DoD5）。

査定と同じシナリオ「原価の列を全部『0円』にして」を、本物の basrun/LibreOffice 経由で
`ailine.py run` の CLI 全体を通して確認する。
①査定と同じ非数値文字列の書き込み → 依存する数式(利益=売上-原価)が本物の LibreOffice
  再計算で #VALUE! になり、(a)(b) 両方の助言が実際に出る。
②同じ列に数値そのもの（"500"）を書く正常系 → LibreOffice の暗黙の数値変換で数式は
  壊れず、(a)(b) どちらも出ない（過剰検出でないことの実機側の裏取り）。

★ fake_apply 経由のモック統合試験（DoD1 の回帰）は tests/test_ailine.py の
「★ 宣言つき挙動変更#1」セクションが担う。ここは「本物の basrun/LibreOffice が実際に
#VALUE! を生成し、それを formula_health が正しく読み戻せるか」という、モックでは
検証できない結線そのものを見る。

★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
（`pytest -m "not local"` で除外される。手動は `pytest -m local` または
`pytest tests/test_formula_health_local.py`）。
★ tests/test_bold_local.py と同じ後始末方針（taskkill 名前一括はしない・
ailine._stop_office() に委譲）。
"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
import ailine  # noqa: E402

from _run_argv import run_argv  # noqa: E402  — C2: cmd_run 直呼び用 Namespace → main(argv) 変換


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1   # 数えられない環境。local マーカーの手動実行前提なので実害は小さい
    return out.lower().count("soffice.bin")


def _genka_book(tmp_path) -> Path:
    """査定と同じ形の検体: 品目/売上/原価/利益(=売上-原価) の7行。"""
    p = tmp_path / "genka.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["品目", "売上", "原価", "利益"])
    rows = [("りんご", 1000, 300), ("みかん", 800, 200), ("ぶどう", 1500, 600),
            ("もも", 1200, 400), ("なし", 900, 350), ("バナナ", 600, 150), ("いちご", 2000, 800)]
    for i, (name, sales, cost) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=sales)
        ws.cell(row=i, column=3, value=cost)
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
    wb.save(p)
    return p


@pytest.mark.local
def test_set_column_value_nonnumeric_breaks_formula_via_real_basrun(tmp_path, monkeypatch, capsys):
    """①査定の再現: 本物の basrun/LibreOffice で『原価』列を『0円』に一括書換すると、
       利益列(=売上-原価)が実際に #VALUE! になり、(a) エラー値増加の網と (b) 型変化の
       助言の両方が出ることを確認する。"""
    before_procs = _soffice_process_count()
    try:
        book = _genka_book(tmp_path)
        monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
        monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
        monkeypatch.setattr(
            ailine, "translate_task",
            lambda model, task, book_meta, temperature=0.1:
                {"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "原価"}}]})

        argv = run_argv(
            book=str(book), task="原価の列を全部『0円』にして", model="qwen2.5-coder:7b",
            refs=None, helpers=None, repair=0, temperature=0.2,
            dry=False, copy=True, json=False, timeout=180.0, ask=False, overwrite=True)
        rc = ailine.main(argv)
        captured = capsys.readouterr()

        assert rc == 0, captured.out
        # ★ C9: 単発の ✓ バナーも読み戻し後の1行に統合された。
        assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out, captured.out
        assert "★ 疑わしい: 適用後にエラー値のセルが増えました" in captured.out, captured.out
        assert "#VALUE!" in captured.out, captured.out
        assert "（確認）列『原価』は元は数値でしたが" in captured.out, captured.out

        # ★ 読み戻しでも実際に #VALUE! になっていることを裏取りする（表示だけでなく実体）。
        out_book = book.with_name(book.stem + ".out" + book.suffix)
        wb2 = openpyxl.load_workbook(out_book, data_only=True)
        ws2 = wb2.active
        assert ws2["D2"].value == "#VALUE!"
        wb2.close()
    finally:
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )


@pytest.mark.local
def test_set_column_value_numeric_looking_text_no_advisory_via_real_basrun(tmp_path, monkeypatch, capsys):
    """②過剰検出でないことの実機側の裏取り: 同じ『原価』列に数値そのものの文字列『500』を
       書く正常系では、LibreOffice の暗黙の数値変換で利益列の数式は壊れず、(a)(b) どちらの
       助言も出ない。"""
    before_procs = _soffice_process_count()
    try:
        book = _genka_book(tmp_path)
        monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
        monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
        monkeypatch.setattr(
            ailine, "translate_task",
            lambda model, task, book_meta, temperature=0.1:
                {"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "原価"}}]})

        argv = run_argv(
            book=str(book), task="原価の列を全部『500』にして", model="qwen2.5-coder:7b",
            refs=None, helpers=None, repair=0, temperature=0.2,
            dry=False, copy=True, json=False, timeout=180.0, ask=False, overwrite=True)
        rc = ailine.main(argv)
        captured = capsys.readouterr()

        assert rc == 0, captured.out
        assert "★ 疑わしい: 適用後にエラー値のセルが増えました" not in captured.out, captured.out
        assert "（確認）列『" not in captured.out, captured.out

        out_book = book.with_name(book.stem + ".out" + book.suffix)
        wb2 = openpyxl.load_workbook(out_book, data_only=True)
        ws2 = wb2.active
        for r in range(2, 9):
            assert ws2.cell(row=r, column=4).value != "#VALUE!", (
                f"D{r} が #VALUE! になっている（LibreOffice の暗黙の数値変換が効いていない可能性）")
        wb2.close()
    finally:
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )
