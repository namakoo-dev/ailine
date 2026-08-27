"""算術恒等の検算 ―― 「上の行を全部足した値」を持つ行を、語を読まずに見つける検体。

★ この検体群は**実装より先に**書いた（先に赤を見る）。後から書くと、実装が通る検体を
書いてしまう。

## なぜ（独立レビューの実測）

`check_append_total` は**期待値を「合計式が生成したのと同じ範囲」から作っていた** ――
検算が被検算と同じ盲点を使う恒真式。既存の合計 300 を持つ表に合計を足すと 600 が書かれ、
「3 行の合計を検証」と言って `✓ 機械検証済み`・exit 0 で原本を上書きした。
並べ替えの事後条件にも同型がある: 値のみの合計行つきの表を「金額で降順」に並べ替えると、
合計行が 2 行目に来ても「5 行を検証（降順）」で通り、✓ が出て原本を上書きした。

この型は repo が既に禁じている（docs/behavior-corpus/nodes/empty-verification-ban.md）。

## 事前登録したバー（ブリーフで凍結・ここでは変えない）

| 検体 | 期待 |
|---|---|
| T6  | 既存の合計が「値」で入っている表に合計を追加 → exit 1・原本無傷・該当行を名指し |
| T7  | 既存の合計が「式」で入っている表に合計を追加 → exit 1・直せる形のメッセージ |
| T8b | 値のみの合計行つき表を「金額で降順」 → ✓ を出さない・原本を上書きしない |
| F6  | データ行が 2 行以下 → 沈黙 |
| F7  | 累計列（各行が上を足し込む列） → 警告 0 本 |
| F8  | factor 1.1（税込み等）の正常系 → 警告 0 本 |
| F9  | 空行で区切られた 2 つの塊 → 誤爆しない |
| F10 | 同梱の `demo/sales.xlsx` に合計を足す（README の quickstart） → exit 0・✓ |

★ バーの言い方: 目標は「二重計上を**検出できる**」ではなく
**「二重計上した値に ✓ 機械検証済みを出さない」**。

## F10 を後から足した理由（実測で足りなかったバー）

最初の実装は「最終行以外で恒等式が成り立つ行」を全部鳴らしていた。同梱の
`demo/sales.xlsx`（部門/金額 = 100,200,**300**,400,500,250）は 300 が開発部門の
ただの売上なのに 100+200 と一致し、**README の quickstart が exit 1 で落ちた** ――
製品の最初の 30 秒で誤爆する状態だった。F6〜F9 は全部緑のまま。
**同梱の検体そのものを走らせるバーが1本も無かったから気づけなかった。**
だから F10 は合成した表ではなく `demo/` の実ファイルを使う。
"""
import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

from _run_argv import run_argv  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402  — 小道具を二重管理しない


# --- 土台 -------------------------------------------------------------------

def _book(tmp_path, rows, name="b.xlsx") -> Path:
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(p)
    return p


def _eof(prompt=""):
    raise EOFError()   # 非対話（パイプ/CI と同じ）


def _run(tmp_path, monkeypatch, book, task, op, args, fake_apply, **over):
    """既定（原本へ直接反映）で1回 run する。翻訳と LibreOffice だけを差し替える。"""
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    (tmp_path / "notice_v2_shown").write_text("shown", encoding="utf-8")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", tmp_path / "run.lock")
    monkeypatch.setattr(ailine, "normalize_book", lambda book_, workdir, timeout=None: book_)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task_, book_meta, temperature=0.1: {"op": op, "args": dict(args)})
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr("builtins.input", _eof)
    return ailine.main(run_argv(book=str(book), task=task, repair=0, **over))


def _cell(path: Path, addr: str, sheet: str = "Sheet"):
    wb = openpyxl.load_workbook(path)
    try:
        return wb[sheet][addr].value
    finally:
        wb.close()


# ===========================================================================
# T6 / T7: APPEND_TOTAL —— 既存の合計を足し込んで二重に数える
# ===========================================================================

def _append_total_fake(total_row: int, col_letter: str, cached, label_cell: str, label="合計",
                        other_caches: dict | None = None):
    """APPEND_TOTAL のヘルパが実際に書くもの（表の下端 +1 に ラベル + 挿入耐性 SUM）。

    ★ other_caches: 表に元からあった式のキャッシュ値。openpyxl は保存時に数式キャッシュを
      落とすが、実機の LibreOffice は往復のたび**全部の式を計算し直して**書く。ブック内の
      既存の式が「キャッシュ無し」で残るのは openpyxl の産物なので、そこは実機に合わせる。
    """
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws[label_cell] = label
        ws[f"{col_letter}{total_row}"] = (
            f"=SUM({col_letter}2:INDEX({col_letter}:{col_letter},ROW()-1))")
        wb.save(out_book)
        caches = dict(other_caches or {})
        caches[f"{col_letter}{total_row}"] = cached
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", caches)
        return True, None, "ok"
    return fake


def test_t6_existing_total_as_value_is_not_certified(tmp_path, monkeypatch, capsys):
    """T6: 既存の合計が「値」で入っている表に合計を追加 → exit 1・原本無傷・該当行を名指し。

    金額列 = [100, 200, 300(既存の合計)]。ヘルパは下端 +1 (row5) に SUM を書き、
    LibreOffice は 600 を計算する ―― 300 が二重に数えられている。
    """
    book = _book(tmp_path, [["品名", "金額"], ["あ", 100], ["い", 200], ["合計", 300]])
    fake = _append_total_fake(total_row=5, col_letter="B", cached=600, label_cell="A5")

    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "✓" not in out, out                      # ✓ 機械検証済みを出さない
    assert _cell(book, "B5") is None, "原本が上書きされている"
    assert _cell(book, "B4") == 300
    # ★ 該当行を名指しする（「エラー」だけの出力は禁止）
    assert "B4" in out, out
    assert "300" in out, out


def test_t7_existing_total_as_formula_is_named(tmp_path, monkeypatch, capsys):
    """T7: 既存の合計が「式」で入っている表に合計を追加 → exit 1・直せる形のメッセージ。

    ★ 実装前も exit 1 にはなる（既存の「式が期待形でない」判定が、**新しく書いた行では
    なく既存の合計行**を掴んで落ちるため）。だがその文面は「4行目: 合計の式が期待形
    (挿入耐性 SUM 型)でない」＝ユーザーが自分で書いた行を責める誤診断で、直しようがない。
    バーは「直せる形のメッセージ」なので、**既存の合計行を名指しすること**を要求する。
    """
    book = _book(tmp_path, [["品名", "金額"], ["あ", 100], ["い", 200], ["合計", None]])
    wb = openpyxl.load_workbook(book)
    wb.active["B4"] = "=SUM(B2:B3)"
    wb.save(book)
    _inject_formula_cache(book, "xl/worksheets/sheet1.xml", {"B4": 300})

    fake = _append_total_fake(total_row=5, col_letter="B", cached=600, label_cell="A5",
                               other_caches={"B4": 300})
    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "✓" not in out, out
    assert _cell(book, "B5") is None, "原本が上書きされている"
    assert "B4" in out, out
    assert "300" in out, out
    assert "合計と一致" in out, out
    assert "期待形" not in out, out   # ユーザーが自分で書いた行を責める誤診断は出さない


# ===========================================================================
# T8b: SORT —— 合計行が最下行でなくなる
# ===========================================================================

_T8_ROWS = [
    ["商品", "数量", "単価", "金額"],
    ["りんご", 10, 300, 3000],
    ["ばなな", 20, 300, 6000],
    ["みかん", 15, 300, 4500],
    ["ぶどう", 30, 300, 9000],
    ["合計", None, None, 22500],          # 金額は値のみ（数式でない）
]


def _sort_desc_fake(order):
    """金額で降順に並べ替える（合計行も一緒に動く ―― 実測した欠陥そのもの）。"""
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for i, row in enumerate(order, start=2):
            for c, v in enumerate(row, start=1):
                ws.cell(row=i, column=c, value=v)
        wb.save(out_book)
        return True, None, "ok"
    return fake


def test_t8b_total_row_moved_off_the_bottom_is_not_certified(tmp_path, monkeypatch, capsys):
    """T8b: 値のみの合計行つき表を「金額で降順」→ ✓ を出さない・原本を上書きしない。"""
    book = _book(tmp_path, _T8_ROWS)
    fake = _sort_desc_fake([
        ["合計", None, None, 22500],
        ["ぶどう", 30, 300, 9000],
        ["ばなな", 20, 300, 6000],
        ["みかん", 15, 300, 4500],
        ["りんご", 10, 300, 3000],
    ])
    rc = _run(tmp_path, monkeypatch, book, "金額で降順に並べ替えて",
              "SORT", {"col": "金額", "order": "desc"}, fake)
    out = capsys.readouterr().out
    assert "✓" not in out, out
    assert rc != 0, out
    assert _cell(book, "A2") == "りんご", "原本が上書きされている"
    assert _cell(book, "A6") == "合計"
    assert "22500" in out, out
    assert "D6" in out or "D2" in out, out   # 動く前/動いた後のどちらかは名指しされる


# ===========================================================================
# F6〜F9: 沈黙しなければならない側（誤爆はバー違反）
# ===========================================================================

def test_f6_two_data_rows_stays_silent(tmp_path, monkeypatch, capsys):
    """F6: データ行が 2 行以下 → 沈黙。

    ★ [100, 100] のような「2 行目が 1 行目と等しい」表は、上を1つだけ足す恒等式なら
    必ず当たる。この線を先に凍結する（合計とみなすには上に2つ以上の項が要る）。
    """
    book = _book(tmp_path, [["品名", "金額"], ["あ", 100], ["い", 100]])
    fake = _append_total_fake(total_row=4, col_letter="B", cached=200, label_cell="A4")
    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out
    assert _cell(book, "B4") is not None


def test_f7_running_total_column_stays_silent(tmp_path, monkeypatch, capsys):
    """F7: 累計列（各行が上を足し込む列）に合計を足しても警告 0 本。"""
    book = _book(tmp_path, [
        ["日付", "売上", "累計"],
        ["4/1", 100, 100],
        ["4/2", 200, 300],
        ["4/3", 150, 450],
        ["4/4", 300, 750],
    ])
    # ★ 2026-08-28: 合計行のラベルは 1 列目（旧: 対象列の左隣）。
    fake = _append_total_fake(total_row=6, col_letter="C", cached=1600, label_cell="A6")
    rc = _run(tmp_path, monkeypatch, book, "累計の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "累計"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out


def test_f7_running_total_column_sort_stays_silent(tmp_path, monkeypatch, capsys):
    """F7（並べ替え側）: 累計列で並べ替えても警告 0 本（最下行は合計ではない）。"""
    rows = [["日付", "売上", "累計"], ["4/1", 100, 100], ["4/2", 200, 300],
            ["4/3", 150, 450], ["4/4", 300, 750]]
    book = _book(tmp_path, rows)
    fake = _sort_desc_fake([["4/4", 300, 750], ["4/3", 150, 450],
                            ["4/2", 200, 300], ["4/1", 100, 100]])
    rc = _run(tmp_path, monkeypatch, book, "累計で降順に並べ替えて",
              "SORT", {"col": "累計", "order": "desc"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out


def test_f8_factor_tax_included_stays_silent(tmp_path, monkeypatch, capsys):
    """F8: factor 1.1（税込み等）の正常系 → 警告 0 本。"""
    book = _book(tmp_path, [["品名", "金額"], ["あ", 1200], ["い", 3400], ["う", 2500]])
    fake = _append_total_fake(total_row=5, col_letter="B", cached=7810,
                               label_cell="A5", label="税込合計")
    rc = _run(tmp_path, monkeypatch, book, "金額の税込合計（消費税10%）を一番下に出して",
              "APPEND_TOTAL", {"col": "金額", "label": "税込合計", "factor": 1.1}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out


def test_f9_two_blocks_separated_by_blank_row_do_not_misfire(tmp_path, monkeypatch, capsys):
    """F9: 空行で区切られた 2 つの塊 → 誤爆しない。

    ★ 2 つの塊を続けて読んでしまうと 4600 == 1200+3400 が当たる。空行で切れていれば
    上の塊しか見ないので黙る（この対比は tests/test_sum_identity_unit.py が直接測る）。
    """
    rows = [["商品", "金額"], ["a", 1200], ["b", 3400], [None, None],
            ["c", 4600], ["d", 800]]
    book = _book(tmp_path, rows)
    fake = _sort_desc_fake([["b", 3400], ["a", 1200]])
    rc = _run(tmp_path, monkeypatch, book, "金額で降順に並べ替えて",
              "SORT", {"col": "金額", "order": "desc"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out
    assert "4600" not in out, out


DEMO_SALES = Path(__file__).resolve().parent.parent / "src" / "ailine" / "demo" / "sales.xlsx"


def test_f10_bundled_demo_book_quickstart_still_passes(tmp_path, monkeypatch, capsys):
    """F10: 同梱の demo/sales.xlsx に合計を足す（README の quickstart）→ exit 0・✓。

    ★ 合成した表ではなく**実ファイル**を使う。100+200==300 は偶然で、300 は開発部門の
    ただの売上 ―― ここで鳴ったら製品の最初の 30 秒で誤爆する。この 1 本が無かったせいで
    最初の実装の誤爆に気づけなかった（この module の docstring 参照）。
    """
    assert DEMO_SALES.exists(), f"同梱の検体が無い: {DEMO_SALES}"
    book = tmp_path / "sales.xlsx"
    shutil.copy2(DEMO_SALES, book)          # 原本の demo/ には触らない
    values = [c.value for c in openpyxl.load_workbook(book).active["B"][1:]]
    assert values == [100, 200, 300, 400, 500, 250], f"検体が変わっている: {values}"

    fake = _append_total_fake(total_row=8, col_letter="B", cached=1750, label_cell="A8")
    rc = _run(tmp_path, monkeypatch, book, "金額の合計を一番下に出して",
              "APPEND_TOTAL", {"col": "金額"}, fake)
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "合計と一致" not in out, out
    assert "✓" in out, out
