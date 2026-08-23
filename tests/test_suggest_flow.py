# W10 便C2: もしかして提案の UX 配線 ── 実装より先に凍結した赤い検体。
# 決裁（Namakoo 2026-08-22 20:00）: 二段構え採用・6 秒は問題でない・対策は「提案の顔」──
# 対応外の部分を黙ったまま提案するのが害であり、機械の残差行で名指しすれば部分対応の
# 提案は誠実なサービス（指示は意図、保証は機械: 7B の自己棄権は 5/6 素通りの実測）。
#
# 契約:
#   S1 頷き→適用→暗黙登録→同じ言い回しの 2 回目は聞かずに直行（登録が育つ層の本体）
#   S2 判定器は第 2 段（厳格一致が空のときだけ呼ばれる）
#   S3 頷き拒否は misclass センサの第 3 信号（suggest_decline）・登録されない
#   S4 非対話は候補表示のみ・登録しない・exit 3
#   S5 残差行: 未消費の内容語を提案の中で名指し（トラップ検体で「印刷」を含む・
#      きれいな検体では残差行なし）
#   S6 判定器プロンプトは第 5 の凍結定数（test_prompt_freeze の番人対象）

import io
import json
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "judge_ops_via_llm"),
    reason="もしかして提案の配線 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


def _sort_book(tmp_path):
    return _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])


def _translate_miss(model, task, book_meta, temperature=0.1):
    # 一次翻訳は語彙外（out_of_vocab）に落ちる ── 提案層の入口
    return {"plan": [{"op": "OUT_OF_VOCAB", "about": "並べ替えのような操作", "args": {}}]}


def _fixed_op_sort(model, op, task, book_meta, temperature=0.1):
    assert op == "SORT"
    return {"op": "SORT", "args": {"col": "金額", "order": "desc"}}


def _sort_apply(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
    wb.save(out_book)
    return True, None, "ok"


def _interactive_yes(monkeypatch, answers="y\n"):
    monkeypatch.setattr("sys.stdin", io.StringIO(answers))
    monkeypatch.setattr("sys.stdin.isatty", lambda: True, raising=False)


@needs_impl
def test_nod_applies_registers_and_second_run_skips_question(tmp_path, monkeypatch, capsys):
    """S1: 厳格一致ヒット→解釈行つき提案→頷き→適用+暗黙登録（開示つき）→
       同じ言い回しの 2 回目は もしかして を出さずに通る。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_miss)
    monkeypatch.setattr(ailine, "translate_task_fixed_op", _fixed_op_sort)
    monkeypatch.setattr(ailine, "basrun_apply", _sort_apply)
    _interactive_yes(monkeypatch)
    # 「並べ替え」は SORT の label そのもの ── 厳格一致（第 1 段）が拾う形
    rc, out = _run_main(["run", str(book), "金額を並べ替えして", "--copy"], capsys)
    assert rc == 0, out
    assert "もしかして" in out, f"提案が出ていない: {out}"
    assert "解釈" in out, f"頷きの対象が解釈行でない（op 名頷きの妥協）: {out}"
    assert "登録しました" in out and "alias undo" in out, f"暗黙登録の開示が無い: {out}"
    aliases = json.loads((tmp_path / "aliases.json").read_text(encoding="utf-8"))
    assert "金額を並べ替えして" in str(aliases), f"言い回しが登録されていない: {aliases}"
    # 2 回目: 同じ言い回し → alias 直行（もしかして を再度出さない）
    book2 = _book(tmp_path, [["商品", "金額"], ["d", 10], ["c", 20]], name="c.xlsx")

    def _apply2(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "c", 20, "d", 10
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", _apply2)
    rc2, out2 = _run_main(["run", str(book2), "金額を並べ替えして", "--copy"], capsys)
    assert rc2 == 0, out2
    assert "もしかして" not in out2, f"登録済みなのに毎回聞いている（連打の禁止違反）: {out2}"


@needs_impl
def test_judge_tier_fires_only_when_strict_empty(tmp_path, monkeypatch, capsys):
    """S2: 厳格一致が空 → 判定器（第 2 段）が呼ばれる。厳格一致が拾えたら呼ばれない。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_miss)
    monkeypatch.setattr(ailine, "translate_task_fixed_op", _fixed_op_sort)
    monkeypatch.setattr(ailine, "basrun_apply", _sort_apply)
    called = {"n": 0}

    def fake_judge(task, about=None):
        called["n"] += 1
        return ["SORT"]
    monkeypatch.setattr(ailine, "judge_ops_via_llm", fake_judge)
    _interactive_yes(monkeypatch)
    # 「大きい順」は label/synonyms の語として現れない ── 厳格一致は空 → 判定器が拾う
    rc, out = _run_main(["run", str(book), "金額の大きい順で頼む", "--copy"], capsys)
    assert called["n"] == 1, f"判定器が呼ばれていない/複数回: {called['n']}\n{out}"
    assert "もしかして" in out and rc == 0, out
    # 厳格一致が拾える言い回しでは判定器を呼ばない
    called["n"] = 0
    book2 = _book(tmp_path, [["商品", "金額"], ["d", 10], ["c", 20]], name="c.xlsx")
    _interactive_yes(monkeypatch, answers="n\n")
    _run_main(["run", str(book2), "金額を並べ替えして", "--copy"], capsys)
    assert called["n"] == 0, "厳格一致で足りるのに判定器（+3秒）を呼んだ"


@needs_impl
def test_decline_records_third_signal_and_no_registration(tmp_path, monkeypatch, capsys):
    """S3: 提案に N → suggest_decline がセンサに載る・登録されない。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_miss)
    monkeypatch.setattr(ailine, "translate_task_fixed_op", _fixed_op_sort)
    _interactive_yes(monkeypatch, answers="n\n")
    rc, out = _run_main(["run", str(book), "金額を並べ替えして", "--copy"], capsys)
    # ★ 検分の締め直し（次元③の指摘）: != 0 は無関係な失敗も通す ── 断りの 3 に固定
    assert rc == 3, out
    entries = [json.loads(ln) for ln in
                (tmp_path / "misclass.jsonl").read_text(encoding="utf-8").splitlines() if ln]
    assert any(e["signal"] == "suggest_decline" for e in entries), \
        f"第 3 信号が記録されていない: {entries}"
    assert not (tmp_path / "aliases.json").exists() or \
        "並べ替え" not in (tmp_path / "aliases.json").read_text(encoding="utf-8"), \
        "N と言ったのに登録された（毒）"


@needs_impl
def test_noninteractive_shows_candidates_but_never_registers(tmp_path, monkeypatch, capsys):
    """S4: 非対話（stdin が tty でない）は候補を見せるだけ・登録しない・exit 3。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_miss)
    # ★ 検体の訂正（凍結漏れ）: fixed_op を mock しないと非対話表示のための接地が
    #   実 ollama を呼ぶ（CI には居ない）。期待値は不変。
    monkeypatch.setattr(ailine, "translate_task_fixed_op", _fixed_op_sort)
    rc, out = _run_main(["run", str(book), "金額を並べ替えして", "--copy"], capsys)
    assert rc == 3, out
    # ★ 恒真の罠回避: 依頼文の残響で当たる語は使わない ── 提案機構の固有語だけを要求
    assert "もしかして" in out, f"候補の表示が無い: {out}"
    assert not (tmp_path / "aliases.json").exists() or \
        "並べ替え" not in (tmp_path / "aliases.json").read_text(encoding="utf-8"), \
        "非対話で登録された（絶対禁止）"


@needs_impl
def test_residue_line_names_unconsumed_words(tmp_path, monkeypatch, capsys):
    """S5: 部分対応の罠 ──「並べ替えてから印刷して」の提案には、未消費の内容語
       『印刷』の名指しが含まれる。きれいな依頼には残差行を出さない。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_miss)
    monkeypatch.setattr(ailine, "translate_task_fixed_op", _fixed_op_sort)
    # ★ 検体の訂正: この言い回しは厳格一致に乗らず判定器の段で拾われる ── 判定器を
    #   決定論 mock に（段の選択は S2 の契約・ここの契約は残差の名指し）
    monkeypatch.setattr(ailine, "judge_ops_via_llm", lambda task, about=None: ["SORT"])
    _interactive_yes(monkeypatch, answers="n\n")
    rc, out = _run_main(["run", str(book), "金額で並べ替えしてから印刷して", "--copy"], capsys)
    assert "もしかして" in out, out
    assert "印刷" in out and "反映されません" in out, f"残差の名指しが無い: {out}"
    # 対照: 全部消費される依頼では残差行を出さない
    book2 = _book(tmp_path, [["商品", "金額"], ["d", 10], ["c", 20]], name="c.xlsx")
    monkeypatch.setattr(ailine, "basrun_apply", _sort_apply)
    _interactive_yes(monkeypatch, answers="n\n")
    rc2, out2 = _run_main(["run", str(book2), "金額を並べ替えして", "--copy"], capsys)
    assert "反映されません" not in out2, f"きれいな依頼に残差行が出た（オオカミ少年）: {out2}"


@needs_impl
def test_judge_prompt_is_fifth_frozen_constant():
    """S6: 判定器プロンプトが第 5 の凍結定数として番人対象に入っている。"""
    assert hasattr(ailine, "SUGGEST_JUDGE_SYSTEM"), "第 5 定数が無い"
    src = (Path(__file__).parent / "test_prompt_freeze.py").read_text(encoding="utf-8")
    assert "SUGGEST_JUDGE_SYSTEM" in src, "判定器プロンプトが凍結番人の対象外"
