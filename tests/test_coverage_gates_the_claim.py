"""走査が表の終わりに届かなかった run は、✓ も △ も名乗らないこと（2026-09-05）。

★★ 実測した事故: 明細の**途中に空行がある**表で「単価の大きい順に並べ替えて」と頼むと、
  **1 行も動いていないのに △「宣言どおりの変化を確認しました」**が出た。
  12000 / 8000 / (空) / 15000 を降順にすれば 15000 が先頭に来るはずだが、
  走査が空行で止まり、機械が見た表は先頭 2 行だけ ── その 2 行は元から降順なので
  事後条件が通る。**分母が縮んだから通った。**

★ それまでの判断（note_unverified の docstring）はこうだった:
    「8 行を本当に検証したことは事実なので、⚠ へは落とさない」
  理屈は通っている。だが画面に出るのは「**宣言どおり**」── 8 行についてではなく
  **操作全体**についての主張で、主張の範囲と確かめた範囲が食い違っていた。

★ 落とすのは「**表の終わりが分からなかった**」場合だけ。数値でない行を並べ替えの
  検算から外した等（範囲は分かっている）は今までどおり ⚠ + △ のまま。
★ 既存の分母（A1 始まり・連続の表）では 1 件も発火しないことを実測してから入れた。
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402


def test_the_flag_is_a_machine_value_not_a_message():
    """★ 表示文から読み取らせない（文言を変えた瞬間に壊れる形にしない）。"""
    args = {}
    ailine._mark_coverage_incomplete(args, "1 列目が空のため走査がそこで止まった")
    assert args["_coverage_incomplete"] == ["1 列目が空のため走査がそこで止まった"]


def test_the_flag_ignores_non_dicts():
    ailine._mark_coverage_incomplete(None, "x")      # 例外を出さない
    ailine._mark_coverage_incomplete("not a dict", "x")


def test_extent_gap_raises_the_flag_for_rows_and_for_columns():
    """★ 行が届かない時も、列が届かない時も旗が立つこと（片方だけにしない）。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src[src.index("def note_extent_gap"):src.index("\ndef check_chart")]
    assert body.count("_mark_coverage_incomplete(") == 2, \
        "行側・列側の両方から旗を立てること（片配線にしない）"


def test_finish_apply_is_the_only_place_that_decides():
    """★ 判断は 1 箇所 ── 呼び出し側 4 箇所は材料を渡すだけ。

    ここは既に同じ式を 4 回書き写していた（machine_verified=(status != "warn" and ...)）。
    同じ形の片配線を新しく作らない。
    """
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    assert src.count("coverage_incomplete=_coverage_sink)") == 4, "4 経路すべてから渡すこと"
    body = src[src.index("def _finish_apply"):src.index("\ndef ", src.index("def _finish_apply") + 10)]
    assert "machine_verified = False" in body, "落とす判断は _finish_apply の中に置く"
    # ★ 呼び出し側が「落とすかどうか」を決めていないこと
    for chunk in src.split("coverage_incomplete=_coverage_sink)")[:-1]:
        tail = chunk[-400:]
        assert "machine_verified=False" not in tail.replace(" ", ""), \
            "呼び出し側で判断している（材料だけ渡すこと）"


def test_the_note_and_the_flag_stay_separate():
    """★ 「見ていない行がある」と「表の終わりが分からない」は別物。

    数値でない行を検算から外した回（範囲は分かっている）は今までどおり △ に留める。
    旗を note_unverified 側に混ぜると、その区別が消える。
    """
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    shared = (Path(ailine.__file__).parent.parent / "ailine_core" / "postconditions"
              / "_shared.py").read_text(encoding="utf-8")
    assert "_coverage_incomplete" not in shared, \
        "note_unverified の側に旗を持ち込まない（区別が消える）"
    assert "_mark_coverage_incomplete" in src


@pytest.mark.local
def test_a_table_with_a_blank_row_inside_does_not_claim_success():
    """★ 実機 ── 途中に空行がある表で並べ替えを頼んでも ✓/△ を名乗らないこと。"""
    import openpyxl, shutil, subprocess, tempfile, os
    tmp = Path(tempfile.mkdtemp())
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "表"
    ws.append(["品名", "数量", "単価"])
    ws.append(["机", 3, 12000]); ws.append(["椅子", 7, 8000])
    ws.append([None, None, None])            # ★ 途中の空行
    ws.append(["棚", 2, 15000])
    src = tmp / "in.xlsx"; wb.save(src)
    repo = Path(ailine.__file__).resolve().parents[2]
    p = subprocess.run([sys.executable, "-m", "ailine", "run", str(src),
                        "単価の大きい順に並べ替えて", "--copy", "--timeout", "150"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       cwd=str(repo), env={**os.environ, "PYTHONPATH": str(repo / "src")})
    assert "✓" not in p.stdout, p.stdout[-500:]
    assert "△" not in p.stdout, p.stdout[-500:]
    assert "機械保証はありません" in p.stdout, p.stdout[-500:]


# --- 忠実度ゲートにも宣言を渡す（2026-09-05）--------------------------------

def test_a_declared_removal_explains_one_lost_validation():
    """★ 「消す」と宣言した回の 1 件減りは、誤報にしない。

    ★★ 実測（弥生の請求書）: 「軽減税率の列を削除して」に対し
      **⚠ 入力規則 1 件が失われています**が出た。だが中を見ると、消えた 1 件は
      **まさにその軽減税率の列に付いていたドロップダウン**で、残る 2 件は列の削除に
      追随して 1 つ左へ正しくずれていた（H→G, J→I・参照先も保たれていた）。
    ★ 件数だけを比べ、「何を消すと宣言したか」を渡していなかった ──
      判定に要る三項（依頼・宣言・実体）のうち**宣言が欠けて**いた。今日 3 度目の形。
    """
    assert ailine._removal_was_declared("DELETE_COLUMN")
    assert ailine._removal_was_declared("DELETE_ROWS")
    assert not ailine._removal_was_declared("SORT")
    assert not ailine._removal_was_declared(None)


def test_the_removal_list_comes_from_the_declaration_not_a_hand_written_list():
    """★ op 名の if を書かない ── 宣言表から引くこと（列挙は漏れる）。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src[src.index("def _removal_was_declared"):src.index("\ndef check_round_trip_fidelity")]
    assert "OP_WRITE_TARGET" in body and "WRITE_REMOVE" in body
    assert "DELETE_COLUMN" not in body, "op 名を手で書いている（宣言から引くこと）"


def test_two_lost_validations_are_still_reported_even_when_removing(tmp_path):
    """★ 宣言で説明がつくのは 1 件まで ── 2 件以上減れば、消す回でも言う。"""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    a = tmp_path / "a.xlsx"; b = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "x"
    for ref in ("B1:B5", "C1:C5", "D1:D5"):
        dv = DataValidation(type="list", formula1='"1,2"'); ws.add_data_validation(dv)
        dv.add(ref)
    wb.save(a)
    wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2["A1"] = "x"
    dv = DataValidation(type="list", formula1='"1,2"'); ws2.add_data_validation(dv)
    dv.add("B1:B5")
    wb2.save(b)
    got = ailine.check_round_trip_fidelity(a, b, op="DELETE_COLUMN")
    assert got["lost"], "2 件減っているのに黙っている"


def test_one_lost_validation_is_still_reported_when_nothing_was_removed(tmp_path):
    """★ 消す宣言をしていない op では、1 件でも言う（今までどおり）。"""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    a = tmp_path / "a.xlsx"; b = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws["A1"] = "x"
    for ref in ("B1:B5", "C1:C5"):
        dv = DataValidation(type="list", formula1='"1,2"'); ws.add_data_validation(dv)
        dv.add(ref)
    wb.save(a)
    wb2 = openpyxl.Workbook(); ws2 = wb2.active; ws2["A1"] = "x"
    dv = DataValidation(type="list", formula1='"1,2"'); ws2.add_data_validation(dv)
    dv.add("B1:B5")
    wb2.save(b)
    assert ailine.check_round_trip_fidelity(a, b, op="SORT")["lost"]
