# 行の入れ替えは、**LLM に聞く前に機械で解く**（2026-09-02）。
#
# ★★ 入れ替えを効果の検体（bench/basic_ops_matrix.py）に載せて初めて見えた:
#   8/31 に見つけた 13 件の欠陥のうち **5 件が入れ替え**だったのに、
#   入れ替えは効果の分母に**1 件も乗っていなかった**。
#
#   載せたら「あかね商事とうえだ物産の行を入れ替えて」が断られた。追うと、
#   読み直しの二段目（op を SWAP に固定して LLM に聞き直す）が
#     a='取引先' b='件数'   ← 人が言っていない**列名**
#   を返していた。しかも実在の列なので `_swap_pair_resolves` は True を返す。
#   止めたのは三項の番人（依頼文の語と照合できない）で**判断は正しい**が、
#   利用者の正当な依頼が通らない。
#   ★ 同じ言い方が 3 表（在庫・名簿・献立）では 6/6 通っていた ── つまり**揺れ**。
#
# ★ 処方は 8/31 にセルでやったものと同じ:
#   **LLM に聞く前に、機械だけで 2 つ解けているならそれを使う。**
#   依頼文と実表しか見ていないので、LLM の返事より確かで、速い。
#
# 契約:
#   ① 依頼文に literal で現れる実在の値が**ちょうど 2 行**に決まるなら、その 2 行
#   ② 決まらなければ None（推測しない ── 別の行を動かすのは取り返しがつかない）
#   ③ 見出しの語・数のように見える値は行の名前にしない（揺れを増幅しない）
#   ④ 列の入れ替えを横取りしない
#   ⑤ 読み直しは **LLM に聞く前**に走る（順番が逆だと揺れに負ける）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine  # noqa: E402
from _product_source import product_text  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む


def _book(tmp_path: Path) -> Path:
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "件数", "単価", "金額"])
    for i, (n, k, t) in enumerate([("あかね商事", 3, 12000), ("いろは工業", 8, 4500),
                                    ("うえだ物産", 2, 30000)], start=2):
        ws.append([n, k, t, f"=B{i}*C{i}"])
    wb.save(p)
    return p


def _rows(tmp_path, task):
    meta = ailine.build_book_meta(_book(tmp_path))
    return ailine.swap_targets_are_rows(task, meta, "請求")


@pytest.mark.parametrize("task", [
    "あかね商事とうえだ物産の行を入れ替えて",
    "あかね商事の行とうえだ物産の行を交換して",
    "あかね商事とうえだ物産の順番を入れ替えて",
])
def test_two_named_rows_resolve_without_the_llm(tmp_path, task):
    """① 依頼文と実表だけで 2 行に決まる。"""
    assert _rows(tmp_path, task) == [2, 4]


@pytest.mark.parametrize("task", [
    "件数と単価の列を入れ替えて",          # ④ 列（見出しの語は行の名前にしない）
    "2行目と4行目を入れ替えて",            # ③ 数は行の名前にしない（別の道が扱う）
    "あかね商事の行を削除して",            # ① 1 つしか出てこない
    "並べ替えて",                          # 名前が 1 つも無い
])
def test_it_does_not_guess(tmp_path, task):
    """② 決まらない時は決めない ── 別の行を動かすのは取り返しがつかない。"""
    assert _rows(tmp_path, task) is None


def test_a_value_that_appears_twice_is_not_a_name(tmp_path):
    """① 同じ値が 2 行に在るなら、それは名前で指せていない（番号でしか言えない）。"""
    p = tmp_path / "d.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "件数"])
    for n, k in [("ヤマノ食品", 3), ("ヤマノ食品", 8), ("うえだ物産", 2)]:
        ws.append([n, k])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    assert ailine.swap_targets_are_rows(
        "ヤマノ食品とうえだ物産の行を入れ替えて", meta, "請求") is None


def test_a_header_word_is_not_a_row_name(tmp_path):
    """③ 見出しの語は行の名前にしない ── 列の話か行の話かが決まらない。

    ★★ 2026-09-02 の変異試験で開いていた穴: 見出し除外を消しても緑のままだった
      （既存の検体には「見出しの語が**行の値としても在る**表」が無かった）。
    ★ 商品名がたまたま『単価』の表 ── 人が「単価と…」と言った時、
      列を指しているのか行を指しているのか決まらない。**決めない**のが正しい。
    ★★ 検体を 1 度作り直した: 初版は相手を『机』（1 文字）にしたため、
      **長さの条件で先に落ちて**いて、見出し除外を消しても緑のままだった
      ── その条件だけが効く入力になっていなかった（今日 3 度目の打ち消し合い）。
    """
    p = tmp_path / "h.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["品名", "単価"])
    for n, v in [("単価", 100), ("事務机", 200), ("書棚", 300)]:
        ws.append([n, v])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    assert ailine.swap_targets_are_rows("単価と事務机の行を入れ替えて", meta, "請求") is None


def test_a_number_like_value_is_not_a_row_name(tmp_path):
    """③ 数は行の名前にしない ── 依頼文に出る数はほぼ常に閾値や個数。

    ★★ 2026-08-31 の実測（Namakoo「LLM の揺れが一番厄介だ」→ 追ったら半分は機械の責任）:
      機械が『60000』を行の名前として解き、**確信をもって間違った操作**に育てていた。
      揺れは消せないが、**増幅しないことはできる**。
    """
    p = tmp_path / "n.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "金額"])
    for n, v in [("あかね商事", 60000), ("いろは工業", 45000), ("うえだ物産", 30000)]:
        ws.append([n, str(v)])          # ★ 文字として入っていても数は数
    wb.save(p)
    meta = ailine.build_book_meta(p)
    assert ailine.swap_targets_are_rows(
        "60000と45000の行を入れ替えて", meta, "請求") is None


def test_the_machine_runs_before_the_llm():
    """⑤ 順番の契約 ── 機械の解決が、LLM への聞き直しより**手前**に在ること。

    ★ ここが逆だと、機械が解けるのに LLM の揺れに負ける（今回の症状そのもの）。
    ★ 語ではなく**位置**で縛る。
    """
    i = product_text().index("swap_targets_are_rows(a.task")
    j = product_text().index('translate_task_fixed_op(a.model, "SWAP"', i - 4000 if i > 4000 else 0)
    assert i < product_text().index('translate_task_fixed_op(a.model, "SWAP"', i), (
        "機械の解決が LLM への聞き直しより後ろに在る")
    assert j >= 0


def test_the_reread_actually_fires_without_the_llm(tmp_path, monkeypatch, capsys):
    """★★ **配線そのものを通す**検体（2026-09-02 の変異試験で開いていた穴）。

      部品（swap_targets_are_rows）を直に叩く試験と、実機の 1 本は在ったが、
      **本番の読み直しの経路**を通る非実機の試験が無かった。そのため
      「読み直しを配線から外す」変異が**緑のまま**だった ── 在っても鳴らない、そのもの。
    ★ 一段目が語彙外を返す回（実測で起きた形）を作り、機械が拾い直すことだけを見る。
      適用はしない（--dry）ので LibreOffice も要らない。
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_golden_transcripts import _isolate, _run_main   # 小道具は二重管理しない

    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "OUT_OF_VOCAB", "about": "行の入れ替え"}]})
    # ★ 二段目（LLM への聞き直し）は**呼ばれてはいけない** ── 呼ばれたら落とす。
    monkeypatch.setattr(ailine, "translate_task_fixed_op",
                         lambda *a, **k: pytest.fail("機械で解けるのに LLM に聞いている"))
    _rc, out = _run_main(["run", str(p), "あかね商事とうえだ物産の行を入れ替えて",
                           "--dry"], capsys)
    assert "読み直しました" in out, out
    assert "2行目" in out and "4行目" in out, out


def test_a_column_swap_is_not_hijacked(tmp_path, monkeypatch, capsys):
    """④ 列の入れ替えを横取りしないこと（配線を通して見る）。

    ★★ 2026-09-02 の変異試験で開いていた穴: 軸の条件を外しても緑のままだった
      ── 使っていた検体が「行が解けない依頼」だったので、**横取りの機会が無かった**。
    ★ ここは行が 2 つ解ける**のに**依頼が「列」と言っている回。
      機械が行を解けるからといって、言われた軸を勝手に変えてはいけない。
    """
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_golden_transcripts import _isolate, _run_main

    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "OUT_OF_VOCAB", "about": "列の入れ替え"}]})
    monkeypatch.setattr(ailine, "translate_task_fixed_op",
                         lambda *a, **k: {"op": "SWAP", "args": {}})
    _rc, out = _run_main(["run", str(p), "あかね商事とうえだ物産の列を入れ替えて",
                           "--dry"], capsys)
    assert "2行目" not in out, ("列の依頼を行の入れ替えに横取りしている: "
                                + out)


@pytest.mark.local
def test_it_really_swaps_on_real_libreoffice(tmp_path):
    """★ 実機 ── 行が入れ替わり、**式は自分の行を指し続ける**。

    ★ 値だけ交換する実装だと、並びは正しく見えるのに各行の金額が他の行の値になる
      （実測で設計が決まった形）。だから式の指す先まで見る。
    """
    import subprocess
    p = _book(tmp_path)
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p),
         "あかね商事とうえだ物産の行を入れ替えて", "--copy"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=900,
        env={**__import__("os").environ, "PYTHONPATH": str(REPO / "src")})
    assert r.returncode == 0, r.stdout[-900:]
    ws = openpyxl.load_workbook(p.with_name(p.stem + ".out.xlsx"))["請求"]
    assert [ws.cell(r_, 1).value for r_ in (2, 3, 4)] == [
        "うえだ物産", "いろは工業", "あかね商事"]
    for r_ in (2, 3, 4):
        assert ws.cell(r_, 4).value == f"=B{r_}*C{r_}", (
            f"{r_}行目の式が自分の行を指していない: {ws.cell(r_, 4).value!r}")
