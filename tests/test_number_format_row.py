# 桁区切りを「行」にも掛ける ── 2026-08-29。Namakoo「合計を金額表示にできなかった」
#
# ★★ 実測: 「合計を金額表示にして」が**合計追加**に読まれ、既にある合計をもう一度
#   書いて ✓ が出た（画面は何も変わらない）。『合計』は**対象**であって操作ではない。
# ★ 道具の側にも穴があった: 数値書式は**列にしか**掛けられなかった
#   ── 行と列の非対称。今日これで 4 度目（列の「間に」／値の引き算の位置語／…）。
#
# 契約:
#   ① 依頼文が「数値の見せ方」を言っているかを機械が見る（『金額』単体では発火しない）
#   ② 掛ける先は機械が決める ── 列名が 1 つなら列、合計行が 1 つなら行
#   ③ 行に掛けた回は**数値のセルだけ**（ラベルの『合計』には掛からない）
#   ★★ 恒真殺し: 「金額列の合計を出して」を書式に横取りしない

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HEADERS = ["取引先", "項目", "件数", "単価", "金額"]


def _book(tmp_path, with_total=True, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    ws.append(["丸和物流", "配送", 12, 4800, 57600])
    ws.append(["ヤマノ食品", "仕入", 28, 1500, 42000])
    if with_total:
        ws.cell(4, 1, "合計")
        ws.cell(4, 3, "=SUM(C2:C3)")
        ws.cell(4, 4, "=SUM(D2:D3)")
        ws.cell(4, 5, "=SUM(E2:E3)")
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(path)}


# --- ① 依頼文の見分け -------------------------------------------------------------------

@pytest.mark.parametrize("task,want", [
    ("合計を金額表示にして", True), ("金額列を桁区切りにして", True),
    ("単価を3桁区切りにして", True), ("合計行をカンマ区切りに", True),
    ("金額列の合計を出して", False),     # ★ 恒真殺し: 『金額』単体では発火しない
    ("合計を出して", False), ("金額で降順に並べ替えて", False),
])
def test_only_a_formatting_request_fires(task, want):
    assert ailine.task_asks_for_number_format(task) is want, task


# --- ② 掛ける先は機械が決める -----------------------------------------------------------

@pytest.mark.parametrize("task,want", [
    ("合計を金額表示にして", ("row", 4)),        # ★ 合計行
    ("合計行をカンマ区切りに", ("row", 4)),
    ("金額列を桁区切りにして", ("col", "金額")),
    ("単価を3桁区切りにして", ("col", "単価")),
])
def test_the_target_is_resolved_from_the_real_table(tmp_path, task, want):
    """★★ 「合計を**金額表示**にして」の『金額』を列名と読んでいた（部分文字列の穴・
       この repo で 3 度目）。操作の語を**先に取り除いてから**列名を探す。"""
    assert ailine.number_format_target(task, _meta(_book(tmp_path)), "請求") == want


def test_two_column_names_are_not_guessed(tmp_path):
    """★ 見出しが 2 つ出てくる依頼は決めない（推測で片方に掛けない）。"""
    assert ailine.number_format_target(
        "単価と金額をカンマ区切りに", _meta(_book(tmp_path)), "請求") is None


def test_no_total_row_means_no_row_target(tmp_path):
    assert ailine.number_format_target(
        "合計を金額表示にして", _meta(_book(tmp_path, with_total=False)), "請求") is None


# --- ③ 行に掛ける（数値だけ）------------------------------------------------------------

def test_a_row_target_is_accepted_without_a_column(tmp_path):
    ok, r, _i, err = ailine.verify_dsl_args(
        "NUMBER_FORMAT", {"row_number": 4, "style": "thousands"}, _meta(_book(tmp_path)))
    assert ok, err
    assert r["_row_index"] == 4 and "col" not in r


def test_the_header_row_is_refused(tmp_path):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "NUMBER_FORMAT", {"row_number": 1, "style": "thousands"}, _meta(_book(tmp_path)))
    assert not ok and "見出し行" in err, err


def test_the_codegen_calls_the_row_helper(tmp_path):
    p = _book(tmp_path)
    ok, r, _i, err = ailine.verify_dsl_args(
        "NUMBER_FORMAT", {"row_number": 4, "style": "thousands"}, _meta(p))
    assert ok, err
    code = ailine.codegen_dsl("NUMBER_FORMAT", r, _meta(p))
    assert "FormatThousandsRow(oDoc, 3," in code, code


def test_the_row_helper_only_touches_numbers():
    """★ ③: ラベルの『合計』に桁区切りを掛けない ── Basic 側の契約。"""
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    i = bas.index("Sub FormatThousandsRow(")
    seg = bas[i:i + 1200]
    assert "CellContentType.VALUE" in seg and "CellContentType.FORMULA" in seg, seg[:400]


def test_the_check_looks_at_the_numeric_cells_of_that_row(tmp_path):
    p = _book(tmp_path, name="out.xlsx")
    wb = openpyxl.load_workbook(p)
    for c in (3, 4, 5):
        wb["請求"].cell(4, c).number_format = "#,##0"
    wb.save(p)
    status, why = ailine.check_number_format(
        p, {"_row_index": 4, "_target_sheet": "請求", "style": "thousands"})
    assert status == "pass", why


def test_a_row_missing_the_format_fails(tmp_path):
    p = _book(tmp_path, name="out2.xlsx")
    wb = openpyxl.load_workbook(p)
    wb["請求"].cell(4, 3).number_format = "#,##0"      # 1 箇所だけ
    wb.save(p)
    status, why = ailine.check_number_format(
        p, {"_row_index": 4, "_target_sheet": "請求", "style": "thousands"})
    assert status == "fail" and "付いていません" in why, why


def test_the_column_path_is_unchanged(tmp_path):
    """★ 黙りすぎない側の対: 列に掛ける今までの道は変えていない。"""
    p = _book(tmp_path, with_total=False, name="out3.xlsx")
    wb = openpyxl.load_workbook(p)
    for r in (2, 3):
        wb["請求"].cell(r, 5).number_format = "#,##0"
    wb.save(p)
    status, why = ailine.check_number_format(
        p, {"col": "金額", "_target_sheet": "請求", "style": "thousands"})
    assert status == "pass", why
