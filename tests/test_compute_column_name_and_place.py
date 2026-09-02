# 計算列の「名前」と「置き場所」の番人（2026-09-02）。
#
# ★★ A（在庫「計算列の置き場所」）を確かめにいって分かったこと:
#   **置き場所は既に動いていた**（横断層 resolve_new_column_placement が 2026-08-27 に
#   入っていた）。README のほうが古かった。
#   代わりに**別の穴**が見つかった ── 依頼が「売上から原価を引いた**利益**の列を作って」
#   なのに、見出しが『売上-原価』（式そのもの）になっていた。
#   ★ 名前の抽出（new_column_name_from_task）は 2 つの理由で届いていなかった:
#     ① **位置の言い回しが在る時しか**働かない作りだった（「〜の右に」の後ろを見る）
#     ② 2 項の演算の枝に**配線されていなかった**（倍率＝税込/税抜の枝にしか無い）
#   どちらも A' 原則（値も名前も依頼文から取る・機械が発明しない）が抜けた形。
#
# 契約:
#   ① 位置を言わない依頼でも、依頼文の名前を拾う
#   ② 名前を言っていない依頼では**拾わない**（発明しない）
#   ③ 既にある見出しと同じなら拾わない（新しい列ではない）
#   ④ 置き場所は「作ってから動かす」── 式は LibreOffice が付け替える

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine                                   # noqa: E402

HEADS = ["商品", "売上", "原価"]


@pytest.mark.parametrize("task,want", [
    # 位置を言う依頼（既定のモード・従来どおり）
    ("金額の右に税込み金額を追加", "税込み金額"),
    # ★ 「〜の間に」は _re_after_position に無い（位置の解決は resolve_col_anchor が
    #   別に持っている）ので、既定のモードでは名前を拾えない ── 実測で確認。
    ("売上と原価の間に、売上から原価を引いた利益の列を作って", None),
    # ★ 位置を言わない依頼は、既定では**拾わない**
    ("売上から原価を引いた利益の列を作って", None),
    ("粗利の列を作って", None),
])
def test_the_default_mode_needs_a_position_phrase(task, want):
    """★ 既定は従来どおり（位置語が在る時だけ）。

    ★★ 2026-09-02 に一度、既定を「位置語なしでも拾う」に変えて**やりすぎた**:
      税の枝（W10c で設計）まで書き換わり、「税込みの列を追加して」の見出しが
      『税込み』になった ── 『税込み』は名前ではなく**修飾語**で、
      機械が組む『税込金額』のほうが良い。**測っていない所まで直しを広げない。**
    """
    assert ailine.new_column_name_from_task(task, HEADS) == want


@pytest.mark.parametrize("task,want", [
    # ① 2 項の演算（ここが欠けていた）
    ("売上から原価を引いた利益の列を作って", "利益"),
    ("粗利の列を作って", "粗利"),
    ("金額を1.1倍した税込金額の列を作って", "税込金額"),
    ("売上と原価の間に、売上から原価を引いた利益の列を作って", "利益"),
    # ② 名前を言っていないなら拾わない（**発明しない**）
    ("売上から原価を引いた列を作って", None),
    ("利益を計算して", None),
    ("並べ替えて", None),
    # ★ 「作る」と言っていない依頼を入口にしない（実測で踏んだ誤爆）
    ("A行G列を「税込み金額」に上書き", None),
    # ③ 既にある見出しは「新しい列の名前」ではない
    ("売上の列を作って", None),
])
def test_the_name_comes_from_the_request(task, want):
    """① 位置語が無くても拾う（2 項の演算の枝が明示的に呼ぶモード）。"""
    assert ailine.new_column_name_from_task(
        task, HEADS, require_position=False) == want


def test_a_clause_ending_does_not_swallow_the_name():
    """★ 初版が踏んだ形: 語尾を先に切らずに走査すると「作**って**」自身が
       修飾節の終わりに当たり、**名前ごと飲み込んで空になる**。"""
    assert ailine.new_column_name_from_task(
        "粗利の列を作って", HEADS, require_position=False) == "粗利"


def test_a_modifying_clause_is_trimmed():
    """★ 逆側 ── 節を切らないと「引いた利益」が丸ごと名前になる。"""
    got = ailine.new_column_name_from_task(
        "売上から原価を引いた利益の列を作って", HEADS, require_position=False)
    assert got == "利益", f"修飾節が残っている: {got!r}"


def test_placement_still_resolves_for_a_compute_column(tmp_path):
    """④ 置き場所（横断層）が計算列にも効いていること。"""
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(HEADS)
    ws.append(["りんご", 1200, 700])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    place = ailine.resolve_new_column_placement(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"},
        meta, "売上と原価の間に、売上から原価を引いた利益の列を作って", "売上表")
    assert place and place["_move_new_col_to"] == 2, place
    assert "『売上』と『原価』の間" in place["_at_basis"], place


def test_no_placement_phrase_means_no_move(tmp_path):
    """★ 陰性対照 ── 位置を言っていないのに動かしたら、静かに違う場所へ入る。"""
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(HEADS)
    ws.append(["りんご", 1200, 700])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    assert ailine.resolve_new_column_placement(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"},
        meta, "売上から原価を引いた利益の列を作って", "売上表") is None


def _book(tmp_path: Path) -> Path:
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(HEADS)
    for n, u, c in [("りんご", 1200, 700), ("みかん", 800, 300)]:
        ws.append([n, u, c])
    wb.save(p)
    return p


def test_both_actually_reach_resolved_through_verify(tmp_path):
    """★★ 配線そのものを通す検体（2026-09-02 の変異試験で開いていた穴）。

      部品（new_column_name_from_task / resolve_new_column_placement）を直に叩く試験は
      在ったが、**本番が通る道**（verify_dsl_args）を通る試験が無かった。
      そのため「配線を外す」変異が**緑のまま**だった ── 在っても鳴らない、そのもの。
    ★ ここでは resolved に両方が載ることだけを見る（実機は別の 1 本が見る）。
    """
    meta = ailine.build_book_meta(_book(tmp_path))
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, meta,
        task="売上と原価の間に、売上から原価を引いた利益の列を作って")
    assert ok, err
    assert resolved.get("_new_col_label") == "利益", resolved.get("_new_col_label")
    assert resolved.get("_move_new_col_to") == 2, resolved.get("_move_new_col_to")


def test_a_request_without_a_name_or_a_place_stays_bare(tmp_path):
    """★ 陰性対照 ── 言っていないものを載せない（発明しない・黙って動かさない）。"""
    meta = ailine.build_book_meta(_book(tmp_path))
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, meta,
        task="売上から原価を引いた列を作って")
    assert ok, err
    assert resolved.get("_new_col_label") is None, resolved.get("_new_col_label")
    assert resolved.get("_move_new_col_to") is None, resolved.get("_move_new_col_to")


# --- 依頼した名前が既に在る回 ── 「作る」ではなく「もう在る」-------------------------

def test_an_existing_name_becomes_the_target_not_a_new_column(tmp_path):
    """★★ 2026-09-02 に実測で捕まえた実害の番人。

      「売上から原価を引いた利益の列を作って」を **2 回**実行すると、
      1 回目は『利益』・2 回目は（『利益』が既に在るので名前として採られず）
      『売上-原価』になり、**見出しが違うので「見出しも値も同一の列」の関所が鳴らない**。
      値がそっくり同じ列が静かに 2 本目として増え、✓ まで出た
      ── 盲検 operator 査定が見つけた「不安でもう一度実行」の事故の再来。
    ★ 意味で考えても「作る」ではなく「**もう在る**」。その列を計算し直す依頼と読み、
      既存の**上書きの関所**に載せる（新しい関所も終了コードも作らない）。
    """
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(["商品", "売上", "原価", "利益"])
    ws.append(["りんご", 1200, 700, 999])
    ws.append(["みかん", 800, 300, 999])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, meta,
        task="売上から原価を引いた利益の列を作って")
    assert ok, err
    assert resolved.get("target") == "利益", (
        f"既に在る名前を『新しい列の名前』として捨てている: {resolved.get('target')!r} / "
        f"{resolved.get('_new_col_label')!r}")
    assert resolved.get("_new_col_label") is None, resolved.get("_new_col_label")


def test_a_new_name_is_still_a_new_column(tmp_path):
    """★ 陰性対照 ── 在らない名前まで既存列扱いしたら、新しい列が作れなくなる。"""
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(HEADS)
    ws.append(["りんご", 1200, 700])
    wb.save(p)
    meta = ailine.build_book_meta(p)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, meta,
        task="売上から原価を引いた利益の列を作って")
    assert ok, err
    assert resolved.get("target") is None, resolved.get("target")
    assert resolved.get("_new_col_label") == "利益", resolved.get("_new_col_label")


# --- 検算そのもの（列がずれても、頼んでいない書き込みは見逃さない）-------------------

def _sheet(path: Path, rows: list) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


def test_an_unrelated_change_is_caught_without_a_move(tmp_path):
    """★ 元の契約 ── 動かさない回で、他の列が変わったら言う。"""
    before = _sheet(tmp_path / "b.xlsx",
                     [["商品", "売上", "原価"], ["りんご", 1200, 700]])
    after = _sheet(tmp_path / "a.xlsx",
                    [["商品", "売上", "原価", "利益"], ["りんご", 1200, 999, 500]])
    got = ailine.only_this_column_changed(after, before, "売上表", 4, 1)
    assert got and "変わっています" in got, got


def test_the_shift_itself_is_not_reported(tmp_path):
    """★ 今回の直し ── 間に挿した回、右へずれた列を「書き換え」と言わない。"""
    before = _sheet(tmp_path / "b.xlsx",
                     [["商品", "売上", "原価"], ["りんご", 1200, 700]])
    after = _sheet(tmp_path / "a.xlsx",
                    [["商品", "売上", "利益", "原価"], ["りんご", 1200, 500, 700]])
    assert ailine.only_this_column_changed(
        after, before, "売上表", 3, 1, inserted_at=3) is None


def test_a_real_overwrite_is_still_caught_after_a_move(tmp_path):
    """★★ ここが芯（2026-09-02 の変異試験が開いていた穴）:
      ずれを許す実装にした以上、**ずれに紛れた本当の書き換え**を見逃さないこと。
      「常に許す」に緩めた変異が、この検体で赤くなる。
    """
    before = _sheet(tmp_path / "b.xlsx",
                     [["商品", "売上", "原価"], ["りんご", 1200, 700]])
    after = _sheet(tmp_path / "a.xlsx",                       # 売上が 1200→9999 に
                    [["商品", "売上", "利益", "原価"], ["りんご", 9999, 500, 700]])
    got = ailine.only_this_column_changed(
        after, before, "売上表", 3, 1, inserted_at=3)
    assert got and "変わっています" in got, f"ずれに紛れた書き換えを見逃した: {got!r}"


@pytest.mark.local
def test_the_name_and_the_place_both_land_on_real_libreoffice(tmp_path):
    """★ 実機 ── 見出しが『利益』で、間に入り、式が付け替わっていること。

    ★ 3 つを 1 本で見る: 名前・位置・式。どれか 1 つでも欠けたら赤。
    """
    import subprocess
    p = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上表"
    ws.append(HEADS)
    for n, u, c in [("りんご", 1200, 700), ("みかん", 800, 300)]:
        ws.append([n, u, c])
    wb.save(p)
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p),
         "売上と原価の間に、売上から原価を引いた利益の列を作って", "--copy"],
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=900,
        env={**__import__("os").environ, "PYTHONPATH": str(REPO / "src")})
    assert r.returncode == 0, r.stdout[-900:]
    ws2 = openpyxl.load_workbook(p.with_name(p.stem + ".out.xlsx"))["売上表"]
    assert [ws2.cell(1, c).value for c in range(1, 5)] == ["商品", "売上", "利益", "原価"]
    assert ws2.cell(2, 3).value == "=B2-D2", ws2.cell(2, 3).value
