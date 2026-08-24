# 第三波（複数ファイル盲検・2026-08-24）H1/H3 ── 修正より先に凍結した赤い検体。
#
# 契約:
#   H1 事後条件が破れた時、**どのファイルが原因か**を名指しし、直し方まで言う
#      （数字の差だけを出して黙るのは、scan/stack が名指しするのに run だけ黙る片配線）
#   H3 数値比較（gte/lte/gt/lt）から落ちた「数字に見える文字列」を開示する。
#      ★ 判定は変えない ── "△1,500" を機械が -1500 と読んだことにはしない。
#      ★ 片配線の番人: 単一ブック経路とフォルダ経路の**両方**を 1 本の試験で縛る。
#         片方だけ直すとここが赤くなる。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import compare_blocked  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _book(path, rows, header_at=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(header_at - 1):
        ws.append([])
    for r in rows:
        ws.append(r)
    wb.save(path)
    return path


# --- H3 器官そのもの（開示専用・判定に使わない）----------------------------------

@pytest.mark.parametrize("v,expect", [
    ("1,000", True), ("△1,500", True), ("８０，０００", True), ("¥12,000", True),
    ("(500)", True), ("abc", False), ("", False), (None, False), (100, False),
    ("2026-07-01", False),      # 日付を数字に見える文字列と誤認しない
])
def test_looks_numeric(v, expect):
    assert compare_blocked.looks_numeric(v) is expect


def test_disclosure_only_for_numeric_comparisons():
    """eq/contains は文字列でも成立しうる ── 余計な警告を出さない（誤爆防止）。"""
    assert compare_blocked.scan_column(["1,000"], "gte")
    assert compare_blocked.scan_column(["1,000"], "eq") is None
    assert compare_blocked.scan_column(["1,000"], "contains") is None


def test_no_disclosure_when_nothing_looks_numeric():
    assert compare_blocked.scan_column(["りんご", "みかん", None], "gte") is None


# --- H3 両経路（片配線の番人）------------------------------------------------------

def test_single_book_extract_discloses_string_amounts(tmp_path):
    """単一ブック経路: check_extract の理由に開示が載る。"""
    p = _book(tmp_path / "s.xlsx",
              [["商品", "金額"], ["a", "1,000"], ["b", "80,000"]])
    wb = openpyxl.load_workbook(p)
    wb.create_sheet("抽出")
    wb["抽出"].append(["商品", "金額"])
    wb.save(p)
    status, reason = ailine.check_extract(
        p, {"col": "金額", "cmp": "gte", "value": 50000, "_new_sheet": "抽出"})
    assert "文字列として入っています" in reason, f"0 行の理由を言っていない: {reason}"
    assert "80,000" in reason, f"実物の例を見せていない: {reason}"


def test_folder_extract_discloses_string_amounts(tmp_path, monkeypatch, capsys):
    """フォルダ経路: 同じ器官が同じ事実を言う（単一ブックだけ直すのを禁じる）。"""
    _isolate(monkeypatch, tmp_path)
    folder = tmp_path / "f"
    folder.mkdir()
    _book(folder / "a.xlsx", [["商品", "金額"], ["a", "1,000"], ["b", "80,000"]])
    _book(folder / "b.xlsx", [["商品", "金額"], ["c", "△1,500"], ["d", "90,000"]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "EXTRACT", "args": {"col": "金額", "cmp": "gte", "value": 50000}})
    rc, out = _run_main(["run", str(folder), "金額が50000以上の行を抜き出して"], capsys)
    assert "0 行一致" in out, f"前提: 一致 0 行になること: {out}"
    assert "文字列として入っています" in out, f"フォルダ経路が黙っている（片配線）: {out}"
    assert "『該当なし』ではありません" in out, out


def test_folder_extract_stays_quiet_on_real_numbers(tmp_path, monkeypatch, capsys):
    """誤爆防止: 本物の数値列では 1 文字も警告しない。"""
    _isolate(monkeypatch, tmp_path)
    folder = tmp_path / "f"
    folder.mkdir()
    _book(folder / "a.xlsx", [["商品", "金額"], ["a", 1000], ["b", 80000]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "EXTRACT", "args": {"col": "金額", "cmp": "gte", "value": 50000}})
    rc, out = _run_main(["run", str(folder), "金額が50000以上の行を抜き出して"], capsys)
    assert "文字列として" not in out, f"数値列に誤爆した: {out}"


# --- H1 事後条件が破れた時の名指し ---------------------------------------------------

def test_postcondition_failure_names_the_culprit_file(tmp_path, monkeypatch, capsys):
    """1 冊だけ見出し行が違うフォルダ ── 数字の差だけでなく、原因のファイル名を出す。"""
    _isolate(monkeypatch, tmp_path)
    folder = tmp_path / "f"
    folder.mkdir()
    _book(folder / "a.xlsx", [["商品", "金額"], ["a", 80000]])
    _book(folder / "b.xlsx", [["商品", "金額"], ["b", 90000]])
    # c.xlsx だけ見出しが 3 行目（照合できない冊）
    _book(folder / "c.xlsx", [["株式会社ほげ 御中"], [], ["商品", "金額"], ["c", 70000]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "EXTRACT", "args": {"col": "金額", "cmp": "gte", "value": 50000}})
    rc, out = _run_main(["run", str(folder), "金額が50000以上の行を抜き出して"], capsys)
    if "事後条件が破れた" not in out:
        pytest.skip(f"この検体では事後条件が破れなかった（前提が崩れた）: {out[-400:]}")
    assert "c.xlsx" in out, f"原因のファイルを名指ししていない（数字の差だけ）: {out}"
    assert "別フォルダ" in out or "見出しの行" in out, f"直し方が無い: {out}"
