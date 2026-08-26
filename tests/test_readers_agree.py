# データの出入口の盲検・致命2（2026-08-26）── repo 内に読み実装が 2 つあり、食い違っていた。
#
#   openpyxl      B2 = datetime.time(9, 0)
#   xml_readback  B2 = datetime.datetime(1899, 12, 30, 9, 0)
#
# export-csv は **xml_readback の読みで書き、同じ読みで検算する**ので、誤読のまま
# 「1 セルも変えずに書いた（欠落0・不一致0・余剰0）」と ✓ が出ていた。
# ★ 検算の分母が、疑うべき対象と同じ 1 回の読みから作られている ──
#   この週に 4 度出た形の 3 度目。
#
# ★ 番人の形: 値を直接 assert すると「その値が正しい」という俺の思い込みを凍結する。
#   代わりに **2 つの実装が一致すること**を測る ── どちらかが動いたら赤くなる。
#   （xml_readback は zip+XML 直読み、openpyxl は別ライブラリ ── 独立している）
#
# 契約:
#   ① 同じセルについて、2 つの読みが同じ型・同じ値を返す
#   ② 時刻・日付・日時・数値・文字列・真偽値の 6 種で成り立つ
#   ③ 片方だけ直しても赤くなる（恒真でない）

import datetime
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core import xml_readback  # noqa: E402


def _book(tmp_path):
    p = tmp_path / "kinds.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "勤怠"
    ws["A1"] = "氏名";   ws["B1"] = "出勤";  ws["C1"] = "日付"
    ws["D1"] = "日時";   ws["E1"] = "金額";  ws["F1"] = "在籍"
    ws["A2"] = "山田"
    ws["B2"] = datetime.time(9, 0);          ws["B2"].number_format = "h:mm"
    ws["C2"] = datetime.date(2026, 1, 2)
    ws["D2"] = datetime.datetime(2026, 1, 2, 18, 30)
    ws["E2"] = 1000
    ws["F2"] = True
    ws["B3"] = datetime.time(18, 30);        ws["B3"].number_format = "h:mm"
    ws["B4"] = datetime.time(0, 0);          ws["B4"].number_format = "h:mm"
    wb.save(p)
    return p


def _openpyxl_grid(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["勤怠"]
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                out[(c.row, c.column)] = c.value
    wb.close()
    return out


def _same_instant(a, b) -> bool:
    """2 つの読みが「同じものを指しているか」。

    ★ 唯一の正規化: **日付**と**その日の 0:00 の日時**は同じ日を指す、とみなす。
      openpyxl は日付セルにも datetime を返し、xml_readback は date を返す。
      どちらが正しいという問題ではなく、CSV へ書くときに
      `2026-01-02` と `2026-01-02T00:00:00` のどちらが人に親切か、という選択で、
      この道具は前者を選んでいる（csv_export._cell_for_csv）。
    ★ ここで緩めるのは**この 1 点だけ**。マイクロ秒のずれ・時刻と日時の取り違え・
      数値と文字列の取り違えは、すべて食い違いとして落とす
      （実際、この番人は 2026-08-26 に `18:30:00.000001` を掴んだ）。
    """
    if type(a) is type(b):
        return a == b
    for x, y in ((a, b), (b, a)):
        if (type(x) is datetime.date and isinstance(y, datetime.datetime)
                and y.time() == datetime.time(0, 0)):
            return x == y.date()
    return False


def test_the_two_readers_agree_cell_by_cell(tmp_path):
    """①② 型も値も一致すること。★ 片方の実装の値を正解として書かない。"""
    p = _book(tmp_path)
    ours = xml_readback.read_grid(p, sheet_name="勤怠")["grid"]
    theirs = _openpyxl_grid(p)
    assert set(ours) == set(theirs), (
        f"読めたセルの集合が違う: 片方だけ {sorted(set(ours) ^ set(theirs))}")
    disagree = [(rc, ours[rc], theirs[rc]) for rc in sorted(theirs)
                 if not _same_instant(ours[rc], theirs[rc])]
    assert not disagree, (
        "2 つの読み実装が食い違っている（この食い違いのまま書いて、同じ読みで検算すると"
        f"『1 セルも変えずに書いた』が恒真になる）: {disagree}")


def test_the_time_cell_is_a_time_in_both(tmp_path):
    """③ 恒真殺し: 実測した壊れ方（時刻が 1899 年の日時になる）を名指しで塞ぐ。"""
    p = _book(tmp_path)
    ours = xml_readback.read_grid(p, sheet_name="勤怠")["grid"]
    v = ours[(2, 2)]
    assert isinstance(v, datetime.time) and not isinstance(v, datetime.datetime), \
        f"時刻セルが時刻でない（実測: 1899-12-30T09:00:00）: {v!r}"
    assert 1899 not in (getattr(v, "year", 0),), v


def test_midnight_is_still_a_time_not_a_date(tmp_path):
    """境界: 0:00（シリアル 0）は日付でなく時刻。"""
    p = _book(tmp_path)
    ours = xml_readback.read_grid(p, sheet_name="勤怠")["grid"]
    assert isinstance(ours[(4, 2)], datetime.time), ours[(4, 2)]


def test_csv_writes_a_time_as_a_time(tmp_path):
    """出口まで通っていること（読みだけ直して書き手を忘れない）。"""
    from ailine_core import csv_export
    assert csv_export._cell_for_csv(datetime.time(9, 0)) == "09:00:00"
    assert csv_export._cell_for_csv(datetime.date(2026, 1, 2)) == "2026-01-02"
