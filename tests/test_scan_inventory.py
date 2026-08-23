"""M1読み `ailine scan <folder>`（棚卸し・書き込みゼロ）の検体。
   ★ 実装前に書いた赤い検体（DESIGN-20260821-multifile v2 §1-§3・specimen-first）。

   契約の要点（需要ノートの $0 条件と信用の条件から機械的に出る）:
   - 分母つき報告: 「N ファイル中 M 照合できた」。★ 分母そのものが検証対象（V7）
   - 基準ファイル方式: 見出しの推定は 1 回（パス辞書順先頭）。残りは基準との機械照合
   - 3 判定: 並び一致=取れた / 名前の多重集合一致・順序違い=取れた(並べ替えを開示) /
     それ以外=取れなかった(名指し+理由)。★ ゆるい「寄せ」はしない
   - 対象外（~$ 一時ファイル・サブフォルダ）は黙って無視せず 1 行開示
   - 読むだけ: 原本は 1 バイトも変わらない・新ファイルも作らない
   - 一部のファイルが読めなくても exit 0（報告が成果物。黙る失敗だけが罪）"""
import hashlib
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent


def _book(path, headers, rows=((1,),)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        padded = list(r) + [None] * (len(headers) - len(r))
        ws.append(padded[: len(headers)])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _scan(folder, *extra):
    return subprocess.run(
        [sys.executable, str(REPO / "ailine.py"), "scan", str(folder), *extra],
        capture_output=True, text=True, timeout=120, encoding="utf-8")


HDRS = ["注文ID", "取引先", "金額"]


def _standard_folder(tmp_path):
    """分母検体（V7）: 直下 .xlsx 3 + .xls 1 が対象。~$ 一時・サブフォルダ・txt は対象外。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS, rows=(("J-1", "甲", 100),))
    _book(folder / "b.xlsx", HDRS, rows=(("J-2", "乙", 200),))
    _book(folder / "c.xlsx", HDRS, rows=(("J-3", "丙", 300),))
    (folder / "old.xls").write_bytes(b"\xd0\xcf\x11\xe0dummy")   # 旧形式（読めない）
    (folder / "~$a.xlsx").write_bytes(b"lockfile")               # Excel の一時ファイル
    _book(folder / "sub" / "d.xlsx", HDRS)                       # サブフォルダ（見ない）
    (folder / "note.txt").write_text("memo", encoding="utf-8")   # Excel でない（黙って無視でよい）
    return folder


def test_denominator_counts_direct_excel_files_and_discloses_exclusions(tmp_path):
    """★ V7 本命: 分母 = 直下の .xlsx + .xls（~$ 除外）。対象外は開示する。"""
    folder = _standard_folder(tmp_path)
    p = _scan(folder)
    assert p.returncode == 0, p.stderr[-500:]
    assert "4 ファイル中" in p.stdout, f"分母が 4 でない:\n{p.stdout}"
    assert "一時ファイル" in p.stdout, "~$ の除外が開示されていない"
    assert "サブフォルダ" in p.stdout, "サブフォルダを見ていないことが開示されていない"


def test_unreadable_xls_is_named_with_reason_not_silently_skipped(tmp_path):
    """$0 条件「黙って失敗する」の正面: .xls は名指し + 理由。exit は 0 のまま。"""
    folder = _standard_folder(tmp_path)
    p = _scan(folder)
    assert p.returncode == 0
    assert "old.xls" in p.stdout, "読めないファイルが名指しされていない"
    assert "旧形式" in p.stdout or ".xls" in p.stdout.replace("old.xls", ""), \
        "理由（旧形式）が書かれていない"


def test_base_file_is_disclosed_and_identical_headers_all_match(tmp_path):
    """基準ファイル方式: 推定は 1 回（辞書順先頭 = a.xlsx）で、それを開示する。"""
    folder = _standard_folder(tmp_path)
    p = _scan(folder)
    assert p.returncode == 0
    assert "a.xlsx" in p.stdout and "基準" in p.stdout, "基準ファイルの開示が無い"
    assert "3 照合できた" in p.stdout or "3 冊" in p.stdout or "3 ファイル照合" in p.stdout, \
        f"照合できた数の分子が読めない:\n{p.stdout}"


def test_reordered_headers_match_with_disclosure(tmp_path):
    """列名の多重集合が一致・順序違い → 取れた + 並べ替えの事実を開示（黙って寄せない）。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS)
    _book(folder / "b.xlsx", ["金額", "注文ID", "取引先"])   # 同じ名前・並びだけ違う
    p = _scan(folder)
    assert p.returncode == 0
    assert "並べ替え" in p.stdout or "順序" in p.stdout, "順序違いの開示が無い"
    assert "取れなかった" not in p.stdout, "順序違いだけで不一致に落としている（厳しすぎ）"


def test_missing_column_is_named_file_and_column(tmp_path):
    """列が欠けたら 取れなかった: どのファイルのどの列かまで名指し（信用の条件②）。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS)
    _book(folder / "b.xlsx", ["注文ID", "取引先"])           # 金額が無い
    p = _scan(folder)
    assert p.returncode == 0
    assert "b.xlsx" in p.stdout and "金額" in p.stdout, \
        f"欠けたファイルと列の名指しが無い:\n{p.stdout}"


def test_scan_is_read_only_no_file_changed_or_created(tmp_path):
    """読むだけ: 全ファイルの sha256 不変・フォルダに新しいファイルを作らない。"""
    folder = _standard_folder(tmp_path)
    before = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
              for p in folder.rglob("*") if p.is_file()}
    r = _scan(folder)
    assert r.returncode == 0
    after = {p.name: hashlib.sha256(p.read_bytes()).hexdigest()
             for p in folder.rglob("*") if p.is_file()}
    assert before == after, "scan がファイルを変更または作成した"


def test_json_output_carries_denominator_and_per_file_status(tmp_path):
    """--json: 分母・基準・ファイルごとの status/reason が機械可読で出る（凍結対象のキー）。"""
    folder = _standard_folder(tmp_path)
    p = _scan(folder, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    assert data["denominator"] == 4
    assert data["base"] == "a.xlsx"
    statuses = {f["name"]: f["status"] for f in data["files"]}
    assert statuses["a.xlsx"] == "取れた"
    assert statuses["old.xls"] == "取れなかった"
    bad = [f for f in data["files"] if f["status"] == "取れなかった"]
    assert all(f.get("reason") for f in bad), "取れなかったに reason が無い"


def test_folder_with_no_readable_base_still_reports_and_exits_zero(tmp_path):
    """★ 実装後に実走で確かめて凍結した検体（implementer の ASSUMED を検体化）:
       読める .xlsx が 1 冊も無くても、基準なしの旨を開示して名指し報告・exit 0。"""
    folder = tmp_path / "books"
    folder.mkdir()
    (folder / "legacy.xls").write_bytes(b"\xd0\xcf\x11\xe0dummy")
    p = _scan(folder)
    assert p.returncode == 0, p.stderr[-500:]
    assert "基準" in p.stdout and ("見つかりません" in p.stdout or "無い" in p.stdout)
    assert "legacy.xls" in p.stdout and "旧形式" in p.stdout


def test_json_reports_total_row_candidates_per_file(tmp_path):
    """単位L の配線（分布測定の口）: --json の各ファイルに total_row_candidates が載る。
       ★ 人間向け出力は変えない（⚠ の総量を増やさない ── D6）。JSON だけの口。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS, rows=(("J-1", "甲", 100), ("J-2", "乙", 200), ("合計", None, 300)))
    _book(folder / "b.xlsx", HDRS, rows=(("J-3", "丙", 50),))
    p = _scan(folder, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    counts = {f["name"]: f.get("total_row_candidates") for f in data["files"]}
    assert counts["a.xlsx"] == 1, f"合計行 1 本が数えられていない: {counts}"
    assert counts["b.xlsx"] == 0


def test_total_row_candidates_respect_reordered_columns(tmp_path):
    """★ implementer 申告の設計の穴を検体化（2026-08-21）: 並べ替えファイルでは
       基準の列位置でなく 列名 で（ラベル列=基準1列目の名前・数値列=基準の数値列の名前）
       引き当てないと、別の列を数えて候補を取り逃がす。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS, rows=(("J-1", "甲", 100),))
    # 並びだけ違うファイルに合計行: ラベル『合計』は 注文ID 列（このファイルでは 2 列目）に居る
    _book(folder / "b.xlsx", ["金額", "注文ID", "取引先"],
          rows=((300, "J-3", "丙"), (300, "合計", None)))
    p = _scan(folder, "--json")
    assert p.returncode == 0, p.stderr[-500:]
    data = json.loads(p.stdout)
    counts = {f["name"]: f.get("total_row_candidates") for f in data["files"]}
    assert counts["b.xlsx"] == 1, f"並べ替えファイルの合計行を取り逃がした: {counts}"


def test_all_matched_files_each_get_their_own_line(tmp_path):
    """★ operator 盲検9回目 CONFUSING①(自分の検体): 全員照合できても「N ファイル中 M
       照合できた」の1行だけで終わらない ── README の「列は揃っているかを分母つきで
       報告」の約束どおり、各ファイル名が1行ずつ名指しされる。"""
    folder = tmp_path / "books"
    _book(folder / "a.xlsx", HDRS, rows=(("J-1", "甲", 100),))
    _book(folder / "b.xlsx", HDRS, rows=(("J-2", "乙", 200),))
    _book(folder / "c.xlsx", HDRS, rows=(("J-3", "丙", 300),))
    p = _scan(folder)
    assert p.returncode == 0, p.stderr[-500:]
    assert "3 ファイル中 3 照合できた" in p.stdout
    for name in ("a.xlsx", "b.xlsx", "c.xlsx"):
        assert f"{name}: 取れた" in p.stdout, \
            f"全員照合できた場合にファイル名が名指しされていない: {name}\n{p.stdout}"


def test_scan_discloses_sheet_fallback_to_first_sheet(tmp_path):
    """★ P2（architect 致命5 前段）: 基準名のシートが無いファイルは1枚目へ落ちる。
       scan の人間向け報告・--json のどちらにもその事実を開示すること。"""
    folder = tmp_path / "books"
    base = folder / "a.xlsx"
    _book(base, HDRS, rows=(("J-1", "甲", 100),))
    wb = openpyxl.load_workbook(base)
    wb.active.title = "明細"
    wb.save(base)
    b = folder / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "集計"
    ws.append(HDRS)
    ws.append(["J-2", "乙", 200])
    wb.save(b)

    p = _scan(folder)
    assert p.returncode == 0, p.stdout
    assert "明細" in p.stdout and "集計" in p.stdout and "b.xlsx" in p.stdout, \
        f"scan にシートのフォールバック開示が無い:\n{p.stdout}"

    p2 = _scan(folder, "--json")
    assert p2.returncode == 0, p2.stdout
    data = json.loads(p2.stdout)
    entry = next(f for f in data["files"] if f["name"] == "b.xlsx")
    fb = entry.get("sheet_fallback")
    assert fb and fb["wanted"] == "明細" and fb["used"] == "集計", \
        f"scan --json にシートのフォールバック開示が無い: {entry}"
