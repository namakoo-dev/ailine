"""★ 挙動変更#3（対象シートの取り違え）の番人。

挙動変更#2 が「依頼文からシート名を機械照合する」ようにした副作用を実測で見つけた:

    sheets=['売上データ','金額'] + 「金額を降順に並べ替えて」→ 2枚目『金額』が対象になる
    sheets=['明細','合計']      + 「合計行を追加して」      → 2枚目『合計』が対象になる

どちらも**列を指したつもりの依頼**。告知は出るので沈黙ではないが、対象シートにも同名列が
あれば**間違ったシートを並べ替えて「成功」してしまう**（エラーで止まるより後で気づきにくい）。

ここで守るのは4つ:
  1. 裸の言及 × その語が実在の列見出しでもある → 既定(1枚目)へ後退し、告知が確定文言どおり
  2. 明示（「金額シートを」「2枚目を」「タブ」）と --sheet は今までどおり効く
  3. **TTY でない時（--json・パイプ・CI）は絶対に入力を待たない**（止めると動いていた
     スクリプトが黙って壊れる）。ここは「input が呼ばれたら即失敗」で機械検証する
  4. 3択の各分岐（1/2/3、②の後の y/N）が決定論的に動く（ollama 不要・stdin は注入）

★ ollama も LibreOffice も使わない（translate_task / basrun_apply は monkeypatch）。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine  # noqa: E402
from ailine_core import ask_choice as ask_choice_mod  # noqa: E402
from ailine_core.target_sheet import (  # noqa: E402
    SheetNameConflict, format_sheet_field, op_verbs, resolve_target_sheet,
    sheet_conflict_choice_lines,
)

from _run_argv import run_argv  # noqa: E402


def _isolate(monkeypatch, tmp_path):
    """本番のユーザーディレクトリ(~/.ailine 等)に一切触れさせない。"""
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", tmp_path / "run.lock")


def _sales_book(tmp_path):
    """★ 再現ブック: 2枚目のシート名『金額』が、1枚目の列名『金額』と衝突する。
       どちらのシートにも『金額』列があるので、②「もう一方を見てみる」も実際に実行できる。"""
    p = tmp_path / "売上.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "売上データ"
    ws1.append(["商品", "金額"])
    for row in [["a", 200], ["b", 300], ["c", 100]]:
        ws1.append(row)
    ws2 = wb.create_sheet("金額")
    ws2.append(["月", "金額"])
    for row in [["1月", 50], ["2月", 80]]:
        ws2.append(row)
    wb.save(p)
    return p


def _detail_book(tmp_path):
    """★ 再現ブック その2: 2枚目のシート名『合計』が、1枚目の列名『合計』と衝突する。"""
    p = tmp_path / "明細.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "明細"
    ws1.append(["商品", "金額", "合計"])
    ws1.append(["a", 100, 100])
    ws2 = wb.create_sheet("合計")
    ws2.append(["区分", "合計"])
    ws2.append(["計", 100])
    wb.save(p)
    return p


def _single_sheet_book(tmp_path):
    p = tmp_path / "単票.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["商品", "金額"])
    for row in [["a", 200], ["b", 300], ["c", 100]]:
        ws.append(row)
    wb.save(p)
    return p


def _sort_desc_apply(out_book, code, workdir, helper_files=(), timeout=None):
    """basrun/LibreOffice の代役。★ 生成された .bas に moveByName("金額", 0)（＝対象シートを
       一時的に先頭へ動かすラップ）が入っているかどうかで、どちらのシートを触るつもりだった
       かを見分ける ── 「どちらのシートが実際に並べ替えられたか」を成果物側でも検証できる。"""
    wb = openpyxl.load_workbook(out_book)
    name = "金額" if 'moveByName("金額"' in code else wb.sheetnames[0]
    ws = wb[name]
    rows = sorted(([ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value]
                   for r in range(2, ws.max_row + 1)),
                  key=lambda pair: pair[1], reverse=True)
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
    wb.save(out_book)
    return True, None, "ok"


def _fixed_sort(monkeypatch, col="金額"):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": col, "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply", _sort_desc_apply)


def _tty(monkeypatch, value: bool):
    monkeypatch.setattr(ailine, "_stdin_isatty", lambda: value)


def _scripted_input(monkeypatch, answers):
    """順に answers を返す input。使い切ったら EOFError（無限待ちの代わり）。"""
    queue = list(answers)
    asked = []

    def fake_input(prompt=""):
        asked.append(prompt)
        if not queue:
            raise EOFError
        return queue.pop(0)

    monkeypatch.setattr("builtins.input", fake_input)
    return asked


def _never_input(monkeypatch):
    """★ DoD5 の機械検証: 呼ばれたら即失敗する input（「待たない」ことを証明する）。"""
    def boom(prompt=""):
        raise AssertionError(f"非対話のはずなのに入力を待った: {prompt!r}")
    monkeypatch.setattr("builtins.input", boom)


# ===========================================================================
# ① 衝突検出そのもの（resolve_target_sheet・純関数レベル）
# ===========================================================================

SALES_HEADERS = {"売上データ": ["商品", "金額"], "金額": ["月", "金額"]}
DETAIL_HEADERS = {"明細": ["商品", "金額", "合計"], "合計": ["区分", "合計"]}


def test_bare_mention_that_is_also_a_column_name_falls_back_to_default_sales():
    """★ DoD1（1例目）: 「金額を降順に並べ替えて」は『金額』シートを指したとは限らない。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "金額を降順に並べ替えて", ["売上データ", "金額"], None, headers=SALES_HEADERS)
    assert (sheet, source, err) == ("売上データ", "default", None)
    assert conflict == SheetNameConflict(word="金額", alternative="金額", chosen="売上データ")


def test_bare_mention_that_is_also_a_column_name_falls_back_to_default_detail():
    """★ DoD1（2例目）: 「合計行を追加して」も同型（『合計』シートへ行を足すのが正とは限らない）。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "合計行を追加して", ["明細", "合計"], None, headers=DETAIL_HEADERS)
    assert (sheet, source, err) == ("明細", "default", None)
    assert conflict == SheetNameConflict(word="合計", alternative="合計", chosen="明細")


@pytest.mark.parametrize("task", [
    "金額シートを降順に並べ替えて",
    "金額タブを降順に並べ替えて",
    "『金額』シートを降順に並べ替えて",
])
def test_explicit_sheet_marker_is_adopted_unconditionally(task):
    """★ DoD2: 「〜シート」「〜タブ」の明示マーカー付きは衝突チェックを免除して採用する。"""
    sheet, source, err, conflict = resolve_target_sheet(
        task, ["売上データ", "金額"], None, headers=SALES_HEADERS)
    assert (sheet, source, err, conflict) == ("金額", "task", None, None)


def test_ordinal_phrase_is_adopted_unconditionally():
    """★ DoD2: 「2枚目を」の序数表現も明示なので衝突チェックの対象外。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "2枚目を降順に並べ替えて", ["売上データ", "金額"], None, headers=SALES_HEADERS)
    assert (sheet, source, err, conflict) == ("金額", "task", None, None)


def test_cli_sheet_always_wins_even_when_the_word_conflicts():
    """★ DoD3: --sheet は衝突チェックと無関係に常に最優先。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "金額を降順に並べ替えて", ["売上データ", "金額"], "金額", headers=SALES_HEADERS)
    assert (sheet, source, err, conflict) == ("金額", "cli", None, None)


def test_without_headers_behaviour_is_unchanged_from_change2():
    """★ 列見出しが取れない場合は挙動変更#2 のまま（退行させない）。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "金額を降順に並べ替えて", ["売上データ", "金額"], None, headers=None)
    assert (sheet, source, err, conflict) == ("金額", "task", None, None)


def test_word_that_is_not_a_column_anywhere_is_not_a_conflict():
    """列名でない語（普通のシート名の言及）は今までどおり採用する（過検知しない）。"""
    sheet, source, err, conflict = resolve_target_sheet(
        "工事台帳の金額を降順に並べ替えて", ["請求書", "工事台帳"],
        None, headers={"請求書": ["宛先", "金額"], "工事台帳": ["取引先名", "金額"]})
    assert (sheet, source, err, conflict) == ("工事台帳", "task", None, None)


# ===========================================================================
# ② 告知文（確定文言）
# ===========================================================================

def test_announcement_wording_on_conflict_fallback(tmp_path, monkeypatch, capsys):
    """★ DoD1: 衝突で後退したときの告知は「操作するシート: 1枚目『売上データ』（このブックは2シート）」。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, False)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_sales_book(tmp_path)),
                              task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "操作するシート: 1枚目『売上データ』（このブックは2シート）" in out


def test_announcement_wording_on_conflict_fallback_detail(tmp_path, monkeypatch, capsys):
    """★ DoD1（2例目）: 『合計』でも同じ形で1枚目へ後退したことが告知される。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "APPEND_TOTAL", "args": {"col": "金額"}})
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    _tty(monkeypatch, False)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_detail_book(tmp_path)),
                              task="合計行を追加して", copy=True, dry=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "操作するシート: 1枚目『明細』（このブックは2シート）" in out


def test_announcement_wording_for_explicit_mention_and_cli(tmp_path, monkeypatch, capsys):
    """★ DoD2/3: 明示指定の2種の告知（確定文言）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, False)
    _never_input(monkeypatch)
    book = _sales_book(tmp_path)
    ailine.main(run_argv(book=str(book), task="金額シートを降順に並べ替えて", copy=True, dry=True))
    assert ("操作するシート: 2枚目『金額』（依頼文から判断・このブックは2シート）"
            in capsys.readouterr().out)
    ailine.main(run_argv(book=str(book), task="降順に並べ替えて", sheet="金額", copy=True, dry=True))
    assert "操作するシート: 2枚目『金額』（--sheet 指定）" in capsys.readouterr().out


# ===========================================================================
# ③ 解釈行のシート欄
# ===========================================================================

def test_confirmation_line_carries_the_sheet_on_multi_sheet_books():
    """★ 確定文言: 解釈: シート:『売上データ』(1枚目) 操作:並べ替え 対象:金額 順:降順"""
    line = ailine.format_confirmation_line(
        "SORT", {"col": "金額", "order": "desc"}, set(),
        sheets=["売上データ", "金額"], target_sheet="売上データ")
    assert line == "解釈: シート:『売上データ』(1枚目) 操作:並べ替え 対象:金額 順:降順"


def test_confirmation_line_is_unchanged_on_single_sheet_books():
    """★ DoD4: 1枚のブックでは出さない（既存ゴールデン f5_confirmation は全検体無変化）。"""
    assert ailine.format_confirmation_line("SORT", {"col": "金額", "order": "desc"}, set(),
                                           sheets=["Sheet"], target_sheet="Sheet") \
        == "解釈: 操作:並べ替え 対象:金額 順:降順"
    assert ailine.format_confirmation_line("SORT", {"col": "金額", "order": "desc"}, set()) \
        == "解釈: 操作:並べ替え 対象:金額 順:降順"


def test_lookup_fill_does_not_repeat_the_sheet_twice():
    """★ 転記の既存「対象シート:」欄は、先頭のシート欄と同じ値なので省く（参照シートは残す）。"""
    resolved = {"target_sheet": "明細", "target_col": "単価",
                "source_sheet": "単価表", "key_col": "商品"}
    line = ailine.format_confirmation_line("LOOKUP_FILL", dict(resolved), set(),
                                           sheets=["明細", "単価表"], target_sheet="明細")
    assert line == "解釈: シート:『明細』(1枚目) 操作:転記 対象列:単価 参照シート:単価表 キー列:商品"
    # 1枚のブック（sheets 未指定）では従来どおり「対象シート:」を出す。
    assert ailine.format_confirmation_line("LOOKUP_FILL", dict(resolved), set()) \
        == "解釈: 操作:転記 対象シート:明細 対象列:単価 参照シート:単価表 キー列:商品"


def test_single_sheet_run_output_has_no_new_lines(tmp_path, monkeypatch, capsys):
    """★ DoD4: 1枚のブックの出力には、今回足した行が1つも現れない。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_single_sheet_book(tmp_path)),
                              task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    for forbidden in ("操作するシート:", "シート:『", "2通りに読めます"):
        assert forbidden not in out


# ===========================================================================
# ④ 「聞かない」ことの機械検証（DoD5）
# ===========================================================================

@pytest.mark.parametrize("flags,label", [
    (dict(json=True), "--json"),
    (dict(dry=True), "--dry"),
])
def test_never_asks_when_json_or_dry_even_on_a_tty(tmp_path, monkeypatch, capsys, flags, label):
    """★ --json / --dry では TTY でも聞かない（入力を待ったら _never_input が落とす）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_sales_book(tmp_path)),
                              task="金額を降順に並べ替えて", copy=True, **flags))
    out = capsys.readouterr().out
    assert rc == 0, f"{label}: {out}"
    assert "2通りに読めます" not in out
    assert "操作するシート: 1枚目『売上データ』（このブックは2シート）" in out


def test_never_asks_when_stdin_is_not_a_tty(tmp_path, monkeypatch, capsys):
    """★ DoD5 本命: パイプ/リダイレクト/CI（stdin が TTY でない）では絶対に待たず、
       既定(1枚目)で進む＝今までの挙動。止めると動いていたスクリプトが黙って壊れる。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, False)
    _never_input(monkeypatch)
    book = _sales_book(tmp_path)
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "2通りに読めます" not in out
    wb = openpyxl.load_workbook(book.with_name(book.stem + ".out" + book.suffix))
    assert [wb["売上データ"].cell(row=r, column=2).value for r in range(2, 5)] == [300, 200, 100]
    assert [wb["金額"].cell(row=r, column=2).value for r in range(2, 4)] == [50, 80]   # 触っていない


def test_no_conflict_means_no_question_even_on_a_tty(tmp_path, monkeypatch, capsys):
    """★ 「はっきり指してきた時は聞かない」（毎回聞くのは確認地獄）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_sales_book(tmp_path)),
                              task="金額シートを降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "2通りに読めます" not in out


# ===========================================================================
# ⑤ 3択の各分岐（DoD6）
# ===========================================================================

EXPECTED_QUESTION = [
    "依頼文の「金額」は2通りに読めます（『金額』という名前のシートもあるため）:",
    "  1) 『売上データ』シートの「金額」列を並べ替える ← 上の解釈のとおり実行する",
    "  2) 『金額』シートを並べ替えた場合を見てみる",
    "  3) やめる",
]


def test_choice_1_executes_the_declared_interpretation(tmp_path, monkeypatch, capsys):
    """①: 上の解釈のとおり（既定=1枚目）実行する。確定文言どおりの3択が出ることも見る。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    asked = _scripted_input(monkeypatch, ["1"])
    book = _sales_book(tmp_path)
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "解釈: シート:『売上データ』(1枚目) 操作:並べ替え 対象:金額 順:降順" in out
    for line in EXPECTED_QUESTION:
        assert line in out
    assert asked == ["> "]
    wb = openpyxl.load_workbook(book.with_name(book.stem + ".out" + book.suffix))
    assert [wb["売上データ"].cell(row=r, column=2).value for r in range(2, 5)] == [300, 200, 100]
    assert [wb["金額"].cell(row=r, column=2).value for r in range(2, 4)] == [50, 80]


def test_choice_2_then_yes_runs_on_the_alternative_sheet(tmp_path, monkeypatch, capsys):
    """②→y: もう一方のシートを対象に翻訳からやり直し、プレビューを見せてから実行する。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    asked = _scripted_input(monkeypatch, ["2", "y"])
    book = _sales_book(tmp_path)
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "解釈: シート:『金額』(2枚目) 操作:並べ替え 対象:金額 順:降順" in out
    assert asked == ["> ", "この内容で実行しますか？ [y/N]: "]
    # ★ プレビューでは「--dry を外して実行」の案内を出さない（直後に y/N を聞くため）。
    assert "--dry を外して実行" not in out
    wb = openpyxl.load_workbook(book.with_name(book.stem + ".out" + book.suffix))
    assert [wb["金額"].cell(row=r, column=2).value for r in range(2, 4)] == [80, 50]
    assert [wb["売上データ"].cell(row=r, column=2).value for r in range(2, 5)] == [200, 300, 100]


def test_choice_2_then_no_aborts_and_leaves_the_original_untouched(tmp_path, monkeypatch, capsys):
    """②→N: やめる。原本は無変更（--copy 無しの既定＝原本直接適用でも1バイトも変わらない）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _scripted_input(monkeypatch, ["2", "n"])
    book = _sales_book(tmp_path)
    before = book.read_bytes()
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて"))
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "× 中止した" in out
    assert book.read_bytes() == before


def test_choice_3_stops_without_doing_anything(tmp_path, monkeypatch, capsys):
    """③: やめる（既存の中止系と同じ exit code=1・原本は無変更）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _scripted_input(monkeypatch, ["3"])
    book = _sales_book(tmp_path)
    before = book.read_bytes()
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて"))
    out = capsys.readouterr().out
    assert rc == 1, out
    assert "× 中止した" in out
    assert book.read_bytes() == before


def test_eof_at_the_question_falls_back_to_the_default(tmp_path, monkeypatch, capsys):
    """聞いたが答えが得られない（EOF）場合も、止めずに既定で進む（今までの挙動）。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _scripted_input(monkeypatch, [])   # 最初の input で即 EOFError
    book = _sales_book(tmp_path)
    rc = ailine.main(run_argv(book=str(book), task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    wb = openpyxl.load_workbook(book.with_name(book.stem + ".out" + book.suffix))
    assert [wb["売上データ"].cell(row=r, column=2).value for r in range(2, 5)] == [300, 200, 100]


def test_the_question_is_asked_only_once(tmp_path, monkeypatch, capsys):
    """②のやり直しでもう一度3択が出たら無限に聞くことになる ── 1回だけであることを見る。"""
    _isolate(monkeypatch, tmp_path)
    _fixed_sort(monkeypatch)
    _tty(monkeypatch, True)
    _scripted_input(monkeypatch, ["2", "y"])
    rc = ailine.main(run_argv(book=str(_sales_book(tmp_path)),
                              task="金額を降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert out.count("2通りに読めます") == 1


def test_composite_plan_shows_the_sheet_but_never_asks(tmp_path, monkeypatch, capsys):
    """★ 複合計画: 各段の「解釈:」行にもシートが載る。ただし3択は**聞かない**
       （ASSUMED・報告参照: 途中の段で対象シートを選び直すと、直前までの段を適用済みの
       作業コピーの上で計画をやり直すことになり「原本にはまだ触れていない」前提が崩れる）。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "SORT", "args": {"col": "金額", "order": "desc"}}]})
    monkeypatch.setattr(ailine, "basrun_apply", _sort_desc_apply)
    _tty(monkeypatch, True)
    _never_input(monkeypatch)
    rc = ailine.main(run_argv(book=str(_sales_book(tmp_path)),
                              task="金額を降順に並べ替えて、もう一度金額で降順に並べ替えて", copy=True))
    out = capsys.readouterr().out
    assert rc == 0, out
    assert "1段目: 解釈: シート:『売上データ』(1枚目) 操作:並べ替え 対象:金額 順:降順" in out
    assert "2通りに読めます" not in out


# ===========================================================================
# ⑥ 部品そのもの（ailine_core.ask_choice・stdin を一切触らない）
# ===========================================================================

CHOICES = [ask_choice_mod.Choice("1", "そのまま"), ask_choice_mod.Choice("2", "別を見る"),
           ask_choice_mod.Choice("3", "やめる")]


def test_ask_choice_is_silent_and_returns_none_when_not_interactive():
    printed = []
    result = ask_choice_mod.ask_choice(["前置き"], CHOICES, interactive=False,
                                       input_fn=lambda p: "1", print_fn=printed.append)
    assert result == ask_choice_mod.ChoiceResult(key=None, asked=False)
    assert printed == []   # 非対話では1行も出さない（既存の出力を汚さない）


def test_ask_choice_renders_block_and_returns_the_key():
    printed = []
    result = ask_choice_mod.ask_choice(["", "前置き:"], CHOICES, interactive=True,
                                       input_fn=lambda p: " 2 ", print_fn=printed.append)
    assert result == ask_choice_mod.ChoiceResult(key="2", asked=True)
    assert printed == ["", "前置き:", "  1) そのまま", "  2) 別を見る", "  3) やめる"]


def test_ask_choice_reprompts_then_gives_up_without_hanging():
    printed, answers = [], ["x", "y", "z", "1"]
    result = ask_choice_mod.ask_choice(["q"], CHOICES, interactive=True,
                                       input_fn=lambda p: answers.pop(0), print_fn=printed.append)
    assert result == ask_choice_mod.ChoiceResult(key=None, asked=True)
    assert answers == ["1"]   # MAX_REPROMPTS 回でやめる（無限ループにしない）


def test_ask_choice_eof_is_not_an_exception():
    def eof(prompt):
        raise EOFError
    assert ask_choice_mod.ask_choice(["q"], CHOICES, interactive=True,
                                     input_fn=eof, print_fn=lambda _: None) \
        == ask_choice_mod.ChoiceResult(key=None, asked=True)


@pytest.mark.parametrize("kwargs,expected", [
    (dict(stdin_isatty=True), True),
    (dict(stdin_isatty=False), False),
    (dict(stdin_isatty=True, json_mode=True), False),
    (dict(stdin_isatty=True, dry=True), False),
])
def test_is_interactive_policy(kwargs, expected):
    assert ask_choice_mod.is_interactive(**kwargs) is expected


def test_ask_yes_no_only_accepts_y():
    for answer, expected in [("y", True), ("Y", True), ("yes", True),
                             ("n", False), ("", False), ("はい", False)]:
        assert ask_choice_mod.ask_yes_no("？", interactive=True,
                                         input_fn=lambda p: answer) is expected
    assert ask_choice_mod.ask_yes_no("？", interactive=False, input_fn=lambda p: "y") is False


def test_format_sheet_field_is_silent_for_single_sheet():
    assert format_sheet_field(["Sheet"], "Sheet") is None
    assert format_sheet_field([], None) is None
    assert format_sheet_field(["a", "b"], "存在しない") is None
    assert format_sheet_field(["a", "b"], "b") == "シート:『b』(2枚目)"


def test_op_verbs_falls_back_to_label_plus_suru():
    assert op_verbs("SORT", "並べ替え") == ("並べ替える", "並べ替えた")
    assert op_verbs("LOOKUP_FILL", "転記") == ("転記する", "転記した")
    assert op_verbs("UNKNOWN_OP", "なにか") == ("なにかする", "なにかした")


def test_sheet_conflict_choice_lines_matches_the_agreed_wording():
    lines, choices = sheet_conflict_choice_lines(
        SheetNameConflict(word="金額", alternative="金額", chosen="売上データ"), "SORT", "並べ替え")
    assert lines == ["", "依頼文の「金額」は2通りに読めます（『金額』という名前のシートもあるため）:"]
    assert choices == [
        ("1", "『売上データ』シートの「金額」列を並べ替える ← 上の解釈のとおり実行する"),
        ("2", "『金額』シートを並べ替えた場合を見てみる"),
        ("3", "やめる"),
    ]
