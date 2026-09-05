"""帳票の形（左余白・結合見出し・見出しの無い列）でも、列が読めること（2026-09-05）。

★★ 実測した壁: 弥生の請求書テンプレートに「金額の大きい順に並べ替えて」と頼むと
  **「列『金額』がありません。ある列: (無し)」**で止まった。見出し行の検出（16 行目）は
  正しく当たっていたのに、列が 1 つも読めていなかった。

★ 真因は `book_columns` の**1 行**だった ── 1 列目から数え、**最初の空欄で打ち切る**。
  帳票はこれを満たさない:
    ・表が B 列から始まる（左が余白）             → 0 列と読む
    ・見出しが結合セル（C:E で 1 つの名前）        → D で打ち切る
    ・単位列など見出しの無い列が途中に挟まる       → そこで打ち切る
  ★ 性質を 1 つずつ分けた検体（_fixtures/ladder）で 3/9 が 0〜1 列になった。

★ 直し方は 3 つ: ①結合は範囲の左上から名前を引く ②空欄で打ち切らず**位置を保ったまま**
  空名を詰める（`headers.index(名前)+1 = 列番号` に依存する箇所が 46 行あるため）
  ③表の外まで拾わないよう、空が続いたら終える。
★ ここは LibreOffice も LLM も要らない（openpyxl だけ）。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402


def _book(tmp_path, cells, merges=()):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "表"
    for ref, v in cells.items():
        ws[ref] = v
    for m in merges:
        ws.merge_cells(m)
    p = tmp_path / "b.xlsx"; wb.save(p)
    return p


def _cols(path, header_row=1):
    return ailine.book_columns(path, header_rows={"表": header_row})["表"]


def test_a_table_that_starts_at_column_b_is_read(tmp_path):
    """★ 左が余白 ── 帳票では普通の形。位置を保つため先頭に空名が入る。"""
    p = _book(tmp_path, {"B1": "品名", "C1": "数量", "D1": "単価", "B2": "机"})
    assert _cols(p) == ["", "品名", "数量", "単価"]


def test_a_gap_in_the_middle_does_not_end_the_table(tmp_path):
    """★ 見出しの無い列（単位列など）が挟まっても、その右を見失わない。"""
    p = _book(tmp_path, {"A1": "品名", "C1": "数量", "D1": "単価", "A2": "机"})
    assert _cols(p) == ["品名", "", "数量", "単価"]


def test_a_merged_header_names_every_column_it_spans(tmp_path):
    """★ 結合は「1 つの名前が複数列に跨る」という**構造の事実**。範囲の左上から引く。"""
    p = _book(tmp_path, {"A1": "品番・品名", "C1": "数量", "A2": "机"}, merges=("A1:B1",))
    assert _cols(p) == ["品番・品名", "品番・品名", "数量"]


def test_a_merged_block_with_an_empty_top_left_gets_no_name(tmp_path):
    """★ 名前が無いものに名前を作らない（実物の A2:A48 ＝ 左余白の結合）。"""
    p = _book(tmp_path, {"B1": "品名", "C1": "数量", "B2": "机"}, merges=("A1:A5",))
    assert _cols(p) == ["", "品名", "数量"]


def test_a_far_away_cell_does_not_become_a_column(tmp_path):
    """★ 表の外まで拾わない ── 空が続いたらそこが右端。"""
    p = _book(tmp_path, {"A1": "品名", "B1": "数量", "A2": "机", "Z1": "遠くの何か"})
    assert _cols(p) == ["品名", "数量"]


def test_trailing_padding_is_dropped(tmp_path):
    """★ 末尾の詰め物は残さない（列数が水増しされない）。"""
    p = _book(tmp_path, {"A1": "品名", "C1": "数量", "A2": "机"})
    assert _cols(p)[-1] != ""


def test_the_position_of_each_name_still_means_its_column(tmp_path):
    """★★ ここが崩れると 46 行が壊れる ── `headers.index(名前)+1` が列番号であること。"""
    p = _book(tmp_path, {"B1": "品名", "D1": "数量", "B2": "机"})
    cols = _cols(p)
    assert cols.index("品名") + 1 == 2, cols       # B 列
    assert cols.index("数量") + 1 == 4, cols       # D 列


def test_the_old_multi_row_header_rescue_still_works(tmp_path):
    """★ 退行の番人 ── 多段見出しで真上を遡る救済（W3）を壊していないこと。"""
    p = _book(tmp_path, {"A1": "商品名", "B2": "数量", "A3": "机"})
    assert _cols(p, header_row=2) == ["商品名", "数量"]


@pytest.mark.local
def test_the_real_invoice_template_becomes_readable():
    """★ 実物（弥生の請求書）── repo には置けないので、在る時だけ測る。"""
    real = Path(r"C:\Dev\_fixtures\misoca_invoice_blackline.xlsx")
    if not real.exists():
        pytest.skip("実物の検体が無い（_fixtures は公開 repo に入れない）")
    cols = ailine.book_columns(real, header_rows={"misoca_invoice": 16})["misoca_invoice"]
    for want in ("日付", "品番・品名", "軽減税率", "数量", "単価（税別）", "税率", "金額"):
        assert want in cols, f"{want} が読めていない: {cols}"
