"""条件つき書換の門は、比較語でなく**実表**に聞くこと（2026-09-05）。

★★ 実測した穴（効果の行列で失敗 2 件として分母に残していたもの）:

    「所属が営業の行のメモに『○』を付けて」  → **行追加に化けて空行が挿さる**

  一段目が ADD_ROW を返し、読み直しの門が開かなかったため。exit 0 で ⚠ は出るが、
  結果は間違い ── **誰でも書く形なのに落ちていた。**

★ 門の鍵が **比較語**（以上・以下・と等しい・を含む…）になっていた。素の
  「が営業の」には比較語が無いので、門は永久に閉じたままだった。
  ★ 当時「〜以外」を条件つき書換で使えないのも、同じ門が原因。

★ 直し方は 2026-09-05 に削除の軸でやったのと同じ ── **実表に聞く**。
  依頼文が実表に在る値をちょうど 1 列で名指しし、書き込む値が引用で囲まれていれば
  条件つき書換。★ 値の照合は task_names_real_values（抽出・軸判定と同じ関数）。

★ 本物の行追加を横取りしない線（既存の材料だけで引く）:
    ① 行追加は書き込む値を引用符で囲まない
    ② 位置の語（の下に・の上に…）が在れば配置の依頼 ── _ANCHOR_* をそのまま使う
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402


@pytest.fixture
def roster(tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "名簿"
    ws.append(["氏名", "所属", "内線", "メモ"])
    for r in [["山田", "営業", 101, None], ["鈴木", "経理", 202, None],
              ["高橋", "総務", 305, None]]:
        ws.append(r)
    p = tmp_path / "b.xlsx"; wb.save(p)
    return ailine.build_book_meta(p)


@pytest.mark.parametrize("task", [
    "所属が営業の行のメモに「○」を付けて",
    "営業の行のメモを「○」にして",
])
def test_a_condition_without_a_comparison_word_opens_the_gate(roster, task):
    """★ 本命 ── 比較語が無くても、実表に在る値を名指ししていれば開く。"""
    assert ailine.task_asks_for_a_conditional_write(task, roster, "名簿"), task


def test_a_comparison_word_still_opens_it(roster):
    """★ 退行の番人 ── 従来の鍵（比較語）を壊していないこと。"""
    assert ailine.task_asks_for_a_conditional_write(
        "内線が200以上の行のメモに「○」を付けて", roster, "名簿")


@pytest.mark.parametrize("task", [
    "営業の行の下に「新品」を追加して",     # ★ 位置の語 → 配置の依頼
    "3行目の下に新品を追加して",            # 引用も実表の値も無い
    "営業の行の上に「新品」を入れて",
])
def test_a_real_row_insertion_is_not_hijacked(roster, task):
    """★ 本物の行追加を横取りしないこと（門が広がって一番こわい所）。"""
    assert not ailine.task_asks_for_a_conditional_write(task, roster, "名簿"), task


def test_without_a_quoted_value_it_stays_shut(roster):
    """★ 書き込む値が引用で囲まれていなければ開かない（値を発明しない）。"""
    assert not ailine.task_asks_for_a_conditional_write(
        "所属が営業の行のメモを確認済にして", roster, "名簿")


def test_a_value_in_two_columns_is_not_decided(tmp_path):
    """★ ちょうど 1 列で名指しできた時だけ ── 複数なら決めない。"""
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "T"
    ws.append(["甲", "乙", "メモ"])
    ws.append(["同じ", "同じ", None])
    p = tmp_path / "b.xlsx"; wb.save(p)
    meta = ailine.build_book_meta(p)
    assert not ailine.task_asks_for_a_conditional_write(
        "同じの行のメモに「○」を付けて", meta, "T")


def test_callers_without_the_table_keep_the_old_behaviour(roster):
    """★ 実表を渡さない呼び出しは従来どおり（既存の経路を 1 ビットも変えない）。"""
    assert not ailine.task_asks_for_a_conditional_write("所属が営業の行のメモに「○」を付けて")
    assert ailine.task_asks_for_a_conditional_write("内線が200以上の行のメモに「○」を付けて")


def test_the_position_words_are_not_re_listed():
    """★ 語を新しく列挙しない ── 既に在る _ANCHOR_* を使うこと（列挙は漏れる）。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    body = src[src.index("def task_asks_for_a_conditional_write"):
               src.index("\ndef ", src.index("def task_asks_for_a_conditional_write") + 10)]
    assert "_ANCHOR_AFTER" in body and "_ANCHOR_BEFORE" in body
    assert '"の下に"' not in body, "位置の語を手で書いている"


def test_the_gate_is_passed_the_table_at_its_only_caller():
    """★ 呼び出し側 1 箇所から実表を渡していること。"""
    src = Path(ailine.__file__).read_text(encoding="utf-8")
    assert src.count("task_asks_for_a_conditional_write(") == 2, "定義 1 + 呼び出し 1 のはず"
    assert "task_asks_for_a_conditional_write(\n                a.task, book_meta, _sheet_h)" in src \
        or "a.task, book_meta, _sheet_h)" in src, "実表を渡していない"


@pytest.mark.local
def test_a_plain_condition_reaches_conditional_write(tmp_path):
    """★ 実機 ── 空行が挿さっていた依頼が、正しい行だけを書き換えること。"""
    import os, subprocess
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "名簿"
    ws.append(["氏名", "所属", "内線", "メモ"])
    for r in [["山田", "営業", 101, None], ["鈴木", "経理", 202, None],
              ["高橋", "総務", 305, None]]:
        ws.append(r)
    src = tmp_path / "in.xlsx"; wb.save(src)
    repo = Path(ailine.__file__).resolve().parents[2]
    p = subprocess.run([sys.executable, "-m", "ailine", "run", str(src),
                        "所属が営業の行のメモに「○」を付けて", "--copy",
                        "--sheet", "名簿", "--timeout", "150"],
                       capture_output=True, text=True, encoding="utf-8", errors="replace",
                       cwd=str(repo), env={**os.environ, "PYTHONPATH": str(repo / "src")})
    assert p.returncode == 0, p.stdout[-500:]
    ws2 = openpyxl.load_workbook(tmp_path / "in.out.xlsx")["名簿"]
    assert ws2.max_row == 4, "行数が変わった（空行が挿さっている）"
    assert ws2.cell(2, 4).value == "○", "営業の行に印が無い"
    assert ws2.cell(3, 4).value is None and ws2.cell(4, 4).value is None, "他の行にも付いた"


# --- ★★ 否定（★ 一度は門を閉じ、同じ日の夜に扱えるようにした）-----------------

@pytest.mark.parametrize("task", [
    "所属が営業以外の行のメモに「○」を付けて",
    "営業を除いた行のメモに「○」を付けて",
    "営業を抜いた行のメモに「○」を付けて",
])
def test_a_negated_condition_now_passes_the_gate(roster, task):
    """★★ 経緯（同じ日に 2 回動いた・どちらも意図した変更）:

    ① 2026-09-04 午後: 門を広げた直後に**静かな嘘**が出た ──
       「所属が営業**以外**の行のメモに『○』」で **営業の行に ○ が付き ✓ が出た**。
       比較語彙に否定が無いので模型は `eq` に潰れ、事後条件も**同じ述語**で数えたので
       逆でも通った（恒真）。★ その日は「否定が在る回は門を開けない」で閉じた。
    ② 2026-09-05 夜: 扱えるようにしたので開けた。
       この試験は当時こう書いてあった ──
         「この試験は『否定が直った』ら赤くなる。その時は nin を語彙に足してから
           この縛りを外すこと（直したなら**意図的な commit として扱う**）」
       そのとおりに赤くなったので、指示どおり外して**逆向きの契約**に置き換える。

    ★ いま逆のことをして ✓ が出ないことは、2 段が守っている（どちらも変異試験で赤）:
        ① 依頼文が「以外」なら比較は**機械が**否定に決める（LLM の語に従わない）
        ② 事後条件は Basic とは**別実装**の Python が数える
      ── Basic だけを『等しい』に差し替えると `×` が出ることを実機で確かめてある
         （tests/test_negated_conditional_write.py）。
    """
    assert ailine.task_asks_for_a_conditional_write(task, roster, "名簿"), task


def test_extraction_with_a_negation_is_untouched(roster):
    """★ 「〜以外を**抜き出して**」（抽出）が、条件つき書換に奪われないこと。

    ★ 門を開けた側の誤爆をここで縛る ── 抽出は書き込む値の引用を持たないので、
      この門の手前（extract_quoted_literal）で外れるのが正しい。
    """
    for task in ("営業以外を抜き出して", "営業以外の行だけ別シートに"):
        assert not ailine.task_asks_for_a_conditional_write(task, roster, "名簿"), task
    # ★ 抽出の読み直しは別の器官（except_extraction_reading）── そちらは触っていない
    assert "except_extraction_reading" in Path(ailine.__file__).read_text(encoding="utf-8")
