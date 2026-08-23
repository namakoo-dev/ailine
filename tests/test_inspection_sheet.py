"""M2.5 検分シート（出力ブック内の視覚的誘導）の検体。
   ★ 実装前に凍結した赤い検体（DESIGN-20260821-multifile M2.5 節・Namakoo 決裁 10:2x）。

   出自: 「374行目が怪しいです、と言われてもそれを追うのは負担」── 誘導は Excel の中に住む。
   原則: ★ 原本には 1 ピクセルも塗らない。全所見は ファイル+シート+セル の 3 座標完全修飾。
   色は「怪しい」の印であって検証の主張ではない（✓ は事後条件の結果からのみ）。"""
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HDRS = ["注文ID", "取引先", "金額"]


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r) + [None] * (len(headers) - len(r)))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _stack(folder, out, *extra):
    return subprocess.run(
        [sys.executable, "-m", "ailine", "stack", str(folder), "--out", str(out), *extra],
        capture_output=True, text=True, timeout=180, encoding="utf-8")


def _made(tmp_path):
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 600), ("J-2", "乙", 400), ("合計", None, 1000)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 300), ("合計", None, 9999)])   # ⚠ 閉じない
    _book(folder / "c.xlsx", ["注文ID", "取引先"], [("J-9", "戊")])              # 取れなかった
    out = tmp_path / "out.xlsx"
    p = _stack(folder, out)
    assert p.returncode == 0, p.stdout + p.stderr[-300:]
    return folder, out


def test_inspection_sheet_exists_with_findings_and_full_addresses(tmp_path):
    """検分シート: 出力 2 枚目に『検分』。全所見が ファイル+シート+セル/行 の 3 座標で載る。"""
    folder, out = _made(tmp_path)
    wb = openpyxl.load_workbook(out)
    assert "検分" in wb.sheetnames, wb.sheetnames
    text = " ".join(str(c.value) for row in wb["検分"].iter_rows() for c in row if c.value is not None)
    assert "b.xlsx" in text and "明細" in text, f"所見に 3 座標が無い: {text[:300]}"
    assert "9999" in text and "300" in text, "両側の数字が無い"
    assert "c.xlsx" in text and "金額" in text, "取れなかったの名指しが無い"
    assert "2 冊" in text or "分母" in text or "3 ファイル中" in text, "分母の要約が無い"


def test_findings_carry_hyperlinks_to_source_cells(tmp_path):
    """所見行の HYPERLINK は元ファイルのシート修飾セルへ飛ぶ（追う負担を殺す装置）。
       ★ v2（2026-08-21 実装後の実 XML 検分で改訂）: OOXML の正式形は
       relationship Target=相対パス（URI）+ hyperlink location='シート'!セル の分離。
       初版設計の『[パス]シート!セル』を Target に丸ごと入れる形は URI でなく、
       Excel が解決できない ── target と location を別々に検査する。"""
    folder, out = _made(tmp_path)
    ws = openpyxl.load_workbook(out)["検分"]
    links = [(c.hyperlink.target or "", c.hyperlink.location or "")
             for row in ws.iter_rows() for c in row if c.hyperlink is not None]
    assert links, "所見に HYPERLINK が 1 本も無い"
    assert any("b.xlsx" in t and "[" not in t and loc == "'明細'!C3"
               for t, loc in links), (
        f"Target=相対パス（URI・角括弧なし）+ location='明細'!C3"
        f"（★ 閉じなかった合計の値セルそのものに着地）の分離形でない: {links}")


def test_suspicious_rows_are_tinted_with_reason_comment_normals_untouched(tmp_path):
    """⚠ 付きファイル由来の行だけ出所列セルに淡色 + 理由コメント。正常行は塗らない。"""
    folder, out = _made(tmp_path)
    ws = openpyxl.load_workbook(out).active
    prov_col = len(HDRS) + 1   # 元ファイル 列
    tinted = {}
    for row in ws.iter_rows(min_row=2):
        cell = row[prov_col - 1]
        has_fill = cell.fill is not None and cell.fill.fgColor.rgb not in (None, "00000000")
        tinted[row[0].value] = (has_fill, cell.comment.text if cell.comment else None)
    assert tinted["J-3"][0], f"⚠ ファイル由来の行が塗られていない: {tinted}"
    assert tinted["J-3"][1] and "9999" in tinted["J-3"][1], "理由コメントに両側の数字が無い"
    assert not tinted["J-1"][0] and not tinted["J-2"][0], f"正常行が塗られている: {tinted}"


def test_legend_disclaims_color_is_not_verification(tmp_path):
    """凡例: 色は『怪しい』の印であって検証の主張でない、が検分シートに明記される。"""
    folder, out = _made(tmp_path)
    text = " ".join(str(c.value) for row in openpyxl.load_workbook(out)["検分"].iter_rows()
                    for c in row if c.value is not None)
    assert "検証" in text and ("印" in text or "主張" in text or "目安" in text), \
        f"凡例が無い: {text[:300]}"


def test_verify_and_determinism_survive_inspection_sheet(tmp_path):
    """検分シートを足しても verify は緑のまま・2 回実行でセル内容（検分込み）が一致。"""
    folder, out = _made(tmp_path)
    assert "検分" in openpyxl.load_workbook(out).sheetnames, "前提: 検分シートが存在すること"
    p = subprocess.run([sys.executable, "-m", "ailine", "verify", str(out), str(folder)],
                       capture_output=True, text=True, timeout=120, encoding="utf-8")
    assert p.returncode == 0, f"検分シートで verify が壊れた:\n{p.stdout}"
    out2 = tmp_path / "out2.xlsx"
    assert _stack(folder, out2).returncode == 0
    def cells(path):
        wb = openpyxl.load_workbook(path)
        return [(ws.title, c.coordinate, str(c.value)) for ws in wb.worksheets
                for row in ws.iter_rows() for c in row if c.value is not None]
    assert cells(out) == cells(out2), "検分シート込みの決定論が破れた"


def test_extract_output_also_gets_inspection_sheet(tmp_path, monkeypatch, capsys):
    """フォルダ抽出（run）の出力にも同じ検分シートが載る（stack だけの飾りにしない）。"""
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 400}}]})
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 600), ("合計", None, 9999)])
    rc = ailine.main(["run", str(folder), "金額が400以上の行を抜き出して"])
    capsys.readouterr()
    assert rc == 0
    out = next(tmp_path.glob("*.xlsx"))
    wb = openpyxl.load_workbook(out)
    assert "検分" in wb.sheetnames, wb.sheetnames


def test_accounting_line_in_inspection_sheet_includes_all_three_buckets(tmp_path, monkeypatch, capsys):
    """★ 実弾検分（2026-08-21 11:4x）で発覚: 検分シートの会計行が『データ行 136 =
       採用 60 + 除外 76』と 不一致 113 行の勘定を丸ごと落とし、その形のまま
       「完全会計成立」を名乗った ── 憲法 2（✓ の絶対性）違反。
       契約: 会計行は 3 勘定すべて（採用/不一致/除外）を持ち、合計が一致すること。"""
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 40000}}]})
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 30000), ("合計", None, 80000)])
    rc = ailine.main(["run", str(folder), "金額が40000以上の行を抜き出して"])
    capsys.readouterr()
    assert rc == 0
    out = next(tmp_path.glob("*.xlsx"))
    text = " ".join(str(c.value) for row in openpyxl.load_workbook(out)["検分"].iter_rows()
                    for c in row if c.value is not None)
    import re
    m = re.search(r"データ行\s*(\d+)\s*=\s*[^0-9]*(\d+)[^0-9]+(\d+)[^0-9]+(\d+)", text)
    assert m, f"3 勘定の会計行が無い: {text[:400]}"
    total, a, b, c = map(int, m.groups())
    assert total == 3 and a + b + c == 3, f"会計が合わない: {m.groups()}"
    assert "不一致" in text, "不一致の勘定が名前ごと消えている"


# ---- UX 磨き（Namakoo 指摘 2026-08-21 12:01: 色が薄い・見切れ・所見が分かりにくい）----


def test_tint_uses_excel_conventional_bad_color(tmp_path):
    """色は Excel の条件付き書式で見慣れた『悪い値』の薄赤（FFC7CE）── 説明ゼロで読める
       意味色を借りる。淡黄は「薄くて気づかない」と実視で却下（Namakoo）。"""
    folder, out = _made(tmp_path)
    ws = openpyxl.load_workbook(out).active
    prov_col = len(HDRS) + 1
    fills = {row[0].value: (row[prov_col - 1].fill.fgColor.rgb or "")
             for row in ws.iter_rows(min_row=2)}
    assert "C7CE" in fills["J-3"], f"薄赤 FFC7CE でない: {fills['J-3']}"


def test_column_widths_are_set_no_hash_marks(tmp_path):
    """列幅を内容から機械算出（CJK=2幅・決定論）。データ面も検分も見切れ・### を作らない。"""
    folder, out = _made(tmp_path)
    wb = openpyxl.load_workbook(out)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        dims = [ws.column_dimensions[chr(ord("A") + i)].width for i in range(ws.max_column)]
        # ★ openpyxl は未設定でも既定 13.0 を返す（恒真の罠・実測）── 「内容由来」を要求する:
        #   全列が一律 13.0 なら未設定と見なして不合格。
        assert not all(w == 13.0 for w in dims), f"{sheet}: 列幅が内容から算出されていない: {dims}"
        assert all(w and w >= 8 for w in dims), f"{sheet}: 狭すぎる列がある: {dims}"
    insp = wb["検分"]
    idims = [insp.column_dimensions[chr(ord("A") + i)].width for i in range(insp.max_column)]
    assert max(idims) >= 25, f"検分の長文列（所見の文章）が広がっていない: {idims}"


def test_findings_are_readable_sentences_with_explicit_next_step(tmp_path):
    """所見は断片でなく 1 所見 1 文。次の手は『リンクをクリック』を含む具体文。"""
    folder, out = _made(tmp_path)
    text = " ".join(str(c.value) for row in openpyxl.load_workbook(out)["検分"].iter_rows()
                    for c in row if c.value is not None)
    assert "クリック" in text, f"次の手に操作（クリック）の言葉が無い: {text[:400]}"
    # 閉じない合計の所見が文章として読めること（両側の数字+動詞を含む 1 文）
    assert ("合計" in text and "合いません" in text) or ("合計" in text and "一致しません" in text), \
        f"所見が文章になっていない: {text[:500]}"


def test_whole_row_is_tinted_not_just_provenance_cell(tmp_path):
    """★ Namakoo 実視 2 巡目（12:14）: 出所セル 1 個の印は目が拾わない ──
       疑わしいファイル由来の行は データセル全部 を薄赤に（行の帯）。正常行は不変。"""
    folder, out = _made(tmp_path)
    ws = openpyxl.load_workbook(out).active
    def tinted(cell):
        return cell.fill is not None and "C7CE" in (cell.fill.fgColor.rgb or "")
    rows = {row[0].value: [tinted(c) for c in row[:len(HDRS) + 2]]
            for row in ws.iter_rows(min_row=2)}
    assert all(rows["J-3"]), f"疑わしい行の帯が欠けている: {rows['J-3']}"
    assert not any(rows["J-1"]) and not any(rows["J-2"]), f"正常行が塗られた: {rows}"


def test_finding_rows_in_inspection_sheet_are_tinted(tmp_path):
    """検分シートの所見行そのものも薄赤 ── 開いた瞬間に目が行く。"""
    folder, out = _made(tmp_path)
    ws = openpyxl.load_workbook(out)["検分"]
    tinted_texts = []
    for row in ws.iter_rows():
        if any(c.fill is not None and "C7CE" in (c.fill.fgColor.rgb or "") for c in row):
            tinted_texts.append(" ".join(str(c.value) for c in row if c.value is not None))
    assert any("合いません" in t or "取れなかった" in t or "合いま" in t for t in tinted_texts), \
        f"所見行が塗られていない: {tinted_texts}"


def test_legend_explains_link_landing_is_the_marker(tmp_path):
    """凡例: 原本には印を付けない・リンクの着地セル（選択状態）が対象、を仕様として明記。"""
    folder, out = _made(tmp_path)
    text = " ".join(str(c.value) for row in openpyxl.load_workbook(out)["検分"].iter_rows()
                    for c in row if c.value is not None)
    assert "原本" in text and ("印" in text or "塗" in text), f"原本不可侵の明記が無い: {text[-400:]}"
    assert "選択" in text or "着地" in text, "リンク着地=対象セルの説明が無い"
    assert "Ctrl" in text, "Ctrl+クリックの案内が無い（Namakoo が実際に引っかかった・2026-08-21）"
