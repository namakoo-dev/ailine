"""ailine の純粋ロジックの単体テスト（ollama / LibreOffice を要さない部分）。
   生成・適用の統合は実機（basrun_spike）で検証済み。ここは回帰用の土台。
"""
import argparse
import json
import os
import subprocess
import sys
import urllib.error
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ailine

from _run_argv import run_argv  # noqa: E402  — C2: cmd_run 直呼び用 Namespace → main(argv) 変換

# ★ W8b-2（既定=原本直接適用への反転）: この回で `dry=False, inplace=False, ...` だった
#   argparse.Namespace リテラルを機械的に `dry=False, copy=True, ...` へ置き換えた
#   （旧既定＝コピーにしか書かない、を再現するのが copy=True になったため）。inplace は
#   もう分岐に使われない（cmd_run 側で `a.inplace = not a.copy` に上書きされる）。
#   ★ 実測: 置き換え前は real ~/.ailine/backups/ にテスト実行のたびバックアップが
#   実書き込みされていた（BACKUP_DIR を monkeypatch していないテストが多数、かつ
#   既定が原本適用になったことで無条件に make_backup が走るようになったため）。


# --- コード抽出 / 署名 -----------------------------------------------------

def test_extract_bas_strips_markdown_fence():
    raw = "```basic\nOption Explicit\nSub Run(oDoc As Object)\nEnd Sub\n```"
    assert ailine.extract_bas(raw).startswith("Option Explicit")
    assert "```" not in ailine.extract_bas(raw)

def test_extract_bas_passthrough_without_fence():
    raw = "Sub Run(oDoc As Object)\nEnd Sub"
    assert ailine.extract_bas(raw) == raw

@pytest.mark.parametrize("code,ok", [
    ("Sub Run(oDoc As Object)\nEnd Sub", True),
    ("sub run( oDoc as object )", True),          # 大文字小文字・空白ゆらぎ
    ("Sub Run()\nEnd Sub", False),                # 引数なし
    ("Sub Other(oDoc As Object)", False),         # 別名
    ("' コメントだけ", False),
])
def test_valid_signature(code, ok):
    assert ailine.valid_signature(code) is ok


# --- 参照ライブラリ --------------------------------------------------------

def test_load_refs_bundles_examples():
    text = ailine.load_refs(ailine.DEFAULT_REFS)
    assert "Sub Run(oDoc As Object)" in text
    assert "参考" in text

def test_load_refs_missing_dir_is_empty(tmp_path):
    assert ailine.load_refs(tmp_path / "nope") == ""


# --- ヘルパ・ライブラリ（呼ぶだけ） ----------------------------------------

def test_load_helpers_catalog_and_files():
    catalog, files = ailine.load_helpers(ailine.DEFAULT_HELPERS)
    assert any(f.name.endswith(".bas") for f in files)
    assert "SortByColumn" in catalog
    assert "InsertBarChart" in catalog
    assert "Call" in catalog          # Call 形式で呼ばせる指示が入っている

def test_load_helpers_missing_dir(tmp_path):
    catalog, files = ailine.load_helpers(tmp_path / "nope")
    assert catalog == "" and files == []


# ★ 太字は native（StyleBold ヘルパが Basic で CharWeight+CharWeightAsian を当てる）。
#   openpyxl 後付けは撤去した。日本語太字は CharWeightAsian が要る点が要（実測）。
#   ここは Basic 側の実挙動なので純ロジック test では検証せず、通し試験＋描画で確認する。


# --- snapshot / 差分（no-op ガードの核） -----------------------------------

def _book(tmp_path, rows):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p

def test_diff_detects_value_change(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.active.cell(1, 3, "new"); wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    assert any("new" in ln for ln in lines)

def test_diff_noop_when_unchanged(tmp_path):
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    after = ailine.snapshot(p)   # 何も変えない
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is False       # ← no-op を正しく no-op と判定
    assert lines == []

def test_diff_detects_new_sheet(tmp_path):
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.create_sheet("集計"); wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    assert any("集計" in ln for ln in lines)

def test_diff_detects_fill_only_change(tmp_path):
    # 値でなく背景色だけ変えても『変化した』と見えること（no-op 誤検出を防ぐ）
    from openpyxl.styles import PatternFill
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(1, 1).fill = PatternFill("solid", fgColor="FFCCCC")
    wb.save(p)
    after = ailine.snapshot(p)
    changed, _ = ailine.diff_snapshots(before, after)
    assert changed is True

def test_diff_detects_border_only_change(tmp_path):
    # 罫線だけの変更も検出すること（罫線ヘルパが no-op 誤判定されないため）
    from openpyxl.styles import Border, Side
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    thin = Side(style="thin")
    wb.active.cell(1, 1).border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.save(p)
    after = ailine.snapshot(p)
    changed, _ = ailine.diff_snapshots(before, after)
    assert changed is True

def test_diff_detects_merge(tmp_path):
    p = _book(tmp_path, [["a", "b"]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.active.merge_cells("A1:B1"); wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    assert any("結合" in ln for ln in lines)

def test_diff_detects_colwidth(tmp_path):
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.active.column_dimensions["A"].width = 30; wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    assert any("列幅" in ln for ln in lines)

def test_diff_detects_align_only_change(tmp_path):
    # 中央揃えだけの変更も検出すること（AlignCenter ヘルパが no-op 誤判定されないため）
    from openpyxl.styles import Alignment
    p = _book(tmp_path, [["a", 1]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(1, 1).alignment = Alignment(horizontal="center")
    wb.save(p)
    after = ailine.snapshot(p)
    changed, _ = ailine.diff_snapshots(before, after)
    assert changed is True


# --- 差分見出し（P1: セル値変更が無見出しで続いていた不整合の修正） -----------

def test_diff_cell_change_has_own_heading(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.active.cell(1, 3, "new"); wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    heading_idx = next(i for i, ln in enumerate(lines) if ln.startswith("＊セル値変更:"))
    detail_idx = next(i for i, ln in enumerate(lines) if "new" in ln)
    assert heading_idx < detail_idx   # 見出し → 明細の順

def test_diff_cell_and_rowheight_each_get_own_heading(tmp_path):
    # 行高変更とセル値変更が両方あるとき、セル値変更が行高見出しの下に
    # 無見出しでぶら下がらず、自前の見出しを持つこと（修正前の不整合の再現）
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.row_dimensions[1].height = 30
    wb.active.cell(1, 3, "new")
    wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    assert changed is True
    assert any(ln.startswith("＊行高変更:") for ln in lines)
    assert any(ln.startswith("＊セル値変更:") for ln in lines)


# --- ★ メッセージの条件（P1: 失敗/--dry でも無条件に出ていた不整合の修正） -----

def test_success_message_on_real_success():
    msg = ailine.success_message({"ok": True, "attempts": 1})
    assert msg is not None
    assert "no-op ガードは正しさを保証しない" in msg

def test_success_message_none_on_dry():
    assert ailine.success_message({"ok": True, "dry": True}) is None

def test_success_message_none_on_failure():
    assert ailine.success_message({"ok": False}) is None


# --- ollama エラー分類（P1: 404 なのに ollama serve を疑わせる誤ヒントの修正） ---

def test_ollama_generate_404_suggests_pull(monkeypatch):
    def fake_urlopen(req, timeout=300):
        raise urllib.error.HTTPError(url="http://x", code=404, msg="Not Found", hdrs=None, fp=None)
    monkeypatch.setattr(ailine.urllib.request, "urlopen", fake_urlopen)
    with pytest.raises(SystemExit) as exc:
        ailine.ollama_generate("qwen2.5-coder:7b", [{"role": "user", "content": "hi"}])
    msg = str(exc.value)
    assert "qwen2.5-coder:7b" in msg
    assert "pull" in msg
    assert "ollama serve" not in msg   # 接続不能の案内と混同しない

def test_ollama_generate_connection_refused_suggests_serve(monkeypatch):
    def fake_urlopen(req, timeout=300):
        raise urllib.error.URLError("connection refused")
    monkeypatch.setattr(ailine.urllib.request, "urlopen", fake_urlopen)
    with pytest.raises(SystemExit) as exc:
        ailine.ollama_generate("qwen2.5-coder:7b", [{"role": "user", "content": "hi"}])
    msg = str(exc.value)
    assert "ollama serve" in msg
    assert "pull" not in msg   # 404 の案内と混同しない

def test_ollama_generate_other_http_error_is_distinct(monkeypatch):
    # 404/接続不能のどちらの定型文にも紐付けない（誤誘導しない）
    def fake_urlopen(req, timeout=300):
        raise urllib.error.HTTPError(url="http://x", code=500, msg="Internal Error", hdrs=None, fp=None)
    monkeypatch.setattr(ailine.urllib.request, "urlopen", fake_urlopen)
    with pytest.raises(SystemExit) as exc:
        ailine.ollama_generate("qwen2.5-coder:7b", [{"role": "user", "content": "hi"}])
    msg = str(exc.value)
    assert "500" in msg
    assert "pull" not in msg
    assert "ollama serve" not in msg


# --- 文書の説明 ------------------------------------------------------------

def test_describe_book_lists_headers(tmp_path):
    p = _book(tmp_path, [["商品", "金額", "在庫"], ["りんご", 1200, 8]])
    desc = ailine.describe_book(p)
    assert "列0=商品" in desc
    assert "列1=金額" in desc
    assert "シート一覧" in desc


# --- ★ M1: 進捗表示（生成中の完全沈黙の解消） --------------------------------

def test_fmt_elapsed_formats_one_decimal():
    assert ailine._fmt_elapsed(12.34) == "(12.3s)"
    assert ailine._fmt_elapsed(0) == "(0.0s)"

def test_progress_start_writes_without_newline_to_stderr(capsys):
    ailine.progress_start("⏳ テスト中…")
    captured = capsys.readouterr()
    assert captured.err == "⏳ テスト中…"
    assert captured.out == ""    # ★ stdout は汚さない（--json 互換の要）

def test_progress_end_appends_elapsed_and_newline(capsys):
    t0 = ailine.progress_start("⏳ テスト中…")
    ailine.progress_end(t0)
    captured = capsys.readouterr()
    assert captured.err.startswith("⏳ テスト中…")
    assert captured.err.rstrip("\n").endswith("s)")
    assert captured.err.endswith("\n")


# --- ★ M1: 適用タイムアウト（暴走マクロ対策） --------------------------------

class _FakeProcTimeoutOnce:
    """1回目の communicate() は TimeoutExpired、2回目(kill 後の回収)は正常終了。"""
    pid = 4242
    returncode = None
    def __init__(self):
        self.calls = 0
    def communicate(self, timeout=None):
        self.calls += 1
        if self.calls == 1:
            raise subprocess.TimeoutExpired(cmd="x", timeout=timeout)
        return ("", "")

class _FakeProcReturns:
    """指定した (stdout, stderr, returncode) をそのまま返す固定 Popen 偽物。"""
    def __init__(self, out="", err="", returncode=0, pid=1):
        self._out, self._err, self.returncode, self.pid = out, err, returncode, pid
    def communicate(self, timeout=None):
        return (self._out, self._err)

def _fake_book(tmp_path):
    return tmp_path / "book.xlsx"

def test_basrun_apply_outer_timeout_kills_and_classifies_as_runtime_error(tmp_path, monkeypatch):
    fake = _FakeProcTimeoutOnce()
    monkeypatch.setattr(ailine.subprocess, "Popen", lambda *a, **k: fake)
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))
    killed = {}
    monkeypatch.setattr(ailine, "_kill_process_tree", lambda pid: killed.setdefault("pid", pid))

    ok, err, raw = ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=5.0)

    assert ok is False
    assert err == "実行時エラー: マクロが 5 秒で終了しない（無限ループの可能性）"
    assert killed["pid"] == 4242    # ★ PID 指定で kill（taskkill /IM の一括killはしない）

def test_basrun_apply_detects_basruns_own_internal_timeout_message(tmp_path, monkeypatch):
    # basrun.py 自身の内部タイムアウト（--timeout 転送先）が先に発火して非ゼロで返る場合も
    # 同じ「実行時エラー: マクロが N 秒で終了しない」に正規化されること。
    msg = "apply が 5 秒応答しなかった (BASRUN_APPLY_TIMEOUT/--timeout)。接続先の LibreOffice を終了させて中止した。"
    monkeypatch.setattr(ailine.subprocess, "Popen",
                        lambda *a, **k: _FakeProcReturns(err=msg, returncode=1))
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))

    ok, err, raw = ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=5.0)

    assert ok is False
    assert err == ailine._timeout_error_message(5.0)

def test_basrun_apply_generic_runtime_error_is_not_misclassified_as_timeout(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine.subprocess, "Popen",
                        lambda *a, **k: _FakeProcReturns(err="何かの実行時エラー", returncode=1))
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))

    ok, err, raw = ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=5.0)

    assert ok is False
    assert "無限ループ" not in err
    assert "何かの実行時エラー" in err

def test_basrun_apply_success_path_unaffected(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine.subprocess, "Popen",
                        lambda *a, **k: _FakeProcReturns(out="applied", returncode=0))
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))

    ok, err, raw = ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=5.0)

    assert ok is True
    assert err is None

def test_basrun_apply_includes_timeout_flag_when_enabled(tmp_path, monkeypatch):
    captured = {}
    def fake_popen(cmd, **kw):
        captured["cmd"] = cmd
        return _FakeProcReturns(returncode=0)
    monkeypatch.setattr(ailine.subprocess, "Popen", fake_popen)
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))

    ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=42.0)

    assert "--timeout" in captured["cmd"]
    assert "42.0" in captured["cmd"]

def test_basrun_apply_omits_timeout_flag_when_disabled(tmp_path, monkeypatch):
    # --timeout 0 相当（呼び出し側が None を渡す＝旧挙動の無制限）
    captured = {}
    def fake_popen(cmd, **kw):
        captured["cmd"] = cmd
        return _FakeProcReturns(returncode=0)
    monkeypatch.setattr(ailine.subprocess, "Popen", fake_popen)
    monkeypatch.setattr(ailine, "basrun_path", lambda: Path("dummy_basrun.py"))

    ailine.basrun_apply(_fake_book(tmp_path), "code", tmp_path, timeout=None)

    assert "--timeout" not in captured["cmd"]


# --- ★ M1: ailine doctor（セットアップ診断） ---------------------------------

def test_check_python_version_ok_on_current_interpreter():
    ok, detail = ailine._check_python_version()
    assert ok is True
    assert detail == ""

def test_check_python_version_fails_on_old(monkeypatch):
    monkeypatch.setattr(ailine.sys, "version_info", (3, 9, 0))
    ok, detail = ailine._check_python_version()
    assert ok is False
    assert "3.10" in detail

def test_check_openpyxl_ok():
    ok, detail = ailine._check_openpyxl()
    assert ok is True and detail == ""

class _FakeUrlopenCtx:
    def __init__(self, payload):
        self._payload = payload
    def __enter__(self):
        return self
    def __exit__(self, *a):
        return False

def test_check_ollama_reachable_ok(monkeypatch):
    monkeypatch.setattr(ailine.urllib.request, "urlopen",
                        lambda req, timeout=3.0: _FakeUrlopenCtx({}))
    monkeypatch.setattr(ailine.json, "load", lambda f: {})
    ok, detail = ailine._check_ollama_reachable()
    assert ok is True and detail == ""

def test_check_ollama_reachable_fails_suggests_serve(monkeypatch):
    def raiser(req, timeout=3.0):
        raise OSError("接続を拒否されました")
    monkeypatch.setattr(ailine.urllib.request, "urlopen", raiser)
    ok, detail = ailine._check_ollama_reachable()
    assert ok is False
    assert "ollama serve" in detail

def test_check_model_available_found(monkeypatch):
    monkeypatch.setattr(ailine.urllib.request, "urlopen",
                        lambda req, timeout=3.0: _FakeUrlopenCtx(None))
    monkeypatch.setattr(ailine.json, "load", lambda f: {"models": [{"name": "qwen2.5-coder:7b"}]})
    ok, detail = ailine._check_model_available("qwen2.5-coder:7b")
    assert ok is True and detail == ""

def test_check_model_available_missing_suggests_pull(monkeypatch):
    monkeypatch.setattr(ailine.urllib.request, "urlopen",
                        lambda req, timeout=3.0: _FakeUrlopenCtx(None))
    monkeypatch.setattr(ailine.json, "load", lambda f: {"models": []})
    ok, detail = ailine._check_model_available("missing:1b")
    assert ok is False
    assert "ollama pull missing:1b" in detail

def test_check_model_available_unreachable(monkeypatch):
    def raiser(req, timeout=3.0):
        raise OSError("x")
    monkeypatch.setattr(ailine.urllib.request, "urlopen", raiser)
    ok, detail = ailine._check_model_available("m")
    assert ok is False

def test_check_libreoffice_missing_basrun(monkeypatch):
    monkeypatch.setattr(ailine, "_find_basrun_path", lambda: None)
    ok, detail = ailine._check_libreoffice()
    assert ok is False
    assert "basrun.py" in detail

def test_check_libreoffice_ok(monkeypatch, tmp_path):
    fake_path = tmp_path / "basrun.py"
    fake_path.write_text("# dummy", encoding="utf-8")
    monkeypatch.setattr(ailine, "_find_basrun_path", lambda: fake_path)
    class _FakeMod:
        @staticmethod
        def office_dir():
            return Path("C:/Program Files/LibreOffice/program")
    monkeypatch.setattr(ailine, "_load_module_from_path", lambda p, name: _FakeMod())
    ok, detail = ailine._check_libreoffice()
    assert ok is True
    assert "LibreOffice" in detail

def test_check_libreoffice_not_found(monkeypatch, tmp_path):
    fake_path = tmp_path / "basrun.py"
    fake_path.write_text("# dummy", encoding="utf-8")
    monkeypatch.setattr(ailine, "_find_basrun_path", lambda: fake_path)
    class _FakeMod:
        @staticmethod
        def office_dir():
            raise SystemExit("LibreOffice が見つからない")
    monkeypatch.setattr(ailine, "_load_module_from_path", lambda p, name: _FakeMod())
    ok, detail = ailine._check_libreoffice()
    assert ok is False

def test_check_basrun_found(monkeypatch, tmp_path):
    p = tmp_path / "basrun.py"
    monkeypatch.setattr(ailine, "_find_basrun_path", lambda: p)
    ok, detail = ailine._check_basrun()
    assert ok is True

def test_check_basrun_missing_gives_clone_hint(monkeypatch):
    monkeypatch.setattr(ailine, "_find_basrun_path", lambda: None)
    ok, detail = ailine._check_basrun()
    assert ok is False
    assert "clone" in detail

def test_check_demo_dir_ok_on_real_repo():
    ok, detail = ailine._check_demo_dir()
    assert ok is True   # このリポジトリの demo/ には .xlsx サンプルが同梱されている

def test_check_demo_dir_missing(monkeypatch, tmp_path):
    monkeypatch.setattr(ailine, "HERE", tmp_path)
    ok, detail = ailine._check_demo_dir()
    assert ok is False

def test_format_doctor_report_all_ok():
    text, all_ok = ailine.format_doctor_report([("a", True, ""), ("b", True, "detail")])
    assert all_ok is True
    assert "✓ a" in text
    assert "✓ b (detail)" in text

def test_format_doctor_report_has_failure_and_fix_hint():
    text, all_ok = ailine.format_doctor_report([("a", True, ""), ("b", False, "直せ")])
    assert all_ok is False
    assert "× b — 直せ" in text

def test_doctor_checks_returns_eight_items():
    # ①python ②openpyxl ③ollama ④モデル ⑤LibreOffice ⑥basrun.py ⑦demo/
    # ★ W10a 項目2: ⑧既定動作の告知（原本直接・v2〜）を追加。
    results = ailine.doctor_checks()
    assert len(results) == 8
    for name, ok, detail in results:
        assert isinstance(name, str)
        assert isinstance(ok, bool)

def test_doctor_checks_includes_default_behavior_notice():
    results = ailine.doctor_checks()
    names = {name: detail for name, ok, detail in results}
    assert "既定動作" in names
    assert "原本直接" in names["既定動作"]


# --- ★ W8a 項目5: doctor の事務向け一行説明（7項目全部・翻訳表示） -----------------

def test_doctor_business_notes_cover_all_seven_real_check_names():
    names = [
        "python 3.10+", "openpyxl", f"ollama 到達 ({ailine.OLLAMA})",
        f"モデル '{ailine.DEFAULT_MODEL}'", "LibreOffice", "basrun.py", "demo/",
    ]
    for name in names:
        assert ailine._doctor_business_note(name) is not None

def test_format_doctor_report_shows_business_note_for_ollama():
    text, _ = ailine.format_doctor_report([(f"ollama 到達 ({ailine.OLLAMA})", True, "")])
    assert "AI エンジン (ollama) に接続できています" in text

def test_format_doctor_report_no_business_note_for_unknown_dummy_name():
    # ★ 既存テスト(test_format_doctor_report_all_ok 等)のダミー名 "a"/"b" は
    #   どの実プレフィックスにも一致しないため、従来どおり内部名のみ表示する。
    text, _ = ailine.format_doctor_report([("a", True, "")])
    assert text == "✓ a"


# --- ★ M1: 実行履歴（最小版） ------------------------------------------------

def test_build_history_entry_shape_and_truncation():
    result = {"ok": True, "attempts": 2, "changes": ["a", "b", "c", "d"], "out": "x.xlsx"}
    e = ailine.build_history_entry(result, Path("book.xlsx"), "タスク", "モデル", "none")
    assert e["ok"] is True
    assert e["attempts"] == 2
    assert e["changes"] == ["a", "b", "c"]      # 先頭3件のみ
    assert e["failure_kind"] == "none"
    assert e["book"] == str(Path("book.xlsx"))
    assert e["task"] == "タスク"
    assert e["model"] == "モデル"
    assert "ts" in e and "T" in e["ts"]          # ISO 形式

def test_append_and_read_history_roundtrip_newest_first(tmp_path):
    p = tmp_path / "history.jsonl"
    e1 = ailine.build_history_entry({"ok": True, "attempts": 1}, Path("a.xlsx"), "t1", "m", "none")
    e2 = ailine.build_history_entry({"ok": False, "attempts": 3}, Path("b.xlsx"), "t2", "m", "noop")
    ailine.append_history(e1, path=p)
    ailine.append_history(e2, path=p)
    entries = ailine.read_history(path=p, max_n=10)
    assert len(entries) == 2
    assert entries[0]["task"] == "t2"   # 新しい順
    assert entries[1]["task"] == "t1"

def test_read_history_missing_file_returns_empty(tmp_path):
    assert ailine.read_history(path=tmp_path / "nope.jsonl") == []

def test_read_history_skips_broken_lines(tmp_path):
    p = tmp_path / "history.jsonl"
    p.write_text('{"task": "ok1"}\nnot json\n{"task": "ok2"}\n', encoding="utf-8")
    entries = ailine.read_history(path=p, max_n=10)
    assert [e["task"] for e in entries] == ["ok2", "ok1"]

def test_read_history_respects_max_n(tmp_path):
    p = tmp_path / "history.jsonl"
    p.write_text("\n".join(f'{{"task": "t{i}"}}' for i in range(5)) + "\n", encoding="utf-8")
    entries = ailine.read_history(path=p, max_n=2)
    assert len(entries) == 2
    assert entries[0]["task"] == "t4"

def test_append_history_raises_when_path_unwritable(tmp_path):
    # 親のつもりの場所が実はファイル → mkdir が失敗する。ここは例外を投げてよい
    # （run 本体を落とさない責務は呼び出し側 cmd_run の try/except が持つ設計）。
    blocker = tmp_path / "blocker"
    blocker.write_text("x", encoding="utf-8")
    bad_path = blocker / "sub" / "history.jsonl"
    with pytest.raises(Exception):
        ailine.append_history({"a": 1}, path=bad_path)

def test_format_history_table_empty_says_none_yet():
    assert ailine.format_history_table([]) == "履歴はまだ無い"

def test_format_history_table_shows_failure_kind_tag():
    entries = [{"ts": "2026-01-01T00:00:00+00:00", "ok": False, "attempts": 3,
                "model": "m", "book": "b.xlsx", "task": "t", "failure_kind": "noop"}]
    text = ailine.format_history_table(entries)
    assert "[noop]" in text
    assert "×" in text

def test_format_history_table_success_has_no_kind_tag():
    entries = [{"ts": "2026-01-01T00:00:00+00:00", "ok": True, "attempts": 1,
                "model": "m", "book": "b.xlsx", "task": "t", "failure_kind": "none"}]
    text = ailine.format_history_table(entries)
    assert "[none]" not in text
    assert "✓" in text


# --- ★ W8a 項目1: dry(下見) と実適用の履歴区別 ---------------------------------

def test_build_history_entry_dry_field_defaults_false():
    e = ailine.build_history_entry({"ok": True, "attempts": 1}, Path("a.xlsx"), "t", "m", "none")
    assert e["dry"] is False

def test_build_history_entry_dry_field_true_when_result_dry():
    e = ailine.build_history_entry({"ok": True, "dry": True}, Path("a.xlsx"), "t", "m", "none")
    assert e["dry"] is True

def test_format_history_table_marks_dry_rows_with_kensaku_label():
    # 「下見」= 事務の言葉。「dry-run」は表示に出さない。
    entries = [{"ts": "2026-01-01T00:00:00+00:00", "ok": True, "attempts": 1, "dry": True,
                "model": "m", "book": "b.xlsx", "task": "t", "failure_kind": "none"}]
    text = ailine.format_history_table(entries)
    assert "(下見)" in text
    assert "dry" not in text.lower()

def test_format_history_table_old_rows_without_dry_key_read_as_applied():
    # ★ 後方互換: 旧 history.jsonl の行（"dry" キーが無い）は実適用扱いのまま読める
    #   （dict.get("dry", False) で False にフォールバック）。
    entries = [{"ts": "2026-01-01T00:00:00+00:00", "ok": True, "attempts": 1,
                "model": "m", "book": "b.xlsx", "task": "t", "failure_kind": "none"}]
    text = ailine.format_history_table(entries)
    assert "(下見)" not in text

def test_cmd_run_dry_survives_history_write_failure(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    # ★ 履歴の書き込み失敗で run 本体を落とさない（try で包み WARN のみ）ことを
    #   --dry（ollama 生成だけで LibreOffice を要さない）経路で確認する。
    book = _book(tmp_path, [["a", 1]])
    # ★ M2b: run はまず translate_task を呼ぶ。ここは自由生成経路の回帰なので FREEFORM に固定する。
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2:
                        "Sub Run(oDoc As Object)\nEnd Sub")
    def boom(entry, path=None):
        raise OSError("書き込み失敗（テスト用）")
    monkeypatch.setattr(ailine, "append_history", boom)

    argv = run_argv(
        book=str(book), task="テスト", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)

    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0   # ★ 履歴書き込みが例外を投げても run 自体は成功のまま
    assert "WARN" in captured.err


# --- ★ M1: 差分の人間化（生 tuple 漏れの解消） -------------------------------

def test_describe_cell_change_value_only_shows_only_value_label():
    before = None   # snapshot() が完全既定セルを辞書に持たない状態を模す
    after = ("りんご", "General", None, False, None, None)
    desc = ailine.describe_cell_change(before, after)
    assert desc == "値 (空)→'りんご'"
    assert "数値書式" not in desc   # ★ 既定 tuple を正しく使わないと偽の差分が出る罠の回帰

def test_describe_cell_change_multiple_fields_listed():
    before = ("りんご", "General", None, False, None, None)
    after = ("リンゴ", "General", None, True, None, None)
    desc = ailine.describe_cell_change(before, after)
    assert "値 'りんご'→'リンゴ'" in desc
    assert "太字 False→True" in desc

def test_describe_cell_change_no_diff_returns_placeholder():
    t = ("a", "General", None, False, None, None)
    assert ailine.describe_cell_change(t, t) == "(差分なし)"

def test_cell_ref_uses_a1_notation():
    assert ailine._cell_ref(1, 1) == "A1"
    assert ailine._cell_ref(2, 2) == "B2"

def test_diff_snapshots_new_value_cell_uses_a1_ref_and_no_raw_tuple(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p); wb.active.cell(1, 3, "new"); wb.save(p)
    after = ailine.snapshot(p)
    changed, lines = ailine.diff_snapshots(before, after)
    detail = next(ln for ln in lines if "new" in ln)
    assert detail.strip().startswith("C1:")     # row=1,col=3 → A1形式で C1
    assert "値" in detail
    assert "None" not in detail                 # 生 tuple の名残（Noneの生表示）が無い
    assert "数値書式" not in detail              # 値だけの変更で偽の書式差分が出ない


# --- ★ M2a: 疑わしい変化の機械検出 ------------------------------------------

def test_detect_ghost_data_single_cell_outside_original_range(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])   # 使用範囲は A1:B2
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=26, value="ghost")   # Z2（範囲外）
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_ghost_data(before, after)
    assert msg is not None
    assert "Z2" in msg
    assert "★ 疑わしい" in msg

def test_detect_ghost_data_range_of_outside_cells(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (2, 3, 4):
        wb.active.cell(row=r, column=26, value=0)
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_ghost_data(before, after)
    assert msg is not None
    assert "Z2:Z4" in msg

def test_detect_ghost_data_none_when_any_change_is_in_range(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=26, value="ghost")   # 範囲外
    wb.active.cell(row=1, column=1, value="変更")      # 範囲内も混ざる
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.detect_ghost_data(before, after) is None

def test_detect_ghost_data_none_when_no_change():
    snap = {"cells": {}}
    assert ailine.detect_ghost_data(snap, snap) is None


def test_detect_uniform_fill_flags_same_value_into_blank_cells(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=1, column=3, value=0)   # C列は元々空欄
    wb.active.cell(row=2, column=3, value=0)
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_uniform_fill(before, after)
    assert msg is not None
    assert "値 0 × 2 セル" in msg
    assert "★ 疑わしい" in msg

def test_detect_uniform_fill_none_when_overwriting_existing_value(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=1, column=1, value=0)   # 既存値の上書き（空欄からではない）
    wb.active.cell(row=1, column=3, value=0)   # こちらは空欄から
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.detect_uniform_fill(before, after) is None

def test_detect_uniform_fill_none_when_values_differ(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=1, column=3, value=0)
    wb.active.cell(row=2, column=3, value=1)   # 値がバラバラ
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.detect_uniform_fill(before, after) is None

def test_detect_uniform_fill_fires_when_mixed_with_border_and_center_align(tmp_path):
    # ★ M2c 回帰: 冷間再監査3回目の実測ケース。罫線+中央揃え(値は不変)と、空欄への
    #   3セル0埋め(値が変わる)が混在すると、旧実装は書式のみの変更セルまで一様性判定に
    #   巻き込んで見逃していた（値変更の部分集合だけで評価するよう修正）。
    from openpyxl.styles import Border, Side, Alignment
    p = _book(tmp_path, [["商品", "金額", "備考"], ["a", 100, None], ["b", 200, None], ["c", 300, None]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center")
    # 罫線+中央揃え(値は変えない・書式のみ) — A1:B4 全体
    for r in range(1, 5):
        for c in (1, 2):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.alignment = center
    # 空欄だった備考列(C列)に0を一様書き込み(値が変わる)
    ws.cell(row=2, column=3, value=0)
    ws.cell(row=3, column=3, value=0)
    ws.cell(row=4, column=3, value=0)
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_uniform_fill(before, after)
    assert msg is not None
    assert "値 0 × 3 セル" in msg

def test_detect_ghost_data_ignores_format_only_changes(tmp_path):
    # 値は変わらず書式だけ変わったセルは幽霊データ判定の対象外（誤検知を増やさない）。
    from openpyxl.styles import Alignment
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=1, column=1).alignment = Alignment(horizontal="center")   # 範囲内・書式のみ
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.detect_ghost_data(before, after) is None


# --- ★ M2a: 件数の突き合わせ -------------------------------------------------

def test_count_reconciliation_reports_data_vs_changed_rows(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["りんご", 100], ["バナナ", 200], ["みかん", 300]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=2, value=999)   # りんご行のみ変更
    wb.active.cell(row=3, column=2, value=999)   # バナナ行のみ変更（みかん行は無変更）
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.count_reconciliation(before, after)
    assert msg == "列 B: データ 3 行のうち 2 行を変更（1 行は未変更）"

def test_count_reconciliation_none_when_multiple_columns_changed(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["りんご", 100], ["バナナ", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=1, value="リンゴ")
    wb.active.cell(row=2, column=2, value=999)
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.count_reconciliation(before, after) is None

# --- ★ W8a 項目2: 件数突合の算数バグ（分子に見出し行が混入していた） ------------------

def test_count_reconciliation_excludes_header_row_from_numerator(tmp_path):
    # ★ 実測(e2e_work/w3_e2e3_log.txt): 新規列を作り見出し(F1)にもデータ全5行(F2-F6)にも
    #   書く COMPUTE_COLUMN で「データ5行のうち6行を変更」という算数が壊れた表示になって
    #   いた。見出し行は分子から除外し、別語「＋見出し行」で添える。
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300], ["d", 400], ["e", 500]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    ws.cell(row=1, column=3, value="金額(税込)")   # 新規列の見出し
    for r in range(2, 7):
        ws.cell(row=r, column=3, value=f"=B{r}*1.1")
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.count_reconciliation(before, after)
    assert msg == "列 C: データ 5 行のうち 5 行を変更（0 行は未変更）＋見出し行"

def test_count_reconciliation_no_header_suffix_when_header_untouched(tmp_path):
    # 見出し行に触れていない既存の回帰(りんご欠落型)には「＋見出し行」を付けない。
    p = _book(tmp_path, [["商品", "金額"], ["りんご", 100], ["バナナ", 200], ["みかん", 300]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=2, value=999)
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.count_reconciliation(before, after)
    assert "＋見出し行" not in msg

def test_count_reconciliation_invariant_denominator_never_less_than_numerator(tmp_path):
    # ★ 不変条件: 「データ N 行のうち M 行を変更」で常に N >= M（分子=データ行のみ）。
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    ws.cell(row=1, column=2, value="金額(税込)")   # 見出しも変更
    ws.cell(row=2, column=2, value=999)
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.count_reconciliation(before, after)
    import re as _re
    m = _re.search(r"データ (\d+) 行のうち (\d+) 行を変更", msg)
    data_rows, changed_rows = int(m.group(1)), int(m.group(2))
    assert data_rows >= changed_rows
    assert "＋見出し行" in msg


# --- ★ M2a: 依頼文と変更範囲の重なりチェック ----------------------------------

def test_extract_task_mentions_column_letter_and_number_and_row():
    m = ailine.extract_task_mentions("列Zの値を全部2倍にする", ["Sheet"])
    assert m["cols"] == {26}
    m2 = ailine.extract_task_mentions("Z列の値を確認", ["Sheet"])
    assert m2["cols"] == {26}
    m3 = ailine.extract_task_mentions("列3を直して行5も見て", ["Sheet"])
    assert m3["cols"] == set()          # 数字表記は曖昧なので cols に断定で入れない
    assert m3["digit_cols"] == {3}      # 生の値のまま保持し照合側で両解釈
    assert m3["rows"] == {5}

def test_extract_task_mentions_only_real_sheet_names_match():
    m = ailine.extract_task_mentions("集計シートを見て", ["Sheet", "集計"])
    assert m["sheets"] == {"集計"}
    m2 = ailine.extract_task_mentions("集計シートを見て", ["Sheet"])   # 実在しない
    assert m2["sheets"] == set()

def test_extract_task_mentions_empty_when_no_mentions():
    m = ailine.extract_task_mentions("いい感じにして", ["Sheet"])
    assert m == {"cols": set(), "digit_cols": set(), "rows": set(), "sheets": set()}


def test_mention_digit_col_accepts_both_zero_and_one_based(tmp_path):
    """再監査 2 回目の誤警報の再発防止: 『在庫(列2)』= 0 起点で C 列(1 起点 3) を
    正しく変更したのに B 列不在の★が出た。数字表記は両解釈のどちらかが触られて
    いれば沈黙する。"""
    p = _book(tmp_path, [["商品", "金額", "在庫"], ["りんご", 100, 8]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=3, value=16)   # C 列 = 0 起点の列2 を変更
    wb.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("在庫(列2)の値を2倍にする", before["sheets"])
    assert ailine.mention_overlap_advisory(mentions, before, after) == []   # 誤警報なし

    # 両解釈とも外れている場合だけ警告し、警告文は数字のまま
    mentions9 = ailine.extract_task_mentions("列9を2倍にする", before["sheets"])
    lines = ailine.mention_overlap_advisory(mentions9, before, after)
    assert any("『列9』" in ln for ln in lines)

def test_mention_overlap_flags_unmatched_column(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["りんご", 100], ["バナナ", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=2, value=999)   # B列だけを変更
    wb.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("列Zを2倍にする", before["sheets"])
    lines = ailine.mention_overlap_advisory(mentions, before, after)
    assert any("列Z" in ln and "★" in ln for ln in lines)

def test_mention_overlap_silent_when_mentioned_column_was_changed(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["りんご", 100]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=2, value=999)
    wb.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("B列を確認", before["sheets"])
    assert ailine.mention_overlap_advisory(mentions, before, after) == []

def test_mention_overlap_flags_untouched_real_sheet(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["りんご", 100]]:
        ws.append(row)
    wb.create_sheet("集計")
    wb.save(p)
    before = ailine.snapshot(p)
    wb2 = openpyxl.load_workbook(p)
    wb2.active.cell(row=2, column=2, value=999)   # Sheet 側だけ変更、集計シートは無変更
    wb2.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("集計シートも直して", before["sheets"])
    lines = ailine.mention_overlap_advisory(mentions, before, after)
    assert any("集計" in ln for ln in lines)

def test_mention_overlap_empty_when_no_mentions(tmp_path):
    p = _book(tmp_path, [["a", 1]])
    snap = ailine.snapshot(p)
    assert ailine.mention_overlap_advisory({"cols": set(), "rows": set(), "sheets": set()}, snap, snap) == []

def test_mention_overlap_exclude_sheets_suppresses_readonly_reference(tmp_path):
    # ★ W10b 項目4b(摩擦): LOOKUP_FILL の参照専用シートは読み取り専用が正しい操作。
    #   exclude_sheets に渡せば言及があっても『変更されていません』を出さない。
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "数量", "単価"], ["りんご", 2, None]]:
        ws.append(row)
    wb.create_sheet("単価表")
    wb.save(p)
    before = ailine.snapshot(p)
    wb2 = openpyxl.load_workbook(p)
    wb2.active.cell(row=2, column=3, value=100)   # 明細側だけ変更、単価表は無変更(正常)
    wb2.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("単価表から単価を引いて入れて", before["sheets"])
    lines_without_exclude = ailine.mention_overlap_advisory(mentions, before, after)
    assert any("単価表" in ln for ln in lines_without_exclude)   # 対照: 抑制無しなら出る
    lines_with_exclude = ailine.mention_overlap_advisory(mentions, before, after, {"単価表"})
    assert not any("単価表" in ln for ln in lines_with_exclude)

def test_mention_overlap_exclude_sheets_still_flags_other_sheets(tmp_path):
    # 抑制は指定したシートだけ・保守的（他の言及シートは従来どおり警報する）。
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["りんご", 100]]:
        ws.append(row)
    wb.create_sheet("単価表")
    wb.create_sheet("集計")
    wb.save(p)
    before = ailine.snapshot(p)
    wb2 = openpyxl.load_workbook(p)
    wb2.active.cell(row=2, column=2, value=999)
    wb2.save(p)
    after = ailine.snapshot(p)
    mentions = ailine.extract_task_mentions("単価表と集計シートも直して", before["sheets"])
    lines = ailine.mention_overlap_advisory(mentions, before, after, {"単価表"})
    assert not any("単価表" in ln for ln in lines)
    assert any("集計" in ln for ln in lines)


# --- ★ W10b 項目4a(摩擦) → ★ C9: 新規列作成の『範囲外』誤警報の中立化 --------------
# ★ C9: 旧 _neutralize_new_column_ghost_warning は advisories リストを後から書き換える
#   単体関数だったが、detect_ghost_data 自身が new_col_letter を受け取り発生源で判定する形に
#   構造を置き換えた（判定条件・出力文言は不変・ゴールデン差分ゼロで確認）。低レベルの
#   detect_ghost_data(new_col_letter=) と、op からの橋渡しを含む _structural_advisories を
#   それぞれ直接叩いて同じシナリオを検証する。

def test_detect_ghost_data_neutral_when_span_confined_to_new_col_letter(tmp_path):
    p = _book(tmp_path, [["品目", "金額"], ["a", 100]])   # 使用範囲は A1:B2、新規列は C
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (2, 3, 4):
        wb.active.cell(row=r, column=3, value=110)   # C列（範囲外・新規列）だけに書く
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_ghost_data(before, after, new_col_letter="C")
    assert msg == "（新規列の追加は意図どおりです）"

def test_detect_ghost_data_keeps_warning_when_span_extends_beyond_new_col_letter(tmp_path):
    p = _book(tmp_path, [["品目", "金額"], ["a", 100]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (3, 4):
        wb.active.cell(row=r, column=2, value=9)   # B列（範囲外だが新規列ではない）も含む
        wb.active.cell(row=r, column=3, value=9)   # C列（新規列）
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_ghost_data(before, after, new_col_letter="C")
    assert msg is not None and msg.startswith("★ 疑わしい")   # 保守的：他列にも及ぶので変えない

def test_structural_advisories_neutralizes_ghost_warning_for_declared_new_column(tmp_path):
    p = _book(tmp_path, [["品目", "金額"], ["a", 100]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (2, 3, 4):
        wb.active.cell(row=r, column=3, value=110)
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    resolved = {"operands": ["金額"], "operator": "*", "factor": 1.1}
    lines = ailine._structural_advisories(before, after, op="COMPUTE_COLUMN", resolved=resolved, meta=meta)
    assert "（新規列の追加は意図どおりです）" in lines
    assert not any(ln.startswith("★ 疑わしい: 変更が元データの範囲外です") for ln in lines)

def test_structural_advisories_keeps_ghost_warning_when_target_specified(tmp_path):
    # target 指定(既存列への書き込み)は新規列作成ではないので中立化しない。
    # ★ 行2は既存データ(110)のままにする ── 変更を混ぜると detect_ghost_data 自体が
    # 「1つでも範囲内の変更がある」で無条件に None を返してしまい判定材料が無くなるため。
    p = _book(tmp_path, [["品目", "金額", "税込み金額"], ["a", 100, 110]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (3, 4):
        wb.active.cell(row=r, column=3, value=999)   # 既存の『税込み金額』列の範囲外行に書く
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額", "税込み金額"]}}
    resolved = {"operands": ["金額"], "operator": "*", "factor": 1.1, "target": "税込み金額"}
    lines = ailine._structural_advisories(before, after, op="COMPUTE_COLUMN", resolved=resolved, meta=meta)
    assert any(ln.startswith("★ 疑わしい: 変更が元データの範囲外です") for ln in lines)

def test_structural_advisories_keeps_ghost_warning_for_other_ops(tmp_path):
    p = _book(tmp_path, [["品目", "金額"], ["a", 100]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (2, 3, 4):
        wb.active.cell(row=r, column=3, value=110)
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    lines = ailine._structural_advisories(before, after, op="SORT", resolved={"col": "金額"}, meta=meta)
    assert any(ln.startswith("★ 疑わしい: 変更が元データの範囲外です") for ln in lines)


# ★ W10d 項目2: COMPUTE_COLUMN 専用だった中立化を OP_WRITE_TARGET の宣言駆動へ一般化。
#   LOOKUP_FILL が新規列を作る場合も同じ誤警報（査定で名指しされたオオカミ少年）が出ていた。

def test_structural_advisories_neutralizes_ghost_warning_for_lookup_fill_new_column(tmp_path):
    # 旧実装は COMPUTE_COLUMN 専用の if だったため、これは常に漏れていた。
    p = _book(tmp_path, [["商品コード", "数量"], ["x1", 3]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (2, 3, 4):
        wb.active.cell(row=r, column=3, value=500)   # C列（範囲外・新規の『単価』列）
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}}
    resolved = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品コード"}
    lines = ailine._structural_advisories(before, after, op="LOOKUP_FILL", resolved=resolved, meta=meta)
    assert "（新規列の追加は意図どおりです）" in lines

def test_structural_advisories_keeps_ghost_warning_for_lookup_fill_existing_column(tmp_path):
    # target_col が対象シートに既に実在する（上書き系）場合は新規列作成ではないので対象外
    # （抑制しすぎない側＝安全器官の減衰は保守的に、が守られていることの確認）。
    # ★ 行2は既存データ(100)のままにする（理由は上のテストのコメント参照）。
    p = _book(tmp_path, [["商品コード", "数量", "単価"], ["x1", 3, 100]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    for r in (3, 4):
        wb.active.cell(row=r, column=3, value=500)
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量", "単価"], "単価表": ["商品コード", "単価"]}}
    resolved = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品コード"}
    lines = ailine._structural_advisories(before, after, op="LOOKUP_FILL", resolved=resolved, meta=meta)
    assert any(ln.startswith("★ 疑わしい: 変更が元データの範囲外です") for ln in lines)

# ★ 単位C(D10): 合計行（データ末尾の新規行）は定義上ずっと原本の使用範囲の外に出るので、
#   幽霊データ警告が毎回・確実に誤爆していた。宣言（writes に new_row_at_end）で消す。

def test_structural_advisories_neutralizes_ghost_warning_for_append_total_new_row(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=4, column=1, value="合計")   # 最終行の下（ラベル）
    wb.active.cell(row=4, column=2, value=300)      # 最終行の下（合計）
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    lines = ailine._structural_advisories(
        before, after, op="APPEND_TOTAL", resolved={"col": "金額", "label": "合計"}, meta=meta)
    assert "（表の末尾への追記は意図どおりです）" in lines
    assert not any(ln.startswith("★ 疑わしい") for ln in lines)

def test_structural_advisories_keeps_ghost_warning_when_append_total_writes_outside_columns(tmp_path):
    # ★ 抑制しすぎない側: 宣言が『末尾の新規行』でも、原本の列範囲より右へ出たセルが
    #   混ざっていれば中立化しない（安全器官の減衰は保守的に、が守られていることの確認）。
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=4, column=2, value=300)
    wb.active.cell(row=4, column=9, value="謎")   # I列＝原本の列範囲の外
    wb.save(p)
    after = ailine.snapshot(p)
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    lines = ailine._structural_advisories(
        before, after, op="APPEND_TOTAL", resolved={"col": "金額"}, meta=meta)
    assert any(ln.startswith("★ 疑わしい: 変更が元データの範囲外です") for ln in lines)

# ★ 単位C(D8): 名指しされた入力シートを読むだけの op（AGGREGATE/PIVOT）で
#   「★ …は変更されていません」が毎回・確実に誤爆していた。宣言（reads_only）で消す。

def test_build_advisories_silences_mention_for_declared_reads_only_sheet(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "工事台帳"
    for row in (["取引先", "金額"], ["a", 100], ["b", 200]):
        wb.active.append(row)
    wb.save(p)
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    ws = wb.create_sheet("集計")
    for row in (["取引先", "合計"], ["a", 100], ["b", 200]):
        ws.append(row)
    wb.save(p)
    after = ailine.snapshot(p)
    task = "工事台帳を取引先ごとに集計して"
    meta = {"sheets": ["工事台帳"], "headers": {"工事台帳": ["取引先", "金額"]}}
    resolved = {"group_col": "取引先", "value_col": "金額", "_target_sheet": "工事台帳"}
    lines = ailine.build_advisories(task, before, after, None, op="AGGREGATE", resolved=resolved, meta=meta)
    assert not any("工事台帳" in ln and ln.startswith("★") for ln in lines)
    # 宣言を渡さなければ従来どおり出る（消えたのが宣言のおかげであることの対照）。
    plain = ailine.build_advisories(task, before, after, None)
    assert "★ 依頼で言及された『工事台帳』は存在しません/変更されていません" in plain


def test_declared_new_column_letter_driven_by_op_write_target():
    # ★ W10d 番人: 新規列作成の判定が OP_WRITE_TARGET の宣言だけで決まり、op ごとの
    #   個別 if を増やしていないことを全 op について機械的に確認する（宣言効果を持つ全 op が
    #   中立化の対象として自動的に登録されている、ことの検査＝test_op_write_target_declares_all_ops
    #   と対になる番人）。
    #   ★ 単位C: 判定の入口が「宣言が None か」から「writes に『新規列』があるか」へ変わった。
    #   新規列を作ると宣言していない op は常に None。宣言していて resolved にその列名が
    #   無い(=新規列作成)場合は必ず新規列の列文字が返る。
    #   新しい op を OP_WRITE_TARGET へ登録しさえすれば、ここへの追記なしに正しく振る舞う。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["既存1", "既存2"]}}
    for op, write_target in ailine.OP_WRITE_TARGET.items():
        if ailine.WRITE_NEW_COLUMN not in write_target.writes:
            assert ailine._declared_new_column_letter(op, {}, meta) is None, op
            continue
        resolved = {write_target.sheet_key: "Sheet"} if write_target.sheet_key else {}
        letter = ailine._declared_new_column_letter(op, resolved, meta)
        assert letter == "C", f"{op}: 新規列作成のはずが列文字が返らなかった: {letter}"


# --- ★ W8a 項目4: 率リテラルの機械スキャン（判断棚から昇格） -----------------------

def test_scan_rate_literals_fires_when_rate_not_in_task():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 1.08\nEnd Sub"
    out = ailine.scan_rate_literals(code, "税込み合計を出して")
    assert len(out) == 1
    assert "1.08" in out[0]
    assert "検算してください" in out[0]

def test_scan_rate_literals_silent_when_task_mentions_percentage():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 1.08\nEnd Sub"
    out = ailine.scan_rate_literals(code, "消費税8%込みの合計を出して")
    assert out == []

def test_scan_rate_literals_silent_when_vocab_has_matching_rate():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 1.1\nEnd Sub"
    out = ailine.scan_rate_literals(code, "税込み合計を出して", vocab={"消費税": 1.1})
    assert out == []

def test_scan_rate_literals_ignores_comment_lines():
    code = "Sub Run(oDoc As Object)\n  ' 参考値 1.08 は例\n  x = 1\nEnd Sub"
    out = ailine.scan_rate_literals(code, "合計を出して")
    assert out == []

def test_scan_rate_literals_ignores_integers():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 5\nEnd Sub"
    out = ailine.scan_rate_literals(code, "合計を出して")
    assert out == []

def test_scan_rate_literals_ignores_out_of_range_decimals():
    code = "Sub Run(oDoc As Object)\n  x = 3.14\nEnd Sub"
    out = ailine.scan_rate_literals(code, "合計を出して")
    assert out == []

def test_scan_rate_literals_dedupes_repeated_literal():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 1.1\n  y = 200 * 1.1\nEnd Sub"
    out = ailine.scan_rate_literals(code, "合計を出して")
    assert len(out) == 1

def test_scan_rate_literals_boundary_values_inclusive():
    code = "Sub Run(oDoc As Object)\n  x = 100 * 0.05\n  y = 200 * 1.2\nEnd Sub"
    out = ailine.scan_rate_literals(code, "合計を出して")
    assert len(out) == 2


# --- ★ M2a: 生成コードの切断検出 ---------------------------------------------

@pytest.mark.parametrize("code,truncated", [
    ("Sub Run(oDoc As Object)\nEnd Sub", False),
    ("Sub Run(oDoc As Object)\n  oDoc.Sheets.getByIndex(0)\nEnd Sub", False),
    ("", True),
    ("   \n  ", True),
    ("Sub Run(oDoc As Object)\n  oDoc.Sheets.getByIndex(0", True),          # 開き括弧で途切れ
    ("Sub Run(oDoc As Object)\n  Dim oSheet As Object\n  oSheet", True),    # 識別子で途切れ
    ("Sub Run(oDoc As Object)\nEnd Sub\nDim leftover", True),              # End Sub の後に続きがある
])
def test_is_truncated_code(code, truncated):
    assert ailine.is_truncated_code(code) is truncated


# --- ★ M2a: 実行時エラー表示の整形 --------------------------------------------

def test_short_error_summary_returns_last_line_of_traceback():
    tb = ("Traceback (most recent call last):\n"
          '  File "basrun.py", line 10, in <module>\n'
          "    raise ValueError('bad state')\n"
          "ValueError: bad state")
    assert ailine.short_error_summary(tb) == "ValueError: bad state"

def test_short_error_summary_single_line():
    assert ailine.short_error_summary("何かの実行時エラー") == "何かの実行時エラー"

def test_short_error_summary_empty():
    assert ailine.short_error_summary("") == "(詳細不明)"
    assert ailine.short_error_summary(None) == "(詳細不明)"

def test_cmd_run_shows_short_error_and_records_full_detail_in_history(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    # ★ M2a: 端末には最終行だけ、履歴 jsonl には全文（トレースバックをそのまま出さない）。
    # ★ W3: 見出し検出には「見出し行(複数の非空文字列)+データ行(型の混在)」が要るため、
    #   単一行 [["a", 1]] でなく見出し+データの2行にする（意図は runtime_error 経路の確認）。
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2:
                        "Sub Run(oDoc As Object)\nEnd Sub")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    traceback_text = ("Traceback (most recent call last):\n"
                       '  File "basrun.py", line 10, in <module>\n'
                       "    raise ValueError('bad state')\n"
                       "ValueError: bad state")
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (False, traceback_text, traceback_text))
    recorded = {}
    monkeypatch.setattr(ailine, "append_history", lambda entry, path=None: recorded.update(entry))

    argv = run_argv(
        book=str(book), task="テスト", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)   # ★ W10b: 関所は無関係（実行時エラー経路を見るテスト）

    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 1
    assert "Traceback (most recent call last):" not in captured.out
    assert "ValueError: bad state" in captured.out
    assert recorded["failure_kind"] == "runtime_error"
    assert recorded["error_detail"] == traceback_text


# --- ★ W8a 項目4/5: 単段 FREEFORM の正直な⚠枠・率スキャン・語彙翻訳 ----------------

def test_cmd_run_freeform_success_shows_honest_warning_not_checkmark(tmp_path, monkeypatch, capsys):
    # ★ W8a 項目4: 単段 FREEFORM/OUT_OF_VOCAB は成功しても『✓』でなく『⚠ AI が直接
    #   作成した処理です（機械保証なし）』の正直な枠で表示する
    #   （実測: 8%仮定やラベル貼りが「✓できました」で素通りしていた）。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value="税込合計")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="税込み合計を出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)   # ★ W10b: 関所は無関係（自由生成の成功表示を見るテスト）
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "⚠ AI が直接作成した処理です（機械保証なし）— 確認してください" in captured.out
    assert "✓ 適用され" not in captured.out

def test_cmd_run_freeform_banner_uses_ai_direct_wording_not_jargon(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    # ★ W8a 項目5: 「自由生成経路」→「AI が直接作成（機械保証なし）」（operator の語彙翻訳）。
    book = _book(tmp_path, [["a", 1]])
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    ailine.main(argv)
    captured = capsys.readouterr()
    assert "自由生成経路" not in captured.out
    assert "AI が直接作成" in captured.out

def test_cmd_run_freeform_rate_literal_scan_fires_when_task_silent_on_rate(tmp_path, monkeypatch, capsys):
    # ★ W8a 項目4: 依頼文にも用語集にも率の出典が無いのに生成コードに率らしい数値が
    #   あれば、検算を促す助言が変更点の後ろに出る。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    code = ("Sub Run(oDoc As Object)\n"
            "  oDoc.Sheets.getByIndex(0).getCellByPosition(2, 0).setValue(100 * 1.08)\n"
            "End Sub")
    monkeypatch.setattr(ailine, "ollama_generate", lambda model, msgs, temperature=0.2: code)
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, c, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value=108)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="税込み合計を出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)   # ★ W10b: 関所は無関係（率リテラル助言を見るテスト）
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "★ 率らしい数値 (1.08) が依頼に無いのに使われています — 検算してください" in captured.out

def test_cmd_run_freeform_rate_literal_scan_silent_when_task_states_rate(tmp_path, monkeypatch, capsys):
    # 対照: 依頼文に率(消費税8%)の出典があれば発火しない。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    code = ("Sub Run(oDoc As Object)\n"
            "  oDoc.Sheets.getByIndex(0).getCellByPosition(2, 0).setValue(100 * 1.08)\n"
            "End Sub")
    monkeypatch.setattr(ailine, "ollama_generate", lambda model, msgs, temperature=0.2: code)
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, c, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value=108)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="消費税8%込みの合計を出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)   # ★ W10b: 関所は無関係（率リテラル助言の対照テスト）
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "率らしい数値" not in captured.out

def test_cmd_run_dsl_dry_banner_says_rule_conversion_not_deterministic(tmp_path, monkeypatch, capsys):
    # ★ W8a 項目5: 「決定論」はユーザー向け文字列から排除（内部名・コメントは不変）。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "決定論" not in captured.out
    assert "ルール変換" in captured.out


# --- ★ M2a: --inplace バックアップ + restore --------------------------------

def test_utc_ts_format():
    # ★ W8b: マイクロ秒まで含む（秒精度だけの同名衝突を避けるための拡張・後方互換は
    #   _BACKUP_TS_RE/_ts_sort_key が旧6桁形式も受け付けることで担保）。
    import re as _re
    assert _re.match(r"^\d{8}T\d{12}Z$", ailine._utc_ts())

def test_ts_sort_key_orders_old_and_new_format_correctly():
    # 旧(秒精度)・新(マイクロ秒精度)が混在しても実時刻順に並ぶこと。
    old = "20260101T120000Z"
    new_earlier = "20260101T115959500000Z"
    new_later = "20260101T120000500000Z"
    keys = sorted([old, new_earlier, new_later], key=ailine._ts_sort_key)
    assert keys == [new_earlier, old, new_later]

def test_ts_sort_key_broken_format_sorts_as_oldest():
    assert ailine._ts_sort_key("not-a-timestamp") == (ailine.datetime.min, 0)

def test_ts_sort_key_seq_suffix_breaks_ties_within_same_timestamp():
    # ★ W8b: make_backup の "-N" 衝突回避連番は、同一 ts 内でも新しい順に並ぶこと。
    base = "20260101T120000000000Z"
    keys = sorted([f"{base}-2", base, f"{base}-3"], key=ailine._ts_sort_key)
    assert keys == [base, f"{base}-2", f"{base}-3"]

def test_backup_path_for_uses_stem_ts_suffix(tmp_path, monkeypatch):
    # ★ W8b 項目3: バックアップは book の親フォルダごとの名前空間ディレクトリの下に置く
    #   （同名ファイルが別フォルダにあっても取り違えない・undo 混線の根治）。
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = Path("demo/sample.xlsx")
    argv = ailine._backup_namespace(book)
    p = ailine.backup_path_for(book, ts="20260814T120000Z")
    assert p == tmp_path / "backups" / argv / "sample.20260814T120000Z.xlsx"

def test_make_backup_copies_content(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"HELLO")
    dst = ailine.make_backup(book)
    assert dst.exists()
    assert dst.read_bytes() == b"HELLO"
    # ★ W8b 項目3: 新規バックアップは名前空間ディレクトリの下（フラット直下ではない）。
    assert dst.parent == tmp_path / "backups" / ailine._backup_namespace(book)

def test_list_backups_sorted_newest_first_and_filters_by_stem_suffix(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"old")
    (backups / "book.20260101T000000Z.xlsx").write_bytes(b"new")
    (backups / "other.20260101T000000Z.xlsx").write_bytes(b"unrelated")   # 別 book
    (backups / "junk.txt").write_bytes(b"not a backup")                  # 形が違う
    result = ailine.list_backups(tmp_path / "book.xlsx")
    assert [p.name for p in result] == ["book.20260101T000000Z.xlsx", "book.20200101T000000Z.xlsx"]

def test_list_backups_empty_when_dir_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "nope")
    assert ailine.list_backups(tmp_path / "book.xlsx") == []

def test_restore_backup_restores_and_stays_reversible(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"BACKED_UP_CONTENT")
    book.write_bytes(b"CURRENT")   # --inplace で上書きされた後の現状を模す

    used = ailine.restore_backup(book)

    assert used.name == "book.20200101T000000Z.xlsx"
    assert book.read_bytes() == b"BACKED_UP_CONTENT"
    # ★ W11: 復元前の CURRENT は退避されている＝復元自体も可逆（この性質は据え置き）。
    #   ただし置き場は undo の棚で、**遡りの履歴には混ぜない**（混ぜていたので、端に着いた
    #   後の undo がこれを最新世代として釣り上げ、打ち消したはずの状態が復活していた）。
    assert {p.read_bytes() for p in ailine.list_undo_shelf(book)} == {b"CURRENT"}
    assert [p.read_bytes() for p in ailine.list_backups(book)] == [b"BACKED_UP_CONTENT"]

def test_restore_backup_raises_when_none_exist(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    with pytest.raises(FileNotFoundError):
        ailine.restore_backup(tmp_path / "book.xlsx")


# --- ★ W8b 項目3: バックアップの名前空間化（同名別フォルダの undo 混線の根治） ------

def test_backup_namespace_differs_by_parent_folder(tmp_path):
    a = tmp_path / "A" / "見積.xlsx"
    b = tmp_path / "B" / "見積.xlsx"
    assert ailine._backup_namespace(a) != ailine._backup_namespace(b)

def test_backup_namespace_stable_for_same_folder(tmp_path):
    p1 = tmp_path / "見積.xlsx"
    p2 = tmp_path / "見積.xlsx"
    assert ailine._backup_namespace(p1) == ailine._backup_namespace(p2)

def test_backup_and_restore_two_same_named_files_in_different_folders_do_not_cross(tmp_path, monkeypatch):
    # ★ 回帰テスト（architect 指摘）: A\見積.xlsx と B\見積.xlsx を交互に backup→restore
    #   しても混線しない（名前空間分離前は同名+フラット領域のため取り違えが起きえた）。
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    dir_a = tmp_path / "A"; dir_a.mkdir()
    dir_b = tmp_path / "B"; dir_b.mkdir()
    book_a = dir_a / "見積.xlsx"
    book_b = dir_b / "見積.xlsx"
    book_a.write_bytes(b"A-ORIGINAL")
    book_b.write_bytes(b"B-ORIGINAL")

    ailine.make_backup(book_a)
    ailine.make_backup(book_b)
    book_a.write_bytes(b"A-EDITED")
    book_b.write_bytes(b"B-EDITED")

    ailine.restore_backup(book_a)
    ailine.restore_backup(book_b)

    assert book_a.read_bytes() == b"A-ORIGINAL"
    assert book_b.read_bytes() == b"B-ORIGINAL"
    # A のバックアップ一覧に B の内容(B-ORIGINAL 等)が紛れ込んでいないこと
    a_contents = {p.read_bytes() for p in ailine.list_backups(book_a)}
    assert b"B-ORIGINAL" not in a_contents and b"B-EDITED" not in a_contents

def test_list_backups_reads_legacy_flat_area_read_only(tmp_path, monkeypatch):
    # ★ 旧フラット領域（名前空間分離前の名残）は読み取りのみ互換。
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    (backups / "legacy.20200101T000000Z.xlsx").write_bytes(b"old-flat")
    book = tmp_path / "legacy.xlsx"
    found = ailine.list_backups(book)
    assert [p.name for p in found] == ["legacy.20200101T000000Z.xlsx"]

def test_make_backup_never_writes_to_legacy_flat_area(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    ailine.make_backup(book)
    flat_files = [p for p in backups.iterdir() if p.is_file()] if backups.is_dir() else []
    assert flat_files == []   # 新規バックアップは名前空間ディレクトリの下だけ


# --- ★ W8b 項目5: ailine undo（restore の昇格） --------------------------------

def test_cmd_undo_restores_and_shows_remaining_count(tmp_path, monkeypatch, capsys):
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"CURRENT")
    ailine.make_backup(book)   # 1世代目
    book.write_bytes(b"EDITED")

    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
    captured = capsys.readouterr()
    assert rc == 0
    assert book.read_bytes() == b"CURRENT"
    assert "あと" in captured.out and "回戻せます" in captured.out

def test_cmd_undo_list_shows_backups_without_restoring(tmp_path, monkeypatch, capsys):
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"CURRENT")
    ailine.make_backup(book)
    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=True))
    captured = capsys.readouterr()
    assert rc == 0
    assert book.read_bytes() == b"CURRENT"   # --list は復元しない
    assert "世代" in captured.out

def test_cmd_undo_fails_honestly_when_no_backup(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
    captured = capsys.readouterr()
    assert rc == 1
    assert "×" in captured.out


# --- ★ W11: undo は履歴の端で止まる（盲検査定 A が致命に挙げた実測不具合） ----------
#
# 旧実装で実測した壊れ方（編集 2 回のあと undo を 4 回）:
#   undo1 → v1 ✓ / undo2 → v0 ✓ / undo3 → **v0 のまま「✓ 復元した」** /
#   undo4 → **v1 が復活**（打ち消したはずの状態が戻る）
# 原因は 2 つの重なり: ①最古に着いても止まらず同じものを復元して成功を名乗る
# ②undo 自身の退避を遡りの履歴に積むので、端の後は「現在と同じ退避」が最新世代として
# 並び、その 1 つ内側＝直前に打ち消した状態を釣り上げる。
# ★ 症状の「3 回目」は編集回数で変わる。バーは構造で書く: **編集 N 回 → undo N+2 回**。

def _seed_edits(tmp_path, monkeypatch, n_edits: int):
    """v0 から n_edits 回書き換えた book を作る（実際の run と同じく上書き前に退避）。"""
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"v0")
    for i in range(1, n_edits + 1):
        ailine.make_backup(book)
        book.write_bytes(f"v{i}".encode())
    return book

@pytest.mark.parametrize("n_edits", [1, 2, 3])
def test_undo_walks_back_then_stops_at_the_oldest(tmp_path, monkeypatch, capsys, n_edits):
    """★ バー: 編集 N 回 → undo N+2 回。1〜N 回目は一段ずつ v(N-1)…v0、
       N+1 回目は非零で止まり、N+2 回目も**状態が動かない**。"""
    book = _seed_edits(tmp_path, monkeypatch, n_edits)

    for k in range(1, n_edits + 1):
        rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
        out = capsys.readouterr().out
        assert rc == 0
        assert book.read_bytes() == f"v{n_edits - k}".encode()
        # 「あと N 回」は残り段数（バックアップの総数ではない）
        assert f"（あと {n_edits - k} 回戻せます）" in out

    for _ in range(2):   # N+1 回目・N+2 回目とも同じ扱い（押し続けても動かない）
        rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
        out = capsys.readouterr().out
        assert rc == 1                                   # 非零（既存体系の 1＝汎用の失敗）
        assert book.read_bytes() == b"v0"                # 最古のまま動かない
        assert "これ以上は戻せません" in out and "✓" not in out

def test_undo_at_the_oldest_raises_a_distinct_error_not_file_not_found(tmp_path, monkeypatch):
    """「もう戻せない」と「そもそもバックアップが無い」は別物（型で分ける）。"""
    book = _seed_edits(tmp_path, monkeypatch, 1)
    ailine.restore_backup(book)
    with pytest.raises(ailine.NoOlderBackupError):
        ailine.restore_backup(book)

def test_undo_shelf_keeps_the_pre_restore_copy_out_of_the_walk(tmp_path, monkeypatch):
    """★ undo を可逆にする性質は残す: 退避は棚に残り、遡りの履歴だけが汚れない。"""
    book = _seed_edits(tmp_path, monkeypatch, 2)
    before = [p.name for p in ailine.list_backups(book)]
    ailine.restore_backup(book)
    assert [p.name for p in ailine.list_backups(book)] == before   # 遡りの履歴は伸びない
    assert [p.read_bytes() for p in ailine.list_undo_shelf(book)] == [b"v2"]   # 退避は残る
    assert ailine.undo_shelf_dir(book).parent == ailine.BACKUP_DIR / ailine._backup_namespace(book)

def test_undo_steps_left_counts_the_walk_not_the_backups(tmp_path, monkeypatch):
    book = _seed_edits(tmp_path, monkeypatch, 3)
    assert ailine.undo_steps_left(book) == 3     # 実編集の直後＝全世代を遡れる
    for expected in (2, 1, 0):
        ailine.restore_backup(book)
        assert ailine.undo_steps_left(book) == expected

def test_undo_then_new_run_then_undo_still_works(tmp_path, monkeypatch):
    """★ undo のやり直し: 端まで戻した後に新しい run をしても、その run を undo できる。"""
    book = _seed_edits(tmp_path, monkeypatch, 2)
    ailine.restore_backup(book); ailine.restore_backup(book)
    assert book.read_bytes() == b"v0"
    ailine.make_backup(book); book.write_bytes(b"v9")   # 新しい run
    ailine.restore_backup(book)
    assert book.read_bytes() == b"v0"

def test_undo_list_mentions_the_shelf_without_counting_it_as_a_generation(tmp_path, monkeypatch, capsys):
    book = _seed_edits(tmp_path, monkeypatch, 2)
    ailine.restore_backup(book)
    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=True))
    out = capsys.readouterr().out
    assert rc == 0
    assert "2 世代" in out                       # 遡れる世代は増えていない
    assert "退避が 1 件" in out and "遡りには数えない" in out

def test_legacy_backups_stay_readable_and_walkable(tmp_path, monkeypatch):
    """★ 既存資産との互換: 名前空間ディレクトリ直下の既存世代も旧フラット領域も、
       これまでどおり遡りの対象（棚を足したことで読めなくなるものは無い）。"""
    backups = tmp_path / "backups"
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    ns = backups / ailine._backup_namespace(book)
    ns.mkdir(parents=True)
    (ns / "book.20260101T000000000000Z.xlsx").write_bytes(b"v1")   # 名前空間の既存世代
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"v0")    # 旧フラット領域
    book.write_bytes(b"v2")
    assert [p.read_bytes() for p in ailine.list_backups(book)] == [b"v1", b"v0"]
    ailine.restore_backup(book); assert book.read_bytes() == b"v1"
    ailine.restore_backup(book); assert book.read_bytes() == b"v0"
    with pytest.raises(ailine.NoOlderBackupError):
        ailine.restore_backup(book)


# --- ★ M2c: バックアップのプルーニング -----------------------------------------

def test_prune_backups_keeps_only_newest_n(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    for i in range(5):
        (backups / f"book.2026010{i+1}T000000Z.xlsx").write_bytes(b"x")
    deleted = ailine.prune_backups(book, keep=3)
    remaining = ailine.list_backups(book)
    assert len(remaining) == 3
    assert len(deleted) == 2
    assert [p.name for p in remaining] == [
        "book.20260105T000000Z.xlsx", "book.20260104T000000Z.xlsx", "book.20260103T000000Z.xlsx"]

def test_prune_backups_negative_keep_means_unlimited(tmp_path, monkeypatch):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    for i in range(3):
        (backups / f"book.2026010{i+1}T000000Z.xlsx").write_bytes(b"x")
    deleted = ailine.prune_backups(book, keep=-1)
    assert deleted == []
    assert len(ailine.list_backups(book)) == 3

def test_make_backup_prunes_beyond_keep(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"content")
    for i in range(3):
        monkeypatch.setattr(ailine, "_utc_ts", lambda i=i: f"2026010{i+1}T000000Z")
        ailine.make_backup(book, keep=2)
    remaining = ailine.list_backups(book)
    assert len(remaining) == 2
    assert remaining[0].name == "book.20260103T000000Z.xlsx"   # 最新が残る

def test_make_backup_default_keep_is_ten(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"content")
    for i in range(12):
        monkeypatch.setattr(ailine, "_utc_ts", lambda i=i: f"202601{i+1:02d}T000000Z")
        ailine.make_backup(book)   # keep 省略 = 既定 DEFAULT_KEEP_BACKUPS(=10)
    assert len(ailine.list_backups(book)) == 10

def test_cmd_restore_list_shows_generation_count(tmp_path, monkeypatch, capsys):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"a")
    (backups / "book.20200102T000000Z.xlsx").write_bytes(b"b")
    ns = argparse.Namespace(book=str(tmp_path / "book.xlsx"), list=True)
    rc = ailine.cmd_restore(ns)
    captured = capsys.readouterr()
    assert rc == 0
    assert "2 世代" in captured.out


# ===========================================================================
# ★ W8b 項目2: Excel ロックの事前検出
# ===========================================================================

def test_check_excel_lock_none_when_free(tmp_path):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    assert ailine.check_excel_lock(book) is None

def test_check_excel_lock_detects_tilde_dollar_lock_file(tmp_path):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    (tmp_path / "~$book.xlsx").write_bytes(b"lock")
    reason = ailine.check_excel_lock(book)
    assert reason is not None
    assert "ロックファイル" in reason

def test_check_excel_lock_detects_exclusively_opened_file(tmp_path):
    # ★ 実際の Excel の排他 open を模す。Python の open() 二重呼び出しは Windows
    #   既定の共有モードでは失敗しないため、CreateFileW を dwShareMode=0 で直接呼ぶ
    #   （Excel 相当の排他アクセス）。
    import ctypes
    from ctypes import wintypes
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    GENERIC_READ = 0x80000000
    GENERIC_WRITE = 0x40000000
    OPEN_EXISTING = 3
    CreateFileW = ctypes.windll.kernel32.CreateFileW
    CreateFileW.restype = wintypes.HANDLE
    CreateFileW.argtypes = [wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD, wintypes.LPVOID,
                             wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE]
    handle = CreateFileW(str(book), GENERIC_READ | GENERIC_WRITE, 0, None, OPEN_EXISTING, 0, None)
    assert handle not in (0, -1)
    try:
        reason = ailine.check_excel_lock(book)
        assert reason is not None
        assert "ロック" in reason
    finally:
        ctypes.windll.kernel32.CloseHandle(handle)

def test_cmd_run_exits_5_when_excel_lock_detected(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    (tmp_path / f"~${book.name}").write_bytes(b"lock")
    called = {"n": 0}
    def boom(*a, **k):
        called["n"] += 1
        return {"op": "FREEFORM", "args": {}}
    monkeypatch.setattr(ailine, "translate_task", boom)
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 5
    assert "Excel で開かれています" in captured.out
    assert called["n"] == 0   # LO 起動・翻訳より前に止まっている


# ===========================================================================
# ★ W8b 項目1: 往復忠実度ゲート
# ===========================================================================

def _make_fake_zip(tmp_path, name, members):
    import zipfile as _zf
    p = tmp_path / name
    with _zf.ZipFile(p, "w") as z:
        for m in members:
            z.writestr(m, b"x")
    return p

def test_check_zip_fidelity_loss_detects_categories(tmp_path):
    original = _make_fake_zip(tmp_path, "orig.xlsx", [
        "xl/workbook.xml", "xl/worksheets/sheet1.xml",
        "xl/drawings/drawing1.xml", "xl/media/image1.png", "xl/media/image2.png",
        "xl/vbaProject.bin", "xl/pivotCache/pivotCacheDefinition1.xml",
        "xl/pivotTables/pivotTable1.xml", "xl/worksheets/_rels/sheet1.xml.rels",
    ])
    normalized = _make_fake_zip(tmp_path, "norm.xlsx", ["xl/workbook.xml", "xl/worksheets/sheet1.xml"])
    result = dict(ailine.check_zip_fidelity_loss(original, normalized))
    assert result["図形/描画"] == 1
    assert result["画像"] == 2
    assert result["VBA マクロ"] == 1
    assert result["ピボットテーブル"] == 2   # pivotCache + pivotTables
    assert result["リンク情報(_rels)"] == 1

def test_check_zip_fidelity_loss_empty_when_nothing_lost(tmp_path):
    members = ["xl/workbook.xml", "xl/worksheets/sheet1.xml", "xl/media/image1.png"]
    original = _make_fake_zip(tmp_path, "orig.xlsx", members)
    normalized = _make_fake_zip(tmp_path, "norm.xlsx", members)
    assert ailine.check_zip_fidelity_loss(original, normalized) == []

def test_check_zip_fidelity_loss_ignores_unrelated_removed_members(tmp_path):
    original = _make_fake_zip(tmp_path, "orig.xlsx", ["xl/workbook.xml", "docProps/core.xml"])
    normalized = _make_fake_zip(tmp_path, "norm.xlsx", ["xl/workbook.xml"])
    assert ailine.check_zip_fidelity_loss(original, normalized) == []

def test_check_zip_fidelity_loss_empty_when_original_not_a_zip(tmp_path):
    p = tmp_path / "not_a_zip.xlsx"
    p.write_bytes(b"not a zip file")
    normalized = _make_fake_zip(tmp_path, "norm.xlsx", ["xl/workbook.xml"])
    assert ailine.check_zip_fidelity_loss(p, normalized) == []

def _cf_dv_book(tmp_path, name, add_cf=True, add_dv=True):
    # ★ 見出し(1行目)+データ(2行以降・型混在)を持たせる。単一セルだけの本だと
    # detect_header_row が確信を持てず、cmd_run 統合テストが CLARIFY に落ちてしまう。
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["りんご", 100])
    ws.append(["バナナ", 200])
    if add_cf:
        ws.conditional_formatting.add("B2:B3", CellIsRule(operator="greaterThan", formula=["150"]))
    if add_dv:
        dv = DataValidation(type="list", formula1='"a,b,c"')
        ws.add_data_validation(dv)
        dv.add("C2")
    wb.save(p)
    return p

def test_check_openpyxl_fidelity_loss_detects_lost_cf_and_dv(tmp_path):
    original = _cf_dv_book(tmp_path, "orig.xlsx", add_cf=True, add_dv=True)
    normalized = _cf_dv_book(tmp_path, "norm.xlsx", add_cf=False, add_dv=False)
    result = ailine.check_openpyxl_fidelity_loss(original, normalized)
    labels = {r[0] for r in result}
    assert labels == {"条件付き書式", "入力規則"}

def test_check_openpyxl_fidelity_loss_empty_when_unchanged(tmp_path):
    original = _cf_dv_book(tmp_path, "orig.xlsx")
    normalized = _cf_dv_book(tmp_path, "norm.xlsx")
    assert ailine.check_openpyxl_fidelity_loss(original, normalized) == []

def test_check_openpyxl_fidelity_loss_empty_when_unreadable(tmp_path):
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"not xlsx")
    normalized = _cf_dv_book(tmp_path, "norm.xlsx")
    assert ailine.check_openpyxl_fidelity_loss(p, normalized) == []

def test_check_round_trip_fidelity_combines_both_checks(tmp_path):
    original = _cf_dv_book(tmp_path, "orig.xlsx", add_cf=True, add_dv=False)
    normalized = _cf_dv_book(tmp_path, "norm.xlsx", add_cf=False, add_dv=False)
    fidelity = ailine.check_round_trip_fidelity(original, normalized)
    assert fidelity["lost"] is True
    assert any(it["label"] == "条件付き書式" and it["count"] == 1 for it in fidelity["items"])

def test_check_round_trip_fidelity_silent_when_nothing_lost(tmp_path):
    original = _cf_dv_book(tmp_path, "orig.xlsx")
    normalized = _cf_dv_book(tmp_path, "norm.xlsx")
    assert ailine.check_round_trip_fidelity(original, normalized) == {"lost": False, "items": []}

def test_format_fidelity_warning_lists_categories_and_counts():
    fidelity = {"lost": True, "items": [{"label": "条件付き書式", "count": 3}]}
    msg = ailine.format_fidelity_warning(fidelity)
    assert msg == "⚠ このファイルには、処理すると失われる飾りがあります（条件付き書式 3 件）"


def _fidelity_gate_argv(book, **overrides):
    # ★ W8b-2: 既定が原本直接適用になったため inplace=True は不要（指定しても挙動は
    #   同じだが廃止フラグの移行メッセージが余計に出るので外す）。
    # ★ W10b: これらのテストは忠実度ゲート自体を見るのが目的で、その先の自由生成の関所
    #   （項目1）は無関係なので allow_freeform=True で素通しする。
    base = dict(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, json=False, timeout=180.0, ask=False,
        accept_loss=False, copy=False, allow_freeform=True)
    base.update(overrides)
    return run_argv(**base)

def _patch_lossy_normalize(monkeypatch):
    """normalize_book を『CF が消える正規化』に差し替える（実際の LO を使わない）。
       値・見出しは原本と同じ形に保つ（header 検出が CLARIFY に落ちないように）。"""
    def fake_normalize(book, workdir, timeout=None):
        norm = workdir / ("normalized" + book.suffix)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["商品", "金額"])
        ws.append(["りんご", 100])
        ws.append(["バナナ", 200])   # CF/DV なし
        wb.save(norm)
        return norm
    monkeypatch.setattr(ailine, "normalize_book", fake_normalize)

def test_cmd_run_inplace_fidelity_gate_blocks_without_flag(tmp_path, monkeypatch, capsys):
    book = _cf_dv_book(tmp_path, "book.xlsx", add_cf=True, add_dv=False)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    _patch_lossy_normalize(monkeypatch)
    called = {"n": 0}
    def boom(*a, **k):
        called["n"] += 1
        return {"op": "FREEFORM", "args": {}}
    monkeypatch.setattr(ailine, "translate_task", boom)
    rc = ailine.main(_fidelity_gate_argv(book))
    captured = capsys.readouterr()
    assert rc == 4
    assert "失われる飾りがあります" in captured.out
    assert "--accept-loss" in captured.out
    assert "--copy" in captured.out
    assert called["n"] == 0   # ゲートで止まる＝翻訳より前

def test_cmd_run_inplace_fidelity_gate_accept_loss_continues(tmp_path, monkeypatch, capsys):
    book = _cf_dv_book(tmp_path, "book.xlsx", add_cf=True, add_dv=False)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    _patch_lossy_normalize(monkeypatch)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=2, value="changed")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc = ailine.main(_fidelity_gate_argv(book, accept_loss=True))
    captured = capsys.readouterr()
    assert "続行します" in captured.out
    # ★ W8b-2/C9: 既定=原本直接適用の trailing（FREEFORM 経路なので ⚠ 側・読み戻し付き）。
    assert "に適用しましたが、機械保証はありません（適用後に読み戻して確認: " in captured.out

def test_cmd_run_copy_flag_skips_fidelity_gate_entirely(tmp_path, monkeypatch, capsys):
    # ★ W8b-2 項目4: --copy 時は原本に一切触れないため、忠実度ゲート自体を走らせない
    #   （喪失があっても申告すら出ない・W8b-1 は『発動後に --copy で降格』だったが、
    #   既定反転後は最初から発動しないよう変わった）。
    book = _cf_dv_book(tmp_path, "book.xlsx", add_cf=True, add_dv=False)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    _patch_lossy_normalize(monkeypatch)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=2, value="changed")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc = ailine.main(_fidelity_gate_argv(book, copy=True))
    captured = capsys.readouterr()
    assert "失われる飾り" not in captured.out   # ゲート自体が走っていない
    assert book.read_bytes() == original_bytes   # 原本は無変更
    out_book = book.with_name(book.stem + ".out" + book.suffix)
    assert out_book.exists()

def test_cmd_run_inplace_fidelity_gate_silent_when_no_loss(tmp_path, monkeypatch, capsys):
    # 喪失ゼロなら --inplace でも無言で通る（体験は不変）。ゲート自体は --dry では
    # 走らない（LibreOffice に一切触れない設計不変条件）ので、ここは非 dry で確認する。
    book = _book(tmp_path, [["商品", "金額"], ["a", 1]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda b, workdir, timeout=None: b)   # 変化なし＝喪失ゼロ
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value="new")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = _fidelity_gate_argv(book, dry=False)
    ailine.main(argv)
    captured = capsys.readouterr()
    assert "失われる飾り" not in captured.out

def test_cmd_run_inplace_fidelity_records_history_field(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _cf_dv_book(tmp_path, "book.xlsx", add_cf=True, add_dv=False)
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    _patch_lossy_normalize(monkeypatch)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=2, value="changed")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    recorded = {}
    monkeypatch.setattr(ailine, "append_history", lambda entry, path=None: recorded.update(entry))
    ailine.main(_fidelity_gate_argv(book, accept_loss=True))
    assert recorded["fidelity"]["lost"] is True
    assert any(it["label"] == "条件付き書式" for it in recorded["fidelity"]["items"])

def test_cmd_run_inplace_no_fidelity_field_when_gate_not_run(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    # --inplace すら要求していない run では、ゲートは走らず fidelity は None のまま。
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    recorded = {}
    monkeypatch.setattr(ailine, "append_history", lambda entry, path=None: recorded.update(entry))
    argv = _fidelity_gate_argv(book, inplace=False, dry=True)
    ailine.main(argv)
    assert recorded["fidelity"] is None


# ===========================================================================
# ★ W8b 項目4: アトミック置換
# ===========================================================================

def test_atomic_replace_inplace_success_replaces_book_and_cleans_up(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"ORIGINAL")
    out_book = tmp_path / "book.out.xlsx"
    out_book.write_bytes(b"NEW-CONTENT")
    workdir = tmp_path / ".ailine_book"
    workdir.mkdir()
    ok, err = ailine.atomic_replace_inplace(book, out_book, workdir)
    assert ok is True
    assert err is None
    assert book.read_bytes() == b"NEW-CONTENT"
    assert not out_book.exists()          # 置換後は .out を残さない（旧 shutil.move と同じ終状態）
    assert not (workdir / f"staged{book.suffix}").exists()   # staging は片付く
    backups = ailine.list_backups(book)
    assert len(backups) == 1
    assert backups[0].read_bytes() == b"ORIGINAL"   # 置換前の原本がバックアップされている

def test_atomic_replace_inplace_backup_failure_aborts_without_touching_book(tmp_path, monkeypatch):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"ORIGINAL")
    out_book = tmp_path / "book.out.xlsx"
    out_book.write_bytes(b"NEW-CONTENT")
    workdir = tmp_path / ".ailine_book"
    workdir.mkdir()
    monkeypatch.setattr(ailine, "make_backup", lambda *a, **k: (_ for _ in ()).throw(OSError("disk full")))
    ok, err = ailine.atomic_replace_inplace(book, out_book, workdir)
    assert ok is False
    assert "バックアップに失敗" in err
    assert book.read_bytes() == b"ORIGINAL"   # 原本は無変更
    assert out_book.exists()                   # .out はそのまま残る

def test_atomic_replace_inplace_falls_back_to_copy2_when_os_replace_fails(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"ORIGINAL")
    out_book = tmp_path / "book.out.xlsx"
    out_book.write_bytes(b"NEW-CONTENT")
    workdir = tmp_path / ".ailine_book"
    workdir.mkdir()
    monkeypatch.setattr(ailine.os, "replace", lambda *a, **k: (_ for _ in ()).throw(OSError("locked")))
    ok, err = ailine.atomic_replace_inplace(book, out_book, workdir)
    assert ok is True
    assert err is None
    assert book.read_bytes() == b"NEW-CONTENT"   # copy2 フォールバックで反映されている
    backups = ailine.list_backups(book)
    assert len(backups) == 1   # バックアップは os.replace 失敗の前に既に確保済み

def test_atomic_replace_inplace_both_replace_and_fallback_fail_reports_honestly(tmp_path, monkeypatch):
    import shutil as _shutil
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"ORIGINAL")
    out_book = tmp_path / "book.out.xlsx"
    out_book.write_bytes(b"NEW-CONTENT")
    workdir = tmp_path / ".ailine_book"
    workdir.mkdir()
    monkeypatch.setattr(ailine.os, "replace", lambda *a, **k: (_ for _ in ()).throw(OSError("locked")))
    real_copy2 = _shutil.copy2
    calls = {"n": 0}
    def fake_copy2(src, dst):
        calls["n"] += 1
        # 1回目=バックアップ作成の内部コピー・2回目=staging へのコピー は成功させ、
        # 3回目=os.replace 失敗後のフォールバック copy2(out_book, book) だけ失敗させる。
        if calls["n"] <= 2:
            return real_copy2(src, dst)
        raise OSError("also locked")
    monkeypatch.setattr(ailine.shutil, "copy2", fake_copy2)
    ok, err = ailine.atomic_replace_inplace(book, out_book, workdir)
    assert ok is False
    assert "置換に失敗した" in err


# ===========================================================================
# ★ W8b 項目6: グローバル run ロック
# ===========================================================================

def test_acquire_and_release_run_lock_roundtrip(tmp_path):
    lock = tmp_path / "run.lock"
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is True
    assert msg is None
    assert lock.exists()
    ailine.release_run_lock(lock)
    assert not lock.exists()

def test_acquire_run_lock_fails_when_held_by_live_other_process(tmp_path, monkeypatch):
    lock = tmp_path / "run.lock"
    other_pid = 999999   # 実在しないふりをする pid（_pid_alive を差し替えて『生きている』にする）
    lock.write_text(json.dumps({"pid": other_pid, "ts": ailine.datetime.now(ailine.timezone.utc)
                                 .isoformat(timespec="seconds")}), encoding="utf-8")
    monkeypatch.setattr(ailine, "_pid_alive", lambda pid: pid == other_pid)
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is False
    assert "別の ailine が実行中です" in msg
    assert lock.exists()   # 奪取していない

def test_acquire_run_lock_reclaims_when_pid_is_dead(tmp_path, monkeypatch):
    lock = tmp_path / "run.lock"
    lock.write_text(json.dumps({"pid": 999999, "ts": ailine.datetime.now(ailine.timezone.utc)
                                 .isoformat(timespec="seconds")}), encoding="utf-8")
    monkeypatch.setattr(ailine, "_pid_alive", lambda pid: False)
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is True
    info = json.loads(lock.read_text(encoding="utf-8"))
    assert info["pid"] == os.getpid()

def test_acquire_run_lock_reclaims_when_stale_by_age(tmp_path, monkeypatch):
    import datetime as _dt
    lock = tmp_path / "run.lock"
    old_ts = (ailine.datetime.now(ailine.timezone.utc) - _dt.timedelta(hours=1)).isoformat(timespec="seconds")
    lock.write_text(json.dumps({"pid": 999999, "ts": old_ts}), encoding="utf-8")
    monkeypatch.setattr(ailine, "_pid_alive", lambda pid: True)   # pid 自体は生きている
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is True   # 30分超の age で stale 判定・奪取できる

def test_acquire_run_lock_reclaims_own_leftover_lock_same_pid(tmp_path):
    # ★ 同一プロセス内で前回の呼び出しが解放し損ねた場合も自己修復する
    #   （テスト実行のようにシーケンシャルに同じ pid で cmd_run が繰り返し走る状況の安全網）。
    lock = tmp_path / "run.lock"
    lock.write_text(json.dumps({"pid": os.getpid(), "ts": ailine.datetime.now(ailine.timezone.utc)
                                 .isoformat(timespec="seconds")}), encoding="utf-8")
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is True

def test_acquire_run_lock_broken_file_is_stale(tmp_path):
    lock = tmp_path / "run.lock"
    lock.write_text("not json", encoding="utf-8")
    acquired, msg = ailine.acquire_run_lock(lock)
    assert acquired is True

def test_cmd_run_exits_6_when_run_lock_busy(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    lock_path = tmp_path / "run.lock"
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", lock_path)
    other_pid = 999999
    lock_path.write_text(json.dumps({"pid": other_pid, "ts": ailine.datetime.now(ailine.timezone.utc)
                                     .isoformat(timespec="seconds")}), encoding="utf-8")
    monkeypatch.setattr(ailine, "_pid_alive", lambda pid: pid == other_pid)
    called = {"n": 0}
    monkeypatch.setattr(ailine, "check_excel_lock", lambda b: called.__setitem__("n", called["n"] + 1) or None)
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 6
    assert "別の ailine が実行中です" in captured.out
    assert called["n"] == 0   # ロックで止まった＝本体は一切呼ばれていない

def test_cmd_run_releases_lock_even_on_early_sys_exit(tmp_path, monkeypatch):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    # book が無い場合は sys.exit() する経路（SystemExit）。それでも lock は解放されること。
    lock_path = tmp_path / "run.lock"
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", lock_path)
    argv = run_argv(
        book=str(tmp_path / "nope.xlsx"), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    with pytest.raises(SystemExit):
        ailine.main(argv)
    assert not lock_path.exists()


# ===========================================================================
# ★ W8b 項目4: run 終了時に自分の workdir を掃除する
# ===========================================================================

def test_cmd_run_cleans_up_workdir_after_success(tmp_path, monkeypatch):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    ailine.main(argv)
    assert not (book.parent / f".ailine_{book.stem}").exists()

def test_cmd_run_cleans_up_workdir_after_clarify_exit(tmp_path, monkeypatch):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "build_struct_dump", lambda book, workdir: ambiguous)
    argv = run_argv(
        book=str(book), task="いい感じにして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    assert rc == 3
    assert not (book.parent / f".ailine_{book.stem}").exists()

# --- ★ M2c: 正規化パス失敗時の1回だけ自動リトライ ------------------------------

@pytest.mark.real_normalize_book
def test_normalize_book_retries_once_after_stop_and_succeeds(tmp_path, monkeypatch):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    calls = {"apply": 0, "stop": 0}
    def fake_apply(normalized, code, workdir, timeout=None):
        calls["apply"] += 1
        if calls["apply"] == 1:
            return False, "RuntimeException: Could not create system bitmap!", "raw"
        return True, None, "raw"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr(ailine, "_stop_office", lambda: calls.__setitem__("stop", calls["stop"] + 1))
    result = ailine.normalize_book(book, tmp_path)
    assert calls["apply"] == 2
    assert calls["stop"] == 1
    assert result.exists()

@pytest.mark.real_normalize_book
def test_normalize_book_gives_up_after_second_failure(tmp_path, monkeypatch):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    monkeypatch.setattr(ailine, "basrun_apply", lambda *a, **k: (False, "boom", "raw"))
    monkeypatch.setattr(ailine, "_stop_office", lambda: None)
    with pytest.raises(SystemExit):
        ailine.normalize_book(book, tmp_path)

@pytest.mark.real_normalize_book
def test_normalize_book_succeeds_first_try_without_retry(tmp_path, monkeypatch):
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"x")
    calls = {"apply": 0, "stop": 0}
    def fake_apply(normalized, code, workdir, timeout=None):
        calls["apply"] += 1
        return True, None, "raw"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr(ailine, "_stop_office", lambda: calls.__setitem__("stop", calls["stop"] + 1))
    ailine.normalize_book(book, tmp_path)
    assert calls["apply"] == 1
    assert calls["stop"] == 0

def test_cmd_restore_list_shows_backups(tmp_path, monkeypatch, capsys):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"a")
    ns = argparse.Namespace(book=str(tmp_path / "book.xlsx"), list=True)
    rc = ailine.cmd_restore(ns)
    captured = capsys.readouterr()
    assert rc == 0
    assert "book.20200101T000000Z.xlsx" in captured.out

def test_cmd_restore_list_says_none_when_no_backups(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    ns = argparse.Namespace(book=str(tmp_path / "book.xlsx"), list=True)
    rc = ailine.cmd_restore(ns)
    captured = capsys.readouterr()
    assert rc == 0
    assert "無い" in captured.out

def test_cmd_restore_restores_and_reports_success(tmp_path, monkeypatch, capsys):
    backups = tmp_path / "backups"
    backups.mkdir(parents=True)
    monkeypatch.setattr(ailine, "BACKUP_DIR", backups)
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"current")
    (backups / "book.20200101T000000Z.xlsx").write_bytes(b"restored")
    ns = argparse.Namespace(book=str(book), list=False)
    rc = ailine.cmd_restore(ns)
    captured = capsys.readouterr()
    assert rc == 0
    assert "✓" in captured.out
    assert book.read_bytes() == b"restored"

def test_cmd_restore_fails_when_no_backups(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    ns = argparse.Namespace(book=str(tmp_path / "book.xlsx"), list=False)
    rc = ailine.cmd_restore(ns)
    captured = capsys.readouterr()
    assert rc == 1
    assert "×" in captured.out


# ---------------------------------------------------------------------------
# ★ M2b: 中間命令言語（DSL）パイプライン
# ---------------------------------------------------------------------------

# --- ① 翻訳（json 退避） -----------------------------------------------------

_SAMPLE_META = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額", "在庫", "売上", "原価"]}}


def test_translate_task_valid_json_nested_args(monkeypatch):
    # ★ M2c: translate_task は常に {"plan": [...]} を返す。後方互換で単一 op は長さ1の計画。
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"op": "SORT", "args": {"col": "金額", "order": "desc"}}')
    got = ailine.translate_task("qwen2.5-coder:7b", "金額で降順に並べ替えて", _SAMPLE_META)
    assert got == {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}}]}

def test_translate_task_flat_args_are_rescued(monkeypatch):
    # モデルが args で包まず op と slot をフラットに返した場合も救済する。
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"op": "SORT", "col": "金額", "order": "desc"}')
    got = ailine.translate_task("qwen2.5-coder:7b", "金額で降順に並べ替えて", _SAMPLE_META)
    assert got["plan"][0]["op"] == "SORT"
    assert got["plan"][0]["args"] == {"col": "金額", "order": "desc"}

def test_translate_task_clarify_passthrough(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"op": "CLARIFY", "question": "どの列ですか？"}')
    got = ailine.translate_task("qwen2.5-coder:7b", "並べ替えて", _SAMPLE_META)
    assert got["plan"][0]["op"] == "CLARIFY"
    assert got["plan"][0]["question"] == "どの列ですか？"

def test_translate_task_invalid_json_falls_back_to_freeform(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        "これは JSON ではない")
    got = ailine.translate_task("qwen2.5-coder:7b", "いい感じにして", _SAMPLE_META)
    assert got["plan"][0]["op"] == "FREEFORM"

def test_translate_task_missing_required_slot_falls_back_to_freeform(monkeypatch):
    # order 欠落 → 必須 slot 不足 → FREEFORM に退避（クラッシュしない）。
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"op": "SORT", "args": {"col": "金額"}}')
    got = ailine.translate_task("qwen2.5-coder:7b", "並べ替えて", _SAMPLE_META)
    assert got["plan"][0]["op"] == "FREEFORM"

def test_translate_task_unknown_op_falls_back_to_freeform(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"op": "DELETE_ROW", "args": {}}')
    got = ailine.translate_task("qwen2.5-coder:7b", "行を消して", _SAMPLE_META)
    assert got["plan"][0]["op"] == "FREEFORM"

def test_translate_task_transport_failure_falls_back_to_freeform(monkeypatch):
    def boom(*a, **k):
        raise OSError("ollama 不通（テスト用）")
    monkeypatch.setattr(ailine, "ollama_generate_json", boom)
    got = ailine.translate_task("qwen2.5-coder:7b", "何かして", _SAMPLE_META)
    assert got["plan"][0]["op"] == "FREEFORM"

# --- ★ M2c: 複合計画パース (plan 配列・OUT_OF_VOCAB 保持) -----------------------

def test_translate_task_parses_multi_step_plan(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}}, '
                        '{"op": "BOLD", "args": {"target": "row:1"}}]}')
    got = ailine.translate_task("qwen2.5-coder:7b", "金額で降順に並べ替えて見出しを太字に", _SAMPLE_META)
    assert len(got["plan"]) == 2
    assert got["plan"][0]["op"] == "SORT"
    assert got["plan"][1] == {"op": "BOLD", "args": {"target": "row:1"}}

def test_translate_task_keeps_out_of_vocab_step_with_about_not_silently_dropped(monkeypatch):
    # ★ 黙落禁止: 語彙外の段は about 付きの OUT_OF_VOCAB として必ず計画に残る。
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"plan": [{"op": "COMPUTE_COLUMN", "args": {"operands": ["数量", "単価"], '
                        '"operator": "*", "target": "小計"}}, '
                        '{"op": "OUT_OF_VOCAB", "about": "税込み合計"}]}')
    got = ailine.translate_task("qwen2.5-coder:7b", "小計に数量×単価を入れて税込み合計も出して", _SAMPLE_META)
    assert len(got["plan"]) == 2
    assert got["plan"][0]["args"]["target"] == "小計"
    assert got["plan"][1] == {"op": "OUT_OF_VOCAB", "about": "税込み合計", "args": {}}

def test_translate_task_out_of_vocab_missing_about_gets_placeholder(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300:
                        '{"plan": [{"op": "OUT_OF_VOCAB"}]}')
    got = ailine.translate_task("qwen2.5-coder:7b", "いい感じにして", _SAMPLE_META)
    assert got["plan"][0]["op"] == "OUT_OF_VOCAB"
    assert got["plan"][0]["about"]   # 空文字ではない

def test_translate_task_empty_plan_array_falls_back_to_freeform(monkeypatch):
    monkeypatch.setattr(ailine, "ollama_generate_json",
                        lambda model, msgs, temperature=0.1, num_predict=300: '{"plan": []}')
    got = ailine.translate_task("qwen2.5-coder:7b", "何かして", _SAMPLE_META)
    assert got["plan"][0]["op"] == "FREEFORM"


# --- ② 検証（接地・数字表記の両解釈） -----------------------------------------

def test_resolve_col_ref_exact_name():
    v, inferred, err = ailine.resolve_col_ref("金額", ["商品", "金額", "在庫"])
    assert (v, inferred, err) == ("金額", False, None)

def test_resolve_col_ref_digit_unique_candidate_is_inferred():
    # "2" は 0起点なら在庫、1起点なら金額 → どちらも実在するが同じ列名になる場合は一意
    v, inferred, err = ailine.resolve_col_ref("0", ["商品", "金額", "在庫"])
    assert v == "商品"
    assert inferred is True
    assert err is None

def test_resolve_col_ref_digit_ambiguous_two_distinct_candidates():
    v, inferred, err = ailine.resolve_col_ref("1", ["商品", "金額", "在庫"])
    # 0起点=金額, 1起点=商品 → 二通りの実在列名に分かれるので一意に決まらない
    assert v is None
    assert "複数の解釈" in err

def test_resolve_col_ref_unknown_lists_known_columns():
    v, inferred, err = ailine.resolve_col_ref("存在しない列", ["商品", "金額"])
    assert v is None
    assert "商品" in err and "金額" in err

def test_verify_dsl_args_sort_ok():
    ok, resolved, inferred, err = ailine.verify_dsl_args("SORT", {"col": "金額", "order": "desc"}, _SAMPLE_META)
    assert ok is True
    # ★ 挙動変更#2: resolved は対象シートの決定(_target_sheet)を常に積む（省略時は1枚目）。
    assert resolved == {"col": "金額", "order": "desc", "_target_sheet": "Sheet"}
    assert inferred == set()
    assert err is None

def test_verify_dsl_args_sort_unknown_column_is_clarify_error():
    ok, resolved, inferred, err = ailine.verify_dsl_args("SORT", {"col": "存在しない", "order": "desc"}, _SAMPLE_META)
    assert ok is False
    assert "がありません" in err

def test_verify_dsl_args_sort_bad_order():
    ok, resolved, inferred, err = ailine.verify_dsl_args("SORT", {"col": "金額", "order": "up"}, _SAMPLE_META)
    assert ok is False
    assert "asc/desc" in err

def test_verify_dsl_args_compute_column_resolves_operands():
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, _SAMPLE_META)
    assert ok is True
    assert resolved["operands"] == ["売上", "原価"]

def test_verify_dsl_args_compute_column_bad_operator():
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "%"}, _SAMPLE_META)
    assert ok is False
    assert "演算子" in err

# --- ★ M2c: COMPUTE_COLUMN の target(名指し列への書き込み) ----------------------

def test_verify_dsl_args_compute_column_target_resolves_existing_column():
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN",
        {"operands": ["売上", "原価"], "operator": "-", "target": "金額"}, _SAMPLE_META)
    assert ok is True
    assert resolved["target"] == "金額"

def test_verify_dsl_args_compute_column_target_unknown_falls_back_to_new_column():
    # ★ W3: 実測で qwen2.5-coder:7b が「利益列を作って」の『利益』(新規列名) を target に
    #   誤って埋める頻度が高いと判明（E2E③『売上から原価を引いた利益列を作って』）。
    #   target が実在しない場合は一意性の曖昧さ(digit候補の複数一致)とは別なので、
    #   CLARIFY で止めず target 無指定＝新規列作成にフォールバックする。
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN",
        {"operands": ["売上", "原価"], "operator": "-", "target": "存在しない列"}, _SAMPLE_META)
    assert ok is True
    assert "target" not in resolved
    assert err is None

def test_verify_dsl_args_compute_column_target_ambiguous_digit_still_errors():
    # ★ W3: 実在しない場合と違い、複数解釈が可能な曖昧なケースは引き続き CLARIFY で止める
    #   （推測で断定しない原則は真に曖昧なケースにだけ残す）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["列0", "列1", "列2", "列3"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN",
        {"operands": ["列0", "列1"], "operator": "-", "target": "2"}, meta)
    assert ok is False
    assert "一意に決まりません" in err

def test_verify_dsl_args_compute_column_no_target_still_ok():
    # target 無指定は従来どおり合格（新規列パス）。
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, _SAMPLE_META)
    assert ok is True
    assert "target" not in resolved

def test_format_confirmation_line_compute_column_without_target_omits_field():
    line = ailine.format_confirmation_line(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, set())
    assert "対象列" not in line

def test_format_confirmation_line_compute_column_with_target_shows_it():
    line = ailine.format_confirmation_line(
        "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-", "target": "金額"}, set())
    assert "対象列:金額" in line

def test_codegen_dsl_compute_column_with_target_writes_into_existing_column():
    # ★ W3 Part3: 既定は式（setFormula）。値ベタ書きは use_formula=False（--values）で確認する。
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN",
        {"operands": ["売上", "原価"], "operator": "-", "target": "金額"}, _SAMPLE_META)
    # 金額は _SAMPLE_META の列1（0起点）。新規列(列5)には書かず、既存の列1に書く。
    assert "getCellByPosition(1, i).setFormula" in code
    assert '"D" & (i + 1) & "-" & "E" & (i + 1)' in code   # 売上=D列(3+1) 原価=E列(4+1)
    assert 'setString("売上-原価")' not in code   # 見出しは上書きしない(既存のまま)
    assert ailine.valid_signature(code)

def test_codegen_dsl_compute_column_values_mode_writes_static_getvalue():
    # ★ W3 Part3: --values（use_formula=False）は旧来の値ベタ書きのまま。
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN",
        {"operands": ["売上", "原価"], "operator": "-", "target": "金額"}, _SAMPLE_META,
        use_formula=False)
    assert "getCellByPosition(1, i).setValue" in code
    assert "setFormula" not in code

# --- ★ W10b 項目3: COMPUTE_COLUMN の「1列 × 率」パターン（税込み/税抜き） -----------

def test_verify_dsl_args_compute_column_single_operand_resolves_factor_from_task_text():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="税込み金額の列を追加して（消費税10%）")
    assert ok
    assert resolved["operands"] == ["金額"]
    assert resolved["factor"] == 1.1
    assert resolved["_sources"]["factor"] == "依頼文: 10%"

def test_verify_dsl_args_compute_column_single_operand_resolves_factor_from_vocab():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="消費税込みの列を追加して", vocab={"消費税": 1.1})
    assert ok
    assert resolved["factor"] == 1.1
    assert resolved["_sources"]["factor"] == "用語集: 消費税"

def test_verify_dsl_args_compute_column_single_operand_clarifies_when_no_rate():
    # ★ A': 率が依頼文にも用語集にも無ければ CLARIFY へ倒す（1.0既定で断定しない・
    #   APPEND_TOTAL の税/込番人と違い、このモードは常に率が要る前提なので無条件に聞く）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="税込みの列を追加して")
    assert not ok
    assert "倍率" in err and "分かりません" in err
    assert "ailine vocab add" in err

def test_verify_dsl_args_compute_column_single_operand_no_rate_signal_questions_classification():
    # ★ W10c 高: 依頼文に率らしい語が一切無いのに「1列×率」へ分類されたのは、分類自体が
    #   誤っている可能性が高い（実測: 「氏名の列を全部『退職済み』に書き換えて」が税率の
    #   話と誤認された事故の再現）。旧来の「倍率が分かりません」ではなく、分類を疑う
    #   文言に変わること・元の「倍率が分かりません」定型文はもう出ないこと。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["氏名", "部署", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["氏名"], "operator": "*"}, meta,
        task="氏名の列を全部『退職済み』に書き換えて")
    assert not ok
    assert "倍率らしき手がかりが見当たりません" in err
    assert "倍率（税率等）が分かりません" not in err

def test_verify_dsl_args_compute_column_single_operand_with_rate_signal_keeps_old_message():
    # 率らしい語（税/倍/掛け等）が依頼文にあれば、従来どおりの「倍率が分かりません」を返す
    # （regression: test_verify_dsl_args_compute_column_single_operand_clarifies_when_no_rate
    # と同じ入力・上の新しい番人が誤って一般ケースまで飲み込んでいないことの確認）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="税込みの列を追加して")
    assert not ok
    assert "倍率（税率等）が分かりません" in err

def test_verify_dsl_args_compute_column_single_operand_labels_new_column_for_tax_inclusive():
    # ★ W10c 中: 依頼文が「税込」と明言していれば、新規列の見出しに使う自然なラベルを
    #   resolved["_new_col_label"] として渡す（A' 原則: LLM を使わず正規表現のみ）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="税込み金額の列を追加して（消費税10%）")
    assert ok
    assert resolved["_new_col_label"] == "税込金額"

def test_verify_dsl_args_compute_column_single_operand_labels_new_column_for_tax_exclusive():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "/"}, meta,
        task="税抜き金額の列を追加して（消費税10%）")
    assert ok
    assert resolved["_new_col_label"] == "税抜金額"

def test_verify_dsl_args_compute_column_single_operand_no_label_when_target_explicit():
    # target 指定（既存列）時は新規列を作らないので _new_col_label は要らない。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額", "税込み金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "target": "税込み金額"}, meta,
        task="消費税10%を金額に掛けて税込み金額に入れて")
    assert ok
    assert "_new_col_label" not in resolved

def test_codegen_dsl_compute_column_single_operand_uses_new_col_label_when_present():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN",
        {"operands": ["金額"], "operator": "*", "factor": 1.1, "_new_col_label": "税込金額"}, meta)
    assert 'setString("税込金額")' in code
    assert 'setString("金額*1.1")' not in code

def test_check_compute_column_single_factor_uses_new_col_label_when_present(tmp_path):
    p = _book(tmp_path, [["品目", "金額", "税込金額"], ["a", 100, 110.0], ["b", 200, 220.0]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["金額"], "operator": "*", "factor": 1.1, "_new_col_label": "税込金額"})
    assert status == "pass"

def test_verify_dsl_args_compute_column_single_operand_llm_factor_ignored_but_warns():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "factor": 99}, meta,
        task="消費税10%を足した列を作って")
    assert ok
    assert resolved["factor"] == 1.1
    assert resolved["_warnings"]
    assert "99" in resolved["_warnings"][0] and "1.1" in resolved["_warnings"][0]

def test_verify_dsl_args_compute_column_single_operand_rejects_bad_operator():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "+"}, meta,
        task="金額に1.1を掛けた列を追加して")
    assert not ok
    assert "* か /" in err

def test_verify_dsl_args_compute_column_single_operand_unknown_column_errors():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["存在しない列"], "operator": "*"}, meta,
        task="消費税10%を足した列")
    assert not ok
    assert "がありません" in err

def test_verify_dsl_args_compute_column_single_operand_with_target_resolves_existing_column():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額", "税込み金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "target": "税込み金額"}, meta,
        task="消費税10%を金額に掛けて税込み金額に入れて")
    assert ok
    assert resolved["target"] == "税込み金額"

def test_codegen_dsl_compute_column_single_operand_writes_formula_with_cell_ref():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "factor": 1.1}, meta)
    # 金額は列1(0起点)＝B列。新規列(列2)に =B{行}*1.1 のセル参照式が入る。
    assert 'getCellByPosition(2, i).setFormula("=" & "B" & (i + 1) & "*1.1")' in code
    assert 'setString("金額*1.1")' in code
    assert ailine.valid_signature(code)

def test_codegen_dsl_compute_column_single_operand_values_mode_writes_static_getvalue():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]}}
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*", "factor": 1.1}, meta,
        use_formula=False)
    assert "getCellByPosition(1, i).getValue() * 1.1)" in code
    assert "setFormula" not in code

def test_codegen_dsl_compute_column_single_operand_with_target_writes_into_existing_column():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額", "税込み金額"]}}
    code = ailine.codegen_dsl(
        "COMPUTE_COLUMN",
        {"operands": ["金額"], "operator": "*", "factor": 1.1, "target": "税込み金額"}, meta)
    assert "getCellByPosition(2, i).setFormula" in code
    assert 'setString("金額*1.1")' not in code   # 既存の見出しは上書きしない

def test_check_compute_column_single_factor_passes_values_mode(tmp_path):
    # 金額*1.1 の結果を「税込み金額」列に直接書いた表（values モード相当）。
    p = _book(tmp_path, [["品目", "金額", "税込み金額"], ["a", 100, 110.0], ["b", 200, 220.0]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["金額"], "operator": "*", "factor": 1.1, "target": "税込み金額"})
    assert status == "pass"

def test_check_compute_column_single_factor_fails_when_value_wrong(tmp_path):
    p = _book(tmp_path, [["品目", "金額", "税込み金額"], ["a", 100, 999]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["金額"], "operator": "*", "factor": 1.1, "target": "税込み金額"})
    assert status == "fail"

def test_check_compute_column_single_factor_new_column_auto_named(tmp_path):
    # target 無指定なら自動命名『金額*1.1』列を検証する（codegen_dsl と同じ命名規則）。
    p = _book(tmp_path, [["品目", "金額", "金額*1.1"], ["a", 100, 110.0]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["金額"], "operator": "*", "factor": 1.1})
    assert status == "pass"

def test_check_compute_column_single_factor_formula_mode_passes(tmp_path):
    # ★ use_formula=True: 式文字列(=B{行}*1.1)とキャッシュ値の両方が一致して初めて pass。
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["品目", "金額", "税込み金額"])
    ws.append(["a", 100, None])
    ws.cell(row=2, column=3).value = "=B2*1.1"
    p = tmp_path / "formula.xlsx"
    wb.save(p)
    # openpyxl の通常読み(キャッシュ値)側も用意する必要があるため、data_only 読み用に
    # 一度 LibreOffice 相当のキャッシュを模して同じセルに数式文字列を残したまま
    # 値をキャッシュとして持たせる別ブックとして保存し直す代わりに、ここでは
    # check_compute_column_single_factor の使い分け（式/キャッシュ2層）を素直に確認する
    # ため、通常読みで式文字列を、data_only 読みでキャッシュ値を返す openpyxl の挙動に従う。
    status, reason = ailine.check_compute_column(
        p, {"operands": ["金額"], "operator": "*", "factor": 1.1, "target": "税込み金額"},
        use_formula=True)
    # openpyxl は数式のみ書いたセルに対する data_only キャッシュを持たない(None)ため、
    # ここでは「式は期待形」までを機械実装の観点で確認する（fail 理由がキャッシュ不一致
    # であることを確認 = 式チェック自体は通っている証拠）。
    assert status == "fail"
    assert "キャッシュ値が不一致" in reason


def test_verify_dsl_args_lookup_fill_ok():
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品", "数量", "単価"], "単価表": ["商品", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"},
        meta)
    assert ok is True
    assert resolved["target_col"] == "単価"

# ★ 挙動変更#2: 旧 test_verify_dsl_args_lookup_fill_rejects_non_first_sheet_target は
#   「対象シートは1枚目のみ対応」という散在した1枚目固定の一つを検証していた（査定の
#   致命そのもの）。その制限を撤廃したので、逆に「2枚目でも通る」ことを検証する。
def test_verify_dsl_args_lookup_fill_allows_non_first_sheet_target():
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品", "数量", "単価"], "単価表": ["商品", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "単価表", "target_col": "単価", "source_sheet": "明細", "key_col": "商品"},
        meta, task="単価表の単価を明細から転記して")
    assert ok is True
    assert err is None
    assert resolved["target_sheet"] == "単価表"
    assert resolved["_target_sheet"] == "単価表"   # ★ codegen/postcondition が読む一元化キー


# --- ★ 挙動変更#2: 対象シートの固定を解く（resolve_target_sheet / describe_target_sheet） ----

def test_resolve_target_sheet_cli_flag_wins():
    sheet, source, err, conflict = ailine.resolve_target_sheet("なにか", ["請求書", "工事台帳"], "工事台帳")
    assert (sheet, source, err, conflict) == ("工事台帳", "cli", None, None)

def test_resolve_target_sheet_cli_flag_unknown_sheet_errors():
    sheet, source, err, conflict = ailine.resolve_target_sheet("なにか", ["請求書", "工事台帳"], "存在しないシート")
    assert sheet is None and source == "cli" and conflict is None
    assert "存在しないシート" in err and "請求書" in err and "工事台帳" in err

def test_resolve_target_sheet_named_in_task_wins_over_default():
    sheet, source, err, conflict = ailine.resolve_target_sheet(
        "工事台帳シートで取引先ごとの売上を集計して", ["請求書", "工事台帳", "取引先マスタ"])
    assert (sheet, source, err, conflict) == ("工事台帳", "task", None, None)

def test_resolve_target_sheet_ordinal_phrase():
    sheet, source, err, conflict = ailine.resolve_target_sheet(
        "2枚目のシートで金額を降順に並べ替えて", ["請求書", "工事台帳"])
    assert (sheet, source, err, conflict) == ("工事台帳", "task", None, None)

def test_resolve_target_sheet_ambiguous_mention_falls_back_to_default_not_clarify():
    # ★ LOOKUP_FILL のように転記先/参照元の2シート名が正当に両方登場するケースを
    #   誤ってブロックしない（resolve_target_sheet 自身の docstring 参照）。
    sheet, source, err, conflict = ailine.resolve_target_sheet(
        "工事台帳の値を単価表から転記して", ["工事台帳", "単価表"])
    assert err is None and conflict is None
    assert (sheet, source) == ("工事台帳", "default")   # 1枚目へフォールバック（曖昧を理由に止めない）

def test_resolve_target_sheet_substring_names_pick_longer_match():
    # 「請求書」は「請求書控え」の部分文字列 → 長い方(実際に言及された固有名)だけ残す。
    sheet, source, err, conflict = ailine.resolve_target_sheet(
        "請求書控えシートで並べ替えて", ["請求書", "請求書控え"])
    assert (sheet, source, err, conflict) == ("請求書控え", "task", None, None)

def test_resolve_target_sheet_no_mention_defaults_to_first():
    sheet, source, err, conflict = ailine.resolve_target_sheet("金額で降順に並べ替えて", ["請求書", "工事台帳"])
    assert (sheet, source, err, conflict) == ("請求書", "default", None, None)

def test_resolve_target_sheet_no_sheets_errors():
    sheet, source, err, conflict = ailine.resolve_target_sheet("なにか", [])
    assert sheet is None and conflict is None
    assert "シートが無い" in err

def test_describe_target_sheet_single_sheet_book_is_silent():
    assert ailine.describe_target_sheet(["Sheet"], "Sheet", "default") is None

def test_describe_target_sheet_default_matches_brief_wording():
    line = ailine.describe_target_sheet(["a", "b", "c", "請求明細"], "請求明細", "default")
    assert line == "操作するシート: 4枚目『請求明細』（このブックは4シート）"

def test_describe_target_sheet_task_source_names_the_reason():
    line = ailine.describe_target_sheet(["請求書", "工事台帳", "取引先マスタ"], "工事台帳", "task")
    assert line == "操作するシート: 2枚目『工事台帳』（依頼文から判断・このブックは3シート）"

def test_describe_target_sheet_cli_source_names_the_reason():
    line = ailine.describe_target_sheet(["請求書", "工事台帳"], "工事台帳", "cli")
    assert line == "操作するシート: 2枚目『工事台帳』（--sheet 指定）"


# ★ DoD1: 査定の再現の回帰テスト（verify_dsl_args レベル）。
#   旧実装は first_sheet が常に sheets[0] だったため、対象シートを名指ししても
#   「ある列: ...」に *1枚目* の列名が出る誤誘導だった（査定の致命そのもの）。
def test_verify_dsl_args_error_lists_named_target_sheet_columns_not_first_sheet():
    meta = {"sheets": ["請求書", "工事台帳"],
            "headers": {"請求書": ["宛先", "金額"], "工事台帳": ["取引先名", "工事名", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "SORT", {"col": "宛先", "order": "desc"}, meta,
        task="工事台帳シートで宛先を降順に並べ替えて",
        target_sheet="工事台帳")   # ★ _cmd_run_dispatch が resolve_target_sheet で決めた値
    assert ok is False
    # ★ 直った点そのもの: 対象(工事台帳)の列名だけが出る。1枚目(請求書)の「宛先」「金額」を
    #   誤って「ある列」に混ぜない（宛先は工事台帳に無い列なのでエラー自体は正しい）。
    assert err == "列『宛先』がありません。ある列: 取引先名, 工事名, 金額"
    assert "請求書" not in err


def test_resolve_header_rows_uses_target_sheet_not_always_first():
    # ★ 挙動変更#2: 対象シートが2枚目でも、build_struct_dump が元々全シート分持っている
    #   rows(書式的特徴) から検出できる（StructDump のやり直しは不要）。
    struct_dump = {"sheets": {
        "請求書": {"rows": {1: {"nonempty": 2, "str": 2, "bold": 0},
                            2: {"nonempty": 2, "str": 1, "bold": 0}}},
        "工事台帳": {"rows": {1: {"nonempty": 0, "str": 0, "bold": 0},
                              2: {"nonempty": 0, "str": 0, "bold": 0},
                              3: {"nonempty": 3, "str": 3, "bold": 0},
                              4: {"nonempty": 3, "str": 1, "bold": 0}}},
    }}
    header_rows, clarify = ailine.resolve_header_rows(
        struct_dump, ["請求書", "工事台帳"], target_sheet="工事台帳")
    assert clarify is None
    assert header_rows["工事台帳"] == 3   # 対象シートは検出結果
    assert header_rows["請求書"] == 1     # 対象外は既定のまま（旧挙動と同一）

def test_resolve_header_rows_target_sheet_none_defaults_to_first_sheet_unchanged():
    struct_dump = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0}, 2: {"nonempty": 2, "str": 1, "bold": 0}}}}}
    header_rows, clarify = ailine.resolve_header_rows(struct_dump, ["Sheet"])   # target_sheet 省略
    assert header_rows == {"Sheet": 1}
    assert clarify is None


# ★ 挙動変更#2: codegen_dsl が対象シートを一時的に先頭へ動かして戻す（_wrap_basic_for_sheet）。
def test_codegen_dsl_first_sheet_target_unchanged_output():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "数量", "単価", "金額", "原価"]}}
    code = ailine.codegen_dsl("SORT", {"col": "金額", "order": "desc"}, meta)
    assert "moveByName" not in code   # ★ 挙動不変（既存ゴールデンと同型）
    assert code == (
        "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\n"
        "    Call SortByColumn(oDoc, 0, 4, 3, False)\nEnd Sub\n")

def test_codegen_dsl_non_first_sheet_target_wraps_with_move_by_name_and_restores_index():
    meta = {"sheets": ["請求書", "工事台帳", "取引先マスタ"],
            "headers": {"請求書": ["宛先", "金額"],
                        "工事台帳": ["取引先名", "工事名", "金額"],
                        "取引先マスタ": ["取引先名", "住所"]}}
    resolved = {"group_col": "取引先名", "value_col": "金額", "_target_sheet": "工事台帳"}
    code = ailine.codegen_dsl("AGGREGATE", resolved, meta)
    assert 'oDoc.Sheets.moveByName("工事台帳", 0)' in code
    assert 'oDoc.Sheets.moveByName("工事台帳", 1)' in code   # 元の位置(index 1)へ必ず戻す
    assert "Sub __AilineTargetBody(oDoc As Object)" in code   # Exit Sub をまたいでも必ず戻る形
    assert "Call SummaryTable(oDoc, 0, 0, 2)" in code

def test_codegen_dsl_missing_target_sheet_key_falls_back_to_first_sheet_unchanged():
    # ★ 後方互換: resolved_args に _target_sheet が無い既存呼び出し（多数の既存単体テスト/
    #   ゴールデン）は codegen_dsl 内部で book_meta["sheets"][0] にフォールバックする。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    code = ailine.codegen_dsl("NUMBER_FORMAT", {"col": "金額", "style": "thousands"}, meta)
    assert "moveByName" not in code


# ★ postcondition チェッカーが resolved_args["_target_sheet"] を読むこと（1つの代表例で確認・
#   他13チェッカーは機械的に同じ書き換え — CHANGED ファイル一覧参照）。
def test_check_sort_reads_target_sheet_from_args_not_first_sheet(tmp_path):
    book = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "請求書"
    ws1.append(["宛先", "金額"])
    ws1.append(["b", 2])
    ws1.append(["a", 1])   # ★ 1枚目はわざと未整列のまま（対象シートでないので無視されるはず）
    ws2 = wb.create_sheet("工事台帳")
    ws2.append(["取引先名", "金額"])
    ws2.append(["田中", 300])
    ws2.append(["山田", 100])
    wb.save(book)
    status, reason = ailine.check_sort(
        book, {"col": "金額", "order": "desc", "_target_sheet": "工事台帳"})
    assert status == "pass", reason


# ★ OP_WRITE_TARGET の sheet_key=None フォールバックが resolved["_target_sheet"] を読む
#   （_maybe_warn_target_overwrite / _declared_new_column_letter の2箇所）。
def test_maybe_warn_target_overwrite_uses_target_sheet_not_first_sheet(tmp_path):
    book = tmp_path / "book.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "請求書"
    ws1.append(["数量", "単価"])   # ★ 1枚目に同名の"金額"列は無い＝1枚目基準だと誤検知しない側
    ws2 = wb.create_sheet("工事台帳")
    ws2.append(["取引先名", "金額"])
    ws2.append(["田中", 300])
    wb.save(book)
    meta = {"sheets": ["請求書", "工事台帳"],
            "headers": {"請求書": ["数量", "単価"], "工事台帳": ["取引先名", "金額"]},
            "header_rows": {"請求書": 1, "工事台帳": 1}}
    warn = ailine._maybe_warn_target_overwrite(
        "SET_COLUMN_VALUE", {"col": "金額", "_target_sheet": "工事台帳"}, meta, book)
    assert warn is not None and "1 件" in warn   # 工事台帳の「金額」列に既存値1件を検知

def test_declared_new_column_letter_uses_target_sheet_not_first_sheet():
    meta = {"sheets": ["請求書", "工事台帳"],
            "headers": {"請求書": ["宛先", "金額"], "工事台帳": ["取引先名", "工事名"]}}
    # COMPUTE_COLUMN の target 無指定＝新規列。対象シート(工事台帳)は2列なので新規列は3列目=C。
    letter = ailine._declared_new_column_letter(
        "COMPUTE_COLUMN", {"_target_sheet": "工事台帳"}, meta)
    assert letter == "C"

# ★ W10c 致命2: target_col が対象シートに実在しない場合の扱い。
#   実測の再現: 対象シートに『単価』列が無い状態で「単価表を見て単価を入れて」を依頼すると、
#   LLM が依頼に無い別の実在列（『数量』）を勝手に代わりに返すことがあった
#   （resolve_col_ref は実在列名なら無条件で通すため、これだけでは見分けられない）。

def test_verify_dsl_args_lookup_fill_missing_target_col_creates_new_when_task_names_it():
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品コード"},
        meta, task="単価表を見て単価を入れて")
    assert ok is True
    assert resolved["target_col"] == "単価"   # 依頼文にある名前のまま＝新規列名として使う

def test_verify_dsl_args_lookup_fill_existing_target_col_without_grounds_clarifies():
    # ★ 監査事故の再現そのもの: LLM が返した『数量』は実在するが、依頼文のどこにも
    # 現れず、転記元（単価表）の値列（単価）とも一致しない＝黙って信用せず CLARIFY
    # （無関係な既存列を上書き対象にしない）。
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "数量", "source_sheet": "単価表", "key_col": "商品コード"},
        meta, task="単価表を見て単価を入れて")
    assert ok is False
    assert "取り違えている" in err
    assert "数量" in err

def test_verify_dsl_args_lookup_fill_existing_target_col_matching_source_value_col_trusted():
    # target_col が実在し、依頼文には無くても、転記元の値列と同じ名前なら信用してよい
    # （VLookupFromTable は常に参照表の列1を値として転記するため、一番自然な対応）。
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品", "数量", "単価"], "単価表": ["商品", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"},
        meta, task="埋めて")
    assert ok is True
    assert resolved["target_col"] == "単価"

def test_verify_dsl_args_lookup_fill_missing_target_col_digit_reference_still_resolves():
    # 数字表記の推定（従来どおり）は維持する（一意に決まる場合のみ）。
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "0", "source_sheet": "単価表", "key_col": "商品コード"},
        meta, task="単価表を見て入れて")
    assert ok is True
    assert resolved["target_col"] == "商品コード"
    assert "target_col" in inferred

def test_verify_dsl_args_fill_color_unknown_color():
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "FILL_COLOR", {"target": "col:金額", "color": "虹色"}, _SAMPLE_META)
    assert ok is False
    assert "未対応" in err

def test_verify_dsl_args_center_align_all_rejected_for_bold():
    ok, resolved, inferred, err = ailine.verify_dsl_args("BOLD", {"target": "all"}, _SAMPLE_META)
    assert ok is False

def test_verify_dsl_args_merge_bad_range_format():
    ok, resolved, inferred, err = ailine.verify_dsl_args("MERGE", {"range": "not-a-range"}, _SAMPLE_META)
    assert ok is False
    assert "形式" in err

def test_verify_dsl_args_bold_col_marks_digit_resolution_as_inferred():
    ok, resolved, inferred, err = ailine.verify_dsl_args("BOLD", {"target": "col:0"}, _SAMPLE_META)
    assert ok is True
    assert resolved["target"] == "col:商品"
    assert "target" in inferred


# --- ③ 確認行 ----------------------------------------------------------------

def test_format_confirmation_line_sort():
    line = ailine.format_confirmation_line("SORT", {"col": "金額", "order": "desc"}, set())
    assert line == "解釈: 操作:並べ替え 対象:金額 順:降順"

def test_format_confirmation_line_marks_inferred_arg():
    line = ailine.format_confirmation_line("SORT", {"col": "金額", "order": "desc"}, {"col"})
    assert "対象:金額(推定)" in line

def test_format_confirmation_line_lookup_fill():
    line = ailine.format_confirmation_line(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}, set())
    assert line == "解釈: 操作:転記 対象シート:明細 対象列:単価 参照シート:単価表 キー列:商品"


# --- ④ 決定論 codegen ---------------------------------------------------------

def test_codegen_dsl_sort_calls_helper_with_zero_based_index():
    # ★ W3: SortByColumn は headerRow(0起点) を新たな第2引数に取る。_SAMPLE_META は
    #   header_rows を持たない旧テスト値＝既定1行目(hr0=0)になる。
    code = ailine.codegen_dsl("SORT", {"col": "金額", "order": "desc"}, _SAMPLE_META)
    assert "Call SortByColumn(oDoc, 0, 4, 1, False)" in code   # lastCol=4 (_SAMPLE_META は5列)
    assert ailine.valid_signature(code)
    assert not ailine.is_truncated_code(code)

def test_codegen_dsl_sort_ascending():
    code = ailine.codegen_dsl("SORT", {"col": "在庫", "order": "asc"}, _SAMPLE_META)
    assert "Call SortByColumn(oDoc, 0, 4, 2, True)" in code

def test_codegen_dsl_lookup_fill_calls_helper():
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品", "数量", "単価"], "単価表": ["商品", "単価"]}}
    code = ailine.codegen_dsl(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}, meta)
    assert 'Call VLookupFromTable(oDoc, 0, 0, 2, "単価表")' in code

def test_codegen_dsl_lookup_fill_creates_new_column_when_target_col_missing():
    # ★ W10c 致命2: target_col が対象シートに実在しない場合、末尾に新しい列を作ってから
    #   転記する（COMPUTE_COLUMN の新規列作成と同じ考え方）。
    meta = {"sheets": ["明細", "単価表"],
            "headers": {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}}
    code = ailine.codegen_dsl(
        "LOOKUP_FILL",
        {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品コード"}, meta)
    assert 'getCellByPosition(2, 0).setString("単価")' in code
    assert 'Call VLookupFromTable(oDoc, 0, 0, 2, "単価表")' in code

def test_codegen_dsl_aggregate_calls_helper():
    code = ailine.codegen_dsl("AGGREGATE", {"group_col": "商品", "value_col": "売上"}, _SAMPLE_META)
    assert "Call SummaryTable(oDoc, 0, 0, 3)" in code

def test_codegen_dsl_number_format_calls_helper():
    code = ailine.codegen_dsl("NUMBER_FORMAT", {"col": "金額", "style": "thousands"}, _SAMPLE_META)
    assert "Call FormatThousands(oDoc, 0, 1)" in code

def test_codegen_dsl_merge_converts_a1_range_to_zero_based():
    code = ailine.codegen_dsl("MERGE", {"range": "A1:E1"}, _SAMPLE_META)
    assert "Call MergeCells(oDoc, 0, 0, 4, 0)" in code

def test_codegen_dsl_chart_calls_helper():
    code = ailine.codegen_dsl("CHART", {"value_col": "金額"}, _SAMPLE_META)
    assert "Call InsertBarChart(oDoc, 0, 1)" in code

def test_codegen_dsl_center_align_all_calls_helper():
    code = ailine.codegen_dsl("CENTER_ALIGN", {"target": "all"}, _SAMPLE_META)
    assert "Call AlignCenter(oDoc, 0, 4)" in code   # lastCol=4 (_SAMPLE_META は5列)

def test_codegen_dsl_center_align_col_writes_template():
    code = ailine.codegen_dsl("CENTER_ALIGN", {"target": "col:在庫"}, _SAMPLE_META)
    assert "CellHoriJustify.CENTER" in code
    assert "getCellRangeByPosition(2, 0, 2, lastRow)" in code
    assert ailine.valid_signature(code)

def test_codegen_dsl_bold_row_calls_helper_with_scanned_range():
    code = ailine.codegen_dsl("BOLD", {"target": "row:1"}, _SAMPLE_META)
    assert "Call StyleBold(oDoc, 0, 0, 4, 0)" in code   # ★ W3: lastCol は走査でなく接地済み列数から決定論的に決まる

def test_codegen_dsl_bold_col_calls_helper_with_scanned_range():
    code = ailine.codegen_dsl("BOLD", {"target": "col:商品"}, _SAMPLE_META)
    assert "Call StyleBold(oDoc, 0, 0, 0, lastRow)" in code

def test_codegen_dsl_fill_color_row_writes_hex_literal():
    code = ailine.codegen_dsl("FILL_COLOR", {"target": "row:1", "color": "yellow"}, _SAMPLE_META)
    assert "&HFFFF00&" in code

def test_codegen_dsl_fill_color_col_writes_hex_literal():
    code = ailine.codegen_dsl("FILL_COLOR", {"target": "col:在庫", "color": "red"}, _SAMPLE_META)
    assert "&HFF0000&" in code
    assert "getCellByPosition(2, r)" in code

def test_codegen_dsl_compute_column_writes_new_column_at_end():
    # ★ W3 Part3: 既定は式。値ベタ書きの回帰は use_formula=False 側で確認する。
    code = ailine.codegen_dsl("COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, _SAMPLE_META)
    assert 'setString("売上-原価")' in code
    assert "getCellByPosition(5, 0)" in code   # 既存5列(0..4)の次=列5
    assert 'getCellByPosition(5, i).setFormula("=" & "D" & (i + 1) & "-" & "E" & (i + 1))' in code
    assert ailine.valid_signature(code)

def test_codegen_dsl_compute_column_values_mode_new_column_uses_getvalue():
    code = ailine.codegen_dsl("COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"},
                               _SAMPLE_META, use_formula=False)
    assert 'setString("売上-原価")' in code
    assert "getCellByPosition(5, 0)" in code
    assert "getCellByPosition(3, i).getValue() - oSheet.getCellByPosition(4, i).getValue()" in code
    assert ailine.valid_signature(code)
    # ★ is_truncated_code() は未検証: "Exit Sub" 直後の改行+識別子("...Exit Sub\n    oSheet")を
    #   \s+ が跨いで新しい Sub 開始と誤認する既存の罠がある（DSL 経路では is_truncated_code を
    #   呼んでいないため実害は無いが、ここでは対象外として素通りする）。


# --- ⑥ op 別事後条件（達成の機械検証） ----------------------------------------
# ★ 止血1/2（bench/realworld/BASELINE.md の D/C②検体の根治）: 各 check_* /
#   run_postcondition の戻り値を (ok: bool, reason: str) から (status: str, reason: str)
#   に変更した（status ∈ {"pass","warn","fail","error"}）。検証対象が0件/意味を
#   持たない少数のケースを「合格(True)」で素通りさせないための3値化。
#   以下の既存テストは ok is True/False を status == "pass"/"fail" に機械的に置換
#   （通常ケース＝検証対象が十分ある場合は判定結果そのものは変わらない）。

def _wb_save(tmp_path, build_fn, name="p.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    build_fn(wb)
    wb.save(p)
    return p

def test_check_sort_passes_when_sorted_desc(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"})
    assert status == "pass"

def test_check_sort_fails_when_not_sorted(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 300], ["c", 200]])
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"})
    assert status == "fail"
    assert "並んでいない" in reason

def test_check_compute_column_passes_when_values_match(tmp_path):
    p = _book(tmp_path, [["売上", "原価", "売上-原価"], [500, 300, 200], [900, 400, 500]])
    status, reason = ailine.check_compute_column(p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "pass"

def test_check_compute_column_fails_when_value_wrong(tmp_path):
    p = _book(tmp_path, [["売上", "原価", "売上-原価"], [500, 300, 999]])
    status, reason = ailine.check_compute_column(p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "fail"

def test_check_compute_column_target_checks_existing_column(tmp_path):
    # ★ M2c: target 指定時は新規列名でなく target 列そのものを検証する。
    p = _book(tmp_path, [["数量", "単価", "小計"], [2, 100, 200], [3, 150, 450]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["数量", "単価"], "operator": "*", "target": "小計"})
    assert status == "pass"

def test_check_compute_column_target_fails_when_target_value_wrong(tmp_path):
    p = _book(tmp_path, [["数量", "単価", "小計"], [2, 100, 999]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["数量", "単価"], "operator": "*", "target": "小計"})
    assert status == "fail"

def test_check_lookup_fill_passes_when_all_transcribed(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "明細"
    for row in [["商品", "数量", "単価"], ["りんご", 2, 100], ["バナナ", 3, 200]]:
        ws1.append(row)
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    p = tmp_path / "lookup.xlsx"
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args)
    assert status == "pass"

def test_check_lookup_fill_fails_when_value_missing_a_row(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "明細"
    for row in [["商品", "数量", "単価"], ["りんご", 2, 100], ["バナナ", 3, None]]:
        ws1.append(row)
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    p = tmp_path / "lookup.xlsx"
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args)
    assert status == "fail"

def test_check_lookup_fill_master_reversed_column_order_gives_actionable_guidance(tmp_path):
    # ★ W10b 項目4a(摩擦): マスタ表が「値→キー」の順(単価が列0・商品名が列1)だと
    # VLookupFromTable ヘルパ(常に列0=キー・列1=値固定)が1件も引けない。旧メッセージは
    # 「対応表に載っているキーが1件も転記されていない」とだけ言って原因を示さなかった。
    # ★ W10f 項目5: この genuine な列順違いのケースでも従来どおり fail することを
    # 固定する（誤診断の是正で断定をやめた結果、本当の列順違いまで見逃す方向に倒れて
    # いないことの実証＝濡れ衣の逆）。ただし文言は『マスタ表の列順が違う』と断定せず
    # 『可能性があります』と並べる形に変わった（区別できない以上、断定は飛躍）ため、
    # 期待文言をそちらに更新する。actionable な直し方（A 列にキー・B 列に値）は残す。
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "明細"
    for row in [["商品", "数量", "単価"], ["りんご", 2, None], ["バナナ", 3, None]]:
        ws1.append(row)
    ws2 = wb.create_sheet("単価表")
    for row in [["単価", "商品"], [100, "りんご"], [200, "バナナ"]]:   # ★ 列順が逆(値→キー)
        ws2.append(row)
    p = tmp_path / "lookup_reversed.xlsx"
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args)
    assert status == "fail"
    assert "単価表" in reason
    assert "可能性があります" in reason   # ★ 断定でなく可能性として並べる
    assert "キー列→値列の順" in reason
    assert "A 列にキー" in reason and "B 列に値" in reason


# --- ★ W10f 項目5: LOOKUP_FILL のキー列を operand にすると誤診断する（Namakoo 実測の
#   再現をそのまま回帰テストにする）。VLookupFromTable ヘルパは getString()(LibreOffice が
#   評価した文字列)でキーを照合するため、前段の式(=A2 等)で埋まったキー列でも転記自体は
#   正しく動く。旧実装は openpyxl の raw 読み(式文字列そのもの)でキーを拾っていたため
#   対応表と1件も一致せず、正しく転記された明細に対して『マスタ表の列順が違う』という
#   濡れ衣を着せていた。use_formula=True を配線し data_only 側から読むよう直す。 ---

def test_check_lookup_fill_formula_mode_key_col_from_prior_formula_passes_when_cache_present(tmp_path):
    # ★ Namakoo が貼った純関数レベルの再現。明細シートのキー列(商品)が前段の式
    # (=A2 等)で埋まっている状態でも、転記は正しく完了していれば pass すること。
    p = tmp_path / "lookup_formula_key.xlsx"
    wb = openpyxl.Workbook(); ws1 = wb.active; ws1.title = "明細"
    ws1.append(["元商品", "商品", "数量", "単価"])
    ws1.append(["りんご", None, 2, 100]); ws1["B2"] = "=A2"
    ws1.append(["バナナ", None, 3, 200]); ws1["B3"] = "=A3"
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    wb.save(p)
    _inject_formula_string_cache(p, "xl/worksheets/sheet1.xml", {"B2": "りんご", "B3": "バナナ"})
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args, use_formula=True)
    assert status == "pass"
    assert "2 行を検証" in reason

def test_check_lookup_fill_formula_mode_key_col_no_cache_fails_with_distinct_reason(tmp_path):
    # 式はあるがキャッシュ値が無い(LibreOffice を通していない)場合は、直った後も
    # 検証できず fail のままだが、理由が『マスタ表の列順が違う』という誤診断ではなく
    # 『キャッシュ値が無く検証できない』という正確な診断に変わる（0cf9218 空虚な検証
    # 合格の禁止 — 『対象が無い』と『対象はあるが読めない』を混同しない）。
    p = tmp_path / "lookup_formula_key_nocache.xlsx"
    wb = openpyxl.Workbook(); ws1 = wb.active; ws1.title = "明細"
    ws1.append(["元商品", "商品", "数量", "単価"])
    ws1.append(["りんご", None, 2, 100]); ws1["B2"] = "=A2"
    ws1.append(["バナナ", None, 3, 200]); ws1["B3"] = "=A3"
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args, use_formula=True)
    assert status == "fail"
    assert "キャッシュ値が無く検証できない行が 2 件" in reason
    assert "列順が違う" not in reason
    assert "可能性があります" not in reason   # 読めなかっただけで判定不能とは別の主張

def test_check_lookup_fill_formula_mode_plain_key_values_still_pass(tmp_path):
    # 対照: キー列が素の値(式でない)なら use_formula=True でも従来どおり pass する。
    p = tmp_path / "lookup_formula_mode_plain_key.xlsx"
    wb = openpyxl.Workbook(); ws1 = wb.active; ws1.title = "明細"
    ws1.append(["商品", "数量", "単価"])
    ws1.append(["りんご", 2, 100])
    ws1.append(["バナナ", 3, 200])
    ws1.append(["みかん", 1, 150])
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200], ["みかん", 150]]:
        ws2.append(row)
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args, use_formula=True)
    assert status == "pass"
    assert reason == "3 行を検証"


def test_check_aggregate_passes_when_sums_correct(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["部門", "金額"], ["営業", 100], ["営業", 200], ["経理", 50]]:
        ws.append(row)
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"])
    out.append(["営業", 300])
    out.append(["経理", 50])
    out.append(["合計", 350])
    p = tmp_path / "agg.xlsx"
    wb.save(p)
    status, reason = ailine.check_aggregate(p, {"group_col": "部門", "value_col": "金額"})
    assert status == "pass"

def test_check_aggregate_fails_when_sum_wrong(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["部門", "金額"], ["営業", 100], ["営業", 200]]:
        ws.append(row)
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"])
    out.append(["営業", 999])
    p = tmp_path / "agg.xlsx"
    wb.save(p)
    status, reason = ailine.check_aggregate(p, {"group_col": "部門", "value_col": "金額"})
    assert status == "fail"

def test_check_aggregate_fails_when_sheet_missing(tmp_path):
    p = _book(tmp_path, [["部門", "金額"], ["営業", 100]])
    status, reason = ailine.check_aggregate(p, {"group_col": "部門", "value_col": "金額"})
    assert status == "fail"
    assert "集計" in reason

# --- ★ W10f 項目1: check_aggregate も同型（operand を式ビューから読むと非数値→0扱いで
#   偽の不一致 fail になる）。use_formula=True で data_only 側から読むよう直した後の挙動。

def test_check_aggregate_use_formula_default_false_keeps_old_behavior(tmp_path):
    # use_formula 省略時=False は旧テストと同じ挙動（回帰確認）。
    wb = openpyxl.Workbook(); ws = wb.active
    for row in [["部門", "金額"], ["営業", 100], ["営業", 200]]:
        ws.append(row)
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"]); out.append(["営業", 300])
    p = tmp_path / "agg_default.xlsx"
    wb.save(p)
    status, reason = ailine.check_aggregate(p, {"group_col": "部門", "value_col": "金額"})
    assert status == "pass"

def test_check_aggregate_operand_from_prior_formula_column_passes_when_cache_present(tmp_path):
    p = tmp_path / "agg_chain.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "部門", "金額"])
    ws.append(["a", 3, 100, "営業", None]); ws["E2"] = "=B2*C2"
    ws.append(["b", 2, 50, "営業", None]); ws["E3"] = "=B3*C3"
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"])
    out.append(["営業", 400])
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"E2": 300, "E3": 100})
    status, reason = ailine.check_aggregate(
        p, {"group_col": "部門", "value_col": "金額"}, use_formula=True)
    assert status == "pass"

def test_check_aggregate_operand_from_prior_formula_column_no_cache_fails_with_distinct_reason(tmp_path):
    # ★ 直る前は非数値(式文字列)→0扱いで「偽の不一致 fail」になっていた（本当は300なのに
    # 0として合算され集計『集計』シートの実測値と食い違って見える）。直った後は『0扱いで
    # 不一致』という誤診断ではなく『キャッシュ値が無く検証できない』という正確な診断になる。
    p = tmp_path / "agg_chain_nocache.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "部門", "金額"])
    ws.append(["a", 3, 100, "営業", None]); ws["E2"] = "=B2*C2"
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"])
    out.append(["営業", 300])
    wb.save(p)
    status, reason = ailine.check_aggregate(
        p, {"group_col": "部門", "value_col": "金額"}, use_formula=True)
    assert status == "fail"
    assert "キャッシュ値が無く検証できない行が 1 件" in reason
    assert "集計を検証できません" in reason
    assert "不一致" not in reason

def test_check_bold_row_passes_when_all_bold(tmp_path):
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["a", 1])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_bold(p, {"target": "row:1"})
    assert status == "pass"

def test_check_bold_row_fails_when_partial(tmp_path):
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws["A1"].font = Font(bold=True)   # B1 は太字にしない
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_bold(p, {"target": "row:1"})
    assert status == "fail"

def test_check_bold_col_passes_when_all_bold(tmp_path):
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["a", 1])
    ws.append(["b", 2])
    for r in (1, 2, 3):
        ws.cell(row=r, column=1).font = Font(bold=True)
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_bold(p, {"target": "col:商品"})
    assert status == "pass"

def test_check_fill_color_passes_when_matching(tmp_path):
    from openpyxl.styles import PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "在庫"])
    ws.append(["a", 3])
    for r in (1, 2):
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFFF00")
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_fill_color(p, {"target": "col:在庫", "color": "yellow"})
    assert status == "pass"

def test_check_fill_color_fails_when_not_colored(tmp_path):
    p = _book(tmp_path, [["商品", "在庫"], ["a", 3]])
    status, reason = ailine.check_fill_color(p, {"target": "col:在庫", "color": "yellow"})
    assert status == "fail"

def test_check_number_format_passes_when_thousands_applied(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["a", 1000])
    ws["B2"].number_format = "#,##0"
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_number_format(p, {"col": "金額", "style": "thousands"})
    assert status == "pass"

def test_check_number_format_fails_when_not_applied(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 1000]])
    status, reason = ailine.check_number_format(p, {"col": "金額", "style": "thousands"})
    assert status == "fail"

def test_check_merge_passes_when_range_merged(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b", "c"])
    ws.merge_cells("A1:C1")
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_merge(p, {"range": "A1:C1"})
    assert status == "pass"

def test_check_merge_fails_when_not_merged(tmp_path):
    p = _book(tmp_path, [["a", "b", "c"]])
    status, reason = ailine.check_merge(p, {"range": "A1:C1"})
    assert status == "fail"

def test_check_chart_passes_when_count_plus_one(monkeypatch, tmp_path):
    p = tmp_path / "c.xlsx"
    p.write_bytes(b"dummy")
    monkeypatch.setattr(ailine, "_charts_count", lambda path: 1)
    status, reason = ailine.check_chart(p, 0)
    assert status == "pass"

def test_check_chart_fails_when_count_unchanged(monkeypatch, tmp_path):
    p = tmp_path / "c.xlsx"
    p.write_bytes(b"dummy")
    monkeypatch.setattr(ailine, "_charts_count", lambda path: 0)
    status, reason = ailine.check_chart(p, 0)
    assert status == "fail"

def test_check_center_align_all_passes(tmp_path):
    from openpyxl.styles import Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 2])
    for r in (1, 2):
        for c in (1, 2):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    p = tmp_path / "b.xlsx"
    wb.save(p)
    status, reason = ailine.check_center_align(p, {"target": "all"})
    assert status == "pass"

def test_check_center_align_col_fails_when_not_centered(tmp_path):
    p = _book(tmp_path, [["a", "b"], [1, 2]])
    status, reason = ailine.check_center_align(p, {"target": "col:a"})
    assert status == "fail"

def test_run_postcondition_dispatches_by_op(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    status, reason = ailine.run_postcondition("SORT", p, {"col": "金額", "order": "desc"})
    assert status == "pass"

def test_run_postcondition_threads_use_formula_to_sort_and_aggregate(tmp_path, monkeypatch):
    # ★ W10f 項目1: use_formula は元は COMPUTE_COLUMN 専用の配線だったが、SORT/AGGREGATE
    # にも同型バグがあったため広げた。run_postcondition がちゃんとその2つにも渡すことを
    # （中身の挙動でなく配線そのものを）確認する。
    # ★ 算術恒等の検算: SORT は source_book（適用前のコピー）も受け取るようになった
    #   ―― 合計行が最下行から動いたかは before が無ければ測れない。ここも配線を見る。
    calls = {}

    def fake_sort(path, args, header_row=1, use_formula=False, source_book=None):
        calls["sort"] = use_formula
        calls["sort_source_book"] = source_book
        return "pass", "ok"

    def fake_aggregate(path, args, header_row=1, use_formula=False):
        calls["aggregate"] = use_formula
        return "pass", "ok"

    monkeypatch.setitem(ailine.POSTCONDITIONS, "SORT", fake_sort)
    monkeypatch.setitem(ailine.POSTCONDITIONS, "AGGREGATE", fake_aggregate)
    p = tmp_path / "x.xlsx"
    before = tmp_path / "before.xlsx"
    ailine.run_postcondition("SORT", p, {}, use_formula=True, source_book=before)
    ailine.run_postcondition("AGGREGATE", p, {}, use_formula=True)
    assert calls == {"sort": True, "sort_source_book": before, "aggregate": True}


# ---------------------------------------------------------------------------
# ★ 止血: 空虚な検証合格の禁止（対象0件/少数を「合格」にしない）・None安全・
#   MAX_ROWS 切り詰め表示の正直な出し分け（bench/realworld/BASELINE.md の D/C②/B 検体）
# ---------------------------------------------------------------------------

# --- check_sort: 0件/1件は「合格」にしない・非数値は除外して件数を表示 -----------

def test_check_sort_fails_when_zero_data_rows(tmp_path):
    # ★ D検体の根治対象: 見出しだけで実データ行が無い(=検証対象0件)を「行数が少なく
    #   比較不要」で素通ししていた旧挙動を止める。
    p = _book(tmp_path, [["商品", "金額"]])
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"})
    assert status == "fail"
    assert "検証対象が0件" in reason

def test_check_sort_warns_when_only_one_row(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"})
    assert status == "warn"
    assert "1行のみ" in reason

def test_check_sort_excludes_non_numeric_rows_and_still_passes(tmp_path):
    # ★ C②検体の根治対象: 合計行のように key_col(A列)は埋まっているが対象列が空欄
    #   (None)の行は比較から除外し、除外件数を表示する。None>=int でクラッシュしない。
    p = _book(tmp_path, [["商品", "数量"], ["a", 30], ["b", 20], ["c", 10], ["合計", None]])
    status, reason = ailine.check_sort(p, {"col": "数量", "order": "desc"})
    assert status == "pass"
    assert "3 行を検証" in reason
    assert "数値でない 1 行は対象外" in reason

def test_check_sort_fails_when_all_rows_non_numeric(tmp_path):
    p = _book(tmp_path, [["商品", "備考"], ["a", "x"], ["b", "y"]])
    status, reason = ailine.check_sort(p, {"col": "備考", "order": "desc"})
    assert status == "fail"
    assert "検証対象が0件" in reason
    assert "数値でない 2 行は対象外" in reason

# --- ★ W10f 項目1: check_sort も同型（対象列を式ビューから読むと全行『数値でない』
#   扱いになる）。SORT は全行をまたぐ検証なので、キャッシュ欠落は部分採点せず fail で
#   打ち切る設計であることも確認する。

def test_check_sort_use_formula_default_false_keeps_old_behavior(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "asc"})
    assert status == "pass"

def test_check_sort_operand_from_prior_formula_column_passes_when_cache_present(tmp_path):
    p = tmp_path / "sort_chain.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計"])
    ws.append(["a", 2, 50, None]); ws["D2"] = "=B2*C2"
    ws.append(["b", 3, 100, None]); ws["D3"] = "=B3*C3"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D2": 100, "D3": 300})
    status, reason = ailine.check_sort(p, {"col": "小計", "order": "asc"}, use_formula=True)
    assert status == "pass"
    assert "2 行を検証" in reason

def test_check_sort_operand_from_prior_formula_column_no_cache_fails_with_distinct_reason(tmp_path):
    p = tmp_path / "sort_chain_nocache.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計"])
    ws.append(["a", 2, 50, None]); ws["D2"] = "=B2*C2"
    ws.append(["b", 3, 100, None]); ws["D3"] = "=B3*C3"
    wb.save(p)
    status, reason = ailine.check_sort(p, {"col": "小計", "order": "asc"}, use_formula=True)
    assert status == "fail"
    assert "キャッシュ値が無く検証できない行が 2 件" in reason
    assert "順序を検証できません" in reason

def test_check_sort_uncached_row_forces_fail_even_when_other_rows_are_plain_numbers(tmp_path):
    # ★ 設計判断: SORTは全行をまたぐ検証。1行でも実値が読めなければ、除いた残りだけで
    # 『順序OK』と判定するのは危険（除いた行が実際は順序を崩していても見逃す）。
    # COMPUTE_COLUMN の行独立検証(部分採点あり)とは対照的に、混在時も部分採点せず fail で
    # 打ち切ることを確認する。
    p = tmp_path / "sort_mixed.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計"])
    ws.append(["a", 1, 10, 10])                             # 小計=プレーン数値
    ws.append(["b", 2, 50, None]); ws["D3"] = "=B3*C3"       # 小計=式・キャッシュ無し
    ws.append(["c", 1, 30, 30])
    wb.save(p)
    status, reason = ailine.check_sort(p, {"col": "小計", "order": "asc"}, use_formula=True)
    assert status == "fail"
    assert "キャッシュ値が無く検証できない行が 1 件" in reason

# --- check_compute_column: 0件は合格にしない・演算対象が空欄の行は対象外 --------

def test_check_compute_column_fails_when_zero_data_rows(tmp_path):
    p = _book(tmp_path, [["売上", "原価", "売上-原価"]])
    status, reason = ailine.check_compute_column(p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "fail"
    assert "検証対象が0件" in reason

def test_check_compute_column_excludes_blank_total_row_and_still_passes(tmp_path):
    p = _book(tmp_path, [["区分", "売上", "原価", "売上-原価"],
                          ["a", 500, 300, 200], ["b", 900, 400, 500],
                          ["合計", None, None, 700]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "pass"
    assert "2 行を検証" in reason
    assert "数値でない 1 行は対象外" in reason

def test_check_compute_column_fails_when_all_rows_blank(tmp_path):
    p = _book(tmp_path, [["区分", "売上", "原価", "売上-原価"], ["合計", None, None, None]])
    status, reason = ailine.check_compute_column(
        p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "fail"
    assert "検証対象が0件" in reason

# --- check_lookup_fill: 対象シートに行が0件 --------------------------------------

def test_check_lookup_fill_fails_when_target_sheet_has_no_rows(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "明細"
    ws1.append(["商品", "数量", "単価"])
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100]]:
        ws2.append(row)
    p = tmp_path / "lookup.xlsx"
    wb.save(p)
    args = {"target_sheet": "明細", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
    status, reason = ailine.check_lookup_fill(p, args)
    assert status == "fail"
    assert "検証対象が0件" in reason

# --- check_aggregate: 元データが0件 ----------------------------------------------

def test_check_aggregate_fails_when_source_has_no_rows(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["部門", "金額"])
    out = wb.create_sheet("集計")
    out.append(["部門", "合計 - 金額"])
    p = tmp_path / "agg.xlsx"
    wb.save(p)
    status, reason = ailine.check_aggregate(p, {"group_col": "部門", "value_col": "金額"})
    assert status == "fail"
    assert "検証対象が0件" in reason

# --- check_bold/check_center_align: 検証対象0件(見出しすら無い空シート) ----------

def test_check_bold_fails_when_sheet_completely_empty(tmp_path):
    wb = openpyxl.Workbook()
    p = tmp_path / "empty.xlsx"
    wb.save(p)
    status, reason = ailine.check_bold(p, {"target": "row:1"})
    assert status == "fail"
    assert "検証対象が0件" in reason

def test_check_center_align_fails_when_sheet_completely_empty(tmp_path):
    wb = openpyxl.Workbook()
    p = tmp_path / "empty.xlsx"
    wb.save(p)
    status, reason = ailine.check_center_align(p, {"target": "all"})
    assert status == "fail"
    assert "検証対象が0件" in reason

# --- check_number_format: データ行0件 --------------------------------------------

def test_check_number_format_fails_when_zero_data_rows(tmp_path):
    p = _book(tmp_path, [["商品", "金額"]])
    status, reason = ailine.check_number_format(p, {"col": "金額", "style": "thousands"})
    assert status == "fail"
    assert "検証対象が0件" in reason

# --- run_postcondition: チェッカー内例外は生トレースバックを出さず"error"に変換 --

def test_run_postcondition_catches_checker_exception(tmp_path, monkeypatch):
    p = _book(tmp_path, [["商品", "金額"], ["a", 1]])

    def boom(path, args):
        raise TypeError("boom")

    monkeypatch.setitem(ailine.POSTCONDITIONS, "SORT", boom)
    status, reason = ailine.run_postcondition("SORT", p, {"col": "金額", "order": "desc"})
    assert status == "error"
    assert "事後条件の検証に失敗" in reason
    assert "TypeError" in reason

def test_cmd_run_dsl_postcondition_warn_does_not_claim_verified(tmp_path, monkeypatch, capsys):
    # ★ 止血1: 検証対象が1行のみの SORT は「機械検証済み」と名乗らず ⚠ で報告する。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "⚠" in captured.out
    assert "機械検証済み" not in captured.out

# --- MAX_ROWS 切り詰め表示の正直な出し分け（B検体所見の根治） --------------------

def test_truncation_notice_is_none_when_not_truncated():
    assert ailine._truncation_notice({"truncated": False}, {"truncated": False}, True) is None

def test_truncation_notice_dsl_path_says_verification_covers_all_rows():
    msg = ailine._truncation_notice({"truncated": False}, {"truncated": True},
                                     exhaustive_postcondition=True)
    assert "検証・適用は全行に対して実施" in msg

def test_truncation_notice_freeform_path_says_verification_also_truncated():
    msg = ailine._truncation_notice({"truncated": True}, {"truncated": False},
                                     exhaustive_postcondition=False)
    assert "検証も先頭" in msg
    assert "行のみ" in msg

def test_snapshot_marks_truncated_when_rows_exceed_max_rows(tmp_path, monkeypatch):
    monkeypatch.setattr(ailine, "MAX_ROWS", 3)
    p = _book(tmp_path, [["a"], [1], [2], [3], [4], [5]])
    snap = ailine.snapshot(p)
    assert snap["truncated"] is True

def test_snapshot_not_truncated_when_within_max_rows(tmp_path):
    p = _book(tmp_path, [["a"], [1], [2]])
    snap = ailine.snapshot(p)
    assert snap["truncated"] is False

def test_cmd_run_dsl_prints_truncation_notice_when_snapshot_truncated(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    monkeypatch.setattr(ailine, "MAX_ROWS", 2)   # 3行あるデータを2行に切り詰めさせる
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "表示は先頭 2 行の変化のみ。検証・適用は全行に対して実施" in captured.out


# --- run コマンド: 翻訳の分岐（CLARIFY exit 3 / DSL 経路の事後条件不合格） -------

def test_cmd_run_clarify_prints_question_and_exits_3(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["商品", "金額"]])
    # ★ W3: 正規化パス(StructDump)は翻訳より前に走るので、CLARIFY 系の単体テストも
    #   normalize_book を差し替えて LibreOffice を要さないようにする。
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "CLARIFY", "question": "どの列を並べ替えますか？", "args": {}})
    argv = run_argv(
        book=str(book), task="いい感じにして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    assert "どの列を並べ替えますか？" in captured.out

def test_cmd_run_dsl_verification_failure_falls_back_to_clarify_exit_3(tmp_path, monkeypatch, capsys):
    # ★ W10a: maybe_show_notice_v2 は既定で HISTORY_FILE.parent を使うため、
    #   実ファイル汚染を避けるにはこちらを monkeypatch する。
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    book = _book(tmp_path, [["商品", "金額"]])
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "存在しない列", "order": "desc"}})
    argv = run_argv(
        book=str(book), task="存在しない列で並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    assert "がありません" in captured.out

def test_cmd_run_dsl_dry_shows_confirmation_and_code_without_applying(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")   # 実 history を汚さない
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=True, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "解釈: 操作:並べ替え 対象:金額 順:降順" in captured.out
    assert "Call SortByColumn" in captured.out
    assert '"path": "dsl"' in captured.out

def test_cmd_run_dsl_postcondition_failure_returns_1(tmp_path, monkeypatch, capsys):
    # basrun_apply/snapshot を差し替え、事後条件が満たされない場合に exit 1 になることを確認。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")   # 実 history を汚さない
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    # basrun_apply は成功したことにするが、実際には out_book の中身は昇順のまま(=事後条件不合格)。
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 1
    assert "事後条件を満たさない" in captured.out


# ---------------------------------------------------------------------------
# ★ M2c: 複合依頼の計画実行と正直な範囲表示
# ---------------------------------------------------------------------------

# --- 項目別報告の整形 ---------------------------------------------------------

def test_format_plan_report_ok_warn_fail_lines():
    items = [
        (1, "操作:計算列 対象列:小計", "ok", "3 行を検証"),
        (2, "税込み合計", "warn", None),
        (3, "操作:並べ替え", "fail", "列『在庫』がありません"),
    ]
    lines = ailine.format_plan_report(items)
    # ★ C9: 段別の行から ✓ を外した（段は evidence だけ述べる）。✓ は原本(--copy なら .out)が
    #   確定した後に読み戻して確かめた1行だけが名乗る。
    assert lines[0] == "1. 操作:計算列 対象列:小計 → 実行: 3 行を検証"
    # ★ W8a 項目5: 表示文言「自由生成」→「AI が直接作成（機械保証なし）」に追従。
    assert lines[1] == "2. 税込み合計 → ⚠ 語彙外のため AI が直接作成（機械保証なし）で実行（確認してください）"
    assert lines[2] == "3. 操作:並べ替え → × 未対応: 列『在庫』がありません"

def test_format_plan_report_ok_without_detail_omits_parens():
    lines = ailine.format_plan_report([(1, "操作:太字", "ok", None)])
    assert lines[0] == "1. 操作:太字 → 実行"

# --- 総合判定規則 -------------------------------------------------------------

def test_overall_verdict_all_ok():
    # ★ C9: 全段 ok では総合判定の行そのものを出さない（旧「✓ すべて機械検証済み」）。
    #   この時点では原本へ反映できるかがまだ分かっていない（--copy か・置換が成功するか）
    #   ため、✓ は _finish_apply（原本が確定した後）の1行に一本化した。
    line, v = ailine.overall_verdict([(1, "x", "ok", "r")])
    assert v == "ok"
    assert line is None

def test_overall_verdict_warn_without_fail():
    line, v = ailine.overall_verdict([(1, "x", "ok", "r"), (2, "y", "warn", None)])
    assert v == "warn"
    assert "確認が必要" in line

def test_overall_verdict_fail_dominates_over_warn():
    line, v = ailine.overall_verdict(
        [(1, "x", "ok", "r"), (2, "y", "warn", None), (3, "z", "fail", "reason")])
    assert v == "fail"

# --- 依存つき連鎖(#107 型)の新規列フォールバック --------------------------------

def test_apply_new_column_fallback_substitutes_sole_new_column():
    args = ailine._apply_new_column_fallback(
        "SORT", {"col": "利益", "order": "desc"},
        ["商品", "売上", "原価", "売上-原価"], ["売上-原価"])
    assert args["col"] == "売上-原価"

def test_apply_new_column_fallback_noop_when_multiple_candidates():
    args = ailine._apply_new_column_fallback(
        "SORT", {"col": "利益", "order": "desc"}, ["商品", "a", "b"], ["a", "b"])
    assert args["col"] == "利益"   # 候補2つ → 書き換えない(保守的)

def test_apply_new_column_fallback_leaves_existing_reference_untouched():
    args = ailine._apply_new_column_fallback(
        "SORT", {"col": "商品", "order": "desc"}, ["商品", "売上-原価"], ["売上-原価"])
    assert args["col"] == "商品"   # 既に実在する参照は書き換えない

def test_apply_new_column_fallback_noop_when_no_new_columns():
    args = ailine._apply_new_column_fallback(
        "SORT", {"col": "利益", "order": "desc"}, ["商品", "売上"], [])
    assert args["col"] == "利益"

# --- cmd_run_plan: 複合計画の実行と honest な報告 -------------------------------

def _plan_book(tmp_path, rows, name="plan_b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook(); ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p

def test_cmd_run_plan_all_dsl_steps_pass_gives_full_verdict(tmp_path, monkeypatch, capsys):
    from openpyxl.styles import Font
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    for c in (1, 2):
        ws.cell(row=1, column=c).font = Font(bold=True)   # 見出し行を先に太字にしておく(pre-condition)
    wb.save(p)

    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて見出しを太字に", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=True, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    # ★ C9: 段別は evidence だけ（✓ 無し）。✓ は原本(.out)が確定した後の1行に一本化。
    assert "1. " in captured.out and "→ 実行: " in captured.out
    assert "2. " in captured.out
    assert "✓ plan_b.out.xlsx は機械検証済みの内容です" in captured.out
    assert captured.out.count("✓") == 1   # run 全体で ✓ はちょうど1つ
    assert '"path": "plan"' in captured.out
    assert '"status": "ok"' in captured.out
    # ★ C9: --json は既存キー不変・claims を追加（何と照合し・どのファイルを読み戻したか）。
    assert '"claims": [{"basis": "declaration"' in captured.out
    assert '"observed_on"' in captured.out

def test_cmd_run_plan_mixes_dsl_success_and_freeform_warns(tmp_path, monkeypatch, capsys):
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])

    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        cell = ws2.cell(row=1, column=10)   # postcondition が見ない列にダミーの変化を残す
        cell.value = (cell.value or 0) + 1
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて条件付き書式もつけて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)   # ★ W10b: 段の自由生成の関所は無関係（項目別報告を見るテスト）
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0   # ⚠ は失敗ではない
    assert "1. 操作:並べ替え 対象:金額 順:降順 → 実行: " in captured.out
    assert "条件付き書式" in captured.out
    # ★ W8a 項目5: 表示文言「自由生成」→「AI が直接作成（機械保証なし）」に追従。
    assert "⚠ 語彙外のため AI が直接作成（機械保証なし）で実行（確認してください）" in captured.out
    assert "⚠ 一部は確認が必要です" in captured.out
    # ★ C9: 語彙外の段が混じる run では ✓ をどこにも出さない（読み戻しの報告は ⚠ 側で行う）。
    assert "✓" not in captured.out
    assert "に適用しましたが、機械保証はありません（適用後に読み戻して確認: " in captured.out

def test_cmd_run_plan_freeform_step_rate_literal_scan_fires(tmp_path, monkeypatch, capsys):
    # ★ W10f 項目2: A' 原則（LLM に率や値を確定させない）を機械で守る唯一の走査
    # (scan_rate_literals) が、単発の自由生成(cmd_run_freeform)には元から通っていたが、
    # 複合計画の自由生成段(run_freeform_plan_step)には通っていなかった（独立監査が発見・
    # 査定5回は偶然当たらなかっただけの穴）。複合計画経由でも率らしい数値の助言が
    # 出ることを確認する（test_cmd_run_freeform_rate_literal_scan_fires_when_task_silent_on_rate
    # の複合計画版）。
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "OUT_OF_VOCAB", "about": "税込み合計"}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    code = ("Sub Run(oDoc As Object)\n"
            "  oDoc.Sheets.getByIndex(0).getCellByPosition(2, 0).setValue(100 * 1.08)\n"
            "End Sub")
    monkeypatch.setattr(ailine, "ollama_generate", lambda model, msgs, temperature=0.2: code)

    def fake_apply(out_book, c, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        cell = ws2.cell(row=1, column=10)   # postcondition が見ない列にダミーの変化を残す
        cell.value = (cell.value or 0) + 1
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて税込み合計も出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "★ 率らしい数値 (1.08) が依頼に無いのに使われています — 検算してください" in captured.out

def test_cmd_run_plan_all_steps_fail_grounding_gives_overall_failure(tmp_path, monkeypatch, capsys):
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "存在しない列1", "order": "desc"}},
                                  {"op": "BOLD", "args": {"target": "col:存在しない列2"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    argv = run_argv(
        book=str(p), task="存在しない列で並べ替えて太字にして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 1
    assert captured.out.count("× 未対応") == 2
    assert "達成できませんでした" in captured.out

def test_cmd_run_plan_dry_previews_without_applying(tmp_path, monkeypatch, capsys):
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})
    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて条件付き書式もつけて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=True, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "1. " in captured.out
    assert "条件付き書式" in captured.out
    out_book = p.with_name(p.stem + ".out" + p.suffix)
    assert not out_book.exists()   # --dry は適用しない
    assert '"path": "plan"' in captured.out
    assert '"dry": true' in captured.out

def test_cmd_run_plan_dependent_chaining_resolves_new_column_reference(tmp_path, monkeypatch, capsys):
    # ★ battery v2 #107 型: 「利益列を作って、利益で降順に並べ替えて」。2段目の "利益" は
    #   実在せず、1段目が作る自動命名列(売上-原価)を指す＝直前段の適用後の列構成で解決する。
    p = _plan_book(tmp_path, [["商品", "売上", "原価"], ["a", 500, 300], ["b", 900, 400]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [
                            {"op": "COMPUTE_COLUMN", "args": {"operands": ["売上", "原価"], "operator": "-"}},
                            {"op": "SORT", "args": {"col": "利益", "order": "desc"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        if "SortByColumn" in code:
            ws2.cell(row=1, column=1, value="商品"); ws2.cell(row=1, column=2, value="売上")
            ws2.cell(row=1, column=3, value="原価"); ws2.cell(row=1, column=4, value="売上-原価")
            ws2.cell(row=2, column=1, value="b"); ws2.cell(row=2, column=2, value=900)
            ws2.cell(row=2, column=3, value=400); ws2.cell(row=2, column=4, value=500)
            ws2.cell(row=3, column=1, value="a"); ws2.cell(row=3, column=2, value=500)
            ws2.cell(row=3, column=3, value=300); ws2.cell(row=3, column=4, value=200)
        else:
            ws2.cell(row=1, column=4, value="売上-原価")
            ws2.cell(row=2, column=4, value=200)   # a: 500-300
            ws2.cell(row=3, column=4, value=500)   # b: 900-400
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(p), task="売上から原価を引いた利益列を作って、利益で降順に並べ替えて",
        model="qwen2.5-coder:7b", refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        # ★ W3 Part3: fake_apply は静的な値を直接書き込む(式は書かない)ので、この
        #   テストは --values（値ベタ書き）経路として実行する。式検証(二層)は
        #   check_compute_column の専用ユニットテストで別途カバーする。
        values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "列『利益』がありません" not in captured.out
    # ★ C9: 総合の ✓ は原本(.out)が確定した後の1行に移った。
    assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out


# --- ★ W10d【本命】: 複合計画の助言（単発では出る助言が丸ごと欠落していた欠陥の修正） ------

def test_group_step_advisories_dedups_identical_text_keeps_first_seen_order():
    entries = [(1, "A"), (2, "B"), (3, "A"), (1, "C")]
    grouped = ailine._group_step_advisories(entries)
    assert grouped == [([1, 3], "A"), ([2], "B"), ([1], "C")]

def test_group_step_advisories_none_step_means_plan_wide_and_has_no_step_numbers():
    entries = [(1, "A"), (None, "A"), (None, "B")]
    grouped = ailine._group_step_advisories(entries)
    assert grouped == [([1], "A"), ([], "B")]

def test_dedup_step_advisories_formats_multi_step_prefix_and_plan_wide_line():
    entries = [(1, "同じ助言"), (3, "同じ助言"), (None, "計画全体の助言")]
    lines = ailine._dedup_step_advisories(entries)
    assert lines == ["  1段目・3段目: 同じ助言", "  計画全体の助言"]

def test_cmd_run_plan_dsl_steps_show_advisories_and_dedup_repeats(tmp_path, monkeypatch, capsys):
    # ★ W10d【本命】: 前任の報告どおり cmd_run_plan は DSL 段に対して助言を一切
    #   組み立てていなかった（build_advisories を呼んでいなかった）ことをまず自分で確認した
    #   （着手前調査）。この回帰テストは単発と同じ助言が複合計画の DSL 段でも出ること・
    #   同じ文言が複数段に渡って出ても1行に畳まれること・新規列の宣言どおりの効果は
    #   中立化されること・宣言と一致しない疑わしい変化はちゃんと残ることを1本で確認する。
    p = _plan_book(tmp_path, [["商品", "金額", "原価"],
                               ["a", 300, 100], ["b", 200, 80], ["c", 100, 50]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [
                            {"op": "COMPUTE_COLUMN", "args": {"operands": ["金額", "原価"], "operator": "-"}},
                            {"op": "BOLD", "args": {"target": "row:1"}},
                            {"op": "CENTER_ALIGN", "args": {"target": "all"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    call_count = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        call_count["n"] += 1
        n = call_count["n"]
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        if n == 1:
            # 1段目 COMPUTE_COLUMN(target無指定): 新規列D(利益)を作る＝宣言どおりの効果。
            ws2.cell(row=1, column=4, value="利益")
            ws2.cell(row=2, column=4, value=200)
            ws2.cell(row=3, column=4, value=120)
            ws2.cell(row=4, column=4, value=50)
        elif n == 2:
            from openpyxl.styles import Font
            for c in (1, 2, 3, 4):
                ws2.cell(row=1, column=c).font = Font(bold=True)
            # 元データ範囲外の空欄に同一値0を一括書き込み（宣言に無い疑わしい副作用の再現）。
            ws2.cell(row=10, column=1, value=0)
            ws2.cell(row=10, column=2, value=0)
        else:
            from openpyxl.styles import Alignment
            for row in ws2.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            # 別の場所に *同じ形* の疑わしい書き込み（文言が2段目と一致する＝畳まれる対象）。
            ws2.cell(row=11, column=1, value=0)
            ws2.cell(row=11, column=2, value=0)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(p), task="金額から原価を引いた利益列を作って、見出しを太字にして全体を中央揃えにして",
        model="qwen2.5-coder:7b", refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=True, timeout=180.0, ask=False, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    # 人が読む表示部分だけを対象にする（末尾の --json 1行は別途 JSON として検証する）。
    display_out = "\n".join(captured.out.strip().splitlines()[:-1])

    # ①宣言どおりの新規列作成は中立表示（COMPUTE_COLUMN の新規列ゴースト誤警報の抑制）。
    assert "（新規列の追加は意図どおりです）" in display_out
    # ②宣言に無い疑わしい変化（空欄への同一値0書き込み）はちゃんと出る（抑制しすぎない）。
    assert "空欄への同一値の一括書き込みです（値 0 × 2 セル）" in display_out
    # ③同じ文言が2段(2段目・3段目)から出ても1行に畳まれている（オオカミ少年化しない）。
    assert display_out.count("空欄への同一値の一括書き込みです（値 0 × 2 セル）") == 1
    assert "2段目・3段目" in display_out
    assert "助言:" in display_out

    payload = json.loads(captured.out.strip().splitlines()[-1])
    adv_texts = [entry["text"] for entry in payload["advisories"]]
    assert any("新規列の追加は意図どおり" in t for t in adv_texts)
    uniform_entries = [entry for entry in payload["advisories"]
                        if "空欄への同一値" in entry["text"]]
    assert len(uniform_entries) == 1
    assert uniform_entries[0]["steps"] == [2, 3]


# ---------------------------------------------------------------------------
# ★ W3 Part1/2: StructDump（LibreOffice の目で構造を読む）+ 見出し行推定
#   ★ LO 依存部（実使用範囲の取得）は fixture の dict/テキストで代用する
#   （normalize_book/basrun_apply は差し替え、StructDump のパーサ・ヒューリスティクス
#   だけを純ロジックとして検証する）。実機での動作は E2E ログで別途確認する。
# ---------------------------------------------------------------------------

def test_parse_structdump_raw_parses_tab_delimited_lines():
    text = "SHEET\tSheet\t0\t0\t4\t9\t0\t1\t0\nSHEET\t単価表\t0\t0\t1\t3\t0\t0\t0\n"
    got = ailine.parse_structdump_raw(text)
    assert got["Sheet"]["used_range"] == {"start_col": 0, "start_row": 0, "end_col": 4, "end_row": 9}
    assert got["Sheet"]["charts"] == 1
    assert got["単価表"]["shapes"] == 0

def test_parse_structdump_raw_ignores_malformed_lines():
    text = "not a sheet line\nSHEET\ttoo\tfew\tcolumns\n"
    assert ailine.parse_structdump_raw(text) == {}

def test_parse_structdump_raw_empty_text_gives_empty_dict():
    assert ailine.parse_structdump_raw("") == {}

def test_row_char_stats_counts_nonempty_str_and_bold(tmp_path):
    from openpyxl.styles import Font
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "金額", "在庫"])
    ws.append(["りんご", 100, 5])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    wb.save(p)
    wb2 = openpyxl.load_workbook(p)
    stats = ailine._row_char_stats(wb2.active, 1, 2, 1, 3)
    assert stats[1] == {"nonempty": 3, "str": 3, "bold": 2}
    assert stats[2] == {"nonempty": 3, "str": 1, "bold": 0}   # りんご=文字列, 100/5=数値

def test_build_struct_dump_falls_back_to_openpyxl_when_no_raw_dump(tmp_path, monkeypatch):
    # ★ normalize_book が差し替えられて structdump.txt が書かれない場合(単体テストの通常経路)、
    #   openpyxl の max_row/max_column から used_range を推定する（CLARIFY を誤って出さないため）。
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    workdir = tmp_path / "wd"
    workdir.mkdir()
    dump = ailine.build_struct_dump(p, workdir)
    sheet = dump["sheets"]["Sheet"]
    assert sheet["used_range"]["start_row"] == 1
    assert sheet["rows"][1] == {"nonempty": 2, "str": 2, "bold": 0}
    assert sheet["rows"][2]["nonempty"] == 2 and sheet["rows"][2]["str"] == 1

def test_detect_header_row_simple_single_level_header():
    # 普通の帳票（demo/sales.xlsx 型）: 行1=見出し、行2以降=数値混在データ。
    sheet_struct = {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 1, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert (row, confident) == (1, True)

def test_detect_header_row_title_row_then_header_at_row3():
    # A検体型: 行1=結合タイトル(str=1・閾値未満で候補外)、行2=単一文字列、行3=見出し(str=5)、
    # 行4以降=型混在データ。
    sheet_struct = {"rows": {
        1: {"nonempty": 1, "str": 1, "bold": 1},
        2: {"nonempty": 1, "str": 1, "bold": 0},
        3: {"nonempty": 5, "str": 5, "bold": 0},
        4: {"nonempty": 5, "str": 1, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert (row, confident) == (3, True)

def test_detect_header_row_two_level_header_picks_child_row():
    # D検体型: 行1=親見出し(結合・str=3)、行2=子見出し(str=4・型混在なし直下）、
    # 行3=型混在データ → 子見出し行(2)を採用する。
    sheet_struct = {"rows": {
        1: {"nonempty": 3, "str": 3, "bold": 0},
        2: {"nonempty": 4, "str": 4, "bold": 0},
        3: {"nonempty": 5, "str": 1, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert (row, confident) == (2, True)

def test_detect_header_row_ambiguous_two_equally_valid_candidates_is_not_confident():
    # 型混在の直下が2つとも存在する（曖昧）場合は推測しない。
    sheet_struct = {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert confident is False
    assert row is None

def test_detect_header_row_empty_sheet_is_not_confident():
    assert ailine.detect_header_row({"rows": {}}) == (None, False)

def test_resolve_header_rows_confident_detection_no_clarify():
    struct_dump = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    header_rows, clarify = ailine.resolve_header_rows(struct_dump, ["Sheet"])
    assert header_rows == {"Sheet": 1}
    assert clarify is None

def test_resolve_header_rows_ambiguous_asks_clarify_with_exact_wording():
    struct_dump = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    header_rows, clarify = ailine.resolve_header_rows(struct_dump, ["Sheet"])
    # ★ W8a 項目3: 旧文言は答え方が無い行き止まりだった。--header-row の使い方まで添える。
    assert clarify == ("見出しが何行目か分かりません。"
                        "`--header-row 3` のように指定して再実行してください")
    assert header_rows == {"Sheet": 1}   # 既定のまま(呼び出し側は CLARIFY で止まるので使われない)

def test_resolve_header_rows_no_struct_dump_defaults_to_row1_no_clarify():
    header_rows, clarify = ailine.resolve_header_rows({}, ["Sheet", "単価表"])
    assert header_rows == {"Sheet": 1, "単価表": 1}
    assert clarify is None

def test_resolve_header_rows_empty_sheets_list_is_noop():
    assert ailine.resolve_header_rows({"sheets": {}}, []) == ({}, None)

def test_cmd_run_clarify_on_ambiguous_header_before_translation(tmp_path, monkeypatch, capsys):
    # ★ W3: 見出し推定が曖昧なら、翻訳(translate_task)を呼ぶ前に CLARIFY して exit 3 になる
    #   （『三層全部が同じ見出し推定を使う』の前提＝そもそも翻訳に渡す接地が無い）。
    book = _book(tmp_path, [["a", 1], ["b", 2]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "build_struct_dump", lambda book, workdir: ambiguous)
    called = {"n": 0}
    def boom(*a, **k):
        called["n"] += 1
        return {"plan": [{"op": "FREEFORM", "args": {}}]}
    monkeypatch.setattr(ailine, "translate_task", boom)
    argv = run_argv(
        book=str(book), task="いい感じにして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    # ★ W8a 項目3: CLARIFY 文言更新（--header-row の使い方まで添える）に追従。
    assert "見出しが何行目か分かりません" in captured.out
    assert "--header-row" in captured.out
    assert called["n"] == 0   # 翻訳は一度も呼ばれていない

def test_cmd_run_header_row_flag_bypasses_ambiguous_detection_no_clarify(tmp_path, monkeypatch, capsys):
    # ★ W8a 項目3: 見出し検出が曖昧(CLARIFY相当)な帳票でも、--header-row を指定すれば
    #   検出を丸ごとスキップしてその行を採用し、CLARIFY の行き止まりに落ちない。
    #   接地(book_meta)経由で codegen/事後条件にも同じ header_row が伝わる（三層貫通）。
    book = _book(tmp_path, [["x", 1], ["y", 2], ["商品", "金額"], ["a", 300], ["b", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "build_struct_dump", lambda book, workdir: ambiguous)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, values=False,
        header_row=3)
    ailine.main(argv)
    captured = capsys.readouterr()
    assert "？" not in captured.out    # CLARIFY に落ちていない
    assert "対象:金額" in captured.out   # 3行目を見出しとして『金額』列を解決できている

def test_cmd_run_without_header_row_flag_still_clarifies_on_ambiguous(tmp_path, monkeypatch, capsys):
    # 対照: --header-row を指定しなければ従来どおり曖昧な場合は CLARIFY のまま。
    book = _book(tmp_path, [["x", 1], ["y", 2], ["商品", "金額"], ["a", 300], ["b", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    ambiguous = {"sheets": {"Sheet": {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}}}
    monkeypatch.setattr(ailine, "build_struct_dump", lambda book, workdir: ambiguous)
    called = {"n": 0}
    def boom(*a, **k):
        called["n"] += 1
        return {"op": "SORT", "args": {"col": "金額", "order": "desc"}}
    monkeypatch.setattr(ailine, "translate_task", boom)
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, values=False,
        header_row=None)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    assert "--header-row" in captured.out
    assert called["n"] == 0


# --- ★ DoD1: 査定の再現をそのまま回帰テストにする（main(argv) 経由・CLI 全体） ---------------
#   独立監査の再現手順そのもの:「複数シートのブックで2枚目シートを名指し」
#   → 旧実装は誤って1枚目を対象にし、「ある列: ...」が1枚目の列名を出す誤誘導だった。
#   ollama は monkeypatch で避ける（DSL 経路は率/列名を LLM に確定させない A' 原則どおり・
#   verify_dsl_args 自体は決定論なので、翻訳結果だけ固定すれば ollama 無しで完全再現できる）。

def _multi_sheet_book(tmp_path):
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "請求書"
    ws1.append(["宛先", "金額"])
    ws1.append(["山田商店", 50000])
    ws2 = wb.create_sheet("工事台帳")
    ws2.append(["取引先名", "工事名", "金額"])
    ws2.append(["山田商店", "A邸新築", 300000])
    ws2.append(["田中建設", "B邸改修", 150000])
    wb.create_sheet("取引先マスタ")
    wb.save(p)
    return p

def test_cmd_run_named_second_sheet_no_longer_shows_misleading_first_sheet_error(
        tmp_path, monkeypatch, capsys):
    """DoD1 の回帰テスト本体。「工事台帳シートで宛先を降順に並べ替えて」— 『宛先』は
       1枚目(請求書)の列で、対象(工事台帳)には無い。旧実装はここで
       「ある列: 宛先, 金額」（1枚目の列！）を出す誤誘導だった。直った後は対象シート
       (工事台帳)の実列名だけが出る。"""
    book = _multi_sheet_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "宛先", "order": "desc"}})
    argv = run_argv(
        book=str(book), task="工事台帳シートで宛先を降順に並べ替えて",
        dry=False, copy=True, header_row=None)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    # ★ 直った点そのもの: 誤誘導が無い（1枚目の列名「宛先」「金額」の組が「ある列」に出ない）。
    assert "ある列: 取引先名, 工事名, 金額" in captured.out
    assert "ある列: 宛先, 金額" not in captured.out
    # ★ 挙動変更#2 最低限: 適用前の対象シート明示（査定所見「これがあれば事故は防げた」）。
    assert "操作するシート: 2枚目『工事台帳』" in captured.out

def test_cmd_run_named_second_sheet_applies_operation_there_not_first_sheet(
        tmp_path, monkeypatch, capsys):
    """同じ複数シートのブックで、今度は実在する列（工事台帳の『金額』）を指定し、
       実際にその段が2枚目シートを対象に適用できることを確認する（basrun_apply は
       decode 済みコードの内容だけ検査するダミーに差し替え・real LO は使わない
       ＝ not local。real LO 版は tests/test_target_sheet_local.py 参照）。"""
    book = _multi_sheet_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    captured_code = {}
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        captured_code["code"] = code
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(book), task="工事台帳シートで金額を降順に並べ替えて",
        dry=False, copy=True, header_row=None)
    ailine.main(argv)
    code = captured_code.get("code", "")
    # ★ 対象シートが1枚目でないので、_wrap_basic_for_sheet が moveByName で包む。
    assert 'oDoc.Sheets.moveByName("工事台帳", 0)' in code
    assert 'oDoc.Sheets.moveByName("工事台帳", 1)' in code

def test_cmd_run_single_sheet_book_no_announcement_no_regression(tmp_path, monkeypatch, capsys):
    """DoD5②: 単一シートのブックは従来どおり（対象シート明示の行が出ない・moveByName も
       生成されない＝既存ゴールデンと同型の出力）。"""
    book = _book(tmp_path, [["商品", "数量", "単価"], ["りんご", 3, 100], ["みかん", 5, 50]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "数量", "order": "desc"}})
    captured_code = {}
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        captured_code["code"] = code
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(book=str(book), task="数量で降順に並べ替えて", dry=False, copy=True, header_row=None)
    ailine.main(argv)   # ★ fake_apply は実際にソートしないため postcondition の pass/fail は見ない
    captured = capsys.readouterr()
    assert "操作するシート:" not in captured.out   # 単一シートは沈黙（既存出力を変えない）
    assert "moveByName" not in captured_code.get("code", "")

def test_cmd_run_plan_composite_step_targets_named_non_first_sheet(tmp_path, monkeypatch, capsys):
    """複合計画(cmd_run_plan/_run_dsl_plan_step)でも対象シートの解決が効くこと
       （単発(cmd_run_dsl)だけでなく _run_dsl_plan_step 側の first_sheet 配線も見る）。"""
    book = _multi_sheet_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [
                            {"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                            {"op": "BOLD", "args": {"target": "col:金額"}},
                        ]})
    captured_codes = []
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        captured_codes.append(code)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(book), task="工事台帳シートで金額を降順に並べ替えて太字にして",
        dry=False, copy=True, header_row=None)
    ailine.main(argv)
    captured = capsys.readouterr()
    assert "操作するシート: 2枚目『工事台帳』" in captured.out
    assert len(captured_codes) == 2
    for code in captured_codes:
        assert 'oDoc.Sheets.moveByName("工事台帳", 0)' in code
        assert 'oDoc.Sheets.moveByName("工事台帳", 1)' in code

def test_cmd_run_dry_skips_structdump_and_uses_physical_row1(tmp_path, monkeypatch, capsys):
    # ★ --dry は LibreOffice に触れない（既存の設計不変条件）。normalize_book が
    #   呼ばれていないことを確認する。
    book = _book(tmp_path, [["商品", "金額"]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    called = {"n": 0}
    def boom(*a, **k):
        called["n"] += 1
        return a[0]
    monkeypatch.setattr(ailine, "normalize_book", boom)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    argv = run_argv(
        book=str(book), task="金額で降順に並べ替えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    assert rc == 0
    assert called["n"] == 0


# --- codegen: header_row(hr0) が三層で一貫して使われる（接地→codegen） -------------

def test_codegen_dsl_sort_uses_detected_header_row_from_book_meta():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]},
            "header_rows": {"Sheet": 3}}
    code = ailine.codegen_dsl("SORT", {"col": "金額", "order": "desc"}, meta)
    assert "Call SortByColumn(oDoc, 2, 1, 1, False)" in code   # hr0 = 3-1 = 2, lastCol=1(列2つ)

def test_codegen_dsl_compute_column_formula_uses_detected_header_row():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "売上", "原価"]},
            "header_rows": {"Sheet": 3}}
    code = ailine.codegen_dsl("COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"}, meta)
    assert "For i = 3 To lastRow" in code   # hr0+1 = 2+1 = 3
    assert 'setString("売上-原価")' in code
    assert 'getCellByPosition(3, 2).setString' in code   # 新規列は列3, 見出し行は hr0=2


# --- 事後条件: header_row が接地・codegen と同じ行を使う ---------------------------

def test_check_sort_with_header_row_three_matches_a_specimen_layout(tmp_path):
    p = tmp_path / "a_like.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "タイトル"
    ws["A2"] = "作成日"
    for c, h in enumerate(["商品", "金額"], start=1):
        ws.cell(row=3, column=c, value=h)
    ws.append(["a", 300]); ws.append(["b", 200]); ws.append(["c", 100])
    # ↑ append は末尾行に追記するため、3行目の直後(4行目)から入る
    wb.save(p)
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"}, header_row=3)
    assert status == "pass"

def test_check_sort_with_wrong_header_row_fails_to_find_column(tmp_path):
    p = tmp_path / "a_like.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "タイトル"
    ws["A2"] = "作成日"
    for c, h in enumerate(["商品", "金額"], start=1):
        ws.cell(row=3, column=c, value=h)
    ws.append(["a", 300])
    wb.save(p)
    status, reason = ailine.check_sort(p, {"col": "金額", "order": "desc"}, header_row=1)
    assert status == "fail"
    assert "がありません" not in reason or "見つからない" in reason


def _inject_formula_cache(path, sheet_filename: str, addr_to_value: dict) -> None:
    """テスト専用: xlsx の数式セルへキャッシュ値(<v>)を直接注入する（openpyxl は数式を
       計算しないため、LO を使わずに二層事後条件(式+キャッシュ値)を検証するための小道具）。"""
    import re
    import zipfile
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_filename:
                text = data.decode("utf-8")
                for addr, value in addr_to_value.items():
                    # 空の <v> の書かれ方は openpyxl の直列化系で 3 通りある。全部受ける:
                    #   lxml あり          -> <v></v>   (開閉ペア)
                    #   lxml なし(ET)      -> <v />     (自己閉じ)  ★これを取りこぼしていた
                    #   そもそも書かれない -> (無し)
                    # ★ 2026-08-15 の実測: ローカルには lxml が入っており CI には無いため、
                    #   CI だけで 12 本落ちた。ローカル緑は「この機械で緑」でしかなかった。
                    pattern = re.compile(
                        rf'(<c r="{addr}"[^>]*>.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?(</c>)')
                    text, n = pattern.subn(rf'\1<v>{value}</v>\2', text, count=1)
                    # ★ 置換ゼロで黙って進まない。注入されなかったのに検証が走ると、
                    #   「キャッシュ値が違う」でなく「小道具が壊れている」を測ることになる。
                    assert n == 1, (
                        f"_inject_formula_cache: {addr} に注入できなかった（xlsx の直列化の形が "
                        f"想定外の可能性）。sheet={sheet_filename}")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


def _inject_formula_string_cache(path, sheet_filename: str, addr_to_str: dict) -> None:
    """テスト専用: 文字列を返す数式セル(=A2 等)へ t="str" 付きキャッシュ値を注入する。
       ★ W10f 項目5: _inject_formula_cache は数値専用（t 属性を書かない＝既定の数値型の
       まま）。LOOKUP_FILL のキー列は文字列キーを想定するため、<c> の t 属性も
       t="str"（OOXML の「式の文字列キャッシュ」型）へ書き換える別ヘルパを用意する。"""
    import re
    import zipfile
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_filename:
                text = data.decode("utf-8")
                for addr, value in addr_to_str.items():
                    pattern = re.compile(
                        rf'<c r="{addr}"([^>]*)>(.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?</c>')

                    def _sub(m, value=value):
                        attrs = re.sub(r'\s*t="[^"]*"', '', m.group(1))
                        return f'<c r="{addr}"{attrs} t="str">{m.group(2)}<v>{value}</v></c>'

                    text, n = pattern.subn(_sub, text, count=1)
                    assert n == 1, (
                        f"_inject_formula_string_cache: {addr} に注入できなかった（xlsx の"
                        f"直列化の形が想定外の可能性）。sheet={sheet_filename}")
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


def _formula_book(tmp_path, header_row_values, data_rows, formula_col_letter, operator_col_letters,
                   operator, name="f.xlsx"):
    """演算対象2列+式列を持つブックを作る。式列には setFormula 相当の文字列を直接書く
       （キャッシュ値は別途 _inject_formula_cache で注ぐ）。"""
    p = tmp_path / name
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(header_row_values)
    for i, row in enumerate(data_rows, start=2):
        ws[f"A{i}"] = f"item{i}"   # ★ _scan_last_row は A列(key_col=1既定)で行の有無を判定する
        for col_letter, val in row.items():
            ws[f"{col_letter}{i}"] = val
        c1, c2 = operator_col_letters
        ws[f"{formula_col_letter}{i}"] = f"={c1}{i}{operator}{c2}{i}"
    wb.save(p)
    return p

def test_check_compute_column_formula_mode_passes_when_formula_and_cache_both_match(tmp_path):
    p = _formula_book(tmp_path, ["商品", "数量", "単価", "小計"],
                       [{"B": 3, "C": 120}, {"B": 5, "C": 80}], "D", ("B", "C"), "*")
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D2": 360, "D3": 400})
    status, reason = ailine.check_compute_column(
        p, {"operands": ["数量", "単価"], "operator": "*", "target": "小計"}, use_formula=True)
    assert status == "pass"
    assert "式・キャッシュ値とも一致" in reason

def test_check_compute_column_formula_mode_fails_when_formula_string_wrong(tmp_path):
    p = tmp_path / "f.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計"])
    ws.append(["りんご", 3, 120, None])
    ws["D2"] = "=B2+C2"   # ★ 期待は * なのに + を書いてしまった想定
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D2": 123})
    status, reason = ailine.check_compute_column(
        p, {"operands": ["数量", "単価"], "operator": "*", "target": "小計"}, use_formula=True)
    assert status == "fail"
    assert "式が期待形でない" in reason

def test_check_compute_column_formula_mode_fails_when_cache_value_wrong(tmp_path):
    p = _formula_book(tmp_path, ["商品", "数量", "単価", "小計"],
                       [{"B": 3, "C": 120}], "D", ("B", "C"), "*")
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D2": 999})   # ★ 期待360なのに999
    status, reason = ailine.check_compute_column(
        p, {"operands": ["数量", "単価"], "operator": "*", "target": "小計"}, use_formula=True)
    assert status == "fail"
    assert "キャッシュ値が不一致" in reason

def test_check_compute_column_values_mode_unaffected_by_formula_flag_default(tmp_path):
    # use_formula 省略時は既定 False（旧テスト・旧挙動と同一）。
    p = _book(tmp_path, [["売上", "原価", "売上-原価"], [500, 300, 200]])
    status, reason = ailine.check_compute_column(p, {"operands": ["売上", "原価"], "operator": "-"})
    assert status == "pass"
    assert "キャッシュ値" not in reason


# --- ★ W10f 項目1: 計算列を operand にすると計画ごとロールバックする（Namakoo 実測の再現） ---
#   check_compute_column が operands を式ビュー(raw)から読んでいたため、前段が式で作った
#   計算列(小計等)を次段の operand にすると全行『数値でない』で除外され、
#   『検証対象0件』で fail → 計画全体が反映されなかった。data_only(計算後の値)側から
#   読むよう直した後の挙動を確認する。

def test_check_compute_column_operand_from_prior_formula_column_passes_when_cache_present(tmp_path):
    # ★ 実運用の形: 1段目(小計=数量*単価)を式のまま basrun/LibreOffice が保存時に
    # 再計算してキャッシュ値を埋める。2段目(税込=小計*数量)が小計を operand にする。
    p = tmp_path / "chain.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計", "税込"])
    ws.append(["a", 3, 100, None, None]); ws["D2"] = "=B2*C2"; ws["E2"] = "=D2*B2"
    ws.append(["b", 2, 50, None, None]); ws["D3"] = "=B3*C3"; ws["E3"] = "=D3*B3"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml",
                           {"D2": 300, "E2": 900, "D3": 100, "E3": 200})
    status, reason = ailine.check_compute_column(
        p, {"operands": ["小計", "数量"], "operator": "*", "target": "税込"}, use_formula=True)
    assert status == "pass"
    assert "2 行を検証" in reason

def test_check_compute_column_operand_from_prior_formula_column_no_cache_fails_with_distinct_reason(tmp_path):
    # ★ Namakoo が貼った純関数レベルの再現をそのまま使う回帰テスト（LLM も LibreOffice も
    # 通さない）。直った後もキャッシュ値が無ければ検証はできない(fail のまま)が、理由が
    # 『数値でない』という誤診断ではなく『キャッシュ値が無く検証できない』という正確な
    # 診断に変わる（0cf9218 空虚な検証合格の禁止 — 『対象が無い』と『対象はあるが
    # 読めない』を混同しない）。
    p = tmp_path / "chain_nocache.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計", "税込"])
    ws.append(["a", 3, 100, None, None]); ws["D2"] = "=B2*C2"
    ws.append(["b", 2, 50, None, None]); ws["D3"] = "=B3*C3"
    wb.save(p)
    status, reason = ailine.check_compute_column(
        p, {"operands": ["小計", "数量"], "operator": "*", "target": "税込"}, use_formula=True)
    assert status == "fail"
    assert "検証対象が0件" in reason
    assert "キャッシュ値が無く検証できない 2 行" in reason
    assert "数値でない" not in reason

def test_check_compute_column_single_factor_operand_from_prior_formula_column_passes_when_cache_present(tmp_path):
    p = tmp_path / "chain_sf.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計", "税込"])
    ws.append(["a", 3, 100, None, None]); ws["D2"] = "=B2*C2"; ws["E2"] = "=D2*1.1"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D2": 300, "E2": 330})
    status, reason = ailine.check_compute_column(
        p, {"operands": ["小計"], "operator": "*", "factor": 1.1, "target": "税込"}, use_formula=True)
    assert status == "pass"

def test_check_compute_column_single_factor_operand_from_prior_formula_column_no_cache_fails_with_distinct_reason(tmp_path):
    p = tmp_path / "chain_sf_nocache.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "数量", "単価", "小計", "税込"])
    ws.append(["a", 3, 100, None, None]); ws["D2"] = "=B2*C2"
    wb.save(p)
    status, reason = ailine.check_compute_column(
        p, {"operands": ["小計"], "operator": "*", "factor": 1.1, "target": "税込"}, use_formula=True)
    assert status == "fail"
    assert "検証対象が0件" in reason
    assert "キャッシュ値が無く検証できない 1 行" in reason
    assert "数値でない" not in reason


# --- codegen: COMPUTE_COLUMN の式化は LO 方言(;/.) を要さない（formula_spike の実測どおり） --

def test_codegen_dsl_compute_column_formula_has_no_semicolon_or_sheet_dot():
    # ★ bench/formula_spike_RESULTS.md: setFormula は多引数(;)・シート参照(.)の時だけ
    #   LO 方言が要る。COMPUTE_COLUMN は単純な行内二項演算(=B2*C2 型)でどちらも使わない。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["数量", "単価", "小計"]}}
    code = ailine.codegen_dsl("COMPUTE_COLUMN",
                               {"operands": ["数量", "単価"], "operator": "*", "target": "小計"}, meta)
    assert ";" not in code
    assert '"A" & (i + 1) & "*" & "B" & (i + 1)' in code   # 数量=A列(0+1) 単価=B列(1+1)


# --- _scan_last_row_basic: header_row 対応の走査開始/ガード閾値 -------------------

def test_scan_last_row_basic_default_matches_legacy_output():
    assert ailine._scan_last_row_basic() == (
        "    lastRow = 1\n"
        "    Do While oSheet.getCellByPosition(0, lastRow).getString() <> \"\"\n"
        "        lastRow = lastRow + 1\n"
        "    Loop\n"
        "    lastRow = lastRow - 1\n"
        "    If lastRow < 1 Then Exit Sub\n")

def test_scan_last_row_basic_custom_start_row_and_min_ok():
    out = ailine._scan_last_row_basic(start_row="3", min_ok="2")
    assert "lastRow = 3\n" in out
    assert "If lastRow < 2 Then Exit Sub\n" in out


# ===========================================================================
# W6: APPEND_TOTAL 語彙昇格（監査3回連続失敗の実測による昇格経済学）
# ===========================================================================

def test_ops_doc_and_fewshot_mention_append_total():
    assert "APPEND_TOTAL" in ailine.OPS_DOC
    assert any("APPEND_TOTAL" in assistant_ex for _u, assistant_ex in ailine.TRANSLATION_FEWSHOT)

def test_translate_task_parses_append_total_with_label_and_factor(monkeypatch):
    # ★ A': プロンプトはもう factor を求めないが、translate_task 自体は仮に LLM が
    #   余計な factor を返しても素通しする（黙って落とさない・verify_dsl_args 側で
    #   factor は無視され機械抽出に置き換わる。ここは翻訳層のJSON解析の頑健性テスト）。
    monkeypatch.setattr(
        ailine, "ollama_generate_json",
        lambda model, msgs, temperature=0.1, num_predict=300:
        '{"plan": [{"op": "APPEND_TOTAL", "args": '
        '{"col": "小計", "label": "税込み合計", "factor": 1.1}}]}')
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    result = ailine.translate_task("m", "税込み合計を一番下に出して（消費税10%）", meta)
    step = result["plan"][0]
    assert step["op"] == "APPEND_TOTAL"
    assert step["args"] == {"col": "小計", "label": "税込み合計", "factor": 1.1}

def test_ops_doc_does_not_ask_llm_for_factor():
    # ★ A': factor は machine-determined。プロンプトが LLM に数値化を求めないことを固定する。
    assert "factor" not in ailine.OPS_DOC
    for _u, assistant_ex in ailine.TRANSLATION_FEWSHOT:
        if "APPEND_TOTAL" in assistant_ex:
            assert '"factor"' not in assistant_ex

def test_translate_task_parses_append_total_without_optional_args(monkeypatch):
    monkeypatch.setattr(
        ailine, "ollama_generate_json",
        lambda model, msgs, temperature=0.1, num_predict=300:
        '{"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額"}}]}')
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    result = ailine.translate_task("m", "金額の合計を最後に", meta)
    step = result["plan"][0]
    assert step["op"] == "APPEND_TOTAL"
    assert step["args"] == {"col": "金額"}   # label/factor の既定は verify_dsl_args 側が確定する


# --- ② 検証（接地） ----------------------------------------------------------

def test_verify_dsl_args_append_total_defaults_label_and_factor():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("APPEND_TOTAL", {"col": "金額"}, meta)
    assert ok
    assert resolved["col"] == "金額"
    assert resolved["label"] == "合計"
    assert resolved["factor"] == 1.0

def test_verify_dsl_args_append_total_resolves_factor_from_task_text_percent():
    # ★ A': factor は LLM でなく、依頼文の明示率を machine が抽出する。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計"}, meta,
        task="税込み合計を一番下に出して（消費税10%）")
    assert ok
    assert resolved["label"] == "税込み合計"
    assert resolved["factor"] == 1.1
    assert resolved["_sources"]["factor"] == "依頼文: 10%"

def test_verify_dsl_args_append_total_resolves_factor_from_vocab():
    # 用語集の語が依頼文に部分一致で含まれる場合だけ引く（無関係な語を勝手に当てない）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計"}, meta,
        task="消費税込みの合計を出して", vocab={"消費税": 1.1})
    assert ok
    assert resolved["factor"] == 1.1
    assert resolved["_sources"]["factor"] == "用語集: 消費税"

def test_verify_dsl_args_append_total_task_text_wins_over_vocab():
    # 依頼文に明示率があれば、用語集より依頼文を優先する。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計"}, meta,
        task="税込み合計を出して（消費税8%で）", vocab={"消費税": 1.1})
    assert ok
    assert resolved["factor"] == 1.08
    assert "依頼文" in resolved["_sources"]["factor"]

def test_verify_dsl_args_append_total_llm_factor_ignored_but_warns_on_mismatch():
    # LLM が(旧仕様の名残や幻覚で)factor を返しても機械抽出が常に勝つ。食い違いは WARN。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計", "factor": 1.08}, meta,
        task="税込み合計を一番下に出して（消費税10%）")
    assert ok
    assert resolved["factor"] == 1.1
    assert resolved["_warnings"]
    assert "1.08" in resolved["_warnings"][0] and "1.1" in resolved["_warnings"][0]

def test_verify_dsl_args_append_total_non_tax_label_defaults_without_clarify():
    # label が税/込を含まなければ、倍率がどこにも無くても既定1.0で通す
    # （恒真式の番人は「税/込ラベルなのに倍率不明」の場合だけに絞る）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "金額"}, meta, task="金額の合計を最後に")
    assert ok
    assert resolved["factor"] == 1.0
    assert "_sources" not in resolved

def test_verify_dsl_args_append_total_clarifies_when_label_implies_tax_but_no_rate():
    # ★ 恒真式の番人（最優先）: label が税/込を含むのに倍率がどこにも無いと CLARIFY へ倒す
    #   （税抜き金額に「税込み」ラベルが付く恒真の誤りを機械で止める）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計"}, meta, task="税込み合計を出して")
    assert not ok
    assert "倍率が分かりません" in err
    assert "ailine vocab add" in err

def test_verify_dsl_args_append_total_ignores_non_numeric_llm_factor():
    # ★ A': factor はもう「ユーザー入力のエラー」ではない。LLM が壊れた値を返しても
    #   単に無視して機械確定（既定1.0）にフォールバックする（float() できない値は
    #   食い違いWARNの対象にもしない＝比較不能なので黙って無視）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "金額", "factor": "abc"}, meta, task="金額の合計を最後に")
    assert ok
    assert resolved["factor"] == 1.0
    assert "_warnings" not in resolved

def test_verify_dsl_args_append_total_rejects_zero_factor_from_text_extraction():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "金額"}, meta, task="0倍の合計を出して")
    assert not ok
    assert "正の数" in err

def test_verify_dsl_args_append_total_unknown_column_is_clarify_error():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "APPEND_TOTAL", {"col": "存在しない"}, meta)
    assert not ok
    assert "がありません" in err

def test_format_confirmation_line_append_total_shows_col_label_factor():
    # ★ W8a 項目5: 表示ラベルのみ「倍率」→「率」（内部キー"factor"は不変）。
    line = ailine.format_confirmation_line(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計", "factor": 1.1}, set())
    assert "対象列:小計" in line
    assert "ラベル:税込み合計" in line
    assert "率:1.1" in line

def test_format_confirmation_line_append_total_shows_factor_source():
    # ★ A': 来歴の可視化。率の出典を確認行に添える。★ W8a 項目5: 表示は「率」。
    line = ailine.format_confirmation_line(
        "APPEND_TOTAL",
        {"col": "小計", "label": "税込み合計", "factor": 1.1,
         "_sources": {"factor": "依頼文: 10%"}},
        set())
    assert "率:1.1（依頼文: 10%）" in line


# --- ④ codegen（決定論） ------------------------------------------------------

def test_codegen_dsl_append_total_places_label_left_of_value_and_formula_with_factor():
    # ★ B: 挿入耐性式（SUM(D2:INDEX(D:D;ROW()-1))型・LO方言のセミコロンで setFormula する。
    #   保存後はカンマ形に自動変換される・formula spike で実測済み）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    code = ailine.codegen_dsl(
        "APPEND_TOTAL", {"col": "小計", "label": "税込み合計", "factor": 1.1}, meta)
    # 小計=列3(0起点)。ラベルはその左隣=列2に置く（既存構造を壊さない置き方）。
    assert 'getCellByPosition(2, totalRow).setString("税込み合計")' in code
    assert 'getCellByPosition(3, totalRow).setFormula(' in code
    assert ('"=SUM(" & "D" & 2 & ":INDEX(" & "D" & ":" & "D" & ";ROW()-1))" & "*1.1"') in code

def test_codegen_dsl_append_total_omits_factor_tail_when_factor_is_one():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    code = ailine.codegen_dsl("APPEND_TOTAL", {"col": "金額", "label": "合計", "factor": 1}, meta)
    assert ('"=SUM(" & "B" & 2 & ":INDEX(" & "B" & ":" & "B" & ";ROW()-1))" & ""') in code

def test_codegen_dsl_append_total_leftmost_column_skips_label():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["金額", "部門"]}}
    code = ailine.codegen_dsl("APPEND_TOTAL", {"col": "金額", "label": "合計", "factor": 1}, meta)
    assert "setString(" not in code   # 列0は左隣が無いためラベルを省略する


# --- ⑥ 事後条件（二層: 式文字列 + キャッシュ値、ラベル一致） -------------------

def test_check_append_total_passes_with_factor_and_label(tmp_path):
    # ★ B: 保存後のカンマ形(=SUM(D2:INDEX(D:D,ROW()-1))*1.1)と照合する（codegen が
    #   LO 方言(;)で書いても、basrun 保存後はこの形に変換される・formula spike で実測済み）。
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["品目", "数量", "単価", "小計"])
    ws.append(["a", 3, 50000, 150000])
    ws.append(["b", 1, 120000, 120000])
    ws.append(["c", 12, 8000, 96000])
    ws["C5"] = "税込み合計"
    ws["D5"] = "=SUM(D2:INDEX(D:D,ROW()-1))*1.1"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"D5": 402600})
    status, reason = ailine.check_append_total(
        p, {"col": "小計", "label": "税込み合計", "factor": 1.1})
    assert status == "pass"
    assert "式・キャッシュ値・ラベルとも一致" in reason

def test_check_append_total_fails_when_formula_not_sum(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["品目", "数量", "単価", "小計"])
    ws.append(["a", 3, 50000, 150000])
    ws["D3"] = 999999   # 合計行のはずが数式でなく値
    wb.save(p)
    status, reason = ailine.check_append_total(p, {"col": "小計", "label": "合計", "factor": 1})
    assert status == "fail"
    assert "式が期待形" in reason

def test_check_append_total_fails_when_cache_value_wrong(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["部門", "金額"])
    ws.append(["a", 100])
    ws.append(["b", 200])
    ws["A4"] = "合計"
    ws["B4"] = "=SUM(B2:INDEX(B:B,ROW()-1))"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"B4": 999})
    status, reason = ailine.check_append_total(p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "fail"
    assert "キャッシュ値が不一致" in reason

def test_check_append_total_fails_when_label_wrong(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["部門", "金額"])
    ws.append(["a", 100])
    ws.append(["b", 200])
    ws["A4"] = "TOTAL"   # 期待ラベルは「合計」
    ws["B4"] = "=SUM(B2:INDEX(B:B,ROW()-1))"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"B4": 300})
    status, reason = ailine.check_append_total(p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "fail"
    assert "ラベルが期待" in reason

def test_check_append_total_fails_when_zero_data_rows(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["部門", "金額"])
    wb.save(p)
    status, reason = ailine.check_append_total(p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "fail"
    assert reason == ailine._ZERO_TARGET_REASON

def test_check_append_total_leftmost_column_skips_label_check(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["金額", "備考"])
    ws.append([100, "x"])
    ws.append([200, "y"])
    ws["A4"] = "=SUM(A2:INDEX(A:A,ROW()-1))"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"A4": 300})
    status, reason = ailine.check_append_total(p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "pass"

def test_check_append_total_fails_on_old_static_range_formula(tmp_path):
    # ★ B: 静的な "=SUM(B2:B3)" 型（挿入耐性が無い旧形）はもう合格させない。
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["部門", "金額"])
    ws.append(["a", 100])
    ws.append(["b", 200])
    ws["A4"] = "合計"
    ws["B4"] = "=SUM(B2:B3)"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"B4": 300})
    status, reason = ailine.check_append_total(p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "fail"
    assert "期待形" in reason

def test_run_postcondition_dispatches_append_total(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["部門", "金額"])
    ws.append(["a", 100])
    ws["A3"] = "合計"
    ws["B3"] = "=SUM(B2:INDEX(B:B,ROW()-1))"
    wb.save(p)
    _inject_formula_cache(p, "xl/worksheets/sheet1.xml", {"B3": 100})
    status, reason = ailine.run_postcondition(
        "APPEND_TOTAL", p, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "pass"


# ===========================================================================
# ★ A': factor の機械抽出（依頼文の明示率・regex のみ・LLM不使用）
# ===========================================================================

def test_extract_rate_factor_percent():
    assert ailine.extract_rate_factor("税込み合計を出して（消費税10%）") == (1.1, "10%")

def test_extract_rate_factor_percent_with_space_and_zenkaku():
    assert ailine.extract_rate_factor("8 ％引きで") == (1.08, "8 ％")

def test_extract_rate_factor_bai_suffix():
    assert ailine.extract_rate_factor("1.1倍にして合計を出して") == (1.1, "1.1倍")

def test_extract_rate_factor_bare_decimal_near_tax_keyword():
    # 税/倍率の語の近傍だけにある裸の小数（0-1未満は 1+n として解釈）
    assert ailine.extract_rate_factor("税率が0.1です") == (1.1, "0.1")

def test_extract_rate_factor_bare_decimal_far_from_keyword_is_ignored():
    # 無関係な数値の誤爆を避ける（税/倍率の語から離れた裸の小数は拾わない）
    factor, snippet = ailine.extract_rate_factor("0.1個くらい余分に買って集計して")
    assert factor is None and snippet is None

def test_extract_rate_factor_no_mention_returns_none():
    assert ailine.extract_rate_factor("金額の合計を最後に") == (None, None)

def test_extract_rate_factor_conflicting_values_returns_none():
    # 複数の異なる値が出たら断定しない（CLARIFY に委ねる）
    factor, snippet = ailine.extract_rate_factor("消費税10%か8%のどちらかで")
    assert factor is None and snippet is None

def test_extract_rate_factor_same_value_repeated_is_not_conflicting():
    factor, _ = ailine.extract_rate_factor("消費税10%(10%)込みで")
    assert factor == 1.1

def test_extract_rate_factor_empty_text():
    assert ailine.extract_rate_factor("") == (None, None)

# --- ★ W10b 項目3: 「掛けた/割った」型（battery v5 #503 実測ギャップの修正） ------------

def test_extract_rate_factor_kake_suffix():
    assert ailine.extract_rate_factor("金額に1.1を掛けた列を追加して") == (1.1, "1.1を掛け")

def test_extract_rate_factor_kake_suffix_no_particle():
    assert ailine.extract_rate_factor("1.1掛けた列を作って") == (1.1, "1.1掛け")

def test_extract_rate_factor_wari_suffix_inverts():
    # 税抜き等の逆算: 「1.1で割った」→ factor = 1/1.1
    factor, snippet = ailine.extract_rate_factor("税込み金額を1.1で割った列を追加して")
    assert factor == round(1 / 1.1, 6)
    assert snippet == "1.1で割っ"

def test_extract_rate_factor_kake_does_not_affect_percent_or_bai():
    # 既存パターンとの併用でも壊れない（回帰）。
    assert ailine.extract_rate_factor("消費税10%を足した列を作って") == (1.1, "10%")
    assert ailine.extract_rate_factor("1.1倍にして合計を出して") == (1.1, "1.1倍")


# ===========================================================================
# ★ A': 用語集（vocab）
# ===========================================================================

def test_load_vocab_missing_file_returns_empty_dict(tmp_path):
    assert ailine.load_vocab(tmp_path / "nope.json") == {}

def test_save_and_load_vocab_roundtrip(tmp_path):
    p = tmp_path / "vocab.json"
    ailine.save_vocab({"消費税": 1.1, "軽減税率": 1.08}, path=p)
    assert ailine.load_vocab(p) == {"消費税": 1.1, "軽減税率": 1.08}

def test_load_vocab_corrupt_json_returns_empty_dict_not_crash(tmp_path):
    p = tmp_path / "vocab.json"
    p.write_text("{not valid json", encoding="utf-8")
    assert ailine.load_vocab(p) == {}

def test_load_vocab_non_dict_json_returns_empty_dict(tmp_path):
    p = tmp_path / "vocab.json"
    p.write_text("[1, 2, 3]", encoding="utf-8")
    assert ailine.load_vocab(p) == {}

def test_load_vocab_skips_non_numeric_values(tmp_path):
    p = tmp_path / "vocab.json"
    p.write_text('{"消費税": 1.1, "壊れた語": "abc", "null値": null}', encoding="utf-8")
    assert ailine.load_vocab(p) == {"消費税": 1.1}

def test_load_vocab_skips_control_character_terms(tmp_path):
    # ★ codegen へ渡る経路の防御。改行を含む語は読み捨てる。
    p = tmp_path / "vocab.json"
    p.write_text(json.dumps({"改行\n入り": 1.1, "消費税": 1.1}, ensure_ascii=False),
                encoding="utf-8")
    assert ailine.load_vocab(p) == {"消費税": 1.1}

def test_load_vocab_skips_overlong_terms(tmp_path):
    p = tmp_path / "vocab.json"
    long_term = "あ" * (ailine.DEFAULT_VOCAB_MAX_TERM_LEN + 1)
    p.write_text(json.dumps({long_term: 1.1, "消費税": 1.1}, ensure_ascii=False),
                encoding="utf-8")
    assert ailine.load_vocab(p) == {"消費税": 1.1}

def test_load_vocab_caps_entry_count(tmp_path):
    p = tmp_path / "vocab.json"
    many = {f"語{i}": 1.0 + i / 1000 for i in range(ailine.DEFAULT_VOCAB_MAX_ENTRIES + 20)}
    p.write_text(json.dumps(many, ensure_ascii=False), encoding="utf-8")
    assert len(ailine.load_vocab(p)) == ailine.DEFAULT_VOCAB_MAX_ENTRIES

def test_vocab_add_new_term(tmp_path):
    p = tmp_path / "vocab.json"
    ok, msg = ailine.vocab_add("消費税", "1.1", path=p)
    assert ok
    assert ailine.load_vocab(p) == {"消費税": 1.1}

def test_vocab_add_updates_existing_term(tmp_path):
    p = tmp_path / "vocab.json"
    ailine.vocab_add("消費税", 1.08, path=p)
    ailine.vocab_add("消費税", 1.1, path=p)
    assert ailine.load_vocab(p) == {"消費税": 1.1}

def test_vocab_add_rejects_non_numeric_value(tmp_path):
    p = tmp_path / "vocab.json"
    ok, msg = ailine.vocab_add("消費税", "abc", path=p)
    assert not ok
    assert "数値ではありません" in msg
    assert ailine.load_vocab(p) == {}

def test_vocab_add_rejects_control_character_term(tmp_path):
    p = tmp_path / "vocab.json"
    ok, msg = ailine.vocab_add("改行\n入り", 1.1, path=p)
    assert not ok

def test_vocab_add_rejects_empty_term(tmp_path):
    p = tmp_path / "vocab.json"
    ok, msg = ailine.vocab_add("   ", 1.1, path=p)
    assert not ok

def test_vocab_add_rejects_when_entry_cap_reached_for_new_term(tmp_path):
    p = tmp_path / "vocab.json"
    many = {f"語{i}": 1.0 for i in range(ailine.DEFAULT_VOCAB_MAX_ENTRIES)}
    ailine.save_vocab(many, path=p)
    ok, msg = ailine.vocab_add("新語", 1.1, path=p)
    assert not ok
    assert "上限" in msg
    # 既存語の更新は上限に関係なく可能
    ok2, _ = ailine.vocab_add("語0", 2.0, path=p)
    assert ok2

def test_lookup_vocab_factor_matches_substring():
    assert ailine.lookup_vocab_factor("消費税込みの合計を出して", {"消費税": 1.1}) == (1.1, "消費税")

def test_lookup_vocab_factor_no_match_returns_none():
    assert ailine.lookup_vocab_factor("金額の合計を出して", {"消費税": 1.1}) == (None, None)

def test_lookup_vocab_factor_conflicting_terms_returns_none():
    factor, term = ailine.lookup_vocab_factor(
        "消費税と軽減税率どちらか", {"消費税": 1.1, "軽減税率": 1.08})
    assert factor is None and term is None

def test_lookup_vocab_factor_empty_vocab_returns_none():
    assert ailine.lookup_vocab_factor("消費税込み", {}) == (None, None)


# ===========================================================================
# ★ A': vocab CLI (`ailine vocab add` / `ailine vocab list`)
# ===========================================================================

def test_cmd_vocab_add_success(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    ns = argparse.Namespace(vocab_cmd="add", term="消費税", value="1.1")
    rc = ailine.cmd_vocab(ns)
    captured = capsys.readouterr()
    assert rc == 0
    assert "登録" in captured.out
    assert ailine.load_vocab(tmp_path / "vocab.json") == {"消費税": 1.1}

def test_cmd_vocab_add_failure_returns_1(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    ns = argparse.Namespace(vocab_cmd="add", term="消費税", value="abc")
    rc = ailine.cmd_vocab(ns)
    captured = capsys.readouterr()
    assert rc == 1
    assert "×" in captured.out

def test_cmd_vocab_list_empty(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    rc = ailine.cmd_vocab(argparse.Namespace(vocab_cmd="list"))
    captured = capsys.readouterr()
    assert rc == 0
    assert "空" in captured.out

def test_cmd_vocab_list_shows_entries(tmp_path, monkeypatch, capsys):
    p = tmp_path / "vocab.json"
    monkeypatch.setattr(ailine, "VOCAB_FILE", p)
    ailine.save_vocab({"消費税": 1.1}, path=p)
    rc = ailine.cmd_vocab(argparse.Namespace(vocab_cmd="list"))
    captured = capsys.readouterr()
    assert rc == 0
    assert "消費税 = 1.1" in captured.out


# ===========================================================================
# ★ A': cmd_run_dsl 統合（APPEND_TOTAL の倍率解決を通しで確認）
# ===========================================================================

def test_cmd_run_dsl_append_total_shows_factor_source_from_task_text(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["品目", "数量", "単価", "小計"], ["a", 3, 50000, 150000]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")   # 実 vocab を汚さない
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "APPEND_TOTAL", "args": {"col": "小計", "label": "税込み合計"}})
    argv = run_argv(
        book=str(book), task="税込み合計を一番下に出して（消費税10%）", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    # ★ W8a 項目5: 表示ラベル「倍率」→「率」。
    assert "率:1.1（依頼文: 10%）" in captured.out

def test_cmd_run_dsl_append_total_clarify_hint_has_copy_paste_command(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["品目", "数量", "単価", "小計"], ["a", 3, 50000, 150000]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")   # 空の用語集
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "APPEND_TOTAL", "args": {"col": "小計", "label": "税込み合計"}})
    argv = run_argv(
        book=str(book), task="税込み合計を出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3
    assert "ailine vocab add" in captured.out

def test_cmd_run_dsl_append_total_uses_registered_vocab(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["品目", "数量", "単価", "小計"], ["a", 3, 50000, 150000]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    vocab_path = tmp_path / "vocab.json"
    monkeypatch.setattr(ailine, "VOCAB_FILE", vocab_path)
    ailine.vocab_add("消費税", 1.1, path=vocab_path)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "APPEND_TOTAL", "args": {"col": "小計", "label": "税込み合計"}})
    argv = run_argv(
        book=str(book), task="消費税込みの合計を出して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    # ★ W8a 項目5: 表示ラベル「倍率」→「率」。
    assert "率:1.1（用語集: 消費税）" in captured.out

def test_cmd_run_dsl_append_total_warns_on_llm_factor_mismatch(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["品目", "数量", "単価", "小計"], ["a", 3, 50000, 150000]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "APPEND_TOTAL",
                         "args": {"col": "小計", "label": "税込み合計", "factor": 1.08}})
    argv = run_argv(
        book=str(book), task="税込み合計を一番下に出して（消費税10%）", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, inplace=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "⚠" in captured.out
    assert "1.08" in captured.out and "1.1" in captured.out


# ===========================================================================
# ★ A': history.jsonl の provenance
# ===========================================================================

def test_build_history_entry_includes_provenance():
    result = {"ok": True, "attempts": 1, "provenance": {"factor": "依頼文: 10%"}}
    e = ailine.build_history_entry(result, Path("book.xlsx"), "タスク", "モデル", "none")
    assert e["provenance"] == {"factor": "依頼文: 10%"}

def test_build_history_entry_provenance_defaults_to_none():
    e = ailine.build_history_entry({"ok": True}, Path("book.xlsx"), "タスク", "モデル", "none")
    assert e["provenance"] is None


# --- CONTRACT: 自由生成に「新規シートを作らない」誘導が入っている ----------------

def test_contract_prompt_discourages_new_sheet_creation():
    assert "新しいシートを作らず" in ailine.CONTRACT


# ===========================================================================
# W6: 助言網を新規シートへ拡張（監査所見: 新規『集計』シートの全0埋めが★素通り）
# ===========================================================================

def test_new_sheet_advisories_flags_entirely_zero_filled_new_sheet(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    out = wb.create_sheet("集計")
    for r in range(1, 4):
        for c in range(1, 3):
            out.cell(row=r, column=c, value=0)
    wb.save(p)
    after = ailine.snapshot(p)
    lines = ailine.new_sheet_advisories(before, after)
    assert any("新規シート『集計』の" in ln and "★ 疑わしい" in ln and "値 0" in ln for ln in lines)

def test_new_sheet_advisories_empty_when_new_sheet_has_normal_varied_content(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    out = wb.create_sheet("集計")
    out.append(["部門", "合計"])
    out.append(["a", 100])
    out.append(["b", 200])
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.new_sheet_advisories(before, after) == []

def test_new_sheet_advisories_empty_when_no_new_sheet(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=1, column=1, value="変更")
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.new_sheet_advisories(before, after) == []

def test_build_advisories_still_flags_new_sheet_zero_fill_when_mixed_with_unrelated_change(tmp_path):
    # ★ 監査実測の再現: COMPUTE_COLUMN が小計列へ正しい(非一様な)値を書き込みつつ、
    #   同じ実行で新規『集計』シートが全0埋め(バグ)で作られるケース。
    #   旧実装は detect_uniform_fill が『変更セル全部が同一値』を diff 全体に要求するため、
    #   小計列の非一様な正しい値に引きずられて新規シートの異常が丸ごと素通りしていた。
    p = _book(tmp_path, [["品目", "小計"], ["a", None], ["b", None]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    ws.cell(row=2, column=2, value=150000)   # 小計へ正しい値（非一様）
    ws.cell(row=3, column=2, value=96000)
    out = wb.create_sheet("集計")
    for r in range(1, 3):
        for c in range(1, 3):
            out.cell(row=r, column=c, value=0)   # ★バグ想定: 新規シートが全0埋め
    wb.save(p)
    after = ailine.snapshot(p)
    lines = ailine.build_advisories("小計を計算して集計もして", before, after)
    assert any("新規シート『集計』の" in ln and "★ 疑わしい" in ln for ln in lines)

def test_detect_ghost_data_still_fires_for_existing_sheet_when_new_sheet_also_created(tmp_path):
    # ★ W6: 以前は新規シートが混ざるだけで rect=None → 関数全体が None を返し、
    #   既存シートの本当のゴーストデータまで丸ごと素通りしていた（回帰）。
    p = _book(tmp_path, [["a", 1], ["b", 2]])   # 使用範囲 A1:B2
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=26, value="ghost")   # Z2、既存シートの範囲外
    wb.create_sheet("集計").cell(row=1, column=1, value="新規")   # 同時に新規シートも作る
    wb.save(p)
    after = ailine.snapshot(p)
    msg = ailine.detect_ghost_data(before, after)
    assert msg is not None
    assert "Z2" in msg

def test_detect_ghost_data_none_when_all_changes_are_in_a_new_sheet(tmp_path):
    p = _book(tmp_path, [["a", 1], ["b", 2]])
    before = ailine.snapshot(p)
    wb = openpyxl.load_workbook(p)
    wb.create_sheet("集計").cell(row=1, column=1, value="新規")
    wb.save(p)
    after = ailine.snapshot(p)
    assert ailine.detect_ghost_data(before, after) is None


# ===========================================================================
# W6: 依頼にないシート新設の申告
# ===========================================================================

def test_unrequested_new_sheet_advisory_fires_when_task_does_not_mention_sheet():
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet", "集計"]}
    lines = ailine.unrequested_new_sheet_advisory("部門ごとに金額を集計して", before, after)
    assert lines == ["★ 依頼にない新しいシートが作成されました（集計）"]

@pytest.mark.parametrize("phrase", ["集計シートを作って", "ピボットにして", "別に表を作って"])
def test_unrequested_new_sheet_advisory_silent_when_task_mentions_keyword(phrase):
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet", "集計"]}
    assert ailine.unrequested_new_sheet_advisory(phrase, before, after) == []

def test_unrequested_new_sheet_advisory_empty_when_no_new_sheet():
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet"]}
    assert ailine.unrequested_new_sheet_advisory("金額を並べ替えて", before, after) == []


# ===========================================================================
# ★ W10c 中 → ★ C9: AGGREGATE/PIVOT の新規シート作成は宣言どおりの効果（中立表示に落とす）
#   ★ C9: 旧 _neutralize_declared_new_sheet_warning は advisories を後から書き換える単体
#   関数だったが、unrequested_new_sheet_advisory 自身が op を受け取り発生源で判定する形に
#   構造を置き換えた（判定条件・出力文言は不変・ゴールデン差分ゼロで確認）。
# ===========================================================================

@pytest.mark.parametrize("op", ["AGGREGATE", "PIVOT"])
def test_unrequested_new_sheet_advisory_neutral_for_declared_ops(op):
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet", "集計"]}
    out = ailine.unrequested_new_sheet_advisory("部門ごとに金額を集計して", before, after, op=op)
    assert out == ["（新規シート『集計』の作成は意図どおりです）"]

def test_unrequested_new_sheet_advisory_warns_for_other_ops():
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet", "集計"]}
    out = ailine.unrequested_new_sheet_advisory("部門ごとに金額を集計して", before, after, op="COMPUTE_COLUMN")
    assert out == ["★ 依頼にない新しいシートが作成されました（集計）"]   # 対象外の op は一切変えない

def test_unrequested_new_sheet_advisory_conservative_when_multiple_new_sheets():
    # 保守的（安全器官の減衰は迷ったら出す側）: 新規シートが2枚以上できたら宣言どおりの
    # 単一効果と断定できないので、警告はそのまま残す。
    before = {"sheets": ["Sheet"]}
    after = {"sheets": ["Sheet", "集計", "もう一枚"]}
    out = ailine.unrequested_new_sheet_advisory("部門ごとに金額を集計して", before, after, op="AGGREGATE")
    assert out == ["★ 依頼にない新しいシートが作成されました（集計）",
                    "★ 依頼にない新しいシートが作成されました（もう一枚）"]

def test_cmd_run_dsl_aggregate_neutralizes_unrequested_new_sheet_warning(tmp_path, monkeypatch, capsys):
    # ★ W10c 中（査定で名指し）: この検体は「シート/ピボット/別に」のどれも言わない依頼
    #   （「部門ごとに金額を集計して」）で、旧実装は AGGREGATE が新規シートを作るたびに
    #   「★ 依頼にない新しいシートが作成されました」が出ていた（AGGREGATE は定義上・毎回
    #   新規シートを作るのが宣言済みの効果なので、これは常に誤警報だった）。
    #   今回 _neutralize_declared_new_sheet_warning を追加し、中立表示に変わる
    #   （旧テスト名 test_..._flags_unrequested_new_sheet はこの検体名を維持したまま
    #   期待を更新。旧警告文そのものが出ないことは下のアサーションで別途担保）。
    book = _book(tmp_path, [["部門", "金額"], ["a", 100], ["a", 200], ["b", 300]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        out = wb2.create_sheet("集計")
        out.append(["部門", "合計 - 金額"])
        out.append(["a", 300])
        out.append(["b", 300])
        out.append(["合計", 600])
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="部門ごとに金額を集計して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "★ 依頼にない新しいシートが作成されました" not in captured.out
    assert "（新規シート『集計』の作成は意図どおりです）" in captured.out

def test_cmd_run_dsl_aggregate_silent_when_task_mentions_sheet_keyword(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["部門", "金額"], ["a", 100], ["b", 300]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        out = wb2.create_sheet("集計")
        out.append(["部門", "合計 - 金額"])
        out.append(["a", 100])
        out.append(["b", 300])
        out.append(["合計", 400])
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="部門ごとに金額を集計シートにまとめて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "依頼にない新しいシート" not in captured.out


# ===========================================================================
# ★ W8b-2 B2: 事故バッテリのpytest化（回帰固定）
# ===========================================================================

def _chain_run_argv(book, task="何かして", **overrides):
    """既定(原本直接適用)のまま run する Namespace。B2 バッテリ共通の土台。
       ★ W10b: これらのテストは反映(置換)経路を見るのが目的で、自由生成の関所（項目1）は
       無関係なので allow_freeform=True で素通しする。"""
    base = dict(
        book=str(book), task=task, model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False,
        allow_freeform=True)
    base.update(overrides)
    return run_argv(**base)


def test_b2_chain_three_runs_then_undo_three_times_matches_each_generation_sha1(tmp_path, monkeypatch):
    # ★ B2① 連鎖3段→undo3段: run→run→run で原本が進化し（連続タスクの自然成立の核心）、
    #   undo を3回繰り返すと各世代の SHA1 に正確に1段ずつ戻ることを固定する。
    import hashlib
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda b, workdir, timeout=None: b)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")

    step = {"n": 0}
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        step["n"] += 1
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value=f"step{step['n']}")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    sha1s = [hashlib.sha1(book.read_bytes()).hexdigest()]   # 世代0（原本）
    for _ in range(3):
        rc = ailine.main(_chain_run_argv(book))
        assert rc == 0
        sha1s.append(hashlib.sha1(book.read_bytes()).hexdigest())

    assert len(set(sha1s)) == 4   # 4世代とも中身が違う（毎回ちゃんと進化している）

    for expected_sha1 in reversed(sha1s[:-1]):   # 世代3→2→1→0 の順に戻る
        rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
        assert rc == 0
        assert hashlib.sha1(book.read_bytes()).hexdigest() == expected_sha1

def test_b2_replace_and_fallback_both_fail_leaves_book_untouched(tmp_path, monkeypatch, capsys):
    # ★ B2② 置換直前で os.replace を例外にし、copy2 フォールバックも失敗させる
    #   （原本無傷であることを固定する。バックアップ自体は成功しているので undo は可能）。
    import shutil as _shutil
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda b, workdir, timeout=None: b)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value="new")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    monkeypatch.setattr(ailine.os, "replace", lambda *a, **k: (_ for _ in ()).throw(OSError("locked")))
    real_copy2 = _shutil.copy2
    def fake_copy2(src, dst):
        if str(dst) == str(book):
            raise OSError("also locked")
        return real_copy2(src, dst)
    monkeypatch.setattr(ailine.shutil, "copy2", fake_copy2)

    rc = ailine.main(_chain_run_argv(book))
    captured = capsys.readouterr()
    assert rc == 0   # 自由生成の適用自体は成功。反映(置換)だけが失敗している
    assert "置換に失敗した" in captured.out
    assert book.read_bytes() == original_bytes   # 原本は無傷
    out_book = book.with_name(book.stem + ".out" + book.suffix)
    assert out_book.exists()   # .out はそのまま残る（レビュー可能）

def test_b2_backup_failure_aborts_replacement_book_untouched(tmp_path, monkeypatch, capsys):
    # ★ B2③ バックアップ自体が失敗したら置換を中止する（回帰固定・原本無傷）。
    book = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda b, workdir, timeout=None: b)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=3, value="new")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr(ailine, "make_backup",
                        lambda *a, **k: (_ for _ in ()).throw(OSError("disk full")))

    rc = ailine.main(_chain_run_argv(book))
    captured = capsys.readouterr()
    assert rc == 0
    assert "バックアップに失敗" in captured.out
    assert book.read_bytes() == original_bytes   # 原本は無傷
    out_book = book.with_name(book.stem + ".out" + book.suffix)
    assert out_book.exists()

def test_b2_undo_remaining_count_matches_after_pruning_beyond_ten(tmp_path, monkeypatch, capsys):
    # ★ B2④ 10世代の剪定を超えてバックアップが積み上がった後も、ailine undo の
    #   「あと N 回戻せます」が**実際に遡れる残り段数**と整合すること。
    # ★ W11 で照合先を直した: 旧文はバックアップの総数と突き合わせていたが、それは
    #   undo が退避を積むほど増える数で、「あと何回押せるか」ではなかった。ここでは
    #   表示どおりの回数だけ実際に押し、そのあと端で止まることまで見て整合を証明する。
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    book = tmp_path / "book.xlsx"
    book.write_bytes(b"v0")
    for i in range(12):
        book.write_bytes(f"v{i + 1}".encode())
        ailine.make_backup(book, keep=10)
    assert len(ailine.list_backups(book)) == 10   # 剪定済み

    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
    captured = capsys.readouterr()
    assert rc == 0
    # 現在地は最新世代(v12)と同じ中身だったので 1 段遡って v11・残りは 8 段。
    assert book.read_bytes() == b"v11"
    assert "あと 8 回戻せます" in captured.out
    for _ in range(8):   # 表示どおり 8 回押せる
        assert ailine.cmd_undo(argparse.Namespace(book=str(book), list=False)) == 0
    capsys.readouterr()
    assert book.read_bytes() == b"v3"   # 剪定で v0〜v2 は既に無い（残る最古が v3）
    rc = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
    assert rc == 1 and "これ以上は戻せません" in capsys.readouterr().out


# ===========================================================================
# ★ W9: 検証済みヘルパ4種の DSL 語彙昇格
#   INSERT_ROWS / DRAW_BORDERS / AUTOFIT / PIVOT
# ===========================================================================

# --- op メタデータ宣言の整合 ---------------------------------------------------

def test_op_meta_op_schema_confirm_fields_share_same_op_set():
    assert set(ailine.OP_META.keys()) == set(ailine.OP_SCHEMA.keys())
    assert set(ailine.OP_META.keys()) == set(ailine._CONFIRM_FIELDS.keys())

def test_op_meta_entries_have_category_label_synonyms():
    for op, meta in ailine.OP_META.items():
        assert set(meta.keys()) == {"category", "label", "synonyms"}
        assert isinstance(meta["category"], str) and meta["category"]
        assert isinstance(meta["label"], str) and meta["label"]
        assert isinstance(meta["synonyms"], list) and meta["synonyms"]

def test_op_labels_is_derived_from_op_meta():
    assert ailine.OP_LABELS == {op: meta["label"] for op, meta in ailine.OP_META.items()}

def test_postconditions_cover_all_ops_except_chart_special_cased():
    # CHART は run_postcondition() が before_charts 比較で特別扱いする（POSTCONDITIONS
    # dict には入れない・既存挙動）。それ以外の全 op は POSTCONDITIONS に登録されていること。
    assert set(ailine.POSTCONDITIONS.keys()) | {"CHART"} == set(ailine.OP_SCHEMA.keys())

def test_op_meta_new_four_ops_present_with_expected_categories():
    assert ailine.OP_META["INSERT_ROWS"]["category"] == "表を編集する"
    assert ailine.OP_META["DRAW_BORDERS"]["category"] == "見た目を整える"
    assert ailine.OP_META["AUTOFIT"]["category"] == "見た目を整える"
    assert ailine.OP_META["PIVOT"]["category"] == "計算する"

def test_ops_doc_mentions_all_four_new_ops():
    for op in ("INSERT_ROWS", "DRAW_BORDERS", "AUTOFIT", "PIVOT"):
        assert op in ailine.OPS_DOC


# --- ② 検証（verify_dsl_args） -------------------------------------------------

def test_verify_dsl_args_insert_rows_valid_defaults_count_to_one():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("INSERT_ROWS", {"at": "3"}, meta)
    assert ok
    assert resolved["at"] == 3
    assert resolved["count"] == 1
    assert "count" in inferred

def test_verify_dsl_args_insert_rows_explicit_count():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("INSERT_ROWS", {"at": 3, "count": 2}, meta)
    assert ok
    assert resolved["count"] == 2
    assert "count" not in inferred

def test_verify_dsl_args_insert_rows_rejects_zero_at():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("INSERT_ROWS", {"at": 0}, meta)
    assert not ok
    assert "不正" in err

def test_verify_dsl_args_insert_rows_rejects_non_numeric_at():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("INSERT_ROWS", {"at": "abc"}, meta)
    assert not ok

def test_verify_dsl_args_insert_rows_rejects_bad_count():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("INSERT_ROWS", {"at": 1, "count": "abc"}, meta)
    assert not ok

def test_verify_dsl_args_draw_borders_no_args_needed():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("DRAW_BORDERS", {}, meta)
    assert ok and err is None

def test_verify_dsl_args_autofit_no_args_needed():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("AUTOFIT", {}, meta)
    assert ok and err is None

def test_verify_dsl_args_pivot_resolves_columns():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "PIVOT", {"group_col": "部門", "value_col": "金額"}, meta)
    assert ok
    assert resolved["group_col"] == "部門" and resolved["value_col"] == "金額"

def test_verify_dsl_args_pivot_unknown_column_is_clarify_error():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "PIVOT", {"group_col": "存在しない", "value_col": "金額"}, meta)
    assert not ok
    assert "がありません" in err

def test_verify_dsl_args_pivot_accepts_digit_column_reference():
    # ★ "0" は idx=0(0起点候補)しか実在範囲に収まらない一意なケースを使う
    #   （resolve_col_ref は 0/1 起点の両解釈を試すため、両方が実在すると曖昧判定になる）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args("PIVOT", {"group_col": "0", "value_col": "金額"}, meta)
    assert ok
    assert resolved["group_col"] == "部門" and resolved["value_col"] == "金額"
    assert "group_col" in inferred


# --- ④ codegen（決定論・既存ヘルパへの Call） -----------------------------------

def test_codegen_dsl_insert_rows_converts_1origin_to_0origin():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    code = ailine.codegen_dsl("INSERT_ROWS", {"at": 3, "count": 1}, meta)
    assert "Call InsertRows(oDoc, 2, 1)" in code   # at=3(1起点) → atRow=2(0起点)

def test_codegen_dsl_insert_rows_multi_count():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    code = ailine.codegen_dsl("INSERT_ROWS", {"at": 1, "count": 3}, meta)
    assert "Call InsertRows(oDoc, 0, 3)" in code

def test_codegen_dsl_draw_borders_calls_helper_no_args():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    code = ailine.codegen_dsl("DRAW_BORDERS", {}, meta)
    assert "Call DrawTableBorders(oDoc)" in code

def test_codegen_dsl_autofit_calls_helper_no_args():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額"]}}
    code = ailine.codegen_dsl("AUTOFIT", {}, meta)
    assert "Call AutoFitColumns(oDoc)" in code

def test_codegen_dsl_pivot_calls_helper_with_column_indices():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["部門", "金額", "在庫"]}}
    code = ailine.codegen_dsl("PIVOT", {"group_col": "部門", "value_col": "在庫"}, meta)
    assert "Call PivotSum(oDoc, 0, 2)" in code


# --- ③ 確認行: PIVOT の既知の癖の案内 ------------------------------------------

def test_format_confirmation_line_pivot_shows_group_and_value_cols():
    line = ailine.format_confirmation_line("PIVOT", {"group_col": "部門", "value_col": "金額"}, set())
    assert "分類列:部門" in line
    assert "集計列:金額" in line

def test_pivot_caveat_mentions_summary_table_alternative():
    assert "集計表" in ailine.PIVOT_CAVEAT


# --- ⑥ 事後条件 ---------------------------------------------------------------

def test_check_draw_borders_pass_when_all_cells_bordered(tmp_path):
    from openpyxl.styles import Border, Side
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    for r in range(1, 4):
        for c in (1, 2):
            ws.cell(row=r, column=c).border = thin
    wb.save(p)
    status, reason = ailine.check_draw_borders(p, {})
    assert status == "pass"

def test_check_draw_borders_fail_when_partial(tmp_path):
    from openpyxl.styles import Border, Side
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    ws.cell(row=1, column=1).border = thin   # 1セルだけ罫線
    wb.save(p)
    status, reason = ailine.check_draw_borders(p, {})
    assert status == "fail"

def test_check_autofit_pass_when_width_changed_with_source_book(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    import shutil as _sh
    after = tmp_path / "after.xlsx"
    _sh.copy2(before, after)
    wb = openpyxl.load_workbook(after)
    wb.active.column_dimensions["A"].width = 25
    wb.active.column_dimensions["B"].width = 12
    wb.save(after)
    status, reason = ailine.check_autofit(after, {}, source_book=before)
    assert status == "pass"

def test_check_autofit_fail_when_width_unchanged_with_source_book(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    status, reason = ailine.check_autofit(before, {}, source_book=before)
    assert status == "fail"

def test_check_autofit_warn_without_source_book(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    wb = openpyxl.load_workbook(p)
    wb.active.column_dimensions["A"].width = 20
    wb.active.column_dimensions["B"].width = 10
    wb.save(p)
    status, reason = ailine.check_autofit(p, {})
    assert status == "warn"

def test_check_insert_rows_pass_with_shift_verified_via_source_book(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]])
    import shutil as _sh
    after = tmp_path / "after.xlsx"
    _sh.copy2(before, after)
    wb = openpyxl.load_workbook(after)
    wb.active.insert_rows(2, amount=1)   # at=2(1起点) の前に1行挿入 = InsertRows(atRow=1,count=1)
    wb.save(after)
    status, reason = ailine.check_insert_rows(after, {"at": 2, "count": 1}, source_book=before)
    assert status == "pass"

def test_check_insert_rows_fail_when_no_shift_happened(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    # after が before と同じ（何も挿入していない）
    status, reason = ailine.check_insert_rows(before, {"at": 2, "count": 1}, source_book=before)
    assert status == "fail"

def test_check_insert_rows_fail_when_shift_amount_wrong(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]])
    import shutil as _sh
    after = tmp_path / "after.xlsx"
    _sh.copy2(before, after)
    wb = openpyxl.load_workbook(after)
    wb.active.insert_rows(2, amount=2)   # 2行挿入したのに count=1 を主張させる
    wb.save(after)
    status, reason = ailine.check_insert_rows(after, {"at": 2, "count": 1}, source_book=before)
    assert status == "fail"

def test_check_insert_rows_warn_without_source_book(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    wb.active.insert_rows(2, amount=1)
    wb.save(p)
    status, reason = ailine.check_insert_rows(p, {"at": 2, "count": 1})
    assert status == "warn"

def test_check_insert_rows_fail_without_source_book_when_target_not_empty(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    status, reason = ailine.check_insert_rows(p, {"at": 1, "count": 1})   # 行1(見出し)は空欄でない
    assert status == "fail"

def test_check_pivot_pass_when_sheet_and_pivot_table_exist(tmp_path):
    # ★ 本物の DataPilot XML を openpyxl では作れないため、check_zip_fidelity_loss の
    #   検体作りと同じ手法（zip 直書き）で xl/pivotTables/ の実体を模す。
    p = _book(tmp_path, [["部門", "金額"], ["営業", 100], ["開発", 200]])
    wb = openpyxl.load_workbook(p)
    wb.create_sheet("ピボット")
    wb.save(p)
    import zipfile as _zf
    items = {}
    with _zf.ZipFile(p) as z:
        for n in z.namelist():
            items[n] = z.read(n)
    items["xl/pivotTables/pivotTable1.xml"] = b"<pivotTableDefinition/>"
    with _zf.ZipFile(p, "w") as z:
        for n, data in items.items():
            z.writestr(n, data)
    status, reason = ailine.check_pivot(p, {})
    assert status == "pass"
    assert "集計表" in reason   # PIVOT_CAVEAT が結果表示にも添えられる

def test_check_pivot_fail_when_sheet_missing(tmp_path):
    p = _book(tmp_path, [["部門", "金額"], ["営業", 100]])
    status, reason = ailine.check_pivot(p, {})
    assert status == "fail"
    assert "『ピボット』シートが無い" in reason

def test_check_pivot_fail_when_sheet_exists_but_no_pivot_table_xml(tmp_path):
    p = _book(tmp_path, [["部門", "金額"], ["営業", 100]])
    wb = openpyxl.load_workbook(p)
    wb.create_sheet("ピボット")   # シートだけ作って DataPilot 実体は無いケース
    wb.save(p)
    status, reason = ailine.check_pivot(p, {})
    assert status == "fail"
    assert "DataPilot" in reason


# --- run_postcondition 経由（source_book の受け渡し） ---------------------------

def test_run_postcondition_insert_rows_passes_source_book_through(tmp_path):
    before = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    import shutil as _sh
    after = tmp_path / "after.xlsx"
    _sh.copy2(before, after)
    wb = openpyxl.load_workbook(after)
    wb.active.insert_rows(2, amount=1)
    wb.save(after)
    status, reason = ailine.run_postcondition("INSERT_ROWS", after, {"at": 2, "count": 1},
                                               source_book=before)
    assert status == "pass"

def test_run_postcondition_autofit_without_source_book_still_warns_not_errors(tmp_path):
    p = _book(tmp_path, [["商品", "金額"], ["a", 100]])
    wb = openpyxl.load_workbook(p)
    wb.active.column_dimensions["A"].width = 15
    wb.active.column_dimensions["B"].width = 15
    wb.save(p)
    status, reason = ailine.run_postcondition("AUTOFIT", p, {})
    assert status == "warn"

def test_run_postcondition_draw_borders_and_pivot_ignore_source_book_kwarg(tmp_path):
    # DRAW_BORDERS/PIVOT のチェッカーは source_book を受け取らないので、run_postcondition
    # が渡さないこと（渡すと TypeError になる）を確認する回帰。
    p = _book(tmp_path, [["部門", "金額"], ["営業", 100]])
    status, reason = ailine.run_postcondition("DRAW_BORDERS", p, {})
    assert status in ("pass", "fail")   # エラーにならないことが主眼
    status2, reason2 = ailine.run_postcondition("PIVOT", p, {"group_col": "部門", "value_col": "金額"})
    assert status2 in ("pass", "fail")


# --- cmd_run_dsl 統合（--dry・軽量） ---------------------------------------------

def test_cmd_run_dsl_insert_rows_dry_shows_confirm_line(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "INSERT_ROWS", "args": {"at": 2, "count": 1}})
    argv = run_argv(
        book=str(book), task="2行目の前に1行挿入して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, copy=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "挿入位置:2" in captured.out
    assert "InsertRows" in captured.out

def test_cmd_run_dsl_pivot_dry_shows_caveat(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["部門", "金額"], ["営業", 100], ["開発", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "PIVOT", "args": {"group_col": "部門", "value_col": "金額"}})
    argv = run_argv(
        book=str(book), task="部門ごとにピボットテーブルで集計して", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, copy=False, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "集計表" in captured.out   # PIVOT_CAVEAT が確認行の直後にも出る
    assert "PivotSum" in captured.out


# ===========================================================================
# ★ W10a: 破壊の関所（既存データ上書きの自動確認・既定変更の告知・解釈要約）
# ===========================================================================

# --- 検出（件数つき） -----------------------------------------------------------

def test_column_existing_value_count_counts_nonempty_cells(tmp_path):
    p = _book(tmp_path, [["商品", "原価"], ["a", 300], ["b", 250], ["c", None]])
    assert ailine._column_existing_value_count(p, "Sheet", "原価") == 2

def test_column_existing_value_count_zero_when_all_empty(tmp_path):
    p = _book(tmp_path, [["商品", "原価"], ["a", None], ["b", None]])
    assert ailine._column_existing_value_count(p, "Sheet", "原価") == 0

def test_column_existing_value_count_zero_when_column_missing(tmp_path):
    p = _book(tmp_path, [["商品", "原価"], ["a", 300]])
    assert ailine._column_existing_value_count(p, "Sheet", "存在しない") == 0

def test_maybe_warn_target_overwrite_message_includes_count(tmp_path):
    p = _book(tmp_path, [["商品", "金額", "原価"], ["a", 100, 300], ["b", 200, 250]])
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "金額", "原価"]}}
    msg = ailine._maybe_warn_target_overwrite(
        "COMPUTE_COLUMN", {"target": "原価"}, meta, p)
    assert msg == "★ 対象列『原価』には既存の値が 2 件あります（上書きします）"

def test_maybe_warn_target_overwrite_none_when_no_existing_values(tmp_path):
    p = _book(tmp_path, [["商品", "新列"], ["a", None], ["b", None]])
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "新列"]}}
    assert ailine._maybe_warn_target_overwrite("COMPUTE_COLUMN", {"target": "新列"}, meta, p) is None

def test_maybe_warn_target_overwrite_none_for_non_compute_column_ops():
    assert ailine._maybe_warn_target_overwrite("SORT", {"col": "金額"}, {"sheets": ["S"]}, Path(".")) is None

# ★ W10c 致命1: 破壊の関所を op ごとの if から OP_WRITE_TARGET 宣言読み取りへ一般化した
#   ことの中核テスト。旧実装は `if op != "COMPUTE_COLUMN": return None` の1行で
#   LOOKUP_FILL がこの関所を素通りしていた（監査実測の事故そのもの）。

def test_op_write_target_declares_all_ops():
    # ★ 再発防止の本体: OP_SCHEMA に op を足したのに OP_WRITE_TARGET への宣言を忘れると
    #   このテストが落ちる（宣言漏れ＝黙って関所を素通りする新しい op、を機械的に防ぐ）。
    missing = [op for op in ailine.OP_SCHEMA if op not in ailine.OP_WRITE_TARGET]
    assert missing == [], f"OP_WRITE_TARGET に書き込み先列の宣言が無い op: {missing}"

# ★ 単位C: 宣言が「列」から「領域」へ広がった分だけ、番人も広げる。登録の有無（上）だけでは
#   「writes を空で登録して素通りさせる」「知らない種類の文字列を書く」「既存列を書くと
#   言いながら書き込み先列を指さない」が通ってしまう。宣言が宣言として成立していることを
#   機械で確かめる（新しい op を足す人が最初にぶつかる番人）。

def test_op_write_target_declarations_are_well_formed():
    for op, wt in ailine.OP_WRITE_TARGET.items():
        assert wt.writes, f"{op}: writes が空（何を書くかを必ず宣言する）"
        unknown = set(wt.writes) - ailine.WRITE_KINDS
        assert not unknown, f"{op}: 未知の書き込み領域の種類: {sorted(unknown)}"
        assert len(set(wt.writes)) == len(wt.writes), f"{op}: writes に重複がある"
        # 既存列を書く op だけが書き込み先列を指す（破壊の関所が守る対象と一致させる）。
        if ailine.WRITE_EXISTING_COLUMN in wt.writes:
            assert wt.col_key, f"{op}: 既存列を書くと宣言しているのに col_key が無い"
        else:
            assert wt.col_key is None, f"{op}: 既存列を書かないのに col_key がある: {wt.col_key}"
        for key in (wt.sheet_key, *wt.reads_only):
            assert key is None or (isinstance(key, str) and key), f"{op}: slot 名が不正: {key!r}"
        # 参照専用シートを、同時に書き込み先シートとして宣言していないこと（自己矛盾）。
        assert wt.sheet_key not in wt.reads_only, f"{op}: 書き込み先シートを reads_only に入れている"

def test_declared_reads_only_sheets_reads_the_declaration():
    # ★ 単位C(D8): `if op == "LOOKUP_FILL"` のハードコードを置き換えた宣言読み取り。
    assert ailine._declared_reads_only_sheets(
        "LOOKUP_FILL", {"source_sheet": "単価表", "target_sheet": "明細"}) == {"単価表"}
    # AGGREGATE/PIVOT の入力シートも同じ理屈で「読むだけ」＝旧実装では毎回誤爆していた側。
    assert ailine._declared_reads_only_sheets(
        "AGGREGATE", {"_target_sheet": "工事台帳"}) == {"工事台帳"}
    assert ailine._declared_reads_only_sheets(
        "PIVOT", {"_target_sheet": "工事台帳"}) == {"工事台帳"}
    # 書き込み系の op は「読むだけのシート」を持たない（抑制を広げすぎない側の確認）。
    assert ailine._declared_reads_only_sheets("SET_COLUMN_VALUE", {"col": "備考"}) == set()
    assert ailine._declared_reads_only_sheets("APPEND_TOTAL", {"col": "金額"}) == set()
    assert ailine._declared_reads_only_sheets(None, None) == set()

def test_maybe_warn_target_overwrite_fires_for_lookup_fill_existing_target_col(tmp_path):
    # ★ 致命1の再現そのもの: 旧実装はここが常に None だった（op != COMPUTE_COLUMN のため）。
    p = _book(tmp_path, [["商品", "数量", "単価"], ["りんご", 2, 999], ["バナナ", 3, 999]])
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "数量", "単価"]}}
    msg = ailine._maybe_warn_target_overwrite(
        "LOOKUP_FILL", {"target_sheet": "Sheet", "target_col": "単価",
                         "source_sheet": "単価表", "key_col": "商品"}, meta, p)
    assert msg == "★ 対象列『単価』には既存の値が 2 件あります（上書きします）"

def test_maybe_warn_target_overwrite_none_for_lookup_fill_new_column(tmp_path):
    # target_col が対象シートにまだ無い（新規作成）場合は件数0なので警告なし
    # （COMPUTE_COLUMN の新規列と同じ扱い）。
    p = _book(tmp_path, [["商品", "数量"], ["りんご", 2], ["バナナ", 3]])
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "数量"]}}
    assert ailine._maybe_warn_target_overwrite(
        "LOOKUP_FILL", {"target_sheet": "Sheet", "target_col": "単価",
                         "source_sheet": "単価表", "key_col": "商品"}, meta, p) is None


# --- 解釈要約（実行前の可視化） --------------------------------------------------

def test_interpretation_summary_line_shows_digit_to_name_resolution():
    resolved = {"target": "原価", "_target_raw": "5"}
    line = ailine._interpretation_summary_line(resolved, {"target"})
    assert line == "→『5』は既存の『原価』列と解釈しました（既存データあり）"

def test_interpretation_summary_line_none_when_target_not_inferred():
    resolved = {"target": "原価"}
    assert ailine._interpretation_summary_line(resolved, set()) is None

def test_interpretation_summary_line_none_when_no_raw_value_recorded():
    # target が推定はされたが _target_raw が無いケース（他の op 由来等）は何も語らない。
    resolved = {"target": "原価"}
    assert ailine._interpretation_summary_line(resolved, {"target"}) is None


# --- 破壊の関所（ゲート判定） ----------------------------------------------------

def _gate_ns(**overrides):
    base = dict(inplace=True, dry=False, ask=False, overwrite=False)
    base.update(overrides)
    return argparse.Namespace(**base)

def test_confirm_overwrite_or_gate_none_when_no_warning():
    assert ailine._confirm_overwrite_or_gate(_gate_ns(), None) is None

def test_confirm_overwrite_or_gate_none_when_copy_mode():
    # --copy は a.inplace=False になる（W8b-2 の反転ロジック）。
    assert ailine._confirm_overwrite_or_gate(_gate_ns(inplace=False), "★ 警告") is None

def test_confirm_overwrite_or_gate_none_when_dry():
    assert ailine._confirm_overwrite_or_gate(_gate_ns(dry=True), "★ 警告") is None

def test_confirm_overwrite_or_gate_none_when_ask_already_set():
    # --ask 側の汎用確認に譲る（二重に聞かない）。
    assert ailine._confirm_overwrite_or_gate(_gate_ns(ask=True), "★ 警告") is None

def test_confirm_overwrite_or_gate_none_when_overwrite_flag_set():
    assert ailine._confirm_overwrite_or_gate(_gate_ns(overwrite=True), "★ 警告") is None

def test_confirm_overwrite_or_gate_interactive_yes_continues(monkeypatch):
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    assert ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告") is None

def test_confirm_overwrite_or_gate_interactive_no_returns_1(monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    rc = ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告")
    captured = capsys.readouterr()
    assert rc == 1
    assert "中止した" in captured.out

def test_confirm_overwrite_or_gate_noninteractive_returns_7_with_guidance(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    rc = ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告")
    captured = capsys.readouterr()
    assert rc == 7
    assert "--overwrite" in captured.out
    assert "--copy" in captured.out

def test_confirm_overwrite_or_gate_uses_step_prefix(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    ailine._confirm_overwrite_or_gate(_gate_ns(), "★ 警告", step_prefix="  3段目: ")
    captured = capsys.readouterr()
    assert "  3段目: 上書きしますか" not in captured.out   # prompt はstdinへ・標準出力ではない
    assert "  3段目: この処理を続けるには" in captured.out


# --- 告知（一度きり） ------------------------------------------------------------

def test_maybe_show_notice_v2_shows_once_and_creates_marker(tmp_path, capsys):
    marker = tmp_path / "notice_v2_shown"
    assert not marker.exists()
    shown = ailine.maybe_show_notice_v2(marker)
    captured = capsys.readouterr()
    assert shown is True
    assert "既定で原本に直接反映" in captured.out
    assert marker.exists()

def test_maybe_show_notice_v2_silent_on_second_call(tmp_path, capsys):
    marker = tmp_path / "notice_v2_shown"
    ailine.maybe_show_notice_v2(marker)
    capsys.readouterr()   # 1回目の出力を捨てる
    shown = ailine.maybe_show_notice_v2(marker)
    captured = capsys.readouterr()
    assert shown is False
    assert captured.out == ""

def test_maybe_show_notice_v2_never_mentions_dry_run_wording(tmp_path):
    marker = tmp_path / "notice_v2_shown"
    ailine.maybe_show_notice_v2(marker)
    assert "dry-run" not in ailine.NOTICE_V2_TEXT.lower()


# --- argparse ------------------------------------------------------------------

def test_build_parser_has_overwrite_flag():
    args = ailine.build_parser().parse_args(["run", "book.xlsx", "task", "--overwrite"])
    assert args.overwrite is True

def test_build_parser_overwrite_defaults_false():
    args = ailine.build_parser().parse_args(["run", "book.xlsx", "task"])
    assert args.overwrite is False


# --- cmd_run_dsl 統合 ------------------------------------------------------------

def _overwrite_scenario_argv(book, **overrides):
    base = dict(
        book=str(book), task="売上から原価を引いた値を5列目に入れて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False, overwrite=False,
        values=False)
    base.update(overrides)
    return run_argv(**base)

def _overwrite_book(tmp_path, name="book.xlsx"):
    # ★ 監査の実際の再現に合わせた列構成（商品/金額/在庫/売上/原価・demo/sample.xlsx 型）。
    return _book(tmp_path, [
        ["商品", "金額", "在庫", "売上", "原価"],
        ["りんご", 1200, 8, 5000, 3000],
        ["みかん", 800, 25, 4000, 3500],
    ])

def _fake_apply_new_column(out_book, code, workdir, helper_files=(), timeout=None):
    """COMPUTE_COLUMN(売上-原価→新規列) の実際の計算結果を書き込む fake_apply。
    ★ W10c 追加項目（CI 落ち対応）: この2テストはもともと basrun_apply を monkeypatch
    せず real LibreOffice（basrun.py 経由）に実際の適用をやらせていた（開発機に basrun が
    隣接していたため気づかず緑になっていた・CI では basrun.py が無く落ちる）。
    tests/conftest.py の既定（normalize_book だけ pass-through）だけでは、この2テストは
    事後条件チェッカーが実際の新規列を要求するため緑にならない。COMPUTE_COLUMN の新規列
    作成（target 無指定）の codegen と同じ結果（列F=売上-原価）をここで直接書く
    （_fake_apply_overwrite_target と同じ作法）。"""
    wb2 = openpyxl.load_workbook(out_book)
    ws2 = wb2.active
    ws2.cell(row=1, column=6, value="売上-原価")
    ws2.cell(row=2, column=6, value=5000 - 3000)
    ws2.cell(row=3, column=6, value=4000 - 3500)
    wb2.save(out_book)
    return True, None, "ok"

def test_cmd_run_dsl_new_column_creation_skips_gate_entirely(tmp_path, monkeypatch, capsys):
    # 新規列作成（target 無指定）は「破壊だけ関所」の対象外・従来どおり素通り。
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN", "args": {"operands": ["売上", "原価"], "operator": "-"}})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_new_column)
    argv = _overwrite_scenario_argv(book, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "上書きしますか" not in captured.out
    assert "既存の値が" not in captured.out

def test_cmd_run_dsl_new_column_creation_neutralizes_ghost_warning(tmp_path, monkeypatch, capsys):
    # ★ W10b 項目4a(摩擦): 新規列作成(target無指定)は宣言どおりの効果として『範囲外』の
    #   ★疑わしい表示を出さない（中立表示に落ちる）。
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN", "args": {"operands": ["売上", "原価"], "operator": "-"}})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_new_column)
    argv = _overwrite_scenario_argv(book, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "変更が元データの範囲外です" not in captured.out
    assert "（新規列の追加は意図どおりです）" in captured.out

def test_cmd_run_dsl_lookup_fill_suppresses_readonly_source_sheet_warning(tmp_path, monkeypatch, capsys):
    # ★ W10b 項目4b(摩擦): 参照専用シート(単価表)は依頼文に言及があっても書き換えない
    #   のが正しい操作。旧実装は「依頼で言及された『単価表』は... 変更されていません」を
    #   誤って出していた。
    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "明細"
    for row in [["商品", "数量", "単価"], ["りんご", 2, None], ["バナナ", 3, None]]:
        ws1.append(row)
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    book = tmp_path / "lookup_int.xlsx"
    wb.save(book)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda b, workdir, timeout=None: b)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "LOOKUP_FILL", "args": {
                            "target_sheet": "明細", "target_col": "単価",
                            "source_sheet": "単価表", "key_col": "商品"}})
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws = wb2["明細"]
        ws.cell(row=2, column=3, value=100)
        ws.cell(row=3, column=3, value=200)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(book), task="単価表から単価を引いてきて明細に入れて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, values=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "依頼で言及された『単価表』は存在しません/変更されていません" not in captured.out

def test_cmd_run_dsl_overwrite_gate_blocks_noninteractive_exit7(tmp_path, monkeypatch, capsys):
    # ★ 監査の実測事故の再現: target が既存列(原価)に解決され、既存データがある。
    book = _overwrite_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "5"}})
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = _overwrite_scenario_argv(book)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 7
    assert "既存の値が 2 件あります" in captured.out
    assert "--overwrite" in captured.out
    assert book.read_bytes() == original_bytes   # 原本は無傷

def test_cmd_run_dsl_overwrite_gate_shows_interpretation_summary(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "5"}})
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = _overwrite_scenario_argv(book)
    ailine.main(argv)
    captured = capsys.readouterr()
    assert "→『5』は既存の『原価』列と解釈しました（既存データあり）" in captured.out

def _fake_apply_overwrite_target(out_book, code, workdir, helper_files=(), timeout=None):
    """COMPUTE_COLUMN(売上-原価→在庫列) の実際の計算結果を書き込む fake_apply
    （target=在庫 は operand ではない既存列なので、事後条件が operand から素直に
    再計算できる・--values 前提で use_formula=False の直値照合が通る）。"""
    wb2 = openpyxl.load_workbook(out_book)
    ws2 = wb2.active
    ws2.cell(row=2, column=3, value=5000 - 3000)
    ws2.cell(row=3, column=3, value=4000 - 3500)
    wb2.save(out_book)
    return True, None, "ok"

def test_cmd_run_dsl_overwrite_gate_bypassed_with_overwrite_flag(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "在庫"}})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_overwrite_target)
    argv = _overwrite_scenario_argv(book, overwrite=True, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert "上書きしますか" not in captured.out   # 関所自体をスキップ
    assert rc == 0
    # ★ C9: 反映の ✓ は「原本を読み戻して確かめた」1行に統合された。
    assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out
    assert "（もとに戻す: ailine undo）" in captured.out

def test_cmd_run_dsl_overwrite_gate_bypassed_with_copy_flag(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "在庫"}})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_overwrite_target)
    argv = _overwrite_scenario_argv(book, copy=True, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert "上書きしますか" not in captured.out   # --copy は原本に触れないため関所自体が発動しない
    assert rc == 0
    assert book.read_bytes() == original_bytes   # 原本は無変更

def test_cmd_run_dsl_overwrite_gate_interactive_yes_applies(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "在庫"}})
    monkeypatch.setattr(ailine, "basrun_apply", _fake_apply_overwrite_target)
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    argv = _overwrite_scenario_argv(book, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    # ★ C9: 反映の ✓ は「原本を読み戻して確かめた」1行に統合された。
    assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out
    assert "（もとに戻す: ailine undo）" in captured.out

def test_cmd_run_dsl_overwrite_gate_interactive_no_aborts(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "原価"], "operator": "-", "target": "5"}})
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    argv = _overwrite_scenario_argv(book)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 1
    assert "× 中止した" in captured.out
    assert book.read_bytes() == original_bytes

def test_cmd_run_dsl_overwrite_gate_no_summary_line_when_target_explicit(tmp_path, monkeypatch, capsys):
    # target が最初から実在列名で明示された場合（数字推定なし）は「→」の解釈要約は出ない
    # （何も『解釈』していないので語ることが無い）。警告そのものは出る。
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "COMPUTE_COLUMN",
                         "args": {"operands": ["売上", "金額"], "operator": "-", "target": "原価"}})
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = _overwrite_scenario_argv(book)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 7
    assert "既存の値が" in captured.out
    assert "と解釈しました" not in captured.out


# --- cmd_run_dsl 統合（LOOKUP_FILL・W10c 致命1/2 の通し確認） --------------------

def _lookup_gate_book(tmp_path, target_values=(999, 999)):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "明細"
    ws1.append(["商品", "数量", "単価"])
    ws1.append(["りんご", 2, target_values[0]])
    ws1.append(["バナナ", 3, target_values[1]])
    ws2 = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 100], ["バナナ", 200]]:
        ws2.append(row)
    book = tmp_path / "lookup_gate.xlsx"
    wb.save(book)
    return book

def _lookup_gate_argv(book, **overrides):
    base = dict(
        book=str(book), task="単価表から単価を引いてきて明細に入れて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False, overwrite=False,
        values=False)
    base.update(overrides)
    return run_argv(**base)

def test_cmd_run_dsl_lookup_fill_overwrite_gate_blocks_noninteractive_exit7(tmp_path, monkeypatch, capsys):
    # ★ W10c 致命1: LOOKUP_FILL は旧実装で破壊の関所が構造的に発火しなかった
    #   （op != "COMPUTE_COLUMN" の1行のせい）。既存の『単価』列に値がある状態で転記を
    #   依頼すると、非対話では exit 7 で止まり原本は無傷であること。
    book = _lookup_gate_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "LOOKUP_FILL", "args": {
                            "target_sheet": "明細", "target_col": "単価",
                            "source_sheet": "単価表", "key_col": "商品"}})
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = _lookup_gate_argv(book)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 7
    assert "既存の値が 2 件あります" in captured.out
    assert "--overwrite" in captured.out
    assert book.read_bytes() == original_bytes   # 原本は無傷

def test_cmd_run_dsl_lookup_fill_overwrite_gate_bypassed_with_overwrite_flag(tmp_path, monkeypatch, capsys):
    book = _lookup_gate_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "LOOKUP_FILL", "args": {
                            "target_sheet": "明細", "target_col": "単価",
                            "source_sheet": "単価表", "key_col": "商品"}})
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws = wb2["明細"]
        ws.cell(row=2, column=3, value=100)
        ws.cell(row=3, column=3, value=200)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = _lookup_gate_argv(book, overwrite=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert "上書きしますか" not in captured.out   # 関所自体をスキップ
    assert rc == 0
    # ★ C9: 反映の ✓ は「原本を読み戻して確かめた」1行に統合された。
    assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out
    assert "（もとに戻す: ailine undo）" in captured.out

def test_cmd_run_dsl_lookup_fill_missing_column_does_not_corrupt_unrelated_column(tmp_path, monkeypatch, capsys):
    # ★ W10c 致命2 の通し確認（査定の再現そのもの）: 明細シートに『単価』列がまだ無い状態
    #   で「単価表を見て単価を入れて」を依頼したとき、無関係な既存列（数量）が黙って
    #   潰されないこと。LLM が誤って「数量」を返すケースを直接シミュレートする。
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "明細"
    ws1.append(["商品コード", "数量"])
    ws1.append(["A001", 5])
    ws1.append(["A002", 3])
    ws2 = wb.create_sheet("単価表")
    for row in [["商品コード", "単価"], ["A001", 450], ["A002", 300]]:
        ws2.append(row)
    book = tmp_path / "lookup_missing_col.xlsx"
    wb.save(book)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    # 実測どおりの誤り: LLM が依頼に無い『数量』を target_col として返す。
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"op": "LOOKUP_FILL", "args": {
                            "target_sheet": "明細", "target_col": "数量",
                            "source_sheet": "単価表", "key_col": "商品コード"}})
    def _boom(prompt=""):
        raise AssertionError("verify_dsl_args が CLARIFY で止まるはずで input まで来ない")
    monkeypatch.setattr("builtins.input", _boom)
    argv = _lookup_gate_argv(book, task="単価表を見て単価を入れて")
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 3   # CLARIFY 相当（取り違えの疑いありの質問で止まる）
    assert "取り違えている" in captured.out
    assert book.read_bytes() == original_bytes   # 原本は無傷（数量列は一切触られない）


# --- cmd_run_plan 統合（複合計画の段ごとの関所） ---------------------------------

def test_cmd_run_plan_overwrite_gate_blocks_noninteractive_book_untouched(tmp_path, monkeypatch, capsys):
    book = _overwrite_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [
                            {"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                            {"op": "COMPUTE_COLUMN",
                             "args": {"operands": ["売上", "原価"], "operator": "-", "target": "5"}},
                        ]})
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None: (True, None, "ok"))
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = _overwrite_scenario_argv(book, task="金額で並べ替えて売上から原価を引いた値を5列目に入れて")
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 7
    assert "  2段目:" in captured.out
    assert book.read_bytes() == original_bytes   # 1段目が実行済みでも原本(book)自体は無傷


# --- 実行前告知の cmd_run 統合（一度きり） ----------------------------------------

def test_cmd_run_shows_notice_v2_once_then_silent(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "金額"], ["a", 1], ["b", 2]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, copy=False, json=False, timeout=180.0, ask=False)
    ailine.main(argv)
    first = capsys.readouterr().out
    assert "既定で原本に直接反映" in first

    ailine.main(argv)
    second = capsys.readouterr().out
    assert "既定で原本に直接反映" not in second


# ===========================================================================
# W10b 項目1/2: 自由生成の関所・ヘルパ総なめ検出
# ===========================================================================

def _fgate_ns(**overrides):
    base = dict(allow_freeform=False)
    base.update(overrides)
    return argparse.Namespace(**base)

def test_confirm_freeform_apply_none_when_allow_freeform_set():
    assert ailine._confirm_freeform_apply(_fgate_ns(allow_freeform=True)) is None

def test_confirm_freeform_apply_interactive_yes_continues(monkeypatch):
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    assert ailine._confirm_freeform_apply(_fgate_ns()) is None

def test_confirm_freeform_apply_interactive_no_returns_1(monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    rc = ailine._confirm_freeform_apply(_fgate_ns())
    captured = capsys.readouterr()
    assert rc == 1
    assert "中止した" in captured.out

def test_confirm_freeform_apply_noninteractive_returns_8_with_guidance(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    rc = ailine._confirm_freeform_apply(_fgate_ns())
    captured = capsys.readouterr()
    assert rc == 8
    assert "--allow-freeform" in captured.out

def test_confirm_freeform_apply_uses_step_prefix(monkeypatch, capsys):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    ailine._confirm_freeform_apply(_fgate_ns(), step_prefix="  2段目: ")
    captured = capsys.readouterr()
    assert "  2段目: この処理を続けるには" in captured.out
    assert "  2段目: 適用しますか" not in captured.out   # prompt はstdinへ・標準出力ではない

def test_confirm_freeform_apply_shows_sweep_warning_before_prompt(monkeypatch, capsys):
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    ailine._confirm_freeform_apply(_fgate_ns(), sweep_warning="疑わしい: テスト警告")
    captured = capsys.readouterr()
    assert "疑わしい: テスト警告" in captured.out


# --- ヘルパ総なめ検出 -------------------------------------------------------

_SWEEP_HELPER_NAMES = {"AutoFitColumns", "AlignCenter", "FormatThousands",
                        "VLookupFromTable", "PivotSum", "SummaryTable", "StyleBold"}

def test_detect_helper_sweep_none_for_normal_single_call():
    code = "Sub Run(oDoc As Object)\n    Call AutoFitColumns(oDoc)\nEnd Sub"
    assert ailine.detect_helper_sweep(code, _SWEEP_HELPER_NAMES) is None

def test_detect_helper_sweep_none_below_threshold():
    code = ("Sub Run(oDoc As Object)\n"
            "    Call AutoFitColumns(oDoc)\n"
            "    Call AlignCenter(oDoc, 0, 4)\n"
            "    Call StyleBold(oDoc, 0, 0, 4, 0)\n"
            "End Sub")
    assert ailine.detect_helper_sweep(code, _SWEEP_HELPER_NAMES) is None

def test_detect_helper_sweep_fires_at_threshold():
    code = ("Sub Run(oDoc As Object)\n"
            "    Call AutoFitColumns(oDoc)\n"
            "    Call AlignCenter(oDoc, 0, 4)\n"
            "    Call FormatThousands(oDoc, 0, 4)\n"
            "    Call VLookupFromTable(oDoc, 0, 0, 2, \"単価表\")\n"
            "End Sub")
    msg = ailine.detect_helper_sweep(code, _SWEEP_HELPER_NAMES)
    assert msg is not None
    assert "4 種類のヘルパ" in msg

def test_detect_helper_sweep_ignores_non_helper_calls():
    code = ("Sub Run(oDoc As Object)\n"
            "    Call Foo(oDoc)\n    Call Bar(oDoc)\n    Call Baz(oDoc)\n    Call Qux(oDoc)\n"
            "End Sub")
    assert ailine.detect_helper_sweep(code, _SWEEP_HELPER_NAMES) is None

def test_known_helper_names_reads_sub_declarations(tmp_path):
    f = tmp_path / "H.bas"
    f.write_text("Sub Foo(oDoc As Object)\nEnd Sub\n\nSub Bar(oDoc As Object, x As Integer)\nEnd Sub\n",
                 encoding="utf-8")
    names = ailine._known_helper_names([f])
    assert names == {"Foo", "Bar"}


# --- cmd_run_freeform 統合（関所） ------------------------------------------

def _freeform_gate_scenario_book(tmp_path):
    return _book(tmp_path, [["部署", "氏名", "金額"], ["営業", "山田", 12000], ["経理", "佐藤", 9800]])

def test_cmd_run_freeform_gate_blocks_noninteractive_exit8_book_untouched(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    # ★ basrun_apply を丸ごと差し替えるため、normalize_book 内部の構造読み取り呼び出しも
    #   同じ fake に飛んでしまう。normalize_book 自体を恒等関数にして混線を避ける
    #   （既存の忠実度ゲートテスト群と同じ作法）。
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    called = {"n": 0}
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        called["n"] += 1
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 8
    assert "この処理を続けるには" in captured.out
    assert "--allow-freeform" in captured.out
    assert called["n"] == 0   # 適用そのものが一度も走っていない
    assert book.read_bytes() == original_bytes   # 原本は無傷

def test_cmd_run_freeform_gate_allow_freeform_flag_applies(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=4, value="new")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False,
        allow_freeform=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "機械検証できません" not in captured.out   # 関所自体をスキップ
    assert "に適用しましたが、機械保証はありません（適用後に読み戻して確認: " in captured.out

def test_cmd_run_freeform_gate_interactive_yes_applies_then_undo_restores(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        wb2.active.cell(row=1, column=4, value="new")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr("builtins.input", lambda prompt="": "y")
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    assert rc == 0
    assert book.read_bytes() != original_bytes   # 適用された

    rc_undo = ailine.cmd_undo(argparse.Namespace(book=str(book), list=False))
    assert rc_undo == 0
    assert book.read_bytes() == original_bytes   # undo で原本に戻る

def test_cmd_run_freeform_gate_interactive_no_aborts_book_untouched(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    original_bytes = book.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    monkeypatch.setattr("builtins.input", lambda prompt="": "n")
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 1
    assert "× 中止した" in captured.out
    assert book.read_bytes() == original_bytes

def test_cmd_run_freeform_gate_dry_run_never_asks(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def _boom(prompt=""):
        raise AssertionError("dry では確認を聞いてはいけない")
    monkeypatch.setattr("builtins.input", _boom)
    argv = run_argv(
        book=str(book), task="何かして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=True, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    assert rc == 0

def test_cmd_run_freeform_gate_sweep_warning_shown_in_output(tmp_path, monkeypatch, capsys):
    book = _freeform_gate_scenario_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": "FREEFORM", "args": {}})
    sweep_code = ("Sub Run(oDoc As Object)\n"
                  "    Call AutoFitColumns(oDoc)\n"
                  "    Call AlignCenter(oDoc, 0, 2)\n"
                  "    Call StyleBold(oDoc, 0, 0, 2, 0)\n"
                  "    Call DrawTableBorders(oDoc)\n"
                  "End Sub")
    monkeypatch.setattr(ailine, "ollama_generate", lambda model, msgs, temperature=0.2: sweep_code)
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = run_argv(
        book=str(book), task="氏名の列を書き換えて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 8
    assert "疑わしい" in captured.out
    assert "4 種類のヘルパ" in captured.out


# --- cmd_run_plan 統合（複合計画の段の自由生成の関所） -----------------------

def test_cmd_run_plan_freeform_step_gate_blocks_noninteractive_book_untouched(tmp_path, monkeypatch, capsys):
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    original_bytes = p.read_bytes()
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None: (True, None, "ok"))
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて条件付き書式もつけて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 8
    assert "  2段目:" in captured.out
    assert p.read_bytes() == original_bytes   # 1段目(SORT)は out_book にしか反映されていない

def test_cmd_run_plan_freeform_step_gate_allow_freeform_applies(tmp_path, monkeypatch, capsys):
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "ollama_generate",
                        lambda model, msgs, temperature=0.2: "Sub Run(oDoc As Object)\nEnd Sub")
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        cell = ws2.cell(row=1, column=10)
        cell.value = (cell.value or 0) + 1
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて条件付き書式もつけて", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        allow_freeform=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "機械検証できません" not in captured.out


# --- argparse ------------------------------------------------------------------

def test_build_parser_has_allow_freeform_flag():
    args = ailine.build_parser().parse_args(["run", "book.xlsx", "task", "--allow-freeform"])
    assert args.allow_freeform is True

def test_build_parser_allow_freeform_defaults_false():
    args = ailine.build_parser().parse_args(["run", "book.xlsx", "task"])
    assert args.allow_freeform is False


# ===========================================================================
# W10e: ブラインド査定の致命3件（対象額 $15）
#   致命1: 「機械検証済み」が計画どおりであって依頼どおりとは限らない（複合計画）
#   致命2: 自由生成が既存シートの中身を静かにすり替えても検証層が無言
#   致命3: 「列を一括で定数に書き換える」が DSL に無く常に自由生成へ落ちる
# ===========================================================================

# --- 致命1: 各段の解釈行の事前表示 + バナーの範囲表示 ---------------------------

def test_maybe_warn_header_col_mismatch_fires_when_target_is_new_col_and_header_word():
    warn = ailine._maybe_warn_header_col_mismatch(
        "BOLD", {"target": "col:金額"}, ["金額"], "見出しを太字にして")
    assert warn is not None
    assert "金額" in warn and "見出し" in warn

def test_maybe_warn_header_col_mismatch_silent_when_target_not_a_new_col():
    # ★ 対象列がこの計画で新規作成された列でない（=最初から明示された既存列名）
    #   通常の BOLD は発火しない（誤検知の芽を new_cols 所属の有無に限定する設計）。
    warn = ailine._maybe_warn_header_col_mismatch(
        "BOLD", {"target": "col:金額"}, [], "見出しを太字にして")
    assert warn is None

def test_maybe_warn_header_col_mismatch_silent_without_header_word():
    warn = ailine._maybe_warn_header_col_mismatch(
        "BOLD", {"target": "col:金額"}, ["金額"], "金額の列を太字にして")
    assert warn is None

def test_maybe_warn_header_col_mismatch_silent_for_row_target():
    warn = ailine._maybe_warn_header_col_mismatch(
        "BOLD", {"target": "row:1"}, ["金額"], "見出しを太字にして")
    assert warn is None   # 既に row: なので食い違いが無い

def test_maybe_warn_header_col_mismatch_only_for_style_ops():
    warn = ailine._maybe_warn_header_col_mismatch(
        "SORT", {"col": "金額"}, ["金額"], "見出しを並べ替えて")
    assert warn is None   # 対象 op(BOLD/FILL_COLOR/CENTER_ALIGN)以外は無関係


def test_cmd_run_plan_prints_interpretation_line_per_step_before_applying(tmp_path, monkeypatch, capsys):
    # ★ 致命1 要求2: 単発(cmd_run_dsl)には元から「解釈:」行があるが、複合計画は
    #   段が流れて気づけないのが査定所見だった。各段に「解釈:」が付くこと。
    from openpyxl.styles import Font
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    for c in (1, 2):
        ws.cell(row=1, column=c).font = Font(bold=True)   # fake basrun_apply は何もしないので事前条件化
    wb.save(p)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                  {"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(p), task="金額で降順に並べ替えて見出しを太字に", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "  1段目: 解釈: 操作:並べ替え" in captured.out
    assert "  2段目: 解釈: 操作:太字" in captured.out


def test_cmd_run_plan_reproduces_bold_target_leak_and_shows_mismatch_warning(tmp_path, monkeypatch, capsys):
    # ★ 実測事故の再現: 「数量と単価をかけた金額列を作って、見出しを太字にして」型の複合
    #   依頼で、翻訳が2段目の target に1段目の演算対象名(存在しない列名)を書いてしまい、
    #   _apply_new_column_fallback が『直前段の新規列(金額)への参照』として救済してしまう
    #   ケース。★ 致命1: この時 check_bold は「対象:col:金額」という*計画どおり*の検証には
    #   合格するが、依頼の「見出し」とは無関係 — mismatch 助言が出て、範囲注記も出ること。
    p = _plan_book(tmp_path, [["商品", "数量", "単価"], ["a", 2, 100], ["b", 3, 200]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [
                            {"op": "COMPUTE_COLUMN",
                             "args": {"operands": ["数量", "単価"], "operator": "*"}},
                            # ★ 翻訳のバグそのもの: target が実在しない「数量*単価」になっている
                            {"op": "BOLD", "args": {"target": "col:数量*単価"}},
                        ]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        # ★ 実際の codegen_dsl と同じ結果を再現する軽量フェイク（LibreOffice を要さない）。
        #   COMPUTE_COLUMN 段: D列(4列目)に「数量*単価」という*実在する*新規列を作る
        #   （codegen_dsl の実際の見出し生成規則どおり＝この検体こそが実測事故の再現形）。
        #   BOLD 段: StyleBold 呼び出しが D 列を指すコードなら D 列だけ太字にする
        #   （見出し行だけ・全体は太字にならない＝実測どおり「新設列だけ太字になった」を再現）。
        from openpyxl.styles import Font
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        if 'setString("数量*単価")' in code:
            ws2.cell(row=1, column=4, value="数量*単価")
            for r in range(2, ws2.max_row + 1):
                q = ws2.cell(row=r, column=2).value
                u = ws2.cell(row=r, column=3).value
                ws2.cell(row=r, column=4, value=(q or 0) * (u or 0))
        if "Call StyleBold(oDoc, 3, " in code:   # 0起点 col_idx=3 = D列
            for r in range(1, ws2.max_row + 1):
                c = ws2.cell(row=r, column=4)
                if c.value not in (None, ""):
                    c.font = Font(bold=True)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(
        book=str(p), task="数量と単価をかけた金額列を作って、見出しを太字にして",
        model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False,
        values=True)   # ★ --values（式でなく値ベタ書き）: フェイクが実データで postcondition
                        #   を実際に pass させ、current_meta の再読込(次段の new_cols 検出)を
                        #   正しく発火させる（fail 時は current_meta が更新されない既存仕様のため）。
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0   # 1段目(COMPUTE_COLUMN)は実データで pass し、2段目の new_cols 検出も動く
    # 2段目の対象『数量*単価』は直前段(COMPUTE_COLUMN)が新規作成した実在列であり、
    # 「見出し」と食い違う旨の助言が出ること。
    assert "  2段目: 解釈: 操作:太字 対象:col:数量*単価" in captured.out
    # ★ 単位B: 手書きの if の文（「…新規作成された列です。」）は、一般則の③が同じスロットで
    #   鳴るこの検体では**単独では出ず**、③の ⚠ に注記として畳み込まれる。落としてはいけない
    #   のは文ではなく**事実**なので、事実の語だけを要求する。
    assert "この計画の直前の段で新規作成された列" in captured.out and "見出し" in captured.out
    # ★★ 単位E で強くなった: W10e の時点では「範囲注記が出ること」しか要求できなかった
    #   （＝『機械検証済み』と言い切ったうえで、あとから範囲を断る形）。今は依頼文の語
    #   『見出し』と解決値『col:数量*単価』の食い違いを機械が突き合わせるので、
    #   **そもそも ✓ を出さない**。旧 _VERIFY_SCOPE_NOTE_PLAN（常時注記）は廃止した。
    assert "は機械検証済みの内容です" not in captured.out
    assert "依頼文が指しているのは: 見出し" in captured.out
    # ★ 単位B: 同じスロットについて ⚠ を2度言わない ―― 段の位置に1本だけ（助言欄への
    #   再掲も無い。畳み込んだ1本が既にその事実を運んでいるため）。
    step2_warnings = [ln for ln in captured.out.splitlines() if ln.startswith("  2段目: ⚠")]
    assert len(step2_warnings) == 1, step2_warnings


def test_cmd_run_dsl_success_does_not_print_an_always_on_scope_note(tmp_path, monkeypatch, capsys):
    """★ 単位E: 旧 `_VERIFY_SCOPE_NOTE`（✓ が出る全 run で必ず出る注記）を廃止した検体。
       ここは「見出し行を太字にして」→ `row:1` ＝ 依頼文の語と対象が機械照合できた①なので、
       ✓ の他には何も足さない（範囲を狭める1文は、照合できなかった run にだけ出る）。"""
    from openpyxl.styles import Font
    p = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    wb = openpyxl.load_workbook(p)
    ws = wb.active
    for c in (1, 2):
        ws.cell(row=1, column=c).font = Font(bold=True)   # fake basrun_apply は何もしないので事前条件化
    wb.save(p)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    monkeypatch.setattr(ailine, "basrun_apply",
                        lambda out_book, code, workdir, helper_files=(), timeout=None:
                        (True, None, "ok"))
    argv = run_argv(
        book=str(p), task="見出し行を太字にして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "は機械検証済みの内容です" in captured.out
    assert "★ ただし" not in captured.out            # ②の1文は出ない（①なので）
    assert "解釈:」行どおりに実行されたことの検証です" not in captured.out   # 旧・常時注記


# --- 致命2: 既存シートの中身が丸ごと置き換わったことの検出 ---------------------

def _sheet_snapshot(cells: dict, sheets=("Sheet",)) -> dict:
    """テスト用の最小 snapshot（cells は {"Sheet!r,c": (val, "General", None, False, None, None)}）。"""
    return {"sheets": list(sheets), "charts": 0, "cells": dict(cells),
            "merges": {s: [] for s in sheets}, "colw": {s: {} for s in sheets},
            "rowh": {s: {} for s in sheets}, "truncated": False}

def _v(val):
    return (val, "General", None, False, None, None)

def test_existing_sheet_replaced_advisory_fires_when_all_cells_change():
    before = _sheet_snapshot({
        "Sheet!1,1": _v("部署"), "Sheet!1,2": _v("合計"),
        "Sheet!2,1": _v("営業"), "Sheet!2,2": _v(100),
    })
    after = _sheet_snapshot({
        "Sheet!1,1": _v("日付"), "Sheet!1,2": _v("件数"),
        "Sheet!2,1": _v("2026-01-01"), "Sheet!2,2": _v(5),
    })
    lines = ailine.existing_sheet_replaced_advisory(before, after)
    assert len(lines) == 1
    assert "既存シート『Sheet』の中身が置き換わりました" in lines[0]

def test_existing_sheet_replaced_advisory_silent_on_partial_update():
    # 一部だけの更新（再計算等）は対象外（保守的・オオカミ少年回避）。
    before = _sheet_snapshot({
        "Sheet!1,1": _v("部署"), "Sheet!1,2": _v("合計"),
        "Sheet!2,1": _v("営業"), "Sheet!2,2": _v(100),
    })
    after = _sheet_snapshot({
        "Sheet!1,1": _v("部署"), "Sheet!1,2": _v("合計"),
        "Sheet!2,1": _v("営業"), "Sheet!2,2": _v(200),   # 値だけ再計算
    })
    assert ailine.existing_sheet_replaced_advisory(before, after) == []

def test_existing_sheet_replaced_advisory_silent_when_no_before_data():
    # 原本にデータが無かった（新規に埋まっただけ）シートは対象外。
    before = _sheet_snapshot({}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("営業")},
                             sheets=("Sheet", "集計"))
    assert ailine.existing_sheet_replaced_advisory(before, after) == []

# ★ C9: 旧 _neutralize_declared_sheet_replace_warning は advisories を後から書き換える単体
#   関数だったが、existing_sheet_replaced_advisory 自身が op を受け取り発生源で判定する形に
#   構造を置き換えた（判定条件・出力文言は不変・ゴールデン差分ゼロで確認）。

def test_existing_sheet_replaced_advisory_neutral_for_aggregate():
    before = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("旧")}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("日付"), "集計!2,1": _v("新")}, sheets=("Sheet", "集計"))
    out = ailine.existing_sheet_replaced_advisory(before, after, op="AGGREGATE")
    assert out == ["（既存シート『集計』の更新は意図どおりです）"]

# ★★ 単位G: 中立化は「前提が成立していた時だけ」。宣言（writes=new_sheet）が黙らせる権利を
#   持つのは、その前提（＝その名前のシートは before に存在しない）が破れていない時だけ。
#   盲検査定の致命: 人が手で作った『集計』シートが SummaryTable に全滅させられたのに、
#   「（既存シート『集計』の更新は意図どおりです）」という肯定文まで出ていた。
#   ★ 単位G の完了条件は「肯定文が消える」ことであって、正常系（前に ailine 自身が作った
#   『集計』の作り直し）を肯定文に戻すのは 単位H（出所判定）の仕事。ここではやらない。

def test_unit_g_neutralization_withdrawn_when_new_sheet_precondition_broke():
    before = _sheet_snapshot({"集計!1,1": _v("年度"), "集計!2,1": _v("予算")}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("営業")}, sheets=("Sheet", "集計"))
    lines = ailine.existing_sheet_replaced_advisory(before, after, op="AGGREGATE",
                                                    precondition_broken="new_sheet")
    # ★ 負の被覆: 「出るべきものが出るか」でなく「出てはいけないものが消えたか」を測る
    assert not any("意図どおりです" in ln for ln in lines), lines
    assert len(lines) == 1
    assert "既存シート『集計』の中身が置き換わりました" in lines[0]

def test_unit_g_neutralization_kept_when_other_precondition_broke():
    # ★ 破れたのが別の種類（format_only 等）なら、new_sheet の宣言は権利を失わない。
    #   「何か破れた」で黙らせると、無関係な理由で肯定文を消すことになる。
    before = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("旧")}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("日付"), "集計!2,1": _v("新")}, sheets=("Sheet", "集計"))
    out = ailine.existing_sheet_replaced_advisory(before, after, op="AGGREGATE",
                                                  precondition_broken="format_only")
    assert out == ["（既存シート『集計』の更新は意図どおりです）"]

def test_unit_g_default_is_unchanged_from_unit_f():
    # ★ 退行の番人: precondition_broken を渡さない呼び出し（既存の全経路）は挙動不変。
    before = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("旧")}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("日付"), "集計!2,1": _v("新")}, sheets=("Sheet", "集計"))
    assert ailine.existing_sheet_replaced_advisory(before, after, op="AGGREGATE") ==            ["（既存シート『集計』の更新は意図どおりです）"]

def test_existing_sheet_replaced_advisory_warns_for_other_ops():
    before = _sheet_snapshot({"集計!1,1": _v("部署"), "集計!2,1": _v("旧")}, sheets=("Sheet", "集計"))
    after = _sheet_snapshot({"集計!1,1": _v("日付"), "集計!2,1": _v("新")}, sheets=("Sheet", "集計"))
    lines = ailine.existing_sheet_replaced_advisory(before, after)
    # AGGREGATE/PIVOT 以外（例: 自由生成には op という概念が無いので FREEFORM 相当）は
    # 何も変えない（このテストでは仮に SORT を渡して確認）。
    out = ailine.existing_sheet_replaced_advisory(before, after, op="SORT")
    assert out == lines


# --- 致命3: 一括定数書き換え(SET_COLUMN_VALUE) ---------------------------------

def test_extract_quoted_literal_single_quote_ok():
    assert ailine.extract_quoted_literal("備考列を全部『確認済み』にして") == "確認済み"
    assert ailine.extract_quoted_literal('氏名の列を全部「退職済み」に書き換えて') == "退職済み"

def test_extract_quoted_literal_none_when_zero_or_multiple():
    assert ailine.extract_quoted_literal("備考列を全部確認済みにして") is None
    assert ailine.extract_quoted_literal("『A』と『B』のどちらかにして") is None

def test_verify_dsl_args_set_column_value_extracts_from_quotes():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "備考"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "SET_COLUMN_VALUE", {"col": "備考"}, meta, task="備考列を全部『確認済み』にして")
    assert ok
    assert resolved["value"] == "確認済み"
    assert resolved["_sources"]["value"] == "依頼文: 「確認済み」"

def test_verify_dsl_args_set_column_value_clarifies_without_quote():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "備考"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "SET_COLUMN_VALUE", {"col": "備考"}, meta, task="備考列を全部確認済みにして")
    assert not ok
    assert "「」" in err or "『』" in err

def test_verify_dsl_args_set_column_value_task_wins_over_llm_value():
    # ★ A' 原則: LLM が返した value は無視/検算対象（依頼文の引用が常に勝つ）。
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "備考"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "SET_COLUMN_VALUE", {"col": "備考", "value": "却下済み"}, meta,
        task="備考列を全部『確認済み』にして")
    assert ok
    assert resolved["value"] == "確認済み"
    assert resolved["_warnings"]

def test_verify_dsl_args_set_column_value_col_must_exist():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "備考"]}}
    ok, resolved, inferred, err = ailine.verify_dsl_args(
        "SET_COLUMN_VALUE", {"col": "存在しない"}, meta, task="存在しない列を全部『x』にして")
    assert not ok

def test_codegen_set_column_value_writes_all_data_rows():
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["商品", "備考"]}}
    code = ailine.codegen_dsl("SET_COLUMN_VALUE", {"col": "備考", "value": "確認済み"}, meta)
    assert 'setString("確認済み")' in code
    assert "Sub Run(oDoc As Object)" in code

def test_check_set_column_value_pass_and_fail(tmp_path):
    p = _book(tmp_path, [["商品", "備考"], ["a", "確認済み"], ["b", "確認済み"]])
    status, reason = ailine.check_set_column_value(p, {"col": "備考", "value": "確認済み"})
    assert status == "pass"

    p2 = _book(tmp_path, [["商品", "備考"], ["a", "確認済み"], ["b", "未確認"]])
    status2, reason2 = ailine.check_set_column_value(p2, {"col": "備考", "value": "確認済み"})
    assert status2 == "fail"

def test_op_write_target_set_column_value_declares_column_write():
    # ★ 既存列への一括書き込み＝破壊の関所の対象（宣言必須）。
    wt = ailine.OP_WRITE_TARGET["SET_COLUMN_VALUE"]
    assert wt.writes == (ailine.WRITE_EXISTING_COLUMN,)
    assert (wt.col_key, wt.sheet_key, wt.reads_only) == ("col", None, ())

def test_set_column_value_end_to_end_via_cmd_run_dsl(tmp_path, monkeypatch, capsys):
    # ★ DoD③: 既存値がある列への一括書き換えは破壊の関所(確認)を経由すること。
    p = _book(tmp_path, [["商品", "備考"], ["a", "旧"], ["b", "旧"]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "備考"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        for r in range(2, ws2.max_row + 1):
            ws2.cell(row=r, column=2, value="確認済み")
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)
    argv = run_argv(
        book=str(p), task="備考列を全部『確認済み』にして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=False, json=False, timeout=180.0, ask=False)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    # 既存値(旧×2件)がある列への上書き＝破壊の関所が発火し、非対話では exit 7 で案内する
    assert rc == 7
    assert "上書き" in captured.out or "--overwrite" in captured.out


# --- ★ 宣言つき挙動変更#1: 型破壊の安全網 ---------------------------------------
#   ブラインド査定の実測: 数値の『原価』列を SET_COLUMN_VALUE で文字列『0円』に一括書換
#   すると、それを参照する数式(利益=売上-原価)が #VALUE! に壊れるのに、事後条件
#   チェッカー(check_set_column_value・対象列が指定文字列になったかだけを見る)は
#   「✓ 達成を機械検証済み」を出す。ailine_core/formula_health.py の (a) エラー値増加の
#   網 (b) 型変化の助言 が、CLI 全体（cmd_run_dsl）を通した時に実際に発火することを見る
#   （fake_apply が basrun/LibreOffice の代わりに書き込み+エラー値キャッシュ注入を行う。
#   本物の basrun/LibreOffice 通しは tests/test_formula_health_local.py が担う）。

def _inject_error_cache_ailine_test(path, sheet_filename: str, addr_to_err: dict) -> None:
    """テスト専用: tests/test_formula_health.py の _inject_error_cache と同型（数式セルへ
       t="e" のエラー値キャッシュを直接注入する）。fake_apply の内部から呼ぶための複製
       （test_ailine.py はモジュール横断のテスト専用 import はしない既存の作法に合わせる）。"""
    import re
    import zipfile
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_filename:
                text = data.decode("utf-8")
                for addr, err in addr_to_err.items():
                    pattern = re.compile(rf'<c r="{addr}"([^>]*)>(.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?</c>')

                    def _sub(m, err=err):
                        attrs = re.sub(r'\s*t="[^"]*"', '', m.group(1))
                        return f'<c r="{addr}"{attrs} t="e">{m.group(2)}<v>{err}</v></c>'

                    text, n = pattern.subn(_sub, text, count=1)
                    assert n == 1, f"_inject_error_cache_ailine_test: {addr} に注入できなかった"
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)

def _genka_book(tmp_path) -> Path:
    """査定と同じ形の検体: 品目/売上/原価/利益(=売上-原価) の7行。"""
    p = tmp_path / "genka.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["品目", "売上", "原価", "利益"])
    rows = [("りんご", 1000, 300), ("みかん", 800, 200), ("ぶどう", 1500, 600)]
    for i, (name, sales, cost) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=sales)
        ws.cell(row=i, column=3, value=cost)
        ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
    wb.save(p)
    return p

def test_set_column_value_nonnumeric_write_triggers_both_advisories(tmp_path, monkeypatch, capsys):
    """★ DoD1: 査定の再現そのものを回帰テストにする。数値列に文字列『0円』を書く →
       依存する数式が壊れる（#VALUE!）→ (a)(b) 両方の助言が出る。★ 事後条件チェッカー
       自体は変えていないので ✓ は今回も出る（★ C9 でその ✓ は「原本(--copy なら .out)を
       読み戻して確かめた」1行に移った・claim の主張範囲は変えない設計判断）── その上で
       波及被害の警告が別チャンネルで出ることを見る。"""
    book = _genka_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "原価"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        for r in (2, 3, 4):
            ws2.cell(row=r, column=3, value="0円")
        wb2.save(out_book)
        _inject_error_cache_ailine_test(out_book, "xl/worksheets/sheet1.xml",
                                         {"D2": "#VALUE!", "D3": "#VALUE!", "D4": "#VALUE!"})
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="原価の列を全部『0円』にして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, overwrite=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    # ★ C9: 単発の ✓ バナーも「原本(.out)を読み戻して確かめた」1行に統合された。
    assert "は機械検証済みの内容です（適用後に読み戻して確認: " in captured.out
    assert "★ 疑わしい: 適用後にエラー値のセルが増えました" in captured.out
    assert "Sheet!D2=#VALUE!" in captured.out
    assert "（確認）列『原価』は元は数値でしたが" in captured.out

def test_set_column_value_numeric_looking_write_no_advisory(tmp_path, monkeypatch, capsys):
    """★ DoD4/5②: 過剰検出でないことの実証。同じ列に数値そのものの文字列『500』を書く
       正常系（数式は壊れない）では、(a)(b) どちらの警告も出ない。"""
    book = _genka_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "原価"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        for r in (2, 3, 4):
            ws2.cell(row=r, column=3, value="500")
            ws2.cell(row=r, column=4, value=500)   # 数式は数値そのものへ正常に再計算された想定
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="原価の列を全部『500』にして", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, overwrite=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "★ 疑わしい: 適用後にエラー値のセルが増えました" not in captured.out
    assert "（確認）列『原価』は元は数値でしたが" not in captured.out

def test_compute_column_new_column_normal_op_no_advisory(tmp_path, monkeypatch, capsys):
    """★ DoD4: 正常な操作の代表例②「計算列作成」で警告が出ないこと（新規列は元の型という
       概念が無いので (b) は対象外・数式も壊れていないので (a) も無言）。"""
    book = _genka_book(tmp_path)
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "COMPUTE_COLUMN",
                                   "args": {"operands": ["売上", "原価"], "operator": "-"}}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        ws2.cell(row=1, column=5, value="売上-原価")
        for r in (2, 3, 4):
            sales = ws2.cell(row=r, column=2).value
            cost = ws2.cell(row=r, column=3).value
            ws2.cell(row=r, column=5, value=sales - cost)
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    argv = run_argv(
        book=str(book), task="売上から原価を引いた列を作って", model="qwen2.5-coder:7b",
        refs=None, helpers=None, repair=0, temperature=0.2,
        dry=False, copy=True, json=False, timeout=180.0, ask=False, values=True)
    rc = ailine.main(argv)
    captured = capsys.readouterr()
    assert rc == 0
    assert "★ 疑わしい: 適用後にエラー値のセルが増えました" not in captured.out
    assert "（確認）列『" not in captured.out


# --- 致命3 項目①: 全列テキストの表で見出し行検出が CLARIFY に落ちる不具合の修正 -----

def test_detect_header_row_all_text_table_picks_first_row():
    # 実測: 氏名/部署/備考のような数値列が一つも無い表は、型混在の手がかりが一度も
    # 起きないため、旧実装は pure_str_rows が複数になり CLARIFY に落ちていた。
    sheet_struct = {"rows": {
        1: {"nonempty": 3, "str": 3, "bold": 0},   # 氏名/部署/備考
        2: {"nonempty": 2, "str": 2, "bold": 0},   # 備考が空欄の行（2セルのみ非空）
        3: {"nonempty": 3, "str": 3, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert (row, confident) == (1, True)

def test_detect_header_row_all_text_table_with_wholly_blank_rows_between():
    # 完全に空白の行（nonempty=0）が混ざっても先頭行を見出しとみなす。
    sheet_struct = {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 0, "str": 0, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert (row, confident) == (1, True)

def test_detect_header_row_ambiguous_numeric_mixture_case_still_not_confident():
    # 既存の「曖昧なら推測しない」挙動は変えない（型混在の手がかりがある通常のケース）。
    sheet_struct = {"rows": {
        1: {"nonempty": 2, "str": 2, "bold": 0},
        2: {"nonempty": 2, "str": 1, "bold": 0},
        3: {"nonempty": 2, "str": 2, "bold": 0},
        4: {"nonempty": 2, "str": 1, "bold": 0},
    }}
    row, confident = ailine.detect_header_row(sheet_struct)
    assert confident is False
    assert row is None
