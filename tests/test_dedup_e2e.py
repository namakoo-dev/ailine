"""DEDUP op（重複行の除去・非破壊形）の検体。
   ★ 実装前に凍結した赤い検体（freeform 廃止バンドルの前段・operator 要望 2 位）。

   形: EXTRACT の兄弟 ── writes=(new_sheet,)。「重複を除いた一覧を 新シート に」。
   ★ 行は消さない（破壊形は関所つきで第二波）。判定キー列は依頼文の名指し ──
   全列一致を既定にしない（「取引先が同じなら重複」は人の意図）。"""
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _book(tmp_path, rows):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(list(r))
    wb.save(p)
    return p


def _translate_dedup(monkeypatch, keys=("取引先",)):
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "DEDUP", "args": {"keys": list(keys)}})


def _fake_apply_dedup(monkeypatch, out_rows, sheet="取引先の重複除去"):
    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.create_sheet(sheet)
        for r in out_rows:
            ws.append(list(r))
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)


ROWS = [["取引先", "金額"], ["甲社", 100], ["甲社", 250], ["乙社", 200], ["甲社", 300]]


def test_dedup_keeps_first_per_key_and_names_dropped_rows(tmp_path, monkeypatch, capsys):
    """本命: キー列で最初の 1 行を残す・落とした行を名指し・分母つき・✓。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, ROWS)
    _translate_dedup(monkeypatch)
    _fake_apply_dedup(monkeypatch,
                      [["取引先", "金額"], ["甲社", 100], ["乙社", 200]])
    rc, out = _run_main(["run", str(book), "取引先が同じ行を重複として除いて", "--copy"], capsys)
    assert rc == 0, out
    assert "✓" in out
    assert "4行中2行を残しました" in out or "重複 2" in out or "2行を除" in out, \
        f"分母つきの報告が無い: {out}"
    assert "3行目" in out and "5行目" in out, f"落とした行の名指しが無い: {out}"
    assert "取引先" in out, "判定キー列の開示が無い"


def test_dedup_postcondition_catches_fake_that_keeps_duplicates(tmp_path, monkeypatch, capsys):
    """★ 両側の検査 その1: 出力に同キーの重複が残っていたら事後条件が落とす（✓ を出さない）。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, ROWS)
    _translate_dedup(monkeypatch)
    _fake_apply_dedup(monkeypatch,
                      [["取引先", "金額"], ["甲社", 100], ["甲社", 250], ["乙社", 200]])
    rc, out = _run_main(["run", str(book), "取引先が同じ行を重複として除いて", "--copy"], capsys)
    assert rc != 0, f"重複が残った出力が通った: {out}"
    assert "✓" not in out
    assert "甲社" in out, "どのキーが重複のままかの名指しが無い"


def test_dedup_postcondition_catches_fabricated_or_overdropped_rows(tmp_path, monkeypatch, capsys):
    """★ 両側の検査 その2: 元に無い行の捏造・残すべき行の落とし過ぎも落とす。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, ROWS)
    _translate_dedup(monkeypatch)
    _fake_apply_dedup(monkeypatch,
                      [["取引先", "金額"], ["甲社", 100]])   # 乙社を落とし過ぎ
    rc, out = _run_main(["run", str(book), "取引先が同じ行を重複として除いて", "--copy"], capsys)
    assert rc != 0, f"落とし過ぎが通った: {out}"
    assert "✓" not in out
    assert "乙社" in out, "落とし過ぎたキーの名指しが無い"


def test_dedup_without_named_key_clarifies_not_全列既定(tmp_path, monkeypatch, capsys):
    """判定キーが依頼文に無い → CLARIFY（全列一致を黙って既定にしない）。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, ROWS)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "DEDUP", "args": {}})
    rc, out = _run_main(["run", str(book), "重複を消して", "--copy"], capsys)
    assert rc == 3, f"キー未指定で走った (rc={rc}): {out}"
    assert "列" in out, f"どの列で重複と見るかの問いが無い: {out}"
