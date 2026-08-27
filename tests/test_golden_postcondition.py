"""C1-F3: run_postcondition(op, fixture, resolved, ...) の (status, reason) を凍結する。

対象: 事後条件が定義された全 op（POSTCONDITIONS の16 op + CHART の計17）について、
pass/warn/fail の各系（warn が定義されている op のみ）と、実行時例外を "error" に
変換する境界を1件収載する。

フィクスチャ（入力ブック）はこのファイル内の Python ビルダ関数で組み立てる
（openpyxl のスタイル/結合/チャート/zip 直挿しはオブジェクトで JSON 化できないため）。
ビルダ自体が git diff で読める「人間可読な入力記述」を兼ねる。ゴールデンとして凍結する
のは (status, reason) の出力側だけ（tests/golden/f3_postcondition/<name>.json）。

use_formula=True 系（式のキャッシュ値二層検証）は、openpyxl では数式のキャッシュ値を
計算できない（LibreOffice/Excel が計算した結果を xlsx の <v> に持たせる必要がある）ため、
_inject_formula_cache() で保存後の sheetN.xml に直接 <v> を挿すことで模擬する。
"""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

from golden._harness import GOLDEN_ROOT, assert_golden_json  # noqa: E402

F3_DIR = GOLDEN_ROOT / "f3_postcondition"

_THIN = Side(style="thin")
_ALL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _book(tmp_path, name, rows, sheet_name="Sheet"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _inject_formula_cache(path: Path, sheet_index: int, updates: dict) -> None:
    """保存済み xlsx の sheetN.xml へ <f>式</f> の直後にある空の <v></v> をキャッシュ値で
       埋める（openpyxl は保存時 <c><f>式</f><v></v></c> という空キャッシュ枠を既に
       書き出しているので、その中身を差し替えるだけでよい）。
       updates: {"D2": "600", ...}（キャッシュ値は文字列で渡す）。
       openpyxl は数式を計算しないため、use_formula=True の二層検証（式text +
       data_only キャッシュ値）を模擬するにはこの直接 XML 編集が要る。"""
    member = f"xl/worksheets/sheet{sheet_index}.xml"
    with zipfile.ZipFile(path) as z:
        xml = z.read(member).decode("utf-8")
        others = {n: z.read(n) for n in z.namelist() if n != member}
    for ref, val in updates.items():
        # ★ CI の長期赤の一因（2026-08-21 実測）: openpyxl の XML 直列化は lxml の有無で変わる。
        #   lxml あり（ローカル）は <f>式</f><v></v> と空の <v> 枠を書くが、lxml なし（CI）は
        #   et_xmlfile 直列化で <v> 枠そのものを書かないことがある。両方の形を受け、無ければ挿す。
        pattern = re.compile(rf'(<c r="{ref}"[^>]*><f>[^<]*</f>)(<v/>|<v>[^<]*</v>)?')
        m = pattern.search(xml)
        assert m, f"{ref} の <f> セルが見つからない（xml 断片は上の pattern 参照）"
        xml = xml[:m.start()] + m.group(1) + f"<v>{val}</v>" + xml[m.end():]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, data in others.items():
            z.writestr(n, data)
        z.writestr(member, xml)


def _inject_zip_stub_member(path: Path, member_name: str, content: bytes = b"<x/>") -> None:
    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as z:
        z.writestr(member_name, content)


def _style_range(ws, r1, c1, r2, c2, *, bold=None, fill_hex=None, center=None,
                  number_format=None, border=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if bold is not None:
                cell.font = Font(bold=bold)
            if fill_hex is not None:
                cell.fill = PatternFill(fgColor=fill_hex, fill_type="solid")
            if center is not None:
                cell.alignment = Alignment(horizontal="center" if center else "left")
            if number_format is not None:
                cell.number_format = number_format
            if border is not None:
                cell.border = border


# 名前 -> (op, args, header_row, use_formula, before_charts, builder(tmp_path)->(out_book, source_book|None))
CASES: dict = {}


def _add(name, op, args, builder, header_row=1, use_formula=False, before_charts=0):
    assert name not in CASES, f"重複した case 名: {name}"
    CASES[name] = (op, args, header_row, use_formula, before_charts, builder)


# --- SORT ---------------------------------------------------------------
def _b_sort_pass(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]]), None


def _b_sort_warn_one_row(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 100]]), None


def _b_sort_fail_not_sorted(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 300], ["b", 100], ["c", 200]]), None


_add("sort_pass_asc", "SORT", {"col": "金額", "order": "asc"}, _b_sort_pass)
_add("sort_warn_single_row", "SORT", {"col": "金額", "order": "asc"}, _b_sort_warn_one_row)
_add("sort_fail_not_sorted", "SORT", {"col": "金額", "order": "asc"}, _b_sort_fail_not_sorted)
_add("sort_fail_unknown_col", "SORT", {"col": "不明", "order": "asc"}, _b_sort_pass)

# --- COMPUTE_COLUMN（2列演算・値モード） -------------------------------------
def _b_cc_pass_values(tmp_path):
    return _book(tmp_path, "b.xlsx", [
        ["商品", "数量", "単価", "数量*単価"], ["a", 2, 100, 200], ["b", 3, 150, 450]]), None


def _b_cc_fail_mismatch(tmp_path):
    return _book(tmp_path, "b.xlsx", [
        ["商品", "数量", "単価", "数量*単価"], ["a", 2, 100, 999], ["b", 3, 150, 450]]), None


_add("compute_column_2op_pass_values", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "*"}, _b_cc_pass_values)
_add("compute_column_2op_fail_mismatch_values", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "*"}, _b_cc_fail_mismatch)


def _b_cc_formula_pass(tmp_path):
    p = _book(tmp_path, "b.xlsx", [
        ["商品", "数量", "単価", "数量*単価"], ["a", 2, 100, "=B2*C2"], ["b", 3, 150, "=B3*C3"]])
    _inject_formula_cache(p, 1, {"D2": "200", "D3": "450"})
    return p, None


_add("compute_column_2op_pass_formula_with_cache", "COMPUTE_COLUMN",
     {"operands": ["数量", "単価"], "operator": "*"}, _b_cc_formula_pass, use_formula=True)

# --- COMPUTE_COLUMN（単列×倍率・委譲パスの通過確認） ---------------------------
def _b_cc_single_pass(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額", "税込金額"], ["a", 100, 110], ["b", 200, 220]]), None


_add("compute_column_single_factor_pass_values", "COMPUTE_COLUMN",
     {"operands": ["金額"], "operator": "*", "factor": 1.1, "_new_col_label": "税込金額"},
     _b_cc_single_pass)

# --- LOOKUP_FILL ------------------------------------------------------------
def _b_lookup_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "単価"], ["a", ""], ["b", ""]]:
        ws.append(row)
    ws["B2"] = 100
    ws["B3"] = 150
    src = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["a", 100], ["b", 150]]:
        src.append(row)
    wb.save(p)
    return p, None


def _b_lookup_fail_mismatch(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "単価"], ["a", 999], ["b", 150]]:
        ws.append(row)
    src = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["a", 100], ["b", 150]]:
        src.append(row)
    wb.save(p)
    return p, None


def _b_lookup_fail_missing_sheet(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "単価"], ["a", 100]]), None


_LOOKUP_ARGS = {"target_sheet": "Sheet", "target_col": "単価", "source_sheet": "単価表", "key_col": "商品"}
_add("lookup_fill_pass", "LOOKUP_FILL", _LOOKUP_ARGS, _b_lookup_pass)
_add("lookup_fill_fail_mismatch", "LOOKUP_FILL", _LOOKUP_ARGS, _b_lookup_fail_mismatch)
_add("lookup_fill_fail_missing_source_sheet", "LOOKUP_FILL", _LOOKUP_ARGS, _b_lookup_fail_missing_sheet)

# --- AGGREGATE --------------------------------------------------------------
def _b_agg_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 100], ["a", 50], ["b", 200]]:
        ws.append(row)
    out = wb.create_sheet("集計")
    out.append(["分類", "合計"])   # ★ check_aggregate は row2 からデータ読み（row1=見出し前提）
    out.append(["a", 150])
    out.append(["b", 200])
    wb.save(p)
    return p, None


def _b_agg_fail_no_sheet(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 100]]), None


_add("aggregate_pass", "AGGREGATE", {"group_col": "商品", "value_col": "金額"}, _b_agg_pass)
_add("aggregate_fail_missing_summary_sheet", "AGGREGATE",
     {"group_col": "商品", "value_col": "金額"}, _b_agg_fail_no_sheet)


def _b_agg_fail_mismatch(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 100], ["a", 50], ["b", 200]]:
        ws.append(row)
    out = wb.create_sheet("集計")
    out.append(["分類", "合計"])
    out.append(["a", 999])
    out.append(["b", 200])
    wb.save(p)
    return p, None


_add("aggregate_fail_value_mismatch", "AGGREGATE",
     {"group_col": "商品", "value_col": "金額"}, _b_agg_fail_mismatch)

# --- BOLD -----------------------------------------------------------------
def _b_bold_col_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    _style_range(ws, 1, 1, 3, 1, bold=True)
    wb.save(p)
    return p, None


def _b_bold_row_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    _style_range(ws, 1, 1, 1, 2, bold=True)
    wb.save(p)
    return p, None


def _b_bold_fail(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 100]]), None


_add("bold_col_pass", "BOLD", {"target": "col:商品"}, _b_bold_col_pass)
_add("bold_row_pass", "BOLD", {"target": "row:1"}, _b_bold_row_pass)
_add("bold_fail_not_bold", "BOLD", {"target": "col:商品"}, _b_bold_fail)

# --- FILL_COLOR ------------------------------------------------------------
def _b_fill_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    _style_range(ws, 1, 2, 2, 2, fill_hex="FF0000")
    wb.save(p)
    return p, None


_add("fill_color_col_pass", "FILL_COLOR", {"target": "col:金額", "color": "red"}, _b_fill_pass)
_add("fill_color_fail_not_filled", "FILL_COLOR", {"target": "col:金額", "color": "red"}, _b_bold_fail)

# --- NUMBER_FORMAT -----------------------------------------------------------
def _b_numfmt_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 1000]]:
        ws.append(row)
    _style_range(ws, 2, 2, 2, 2, number_format="#,##0")
    wb.save(p)
    return p, None


_add("number_format_pass", "NUMBER_FORMAT", {"col": "金額", "style": "thousands"}, _b_numfmt_pass)
_add("number_format_fail_not_formatted", "NUMBER_FORMAT", {"col": "金額", "style": "thousands"},
     _b_bold_fail)

# --- MERGE --------------------------------------------------------------
def _b_merge_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    ws.merge_cells("A1:B1")
    wb.save(p)
    return p, None


_add("merge_pass", "MERGE", {"range": "A1:B1"}, _b_merge_pass)
_add("merge_fail_not_merged", "MERGE", {"range": "A1:B1"}, _b_bold_fail)

# --- CHART --------------------------------------------------------------
def _b_chart_pass(tmp_path):
    # ★ グラフ段: 本物のヘルパ(InsertChart)は項目名列(c:cat)も必ずセットする。
    #   check_chart_series の恒真殺しがそこまで見るので、fake もその形に合わせる
    #   （openpyxl の set_categories() は常に numRef を作るので、実 LO と同じ strRef を
    #   手で組む ── test_write_precondition.py の _f5_chart と同じ理由・同じ作法）。
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.data_source import AxDataSource, StrRef
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    chart.add_data(data, titles_from_data=True)
    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=3)
    for s in chart.series:
        s.cat = AxDataSource(strRef=StrRef(f=str(cat_ref)))
    ws.add_chart(chart, "D1")
    wb.save(p)
    return p, None


_add("chart_pass", "CHART", {"value_col": "金額"}, _b_chart_pass, before_charts=0)
_add("chart_fail_no_new_chart", "CHART", {"value_col": "金額"}, _b_bold_fail, before_charts=0)

# --- CENTER_ALIGN ---------------------------------------------------------
def _b_center_all_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    _style_range(ws, 1, 1, 2, 2, center=True)
    wb.save(p)
    return p, None


_add("center_align_all_pass", "CENTER_ALIGN", {"target": "all"}, _b_center_all_pass)
_add("center_align_fail_not_centered", "CENTER_ALIGN", {"target": "all"}, _b_bold_fail)

# --- APPEND_TOTAL ----------------------------------------------------------
def _b_append_total_pass_factor1(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    ws.append(["合計", "=SUM(B2:INDEX(B:B,ROW()-1))"])   # ★ 総計行はデータ行の後(row4)に足す
    wb.save(p)
    _inject_formula_cache(p, 1, {"B4": "300"})
    return p, None


_add("append_total_pass_factor1", "APPEND_TOTAL",
     {"col": "金額", "label": "合計", "factor": 1.0}, _b_append_total_pass_factor1)


def _b_append_total_pass_factor1_1(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100], ["b", 200]]:
        ws.append(row)
    ws.append(["税込合計", "=SUM(B2:INDEX(B:B,ROW()-1))*1.1"])
    wb.save(p)
    _inject_formula_cache(p, 1, {"B4": "330"})
    return p, None


_add("append_total_pass_factor1_1", "APPEND_TOTAL",
     {"col": "金額", "label": "税込合計", "factor": 1.1}, _b_append_total_pass_factor1_1)
_add("append_total_fail_zero_target_no_sum_formula", "APPEND_TOTAL",
     {"col": "金額", "label": "合計", "factor": 1.0}, _b_bold_fail)

# --- INSERT_ROWS -----------------------------------------------------------
def _b_insert_rows_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx",
                    [["商品", "金額"], ["a", 100], ["b", 200], ["c", 300]])
    after = tmp_path / "after.xlsx"
    wb = openpyxl.load_workbook(before)
    wb.save(after)
    ws = openpyxl.load_workbook(after)["Sheet"]
    # at=2(1起点)に1行挿入 → 2行目以降が1行下にシフトし、2行目が空欄になる。
    ws.insert_rows(2, amount=1)
    wb2 = ws.parent
    wb2.save(after)
    return after, before


def _b_insert_rows_warn_no_source(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append([])   # 2行目(挿入位置)が空欄
    ws.append(["a", 100])
    wb.save(p)
    return p, None


def _b_insert_rows_fail_mismatch(tmp_path):
    before = _book(tmp_path, "before.xlsx",
                    [["商品", "金額"], ["a", 100], ["b", 200]])
    after = tmp_path / "after.xlsx"
    # 挿入せず適用前と同一内容のまま保存 → シフトが一致しない。
    wb = openpyxl.load_workbook(before)
    wb.save(after)
    return after, before


_add("insert_rows_pass_with_source_book", "INSERT_ROWS", {"at": 2, "count": 1}, _b_insert_rows_pass)
_add("insert_rows_warn_no_source_book", "INSERT_ROWS", {"at": 2, "count": 1},
     _b_insert_rows_warn_no_source)
_add("insert_rows_fail_shift_mismatch", "INSERT_ROWS", {"at": 2, "count": 1},
     _b_insert_rows_fail_mismatch)

# --- ★ 2026-08-26: 表の基本操作 3 種 ----------------------------------------
_TB = [["商品", "金額"], ["りんご", 100], ["みかん", 200], ["ぶどう", 300]]


def _b_add_row_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["梨", 600],
                    ["みかん", 200], ["ぶどう", 300]])
    return after, before


def _b_add_row_overwrite(tmp_path):
    """押し下げずに上書きした（みかんが消えた）。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["梨", 600], ["ぶどう", 300]])
    return after, before


def _b_delete_rows_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["ぶどう", 300]])
    return after, before


def _b_delete_rows_left_blank(tmp_path):
    """詰めずに空行を残した。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], [None, None], ["ぶどう", 300]])
    return after, before


def _b_delete_column_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品"], ["りんご"], ["みかん"], ["ぶどう"]])
    return after, before


def _b_delete_column_took_neighbour(tmp_path):
    """隣の列を巻き込んだ（商品が消えて金額が残った）。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["金額"], [100], [200], [300]])
    return after, before


_ADD = {"at": 3, "values": {"商品": "梨", "金額": 600}}
_add("add_row_pass", "ADD_ROW", dict(_ADD), _b_add_row_pass)
_add("add_row_fail_overwrote_a_row", "ADD_ROW", dict(_ADD), _b_add_row_overwrite)
_add("delete_rows_pass", "DELETE_ROWS", {"at": 3, "count": 1}, _b_delete_rows_pass)
_add("delete_rows_fail_blank_left", "DELETE_ROWS", {"at": 3, "count": 1},
     _b_delete_rows_left_blank)
_add("delete_column_pass", "DELETE_COLUMN", {"col": "金額"}, _b_delete_column_pass)
_add("delete_column_fail_took_neighbour", "DELETE_COLUMN", {"col": "金額"},
     _b_delete_column_took_neighbour)


def _b_set_cell_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["みかん", 2000], ["ぶどう", 300]])
    return after, before


def _b_set_cell_whole_column(tmp_path):
    """★ 1 セルのはずが列全体 ── 列全体の番人なら pass してしまう形。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 2000], ["みかん", 2000], ["ぶどう", 2000]])
    return after, before


_SC = {"row": "みかん", "col": "金額", "value": "2000",
        "_write_numeric": True, "_write_numeric_value": 2000.0}
_add("set_cell_pass", "SET_CELL_VALUE", dict(_SC), _b_set_cell_pass)
_add("set_cell_fail_whole_column", "SET_CELL_VALUE", dict(_SC), _b_set_cell_whole_column)

# --- ADD_COLUMN（2026-08-27）--------------------------------------------------
def _b_add_column_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額", "備考"], ["りんご", 100], ["みかん", 200], ["ぶどう", 300]])
    return after, before


def _b_add_column_overwrote(tmp_path):
    """★ 押し出さずに既存列を潰した（列数は増えているが中身が違う）。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額", "備考"], ["りんご", 999], ["みかん", 200], ["ぶどう", 300]])
    return after, before


def _b_add_column_filled(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額", "備考"], ["りんご", 100, "?"], ["みかん", 200, "?"],
                    ["ぶどう", 300, "?"]])
    return after, before


_AC = {"name": "備考", "_at_col": 3, "_header_row": 1}
_add("add_column_pass", "ADD_COLUMN", dict(_AC), _b_add_column_pass)
_add("add_column_fail_overwrote", "ADD_COLUMN", dict(_AC), _b_add_column_overwrote)
_add("add_column_fail_filled", "ADD_COLUMN", dict(_AC), _b_add_column_filled)

# --- SWAP（2026-08-27）--------------------------------------------------------
def _b_swap_pass(tmp_path):
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["ぶどう", 300], ["みかん", 200]])
    return after, before


def _b_swap_not_done(tmp_path):
    """入れ替えを頼まれて何も動いていない（マクロが走っていない形）。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx", _TB)
    return after, before


def _b_swap_only_names(tmp_path):
    """★ 名前だけ入れ替えて金額を置いていった ── 数字の意味が入れ替わる最悪の形。"""
    before = _book(tmp_path, "before.xlsx", _TB)
    after = _book(tmp_path, "after.xlsx",
                   [["商品", "金額"], ["りんご", 100], ["ぶどう", 200], ["みかん", 300]])
    return after, before


_SW = {"a": "みかん", "b": "ぶどう", "_axis": "row", "_a_pos": 3, "_b_pos": 4}
_add("swap_pass", "SWAP", dict(_SW), _b_swap_pass)
_add("swap_fail_not_done", "SWAP", dict(_SW), _b_swap_not_done)
_add("swap_fail_only_names_moved", "SWAP", dict(_SW), _b_swap_only_names)

# --- DRAW_BORDERS -----------------------------------------------------------
def _b_borders_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    _style_range(ws, 1, 1, 2, 2, border=_ALL_BORDER)
    wb.save(p)
    return p, None


_add("draw_borders_pass", "DRAW_BORDERS", {}, _b_borders_pass)
_add("draw_borders_fail_missing", "DRAW_BORDERS", {}, _b_bold_fail)

# --- AUTOFIT -----------------------------------------------------------
def _b_autofit_pass_with_source(tmp_path):
    before = _book(tmp_path, "before.xlsx", [["商品", "金額"], ["a", 100]])
    after = tmp_path / "after.xlsx"
    wb = openpyxl.load_workbook(before)
    wb["Sheet"].column_dimensions["A"].width = 20
    wb.save(after)
    return after, before


def _b_autofit_warn_no_source(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    wb.save(p)
    return p, None


_add("autofit_pass_with_source_book", "AUTOFIT", {}, _b_autofit_pass_with_source)
_add("autofit_warn_no_source_book", "AUTOFIT", {}, _b_autofit_warn_no_source)

# --- PIVOT --------------------------------------------------------------
def _b_pivot_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["商品", "金額"], ["a", 100]]:
        ws.append(row)
    wb.create_sheet("ピボット")
    wb.save(p)
    _inject_zip_stub_member(p, "xl/pivotTables/pivotTable1.xml")
    return p, None


def _b_pivot_fail_no_sheet(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 100]]), None


_add("pivot_pass", "PIVOT", {"group_col": "商品", "value_col": "金額"}, _b_pivot_pass)
_add("pivot_fail_missing_sheet", "PIVOT", {"group_col": "商品", "value_col": "金額"},
     _b_pivot_fail_no_sheet)

# --- SET_COLUMN_VALUE --------------------------------------------------------
def _b_setcol_pass(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "備考"], ["a", "確認済み"], ["b", "確認済み"]]), None


def _b_setcol_fail(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "備考"], ["a", "確認済み"], ["b", "未確認"]]), None


_add("set_column_value_pass", "SET_COLUMN_VALUE", {"col": "備考", "value": "確認済み"}, _b_setcol_pass)
_add("set_column_value_fail_not_uniform", "SET_COLUMN_VALUE", {"col": "備考", "value": "確認済み"},
     _b_setcol_fail)

# --- EXTRACT --------------------------------------------------------------
def _b_extract_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["商品", "金額"], ["a", 30000], ["b", 50000], ["c", 45000]]:
        ws.append(row)
    out = wb.create_sheet("金額40000以上")
    out.append(["商品", "金額"])
    out.append(["b", 50000])
    out.append(["c", 45000])
    wb.save(p)
    return p, None


def _b_extract_fail_missing_sheet(tmp_path):
    return _book(tmp_path, "b.xlsx", [["商品", "金額"], ["a", 30000], ["b", 50000]]), None


_EXTRACT_ARGS = {"col": "金額", "cmp": "gte", "value": 40000.0,
                  "_target_sheet": "Sheet", "_new_sheet": "金額40000以上"}
_add("extract_pass", "EXTRACT", _EXTRACT_ARGS, _b_extract_pass)
_add("extract_fail_missing_sheet", "EXTRACT", _EXTRACT_ARGS, _b_extract_fail_missing_sheet)

# --- DEDUP（EXTRACT の兄弟）--------------------------------------------------
def _b_dedup_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in [["取引先", "金額"], ["甲社", 100], ["甲社", 250], ["乙社", 200]]:
        ws.append(row)
    out = wb.create_sheet("取引先の重複除去")
    out.append(["取引先", "金額"])
    out.append(["甲社", 100])
    out.append(["乙社", 200])
    wb.save(p)
    return p, None


def _b_dedup_fail_missing_sheet(tmp_path):
    return _book(tmp_path, "b.xlsx",
                 [["取引先", "金額"], ["甲社", 100], ["甲社", 250], ["乙社", 200]]), None


_DEDUP_ARGS = {"keys": ["取引先"], "_target_sheet": "Sheet", "_new_sheet": "取引先の重複除去"}
_add("dedup_pass", "DEDUP", _DEDUP_ARGS, _b_dedup_pass)
_add("dedup_fail_missing_sheet", "DEDUP", _DEDUP_ARGS, _b_dedup_fail_missing_sheet)

# --- REPORT_PER_ROW（帳票段）--------------------------------------------------
_REPORT_ARGS = {
    "template_sheet": "雛形", "name_col": "取引先", "_target_sheet": "売上",
    "_inspection_sheet": "検分",
    "_report_rows": [{"row": 2, "sheet": "甲社"}, {"row": 3, "sheet": "乙社"}],
    "_placeholders": [
        {"cell": "B1", "row": 1, "col": 2, "column_name": "取引先", "whole": True,
         "raw": "{{取引先}}", "col_idx": 1},
        {"cell": "B2", "row": 2, "col": 2, "column_name": "金額", "whole": True,
         "raw": "{{金額}}", "col_idx": 2},
    ],
}


def _b_report_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [["取引先", "金額"], ["甲社", 100], ["乙社", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("雛形")
    tpl["B1"] = "{{取引先}}"
    tpl["B2"] = "{{金額}}"
    sh1 = wb.copy_worksheet(tpl)
    sh1.title = "甲社"
    sh1["B1"] = "甲社"
    sh1["B2"] = 100
    sh2 = wb.copy_worksheet(tpl)
    sh2.title = "乙社"
    sh2["B1"] = "乙社"
    sh2["B2"] = 200
    insp = wb.create_sheet("検分")
    insp.append(["シート名", "元の行", "埋めた印の数"])
    insp.append(["甲社", 2, 2])
    insp.append(["乙社", 3, 2])
    wb.save(p)
    return p, None


def _b_report_fail_mismatch(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [["取引先", "金額"], ["甲社", 100], ["乙社", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("雛形")
    tpl["B1"] = "{{取引先}}"
    tpl["B2"] = "{{金額}}"
    sh1 = wb.copy_worksheet(tpl)
    sh1.title = "甲社"
    sh1["B1"] = "甲社"
    sh1["B2"] = 999   # ★ 元の行(100)と不一致
    sh2 = wb.copy_worksheet(tpl)
    sh2.title = "乙社"
    sh2["B1"] = "乙社"
    sh2["B2"] = 200
    insp = wb.create_sheet("検分")
    insp.append(["シート名", "元の行", "埋めた印の数"])
    insp.append(["甲社", 2, 2])
    insp.append(["乙社", 3, 2])
    wb.save(p)
    return p, None


_add("report_per_row_pass", "REPORT_PER_ROW", _REPORT_ARGS, _b_report_pass)
_add("report_per_row_fail_mismatch", "REPORT_PER_ROW", _REPORT_ARGS, _b_report_fail_mismatch)

# --- FORMAT_MAP（様式写像段。REPORT_PER_ROW の兄弟・縦の展開）------------------
_FORMAT_MAP_ARGS = {
    "template_sheet": "様式", "_target_sheet": "売上", "_output_sheet": "様式_出力",
    "_inspection_sheet": "検分",
    "_data_rows": [2, 3],
    "_placeholders": [
        {"cell": "A2", "row": 2, "col": 1, "column_name": "取引先", "whole": True,
         "raw": "{{取引先}}", "col_idx": 1, "out_col": 1},
        {"cell": "B2", "row": 2, "col": 2, "column_name": "金額", "whole": True,
         "raw": "{{金額}}", "col_idx": 2, "out_col": 2},
    ],
    "_header_texts": ["取引先名", "金額"],
}


def _b_format_map_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [["取引先", "金額"], ["甲社", 100], ["乙社", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "取引先名"
    tpl["B1"] = "金額"
    tpl["A2"] = "{{取引先}}"
    tpl["B2"] = "{{金額}}"
    out = wb.create_sheet("様式_出力")
    out.append(["取引先名", "金額"])
    out.append(["甲社", 100])
    out.append(["乙社", 200])
    insp = wb.create_sheet("検分")
    insp.append(["出力シート", "出力行", "元の行", "埋めた印の数"])
    insp.append(["様式_出力", 2, 2, 2])
    insp.append(["様式_出力", 3, 3, 2])
    wb.save(p)
    return p, None


def _b_format_map_fail_mismatch(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for row in [["取引先", "金額"], ["甲社", 100], ["乙社", 200]]:
        ws.append(row)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "取引先名"
    tpl["B1"] = "金額"
    tpl["A2"] = "{{取引先}}"
    tpl["B2"] = "{{金額}}"
    out = wb.create_sheet("様式_出力")
    out.append(["取引先名", "金額"])
    out.append(["甲社", 999])   # ★ 元の行(100)と不一致
    out.append(["乙社", 200])
    insp = wb.create_sheet("検分")
    insp.append(["出力シート", "出力行", "元の行", "埋めた印の数"])
    insp.append(["様式_出力", 2, 2, 2])
    insp.append(["様式_出力", 3, 3, 2])
    wb.save(p)
    return p, None


_add("format_map_pass", "FORMAT_MAP", _FORMAT_MAP_ARGS, _b_format_map_pass)
_add("format_map_fail_mismatch", "FORMAT_MAP", _FORMAT_MAP_ARGS, _b_format_map_fail_mismatch)

# --- SPLIT_CELL（繋ぎ直して元と一致するか）------------------------------------
_SPLIT_ARGS = {"col": "URL", "sep": ",", "_target_sheet": "一覧",
                "_new_cols": ["URL_1", "URL_2"]}


def _b_split_pass(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["会社", "URL", "URL_1", "URL_2"])
    ws.append(["甲社", "a,b", "a", "b"])
    ws.append(["乙社", "c", "c", ""])
    wb.save(p)
    return p, None


def _b_split_fail_lost_fragment(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["会社", "URL", "URL_1", "URL_2"])
    ws.append(["甲社", "a,b", "a", ""])   # ★ 2つ目の断片が落ちている
    wb.save(p)
    return p, None


def _b_split_fail_source_column_gone(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws.append(["会社", "URL_1", "URL_2"])   # ★ 元の列が消えている
    ws.append(["甲社", "a", "b"])
    wb.save(p)
    return p, None


_add("split_cell_pass", "SPLIT_CELL", _SPLIT_ARGS, _b_split_pass)
_add("split_cell_fail_lost_fragment", "SPLIT_CELL", _SPLIT_ARGS, _b_split_fail_lost_fragment)
_add("split_cell_fail_source_column_gone", "SPLIT_CELL", _SPLIT_ARGS,
     _b_split_fail_source_column_gone)

# --- error 状態（事後条件チェッカー自身の例外をキャッチして "error" に変換する境界） -----
_add("error_missing_required_arg_key", "SORT", {}, _b_sort_pass)   # args["col"] で KeyError


def _case_ids():
    return sorted(CASES.keys())


def test_run_postcondition_golden_coverage_matches_declared_ops():
    """★網羅性の自己検査: POSTCONDITIONS + CHART の全 op が最低1ケース入っているか。"""
    covered_ops = {op for op, *_ in CASES.values()}
    all_ops = set(ailine.POSTCONDITIONS.keys()) | {"CHART"}
    missing = all_ops - covered_ops
    assert not missing, f"run_postcondition golden が網羅していない op: {sorted(missing)}"


@pytest.mark.parametrize("name", _case_ids())
def test_run_postcondition_golden(tmp_path, name):
    op, args, header_row, use_formula, before_charts, builder = CASES[name]
    out_book, source_book = builder(tmp_path)
    status, reason = ailine.run_postcondition(
        op, out_book, dict(args), before_charts=before_charts, header_row=header_row,
        use_formula=use_formula, source_book=source_book)
    payload = {"status": status, "reason": reason}
    assert_golden_json(F3_DIR / f"{name}.json", payload, label=name)
