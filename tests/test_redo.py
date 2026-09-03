# `ailine redo` の番人（2026-09-02）── 戻しすぎた時に進めない、を塞ぐ。
#
# ★★ README の「既知の問題」に自分で書いていた欠け: undo に redo が無い。
# ★ やり直しの材料は**既に積まれていた** ── undo は毎回「戻す前の中身」を棚
#   （undo_shelf_dir）へ退避している（`_shelve_bytes`・「undo 自体も可逆」）。
#   足したのは**取り出す口**だけで、新しい保管場所も新しい概念も作っていない。
#
# 契約:
#   ① undo → redo で元に戻る（往復して同じ中身になる）
#   ② redo は棚を 1 つ下ろす（同じ状態を何度もやり直せない＝残り回数が嘘にならない）
#   ③ やり直せるものが無ければ、非零で止まり**原本に触らない**
#   ④ redo のあと undo すると、また戻れる（対称）
#   ⑤ 壊れた退避を原本に被せない（undo と同じ物差し）
#   ⑥ 原本を書き換えるので、run と同じ実行ロックを取る

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402
from _product_source import product_text  # noqa: E402 ── ★ 番人は本体決め打ちでなく製品コード全体を読む


def _book(tmp_path: Path, first: str) -> Path:
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "売上"])
    ws.append([first, 100])
    wb.save(p)
    return p


def _first(p: Path):
    return openpyxl.load_workbook(p).active.cell(row=2, column=1).value


def _edit(p: Path, value: str):
    """人の編集を模して、バックアップを取ってから書き換える（run と同じ順）。"""
    ailine.make_backup(p)
    wb = openpyxl.load_workbook(p)
    wb.active.cell(row=2, column=1, value=value)
    wb.save(p)


def test_undo_then_redo_comes_back(tmp_path, monkeypatch):
    """① 往復して同じ中身になる。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    assert _first(p) == "みかん"
    ailine.restore_backup(p)
    assert _first(p) == "りんご", "undo が効いていない"
    ailine.redo_last_undo(p)
    assert _first(p) == "みかん", "redo で戻っていない"


def test_redo_takes_one_off_the_shelf(tmp_path, monkeypatch):
    """② 棚を 1 つ下ろす ── 外さないと同じ状態を何度もやり直せてしまう。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)
    assert len(ailine.list_undo_shelf(p)) == 1
    ailine.redo_last_undo(p)
    assert ailine.list_undo_shelf(p) == [], "棚に残っている（残り回数が嘘になる）"


def test_nothing_to_redo_does_not_touch_the_book(tmp_path, monkeypatch):
    """③ やり直せるものが無ければ止まる。★ 原本に触らないことまで見る。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    before = p.read_bytes()
    with pytest.raises(ailine.NothingToRedoError):
        ailine.redo_last_undo(p)
    assert p.read_bytes() == before, "原本に触っている"


def test_undo_after_redo_goes_back_again(tmp_path, monkeypatch):
    """④ 対称 ── redo のあとの undo が、また 1 段戻る。

    ★ ここが崩れると「進んだら戻れない」という逆の行き止まりを作る。
    ★★ 注意（2026-09-02 の変異試験）: この検体は**対称そのもの**を見ているが、
      `redo` が印を消す 1 行までは縛れていない ── 印が指す世代と中身が違えば
      `_undo_position` が無効と見るので、消しても消さなくても同じ場所へ行く。
      効くことを示す検体を作れなかったので、**番人は付いていない**と明記しておく
      （「守られているつもり」を残さない）。
    """
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)
    ailine.redo_last_undo(p)
    assert _first(p) == "みかん"
    ailine.restore_backup(p)
    assert _first(p) == "りんご", "redo のあと undo で戻れない"


def test_an_edit_made_after_the_undo_is_not_lost(tmp_path, monkeypatch):
    """★★ 自作 review が見つけた**致命**（2026-09-02・敵対検証 2 レンズとも再現）。

      undo のあとに**別の編集を挟んでから** redo すると、その編集内容が
      **警告なしに完全消失**していた ── backups にも undo 棚にも残らず、
      画面には「✓ やり直しました」しか出ない。
    ★ 姉妹関数 `restore_backup` は上書き前に現在の中身を退避している ── **片配線**。
    ★ 直しは「棚に積む」ではなく**世代に積む**（make_backup）。棚に積むと redo が
      自分の直前状態を拾い直して往復が壊れる。世代なら `ailine undo` で取り戻せる。
    ★ 元の検体（6 本）は「undo の直後にすぐ redo」しか見ていなかった ──
      **間に何かを挟む**という形が 1 件も無かった。
    """
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)                    # undo → りんご（棚に みかん）
    _edit(p, "だいじな編集")                     # ★ ここが消えていた
    assert _first(p) == "だいじな編集"
    ailine.redo_last_undo(p)
    assert _first(p) == "みかん", "redo が効いていない"
    # ★ 消えていないこと ── undo で取り戻せる
    ailine.restore_backup(p)
    assert _first(p) == "だいじな編集", "redo が、あとから足した編集を消した"


def test_redo_is_stopped_by_the_lock_gate(tmp_path, monkeypatch, capsys):
    """★★ 自作 review が見つけた**致命**（2026-09-02）: redo が
      ロックの関所（refuse_if_locked）を**一度も通っていなかった**。

    ★ この repo は「run は Excel ロックで止まるのに undo は素通り」（復元の致命5）を
      既に踏み、番人を「1 本で 4 経路を縛る」形にしていた。
      **5 本目の経路を作って配線しなかった** ── 在っても鳴らない、そのもの。
    """
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)
    (p.parent / f"~${p.name}").write_bytes(b"excel lock")   # Excel が開いている印
    before = p.read_bytes()
    rc, out = _run_main(["redo", str(p)], capsys)
    assert rc != 0, out
    assert p.read_bytes() == before, "ロックを無視して上書きしている"


def test_every_write_path_goes_through_the_lock_gate():
    """★ 経路を数え上げる ── 原本を書き換える口が、全部この関所を通ること。

    ★ 「1 本で N 経路を縛る」と docstring が宣言している以上、**N を数える側**が要る。
      新しい口（今回の redo）を足した時に、配線忘れがここで赤くなる。
    """
    for fn in ("_cmd_undo_body", "cmd_redo"):
        i = product_text().index(f"def {fn}(")
        j = product_text().index(chr(10) + "def ", i + 10)
        assert "refuse_if_locked" in product_text()[i:j], f"{fn} がロックの関所を通っていない"


def test_a_broken_shelf_item_is_refused(tmp_path, monkeypatch):
    """⑤ 壊れた退避を原本に被せない（undo と同じ物差し）。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)
    shelved = ailine.list_undo_shelf(p)
    assert shelved
    shelved[0].write_bytes(b"not an xlsx")
    before = p.read_bytes()
    with pytest.raises(ailine.BrokenBackupError):
        ailine.redo_last_undo(p)
    assert p.read_bytes() == before, "壊れた退避を被せている"


def test_the_command_takes_the_run_lock():
    """⑥ 原本を書き換えるので run と同じロックを取る（undo と同じ作法）。"""
    i = product_text().index("def cmd_redo(")
    j = product_text().index("\ndef ", i + 10)
    assert "under_run_lock" in product_text()[i:j], "redo が実行ロックを取っていない"


def test_the_cli_exposes_redo(tmp_path, monkeypatch, capsys):
    """★ 配線 ── `ailine redo` が本当に叩けること（在るのに呼べない、を作らない）。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    _edit(p, "みかん")
    ailine.restore_backup(p)
    rc, out = _run_main(["redo", str(p)], capsys)
    assert rc == 0, out
    assert _first(p) == "みかん", out
    assert "やり直しました" in out, out


def test_the_cli_says_how_to_go_back_when_there_is_nothing(tmp_path, monkeypatch, capsys):
    """★ 断りを行き止まりにしない（この repo の作法）。"""
    _isolate(monkeypatch, tmp_path)
    p = _book(tmp_path, "りんご")
    rc, out = _run_main(["redo", str(p)], capsys)
    assert rc != 0
    assert "ailine undo" in out, out
