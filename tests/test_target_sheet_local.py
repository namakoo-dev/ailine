"""★ 挙動変更#2 実機E2E: 対象シートの固定を解いた fix を、本物の basrun/LibreOffice
   経由で通す通し試験（DoD5）。

   ① 複数シートのブックで2枚目シートを対象にした依頼が、本物の LibreOffice 上でも
      正しく2枚目に書き込まれること（_wrap_basic_for_sheet の moveByName が実際に効くこと・
      ailine.py のコメント/openpyxl 読み戻しだけでは検証できない「本当に LO 上でシートの
      並び替えが安全に戻るか」を見る）。
   ② 単一シートのブックは従来どおり動くこと（退行が無いこと）。

   translate_task は monkeypatch で固定する（DSL 経路は率/列名を LLM に確定させない
   A' 原則どおり・ollama 自体の応答ゆらぎはこのテストの関心事ではない。実際に ollama を
   通す確認は手動実行で別途行った — 変更差分の報告参照）。basrun/LibreOffice は本物を使う。

   ★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
   （`pytest -m "not local"` で除外される。手動は `pytest -m local` または
   `pytest tests/test_target_sheet_local.py`）。
   ★ tests/test_bold_local.py と同じ後始末方針（taskkill 名前一括はしない・
   ailine._stop_office() に委譲）。
"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))
import ailine  # noqa: E402

from _run_argv import run_argv  # noqa: E402  — C2: cmd_run 直呼び用 Namespace → main(argv) 変換


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1
    return out.lower().count("soffice.bin")


def _multi_sheet_book(tmp_path):
    p = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "請求書"
    ws1.append(["宛先", "金額"])
    ws1.append(["山田商店", 50000])
    ws1.append(["鈴木工務店", 80000])
    ws2 = wb.create_sheet("工事台帳")
    ws2.append(["取引先名", "工事名", "金額"])
    ws2.append(["山田商店", "A邸新築", 120000])   # ★ わざと降順になっていない並びにする
    ws2.append(["鈴木工務店", "B邸改修", 300000])  # （SORT が本当に効いたかを diff で見るため）
    ws2.append(["山田商店", "C邸外構", 150000])
    ws3 = wb.create_sheet("取引先マスタ")
    ws3.append(["取引先名", "住所"])
    ws3.append(["山田商店", "東京都"])
    wb.save(p)
    return p


@pytest.mark.local
def test_e2e_multi_sheet_book_writes_to_named_second_sheet_via_real_basrun(
        tmp_path, monkeypatch, capsys):
    """① DoD5: 「工事台帳シートで金額を降順に並べ替えて」を本物の basrun/LibreOffice で
       通し、①1枚目(請求書)が一切変更されない ②対象(工事台帳)が実際に降順になる
       ③保存後もシート順が元のまま(請求書/工事台帳/取引先マスタ)であることを確認する
       （_wrap_basic_for_sheet の moveByName が実行後に必ず元へ戻ることの実機証拠）。"""
    before_procs = _soffice_process_count()
    try:
        book = _multi_sheet_book(tmp_path)
        monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
        monkeypatch.setattr(ailine, "translate_task",
                            lambda model, task, book_meta, temperature=0.1:
                            {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
        argv = run_argv(
            book=str(book), task="工事台帳シートで金額を降順に並べ替えて",
            dry=False, copy=True, timeout=180.0)
        rc = ailine.main(argv)
        captured = capsys.readouterr()
        assert rc == 0, captured.out
        assert "操作するシート: 2枚目『工事台帳』" in captured.out

        out_book = book.with_name(book.stem + ".out" + book.suffix)
        wb2 = openpyxl.load_workbook(out_book)
        # ★ シート順が元のまま保たれている（moveByName の戻しが本物の LO でも効いている証拠）。
        assert wb2.sheetnames == ["請求書", "工事台帳", "取引先マスタ"]
        # 1枚目は一切変更されていない。
        ws1 = wb2["請求書"]
        assert [c.value for c in ws1[1]] == ["宛先", "金額"]
        assert [c.value for c in ws1[2]] == ["山田商店", 50000]
        assert [c.value for c in ws1[3]] == ["鈴木工務店", 80000]
        # 対象(工事台帳)は金額の降順になっている。
        ws2 = wb2["工事台帳"]
        amounts = [ws2.cell(row=r, column=3).value for r in range(2, 5)]
        assert amounts == sorted(amounts, reverse=True), amounts
        wb2.close()
    finally:
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )


@pytest.mark.local
def test_e2e_single_sheet_book_still_works_unchanged_no_regression(tmp_path, monkeypatch, capsys):
    """② DoD5: 単一シートのブックは従来どおり本物の LibreOffice 経由で正しく動く
       （退行が無いこと・対象シート明示の行は出ない）。"""
    before_procs = _soffice_process_count()
    try:
        book = tmp_path / "single.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet"
        ws.append(["商品", "数量", "単価"])
        ws.append(["りんご", 3, 100])
        ws.append(["みかん", 5, 50])
        wb.save(book)

        monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
        monkeypatch.setattr(ailine, "translate_task",
                            lambda model, task, book_meta, temperature=0.1:
                            {"op": "SORT", "args": {"col": "数量", "order": "desc"}})
        argv = run_argv(
            book=str(book), task="数量で降順に並べ替えて", dry=False, copy=True, timeout=180.0)
        rc = ailine.main(argv)
        captured = capsys.readouterr()
        assert rc == 0, captured.out
        assert "操作するシート:" not in captured.out   # 単一シートは沈黙（既存挙動を変えない）

        out_book = book.with_name(book.stem + ".out" + book.suffix)
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        qty = [ws2.cell(row=r, column=2).value for r in range(2, 4)]
        assert qty == [5, 3]   # 降順
        wb2.close()
    finally:
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )
