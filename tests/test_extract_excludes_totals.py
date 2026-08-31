# 抜き出しは合計行を持っていかない／行番号では抜き出せない ── 2026-08-31。
# Namakoo「この操作をするとヤマノ食品が両方とも抜き出される。5行目の の指示が無視される」
#
# ★★ ここから 3 つの欠陥が出た。
#
# ① 行番号での限定が無視されていた
#   「**5行目の**ヤマノ食品を抜き出して」→ ヤマノ食品が 2 行とも抜き出された。
#   三項の番人は鳴っていた（「依頼文が指しているのは: 5行目」）が、**⚠ を出して進んで**
#   いた ── 何が無視されたかを言わないので、利用者には「効かなかった」に見える。
#   ★ そもそも**行番号での抜き出しは語彙に無い**（EXTRACT は列×比較×値だけ）。
#     できないことは ⚠ でなく**断る**。
#
# ② 合計行まで抜き出していた
#   「金額が60000以上の行を抜き出して」→ 合計 356400 も条件に合うので抜き出された。
#   ★ 並べ替え・条件つき書換では既に外していたのに、**抽出だけ外していなかった**。
#   ★ 直す途中で 3 回つまずいた:
#     ・判定を**早い出口より後ろ**に置いて素通りした（この分岐には出口が 2 つある）
#     ・Basic のループは **0 起点**なのに 1 起点の行番号を渡して 1 行ずれた
#     ・**分母を縮め忘れ**て「行数が期待と不一致」になった（並べ替えで 2 度踏んだ形）
#
# ③ 機械が LLM の揺れを増幅していた（Namakoo「LLM の揺れが一番厄介だ」）
#   同じ依頼で 3 回中 2 回が「行追加」に化けていた。追うと、機械が『60000』を
#   **行の名前**として解いていた（金額列に 60000 が在るため）。
#   ★ 揺れは消せないが**増幅は消せる** ── 依頼文に出る数は閾値や個数で、行の名前ではない。
#   ★ 判定は既にある `_is_number_like`（「依頼文に出る数と、行の名前を混同しないため」）
#     を借りた ── 1 箇所でしか使われておらず、**合流点に配られていなかった**。
#   実測: 直したら 4/4 で「抽出」に安定した。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HEADERS = ["取引先", "項目", "金額"]
ROWS = [["丸和物流", "配送", 57600], ["近江スチール", "鋼材", 60000],
        ["ヤマノ食品", "食品", 42000], ["ヤマノ食品", "冷蔵", 18000],
        ["合計", None, 177600]]


@pytest.fixture()
def meta(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(p)
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(p)}


# --- ① 行番号での抜き出しは断る ------------------------------------------------------------

def test_a_row_number_is_refused(meta):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "取引先", "cmp": "eq", "value": "ヤマノ食品"}, meta,
        task="5行目のヤマノ食品を抜き出して")
    assert not ok
    assert "行番号での抜き出し" in err and "5行目" in err, err


def test_a_normal_extract_still_passes(meta):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "取引先", "cmp": "eq", "value": "ヤマノ食品"}, meta,
        task="ヤマノ食品を抜き出して")
    assert ok, err


# --- ② 合計行を外す（宣言・生成・検算の 3 つに同じ範囲が届くこと）--------------------------

def test_the_total_row_is_declared_as_skipped(meta):
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "金額", "cmp": "gte", "value": 60000}, meta,
        task="金額が60000以上の行を抜き出して")
    assert r.get("_skip_rows") == [6], r.get("_skip_rows")
    assert "合計行" in r.get("_skip_label", "")


def test_the_codegen_passes_zero_based_rows(meta):
    """★★ Basic のループは 0 起点 ── 1 起点で渡して 1 行ずれた（実測）。"""
    _ok, r, _i, _err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "金額", "cmp": "gte", "value": 60000}, meta,
        task="金額が60000以上の行を抜き出して")
    code = ailine.codegen_dsl("EXTRACT", r, meta)
    assert '"5")' in code, code          # 6 行目 → Basic では 5


def test_the_check_shrinks_the_denominator_too(tmp_path):
    """★★ 片側だけ縮めると「行数が期待と不一致」になる（並べ替えで 2 度踏んだ形）。
       宣言（_skip_rows）を同じ 1 箇所から生成にも検算にも配る。"""
    src = tmp_path / "s.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(src)
    out = tmp_path / "o.xlsx"
    wb2 = openpyxl.Workbook()
    s1 = wb2.active
    s1.title = "請求"
    s1.append(HEADERS)
    for r in ROWS:
        s1.append(r)
    s2 = wb2.create_sheet("金額60000以上")
    s2.append(HEADERS)
    s2.append(["近江スチール", "鋼材", 60000])
    wb2.save(out)
    args = {"col": "金額", "cmp": "gte", "value": 60000, "_target_sheet": "請求",
            "_new_sheet": "金額60000以上", "_skip_rows": [6],
            "_source_headers": tuple(HEADERS)}
    st, why = ailine.check_extract(out, args)
    assert st == "pass", why
    assert "4行中1行" in why, why        # ★ 合計行は分母にも入らない


def test_the_disclosure_is_on_the_interpretation_line():
    """★ 外したことを画面に出す（08-29 に SORT で同じ形を踏んだ）。"""
    keys = {k for _l, k, _f in ailine._CONFIRM_FIELDS["EXTRACT"]}
    assert "_skip_label" in keys


# --- ③ 数を行の名前にしない（揺れを増幅しない）--------------------------------------------

def test_a_threshold_number_is_not_a_row_name(meta):
    """★★ 「金額が60000以上の…」で『60000』を行の名前として解いていた。"""
    rows, heads = ailine._table_rows_for_anchor(meta, "請求", 1)
    assert ailine._row_named_anywhere_in_task(
        "金額が60000以上の行を抜き出して", rows, heads) is None


def test_a_real_name_is_still_found(meta):
    """★ 黙りすぎていないこと。"""
    rows, heads = ailine._table_rows_for_anchor(meta, "請求", 1)
    got = ailine._row_named_anywhere_in_task("丸和物流を削除して", rows, heads)
    assert got and got[0] == 2, got


def test_the_numeric_guard_is_at_the_junction():
    """★ 4 つの呼び出しの合流点に置く（呼び出し側に配らない）。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index("def _row_named_anywhere_in_task(")
    seg = src[i:i + 2600]
    assert "_is_number_like(v)" in seg, "合流点で数を弾いていない"


# --- ④ 実物で（LibreOffice を通す）--------------------------------------------------------

@pytest.mark.local
def test_the_total_row_is_not_extracted_for_real(tmp_path):
    import subprocess
    p = tmp_path / "r.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(p)
    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}
    r = subprocess.run(
        [sys.executable, "-m", "ailine", "run", str(p),
         "金額が60000以上の行を抜き出して", "--copy", "--sheet", "請求"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=900, cwd=str(REPO), env=env)
    assert "抽出" in r.stdout, r.stdout[-1200:]
    wb2 = openpyxl.load_workbook(tmp_path / "r.out.xlsx", data_only=True)
    out = [s for s in wb2.sheetnames if "金額" in s]
    assert out, wb2.sheetnames
    vals = [row[0] for row in wb2[out[-1]].iter_rows(values_only=True)]
    assert "合計" not in vals, vals
