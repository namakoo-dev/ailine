# 表の基本操作 3 種（2026-08-26）── ADD_ROW / DELETE_ROWS / DELETE_COLUMN。
#
# ★ きっかけは Namakoo が GUI を触った実測:
#     「5行目に商品として梨を追加して。売上は600　原価は300」
#   → 行挿入 + 一括書換×3 に分解され、4 段とも別々の理由で落ちた。
#   21 op のどれにも「データを 1 行足す」「行/列を消す」が無かった。
#
# ★ Namakoo の指摘（設計の芯）:
#   「最後尾に追加するだけとは限らない。途中に行を追加する必要もあるし、
#     削除した場合はそこを詰めないといけない」
#   → insertByIndex / removeByIndex を使う（clearContents だと空行が残る）。
#     そして**そう書いたから正しい、では済まない** ── 事後条件で証明する。
#
# 契約:
#   ① 追加: 行数 +1・その行が宣言どおり・**at より下が 1 行ずれてそのまま**（押し下げた証拠）
#   ② 追加: at より上が 1 セルも変わらない
#   ③ 削除: 残りが「消した分を抜いた並び」と**順序ごと連続で**一致（詰めた証拠）
#   ④ 列削除: 他の列が 1 セルも変わらない
#   ⑤ 消した中身を機械の値として持ち、画面に出す（削除は差分に何も出ない操作）
#   ⑥ 実在しない列名・見出し行より上は断る（幻覚と骨格破壊の封鎖）
#   ⑦ 数値は数値のまま入る（文字列化すると下流の SUM が静かに壊れる）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402

META = {"sheets": ["売上"], "headers": {"売上": ["商品", "売上", "原価"]},
         "header_rows": {"売上": 1}}
ROWS = [["商品", "売上", "原価"], ["りんご", 1200, 700],
         ["みかん", 800, 300], ["ぶどう", 1500, 900]]


def _book(tmp_path, rows=None, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in (rows or ROWS):
        ws.append(r)
    wb.save(p)
    return p


def _grid(p):
    ws = openpyxl.load_workbook(p)["売上"]
    return [[c.value for c in row] for row in ws.iter_rows()]


# --- ⑥ 関所 ---------------------------------------------------------------------------

@pytest.mark.parametrize("op,args,expect", [
    ("ADD_ROW", {"at": 5, "values": {"存在しない列": 1}}, "がこの表にありません"),
    ("ADD_ROW", {"at": 1, "values": {"商品": "梨"}}, "見出し行"),
    ("ADD_ROW", {"at": 5, "values": {}}, "読み取れません"),
    ("DELETE_ROWS", {"at": 1}, "見出し行"),
    ("DELETE_COLUMN", {"col": "利益"}, "がこの表にありません"),
])
def test_refusals(op, args, expect):
    ok, _resolved, _inf, err = ailine.verify_dsl_args(op, dict(args), META, task="t")
    assert not ok, f"{op} {args} を通した"
    assert expect in err, err


def test_a_real_request_passes():
    """誤爆防止: 実在する列・見出しより下なら通る。"""
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 3, "values": {"商品": "梨", "売上": 600}}, META, task="t")
    assert ok, err
    assert resolved["_values_label"] == "商品=梨／売上=600"


# --- ①②⑦ 追加（途中に挿す）------------------------------------------------------------

def test_add_row_in_the_middle_pushes_the_rest_down(tmp_path):
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["りんご", 1200, 700],
                              ["梨", 600, 300],           # ← 3 行目に挿さった
                              ["みかん", 800, 300], ["ぶどう", 1500, 900]])
    args = {"at": 3, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, reason = ailine.check_add_row(after, args, source_book=before)
    assert status == "pass", reason
    assert "元のまま" in reason


def test_add_row_catches_an_overwrite(tmp_path):
    """★ 恒真殺し: 押し下げずに**上書き**したら落ちること。"""
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["りんご", 1200, 700],
                              ["梨", 600, 300],           # みかんを潰した
                              ["ぶどう", 1500, 900]])
    args = {"at": 3, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, reason = ailine.check_add_row(after, args, source_book=before)
    assert status == "fail", reason


def test_add_row_catches_a_wrong_value(tmp_path):
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["りんご", 1200, 700],
                              ["梨", 999, 300],
                              ["みかん", 800, 300], ["ぶどう", 1500, 900]])
    args = {"at": 3, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, reason = ailine.check_add_row(after, args, source_book=before)
    assert status == "fail" and "売上" in reason, reason


def test_added_numbers_stay_numbers(tmp_path):
    """⑦ 数値が文字列で入ると下流の SUM が静かに壊れる。"""
    after = _book(tmp_path, [["商品", "売上", "原価"], ["梨", 600, 300]])
    args = {"at": 2, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, _ = ailine.check_add_row(after, args, source_book=None)
    assert status == "warn"          # before 無しなら断定しない
    after2 = _book(tmp_path, [["商品", "売上", "原価"], ["梨", "600", 300]], name="s.xlsx")
    status2, reason2 = ailine.check_add_row(after2, args, source_book=None)
    assert status2 == "fail", f"文字列の 600 を通した: {reason2}"


# --- ③⑤ 行削除（詰める）----------------------------------------------------------------

def test_delete_rows_closes_the_gap(tmp_path):
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["りんご", 1200, 700],
                              ["ぶどう", 1500, 900]])       # みかんが消えて詰まった
    args = {"at": 3, "count": 1}
    status, reason = ailine.check_delete_rows(after, args, source_book=before)
    assert status == "pass", reason
    assert args["_deleted"] == [["みかん", 800, 300]], args.get("_deleted")


def test_delete_rows_catches_a_blank_left_behind(tmp_path):
    """★ 恒真殺し: 詰めずに空行を残したら落ちること（clearContents 実装への番人）。"""
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["りんご", 1200, 700],
                              [None, None, None], ["ぶどう", 1500, 900]])
    status, reason = ailine.check_delete_rows(after, {"at": 3, "count": 1}, source_book=before)
    assert status == "fail", reason


def test_delete_rows_catches_deleting_the_wrong_row(tmp_path):
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価"], ["みかん", 800, 300],
                              ["ぶどう", 1500, 900]])       # りんごを消した
    status, reason = ailine.check_delete_rows(after, {"at": 3, "count": 1}, source_book=before)
    assert status == "fail", reason


# --- ④⑤ 列削除 -------------------------------------------------------------------------

def test_delete_column_leaves_the_others_untouched(tmp_path):
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上"], ["りんご", 1200],
                              ["みかん", 800], ["ぶどう", 1500]])
    args = {"col": "原価"}
    status, reason = ailine.check_delete_column(after, args, source_book=before)
    assert status == "pass", reason
    assert args["_deleted"] == [[700], [300], [900]], args.get("_deleted")


def test_delete_column_catches_taking_a_neighbour(tmp_path):
    """★ 恒真殺し: 隣の列を巻き込んだら落ちること。"""
    before = _book(tmp_path, name="before.xlsx")
    after = _book(tmp_path, [["商品", "原価"], ["りんご", 700],
                              ["みかん", 300], ["ぶどう", 900]])   # 売上を消してしまった
    status, reason = ailine.check_delete_column(after, {"col": "原価"}, source_book=before)
    assert status == "fail", reason


# --- ⑤ 消した中身が画面に出る -----------------------------------------------------------

def test_deleted_content_reaches_the_report():
    from ailine_core import dsl_step
    # 本番の合流点が組み立てる助言に、消した中身が載ること
    src = Path(REPO / "src" / "ailine_core" / "dsl_step.py").read_text(encoding="utf-8")
    assert '"_deleted"' in src, "本番の合流点が消した中身を読んでいない"
    assert "戻すなら ailine undo" in src, "取り返しがつくことを言っていない"


# --- ⑧ 位置は相対で言われる（Namakoo が実測）--------------------------------------------
#
# 「みかんの下に梨を追加して」「みかんとぶどうの間に梨を追加して」がどちらも動かなかった。
# ★ 根: ADD_ROW は位置を**行番号**でしか受け取れないのに、人は相対で言う。
#   LLM に数えさせると外し、空行だけの INSERT_ROWS に落ちていた。
# ★ 分担を変えた: LLM は「誰の隣か」を言うだけ／**行番号は機械が実表を数えて決める**
#   （列名の解決を機械 3 段でやっているのと同じ形）。

def _anchor_meta(tmp_path):
    p = _book(tmp_path, name="anchor.xlsx")
    return p, {"sheets": ["売上"], "headers": {"売上": ["商品", "売上", "原価"]},
                "header_rows": {"売上": 1}, "path": str(p)}


@pytest.mark.parametrize("task,expect_at", [
    ("みかんの下に梨を追加して", 4),
    ("みかんの後に梨を追加して", 4),
    ("みかんとぶどうの間に梨を追加して", 4),
    ("りんごの上に梨を追加して", 2),
    ("ぶどうの前に梨を追加して", 4),
])
def test_relative_positions_resolve_to_a_row_number(tmp_path, task, expect_at):
    _p, meta = _anchor_meta(tmp_path)
    at, note = ailine.resolve_row_anchor(task, meta, "売上")
    assert at == expect_at, f"{task} → {at} / {note}"
    assert note and "行目" in note, note


def test_an_unknown_anchor_is_refused_by_name(tmp_path):
    """★ 推測で行を挿さない（静かに別の場所へ入るのが一番こわい）。"""
    _p, meta = _anchor_meta(tmp_path)
    at, note = ailine.resolve_row_anchor("すいかの下に梨を追加して", meta, "売上")
    assert at is None and "すいか" in note, note


def test_an_ambiguous_anchor_is_refused(tmp_path):
    """同じ名前が 2 行あるなら決めない。"""
    p = _book(tmp_path, [["商品", "売上", "原価"], ["みかん", 1, 1], ["みかん", 2, 2]],
               name="dup.xlsx")
    meta = {"sheets": ["売上"], "headers": {"売上": ["商品", "売上", "原価"]},
             "header_rows": {"売上": 1}, "path": str(p)}
    at, note = ailine.resolve_row_anchor("みかんの下に足して", meta, "売上")
    assert at is None and "2 行" in note, note


def test_the_machine_position_beats_the_model(tmp_path):
    """★ 実表を見た側が正しい ── LLM が別の行番号を言っても機械が上書きし、根拠を残す。"""
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 99, "values": {"商品": "梨"}}, meta,
        task="みかんの下に梨を追加して")
    assert ok, err
    assert resolved["at"] == 4, resolved
    assert "みかん" in resolved["_at_basis"], resolved


def test_positional_values_are_named_by_the_machine(tmp_path):
    """★ LLM は values を**並び**で返すことがある。列名は機械が付ける（決めた対応は出す）。"""
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 3, "values": ["梨", 600, 300]}, meta, task="3行目に足して")
    assert ok, err
    assert resolved["values"] == {"商品": "梨", "売上": 600, "原価": 300}, resolved
    assert resolved["_values_label"] == "商品=梨／売上=600／原価=300"


def test_too_many_positional_values_are_refused(tmp_path):
    _p, meta = _anchor_meta(tmp_path)
    ok, _r, _i, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 3, "values": [1, 2, 3, 4]}, meta, task="3行目に足して")
    assert not ok and "4 個" in err, err


def test_insert_rows_with_values_in_the_task_is_re_read():
    """★ 「値の指定が在る」という証拠がある時だけ、行追加として読み直す（黙って変えない）。"""
    assert ailine.insert_rows_should_have_been_add_row("みかんの下に梨を追加して。売上は600", {})
    assert ailine.insert_rows_should_have_been_add_row("3行目に1行挿入して", {}) is None


def test_empty_values_are_not_written_as_the_word_none(tmp_path):
    """★ 2026-08-27（実測・俺が入れた壊し方）: LLM が埋まらない列を None で返すと、
       codegen が `str(None)` を書き、セルに**文字列 "None"** が入った。

    ★ 事後条件はこの壊れ方を捕まえていた（宣言と実物が食い違うので rc=1）── 番人は
      効いていたが、**壊れた物を作ってから**気づく形だったので、入口で落とす。
    ★ 指定の無い列には何も書かない（空欄のまま）。
    """
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 3, "values": ["梨", None, None]}, meta, task="3行目に足して")
    assert ok, err
    assert resolved["values"] == {"商品": "梨"}, resolved["values"]
    assert "None" not in resolved["_values_label"], resolved["_values_label"]


def test_named_empty_values_are_dropped_too(tmp_path):
    """名前つきで来た時も同じ（片配線を作らない）。"""
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "ADD_ROW", {"at": 3, "values": {"商品": "梨", "売上": None, "原価": ""}},
        meta, task="3行目に足して")
    assert ok, err
    assert resolved["values"] == {"商品": "梨"}, resolved["values"]


@pytest.mark.parametrize("task,fires", [
    ("みかんとぶどうの間に梨を追加して", True),      # 数字が無くても場所で分かる
    ("みかんの下に梨を追加して。売上は600", True),   # 値の指定がある
    ("3行目に1行挿入して", False),                  # 空行が欲しい
    ("みかんの下に空行を入れて", False),            # 明示的に空行
    ("行間を空けて", False),
])
def test_when_insert_rows_should_be_re_read(tmp_path, task, fires):
    """★ 証拠がある時だけ読み直す（黙って op を書き換えない・誤爆させない）。

    ★ 2 度直した: 初版は「値の代入がある」だけを証拠にしていたので、
      「みかんとぶどうの間に梨を追加して」（数字が無い）で発火しなかった。
      **相対位置が実表で解けること自体**を証拠に加えた ── 空行を「みかんとぶどうの
      間に」挿してくれ、という依頼は考えにくい。
    """
    _p, meta = _anchor_meta(tmp_path)
    got = ailine.insert_rows_should_have_been_add_row(task, {}, meta, "売上")
    assert bool(got) is fires, f"{task} → {got}"


# --- ⑨ 空行の挿入と record の追加を、言い回しで取り違えない ------------------------------
#
# ★ Namakoo「少なくとも空白行の追加は出来ないといけない」＋「揺れ無しで追加するには？」
# ★ 実測で踏んだ俺の誤爆: 「みかんとぶどうの間に1行足して」を record の追加と誤解し、
#   `商品=みかんとぶどう` という**値をでっち上げた**。
#   → 足そうとしているのが「**行**」そのものなら、それは空行の挿入（読み直さない）。

@pytest.mark.parametrize("task,should_reread", [
    ("みかんとぶどうの間に1行足して", False),   # 足すのは「行」
    ("みかんの下に1行挿入して", False),
    ("みかんの下に空行を入れて", False),
    ("みかんの下に空白行を追加して", False),
    ("3行目に1行挿入して", False),
    ("みかんの下に梨を追加して", True),         # 足すのは record
    ("みかんとぶどうの間に梨を追加して", True),
])
def test_row_versus_record(tmp_path, task, should_reread):
    _p, meta = _anchor_meta(tmp_path)
    got = ailine.insert_rows_should_have_been_add_row(task, {}, meta, "売上")
    assert bool(got) is should_reread, f"{task} → {got}"


def test_blank_row_insertion_also_uses_the_machine_position(tmp_path):
    """★ 位置は op に関係なく位置 ── INSERT_ROWS にも同じ解決を通す（片配線を作らない）。

    実測: 「みかんの下に空行を入れて」で LLM が 3 行目と言った（みかんが 3 行目なので、
    下は 4 行目）。
    """
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "INSERT_ROWS", {"at": 3, "count": 1}, meta, task="みかんの下に空行を入れて")
    assert ok, err
    assert resolved["at"] == 4, resolved
    assert "みかん" in resolved.get("_at_basis", ""), resolved


def test_plain_row_numbers_are_left_alone(tmp_path):
    """誤爆防止: 相対の言い回しが無ければ、LLM の行番号をそのまま使う。"""
    _p, meta = _anchor_meta(tmp_path)
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "INSERT_ROWS", {"at": 3, "count": 1}, meta, task="3行目に1行挿入して")
    assert ok, err
    assert resolved["at"] == 3 and "_at_basis" not in resolved, resolved


# --- ⑩ 表の途中に空行があっても壊れない（Namakoo が実測・今日 3 箇所目）-------------------
#
# ★ `_scan_last_row` は 1 列目を上から見て**最初の空で止まる**。下書きに空行が 1 本
#   あるだけで、その下が全部消えていた:
#     ・位置の解決が「『みかん』が見つかりません」になり、黙って LLM の行番号へ落ちた
#     ・事後条件が「検証対象が 0 件」になった
# ★ 3 箇所で同じ穴を開けたので、**器官を 1 つにした**（data_extent）。

def _gappy(tmp_path):
    p = _book(tmp_path, [["商品", "売上", "原価"], [None, None, None],
                          ["りんご", 1200, 700], ["みかん", 800, 300], ["ぶどう", 1500, 900]],
               name="gappy.xlsx")
    return p, {"sheets": ["売上"], "headers": {"売上": ["商品", "売上", "原価"]},
                "header_rows": {"売上": 1}, "path": str(p)}


def test_the_extent_organ_sees_past_a_blank_row(tmp_path):
    p, _m = _gappy(tmp_path)
    ws = openpyxl.load_workbook(p)["売上"]
    assert ailine._scan_last_row(ws, header_row=1) == 1, "前提: 走査は 1 行で止まる"
    last, cols = ailine.data_extent(ws, header_row=1)
    assert (last, cols) == (5, 3), (last, cols)


@pytest.mark.parametrize("task,expect_at", [
    ("みかんの下に梨を追加して", 5),
    ("みかんの下に新しい行を挿入して", 5),
    ("りんごの行を削除して", 3),
])
def test_positions_resolve_past_a_blank_row(tmp_path, task, expect_at):
    _p, meta = _gappy(tmp_path)
    at, note = ailine.resolve_row_anchor(task, meta, "売上")
    assert at == expect_at, f"{task} → {at} / {note}"


def test_a_row_named_by_its_content(tmp_path):
    """★「りんごの行を削除して」── 人は行を**中身**で指す。"""
    _p, meta = _gappy(tmp_path)
    at, note = ailine.resolve_row_anchor("りんごの行を削除して", meta, "売上")
    assert at == 3 and "『りんご』の行" in note, note


def test_a_column_deletion_is_not_mistaken_for_a_row(tmp_path):
    """誤爆防止: 「金額の列を削除して」を行の指定と読まない。"""
    _p, meta = _gappy(tmp_path)
    assert ailine.resolve_row_anchor("原価の列を削除して", meta, "売上") == (None, None)


def test_postconditions_do_not_report_zero_targets_on_a_gappy_sheet(tmp_path):
    """★ 事後条件も同じ器官を通る（「検証対象が 0 件」で止まらない）。"""
    before, _m = _gappy(tmp_path)
    after = _book(tmp_path, [["商品", "売上", "原価"], [None, None, None],
                              ["りんご", 1200, 700], ["みかん", 800, 300],
                              ["梨", 600, 300], ["ぶどう", 1500, 900]], name="after.xlsx")
    status, reason = ailine.check_add_row(
        after, {"at": 5, "values": {"商品": "梨", "売上": 600, "原価": 300}},
        source_book=before)
    assert status == "pass", reason


@pytest.mark.parametrize("task", [
    "2行目の前に1行挿入して",
    "3行目に1行挿入して",
    "5行目を削除して",
])
def test_row_numbers_are_not_content_anchors(tmp_path, task):
    """★ 2026-08-27（自分で入れた誤爆・既存の検体が捕まえた）。

    「**2行目の前に**1行挿入して」の「2行目」を中身の名前として探し、見つからず
    断っていた。**行番号は名前ではない** ── 数字の指定はそのまま通す。
    ★ 相対の言い回しを足すと、絶対の言い回しを壊しうる。両方を同じ試験で見張る。
    """
    _p, meta = _gappy(tmp_path)
    assert ailine.resolve_row_anchor(task, meta, "売上") == (None, None), task


# --- ⑧ 式の列がある表で、行を動かしても落ちないこと（2026-08-27・実測で出た欠陥）------
#
# ★ 実測: README の手順どおり ①利益列を作る → ②「みかんの下に梨を追加して」と続けたら、
#   **正しく押し下げているのに ×**（「押し下げずに上書きした疑い」）が出た。
#   原因は、式は行が動くと参照が追随する（`=B4-C4` → `=B5-C5`）のに、
#   事後条件が**式を文字で比べていた**こと。デモの経路そのものだった。
#
# ★ 直しは 3 経路（追加・行削除・列削除）**すべて**に効く 1 箇所（compare_moved_rows）。
#   だから試験も 3 経路ぶん置く ── 1 本だけ直すと、次に同じ形で刺される。

_F_ROWS = [["商品", "売上", "原価", "利益"],
            ["りんご", 1200, 700, "=B2-C2"],
            ["みかん", 800, 300, "=B3-C3"],
            ["ぶどう", 1500, 900, "=B4-C4"]]


def test_add_row_does_not_fail_just_because_formulas_followed_the_shift(tmp_path):
    before = _book(tmp_path, _F_ROWS, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価", "利益"],
                              ["りんご", 1200, 700, "=B2-C2"],
                              ["みかん", 800, 300, "=B3-C3"],
                              ["梨", 600, 300, None],          # ← 4 行目に挿さった
                              ["ぶどう", 1500, 900, "=B5-C5"]])  # ← 式が追随した
    args = {"at": 4, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, reason = ailine.check_add_row(after, args, source_book=before)
    assert status == "pass", f"追随した式を「上書き」と誤断した: {reason}"


def test_add_row_still_catches_an_overwrite_when_the_table_has_formulas(tmp_path):
    """★ 恒真殺し: 緩めた分で、本当の上書きを見逃していないこと。"""
    before = _book(tmp_path, _F_ROWS, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価", "利益"],
                              ["りんご", 1200, 700, "=B2-C2"],
                              ["みかん", 800, 300, "=B3-C3"],
                              ["梨", 600, 300, None]])          # ぶどうを潰した
    args = {"at": 4, "values": {"商品": "梨", "売上": 600, "原価": 300}}
    status, reason = ailine.check_add_row(after, args, source_book=before)
    assert status == "fail", f"上書きを通した: {reason}"


def test_delete_rows_does_not_fail_just_because_formulas_followed_the_shift(tmp_path):
    before = _book(tmp_path, _F_ROWS, name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "原価", "利益"],
                              ["りんご", 1200, 700, "=B2-C2"],
                              ["ぶどう", 1500, 900, "=B3-C3"]])   # みかんを消して上へ詰めた
    args = {"at": 3, "count": 1}
    status, reason = ailine.check_delete_rows(after, args, source_book=before)
    assert status == "pass", f"追随した式を「詰め方が正しくない」と誤断した: {reason}"


def test_delete_column_does_not_fail_just_because_formulas_followed_the_shift(tmp_path):
    before = _book(tmp_path, [["商品", "原価", "売上", "利益"],
                               ["りんご", 700, 1200, "=C2-B2"],
                               ["みかん", 300, 800, "=C3-B3"]], name="before.xlsx")
    after = _book(tmp_path, [["商品", "売上", "利益"],
                              ["りんご", 1200, "=B2-999"],
                              ["みかん", 800, "=B3-999"]])        # 参照が左へ寄った（形は変わる）
    args = {"col": "原価"}
    status, reason = ailine.check_delete_column(after, args, source_book=before)
    assert status == "pass", f"追随した式を「別の列を巻き込んだ」と誤断した: {reason}"


# --- ⑨ 比べ方そのものの単体（判断は 1 箇所にしか無い）------------------------------
def _cell(raw, cached=None):
    is_f = isinstance(raw, str) and raw.startswith("=")
    return (is_f, raw, cached if is_f else raw)


def test_compare_moved_rows_is_strict_about_literals():
    a = [(_cell("みかん"), _cell(800))]
    b = [(_cell("みかん"), _cell(801))]
    st, why = ailine.compare_moved_rows(a, b, "下")
    assert st == "broken" and "2 列目" in why, why


def test_compare_moved_rows_refuses_when_a_formula_disappears():
    a = [(_cell("みかん"), _cell(500))]
    b = [(_cell("みかん"), _cell("=B3-C3", 500))]
    st, why = ailine.compare_moved_rows(a, b, "下")
    assert st == "broken" and "消えました" in why, why


def test_compare_moved_rows_accepts_a_formula_that_only_changed_shape():
    """★ これが直したかった形: 式の文字は変わったが、計算後の値は同じ。"""
    a = [(_cell("みかん"), _cell("=B5-C5", 500))]
    b = [(_cell("みかん"), _cell("=B4-C4", 500))]
    st, notes = ailine.compare_moved_rows(a, b, "下")
    assert st == "ok" and notes == [], notes


def test_compare_moved_rows_discloses_when_the_computed_value_changed():
    """★ 落とさずに開示する ── 挿入した行を巻き込む合計式なら正当に変わりうる。
       ただし ✓ は出さない（呼び側が warn を返す）。"""
    a = [(_cell("合計"), _cell("=SUM(B2:B5)", 4100))]
    b = [(_cell("合計"), _cell("=SUM(B2:B4)", 3500))]
    st, notes = ailine.compare_moved_rows(a, b, "下")
    assert st == "ok" and len(notes) == 1 and "3500" in notes[0], notes
    assert "✓" not in ailine._moved_rows_note(notes)
