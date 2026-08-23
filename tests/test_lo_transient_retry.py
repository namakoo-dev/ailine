# operator 摩擦⑥（UNO DisposedException の自動 1 回リトライ）── 実装より先に凍結した赤い検体。
# 価値の根拠: operator8 ③の真因が「StructDump 不調→無言の 1 行目」= LO の一時不調だった。
# 正規化+StructDump 側には既に復元つきリトライがある（M2c・normalize_book）── 適用側が
# 未配線（片配線の形）。
#
# 契約:
#   ① LO の一時不調（DisposedException 等の凍結マーカー）で適用が失敗したら、
#      LO を stop → **原本から作り直して** 1 回だけ再試行し、再試行した事実を開示する
#   ② 再試行に渡る book は無垢であること（半適用の上に再実行しない ── mock 内 assert で凍結）
#   ③ 一時不調でない普通の実行時エラーはこの層で再試行しない（盲目リトライをしない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402


def _sort_book(tmp_path):
    return _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])


def _translate_sort(model, task, book_meta, temperature=0.1):
    return {"op": "SORT", "args": {"col": "金額", "order": "desc"}}


def test_disposed_exception_retries_once_with_pristine_book(tmp_path, monkeypatch, capsys):
    """①+②: 1 回目は半適用のまま DisposedException で死ぬ → 復元+再試行で成功。
       再試行が受け取る book が無垢であることを mock 内で機械検査する。"""
    _isolate(monkeypatch, tmp_path)
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_sort)
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if calls["n"] == 1:
            # 半適用: A2 だけ書き換えて死ぬ（LO がマクロ途中で落ちた形）
            ws["A2"] = "a"
            wb.save(out_book)
            return False, "com.sun.star.lang.DisposedException: Binary URP bridge disposed", "raw"
        # ★ 契約②: 再試行は無垢の book から始まること（半適用の残骸が無い）
        assert ws["A2"].value == "b", f"再試行が半適用の上に走っている: A2={ws['A2'].value!r}"
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr(ailine, "_stop_office", lambda: None)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert calls["n"] == 2, f"再試行していない（呼び出し {calls['n']} 回）: {out}"
    assert rc == 0, out
    assert "再試行" in out, f"再試行の開示が無い: {out}"
    assert "✓" in out, out


def test_ordinary_runtime_error_does_not_retry_here(tmp_path, monkeypatch, capsys):
    """③: 普通の実行時エラー（マクロの論理エラー）はこの層で再試行しない。"""
    _isolate(monkeypatch, tmp_path)
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_sort)
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        return False, "BASIC runtime error: Sub or Function not defined", "raw"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert calls["n"] == 1, f"一時不調でないのに再試行している（盲目リトライ）: {calls['n']} 回"
    assert rc != 0


def test_disposed_twice_fails_honestly(tmp_path, monkeypatch, capsys):
    """再試行は 1 回だけ ── 2 回目も一時不調なら正直に失敗する（無限リトライしない）。"""
    _isolate(monkeypatch, tmp_path)
    book = _sort_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_sort)
    calls = {"n": 0}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        return False, "com.sun.star.lang.DisposedException: Binary URP bridge disposed", "raw"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr(ailine, "_stop_office", lambda: None)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert calls["n"] == 2, f"再試行が 1 回でない: {calls['n']} 回"
    assert rc != 0, f"2 連続の一時不調で成功を名乗っている: {out}"


def test_freeform_step_disposed_retry_with_pristine_source(tmp_path, monkeypatch, capsys):
    """両経路分（幹の系譜）: 自由生成の段でも 一時不調→復元つき 1 回再試行。
       再試行が受け取る book は段の無垢な入力（前段の適用結果）であること。"""
    from test_ailine import _plan_book, run_argv  # noqa: E402
    p = _plan_book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                    {"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]})
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)
    code = ("Sub Run(oDoc As Object)\n"
            "  oDoc.Sheets.getByIndex(0).getCellByPosition(9, 0).setValue(1)\n"
            "End Sub")
    monkeypatch.setattr(ailine, "ollama_generate", lambda model, msgs, temperature=0.2: code)
    monkeypatch.setattr(ailine, "_stop_office", lambda: None)
    calls = {"n": 0}

    def fake_apply(out_book, c, workdir, helper_files=(), timeout=None):
        calls["n"] += 1
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        if calls["n"] == 1:            # SORT 段: 素直に成功（既に降順なので無変化はまずい）
            cell = ws2.cell(row=1, column=9)
            cell.value = (cell.value or 0) + 1
            wb2.save(out_book)
            return True, None, "ok"
        if calls["n"] == 2:            # 自由生成段 1 回目: 半適用で死ぬ
            ws2["A2"] = "汚染"
            wb2.save(out_book)
            return False, "com.sun.star.lang.DisposedException: bridge disposed", "raw"
        # 再試行: 半適用の残骸が無いこと（段の無垢な入力から作り直されている）
        assert ws2["A2"].value != "汚染", "自由生成段の再試行が半適用の上に走っている"
        cell = ws2.cell(row=1, column=10)
        cell.value = (cell.value or 0) + 1
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    argv = run_argv(book=str(p), task="金額で降順に並べ替えて条件付き書式も付けて",
                     model="qwen2.5-coder:7b", refs=None, helpers=None, repair=0,
                     temperature=0.2, dry=False, copy=True, json=False, timeout=180.0,
                     ask=False, allow_freeform=True)
    rc = ailine.main(argv)
    out = capsys.readouterr().out
    assert calls["n"] == 3, f"自由生成段が再試行していない: {calls['n']} 回\n{out}"
    assert "再試行" in out, f"再試行の開示が無い: {out}"
    assert rc == 0, out
