# 誤分類の実例台帳センサ ── 実装より先に凍結した赤い検体。
# 出典: cookbooks ②-C（動的 few-shot 分類）の発火条件「誤分類の実例 10 件」を測れるように
# する前置き（Namakoo 興味あり印・ABSORB-20260822-claude-cookbooks.md）。vocab_miss と同じ
# 需要センサ方式 ── 記録するだけ。分析・提案・表示は 10 件貯まってから設計する。
#
# 契約:
#   ① 関所で N（解釈を見て人が止めた）→ ~/.ailine/misclass.jsonl に 1 行追記
#      {ts, signal:"gate_decline", task, plan, book}
#   ② ailine undo → 直近 run の task を添えて signal:"undo" で 1 行追記
#   ③ 普通に成功した run は何も記録しない（センサは容疑だけを拾う）

import json
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "MISCLASS_FILE"),
    reason="誤分類センサ 未実装（契約は凍結済み・実装が来たら自動で実測に切り替わる）",
    strict=True,
)


def _read_entries():
    p = ailine.MISCLASS_FILE
    if not Path(p).exists():
        return []
    return [json.loads(ln) for ln in Path(p).read_text(encoding="utf-8").splitlines() if ln]


def _overwrite_book(tmp_path):
    # SET_COLUMN_VALUE で既存値のある列に書く → 破壊の関所（y/N）が必ず開く形
    return _book(tmp_path, [["商品", "備考"], ["a", "旧値1"], ["b", "旧値2"]])


@needs_impl
def test_gate_decline_records_suspect(tmp_path, monkeypatch, capsys):
    """①: 関所で N → 誤分類容疑として task と plan が記録される。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _overwrite_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SET_COLUMN_VALUE", "args": {"col": "備考", "value": "確認済み"}})
    monkeypatch.setattr("sys.stdin", __import__("io").StringIO("n\n"))
    monkeypatch.setattr("sys.stdin.isatty", lambda: True, raising=False)
    rc, out = _run_main(["run", str(book), "備考列を全部「確認済み」にして"], capsys)
    entries = _read_entries()
    assert len(entries) == 1, f"関所 N が記録されていない: {entries}\n{out}"
    e = entries[0]
    assert e["signal"] == "gate_decline"
    assert "備考列" in e["task"]
    assert e["plan"], "plan が空"


@needs_impl
def test_undo_records_suspect_with_last_task(tmp_path, monkeypatch, capsys):
    """②: 適用 → undo で、直近 run の task を添えた容疑が記録される。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて"], capsys)
    assert rc == 0, out
    assert _read_entries() == [], "成功しただけで記録された（契約③違反）"
    rc2, out2 = _run_main(["undo", str(book)], capsys)
    assert rc2 == 0, out2
    entries = _read_entries()
    assert len(entries) == 1, f"undo が記録されていない: {entries}\n{out2}"
    e = entries[0]
    assert e["signal"] == "undo"
    assert "並べ替え" in e["task"], f"直近 run の task が引かれていない: {e}"


@needs_impl
def test_clean_success_records_nothing(tmp_path, monkeypatch, capsys):
    """③: 関所も undo も無い成功 run は 1 行も記録しない。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    assert _read_entries() == []
