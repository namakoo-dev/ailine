# 同名の列で値が別列に化けるのを止める（2026-08-24・第二波 M3）。
#
# ★ 実測（盲検 2 者が独立に再現・俺も実物で再現）:
#     見出し `品番 / 備考 / 金額 / 備考`（社内用メモ と 客先提出用）
#     出力  ('0012','社内用メモ',1000,'社内用メモ',…)   ← **客先提出用が社内用に化けた**
#     exit 0 ・ ⚠ なし ・ Σ金額 元2000/出力2000 ✓
#   しかも**やる側（stack）と見る側（verify）が同じ「名前→列」の辞書を別々に組む**ので、
#   照合が同じように潰れて**恒真**になる ── 誰も気づけない。
#
# ★ 根: **列を「名前」で引いている**。名前は表示のためのラベルであって、
#   同一性の担保ではない（同じ名前の列は実在する）。
#
# 契約:
#   ① 同名の列があっても、**位置ごとに別の列として扱う**（値が化けない）
#   ② 名前で引く経路が残っているなら、**同名を検出して名指しする**（黙って潰さない）
#   ③ 誤爆しない: 同名が無い普通の表は今までどおり

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from ailine_core.multifile import _column_index, duplicate_header_names  # noqa: E402


def test_duplicate_names_are_detected():
    """② 同名の列を名指しできる（黙って潰さない）。"""
    dupes = duplicate_header_names(["品番", "備考", "金額", "備考"])
    assert dupes == ["備考"], dupes


def test_no_duplicates_is_empty():
    """③ 誤爆防止。"""
    assert duplicate_header_names(["品番", "備考", "金額"]) == []


def test_empty_names_are_not_counted_as_duplicates():
    """★ 位置保持のための空文字（結合見出しの直し）を同名扱いしない。"""
    assert duplicate_header_names(["品番", "", "金額", ""]) == []


def test_column_index_still_finds_the_first_for_unique_names():
    """③ 一意な名前は今までどおり（1 起点）。"""
    assert _column_index(["品番", "備考", "金額"], "金額") == 3
    assert _column_index(["品番", "備考"], "無い") is None


# --- ★ 器を作って呼ばない、をやらない（配線を検体で縛る）----------------------------

def test_stack_refuses_a_book_with_duplicate_columns(tmp_path):
    """① 同名の列がある冊は**積まずに名指しで断る**（値が化けるより断られる方が安い）。"""
    import openpyxl, subprocess, sys as _s, os
    f = tmp_path / "d"; f.mkdir()
    for n in ("a.xlsx", "b.xlsx"):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["品番", "備考", "金額", "備考"])
        ws.append(["0012", "社内用メモ", 1000, "客先提出用"])
        wb.save(f / n)
    out = tmp_path / "out.xlsx"
    r = subprocess.run([_s.executable, "-m", "ailine", "stack", str(f), "--out", str(out)],
                        capture_output=True, text=True, encoding="utf-8", errors="replace",
                        cwd=str(Path(__file__).resolve().parent.parent),
                        env={**os.environ, "PYTHONPATH": "src"})
    assert "備考" in r.stdout, f"同名の列を名指ししていない:\n{r.stdout}"
    assert r.returncode != 0, f"値が化けるのに exit 0:\n{r.stdout}"
    if out.exists():
        ws2 = openpyxl.load_workbook(out).active
        rows = [x for x in ws2.iter_rows(values_only=True)][1:]
        for row in rows:
            assert not (row[1] == row[3] == "社内用メモ"), \
                f"客先提出用が社内用に化けたまま書かれた: {row}"


def test_stack_still_works_without_duplicates(tmp_path):
    """③ 誤爆防止: 同名が無ければ今までどおり通る。"""
    import openpyxl, subprocess, sys as _s, os
    f = tmp_path / "ok"; f.mkdir()
    for n in ("a.xlsx", "b.xlsx"):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["品番", "備考", "金額"]); ws.append(["0012", "メモ", 1000])
        wb.save(f / n)
    out = tmp_path / "out2.xlsx"
    r = subprocess.run([_s.executable, "-m", "ailine", "stack", str(f), "--out", str(out)],
                        capture_output=True, text=True, encoding="utf-8", errors="replace",
                        cwd=str(Path(__file__).resolve().parent.parent),
                        env={**os.environ, "PYTHONPATH": "src"})
    assert r.returncode == 0, r.stdout
