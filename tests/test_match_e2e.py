"""M3 `ailine run <入金.xlsx> <請求.xlsx> "<依頼>"`（2 冊の照合）の検体。
   ★ 実装前に凍結した赤い検体（DESIGN-20260821-multifile M3 v2・specimen-first）。

   芯（買い手の言葉・凍結）: 「候補を並べて差額だけ機械で保証、決めるのは人。
   消し込みまで自動と言われた瞬間、価値 0」── 判断しない。並べて、差額を保証して、渡す。
   ★ 列名は実物寄り（共通列名 0 個)── 列対応は依頼文の名指し（機械照合・LLM ゼロ）で解決。"""
import datetime
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

NYUKIN_HDRS = ["取引日", "振込人名義", "お預り金額", "摘要"]
SEIKYU_HDRS = ["請求日", "取引先名", "請求金額", "請求番号"]
TASK = "振込人名義と取引先名をキーに、お預り金額と請求金額を突き合わせて"


def _nyukin(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入金明細"
    ws.append(NYUKIN_HDRS)
    for r in rows:
        ws.append(list(r))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _seikyu(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求一覧"
    ws.append(SEIKYU_HDRS)
    for r in rows:
        ws.append(list(r))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _five_forms(tmp_path):
    """5 つの現実形: 完全一致・手数料引き・請求のみ・入金のみ・合算入金。"""
    a = tmp_path / "in" / "入金明細_7月.xlsx"
    b = tmp_path / "in" / "請求_7月.xlsx"
    _nyukin(a, [
        (datetime.date(2026, 7, 31), "甲社", 220000, "7月分"),
        (datetime.date(2026, 7, 31), "乙社", 109340, "手数料引き"),
        (datetime.date(2026, 8, 1),  "丁社", 55000, "台帳に無い"),
        (datetime.date(2026, 8, 2),  "戊社", 330000, "2件合算"),
    ])
    _seikyu(b, [
        (datetime.date(2026, 7, 10), "甲社", 220000, "INV-101"),
        (datetime.date(2026, 7, 12), "乙社", 110000, "INV-102"),
        (datetime.date(2026, 7, 15), "丙社", 88000, "INV-103"),
        (datetime.date(2026, 7, 20), "戊社", 110000, "INV-104"),
        (datetime.date(2026, 7, 25), "戊社", 220000, "INV-105"),
    ])
    return a, b


def _run(a, b, task, *extra, capsys=None):
    rc = ailine.main(["run", str(a), str(b), task, *extra])
    out = capsys.readouterr().out if capsys else ""
    return rc, out


def _sheet_rows(path, title):
    ws = openpyxl.load_workbook(path)[title]
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_five_forms_land_in_arithmetic_states(tmp_path, capsys):
    """本命: 5 形が算術の状態語に振り分けられる。『一致』の語は使わない（合算・相殺で嘘になる）。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0, out
    outs = list((tmp_path / "in").glob("*照合*.xlsx")) or list((tmp_path / "in").glob("*突き合わせ*.xlsx"))
    assert outs, f"出力ブックが A の親に無い: {list((tmp_path/'in').glob('*.xlsx'))}"
    rows = _sheet_rows(outs[0], "照合")
    by_key = {r[0]: r for r in rows[1:] if r and r[0]}
    assert by_key["乙社"][5] == -660, f"手数料の差額: {by_key['乙社']}"
    assert by_key["甲社"][5] == 0
    assert by_key["戊社"][5] == 0 and by_key["戊社"][1] == 1 and by_key["戊社"][3] == 2, \
        f"合算: 差額 0 でも件数 A1/B2 が同じ行に見えること: {by_key['戊社']}"
    assert by_key["丙社"][6] == "B のみ" and by_key["丁社"][6] == "A のみ"
    text = " ".join(str(c) for r in rows for c in r if c is not None)
    assert "一致" not in text, "『一致』は ✓ の嘘（憲法2）── 状態語は算術のみ"
    assert "660" in out and "乙社" in out, f"端末報告に差額の名指しが無い:\n{out}"


def test_symmetry_swapping_books_flips_only_the_sign(tmp_path, capsys):
    """★ M2 バー（最強の不変条件）: A/B を入れ替えると差額の符号だけ反転し、
       キー集合・件数・Σ・状態の対応は同一。片肺実装を 1 本で捕る。"""
    a, b = _five_forms(tmp_path)
    rc1, _ = _run(a, b, TASK, capsys=capsys)
    assert rc1 == 0
    out1 = next((tmp_path / "in").glob("*.xlsx.tmp"), None) or \
        max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    rows1 = {r[0]: r for r in _sheet_rows(out1, "照合")[1:] if r and r[0]}
    task_rev = "取引先名と振込人名義をキーに、請求金額とお預り金額を突き合わせて"
    rc2, _ = _run(b, a, task_rev, capsys=capsys)
    assert rc2 == 0
    out2 = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    assert out2 != out1
    rows2 = {r[0]: r for r in _sheet_rows(out2, "照合")[1:] if r and r[0]}
    assert set(rows1) == set(rows2), "キー集合が入れ替えで変わった"
    for k in rows1:
        assert rows1[k][5] == -(rows2[k][5] or 0) if rows1[k][5] is not None else True, \
            f"{k}: 符号だけの反転になっていない: {rows1[k][5]} vs {rows2[k][5]}"


def test_self_match_all_zero_with_disclosure(tmp_path, capsys):
    """M3 バー: 同一ブックを 2 冊に渡すと全キー差額 0 + sha 一致の 1 行開示（拒否しない）。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, a, "振込人名義と振込人名義をキーに、お預り金額とお預り金額を突き合わせて",
                   capsys=capsys)
    assert rc == 0, out
    assert "同じ" in out or "同一" in out, f"同一ブックの開示が無い:\n{out}"


def test_blank_key_row_becomes_fifth_state_not_silently_dropped(tmp_path, capsys):
    """★ 致命3: キー空白の入金（誰からか分からない ── 人が一番見たい行）は
       『キー不明』の区分で必ず表に出る。単位L の構造トリガで黙って消さない。
       『合計商事』もキーとして生きる（ラベル語の部分一致で除外しない）。"""
    a = tmp_path / "in" / "入金.xlsx"
    b = tmp_path / "in" / "請求.xlsx"
    _nyukin(a, [
        (datetime.date(2026, 7, 1), "合計商事", 10000, ""),
        (datetime.date(2026, 7, 2), None, 5000, "名義不明"),
    ])
    _seikyu(b, [(datetime.date(2026, 7, 1), "合計商事", 10000, "INV-1")])
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0, out
    outp = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    rows = _sheet_rows(outp, "照合")
    text = " ".join(str(c) for r in rows for c in r if c is not None)
    assert "キー不明" in text, f"キー空白行が消えた: {text[:300]}"
    assert "5000" in text, "キー不明の金額が表に出ていない"
    assert "合計商事" in text, "『合計商事』がラベル語トリガで消えた"


def test_ambiguous_amount_column_stops_with_exit_3(tmp_path, capsys):
    """★ 致命5: 金額列に既定を持たせない。依頼文が列を名指しせず候補が複数なら
       exit 3 + 候補の列挙（数量を金額として ✓ を出すのが最悪の形）。"""
    a = tmp_path / "in" / "入金.xlsx"
    b = tmp_path / "in" / "請求.xlsx"
    _nyukin(a, [(datetime.date(2026, 7, 1), "甲社", 100, "")])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求一覧"
    ws.append(["請求日", "取引先名", "数量", "単価", "請求金額"])
    ws.append([datetime.date(2026, 7, 1), "甲社", 2, 50, 100])
    b.parent.mkdir(parents=True, exist_ok=True)
    wb.save(b)
    rc, out = _run(a, b, "入金と請求を突き合わせて", capsys=capsys)
    assert rc == 3, f"曖昧な金額列で走った (rc={rc}):\n{out}"
    assert "数量" in out and "請求金額" in out, f"候補の列挙が無い:\n{out}"


def test_nonexistent_second_path_is_asked_not_swallowed_as_task(tmp_path, capsys):
    """★ 代替1(a): 2 冊目のパスを打ち間違えたら、黙って依頼文として単一ブック経路に
       流さず、名指しで聞いて止まる（exit 3）。"""
    a, b = _five_forms(tmp_path)
    typo = str(b).replace("請求", "誠求")
    rc, out = _run(a, typo, TASK, capsys=capsys)
    assert rc == 3, f"打ち間違いが黙って依頼文扱いされた (rc={rc}):\n{out}"
    assert "誠求" in out or "見つかりません" in out, f"名指しの確認が無い:\n{out}"


def test_detail_sheet_lists_all_rows_with_provenance_and_bands(tmp_path, capsys):
    """代替3: ②明細シート ── 両冊の全データ行を キー順・出所つきで並べる（「候補を並べて」の充足）。
       差額 0 でないキーの行は薄赤の帯。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0
    outp = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    wb = openpyxl.load_workbook(outp)
    assert "明細" in wb.sheetnames, wb.sheetnames
    ws = wb["明細"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "INV-102" in text and "109340" in text, "両冊の行が明細に並んでいない"
    tinted = [c for row in ws.iter_rows(min_row=2) for c in row
              if c.fill is not None and "C7CE" in (c.fill.fgColor.rgb or "")]
    assert tinted, "差額ありキーの帯が無い"


# ---- 実弾検分の差し戻し（2026-08-21 15:3x・実物寄り検体で確認した実害）----


def _with_bank_total(tmp_path):
    """銀行明細の実物形: 合計行（1 列目に『合計』・キー列は空・金額 = 他行の和）入り。"""
    a = tmp_path / "in" / "入金.xlsx"
    b = tmp_path / "in" / "請求.xlsx"
    _nyukin(a, [
        (datetime.date(2026, 7, 31), "甲社", 220000, ""),
        (datetime.date(2026, 8, 2), None, 5000, "名義読めず"),
        ("合計", None, 225000, ""),
    ])
    _seikyu(b, [(datetime.date(2026, 7, 10), "甲社", 220000, "INV-101")])
    return a, b


def test_bank_total_row_gets_note_not_silent_pollution(tmp_path, capsys):
    """★ 実害: 合計行がキー不明の袋に混ざり、本物の名義不明 5,000 が 225,000 の中に沈む。
       設計どおり除外はしないが、『金額が他の行の和と一致 ── 合計行かもしれません』の注記を
       両側の数字つきで出す（報告と検分の両方）。"""
    a, b = _with_bank_total(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0, out
    assert "合計行かもしれません" in out, f"注記が端末に無い:\n{out}"
    assert "225000" in out.replace(",", ""), "注記に当の金額が無い"
    outp = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    text = " ".join(str(c.value) for row in openpyxl.load_workbook(outp)["検分"].iter_rows()
                    for c in row if c.value is not None)
    assert "合計行かもしれません" in text, "注記が検分に無い"


def test_terminal_unknown_key_line_reads_naturally(tmp_path, capsys):
    """文言: 『⚠ キー不明: キー不明（…）』の重複をやめ、行数つきで読める形に。"""
    a, b = _with_bank_total(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0
    assert "キー不明: キー不明" not in out, f"重複文言:\n{out}"


def test_inspection_findings_link_to_detail_sheet(tmp_path, capsys):
    """★ gap#1（誘導の憲法）: 差額あり・片側のみ・キー不明の各所見が検分の所見表に載り、
       ブック内リンク（#'明細'!…）で明細シートの該当キーの行へ飛べる。
       ブック内リンクなら 2 冊が別フォルダでも切れない。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0
    outp = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    ws = openpyxl.load_workbook(outp)["検分"]
    links = [(c.hyperlink.target or "", c.hyperlink.location or "")
             for row in ws.iter_rows() for c in row if c.hyperlink is not None]
    assert links, "所見にリンクが 1 本も無い"
    assert any("明細" in loc for t, loc in links if not t), \
        f"明細シートへのブック内リンクが無い: {links}"
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "乙社" in text and "660" in text, "差額ありの所見が検分の表に無い"


def test_detail_sheet_dates_keep_date_format(tmp_path, capsys):
    """明細シートの日付セルが 00:00:00 の尻尾つきで出る再発 ── 元の number_format を運ぶ
       （M2.5 の日付書式引き継ぎと同じ線・照合の明細にも適用）。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc == 0
    outp = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    ws = openpyxl.load_workbook(outp)["明細"]
    import datetime as _dt
    date_cells = [c for row in ws.iter_rows(min_row=2) for c in row
                  if isinstance(c.value, (_dt.date, _dt.datetime))]
    assert date_cells, "明細に日付セルが無い（検体の前提崩れ）"
    bad = [c.number_format for c in date_cells if "h" in c.number_format or ":" in c.number_format
           or c.number_format == "General"]
    assert not bad, f"時刻の尻尾つき書式が残る: {bad[:3]}"


# ---- jisaku-review 5 戦目 確定 4 件の凍結（2026-08-21 16:0x）----


def test_postcondition_recomputes_sums_from_source_books(tmp_path, monkeypatch, capsys):
    """★ review5#1 critical（凍結予測③的中）: 事後条件が a_sum/b_sum を原本から独立に
       再集計しておらず、内部整合（diff = a−b）さえ保てば虚偽の金額が exit 0 で出荷された。
       『差額だけ機械で保証』の芯そのもの。契約: compute_match の金額改竄（diff 整合込み）は
       事後条件が両側の数字つきで捕まえ exit 1・出力を本置き場に残さない。"""
    from ailine_core import match as mm
    a, b = _five_forms(tmp_path)
    real = mm.compute_match

    def tampered(*args, **kw):
        groups, extras = real(*args, **kw) if isinstance(real(*args, **kw), tuple) else (real(*args, **kw), None)
        # 甲社の A 側金額を膨らませ、diff も式どおり再計算して内部整合は保つ
        for g in (groups if extras is None else groups):
            if getattr(g, "key", None) == "甲社":
                object.__setattr__(g, "a_sum", 999999) if hasattr(g, "__dataclass_fields__") else None
                try:
                    g.a_sum = 999999
                    g.diff = g.a_sum - g.b_sum
                except Exception:
                    pass
        return (groups, extras) if extras is not None else groups

    monkeypatch.setattr(mm, "compute_match", tampered)
    rc, out = _run(a, b, TASK, capsys=capsys)
    assert rc != 0, f"虚偽の金額 999999 が事後条件を素通りした:\n{out}"
    assert "999999" in out or "事後条件" in out, f"両側の数字つきの破れ報告が無い:\n{out}"


def test_xlsm_typo_second_path_is_asked_not_swallowed(tmp_path, capsys):
    """★ review5#2: 拡張子リストに .xlsm が無く、裸のファイル名の打ち間違いが黙って
       依頼文として単一ブック経路に落ちた。表計算らしい拡張子は広く『2 冊目のパスらしい』
       と見て、実在しなければ名指しで聞く（exit 3）。"""
    a, b = _five_forms(tmp_path)
    rc, out = _run(a, "請求_7月.xlsm", TASK, capsys=capsys)
    assert rc == 3, f".xlsm の打ち間違いが黙って流れた (rc={rc}):\n{out}"
    assert "請求_7月.xlsm" in out, f"名指しが無い:\n{out}"


def test_existing_extensionless_file_does_not_hijack_single_book_task(tmp_path, monkeypatch, capsys):
    """★ review5#3: 依頼文の先頭語がたまたま実在ファイル名と一致すると、単一ブックの依頼が
       黙って 2 冊照合に誤読された。2 冊分岐は『実在 かつ 表計算らしい形（拡張子/区切り）』の
       両方を要求 ── 拡張子なしの偶然の一致では発火しない。"""
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1:
                        {"plan": [{"op": "CLARIFY", "question": "確認"}]})
    a, _ = _five_forms(tmp_path)
    trap = tmp_path / "金額データ"
    trap.write_text("罠", encoding="utf-8")
    monkeypatch.chdir(tmp_path)
    rc = ailine.main(["run", str(a), "金額データ", "を集計して"])
    out = capsys.readouterr().out
    # ★ 誤読の顔は 2 つある: 「2冊の照合」バナー / 罠ファイルをブックとして読もうとする
    #   「読めませんでした」。どちらも出ず、単一ブック経路（モックの CLARIFY）に到達すること。
    assert "2冊の照合" not in out and "読めませんでした" not in out, \
        f"単一ブックの依頼が 2 冊照合に誤読された:\n{out}"
    assert rc == 3 and "確認" in out, f"単一ブック経路（CLARIFY モック）に到達していない (rc={rc}):\n{out}"


def test_longest_named_header_wins_over_its_fragment(tmp_path, capsys):
    """★ review5#4（凍結予測①的中・断片問題の再演）: 『金額』⊂『税込金額』の両列があるとき、
       依頼文が『税込金額』だけを名指ししても部分文字列包含で両方が名指し扱いになり
       exit 3 に落ちた。飲み込まれ判定（単位B の轍）: 長い名指しに包含される短い列名は
       名指しと数えない ── 税込金額 に一意解決すること。"""
    a = tmp_path / "in" / "入金.xlsx"
    b = tmp_path / "in" / "請求.xlsx"
    _nyukin(a, [(datetime.date(2026, 7, 1), "甲社", 110000, "")])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求一覧"
    ws.append(["請求日", "取引先名", "金額", "税込金額"])
    ws.append([datetime.date(2026, 7, 1), "甲社", 100000, 110000])
    b.parent.mkdir(parents=True, exist_ok=True)
    wb.save(b)
    rc, out = _run(a, b, "振込人名義と取引先名をキーに、お預り金額と税込金額を突き合わせて",
                   capsys=capsys)
    assert rc == 0, f"断片の飲み込まれで曖昧扱いになった (rc={rc}):\n{out}"
    assert "税込金額" in out, f"解決された列名の開示が無い:\n{out}"
