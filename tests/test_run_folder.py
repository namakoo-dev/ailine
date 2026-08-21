"""M2 `ailine run <フォルダ> "<依頼>"`（抽出集約）の検体。
   ★ 実装前に凍結した赤い検体（DESIGN-20260821-multifile M2 節・E3 が筆頭）。

   ★ 7B を使わない: 翻訳は translate_task の monkeypatch（f9 transcripts と同じ作法・
   製品コードにテスト用の口を彫らない）。測るのはフォルダ分岐の配管と事後条件であって、
   翻訳の質は battery の仕事。"""
import json
import sys
from pathlib import Path

import openpyxl
import pytest

# ★ M2 実装前の凍結検体: 実装が通ったら strict xfail が XPASS で赤くなり、
# この行を外す変更が同じ diff に必ず現れる（黙って通過できない）。

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

HDRS = ["注文ID", "取引先", "金額"]


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r) + [None] * (len(headers) - len(r)))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _mock_translation(monkeypatch, plan):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"plan": plan})


_EXTRACT_40000 = [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 40000}}]


def _run(folder, task, *extra, capsys=None):
    rc = ailine.main(["run", str(folder), task, *extra])
    out = capsys.readouterr().out if capsys else ""
    return rc, out


def test_e3_total_rows_do_not_leak_into_condition_matches(tmp_path, monkeypatch, capsys):
    """★ E3（検体の筆頭）: 『金額 40000 以上』の抽出では合計行が必ず条件を満たす。
       単位L の除外が条件適用の 前 に回っていないと、合計行が混ざって二重計上。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 30000), ("合計", None, 80000)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 45000)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc == 0, out
    out_files = list(tmp_path.glob("*.xlsx"))
    assert out_files, f"出力ブックがフォルダの親に無い:\n{out}"
    ws = openpyxl.load_workbook(out_files[0]).active
    labels = [r[0].value for r in ws.iter_rows(min_row=2)]
    assert "合計" not in labels, f"★ 合計行 80000 が条件を通って混ざった: {labels}"
    assert sorted(x for x in labels if x) == ["J-1", "J-3"], labels


def test_folder_run_never_reports_excel_lock_lie(tmp_path, monkeypatch, capsys):
    """★ architect 致命4 の実バグの凍結: 今日の実装は run <フォルダ> に
       「Excel で開かれています」という嘘の診断を返す（フォルダの open(r+b) を
       ロックと誤読）。フォルダ分岐後はこの文言がフォルダに対して出ないこと。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert "Excel で開かれています" not in out, f"フォルダに Excel ロックの嘘の診断:\n{out}"


def test_unsupported_op_on_folder_is_refused_by_name(tmp_path, monkeypatch, capsys):
    """★ E11: フォルダに未対応 op（並べ替え）→ 名指しの断り + 次の手 + exit 3。
       ★ 黙って 1 冊目に適用が最悪の形（原本無変更を機械で確認）。"""
    _mock_translation(monkeypatch, [{"op": "SORT", "args": {"column": "金額", "order": "desc"}}])
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000)])
    before = (folder / "a.xlsx").read_bytes()
    rc, out = _run(folder, "金額で並べ替えて", capsys=capsys)
    assert rc == 3, f"exit={rc}\n{out}"
    assert "並べ替え" in out and ("抽出" in out or "1 冊" in out), f"名指しの断りと次の手が無い:\n{out}"
    assert (folder / "a.xlsx").read_bytes() == before, "断ったのに原本に触った"


def test_row_accounting_identity_is_reported(tmp_path, monkeypatch, capsys):
    """★ 憲法（行の完全会計・⑨）: 読めた各ファイルで データ行 = 一致 + 不一致 + 除外 の
       恒等が --json の multifile 節に出る（run の既存キーは壊さず入れ子で足す）。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 30000), ("合計", None, 80000)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", "--json", capsys=capsys)
    assert rc == 0, out
    data = json.loads(out)
    mf = data.get("multifile")
    assert mf, f"multifile 節が --json に無い: {list(data)}"
    f = {x["name"]: x for x in mf["files"]}["a.xlsx"]
    assert f["rows_matched"] == 1 and f["rows_unmatched"] == 1 and f["total_rows_excluded"] == 1
    assert f["rows_matched"] + f["rows_unmatched"] + f["total_rows_excluded"] == 3


def test_e4_zero_matches_is_loud_and_never_claims_check(tmp_path, monkeypatch, capsys):
    """★ E4: 全ファイルで一致 0 行 → 出力は書く（見出し+出所列のみ・決定論）が、
       分母つきで「0 行」を明示し ✓ を名乗らない。黙る 0 件は「黙って失敗」の隣。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 100), ("J-2", "乙", 200)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc == 0, out
    assert "0 行" in out, f"一致 0 行の明示が無い:\n{out}"
    assert "✓" not in out, f"0 件の出力に ✓ を名乗った:\n{out}"


def test_multiple_defect_kinds_are_all_reported_in_one_run(tmp_path, monkeypatch, capsys):
    """★ 一括検出（Namakoo 決裁 09:22・1 シート 1 ミスと限らない）: 欠け・合計不一致・
       一致 0 が同居するフォルダで、1 回の実行で全部名指し。⚠ の件数は等号で数える。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("合計", None, 99999)])   # 閉じない合計
    _book(folder / "b.xlsx", ["注文ID", "取引先"], [("J-9", "戊")])                  # 金額が欠け
    _book(folder / "c.xlsx", HDRS, [("J-3", "丙", 45000)])                           # 正常
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc == 0, out
    assert "b.xlsx" in out and "金額" in out, "欠けの名指しが無い"
    assert "99999" in out and "50000" in out, "合計不一致の両側の数字が無い"
    assert out.count("⚠") == 2, f"⚠ は等号 2 のはず（欠け+閉じない合計）:\n{out}"


def test_normal_files_are_not_named_in_human_report_only_anomalies(tmp_path, monkeypatch, capsys):
    """★ 実弾検分の差し戻し（2026-08-21・設計 D6）: 正常ファイル（並べ替えで取れた・
       合計行が正しく閉じて除外された・行の完全会計が成立）は人間向け報告で**名指ししない**
       ── 名指しは異常（取れなかった／閉じる検査の不一致／シート fallback）だけ。
       正常分は「合計行 N 行を M 冊で除外」のような 1 行の集計に畳む（内訳は --json）。
       ⚠ は本物の異常（欠け1件・閉じない合計1件）の 2 本だけのはず。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    # a.xlsx: 正常 ── 合計行はあるが正しく閉じて除外される（異常ではない）。
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 30000), ("合計", None, 80000)])
    # b.xlsx: 正常 ── 列の並び替え（多重集合は一致）。
    _book(folder / "b.xlsx", ["金額", "取引先", "注文ID"],
          [(60000, "己", "J-6")])
    # c.xlsx: 正常 ── 何の異常も無い。
    _book(folder / "c.xlsx", HDRS, [("J-3", "丙", 45000)])
    # d.xlsx: 異常 ── 金額列が欠けて取れない。
    _book(folder / "d.xlsx", ["注文ID", "取引先"], [("J-9", "戊")])
    # e.xlsx: 異常 ── 合計行があるが値が閉じない。
    _book(folder / "e.xlsx", HDRS, [("J-5", "庚", 50000), ("合計", None, 99999)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc == 0, out
    for normal in ("a.xlsx", "b.xlsx", "c.xlsx"):
        assert normal not in out, f"正常ファイル {normal} が人間向け報告に名指しされている:\n{out}"
    assert "d.xlsx" in out, f"異常ファイル d.xlsx の名指しが無い:\n{out}"
    assert "e.xlsx" in out, f"異常ファイル e.xlsx の名指しが無い:\n{out}"
    assert out.count("⚠") == 2, f"⚠ は等号 2 のはず（欠け1件+閉じない合計1件）:\n{out}"
    assert "99999" in out and "50000" in out, f"閉じる検査の両側の数字が無い:\n{out}"
    assert "並べ替え" in out and "1" in out, f"並べ替えの集計が無い:\n{out}"
    assert "合計行" in out and "冊" in out, f"合計行の除外の集計が無い:\n{out}"
    assert "行の完全会計" in out, f"行の完全会計の集計が無い:\n{out}"


def test_verify_extract_kind_mismatch_exits_5(tmp_path, monkeypatch, capsys):
    """★ 実弾検分で発覚した実バグ: 抽出出力を改竄すると verify は ⚠ を出すのに exit=0。
       終了コードだけを見る自動化経路が改竄を見逃す。stack 出力と同じく exit 5。"""
    import subprocess, sys
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000), ("J-2", "乙", 45000)])
    rc, out = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc == 0, out
    out_book = next(tmp_path.glob("*.xlsx"))
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ws.cell(row=2, column=3).value = 99999999
    wb.save(out_book)
    p = subprocess.run([sys.executable, str(REPO / "ailine.py"), "verify", str(out_book), str(folder)],
                       capture_output=True, text=True, timeout=120, encoding="utf-8")
    assert "⚠" in p.stdout, p.stdout
    assert p.returncode == 5, f"⚠ を出したのに exit={p.returncode}（自動化経路が見逃す）:\n{p.stdout}"


def test_long_folder_name_different_conditions_do_not_collide(tmp_path, monkeypatch, capsys):
    """★ review3#1/#5 critical: フォルダ名が長いと切り詰めで別条件の出力が同名に潰れ、
       2 回目が 1 回目を「作り直しました」の顔で無警告消去した（実機再現済み）。
       契約: 作り直してよいのは 印+条件 が両方一致する時だけ。条件が違えば別ファイル
       （切り詰め時は条件のハッシュで分ける）── 両方の結果が生き残ること。"""
    _mock_translation(monkeypatch, _EXTRACT_40000)
    folder = tmp_path / ("あ" * 100)
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 100)])
    rc1, out1 = _run(folder, "金額が40000以上の行を抜き出して", capsys=capsys)
    assert rc1 == 0, out1
    _mock_translation(monkeypatch, [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "lt", "value": 200}}])
    rc2, out2 = _run(folder, "金額が200未満の行を抜き出して", capsys=capsys)
    assert rc2 == 0, out2
    assert "作り直しました" not in out2, "別条件なのに前回出力の作り直しを名乗った"
    outs = sorted(tmp_path.glob("*.xlsx"))
    assert len(outs) == 2, f"別条件の出力が同名に潰れた: {[o.name for o in outs]}"
    all_rows = []
    for o in outs:
        ws = openpyxl.load_workbook(o).active
        all_rows += [r[0].value for r in ws.iter_rows(min_row=2)]
    assert "J-1" in all_rows and "J-2" in all_rows, f"片方の結果が消えた: {all_rows}"


@pytest.mark.parametrize("plan,expect_snippet", [
    ([{"op": "CLARIFY", "question": "対象の列が分かりません"}], "対象の列が分かりません"),
    ([{"op": "EXTRACT", "args": {"column": "金額", "cmp": "between", "value": 100}},
      {"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 100}}],
     "複数の操作をまとめた依頼"),
    ([{"op": "EXTRACT", "args": {"column": "金額", "cmp": "between", "value": 100}}], "比較"),
    ([{"op": "EXTRACT", "args": {"column": "存在しない列", "cmp": "gte", "value": 100}}],
     "列『存在しない列』"),
    ([{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": ""}}], "抽出する値"),
    ([{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": "たくさん"}}],
     "数値として読めない"),
])
def test_folder_condition_validation_paths_exit_3(tmp_path, monkeypatch, capsys,
                                                    plan, expect_snippet):
    """★ review3#4 の裏取り: tests/golden/f6_exit_codes.md の exit=3 行が
       『cmd_run_folder の断りはすべて _run_folder_refuse』と言い過ぎていた ──
       CLARIFY・複数段計画・cmp/col/value の読み取り不能は _run_folder_refuse を
       経由しない直接の return 3（合わせて6経路、この検体で全部を裏取りする）。"""
    _mock_translation(monkeypatch, plan)
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS, [("J-1", "甲", 50000)])
    rc, out = _run(folder, "何かを抜き出して", capsys=capsys)
    assert rc == 3, out
    assert expect_snippet in out, out
