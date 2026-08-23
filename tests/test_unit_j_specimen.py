"""単位J 端から端の検体: 「同じ依頼を2回実行すると同じ列がもう1本できる」事故が
非対話で exit 7 になり、原本に重複列ができないことを、ailine.main を実際に叩いて確かめる。

★ 発端（盲検 operator 査定・実際に起きた再現手順）: title_rows.xlsx に「売上から原価を
引いた利益の列を作って」を実行 → F 列に「売上-原価」列ができる。同じ依頼をもう一度実行
→ G 列に同じ見出し・同じ値の列がもう一つでき、警告ゼロで ✓ 機械検証済み まで出た。
「反映されたか不安でもう一回実行」は事務職の最もありがちな操作。

★ 小道具は test_golden_transcripts.py の _isolate/_run_main をそのまま流用する
（test_unit_h_specimens.py と同じ作法・小道具を二重管理しない）。
"""
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _translate(monkeypatch, op, args):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": op, "args": args})


def _noninteractive(monkeypatch):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)


def _write_profit_column():
    """basrun_apply の差し替え: シート右端の新規列に「売上-原価」を書く（1回目は列D・
       2回目は列D が既に埋まっているので列E ── 実際の LibreOffice/codegen が右端の
       空列を選ぶ挙動を、右端に足すだけの単純なコードで模する）。"""
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value="売上-原価")
        ws.cell(row=2, column=col, value=200)
        ws.cell(row=3, column=col, value=300)
        wb.save(out_book)
        return True, None, "ok"
    return fake


def test_unit_j_repeated_request_is_gated_and_original_stays_clean(tmp_path, monkeypatch, capsys):
    """1回目: 新規列(D)ができて exit 0。
       2回目（同じ依頼をもう一度）: 見出しも値も同一の列を作ろうとして前提が破れ、
       非対話で exit 7・原本に重複列(E)ができない。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "売上", "原価"], ["a", 300, 100], ["b", 500, 200]])
    _translate(monkeypatch, "COMPUTE_COLUMN", {"operands": ["売上", "原価"], "operator": "-"})
    monkeypatch.setattr(ailine, "basrun_apply", _write_profit_column())

    # ★ --values: 式でなく値ベタ書きの事後条件にする（fake basrun_apply が単純な値を
    #   書くだけで済むよう、検体の焦点(単位J の前提検査)から式検証の複雑さを外す）。
    # 1回目: 正常に列D「売上-原価」ができる
    rc1, out1 = _run_main(["run", str(book), "売上から原価を引いた利益の列を作って", "--values"], capsys)
    assert rc1 == 0, f"rc={rc1}\n{out1}"
    wb1 = openpyxl.load_workbook(book)
    assert wb1.active.cell(row=1, column=4).value == "売上-原価"
    assert wb1.active.cell(row=2, column=4).value == 200

    # 2回目: 同じ依頼をもう一度。★ 事故の再現 ── 見出しも値も同一の列(E)を作ろうとする。
    _noninteractive(monkeypatch)
    rc2, out2 = _run_main(["run", str(book), "売上から原価を引いた利益の列を作って", "--values"], capsys)
    assert rc2 == 7, f"rc={rc2}\n{out2}"
    assert "見出しも値も同一の列を作りました" in out2, out2

    # 原本は無変更（重複列(E)ができていない）── 破壊の関所が止めたので atomic_replace は
    # 起きていないはず。
    # ★ ws.cell(row,col) は読むだけでもそのセルを内部に materialize する（openpyxl の仕様）
    #   ので、max_column を**先に**見る（後にすると cell() 呼び出し自体が列を作ってしまい
    #   誤検知になる）。
    wb2 = openpyxl.load_workbook(book)
    assert wb2.active.max_column == 4, f"原本の列数が変わっている: {wb2.active.max_column}"
    assert wb2.active.cell(row=1, column=5).value is None, "原本に重複列ができている"
