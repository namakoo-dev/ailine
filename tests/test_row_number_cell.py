# 行番号で 1 セルを指す ── 2026-08-28。Namakoo「指示文が通らない」
#
# ★★ 実測した壊れ方（今日いちばん悪い）: 「7行目の担当を『佐藤』にして」で
#   **担当列が全行『佐藤』になり、✓ が出た**。
#   一括書換の契約としては ✓ は正しい ── 宣言（列全体）と実体は一致していた。
#   見ていなかったのは **依頼**。三項（依頼・宣言・実体）のうち 1 項を代用すると
#   検算は恒真になる、というこの repo の型がそのまま出た。
#
# ★ 二次の壊れ方: 「7 行F列に『佐藤』を追加」は一段目が 3/3 で **ADD_ROW**。
#   読み直しの門を **op 名** で作っていたので素通りした ──
#   op 名の数え上げが破れるのは今日 3 度目なので、門は **宣言**（OP_WRITE_TARGET が
#   書いている「書く領域」）で作る。新しい op は宣言を書けば自動で門に載る。
#
# 契約:
#   ① 依頼文の行番号を機械が読む（第/目/全角も）
#   ② A1 の列名（F・F列）を実表の見出しへ解決する
#   ③ 行番号と名前が食い違ったら**書かない**
#   ④ 事後条件は行番号を正とする（同名が 2 行ある表でも決まる）
#   ★★ 恒真殺し: 行を指した依頼が**列全体を書いて ✓** になる経路が無いこと

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

ROWS = [["取引先", "件数", "担当"], ["丸和物流", 10, "田中"], ["ヤマノ食品", 20, "鈴木"],
         ["北斗精機", 30, ""], ["ヤマノ食品", 40, ""]]


def _book(tmp_path, rows=None, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    for r in (rows or ROWS):
        ws.append(r)
    wb.save(p)
    return p


def _meta(path):
    return {"sheets": ["請求"], "headers": {"請求": ["取引先", "件数", "担当"]},
            "header_rows": {"請求": 1}, "path": str(path)}


# --- ① 行番号を読む -------------------------------------------------------------------

@pytest.mark.parametrize("task,want", [
    ("7行目の担当を「佐藤」にして", 7),
    ("7 行F列に「佐藤」を追加", 7),
    ("第12行を消して", 12),
    ("１２行目の担当", 12),
    ("担当を全部「佐藤」にして", None),
    ("金額で降順に並べ替えて", None),
])
def test_the_row_number_is_read_from_the_request(task, want):
    assert ailine.task_names_a_row_number(task) == want


# --- ② A1 の列名 ---------------------------------------------------------------------

@pytest.mark.parametrize("ref,guessed", [("C", True), ("C列", True), ("c", True),
                                          ("担当", False)])
def test_a1_letters_resolve_to_a_real_header(ref, guessed):
    """★ 2 つ目の返り値は「推定だったか」── A1 の列名は推定なので画面に「(推定)」が出る。"""
    name, is_guess, note = ailine.resolve_col_ref(ref, ["取引先", "件数", "担当"])
    assert (name, is_guess, note) == ("担当", guessed, None)


@pytest.mark.parametrize("ref", ["Z", "URL", "ZZ列"])
def test_a_letter_beyond_the_table_is_refused(ref):
    _name, ok, _note = ailine.resolve_col_ref(ref, ["取引先", "件数", "担当"])
    assert not ok, ref


# --- ③ 食い違いは書かない -------------------------------------------------------------

def test_a_row_number_that_disagrees_with_the_name_is_refused(tmp_path):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "SET_CELL_VALUE", {"row": "丸和物流", "col": "担当", "value": "佐藤",
                            "row_number": 3}, _meta(_book(tmp_path)))
    assert not ok and "食い違" in err, err


def test_a_row_number_wins_and_a_numeric_row_is_not_a_name(tmp_path):
    """★ 実測: 第二段は row に**行番号そのもの**（"7"）を入れてくる回がある。
       それを名前として突き合わせると、正しい依頼が全部断られる。"""
    ok, r, _i, err = ailine.verify_dsl_args(
        "SET_CELL_VALUE", {"row": "3", "col": "担当", "value": "佐藤", "row_number": 3},
        _meta(_book(tmp_path)), task="3行目の担当を「佐藤」にして")
    assert ok, err
    assert r["_row_index"] == 3, r.get("_row_index")


def test_a_row_number_outside_the_table_is_refused(tmp_path):
    ok, _r, _i, err = ailine.verify_dsl_args(
        "SET_CELL_VALUE", {"col": "担当", "value": "佐藤", "row_number": 99},
        _meta(_book(tmp_path)))
    assert not ok and "範囲外" in err, err


# --- ④ 事後条件は行番号を正とする -----------------------------------------------------

def _after(tmp_path, changes, name="out.xlsx"):
    p = _book(tmp_path, name=name)
    wb = openpyxl.load_workbook(p)
    ws = wb["請求"]
    for (r, c), v in changes.items():
        ws.cell(r, c, v)
    wb.save(p)
    return p


_ARGS = {"row": "ヤマノ食品", "col": "担当", "value": "佐藤", "row_number": 5,
          "_row_index": 5, "_target_sheet": "請求"}


def test_the_right_single_cell_passes(tmp_path):
    src = _book(tmp_path, name="src.xlsx")
    out = _after(tmp_path, {(5, 3): "佐藤"})
    status, reason = ailine.check_set_cell_value(out, dict(_ARGS), source_book=src)
    assert status == "pass", reason


def test_a_duplicated_name_no_longer_blocks_the_check(tmp_path):
    """★ ヤマノ食品は 2 行ある ── 名前で探し直す実装だと『1 件に決まらない』で落ちる。
       依頼は行番号だったのだから、番号を正とする。"""
    src = _book(tmp_path, name="src2.xlsx")
    out = _after(tmp_path, {(5, 3): "佐藤"}, name="out2.xlsx")
    status, reason = ailine.check_set_cell_value(out, dict(_ARGS), source_book=src)
    assert status == "pass", reason
    assert "1 件に決まらない" not in reason


def test_writing_the_whole_column_fails_the_check(tmp_path):
    """★★ 恒真殺し: 今日の事故そのもの。1 セルの宣言で列を潰したら落ちること。"""
    src = _book(tmp_path, name="src3.xlsx")
    out = _after(tmp_path, {(2, 3): "佐藤", (3, 3): "佐藤", (4, 3): "佐藤", (5, 3): "佐藤"},
                  name="out3.xlsx")
    status, reason = ailine.check_set_cell_value(out, dict(_ARGS), source_book=src)
    assert status == "fail", reason


# --- ★★ 門は宣言で作る（op 名を数え上げない）------------------------------------------

def test_the_gate_is_declaration_driven_not_a_list_of_op_names():
    """★ 「1 セルより広く書く」は各 op が OP_WRITE_TARGET に自分で書いている。
       ADD_ROW（実測でここに落ちてきた op）も、名前を足さずに門に載る。"""
    assert ailine.plan_writes_beyond_one_cell([{"op": "ADD_ROW"}])
    assert ailine.plan_writes_beyond_one_cell([{"op": "SET_COLUMN_VALUE"}])
    # 自分で「1 セル」と宣言している op は読み直さない（無駄な往復をしない）
    assert not ailine.plan_writes_beyond_one_cell([{"op": "SET_CELL_VALUE"}])
    # 書式・並べ替え・削除は値を書かない
    assert not ailine.plan_writes_beyond_one_cell([{"op": "BOLD"}])
    assert not ailine.plan_writes_beyond_one_cell([{"op": "SORT"}])


def test_every_value_writing_op_is_covered_by_the_gate():
    """★ 変異試験のかわり: 既存列/新規列/末尾行に書くと宣言した op が
       1 つでも門から漏れていたら落ちる（宣言と門が別実装でずれないこと）。"""
    wide = {ailine.WRITE_EXISTING_COLUMN, ailine.WRITE_NEW_COLUMN,
             ailine.WRITE_NEW_ROW_AT_END, ailine.WRITE_ROW_SHIFT}
    for op, wt in ailine.OP_WRITE_TARGET.items():
        expect = bool(set(wt.writes) & wide) and ailine.WRITE_SINGLE_CELL not in wt.writes
        assert ailine.plan_writes_beyond_one_cell([{"op": op}]) is expect, op


def test_a_request_without_a_quoted_value_is_never_stolen():
    """★ 恒真殺し: 「みかんの利益を計算して」を 1 セル書換に化けさせない
       （行は名指しされているが、書く値はどこにも書かれていない）。"""
    assert ailine.task_quotes_a_value("みかんの利益を計算して") is None
    assert ailine.task_quotes_a_value("7行目の担当を「佐藤」にして") == "佐藤"


def test_a_conditional_request_is_never_stolen():
    """★ 「原価が500以上の…に『◎』を付けて」は 1 行ではない ── 比較語で止める。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index("if not _reread_done and plan_writes_beyond_one_cell(plan):")
    assert "extract_cmp_from_task(a.task) is None" in src[i:i + 900], \
        "1 セルの読み直しが、条件つき書換の依頼まで拾う形になっている"


def test_pointing_at_a_row_never_falls_back_to_the_whole_column():
    """★★ この repo で一番効いた形の再演: 落とせなかったら**断る**。
       ここを『列全体を書く』に戻すと、依頼を見ない ✓ が復活する。"""
    src = (REPO / "src" / "ailine" / "__init__.py").read_text(encoding="utf-8")
    i = src.index("if _points and not _one_cell:")
    seg = src[i:i + 500]
    assert "列全体は書き換えません" in seg and "return 3" in seg, seg[:300]
