# 第三波 #9 ── openpyxl の往復で消える図形（描かれた角印・社判）。
#
# ★ この repo の中で実測した事実（推測ではない）:
#     画像（xdr:pic / xl/media）                    → load→save で **残る**
#     画像でない図形（sp・テキストボックス・オートシェイプ）→ load→save で **消える**
#   `ailine export-pdf` は指定シートの抽出とページ設定のため原本を openpyxl で
#   一度書き直すので、描かれた角印が PDF から消える。出来上がりは完成品に見える。
#
# 契約:
#   ① 図形が在れば、変換の**前に**名指しする（消えたものは差分に出ない）
#   ② 画像だけの冊では 1 文字も警告しない（誤爆防止）
#   ③ 検出は zip の drawing XML を直接読む ── openpyxl に聞くと、openpyxl が
#      読めない図形は最初から見えないので恒真になる（別実装で測る）

import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core import pdf_export  # noqa: E402

NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
_SHAPE = ('<oneCellAnchor><from><col>5</col><colOff>0</colOff><row>1</row><rowOff>0</rowOff></from>'
          '<ext cx="500000" cy="500000"/>'
          '<sp macro="" textlink=""><nvSpPr><cNvPr id="99" name="KAKUIN"/><cNvSpPr/></nvSpPr>'
          f'<spPr><a:prstGeom xmlns:a="{NS}" prst="ellipse"/></spPr></sp><clientData/></oneCellAnchor>')


def _png(w=8, h=8):
    """8x8 の赤い PNG を手で作る（Pillow を使わない ── CI に Pillow は無い）。"""
    import struct, zlib
    def chunk(tag, data):
        body = tag + data
        return struct.pack(">I", len(data)) + body + struct.pack(">I", zlib.crc32(body) & 0xffffffff)
    raw = b"".join(bytes([0]) + bytes((200, 0, 0)) * w for _ in range(h))
    return (bytes([137, 80, 78, 71, 13, 10, 26, 10])
            + chunk(b"IHDR", struct.pack(">IIBBBBB", w, h, 8, 2, 0, 0, 0))
            + chunk(b"IDAT", zlib.compress(raw)) + chunk(b"IEND", b""))


# ★ 2026-08-24 の CI 赤で学んだ: 初版は openpyxl.drawing.image.Image を使っていて
#   **Pillow を要求**した。手元には在り CI には無い ──「居るから見えない」の 3 度目。
#   ここでは drawing 一式を手で組んで zip に入れる（依存ゼロ・openpyxl の書き手にも依存
#   しないので、検出器を openpyxl とは別実装で測るという狙いにも合う）。

_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

_PIC = ('<oneCellAnchor><from><col>3</col><colOff>0</colOff><row>1</row><rowOff>0</rowOff></from>'
        '<ext cx="76200" cy="76200"/>'
        '<pic><nvPicPr><cNvPr id="1" name="SEAL_IMAGE"/><cNvPicPr/></nvPicPr>'
        f'<blipFill><a:blip xmlns:a="{NS}" xmlns:r="{_REL}" r:embed="rId1"/><a:stretch xmlns:a="{NS}">'
        '<a:fillRect/></a:stretch></blipFill>'
        f'<spPr><a:prstGeom xmlns:a="{NS}" prst="rect"/></spPr></pic><clientData/></oneCellAnchor>')


def _write_drawing(path, body):
    """既存の .xlsx に drawing 一式（media + drawing.xml + rels + シートからの参照）を足す。"""
    import os
    tmp = str(path) + ".t"
    drawing = f'<wsDr xmlns="{_XDR}">{body}</wsDr>'
    rels = (f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{_REL}/image" Target="../media/image1.png"/>'
            f'</Relationships>')
    sheet_rels = (f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                  f'<Relationship Id="rIdD" Type="{_REL}/drawing" Target="../drawings/drawing1.xml"/>'
                  f'</Relationships>')
    with zipfile.ZipFile(path) as zi, zipfile.ZipFile(tmp, "w") as zo:
        for item in zi.infolist():
            data = zi.read(item.filename)
            if item.filename == "[Content_Types].xml":
                text = data.decode("utf-8").replace(
                    "</Types>",
                    '<Default Extension="png" ContentType="image/png"/>'
                    '<Override PartName="/xl/drawings/drawing1.xml" ContentType='
                    '"application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
                data = text.encode("utf-8")
            elif item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8").replace(
                    "</worksheet>", f'<drawing xmlns:r="{_REL}" r:id="rIdD"/></worksheet>')
                data = text.encode("utf-8")
            zo.writestr(item, data)
        zo.writestr("xl/media/image1.png", _png())
        zo.writestr("xl/drawings/drawing1.xml", drawing)
        zo.writestr("xl/drawings/_rels/drawing1.xml.rels", rels)
        zo.writestr("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels)
    os.replace(tmp, path)
    return path


def _book_with_image(tmp_path):
    """画像だけの冊（写真として貼られた印 ── これは消えない）。"""
    p = tmp_path / "img.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "請求書"
    wb.save(p)
    return _write_drawing(p, _PIC)


def _book_with_shape(tmp_path):
    """画像 + 画像でない図形（描かれた角印 ── これが消える）。"""
    p = tmp_path / "shape.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "請求書"
    wb.save(p)
    _write_drawing(p, _PIC + _SHAPE)
    # ★ 治具が本当に仕込めたかを先に確かめる（初版は仕込めておらず、自分の壊れた
    #   治具を測って「消えない」と誤結論しかけた）。
    with zipfile.ZipFile(p) as z:
        assert "KAKUIN" in z.read("xl/drawings/drawing1.xml").decode("utf-8"), "治具が仕込めていない"
    return p


# --- 前提そのものの測定（この試験が守る事実）------------------------------------------

def test_openpyxl_roundtrip_keeps_images_but_drops_shapes(tmp_path):
    """★ 前提の凍結: openpyxl の版が変わってこの事実が変われば、ここが最初に鳴る。"""
    if not pdf_export.pillow_available():
        pytest.skip("Pillow が無い環境 ── この向き（画像は残る）は測れない"
                     "（測れない回は skip と書く。緑にしない）")
    shaped = _book_with_shape(tmp_path)
    out = tmp_path / "after.xlsx"
    openpyxl.load_workbook(shaped).save(out)
    with zipfile.ZipFile(out) as z:
        names = z.namelist()
        assert any("media" in n for n in names), "画像まで落ちた（前提が変わった）"
        drawing = z.read("xl/drawings/drawing1.xml").decode("utf-8") if \
            "xl/drawings/drawing1.xml" in names else ""
    assert "KAKUIN" not in drawing, \
        "図形が残った ── 前提が変わったので、警告そのものを見直すこと"


# --- ①② 検出器 ----------------------------------------------------------------------

def test_detects_the_vanishing_shape(tmp_path):
    """★ 環境に依らず決まるよう with_pillow を明示で渡す（実測の再現性）。"""
    got = pdf_export.vanishing_shapes(_book_with_shape(tmp_path), with_pillow=True)
    assert got == ["KAKUIN"]


def test_image_only_book_is_silent_when_pillow_is_present(tmp_path):
    got = pdf_export.vanishing_shapes(_book_with_image(tmp_path), with_pillow=True)
    assert got == [], "画像しか無い冊に誤爆した（写真の印は消えないのに消えると言った）"


def test_image_also_vanishes_without_pillow(tmp_path):
    """★ Pillow が無ければ写真の印も消える ── 黙っていると嘘の安心を与える。

    2026-08-24、CI と同じ素の環境で走らせて初めて出た。ailine が宣言している
    依存は openpyxl だけなので、買い手の環境に Pillow が在る保証はない。
    """
    got = pdf_export.vanishing_shapes(_book_with_image(tmp_path), with_pillow=False)
    assert got, "Pillow 不在なのに画像を安全だと言った"


def test_warning_wording_depends_on_pillow():
    joined = lambda flag: chr(10).join(
        pdf_export.vanishing_shapes_warning(["K"], with_pillow=flag))
    assert "写真として貼られた印は消えません" in joined(True)
    assert "写真として貼られた印は消えません" not in joined(False), (
        "Pillow が無い環境で『写真の印は消えない』と嘘をついた")
    assert "pip install Pillow" in joined(False), "逃げ道を出していない"


def test_unreadable_file_is_silent(tmp_path):
    p = tmp_path / "broken.xlsx"
    p.write_bytes(b"not a zip")
    assert pdf_export.vanishing_shapes(p, with_pillow=True) == []


def test_warning_names_the_shape_and_offers_a_way_out(tmp_path):
    lines = pdf_export.vanishing_shapes_warning(["KAKUIN"])
    text = "\n".join(lines)
    assert "KAKUIN" in text and "消えます" in text
    assert "LibreOffice" in text or "Excel" in text, f"逃げ道が無い: {text}"
    assert pdf_export.vanishing_shapes_warning([]) == []


# --- ③ 別実装であること -----------------------------------------------------------------

def test_detection_does_not_ask_openpyxl():
    src = (REPO / "src" / "ailine_core" / "pdf_export.py").read_text(encoding="utf-8")
    body = src[src.index("def vanishing_shapes("):]
    # 物差しの訂正（封印者ナギ）: 初版は "openpyxl" の文字列を見ていて docstring の
    # 説明文まで掴んだ（測定器が粗い）。意図は「openpyxl に聞いていない」なので
    # 呼び出しの形で測る。assert の意図は変えていない。
    assert "load_workbook" not in body.split("def vanishing_shapes_warning")[0], \
        "openpyxl に聞いている（読めない図形は見えないので恒真になる）"
