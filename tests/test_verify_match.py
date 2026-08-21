"""_verify_match（照合出力の単独検算）の検体。
   ★ 実装前に凍結した赤い検体。信用の条件⑥「検算側を単独で再実行できること」の M3 版 ──
   出力ブック + 元 2 冊だけで、キー勘定（件数・Σ・差額）を独立に検算し直す。

   CLI: ailine verify <照合出力.xlsx> <元A.xlsx> <元B.xlsx>（ファイル 2 冊形）。
   既存の <出力> <フォルダ> 形（stack/extract）は不変。"""
import datetime
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

NYUKIN_HDRS = ["取引日", "振込人名義", "お預り金額", "摘要"]
SEIKYU_HDRS = ["請求日", "取引先名", "請求金額", "請求番号"]
TASK = "振込人名義と取引先名をキーに、お預り金額と請求金額を突き合わせて"


def _book(path, title, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _made(tmp_path, capsys):
    a = tmp_path / "in" / "入金.xlsx"
    b = tmp_path / "in" / "請求.xlsx"
    _book(a, "入金明細", NYUKIN_HDRS, [
        (datetime.date(2026, 7, 31), "甲社", 220000, ""),
        (datetime.date(2026, 7, 31), "乙社", 109340, "手数料引き"),
    ])
    _book(b, "請求一覧", SEIKYU_HDRS, [
        (datetime.date(2026, 7, 10), "甲社", 220000, "INV-101"),
        (datetime.date(2026, 7, 12), "乙社", 110000, "INV-102"),
    ])
    rc = ailine.main(["run", str(a), str(b), TASK])
    capsys.readouterr()
    assert rc == 0
    out = max((tmp_path / "in").glob("*照合*.xlsx"), key=lambda p: p.stat().st_mtime)
    return a, b, out


def _verify(out, a, b):
    return subprocess.run(
        [sys.executable, str(REPO / "ailine.py"), "verify", str(out), str(a), str(b)],
        capture_output=True, text=True, timeout=120, encoding="utf-8")


def test_verify_match_passes_with_both_side_numbers(tmp_path, capsys):
    a, b, out = _made(tmp_path, capsys)
    p = _verify(out, a, b)
    assert p.returncode == 0, p.stdout + p.stderr[-300:]
    assert "329340" in p.stdout.replace(",", ""), f"A 側 Σ（220000+109340）の表示が無い:\n{p.stdout}"
    assert "330000" in p.stdout.replace(",", ""), "B 側 Σ の表示が無い"


def test_verify_match_catches_tampered_sum_with_key_named(tmp_path, capsys):
    """照合シートの a_sum を後から編集 → どのキーがいくつ→いくつ、まで名指しで exit 5。"""
    a, b, out = _made(tmp_path, capsys)
    wb = openpyxl.load_workbook(out)
    ws = wb["照合"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "甲社":
            row[2].value = 999999   # A側 合計
            break
    wb.save(out)
    p = _verify(out, a, b)
    assert p.returncode == 5, f"改竄が素通り (exit={p.returncode}):\n{p.stdout}"
    assert "甲社" in p.stdout and "999999" in p.stdout and "220000" in p.stdout, \
        f"キーの名指しと両側の数字が無い:\n{p.stdout}"


def test_verify_match_catches_tampered_diff_cell(tmp_path, capsys):
    a, b, out = _made(tmp_path, capsys)
    wb = openpyxl.load_workbook(out)
    ws = wb["照合"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "乙社":
            row[5].value = 0        # 差額 −660 を 0 に偽装
            break
    wb.save(out)
    p = _verify(out, a, b)
    assert p.returncode == 5
    assert "乙社" in p.stdout and "660" in p.stdout, f"差額の偽装が名指しされない:\n{p.stdout}"


def test_verify_match_refuses_wrong_source_books_not_silent_pass(tmp_path, capsys):
    """★ 空虚な合格の禁止: 違う元 2 冊を渡されたら、一致しない事実が名指しで出て exit 5
       （0 件照合で合格を名乗らない）。"""
    a, b, out = _made(tmp_path, capsys)
    c = tmp_path / "in" / "別の入金.xlsx"
    _book(c, "入金明細", NYUKIN_HDRS, [(datetime.date(2026, 7, 1), "丙社", 1, "")])
    p = _verify(out, c, b)
    assert p.returncode == 5, f"違う元帳で合格した (exit={p.returncode}):\n{p.stdout}"


def test_verify_match_is_read_only_on_all_three(tmp_path, capsys):
    import hashlib
    a, b, out = _made(tmp_path, capsys)
    before = {p.name: hashlib.sha256(p.read_bytes()).hexdigest() for p in (a, b, out)}
    assert _verify(out, a, b).returncode == 0
    after = {p.name: hashlib.sha256(p.read_bytes()).hexdigest() for p in (a, b, out)}
    assert before == after


def test_missing_out_file_says_not_found_not_unmarked(tmp_path, capsys):
    """★ 実弾検分（2026-08-21 16:5x）: 存在しない出力パスに「印がありません」と誤診した。
       無いなら無いと言う（誤診は次の手を間違わせる）。exit は 4 のまま。"""
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _book(a, "入金明細", NYUKIN_HDRS, [(datetime.date(2026, 7, 1), "甲社", 1, "")])
    _book(b, "請求一覧", SEIKYU_HDRS, [(datetime.date(2026, 7, 1), "甲社", 1, "x")])
    p = _verify(tmp_path / "居ない.xlsx", a, b)
    assert p.returncode == 4
    assert "見つかりません" in p.stdout, f"誤診（印がありません）:\n{p.stdout}"
    assert "印がありません" not in p.stdout
