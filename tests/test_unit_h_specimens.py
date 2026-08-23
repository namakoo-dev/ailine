"""単位 H の事前登録検体（T1〜T5 / F1〜F5）。★ 実装の前に commit して「先に赤を見る」。

★ なぜ実装より先に書くか（査定 2026-08-18 の指示）:
   「実装後に検体を作ると、実装が通る検体を作ってしまう。」
   単位 C の commit が正直に記録したとおり、凍結コーパスが誤爆を 1 件も含んでいなかったため
   CI は誤爆を一度も赤にできなかった。同じ穴を繰り返さないための先出しコーパス。

★ GO/NO-GO はこの 2 数字だけ（査定が事前登録した）:
     F1〜F5 の ★ 合計本数 = 0 本
     T1〜T5 の 止まった件数 = 5/5

★ 現在赤いものは xfail(strict=True) で凍結してある。単位 H が直したら XPASS になって
   **必ず落ちる** ので、直したのに印を外し忘れることができない（黙って緑にならない）。

★ T4（PIVOT）と T5（『集計』自身を対象に集計）は査定が「未測定・落ちる可能性を織り込む」と
   明記した検体。実測の結果はこのファイルの xfail の付き方がそのまま示す。
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402 — 小道具を二重管理しない（test_sum_identity.py と同じ流用）

# 一度きりの既定変更の告知は ★ で始まるが警告ではない。★ の本数を数える時は必ず除く。
_ONE_TIME_NOTICE = "このバージョンから、既定で原本に直接反映します"


def _warn_stars(out: str) -> list:
    """警告としての ★ 行だけを返す（一度きりの告知を除く）。"""
    return [ln.strip() for ln in out.splitlines()
            if ln.strip().startswith("★") and _ONE_TIME_NOTICE not in ln]


def _translate(monkeypatch, op, args):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"op": op, "args": args})


def _noninteractive(monkeypatch):
    def _raise_eof(prompt=""):
        raise EOFError()
    monkeypatch.setattr("builtins.input", _raise_eof)


def _apply(monkeypatch, fn):
    monkeypatch.setattr(ailine, "basrun_apply", fn)


def _nop_apply(out_book, code, workdir, helper_files=(), timeout=None):
    return True, None, "ok"


# ---------------------------------------------------------------- 検体ブック

def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _book_with_summary(tmp_path, summary_rows, sheet_name="集計", data_rows=None,
                       data_sheet="Sheet"):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data_sheet
    for r in (data_rows or [["部門", "金額"], ["営業", 100], ["営業", 200], ["開発", 300]]):
        ws.append(r)
    s = wb.create_sheet(sheet_name)
    for r in summary_rows:
        s.append(r)
    wb.save(p)
    return p


def _rewrite_sheet(sheet_name, rows):
    """basrun_apply の差し替え: 宣言どおり sheet を作り直す（removeByName → 再作成を模す）。"""
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        s = wb.create_sheet(sheet_name)
        for r in rows:
            s.append(r)
        wb.save(out_book)
        return True, None, "ok"
    return fake


# ★ APPEND_TOTAL の事後条件は「挿入耐性 SUM 型」の式(ailine.py:_APPEND_TOTAL_FORMULA_RE)
#   **かつ** ラベル一致 **かつ** キャッシュ値一致 の三点を要求する
#   (ailine.py:check_append_total 3380-3454)。式の形自体は当初の当て推量で合っていたが、
#   openpyxl で書いた数式にはキャッシュ値(<v>)が無い（LibreOffice を経由しないため）ので
#   キャッシュ値一致だけが必ず落ちていた。tests/test_sum_identity.py が同じ壁を
#   _inject_formula_cache（xl/worksheets/sheetN.xml へ <v> を直接注ぐ, test_ailine.py:4711）
#   で越えているので、その小道具をそのまま流用する。
_TOTAL_FORMULA = "=SUM(B2:INDEX(B:B,ROW()-1))"


def _append_total_row(label="合計"):
    """basrun_apply の差し替え: 1 枚目の末尾に 合計行 を足す（式で書く・キャッシュ値も注入）。"""
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        r = ws.max_row + 1
        total = sum(v for v in (ws.cell(row=rr, column=2).value for rr in range(2, r))
                    if isinstance(v, (int, float)))
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=_TOTAL_FORMULA)
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {f"B{r}": total})
        return True, None, "ok"
    return fake


def _overwrite_last_row_with_formula():
    """basrun_apply の差し替え: 既存の最終行（＝既存の合計行）を SUM 式で上書きする。
       ★ 盲検査定の致命1の実物: 116600 が =SUM(...) に置き換わって 106000 になった。
       ★ 単位H: 致命1のバグの実物そのものを再現する ── 末尾に新しい行を足すはずが
       既存行(ラベルも値も)を潰す。ラベルと値(キャッシュ)を書き込むのは、これが
       事後条件(三点)を満たしてしまう＝「反映は正しく通った」上で単位F/G の
       破壊の関所(_check_new_row_at_end)が鳴る、という T2 が測りたい経路そのものだから
       （postcondition が先に fail すると関所まで到達しない）。"""
    def fake(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        r = ws.max_row
        total = sum(v for v in (ws.cell(row=rr, column=2).value for rr in range(2, r))
                    if isinstance(v, (int, float)))
        ws.cell(row=r, column=1, value="合計")
        ws.cell(row=r, column=2, value=_TOTAL_FORMULA)
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {f"B{r}": total})
        return True, None, "ok"
    return fake


# ================================================================ 真陽性 T1〜T5
# 期待: 止まる（exit 0 で無言は不可）

def test_t1_handwritten_total_row_with_label(tmp_path, monkeypatch, capsys):
    """T1: 手入力の合計行（A 列にラベル『合計』有）があるのに『合計を出して』。
       期待: 関所 or 解釈が見える表示。★ exit 0 で無言は不可。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["合計", 500]])
    _translate(monkeypatch, "APPEND_TOTAL", {"col": "金額"})
    _apply(monkeypatch, _append_total_row())
    _noninteractive(monkeypatch)
    rc, out = _run_main(["run", str(book), "合計を出して"], capsys)
    assert not (rc == 0 and not _warn_stars(out)), f"exit 0 で無言だった:\n{out}"


def test_t2_existing_total_row_blank_label(tmp_path, monkeypatch, capsys):
    """T2: 既存の合計行（A 列が空）を上書きする。期待: exit 7・既存値 116600 が保存されている。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], [None, 116600]])
    _translate(monkeypatch, "APPEND_TOTAL", {"col": "金額"})
    _apply(monkeypatch, _overwrite_last_row_with_formula())
    _noninteractive(monkeypatch)
    rc, out = _run_main(["run", str(book), "合計を出して"], capsys)
    assert rc == 7, f"rc={rc}\n{out}"
    assert openpyxl.load_workbook(book).active.cell(row=4, column=2).value == 116600


def test_t3_unrelated_summary_sheet_aggregate(tmp_path, monkeypatch, capsys):
    """T3: 無関係な手作りの『集計』（年度/予算）に AGGREGATE。期待: exit 7・元データ保存。
       ★ 端から端の凍結は golden 側 (unit_g_declared_sheet_premise_broken) にもある。"""
    _isolate(monkeypatch, tmp_path)
    book = _book_with_summary(tmp_path, [["年度", "予算"], [2025, 5000], [2026, 6000]])
    _translate(monkeypatch, "AGGREGATE", {"group_col": "部門", "value_col": "金額"})
    _apply(monkeypatch, _rewrite_sheet("集計", [["部門", "合計"], ["営業", 300], ["開発", 300]]))
    _noninteractive(monkeypatch)
    rc, out = _run_main(["run", str(book), "部門ごとに金額をまとめて"], capsys)
    assert rc == 7, f"rc={rc}\n{out}"
    assert openpyxl.load_workbook(book)["集計"].cell(row=2, column=1).value == 2025


@pytest.mark.xfail(strict=True, reason="★ 未確定: rc=3(CLARIFY) で止まる。査定自身が「T4 は未測定・落ちる可能性を織り込む」と明記した検体。AGGREGATE の同型(T3)は通るので、PIVOT 経路だけが CLARIFY する理由が本物の発見かもしれない。★ 切り分けてから印を外す")
def test_t4_unrelated_pivot_sheet(tmp_path, monkeypatch, capsys):
    """T4: 同じことを PIVOT で（出力先『ピボット』）。★ 査定が未測定と明記した検体。"""
    _isolate(monkeypatch, tmp_path)
    book = _book_with_summary(tmp_path, [["年度", "予算"], [2025, 5000], [2026, 6000]],
                              sheet_name="ピボット")
    _translate(monkeypatch, "PIVOT", {"group_col": "部門", "value_col": "金額"})
    _apply(monkeypatch, _rewrite_sheet("ピボット", [["部門", "合計"], ["営業", 300], ["開発", 300]]))
    _noninteractive(monkeypatch)
    rc, out = _run_main(["run", str(book), "部門ごとに金額をピボットでまとめて"], capsys)
    assert rc == 7, f"rc={rc}\n{out}"
    assert openpyxl.load_workbook(book)["ピボット"].cell(row=2, column=1).value == 2025


def test_t5_aggregate_targeting_the_summary_sheet_itself(tmp_path, monkeypatch, capsys):
    """T5: 『集計』シート自身を対象に集計する。期待: exit 7 または CLARIFY。
       ★ 査定が未測定と明記した検体。exit 0 で黙って通るのが不可。"""
    _isolate(monkeypatch, tmp_path)
    book = _book_with_summary(tmp_path, [["部門", "金額"], ["営業", 100], ["開発", 200]])
    _translate(monkeypatch, "AGGREGATE", {"group_col": "部門", "value_col": "金額"})
    _apply(monkeypatch, _rewrite_sheet("集計", [["部門", "合計"], ["営業", 100], ["開発", 200]]))
    _noninteractive(monkeypatch)
    rc, out = _run_main(["run", str(book), "集計シートを部門ごとにまとめて", "--sheet", "集計"], capsys)
    assert rc != 0, f"exit 0 で黙って通った:\n{out}"


# ================================================================ 偽陽性 F1〜F5
# 期待: ★ が 1 本も出ない（誤爆の復活を防ぐ）

def test_f1_plain_table_append_total(tmp_path, monkeypatch, capsys):
    """F1: 合計行の無い普通の表 + 合計を出して。期待: 中立表示のみ・★ 0 本・exit 0。
       ★ 単位H: 依頼文に対象列『金額』を明記する（単位E: 依頼文が無言の対象スロットは
       ②UNSPOKEN として ✓ の直後に1文が付く ―― これは誤爆ではなく正しい仕分けなので、
       ★0本を測る F1 では対象を名指しして UNSPOKEN を起こさない。同じ言い方は
       tests/test_sum_identity.py の APPEND_TOTAL 検体（「金額の合計を一番下に出して」）
       にも既にある）。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    _translate(monkeypatch, "APPEND_TOTAL", {"col": "金額"})
    _apply(monkeypatch, _append_total_row())
    rc, out = _run_main(["run", str(book), "金額の合計を出して", "--copy"], capsys)
    assert _warn_stars(out) == [], f"rc={rc} stars={_warn_stars(out)}"
    assert rc == 0, f"rc={rc} stars={_warn_stars(out)}"


def test_f2_single_sheet_new_aggregate(tmp_path, monkeypatch, capsys):
    """F2: 単一シートに初めて『集計』を作る。期待: 新設の中立表示のみ・★ 0 本。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["取引先", "金額"], ["甲", 100], ["乙", 200]])
    _translate(monkeypatch, "AGGREGATE", {"group_col": "取引先", "value_col": "金額"})
    _apply(monkeypatch, _rewrite_sheet("集計", [["取引先", "合計"], ["甲", 100], ["乙", 200]]))
    rc, out = _run_main(["run", str(book), "工事台帳を取引先ごとに集計して", "--copy"], capsys)
    # ★ 凍結バーの文言に忠実に: 「★ …変更されていません」が 0 本・新設の中立表示が出ること。
    #   対象が無言だったことの注記（単位E の②）はこのバーの対象外（俺の最初の判定はバーより厳しかった）。
    assert [s for s in _warn_stars(out) if "変更されていません" in s] == [], _warn_stars(out)
    assert any("新規シート『集計』の作成は意図どおりです" in ln for ln in out.splitlines()), out[-400:]


def test_f3_sheet_name_column_name_conflict(tmp_path, monkeypatch, capsys):
    """F3: シート名と列名が衝突するブック（誤爆#3 の検体）。期待: ★ 0 本。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in [["商品", "売上"], ["a", 300], ["b", 200]]:
        ws.append(r)
    wb.create_sheet("金額")
    wb.save(p)
    _translate(monkeypatch, "SORT", {"col": "売上", "order": "desc"})
    _apply(monkeypatch, _nop_apply)
    rc, out = _run_main(["run", str(p), "売上で降順に並べ替えて", "--copy"], capsys)
    assert _warn_stars(out) == [], f"rc={rc} stars={_warn_stars(out)}"


def test_f4_second_aggregate_on_same_book(tmp_path, monkeypatch, capsys):
    """★★ F4: AGGREGATE を同じブックで 2 回連続（前回の『集計』は ailine 自身の出力で、
       その後 元データが増えた状態）。期待: 2 回目も exit 0・★ 0 本。
       ★ 実測 2026-08-19: 単位 F/G だけの状態では exit 7 / ★ 1 本 = 赤。単位 H が直す対象。"""
    _isolate(monkeypatch, tmp_path)
    book = _book_with_summary(
        # ★ 前回の出力は SummaryTable の 本物の署名 で作る（helpers/AiLineHelpers.bas:374-375）
        #   A1=分類列名 / B1="合計 - 集計列名"。ここを想像で書くと H が効かず別の理由で赤くなる。
        tmp_path, [["部門", "合計 - 金額"], ["営業", 300], ["開発", 300]],
        data_rows=[["部門", "金額"], ["営業", 100], ["営業", 200], ["営業", 200],
                   ["開発", 300], ["開発", 400]],
        data_sheet="工事台帳")
    _translate(monkeypatch, "AGGREGATE", {"group_col": "部門", "value_col": "金額"})
    _apply(monkeypatch, _rewrite_sheet("集計", [["部門", "合計 - 金額"], ["営業", 500], ["開発", 700]]))
    _noninteractive(monkeypatch)
    # ★ --sheet を明示する: 2 シートあると対象が無言で機械決定され、単位E の②注記（★ 1 本）が
    #   出る。それは誤爆ではないが、この検体が凍結したい数字（★ 0 本）と混ざる。
    #   ★ 凍結バーを緩めるのでなく、検体から無関係な変数を外す。
    # ★ シートを依頼文で名指しする。--sheet フラグでは単位E の②（依頼文の語と照合できたか）が
    #   「機械決定」のままで注記が出る ── 検体から無関係な変数を外すには 語 で指す必要がある。
    rc, out = _run_main(["run", str(book), "工事台帳を部門ごとに金額をまとめて"], capsys)
    assert _warn_stars(out) == [], f"rc={rc} stars={_warn_stars(out)}"
    assert rc == 0, f"rc={rc} stars={_warn_stars(out)}"


@pytest.mark.parametrize("op,args,task", [
    ("BOLD", {"target": "row:1"}, "見出しを太字にして"),
    # ★ 単位I: CENTER_ALIGN×row: は verify_dsl_args が拒否するようになった（契約文(1743)・
    #   codegen(2588) と合わせた）。fail で静かに終わる（traceback にならない）ので、
    #   BOLD/AUTOFIT と同じ健全系パラメータとして扱える（かつて xfail(strict=True,
    #   raises=ValueError) だったが、traceback が出なくなったので凍結は不要になった）。
    ("CENTER_ALIGN", {"target": "row:1"}, "見出しを中央揃えにして"),
    ("AUTOFIT", {"target": "col:金額"}, "金額の列幅を自動調整して"),
])
def test_f5_format_only_ops_stay_quiet(tmp_path, monkeypatch, capsys, op, args, task):
    """F5: format_only を宣言する op を健全系で 1 回ずつ。期待: ★ 0 本。
       ★ ここで鳴ったら『新設した前提検査が既存 op を鳴かせた』= 新しい発見として別扱い。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200]])
    _translate(monkeypatch, op, args)
    _apply(monkeypatch, _nop_apply)
    rc, out = _run_main(["run", str(book), task, "--copy"], capsys)
    assert _warn_stars(out) == [], f"rc={rc} stars={_warn_stars(out)}"
