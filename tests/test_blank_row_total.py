# 第三波 S4 ── 修正より先に凍結した赤い検体。
#
# 実測した事故: [a,100] / [空白だけの行] / [b,200] の表を stack すると、b が
# **合計行として除外され 200 が黙って消えた**。しかも報告は
# 「Σ金額: 元 100 / 出力 100」── 除外後の値を『元』と呼んでいるので、間違い同士で
# 一致していた（数字は合っているのに事実は間違っている、という形）。
#
# 契約:
#   ① トリガ c（直上が空行）は**裏取り**を要求する ── 算術が閉じる時だけ除外する
#      （a=ラベル語 / b=ラベル空白 は構造そのものが「合計」と言っているので無条件のまま）
#   ② 単一列版と複数列版が**同じ関数**で裏取りする（片配線の禁止）
#   ③ 複数列版は「その行が数字を持つ全ての列」が閉じた時だけ除外（1 列でも説明が
#      つかなければ本物のデータ行として残す＝安全側）
#   ④ Σ の『元』が除外後の値であることを明示する

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
from ailine_core.total_row import (  # noqa: E402
    closes_as_total, split_total_rows, split_total_rows_multi)
from ailine_core.cli_render import sum_line  # noqa: E402


def _rows_single(triples):
    return split_total_rows(triples)


def _reasons(verdict):
    return [(e.row, e.reason) for e in verdict.excluded]


# --- ① 弱い証拠は裏取りを要求する ---------------------------------------------------

def test_real_data_row_after_blank_line_survives_single():
    """和が閉じない ── 本物のデータ行なので消さない（実測した事故そのもの）。"""
    v = _rows_single([(2, "a", 100), (3, None, None), (4, "b", 200)])
    assert _reasons(v) == [], f"本物のデータ行を合計行として消した: {_reasons(v)}"


def test_real_total_row_after_blank_line_is_still_excluded_single():
    """裏返し: 語彙外のラベル（『総額』）でも、和が閉じるなら従来どおり除外する。"""
    v = _rows_single([(2, "a", 100), (3, None, None), (4, "総額", 100)])
    assert _reasons(v) == [(4, "直上空行")], f"本物の合計行を見逃した: {_reasons(v)}"


@pytest.mark.parametrize("label", ["合計", "小計", "計"])
def test_label_word_trigger_is_unconditional(label):
    """誤爆防止: ラベル語は裏取り無しで従来どおり除外（和が閉じなくても）。"""
    v = _rows_single([(2, "a", 100), (3, label, 999)])
    assert _reasons(v) == [(3, "ラベル語")], f"{label} が除外されなくなった: {_reasons(v)}"


def test_blank_label_trigger_is_unconditional():
    v = _rows_single([(2, "a", 100), (3, None, 999)])
    assert _reasons(v) == [(3, "ラベル空白")]


# --- ②③ 複数列版（片配線の番人）-----------------------------------------------------

def test_multi_version_shares_the_same_narrowing():
    """単一列版だけ直して複数列版が素通りするのを禁じる（実測でそうなりかけた）。"""
    rows = [(2, "a", {"金額": 100}), (3, None, {"金額": None}), (4, "b", {"金額": 200})]
    assert _reasons(split_total_rows_multi(rows)) == [], "複数列版が片配線のまま"
    rows_ok = [(2, "a", {"金額": 100}), (3, None, {"金額": None}), (4, "総額", {"金額": 100})]
    assert _reasons(split_total_rows_multi(rows_ok)) == [(4, "直上空行")]


def test_multi_requires_every_numeric_column_to_close():
    """③ 1 列でも閉じなければ残す（安全側 ── 消すほうが取り返しがつかない）。"""
    rows = [(2, "a", {"金額": 100, "数量": 5}), (3, None, {"金額": None, "数量": None}),
            (4, "x", {"金額": 100, "数量": 9})]     # 金額は閉じるが数量は閉じない
    assert _reasons(split_total_rows_multi(rows)) == [], "1 列閉じないのに消した"


def test_closes_as_total_is_one_implementation():
    """② 裏取りの実装が 1 つであること（両版がこの関数を呼ぶ）。"""
    adopted = [(2, 100.0), (3, 200.0)]
    assert closes_as_total(300.0, adopted, None) is True      # 累積和
    assert closes_as_total(200.0, adopted, 2) is True         # 区間和
    assert closes_as_total(250.0, adopted, None) is False


# --- ④ Σ の『元』の開示 --------------------------------------------------------------

def test_sum_line_discloses_that_source_excludes_total_rows():
    both = {"source": 300, "output": 300}
    assert "除いた後" in sum_line("金額", both, excluded_rows=1)
    assert "除いた後" not in sum_line("金額", both, excluded_rows=0), \
        "除外が無いのに注記した（誤爆）"


# --- e2e（実物のブックで消えないこと）-------------------------------------------------

def test_stack_keeps_the_row_after_a_whitespace_row(tmp_path, monkeypatch, capsys):
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from test_golden_transcripts import _isolate, _run_main
    import ailine
    _isolate(monkeypatch, tmp_path)
    folder = tmp_path / "f"
    folder.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [["商品", "金額"], ["a", 100], ["   ", None], ["b", 200]]:
        ws.append(r)
    wb.save(folder / "a.xlsx")
    out = tmp_path / "out.xlsx"
    rc, text = _run_main(["stack", str(folder), "--out", str(out)], capsys)
    got = openpyxl.load_workbook(out).active
    vals = [row[1] for row in got.iter_rows(min_row=2, values_only=True)]
    assert 200 in vals, f"空白行の次の本物の行が消えた: {vals}\n{text}"
