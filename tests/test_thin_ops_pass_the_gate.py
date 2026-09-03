# 判定の段（verify_dsl_args）を**通る**検体 ── 薄い 5 op ぶん（2026-09-03）。
#
# ★★ なぜ在るか: テスト全体で verify_dsl_args がどの op を何回通るかを測ったら、
#   29 op すべてが通ってはいたが、**回数が 20 倍ちがった**（SORT 149 / DELETE_COLUMN 3）。
#   そして断りの理由を全部読んだところ、**予期せぬ失敗は 1 件も無かった** ──
#   「存在しない列を指した」「見出しを壊す操作」「全部の列を指定した」等、
#   どれもテストがわざと壊れた入力を渡しているか、製品の規則が正しく働いた断り。
#
# ★★★ だから薄さの正体は「合格率が低い」ではなく **「通す検体が足りない」** だった:
#     SORT        149 呼出 / 断り 11 → 通す検体 138
#     FORMAT_MAP    7 呼出 / 断り  5 → 通す検体  2   ★ ここ
#   ★ 合格率だけ見て「判定が厳しすぎる」と読んで断りを緩めていたら、**製品を壊していた**。
#   成功例と比べたから向きが分かった（Namakoo「成功例との比較も行ってほしい」）。
#
# ★ この検体が守るもの: verify_dsl_args（1,735 行・op 分岐 30 個・5 本の入れ子チェーン）を
#   これから組み替える。その時、**薄い op から静かに壊れる**のを防ぐ。
#   ★ 断りではなく「通ること」を見る ── 通る道が消えたら鳴る。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))

import ailine  # noqa: E402


def _book(tmp_path, name="表.xlsx"):
    """列も行も揃った、素直な表（合計行つき）。"""
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in [["商品", "売上", "原価"],
               ["みかん", 1000, 400],
               ["ぶどう", 2000, 900],
               ["合計", 3000, 1300]]:
        ws.append(r)
    wb.save(p)
    return p


def _template_book(tmp_path):
    """様式写像用 ── データシートと、人が並べた雛形シート。"""
    p = tmp_path / "様式つき.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "元"
    for r in [["日付", "摘要", "入金"],
               ["2026-07-01", "売上入金", 50000],
               ["2026-07-03", "追加入金", 12000]]:
        ws.append(r)
    tpl = wb.create_sheet("様式")
    tpl["A1"] = "取引日"; tpl["B1"] = "内容"; tpl["C1"] = "金額"
    tpl["A2"] = "{{日付}}"; tpl["B2"] = "{{摘要}}"; tpl["C2"] = "{{入金}}"
    wb.save(p)
    return p


def _gate(op, args, book, task=""):
    ok, resolved, _inferred, err = ailine.verify_dsl_args(
        op, dict(args), ailine.build_book_meta(book), task=task)
    return ok, resolved, err


# --- 薄い 5 op が「通る」ことを見る -----------------------------------------------------

def test_delete_column_passes_with_a_real_column(tmp_path):
    """列削除: 実在する列を指せば通る（★ 通す検体は 1 件しか無かった）。"""
    ok, resolved, err = _gate("DELETE_COLUMN", {"col": "原価"}, _book(tmp_path),
                               task="原価の列を削除して")
    assert ok, err
    assert resolved.get("col") == "原価"


def test_delete_rows_passes_below_the_header(tmp_path):
    """行削除: 見出しより下の行なら通る（1 行目は規則で断られる ── それは正しい）。"""
    ok, resolved, err = _gate("DELETE_ROWS", {"at": 3, "count": 1}, _book(tmp_path),
                               task="3行目を削除して")
    assert ok, err
    assert resolved.get("at") == 3


def test_extract_columns_passes_with_a_subset(tmp_path):
    """列抽出: 一部の列を名指しすれば通る（全部を指すと『抜き出す意味がない』で断り）。"""
    ok, resolved, err = _gate("EXTRACT_COLUMNS", {"cols": ["商品", "売上"]}, _book(tmp_path),
                               task="商品と売上の列だけ抜き出して")
    assert ok, err
    assert resolved.get("cols") == ["商品", "売上"]


def test_split_cell_passes_with_a_separator_that_exists(tmp_path):
    """セル分割: 区切りが実際に在る列なら通る。"""
    p = tmp_path / "URL.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "一覧"
    ws.append(["名前", "URL"])
    ws.append(["A社", "http://a.example\nhttp://b.example"])
    wb.save(p)
    ok, resolved, err = _gate("SPLIT_CELL", {"col": "URL", "sep": "改行"}, p,
                               task="URL の列を改行で分けて")
    assert ok, err
    assert resolved.get("col") == "URL"


def test_format_map_passes_with_a_template_sheet(tmp_path):
    """様式写像: 雛形シートが別に在り、印が実在する列を指していれば通る。

    ★ B（verify_dsl_args の組み替え）で最大の分岐（136 行）を持つ op。
      通す検体が 2 件しか無かったので、ここが静かに壊れる筆頭だった。
    """
    ok, resolved, err = _gate("FORMAT_MAP", {"template_sheet": "様式"},
                               _template_book(tmp_path), task="様式に合わせて出力して")
    assert ok, err
    assert resolved.get("template_sheet") == "様式"


# --- ★ 陽性対照: この検体が「通ること」を見ていると確かめる ------------------------------

def test_the_gate_still_refuses_what_it_should(tmp_path):
    """★ 通す検体を足したせいで、断るべきものまで通していないこと。

    ★ 断りは製品の正しい働き ── 合格率を上げたくて緩めたら本末転倒になる。
      この試験は「通す検体を足す」作業の**対**として置く。
    """
    book = _book(tmp_path)
    ok, _r, err = _gate("DELETE_COLUMN", {"col": "存在しない列"}, book)
    assert not ok and err, "存在しない列を通した"
    ok, _r, err = _gate("DELETE_ROWS", {"at": 1, "count": 1}, book)
    assert not ok and err, "見出し行の削除を通した"
    ok, _r, err = _gate("EXTRACT_COLUMNS", {"cols": ["商品", "売上", "原価"]}, book)
    assert not ok and err, "全部の列の抽出を通した（抜き出す意味がない）"
