# 完成度レビュー（自作 ultrareview × 本家 ultra・2026-08-23）の致命 6 件 ── 修正より先に
# 凍結した赤い検体。Namakoo「ＯＫだ」(12:03)。出典: SEALED-20260823-jisaku-ultra.md。
#
# 契約:
#   ① CSV: cmd_run_csv は照合結果（欠落/不一致/余剰）が 0 でなければ ✓ を名乗らない
#      （実測: float ドリフトで「✓ …不一致1…」の自己矛盾・暗黙前段は正しく見ている=片配線）
#   ② CSV: EUC-JP は cp932 で黙って復号せず「対象外」と正直に断る（README の約束）
#   ③ 税語彙: 敗者復活（key に税を含む語彙の単一値採用）を COMPUTE_COLUMN 側にも配線
#   ④ チャート: 事後条件は「今回増えた 1 個」を同定して検証（既存グラフの誤 fail/誤 pass 両殺し）
#   ⑤ CSV: カンマ桁区切りでも整数部 16 桁以上は digit_overflow 拒否権（真理値表 rule d の穴）
#   ⑥ 提案の暗黙登録: save_alias が失敗したら「登録しました」と言わず理由を正直に開示

import io
import json
import shutil
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _book, _isolate, _run_main  # noqa: E402

FIXTURES = Path(__file__).parent / "fixtures" / "charts"


# --- ① CSV 照合無視の ✓（片配線）--------------------------------------------------

def test_csv_mismatch_never_wears_checkmark(tmp_path, monkeypatch, capsys):
    """float ドリフトで不一致 1 が出る実物 ── ✓ を名乗らない。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "in.csv"
    p.write_bytes(b"a\n1.1234567890123456789\n")
    rc, out = _run_main(["csv", str(p)], capsys)
    assert "不一致1" in out.replace(" ", "") or "不一致 1" in out, f"前提: 不一致が出ること: {out}"
    assert "✓" not in out, f"照合が合っていないのに ✓ を名乗った: {out}"


# --- ② EUC の黙読 ---------------------------------------------------------------

def test_euc_csv_is_refused_honestly(tmp_path, monkeypatch, capsys):
    """EUC-JP の CSV ── cp932 の文字化けで ✓ を出さず、対象外と名指しで断る。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "euc.csv"
    p.write_bytes("品番,金額\n0123,100\n".encode("euc-jp"))
    rc, out = _run_main(["csv", str(p)], capsys)
    assert rc != 0, f"文字化けのまま成功した: {out}"
    assert "✓" not in out
    assert "対象外" in out or "EUC" in out, f"正直な断りが無い: {out}"


def test_cp932_csv_still_reads_fine(tmp_path, monkeypatch, capsys):
    """誤爆防止: 本物の cp932 は従来どおり読める（EUC 判定が cp932 を巻き込まない）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "sjis.csv"
    p.write_bytes("品番,金額\n0123,100\n".encode("cp932"))
    rc, out = _run_main(["csv", str(p)], capsys)
    assert rc == 0, out
    assert "cp932" in out and "✓" in out, out


# --- ③ 税語彙の双子片配線 ---------------------------------------------------------

def test_tax_rescue_wired_into_compute_column():
    """語彙 {消費税:1.1} 登録済み ── COMPUTE_COLUMN の税込みでも敗者復活が効く
       （APPEND_TOTAL と同じ・「登録してください」の嘘を止める）。"""
    meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "金額"]},
             "header_rows": {"Sheet": 1}}
    ok, resolved, _inf, err = ailine.verify_dsl_args(
        "COMPUTE_COLUMN", {"operands": ["金額"], "operator": "*"}, meta,
        task="金額を税込みにして", vocab={"消費税": 1.1})
    assert ok, f"登録済みなのに失敗: {err}"
    assert resolved.get("factor") == 1.1


# --- ④ チャートの「今回増えた 1 個」の同定 -----------------------------------------

def _merge_chart_into(base_xlsx: Path, donor_xlsx: Path, out: Path):
    """base の zip に donor の chart XML を chart2 として合成（既存+新規の 2 枚本を作る）。"""
    shutil.copy2(base_xlsx, out)
    with zipfile.ZipFile(donor_xlsx) as zd:
        donor_chart = next(n for n in zd.namelist()
                            if n.startswith("xl/charts/chart") and n.endswith(".xml")
                            and "/charts/chart" in n)
        donor_xml = zd.read(donor_chart)
    import os
    tmp = str(out) + ".tmp"
    with zipfile.ZipFile(out) as zi, zipfile.ZipFile(tmp, "w") as zo:
        for item in zi.infolist():
            zo.writestr(item, zi.read(item.filename))
        zo.writestr("xl/charts/chart2.xml", donor_xml)
    os.replace(tmp, out)


def test_chart_check_verifies_the_new_chart_not_the_first(tmp_path):
    """誤 fail 殺し: 既存 bar(chart1) + 新規 line(chart2) ── 新規を検証して pass。"""
    merged = tmp_path / "two.xlsx"
    _merge_chart_into(FIXTURES / "bar.xlsx", FIXTURES / "line.xlsx", merged)
    from ailine_core import chart_check
    status, reason = chart_check.check_chart_series(
        merged, kind="line", value_col_letter="B",
        before_chart_paths={"xl/charts/chart1.xml"})
    assert status == "pass", f"既存グラフを見て誤 fail: {reason}"


def test_chart_check_fails_when_new_chart_is_wrong_even_if_old_matches(tmp_path):
    """誤 pass 殺し: 既存 bar(chart1・条件に適合) + 新規 pie(chart2・種別違い=壊れた新規)
       ── 既存に惑わされず fail。"""
    merged = tmp_path / "two.xlsx"
    _merge_chart_into(FIXTURES / "bar.xlsx", FIXTURES / "pie.xlsx", merged)
    from ailine_core import chart_check
    status, reason = chart_check.check_chart_series(
        merged, kind="bar", value_col_letter="B",
        before_chart_paths={"xl/charts/chart1.xml"})
    assert status == "fail", f"既存の適合グラフで壊れた新規を隠した（偽の ✓）: {reason}"


# --- ⑤ カンマ桁区切りの 16 桁（真理値表 rule d の穴・レビュー実測で追補）----------------

def test_comma_grouped_16_digits_is_vetoed():
    """"1,234,567,890,123,456"（16 桁）── カンマ形でも digit_overflow で文字列保持。"""
    from ailine_core import csv_quarantine as cq
    v = cq.classify_column(["1,234,567,890,123,456"])
    assert v.kind == "string", f"16 桁がカンマで素通り: {v}"
    assert "digit_overflow" in v.reasons


def test_comma_grouped_normal_still_number():
    """誤爆防止: 15 桁以内のカンマ列は従来どおり数値。"""
    from ailine_core import csv_quarantine as cq
    v = cq.classify_column(["1,234", "2,345"])
    assert v.kind == "number" and "comma_grouped" in v.reasons


# --- ⑥ 暗黙登録の偽成功 -----------------------------------------------------------

def test_long_phrase_nod_discloses_registration_failure(tmp_path, monkeypatch, capsys):
    """40 字超の言い回しで頷いた ── 「登録しました」と言わず、登録できなかった理由を開示。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "ALIASES_FILE", tmp_path / "aliases.json")
    monkeypatch.setattr(ailine, "MISCLASS_FILE", tmp_path / "misclass.jsonl")
    book = _book(tmp_path, [["商品", "金額"], ["b", 50], ["a", 100]])
    # ★ 治具の訂正（封印者ナギ・2026-08-23）: 初版はちょうど 40 字で自らの前提 len>40 を
    #   満たさないオフバイワンだった（implementer が正しく検出・封印不変のまま報告）。
    #   assert は不変・言い回しに 3 字足して 43 字に。
    long_task = "金額で降順に並べ替えてから合計をどうかこちらのシートに一番下に書いて頂けますでしょうか"
    assert len(long_task) > 40
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "OUT_OF_VOCAB", "about": "並べ替えのような操作", "args": {}}]})
    monkeypatch.setattr(
        ailine, "translate_task_fixed_op",
        lambda model, op, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "a", 100, "b", 50
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    monkeypatch.setattr("sys.stdin", io.StringIO("y\n"))
    monkeypatch.setattr("sys.stdin.isatty", lambda: True, raising=False)
    rc, out = _run_main(["run", str(book), long_task, "--copy"], capsys)
    assert rc == 0, out
    assert "登録しました" not in out, f"登録できていないのに成功を騙った: {out}"
    assert "登録できません" in out or "登録でき" in out, f"理由の開示が無い: {out}"
