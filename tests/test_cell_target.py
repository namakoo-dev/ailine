# セル指定（1 セルだけ飾る）── 2026-08-27。Namakoo「セル指定は出来る？『商品』セルに色を付けて」
#
# ★★ 実測した壊れ方: 「『商品』セルに色を付けて」は 3/3 で `col:商品`（**列ぜんぶ**）に
#   化けていた。頼んでいない範囲へ静かに広がる形で、この道具が最も嫌うもの ──
#   「動かない」より悪い。**それらしく動いて範囲だけ違う**からだ。
#
# 契約:
#   ① 依頼文が「セル」「見出しだけ」と言い、機械がその値を**1 箇所に**特定できたら 1 セル
#   ② 見つからない・複数ある時は決めない（推測で別のセルを塗らない）
#   ③ 座標は LLM に出させない（cell:R,C は機械が作る形）
#   ④ 事後条件は**両方向**: 宣言したセルが飾られ、**他のセルの飾りが 1 つも変わらない**
#   ⑤ 適用前が無ければ ④ は主張しない（言えないことは言わない）

import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

ROWS = [["商品", "売上", "原価"], ["りんご", 1200, 700], ["みかん", 800, 300]]
META_NP = {"sheets": ["売上"], "headers": {"売上": ["商品", "売上", "原価"]},
            "header_rows": {"売上": 1}}


def _book(tmp_path, name="b.xlsx", paint=(), bold=()):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    for r in ROWS:
        ws.append(r)
    for rc in paint:
        ws.cell(*rc).fill = PatternFill("solid", fgColor="FFFF00")
    for rc in bold:
        ws.cell(*rc).font = Font(bold=True)
    wb.save(p)
    return p


def _meta(path):
    return dict(META_NP, path=str(path))


# --- ①②③ -----------------------------------------------------------------------------

@pytest.mark.parametrize("name,rc", [("商品", (1, 1)), ("みかん", (3, 1)), ("売上", (1, 2))])
def test_a_named_cell_resolves(tmp_path, name, rc):
    r, c, note = ailine.resolve_named_cell(_meta(_book(tmp_path)), "売上", name)
    assert (r, c) == rc, note


def test_a_name_that_is_not_there_is_refused(tmp_path):
    r, c, note = ailine.resolve_named_cell(_meta(_book(tmp_path)), "売上", "すいか")
    assert (r, c) == (None, None) and "見つかりません" in note


def test_a_name_in_two_places_is_refused(tmp_path):
    """② 複数あるなら決めない ── 推測で別のセルを塗るのが一番こわい。"""
    rows = [["商品", "売上"], ["みかん", 1], ["みかん", 2]]
    p = tmp_path / "d.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    for r in rows:
        ws.append(r)
    wb.save(p)
    r, c, note = ailine.resolve_named_cell(dict(META_NP, path=str(p)), "売上", "みかん")
    assert (r, c) == (None, None) and "2 箇所あります" in note, note


@pytest.mark.parametrize("task,ok", [
    ("「商品」セルに色を付けて", True),
    ("商品の見出しだけ太字にして", True),
    ("みかんのセルを黄色にして", True),
    ("金額で降順に並べ替えて", False),
    ("商品列を全部太字にして", False),
    # ★ 自分で開けた穴（既存の検体が捕まえた）: 「見出しを太字にして」は**見出し行ぜんぶ**の
    #   意味でもありうる。曖昧なものを勝手に狭めない（広げないのと同じくらい大事）。
    ("見出しを太字にして", False),
    ("商品の見出しを太字にして", False),
])
def test_the_one_cell_phrasing_is_recognised(task, ok):
    assert ailine.task_asks_for_one_cell(task) is ok


def test_the_format_kind_comes_from_the_words():
    assert ailine.format_op_from_task("みかんのセルを黄色にして") == "FILL_COLOR"
    assert ailine.format_op_from_task("見出しセルだけ太字にして") == "BOLD"
    assert ailine.format_op_from_task("商品セルを中央揃えに") == "CENTER_ALIGN"
    assert ailine.format_op_from_task("金額で並べ替えて") is None


def test_a_cell_target_is_accepted_and_a_broken_one_is_not():
    ok, _r, _i, err = ailine.verify_dsl_args(
        "FILL_COLOR", {"target": "cell:1,1", "color": "yellow"}, META_NP, task="t")
    assert ok, err
    ok2, _r2, _i2, err2 = ailine.verify_dsl_args(
        "FILL_COLOR", {"target": "cell:x", "color": "yellow"}, META_NP, task="t")
    assert not ok2 and "読めません" in err2, err2


@pytest.mark.parametrize("op,args,want", [
    ("BOLD", {"target": "cell:1,1"}, "Call StyleBold(oDoc, 0, 0, 0, 0)"),
    ("FILL_COLOR", {"target": "cell:3,1", "color": "yellow"},
     "getCellByPosition(0, 2).CellBackColor = &HFFFF00&"),
    ("CENTER_ALIGN", {"target": "cell:1,2"}, "getCellByPosition(1, 0).HoriJustify"),
])
def test_codegen_touches_exactly_one_cell(op, args, want):
    ok, r, _i, err = ailine.verify_dsl_args(op, dict(args), META_NP, task="t")
    assert ok, err
    code = ailine.codegen_dsl(op, r, META_NP)
    assert want in code, code
    # ★ 1 セルの生成に走査（lastRow）が混ざっていないこと ── 混ざれば列へ広がる
    assert "lastRow" not in code, code


# --- ④⑤ 事後条件（両方向）---------------------------------------------------------------

def test_only_that_cell_passes(tmp_path):
    before = _book(tmp_path, "before.xlsx")
    after = _book(tmp_path, "after.xlsx", paint=[(1, 1)])
    status, reason = ailine.check_fill_color(
        after, {"target": "cell:1,1", "color": "yellow"}, source_book=before)
    assert status == "pass" and "他のセルは 1 つも変わらず" in reason, reason


def test_spreading_to_the_whole_column_fails(tmp_path):
    """★★ 恒真殺しの本命: これが実測で起きていた形。「付いたか」だけ見る番人なら通る。"""
    before = _book(tmp_path, "before.xlsx")
    after = _book(tmp_path, "after.xlsx", paint=[(1, 1), (2, 1), (3, 1)])
    status, reason = ailine.check_fill_color(
        after, {"target": "cell:1,1", "color": "yellow"}, source_book=before)
    assert status == "fail" and "広がった疑い" in reason, reason


def test_not_painted_at_all_fails(tmp_path):
    before = _book(tmp_path, "before.xlsx")
    after = _book(tmp_path, "after.xlsx")
    status, reason = ailine.check_fill_color(
        after, {"target": "cell:1,1", "color": "yellow"}, source_book=before)
    assert status == "fail" and "付いていません" in reason, reason


def test_a_cell_already_painted_before_is_not_called_spreading(tmp_path):
    """★ 元から飾ってあったセルを「広がった」と誤らない（だから適用前と比べる）。"""
    before = _book(tmp_path, "before.xlsx", paint=[(3, 1)])
    after = _book(tmp_path, "after.xlsx", paint=[(1, 1), (3, 1)])
    status, reason = ailine.check_fill_color(
        after, {"target": "cell:1,1", "color": "yellow"}, source_book=before)
    assert status == "pass", reason


def test_without_the_before_file_it_does_not_claim(tmp_path):
    """⑤ 言えないことは言わない ── 広がっていないかは主張しない。"""
    after = _book(tmp_path, "after.xlsx", paint=[(1, 1)])
    status, reason = ailine.check_fill_color(after, {"target": "cell:1,1", "color": "yellow"})
    assert status == "warn" and "未確認" in reason, reason


def test_bold_one_cell(tmp_path):
    before = _book(tmp_path, "before.xlsx")
    after = _book(tmp_path, "after.xlsx", bold=[(1, 1)])
    assert ailine.check_bold(after, {"target": "cell:1,1"}, source_book=before)[0] == "pass"
    wide = _book(tmp_path, "wide.xlsx", bold=[(1, 1), (2, 1), (3, 1)])
    assert ailine.check_bold(wide, {"target": "cell:1,1"}, source_book=before)[0] == "fail"
