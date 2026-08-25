"""★ C9 の事前登録した合否バー（この1本が単位 A の測定器）。

## バー（実装より先に凍結した文）

> **「`✓` が出た全 run について、run 終了後に別プロセスが openpyxl で原本を開き、
> claim の evidence を独立に再導出できる」— 例外 0 件。**

ブラインド査定2本のうちの一人の言葉「**毎回 openpyxl で開くまでファイルがどうなったか
分からなかった。この道具の一番の売りは検証なのに。**」を、そのままテストに翻訳したもの。
「独立に」を守るため、再導出は **subprocess で別の python を起動して行う**（ailine を
import しない・openpyxl だけを使う）。期待文字列の組み立てもこのテストが自前で書く
（実装の整形関数を呼ばない＝実装が形を変えたらここが赤くなる）。

## 電池に含める場合分け（ブリーフが列挙したもの）

- `--dry`（単発 DSL / 複合計画）
- `--copy`
- 複合計画の途中失敗（全段破棄）
- 単発の成功（既定＝原本直接適用）
- `APPEND_TOTAL` → `SORT` の順（「最終ファイルで全段の事後条件を再実行する版を選ばなかった」
  ことの回帰試験。再実行版だと `check_append_total` が `"=SUM("` の初出行で合計行を探すため、
  SORT で合計行が上へ移動した**正しい** run が偽 fail になる）
- 原本適用のフォールバック経路（`atomic_replace_inplace` の copy2 側・ゴールデンが1本も
  無かった経路）

★ 実機 LibreOffice/ollama は使わない（translate_task / basrun_apply を差し替え、
それ以外は本物のパイプラインを通す＝tests/test_golden_transcripts.py と同じ流儀）。
"""
import json
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import ailine  # noqa: E402

# ★ ✓ の行の形（実装の文字列を import せず、このテストが独立に持つ）。
_CLAIM_RE = re.compile(
    r"^✓ (?P<name>.+?) は機械検証済みの内容です（適用後に読み戻して確認: (?P<evidence>.+)）$")

_REDERIVE_SCRIPT = """
import json, sys
import openpyxl
wb = openpyxl.load_workbook(sys.argv[1])
out = []
for ws in wb.worksheets:
    filled = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
    out.append([ws.title, ws.max_row, ws.max_column, filled])
print(json.dumps(out, ensure_ascii=False))
"""


def _rederive_evidence(path: Path) -> str:
    """★ バーの本体: **別プロセス**で openpyxl だけを使ってファイルを開き、evidence を
       独立に組み立て直す（ailine も ailine_core も import しない）。"""
    proc = subprocess.run([sys.executable, "-c", _REDERIVE_SCRIPT, str(path)],
                           capture_output=True, text=True, encoding="utf-8")
    assert proc.returncode == 0, f"別プロセスでの読み戻しに失敗した: {proc.stderr}"
    sheets = json.loads(proc.stdout.strip())
    return "・".join(f"{title}: {rows}行×{cols}列・値のあるセル {filled}"
                      for title, rows, cols, filled in sheets)


def _claim_lines(out: str) -> list:
    return [ln for ln in out.splitlines() if ln.startswith("✓ ")]


def _assert_claim_is_independently_rederivable(out: str, candidates: dict) -> str:
    """✓ が出ているなら、その1行が名指ししたファイルを別プロセスで開き直して evidence を
       突き合わせる。戻り値は ✓ が名指ししたファイル名（✓ が無ければ ""）。"""
    checks = _claim_lines(out)
    assert len(checks) <= 1, f"✓ が run 中に複数出ている（発生点は1箇所のはず）: {checks}"
    if not checks:
        assert "✓" not in out, f"✓ が claim の行以外の場所に出ている: {out}"
        return ""
    m = _CLAIM_RE.match(checks[0])
    assert m, f"✓ の行が claim の形をしていない: {checks[0]!r}"
    name = m.group("name")
    assert name in candidates, f"✓ が知らないファイルを名指ししている: {name}"
    path = candidates[name]
    assert path.exists(), f"✓ が名指ししたファイルが存在しない: {path}"
    assert m.group("evidence") == _rederive_evidence(path), (
        "別プロセスで読み戻した観測と claim の evidence が食い違う（＝✓ が最終ファイルを"
        "見ていない）")
    return name


# --- 土台（tests/test_golden_transcripts.py と同じ流儀） ------------------------------

def _book(tmp_path, rows, name="b.xlsx"):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return p


def _isolate(monkeypatch, tmp_path):
    monkeypatch.setattr(ailine, "HISTORY_FILE", tmp_path / "history.jsonl")
    monkeypatch.setattr(ailine, "BACKUP_DIR", tmp_path / "backups")
    monkeypatch.setattr(ailine, "VOCAB_FILE", tmp_path / "vocab.json")
    monkeypatch.setattr(ailine, "RUN_LOCK_FILE", tmp_path / "run.lock")
    monkeypatch.setattr(ailine, "normalize_book", lambda book, workdir, timeout=None: book)


def _translate(monkeypatch, payload):
    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1: payload)


def _candidates(book: Path) -> dict:
    return {book.name: book,
            book.stem + ".out" + book.suffix: book.with_name(book.stem + ".out" + book.suffix)}


def _inject_formula_cache(path, sheet_filename: str, addr_to_value: dict) -> None:
    """テスト専用: xlsx の数式セルへキャッシュ値(<v>)を直接注入する（openpyxl は数式を
       計算しないため、LO を使わずに二層事後条件(式+キャッシュ値)を通すための小道具。
       tests/test_ailine.py の同名ヘルパと同じ実装）。"""
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_filename:
                text = data.decode("utf-8")
                for addr, value in addr_to_value.items():
                    pattern = re.compile(
                        rf'(<c r="{addr}"[^>]*>.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?(</c>)')
                    text, n = pattern.subn(rf'\1<v>{value}</v>\2', text, count=1)
                    assert n == 1, f"_inject_formula_cache: {addr} に注入できなかった"
                data = text.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


# ===========================================================================
# ① --dry: そもそも実行していない run は ✓ を1つも出さない
# ===========================================================================

@pytest.mark.parametrize("payload,task", [
    ({"op": "SORT", "args": {"col": "金額", "order": "desc"}}, "金額で降順に並べ替えて"),
    ({"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                {"op": "BOLD", "args": {"target": "row:1"}}]}, "並べ替えて見出しを太字に"),
])
def test_dry_run_never_claims_anything(tmp_path, monkeypatch, capsys, payload, task):
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 100], ["b", 200]])
    before = book.read_bytes()
    _translate(monkeypatch, payload)
    rc = ailine.main(["run", str(book), task, "--dry"])
    out = capsys.readouterr().out
    assert rc == 0
    assert "✓" not in out, f"--dry なのに ✓ が出ている（査定が挙げた3つ目の意味）: {out}"
    assert "--dry: 適用しない" in out
    assert book.read_bytes() == before


# ===========================================================================
# ② --copy: ✓ は .out について言い、原本無変更をその場で告げる
# ===========================================================================

def test_copy_mode_claim_names_the_out_file_and_says_original_untouched(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 200], ["b", 300], ["c", 100]])
    before = book.read_bytes()
    _translate(monkeypatch, {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for i, (name, val) in enumerate([("b", 300), ("a", 200), ("c", 100)], start=2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=val)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    rc = ailine.main(["run", str(book), "金額で降順に並べ替えて", "--copy"])
    out = capsys.readouterr().out
    assert rc == 0
    named = _assert_claim_is_independently_rederivable(out, _candidates(book))
    assert named == book.stem + ".out" + book.suffix, "✓ が .out でなく原本を名指ししている"
    assert f"（原本 {book.name} は変更していません）" in out
    assert book.read_bytes() == before


# ===========================================================================
# ③ 複合計画の途中失敗（全段破棄）: ✓ を出さず、原本無変更を必ず言う
# ===========================================================================

def test_partially_failed_plan_claims_nothing_and_reports_the_original_is_untouched(
        tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path, [["商品", "金額"], ["a", 300], ["b", 200], ["c", 100]])
    before = book.read_bytes()
    _translate(monkeypatch, {"plan": [{"op": "SORT", "args": {"col": "金額", "order": "desc"}},
                                       {"op": "BOLD", "args": {"target": "row:1"}}]})
    monkeypatch.setattr(ailine, "basrun_apply",
                         lambda out_book, code, workdir, helper_files=(), timeout=None:
                         (True, None, "ok"))   # BOLD は何もしない → 2段目 fail

    rc = ailine.main(["run", str(book), "金額で降順に並べ替えて見出しを太字に"])
    out = capsys.readouterr().out
    assert rc == 1
    # ★ 査定が名指しした「意味②」: 成功した段に ✓ が出て、原本がどうなったかは無言だった。
    assert "✓" not in out, f"全段破棄なのに ✓ が出ている: {out}"
    assert f"（原本 {book.name} は変更していません。作業結果は" in out
    assert book.read_bytes() == before


# ===========================================================================
# ④ 単発の成功（既定＝原本直接適用）
# ===========================================================================

def _single_success_run(tmp_path, monkeypatch, capsys):
    book = _book(tmp_path, [["商品", "数量", "単価", "金額"], ["a", 2, 100, 999], ["b", 3, 150, 999]])
    _translate(monkeypatch, {"op": "COMPUTE_COLUMN",
                              "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        ws.cell(row=2, column=4, value=200)
        ws.cell(row=3, column=4, value=450)
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc = ailine.main(["run", str(book), "金額を数量×単価で上書きして", "--values", "--overwrite"])
    return book, rc, capsys.readouterr().out


def test_single_success_claim_names_the_original_and_is_rederivable(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    book, rc, out = _single_success_run(tmp_path, monkeypatch, capsys)
    assert rc == 0
    named = _assert_claim_is_independently_rederivable(out, _candidates(book))
    assert named == book.name, "既定（原本直接適用）なのに ✓ が原本を名指ししていない"
    assert "（もとに戻す: ailine undo）" in out
    # ★ undo の案内と ✓ の重複整理: 「反映しました」型の別バナーはもう出ない。
    assert out.count("ailine undo") == 1


# ===========================================================================
# ⑤ APPEND_TOTAL → SORT（「事後条件を再実行する版を選ばなかった」ことの回帰試験）
# ===========================================================================

def test_append_total_then_sort_still_claims_and_rerunning_postconditions_would_have_lied(
        tmp_path, monkeypatch, capsys):
    """★ 不健全な代替案の回帰試験。SORT で合計行が最上段へ移った**正しい** run に対して
       `check_append_total` を最終ファイルで再実行すると偽 fail になることを、この
       テスト自身が実演したうえで、実装の ✓ はそれに影響されないことを固定する。

       ★★ 検体を組み替えた（算術恒等の検算・tests/test_sum_identity.py）。2つ理由がある:

       1. **並べ替える列を、合計を足した列とは別の列にした。** 元の検体は「金額の合計を
          足して、金額で降順」だった ―― その並びは今や T8b そのもの（合計行が最下行から
          動いたら ✓ を出さない）で、run が正しく fail する。
          ★★ **元の検体が「合計を足して、その同じ列で降順」＝ T8b と同型の並びを
          「正しい run」として ✓ ごと凍結していたこと自体が、この repo が二重計上
          （合計行がデータに混ざった出力）を正しい結果として凍結していた証拠**である
          ―― 検体は在ったが、測っていた性質が違った。
          この回帰試験が測りたいのはそこではなく「最終ファイルで事後条件を再実行する版を
          選ばなかったこと」なので、✓ が出る形を保ったまま合計行が上へ移る組み合わせに替えた。
          ★ 裏を返すと、**並べ替えた列以外にある合計行が動いても今の検算は鳴らない** ――
          この検体はその穴の上に立っている（報告済みの既知の限界）。
       2. **データを 100/200/300 から 100/250/400 に替えた。** 100+200==300 は
          「上の全部の合計」の恒等式に当たってしまう ―― 既存の合計行なのか、ただの
          データなのかは算術だけでは区別できない（T6 の真陽性と同じ数の並びになる）。
          ★ その後この検算は「足し込んだ範囲の最終行だけ」に絞ったので、今のこの並びは
          絞りの前後どちらでも黙る。検体はそのまま（余裕を持たせた側に倒しておく）。
       """
    _isolate(monkeypatch, tmp_path)
    # ★ 2026-08-25（塊①）: 列の並びを 部門/件数/金額 → 部門/金額/件数 に替えた。
    #   合計ラベルは**対象列の左隣**に置かれるので、元の並びでは 件数 が "合計"（文字列）に
    #   なる。検証できない行が 1 行でもあると ✓ を名乗らなくなったため、この検体の
    #   **乗り物**（✓ が出る run）が成立しなくなった。
    #   並びを替えるとラベルは 部門 に載り、件数 は数値のままでいられる。
    #   ★ 検体の設計（合計を足した列とは**別の列**で並べる／合計行が上へ移る／T8b は鳴らない）
    #     は 1 ミリも変えていない。
    book = _book(tmp_path, [["部門", "金額", "件数"], ["a", 100, 3], ["b", 250, 1], ["c", 400, 2]])
    _translate(monkeypatch, {"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額"}},
                                       {"op": "SORT", "args": {"col": "件数", "order": "desc"}}]})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        if "SortByColumn" in code:   # 2段目: 件数が空欄の合計行(750)が降順の先頭へ移動する
            rows = [("合計", "=SUM(B2:INDEX(B:B,ROW()-1))", 9),
                    ("a", 100, 3), ("c", 400, 2), ("b", 250, 1)]
            for i, row in enumerate(rows, start=2):
                for c, v in enumerate(row, start=1):
                    ws.cell(row=i, column=c, value=v)
            wb.save(out_book)
            _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"B2": 750})
            return True, None, "ok"
        ws.cell(row=5, column=1, value="合計")            # 1段目: データ末尾の下に合計行
        ws.cell(row=5, column=2, value="=SUM(B2:INDEX(B:B,ROW()-1))")
        # ★ 9 にする理由: 3+1+2=6 は「上の全部の合計」の恒等式に当たり、件数側にも
        #   合計行が在ると算術の検算が正しく鳴る（この検体の docstring が
        #   100+200==300 で同じ罠を踏んだと警告している）。件数は合計でない数にする。
        ws.cell(row=5, column=3, value=9)                 # 数値なので除外されない
        wb.save(out_book)
        _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"B5": 750})
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)

    rc = ailine.main(["run", str(book), "金額の合計を最後に足して、件数で降順に並べ替えて", "--copy"])
    out = capsys.readouterr().out
    assert rc == 0, out
    final = book.with_name(book.stem + ".out" + book.suffix)
    named = _assert_claim_is_independently_rederivable(out, _candidates(book))
    assert named == final.name

    # ★ 実演: 最終ファイルで1段目の事後条件を再実行すると fail になる（＝再実行版を選んで
    #   いたら、この正しい run が ✓ を失っていた）。
    status, _reason = ailine.check_append_total(final, {"col": "金額", "label": "合計", "factor": 1})
    assert status == "fail", (
        "前提が崩れている: SORT 後に check_append_total を再実行しても通ってしまうため、"
        "この回帰試験は『再実行版を選ばなかったこと』を測れていない")


# ===========================================================================
# ⑥ 原本適用のフォールバック経路（atomic_replace_inplace の copy2 側）
# ===========================================================================

def test_claim_holds_on_the_copy2_fallback_path_of_atomic_replace(tmp_path, monkeypatch, capsys):
    """★ ゴールデンが1本も無かった経路。os.replace が落ちて copy2 へ落ちた run でも、
       ✓ は「原本を読み戻して確かめた」という同じ意味でなければならない。"""
    _isolate(monkeypatch, tmp_path)
    real_replace = os.replace

    def flaky_replace(src, dst, *args, **kwargs):
        if str(dst).endswith("b.xlsx"):
            raise OSError("staging の置換に失敗した（テスト用の擬似障害）")
        return real_replace(src, dst, *args, **kwargs)
    monkeypatch.setattr(ailine.os, "replace", flaky_replace)

    book, rc, out = _single_success_run(tmp_path, monkeypatch, capsys)
    assert rc == 0, out
    assert "copy2 へフォールバックした" in out, out
    named = _assert_claim_is_independently_rederivable(out, _candidates(book))
    assert named == book.name
    # 原本に実際に反映されている（フォールバックでも中身が入っている）ことを裏取りする。
    ws = openpyxl.load_workbook(book).active
    assert [ws.cell(row=r, column=4).value for r in (2, 3)] == [200, 450]


# ===========================================================================
# ⑦ 読み戻せなかったら ✓ は出さない（✓ の意味の下限を固定する）
# ===========================================================================

def test_unreadable_final_file_downgrades_the_claim(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(ailine, "observe_book_state",
                         lambda path: (None, "BadZipFile: File is not a zip file"))
    book, rc, out = _single_success_run(tmp_path, monkeypatch, capsys)
    assert rc == 0
    assert "✓" not in out, f"読み戻せていないのに ✓ が出ている: {out}"
    assert "読み戻して確認できませんでした" in out
