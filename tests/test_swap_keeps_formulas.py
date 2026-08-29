# 入れ替えても式の意味が変わらないこと ── 2026-08-29→30。
# Namakoo「でもこれのユーザの意図としては、合計行ごと参照を変えずに追記したいって
# ことじゃないの？」
#
# ★★ それまでの ailine は「壊れたことを見つけて止める」で終わっていた。断りは正しかった
#   （実測: 列を入れ替えると合計式が二列にまたがり、両方 1,000,440 ＝ 金額＋税込み金額）。
#   ★ だが利用者が欲しいのは**意味を保ったまま位置だけ入れ替わった表**で、それは
#     機械が全部言える ── 操作前の式と写像 π が分かっているのだから、操作後の式は
#     π(操作前) でしかない。「後から直す」のでなく **最初から正しく書く**。
#
# ★★ そこで実測が 2 つ、設計を決めた:
#
#  (1) `setFormula` の引数区切りは **`;`**。同じ式を 3 通り書いて読み戻した:
#        =SUM(E2:INDEX(E:E,ROW()-1))  → #VALUE!
#        =SUM(E2:INDEX(E:E;ROW()-1))  → 通る
#        =SUM(E2:E8)                  → 通る（区切りが無いので影響なし）
#      ★ 文字は正しく見えるのに値だけ壊れる ── 一番たちが悪い形。
#
#  (2) 列と行で**逆のことが起きた**:
#        列を入れ替え → 合計セルは列ごと動いた。動いた先で自分の列を指すのが正しい。
#        行を入れ替え → 合計行は動かないのに LibreOffice が範囲の始まり B2 を B4 に
#                       付け替え、合計が 3500 → 1200 になった。
#      ★ 一つの規則で両方説明できる:
#        **動いた式は隣を指し続ける／動かなかった式は同じ番地を指し続ける。**
#      ★ ただしこれは**並べ替えだけ**の規則。挿す/消す時は番地そのものがずれるので、
#        動かない式も追従しなければならない（片方の規則を両方に当てない）。

import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402
from ailine_core import cellmap as cm  # noqa: E402


# --- ① Basic に渡す形（実測した環境の事実）------------------------------------------

@pytest.mark.parametrize("src, want", [
    ("=SUM(E2:INDEX(E:E,ROW()-1))", "=SUM(E2:INDEX(E:E;ROW()-1))"),
    ("=SUM(E2:E8)", "=SUM(E2:E8)"),
    ("=B2-C2", "=B2-C2"),
])
def test_the_argument_separator_becomes_a_semicolon(src, want):
    assert ailine.formula_for_basic(src) == want


def test_commas_inside_text_are_left_alone():
    """★ 文字列の中まで書き換えると、人が入れた文言が壊れる。"""
    f = '=IF(A1>0,"a,b","c,d")'
    assert ailine.formula_for_basic(f) == '=IF(A1>0;"a,b";"c,d")'


# --- ② 列まるごとの参照（それまで見えていなかった）----------------------------------

def test_a_whole_column_reference_follows_a_column_swap():
    """★★ `_REF` は行番号を要求するので `INDEX(E:E,…)` が**見えていなかった**。
       そのせいで map_formula は LibreOffice と同じ壊し方を再現していた ──
       直す側と壊す側が同じ盲点を持っていた。"""
    got = cm.map_formula("=SUM(E2:INDEX(E:E,ROW()-1))", cm.swap_cols(5, 6))
    assert got == "=SUM(F2:INDEX(F:F,ROW()-1))", got


def test_a_whole_row_reference_follows_a_row_insert():
    assert cm.map_formula("=INDEX(3:3,1)", cm.insert_rows(3)) == "=INDEX(4:4,1)"


def test_a_whole_column_reference_is_untouched_by_a_row_shift():
    """★ 別の軸の操作では動かない（片方だけ賢くしない、の裏返し）。"""
    assert cm.map_formula("=SUM(E2:INDEX(E:E,ROW()-1))",
                           cm.insert_rows(3)) == "=SUM(E2:INDEX(E:E,ROW()-1))"


def test_a_sheet_qualified_reference_is_left_alone():
    assert cm.map_formula("=売上!E:E", cm.swap_cols(5, 6)) == "=売上!E:E"


# --- ③ 動いたか動かなかったかで規則が変わる ------------------------------------------

def test_a_formula_that_moves_keeps_pointing_at_its_neighbours():
    """列の入れ替え: 合計セルは列ごと動く ── 動いた先で自分の列を指す。"""
    out, lost = cm.formulas_after({(9, 6): "=SUM(F2:INDEX(F:F,ROW()-1))"}, cm.swap_cols(5, 6))
    assert not lost
    assert out == {(9, 5): "=SUM(E2:INDEX(E:E,ROW()-1))"}, out


def test_a_formula_that_stays_keeps_pointing_at_the_same_address():
    """★★ 行の入れ替え: 合計行は動かない ── LibreOffice はここで B2 を B4 に
       付け替え、合計を 3500 から 1200 に変えた（実測）。"""
    out, lost = cm.formulas_after({(5, 2): "=SUM(B2:INDEX(B:B,ROW()-1))"}, cm.swap_rows(2, 4))
    assert not lost
    assert out == {(5, 2): "=SUM(B2:INDEX(B:B,ROW()-1))"}, out


def test_moving_rows_still_carry_their_formulas():
    out, _lost = cm.formulas_after({(2, 4): "=B2-C2"}, cm.swap_rows(2, 4))
    assert out == {(4, 4): "=B4-C4"}, out


def test_an_insert_moves_even_the_formulas_that_stay_put():
    """★ 並べ替えの規則を挿入に当てない ── 番地そのものがずれるので追従が要る。"""
    out, _lost = cm.formulas_after({(1, 2): "=SUM(B2:B4)"}, cm.insert_rows(3))
    assert out == {(1, 2): "=SUM(B2:B5)"}, out


# --- ④ 生成コードに積まれること（順序が意味を持つ）------------------------------------

def test_the_rewrites_come_after_the_swap():
    meta = {"sheets": ["S"], "headers": {"S": ["a", "b"]}, "header_rows": {"S": 1}}
    args = {"a": "金額", "b": "税込み金額", "_axis": "column", "_header_row": 1,
            "_formula_rewrites": [(9, 5, "=SUM(E2:INDEX(E:E,ROW()-1))")]}
    code = ailine.codegen_dsl("SWAP", args, meta)
    i, j = code.index("SwapColumnsByName"), code.index("SetFormulaAt")
    assert i < j, "式の書き直しが入れ替えより先に出ている"
    assert 'SetFormulaAt(oDoc, 8, 4, "=SUM(E2:INDEX(E:E;ROW()-1))")' in code, code


def test_the_basic_helper_writes_a_formula_not_a_string():
    """★ SetCellAt は setString なので式にならない ── 別の Sub が要る。"""
    bas = (REPO / "src" / "ailine" / "helpers" / "AiLineHelpers.bas").read_text(encoding="utf-8")
    i = bas.index("Sub SetFormulaAt(")
    assert "setFormula" in bas[i:i + 400], bas[i:i + 400]


# --- ⑤ 実物で（LibreOffice を通す）-----------------------------------------------------

def _run(book: Path, task: str, sheet: str | None = None):
    cmd = [sys.executable, "-m", "ailine", "run", str(book), task, "--copy"]
    if sheet:
        cmd += ["--sheet", sheet]
    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}
    return subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8",
                           errors="replace", timeout=900, cwd=str(REPO), env=env)


@pytest.mark.local
def test_swapping_columns_keeps_both_totals(tmp_path):
    """★★ Namakoo の実例そのもの。入れ替えたあとも、合計は操作前と同じ値のまま
       （位置だけが入れ替わる）。"""
    src = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(["取引先", "金額", "税込み金額"])
    for i, (n, v) in enumerate([("丸和", 57600), ("近江", 60000), ("ヤマノ", 42000)], start=2):
        ws.append([n, v, f"=B{i}*1.1"])
    ws.append(["合計", "=SUM(B2:INDEX(B:B,ROW()-1))", "=SUM(C2:INDEX(C:C,ROW()-1))"])
    wb.save(src)
    r = _run(src, "税込み金額と金額を入れ替えて")
    assert "✓" in r.stdout, r.stdout[-1500:]
    out = tmp_path / "b.out.xlsx"
    f = openpyxl.load_workbook(out)["請求"]
    v = openpyxl.load_workbook(out, data_only=True)["請求"]
    assert [f.cell(1, c).value for c in (2, 3)] == ["税込み金額", "金額"]
    # ★ 操作前: 金額 159600 ／ 税込み 175560。入れ替え後も同じ値が、入れ替わった位置に。
    assert v.cell(5, 2).value == pytest.approx(175560)
    assert v.cell(5, 3).value == pytest.approx(159600)


@pytest.mark.local
def test_swapping_rows_keeps_the_total_row(tmp_path):
    """★★ 逆向きの検体: 合計行は**動かない**ので、同じ番地を指し続けるのが正しい。
       ここを写像に通すと合計が 3500 → 1200 になる（実測した壊れ方）。"""
    src = tmp_path / "r.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価", "利益"])
    for i, (n, s, c) in enumerate([("りんご", 1200, 700), ("みかん", 800, 300),
                                    ("ぶどう", 1500, 900)], start=2):
        ws.append([n, s, c, f"=B{i}-C{i}"])
    ws.append(["合計", "=SUM(B2:INDEX(B:B,ROW()-1))", "=SUM(C2:INDEX(C:C,ROW()-1))", None])
    wb.save(src)
    r = _run(src, "りんごの行とぶどうの行を入れ替えて")
    assert "✓" in r.stdout, r.stdout[-1500:]
    out = tmp_path / "r.out.xlsx"
    v = openpyxl.load_workbook(out, data_only=True)["売上"]
    assert [v.cell(r_, 1).value for r_ in (2, 3, 4)] == ["ぶどう", "みかん", "りんご"]
    assert v.cell(2, 4).value == 600 and v.cell(4, 4).value == 500   # 利益は隣を指し続ける
    assert v.cell(5, 2).value == 3500 and v.cell(5, 3).value == 1900  # 合計は変わらない
