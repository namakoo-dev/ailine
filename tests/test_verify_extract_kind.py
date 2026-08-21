"""`ailine verify` が M2（フォルダ抽出）の出力も検算できること（種類判定・architect 致命3）。

★ 何を測るか: 出力ブックに焼いた条件（docProps の creator=印 / description=col・cmp・value）
   だけを頼りに、verify が同じ除外規則と同じ述語で元フォルダを読み直し、run の言い分と
   一致すること。一致するだけでなく、改竄には落ちること（両側）。
★ run は in-process で monkeypatch（subprocess 越しには翻訳の差し替えが効かない）。
"""
import argparse
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402

HDRS = ["注文ID", "取引先", "金額"]
_EXTRACT_40000 = [{"op": "EXTRACT", "args": {"column": "金額", "cmp": "gte", "value": 40000}}]
TASK = "金額が40000以上の行を抜き出して"


def _book(path, headers, rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r) + [None] * (len(headers) - len(r)))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _made(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(ailine, "translate_task",
                        lambda model, task, book_meta, temperature=0.1: {"plan": _EXTRACT_40000})
    folder = tmp_path / "src"
    _book(folder / "a.xlsx", HDRS,
          [("J-1", "甲", 50000), ("J-2", "乙", 30000), ("合計", None, 80000)])
    _book(folder / "b.xlsx", HDRS, [("J-3", "丙", 45000)])
    rc = ailine.main(["run", str(folder), TASK])
    out_text = capsys.readouterr().out
    assert rc == 0, out_text
    out = next(iter(tmp_path.glob("*.xlsx")))
    return folder, out


def _verify(out, folder, capsys):
    rc = ailine.cmd_verify(argparse.Namespace(out=str(out), srcfolder=str(folder)))
    return rc, capsys.readouterr().out


def test_verify_agrees_with_a_folder_extract_output(tmp_path, monkeypatch, capsys):
    folder, out = _made(tmp_path, monkeypatch, capsys)
    rc, text = _verify(out, folder, capsys)
    assert rc == 0, f"正当な抽出出力で verify が落ちた:\n{text}"
    assert "⚠" not in text, f"正当な出力に偽 ⚠:\n{text}"
    assert text.count("95000") >= 2, f"Σ金額の両側（50000+45000）が無い:\n{text}"


def test_verify_catches_a_tampered_value_in_an_extract_output(tmp_path, monkeypatch, capsys):
    """★ 敵対側: 出力の金額 1 セルを改竄したら、列名と両側の数字つきで exit 5。"""
    folder, out = _made(tmp_path, monkeypatch, capsys)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[2].value == 50000:
            row[2].value = 1
            break
    wb.save(out)
    rc, text = _verify(out, folder, capsys)
    assert rc == 5, f"exit={rc}\n{text}"
    assert "金額" in text and "95000" in text and "45001" in text, text


def test_verify_catches_a_row_that_never_satisfied_the_condition(tmp_path, monkeypatch, capsys):
    """★ 条件の両側: 条件を満たさない行を出力に足したら行数が合わずに落ちる
       （焼いた条件から述語を復元できていなければ、この行は素通りする）。"""
    folder, out = _made(tmp_path, monkeypatch, capsys)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    ws.append(["J-2", "乙", 30000, "a.xlsx", 3])
    wb.save(out)
    rc, text = _verify(out, folder, capsys)
    assert rc == 5, f"条件を満たさない行が素通りした:\n{text}"
    assert "2" in text and "3" in text, text
