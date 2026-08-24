# EXTRACT の日付範囲比較 ── 実装より先に凍結した赤い検体（2026-08-24）。
# 設計: ~/.nagi/plans/DESIGN-20260824-ledger-rest.md
#
# ★ 出所: 台帳の DATE_RANGE_AGG 2 件（4976755 日報の月次/年次・2896855 工程表の期間指定）。
#   実測（実 7B + 実 LibreOffice）で、複合計画は正しく出る（EXTRACT×2 → PIVOT）のに
#   EXTRACT が「比較『gte』には数値の値が必要」で止まっていた。**新 op ではなく既存 op の穴**。
#
# 契約:
#   ① 日付リテラル（2026/3/26・2026-03-26・2026年3月26日）を日付として読む。年が無い
#      「3/26」は**読まない**（どの年かは機械が決めてよい話ではない）
#   ② 日付列に対する gte/lte は **両端を含む**（締め日 3/26〜4/25 は 4/25 を含む）
#   ③ ★ 時刻つきの列でも 4/25 23:00 が lte 4/25 に**入る**（日付だけ書いた人の意図は
#      「その日いっぱい」── ここを取りこぼすと締め日の売上が静かに落ちる）
#   ④ 文字列として日付が入っている列は、日付として比較できないと**正直に断る**
#      （黙って辞書順で比べない ── 2026/3/26 と 2026/12/1 は辞書順では逆転する）

import datetime as dt
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

needs_impl = pytest.mark.xfail(
    not hasattr(ailine, "parse_date_literal"),
    reason="日付比較 未実装（契約は凍結済み・実装が来たら自動実測化）",
    strict=True,
)


def _daybook(tmp_path, name="日報.xlsx", with_time=False, as_text=False):
    p = tmp_path / name
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日報"
    ws.append(["日付", "現場", "売上"])
    rows = [(dt.date(2026, 3, 20), "A", 50),    # 期間外（前）
            (dt.date(2026, 3, 26), "A", 10),    # 境界（含む）
            (dt.date(2026, 3, 27), "A", 100),
            (dt.date(2026, 4, 25), "B", 70),    # 境界（含む）
            (dt.date(2026, 5, 1), "B", 200)]    # 期間外（後）
    for d, site, amt in rows:
        v = d.isoformat() if as_text else (dt.datetime(d.year, d.month, d.day, 23, 0)
                                            if with_time else dt.datetime(d.year, d.month, d.day))
        ws.append([v, site, amt])
    wb.save(p)
    return p


# --- ① 日付リテラルの読み取り -------------------------------------------------------

@needs_impl
@pytest.mark.parametrize("raw,expected", [
    ("2026/3/26", dt.date(2026, 3, 26)),
    ("2026-03-26", dt.date(2026, 3, 26)),
    ("2026年3月26日", dt.date(2026, 3, 26)),
    ("2026.3.26", dt.date(2026, 3, 26)),
    ("3/26", None),                 # ★ 年が無い ── 機械が年を決めない
    ("令和8年3月26日", None),        # 和暦は未対応（黙って誤変換しない）
    ("abc", None),
    ("100", None),                  # 数値は数値のまま（日付に化けさせない）
])
def test_parse_date_literal(raw, expected):
    assert ailine.parse_date_literal(raw) == expected


@needs_impl
def test_serial_matches_libreoffice_epoch():
    """LibreOffice/Excel の既定 null date は 1899-12-30。ここがずれると全部 1 日ずれる。"""
    assert ailine.date_to_serial(dt.date(1899, 12, 31)) == 1
    assert ailine.date_to_serial(dt.date(1900, 1, 1)) == 2
    assert ailine.date_to_serial(dt.date(2026, 3, 26)) == 46107


# --- ②③ 締め日の範囲（両端を含む・時刻つきでも落とさない）-----------------------------

@needs_impl
@pytest.mark.parametrize("with_time", [False, True])
def test_closing_period_includes_both_ends(tmp_path, with_time):
    """3/26〜4/25 は両端を含む。★ 時刻つきの列でも 4/25 23:00 を落とさない。"""
    book = _daybook(tmp_path, with_time=with_time)
    meta = ailine.build_book_meta(book)
    ok_lo, res_lo, _i, err_lo = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "gte", "value": "2026/3/26"}, meta,
        task="2026/3/26以降を抽出して")
    ok_hi, res_hi, _i2, err_hi = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "lte", "value": "2026/4/25"}, meta,
        task="2026/4/25以前を抽出して")
    assert ok_lo, err_lo
    assert ok_hi, err_hi
    # ★ 治具の訂正（封印者ナギ・2026-08-24）: 初版は閾値が resolved["value"] に入ると
    #   仮定していた。実装は **表示用に元の文字列を value に残し、閾値を _value_serial に
    #   置く**（解釈行と出力シート名が「46107」になるのを防ぐため）。assert（両端を含む・
    #   時刻つきでも落とさない）は 1 文字も変えていない ── 変えたのは閾値の読み出し先だけ。
    lo, hi = res_lo["_value_serial"], res_hi["_value_serial"]
    # ★ 同時に、人が読む側の契約もここで凍結する（片配線の防止）。
    assert res_lo["value"] == "2026/3/26" and res_hi["value"] == "2026/4/25",         f"表示用の値がシリアル値に化けた: {res_lo['value']!r} / {res_hi['value']!r}"
    def serial(d, h=0):
        return ailine.date_to_serial(d) + h / 24.0
    h = 23 if with_time else 0
    inside = [dt.date(2026, 3, 26), dt.date(2026, 3, 27), dt.date(2026, 4, 25)]
    outside = [dt.date(2026, 3, 20), dt.date(2026, 5, 1)]
    for d in inside:
        assert serial(d, h) >= lo and serial(d, h) <= hi, f"{d}(時刻{h}時) が期間から落ちた"
    for d in outside:
        assert not (serial(d, h) >= lo and serial(d, h) <= hi), f"{d} が期間に入った"


# --- ④ 文字列の日付列は正直に断る ---------------------------------------------------

@needs_impl
def test_text_date_column_is_refused_honestly(tmp_path):
    """日付が文字列で入っている列 ── 辞書順で黙って比べず、理由を名指しして断る。"""
    book = _daybook(tmp_path, as_text=True)
    meta = ailine.build_book_meta(book)
    ok, _res, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "gte", "value": "2026/3/26"}, meta,
        task="2026/3/26以降を抽出して")
    assert not ok, "文字列の日付列を黙って比較した"
    assert "文字列" in (err or "") or "日付" in (err or ""), f"理由が名指しされていない: {err}"


# ★ この 1 本だけ needs_impl を付けない ── 実装前から緑（今の挙動を守る誤爆防止の柵）。
def test_numeric_column_still_compares_numerically(tmp_path):
    """誤爆防止: ただの数値列の gte は今までどおり数値比較（日付に化けない）。"""
    book = _daybook(tmp_path)
    meta = ailine.build_book_meta(book)
    ok, res, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "売上", "cmp": "gte", "value": "100"}, meta,
        task="売上が100以上を抽出して")
    assert ok, err
    assert res["value"] == 100.0


# --- ★ 片配線の追補（2026-08-24 夕・盲検の使い勝手レビューで捕獲）---------------------
#
# verify と codegen だけ日付に対応させて、**事後条件を忘れていた**（俺の落とし物）。
# 実測: 「3月26日から4月25日までの分だけ集計して」で抽出は正しく 5 行を出しているのに
#   「× 未対応: 6行中0行が一致 → 出力は5行（行数が期待と不一致）」
# という**自己矛盾**を出して run 全体が exit 1 で失敗していた。ユーザーには手の打ちようがない。

@needs_impl
def test_extract_postcondition_counts_date_matches(tmp_path):
    """事後条件も日付で数える（シリアル値の閾値と、日付セルの数値化の両方が要る）。"""
    import openpyxl as _op
    book = _daybook(tmp_path)
    meta = ailine.build_book_meta(book)
    ok, res, _i, err = ailine.verify_dsl_args(
        "EXTRACT", {"col": "日付", "cmp": "gte", "value": "2026/3/26"}, meta,
        task="2026/3/26以降を抽出して")
    assert ok, err
    # 期待: 3/26・3/27・4/25・5/1 の 4 行が一致する
    wb = _op.load_workbook(book)
    src = wb["日報"]
    out = wb.create_sheet(res["_new_sheet"])
    out.append(["日付", "現場", "売上"])
    for r in range(2, src.max_row + 1):
        if src.cell(r, 1).value and src.cell(r, 1).value >= dt.datetime(2026, 3, 26):
            out.append([src.cell(r, c).value for c in (1, 2, 3)])
    wb.save(book)
    res["_target_sheet"] = "日報"
    status, reason = ailine.check_extract(book, res, header_row=1)
    assert status == "pass", f"日付の抽出を事後条件が数えられない（自己矛盾の再現）: {reason}"
    assert "0行が一致" not in (reason or ""), reason
