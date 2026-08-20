"""helpers/AiLineHelpers.bas の CharWeightAsian が実際に効いているかを、basrun で本物に
   太字を当てて openpyxl で読み戻して確認する通し試験。

   ★ 背景（全域監査 M1/M2 で発覚）: tests/test_ailine.py の check_bold 系テストは、検体を
   openpyxl の Font(bold=True) で直接作り、check_bold も openpyxl で読むだけ。実際に太字を
   当てている helpers/AiLineHelpers.bas の BoldRange（CharWeightAsian の1行）を一度も通らず、
   この1行を削除しても現行テストは全部緑のまま通っていた。この試験はその欠陥を埋める。

   ★ LibreOffice・basrun（sibling repo）が要るため @pytest.mark.local。CI では走らない
   （`pytest -m "not local"` で除外される。手動は `pytest -m local` または
   `pytest tests/test_bold_local.py`）。
   ★ tests/test_helper_catalog_sync.py の静的検査（CharWeight/CharWeightAsian が両方
   「書いてある」かのチェック）と役割が違う: あちらは書いてあることしか保証しない。
   こちらは実際に xlsx へ効いていることそのものを、本物の basrun 経由で見る。"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))
import ailine

HELPERS_DIR = REPO_ROOT / "helpers"


def _soffice_process_count() -> int:
    """soffice.bin プロセス数を数える（Windows tasklist）。テスト前後の残存確認用の観測のみ。
       ★ ここでは何も kill しない（名前一括 kill の事故を避ける。後始末は ailine._stop_office()
       ＝basrun.py stop に委譲し、接続先だけを閉じる既存機構を使う）。"""
    try:
        out = subprocess.run(
            ["tasklist", "/FI", "IMAGENAME eq soffice.bin"],
            capture_output=True, text=True, timeout=10).stdout
    except Exception:
        return -1   # 数えられない環境。local マーカーの手動実行前提なので実害は小さい
    return out.lower().count("soffice.bin")


@pytest.mark.local
def test_style_bold_actually_bolds_japanese_header_via_basrun(tmp_path):
    """StyleBold ヘルパを実際に basrun で日本語見出しに適用し、openpyxl の読み戻しで
       Font.bold が立つことを確認する。

       ★ この試験の存在意義: helpers/AiLineHelpers.bas の BoldRange から
       `oCell.CharWeightAsian = com.sun.star.awt.FontWeight.BOLD` の1行を消すと、
       この試験は赤くなる（実測で確認済み・PR 説明参照）。"""
    before_procs = _soffice_process_count()
    try:
        book = tmp_path / "bold.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["商品", "金額"])          # 日本語見出し（太字を当てる対象）
        ws.append(["りんご", 100])
        wb.save(book)

        workdir = tmp_path / "work"
        workdir.mkdir()
        _catalog, helper_files = ailine.load_helpers(HELPERS_DIR)
        assert helper_files, "helpers/*.bas が見つからない（HELPERS_DIR の場所を確認）"

        code = (
            "Option VBASupport 1\nOption Explicit\n\n"
            "Sub Run(oDoc As Object)\n"
            "    Call StyleBold(oDoc, 0, 0, 1, 0)\n"   # 見出し行(行0, 列0〜1)を太字に
            "End Sub\n"
        )
        ok, err, raw = ailine.basrun_apply(book, code, workdir, helper_files)
        assert ok, f"basrun_apply が失敗した: {err}\n{(raw or '')[-500:]}"

        wb2 = openpyxl.load_workbook(book)
        ws2 = wb2.active
        assert ws2["A1"].font.bold is True, "日本語見出し「商品」に太字が反映されていない"
        assert ws2["B1"].font.bold is True, "日本語見出し「金額」に太字が反映されていない"
        wb2.close()
    finally:
        # ★ 後片付け: basrun.py stop（接続先の LibreOffice だけを終了・ポートが閉じるまで
        #   待つ既存機構）。taskkill 一括はしない。
        ailine._stop_office()

    after_procs = _soffice_process_count()
    assert after_procs <= before_procs, (
        f"basrun 終了後も soffice.bin が残っている可能性がある（前={before_procs} 後={after_procs}）。"
        " ★ ここから先の後始末は名前一括 kill をせず、PID を特定して個別に kill すること。"
    )

@pytest.mark.local
def test_style_bold_bolds_mixed_script_header_via_basrun(tmp_path):
    """★ 混在文字（日本語+英字）の見出しセルに太字が効くか。

    実測 (2026-08-19 デモ制作): 見出し行 7 セルのうち『注文ID』だけが LO 往復後に
    openpyxl の Font.bold=False のまま残り、事後条件が正しく fail した。純日本語の
    6 セル（受注日・取引先…）は太字になっていた ── 7 セルで唯一の混在文字セルだけが落ちた。
    このテストはその再現。緑になるまでが修正。"""
    book = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["注文ID", "受注日", "金額"])   # ★ 先頭が混在文字（CJK+ASCII）
    ws.append(["J-1041", "2026-08-03", 100])
    wb.save(book)

    workdir = tmp_path / "work"
    workdir.mkdir()
    _catalog, helper_files = ailine.load_helpers(HELPERS_DIR)

    code = (
        "Option VBASupport 1\nOption Explicit\n\n"
        "Sub Run(oDoc As Object)\n"
        "    Call StyleBold(oDoc, 0, 0, 2, 0)\n"
        "End Sub\n"
    )
    # ★ 再現の鍵: 太字の前に LO の保存を一度通す（実機の事故は、先行 op で何度も
    #   LO 保存を経たブックで起きた。openpyxl 直後のブックでは再現しない ── 実測）。
    noop = (
        "Option VBASupport 1\nOption Explicit\n\n"
        "Sub Run(oDoc As Object)\n"
        "    oDoc.Sheets.getByIndex(0).getCellByPosition(5, 5).setString(\"x\")\n"
        "End Sub\n"
    )
    ok, err, raw = ailine.basrun_apply(book, noop, workdir, helper_files)
    assert ok, f"前段の LO 保存が失敗した: {err}"

    ok, err, raw = ailine.basrun_apply(book, code, workdir, helper_files)
    assert ok, f"basrun_apply が失敗した: {err}"

    ws2 = openpyxl.load_workbook(book).active
    got = {c.coordinate: c.font.bold for c in ws2[1] if c.value}   # noop が広げた空セルは対象外
    assert got == {"A1": True, "B1": True, "C1": True}, f"太字の読み戻し: {got}"

