# まだ表に無い行を、行番号で置く ── 2026-08-31。
# Namakoo が提案した実演の 1 幕目（丸山工業を追加して値を入れていく）を、
# 本番の前に俺が先に通したところ **5 件中 5 件が落ちた**。
#
# ★★ ① 名前の切り出しが行番号をまたいで飲み込んでいた
#     「**8行目に**丸山工業の行を作って」
#       → 『8行目に丸山工業』という行が見つかりません
#     `_re_row_of`（「〜の行」）の区切りが空白と読点しか無く、助詞をまたいで拾っていた。
#     ★ そのとき task_names_a_row_number は正しく 8 を返していた ──
#       **行番号が分かっているのに、名前の切り出しがそれを無視していた。**
#
# ★★ ② 名前が表に無いことを、断りの理由にしていた
#     「これから置く」行なのだから、名前が表に無いのは**当たり前**。
#     依頼文が行番号を名指ししているなら、それが場所。
#
# ★ ①の直しは語彙ではなく**文法の線**（助詞は名前に含まれない）。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HEADERS = ["取引先", "項目", "件数", "単価", "金額"]
ROWS = [["丸和物流", "配送", 12, 4800, 57600],
        ["近江スチール", "鋼材", 5, 12000, 60000],
        ["ヤマノ食品", "食品", 28, 1500, 42000]]


@pytest.fixture()
def meta(tmp_path):
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(p)
    return {"sheets": ["請求"], "headers": {"請求": list(HEADERS)},
            "header_rows": {"請求": 1}, "path": str(p)}


# --- ① 名前に助詞を飲み込まない ------------------------------------------------------------

@pytest.mark.parametrize("task, want", [
    ("8行目に丸山工業の行を作って", "丸山工業"),
    ("丸和物流の行を削除して", "丸和物流"),
    ("5行目にヤマノ食品の行を追加して", "ヤマノ食品"),
])
def test_the_row_name_stops_at_a_particle(task, want):
    m = ailine._re_row_of.search(task)
    assert m and m.group(1) == want, m.group(1) if m else None


# --- ② これから置く行は、行番号で場所が決まる ----------------------------------------------

def test_a_new_name_with_a_row_number_resolves(meta):
    """★★ 実演の 1 幕目そのもの ── 名前が表に無くても、行番号が在れば置ける。"""
    at, note = ailine.resolve_row_anchor("8行目に丸山工業の行を作って", meta, "請求")
    assert at == 8, (at, note)
    assert "8行目" in note


def test_a_new_name_without_a_row_number_still_refuses(meta):
    """★ 黙りすぎていないこと: 場所の手掛かりが無ければ、今までどおり断る。"""
    at, note = ailine.resolve_row_anchor("丸山工業の行を作って", meta, "請求")
    assert at is None
    assert "見つかりません" in (note or ""), note


def test_an_existing_name_still_wins(meta):
    """★ 表に在る名前は今までどおり中身で解く（行番号に横取りされない）。"""
    at, note = ailine.resolve_row_anchor("丸和物流の行を削除して", meta, "請求")
    assert at == 2, (at, note)


def test_the_header_row_is_not_a_place(meta):
    """★ 見出し行を場所にしない。"""
    at, _note = ailine.resolve_row_anchor("1行目に丸山工業の行を作って", meta, "請求")
    assert at != 1


# --- ③ 実物で（実演の 1 幕目を通す）--------------------------------------------------------

@pytest.mark.local
def test_the_first_act_of_the_demo_runs(tmp_path):
    """★★ 行を作って、セルを 1 つ埋めるところまで（実演の入口）。"""
    import subprocess
    p = tmp_path / "d.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請求"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)
    wb.save(p)
    env = {**__import__("os").environ, "PYTHONPATH": str(REPO / "src")}

    def _run(task):
        return subprocess.run(
            [sys.executable, "-m", "ailine", "run", str(p), task, "--sheet", "請求"],
            capture_output=True, text=True, encoding="utf-8", errors="replace",
            timeout=900, cwd=str(REPO), env=env)

    r1 = _run("5行目に丸山工業の行を作って")
    assert "✓" in r1.stdout, r1.stdout[-1200:]
    r2 = _run("丸山工業の件数を20にして")
    assert "✓" in r2.stdout, r2.stdout[-1200:]
    v = openpyxl.load_workbook(p, data_only=True)["請求"]
    assert v.cell(5, 1).value == "丸山工業"
    assert v.cell(5, 3).value == 20
