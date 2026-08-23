# CSV 検疫接続の自分の検体（凍結された tests/test_csv_cmd.py の3本の外側）。
#
# 対象: ⚠ 経路（undecidable 列・制御文字・長大セル・Excel 破壊済み検出・行数上限）・
# run の暗黙前段（成功時は既存機械へ継続／⚠ 時は続行しない）・own 印の再実行/衝突・
# フォルダ実行（scan/stack）の .csv 名指し。
#
# ★ tests/test_csv_truth_table.py・tests/test_csv_cmd.py の期待値・文言は変更しない。
# ここは自分の追加検体専用（brief の指示どおり別ファイル）。

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
import ailine  # noqa: E402
from ailine_core import csv_quarantine, xml_readback  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _write_csv(tmp_path, name, text, encoding="utf-8"):
    p = tmp_path / name
    p.write_bytes(text.encode(encoding))
    return p


# ===========================================================================
# ⚠ 経路: undecidable 列 → △（✓ は名乗らない）
# ===========================================================================

def test_csv_command_undecidable_column_demotes_check_to_triangle(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    # 日付と数値が同居する列（mixed_confident）── 確信クラス同士の衝突で undecidable。
    csv_path = _write_csv(tmp_path, "混在.csv",
                          "キー,値\n1,2026-01-02\n2,123\n")
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc == 0, out   # ★ 明示コマンドは常に書く（続行しないのは run の暗黙前段だけ）
    assert "△" in out, out
    assert "✓ 読み取った" not in out, out   # ✓ の転送claim行そのものは出ていない
    assert "⚠" in out, out
    assert "正しく読み込みました" not in out, out


# ===========================================================================
# セル単位の隔離: 制御文字・長大セル
# ===========================================================================

def test_csv_command_control_char_cell_is_removed_and_disclosed(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "制御文字.csv", "メモ\nok\na\x01b\n")
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc == 0, out
    assert "制御文字" in out and "U+0001" in out, out
    out_x = csv_path.with_suffix(".xlsx")
    grid = xml_readback.read_grid(out_x)["grid"]
    # 除去済み（"ab" になっている・落ちて save が死んでいない）
    assert grid[(3, 1)] == "ab", grid.get((3, 1))


def test_csv_command_overlong_cell_is_disclosed(tmp_path, monkeypatch, capsys):
    # ★ 致命①(2026-08-23レビュー)による golden 更新: 32,768文字セルは xlsx 書き込み時に
    # 32,767文字へ実際に切り詰められる（実測: 転送検算が不一致1を検出）。旧実装は
    # warn_count（この切り詰めの ⚠ 開示）だけを見て △ rc=0 を返し、compare_result（実際に
    # 1セルも変えずに書けたか）を見ていなかった ── 転送の主張自体が成立していないので、
    # ① の修正後は △ でなく × rc!=0 に倒す（該当セルを名指し）。
    _isolate(monkeypatch, tmp_path)
    long_cell = "x" * 32768
    csv_path = _write_csv(tmp_path, "長大.csv", f"メモ\n{long_cell}\n")
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc != 0, out
    assert "32,767" in out or "32768" in out, out
    assert "×" in out and "✓" not in out and "△" not in out, out
    assert "不一致1" in out.replace(" ", ""), out


# ===========================================================================
# Excel 破壊済み検出（設計 v2「壊れた後にも言える唯一の機能」）
# ===========================================================================

def test_csv_command_detects_digit_length_variance_as_excel_damage(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    # 桁数がバラバラな純数字列（先頭ゼロがすでに失われた痕跡の疑い）。
    csv_path = _write_csv(tmp_path, "コード変.csv", "コード\n12\n3\n456\n")
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc == 0, out
    assert "桁数がバラバラ" in out, out


# ===========================================================================
# 行数上限（先に数えて断る）
# ===========================================================================

def test_csv_command_refuses_when_row_count_exceeds_cap(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(csv_quarantine, "MAX_ROWS", 2)
    csv_path = _write_csv(tmp_path, "大量.csv", "a\n1\n2\n3\n4\n")
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc == 3, out
    assert "行数が多すぎます" in out, out
    assert not csv_path.with_suffix(".xlsx").exists()


# ===========================================================================
# own 印: 再実行（同じ原本）は上書き可・他人のファイルとの衝突は exit 7
# ===========================================================================

def test_csv_command_rerun_same_source_overwrites_without_refusal(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "再実行.csv", "a,b\n1,2\n")
    rc1, out1 = _run_main(["csv", str(csv_path)], capsys)
    assert rc1 == 0, out1
    rc2, out2 = _run_main(["csv", str(csv_path)], capsys)
    assert rc2 == 0, out2
    assert "出力先に書けません" not in out2, out2


def test_csv_command_refuses_foreign_output_conflict(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "衝突.csv", "a,b\n1,2\n")
    foreign = csv_path.with_suffix(".xlsx")
    wb = openpyxl.Workbook()
    wb.active.append(["人の", "ファイル"])
    wb.save(foreign)   # ailine 印なし
    rc, out = _run_main(["csv", str(csv_path)], capsys)
    assert rc == 7, out
    assert "出力先に書けません" in out, out
    # 人のファイルは無傷
    wb2 = openpyxl.load_workbook(foreign)
    assert wb2.active.cell(row=1, column=1).value == "人の"


# ===========================================================================
# run の暗黙前段（設計 v2 B）: ⚠ があれば続行しない／無ければ既存機械へ継続
# ===========================================================================

def test_run_csv_prestage_refuses_on_warning_without_touching_lo(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "確認要.csv", "キー,値\n1,2026-01-02\n2,123\n")

    def _trap_normalize(path, *a, **kw):
        raise AssertionError(f"normalize_book が csv 暗黙前段で呼ばれた: {path}")

    def _trap_basrun(*a, **kw):
        raise AssertionError("basrun_apply が csv 暗黙前段で呼ばれた")

    monkeypatch.setattr(ailine, "normalize_book", _trap_normalize)
    monkeypatch.setattr(ailine, "basrun_apply", _trap_basrun)
    monkeypatch.setattr(ailine, "translate_task",
                         lambda *a, **kw: (_ for _ in ()).throw(
                             AssertionError("translate_task が呼ばれた（続行してしまった）")))
    rc, out = _run_main(["run", str(csv_path), "値で並べ替えて"], capsys)
    assert rc == 3, out
    assert "⚠" in out, out
    assert "続行しません" in out, out
    # 検疫結果自体は書く（原本は無変更・失敗ではない）が、既存機械へは進まない。
    assert csv_path.with_suffix(".xlsx").exists()


def test_run_csv_prestage_continues_to_existing_pipeline_on_success(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "きれい.csv", "商品,金額\na,200\nb,300\nc,100\n")

    monkeypatch.setattr(ailine, "translate_task",
                         lambda model, task, book_meta, temperature=0.1:
                         {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    applied = {}

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        applied["path"] = Path(out_book)
        wb = openpyxl.load_workbook(out_book)
        ws = wb.active
        for i, (name, val) in enumerate([("b", 300), ("a", 200), ("c", 100)], start=2):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=val)
        wb.save(out_book)
        return True, None, "ok"

    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(csv_path), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    xlsx_path = csv_path.with_suffix(".xlsx")
    assert xlsx_path.exists(), out
    assert applied.get("path") is not None, "basrun_apply が呼ばれていない（既存機械に継続していない）"
    # 原本 CSV は無変更（バイト完全一致）。
    assert csv_path.read_bytes() == "商品,金額\na,200\nb,300\nc,100\n".encode("utf-8")


# ===========================================================================
# フォルダ実行（scan/stack）: .csv を分母に数えて名指しで断る
# ===========================================================================

def test_scan_reports_csv_count_and_ignores_them(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    wb = openpyxl.Workbook()
    wb.active.append(["a", "b"])
    wb.active.append([1, 2])
    wb.save(tmp_path / "本命.xlsx")
    _write_csv(tmp_path, "混じった.csv", "a,b\n1,2\n")
    rc, out = _run_main(["scan", str(tmp_path)], capsys)
    assert rc == 0, out
    assert ".csv" in out and "ailine csv" in out, out
    assert "1 件" in out, out


def test_stack_reports_csv_count_and_excludes_from_stacking(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    wb = openpyxl.Workbook()
    wb.active.append(["a", "b"])
    wb.active.append([1, 2])
    wb.save(tmp_path / "本命.xlsx")
    _write_csv(tmp_path, "混じった.csv", "a,b\n1,2\n")
    out_path = tmp_path / "stacked.xlsx"
    rc, out = _run_main(["stack", str(tmp_path), "--out", str(out_path)], capsys)
    assert rc == 0, out
    assert ".csv" in out and "ailine csv" in out, out
    # .csv は積まれていない（分母1・積んだ1 ── xlsx 1冊だけが対象）。
    assert "1 ファイル中 1 積んだ" in out, out


def test_stack_self_excludes_own_csv_output(tmp_path, monkeypatch, capsys):
    """own 印つき csv 出力（.xlsx）は V6 の自己参照除外の対象（種類を問わず ailine 産は除外）。"""
    _isolate(monkeypatch, tmp_path)
    csv_path = _write_csv(tmp_path, "元.csv", "a,b\n1,2\n")
    rc0, _ = _run_main(["csv", str(csv_path)], capsys)
    assert rc0 == 0
    wb = openpyxl.Workbook()
    wb.active.append(["c", "d"])
    wb.active.append([9, 9])
    wb.save(tmp_path / "本命.xlsx")
    out_path = tmp_path / "stacked.xlsx"
    rc, out = _run_main(["stack", str(tmp_path), "--out", str(out_path)], capsys)
    assert rc == 0, out
    assert "元" in out and "除外しました" in out, out
