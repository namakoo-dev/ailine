# operator 盲検 8 回目（$45・PRED-20260822-operator8.md）の BROKEN 3 件 ── 修正より先に
# 凍結した赤い検体。Namakoo「順次GO」(2026-08-22 09:59)。
#
# ①転記(LOOKUP_FILL): 事前行「操作するシート: 2枚目『単価表』」と解釈行「シート:『明細』」が
#   自己矛盾し、参照シートへの言及が ③ の ⚠（機械照合できません）として誤爆 → 自然な
#   言い回しが毎回確認往復になる（旗艦機能・2 回再現）
# ②税込 CLARIFY: label『税込み合計』が語彙 key『消費税』を字面に含まず照合失敗、なのに
#   「登録してください」と既に済んだ手を案内（誤誘導）
# ③見出し検出失敗: タイトル行つき請求書で「？ 列『数量』がありません。ある列: 2026年6月分
#   請求書」── --header-row への導線ゼロ（約束の CLARIFY でもない）

import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent))
from test_golden_transcripts import _isolate, _run_main  # noqa: E402
from test_ailine import _inject_formula_cache  # noqa: E402


# --- ① 転記: 参照シートだけを名指しした自然な依頼が、矛盾なく ✓ まで通る -------------

def _lookup_book(tmp_path):
    p = tmp_path / "lookup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.append(["商品", "単価"])
    ws.append(["りんご", None])
    ws.append(["みかん", None])
    ref = wb.create_sheet("単価表")
    ref.append(["商品", "単価"])
    ref.append(["りんご", 120])
    ref.append(["みかん", 80])
    wb.save(p)
    return p


def test_lookup_natural_phrasing_no_contradiction_no_gate(tmp_path, monkeypatch, capsys):
    """『単価表シートから単価を引っ張ってきて』（参照シートのみ言及の自然形）:
       - どの行も『単価表』を操作する(書き込む)シートだと主張しない
       - 参照シートへの言及は ③ の ⚠（機械照合できません）にならない
       - 確認往復なしで ✓ まで通る"""
    _isolate(monkeypatch, tmp_path)
    book = _lookup_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "LOOKUP_FILL",
         "args": {"target_sheet": "明細", "target_col": "単価",
                   "source_sheet": "単価表", "key_col": "商品"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["明細"]
        ws["B2"], ws["B3"] = 120, 80
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "単価表シートから単価を引っ張ってきて", "--copy"],
                         capsys)
    assert "機械照合できません" not in out, f"参照シートへの言及が ③ 誤爆している: {out}"
    assert "操作するシート: 2枚目『単価表』" not in out, \
        f"参照シートを操作対象と主張する行が残っている: {out}"
    assert rc == 0, f"確認往復（gate）に落ちている: {out}"
    assert "✓" in out, f"✓ まで通っていない: {out}"


# --- ② 税込: 登録済みの税率は label の字面照合に頼らず引き当てる -----------------------

def _tax_book(tmp_path):
    p = tmp_path / "inv.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["品目", "数量", "単価", "小計"])
    ws.append(["a", 1, 100, 100])
    ws.append(["b", 2, 50, 100])
    wb.save(p)
    return p


def _tax_apply(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ws["C4"] = "税込み合計"
    ws["D4"] = "=SUM(D2:INDEX(D:D,ROW()-1))*1.1"
    wb.save(out_book)
    _inject_formula_cache(out_book, "xl/worksheets/sheet1.xml", {"D4": 200 * 1.1})
    return True, None, "ok"


def _translate_append_total(model, task, book_meta, temperature=0.1):
    return {"op": "APPEND_TOTAL", "args": {"col": "小計", "label": "税込み合計"}}


def test_registered_tax_rate_is_found_for_natural_label(tmp_path, monkeypatch, capsys):
    """語彙に 消費税=1.1 が登録済みなら、『税込み合計を出して』（key を字面に含まない
       自然形）でも CLARIFY に落ちず、用語集由来の 1.1 を開示つきで適用する。"""
    _isolate(monkeypatch, tmp_path)
    rc0, _ = _run_main(["vocab", "add", "消費税", "1.1"], capsys)
    assert rc0 == 0
    book = _tax_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_append_total)
    monkeypatch.setattr(ailine, "basrun_apply", _tax_apply)
    rc, out = _run_main(["run", str(book), "税込み合計を出して", "--copy"], capsys)
    assert "倍率が分かりません" not in out, f"登録済みなのに CLARIFY に落ちた: {out}"
    assert rc == 0, out
    assert "1.1" in out and "用語集" in out, f"適用と出所の開示が無い: {out}"


def test_two_registered_tax_rates_clarify_names_candidates(tmp_path, monkeypatch, capsys):
    """税っぽい登録が 2 値あるなら勝手に選ばず CLARIFY ── ただし候補を名指しし、
       「登録してください」（既に済んだ手）とは言わない。"""
    _isolate(monkeypatch, tmp_path)
    rc0, _ = _run_main(["vocab", "add", "消費税", "1.1"], capsys)
    rc1, _ = _run_main(["vocab", "add", "軽減税率", "1.08"], capsys)
    assert rc0 == 0 and rc1 == 0
    book = _tax_book(tmp_path)
    monkeypatch.setattr(ailine, "translate_task", _translate_append_total)
    rc, out = _run_main(["run", str(book), "税込み合計を出して", "--copy"], capsys)
    assert rc == 3, out
    assert "消費税" in out and "軽減税率" in out, f"候補の名指しが無い: {out}"
    assert "登録してください" not in out, f"既に済んだ手を案内している: {out}"


# --- ③ 見出し検出失敗: 誤誘導でなく --header-row への導線を出す ------------------------

# ★ 検体の差し替え（2026-08-22 10:4x・測定器の修正）: 初版の凍結形は「検出が正しく 3 行目を
#   当てる形」で、operator の実事故（silently assumed row 1）を再現していなかった
#   （operator の実物ファイル タイトル付き請求書.xlsx は現行コードで正しく検出することを
#   実 LO で確認済み）。実事故の機構は resolve_header_rows の「StructDump に対象シートが
#   無ければ無言で 1 行目」フォールバック（LO の一時不調で発生しうる・バックログ⑥と同根）。
#   契約は「検出の成功している道に確認を足す」ことでは**なく**、
#   ①フォールバックで 1 行目とみなした事実を開示する
#   ②列解決が失敗したとき、他の行に見出しらしき同名セルを見つけたら --header-row R へ誘導する

def test_column_not_found_points_to_header_row_when_struct_dump_is_unavailable(tmp_path, monkeypatch, capsys):
    """StructDump が使えず無言で 1 行目とみなした結果『列がありません』になる実事故の形:
       誤誘導で終わらず、3 行目の見出しらしき『数量』を見つけて --header-row 3 へ誘導する。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "請求書.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["2026年6月分請求書"])
    ws.append([])
    ws.append(["商品", "数量", "単価"])
    ws.append(["a", 2, 100])
    ws.append(["b", 3, 200])
    wb.save(p)
    monkeypatch.setattr(ailine, "build_struct_dump",
                         lambda normalized_book, workdir: {})   # LO 不調の再現（無言の 1 行目化）
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "数量", "order": "desc"}})
    rc, out = _run_main(["run", str(p), "数量で降順に並べ替えて", "--copy"], capsys)
    assert rc == 3, out
    assert "--header-row 3" in out, f"見出し行への導線が無い: {out}"


def test_confident_detection_stays_silent_no_new_gate(tmp_path, monkeypatch, capsys):
    """検出が確信を持って成功した道（同じタイトル付きの形・StructDump あり）は、
       確認を挟まずそのまま通る ── 成功している道に新しい摩擦を足さない。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "請求書.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["2026年6月分請求書"])
    ws.append([])
    ws.append(["商品", "数量", "単価"])
    ws.append(["a", 2, 100])
    ws.append(["b", 3, 200])
    wb.save(p)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "数量", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        ws2["A4"], ws2["B4"], ws2["C4"] = "b", 3, 200
        ws2["A5"], ws2["B5"], ws2["C5"] = "a", 2, 100
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "数量で降順に並べ替えて", "--copy"], capsys)
    assert "--header-row" not in out, f"確信のある検出に確認を挟んでいる（新しい摩擦）: {out}"
    assert rc == 0, out


def test_struct_dump_unavailable_discloses_silent_row1_assumption(tmp_path, monkeypatch, capsys):
    """③ の契約補完（自分の検体・文言は縛らない）: StructDump が使えず無言で1行目と
       みなした run は、たとえ見出しが実際に1行目にあって列解決が成功しても、
       その仮定を開示すること ── 無言の仮定は開示する（成功/失敗どちらの道でも）。"""
    _isolate(monkeypatch, tmp_path)
    p = tmp_path / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"])
    ws.append(["a", 300])
    ws.append(["b", 200])
    wb.save(p)
    monkeypatch.setattr(ailine, "build_struct_dump",
                         lambda normalized_book, workdir: {})   # LO 不調の再現
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb2 = openpyxl.load_workbook(out_book)
        ws2 = wb2.active
        ws2["A2"], ws2["B2"] = "a", 300
        ws2["A3"], ws2["B3"] = "b", 200
        wb2.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(p), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
    assert "1行目" in out and "自動検出" in out, f"無言の仮定が開示されていない: {out}"


def test_compound_plan_lookup_announce_no_contradiction(tmp_path, monkeypatch, capsys):
    """① の複合計画経路（片配線防止・幹の系譜）: 計画経由の LOOKUP_FILL でも
       参照シートを操作対象と主張する行を出さず、③ 誤爆もしない。"""
    _isolate(monkeypatch, tmp_path)
    book = _lookup_book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"plan": [{"op": "LOOKUP_FILL",
                    "args": {"target_sheet": "明細", "target_col": "単価",
                              "source_sheet": "単価表", "key_col": "商品"}}]})

    def fake_apply(out_book, code, workdir, helper_files=(), timeout=None):
        wb = openpyxl.load_workbook(out_book)
        ws = wb["明細"]
        ws["B2"], ws["B3"] = 120, 80
        wb.save(out_book)
        return True, None, "ok"
    monkeypatch.setattr(ailine, "basrun_apply", fake_apply)
    rc, out = _run_main(["run", str(book), "単価表シートから単価を引っ張ってきて", "--copy"],
                         capsys)
    assert "機械照合できません" not in out, f"複合経路で ③ 誤爆（片配線）: {out}"
    assert "操作するシート: 2枚目『単価表』" not in out,         f"複合経路で参照シートを操作対象と主張（片配線）: {out}"
    assert rc == 0, out
