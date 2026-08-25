# 復元の重大6（2026-08-24 の盲検）── 原本への反映に失敗しても exit 0 だった。
#
# ★ 実測: バックアップ先が書けない状態で run すると
#     × バックアップに失敗したため --inplace を中止した（原本は無変更）: [WinError 183] …
#     （原本 nb.xlsx は変更していません。作業結果は nb.out.xlsx に残っています）
#     REAL exit=0        ← パイプ無しで測定
#   `_finish_apply` は置換の成否を bool で返すのに、**呼び出し元 4 箇所が全部捨てていた**。
#   ★ 表示は正直だった（「原本は無変更」と言っている）。嘘だったのは**終了コード**だけ。
#   スクリプトから ailine を回す利用者は「反映されなかった」を検出できない。
#
# 契約:
#   ① 反映が中止されたら非零で返す
#   ② 原本は 1 バイトも変わらない（従来どおり）
#   ③ --copy（原本に触らない経路）は影響を受けない

import argparse
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _book(tmp_path):
    p = tmp_path / "nb.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品", "金額"]); ws.append(["a", 100]); ws.append(["b", 250])
    wb.save(p)
    return p


def _sorted_apply(out_book, code, workdir, helper_files=(), timeout=None):
    wb = openpyxl.load_workbook(out_book)
    ws = wb.active
    ws["A2"], ws["B2"], ws["A3"], ws["B3"] = "b", 250, "a", 100
    wb.save(out_book)
    return True, None, "ok"


def test_apply_failure_is_not_reported_as_success(tmp_path, monkeypatch, capsys):
    """①② 反映が中止されたら非零・原本は無変更。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    before = book.read_bytes()
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply", _sorted_apply)
    monkeypatch.setattr(ailine, "atomic_replace_inplace",
                        lambda *a, **k: (False, "バックアップに失敗したため中止した"))
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて"], capsys)
    assert rc != 0, f"反映していないのに成功を名乗った: exit={rc} / {out}"
    assert book.read_bytes() == before, "中止したのに原本が変わった"


def test_copy_path_is_unaffected(tmp_path, monkeypatch, capsys):
    """③ --copy は原本に触らないので、この関所の対象外。"""
    _isolate(monkeypatch, tmp_path)
    book = _book(tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "SORT", "args": {"col": "金額", "order": "desc"}})
    monkeypatch.setattr(ailine, "basrun_apply", _sorted_apply)
    rc, out = _run_main(["run", str(book), "金額で降順に並べ替えて", "--copy"], capsys)
    assert rc == 0, out
