# 土台固め（2026-08-24）── 「運ぶもの」と「運ばないもの」を正確に言う。
#
# ★ Namakoo の問い:「この全滅のうち VBA を除いたものは仕様で問題はないということか？」
#   「むしろ全滅が仕様というと少し複雑な感じもする」
#   → 測ったら **2 つ崩れた**。
#
#   ① 俺が書いた「元の書式は持ち越しません」は**嘘**だった。数値書式（#,##0 も
#      yyyy-mm-dd も）は運んでいる。設計文書にも「日付書式の引き継ぎ」と書いてあった。
#      測らずに一般化して、その嘘を製品の表示文にした。
#      ★ 入口は測定器の欠落 ── 生存表に「数値書式」の列が無かった。
#        測っていない列は、無いのではなく**見えていない**。
#   ② コメントとハイパーリンクは**飾りでなく中身**（「要確認: 入金待ち」のような、
#      人が打った情報）。持ち越さないのは同じでも、黙って消すのは違う。
#
# 契約:
#   A 縦積み/フォルダ抽出は数値書式を運ぶ（「運ばない」と言わない）
#   B コメント/ハイパーリンクが元に在れば、運んでいないと**名指しで**言う
#   C 両経路（stack / run フォルダ抽出）が同じことを言う（片配線の禁止）
#   D 元に無ければ 1 文字も言わない（誤爆しない）

import datetime
import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parent))
import ailine  # noqa: E402
from test_golden_transcripts import _isolate, _run_main  # noqa: E402


def _folder(tmp_path, with_notes=True):
    d = tmp_path / "books"
    d.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["取引先", "金額", "日付", "備考"])
    ws.append(["あかつき商事", 12000, datetime.date(2026, 7, 31), "リンク"])
    ws.append(["みどり工業", 8500, datetime.date(2026, 8, 15), "メモ"])
    ws["B2"].number_format = "#,##0"
    ws["B3"].number_format = "#,##0"
    if with_notes:
        ws["D2"].hyperlink = "https://example.invalid/inv1"
        ws["D3"].comment = Comment("要確認: 入金待ち", "経理")
    wb.save(d / "a.xlsx")
    return d


# --- A 数値書式は運ぶ（「運ばない」と言わない）----------------------------------------

def test_stack_carries_number_formats(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    out = tmp_path / "stacked.xlsx"
    rc, text = _run_main(["stack", str(_folder(tmp_path)), "--out", str(out)], capsys)
    ws = openpyxl.load_workbook(out).active
    fmts = {c.number_format for row in ws.iter_rows(min_row=2) for c in row}
    assert "#,##0" in fmts, f"桁区切りが落ちた: {fmts}"
    assert any("yy" in f for f in fmts), f"日付書式が落ちた: {fmts}"


def test_stack_does_not_claim_it_drops_number_formats(tmp_path, monkeypatch, capsys):
    """★ 嘘の再発防止: 「元の書式は持ち越しません」と一般化して言わないこと。"""
    _isolate(monkeypatch, tmp_path)
    out = tmp_path / "stacked.xlsx"
    rc, text = _run_main(["stack", str(_folder(tmp_path)), "--out", str(out)], capsys)
    assert "元の書式・数式・図形は持ち越しません" not in text, \
        f"数値書式を運んでいるのに『書式は持ち越さない』と言った: {text}"
    assert "値と数値書式は運びます" in text, text


# --- B/C コメントとリンクは名指しで開示（両経路）-----------------------------------------

def test_stack_names_dropped_comments_and_links(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    out = tmp_path / "stacked.xlsx"
    rc, text = _run_main(["stack", str(_folder(tmp_path)), "--out", str(out)], capsys)
    assert "コメント 1 件" in text, f"人が打ったメモが黙って消えた: {text}"
    assert "ハイパーリンク 1 件" in text, text


def test_folder_extract_names_the_same_things(tmp_path, monkeypatch, capsys):
    """C: 片配線の禁止 ── stack だけ直して run が黙るのを止める（実測でそうなった）。"""
    _isolate(monkeypatch, tmp_path)
    monkeypatch.setattr(
        ailine, "translate_task",
        lambda model, task, book_meta, temperature=0.1:
        {"op": "EXTRACT", "args": {"col": "金額", "cmp": "gte", "value": 5000}})
    rc, text = _run_main(["run", str(_folder(tmp_path)), "金額が5000以上の行を抜き出して"], capsys)
    assert "コメント 1 件" in text, f"フォルダ抽出が黙っている（片配線）: {text}"
    assert "ハイパーリンク 1 件" in text, text


# --- D 誤爆しない -------------------------------------------------------------------

def test_silent_when_there_is_nothing_to_drop(tmp_path, monkeypatch, capsys):
    _isolate(monkeypatch, tmp_path)
    out = tmp_path / "stacked.xlsx"
    rc, text = _run_main(["stack", str(_folder(tmp_path, with_notes=False)),
                           "--out", str(out)], capsys)
    assert "運んでいません" not in text, f"何も無いのに警告した: {text}"
