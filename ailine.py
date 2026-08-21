#!/usr/bin/env python
"""ailine — 自然言語のタスクを、ローカル LLM が LibreOffice Basic に書き起こし、
   basrun で文書に適用し、★ 効果を読み戻して検証する（「走った ≠ できた」）。

    ailine run  <book> "<タスク>"          生成 → 検品ゲート → 原本に反映 → 変化検証 → 修復 → 差分表示
    ailine run  <book> "<タスク>" --dry     生成して見せるだけ（適用しない・レビュー用）
    ailine run  <book> "<タスク>" --copy    原本には触らず <book>.out に結果を作る（旧既定）
    ailine undo <book>                      直前の反映を取り消す（世代バックアップから復元）
    ailine stop                             起動した LibreOffice を落とす（basrun に委譲）

## 設計判断（basrun_spike 2026-08-10 の実証に基づく）

- **モデル非依存**: `--model` で差し替え（既定 qwen2.5-coder:7b）。天井はモデルの大きさでなく
  「参照例の供給＋効果の検証」で上げる（7B が正解例1本で苦手層 0%→67% になった）。
- ★ **検証をループに**: 適用の前後で文書が変化したかを見る **no-op ガード**。
  LibreOffice + LLM は「実行時エラー無しで成功と報告し、実際は何もしない」ことがある
  （もっともらしい UNO の幻覚）。変化ゼロなら失敗として修復に回す。
- ★ **正規化パス（2026-08-14 修正・製品の心臓）**: before スナップショットの前に、
  コピーを LibreOffice で一度（空マクロで）開いて保存する。openpyxl 製ブックは LO の
  初回保存で行高（時に列幅）を実体化する副作用があり、これを先に済ませておかないと
  no-op ガードが偽陽性になる（何もしないマクロでも「変化した」と誤って成功表示していた）。
  コストは LO 往復 1 回（数秒）— 正しさ優先で受け入れる。
- ★ **既定=原本にそのまま反映（W8b-2）**: `<book>` に直接書く。「壊さない」の担保は
  もう「コピーにしか書かない」ことではなく、3重の安全網でまかなう:
  ①往復忠実度ゲート（LO 往復だけで失われる飾りを検品して申告・`--accept-loss`/`--copy` で選ぶ）
  ②世代バックアップ+`ailine undo`（反映前に必ず退避・いつでも戻せる）
  ③ no-op ガード＋事後条件などの機械検証（変化・正しさを見る）。
  従来の「コピーにしか書かない」挙動は `--copy` で選べる（原本は無変更・`<book>.out.xlsx` に生成）。
- **参照ライブラリ**: `refs/*.bas` を few-shot に供給。苦手層（新シート・色）を補う。
  並べ替え・グラフ・ピボットなどの難所は `helpers/*.bas` の検証済みヘルパを `Call` で呼ばせる。
- **レビュー導線**: 生成した .bas と、変わったセルの差分を必ず表示する。

## 正直な限界

- LibreOffice Basic + UNO は学習データが薄く、珍しい操作は外しやすい。参照とヘルパで補う設計
  （太字も当初「環境不可」と誤断したが、実際は `CharWeight`+`CharWeightAsian` の native 書きで解決済み）。
- ローカル LLM(ollama) と LibreOffice(basrun 経由) が要る。**外部送信はしない。**
- no-op ガードが保証するのは「変化したこと」だけ。「**正しいか**」は差分を人が見て判断する。
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    sys.exit("openpyxl が要る:  pip install openpyxl")

# ★ C4（再設計 分割の一歩目）: 新しい単位は ailine.py に足さず ailine_core/ に置き、
#   ここからは import するだけにする（tests/ailine_py_line_budget.txt が単調減少を見張る）。
#   ailine_core は ailine.py と同じディレクトリの兄弟パッケージなので、python が script
#   実行時に自動で sys.path[0] に入れる ailine.py の所在ディレクトリからそのまま拾える
#   （別の作業ディレクトリから `python C:\...\ailine.py run ...` と叩いても同じ）。
from ailine_core.book_view import BookView
from ailine_core.claim import (   # ★ C5/C9: Claim 型と『✓』の一元レンダラ（✓ は反映後の1箇所だけ）
    Claim, format_plan_report, format_plan_preview, overall_verdict,
    render_applied_claim, render_applied_unverified, render_applied_unobservable,
    render_scope_notes,   # ★ 単位E: 常時注記を廃止し、その run 固有の②の1文に置き換えた
)
from ailine_core.dsl_step import (   # ★ C7: 単発 DSL / 複合計画の DSL 段が共有する実行エンジン
    DslStepDeps, resolve_dsl_step_args, print_dsl_confirmation, apply_dsl_step, compose_dsl_step_advisories,
    NEW_COLUMN_ORIGIN,   # ★ 単位B: 「直前の段が作った列」の文言の唯一の出どころ
)
from ailine_core.cli_render import (   # ★ C8: 複数経路が同じ形を手書きしていた表示の純関数化
    render_code_block, render_retry_options, render_aborted, render_run_header,
    render_backup_list, render_restore_done, render_vocab_add_result, render_vocab_listing,
    render_ops_table,
    freeform_notice_reason, render_freeform_notice, render_freeform_notice_compact,   # ★ K-1
    render_scan_report,   # ★ M1読み: `ailine scan`
    render_stack_report, render_verify_report,   # ★ M1書き: `ailine stack` / `ailine verify`
)
from ailine_core import multifile   # ★ M1読み: 多ファイル棚卸し（DESIGN-20260821-multifile.md）
from ailine_core import stack as multifile_stack   # ★ M1書き: 縦積み本体（DESIGN v2 §1 M1書き）
from ailine_core import verify as multifile_verify   # ★ M1書き: `ailine verify` の検算本体
from ailine_core import xml_readback   # ★ 検算の独立読み実装（openpyxl を import しない別実装）
from ailine_core import extract_multi   # ★ M2: `ailine run <フォルダ>`（抽出集約）の本体
from ailine_core.formula_health import formula_error_advisory, detect_write_target_type_change   # ★ 挙動変更#1(a)(b)
from ailine_core.write_precondition import (   # ★ 単位F/G: 宣言した領域の前提（破れた種類つき）
    check_write_preconditions_detail,
    own_prior_output_notice_lines,   # ★ 単位H 開示: 関所が黙った理由を1行で見せる
)
from ailine_core.sum_identity import rows_matching_sum_above   # ★ 算術恒等の検算（二重計上）
from ailine_core.target_sheet import (   # ★ 挙動変更#2/#3: 対象シートの決定を一箇所に閉じ込める
    resolve_target_sheet, describe_target_sheet, wrap_basic_for_sheet,
    format_sheet_field, sheet_conflict_choice_lines, conflict_excluded_sheets,
    sheet_names_mentioned_in,   # ★ 単位E: シート名照合の素材（決定側と助言側が共有する）
)
from ailine_core.subject import (   # ★ 単位E: A' 原則を「値」から「対象スロット」へ広げる
    Slot, Consumed as SubjectConsumed, classify_slots,
    COLUMN as SUBJ_COLUMN, REGION as SUBJ_REGION, ROW as SUBJ_ROW,
    SHEET as SUBJ_SHEET, LABEL as SUBJ_LABEL, INPUT as SUBJ_INPUT,
    name_matches_task,   # ★ W3 改定(2026-08-20): 実在しない target が「依頼文の名指し」か
                          #   「翻訳の捏造」かの照合に、単位B の部分文字列規律を再利用する
)
from ailine_core.interpretation import build_interpretation   # ★ 段1: 解釈を機械可読で出す（--json の interpretation/provenance）
from ailine_core.ask_choice import (   # ★ 挙動変更#3: 「選択肢を出して選ばせる」対話部品
    Choice, ask_choice, ask_yes_no, is_interactive,
)
HERE = Path(__file__).resolve().parent
DEFAULT_REFS = HERE / "refs"
DEFAULT_HELPERS = HERE / "helpers"
OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
DEFAULT_MODEL = os.environ.get("AILINE_MODEL", "qwen2.5-coder:7b")
DEFAULT_APPLY_TIMEOUT = 180.0  # M1: 暴走マクロで無限ハングしないよう既定 ON（--timeout 0 で無効化）

HISTORY_DIR = Path.home() / ".ailine"
HISTORY_FILE = HISTORY_DIR / "history.jsonl"
BACKUP_DIR = HISTORY_DIR / "backups"
DEFAULT_KEEP_BACKUPS = 10   # M2c: book ごとにこの世代数を超えたら古い順に削除する

# ★ W8b 項目6: グローバル run ロック。基盤の LibreOffice(basrun 経由)が単一インスタンス
#   (port 2002)前提のため、ブック単位でなく `ailine run` 全体で1本にする。
RUN_LOCK_FILE = HISTORY_DIR / "run.lock"
RUN_LOCK_STALE_SECONDS = 30 * 60   # 30分超のロックは stale とみなして奪取する

# ★ W10a 項目2: 既定変更(W8b-2・原本直接反映)の一度きり告知。marker ファイルの有無だけで
#   判定する（history 等には依存しない・単純に「見せたか」の1ビット）。
NOTICE_V2_FILE = HISTORY_DIR / "notice_v2_shown"
NOTICE_V2_TEXT = (
    "★ このバージョンから、既定で原本に直接反映します（自動バックアップ+undo つき）。"
    "従来のコピー方式は --copy。この告知は一度だけ表示されます"
)

# ★ A': 用語集（税率等の「現場の取り決め値」）。グローバルのみ・ブック別上書きは作らない
#   （YAGNI・受信ファイル注入の予防。サイドカー自動読みは絶対にしない）。
VOCAB_FILE = HISTORY_DIR / "vocab.json"
DEFAULT_VOCAB_MAX_ENTRIES = 200     # 個人利用の上限。無制限にしない
DEFAULT_VOCAB_MAX_TERM_LEN = 40     # 語（キー）の最大長


def _find_basrun_path() -> Path | None:
    """basrun.py の場所。環境変数 BASRUN > ailine と並びの checkout の順で探す。
       見つからなければ None（sys.exit しない版。doctor から非致命的に使う）。"""
    env = os.environ.get("BASRUN")
    if env:
        p = Path(env)
        return p if p.exists() else None
    for name in ("basrun", "nagi-bas"):  # 公開 repo 名 / 作者ローカルの旧ディレクトリ名
        p = HERE.parent / name / "basrun.py"
        if p.exists():
            return p
    return None


def basrun_path() -> Path:
    """basrun.py の場所。無ければ理由つきで落とす（run から使う致命版）。"""
    p = _find_basrun_path()
    if p is None:
        sys.exit("basrun.py が見つからない: ailine と並びに"
                 " https://github.com/namakoo-dev/basrun を clone するか、"
                 "環境変数 BASRUN でパスを指定する")
    return p


# ---------------------------------------------------------------------------
# 進捗表示（M1: 生成中の完全沈黙を解消。凝った演出はしない — Windows コンソール安全第一）
# ---------------------------------------------------------------------------

def _fmt_elapsed(seconds: float) -> str:
    """経過秒を表示用に整形する（例: 12.34 → '(12.3s)'）。"""
    return f"({seconds:.1f}s)"


def progress_start(label: str) -> float:
    """進捗の開始行を stderr に出す（改行しない＝完了時に経過秒を追記する）。
       --json のときも常に stderr（stdout の機械出力を汚さない）。
       開始時刻(time.monotonic())を返す。"""
    print(label, end="", file=sys.stderr, flush=True)
    return time.monotonic()


def progress_end(start: float) -> None:
    """開始行に経過秒を追記して改行する。"""
    print(" " + _fmt_elapsed(time.monotonic() - start), file=sys.stderr)

CONTRACT = """あなたは LibreOffice Basic を書く。出力は .bas のコードだけ。説明・markdown 柵は禁止。

厳守する契約:
- 先頭に `Option VBASupport 1` と `Option Explicit`。
- 手続きは **ちょうど1つ**、必ず `Sub Run(oDoc As Object)` という名前と署名。
- `ThisComponent` は使わない。文書は引数 oDoc で受け取る。
- シート: `oDoc.Sheets.getByIndex(0)`。セルは 0 起点 `getCellByPosition(列, 行)`。
  値は `.getValue()/.setValue()`、文字は `.getString()/.setString()`。
- 数値書式は `oDoc.getNumberFormats()` の queryKey/addNew を取り `範囲.NumberFormat = nFmt`。
- 列を文字("A")で指す API は使わない（例外で静かに止まる）。数値の列番号だけ。
- ★ 新しいシートを作らず、既存シートの中の最小の変更で達成することを最優先する
  （依頼が明示的にシート新設・ピボットを求めていない限り）。
"""

# セルの状態を snapshot する際の使用範囲の上限（病的に巨大な文書での暴走を防ぐ）
MAX_ROWS = 1000
MAX_COLS = 64


# ---------------------------------------------------------------------------
# ローカル LLM（ollama）
# ---------------------------------------------------------------------------

def ollama_generate(model: str, messages: list, temperature: float = 0.2) -> str:
    body = json.dumps({
        "model": model, "messages": messages, "stream": False,
        "options": {"temperature": temperature, "num_predict": 1600, "num_ctx": 8192},
    }).encode()
    req = urllib.request.Request(f"{OLLAMA}/api/chat", data=body,
                                 headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=300) as r:
            d = json.load(r)
    except urllib.error.HTTPError as e:
        # ★ HTTPError は URLError のサブクラスなので先に拾う。
        #   404 は「繋がっているがモデルが無い」で、接続不能とは原因も対処も別。
        if e.code == 404:
            sys.exit(f"ollama にモデル '{model}' が見つからない (HTTP 404)。\n"
                     f"★ `ollama pull {model}` で取得してから再実行して。")
        sys.exit(f"ollama がエラーを返した ({OLLAMA}): HTTP {e.code} {e.reason}")
    except urllib.error.URLError as e:
        sys.exit(f"ollama に繋がらない ({OLLAMA}): {e}\n"
                 "★ `ollama serve` が動いているか確認。外部送信はしない設計。")
    return d.get("message", {}).get("content", "")


def load_refs(refs_dir: Path) -> str:
    """refs/*.bas を few-shot テキストに連結する。"""
    if not refs_dir.is_dir():
        return ""
    chunks = []
    for f in sorted(refs_dir.glob("*.bas")):
        chunks.append(f"--- 参考例: {f.stem} ---\n{f.read_text(encoding='utf-8').strip()}")
    if not chunks:
        return ""
    return "\n\nこれらは正しい書き方の参考（別タスク）:\n" + "\n".join(chunks) + "\n--- 参考ここまで ---\n"


def load_helpers(helpers_dir: Path) -> tuple:
    """helpers/*.bas を (プロンプト用カタログ, ファイル一覧) にする。
       ★ 難所（ソートの ContainsHeader 等）は人が検証したヘルパに閉じ込め、
          モデルには『Call で呼ぶだけ』させる。arcane 層の確度を上げる要。"""
    files = sorted(helpers_dir.glob("*.bas")) if helpers_dir.is_dir() else []
    if not files:
        return "", []
    srcs = "\n".join(f.read_text(encoding="utf-8").strip() for f in files)
    catalog = (
        "\n\n## 定義済みヘルパ（★ 呼ぶだけ・再定義しない）\n"
        "arcane な操作（並べ替え等）は、自分で書かず次のヘルパを使うこと。\n"
        "★ 呼び方は必ず `Call 名前(引数)` の形（Call を付ける。括弧つきで Call 無しは誤動作する）。\n"
        "★ ヘルパの中身は絶対に書き写すな（SummaryTable 等が長くても）。必ず `Call 名前(...)` の1行だけで呼ぶ。\n"
        "★ W10b: 依頼に無い操作のヘルパは絶対に呼ぶな（『何か効くかもしれない』で総当たりに"
        "色々呼ぶのは禁止。依頼を達成するのに要る分だけを呼ぶ）。\n"
        "★ headerRow 引数は見出し行（0起点）。見出しが物理1行目ならほぼ常に 0。\n"
        "例: 列0〜4の表で金額が列1なら、金額で降順に並べ替え"
        " → `Call SortByColumn(oDoc, 0, 4, 1, False)`（第2引数 lastCol=表の最終列）\n"
        "例: 金額(列1)の棒グラフ（項目名は先頭列に自動）→ `Call InsertBarChart(oDoc, 0, 1)`\n"
        "例: A1とB1を結合 → `Call MergeCells(oDoc, 0, 0, 1, 0)`\n"
        "例: 先頭データ行(2行目)の前に1行挿入 → `Call InsertRows(oDoc, 1, 1)`\n"
        "例: 表に罫線を引く → `Call DrawTableBorders(oDoc)`\n"
        "例: 各列の幅を内容に合わせる → `Call AutoFitColumns(oDoc)`\n"
        "例: C列(列2)に、商品名(列0)をキーに『単価表』から値を引く（VLOOKUP相当）"
        " → `Call VLookupFromTable(oDoc, 0, 0, 2, \"単価表\")`（参照表は 列0=キー・列1=値）\n"
        "例: 『ピボット』で部門(列0)ごとに金額(列1)を集計（本物の DataPilot・Excel で操作可）"
        " → `Call PivotSum(oDoc, 0, 1)`\n"
        "例: 『集計表／まとめ』を作る＝部門(列0)ごとの金額(列1)を見栄えのする普通の表に"
        "（罫線・カンマ・太字つき。★『ピボット』と明示されない集計は基本こちら）"
        " → `Call SummaryTable(oDoc, 0, 0, 1)`\n"
        "例: 見出し行(行0, 列0〜4)を太字に → `Call StyleBold(oDoc, 0, 0, 4, 0)`\n"
        f"--- 定義済み（この通り既に存在する。再定義するな）---\n{srcs}\n--- ここまで ---\n")
    return catalog, files


def extract_bas(text: str) -> str:
    m = re.search(r"```(?:\w+)?\s*(.*?)```", text, re.S)
    return (m.group(1) if m else text).strip()


def valid_signature(code: str) -> bool:
    return re.search(r"Sub\s+Run\s*\(\s*oDoc\s+As\s+Object\s*\)", code, re.I) is not None


#  ★ 開始の "Sub 名前(...)" だけを数える（"End Sub" 内の "Sub" を誤って開始として
#     二重カウントしないよう、識別子が続く形に絞る）。
_SUB_OPEN_RE = re.compile(r"\bSub\s+[A-Za-z_]\w*", re.I)
_ENDSUB_RE = re.compile(r"\bEnd\s+Sub\b", re.I)


def is_truncated_code(code: str) -> bool:
    """生成 Basic が途中で切断された兆候を、構造だけを見て検出する（M2a）。
       ★ 正しさは判定しない。CONTRACT は『手続きはちょうど1つ、必ず End Sub で閉じる』
       前提のため、Sub/End Sub の対応と『最後の行がちょうど End Sub か』だけを見れば、
       開き括弧や識別子で途中生成が止まった典型パターン（＝最後の行が End Sub でない）
       を漏れなく拾える。bad_signature（署名自体が無い）とは別の失敗分類。"""
    stripped = code.strip()
    if not stripped:
        return True
    sub_count = len(_SUB_OPEN_RE.findall(stripped))
    endsub_count = len(_ENDSUB_RE.findall(stripped))
    if sub_count == 0 or sub_count != endsub_count:
        return True
    last_line = stripped.splitlines()[-1].strip()
    return not bool(_ENDSUB_RE.fullmatch(last_line))


# ---------------------------------------------------------------------------
# 文書の説明と snapshot（検証の土台）
# ---------------------------------------------------------------------------

def describe_book(path: Path) -> str:
    wb = openpyxl.load_workbook(path, read_only=True)
    lines = [f"シート一覧: {wb.sheetnames}（1枚目 = {wb.sheetnames[0]!r}）"]
    ws = wb[wb.sheetnames[0]]
    nrow, ncol = ws.max_row or 0, ws.max_column or 0
    lines.append(f"1枚目のデータ範囲: 約 {nrow} 行 x {ncol} 列（列は 0 起点で 0..{max(ncol-1,0)}）。")
    headers = []
    for c in range(1, min(ncol, MAX_COLS) + 1):
        v = ws.cell(row=1, column=1 + (c - 1)).value
        if v not in (None, ""):
            headers.append(f"列{c-1}={v}")
    if headers:
        lines.append("行0(見出し): " + ", ".join(headers))
    lines.append("行1以降がデータ。")
    wb.close()
    return "\n".join(lines)


def book_columns(path: Path, header_rows: dict | None = None) -> dict:
    """全シートの見出し行を {シート名: [列名,...]} で返す。
       ★ describe_book は1枚目だけの人間可読版。こちらは M2b の翻訳・検証が使う
       機械可読の接地情報（列は最初の空欄で打ち切る＝連続した見出しだけを列とみなす）。
       ★ W3: header_rows({シート名: 見出し行(1起点)}) を渡すと、そのシートだけは
       物理1行目でなく指定行を見出しとして読む（StructDump の見出し検出結果を反映する）。
       未指定のシートは従来どおり1行目（後方互換・引数省略時は完全に旧挙動）。"""
    header_rows = header_rows or {}
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        hr = header_rows.get(name, 1)
        ncol = min(ws.max_column or 0, MAX_COLS)
        headers = []
        for c in range(1, ncol + 1):
            v = ws.cell(row=hr, column=c).value
            if v in (None, "") and hr > 1:
                # ★ W3: 多段見出し対策。子見出し行(hr>1)で空欄の列は、真上の行を遡って
                #   最初に見つかる非空値を引き継ぐ（D検体: 商品名は1行目だけにあり
                #   2行目(子見出し)の同じ列は空。無いと打ち切り誤検知で列挙が0件になる）。
                for up in range(hr - 1, 0, -1):
                    uv = ws.cell(row=up, column=c).value
                    if uv not in (None, ""):
                        v = uv
                        break
            if v in (None, ""):
                break
            headers.append(str(v))
        out[name] = headers
    wb.close()
    return out


def build_book_meta(path: Path, header_rows: dict | None = None) -> dict:
    """{"sheets": [...], "headers": {シート名: [列名,...]}, "header_rows": {シート名: 行(1起点)}}。
       M2b 翻訳・検証の接地情報。★ W3: header_rows 省略時は全シート1行目（旧挙動と同一）。"""
    header_rows = dict(header_rows or {})
    headers = book_columns(path, header_rows)
    resolved_header_rows = {name: header_rows.get(name, 1) for name in headers}
    return {"sheets": list(headers.keys()), "headers": headers, "header_rows": resolved_header_rows}


# ---------------------------------------------------------------------------
# W3 Part1/2: StructDump（LibreOffice の目で構造を読む）+ 見出し行推定
#
# ★ 正規化パス(normalize_book)の LO 往復に同乗させる（二役）。normalize_book が実行する
#   マクロを「何もしない空マクロ」から「何もしないが構造をテキストへ書き出すマクロ」に
#   差し替えるだけで、追加の LO 起動は発生しない。
# ★ LibreOffice の Cursor.gotoEndOfUsedArea・図形/グラフ/DataPilot 個数は LO でしか
#   正確に取れない（openpyxl の ws.max_row は書式だけ残った幽霊セルで過大評価しうる）。
#   一方、行ごとの書式的特徴(太字数等)・結合範囲は、LO が保存し終えた同じファイルを
#   openpyxl で読むだけで求まる（もう一度 LO を起動しない）。
# ---------------------------------------------------------------------------

STRUCT_HEADER_SCAN_ROWS = 20   # 見出し検出に使う先頭行数（多段見出し・タイトル行を含めても十分な余裕）
STRUCTDUMP_FILENAME = "structdump.txt"   # Basic → Python の受け渡し用（生のタブ区切りテキスト）


def _structdump_macro(out_path: Path) -> str:
    """StructDump 用の Basic マクロ。文書には一切手を触れない（正規化パスの no-op 性質を
       保つ）。シート毎に「実使用範囲(Cursor.gotoEndOfUsedArea)・図形/グラフ/DataPilot 個数」
       だけをタブ区切りテキストで書き出す（Basic に JSON エンコードは無いため、
       機械可読 dict への組み立ては Python 側 build_struct_dump() が担う）。"""
    out_str = str(out_path).replace('"', '""')
    return f'''Option VBASupport 1
Option Explicit

Sub Run(oDoc As Object)
    Dim iFile As Integer, i As Integer, n As Integer
    Dim oSheet As Object, oCursor As Object, oAddr As Object
    iFile = FreeFile
    Open "{out_str}" For Output As #iFile
    n = oDoc.Sheets.Count
    For i = 0 To n - 1
        oSheet = oDoc.Sheets.getByIndex(i)
        oCursor = oSheet.createCursor()
        oCursor.gotoEndOfUsedArea(True)
        oAddr = oCursor.RangeAddress
        Print #iFile, "SHEET" & Chr(9) & oSheet.Name & Chr(9) & oAddr.StartColumn & Chr(9) _
            & oAddr.StartRow & Chr(9) & oAddr.EndColumn & Chr(9) & oAddr.EndRow & Chr(9) _
            & oSheet.DrawPage.Count & Chr(9) & oSheet.Charts.Count & Chr(9) & oSheet.DataPilotTables.Count
    Next i
    Close #iFile
End Sub
'''


def parse_structdump_raw(text: str) -> dict:
    """StructDump の生テキスト（"SHEET"行群・タブ区切り）を
       {シート名: {"used_range": {...0起点...}, "shapes":n, "charts":n, "datapilots":n}} にパースする。
       壊れた/欠けた行は無視する（安全側・落とさない）。"""
    sheets: dict = {}
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) != 9 or parts[0] != "SHEET":
            continue
        try:
            name = parts[1]
            sc, sr, ec, er = int(parts[2]), int(parts[3]), int(parts[4]), int(parts[5])
            shapes, charts, pivots = int(parts[6]), int(parts[7]), int(parts[8])
        except ValueError:
            continue
        sheets[name] = {
            "used_range": {"start_col": sc, "start_row": sr, "end_col": ec, "end_row": er},
            "shapes": shapes, "charts": charts, "datapilots": pivots,
        }
    return sheets


def _row_char_stats(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> dict:
    """StructDump の行ごとの書式的特徴（見出し検出のヒューリスティクスに足る最小限）。
       {行(1起点): {"nonempty": 非空セル数, "str": うち文字列セル数, "bold": 太字セル数}}。"""
    stats = {}
    for r in range(start_row, end_row + 1):
        nonempty = 0
        strcnt = 0
        boldcnt = 0
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v in (None, ""):
                continue
            nonempty += 1
            if isinstance(v, str):
                strcnt += 1
            if cell.font and cell.font.bold:
                boldcnt += 1
        stats[r] = {"nonempty": nonempty, "str": strcnt, "bold": boldcnt}
    return stats


def build_struct_dump(normalized_book: Path, workdir: Path) -> dict:
    """StructDump 本体。normalize_book() が同じ LO 往復で書き出した生ダンプ
       (workdir/STRUCTDUMP_FILENAME) をパースし、正規化済みブックを openpyxl で読んで
       行の書式的特徴・結合範囲を足した機械可読 dict にする（もう一度 LO を起動しない）。
       生ダンプが無い/壊れている場合はそのシートだけ openpyxl の推定で代用する
       （テストでの normalize_book 差し替え等・呼び出し側は既定 header_row=1 に退避できる）。
       戻り値: {"sheets": {シート名: {"used_range":{1起点}, "merges":[...], "shapes":n,
                "charts":n, "datapilots":n, "rows": {行(1起点): {...}}}}}"""
    dump_path = workdir / STRUCTDUMP_FILENAME
    raw_sheets: dict = {}
    if dump_path.exists():
        try:
            raw_sheets = parse_structdump_raw(dump_path.read_text(encoding="utf-8"))
        except Exception:
            raw_sheets = {}
    wb = openpyxl.load_workbook(normalized_book)
    out: dict = {"sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        raw = raw_sheets.get(name)
        if raw is not None:
            ur = raw["used_range"]
            start_row, end_row = ur["start_row"] + 1, ur["end_row"] + 1     # 0起点→1起点
            start_col, end_col = ur["start_col"] + 1, ur["end_col"] + 1
        else:
            start_row, start_col = 1, 1
            end_row = min(ws.max_row or 1, MAX_ROWS)
            end_col = min(ws.max_column or 1, MAX_COLS)
        scan_end = min(end_row, start_row + STRUCT_HEADER_SCAN_ROWS - 1)
        rows = _row_char_stats(ws, start_row, scan_end, start_col, min(end_col, MAX_COLS))
        out["sheets"][name] = {
            "used_range": {"start_row": start_row, "end_row": end_row,
                            "start_col": start_col, "end_col": end_col},
            "merges": sorted(str(r) for r in ws.merged_cells.ranges),
            "shapes": raw["shapes"] if raw else 0,
            "charts": raw["charts"] if raw else 0,
            "datapilots": raw["datapilots"] if raw else 0,
            "rows": rows,
        }
    wb.close()
    return out


def detect_header_row(sheet_struct: dict) -> tuple:
    """(見出し行(1起点) or None, confident: bool)。
       ヒューリスティクス: 「上に結合セルやタイトル行があっても、列方向に複数(2以上)の
       非空文字列セルが並び、その行自体は非空セルが全部文字列（=見出しらしい）で、
       直下の行に文字列でない非空セル(数値等)が混ざる（=データが始まる）」行を強い候補とする。
       ★ D検体（2段見出し）対策: 親見出し行(結合)も str_count>=2 を満たしうるが、その直下も
       まだ見出し（型の混在なし）なので候補から外れる。子見出し行の直下でようやくデータの
       型混在が起き、そこで初めて候補になる（追加ルール不要でこの一般則から自然に解ける）。
       ★ フォールバック: 型混在チェックで候補0件（例: 全列が文字列のみの表）でも、
       「非空文字列セル2以上・行自体は純文字列」を満たす行が候補全体でちょうど1つなら
       それを確信とする（曖昧さが無いケース）。
       強い候補が0個か2個以上（=曖昧）なら (None, False) を返す＝呼び出し側は推測せず CLARIFY。"""
    rows = sheet_struct.get("rows", {})
    if not rows:
        return None, False
    sorted_rows = sorted(rows.keys())
    pure_str_rows = [r for r in sorted_rows
                      if rows[r]["str"] >= 2 and rows[r]["nonempty"] == rows[r]["str"]]

    def _mixture_below(r: int) -> bool:
        nxt = rows.get(r + 1)
        return nxt is not None and nxt["nonempty"] > nxt["str"]

    with_mixture = [r for r in pure_str_rows if _mixture_below(r)]
    if len(with_mixture) == 1:
        return with_mixture[0], True
    if not with_mixture and len(pure_str_rows) == 1:
        return pure_str_rows[0], True
    # ★ 致命3(W10e) 項目①: 表の全列が文字列型（数値/日付列が一つも無い）の場合、
    #   「型混在」の手がかりが構造的に一度も発生しない（with_mixture は必ず空）ため、
    #   pure_str_rows が『先頭行から続く全データ行』になり、見出し行が一意に決まらず
    #   毎回 CLARIFY に落ちていた（実測: 氏名/部署/備考のようなテキストのみの表）。
    #   スキャン範囲内の非空行が『例外なく全部』文字列のみ（数値/日付の手がかりが
    #   一つも無い）と確認できた場合に限り、最初の候補行（先頭・str>=2）を見出しと
    #   みなす（型混在の手がかりが少しでもある通常のケースはこの分岐に来ない＝
    #   test_detect_header_row_ambiguous_two_equally_valid_candidates_is_not_confident 等の
    #   既存挙動は変えない）。
    if not with_mixture and pure_str_rows:
        no_type_mixture_possible = all(
            info["nonempty"] == info["str"] for info in rows.values() if info["nonempty"] > 0)
        if no_type_mixture_possible:
            return pure_str_rows[0], True
    return None, False


# ★ W8a 項目3: 旧文言は「答えて」と聞くだけで答え方(CLI で何を打てばいいか)が無い行き止まり
#   だった（architect 発見）。--header-row フラグの使い方まで添えて、次のコマンドが打てる形にする。
CLARIFY_HEADER_ROW_QUESTION = ("見出しが何行目か分かりません。"
                                "`--header-row 3` のように指定して再実行してください")


def resolve_header_rows(struct_dump: dict, sheets: list, target_sheet: str | None = None) -> tuple:
    """全シートの見出し行(1起点)を決める。(header_rows: {シート名: 行}, clarify_question|None)。
       ★ 挙動変更#2: 対象シート（target_sheet・省略時は1枚目＝旧挙動と同一）だけ StructDump の
       ヒューリスティクスで推定する（DSL 操作の書き込み対象はこのシート1枚だけのため）。
       他シート（LOOKUP_FILL の参照表等）は物理1行目を既定にする。
       ★ build_struct_dump は元々全シート分の rows(書式的特徴) を計算済み（_structdump_macro が
       `For i = 0 To n - 1` で全シートを走査する）。対象シートが1枚目でなくても StructDump 自体の
       やり直しは不要 — ここで見る辞書のキーを sheets[0] から target_sheet に差し替えるだけでよい。
       StructDump が無い（テストでの normalize_book 差し替え等）場合は全シート1行目のまま
       （旧挙動と同一・CLARIFY は出さない）。自信が持てない場合だけ対象シートについて
       CLARIFY 質問を返す（推測で進まない）。"""
    header_rows = {s: 1 for s in sheets}
    if not sheets:
        return header_rows, None
    sd_sheets = (struct_dump or {}).get("sheets", {})
    target = target_sheet if target_sheet in sheets else sheets[0]
    info = sd_sheets.get(target)
    if info is None:
        return header_rows, None
    row, confident = detect_header_row(info)
    if confident:
        header_rows[target] = row
        return header_rows, None
    return header_rows, CLARIFY_HEADER_ROW_QUESTION


def _charts_count(path: Path) -> int:
    try:
        with zipfile.ZipFile(path) as z:
            return sum(1 for n in z.namelist()
                       if "chart" in n.lower() and n.lower().endswith(".xml")
                       and "/charts/chart" in n.lower())
    except Exception:
        return 0


def snapshot(path: Path) -> dict:
    """文書の状態を取る。★ 値/数値書式/背景色/太字 に加え、罫線・結合・列幅・行高・
       水平配置も捉える。これで『書式のみ・罫線のみ・列幅のみ・結合のみ・中央揃えのみ』
       の変更も『変化した』と検出でき、no-op 誤検出（＝効いているのに失敗扱い）を防ぐ。"""
    wb = openpyxl.load_workbook(path)
    # ★ 止血3: truncated は「この snapshot が MAX_ROWS/MAX_COLS で切り詰められたか」を
    #   保持する。diff_snapshots 後の表示で「先頭1000行しか見せていない」ことを正直に
    #   注記するために使う（bench/realworld/BASELINE.md の B 検体所見の根治）。
    # ★ operator 指摘②(2026-08-19): true_rows はシート毎の実際の行数（切り詰め前）。
    #   count_reconciliation が「データ 999 行のうち…」と 嘘の分母 を出していた根治に使う。
    #   値は既にこのループで計算していた（捨てていただけ）。
    snap = {"sheets": list(wb.sheetnames), "charts": _charts_count(path),
            "cells": {}, "merges": {}, "colw": {}, "rowh": {}, "truncated": False,
            "true_rows": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        true_nrow = ws.max_row or 0
        snap["true_rows"][name] = true_nrow
        true_ncol = ws.max_column or 0
        if true_nrow > MAX_ROWS or true_ncol > MAX_COLS:
            snap["truncated"] = True
        nrow = min(true_nrow, MAX_ROWS)
        ncol = min(true_ncol, MAX_COLS)
        for r in range(1, nrow + 1):
            for c in range(1, ncol + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                fill = None
                if cell.fill is not None and cell.fill.patternType:
                    fill = str(cell.fill.start_color.rgb)
                bold = bool(cell.font.bold) if cell.font else False
                numfmt = cell.number_format
                bd = cell.border
                bsig = (bd.left.style, bd.right.style, bd.top.style, bd.bottom.style) if bd else None
                if bsig == (None, None, None, None):
                    bsig = None
                # ★ 水平配置（中央揃え等）。既定は None/'general' 扱い。
                align = cell.alignment.horizontal if cell.alignment else None
                if align == "general":
                    align = None
                if (val in (None, "") and fill is None and not bold
                        and numfmt == "General" and bsig is None and align is None):
                    continue
                snap["cells"][f"{name}!{r},{c}"] = (val, numfmt, fill, bold, bsig, align)
        snap["merges"][name] = sorted(str(rng) for rng in ws.merged_cells.ranges)
        snap["colw"][name] = {k: round(d.width, 2) for k, d in ws.column_dimensions.items() if d.width}
        snap["rowh"][name] = {k: round(d.height, 2) for k, d in ws.row_dimensions.items() if d.height}
    wb.close()
    return snap


# ★ snapshot() の cell tuple の並びと対応する日本語ラベル（M1: 生 tuple 漏れの解消）。
_CELL_FIELD_LABELS = ("値", "数値書式", "背景", "太字", "罫線", "配置")
# snapshot() は完全に既定状態のセルを cells 辞書に入れない（skip 条件参照）。
# 片側が辞書に無い(=完全既定)ときは、この既定 tuple で埋める（全 None で埋めると
# 数値書式が『None→General』という偽の差分を作ってしまう＝実測で見つけた罠）。
_CELL_DEFAULT = (None, "General", None, False, None, None)


def _cell_ref(row: int, col: int) -> str:
    """0起点の内部行/列表記(r,c)を、人が読める A1 形式（例: B2）にする。"""
    return f"{get_column_letter(col)}{row}"


def _fmt_cell_value(v) -> str:
    """人が読める値表示（None は '(空)'、文字列はクォート）。"""
    if v is None:
        return "(空)"
    if isinstance(v, str):
        return f"'{v}'"
    return str(v)


def describe_cell_change(before: tuple | None, after: tuple | None) -> str:
    """(値,数値書式,背景,太字,罫線,配置) の tuple 差分を、★ 変わったフィールドだけ
       日本語ラベルで列挙する人間可読の文字列にする（tuple の生表示は出さない）。
       例: 値 'りんご'→'リンゴ', 太字 (空)→True"""
    b = before if before is not None else _CELL_DEFAULT
    a = after if after is not None else _CELL_DEFAULT
    parts = []
    for label, bv, av in zip(_CELL_FIELD_LABELS, b, a):
        if bv != av:
            parts.append(f"{label} {_fmt_cell_value(bv)}→{_fmt_cell_value(av)}")
    return ", ".join(parts) if parts else "(差分なし)"


def diff_snapshots(before: dict, after: dict) -> tuple:
    """(changed: bool, lines: [str])。人が読める変更点も返す。
       セル値/書式/色/太字/罫線・結合・列幅・行高・シート・グラフの変化を見る。"""
    lines = []
    added = [s for s in after["sheets"] if s not in before["sheets"]]
    removed = [s for s in before["sheets"] if s not in after["sheets"]]
    if added:
        lines.append(f"＋シート追加: {added}")
    if removed:
        lines.append(f"－シート削除: {removed}")
    if after["charts"] != before["charts"]:
        lines.append(f"＊グラフ数: {before['charts']} → {after['charts']}")

    # 結合セル
    merge_changes = 0
    for name in after.get("merges", {}):
        b = set(before.get("merges", {}).get(name, []))
        a = set(after["merges"][name])
        for m in sorted(a - b):
            merge_changes += 1; lines.append(f"＋結合 {name}!{m}")
        for m in sorted(b - a):
            merge_changes += 1; lines.append(f"－結合解除 {name}!{m}")

    # 列幅・行高（変わったシート数で示す）
    dim_changes = 0
    for key, label in (("colw", "列幅"), ("rowh", "行高")):
        for name in after.get(key, {}):
            if before.get(key, {}).get(name) != after[key].get(name):
                dim_changes += 1; lines.append(f"＊{label}変更: {name}")

    # セル（値/書式/色/太字/罫線）－ シートごとに自前の見出し→明細（他の種別と揃える）
    keys = set(before["cells"]) | set(after["cells"])
    changed_by_sheet: dict = {}
    cell_changes = 0
    for k in sorted(keys):
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        if b == a:
            continue
        cell_changes += 1
        sheet, rc = k.split("!", 1)
        r_str, c_str = rc.split(",")
        ref = _cell_ref(int(r_str), int(c_str))
        changed_by_sheet.setdefault(sheet, []).append(f"  {ref}: {describe_cell_change(b, a)}")

    shown = 0
    for sheet, clines in changed_by_sheet.items():
        if shown >= 12:  # 全部は出さない。多いときは件数で示す
            break
        lines.append(f"＊セル値変更: {sheet}")
        for cl in clines:
            if shown >= 12:
                break
            lines.append(cl)
            shown += 1
    if cell_changes > shown:
        lines.append(f"  …ほか {cell_changes - shown} セル")

    changed = bool(added or removed or (after["charts"] != before["charts"])
                   or merge_changes or dim_changes or cell_changes)
    return changed, lines


def _truncation_notice(before: dict, after: dict, exhaustive_postcondition: bool) -> str | None:
    """★ 止血3: before/after どちらかの snapshot が MAX_ROWS/MAX_COLS で切り詰められて
       いたら、無言で切らず正直な1行を返す（bench/realworld/BASELINE.md の B 検体所見）。
       ★ 経路で文言を出し分ける:
       - exhaustive_postcondition=True（DSL経路・cmd_run_dsl / cmd_run_plan の DSL 段）:
         事後条件チェッカーは openpyxl で out ファイルを直接開き _scan_last_row で全行を
         走査する（snapshot() の MAX_ROWS とは無関係）。表示だけが切り詰められている。
       - exhaustive_postcondition=False（FREEFORM経路）: no-op ガード・advisories も
         snapshot() 頼みなので、検証自体も先頭1000行までしか見ていない。
       切り詰めが無ければ None。"""
    if not (before.get("truncated") or after.get("truncated")):
        return None
    if exhaustive_postcondition:
        return f"（表示は先頭 {MAX_ROWS} 行の変化のみ。検証・適用は全行に対して実施）"
    return f"（表示は先頭 {MAX_ROWS} 行の変化のみ。検証も先頭 {MAX_ROWS} 行のみ）"


# ---------------------------------------------------------------------------
# 疑わしい変化の機械検出（M2a: 冷間再監査が実測した「✓ の下の失敗」への対抗）
#
# ★ 全部「差分の後の助言」であってブロックしない。誤検知を恐れて条件は保守的に
#   （両条件とも『変更セルの全部がそれに該当』した時だけ発火）。
# ---------------------------------------------------------------------------

def _changed_cells(before: dict, after: dict) -> list:
    """(sheet, row, col) のリスト。値/書式/色/太字/罫線/配置のいずれかが変わったセル。"""
    keys = set(before["cells"]) | set(after["cells"])
    out = []
    for k in keys:
        if before["cells"].get(k) != after["cells"].get(k):
            sheet, rc = k.split("!", 1)
            r_str, c_str = rc.split(",")
            out.append((sheet, int(r_str), int(c_str)))
    return out


def _value_changed_cells(before: dict, after: dict) -> list:
    """(sheet, row, col) のリスト。★ M2c: 値(idx0)が実際に変わったセルだけ
       （書式(罫線・中央揃え等)だけが変わったセルは含めない）。
       幽霊データ/一様埋め検出を『値変更』の部分集合だけで評価するための土台。
       ★ 冷間再監査3回目の実測: 罫線+中央揃え+0埋めが混在すると、書式だけ変わった
       セル（値は不変）が uniform 判定の対象に紛れ込み、一様埋めが見逃されていた。"""
    keys = set(before["cells"]) | set(after["cells"])
    out = []
    for k in keys:
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        b_val = b[0] if b is not None else None
        a_val = a[0] if a is not None else None
        if b_val != a_val:
            sheet, rc = k.split("!", 1)
            r_str, c_str = rc.split(",")
            out.append((sheet, int(r_str), int(c_str)))
    return out


def _used_range(before: dict, sheet: str) -> tuple | None:
    """原本(before snapshot)においてそのシートで実際に値が入っていたセルの矩形
       (min_row, max_row, min_col, max_col)。★ 近似: 値(idx0)が非空のセルだけを対象に
       する（書式のみのセルは『データ』に数えない）。値が1つも無ければ None。"""
    rows, cols = [], []
    prefix = sheet + "!"
    for k, v in before["cells"].items():
        if not k.startswith(prefix):
            continue
        if v[0] in (None, ""):
            continue
        r_str, c_str = k[len(prefix):].split(",")
        rows.append(int(r_str)); cols.append(int(c_str))
    if not rows:
        return None
    return (min(rows), max(rows), min(cols), max(cols))


def detect_ghost_data(before: dict, after: dict, *, new_col_letter: str | None = None,
                       new_row_at_end: bool = False) -> str | None:
    """★ 幽霊データ検出: 変更セルが全部、原本の使用範囲（データが存在した矩形）の
       外に集中している場合だけ疑わしい旨を返す。1セルでも範囲内なら何も言わない
       （保守的。使用範囲が不明なシートが混ざる場合も判定を保留する）。
       ★ M2c: 判定対象は『値変更』の部分集合だけ（書式のみの変更は無視・保守性は部分集合内で維持）。
       ★ W6: 実行中に新規作成されたシート（before["sheets"] に無い）のセルはここでは無視する
       （new_sheet_advisories が別途担当）。以前は『使用範囲が不明なシートが1つでも混ざると
       関数全体が判定を保留する』実装だったため、AGGREGATE/CHART/PivotSum のように新規シート
       を作る操作が絡むたび、他シートの本当のゴーストデータ検出まで丸ごと素通りしていた
       （監査実測。旧シート範囲が不明＝原本に無い＝新規シート、の場合に限って除外することで
       既存シートに対する検出力は変えずに直す）。
       ★ C9: new_col_letter（呼び出し側が op の宣言（OP_WRITE_TARGET）から求めた、今回
       新規に作る列の文字）が与えられ、かつ検出範囲が丸ごとその1列に収まる場合は、
       警告でなく中立表示を返す（旧 _neutralize_new_column_ghost_warning が『出してから
       打ち消す』後処理でやっていたのと同じ判定を、発生源で先取りする）。
       ★ 単位C(D10): new_row_at_end（op の宣言 writes に WRITE_NEW_ROW_AT_END がある＝
       データ最終行の下に行を足す op）が真で、検出セルが全部『原本の最終行より下・かつ
       原本の列範囲の中』に収まる場合も同じく中立表示にする。合計行は定義上ずっと原本の
       使用範囲の外に出るので、この誤警報は毎回・確実に再現していた（APPEND_TOTAL）。
       列の外（右）へ出たセルが混ざる場合は中立化しない＝検出力は保守的に温存する。"""
    changed = _value_changed_cells(before, after)
    if not changed:
        return None
    outside = []
    below_only = True   # ★ 単位C(D10): 範囲外セルが全部「原本の最終行より下・列範囲の中」か
    for sheet, r, c in changed:
        if sheet not in before["sheets"]:
            continue   # ★ W6: 新規作成されたシート＝ここでの判定対象外
        rect = _used_range(before, sheet)
        if rect is None:
            return None  # このシートの原本データ範囲が不明 → 判定を保留
        min_r, max_r, min_c, max_c = rect
        if min_r <= r <= max_r and min_c <= c <= max_c:
            return None  # 1つでも範囲内 → 発火しない
        if not (r > max_r and min_c <= c <= max_c):
            below_only = False
        outside.append((r, c))
    if not outside:
        return None   # 変更が全部、新規シートのセルだけだった
    rows = [r for r, _ in outside]
    cols = [c for _, c in outside]
    min_c, max_c = min(cols), max(cols)
    if new_col_letter is not None:
        try:
            new_col_idx = column_index_from_string(new_col_letter)
        except ValueError:
            new_col_idx = None
        if new_col_idx is not None and min_c == max_c == new_col_idx:
            return "（新規列の追加は意図どおりです）"
    if new_row_at_end and below_only:
        return "（表の末尾への追記は意図どおりです）"
    top_left = _cell_ref(min(rows), min_c)
    bot_right = _cell_ref(max(rows), max_c)
    span = top_left if len(outside) == 1 else f"{top_left}:{bot_right}"
    return f"★ 疑わしい: 変更が元データの範囲外です（{span}）"


# ★ W10b 項目4a(摩擦): 「★疑わしい: 変更が元データの範囲外」が、新規列の作成という
#   *意図した*操作でも毎回出ていた実測所見への対応。DSL 経路(COMPUTE_COLUMN・target
#   無指定=新規列作成)に限り、detect_ghost_data が指す範囲が丸ごとその新規列に収まって
#   いれば中立表示に落とす（新規列は原本の使用範囲外に出るのが当然＝誤警報）。
# ★ W10d: COMPUTE_COLUMN 専用の if だったものを OP_WRITE_TARGET の宣言駆動へ一般化した
#   （査定で名指しされたオオカミ少年: LOOKUP_FILL が新規列を作っても同じ誤警報が出ていた）。
#   op ごとの if は増やさない — OP_WRITE_TARGET に既に登録済みの「書き込み先列」宣言を
#   そのまま読み、その列が対象シートの既存見出しに無ければ『新規列を作る効果』とみなす
#   （COMPUTE_COLUMN の target 無指定＝キーが空、LOOKUP_FILL の target_col 有指定だが
#   対象シートにまだ無い列名＝どちらも実行時にその名前で新しい列を作るという同じ意味）。
# ★ C9: 以前は detect_ghost_data がまず警告を出し、_neutralize_new_column_ghost_warning が
#   advisories の該当行を後から中立表示に置き換えていた（『出してから打ち消す』）。W10c/W10d で
#   宣言（OP_WRITE_TARGET）が入力として取れるようになったので、detect_ghost_data 自身が
#   new_col_letter を受け取り、発生源で中立表示を返す形にした（_neutralize_new_column_ghost_warning
#   は削除。判定条件・出力文言は一切変えていない＝ゴールデン差分ゼロで確認済み）。


def _declared_new_column_letter(op: str, resolved: dict, book_meta: dict) -> str | None:
    """op が今回、宣言済みの効果として新規列を作るなら、その列の文字（"C" 等）を返す。
       作らない/対象シートが分からない場合は None。
       ★ W10d 番人の土台: OP_WRITE_TARGET だけを見る。新しい op を足しても
       OP_WRITE_TARGET へ登録さえすれば、ここへの追記なしで正しく判定される
       （test_op_write_target_declares_all_ops が登録漏れ自体を防ぐ）。
       ★ 単位C: 宣言が領域を持つようになったので、まず writes に『新規列』があるかを見る
       （旧形は col_key の有無だけが手掛かりで、新規列を作らない op でも列名が resolved に
       無ければ列文字を返してしまう形だった）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or WRITE_NEW_COLUMN not in write_target.writes:
        return None
    col_key, sheet_key = write_target.col_key, write_target.sheet_key
    if not col_key:
        return None
    if sheet_key:
        sheet = resolved.get(sheet_key)
    else:
        # ★ 挙動変更#2: sheets[0] 決め打ちをやめ、verify_dsl_args が一箇所で決めた
        # resolved["_target_sheet"] を読む（旧値と後方互換のフォールバック付き）。
        sheets = book_meta.get("sheets") or []
        sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    if not sheet:
        return None
    headers = book_meta.get("headers", {}).get(sheet, [])
    col_name = resolved.get(col_key)
    if col_name and col_name in headers:
        return None   # 既存列への書き込み（上書き側の話・新規列ではない）
    new_col_idx = len(headers)   # 0起点・新規列は既存見出しの直後
    return get_column_letter(new_col_idx + 1)


def detect_uniform_fill(before: dict, after: dict) -> str | None:
    """★ 一様埋め検出: 変更セルの全部で『変化前が空欄』かつ『変化後が全部同一値』
       （特に 0/空文字）の場合だけ疑わしい旨を返す（保守的）。
       ★ M2c: 判定対象は『値変更』の部分集合だけ（罫線・中央揃えなど書式のみが変わった
       セルは対象外にする — 混ざっていると後方の値だけ均一でも見逃していた実測不具合の修正）。"""
    keys = set(before["cells"]) | set(after["cells"])
    after_vals = []
    for k in keys:
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        b_val = b[0] if b is not None else None
        a_val = a[0] if a is not None else None
        if b_val == a_val:
            continue  # 値は変わっていない（書式だけの変更は対象外）
        if b_val not in (None, ""):
            return None  # 変化前が空欄でないセルが1つでもある → 発火しない
        after_vals.append(a_val)
    if not after_vals:
        return None
    if len(set(after_vals)) != 1:
        return None
    val = after_vals[0]
    if val in (None, ""):
        return None  # 空欄→空欄は『埋めた』ことにならない
    return f"★ 疑わしい: 空欄への同一値の一括書き込みです（値 {_fmt_cell_value(val)} × {len(after_vals)} セル）"


def _data_row_count(before: dict, sheet: str, key_col: int) -> int:
    """★ 近似: そのシートの原本使用範囲のうち見出し行(1行目)を除いた行で、
       key_col（左隣接のキー列）が埋まっている行数をデータ行数の目安とする。
       結合セルや空行の扱いまでは厳密化しない、素直な近似。"""
    rect = _used_range(before, sheet)
    if rect is None:
        return 0
    min_r, max_r, _min_c, _max_c = rect
    start = max(min_r + 1, 2)
    prefix = sheet + "!"
    count = 0
    for r in range(start, max_r + 1):
        v = before["cells"].get(f"{prefix}{r},{key_col}")
        if v is not None and v[0] not in (None, ""):
            count += 1
    return count


def count_reconciliation(before: dict, after: dict) -> str | None:
    """変更が単一シート・単一列に集中している場合だけ「データ N 行のうち M 行を変更」
       を添える（りんご欠落型のような『1行だけ抜けている』変更を1秒で見えるように）。
       ★ W8a 項目2: 分子(M)は必ずデータ行のみに数える。_used_range/_data_row_count は
       先頭行(min_r)を見出し行とみなして分母を数えているのに、分子側(changed_rows)は
       全変更行をそのまま数えていたため、見出し行も対象列で変わった場合（例: 新規列を
       作って見出し+データ全行に書く COMPUTE_COLUMN）に『データ5行のうち6行を変更』の
       ような算数が壊れた表示になっていた（実測: e2e_work/w3_e2e3_log.txt）。
       見出し行(min_r)自体の変更・原本の使用範囲より下の新規行の変更は、分子(データ行数)
       には数えず、見出し行の変更だけ別語「＋見出し行」で添える。"""
    changed = _changed_cells(before, after)
    if not changed:
        return None
    sheets = {s for s, _, _ in changed}
    cols = {c for _, _, c in changed}
    if len(sheets) != 1 or len(cols) != 1:
        return None
    sheet = next(iter(sheets))
    col = next(iter(cols))
    rect = _used_range(before, sheet)
    if rect is None:
        return None
    min_r, max_r, min_c, _max_c = rect
    key_col = col - 1 if col > min_c else col
    data_rows = _data_row_count(before, sheet, key_col)
    all_changed_rows = {r for _, r, _ in changed}
    header_changed = min_r in all_changed_rows
    # ★ 分子=データ行のみ: 見出し行(min_r)と、原本の使用範囲より下の新規行(> max_r)は
    #   ここでは「データ行」として数えない（分母(data_rows)も同じ範囲の定義に揃える）。
    data_changed_rows = {r for r in all_changed_rows if min_r < r <= max_r}
    changed_rows = len(data_changed_rows)
    unchanged_rows = max(data_rows - changed_rows, 0)
    col_letter = get_column_letter(col)
    # ★ operator 指摘②(2026-08-19): 1500 行のブックで「データ 999 行のうち 999 行を変更」と
    #   出ていた ── snapshot が MAX_ROWS で切れているのに、分母を snapshot から数えていた。
    #   ★ 分母を実データ行数に置き換えることは しない: 1000 行目より下の空/非空は snapshot に
    #   無いので、その数字はでっち上げになる。代わりに 主張を狭める（単位I と同じ型）──
    #   確認した範囲を言い、確認していない残りを 物理行数（true_rows・これは実測）で正直に述べる。
    true_rows = (before.get("true_rows") or {}).get(sheet, 0)
    if before.get("truncated") and true_rows > MAX_ROWS:
        hidden = true_rows - MAX_ROWS
        msg = (f"列 {col_letter}: 確認した先頭 {data_rows} 行のうち {changed_rows} 行を変更"
               f"（{unchanged_rows} 行は未変更）★ {MAX_ROWS} 行目より下の {hidden} 行は確認していない")
    else:
        msg = (f"列 {col_letter}: データ {data_rows} 行のうち {changed_rows} 行を変更"
               f"（{unchanged_rows} 行は未変更）")
    if header_changed:
        msg += "＋見出し行"
    return msg


# --- 依頼文言と変更範囲の重なりチェック（保守的・明示言及がある時だけ）--------

_RE_COL_KANJI_NUM = re.compile(r"列\s*(\d+)")
_RE_COL_LETTER_PREFIX = re.compile(r"列\s*([A-Za-z]{1,3})(?![A-Za-z0-9])")
_RE_COL_LETTER_SUFFIX = re.compile(r"(?<![A-Za-z0-9])([A-Za-z]{1,3})\s*列")
_RE_ROW = re.compile(r"行\s*(\d+)")


def extract_task_mentions(task: str, sheet_names: list) -> dict:
    """タスク文言から明示的な言及だけを正規表現で抜き出す（保守的・誤検知回避優先）。
       戻り値: {"cols": {1起点の列番号(文字表記=曖昧さなし)}, "digit_cols": {数字表記の生の値},
               "rows": {1起点の行番号}, "sheets": {原本に実在するシート名}}
       ★ 列の数字表記(『列2』)は曖昧 — 本ツールの規約(README・ヘルパ)は 0 起点だが、
       人が普段言うのは 1 起点。どちらか一方に断定すると正しい結果へ誤警報を出す
       (2026-08-14 再監査 2 回目で実測:『在庫(列2)』= 0 起点 C 列への正しい変更に
       B 列不在の★が出た)。生の値のまま保持し、照合側で両解釈を許す。"""
    cols = set()
    digit_cols = {int(m.group(1)) for m in _RE_COL_KANJI_NUM.finditer(task)
                  if int(m.group(1)) >= 0}
    for pat in (_RE_COL_LETTER_PREFIX, _RE_COL_LETTER_SUFFIX):
        for m in pat.finditer(task):
            try:
                cols.add(column_index_from_string(m.group(1).upper()))
            except ValueError:
                pass
    rows = {int(m.group(1)) for m in _RE_ROW.finditer(task) if int(m.group(1)) >= 1}
    # ★ 単位E: 「このシート名は依頼文に含まれるか」の照合は、決定側(resolve_target_sheet)と
    #   ここ（助言側）が独立に同じ文字列照合を書いていた。素材を1つに寄せ、決定側は1つに
    #   絞る／助言側は全部使う、という役割の違いだけを残す（判定規則も戻り値も不変）。
    sheets = set(sheet_names_mentioned_in(task, list(sheet_names or ())))
    return {"cols": cols, "digit_cols": digit_cols, "rows": rows, "sheets": sheets}


def _new_sheets(before: dict, after: dict) -> list:
    """before に無く after に有るシート名（実行中に新規作成されたシート）。順序維持。"""
    return [s for s in after["sheets"] if s not in before["sheets"]]


def _sheet_only_snapshot(snap: dict, sheet: str) -> dict:
    """snap のうち sheet に属する情報だけを持つ縮小 snapshot。
       ★ detect_ghost_data/detect_uniform_fill/count_reconciliation は snapshot() が返す
       dict の形（cells/merges/colw/rowh/sheets）をそのまま前提にした既存ロジックなので、
       新規シート専用に書き直さず、この縮小 snapshot 越しに『再利用』する（new_sheet_advisories
       が使う）。"""
    prefix = sheet + "!"
    return {
        "sheets": [sheet],
        "charts": 0,
        "cells": {k: v for k, v in snap["cells"].items() if k.startswith(prefix)},
        "merges": {sheet: snap.get("merges", {}).get(sheet, [])},
        "colw": {sheet: snap.get("colw", {}).get(sheet, {})},
        "rowh": {sheet: snap.get("rowh", {}).get(sheet, {})},
        "truncated": snap.get("truncated", False),
    }


def new_sheet_advisories(before: dict, after: dict) -> list:
    """★ W6 項目2: 監査所見「新規『集計』シートの全 0 埋めが★素通り」の根治。
       detect_ghost_data / detect_uniform_fill / count_reconciliation は『変更セル全部が
       該当した時だけ発火する』設計のため、新規シートの異常が他シートの正常な変更と混ざると
       『全部該当』が崩れ、丸ごと素通りしていた（例: COMPUTE_COLUMN が本シートの小計に書いた
       正しい値と、AGGREGATE が作った『集計』シートの壊れた全0埋めが同じ diff に同居する）。
       新規シートごとに単独の縮小 before/after（空 before・そのシートだけの after）を作って
       同じ3関数へ通すことで、他シートの変更に影響されず判定する。
       ★ detect_ghost_data は『原本の使用範囲外か』を測る関数だが、新規シートに原本の
       使用範囲は存在しない（何もかもが『新規』であって『範囲外』ではない）。そのため
       常に None を返す（適用はするが構造的に無言＝その関数の既存の誠実さをそのまま踏襲）。"""
    lines = []
    for sheet in _new_sheets(before, after):
        empty_before = {"sheets": [sheet], "charts": 0, "cells": {}, "merges": {sheet: []},
                         "colw": {sheet: {}}, "rowh": {sheet: {}}, "truncated": False}
        sheet_after = _sheet_only_snapshot(after, sheet)
        for fn in (detect_ghost_data, detect_uniform_fill):
            msg = fn(empty_before, sheet_after)
            if msg:
                lines.append(msg.replace("★ 疑わしい: ", f"★ 疑わしい: 新規シート『{sheet}』の", 1))
        recon = count_reconciliation(empty_before, sheet_after)
        if recon:
            lines.append(f"新規シート『{sheet}』 {recon}")
    return lines


_NEW_SHEET_MENTION_RE = re.compile(r"シート|ピボット|別に")


def unrequested_new_sheet_advisory(task: str, before: dict, after: dict, *,
                                    op: str | None = None) -> list:
    """★ W6 項目3（機械側）: 依頼文にシート新設の明示的な言及（『シート』『ピボット』
       『別に』のいずれか）が無いのに新規シートが作られたら申告する。
       ★ 保守的: 言及があれば（AGGREGATE/CHART/PivotSum 等が意図どおり新設したと見なし）沈黙。
       プロンプト側の抑制（CONTRACT の追記）はあくまで誘導であって保証にならないため、
       この機械申告が最終防衛線（feedback_intent_vs_guarantee: 指示は意図、保証は機械）。
       ★ C9: op が『新規シートを作る』と宣言していて（OP_WRITE_TARGET の writes に
       WRITE_NEW_SHEET・AGGREGATE(SummaryTable)/PIVOT(DataPilot)）、かつ今回ちょうど1枚だけ
       新規シートができた場合は、その1枚については警告でなく中立表示を返す
       （旧 _neutralize_declared_new_sheet_warning の後処理を発生源へ先取り。
       2枚以上できた場合は宣言どおりと断定できないので従来どおり全部警告する＝保守的）。
       ★ 単位C: 以前はここ専用の op 名集合 OP_DECLARED_SHEET_EFFECT を別に持っていたが、
       「新規シートを作る」は OP_WRITE_TARGET の宣言そのものなので、宣言を1つに畳んだ
       （宣言が2箇所にあると片方だけ更新されて食い違う）。集合の中身は変わっていない。"""
    new_sheets = _new_sheets(before, after)
    if not new_sheets or _NEW_SHEET_MENTION_RE.search(task):
        return []
    if _op_writes(op, WRITE_NEW_SHEET) and len(new_sheets) == 1:
        return [f"（新規シート『{new_sheets[0]}』の作成は意図どおりです）"]
    return [f"★ 依頼にない新しいシートが作成されました（{s}）" for s in new_sheets]


# ★ 致命2(W10e): 「既存シートの中身が置き換わった」検出。自由生成が依頼と無関係な
#   内容で既存シートを丸ごと上書きした実測事故（「集計」シートが日付別の無関係な内容に
#   すり替わった）への対抗。detect_ghost_data/detect_uniform_fill と同じ保守的な方針
#   （両条件とも『原本にあった非空セルが全部』変わった時だけ発火＝一部だけの更新・
#   再計算は対象外＝誤検知回避優先）。
def existing_sheet_replaced_advisory(before: dict, after: dict, *, op: str | None = None,
                                      precondition_broken: str | None = None) -> list:
    """before・after の両方に実在するシート（新規作成ではない）のうち、原本の使用範囲に
       あった非空セルが【全部】別の値に変わっている場合だけ「中身が置き換わった」を返す。
       一部のセルだけが変わった（値の再計算・部分更新等）場合は対象外（保守的）。
       ★ 空欄への一様書き込み等は detect_uniform_fill が別途担当するので、ここでは
       『置き換え後も何かしら値が残っている』ケースだけを見る（全消去は別の懸念）。
       ★ C9: op が『新規シートを作る』と宣言していて（OP_WRITE_TARGET の writes に
       WRITE_NEW_SHEET）、かつそのシートが OP_DECLARED_SHEET_NAME の宣言どおりの出力先
       （例: AGGREGATE→『集計』）なら、警告でなく中立表示を返す
       （旧 _neutralize_declared_sheet_replace_warning の後処理を発生源へ先取り）。"""
    # ★★ 単位G: 中立化は「前提が成立していた時だけ」。宣言(writes=new_sheet)は警告を黙らせる
    #   権利を持つが、その前提（＝その名前のシートは before に存在しない）が単位F の検査で
    #   破れていたなら、権利を失う ── 破れた宣言に「意図どおりです」と言わせない。
    #   ★ 破れた種類を見る（「何か破れた」では、format_only や reorder の破れで無関係に黙らせる）。
    #   ★ ここで警告に戻すだけで、正常系（前に ailine 自身が作った『集計』の作り直し）を
    #   肯定文に戻すのは 単位H（出所判定）の仕事。G で H をやると、また恒真を踏む。
    declared_sheet = (OP_DECLARED_SHEET_NAME.get(op)
                      if _op_writes(op, WRITE_NEW_SHEET) and precondition_broken != WRITE_NEW_SHEET
                      else None)
    lines = []
    for sheet in before["sheets"]:
        if sheet not in after["sheets"]:
            continue   # シート削除は diff_snapshots が別途拾う
        rect = _used_range(before, sheet)
        if rect is None:
            continue   # 原本にデータが無かった（新規に埋まっただけ）→ここでは判定しない
        min_r, max_r, min_c, max_c = rect
        prefix = sheet + "!"
        total = 0
        changed = 0
        after_has_content = False
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                key = f"{prefix}{r},{c}"
                b = before["cells"].get(key)
                b_val = b[0] if b is not None else None
                if b_val in (None, ""):
                    continue
                total += 1
                a = after["cells"].get(key)
                a_val = a[0] if a is not None else None
                if a_val not in (None, ""):
                    after_has_content = True
                if a_val != b_val:
                    changed += 1
        if total == 0 or changed != total or not after_has_content:
            continue   # 一部だけの変更、または全消去（置き換えではない）→ 発火しない
        if sheet == declared_sheet:
            lines.append(f"（既存シート『{sheet}』の更新は意図どおりです）")
            continue
        lines.append(f"★ 疑わしい: 既存シート『{sheet}』の中身が置き換わりました"
                      f"（元データ {total} セル分が別の内容に変わっています）")
    return lines


# ★ 致命2(W10e): AGGREGATE(集計)/PIVOT(ピボット) が既存の同名シートを宣言どおり
#   再生成する場合（例: 既に「集計」シートがある状態で再度「部門ごとにまとめて」を頼む
#   正常系）は、上の検出と同じ理屈で『中身が置き換わった』が誤って出る。W10c/W10d で
#   確立した「op の宣言済み効果と一致する変化は中立」に乗せて中立化する
#   （helpers/*.bas は出力シート名を固定で決め打つため、ここも固定表で対応づける）。
OP_DECLARED_SHEET_NAME = {"AGGREGATE": "集計", "PIVOT": "ピボット"}


def _changed_sheets(before: dict, after: dict) -> set:
    """何かしら変わったシート名の集合（セル・結合・列幅・行高・追加/削除）。"""
    changed = set()
    for name in set(before["sheets"]) | set(after["sheets"]):
        if name not in before["sheets"] or name not in after["sheets"]:
            changed.add(name)
            continue
        if before.get("merges", {}).get(name) != after.get("merges", {}).get(name):
            changed.add(name)
        elif before.get("colw", {}).get(name) != after.get("colw", {}).get(name):
            changed.add(name)
        elif before.get("rowh", {}).get(name) != after.get("rowh", {}).get(name):
            changed.add(name)
    for sheet, _r, _c in _changed_cells(before, after):
        changed.add(sheet)
    return changed


def mention_overlap_advisory(mentions: dict, before: dict, after: dict,
                              exclude_sheets: set | None = None) -> list:
    """言及があるのに変更範囲と全く重ならない場合だけ警告する（保守的）。
       数字表記の列は 0 起点/1 起点の両解釈を許し、どちらかが触られていれば沈黙する。
       ★ W10b 項目4b(摩擦): exclude_sheets に載るシート（例: LOOKUP_FILL の参照専用
       source_sheet）は、依頼文に言及があっても『変更されていません』を出さない
       （読み取り専用が正しい操作で、変更が無いのが正常なため誤警報だった）。
       ★ 安全器官の減衰なので保守的に: 抑制は呼び出し側が op から明示的に渡す時だけ・
       既定(None)は従来どおり無抑制。"""
    exclude_sheets = exclude_sheets or set()
    if not (mentions["cols"] or mentions.get("digit_cols") or mentions["rows"]
            or mentions["sheets"]):
        return []
    changed = _changed_cells(before, after)
    changed_cols = {c for _, _, c in changed}
    changed_rows = {r for _, r, _ in changed}
    changed_sheets = _changed_sheets(before, after)

    lines = []
    for col in sorted(mentions["cols"]):
        if col not in changed_cols:
            letter = get_column_letter(col)
            lines.append(f"★ 依頼で言及された『列{letter}』は存在しません/変更されていません")
    for n in sorted(mentions.get("digit_cols", set())):
        # 1 起点読み = 列 n / 0 起点読み = 列 n+1 (1 起点換算)。両方外れた時だけ警告し、
        # 警告文は文字に変換せずユーザーの書いた数字のまま返す (推定で上書きしない)
        candidates = {c for c in (n, n + 1) if c >= 1}
        if candidates and not (candidates & changed_cols):
            lines.append(f"★ 依頼で言及された『列{n}』は存在しません/変更されていません")
    for row in sorted(mentions["rows"]):
        if row not in changed_rows:
            lines.append(f"★ 依頼で言及された『行{row}』は存在しません/変更されていません")
    for sheet in sorted(mentions["sheets"]):
        if sheet in exclude_sheets:
            continue
        if sheet not in changed_sheets:
            lines.append(f"★ 依頼で言及された『{sheet}』は存在しません/変更されていません")
    return lines


def _structural_advisories(before: dict, after: dict, *, op: str | None = None,
                            resolved: dict | None = None, meta: dict | None = None,
                            precondition_broken: str | None = None) -> list:
    """助言のうち『この差分そのものが疑わしいか』を判定する部分だけ
       （①幽霊データ ②一様埋め ③件数の突き合わせ ⑤新規シートの中身・★ W6）。
       依頼文言との重なり(④ mention_overlap_advisory)は含めない。
       ★ W10d: 複合計画(cmd_run_plan)が段ごとの before/after にこの部分だけを
       再利用するために build_advisories から切り出した。④は依頼文全体に対する
       充足を問う質問なので、段ごとの局所的な before/after では判定できない
       （他段が担当する言及まで『この段で変更されていない』と誤検知する）。
       単発 op(build_advisories 経由)ではこれまでどおり④も同じ before/after で
       評価する（そちらは1段しかないため局所=全体で一致し、挙動は不変）。
       ★ C9: op/resolved/meta（呼び出し側が今回の段の宣言済み効果を渡す・省略時は None）は
       detect_ghost_data の new_col_letter 判定と existing_sheet_replaced_advisory の宣言シート判定・
       detect_write_target_type_change（★宣言つき挙動変更#1(b)）にそのまま渡す（旧 _neutralize_* 三兄弟の後処理を発生源へ先取り）。"""
    lines = []
    new_col_letter = _declared_new_column_letter(op, resolved, meta) if (op and resolved is not None and meta is not None) else None
    new_row_at_end = _op_writes(op, WRITE_NEW_ROW_AT_END)   # ★ 単位C(D10): 合計行は宣言済みの効果
    for fn, kwargs in ((detect_ghost_data, {"new_col_letter": new_col_letter, "new_row_at_end": new_row_at_end}), (detect_uniform_fill, {})):
        msg = fn(before, after, **kwargs)
        if msg:
            lines.append(msg)
    recon = count_reconciliation(before, after)
    if recon:
        lines.append(recon)
    lines.extend(new_sheet_advisories(before, after))
    lines.extend(existing_sheet_replaced_advisory(before, after, op=op, precondition_broken=precondition_broken) + [m for m in [detect_write_target_type_change(before, after, op=op, resolved=resolved, meta=meta, op_write_target=OP_WRITE_TARGET, is_number=_is_number)] if m])   # ★ 致命2(W10e) + 挙動変更#1(b)
    return lines


def build_advisories(task: str, before: dict, after: dict, exclude_sheets: set | None = None, *,
                      op: str | None = None, resolved: dict | None = None,
                      meta: dict | None = None, sheet_conflict=None,
                      precondition_broken: str | None = None) -> list:
    """diff の後に表示する助言行を全部集める。
       ①幽霊データ ②一様埋め ③件数の突き合わせ ⑤新規シートの中身（★ W6・
       _structural_advisories が担当） ⑥依頼にないシート新設の申告（★ W6）
       ④依頼文言との重なり。
       ★ C9: op/resolved/meta は _structural_advisories/unrequested_new_sheet_advisory へ
       そのまま横流しする（宣言済み効果の中立化を発生源で先取りするための追加引数・
       省略時は従来どおり無条件で全部発火する）。
       ★ 単位C(D8): 参照専用シート（OP_WRITE_TARGET の reads_only 宣言）は、呼び出し側が
       op ごとの if で渡すのをやめ、ここが宣言から自分で求めて exclude_sheets に足す
       （明示の exclude_sheets は自由生成経路のために残す＝宣言と和を取る）。
       ★ 誤爆#3: sheet_conflict（resolve_target_sheet が「この語は列名とも一致したので
       曖昧＝既定へ後退した」と決めた記録）に載る同名シートも、同じ和に足す ── 助言側で
       「曖昧かどうか」を判定し直さず、決めた側の結果をそのまま運ぶ
       （ailine_core.target_sheet.conflict_excluded_sheets 参照）。"""
    lines = list(_structural_advisories(before, after, op=op, resolved=resolved, meta=meta,
                                        precondition_broken=precondition_broken))
    lines.extend(unrequested_new_sheet_advisory(task, before, after, op=op))
    mentions = extract_task_mentions(task, before["sheets"])
    excluded = (set(exclude_sheets or ()) | _declared_reads_only_sheets(op, resolved)
                | conflict_excluded_sheets(sheet_conflict))
    lines.extend(mention_overlap_advisory(mentions, before, after, excluded or None))
    return lines


# ---------------------------------------------------------------------------
# ★ A': 用語集（vocab）— 税率等の「現場の取り決め値」を LLM でなく辞書に持たせる。
#   ~/.ailine/vocab.json（グローバルのみ）に {"語": 値} の平坦な dict で持つ。
#   設定ファイル経由で唯一「自由入力」がコード生成に届く経路なので、他のどの入力より
#   厳しく検疫する（float() 関門・制御文字禁止・件数/長さ上限。壊れたファイルは
#   クラッシュせず空辞書として扱う）。
# ---------------------------------------------------------------------------

def _sanitize_vocab_term(term) -> str | None:
    """語（キー）が登録可能な形か。空・制御文字・長すぎるものは None（拒否）。"""
    s = str(term).strip()
    if not s or len(s) > DEFAULT_VOCAB_MAX_TERM_LEN:
        return None
    if re.search(r"[\x00-\x1f\x7f]", s):   # 改行等の制御文字禁止（codegen へ渡る経路の防御）
        return None
    return s


def load_vocab(path: Path | None = None) -> dict:
    """~/.ailine/vocab.json を読む。無い/壊れている/形が違う場合は空の辞書を返す
       （★ クラッシュしない・battery や pytest からは path 明示で差し替えて再現性を保つ）。
       エントリごとに語をサニタイズし、値は float() を通ったものだけを採用する。
       件数が上限を超えた分は読み捨てる（先着順・ファイルの並び順に依存）。"""
    p = path or VOCAB_FILE
    if not p.is_file():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(raw, dict):
        return {}
    vocab: dict = {}
    for term, value in raw.items():
        if len(vocab) >= DEFAULT_VOCAB_MAX_ENTRIES:
            break
        clean_term = _sanitize_vocab_term(term)
        if clean_term is None:
            continue
        try:
            vocab[clean_term] = float(value)
        except (TypeError, ValueError):
            continue
    return vocab


def save_vocab(vocab: dict, path: Path | None = None) -> None:
    """vocab を ~/.ailine/vocab.json に上書き保存する。"""
    p = path or VOCAB_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(vocab, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def vocab_add(term: str, value, path: Path | None = None) -> tuple:
    """(ok, message)。term/value を検証してから登録・保存する。
       未登録の新規語で件数が上限に達している場合は拒否する（既存語の値更新は上限に関係なく可）。"""
    clean_term = _sanitize_vocab_term(term)
    if clean_term is None:
        return False, f"語『{term}』は登録できません（空/制御文字/{DEFAULT_VOCAB_MAX_TERM_LEN}文字超）"
    try:
        fval = float(value)
    except (TypeError, ValueError):
        return False, f"値『{value}』が数値ではありません"
    vocab = load_vocab(path)
    if clean_term not in vocab and len(vocab) >= DEFAULT_VOCAB_MAX_ENTRIES:
        return False, f"用語集が上限（{DEFAULT_VOCAB_MAX_ENTRIES}件）に達しています"
    vocab[clean_term] = fval
    save_vocab(vocab, path)
    return True, f"登録: {clean_term} = {fval:g}"


# --- ★ A': APPEND_TOTAL の倍率(factor)を LLM から切り離し、機械が確定する ------------
#   ①依頼文の明示率を regex で抽出 ②無ければ用語集を引く ③どちらも無く label が税/込を
#   含むなら CLARIFY。LLM が factor を返しても、ここが常に勝つ（verify_dsl_args 側で
#   食い違いを WARN として記録する）。

_RATE_PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[%％]")
_RATE_BAI_RE = re.compile(r"(\d+(?:\.\d+)?)\s*倍")
# ★ W10b 項目3: 「1.1を掛けた/掛ける」「1.1で割った/割る」型（税込み/税抜きの言い換えで
#   「倍」を伴わない場合の実測ギャップ・battery v5 #503 で発覚）。掛けるは n そのまま、
#   割るは 1/n（税抜き＝税込み金額から逆算する倍率）。
_RATE_KAKE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:を|に)?\s*掛け")
_RATE_WARI_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:で|に)?\s*割っ")
_RATE_KEYWORD_RE = re.compile(r"税|倍率")
_RATE_BARE_NUM_RE = re.compile(r"(\d+(?:\.\d+)?)")
# ★ W10c 高: 依頼文に「率らしい語」が一切無いのに COMPUTE_COLUMN の「1列×率」パターン
#   （税込み/税抜き専用）へ誤分類された時に分類そのものを疑うための、より広い信号語。
#   上の各 _RATE_*_RE より緩い（数値を伴わなくても良い）— 「倍率を求める話かどうか」
#   だけを見る。
_RATE_SIGNAL_RE = re.compile(r"[%％]|倍|掛け|割っ|税")
# ★ W10c 中: 新規列の見出しを自然な日本語にするための語（A' 原則: LLM を使わず正規表現の
#   有無判定だけで決める。査定で名指しされた「金額*1.1」という数式風の見出しの対応）。
_TAX_INCLUSIVE_RE = re.compile(r"税込")
_TAX_EXCLUSIVE_RE = re.compile(r"税抜")


def extract_rate_factor(text: str) -> tuple:
    """依頼文から明示の倍率を抽出する。戻り値は (factor, 出典スニペット) か (None, None)。
       ①「10%」「8 ％」型 → 1+n/100 ②「1.1倍」型 → n そのまま ②'「1.1を掛けた」型 → n
       そのまま ②''「1.1で割った」型 → 1/n（税抜き等の逆算） ③「税」「倍率」という語の
       前後8文字だけにある裸の小数（例:「税率0.1」）→ 1未満なら 1+n・1以上ならそのまま
       （無関係な数値の誤爆を避けるため、③だけは税/倍率の語の近傍に絞る）。
       複数の異なる値が見つかった場合は断定しない（None, None・CLARIFY に委ねる）。"""
    if not text:
        return None, None
    candidates: dict = {}   # factor -> 出典スニペット（最初に見つかったもの）
    for m in _RATE_PCT_RE.finditer(text):
        f = round(1 + float(m.group(1)) / 100, 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_BAI_RE.finditer(text):
        f = round(float(m.group(1)), 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_KAKE_RE.finditer(text):
        f = round(float(m.group(1)), 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_WARI_RE.finditer(text):
        n = float(m.group(1))
        if n > 0:
            f = round(1 / n, 6)
            candidates.setdefault(f, m.group(0))
    if not candidates:
        for km in _RATE_KEYWORD_RE.finditer(text):
            window = text[max(0, km.start() - 8): km.end() + 8]
            nm = _RATE_BARE_NUM_RE.search(window)
            if nm:
                n = float(nm.group(1))
                f = round((1 + n) if n < 1 else n, 6)
                candidates.setdefault(f, nm.group(0))
    if len(candidates) == 1:
        f, snippet = next(iter(candidates.items()))
        return f, snippet
    return None, None


def lookup_vocab_factor(text: str, vocab: dict) -> tuple:
    """依頼文に用語集の語が部分一致で含まれるかを見る。戻り値は (factor, 用語) か
       (None, None)。複数の異なる語（異なる値）がヒットした場合は断定しない。"""
    if not text or not vocab:
        return None, None
    hits: dict = {}
    for term, value in vocab.items():
        if term and term in text:
            hits.setdefault(value, term)
    if len(hits) == 1:
        value, term = next(iter(hits.items()))
        return value, term
    return None, None


# --- ★ A' 原則(致命3・W10e): SET_COLUMN_VALUE が書き込む定数値を LLM から切り離す ------
#   依頼文の引用符（「」『』""''）で囲まれた文字列を機械抽出する。extract_rate_factor と
#   同じ考え方 — ちょうど1つに絞れる時だけ確定・0件/2件以上は CLARIFY に委ねる（None）。
_QUOTE_PATTERNS = (
    re.compile(r"「([^」]+)」"),
    re.compile(r"『([^』]+)』"),
    re.compile(r'"([^"]+)"'),
    re.compile(r"'([^']+)'"),
)


def extract_quoted_literal(text: str) -> str | None:
    """依頼文全体を通して、引用符で囲まれた文字列がちょうど1つだけ見つかった場合に
       その中身を返す。0個/2個以上は曖昧とみなし None（機械確定を諦める＝呼び出し側が
       CLARIFY にする）。"""
    if not text:
        return None
    found = []
    for pat in _QUOTE_PATTERNS:
        found.extend(m.group(1) for m in pat.finditer(text))
    if len(found) == 1:
        return found[0]
    return None


# ---------------------------------------------------------------------------
# M2b: 中間命令言語（DSL）パイプライン
#   ①翻訳(LLM) → ②検証(接地) → ③確認行 → ④決定論 codegen(LLM不使用) → ⑤適用(既存機構)
#   → ⑥op別事後条件(openpyxl で機械検証)
#   ★ 天井は bench/translation_battery.json で実測（凍結合格線: op90%/slot80%/誤断定20%）。
#   CLARIFY・語彙外(FREEFORM)・翻訳失敗は現行の自由生成経路（M2a 助言つき）へ retreat する。
# ---------------------------------------------------------------------------

# ★ W9: op メタデータ（接地フォーム設計の確定分・今回は宣言だけで UI は変えない）。
#   各 op に category(大分類・事務の言葉)・label(操作名・事務の言葉)・synonyms(依頼文で
#   よく使う言い方の例・宣言のみ＝翻訳プロンプトへ機械的に注入はしない) を一箇所で持つ。
#   OP_LABELS は後方互換のためこの dict から導出する（既存コードの OP_LABELS.get(op, op)
#   はそのまま・値は完全に同じ）。
OP_META = {
    "SORT": {"category": "並べ替える", "label": "並べ替え", "folder": False,
              "synonyms": ["並べ替え", "ソート", "順に並べる"]},
    "COMPUTE_COLUMN": {"category": "計算する", "label": "計算列", "folder": False,
                         "synonyms": ["計算", "掛け算・割り算", "列同士の演算"]},
    "LOOKUP_FILL": {"category": "表を編集する", "label": "転記", "folder": False,
                     "synonyms": ["引っ張ってくる", "転記", "VLOOKUP"]},
    "AGGREGATE": {"category": "計算する", "label": "集計", "folder": False,
                   "synonyms": ["集計", "まとめる", "グループごとに小計"]},
    "BOLD": {"category": "見た目を整える", "label": "太字", "folder": False,
              "synonyms": ["太字", "ボールド", "強調"]},
    "FILL_COLOR": {"category": "見た目を整える", "label": "背景色", "folder": False,
                    "synonyms": ["色を付ける", "塗りつぶす", "ハイライト"]},
    "NUMBER_FORMAT": {"category": "見た目を整える", "label": "数値書式", "folder": False,
                        "synonyms": ["桁区切り", "カンマ区切り", "3桁区切り"]},
    "MERGE": {"category": "表を編集する", "label": "セル結合", "folder": False,
               "synonyms": ["結合", "セルを繋げる", "セルをまとめる"]},
    "CHART": {"category": "グラフを作る", "label": "グラフ", "folder": False,
               "synonyms": ["グラフ", "棒グラフ", "チャート"]},
    "CENTER_ALIGN": {"category": "見た目を整える", "label": "中央揃え", "folder": False,
                       "synonyms": ["中央揃え", "センタリング", "真ん中に寄せる"]},
    "APPEND_TOTAL": {"category": "計算する", "label": "合計追加", "folder": False,
                       "synonyms": ["合計を出す", "税込み合計", "一番下に合計"]},
    # ★ W9: 検証済みヘルパ4種の DSL 語彙昇格。
    "INSERT_ROWS": {"category": "表を編集する", "label": "行挿入", "folder": False,
                      "synonyms": ["行を挿入", "行を追加", "行を足す"]},
    "DRAW_BORDERS": {"category": "見た目を整える", "label": "けい線", "folder": False,
                       "synonyms": ["けい線を引く", "罫線を引く", "枠線を付ける"]},
    "AUTOFIT": {"category": "見た目を整える", "label": "列幅自動調整", "folder": False,
                 "synonyms": ["幅を内容に合わせる", "列幅調整", "列を自動調整"]},
    "PIVOT": {"category": "計算する", "label": "ピボット", "folder": False,
               "synonyms": ["ピボットテーブル", "ピボットで集計", "クロス集計"]},
    # ★ 致命3(W10e): 「列を一括で定数に書き換える」の DSL 昇格（査定所見:総務事務が
    #   最も頻繁に行う操作に信頼できる経路が無かった）。
    "SET_COLUMN_VALUE": {"category": "表を編集する", "label": "一括書換", "folder": False,
                           "synonyms": ["全部同じ値にする", "一括で書き換える", "列を統一する"]},
    # ★ 生まれた時から検証つきの1例目（コミット 2edcb08「EXTRACT op」参照）: 単一条件
    #   （列×比較×値）に一致する行を新シートへ抜き出す。自由生成の実弾2件（全セル文字列化・
    #   空シートで exit 0）を事後条件(check_extract)が直接殺す形で op に昇格させる。
    "EXTRACT": {"category": "表を編集する", "label": "抽出", "folder": True,
                 "synonyms": ["抜き出す", "抽出", "絞り込んでコピー"]},
}

OP_LABELS = {op: meta["label"] for op, meta in OP_META.items()}

# op → 必須 slot 名のタプル。翻訳直後の slot 欠落チェックと確認行の項目順を兼ねる。
OP_SCHEMA = {
    "SORT": ("col", "order"),
    "COMPUTE_COLUMN": ("operands", "operator"),
    "LOOKUP_FILL": ("target_sheet", "target_col", "source_sheet", "key_col"),
    "AGGREGATE": ("group_col", "value_col"),
    "BOLD": ("target",),
    "FILL_COLOR": ("target", "color"),
    "NUMBER_FORMAT": ("col", "style"),
    "MERGE": ("range",),
    "CHART": ("value_col",),
    "CENTER_ALIGN": ("target",),
    # ★ W6: label/factor は既定値がある任意項目（無指定でも FREEFORM に退避させない）。
    "APPEND_TOTAL": ("col",),
    # ★ W9: INSERT_ROWS は count が既定値ありの任意項目（at だけ必須）。
    #   DRAW_BORDERS/AUTOFIT は引数無し（空タプル＝ any(...) が常に False で FREEFORM に落ちない）。
    "INSERT_ROWS": ("at",),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": ("group_col", "value_col"),
    # ★ 致命3(W10e): value は必須 slot に入れない（LLM に確定させない・A' 原則）。
    #   実際に書き込む値は verify_dsl_args が依頼文の引用符から機械抽出する
    #   （抽出できなければ CLARIFY）。翻訳直後の slot 欠落チェックは col だけを見る。
    "SET_COLUMN_VALUE": ("col",),
    # ★ EXTRACT: col(対象列)・cmp(比較。gte/lte/gt/lt/eq/contains)・value(比較する値)。
    #   出力シート名は verify_dsl_args が機械で決め打ちする（LLM に決めさせない・A' 原則）。
    "EXTRACT": ("col", "cmp", "value"),
}

# ★ W10c 致命1: 「破壊の関所」（既存列への上書き検知・下の _maybe_warn_target_overwrite）が
#   守る対象を op ごとの if 分岐で持たず宣言駆動にする。旧実装は
#   `if op != "COMPUTE_COLUMN": return None` の1行で、COMPUTE_COLUMN 以外（LOOKUP_FILL 等）
#   は関所が構造的に発火しなかった（監査実測: 存在しない転記先列が無関係な既存列へ
#   解決され、確認なしで上書きされた事故）。
#   ★ 単位C: 宣言の形を「列」から「領域」へ広げた。旧形は (col_key, sheet_key) か None の
#   2択で、「どこに書くか」を列でしか言えなかった。そのため助言の側は宣言を読めず、
#   依頼文の表層語や列の形で代用して毎回誤爆していた（実測2件: APPEND_TOTAL の合計行が
#   必ず「★ 疑わしい: 変更が元データの範囲外です」を出す／AGGREGATE・PIVOT が名指しの
#   入力シートを読むだけなのに「★ …は変更されていません」を出す）。誤爆は条件分岐でなく
#   宣言で消す ── writes（書く領域の種類）と reads_only（参照専用シートの slot 名）を足した。
#   None は廃止し、全 op が WriteTarget を持つ（「安全だから省略した」でなく「対象が無いと
#   確認した」の明示宣言、という旧 None の思想はそのまま writes/col_key=None が担う）。
#   sheet_key が None のときは verify_dsl_args が決めた resolved["_target_sheet"]
#   （後方互換で book_meta の先頭シート）を指す。LOOKUP_FILL だけ target_sheet で明示する。
#   番人: test_op_write_target_declares_all_ops（OP_SCHEMA の全 op に宣言があるか）と
#   test_op_write_target_declarations_are_well_formed（未知の種類・矛盾した組み合わせ）。

# 書く領域の種類の語彙（宣言に書けるのはここに載る種類だけ）。
WRITE_EXISTING_COLUMN = "existing_column"   # 既存列の値を書き換える（＝破壊の関所の対象）
WRITE_NEW_COLUMN = "new_column"             # データの右端に新しい列を作る
WRITE_NEW_ROW_AT_END = "new_row_at_end"     # データ最終行の下に新しい行を足す
WRITE_NEW_SHEET = "new_sheet"               # 新しいシートを作る
WRITE_FORMAT_ONLY = "format_only"           # セルの値は変えない（書式・罫線・列幅・埋め込みグラフ）
WRITE_ROW_SHIFT = "row_shift"               # 行を挿入して既存行を下へずらす（値そのものは残る）
WRITE_REORDER = "reorder"                   # 行を並べ替える（値の集合は保存される）
WRITE_KINDS = frozenset({
    WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN, WRITE_NEW_ROW_AT_END, WRITE_NEW_SHEET,
    WRITE_FORMAT_ONLY, WRITE_ROW_SHIFT, WRITE_REORDER})


@dataclass(frozen=True)
class WriteTarget:
    """op が「どこに書くか / どこを読むだけか」の宣言。
       writes: 書く領域の種類（WRITE_KINDS の部分集合・空は不可＝必ず何かを宣言する）。
       col_key: 書き込み先列を指す resolved args のキー（既存列を書く op だけが持つ）。
       sheet_key: 対象シート名を指す resolved args のキー（None = resolved["_target_sheet"]）。
       reads_only: 参照専用シートを指す resolved args のキー（そのシートが無変更なのは
                   正常なので、助言側は「変更されていません」を言ってはいけない）。"""
    writes: tuple = ()
    col_key: str | None = None
    sheet_key: str | None = None
    reads_only: tuple = ()


OP_WRITE_TARGET = {
    # 並べ替えのみ・値そのものは保存される（書き込み先列という対象は無いと確認した）
    "SORT": WriteTarget(writes=(WRITE_REORDER,)),
    # target 有指定なら既存列・無指定なら新規列（resolved に無い＝関所は素通り）
    "COMPUTE_COLUMN": WriteTarget(writes=(WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN), col_key="target"),
    # target_col が対象シートに実在すれば既存列・無ければその名前で新規列を作る（codegen 参照）
    "LOOKUP_FILL": WriteTarget(writes=(WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN),
                                col_key="target_col", sheet_key="target_sheet",
                                reads_only=("source_sheet",)),
    # 新規シート（SummaryTable）を作るだけ。入力シートは読むだけ＝無変更が正常
    "AGGREGATE": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    "BOLD": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "FILL_COLOR": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "NUMBER_FORMAT": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "MERGE": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # 埋め込みグラフを足すだけ（セルの値も見出しも書かない＝値の書き込み先は無いと確認した）
    "CHART": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "CENTER_ALIGN": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # ★ 単位C(D10): データ末尾の新規行に [ラベル|合計式] を書く（W6・既存列は不可侵）。
    #   合計行は定義上ずっと元データの使用範囲の外に出る＝幽霊データの誤爆源だった。
    #   args の col は「合計する列」であって書き込み先列ではない（col_key に入れない）。
    "APPEND_TOTAL": WriteTarget(writes=(WRITE_NEW_ROW_AT_END,)),
    # 行を挿入するだけ・既存値は下にずれるだけで残る
    "INSERT_ROWS": WriteTarget(writes=(WRITE_ROW_SHIFT,)),
    "DRAW_BORDERS": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "AUTOFIT": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # 新規シート（DataPilot）を作るだけ。入力シートは読むだけ＝無変更が正常
    "PIVOT": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    # ★ 致命3(W10e): 既存列への一括書き込み＝破壊の関所の対象そのもの（宣言必須）。
    "SET_COLUMN_VALUE": WriteTarget(writes=(WRITE_EXISTING_COLUMN,), col_key="col"),
    # ★ EXTRACT: AGGREGATE/PIVOT と同じ形（新規シートを作るだけ・入力シートは読むだけ）。
    #   出力シート名は動的（col/cmp/value から機械決定）なので OP_DECLARED_SHEET_NAME の
    #   固定表には乗らない ── 単位H(_own_output_headers)側で動的な名前を扱う。
    "EXTRACT": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
}


def _op_writes(op: str | None, kind: str) -> bool:
    """op が kind の領域に書くと宣言しているか（op ごとの if を増やさないための問い合わせ口）。"""
    wt = OP_WRITE_TARGET.get(op)
    return bool(wt and kind in wt.writes)


def _declared_reads_only_sheets(op: str | None, resolved: dict | None) -> set:
    """op が「読むだけ」と宣言したシート名の集合（reads_only の slot を resolved で引く）。
       ★ 単位C(D8): 助言側はこれを exclude_sheets として受け取り、読むだけのシートに
       『変更されていません』を言わない。旧実装は `if op == "LOOKUP_FILL"` のハードコード
       2箇所で、AGGREGATE/PIVOT の入力シートは同じ理屈なのに毎回誤爆していた。"""
    wt = OP_WRITE_TARGET.get(op)
    if not wt or not wt.reads_only or resolved is None:
        return set()
    return {resolved[k] for k in wt.reads_only if resolved.get(k)}


# ★★ 単位E: 「利用者が依頼文で名指ししうる対象」の宣言。OP_WRITE_TARGET（どこに**書く**か）
#   とは別の事実なので別の表にする ―― 実証: APPEND_TOTAL は writes=new_row_at_end（書く先は
#   末尾の新規行）だが、利用者が名指しするのは合計する列(col)。同じ事実を2箇所に書くのが
#   禁じ手であって、違う事実を別々に宣言するのは正しい（両者が食い違うことはありえない）。
#   ★ 種別は「解決値の形」で決まる: column=列名 / region=col:列名|row:N|all /
#     row=行番号 / label=書き込むラベル。判定本体は ailine_core/subject.py（純ロジック）。
#   ★ 空タプルは「名指しできる対象が無いと確認した」宣言（DRAW_BORDERS/AUTOFIT は引数無しで
#     表全体・MERGE の range は A1:C1 という機械形式で実在名との照合対象が無い）。
#   ★ 対象シート(_target_sheet)はどの op も持つので、この表でなく共通処理として足す
#     （複数シートのブックだけ・1枚のブックには曖昧さが存在しない＝出力は従来どおり不変）。
#   番人: test_op_subject_slots_declares_all_ops / test_op_subject_slots_are_well_formed。
OP_SUBJECT_SLOTS = {
    "SORT": (("col", SUBJ_COLUMN),),
    # target 無指定（新規列作成）なら resolved に無い＝判定対象そのものが無い。
    # ★ operands は「対象」ではないが計画が実際に使った実在列なので、依頼文の語を消費する
    #   （input 種別・subject.py 参照）。これが無いと「売上から原価を引いた利益列を作って、
    #   利益で降順に」の 2 段目が、誰にも拾われない『売上』『原価』を反証と誤読する。
    "COMPUTE_COLUMN": (("operands", SUBJ_INPUT), ("target", SUBJ_COLUMN)),
    "LOOKUP_FILL": (("target_col", SUBJ_COLUMN), ("key_col", SUBJ_COLUMN)),
    "AGGREGATE": (("group_col", SUBJ_COLUMN), ("value_col", SUBJ_COLUMN)),
    "BOLD": (("target", SUBJ_REGION),),
    "FILL_COLOR": (("target", SUBJ_REGION),),
    "CENTER_ALIGN": (("target", SUBJ_REGION),),
    "NUMBER_FORMAT": (("col", SUBJ_COLUMN),),
    "MERGE": (),
    "CHART": (("value_col", SUBJ_COLUMN),),
    # ★ label は「金額の性質の限定（税込み/税抜き）」が依頼文にある時だけ問う（subject.py 参照）。
    "APPEND_TOTAL": (("col", SUBJ_COLUMN), ("label", SUBJ_LABEL)),
    "INSERT_ROWS": (("at", SUBJ_ROW),),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": (("group_col", SUBJ_COLUMN), ("value_col", SUBJ_COLUMN)),
    "SET_COLUMN_VALUE": (("col", SUBJ_COLUMN),),
    # ★ EXTRACT: cmp/value は SET_COLUMN_VALUE の value と同じ理由で対象に含めない
    #   （依頼文が名指しうる「対象」は列だけ・比較の種類や閾値は列名と同種の実在物ではない）。
    "EXTRACT": (("col", SUBJ_COLUMN),),
}


def _subject_slots(op: str, resolved: dict, sheets: list) -> list:
    """宣言(OP_SUBJECT_SLOTS)と resolved から判定対象のスロットを組む。
       ★ 対象シートは全 op 共通で足す ―― ただし**複数シートのブックだけ**
       （1枚しか無いブックに『どのシートか』の曖昧さは存在しない。format_sheet_field/
       describe_target_sheet が沈黙するのと同じ線引き＝単一シート帳票の出力は不変）。"""
    slots = []
    for key, kind in OP_SUBJECT_SLOTS.get(op, ()):
        value = resolved.get(key)
        if value in (None, ""):
            continue
        if isinstance(value, (list, tuple)):   # operands のような複数値の slot は1件ずつ
            slots.extend(Slot(key=key, value=str(v), kind=kind) for v in value if v not in (None, ""))
            continue
        context = ""
        if kind == SUBJ_LABEL:
            # ★ 率が機械確定していれば（＝1.0 でない）、限定は「率:1.1（依頼文: 消費税10%）」
            #   として既に解釈行に出ている ―― 二重に問わない（判定対象そのものにしない）。
            if resolved.get("factor") not in (None, 1.0):
                continue
            # 限定語が対象列の名前に現れていれば、解釈は限定を運んでいる
            # （例: 対象列『税込金額』の合計にラベル『合計』は正しい）。
            context = str(resolved.get("col") or "")
        slots.append(Slot(key=key, value=str(value), kind=kind, context=context))
    target_sheet = resolved.get("_target_sheet")
    if target_sheet and len(sheets or []) > 1:
        slots.append(Slot(key="_target_sheet", value=str(target_sheet), kind=SUBJ_SHEET))
    return slots


def classify_subject_provenance(op: str, resolved: dict, meta: dict, task: str, a=None) -> list:
    """★ 単位E の入口: この段の対象スロットを①②③に仕分ける（SubjectVerdict のリスト）。
       照合の材料は実在物だけ ―― 対象シートの実在列名・ブックの実在シート名・見出し行。
       ★ 複合計画では meta が『直前までの段を適用した後』の実体なので、前段が作った新規列も
       実在列として材料に入る（症状の検体そのもの: `col:数量*単価` は実在するが、依頼文の
       どの語とも照合できない ―― 断片『数量』『単価』は他の実在列も指しうるので証拠にならない）。
       ★ 「誰も拾わなかった語だけが反証」を段またぎで成立させるため、消費の台帳(Consumed)は
       a（run 全体を通して1つ）に持たせる ―― _sheet_conflict 等と同じ置き場所。a を渡さない
       直接呼び出し（単体テスト）では毎回まっさらな台帳で判定する。"""
    sheets = list((meta or {}).get("sheets") or [])
    target_sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    columns = list(((meta or {}).get("headers") or {}).get(target_sheet, []))
    header_row = ((meta or {}).get("header_rows") or {}).get(target_sheet, 1)
    qualifier = bool(_TAX_INCLUSIVE_RE.search(task or "") or _TAX_EXCLUSIVE_RE.search(task or ""))
    consumed = getattr(a, "_subject_consumed", None) if a is not None else None
    if a is not None and consumed is None:
        consumed = SubjectConsumed()
        a._subject_consumed = consumed
    return classify_slots(_subject_slots(op, resolved, sheets), task=task or "",
                           columns=columns, header_row=header_row, sheets=sheets,
                           qualifier_signal=qualifier, consumed=consumed)


# ★ bench/translation_spike.py（実測 v1）と同じ語彙定義（bench 側は比較用に据え置き、
#   本番プロンプトはここが唯一の元）。
OPS_DOC = """SORT: 並べ替え。args: col(列名), order(asc|desc)
COMPUTE_COLUMN: 既存列同士の計算。args: operands(列名2つ), operator(+,-,*,/), target(省略可・実在する列名。
  依頼が「〜に」のように既存列を名指ししたらその列名を入れる。無指定なら新しい列を作る)
  ★ 税込み/税抜き等「1列 × 率」の場合は operands を既存列名1つだけの配列にする。
  operator は * (税込み等、掛ける) か / (税抜き等、割る)。倍率(税率等)の数値はここに入れない
  （数値化はここでは行わない・機械が別途確定する。APPEND_TOTAL の倍率と同じ扱い）
LOOKUP_FILL: 別シートの対応表から値を転記。args: target_sheet, target_col, source_sheet, key_col
  ★ target_col は target_sheet に実在する列名が望ましいが、無ければ依頼文に書かれている
  そのままの列名を入れてよい（実行時にその名前で新しい列を作る）。実在しない列名を
  依頼文に無い別の実在列名に置き換えて誤魔化してはいけない（無関係な列を上書きする事故になる）
AGGREGATE: グループ別に集計表を作る。args: group_col, value_col
BOLD: 太字。args: target("row:行番号" か "col:列名")
FILL_COLOR: 背景色。args: target("row:N"か"col:列名"), color(英語色名)
NUMBER_FORMAT: 数値書式。args: col(列名), style("thousands")
MERGE: セル結合。args: range("A1:C1"形式)
CHART: 棒グラフ。args: value_col(列名)
CENTER_ALIGN: 中央揃え。args: target("all" か "col:列名")
APPEND_TOTAL: 列の合計(SUM)を表の最終行の下に追加する（税込み合計等）。args: col(合計する列名),
  label(省略可・既定"合計"。表示ラベル。「税込み合計」等、依頼の言い方をそのまま入れる)
  ★ 合計(SUM)専用。平均・最大・最小など他の統計量は語彙に無い（OUT_OF_VOCAB にする）。
  倍率(税率等)は入れない。数値化はここでは行わない（機械が別途確定する）
INSERT_ROWS: 行を挿入する。依頼文に具体的な行番号がある時だけ使う。
  args: at(1起点の行番号。この行の位置に挿入され、既存行は下にずれる), count(省略可・既定1・挿入する行数)
DRAW_BORDERS: 依頼文に「けい線/罫線/枠線」という言葉が明示された時だけ使う。表にけい線(格子線)を
  引く。args不要（表全体が対象）★「整えて」「いい感じに」のような具体性の無い依頼には
  絶対に使わない（曖昧なら CLARIFY で確認する）
AUTOFIT: 依頼文に「列幅/幅」という言葉が明示された時だけ使う。列幅を内容に合わせて自動調整する。
  args不要（全列が対象）★ 具体性の無い依頼には使わない（曖昧なら CLARIFY で確認する）
PIVOT: 依頼文に「ピボット」「ピボットテーブル」という言葉が明示された時だけ使う。
  本物のピボットテーブル(DataPilot)を作る。args: group_col(分類する列), value_col(合計する列)
  ★「ピボット」の語が無いグループ別集計（「集計」「まとめる」「小計」「合計がみたい」等、
  ピボットという語を伴わない言い方はすべて）は AGGREGATE を使う（PIVOT は LibreOffice で
  開き直すたび書式が消える癖があるため、書式つきの見栄えが要るなら AGGREGATE の方が適する）
SET_COLUMN_VALUE: 既存列の値を全部、同じ1つの値に書き換える。args: col(書き換える既存列名)
  ★ 実際に書き込む値(value)は依頼文の「」または『』で囲まれた引用を機械が抽出する
  （ここに書いてもよいが、依頼文の引用と食い違えば依頼文側が優先される）。税率等の
  倍率計算(COMPUTE_COLUMN)とは別物 — 依頼が「〜を掛けた」のような計算でなく、
  同じ文字列/値をそのまま代入するだけの依頼はこちらを使う
EXTRACT: 単一条件（列×比較×値）に一致する行だけを新しいシートへ抜き出す。
  args: col(列名), cmp(比較。gte=以上, lte=以下, gt=超, lt=未満, eq=等しい, contains=を含む),
  value(比較する値。数値または文字列。ここでは数値化しなくてよい・機械が確定する)
  ★ 出力シート名は機械が決める（LLM は考えなくてよい）。一部の列だけを残す絞り込みや
  複数条件(AND/OR)、グループごとに分けての抽出は語彙に無い（OUT_OF_VOCAB にする）"""

# ★ M2c: battery(v1) が実測で取り違えた基本パターンに加え、複合依頼(battery v2)を few-shot で
#   教える。同じ混同/構造を別の言い回しで示す（battery の項目文そのままは使わない＝暗記でなく
#   汎化を確かめる）。
#   ①「引いてくる/転記」は LOOKUP_FILL であって COMPUTE_COLUMN（四則演算）ではない。
#   ②③ 条件付き書式・集計行の追記は語彙に無い操作＝ OUT_OF_VOCAB（曖昧ではないので CLARIFY ではない）。
#   ④ 語彙内(target 付き COMPUTE_COLUMN)＋語彙外の混在。⑤ 語彙内どうしの2連(依存なし)。
TRANSLATION_FEWSHOT = [
    ('対象ブックの構成: {"Sheet": ["商品", "単価"], "商品マスタ": ["商品", "単価"]}\n'
     '依頼: 「商品マスタから商品名を引っ張ってきて明細に入れて」',
     '{"plan": [{"op": "LOOKUP_FILL", "args": {"target_sheet": "Sheet", "target_col": "商品", '
     '"source_sheet": "商品マスタ", "key_col": "商品"}}]}'),
    # ★ W10c 致命2: target_col が対象シートにまだ無いケース（監査実測: これを教えないと
    #   LLM が依頼に無い別の実在列名（この例なら「数量」）を勝手に代入することがあった）。
    ('対象ブックの構成: {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}\n'
     '依頼: 「単価表を見て単価を入れて」',
     '{"plan": [{"op": "LOOKUP_FILL", "args": {"target_sheet": "明細", "target_col": "単価", '
     '"source_sheet": "単価表", "key_col": "商品コード"}}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "金額", "在庫"]}\n'
     '依頼: 「金額が1000円未満の行を薄い黄色にして」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「合計を一番下の行に追加して」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "集計行の追記"}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "数量", "単価", "金額"]}\n'
     '依頼: 「金額に数量×単価を入れて、割引後の金額も出して」',
     '{"plan": ['
     '{"op": "COMPUTE_COLUMN", "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}}, '
     '{"op": "OUT_OF_VOCAB", "about": "割引後の金額"}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額で昇順に並べ替えて、見出し行を太字にして」',
     '{"plan": ['
     '{"op": "SORT", "args": {"col": "金額", "order": "asc"}}, '
     '{"op": "BOLD", "args": {"target": "row:1"}}]}'),
    # ★ W6: APPEND_TOTAL 語彙昇格（監査3回連続失敗の実測による）。
    #   ①明示的な倍率(消費税等)つき ②単純な合計、の両方を教える。
    #   ★ A': 倍率の数値化(1.1等)は LLM に求めない（factor は machine-determined。
    #   verify_dsl_args の extract_rate_factor/lookup_vocab_factor が依頼文から確定する）。
    ('対象ブックの構成: {"Sheet": ["品目", "数量", "単価", "小計"]}\n'
     '依頼: 「税込み合計を一番下に出して（消費税10%）」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": '
     '{"col": "小計", "label": "税込み合計"}}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額の合計を最後に」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額"}}]}'),
    # ★ C9: items_v7 の実測（2026-08-20・round0）で見つかった穴 ── 「税込み」と言っている
    #   のに依頼文に倍率の数字が一切無い依頼（例:「消費税込みでいくらになるか教えて」）を、
    #   モデルが APPEND_TOTAL ではなく OUT_OF_VOCAB に退避させることがあった（既存の
    #   #304/#704 で再現）。倍率が無いなら CLARIFY で聞き返すのが正しい挙動だが、それは
    #   APPEND_TOTAL に翻訳された後でしか起動しない（factor は machine-determined・
    #   verify_dsl_args 側の仕事）。ここでは「倍率の数字が無くても税込み/消費税の言及だけで
    #   op は APPEND_TOTAL」であることを教える（label に税/込を残せば、機械側が倍率不明を
    #   検出して CLARIFY に倒す）。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「消費税込みでいくらになるか教えて」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額", "label": "消費税込み合計"}}]}'),
    # ★ W9: 検証済みヘルパ4種の語彙昇格。PIVOT/AGGREGATE の分岐（「ピボット」と明示
    #   された時だけ PIVOT・それ以外の集計語は AGGREGATE）と INSERT_ROWS を1例ずつ教える。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門ごとにピボットテーブルで集計して」',
     '{"plan": [{"op": "PIVOT", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門ごとに金額をまとめて」',
     '{"plan": [{"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "金額"]}\n'
     '依頼: 「3行目の前に1行挿入して」',
     '{"plan": [{"op": "INSERT_ROWS", "args": {"at": 3, "count": 1}}]}'),
    # ★ W9 回帰対応（battery v1 再実測で実測）: DRAW_BORDERS/AUTOFIT が「引数不要」なぶん
    #   曖昧な依頼の当てはめ先として誤って選ばれやすかった（「整えて」「いい感じにして」→
    #   誤って DRAW_BORDERS を選ぶ誤断定が発生）。CLARIFY が正しい例を明示する。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「整えて」',
     '{"plan": [{"op": "CLARIFY", "question": "「整える」とは具体的に何をしますか"'
     '"（例: けい線を引く／列幅を合わせる／太字にする）"}]}'),
    # ★ W9 回帰対応: 「ピボット」の語が無いのに漠然とした集計依頼を PIVOT と誤断定した
    #   実測（#14「部署別の売上合計がみたい」→ AGGREGATE が正しい）。同型の言い回しを直接教える。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門別の金額合計がみたい」',
     '{"plan": [{"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    # ★ W9 回帰対応: APPEND_TOTAL は合計(SUM)専用。平均等の他統計量は語彙外
    #   （実測: 「平均値を追加」を誤って APPEND_TOTAL にした）。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額の平均値を一番下に追加して」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "平均値の追加"}]}'),
    # ★ W10b 項目3: 税込み/税抜き等「1列 × 率」パターンの語彙昇格（operator 第6回査定の
    #   実測: この言い回しが4種すべて DSL 分類に失敗し自由生成へ退避していた）。
    #   few-shot は最小限（W9 の実測: 足しすぎると別 op の誤断定回帰が出る）。
    #   ★ A': 倍率の数値化は LLM に求めない（factor は machine-determined。verify_dsl_args の
    #   extract_rate_factor/lookup_vocab_factor が依頼文/用語集から機械確定する）。
    #   ★ battery v3 再走で実測した回帰: 「税込み○○を出して」型の言い回しで例を作ると、
    #   APPEND_TOTAL の label 例（「税込み合計を一番下に出して」）と表現が似すぎて、
    #   モデルが APPEND_TOTAL の label フィールドまで省略するようになった(#305 CLARIFY
    #   すべきが確定に化けた)。「〜を追加して」型（新規列作成が明確な言い回し）の例だけ
    #   残し、APPEND_TOTAL の言い回しと重ならないようにする。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額に1.1を掛けた列を追加して」',
     '{"plan": [{"op": "COMPUTE_COLUMN", "args": {"operands": ["金額"], "operator": "*"}}]}'),
    # ★ W10c 高 → 致命3(W10e) で語彙昇格: 「列を全部Xに書き換える」（数値の倍率ではなく
    #   文字列を一律に代入する）は、税込み/税抜き（COMPUTE_COLUMN の1列×率パターン）と
    #   表現が似ているが別物（実測: 「氏名の列を全部『退職済み』に書き換えて」が税率の話と
    #   誤認された事故）。以前は語彙に無く OUT_OF_VOCAB へ退避させていたが、W10e で
    #   SET_COLUMN_VALUE を新設したため、この例も新しい正解へ差し替える（既存 fewshot の
    #   再利用＝件数を増やさない。W9 の教訓＝足しすぎは別 op の誤断定回帰を招く）。
    ('対象ブックの構成: {"Sheet": ["氏名", "部署", "金額"]}\n'
     '依頼: 「氏名の列を全部『退職済み』に書き換えて」',
     '{"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "氏名", "value": "退職済み"}}]}'),
]

TRANSLATION_SYSTEM = """あなたは表計算操作の翻訳係。日本語の依頼を、下の操作語彙を使った「計画」の JSON に翻訳する。
出力形式は必ず {{"plan": [ {{...}}, {{...}}, ... ]}}。それ以外は書かない。
依頼が複数の操作を含むなら、その全部を計画に順番どおり列挙すること。一部を省略してはいけない
（黙って落とすことを禁止・単一の依頼なら要素数1の計画にする）。

計画の各要素は次のどれか一つ:
- {{"op": "<語彙>", "args": {{...}}}}  操作語彙のどれかに当てはまる場合
- {{"op": "OUT_OF_VOCAB", "about": "<何についての依頼か、短く>"}}  操作語彙のどれにも当てはまらない部分
  （条件付き書式・行/シートの削除やコピー・集計行の追記などは語彙外。必ずこの形で計画に残す）
- {{"op": "CLARIFY", "question": "確認文"}}  依頼が曖昧で必須引数を確定できない場合。推測で断定しない

重要な規則:
- 列は必ず「対象ブックの構成」に実在する列名で指定する（番号ではなく）。ただし直前の段が新規作成する
  列を後続の段が参照する場合は、依頼文の言い方のままでよい（実行時に解決する）
- 依頼が既存の列を名指し（「小計に」等）して値を入れる/書き換える場合、COMPUTE_COLUMN の args に
  target(その実在列名) を入れる
- 「整えて」「いい感じにして」のような、何をすればよいか具体的に書いていない依頼は、
  引数が要らない操作(DRAW_BORDERS/AUTOFIT 等)であっても絶対に推測で当てはめない。
  必ず CLARIFY で何をしたいか確認する
- JSON のみ出力（説明・markdown 柵は禁止）

操作語彙:
{ops}"""

TRANSLATION_USER = "対象ブックの構成: {book}\n依頼: 「{text}」"


def build_translation_messages(task: str, book_meta: dict) -> list:
    """翻訳リクエストの messages（system + few-shot + 実クエリ）を組む。"""
    msgs = [{"role": "system", "content": TRANSLATION_SYSTEM.format(ops=OPS_DOC)}]
    for user_ex, assistant_ex in TRANSLATION_FEWSHOT:
        msgs.append({"role": "user", "content": user_ex})
        msgs.append({"role": "assistant", "content": assistant_ex})
    book_desc = json.dumps(book_meta.get("headers", {}), ensure_ascii=False)
    msgs.append({"role": "user", "content": TRANSLATION_USER.format(book=book_desc, text=task)})
    return msgs


def ollama_generate_json(model: str, messages: list, temperature: float = 0.1,
                          num_predict: int = 300) -> str:
    """M2b 翻訳層専用。/api/chat を format="json" 付きで叩く（bench/translation_spike.py
       で実測済みの型。qwen3 系は think:false を足す）。★ 通常生成の ollama_generate とは
       あえて別関数にする — format="json" を全体に効かせると Basic 生成側が壊れるため。
       戻り値は content 文字列（json.loads は呼び出し側 translate_task が行う）。"""
    body = {
        "model": model, "messages": messages, "stream": False, "format": "json",
        "options": {"temperature": temperature, "num_predict": num_predict, "num_ctx": 8192},
    }
    if "qwen3" in model:
        body["think"] = False
    req = urllib.request.Request(f"{OLLAMA}/api/chat", data=json.dumps(body).encode(),
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=300) as r:
        d = json.load(r)
    return d.get("message", {}).get("content", "")


def _normalize_plan_step(data) -> dict:
    """計画の1要素を正規化する（① 翻訳の per-step 版・translate_task の旧 per-op ロジックを継承）。
       ★ 黙って落とさない: 不明な形・語彙外・必須 slot 欠落は必ず何らかの op を持つ dict として
       残す（呼び出し側がその段を FREEFORM(自由生成)/OUT_OF_VOCAB として扱えるように）。
       - モデルが明示した op="OUT_OF_VOCAB" は about を保って素通し。
       - op が語彙外・必須 slot 欠落・非 dict 要素は op="FREEFORM" にする
        （旧 translate_task の『退避は FREEFORM』という分類をそのまま踏襲。単一依頼のときの
        後方互換＝cmd_run の全面自由生成 retreat が変わらないようにするため）。"""
    if not isinstance(data, dict):
        return {"op": "FREEFORM", "args": {}}
    op = str(data.get("op", "")).upper()
    if op == "CLARIFY":
        question = data.get("question")
        return {"op": "CLARIFY", "question": question or "確認が必要です", "args": {}}
    if op == "OUT_OF_VOCAB":
        about = data.get("about")
        return {"op": "OUT_OF_VOCAB", "about": str(about) if about else "内容不明の依頼", "args": {}}
    if op not in OP_SCHEMA:
        return {"op": "FREEFORM", "args": {}}
    args = data.get("args")
    if not isinstance(args, dict):
        # モデルが args で包まず op と slot をフラットに返した場合の救済（寛容に受ける）。
        args = {k: v for k, v in data.items() if k not in ("op", "about", "question")}
    required = OP_SCHEMA[op]
    if any(k not in args or args[k] in (None, "") for k in required):
        return {"op": "FREEFORM", "args": {}}
    return {"op": op, "args": args}


def translate_task(model: str, task: str, book_meta: dict, temperature: float = 0.1) -> dict:
    """① 翻訳。自然言語タスクを命令言語の「計画」 {"plan": [step, ...]} に翻訳する（M2c）。
       接地: book_meta の実在シート/列名だけを few-shot と一緒に渡す。
       ★ 複合依頼（複数の操作を含む依頼）は plan に全段を列挙する。★ 後方互換: 単一依頼は
       長さ1の計画になる（モデルが "plan" で包まず1操作だけ返した場合もここで長さ1に正規化する）。
       各 step は {"op": <語彙>, "args": {...}} / {"op": "OUT_OF_VOCAB", "about": ...} /
       {"op": "CLARIFY", "question": ...} / {"op": "FREEFORM"}（退避）のいずれか。
       ★ 失敗（API 不通・JSON 不正・空応答）は例外を投げずクラッシュさせない。
       op="FREEFORM" 一段の計画に退避し、呼び出し側が現行の自由生成経路（M2a 助言つき）
       へフォールバックできるようにする。"""
    try:
        messages = build_translation_messages(task, book_meta)
        raw = ollama_generate_json(model, messages, temperature=temperature, num_predict=700)
        data = json.loads(raw)
    except Exception:
        return {"plan": [{"op": "FREEFORM", "args": {}}]}
    steps_raw = None
    if isinstance(data, dict) and isinstance(data.get("plan"), list):
        steps_raw = data["plan"]
    elif isinstance(data, dict):
        steps_raw = [data]   # 後方互換: モデルが plan で包まず単一 op を直接返した
    elif isinstance(data, list):
        steps_raw = data
    if not steps_raw:
        return {"plan": [{"op": "FREEFORM", "args": {}}]}
    return {"plan": [_normalize_plan_step(s) for s in steps_raw]}


# --- ② 検証（接地：実在するシート/列名かを機械照合） -------------------------

COLOR_MAP = {
    "red": "FF0000", "green": "00B050", "blue": "0000FF", "yellow": "FFFF00",
    "orange": "FFA500", "purple": "800080", "pink": "FFC0CB", "black": "000000",
    "white": "FFFFFF", "gray": "808080", "grey": "808080",
    "lightblue": "ADD8E6", "lightgreen": "90EE90", "lightyellow": "FFFFE0",
    "lightred": "FFCCCC", "lightgray": "D3D3D3", "lightgrey": "D3D3D3",
}


def _digit_candidates(raw: str, headers: list) -> list:
    """数字表記の列参照を 0 起点/1 起点の両解釈で実在列名に変換した候補（重複除去・順序維持）。
       数字表記でなければ空リスト。"""
    s = str(raw).strip()
    if not re.fullmatch(r"\d+", s):
        return []
    n = int(s)
    cands = []
    for idx in (n, n - 1):        # 0起点読み(n) と 1起点読み(n-1) の両方を試す
        if 0 <= idx < len(headers) and headers[idx] not in cands:
            cands.append(headers[idx])
    return cands


def resolve_col_ref(raw, headers: list) -> tuple:
    """(実在列名 or None, 推定だったか, エラー文 or None)。
       ★ 列名を正とする。実在すればそのまま。数字表記なら 0/1 起点の両候補を試し、
       一意に決まればそれを『推定』として解決、決まらなければ CLARIFY 相当のエラーを返す。"""
    s = str(raw)
    if s in headers:
        return s, False, None
    cands = _digit_candidates(s, headers)
    if len(cands) == 1:
        return cands[0], True, None
    if len(cands) > 1:
        return None, False, f"列『{s}』は複数の解釈が可能で一意に決まりません: {cands}"
    known = ", ".join(headers) if headers else "(無し)"
    return None, False, f"列『{s}』がありません。ある列: {known}"


# ★ 単位B 照合の断片ガード（呼び出し側で・単位B 本体＝ailine_core/subject.py は変更しない）。
#   name_matches_task の呼び出し元は「実在するかどうか未確認の raw_target」を渡すことがあり
#   （下の COMPUTE_COLUMN target 経路）、_standalone_occurrence は「他の“実在名”の一部でしか
#   ない出現」しか除外しない ── raw_target が「実在しない・別の複合語」の断片（例:『小計』の
#   『計』）であっても素通りする。ailine_core/subject.py の _MIN_FRAGMENT=2（「1文字の漢字は
#   偶然一致しすぎる」）と同じ理由で、①長さ2未満は最初から証拠にしない、②2文字以上でも
#   依頼文中の全出現が「より長い連続した漢字の内部」（＝別の複合語の内側）でしかないなら
#   証拠にしない。
_CJK_KANJI_RE = re.compile(u"[㐀-䶿一-鿿豈-﫿]")


def _raw_target_not_embedded_in_task(raw_target: str, task: str) -> bool:
    """raw_target の依頼文中の出現のうち、少なくとも1つが「より長い連続した漢字の内部」
       ではない（＝独立した語としての出現がある）なら True。ひらがな/カタカナ/記号は
       日本語の語境界（助詞など）として扱う ―― 漢字が両隣にも続く場合だけ『内部』とみなす。
       出現が無ければ False（そもそも証拠が無い）。"""
    if not raw_target or not task:
        return False
    at = task.find(raw_target)
    if at < 0:
        return False
    n = len(raw_target)
    while at >= 0:
        before_ok = at == 0 or not _CJK_KANJI_RE.match(task[at - 1])
        after_ok = (at + n) >= len(task) or not _CJK_KANJI_RE.match(task[at + n])
        if before_ok and after_ok:
            return True
        at = task.find(raw_target, at + 1)
    return False


def verify_dsl_args(op: str, args: dict, book_meta: dict, task: str = "", vocab: dict | None = None,
                     target_sheet: str | None = None) -> tuple:
    """② 検証。(ok, resolved_args, inferred_keys, error_message)。
       args のシート/列名が実在するかを機械照合し、実在名に解決する。実在しなければ
       CLARIFY 相当のエラーメッセージを返す（呼び出し側が確認質問として表示する）。
       ★ A': task/vocab は APPEND_TOTAL の倍率(factor)確定専用（他の op は使わない・
       既定値のままで後方互換）。倍率の出典は resolved["_sources"]["factor"] に、
       LLM 由来の値との食い違いは resolved["_warnings"] に積む（戻り値のタプル形は
       変えない＝呼び出し側の unpack を壊さない）。
       ★ 挙動変更#2: target_sheet（省略時は1枚目＝旧挙動と同一）が「対象シート」の
       決定そのもの（resolve_target_sheet が一箇所で決めた値・呼び出し側が渡す）。
       resolved["_target_sheet"] に必ず積む — codegen_dsl・_maybe_warn_target_overwrite・
       _declared_new_column_letter・postcondition の各チェッカーは、みなここを読むだけで
       個別に sheets[0] を仮定しない（LOOKUP_FILL だけは自分自身の target_sheet slot が
       別途あるので、それを resolved["_target_sheet"] にも複製する）。"""
    sheets = book_meta["sheets"]
    headers = book_meta["headers"]
    if not sheets:
        return False, dict(args), set(), "ブックにシートが無い"
    first_sheet = target_sheet if target_sheet in sheets else sheets[0]
    resolved = dict(args)
    resolved["_target_sheet"] = first_sheet
    inferred: set = set()

    def resolve_in(key: str, sheet_name: str):
        val, was_inferred, err = resolve_col_ref(resolved.get(key), headers.get(sheet_name, []))
        if err:
            return err
        resolved[key] = val
        if was_inferred:
            inferred.add(key)
        return None

    def check_sheet(key: str):
        name = resolved.get(key)
        if name not in sheets:
            return f"シート『{name}』がありません。あるシート: {', '.join(sheets)}"
        return None

    if op == "SORT":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        if resolved.get("order") not in ("asc", "desc"):
            return False, resolved, inferred, f"順序『{resolved.get('order')}』は asc/desc のどちらでもありません"

    elif op == "COMPUTE_COLUMN":
        operands = resolved.get("operands")
        # ★ W10b 項目3: 税込み/税抜き等「1列 × 率」パターン。operands が列名1つだけの
        #   配列なら『既存列×倍率』とみなす（2列の四則演算とは別モード）。倍率(factor)は
        #   APPEND_TOTAL と同じ A' 原則で LLM から受け取らず機械確定する
        #   （extract_rate_factor/lookup_vocab_factor・regex のみ）。
        single_factor_mode = isinstance(operands, list) and len(operands) == 1
        if not single_factor_mode and not (isinstance(operands, list) and len(operands) == 2):
            return False, resolved, inferred, "演算対象が2つの列名になっていません"

        if single_factor_mode:
            v, was_inferred, err = resolve_col_ref(operands[0], headers.get(first_sheet, []))
            if err:
                return False, resolved, inferred, err
            resolved["operands"] = [v]
            if was_inferred:
                inferred.add("operands")
            if resolved.get("operator") not in ("*", "/"):
                return False, resolved, inferred, (
                    f"演算子『{resolved.get('operator')}』は列1つの計算（税込み/税抜き等）"
                    "では * か / のみ対応です")

            llm_factor_raw = resolved.pop("factor", None)
            text_factor, text_snippet = extract_rate_factor(task)
            vocab_factor, vocab_term = (None, None)
            if text_factor is None:
                vocab_factor, vocab_term = lookup_vocab_factor(task, vocab or {})

            sources: dict = {}
            if text_factor is not None:
                resolved["factor"] = text_factor
                sources["factor"] = f"依頼文: {text_snippet}"
            elif vocab_factor is not None:
                resolved["factor"] = vocab_factor
                sources["factor"] = f"用語集: {vocab_term}"
            else:
                # ★ W10c 高: 依頼文に率らしい語が一切無いのに「1列×率」（税込み/税抜き専用）へ
                #   分類されているのは、分類そのものが誤っている可能性が高い（実測: 「氏名の
                #   列を全部『退職済み』に書き換えて」のような値の一括書き換え依頼が、税率の
                #   話と誤認されて COMPUTE_COLUMN の単列モードに落ちることがあった）。
                #   その場合は「倍率が分からない」でなく、分類自体を疑う文言に変える
                #   （率を要求する op に分類されたのに率の手がかりが無い＝CLARIFY の理由を
                #   正直に言い換える。指示は意図・保証は機械＝プロンプト側だけに頼らない）。
                if not _RATE_SIGNAL_RE.search(task or ""):
                    return False, resolved, inferred, (
                        f"依頼「{task}」は『{v}』列に何らかの倍率（税率等）を掛ける操作として"
                        "解釈しましたが、依頼文に倍率らしき手がかりが見当たりません。"
                        "列の値をそのまま書き換える操作は今のところ対応していません。"
                        "倍率を掛ける処理であれば、依頼文に率を書く（例:「消費税10%」）か、"
                        "用語集に登録してください（例: ailine vocab add 消費税 1.1）"
                    )
                return False, resolved, inferred, (
                    "倍率（税率等）が分かりません。依頼文に率を書く（例:「消費税10%」）か、"
                    "用語集に登録してください（例: ailine vocab add 消費税 1.1）"
                )
            if resolved["factor"] <= 0:
                return False, resolved, inferred, f"倍率『{resolved['factor']}』は正の数でなければなりません"
            if sources:
                resolved["_sources"] = sources
            # ★ W10c 中: 新規列の見出しの自然化。旧実装は見出しを f"{op1}{operator}{factor:g}"
            #   （例:「金額*1.1」）という数式風の文字列にしていた（査定で名指し）。target
            #   無指定(新規列作成)かつ依頼文が税込み/税抜きと分かる言い方の場合だけ、
            #   その日本語ラベルを見出しに使う（A' 原則: LLM を使わず正規表現の有無のみで
            #   決める。手がかりが無ければ従来どおりの数式風見出しにフォールバック）。
            if not resolved.get("target"):
                if _TAX_INCLUSIVE_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税込{v}"
                elif _TAX_EXCLUSIVE_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税抜{v}"
                elif _RATE_KEYWORD_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税込{v}" if resolved["operator"] == "*" else f"税抜{v}"
            if llm_factor_raw not in (None, ""):
                try:
                    llm_factor = float(llm_factor_raw)
                except (TypeError, ValueError):
                    llm_factor = None
                if llm_factor is not None and abs(llm_factor - resolved["factor"]) > 1e-9:
                    mfactor = resolved["factor"]
                    resolved["_warnings"] = [
                        f"LLM が返した倍率({llm_factor:g})と機械抽出の倍率({mfactor:g})が"
                        f"食い違うため機械抽出({mfactor:g})を採用しました"
                    ]
        else:
            new_operands = []
            for o in operands:
                v, was_inferred, err = resolve_col_ref(o, headers.get(first_sheet, []))
                if err:
                    return False, resolved, inferred, err
                new_operands.append(v)
                if was_inferred:
                    inferred.add("operands")
            resolved["operands"] = new_operands
            if resolved.get("operator") not in ("+", "-", "*", "/"):
                return False, resolved, inferred, f"演算子『{resolved.get('operator')}』が不明です"
        # ★ M2c: target(任意) — 依頼が既存列を名指し（「小計に」等）した場合はその列に書く。
        #   無指定なら従来どおり新規列（codegen_dsl 側で分岐）。
        # ★ W3: target が実在しない場合、翻訳が「新しい列の名前」（例:「利益列を作って」の
        #   『利益』）を target と誤って埋めていることが多いと実測された（qwen2.5-coder:7b が
        #   『既存列に書く/新規に作る』の区別を安定して守らない）。実在しない＝一意に決まらない
        #   （複数解釈で曖昧）のとは別の理由なので、その場合だけ target を無指定として扱い
        #   新規列作成にフォールバックする（推測で断定しない CLARIFY の原則は、真に曖昧な
        #   ケース＝digit_candidates の複数一致にだけ残す）。
        # ★ W3 改定(2026-08-20): 上の前提（実在しない target＝ほぼ捏造）が古くなった実測が
        #   出た。「金額を数量×単価で埋めて」（金額列がまだ実在しない構成）で翻訳は正しく
        #   target:"金額" を返す。だが無条件に捨てる旧実装はそれも落とし、新規列が
        #   『数量*単価』に自動命名されていた（利用者が名前を言っているのに無視される・
        #   2026-08-19 のデモ制作で3回踏んだ実害）。★ 依頼文を判定者にする: raw_target が
        #   依頼文に実在する語として機械照合できるなら（単位B の name_matches_task を再利用
        #   ── 素朴な in 判定はしない。「税込金額を…」+target「金額」のような片方向の部分
        #   文字列の穴は単位B が塞いだ形そのものなので、同じ判定を2箇所に書かず呼ぶ）、
        #   捏造ではなく利用者の指名とみなして新規列の名前として使う。依頼文に語が無ければ
        #   従来どおり捏造とみなして捨てる（W3 本来の防御は生きている）。
        if resolved.get("target"):
            raw_target = resolved["target"]
            v, was_inferred, err = resolve_col_ref(raw_target, headers.get(first_sheet, []))
            if err:
                if "一意に決まりません" in err:
                    return False, resolved, inferred, err
                # ★ 単位B 照合の断片ガード（呼び出し側・上の _raw_target_not_embedded_in_task
                #   docstring 参照）: raw_target は「実在するか未確認」の生の文字列なので、
                #   name_matches_task を素通しに使う前に (1) 1文字を弾き (2) 依頼文中の全出現が
                #   「他の複合語（実在列とは限らない）の内部」でしかないなら弾く。
                if (len(raw_target) >= 2
                        and _raw_target_not_embedded_in_task(raw_target, task)
                        and name_matches_task(raw_target, task, others=headers.get(first_sheet, []))):
                    resolved["_new_col_label"] = raw_target
                del resolved["target"]
            else:
                resolved["target"] = v
                if was_inferred:
                    inferred.add("target")
                    # ★ W10a 項目3: 数字指定→列名解決の元の表記を残す（解釈要約の表示用・
                    #   例:「列5」→「在庫」列と解決した時、確認行の直後にその経緯を見せる）。
                    resolved["_target_raw"] = raw_target

    elif op == "LOOKUP_FILL":
        if (err := check_sheet("target_sheet")):
            return False, resolved, inferred, err
        if (err := check_sheet("source_sheet")):
            return False, resolved, inferred, err
        # ★ 挙動変更#2: 旧実装はここで「対象シートは1枚目のみ対応しています」と拒否していた
        #   （散在した『1枚目固定』の一つ・査定の致命そのもの）。LOOKUP_FILL は元々
        #   target_sheet を自分の必須 slot として名前で受け取り check_sheet で実在確認まで
        #   済ませているので、この制限を外すだけで対応できる。resolved["_target_sheet"] は
        #   LOOKUP_FILL 自身の target_sheet を正とする（他 op 用の一般解決 first_sheet より
        #   こちらを優先 — 依頼文に転記先/参照元の2シート名が両方出て一般解決が曖昧に
        #   フォールバックしていても、LOOKUP_FILL のここでの解決は影響を受けない）。
        resolved["_target_sheet"] = resolved["target_sheet"]
        # ★ W10c 致命2: target_col は COMPUTE_COLUMN の target と違い OP_SCHEMA 上は必須
        #   slot なので、LLM は「存在しないなら空にする」を選べない。実測（監査再現）:
        #   対象シートに『単価』列がまだ無いのに転記を頼むと、LLM がそれと無関係な
        #   *実在する*既存列（例:「数量」）の名前を代わりに返すことがある。resolve_col_ref
        #   は実在列名なら無条件で素通しするため、これだけでは見分けられない（そのまま
        #   進めると「数量」が確認なしで上書きされる事故になる）。
        #   ここでは「実在するから信用する」をやめ、根拠を要求する:
        #   ①依頼文にその列名が書かれている ②転記元（source_sheet）の値列
        #   （VLookupFromTable ヘルパの仕様どおり常に列1＝2番目の列）と同じ名前
        #   のどちらかが無いと、実在列であっても信用しない。
        target_headers = headers.get(resolved["target_sheet"], [])
        source_headers = headers.get(resolved["source_sheet"], [])
        value_col_hint = source_headers[1] if len(source_headers) > 1 else None
        raw_target_col = resolved.get("target_col")
        raw_str = str(raw_target_col) if raw_target_col not in (None, "") else ""
        exists = raw_str in target_headers
        mentioned = bool(raw_str) and raw_str in task
        matches_value_col = value_col_hint is not None and raw_str == value_col_hint

        if exists and (mentioned or matches_value_col):
            pass   # 根拠つきで実在列を指名＝そのまま使う（上書き注意は破壊の関所が別途担当）
        elif not exists:
            cands = _digit_candidates(raw_str, target_headers)
            if len(cands) == 1:
                resolved["target_col"] = cands[0]   # 数字表記の推定は従来どおり許容
                inferred.add("target_col")
            elif mentioned:
                # ★ 依頼文にも同じ列名が書かれている＝新規作成が正しい解釈（COMPUTE_COLUMN の
                #   target 無指定＝新規列と同じ考え方）。target_col はそのまま残し、
                #   codegen_dsl 側で新規列として作る。
                pass
            else:
                known = ", ".join(target_headers) if target_headers else "(無し)"
                return False, resolved, inferred, (
                    f"転記先の列『{raw_target_col}』が『{resolved['target_sheet']}』シートに"
                    f"見つかりません。ある列: {known}。新しい列として作る場合は、依頼文に"
                    f"その列名を書いてください（例:「{raw_target_col}という列を作って転記して」）"
                )
        else:
            # ★ 実測の事故そのもの: exists=True だが根拠が無い（依頼文にも書かれておらず、
            #   転記元の値列とも一致しない）＝上書き対象を取り違えている可能性が高い。
            hint = f"（参照表『{resolved['source_sheet']}』の値の列は『{value_col_hint}』です）" \
                if value_col_hint else ""
            return False, resolved, inferred, (
                f"転記先の列『{raw_target_col}』は実在しますが、依頼文にその列名が見当たらず、"
                f"転記元の値とも対応が確認できません{hint}。上書き対象を取り違えている"
                "可能性があるため、意図した列名を依頼文に明記してください"
            )
        if (err := resolve_in("key_col", resolved["target_sheet"])):
            return False, resolved, inferred, err

    elif op == "AGGREGATE":
        if (err := resolve_in("group_col", first_sheet)):
            return False, resolved, inferred, err
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err

    elif op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        target = str(resolved.get("target", ""))
        if target == "all":
            if op != "CENTER_ALIGN":
                return False, resolved, inferred, f"対象『all』は {OP_LABELS[op]} では未対応です"
        elif target.startswith("row:"):
            n = target[4:]
            if not (n.isdigit() and int(n) >= 1):
                return False, resolved, inferred, f"行番号『{n}』が不正です"
            # ★ 単位I: 契約文(1743)・codegen(2588)は CENTER_ALIGN に row: を認めておらず、
            #   ここだけが少数派だった（実測: row:N を通した先で codegen が素の traceback）。
            #   codegen に row: を実装するのは別作業 ―― ここは契約文に合わせて拒否するだけ。
            if op == "CENTER_ALIGN":
                return False, resolved, inferred, (
                    f"対象『{target}』は {OP_LABELS[op]} では未対応です"
                    "（col:列名 か all を使ってください）"
                )
        elif target.startswith("col:"):
            colname = target[4:]
            v, was_inferred, err = resolve_col_ref(colname, headers.get(first_sheet, []))
            if err:
                return False, resolved, inferred, err
            resolved["target"] = f"col:{v}"
            if was_inferred:
                inferred.add("target")
        else:
            return False, resolved, inferred, f"対象『{target}』の形式が不明です（row:N / col:列名 / all）"
        if op == "FILL_COLOR":
            color = str(resolved.get("color", "")).lower()
            if color not in COLOR_MAP:
                return False, resolved, inferred, f"色『{color}』は未対応です。使える色: {', '.join(sorted(COLOR_MAP))}"
            resolved["color"] = color

    elif op == "NUMBER_FORMAT":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        if resolved.get("style") != "thousands":
            return False, resolved, inferred, f"書式『{resolved.get('style')}』は未対応です（対応: thousands）"

    elif op == "MERGE":
        if not re.fullmatch(r"[A-Za-z]{1,3}\d+:[A-Za-z]{1,3}\d+", str(resolved.get("range", ""))):
            return False, resolved, inferred, f"範囲『{resolved.get('range')}』の形式が不正です（例: A1:C1）"

    elif op == "CHART":
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err

    elif op == "APPEND_TOTAL":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        # ★ W6: label は既定値を持つ任意項目。ここで確定させ、codegen/事後条件/
        #   確認行の全部に同じ既定解決を一貫して渡す。
        resolved["label"] = str(resolved.get("label") or "合計")
        label = resolved["label"]

        # ★ A': factor は LLM から受け取らない。LLM が返した値(あれば)はいったん取り出して
        #   おき、機械抽出/用語集の結果と食い違う場合だけ WARN として記録する（常に機械が勝つ）。
        llm_factor_raw = resolved.pop("factor", None)

        text_factor, text_snippet = extract_rate_factor(task)
        vocab_factor, vocab_term = (None, None)
        if text_factor is None:
            vocab_factor, vocab_term = lookup_vocab_factor(task, vocab or {})

        sources: dict = {}
        if text_factor is not None:
            resolved["factor"] = text_factor
            sources["factor"] = f"依頼文: {text_snippet}"
        elif vocab_factor is not None:
            resolved["factor"] = vocab_factor
            sources["factor"] = f"用語集: {vocab_term}"
        else:
            resolved["factor"] = 1.0

        if resolved["factor"] <= 0:
            return False, resolved, inferred, f"倍率『{resolved['factor']}』は正の数でなければなりません"

        # ★ 恒真式の番人（最優先）: label が「税」/「込」を含むのに倍率が確定できず既定
        #   1.0 のままだと、税抜き金額に「税込み」ラベルが付いた恒真の誤りを事後条件が
        #   pass にしてしまう（args 基準の検証だから）。ここで機械的に CLARIFY へ倒す
        #   （語リストは 税/込 の2語で凍結・むやみに増やさない）。
        if resolved["factor"] == 1.0 and any(k in label for k in ("税", "込")):
            return False, resolved, inferred, (
                f"ラベル『{label}』は税/込を含みますが倍率が分かりません。"
                "依頼文に税率を書く（例:「消費税10%」）か、用語集に登録してください"
                "（例: ailine vocab add 消費税 1.1）"
            )

        if sources:
            resolved["_sources"] = sources
        if llm_factor_raw not in (None, ""):
            try:
                llm_factor = float(llm_factor_raw)
            except (TypeError, ValueError):
                llm_factor = None
            if llm_factor is not None and abs(llm_factor - resolved["factor"]) > 1e-9:
                mfactor = resolved["factor"]
                resolved["_warnings"] = [
                    f"LLM が返した倍率({llm_factor:g})と機械抽出の倍率({mfactor:g})が"
                    f"食い違うため機械抽出({mfactor:g})を採用しました"
                ]

    # --- ★ W9: 検証済みヘルパ4種の語彙昇格 -----------------------------------
    elif op == "INSERT_ROWS":
        at_raw = str(resolved.get("at", "")).strip()
        if not (at_raw.isdigit() and int(at_raw) >= 1):
            return False, resolved, inferred, f"行番号『{resolved.get('at')}』が不正です（1以上の整数）"
        resolved["at"] = int(at_raw)
        count_raw = resolved.get("count")
        if count_raw in (None, ""):
            resolved["count"] = 1
            inferred.add("count")
        else:
            count_str = str(count_raw).strip()
            if not (count_str.isdigit() and int(count_str) >= 1):
                return False, resolved, inferred, f"挿入行数『{count_raw}』が不正です（1以上の整数）"
            resolved["count"] = int(count_str)

    elif op == "DRAW_BORDERS":
        pass   # 引数無し・表全体が対象（検証することが無い）

    elif op == "AUTOFIT":
        pass   # 引数無し・全列が対象（検証することが無い）

    elif op == "PIVOT":
        # ★ AGGREGATE と同じ2 slot（group_col/value_col）。列名の実在確認だけ共有する。
        if (err := resolve_in("group_col", first_sheet)):
            return False, resolved, inferred, err
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err

    # ★ 致命3(W10e): 一括定数書き換え。値は LLM に確定させない（A' 原則）。
    elif op == "SET_COLUMN_VALUE":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        llm_value_raw = resolved.pop("value", None)
        quoted = extract_quoted_literal(task)
        if quoted is None:
            return False, resolved, inferred, (
                "書き込む値が依頼文から一意に読み取れません。値を「」または『』で囲んで"
                "書いてください（例:「備考列を全部『確認済み』にして」）"
            )
        resolved["value"] = quoted
        resolved["_sources"] = {**resolved.get("_sources", {}), "value": f"依頼文: 「{quoted}」"}
        if llm_value_raw not in (None, "") and str(llm_value_raw) != quoted:
            resolved["_warnings"] = resolved.get("_warnings", []) + [
                f"LLM が返した値('{llm_value_raw}')と依頼文の引用('{quoted}')が食い違うため"
                f"依頼文側('{quoted}')を採用しました"
            ]

    # ★ EXTRACT: 単一条件（col × cmp × value）に一致する行を新シートへ抜き出す
    #   （コミット 2edcb08「EXTRACT op」参照）。col は実在検証、cmp は語彙の6値に限定、value は
    #   gte/lte/gt/lt なら数値必須・eq は数値化できればそのまま数値・できなければ文字列・
    #   contains は常に文字列（A' 原則: 数値化は機械が行う。LLM の言い分をそのまま信じない）。
    elif op == "EXTRACT":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        cmp = str(resolved.get("cmp", "")).strip().lower()
        if cmp not in _EXTRACT_CMPS:
            return False, resolved, inferred, (
                f"比較『{resolved.get('cmp')}』は {'/'.join(_EXTRACT_CMPS)} のどれでもありません"
            )
        resolved["cmp"] = cmp
        raw_value = resolved.get("value")
        if raw_value in (None, ""):
            return False, resolved, inferred, "抽出する値(value)が依頼文から読み取れません"
        if cmp == "contains":
            resolved["value"] = str(raw_value)
        else:
            try:
                resolved["value"] = float(raw_value)
            except (TypeError, ValueError):
                if cmp != "eq":
                    return False, resolved, inferred, (
                        f"比較『{cmp}』には数値の値が必要ですが『{raw_value}』は数値に変換できません"
                    )
                resolved["value"] = str(raw_value)
        # ★ 単位H: 出力シートの見出し署名(= 元シートの見出し行そのもの)を _own_output_headers
        #   が組めるよう、決めた材料をここで resolved に積む（他 op の _target_sheet と同じ作法）。
        resolved["_source_headers"] = tuple(headers.get(first_sheet, []))
        resolved["_new_sheet"] = _extract_output_sheet_name(resolved["col"], cmp, resolved["value"])

    else:
        return False, resolved, inferred, f"未対応の操作: {op}"

    return True, resolved, inferred, None


# --- ③ 確認行（命令言語形式） -------------------------------------------------

_CONFIRM_FIELDS = {
    "SORT": (("対象", "col", None), ("順", "order", lambda v: "降順" if v == "desc" else "昇順")),
    "COMPUTE_COLUMN": (("演算対象", "operands", lambda v: " と ".join(v)), ("演算子", "operator", None),
                        ("対象列", "target", None)),
    "LOOKUP_FILL": (("対象シート", "target_sheet", None), ("対象列", "target_col", None),
                     ("参照シート", "source_sheet", None), ("キー列", "key_col", None)),
    "AGGREGATE": (("分類列", "group_col", None), ("集計列", "value_col", None)),
    "BOLD": (("対象", "target", None),),
    "FILL_COLOR": (("対象", "target", None), ("色", "color", None)),
    "NUMBER_FORMAT": (("対象列", "col", None), ("書式", "style", None)),
    "MERGE": (("範囲", "range", None),),
    "CHART": (("値列", "value_col", None),),
    "CENTER_ALIGN": (("対象", "target", None),),
    # ★ W8a 項目5: 表示ラベルのみ「倍率」→「率」（税率・掛け率の文脈での事務向け言い換え）。
    #   内部キー("factor")・関数名・コメントは不変。
    "APPEND_TOTAL": (("対象列", "col", None), ("ラベル", "label", None), ("率", "factor", None)),
    # ★ W9: 検証済みヘルパ4種の語彙昇格。
    "INSERT_ROWS": (("挿入位置", "at", None), ("行数", "count", None)),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": (("分類列", "group_col", None), ("集計列", "value_col", None)),
    "SET_COLUMN_VALUE": (("対象列", "col", None), ("値", "value", None)),
    "EXTRACT": (("対象列", "col", None), ("条件", "cmp", lambda v: _EXTRACT_CMP_LABELS.get(v, v)),
                 ("値", "value", lambda v: _format_extract_value(v))),
}

# ★ EXTRACT: 比較の語彙（設計書どおり6種）。gte/lte/gt/lt は数値比較・eq は値の型に応じて
#   数値/文字列どちらでも・contains は常に文字列の部分一致。
_EXTRACT_CMPS = ("gte", "lte", "gt", "lt", "eq", "contains")
_EXTRACT_CMP_LABELS = {"gte": "以上", "lte": "以下", "gt": "超", "lt": "未満",
                        "eq": "等しい", "contains": "を含む"}
_EXTRACT_CMP_CODE = {"gte": 0, "lte": 1, "gt": 2, "lt": 3, "eq": 4, "contains": 5}
_EXTRACT_SHEET_NAME_FORBIDDEN_RE = re.compile(r'[:\\/?*\[\]]')


def _format_extract_value(value) -> str:
    """EXTRACT のシート名/確認行に使う値の表示形。整数相当の float は小数点を付けない
       （40000.0 でなく 40000）。"""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:g}"
    return str(value)


def _extract_output_sheet_name(col: str, cmp: str, value) -> str:
    """★ A': 出力シート名は機械が決め打ちで組む（LLM に名前を決めさせない・設計書の例そのまま
       ―― 列『金額』・cmp gte・value 40000 → 『金額40000以上』）。
       Excel が禁じる文字(: \\ / ? * [ ])は '_' に置き換え、31文字上限（Excel のシート名
       制限）で切り詰める。"""
    label = _EXTRACT_CMP_LABELS.get(cmp, cmp)
    name = f"{col}{_format_extract_value(value)}{label}"
    return _EXTRACT_SHEET_NAME_FORBIDDEN_RE.sub("_", name)[:31]


# ★ W9 項目4: PIVOT(DataPilot) の既知の癖（README 記載・再描画で書式が撥ねる）を
#   確認行・結果表示の両方に一言添える。AGGREGATE(SummaryTable) との使い分けを促す。
PIVOT_CAVEAT = "書式なしの素の表になります。書式つきは『集計表』"


def format_confirmation_line(op: str, resolved_args: dict, inferred: set,
                              sheets: list | None = None, target_sheet: str | None = None) -> str:
    """命令言語形式の確認行を1行で組む（例: 解釈: 操作:並べ替え 対象:金額 順:降順）。
       推定で埋めた（数字表記から解決した等）引数には (推定) を付ける。
       ★ M2c: キー自体が resolved_args に無い任意項目（COMPUTE_COLUMN の target 等）は
       そのフィールドを丸ごと省略する（必須項目は常に存在するので既存の表示は変わらない）。
       ★ A': resolved_args["_sources"] に該当キーの出典があれば（例: 倍率:1.1）
       末尾に「（用語集: 消費税）」のように出典を添える（verify_dsl_args の APPEND_TOTAL が積む）。
       ★★ 挙動変更#3: sheets/target_sheet を渡すと、複数シートのブックでは先頭に
       `シート:『売上データ』(1枚目)` を載せる（本体は
       ailine_core.target_sheet.format_sheet_field・1枚のブックでは None で従来どおり）。
       その場合 LOOKUP_FILL の「対象シート:」欄は同じ値の二重表示になるので省く
       （転記の「参照シート:」はもう一方のシートなので残す）。"""
    sources = resolved_args.get("_sources") or {}
    sheet_field = format_sheet_field(sheets or [], target_sheet)
    parts = ([sheet_field] if sheet_field else []) + [f"操作:{OP_LABELS.get(op, op)}"]
    for label, key, transform in _CONFIRM_FIELDS.get(op, ()):
        if key not in resolved_args or (sheet_field and key == "target_sheet"):
            continue
        val = resolved_args.get(key)
        shown = transform(val) if transform else val
        tag = "(推定)" if key in inferred else ""
        if key in sources:
            tag += f"（{sources[key]}）"
        parts.append(f"{label}:{shown}{tag}")
    return "解釈: " + " ".join(parts)


# --- ④ 決定論 codegen（op → Basic テンプレ。LLM は一切使わない） --------------

def _wrap_basic(body: str) -> str:
    """CONTRACT と同じ骨格（Option 2行 + Sub Run(oDoc As Object) 1つ）で包む。"""
    return "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\n" + body + "End Sub\n"


def _wrap_basic_for_sheet(body: str, book_meta: dict, target_sheet: str | None) -> str:
    """★ 挙動変更#2: 薄い配線。本体（対象シートの一時的な並べ替えロジック）は
       ailine_core.target_sheet.wrap_basic_for_sheet に置く（移植可能性の番人 —
       tests/ailine_py_line_budget.txt 参照）。_wrap_basic を「対象シートが1枚目のときの
       既定ラップ」としてコールバックで渡す（ailine_core → ailine の逆流を避けるため）。"""
    return wrap_basic_for_sheet(body, _wrap_basic, book_meta.get("sheets") or [], target_sheet)


def _scan_last_row_basic(var: str = "oSheet", key_col: str = "0",
                          start_row: str = "1", min_ok: str | None = None) -> str:
    """走査ループの定型（refs の作法どおり：A列を上から走査して最終データ行を探す）。
       ★ W3: start_row（Basic 0起点の走査開始行＝見出し行の直下）を渡すと、見出しが
       物理1行目でない帳票でも正しい行から走査する（既定 "1" は旧挙動と同一）。
       min_ok（"lastRow がこの値未満なら Exit Sub"の閾値）を渡すと保存境界を調整できる
       （既定は start_row と同じ＝『データ0行なら何もしない』。見出しを含めて範囲を
       スタイリングする操作は min_ok に start_row-1 相当を渡す＝データ0行でも見出しは扱える）。"""
    if min_ok is None:
        min_ok = start_row
    return (f"    lastRow = {start_row}\n"
            f"    Do While {var}.getCellByPosition({key_col}, lastRow).getString() <> \"\"\n"
            f"        lastRow = lastRow + 1\n"
            f"    Loop\n"
            f"    lastRow = lastRow - 1\n"
            f"    If lastRow < {min_ok} Then Exit Sub\n")


def codegen_dsl(op: str, resolved_args: dict, book_meta: dict, use_formula: bool = True) -> str:
    """④ 決定論 codegen。既存ヘルパへの Call を最優先し、無い操作だけテンプレ Basic を書く。
       ★ W3: book_meta["header_rows"] があれば対象シートの見出し行(1起点)をそこから読み
       （★ 挙動変更#2 より前は対象シート＝常に1枚目だった。下記参照）、
       0起点(hr0)に変換して全 op に一貫して渡す（『三層全部が同じ見出し推定を使う』の codegen 側）。
       header_rows が無い/キーが無い book_meta（_SAMPLE_META 等の旧テスト値）は既定1行目
       （hr0=0）＝旧挙動と完全に同一の Basic を生成する。
       use_formula: COMPUTE_COLUMN の既定を式（=B2*C2 等）にする（★ W3 Part3）。False で
       従来の値ベタ書きに戻す（--values）。
       ★ 挙動変更#2: 対象シートは resolved_args["_target_sheet"]（verify_dsl_args が一箇所で
       決めた値）を読む。無ければ従来どおり1枚目（後方互換 — codegen_dsl を直接呼ぶ既存の
       単体テスト・golden はこのキーを持たない args を渡しており、その挙動は変えない）。
       wrap() は _wrap_basic の対象シート対応版（_wrap_basic_for_sheet）に book_meta/
       first_sheet を閉じ込めたショートハンド — 対象シートの決定を codegen 側でも
       一箇所（_wrap_basic_for_sheet）に寄せるための配線。"""
    headers = book_meta["headers"]
    first_sheet = resolved_args.get("_target_sheet") or book_meta["sheets"][0]
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    hr0 = header_row - 1   # Basic 0起点の見出し行

    def wrap(body: str) -> str:
        return _wrap_basic_for_sheet(body, book_meta, first_sheet)

    if op == "SORT":
        col_idx = headers[first_sheet].index(resolved_args["col"])
        asc = "True" if resolved_args["order"] == "asc" else "False"
        last_col = len(headers[first_sheet]) - 1
        return wrap(f"    Call SortByColumn(oDoc, {hr0}, {last_col}, {col_idx}, {asc})\n")

    if op == "LOOKUP_FILL":
        theaders = headers[resolved_args["target_sheet"]]
        key_idx = theaders.index(resolved_args["key_col"])
        src = resolved_args["source_sheet"].replace('"', '""')
        target_col_name = resolved_args["target_col"]
        # ★ W10c 致命2: target_col が対象シートに実在しない場合（verify_dsl_args が依頼文に
        #   同じ列名があると確認済み）は、COMPUTE_COLUMN の新規列作成と同じ考え方で
        #   末尾に新しい列を作ってから転記する（無関係な既存列を上書きしない）。
        if target_col_name in theaders:
            tgt_idx = theaders.index(target_col_name)
            header_write = ""
        else:
            tgt_idx = len(theaders)   # 0起点・次の空き列
            header_name = str(target_col_name).replace('"', '""')
            header_write = (f'    oDoc.Sheets.getByIndex(0).getCellByPosition({tgt_idx}, {hr0})'
                             f'.setString("{header_name}")\n')
        return wrap(header_write +
                            f'    Call VLookupFromTable(oDoc, {hr0}, {key_idx}, {tgt_idx}, "{src}")\n')

    if op == "AGGREGATE":
        g_idx = headers[first_sheet].index(resolved_args["group_col"])
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        return wrap(f"    Call SummaryTable(oDoc, {hr0}, {g_idx}, {v_idx})\n")

    if op == "NUMBER_FORMAT":
        col_idx = headers[first_sheet].index(resolved_args["col"])
        return wrap(f"    Call FormatThousands(oDoc, {hr0}, {col_idx})\n")

    if op == "MERGE":
        c1s, r1s, c2s, r2s = re.match(
            r"([A-Za-z]{1,3})(\d+):([A-Za-z]{1,3})(\d+)", resolved_args["range"]).groups()
        col1 = column_index_from_string(c1s.upper()) - 1
        col2 = column_index_from_string(c2s.upper()) - 1
        row1, row2 = int(r1s) - 1, int(r2s) - 1
        return wrap(f"    Call MergeCells(oDoc, {col1}, {row1}, {col2}, {row2})\n")

    if op == "CHART":
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        return wrap(f"    Call InsertBarChart(oDoc, {hr0}, {v_idx})\n")

    if op == "APPEND_TOTAL":
        # ★ W6: ヘルパは無し（罫線・カンマ等の見栄えまでは踏み込まない・素の SUM 式だけ）。
        #   データ最終行の直下に [ラベル文字列 | 合計式] を書く。
        #   ラベルは対象列の左隣に置く（既存の帳票では『合計』の文字が金額の左に来るのが自然
        #   で、対象列自体を上書きしない＝既存構造を壊さない置き方）。対象列が表の最左端
        #   （col_idx=0）の場合は左隣が無いため、値のみを書きラベルは省略する。
        # ★ B: 挿入耐性式（bench/formula_spike_work2 で実測）。SUM(D2:INDEX(D:D;ROW()-1))
        #   型で書く。INDEX(D:D;ROW()-1) は「この式自身の1行上」を指すので、後でデータ行を
        #   1本挿入しても SUM 範囲が自動で追従する（静的な "D2:D5" 型は追従しない）。
        #   ★ setFormula は LO 方言＝INDEX の2引数区切りはセミコロン(;)。カンマ(,)で
        #   書くと #VALUE!/#NAME? になる（実測: formula_spike_RESULTS.md 追記分）。
        #   保存後は自動でカンマ形 "=SUM(D2:INDEX(D:D,ROW()-1))*1.1" に変換される
        #   （check_append_total はこの保存後カンマ形と照合する）。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        label = str(resolved_args.get("label", "合計")).replace('"', '""')
        factor = float(resolved_args.get("factor", 1) or 1)
        col_letter = get_column_letter(col_idx + 1)
        start_excel_row = hr0 + 2   # データ先頭行（Basic 0起点 hr0+1）の Excel(1起点) 行
        factor_tail = "" if factor == 1 else f"*{factor:g}"
        body = ("    Dim oSheet As Object, lastRow As Long, totalRow As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1))
                + "    totalRow = lastRow + 1\n")
        if col_idx > 0:
            body += f'    oSheet.getCellByPosition({col_idx - 1}, totalRow).setString("{label}")\n'
        body += (f'    oSheet.getCellByPosition({col_idx}, totalRow).setFormula('
                 f'"=SUM(" & "{col_letter}" & {start_excel_row} & ":INDEX(" & "{col_letter}" & '
                 f'":" & "{col_letter}" & ";ROW()-1))" & "{factor_tail}")\n')
        return wrap(body)

    if op == "CENTER_ALIGN":
        if resolved_args["target"] == "all":
            last_col = len(headers[first_sheet]) - 1
            return wrap(f"    Call AlignCenter(oDoc, {hr0}, {last_col})\n")
        # col:NAME はヘルパ無し → refs の作法（走査して範囲を求め HoriJustify）でテンプレを書く。
        col_idx = headers[first_sheet].index(resolved_args["target"][4:])
        body = ("    Dim oSheet As Object, oRange As Object, lastRow As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                + f"    oRange = oSheet.getCellRangeByPosition({col_idx}, {hr0}, {col_idx}, lastRow)\n"
                "    oRange.HoriJustify = com.sun.star.table.CellHoriJustify.CENTER\n")
        return wrap(body)

    if op == "BOLD":
        target = resolved_args["target"]
        if target.startswith("row:"):
            row_idx = int(target[4:]) - 1
            # ★ W3: 列幅は Basic で走査せず、接地済みの見出し列数(headers)から決定論的に
            #   決める（多段見出しで先頭列が空欄のケースで走査が誤って -1 になるのを回避）。
            last_col = len(headers[first_sheet]) - 1
            body = (f"    Call StyleBold(oDoc, 0, {row_idx}, {last_col}, {row_idx})\n")
        else:
            col_idx = headers[first_sheet].index(target[4:])
            body = ("    Dim oSheet As Object, lastRow As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                    + f"    Call StyleBold(oDoc, {col_idx}, {hr0}, {col_idx}, lastRow)\n")
        return wrap(body)

    if op == "FILL_COLOR":
        target = resolved_args["target"]
        hexcolor = COLOR_MAP[resolved_args["color"]]
        if target.startswith("row:"):
            row_idx = int(target[4:]) - 1
            # ★ W3: 列幅は Basic で走査せず、接地済みの見出し列数(headers)から決定論的に
            #   決める（多段見出しで先頭列が空欄のケースで走査が誤って -1 になるのを回避）。
            last_col = len(headers[first_sheet]) - 1
            body = ("    Dim oSheet As Object, c As Integer\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    f"    For c = 0 To {last_col}\n"
                    f"        oSheet.getCellByPosition(c, {row_idx}).CellBackColor = &H{hexcolor}&\n"
                    "    Next c\n")
        else:
            col_idx = headers[first_sheet].index(target[4:])
            body = ("    Dim oSheet As Object, lastRow As Long, r As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                    + f"    For r = {hr0} To lastRow\n"
                    f"        oSheet.getCellByPosition({col_idx}, r).CellBackColor = &H{hexcolor}&\n"
                    "    Next r\n")
        return wrap(body)

    if op == "COMPUTE_COLUMN":
        operands = resolved_args["operands"]
        operator = resolved_args["operator"]
        target = resolved_args.get("target")

        if len(operands) == 1:
            # ★ W10b 項目3: 税込み/税抜き等「列 × 率」パターン（1列 + 機械確定した factor）。
            #   codegen は既存の2列版と同じ書き方（式=セル参照、値ベタ書きの選択も同じ）。
            op1 = operands[0]
            i1 = headers[first_sheet].index(op1)
            factor = float(resolved_args["factor"])
            if target:
                new_col = headers[first_sheet].index(target)
                header_write = ""
            else:
                new_col = len(headers[first_sheet])
                # ★ W10c 中: verify_dsl_args が税込み/税抜きと判定できていれば自然な
                #   日本語見出し（例:「税込金額」）を使う。無ければ従来どおりの数式風見出し。
                header_name = str(resolved_args.get("_new_col_label")
                                   or f"{op1}{operator}{factor:g}").replace('"', '""')
                header_write = f"    oSheet.getCellByPosition({new_col}, {hr0}).setString(\"{header_name}\")\n"
            if use_formula:
                col1_letter = get_column_letter(i1 + 1)
                write_line = (f'        oSheet.getCellByPosition({new_col}, i).setFormula('
                              f'"=" & "{col1_letter}" & (i + 1) & "{operator}{factor:g}")\n')
            else:
                write_line = (f"        oSheet.getCellByPosition({new_col}, i).setValue("
                              f"oSheet.getCellByPosition({i1}, i).getValue() {operator} {factor:g})\n")
            body = ("    Dim oSheet As Object, lastRow As Long, i As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1))
                    + header_write
                    + f"    For i = {hr0 + 1} To lastRow\n"
                    + write_line
                    + "    Next i\n")
            return wrap(body)

        op1, op2 = operands
        i1 = headers[first_sheet].index(op1)
        i2 = headers[first_sheet].index(op2)
        # ★ M2c: target(実在列名) 指定時はその列に書く（新規列を作らない）。
        #   無指定なら従来どおり次の空き列に新規列を作る。
        if target:
            new_col = headers[first_sheet].index(target)
            header_write = ""   # 既存の見出しはそのまま（上書きしない）
        else:
            new_col = len(headers[first_sheet])   # 0起点で次の空き列
            # ★ W3 改定(2026-08-20): verify_dsl_args が「target は実在しないが依頼文に
            #   実在する語」と判定していれば _new_col_label に利用者の指名が入っている
            #   （単列×率パターンの税込/税抜ラベルと同じ仕組み）。無ければ従来どおりの
            #   数式風見出し。
            header_name = str(resolved_args.get("_new_col_label")
                               or f"{op1}{operator}{op2}").replace('"', '""')
            header_write = f"    oSheet.getCellByPosition({new_col}, {hr0}).setString(\"{header_name}\")\n"
        # ★ W3 Part3: 既定は式（setFormula）。formula_spike の実測どおり、単純な行内の
        #   二項演算（=B2*C2 型）は区切り記号もシート参照も無いため LO 方言(;/.)の
        #   気遣いが不要（多引数関数・シート跨ぎ参照だけが方言の対象＝ここでは無縁）。
        #   行番号だけが Basic 側ループ変数(i)で動くので、その部分だけ実行時に連結する。
        if use_formula:
            col1_letter = get_column_letter(i1 + 1)
            col2_letter = get_column_letter(i2 + 1)
            write_line = (f'        oSheet.getCellByPosition({new_col}, i).setFormula('
                          f'"=" & "{col1_letter}" & (i + 1) & "{operator}" & "{col2_letter}" & (i + 1))\n')
        else:
            write_line = (f"        oSheet.getCellByPosition({new_col}, i).setValue("
                          f"oSheet.getCellByPosition({i1}, i).getValue() {operator} "
                          f"oSheet.getCellByPosition({i2}, i).getValue())\n")
        body = ("    Dim oSheet As Object, lastRow As Long, i As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1))
                + header_write
                + f"    For i = {hr0 + 1} To lastRow\n"
                + write_line
                + "    Next i\n")
        return wrap(body)

    # --- ★ W9: 検証済みヘルパ4種の語彙昇格。いずれも helpers/*.bas のヘルパは headerRow
    #   引数を取らない（物理1行目を前提に自前走査する既存実装・ここでは変更しない）ため、
    #   codegen 側も hr0 を渡さずそのまま Call するだけ。
    if op == "INSERT_ROWS":
        at0 = int(resolved_args["at"]) - 1   # 1起点(Excel行番号) → 0起点(Basic)
        count = int(resolved_args.get("count", 1) or 1)
        return wrap(f"    Call InsertRows(oDoc, {at0}, {count})\n")

    if op == "DRAW_BORDERS":
        return wrap("    Call DrawTableBorders(oDoc)\n")

    if op == "AUTOFIT":
        return wrap("    Call AutoFitColumns(oDoc)\n")

    if op == "PIVOT":
        g_idx = headers[first_sheet].index(resolved_args["group_col"])
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        return wrap(f"    Call PivotSum(oDoc, {g_idx}, {v_idx})\n")

    if op == "SET_COLUMN_VALUE":
        # ★ 致命3(W10e): ヘルパ無し・既存列のデータ行全部に同じ文字列を setString する
        #   （CENTER_ALIGN の col: 分岐と同じ「走査してヘッダ直下から最終行まで」の作法）。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        value = str(resolved_args["value"]).replace('"', '""')
        body = ("    Dim oSheet As Object, lastRow As Long, r As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1))
                + f"    For r = {hr0 + 1} To lastRow\n"
                f"        oSheet.getCellByPosition({col_idx}, r).setString(\"{value}\")\n"
                "    Next r\n")
        return wrap(body)

    if op == "EXTRACT":
        # ★ ヘルパへの Call 1行だけ（helpers/AiLineHelpers.bas:ExtractRows）。
        #   型を保つコピー（getValue/setValue・getString/setString の分岐）は helper 側。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        cmp_code = _EXTRACT_CMP_CODE[resolved_args["cmp"]]
        value = resolved_args["value"]
        if isinstance(value, str):
            value_lit = '"' + value.replace('"', '""') + '"'
        else:
            value_lit = f"{float(value):g}"
        dst_name = str(resolved_args["_new_sheet"]).replace('"', '""')
        return wrap(f'    Call ExtractRows(oDoc, {hr0}, {col_idx}, {cmp_code}, {value_lit}, "{dst_name}")\n')

    raise ValueError(f"未対応の op: {op}")


# --- ⑥ op 別事後条件（達成の機械検証。openpyxl で out ファイルを読むだけ・LO 不要） ----

def _col_index_by_header(ws, name: str, header_row: int = 1):
    """見出し行(既定は物理1行目・header_row で指定可)を左から走査して name に一致する
       列の1起点インデックスを返す。無ければ None。★ W3: header_row 省略時は旧挙動と同一。
       ★ W3: header_row>1 の子見出し行で空欄の列は、book_columns() と同じ規則で
       真上の行を遡って引き継ぐ（多段見出しの先頭列対策。無いと『7月』等キー列より
       手前の空欄列で走査が誤って打ち切られる）。"""
    c = 1
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v in (None, "") and header_row > 1:
            for up in range(header_row - 1, 0, -1):
                uv = ws.cell(row=up, column=c).value
                if uv not in (None, ""):
                    v = uv
                    break
        if v in (None, ""):
            return None
        if str(v) == name:
            return c
        c += 1


def _scan_last_row(ws, key_col: int = 1, header_row: int = 1) -> int:
    """key_col(1起点)を上から走査した最終データ行（見出し行を除く）。データが無ければ
       header_row。★ W3: header_row 省略時は旧挙動（見出し=1行目・データ開始=2行目）と同一。"""
    r = header_row + 1
    while ws.cell(row=r, column=key_col).value not in (None, ""):
        r += 1
    return r - 1


def _scan_last_col(ws, header_row: int = 1) -> int:
    """見出し行(既定は物理1行目)を左から走査した最終列（1起点）。
       ★ W3: _col_index_by_header と同じ規則で、子見出し行(header_row>1)の空欄列は
       真上の行を遡って引き継ぐ（多段見出しの先頭列対策）。"""
    def _effective(c: int):
        v = ws.cell(row=header_row, column=c).value
        if v in (None, "") and header_row > 1:
            for up in range(header_row - 1, 0, -1):
                uv = ws.cell(row=up, column=c).value
                if uv not in (None, ""):
                    return uv
        return v
    c = 1
    while _effective(c) not in (None, ""):
        c += 1
    return c - 1


def _apply_operator(a, b, operator: str):
    a = a or 0
    b = b or 0
    if operator == "+":
        return a + b
    if operator == "-":
        return a - b
    if operator == "*":
        return a * b
    if operator == "/":
        return a / b if b else 0
    raise ValueError(operator)


def _is_number(v) -> bool:
    """★ 止血2: bool は int のサブクラスだが数値セルとしては扱わない（True/False混入対策）。"""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ★ 止血1/2 共通の文言。事後条件チェッカーは検証対象0件を絶対に「合格」にしない
#   （D検体: no-opを『行数が少なく比較不要』で素通ししていた根治）。
_ZERO_TARGET_REASON = "事後条件の検証対象が0件（何も検証できていない）"


# --- ★ 算術恒等の検算（二重計上・合計行の位置） -------------------------------
#   独立レビューの実測: check_append_total は期待値を「合計式が生成したのと同じ範囲」から
#   作っていた ―― 検算が被検算と同じ盲点を使う恒真式。既存の合計 300 を持つ表に合計を
#   足すと 600 が書かれ「3 行の合計を検証」と言って ✓ が出た。並べ替えにも同型がある
#   （合計行が2行目に来ても「5 行を検証（降順）」で通る）。
#   判定は ailine_core/sum_identity.py（語も書式も読まない純粋な算術）に閉じ込め、
#   ここは「どのセルか」を人が読める文にするだけにする。

def _fmt_amount(v) -> str:
    """恒等式の説明に出す数値（整数なら小数点を付けない・指数表記にしない）。"""
    return str(int(v)) if float(v).is_integer() else str(v)


def _column_block_values(bv, ws, idx: int, header_row: int, sheet: str | None) -> list:
    """対象列を見出しの次の行から**空欄まで**走査した (行番号, 計算後の値) の並び。

    ★ 行の広がりは式ビュー（そのセルが空欄か）で決め、値は値ビュー（キャッシュ）から取る
      ―― 式で埋まった列を『空』と誤読せず、式文字列を数値と誤読しないための組み合わせ。
    ★ 空欄で切れる＝空行で区切られた2つ目の塊は読まない（別の表を続けて足すと、
      無関係な行が『上の合計』に当たってしまう）。"""
    out: list = []
    r = header_row + 1
    while ws.cell(row=r, column=idx).value not in (None, ""):
        out.append((r, bv.cell_value(r, idx, sheet=sheet)))
        r += 1
    return out


def _nested_total_reason(values: list, sheet_name: str, idx: int) -> str | None:
    """**足し込んだ範囲の最終行**が『自分より上の全部の合計』なら、その1行を名指しする。

    ★ 位置で判定する（2段構え）:
      1. 並びの最後の数値行は、今この op が書いた合計そのもの。範囲から外す
         （自分自身と照合したら恒真式に戻る）。
      2. 残った範囲の中でも、**最終行**で一致したときだけ鳴らす。既にある合計は
         その塊の一番下に在り、二重計上はそういう起き方をする。範囲の**真ん中**で
         一致するのは偶然の側 ―― 実測: demo/sales.xlsx の 部門/金額 =
         100,200,**300**,400,500,250 は 300 が開発部門のただの売上なのに
         100+200 と一致し、README の quickstart が exit 1 で落ちていた。
         手元の実検体 70 数値列で 発火 5 件（うち偽陽性 3）→ 2 件（どちらも本物の
         合計行）に下がった。語も書式も読まない性質は保たれる（算術＋位置だけ）。
    ★ 代償は取り逃がし: 『本体→小計→本体→小計→合計』のように**最終行でない**位置に
      合計がある帳票は、二重に数えていても鳴らない（見えない側に倒した）。"""
    numeric = [(row, value) for row, value in values if _is_number(value)]
    for hit in rows_matching_sum_above(numeric[:-1]):
        if not hit.is_last:
            continue
        span = f"{_cell_ref(hit.term_rows[0], idx)}:{_cell_ref(hit.term_rows[-1], idx)}"
        return (f"{sheet_name}!{_cell_ref(hit.row, idx)} の {_fmt_amount(hit.value)} は "
                f"{span} の合計と一致します"
                f"（この行まで足し込むと既にある合計を二重に数えます）")
    return None


def _total_row_left_the_bottom_reason(path: Path, source_book: Path | None, args: dict,
                                       header_row: int, idx: int, sheet_name: str) -> str | None:
    """並べ替えの前は最下行が『上の全部の合計』だったのに、後ではそうでなくなったら1行返す。

    ★ 判定は「存在」でなく「位置」: 並べ替え前の表にも合計行は在る（それが正常）。
      危険なのは合計と一致する行が**最下行でなくなった**とき ―― 合計行がただのデータとして
      一緒に並べ替えられ、表の途中に混ざった状態。並べ替え前は沈黙し、後にだけ鳴る。
    ★ source_book（適用前のコピー）が無い経路では黙る ―― before が無ければ位置の変化は
      測れない。測れないことを鳴らす側にも合格側にも寄せない（断定しない）。"""
    sheet = args.get("_target_sheet")
    if source_book is None or not Path(source_book).exists():
        return None
    with BookView(source_book) as bv_before:
        ws_before = bv_before.sheet(sheet)
        idx_before = _col_index_by_header(ws_before, args["col"], header_row=header_row)
        if idx_before is None:
            return None
        before_values = _column_block_values(bv_before, ws_before, idx_before, header_row, sheet)
    was_total = next((h for h in rows_matching_sum_above(before_values) if h.is_last), None)
    if was_total is None:
        return None                      # 元から最下行に合計は無い＝この検査の対象外
    with BookView(path) as bv_after:
        ws_after = bv_after.sheet(sheet)
        after_values = _column_block_values(bv_after, ws_after, idx, header_row, sheet)
    if any(h.is_last for h in rows_matching_sum_above(after_values)):
        return None                      # 合計はまだ最下行にある
    moved_to = next((r for r, v in after_values
                     if _is_number(v) and abs(v - was_total.value) <= 1e-6), None)
    where = (f"{sheet_name}!{_cell_ref(moved_to, idx)}" if moved_to
             else "並べ替えた後の表の中に見つかりません")
    return (f"並べ替える前は最下行 {sheet_name}!{_cell_ref(was_total.row, idx_before)} の "
            f"{_fmt_amount(was_total.value)} が上の {len(was_total.term_rows)} 行の合計でしたが、"
            f"並べ替えた後は {where} にあり最下行ではありません"
            f"（合計行がデータとして一緒に並べ替えられています）")


def check_sort(path: Path, args: dict, header_row: int = 1, use_formula: bool = False,
                source_book: Path | None = None) -> tuple:
    """SORT の事後条件。戻り値は (status, reason)。status ∈ {"pass","warn","fail"}。
       ★ 止血1: 検証対象が0件なら fail、1件（順序が定義できない）なら warn とし、
       どちらも「機械検証済み」とは名乗らない。
       ★ 止血2: 合計行等の非数値/None セルは比較から除外し、除外件数を表示する
       （C②: None >= int の生トレースバックの根治）。全部除外なら0件と同じ扱い。
       ★ W3: header_row(1起点、省略時1) が『接地・codegen』と同じ見出し行を指す。
       ★ W10f 項目1: use_formula のとき対象列を data_only(計算後の値)側から読む
       （check_compute_column と同型のバグ — 前段が式で作った計算列(小計等)を SORT の
       対象列にした場合、raw 側は式文字列のままで全行『数値でない』扱いになっていた）。
       SORT は相対順序という『全行をまたぐ』検証のため、式にキャッシュ値が無く読めない
       行が1件でもあれば、その行を除いた残りだけで『順序OK』と判定するのは危険
       （除いた行が実際は順序を崩していても見逃す＝COMPUTE_COLUMN の行独立検証とは違い
       部分採点できない）。0cf9218 空虚な検証合格の禁止の趣旨のまま fail で打ち切る。
       ★ 算術恒等の検算: 並び順が合っていても、合計行がデータとして一緒に並べ替えられて
       表の途中に混ざったら ✓ は出さない（source_book が渡された経路のみ・
       _total_row_left_the_bottom_reason 参照）。"""
    sheet_name = args.get("_target_sheet")
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        sheet_name = sheet_name or ws.title
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        raw_vals = [ws.cell(row=r, column=idx).value for r in range(header_row + 1, last + 1)]
        if use_formula:
            eff_vals = [bv.cell_value(r, idx, sheet=args.get("_target_sheet")) for r in range(header_row + 1, last + 1)]
        else:
            eff_vals = raw_vals
    vals = []
    excluded = 0
    uncached = 0
    for rv, ev in zip(raw_vals, eff_vals):
        if _is_number(ev):
            vals.append(ev)
        elif use_formula and isinstance(rv, str) and rv.startswith("="):
            uncached += 1   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
        else:
            excluded += 1
    note = f"（数値でない {excluded} 行は対象外）" if excluded else ""
    if uncached:
        return "fail", (f"並び順の検証対象に式はあるがキャッシュ値が無く検証できない行が "
                         f"{uncached} 件あり、順序を検証できません"
                         f"（LibreOffice を通していない可能性）{note}")
    if len(vals) == 0:
        return "fail", _ZERO_TARGET_REASON + note
    if len(vals) == 1:
        return "warn", f"検証対象が1行のみ（並べ替えの意味がありません）{note}"
    asc = args["order"] == "asc"
    ok = (all(vals[i] <= vals[i + 1] for i in range(len(vals) - 1)) if asc
          else all(vals[i] >= vals[i + 1] for i in range(len(vals) - 1)))
    if not ok:
        return "fail", f"列『{args['col']}』が指定順（{args['order']}）に並んでいない{note}"
    moved = _total_row_left_the_bottom_reason(path, source_book, args, header_row, idx, sheet_name)
    if moved:
        return "fail", moved
    return "pass", f"{len(vals)} 行を検証（{'昇順' if asc else '降順'}）{note}"


def check_compute_column(path: Path, args: dict, header_row: int = 1,
                          use_formula: bool = False) -> tuple:
    """★ 止血1/2: 演算対象が非数値/None の行（合計行等）は「対象外」として除外し件数を
       表示する。除外後に検証できた行が0件なら fail（機械検証済みと名乗らない）。
       対象列(target)自体が非数値なのは演算が本当に効いていない証拠なので除外せず fail。
       ★ W3 Part3: use_formula=True のとき事後条件を二層化する — ①通常の openpyxl 読み
       で保存された式文字列が期待形か ②data_only 読みでキャッシュ値が演算結果と一致するか。
       両方合格して初めて pass にする（式だけ合っていて未計算/値だけ合っていて式が
       無いケースの両方を見逃さない）。
       ★ W10b 項目3: operands が1つだけ（税込み/税抜き等「列 × 率」パターン）の場合は
       check_compute_column_single_factor に委譲する。
       ★ W10f 項目1: use_formula のとき operands は data_only(計算後の値)側から読む
       （旧実装は raw 側=式ビューから読んでいたため、前段が式で作った計算列(小計等)を
       operand にすると raw 値が式文字列のままで『数値でない』扱いになり、全行除外→
       検証対象0件failで計画ごとロールバックしていた＝実測の再現形そのもの）。
       operand 自体が式なのにキャッシュ値が無い(LibreOffice を通していない)行は、
       『対象が無い』(合計行等の非数値/空欄)とは別に数える（0cf9218 空虚な検証合格の
       禁止の趣旨 — 対象は本当は有るのに読めていないだけ、と混同しない）。COMPUTE_COLUMN
       は行ごとに独立な検証なので、うち検証できた行だけを checked に数えても他行の結果は
       歪めない（AGGREGATE/SORT の全行をまたぐ検証とは違い部分採点できる）。"""
    if len(args["operands"]) == 1:
        return check_compute_column_single_factor(path, args, header_row=header_row,
                                                    use_formula=use_formula)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        op1, op2 = args["operands"]
        i1 = _col_index_by_header(ws, op1, header_row=header_row)
        i2 = _col_index_by_header(ws, op2, header_row=header_row)
        # ★ M2c: target(実在列名) 指定時はその列を検証する。無指定なら従来どおり自動命名の新列。
        # ★ W3 改定(2026-08-20): codegen_dsl と同じ見出し名決定（_new_col_label があれば
        #   利用者が依頼文で名指しした新規列名）。
        target = args.get("target")
        newname = target or args.get("_new_col_label") or f"{op1}{args['operator']}{op2}"
        inew = _col_index_by_header(ws, newname, header_row=header_row)
        if i1 is None or i2 is None or inew is None:
            return "fail", f"演算対象または対象列『{newname}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        col1_letter = get_column_letter(i1)
        col2_letter = get_column_letter(i2)
        checked = 0
        excluded = 0
        uncached = 0
        for r in range(header_row + 1, last + 1):
            a_raw = ws.cell(row=r, column=i1).value
            b_raw = ws.cell(row=r, column=i2).value
            got = ws.cell(row=r, column=inew).value
            a = bv.cell_value(r, i1, sheet=args.get("_target_sheet")) if use_formula else a_raw
            b = bv.cell_value(r, i2, sheet=args.get("_target_sheet")) if use_formula else b_raw
            if not _is_number(a) or not _is_number(b):
                # ★ W10f: operand 自体が式(前段の計算列)で、かつキャッシュ値が無い行は
                #   『数値でない対象外』と別カウントする（下の note で区別して表示）。
                if use_formula and (
                    (not _is_number(a) and isinstance(a_raw, str) and a_raw.startswith("="))
                    or (not _is_number(b) and isinstance(b_raw, str) and b_raw.startswith("="))
                ):
                    uncached += 1
                else:
                    excluded += 1   # 例: 合計行で演算対象セルが空欄
                continue
            want = _apply_operator(a, b, args["operator"])
            if use_formula:
                expect_formula = f"={col1_letter}{r}{args['operator']}{col2_letter}{r}"
                if not isinstance(got, str) or got.replace(" ", "") != expect_formula:
                    return "fail", f"{r}行目: 式が期待形でない (期待 {expect_formula} 実際 {got!r})"
                got_cached = bv.cell_value(r, inew, sheet=args.get("_target_sheet"))
                if not _is_number(got_cached) or abs(got_cached - want) > 1e-6:
                    return "fail", f"{r}行目: 式のキャッシュ値が不一致 (期待 {want} 実際 {got_cached!r})"
            else:
                if not _is_number(got) or abs(got - want) > 1e-6:
                    return "fail", f"{r}行目: 期待 {want} 実際 {got}"
            checked += 1
    note_parts = []
    if excluded:
        note_parts.append(f"数値でない {excluded} 行は対象外")
    if uncached:
        note_parts.append(f"演算対象の式にキャッシュ値が無く検証できない {uncached} 行")
    note = f"（{'・'.join(note_parts)}）" if note_parts else ""
    if checked == 0:
        return "fail", _ZERO_TARGET_REASON + note
    if use_formula:
        return "pass", f"{checked} 行を検証（式・キャッシュ値とも一致）{note}"
    return "pass", f"{checked} 行を検証{note}"


def check_compute_column_single_factor(path: Path, args: dict, header_row: int = 1,
                                        use_formula: bool = False) -> tuple:
    """★ W10b 項目3: COMPUTE_COLUMN の「1列 × 率」パターン（税込み/税抜き等）専用の事後条件。
       check_compute_column（2列版）と同じ二層検証（式の期待形・data_only キャッシュ値）を
       1列 + factor に合わせて行う。factor は verify_dsl_args が機械確定済みの値
       （resolved["factor"]）をそのまま受け取る前提。
       ★ W10f 項目1: use_formula のとき operand は data_only(計算後の値)側から読む
       （check_compute_column と同型・前段が式で作った計算列(小計等)を operand にした
       場合の同じバグをここでも直す）。キャッシュ値が無い(式はあるが未計算)行は
       『数値でない対象外』と別カウントする（0cf9218 空虚な検証合格の禁止の趣旨）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        op1 = args["operands"][0]
        operator = args["operator"]
        factor = float(args.get("factor", 1) or 1)
        i1 = _col_index_by_header(ws, op1, header_row=header_row)
        target = args.get("target")
        # ★ W10c 中: codegen_dsl と同じ見出し名決定（_new_col_label があれば自然な日本語見出し）。
        newname = target or args.get("_new_col_label") or f"{op1}{operator}{factor:g}"
        inew = _col_index_by_header(ws, newname, header_row=header_row)
        if i1 is None or inew is None:
            return "fail", f"演算対象または対象列『{newname}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        col1_letter = get_column_letter(i1)
        checked = 0
        excluded = 0
        uncached = 0
        for r in range(header_row + 1, last + 1):
            a_raw = ws.cell(row=r, column=i1).value
            got = ws.cell(row=r, column=inew).value
            a = bv.cell_value(r, i1, sheet=args.get("_target_sheet")) if use_formula else a_raw
            if not _is_number(a):
                if use_formula and isinstance(a_raw, str) and a_raw.startswith("="):
                    uncached += 1   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
                else:
                    excluded += 1   # 例: 合計行で演算対象セルが空欄
                continue
            want = _apply_operator(a, factor, operator)
            if use_formula:
                expect_formula = f"={col1_letter}{r}{operator}{factor:g}"
                if not isinstance(got, str) or got.replace(" ", "") != expect_formula:
                    return "fail", f"{r}行目: 式が期待形でない (期待 {expect_formula} 実際 {got!r})"
                got_cached = bv.cell_value(r, inew, sheet=args.get("_target_sheet"))
                if not _is_number(got_cached) or abs(got_cached - want) > 1e-6:
                    return "fail", f"{r}行目: 式のキャッシュ値が不一致 (期待 {want} 実際 {got_cached!r})"
            else:
                if not _is_number(got) or abs(got - want) > 1e-6:
                    return "fail", f"{r}行目: 期待 {want} 実際 {got}"
            checked += 1
    note_parts = []
    if excluded:
        note_parts.append(f"数値でない {excluded} 行は対象外")
    if uncached:
        note_parts.append(f"演算対象の式にキャッシュ値が無く検証できない {uncached} 行")
    note = f"（{'・'.join(note_parts)}）" if note_parts else ""
    if checked == 0:
        return "fail", _ZERO_TARGET_REASON + note
    if use_formula:
        return "pass", f"{checked} 行を検証（式・キャッシュ値とも一致）{note}"
    return "pass", f"{checked} 行を検証{note}"


def check_lookup_fill(path: Path, args: dict, header_row: int = 1,
                       use_formula: bool = False) -> tuple:
    """★ W3: header_row は対象シート(target_sheet。★ 挙動変更#2 より前は常に1枚目だった)の
       見出し行。参照表(source_sheet)は常に「列0=キー・列1=値」の物理1行目見出し前提
       （VLookupFromTable ヘルパの仕様どおり・検出対象外）。
       ★ W10f 項目5: use_formula のとき対象シートのキー列を data_only(計算後の値)側から
       読む。VLookupFromTable ヘルパ自体は getString()（LibreOffice が式を評価した文字列）
       でキーを照合するため、前段の式(=A2 等)で埋まったキー列でも転記そのものは正しく
       動く。旧実装は openpyxl の raw 読み（式文字列そのもの）でキーを拾っていたため
       対応表と1件も一致せず、正しく転記されたブックに対して『マスタ表の列順が違う』と
       いう誤診断（濡れ衣）を出していた（Namakoo の純関数レベルの実測をそのまま回帰
       テストにする）。LOOKUP_FILL は check_compute_column と同型の行独立の検証（他行の
       結果に影響しない）なので、キーが読めなかった行は checked から除外する部分採点に
       する（0cf9218 空虚な検証合格の禁止の趣旨 — 読めた行だけを『検証済み』と名乗る）。
       ★ 誤診断の是正: 『対応表のキーが1件も転記されていない』という観測から即座に
       『マスタ表の列順が違う』と断定するのは飛躍だった。まずキーが読めなかった行を
       別集計し（『対象が無い』と『対象はあるが読めない』を混同しない）、キーが全部
       読めたのに1件も一致しない場合でも、本当に列順が違うのか単にキー値そのものが
       対応表と食い違うのかはこのデータだけでは区別できないため、断定をやめて
       『可能性』として並べる。"""
    with BookView(path) as bv:
        if (args["target_sheet"] not in bv.sheetnames
                or args["source_sheet"] not in bv.sheetnames):
            return "fail", "対象/参照シートが無い"
        tws = bv.sheet(args["target_sheet"])
        sws = bv.sheet(args["source_sheet"])
        key_idx = _col_index_by_header(tws, args["key_col"], header_row=header_row)
        tgt_idx = _col_index_by_header(tws, args["target_col"], header_row=header_row)
        if key_idx is None or tgt_idx is None:
            return "fail", "対象シートにキー列/対象列が無い"
        lookup = {}
        r = 2
        while sws.cell(row=r, column=1).value not in (None, ""):
            lookup[sws.cell(row=r, column=1).value] = sws.cell(row=r, column=2).value
            r += 1
        scanned = 0   # ★ 止血1: 対象シートに行が1件も無い(0件)場合と、行はあるが1件も
                      #   対応表に載っていない場合を別のメッセージで区別する。
        checked = 0
        uncached = 0
        r = header_row + 1
        while tws.cell(row=r, column=key_idx).value not in (None, ""):
            scanned += 1
            key_raw = tws.cell(row=r, column=key_idx).value
            key = bv.cell_value(r, key_idx, sheet=args["target_sheet"]) if use_formula else key_raw
            if use_formula and key is None and isinstance(key_raw, str) and key_raw.startswith("="):
                uncached += 1   # ★ W10f: キー列に式はあるがキャッシュ値が無い（『対象が無い』とは別）
                r += 1
                continue
            if key in lookup:
                got = tws.cell(row=r, column=tgt_idx).value
                want = lookup[key]
                if got != want:
                    return "fail", f"{r}行目: キー『{key}』の転記値が不一致 (期待 {want!r} 実際 {got!r})"
                checked += 1
            r += 1
    if scanned == 0:
        return "fail", _ZERO_TARGET_REASON
    if uncached == scanned:
        return "fail", (f"対象シートのキー列に式はあるがキャッシュ値が無く検証できない行が "
                         f"{uncached} 件あり、転記結果を検証できません"
                         f"（LibreOffice を通していない可能性）")
    note = f"（キー列に式はあるがキャッシュ値が無く検証できない {uncached} 行は対象外）" if uncached else ""
    if checked == 0:
        # ★ W10b 項目4a → W10f 項目5: ここまで来た行はキーが読めている（『読めなかった』
        #   場合とは別集計済み）。ただし『本当に列順が違う』のか『キー値そのものが対応表と
        #   一致しない』のかはこのデータだけでは区別できないため、断定せず可能性として
        #   並べる（マスタ表が実際にキー列→値列の順で書かれた具体的な直し方は残す）。
        return "fail", (
            f"対応表『{args['source_sheet']}』に載っているキーが1件も転記されていません。"
            f"マスタ表の列順が違う可能性があります。または『{args['target_sheet']}』シートの"
            f"『{args['key_col']}』列の値が対応表のキーと一致していない可能性があります。"
            f"マスタ表がキー列→値列の順であれば、『{args['source_sheet']}』シートの A 列に"
            f"キー（{args['key_col']} に対応する値）、B 列に値を置いてください{note}"
        )
    return "pass", f"{checked} 行を検証{note}"


def check_aggregate(path: Path, args: dict, header_row: int = 1, use_formula: bool = False) -> tuple:
    """★ W3: header_row は集計元(対象シート。★ 挙動変更#2 より前は常に1枚目だった)の
       見出し行。出力の「集計」シートは SummaryTable ヘルパが毎回新規作成し常に物理1行目が
       見出し（検出対象外・そのまま）。
       ★ W10f 項目1: use_formula のとき分類列/集計列を data_only(計算後の値)側から読む
       （check_compute_column と同型のバグ。前段が式で作った計算列を group_col/value_col
       にした場合、raw 側は式文字列のままで value_col が『数値でない→0扱い』に落ち、
       偽の不一致 fail になっていた。value_col が0扱いされる止血2 の丸めは残すが、
       それは『本当に非数値/空欄』の行専用にする）。
       AGGREGATE は合計という『全行をまたぐ』検証のため、式にキャッシュ値が無く読めない
       行が1件でもあれば期待値そのものが信頼できない。COMPUTE_COLUMN の行独立検証と違い
       部分採点はせず、fail で理由を示して打ち切る（0cf9218 空虚な検証合格の禁止の趣旨）。"""
    with BookView(path) as bv:
        if "集計" not in bv.sheetnames:
            return "fail", "『集計』シートが無い"
        src = bv.sheet(args.get("_target_sheet"))
        gi = _col_index_by_header(src, args["group_col"], header_row=header_row)
        vi = _col_index_by_header(src, args["value_col"], header_row=header_row)
        if gi is None or vi is None:
            return "fail", "分類列/集計列が見つからない"
        expect: dict = {}
        uncached = 0
        r = header_row + 1
        while src.cell(row=r, column=1).value not in (None, ""):
            k_raw = src.cell(row=r, column=gi).value
            v_raw = src.cell(row=r, column=vi).value
            if use_formula:
                k = bv.cell_value(r, gi, sheet=args.get("_target_sheet"))
                v = bv.cell_value(r, vi, sheet=args.get("_target_sheet"))
            else:
                k, v = k_raw, v_raw
            k_uncached = use_formula and k is None and isinstance(k_raw, str) and k_raw.startswith("=")
            v_uncached = use_formula and not _is_number(v) and isinstance(v_raw, str) and v_raw.startswith("=")
            if k_uncached or v_uncached:   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
                uncached += 1
                r += 1
                continue
            v = v if _is_number(v) else 0   # ★ 止血2: 非数値/None は0扱い（クラッシュさせない）
            expect[k] = expect.get(k, 0) + v
            r += 1
        if uncached:
            return "fail", (f"分類列/集計列に式はあるがキャッシュ値が無く検証できない行が "
                             f"{uncached} 件あり、集計を検証できません"
                             f"（LibreOffice を通していない可能性）")
        if not expect:
            return "fail", _ZERO_TARGET_REASON   # ★ 止血1: 集計元データが0件を「合格」にしない
        out = bv.sheet("集計")
        seen = set()
        r = 2
        while True:
            k = out.cell(row=r, column=1).value
            if k in (None, "") or k == "合計":
                break
            v = out.cell(row=r, column=2).value
            v = v if _is_number(v) else 0
            if k not in expect or abs(v - expect[k]) > 1e-6:
                return "fail", f"グループ『{k}』の合計が不一致 (期待 {expect.get(k)} 実際 {v})"
            seen.add(k)
            r += 1
    if seen != set(expect.keys()):
        return "fail", "集計に含まれないグループがある"
    return "pass", f"{len(expect)} グループを検証"


def check_bold(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ W3: "col:" 対象は見出し(header_row)を含めて検証する（codegen の
       StyleBold(oDoc, col, hr0, col, lastRow) が見出しも含めて太字にするため）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        kind, val = args["target"].split(":", 1)
        if kind == "row":
            last_col = _scan_last_col(ws, header_row=header_row)
            row = int(val)
            cells = [ws.cell(row=row, column=c) for c in range(1, last_col + 1)]
            label = f"{row}行目"
        else:
            idx = _col_index_by_header(ws, val, header_row=header_row)
            if idx is None:
                return "fail", f"列『{val}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{val}』"
        if not cells:   # ★ 止血1: 検証対象0件（見出しすら無い空シート等）を合格にしない
            return "fail", _ZERO_TARGET_REASON
        ok = all(c.font and c.font.bold for c in cells)
    if not ok:
        return "fail", f"{label} に太字でないセルがある"
    return "pass", f"{len(cells)} セルが太字"


def check_fill_color(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ W3: "col:" 対象は見出し(header_row)を含めて検証する（codegen が見出しも
       含めて塗るため）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        want_hex = COLOR_MAP[args["color"]].upper()
        kind, val = args["target"].split(":", 1)
        if kind == "row":
            last_col = _scan_last_col(ws, header_row=header_row)
            row = int(val)
            cells = [ws.cell(row=row, column=c) for c in range(1, last_col + 1)]
            label = f"{row}行目"
        else:
            idx = _col_index_by_header(ws, val, header_row=header_row)
            if idx is None:
                return "fail", f"列『{val}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{val}』"
        if not cells:   # ★ 止血1
            return "fail", _ZERO_TARGET_REASON

        def _matches(cell) -> bool:
            if cell.fill is None or not cell.fill.patternType:
                return False
            return str(cell.fill.start_color.rgb).upper().endswith(want_hex)

        ok = all(_matches(c) for c in cells)
    if not ok:
        return "fail", f"{label} に色『{args['color']}』が付いていないセルがある"
    return "pass", f"{len(cells)} セルの背景色を確認"


def check_number_format(path: Path, args: dict, header_row: int = 1) -> tuple:
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        if last < header_row + 1:   # ★ 止血1: データ行0件を合格にしない
            return "fail", _ZERO_TARGET_REASON
        ok = all("#,##0" in (ws.cell(row=r, column=idx).number_format or "")
                 for r in range(header_row + 1, last + 1))
    if not ok:
        return "fail", f"列『{args['col']}』に桁区切り書式が付いていないセルがある"
    return "pass", f"{last - header_row} 行に桁区切り書式を確認"


def check_merge(path: Path, args: dict, header_row: int = 1) -> tuple:
    with BookView(path) as bv:
        ranges = {str(r) for r in bv.sheet(args.get("_target_sheet")).merged_cells.ranges}
    if args["range"] not in ranges:
        return "fail", f"範囲『{args['range']}』が結合されていない"
    return "pass", f"{args['range']} の結合を確認"


def check_chart(path: Path, before_charts: int) -> tuple:
    after = _charts_count(path)
    if after != before_charts + 1:
        return "fail", f"グラフ数が +1 でない（{before_charts} → {after}）"
    return "pass", f"グラフ数 {before_charts} → {after}"


def check_center_align(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ W3: "all"/"col:" とも見出し(header_row)を含めて検証する（codegen の
       AlignCenter/inline テンプレが見出しも含めて中央揃えにするため）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        target = args["target"]
        if target == "all":
            last_row = _scan_last_row(ws, header_row=header_row)
            last_col = _scan_last_col(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=c) for r in range(header_row, last_row + 1)
                     for c in range(1, last_col + 1)]
            label = "表全体"
        else:
            colname = target[4:]
            idx = _col_index_by_header(ws, colname, header_row=header_row)
            if idx is None:
                return "fail", f"列『{colname}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{colname}』"
        if not cells:   # ★ 止血1
            return "fail", _ZERO_TARGET_REASON
        ok = all(c.alignment and c.alignment.horizontal == "center" for c in cells)
    if not ok:
        return "fail", f"{label} に中央揃えでないセルがある"
    return "pass", f"{len(cells)} セルの中央揃えを確認"


_APPEND_TOTAL_FORMULA_RE = re.compile(
    r"^=SUM\([A-Za-z]{1,3}\d+:INDEX\([A-Za-z]{1,3}:[A-Za-z]{1,3},ROW\(\)-1\)\)(\*[\d.]+)?$",
    re.IGNORECASE)


def check_append_total(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ W6: APPEND_TOTAL の事後条件（二層・COMPUTE_COLUMN の use_formula 版と同じ考え方）。
       ①式文字列が期待形『=SUM(<列><行>:INDEX(<列>:<列>,ROW()-1))』（★ B: 挿入耐性式・
       factor!=1 のときは末尾に *factor が付く）②data_only 読みのキャッシュ値が
       「列合計×factor」と一致（浮動小数は丸め許容）。対象列が表の最左端でない場合は、
       その左隣セルに期待ラベルが立っていることも確認する（codegen_dsl の配置と対）。
       ★ B: codegen は LO 方言（INDEX の引数区切りがセミコロン）で書くが、保存後は
       カンマ区切りに自動変換される（実測）ので、ここではその保存後のカンマ形と照合する。

       ★ 合計行の位置は『対象列そのものに "=SUM(" で始まる式が最初に現れた行』として直接
       探す（生読み・型判定はしない）。実測で分かった罠が2つあり、どちらもこの探し方だけ
       で同時に避けられる:
       - COMPUTE_COLUMN の式化（W3 Part3・既定）で対象列自体が "=B2*C2" 型の式で埋まって
         いることがある。生セル値の型（数値かどうか）だけで『データ行/合計行』を分けようと
         すると、データ行の式もキャッシュ値未読の合計行の式も同じ『文字列』に見えて区別
         できない（E2E 実測: 小計列を先に式化してから APPEND_TOTAL するケースで発覚）。
       - 最終データ行の判定に列1(左端)を走査する既存の _scan_last_row は使わない。左端が
         分類列で、かつ対象列がその右隣（＝ラベルの置き場所が偶然その分類列と重なる、例:
         2列だけの帳票）だと、書き込んだラベルの文字列自体が『データ行』として誤って
         数えられ off-by-one になる（実測）。
       "=SUM(" という固有の目印を対象列自身の中だけで探すので、他列の中身にも
       COMPUTE_COLUMN の式の形にも影響されない。

       ★★ 算術恒等の検算（二重計上）を**最初に**行う: 下の①②は期待値を「合計式が
       生成したのと同じ範囲」から作るので、既存の合計を足し込んでいても両方通ってしまう
       （検算が被検算と同じ盲点を使う恒真式）。しかも既存の合計が『式』で入っている表では
       "=SUM(" の初出行がその既存行を掴み、①が「合計の式が期待形でない」と**ユーザーが
       自分で書いた行を責める**誤診断を出す。どちらの誤りも、対象列全体を数値として見る
       この検算が先に立てば正しい行を名指しできる（_nested_total_reason 参照）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        nested = _nested_total_reason(
            _column_block_values(bv, ws, idx, header_row, args.get("_target_sheet")),
            args.get("_target_sheet") or ws.title, idx)
        if nested:
            return "fail", nested
        r = header_row + 1
        while True:
            v = ws.cell(row=r, column=idx).value
            if v is None:
                break
            if isinstance(v, str) and v.replace(" ", "").startswith("=SUM("):
                break
            r += 1
        last = r - 1
        if last < header_row + 1:
            return "fail", _ZERO_TARGET_REASON
        total_row = r
        got_formula = ws.cell(row=total_row, column=idx).value
        got_formula_norm = got_formula.replace(" ", "") if isinstance(got_formula, str) else ""
        if not _APPEND_TOTAL_FORMULA_RE.match(got_formula_norm):
            detail = f"{total_row}行目: 合計の式が期待形(挿入耐性 SUM 型)でない (実際 {got_formula!r})"
            return "fail", detail
        label_ok = True
        want_label = str(args.get("label") or "合計")
        got_label = None
        if idx > 1:
            got_label = ws.cell(row=total_row, column=idx - 1).value
            label_ok = got_label == want_label
        if not label_ok:
            return "fail", f"{total_row}行目: ラベルが期待『{want_label}』と不一致 (実際 {got_label!r})"

        raw_vals = [bv.cell_value(rr, idx, sheet=args.get("_target_sheet")) for rr in range(header_row + 1, last + 1)]
        nums = [v for v in raw_vals if _is_number(v)]
        got_cached = bv.cell_value(total_row, idx, sheet=args.get("_target_sheet"))
    if not nums:
        return "fail", _ZERO_TARGET_REASON
    factor = float(args.get("factor", 1) or 1)
    want_total = sum(nums) * factor
    if not _is_number(got_cached) or abs(got_cached - want_total) > 1e-6:
        return "fail", f"{total_row}行目: 合計のキャッシュ値が不一致 (期待 {want_total} 実際 {got_cached!r})"
    # ★★ 単位I: 主張の範囲を狭める ── **どの行に合計が立っているか**を pass でも言う。
    #   非対称だった: 失敗時は「5行目: …」と行番号を言うのに、成功時は言っていなかった。
    #   ★ 見せる必要が一番あるのは成功したときの方（盲検査定の致命①: 手入力の合計行がある
    #   ブックで、書かれた先が見えないまま ✓ が出ていた）。
    #   ★ 「何を検証したか」でなく「どこを検証したか」まで述べる。件数だけでは範囲を主張できない。
    return "pass", (f"{total_row}行目の合計を検証"
                    f"（{len(nums)} 行分・式・キャッシュ値・ラベルとも一致）")


# --- ★ W9: 検証済みヘルパ4種の事後条件 ---------------------------------------
#   ヘルパ本体(helpers/*.bas)が headerRow を取らないのと同じ理由で、これらのチェッカーは
#   openpyxl の生スキャン(_scan_last_row/_scan_last_col)で実データ範囲を都度見つける
#   （header_row を渡しはするが、ヘルパが物理1行目前提で動く以上、通常は 1 のまま使う想定）。

def check_insert_rows(path: Path, args: dict, header_row: int = 1,
                       source_book: Path | None = None) -> tuple:
    """INSERT_ROWS の事後条件。source_book(適用前のコピー)が渡されれば、
       ①データ最終行が count 増えている ②挿入位置(at)以降の内容が、適用前の対応行から
       count 行分ずれて現れている（シフト）③挿入された行自体が空欄、を突き合わせる。
       ★ source_book が無い（複合計画の途中段等で before が用意できない経路）場合は、
       挿入位置が空欄であることだけを見る warn 判定に落とす（保守的・断定しない）。"""
    at = int(args["at"])
    count = int(args.get("count", 1) or 1)

    if source_book is None or not Path(source_book).exists():
        with BookView(path) as bv:
            ws = bv.sheet(args.get("_target_sheet"))
            last_col = max(_scan_last_col(ws, header_row=header_row), 1)
            row_cells = [ws.cell(row=at, column=c).value for c in range(1, last_col + 1)]
        if all(v in (None, "") for v in row_cells):
            return "warn", "挿入位置が空欄であることのみ確認（適用前ファイルとの突き合わせ無し）"
        return "fail", f"{at}行目が空欄でない（挿入されていない可能性）"

    with BookView(source_book) as bv_before, BookView(path) as bv_after:
        ws_before = bv_before.sheet(args.get("_target_sheet"))
        last_before = _scan_last_row(ws_before, header_row=header_row)
        last_col = _scan_last_col(ws_before, header_row=header_row)
        if last_col < 1 or last_before < header_row + 1:
            return "fail", _ZERO_TARGET_REASON

        ws_after = bv_after.sheet(args.get("_target_sheet"))

        # ★ 挿入は AFTER 側に意図的な空行（挿入行そのもの）を作るため、連続データを前提にする
        #   _scan_last_row を AFTER 側の「最終行」検出には使わない（挿入行で即座に打ち切られて
        #   しまう）。BEFORE の各データ行が「count 行下」に正しく現れているかを直接照合する。
        mismatches = checked = 0
        for r in range(max(at, header_row + 1), last_before + 1):
            for c in range(1, last_col + 1):
                checked += 1
                if ws_before.cell(row=r, column=c).value != ws_after.cell(row=r + count, column=c).value:
                    mismatches += 1
        inserted_ok = all(ws_after.cell(row=r, column=c).value in (None, "")
                           for r in range(at, at + count) for c in range(1, last_col + 1))
        # 期待される最終データ行のさらに1行下が空欄であること（想定外の余剰データが無いこと）。
        expect_last_after = last_before + count
        extra_row_empty = all(ws_after.cell(row=expect_last_after + 1, column=c).value in (None, "")
                               for c in range(1, last_col + 1))

    if checked == 0:
        return "fail", _ZERO_TARGET_REASON
    if mismatches:
        return "fail", f"挿入位置以降のシフトが一致しない（不一致 {mismatches}/{checked} セル）"
    if not inserted_ok:
        return "fail", f"挿入された{count}行が空欄でない"
    if not extra_row_empty:
        return "fail", f"期待より多くの行にデータがある（{expect_last_after}行より下に想定外のデータ）"
    return "pass", f"{count}行挿入・{checked}セル分のシフトを確認（挿入行は空欄・行数も+{count}ぴったり）"


def check_draw_borders(path: Path, args: dict, header_row: int = 1) -> tuple:
    """DRAW_BORDERS の事後条件。使用範囲の全セルに上下左右の罫線属性が付いているかを見る
       （DrawTableBorders ヘルパは格子罫線を範囲全体に一括で付けるため、1セルでも
       欠けていれば ヘルパが動いていないか範囲がずれている）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_row = _scan_last_row(ws, header_row=header_row)
        last_col = _scan_last_col(ws, header_row=header_row)
        if last_col < 1 or last_row < header_row:
            return "fail", _ZERO_TARGET_REASON
        cells = [ws.cell(row=r, column=c) for r in range(header_row, last_row + 1)
                 for c in range(1, last_col + 1)]
    if not cells:
        return "fail", _ZERO_TARGET_REASON

    def _has_border(cell) -> bool:
        bd = cell.border
        return bool(bd and bd.top and bd.top.style and bd.bottom and bd.bottom.style
                     and bd.left and bd.left.style and bd.right and bd.right.style)

    ok = all(_has_border(c) for c in cells)
    if not ok:
        return "fail", "使用範囲の一部に罫線が無い"
    return "pass", f"{len(cells)} セルの罫線を確認"


def check_autofit(path: Path, args: dict, header_row: int = 1,
                   source_book: Path | None = None) -> tuple:
    """AUTOFIT の事後条件。source_book が渡されれば、使用中の各列の幅が適用前後で
       変化した列数を数える（1列でも変化していれば pass）。source_book が無ければ、
       AFTER 側で幅が明示的に設定されている列があることだけを見る warn 判定に落とす。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_col = _scan_last_col(ws, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON
        after_widths = {}
        for c in range(1, last_col + 1):
            letter = get_column_letter(c)
            dim = ws.column_dimensions.get(letter)
            after_widths[letter] = dim.width if dim and dim.width else None

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            before_widths = {}
            for c in range(1, last_col + 1):
                letter = get_column_letter(c)
                dim = ws_b.column_dimensions.get(letter)
                before_widths[letter] = dim.width if dim and dim.width else None
        changed = sum(1 for k in after_widths if after_widths[k] != before_widths.get(k))
        if changed == 0:
            return "fail", "列幅が変化していない"
        return "pass", f"{changed}/{len(after_widths)} 列の幅が変化"

    set_count = sum(1 for v in after_widths.values() if v)
    if set_count == 0:
        return "fail", "列幅が設定されていない（AutoFitColumns が効いていない可能性）"
    return "warn", f"{set_count} 列に幅が設定されていることのみ確認（適用前ファイルとの突き合わせ無し）"


def check_pivot(path: Path, args: dict, header_row: int = 1) -> tuple:
    """PIVOT の事後条件。①『ピボット』シートが存在 ②xlsx zip 内に本物の DataPilot
       (xl/pivotTables/) が実在するか、を見る（W8b の忠実度ゲートと同じ zip 直読み技法・
       LO を再起動しない）。★ W9 項目4: DataPilot は開き直すたび再描画で値キャッシュが
       変わりうる既知の癖があるため、集計値そのものの照合（check_aggregate 相当）はせず
       構造の存在確認に留める（PIVOT_CAVEAT として案内も添える）。"""
    with BookView(path, read_only=True) as bv:
        has_sheet = "ピボット" in bv.sheetnames
    if not has_sheet:
        return "fail", "『ピボット』シートが無い"
    try:
        with zipfile.ZipFile(path) as z:
            names = [n.lower() for n in z.namelist()]
    except Exception as e:
        return "fail", f"xlsx を zip として読めない: {e}"
    has_pivot_table = any(n.startswith("xl/pivottables/") for n in names)
    if not has_pivot_table:
        return "fail", "『ピボット』シートはあるが DataPilot(ピボットテーブル)の実体が無い"
    return "pass", f"『ピボット』シートと DataPilot を確認（{PIVOT_CAVEAT}）"


def check_set_column_value(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ 致命3(W10e): SET_COLUMN_VALUE の事後条件。対象列のデータ行が全部、機械抽出した
       定数値(args["value"])と一致するかを見る（型を問わず文字列表現で比較 — codegen は
       setString で書くため、読み戻しも文字列として揃える）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        if last < header_row + 1:   # ★ 止血1: データ行0件を合格にしない
            return "fail", _ZERO_TARGET_REASON
        value = args["value"]
        vals = [ws.cell(row=r, column=idx).value for r in range(header_row + 1, last + 1)]
    if not all(str(v) == str(value) for v in vals):
        return "fail", f"列『{args['col']}』に『{value}』でないセルがある"
    return "pass", f"{len(vals)} 行を『{value}』に統一"


def _extract_predicate(cmp: str, threshold):
    """EXTRACT の判定を Basic 側(ExtractRows/helpers/AiLineHelpers.bas)とは別実装で
       もう一度書く（同じ勘定を2箇所が違う実装で書いて一致を見る・独立測定）。
       ★ M2（2026-08-21・宣言済みの挙動変更）: 意味論を tests/test_predicate_truth_table.py
       の手書きの表に合わせた。① eq は両辺が数値なら**許容誤差 1e-6**（浮動小数の完全一致は
       表計算の実データで偽陰性になる）② contains は**文字列セルのみ**（数値 140000 を
       黙って "140000" に文字列化して『40 を含む』としない ── 型の保存の哲学）。
       単一ブック EXTRACT（check_extract）の挙動もこの線に揃う。"""
    def _match(cell_value) -> bool:
        if cmp == "contains":
            return (isinstance(cell_value, str) and threshold is not None
                    and str(threshold) in cell_value)
        if cmp == "eq":
            if isinstance(threshold, (int, float)) and not isinstance(threshold, bool):
                return (_is_number(cell_value)
                        and abs(float(cell_value) - float(threshold)) <= 1e-6)
            return str(cell_value) == str(threshold)
        if not _is_number(cell_value):   # gte/lte/gt/lt は数値比較のみ
            return False
        v, t = float(cell_value), float(threshold)
        if cmp == "gte":
            return v >= t
        if cmp == "lte":
            return v <= t
        if cmp == "gt":
            return v > t
        if cmp == "lt":
            return v < t
        return False
    return _match


def check_extract(path: Path, args: dict, header_row: int = 1,
                   source_book: Path | None = None) -> tuple:
    """EXTRACT の事後条件（コミット 2edcb08「EXTRACT op」―「番人が入場料を徴収した初の op」）。
       ①行数一致 ②値と型の保存 ③両側の網羅（多い/少ないの両方を落とす）を、1つの位置対応
       比較で同時に見る: 元シートを独立に走査して『条件に一致する行』を上から順に集めた
       ものが expected。出力の各行が expected と完全に同じ順・同じ内容で並んでいれば、
       行数も一致し(①)・値も型も保存され(②。取り違えれば != が拾う ── 数値 59400 と
       文字列 '59400' は Python の != では等しくない＝昨夜の全セル文字列化バグを直接殺す)・
       多く含めても少なく埋めても位置がずれて不一致になる(③)。
       ④ 元シートが無変更（source_book が渡された時だけ突き合わせられる・読むだけの op。
       source_book が無ければ①②③だけを見る＝INSERT_ROWS/AUTOFIT と同じ劣化フォールバック）。"""
    dst_name = args.get("_new_sheet")
    col_name = args.get("col")
    cmp = args.get("cmp")
    threshold = args.get("value")
    if not dst_name:
        return "fail", "出力シート名が決まっていません（verify_dsl_args を経由していない可能性）"
    match = _extract_predicate(cmp, threshold)

    with BookView(path) as bv:
        src = bv.sheet(args.get("_target_sheet"))
        src_name = src.title
        if dst_name not in bv.sheetnames:
            return "fail", f"出力シート『{dst_name}』が作られていません"
        col_idx = _col_index_by_header(src, col_name, header_row=header_row)
        if col_idx is None:
            return "fail", f"対象列『{col_name}』が元シート『{src_name}』に見つかりません"
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        total = 0
        expected_rows = []
        r = header_row + 1
        while src.cell(row=r, column=1).value not in (None, ""):
            total += 1
            if match(src.cell(row=r, column=col_idx).value):
                expected_rows.append([src.cell(row=r, column=c).value for c in range(1, last_col + 1)])
            r += 1
        if total == 0:
            return "fail", _ZERO_TARGET_REASON

        out = bv.sheet(dst_name)
        out_rows = []
        r = 2   # 出力は ExtractRows の仕様どおり常に物理1行目が見出し
        while out.cell(row=r, column=1).value not in (None, ""):
            out_rows.append([out.cell(row=r, column=c).value for c in range(1, last_col + 1)])
            r += 1

    denom = f"{total}行中{len(expected_rows)}行が一致"
    if len(out_rows) != len(expected_rows):
        return "fail", f"{denom} → 出力は{len(out_rows)}行（行数が期待と不一致）"
    for i, (want, got) in enumerate(zip(expected_rows, out_rows), start=1):
        if want == got:
            continue
        for c, (wv, gv) in enumerate(zip(want, got), start=1):
            if wv != gv:
                letter = get_column_letter(c)
                return "fail", (
                    f"{denom} → 出力{i}行目 {letter}列が元と不一致"
                    f"（元 {wv!r}（{type(wv).__name__}） 出力 {gv!r}（{type(gv).__name__}））"
                )

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            src_before = bv_before.sheet(args.get("_target_sheet"))
            src_after = bv_after.sheet(args.get("_target_sheet"))
            last_row_before = _scan_last_row(src_before, header_row=header_row)
            mismatches = sum(
                1 for r in range(header_row, last_row_before + 1) for c in range(1, last_col + 1)
                if src_before.cell(row=r, column=c).value != src_after.cell(row=r, column=c).value)
        if mismatches:
            return "fail", (f"{denom} を抽出しましたが、元シート『{src_name}』が {mismatches} セル"
                             "変更されています（読むだけのはず）")
        return "pass", f"{denom} → {len(expected_rows)}行を抽出（値・型とも保存・元シート無変更）"

    return "pass", f"{denom} → {len(expected_rows)}行を抽出（値・型とも保存。元シートとの突き合わせ無し）"


POSTCONDITIONS = {
    "SORT": check_sort, "COMPUTE_COLUMN": check_compute_column,
    "LOOKUP_FILL": check_lookup_fill, "AGGREGATE": check_aggregate,
    "BOLD": check_bold, "FILL_COLOR": check_fill_color,
    "NUMBER_FORMAT": check_number_format, "MERGE": check_merge,
    "CENTER_ALIGN": check_center_align, "APPEND_TOTAL": check_append_total,
    # ★ W9: 検証済みヘルパ4種。
    "INSERT_ROWS": check_insert_rows, "DRAW_BORDERS": check_draw_borders,
    "AUTOFIT": check_autofit, "PIVOT": check_pivot,
    # ★ 致命3(W10e):
    "SET_COLUMN_VALUE": check_set_column_value,
    "EXTRACT": check_extract,
}


def run_postcondition(op: str, out_book: Path, resolved_args: dict, before_charts: int = 0,
                       header_row: int = 1, use_formula: bool = False,
                       source_book: Path | None = None) -> tuple:
    """⑥ op 別事後条件。(status, reason)。status ∈ {"pass","warn","fail","error"}。
       CHART だけ before_charts と比較する専用の形。
       ★ W3: header_row（1起点、省略時1）を全チェッカーに一貫して渡す（『三層全部が
       同じ見出し推定を使う』の事後条件側）。
       ★ W10f 項目1: use_formula は元は COMPUTE_COLUMN 専用（W3 Part3）だったが、
       operand を式ビューから読む同型バグが SORT/AGGREGATE にもあったため、その2つにも
       渡すよう広げた（同じ理由で check_compute_column_single_factor は
       check_compute_column からの委譲で既に受け取っている）。
       ★ W10f 項目5: LOOKUP_FILL のキー列にも同型バグ(対象シートのキー列を式ビューから
       読む)があったため渡すよう広げた。
       ★ W9: source_book（適用前のコピー・無ければ None）は INSERT_ROWS/AUTOFIT だけが使う
       （before/after の突き合わせが要る2op・他は無視される安全なキーワード引数）。
       ★ 止血2: チェッカー内で予期しない例外が起きても生の Python トレースバックを
       出さない。ここで必ず捕まえて "error" ステータス + 要約1行に変換する
       （C②の教訓: 事後条件チェッカー自身のクラッシュがユーザーに未捕捉のまま漏れていた）。"""
    try:
        if op == "CHART":
            return check_chart(out_book, before_charts)
        fn = POSTCONDITIONS.get(op)
        if fn is None:
            return "fail", f"未対応の op: {op}"
        if op == "COMPUTE_COLUMN":
            return fn(out_book, resolved_args, header_row, use_formula)
        if op == "SORT":
            # ★ 算術恒等の検算: 合計行が最下行から動いたかは before が無いと測れない。
            return fn(out_book, resolved_args, header_row, use_formula=use_formula,
                       source_book=source_book)
        if op in ("AGGREGATE", "LOOKUP_FILL"):
            return fn(out_book, resolved_args, header_row, use_formula=use_formula)
        if op in ("INSERT_ROWS", "AUTOFIT", "EXTRACT"):
            return fn(out_book, resolved_args, header_row, source_book=source_book)
        return fn(out_book, resolved_args, header_row)
    except Exception as e:
        return "error", f"事後条件の検証に失敗: {type(e).__name__}: {e}"


# ---------------------------------------------------------------------------
# basrun 経由の適用
# ---------------------------------------------------------------------------

def _timeout_error_message(timeout: float) -> str:
    return f"実行時エラー: マクロが {timeout:.0f} 秒で終了しない（無限ループの可能性）"


def short_error_summary(err: str) -> str:
    """★ M2a: obasync 由来の生 Python トレースバックを端末にそのまま出さないための
       整形。最終行（例外名+メッセージ）だけを取り出す。全文は履歴 jsonl 側に残す
       （修復ループへは従来どおり err の全文を渡す＝モデルへの情報は減らさない）。"""
    lines = [ln for ln in (err or "").splitlines() if ln.strip()]
    return lines[-1].strip() if lines else "(詳細不明)"


def _kill_process_tree(pid: int) -> None:
    """PID 指定でプロセスツリーを kill する。
       ★ taskkill /IM（名前一括）は使わない — 無関係な他プロセスも巻き込む
       （2026-08 の Senior MCP 事故の教訓）。必ず特定した PID だけを狙う。"""
    if os.name == "nt":
        subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"],
                        capture_output=True)
    else:
        import signal
        try:
            os.killpg(os.getpgid(pid), signal.SIGKILL)
        except Exception:
            pass


def basrun_apply(book: Path, code: str, workdir: Path, helper_files=(),
                  timeout: float | None = DEFAULT_APPLY_TIMEOUT) -> tuple:
    """生成コードを book に適用する。(ok, error_or_None, raw_output)。
       helper_files があれば同じ src に置く＝同じライブラリに同期され、Gen から呼べる。

    ★ M1: timeout（既定 180 秒、None/0 で無効）を basrun 側の `apply --timeout` に
       転送する。basrun は自分の内部タイムアウトで、ハングした接続先の LibreOffice
       だけを stop_office() で終了させる（taskkill 一括はしない、既存機構）。
       ここではさらに外側の安全網として、basrun 自身が固まった場合に備えて
       ゆとり(+30秒)を持たせた outer timeout を持ち、それも超えたら basrun.py の
       プロセスを PID 指定で kill する。どちらの経路で落ちても、修復ループには
       同じ「実行時エラー: マクロが N 秒で終了しない」で渡す（正直な分類）。"""
    src = workdir / "src"
    if src.exists():
        shutil.rmtree(src)
    src.mkdir(parents=True)
    for hf in helper_files:
        shutil.copy2(hf, src / hf.name)
    (src / "Gen.bas").write_text(code, encoding="utf-8")

    cmd = [sys.executable, str(basrun_path()), "apply", str(book), str(src), "AiLine", "Gen.Run"]
    if timeout:
        cmd += ["--timeout", str(timeout)]
    outer_timeout = (timeout + 30) if timeout else None

    proc = subprocess.Popen(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        text=True, encoding="utf-8", errors="replace")  # ★ cp932 事故を避ける
    try:
        out, err_out = proc.communicate(timeout=outer_timeout)
    except subprocess.TimeoutExpired:
        # ★ basrun 自身の内部タイムアウト＋stop_office() が働かなかった場合の安全網。
        _kill_process_tree(proc.pid)
        try:
            out, err_out = proc.communicate(timeout=5)
        except Exception:
            out, err_out = "", ""
        raw = (out or "") + "\n" + (err_out or "")
        return False, _timeout_error_message(timeout), raw

    raw = (out or "") + "\n" + (err_out or "")
    if proc.returncode != 0:
        # basrun 自身の内部タイムアウトで落ちた場合も、同じ「実行時エラー」分類に正規化する。
        if timeout and "秒応答しなかった" in raw:
            return False, _timeout_error_message(timeout), raw
        return False, raw.strip()[-800:], raw
    return True, None, raw


# ★ 正規化パス専用の空マクロ。何もしない（Call すら書かない）が、basrun_apply の
#   `doc.store()` を一度通すことで LibreOffice 側の初回保存の実体化を先に済ませる。
NOOP_MACRO = "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\nEnd Sub\n"
#  ★ W3: 正規化パスは NOOP_MACRO でなく _structdump_macro() を使う（同じ LO 往復で
#     構造読み取りを二役させる）。文書へは何も書かない点は NOOP_MACRO と同じ。


def _stop_office() -> None:
    """basrun.py stop を呼んで LibreOffice を落とす（M2c: normalize_book の自動リトライ用）。
       ★ ここでの失敗は無視する（次の apply がどのみち再起動を試みる。taskkill 一括はしない
       既存機構の外側なので、ここでは basrun 自身の stop に委譲するだけ）。"""
    try:
        subprocess.run([sys.executable, str(basrun_path()), "stop"],
                        capture_output=True, encoding="utf-8", errors="replace")
    except Exception:
        pass


def normalize_book(book: Path, workdir: Path,
                    timeout: float | None = DEFAULT_APPLY_TIMEOUT) -> Path:
    """コピーを LibreOffice で一度（空マクロで）開いて保存する ＝ P0 の正規化パス。

    LibreOffice は openpyxl 製（＝ LO で保存されたことがない）ブックを初回保存する際、
    行高（時に列幅）を実体化する。before スナップショットをこの実体化の**前**に取ると、
    その副作用が「マクロが変化させた」と誤検出される（no-op ガードの偽陽性・製品の心臓）。
    先にこの正規化を一度済ませておけば、以降の before/after 比較はマクロの実際の効果
    だけを見る。コストは LO 往復 1 回（数秒）— 正しさ優先で受け入れる。
    参考: ailine-ts の tests/e/_harness.ts normalizeThroughLibreOffice が同じ手当てを
    テスト側で先に実装していた（挙動の参考。製品経路に入れるのはこちらが初）。

    ★ M2c: 監査2回で2回再現した既知の摩擦（RuntimeException: Could not create system
    bitmap! 等、LibreOffice 側の一時的な描画/接続不調）への低リスク対処。1回失敗しても
    即座に落とさず、stop（LibreOffice を落とす）→ 再起動を挟んで1回だけ自動リトライする。
    それでも失敗したら現行どおりのエラーで落ちる（無限リトライはしない）。"""
    normalized = workdir / ("normalized" + book.suffix)
    shutil.copy2(book, normalized)
    dump_path = workdir / STRUCTDUMP_FILENAME
    struct_code = _structdump_macro(dump_path)
    ok, err, _ = basrun_apply(normalized, struct_code, workdir, timeout=timeout)
    if not ok:
        _stop_office()
        shutil.copy2(book, normalized)   # 中途半端な保存状態を残さず作り直す
        ok, err, _ = basrun_apply(normalized, struct_code, workdir, timeout=timeout)
    if not ok:
        sys.exit(f"正規化パスに失敗した（LibreOffice で開けなかった）: {err}")
    return normalized


def success_message(result: dict) -> str | None:
    """★ の注意書きは『変化を検出して適用が成功した』ときだけ出す。
       失敗(exit 1)や --dry（何も適用していない）で出すのは不誠実（P1）。"""
    if result.get("ok") and not result.get("dry"):
        return "★ 変化は検出したが『正しいか』は上の差分を見て判断してください（no-op ガードは正しさを保証しない）。"
    return None


# ---------------------------------------------------------------------------
# --inplace バックアップ + restore（M2a）
#
# ★ --inplace は原本を上書きするので、上書き前に必ず ~/.ailine/backups/ へコピーする。
#   バックアップに失敗したら --inplace 自体を中止する（安全側。原本は無変更のまま）。
#   restore は復元前の現状も退避してから上書きする＝復元自体も取り消せる。
# ---------------------------------------------------------------------------

# ★ W8b: 秒精度(旧・6桁)・マイクロ秒精度(新・12桁)・衝突回避の "-N" 連番つき、の
#   いずれも受け付ける（後方互換）。
_BACKUP_TS_RE = re.compile(r"^\d{8}T\d{6}(?:\d{6})?Z(?:-\d+)?$")
_BACKUP_TS_SEQ_RE = re.compile(r"-(\d+)$")

# ★ W11: undo が「復元前の現状」を退避する棚。名前空間ディレクトリの**さらに下**に置く
#   （BACKUP_DIR/<ns>/undo/）。list_backups は名前空間ディレクトリ直下のファイルしか
#   見ないので、ここに置いたものは自動的に「遡りの履歴」から外れる。
#   なぜ分けるか（実測した不具合）: 退避を実編集の世代と同じ棚に積むと、undo のたびに
#   履歴が伸び、しかも最も古い状態に着いた後は「現在の中身と同じ退避」が最新世代として
#   並ぶため、次の undo が『その1つ内側』＝直前に打ち消したはずの新しい状態を釣り上げて
#   いた（N 回編集して N+2 回 undo すると v1 が復活する）。退避を消すのでなく、
#   **数える場所と遡りの参照点から外す**ことで直す（安全網は減らさない）。
UNDO_SHELF_DIRNAME = "undo"


class NoOlderBackupError(Exception):
    """★ W11: 最も古い世代に着いていて、これ以上遡れない（＝undo の端）。
       「バックアップが1つも無い」(FileNotFoundError)とは別物なので型で分ける。"""


def _utc_ts() -> str:
    """ファイル名に使える UTC タイムスタンプ（例: 20260814T120000123456Z）。
       ★ W8b: マイクロ秒まで含める（実測: それでも Windows の壁時計分解能は粗く
       20万回の連続呼び出しで56通りしか値が変わらない。衝突自体は make_backup 側の
       "-N" 連番フォールバックで最終的に防ぐ・ここは表示上の精度向上に留める）。"""
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%f") + "Z"


def _ts_sort_key(ts: str):
    """ts 文字列（秒/マイクロ秒精度・衝突回避の "-N" 連番つき）を新しい順ソート用の
       比較可能な値 (datetime, 連番) にする。壊れた形式は最古扱い。"""
    body = ts
    seq = 0
    m = _BACKUP_TS_SEQ_RE.search(body)
    if m:
        seq = int(m.group(1))
        body = body[:m.start()]
    if body.endswith("Z"):
        body = body[:-1]
    for fmt in ("%Y%m%dT%H%M%S%f", "%Y%m%dT%H%M%S"):
        try:
            return (datetime.strptime(body, fmt), seq)
        except ValueError:
            continue
    return (datetime.min, seq)


def _backup_namespace(book: Path) -> str:
    """book が置かれているフォルダごとのバックアップ名前空間（sha1 の先頭8桁）。
       ★ W8b 項目3: 同名ファイルが別フォルダにある場合（A\\見積.xlsx / B\\見積.xlsx）、
       旧・フラットな BACKUP_DIR では stem+suffix だけで一致させていたため取り違え
       （undo 混線）が起きえた。フォルダの絶対パスでハッシュ化した名前空間ディレクトリへ
       分離する（book が相対パスで渡されても resolve() してから使う）。"""
    return hashlib.sha1(str(Path(book).resolve().parent).encode("utf-8")).hexdigest()[:8]


def undo_shelf_dir(book: Path) -> Path:
    """★ W11: undo が取る「復元前の現状」の退避先（遡りの履歴には数えない棚）。"""
    return BACKUP_DIR / _backup_namespace(book) / UNDO_SHELF_DIRNAME


def backup_path_for(book: Path, ts: str | None = None, shelf: bool = False) -> Path:
    ts = ts or _utc_ts()
    base = undo_shelf_dir(book) if shelf else BACKUP_DIR / _backup_namespace(book)
    return base / f"{book.stem}.{ts}{book.suffix}"


def prune_backups(book: Path, keep: int = DEFAULT_KEEP_BACKUPS, shelf: bool = False) -> list:
    """★ M2c: book の世代のうち keep 件を超える古いもの（list_backups は新しい順）を削除する。
       戻り値は削除したパスのリスト。keep < 0 は「無制限（削除しない）」扱い。
       ★ W11: shelf=True のときは undo の退避棚を対象にする（棚も同じ上限で剪定する）。"""
    if keep < 0:
        return []
    backups = list_undo_shelf(book) if shelf else list_backups(book)
    stale = backups[keep:]
    deleted = []
    for p in stale:
        try:
            p.unlink()
            deleted.append(p)
        except OSError:
            pass
    return deleted


def make_backup(book: Path, keep: int = DEFAULT_KEEP_BACKUPS, shelf: bool = False) -> Path:
    """book のバックアップを ~/.ailine/backups/<名前空間>/ に作る。戻り値はバックアップ先。
       ★ 失敗したら例外を投げる（呼び出し側が --inplace 中止の判断に使う）。
       ★ M2c: 新しいバックアップを作った後、keep 世代を超えた古いものを剪定する
       （既定 DEFAULT_KEEP_BACKUPS=10。無制限にすると個人開発機のディスクを静かに食う）。
       ★ W8b 項目3: 新規のバックアップは必ず名前空間ディレクトリへ書く（フラット領域には
       もう書かない・読み取り専用の後方互換は list_backups 側で担う）。
       ★ W8b: Windows の壁時計分解能は実測で粗く（20万回の連続呼び出しで56通りしか
       値が変わらない）、restore_backup が「復元前の現状」を退避する高速な連続呼び出し
       等でファイル名が衝突しうる。衝突したら "-N" 連番を足して必ず別ファイルにする
       （既存の世代を上書きで消さない・回帰テストで自己顕在化した実際の不具合の修正）。
       ★ W11: shelf=True は undo が「復元前の現状」を退避するときの置き場（BACKUP_DIR/
       <ns>/undo/）。遡りの履歴には数えないが、ファイルとしては同じ作法で残す
       （undo を可逆にする性質＝退避そのものは減らさない）。"""
    ts = _utc_ts()
    dst = backup_path_for(book, ts=ts, shelf=shelf)
    dst.parent.mkdir(parents=True, exist_ok=True)
    n = 2
    while dst.exists():
        dst = backup_path_for(book, ts=f"{ts}-{n}", shelf=shelf)
        n += 1
    shutil.copy2(book, dst)
    prune_backups(book, keep=keep, shelf=shelf)
    return dst


def _parse_backup_name(name: str, stem: str, suffix: str) -> str | None:
    """バックアップのファイル名が `<stem>.<ts><suffix>` の形かを見て、ts を返す
       （形が違えば None）。"""
    prefix = stem + "."
    if not (name.startswith(prefix) and name.endswith(suffix)):
        return None
    ts = name[len(prefix):len(name) - len(suffix)]
    return ts if _BACKUP_TS_RE.match(ts) else None


def _gather_backups(directory: Path, stem: str, suffix: str) -> list:
    """directory 直下（再帰しない）の `<stem>.<ts><suffix>` 形のファイルを (ts, Path) で集める。
       iterdir で拾うのはファイルだけ＝サブディレクトリ（名前空間・undo の退避棚）自体を
       誤ってバックアップと数えない。並べ替えは呼び出し側の責任。"""
    found = []
    if directory.is_dir():
        for p in directory.iterdir():
            if not p.is_file():
                continue
            ts = _parse_backup_name(p.name, stem, suffix)
            if ts is not None:
                found.append((ts, p))
    return found


def _sorted_newest_first(found: list) -> list:
    # ★ W8b: 秒精度(旧)とマイクロ秒精度(新)が混在しうるため、生文字列の辞書順ではなく
    #   _ts_sort_key() でパースした実時刻順に並べる（桁数違いの文字列比較は時刻順にならない）。
    found.sort(key=lambda pair: _ts_sort_key(pair[0]), reverse=True)
    return [p for _ts, p in found]


def list_backups(book: Path) -> list:
    """book に対応するバックアップ（＝遡れる実編集の世代）を新しい順で返す。
       ★ W8b 項目3: 名前空間ディレクトリ BACKUP_DIR/<ns>/ を主として見る。
       旧フラット領域（BACKUP_DIR 直下・名前空間分離前の名残）も読み取り専用互換で
       あわせて見る（新規はもう書かない）。
       ★ W11: undo の退避棚 BACKUP_DIR/<ns>/undo/ は**含めない**（直下しか見ないので
       自動的に外れる）。棚を見たいときは list_undo_shelf() を使う。"""
    stem, suffix = book.stem, book.suffix
    found = _gather_backups(BACKUP_DIR / _backup_namespace(book), stem, suffix)
    found += _gather_backups(BACKUP_DIR, stem, suffix)
    return _sorted_newest_first(found)


def list_undo_shelf(book: Path) -> list:
    """★ W11: undo が取った「復元前の現状」の退避を新しい順で返す（遡りには数えない）。"""
    return _sorted_newest_first(_gather_backups(undo_shelf_dir(book), book.stem, book.suffix))


def _undo_position(book: Path, backups: list) -> int | None:
    """book の現在の中身が世代列(新しい順)のどこにいるかの添字。どこにも無ければ None
       （＝実編集の直後で、まだ 1 段も遡っていない）。"""
    try:
        current = book.read_bytes()
    except OSError:
        return None
    for i, p in enumerate(backups):
        try:
            if p.read_bytes() == current:
                return i
        except OSError:
            continue
    return None


def undo_steps_left(book: Path, backups: list | None = None) -> int:
    """★ W11: 今の位置から**あと何回 undo できるか**。バックアップの総数ではない
       （総数を数えていたので、undo が退避を積むたびに『あと N 回』が増えていた）。
       現在地が世代列の i 番目なら、残りは i より古い世代の数 = len-1-i。
       現在地がどこにも無い（＝実編集の直後）なら全世代を遡れるので len。"""
    backups = list_backups(book) if backups is None else backups
    if not book.exists():
        return len(backups)
    i = _undo_position(book, backups)
    return len(backups) if i is None else len(backups) - 1 - i


def restore_backup(book: Path) -> Path:
    """book を「1つ前の世代」から復元する。戻り値は使ったバックアップの Path。
       バックアップが1つも無ければ例外を投げる。
       ★ W8b-2: 連続 undo が実編集履歴を1段ずつ正しく遡れるようにする（B2 事故バッテリ
       ①で実測: 素朴に『常に最新のバックアップを使う』実装だと、undo が作る『復元前の
       現状の退避』が次回の undo で『最新』として選ばれてしまい、2回目の undo が
       1段先(=1回目の undo を打ち消す)に進んでしまっていた＝多段 undo が成立しない）。
       book の現在の中身と一致するバックアップがあれば、そのすぐ内側(より古い方)を
       復元先にする（＝現在地がバックアップ履歴のどこかに『既にいる』とみなし、そこから
       もう1段遡る）。一致するものが無ければ（＝直前の実編集の直後・通常の最初の undo）
       最新のバックアップを復元する。復元前の現状は必ず退避する（undo 自体も可逆）。
       ★ W11: 最も古い世代に着いていたら NoOlderBackupError を投げて**止まる**
       （旧実装は同じものを復元して『✓ 復元した』と名乗っていた＝何もしていないのに成功）。
       ★ W11: 退避先は undo の棚（undo_shelf_dir）で、遡りの履歴には混ぜない
       （混ぜていたので、端に着いた後の undo が退避を最新世代として釣り上げていた）。"""
    backups = list_backups(book)   # 新しい順
    if not backups:
        raise FileNotFoundError(f"{book.name} のバックアップが無い")

    target = backups[0]
    if book.exists():
        i = _undo_position(book, backups)
        if i is not None:
            if i + 1 >= len(backups):
                raise NoOlderBackupError(
                    f"{book.name} をこれ以上は戻せません（最も古い状態です）")
            target = backups[i + 1]
        make_backup(book, shelf=True)   # 復元前の現状も退避＝restore 自体も可逆にする
    shutil.copy2(target, book)
    return target


def cmd_restore(a: argparse.Namespace) -> int:
    book = Path(a.book).resolve()
    if a.list:
        backups = list_backups(book)
        for ln in render_backup_list(book.name, backups):
            print(ln)
        return 0
    try:
        used = restore_backup(book)
    except (FileNotFoundError, NoOlderBackupError) as e:
        print(f"× {e}")
        return 1
    print(render_restore_done(book.name, used.name))
    return 0


def cmd_undo(a: argparse.Namespace) -> int:
    """★ W8b 項目5: `restore` の昇格。真実の源はバックアップファイル自体
       （history.jsonl には依存しない＝history が壊れていても undo できる）。
       名前空間対応(item3)は list_backups/restore_backup 経由でそのまま効く。
       復元後、まだ戻せる回数（★ W11: バックアップの総数ではなく**あと何段遡れるか**）を
       添える。端（最も古い状態）に着いたら復元せずに非零で止まる。
       ★ M2: フォルダ抽出（run <フォルダ>）は原本を読むだけでバックアップも履歴も作らない
       ── 戻す対象が構造的に存在しないので、フォルダには「無い」と正直に言って止まる。"""
    book = Path(a.book).resolve()
    if book.is_dir():
        print(f"× フォルダに対する undo はありません"
              f"（原本は読んでいません。消せるのは出力ブックだけです）: {book}")
        return 1
    if a.list:
        backups = list_backups(book)
        for ln in render_backup_list(book.name, backups, shelved=len(list_undo_shelf(book))):
            print(ln)
        return 0
    try:
        used = restore_backup(book)
    except (FileNotFoundError, NoOlderBackupError) as e:
        print(f"× {e}")
        return 1
    print(render_restore_done(book.name, used.name, remaining=undo_steps_left(book)))
    return 0


# ---------------------------------------------------------------------------
# ★ W8b: 安全器官（既定の反転は次コミット。今回は原本を直接書く危険を減らす下ごしらえ）
# ---------------------------------------------------------------------------

def check_excel_lock(book: Path) -> str | None:
    """book が Excel 等で開かれている兆候を機械的に見る。開かれていそうなら理由の
       文字列（人間可読）、そうでなければ None。
       ★ W8b 項目2: ①同フォルダの Excel ロックファイル(~$<name>) の存在
       ②open(book, 'r+b') を試みて PermissionError になるか、の2つを見る（保守的
       ＝どちらかに該当したら『開かれている可能性』として止める。誤検知より、
       書き込み中の文書を壊さない方を優先する）。run の最初（LO 起動・翻訳より前）
       に呼ぶ（--copy 時も含め常に同じ判定にする＝整合性の観点で経路を分けない）。"""
    lock_file = book.parent / f"~${book.name}"
    if lock_file.exists():
        return f"ロックファイル {lock_file.name} が存在します"
    try:
        with open(book, "r+b"):
            pass
    except PermissionError:
        return "書き込みロックされています（他のアプリが開いている可能性）"
    except OSError:
        return None   # その他の I/O エラーはここでは判定しない（保守的・誤検知回避）
    return None


# --- ★ W8b 項目1: 往復忠実度ゲート -------------------------------------------
#   normalize_book の LO 往復『だけ』で失われる飾り（原本にはあり、正規化後には無い）を
#   検出する。マクロの効果とは無関係（normalize は何もしない空マクロで一度 LO を通す
#   だけの工程）。喪失ゼロなら無言（体験は不変）。喪失があれば --inplace の直前で
#   申告し、--accept-loss / --copy のどちらかを選ばせる（exit 4）。

_FIDELITY_RELS_RE = re.compile(r"^xl/worksheets/_rels/.*\.rels$", re.I)


def _fidelity_zip_members(path: Path) -> set:
    try:
        with zipfile.ZipFile(path) as z:
            return {n.lower() for n in z.namelist()}
    except Exception:
        return set()


def _classify_fidelity_member(name: str) -> str | None:
    """xlsx zip 内の1エントリが『LO 往復で失われがちな飾り』のどのカテゴリか。
       対象外（本文と無関係な構造ファイル等）なら None。"""
    if name.startswith("xl/pivotcache") or name.startswith("xl/pivottables/"):
        return "ピボットテーブル"
    if name.startswith("xl/drawings/"):
        return "図形/描画"
    if name.startswith("xl/media/"):
        return "画像"
    if name == "xl/vbaproject.bin":
        return "VBA マクロ"
    if _FIDELITY_RELS_RE.match(name):
        return "リンク情報(_rels)"
    return None


def check_zip_fidelity_loss(original: Path, normalized: Path) -> list:
    """(a) original の zip 構成要素のうち normalized で消えたものをカテゴリ別に集計する。
       戻り値: [(カテゴリ, 消えた件数), ...]（件数0のカテゴリは含めない・カテゴリ名順）。
       ★ zip として読めない（xlsx でない・壊れている）場合は空リスト（保守的・誤検知しない）。"""
    before = _fidelity_zip_members(original)
    if not before:
        return []
    after = _fidelity_zip_members(normalized)
    lost = before - after
    counts: dict = {}
    for name in lost:
        cat = _classify_fidelity_member(name)
        if cat:
            counts[cat] = counts.get(cat, 0) + 1
    return sorted(counts.items())


def _count_cf_and_dv(wb) -> tuple:
    """ブック全体の(条件付き書式の件数, 入力規則の件数)を合計する。"""
    cf_total = 0
    dv_total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        cf = getattr(ws, "conditional_formatting", None)
        if cf is not None:
            cf_total += sum(1 for _ in cf)
        dv = getattr(ws, "data_validations", None)
        if dv is not None:
            dv_total += len(dv.dataValidation)
    return cf_total, dv_total


def check_openpyxl_fidelity_loss(original: Path, normalized: Path) -> list:
    """(b) openpyxl で条件付き書式/入力規則の件数が正規化後に減っていないかを見る。
       戻り値: [(カテゴリ, before件数, after件数), ...]（減っていないカテゴリは含めない）。
       ★ どちらか一方でも開けない場合は空リスト（保守的・誤検知しない）。"""
    try:
        wb_b = openpyxl.load_workbook(original)
        cf_b, dv_b = _count_cf_and_dv(wb_b)
        wb_b.close()
        wb_a = openpyxl.load_workbook(normalized)
        cf_a, dv_a = _count_cf_and_dv(wb_a)
        wb_a.close()
    except Exception:
        return []
    out = []
    if cf_a < cf_b:
        out.append(("条件付き書式", cf_b, cf_a))
    if dv_a < dv_b:
        out.append(("入力規則", dv_b, dv_a))
    return out


def check_round_trip_fidelity(original: Path, normalized: Path) -> dict:
    """往復忠実度ゲート本体。{"lost": bool, "items": [{"label":str,"count":int}, ...]}。
       ★ history.jsonl の fidelity フィールドにそのまま記録する形（機械可読）。"""
    items = []
    for cat, n in check_zip_fidelity_loss(original, normalized):
        items.append({"label": cat, "count": n})
    for cat, b, a in check_openpyxl_fidelity_loss(original, normalized):
        items.append({"label": cat, "count": b - a, "before": b, "after": a})
    return {"lost": bool(items), "items": items}


def format_fidelity_warning(fidelity: dict) -> str:
    """人間可読の申告文（例:「⚠ このファイルには、処理すると失われる飾りがあります
       （条件付き書式 3 件・図形/描画 1 件）」）。"""
    parts = "・".join(f"{it['label']} {it['count']} 件" for it in fidelity.get("items", []))
    return f"⚠ このファイルには、処理すると失われる飾りがあります（{parts}）"


# --- ★ W8b 項目4: アトミック置換（--inplace の torn-write 窓の根治） --------------

def atomic_replace_inplace(book: Path, out_book: Path, workdir: Path,
                            keep_backups: int = DEFAULT_KEEP_BACKUPS) -> tuple:
    """--inplace の実体。(ok: bool, error_message: str|None)。ok=False の場合、
       out_book はそのまま残り、原本(book)は無変更（呼び出し側が『--inplace は中止』
       として表示する）。
       ★ 手順: ①バックアップ（失敗したら即中止・原本は触らない）
       ②原本と同じボリューム上の staging(workdir/staged<suffix>)へコピー
       ③os.replace(staging, book) で原子的に置換（POSIX rename(2)/Windows
       MoveFileEx の原子性保証＝torn write の窓が無い。同一ボリュームである
       ことは staging を book と同じ親フォルダ配下(workdir)に置くことで保証する）。
       os.replace が失敗したら（バックアップは既に確保済みの上で）shutil.copy2
       へフォールバックし、その旨を1行表示する。成功時は out_book を削除する
       （原本に反映済みの内容と同じものを残しておく理由が無い＝旧 shutil.move
       と同じ最終状態）。"""
    try:
        make_backup(book, keep=keep_backups)
    except Exception as e:
        return False, f"バックアップに失敗したため --inplace を中止した（原本は無変更）: {e}"

    staging = workdir / f"staged{book.suffix}"
    try:
        shutil.copy2(out_book, staging)
        os.replace(staging, book)
    except OSError as e:
        try:
            shutil.copy2(out_book, book)
        except OSError as e2:
            return False, f"置換に失敗した（バックアップは確保済み）: {e2}"
        print(f"⚠ 原子的な置換に失敗したため copy2 へフォールバックした（バックアップは確保済み）: {e}")
    finally:
        if staging.exists():
            try:
                staging.unlink()
            except OSError:
                pass

    try:
        if out_book.exists() and out_book != book:
            out_book.unlink()
    except OSError:
        pass
    return True, None


# --- ★ W8b 項目6: グローバル run ロック --------------------------------------

def _pid_alive(pid: int) -> bool:
    """PID が生きているか（確実な保証は無いが十分・追加の依存(psutil 等)は増やさない）。
       判定できない場合は「生きている」扱い（安全側＝奪取しない）。"""
    if os.name == "nt":
        try:
            out = subprocess.run(["tasklist", "/FI", f"PID eq {pid}"],
                                  capture_output=True, text=True, encoding="utf-8", errors="replace")
            return str(pid) in out.stdout
        except Exception:
            return True
    else:
        try:
            os.kill(pid, 0)
            return True
        except ProcessLookupError:
            return False
        except OSError:
            return True


def _read_lock_info(path: Path) -> dict | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _lock_is_stale(info: dict | None) -> bool:
    """既存ロックが奪取してよい(stale)か。①壊れた/読めないロック ②pid が自分自身
       （同一プロセス内の前回呼び出しが解放し損ねた・テスト等で頻出）
       ③pid が既に存在しない ④30分超、のいずれか。"""
    if info is None:
        return True
    pid = info.get("pid")
    if not isinstance(pid, int):
        return True
    if pid == os.getpid():
        return True
    if not _pid_alive(pid):
        return True
    try:
        ts = datetime.fromisoformat(info.get("ts", ""))
        age = (datetime.now(timezone.utc) - ts).total_seconds()
        if age > RUN_LOCK_STALE_SECONDS:
            return True
    except Exception:
        pass
    return False


def acquire_run_lock(path: Path | None = None) -> tuple:
    """(acquired: bool, message: str|None)。O_EXCL でグローバル実行ロックを取る。
       ★ W8b 項目6: 基盤の LibreOffice は単一インスタンス(port 2002)前提のため、
       ブック単位でなく ailine run 全体で1本にする。stale（pid が既に無い/自分自身/
       30分超）なら奪取して新規取得する。"""
    p = path or RUN_LOCK_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    info = {"pid": os.getpid(), "ts": datetime.now(timezone.utc).isoformat(timespec="seconds")}

    def _try_create() -> bool:
        try:
            fd = os.open(str(p), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        except FileExistsError:
            return False
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(json.dumps(info))
        return True

    if _try_create():
        return True, None

    existing = _read_lock_info(p)
    if not _lock_is_stale(existing):
        pid = existing.get("pid") if existing else "?"
        ts = existing.get("ts") if existing else "?"
        return False, f"別の ailine が実行中です（pid={pid}・{ts}）"

    try:
        p.unlink()
    except OSError:
        pass
    if _try_create():
        return True, None
    return False, "別の ailine が実行中です（ロック取得の競合）"


def release_run_lock(path: Path | None = None) -> None:
    p = path or RUN_LOCK_FILE
    try:
        p.unlink()
    except OSError:
        pass


def maybe_show_notice_v2(path: Path | None = None) -> bool:
    """★ W10a 項目2: 既定変更(原本直接反映)の一度きり告知。marker ファイルが無ければ
       表示して作る（以後は無言）。戻り値は「今回表示したか」（テスト用）。
       marker の読み書き失敗（権限等）で run 本体を落とさないよう、書き込み失敗は無視する
       （その場合は次回また表示されうる＝安全側に倒す）。
       ★ path 省略時は HISTORY_FILE.parent（＝実運用では NOTICE_V2_FILE と同じ
       ~/.ailine/）から求める（固定の NOTICE_V2_FILE を直接使わない）。実測: この関数は
       cmd_run のたび無条件に呼ぶため、HISTORY_FILE を monkeypatch 済みの大量のテストが
       それに便乗して安全になる（このファイルは backups と違って『一度書いたら消えない』
       ため、実ユーザーの ~/.ailine/notice_v2_shown をテストが汚すと『初回表示』が二度と
       出せなくなる＝汚染の被害が他の一時ファイルより重い）。"""
    p = path or (HISTORY_FILE.parent / "notice_v2_shown")
    if p.exists():
        return False
    print(NOTICE_V2_TEXT)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(datetime.now(timezone.utc).isoformat(timespec="seconds"), encoding="utf-8")
    except OSError:
        pass
    return True


# ---------------------------------------------------------------------------
# doctor（M1: セットアップ診断）
# ---------------------------------------------------------------------------

def _load_module_from_path(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _check_python_version() -> tuple:
    ok = sys.version_info >= (3, 10)
    detail = "" if ok else f"現在 {sys.version.split()[0]}。3.10 以上へ更新して"
    return ok, detail


def _check_openpyxl() -> tuple:
    try:
        import openpyxl as _op  # noqa: F401 — 到達確認のみ
        return True, ""
    except ImportError:
        return False, "pip install openpyxl"


def _check_ollama_reachable(timeout: float = 3.0) -> tuple:
    try:
        with urllib.request.urlopen(f"{OLLAMA}/api/tags", timeout=timeout) as r:
            json.load(r)
        return True, ""
    except Exception as e:
        return False, f"`ollama serve` を起動して（{e}）"


def _check_model_available(model: str, timeout: float = 3.0) -> tuple:
    try:
        with urllib.request.urlopen(f"{OLLAMA}/api/tags", timeout=timeout) as r:
            d = json.load(r)
    except Exception:
        return False, "ollama に繋がらないため確認できない（先に ollama 到達を直して）"
    names = {m.get("name") for m in d.get("models", [])}
    if model in names:
        return True, ""
    return False, f"`ollama pull {model}` で取得して"


def _check_libreoffice() -> tuple:
    p = _find_basrun_path()
    if p is None:
        return False, "basrun.py が無いため確認できない"
    try:
        mod = _load_module_from_path(p, "_ailine_basrun_probe")
        d = mod.office_dir()
        return True, str(d)
    except SystemExit as e:
        return False, str(e)
    except Exception as e:
        return False, f"検出に失敗: {e}"


def _check_basrun() -> tuple:
    p = _find_basrun_path()
    if p is None:
        return False, ("ailine と並びに https://github.com/namakoo-dev/basrun を"
                       " clone するか、環境変数 BASRUN でパスを指定して")
    return True, str(p)


def _check_demo_dir() -> tuple:
    d = HERE / "demo"
    ok = d.is_dir() and any(d.glob("*.xlsx"))
    detail = "" if ok else f"{d} に .xlsx サンプルが無い"
    return ok, detail


def doctor_checks(model: str = DEFAULT_MODEL) -> list:
    """(項目名, ok, 詳細/直し方) のリスト。判定ロジックだけを持ち、副作用(print)は
       cmd_doctor 側に置く（テストしやすくするため分離）。"""
    return [
        ("python 3.10+", *_check_python_version()),
        ("openpyxl", *_check_openpyxl()),
        (f"ollama 到達 ({OLLAMA})", *_check_ollama_reachable()),
        (f"モデル '{model}'", *_check_model_available(model)),
        ("LibreOffice", *_check_libreoffice()),
        ("basrun.py", *_check_basrun()),
        ("demo/", *_check_demo_dir()),
        # ★ W10a 項目2: 動作可否のチェックではなく設定の告知（常に ok=True）。
        #   既定が原本直接反映に変わったこと(W8b-2〜)を doctor でも確認できるようにする。
        ("既定動作", True, "原本直接（v2〜）。--copy で従来のコピー方式"),
    ]


# ★ W8a 項目5: doctor の内部名（"ollama 到達 (URL)" 等）は operator（事務職）には
#   意味が伝わらない。内部名は残したまま(技術者が追える)、事務向けの一行説明を併記する。
#   prefix 一致（name には URL/モデル名など動的な値が混ざるため完全一致にしない）。
#   ★ ダミー名（テストの "a"/"b" 等）はどれにも一致しないため従来どおり内部名のみ表示する。
_DOCTOR_BUSINESS_NOTES = (
    ("python 3.10+", "実行に必要なプログラム言語が使えます"),
    ("openpyxl", "Excel ファイルを読み書きする部品が使えます"),
    ("ollama 到達", "AI エンジン (ollama) に接続できています"),
    ("モデル", "使う AI モデルの準備ができています"),
    ("LibreOffice", "文書を開いて処理する土台があります"),
    ("basrun.py", "文書に処理を適用する仕組みがあります"),
    ("demo/", "動作確認用のサンプル文書があります"),
)


def _doctor_business_note(name: str) -> str | None:
    for prefix, note in _DOCTOR_BUSINESS_NOTES:
        if name.startswith(prefix):
            return note
    return None


def format_doctor_report(results: list) -> tuple:
    """(表示テキスト, all_ok)。"""
    lines = []
    all_ok = True
    for name, ok, detail in results:
        mark = "✓" if ok else "×"
        note = _doctor_business_note(name)
        shown = f"{name}（{note}）" if note else name
        if ok:
            line = f"{mark} {shown}" + (f" ({detail})" if detail else "")
        else:
            all_ok = False
            line = f"{mark} {shown}" + (f" — {detail}" if detail else "")
        lines.append(line)
    return "\n".join(lines), all_ok


def cmd_doctor(a: argparse.Namespace) -> int:
    text, all_ok = format_doctor_report(doctor_checks(a.model))
    print(text)
    return 0 if all_ok else 1


def cmd_ops(a: argparse.Namespace) -> int:
    """★ `ailine ops`: 頼める操作の一覧。盲検査定 2 本が独立に MISSING の筆頭へ置いた
       「こう頼めばこれができる」の対応表。中身は登録簿から生成する（手書きしない）。"""
    for line in render_ops_table(OP_META, OP_SCHEMA, _CONFIRM_FIELDS):
        print(line)
    return 0


# ---------------------------------------------------------------------------
# 実行履歴（M1: 最小版・将来の需要センサー）
# ---------------------------------------------------------------------------

def build_history_entry(result: dict, book: Path, task: str, model: str, failure_kind: str,
                         error_detail: str | None = None) -> dict:
    """1 run の結果を history.jsonl の 1 行分の dict にする（純ロジック・テスト用に分離）。
       ★ M2a: error_detail は runtime_error 時の生エラー全文（端末には最終行だけ出す
       ため、詳細を追いたい時の唯一の入り口になる）。
       ★ W8a 項目1: "dry" は --dry（見せるだけ・未適用）で走ったかどうか。result["dry"] を
       そのまま bool 化する（result に無ければ False＝実適用と同じ扱い）。旧 history.jsonl
       の行（このキーが無い）は read_history/format_history_table 側で dict.get("dry", False)
       により実適用扱いのまま読める（後方互換・新旧を混在させても壊れない）。
       ★ W8b 項目1: "fidelity" は往復忠実度ゲートの検出結果（--inplace が要求され、
       かつ --dry でない run でのみ実際に計算される）。ゲートを走らせなかった run は
       None のまま（既存キーと同じ形＝無ければ None）。"""
    return {
        "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "book": str(book),
        "task": task,
        "model": model,
        "ok": bool(result.get("ok")),
        "dry": bool(result.get("dry")),
        "attempts": result.get("attempts", 0),
        "failure_kind": failure_kind,
        "error_detail": error_detail,
        "changes": (result.get("changes") or [])[:3],
        "out": result.get("out"),
        # ★ M2b: DSL 経路(path="dsl")では命令言語の確認文(command)と事後条件の合否を残す。
        #   自由生成経路(path="freeform")では両方 None のまま（既存キーは不変）。
        "path": result.get("path", "freeform"),
        "command": result.get("command"),
        "postcondition": result.get("postcondition"),
        # ★ A': 用語集/依頼文から機械確定した値（APPEND_TOTAL の倍率等）の出典。無ければ None。
        "provenance": result.get("provenance"),
        "fidelity": result.get("fidelity"),
    }


def append_history(entry: dict, path: Path | None = None) -> None:
    """history.jsonl に 1 行 append する。★ 失敗したら例外を投げる（run 本体を落とさ
       ないための try は呼び出し側(cmd_run)が持つ。ここでは書き込みロジックだけ）。"""
    p = path or HISTORY_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def read_history(path: Path | None = None, max_n: int = 10) -> list:
    """新しい順に最大 max_n 件を返す。壊れた行は読み飛ばす。"""
    p = path or HISTORY_FILE
    if not p.exists():
        return []
    entries = []
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            entries.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return list(reversed(entries))[:max_n]


def format_history_table(entries: list) -> str:
    """人が読める表形式。履歴が無ければ「履歴はまだ無い」を返す。
       ★ W8a 項目1: dry(下見・未適用)の行は末尾に「(下見)」を付けて実適用と区別する
       （「dry-run」は事務の言葉ではないため表示には出さない）。dict にキーが無い旧行は
       dict.get("dry", False) で実適用扱いのまま読める（後方互換）。"""
    if not entries:
        return "履歴はまだ無い"
    header = f"{'日時':<20} {'結果':<4} {'試行':<4} {'モデル':<20} {'文書':<20} タスク"
    lines = [header]
    for e in entries:
        mark = "✓" if e.get("ok") else "×"
        ts = str(e.get("ts", ""))
        attempts = str(e.get("attempts", ""))
        model = str(e.get("model", ""))
        book = Path(str(e.get("book", ""))).name
        task = str(e.get("task", ""))
        line = f"{ts:<20} {mark:<4} {attempts:<4} {model:<20} {book:<20} {task}"
        if e.get("dry", False):
            line += "  (下見)"
        kind = e.get("failure_kind")
        if not e.get("ok") and kind not in (None, "none"):
            line += f"  [{kind}]"
        lines.append(line)
    return "\n".join(lines)


def cmd_history(a: argparse.Namespace) -> int:
    entries = read_history(max_n=a.max)
    print(format_history_table(entries))
    return 0


# --- ★ A': 用語集(vocab) コマンド --------------------------------------------

def cmd_vocab(a: argparse.Namespace) -> int:
    """`ailine vocab add <語> <値>` / `ailine vocab list`。
       ★ remove は作らない（今の実需では add/list の2本で足りる。壊れたら vocab.json を
       直接編集すればよい）。"""
    if a.vocab_cmd == "add":
        ok, msg = vocab_add(a.term, a.value)
        print(render_vocab_add_result(ok, msg))
        return 0 if ok else 1
    # list
    vocab = load_vocab()
    for ln in render_vocab_listing(vocab, VOCAB_FILE):
        print(ln)
    return 0


# ---------------------------------------------------------------------------
# run コマンド本体
# ---------------------------------------------------------------------------

def _finish_run(a: argparse.Namespace, book: Path, result: dict, failure_kind: str,
                 error_detail: str | None = None) -> None:
    """--json 出力・成功時の注意書き・履歴の記録。cmd_run_freeform / cmd_run_dsl / cmd_run_plan
       の共通末尾。
       ★ DSL 経路(path="dsl")・複合計画経路(path="plan")は達成/総合判定の行を既に自分で
       出しているので、success_message() の『正しいかは差分を見て判断』（自由生成向けの
       注意書き）はここでは出さない。"""
    # ★ W8b 項目1: 往復忠実度ゲートの検出結果（cmd_run 側で計算済みなら a._fidelity に
    #   乗っている）を history に写す。ゲートを走らせなかった run は None のまま。
    if "fidelity" not in result:
        result["fidelity"] = getattr(a, "_fidelity", None)
    # ★ C9: --json の新キー。既存キーの意味は変えず、「何と照合し・どのファイルを読み戻して
    #   ✓ と言ったか」を足すだけ。✓ を出さなかった run（--dry・失敗・機械保証なし）は空リスト
    #   ＝『主張していない』が機械可読に残る。
    result.setdefault("claims", [])
    if a.json:
        print("\n" + json.dumps(result, ensure_ascii=False))
    if result.get("path") not in ("dsl", "plan"):
        msg = success_message(result)
        if msg:
            print("\n" + msg)
    try:
        detail = error_detail if error_detail is not None else (
            result.get("last_error_full") if failure_kind == "runtime_error" else None)
        append_history(build_history_entry(result, book, a.task, a.model, failure_kind,
                                            error_detail=detail))
    except Exception as e:
        print(f"WARN: 履歴の記録に失敗した: {e}", file=sys.stderr)


# ★ C9: ✓ の唯一の根拠。適用が全部終わった「最終ファイル」を openpyxl で開き直し、
#   機械が読み取れる事実だけで「今このファイルはこうなっている」を述べる。
#   ★★ ここで事後条件を再実行しない（明示的に否定された設計）: APPEND_TOTAL の後に SORT が
#   来れば合計行の位置は正当に動き（check_append_total は "=SUM(" の初出行で合計行を探す）、
#   再実行すると正しい run が偽 fail になる。読み戻すのは「反映が成功したこと」だけでよい。
def observe_book_state(path: Path) -> tuple:
    """(evidence: str|None, error: str|None)。全シート・全行を走査する（切り詰めない＝
       Claim.observation_complete=True の根拠）。evidence は最終ファイルだけから独立に
       再導出できる事実（シート名・行数・列数・値のあるセル数）に限る。"""
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return None, short_error_summary(str(e))
    try:
        parts = []
        for ws in wb.worksheets:
            filled = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
            parts.append(f"{ws.title}: {ws.max_row}行×{ws.max_column}列・値のあるセル {filled}")
        return "・".join(parts), None
    finally:
        wb.close()


# ★ W8b-2 項目1: 既定=原本直接適用の終端メッセージを一箇所に集約する
#   （cmd_run_dsl/cmd_run_freeform/cmd_run_plan の3箇所が同じ形だったのを統合）。
#   pending/confirm の中間状態は作らない — undo 一本（architect 判定）。
# ★★ C9: 『✓』の発生点をここ1箇所へ動かした。原本(--copy なら .out)が確定した後に
#   読み戻し、その結果だけを Claim にして描く。段別 ✓・--dry の ✓・反映前の ✓ は廃止。
def _finish_apply(a: argparse.Namespace, book: Path, out_book: Path, workdir: Path,
                   result: dict, machine_verified: bool, scope: str = "",
                   scope_note: str = "") -> bool:
    """--copy（a.inplace が False）なら .out のまま（原本は無変更）。既定(a.inplace)なら
       backup+原子的置換(atomic_replace_inplace)で原本へ反映する。そのうえで**最終ファイルを
       読み戻し**、machine_verified=True なら ✓ の1行を、False（自由生成・検証対象不足の段を
       含む計画）なら ⚠ の1行を出す。読み戻せなかったら ✓ は出さない。
       scope は照合した宣言（Claim.scope・machine_verified=True のとき必須）、
       scope_note は経路別の範囲注記（単発/複合計画）。
       戻り値: 置換が成功した(または --copy で置換不要だった)か。"""
    if a.inplace:
        ok_ip, err_ip = atomic_replace_inplace(
            book, out_book, workdir, keep_backups=getattr(a, "keep_backups", DEFAULT_KEEP_BACKUPS))
        if not ok_ip:
            print(f"× {err_ip}")
            print(_untouched_original_line(book, out_book))
            result["out"] = str(out_book)
            return False
        final, trailer = book, "（もとに戻す: ailine undo）"
        result["out"] = str(book)
    else:
        final, trailer = out_book, f"（原本 {book.name} は変更していません）"
        result["out"] = str(out_book)

    evidence, err = observe_book_state(final)
    if err is not None:
        for ln in render_applied_unobservable(final.name, err):
            print(ln)
    elif machine_verified:
        claim = Claim(verified=True, basis="declaration", scope=scope, evidence=evidence,
                       observation_complete=True, observed_on=str(final), observed_after_apply=True)
        for ln in render_applied_claim(claim, final.name):
            print(ln)
        if scope_note:
            print(scope_note)
        # ★ --json: 既存キーの意味は変えず、claims を足すだけ（何と照合し・どのファイルを
        #   読み戻して言っているのかを機械可読にする）。
        result["claims"] = [{"basis": claim.basis, "compared_with": claim.scope,
                              "observed_on": claim.observed_on}]
    else:
        for ln in render_applied_unverified(final.name, evidence):
            print(ln)
    print(trailer)
    return True


def _untouched_original_line(book: Path, out_book: Path) -> str:
    """★ C9: 失敗して終わるときに必ず出す1行。査定2本が「原本がどうなったか分からない」と
       書いた沈黙の穴 ―― 途中で止まった run は原本無変更を名乗らず、`.out` が黙って隣に
       残ることも告げていなかった。"""
    return f"（原本 {book.name} は変更していません。作業結果は {out_book.name} に残っています）"


def cmd_run(a: argparse.Namespace) -> int:
    """run コマンドの入口。★ W8b 項目6: 実処理(_cmd_run_body)の前後をグローバル run
       ロックで挟む（基盤の LibreOffice が単一インスタンス前提のため、ブック単位でなく
       ailine run 全体で1本）。取得できなければ即 exit 6（LO 起動・翻訳より前）。
       finally で確実に解放する（sys.exit も SystemExit 例外なので finally は通る）。"""
    acquired, msg = acquire_run_lock()
    if not acquired:
        print(f"× {msg}")
        return 6
    try:
        return _cmd_run_body(a)
    finally:
        release_run_lock()


def _cmd_run_body(a: argparse.Namespace) -> int:
    """run コマンドの本体。★ W3: 正規化パス(＋StructDump による見出し行推定)を
       翻訳より前に一度だけ行う（『三層全部が同じ見出し推定を使う』ための土台。
       translate_task 自身の接地(book_meta)も検出した見出し行を使う）。
       --dry は従来どおり LibreOffice に触れない（見出しは物理1行目のまま・E2E 対象外）。
       ★ W8a 項目3: --header-row N（1起点）が指定されていれば StructDump 検出を丸ごと
       スキップしてその行を採用する（CLARIFY の行き止まりから抜ける唯一の入り口）。
       ★ W8b 項目2: Excel ロック検出は LO 起動・翻訳より前（一番最初）に行う。
       ★ W8b 項目1: --inplace が要求され --dry でない場合だけ、正規化直後に往復忠実度
       ゲートを見る（原本にまだ一切触れていない段階＝『原本に触る前に申告と選択』）。
       ★ W8b 項目4: 自分の workdir(.ailine_<stem>/) は run の終了時に必ず掃除する
       （成功・失敗とも・GC の作り込みはしない＝自分の後始末だけ）。
       ① 見出し行推定（StructDump） → 自信不足なら CLARIFY して exit 3
       ② 翻訳（計画）→
       - 計画が空/1段で CLARIFY → 質問して exit 3
       - 計画が空/1段で DSL 語彙 → ③〜⑥の決定論パイプライン(cmd_run_dsl)
       - 計画が空/1段でそれ以外(FREEFORM・翻訳失敗) → 現行の自由生成経路(cmd_run_freeform)
       - 計画が2段以上(複合依頼) → 段ごとに honest な項目別実行(cmd_run_plan)（M2c）
       ★ 後方互換: translate_task が "plan" で包まない旧形式（bare {"op":...}）を返した場合
       （テストの monkeypatch を含む）も、その dict をそのまま単一段として扱う。"""
    # ★ M2（architect 致命4）: book の位置がディレクトリなら多ファイル分岐へ ── **一番最初**に
    #   分ける。ここから下は1冊のブック前提の器官（ロック検出・正規化・バックアップ・undo）で、
    #   フォルダを渡すと check_excel_lock の open(r+b) が PermissionError になり
    #   「Excel で開かれています」という嘘の診断を返していた（凍結検体あり）。
    if Path(a.book).is_dir():
        return cmd_run_folder(a)
    maybe_show_notice_v2()   # ★ W10a 項目2: 既定変更の一度きり告知（run の一番最初）

    book = Path(a.book).resolve()
    if not book.exists():
        sys.exit(f"文書が無い: {book}")

    lock_reason = check_excel_lock(book)
    if lock_reason:
        print(f"× Excel で開かれています。閉じてから実行してください。（{lock_reason}）")
        return 5

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    try:
        return _cmd_run_dispatch(a, book, workdir)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _cmd_run_dispatch(a: argparse.Namespace, book: Path, workdir: Path) -> int:
    # ★ W8b-2 項目2: 既定 = 原本直接適用（旧 --inplace 相当）。--copy で旧 .out 挙動を
    #   温存する。--inplace は廃止した旧フラグだが、互換のため受理はして移行メッセージ
    #   だけ出す（指定してもしなくても挙動は同じ＝既定が原本適用）。
    if getattr(a, "inplace", False):
        print("★ --inplace は廃止されました。既定で原本に直接適用します"
              "（従来の .out 挙動が欲しい場合は --copy を使ってください）。")
    a.inplace = not getattr(a, "copy", False)

    apply_timeout = a.timeout if a.timeout else None   # 0 で無効化（旧挙動 = 無制限）

    source_book = book
    struct_dump: dict = {}
    if not a.dry:
        t0 = progress_start("⏳ 初回準備（文書の正規化+構造読み取り）…")
        source_book = normalize_book(book, workdir, timeout=apply_timeout)
        progress_end(t0)
        struct_dump = build_struct_dump(source_book, workdir)

        # ★ W8b 項目1: 原本に実際に適用する時だけ、まだ触れていないこの時点で往復忠実度
        #   を見る（喪失ゼロなら無言・体験は不変）。★ W8b-2 項目4: --copy 時は原本に
        #   一切触れないためゲート自体を走らせない（無駄な zip/openpyxl 比較も省く）。
        if a.inplace:
            fidelity = check_round_trip_fidelity(book, source_book)
            a._fidelity = fidelity
            if fidelity["lost"]:
                print(format_fidelity_warning(fidelity))
                if getattr(a, "accept_loss", False):
                    print("→ --accept-loss 指定のため続行します（失われても ailine undo で元に戻せます）")
                else:
                    for ln in render_retry_options("", [
                            ("--accept-loss", "失われてよい（バックアップから ailine undo で復元可能）"),
                            ("--copy", "原本には触らず .out に結果を作る（原本は無変更）")]):
                        print(ln)
                    return 4

    # ★ 挙動変更#3: 衝突検出のため headers も使う。build_book_meta は元から1回呼んでいて
    #   sheets しか読んでいなかった ── 同じ戻り値の headers を渡すだけ（ブックは開き直さない）。
    probe_meta = build_book_meta(source_book)
    sheets = probe_meta.get("sheets", [])

    # ★ 挙動変更#2: 対象シートの決定はここ1箇所（resolve_target_sheet）。それより後の
    #   全処理（見出し行検出・翻訳→検証→codegen→事後条件）は a._target_sheet を読むだけで
    #   個別に「1枚目」を仮定しない。--header-row/翻訳より前＝原本の実処理が始まる前に
    #   決め、実処理の前に必ず宣言する（査定所見:「これがあれば事故は防げた」）。
    target_sheet, sheet_source, sheet_err, sheet_conflict = resolve_target_sheet(
        a.task, sheets, getattr(a, "sheet", None), headers=probe_meta.get("headers"))
    if sheet_err:
        print(f"？ {sheet_err}")
        return 3
    a._target_sheet = target_sheet
    a._sheet_conflict = sheet_conflict      # ★ 挙動変更#3: 3択の関門(_sheet_conflict_gate)が読む
    a._rerun_ctx = (book, source_book, struct_dump, sheets)
    announce = describe_target_sheet(sheets, target_sheet, sheet_source)
    if announce:
        print(announce)
    return _translate_and_dispatch(a, book, source_book, struct_dump, sheets)


def _translate_and_dispatch(a: argparse.Namespace, book: Path, source_book: Path,
                             struct_dump: dict, sheets: list) -> int:
    """対象シートが決まった後の残り（見出し行 → 翻訳 → 計画の振り分け）。
       ★ 挙動変更#3 のために _cmd_run_dispatch から切り出した: シート名の衝突の3択で
       「もう一方のシートを見てみる」が選ばれたとき、a._target_sheet を差し替えて
       **翻訳からやり直す**ため（見出し行の検出も翻訳結果も対象シートに依存するので、
       翻訳を使い回すのでは「やり直し」にならない）。
       ★ a._reuse_translation が積まれている場合だけ翻訳を省く ── 3択の②で
       プレビュー用に一度翻訳したものを、y の後の本番実行で使い回すため（同じ対象シート・
       同じ依頼文に対して ollama を2回叩かない）。"""
    target_sheet = getattr(a, "_target_sheet", None)
    forced_header_row = getattr(a, "header_row", None)
    if forced_header_row:
        # ★ W8a 項目3: --header-row 指定時は検出(StructDump ヒューリスティクス)を丸ごと
        #   スキップし、その行を対象シートの見出しとして採用する（他シートは既定1行目のまま
        #   ＝ resolve_header_rows の既定と同じ扱い）。CLARIFY には絶対に落ちない。
        header_rows = {s: 1 for s in sheets}
        if target_sheet:
            header_rows[target_sheet] = forced_header_row
        clarify_q = None
    else:
        header_rows, clarify_q = resolve_header_rows(struct_dump, sheets, target_sheet=target_sheet)
    if clarify_q:
        print(f"？ {clarify_q}")
        return 3

    book_meta = build_book_meta(source_book, header_rows=header_rows)
    translation = getattr(a, "_reuse_translation", None)
    a._reuse_translation = None
    if translation is None:
        t0 = progress_start(f"⏳ 翻訳中 ({a.model})…")
        translation = translate_task(a.model, a.task, book_meta, temperature=0.1)
        progress_end(t0)
    a._last_translation = translation

    plan = translation.get("plan") if isinstance(translation, dict) else None
    if not isinstance(plan, list) or not plan:
        if isinstance(translation, dict) and translation.get("op"):
            plan = [translation]
        else:
            plan = [{"op": "FREEFORM", "args": {}}]

    if len(plan) == 1:
        step = plan[0]
        op = step.get("op")
        if op == "CLARIFY":
            question = step.get("question") or "確認が必要です"
            print(f"？ {question}")
            # ★ 行き止まりに出口を置く（盲検査定 A の実測: 語彙外の依頼を 4 回言い直して
            #   4 回とも質問返しになり「普通の購入検討者ならここで評価を終える」）。
            #   聞き返しは「言い方が悪い」場合と「そもそも対応していない」場合を
            #   区別できない ── 区別する手段を毎回そえる。
            print("  （頼める操作の一覧: ailine ops）")
            return 3
        if op in OP_SCHEMA:
            return cmd_run_dsl(a, book, source_book, book_meta, op, step.get("args", {}))
        return cmd_run_freeform(a, book, source_book, step)

    return cmd_run_plan(a, book, source_book, book_meta, plan)


def _column_existing_value_count(book_path: Path, sheet_name: str, col_name: str,
                                  header_row: int = 1) -> int:
    """★ M2c / W10a: target(既存列指定)列に、見出し行を除いて値が入っているセルの件数。
       上書き検知の明示（確認行の注意書き）と W10a の破壊の関所（確認メッセージの件数）が
       共有する。読めない/列やシートが見つからない場合は 0（保守的に『無い』扱い＝誤って
       止めない）。★ W3: header_row(1起点)で見出しの実位置を受け取る
       （省略時は物理1行目・旧挙動と同一）。"""
    try:
        wb = openpyxl.load_workbook(book_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb[sheet_name]
        idx = _col_index_by_header(ws, col_name, header_row=header_row)
        if idx is None:
            wb.close()
            return 0
        last = _scan_last_row(ws, header_row=header_row)
        count = sum(1 for r in range(header_row + 1, last + 1)
                    if ws.cell(row=r, column=idx).value not in (None, ""))
        wb.close()
        return count
    except Exception:
        return 0


def _maybe_warn_target_overwrite(op: str, resolved: dict, book_meta: dict, book_path: Path) -> str | None:
    """★ M2c 項目2 / W10c 致命1: OP_WRITE_TARGET が宣言する書き込み先列に既存値がある場合、
       上書きになる旨の1行を返す（無ければ None・確認行に明示するため）。
       ★ W10c: 対象を COMPUTE_COLUMN 専用の if から OP_WRITE_TARGET の宣言読み取りへ
       一般化した（監査実測: LOOKUP_FILL がこの関所を素通りしていた事故の再発防止。
       OP_WRITE_TARGET のコメント参照）。
       ★ W10a 項目1: この検出（と件数）を「破壊の関所」（原本適用時に確認を挟む・
       cmd_run_dsl/cmd_run_plan 側）がそのまま流用する（検出ロジックを二重管理しない）。
       ★ 単位C: 宣言が領域を持つようになったので、『既存列を書く』と宣言した op だけを
       対象にする（col_key を持つのはその op だけ＝番人テストが両者の一致を検査する）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or not write_target.col_key:
        return None
    col_key, sheet_key = write_target.col_key, write_target.sheet_key
    col_name = resolved.get(col_key)
    if not col_name:
        return None
    if sheet_key:
        sheet_name = resolved.get(sheet_key)
    else:
        # ★ 挙動変更#2: sheets[0] 決め打ちをやめ、verify_dsl_args が一箇所で決めた
        # resolved["_target_sheet"] を読む（旧値と後方互換のフォールバック付き）。
        sheets = book_meta.get("sheets") or []
        sheet_name = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    if not sheet_name:
        return None
    header_row = book_meta.get("header_rows", {}).get(sheet_name, 1)
    count = _column_existing_value_count(book_path, sheet_name, col_name, header_row=header_row)
    if count > 0:
        return f"★ 対象列『{col_name}』には既存の値が {count} 件あります（上書きします）"
    return None


def _own_output_headers(op: str, resolved: dict | None):
    """★★ 単位H: 「宣言した出力先に在るのは、自分の前回の出力か」を見分ける見出し署名。

    ★ 署名は実装から取る（想像しない）: `helpers/AiLineHelpers.bas` の SummaryTable は
      A1 に分類列の見出し名を、B1 に "合計 - " & 集計列名 を書く（同ファイル 374-375 行）。
    ★ PIVOT は対象外。PivotSum は本物の DataPilot を LibreOffice に描かせるので、
      出力の見出しがコード上に無く、静的に決められない。
      → 発火条件つきで残す: DataPilot の実出力の見出しを 1 度実機で観測できたら、ここに足す。
        それまで PIVOT の 2 回目は関所が鳴る（＝安全側で、うるさい側に倒れている）。
    ★★ EXTRACT（単位H の2例目・コミット 2edcb08「EXTRACT op」）: ExtractRows ヘルパは
      出力の1行目に元シートの見出し行をそのままコピーする（helpers/AiLineHelpers.bas 参照）。
      出力シート名は固定でなく col/cmp/value から動的に決まる（_extract_output_sheet_name。
      verify_dsl_args が resolved["_new_sheet"] に積む）ので OP_DECLARED_SHEET_NAME の
      固定表には乗らない ── 署名の材料(見出し全体)は resolved["_source_headers"]（同じく
      verify_dsl_args が積む）を読む。
    """
    if not resolved:
        return None
    if op == "AGGREGATE":
        sheet = OP_DECLARED_SHEET_NAME.get(op)
        group, value = resolved.get("group_col"), resolved.get("value_col")
        if not (sheet and group and value):
            return None
        return {sheet: (group, f"合計 - {value}")}
    if op == "EXTRACT":
        sheet = resolved.get("_new_sheet")
        source_headers = resolved.get("_source_headers")
        if not (sheet and source_headers):
            return None
        return {sheet: tuple(source_headers)}
    return None


def _maybe_warn_write_precondition(op: str, before: dict, after: dict, resolved: dict | None = None):
    """★★ 単位F: op が宣言した書き込み領域(OP_WRITE_TARGET.writes)の**前提**が、適用後の
       実測で破れていれば1行を返す（無ければ None）。判定本体は
       ailine_core/write_precondition.py（純ロジック・前提の一覧はそこの docstring）。
       ★ 上の _maybe_warn_target_overwrite（適用**前**に列の既存値を数える）と、この関数
       （適用**後**に before/after を突き合わせる）は、同じ破壊の関所の2つの入力にすぎない
       ―― 返すのは同じ形の1行で、止める判断は両方とも _confirm_overwrite_or_gate が行う。
       col_key を持つ 3 op しか守れていなかった関所に、行・シート・書式・並べ替えの腕を
       足すのがこの関数（関所そのものは増やさない）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target:
        return None
    return check_write_preconditions_detail(write_target.writes, before, after,
                                             cell_ref=_cell_ref, fmt_value=_fmt_cell_value,
                                             own_output_headers=_own_output_headers(op, resolved))


def _maybe_own_prior_output_notice(op: str, before: dict, after: dict, resolved: dict | None = None) -> list:
    """★★ 単位H 開示: _looks_like_own_prior_output が真になって new_sheet の前提検査を
       スキップした（＝関所が黙った）ときの理由を、1行ずつ開示する（無ければ空リスト）。
       助言ではない ── 前提検査（_maybe_warn_write_precondition）が黙って通した理由の
       開示そのもの。呼び出し側は前提検査の直後で呼ぶ（単発・複合計画の両経路）。
       ★ new_sheet を writes に宣言していない op は最初から対象外（own 判定自体が起きない）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or "new_sheet" not in write_target.writes:
        return []
    return own_prior_output_notice_lines(before, after, own_output_headers=_own_output_headers(op, resolved))


def _interpretation_summary_line(resolved: dict, inferred: set) -> str | None:
    """★ W10a 項目3: 実行前の解釈要約。target が数字表記から列名へ推定解決され、かつ
       （呼び出し側が別途 _maybe_warn_target_overwrite で）既存データありと分かっている
       場合だけ、その経緯を1文で見せる（「監査要望3」＝数字指定→列名解決の可視化）。
       それ以外（target が最初から実在列名で指定された等）は None（何も語ることが無い）。"""
    if "target" not in inferred or resolved.get("_target_raw") is None:
        return None
    return f"→『{resolved['_target_raw']}』は既存の『{resolved['target']}』列と解釈しました（既存データあり）"


def _confirm_overwrite_or_gate(a: argparse.Namespace, warn_overwrite: str | None,
                                step_prefix: str = "", subject_mismatch: bool = False) -> int | None:
    """★ W10a 項目1: 破壊の関所。既定(原本へ直接反映)で、既存データへの上書きが起きる
       操作（_maybe_warn_target_overwrite が検出）は、--ask 無指定でも確認を挟む
       （監査実測: target が誤って既存列に解決され、確認なしで実データが上書きされた
       事故の再発防止）。--copy/--dry 時は原本に触れない/何もしないため素通し。
       --overwrite が既に立っていれば承知の上として素通し。--ask が既に立っている場合は
       cmd_run_dsl 側の汎用確認で兼ねる（二重に聞かない）。
       戻り値: 続行してよければ None、中断すべきなら呼び出し側がそのまま return すべき
       exit code（対話で拒否=1・非対話で確認できない=7）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。単発 DSL では空文字。
       ★★ 単位E: subject_mismatch（依頼文の語と解決値が矛盾する対象がある＝③）も**この関所を
       再利用**して確認する。新しい関所も新しい exit code も作らない ―― 止める条件
       （原本へ直接反映する時だけ・--dry/--copy/--ask/--overwrite では素通し）も、拒否 1 /
       非対話 7 も、上書きの関所とまったく同じ。違うのは聞く文と、非対話時に示す逃げ道だけ。
       ★ 対象シートの3択（_sheet_conflict_gate）を既に出した run では重ねて聞かない
       （同じ対象について2度尋ねることになる。✓ の抑制と ⚠ の印字はそのまま行う）。"""
    reason = warn_overwrite or (subject_mismatch and not getattr(a, "_sheet_conflict_asked", False))
    if not (reason and getattr(a, "inplace", False) and not getattr(a, "dry", False)
            and not getattr(a, "ask", False) and not getattr(a, "overwrite", False)):
        return None
    prompt = "上書きしますか？" if warn_overwrite else "この対象で実行しますか？"
    try:
        ans = input(f"{step_prefix}{prompt} [y/N]: ").strip().lower()
    except EOFError:
        options = [("--overwrite", "上書きを承知して続行する（バックアップから ailine undo で戻せる）"),
                   ("--copy", "原本には触らず .out に結果を作る（原本は無変更）")] if warn_overwrite \
            else [("--copy", "原本には触らず .out に結果を作る（原本は無変更）"),
                  ("--sheet / 列名を依頼文に明記", "対象を依頼文で名指しして、もう一度実行する")]
        for ln in render_retry_options(step_prefix, options):
            print(ln)
        return 7
    if ans not in ("y", "yes"):
        print(render_aborted(step_prefix))
        return 1
    return None


def _stdin_isatty() -> bool:
    """標準入力が端末か（対話してよいか）。★ ここを関数にしてあるのはテストから
       差し替えるため（stdin そのものを触らずに TTY/非 TTY を作り分ける）。"""
    try:
        return bool(sys.stdin) and sys.stdin.isatty()
    except Exception:
        return False


def _sheet_conflict_gate(a: argparse.Namespace, op: str) -> int | None:
    """★ 挙動変更#3: シート名の衝突で既定(1枚目)へ後退した時だけ、「解釈:」行の直後に
       3択を出す薄い配線（選択肢の文言は ailine_core.target_sheet、対話の仕組みは
       ailine_core.ask_choice が本体）。
       戻り値: そのまま続行してよければ None、呼び出し側がそのまま return すべき
       exit code なら int（②の実行結果もここに含まれる）。
       ★ 聞くのは衝突で後退した時だけ・TTY の時だけ（--json/パイプ/CI/--dry では
       絶対に入力を待たない ── 止めると動いていたスクリプトが黙って壊れる）。"""
    conflict = getattr(a, "_sheet_conflict", None)
    if not conflict or getattr(a, "_sheet_conflict_asked", False):
        return None
    # ★ 誤爆#3: ここで a._sheet_conflict は消さない ── 「曖昧なので既定へ後退した」という
    #   判定結果は実行の最後（助言）まで運ぶ必要がある。消してよいのは『もう聞いた』という
    #   対話側の状態だけなので、別のフラグに分けた（②で対象シートを差し替えたときだけは
    #   後退そのものが取り消されるので、_preview_and_run_on_alternative_sheet が記録を落とす）。
    a._sheet_conflict_asked = True   # 二重に聞かない（②のやり直しでも再発火させない）
    lines, choices = sheet_conflict_choice_lines(conflict, op, OP_LABELS.get(op, op))
    result = ask_choice(lines, [Choice(key=k, text=t) for k, t in choices],
                        interactive=is_interactive(stdin_isatty=_stdin_isatty(),
                                                   json_mode=getattr(a, "json", False),
                                                   dry=getattr(a, "dry", False)))
    if result.key in (None, "1"):
        return None            # 聞かなかった/①＝上の解釈のとおり（既定＝今までの挙動）
    if result.key == "3":
        print(render_aborted())
        return 1               # 既存の中止系（対話で拒否）と同じ exit code
    return _preview_and_run_on_alternative_sheet(a, conflict)


def _preview_and_run_on_alternative_sheet(a: argparse.Namespace, conflict) -> int:
    """★ 挙動変更#3 の②: もう一方のシートを対象に**翻訳からやり直し**、その解釈行と
       --dry 相当のプレビューを見せた上で y/N を聞く。y ならそちらで実行、N ならやめる
       （原本は無変更）。この時点で原本にはまだ一切触れていない。"""
    book, source_book, struct_dump, sheets = a._rerun_ctx
    a._target_sheet = conflict.alternative
    # ★ 誤爆#3: ②を選んだ時点で「既定へ後退した」という判定は取り消された（そのシート
    #   自身が対象になる）。記録を落として、助言側のシート言及の抑制も効かせない。
    a._sheet_conflict = None
    original_dry, a.dry, a._preview_only = getattr(a, "dry", False), True, True
    rc = _translate_and_dispatch(a, book, source_book, struct_dump, sheets)
    a.dry, a._preview_only = original_dry, False
    if rc != 0:
        return rc
    if not ask_yes_no("この内容で実行しますか？ [y/N]: ",
                      interactive=is_interactive(stdin_isatty=_stdin_isatty(),
                                                 json_mode=getattr(a, "json", False),
                                                 dry=original_dry)):
        print(render_aborted())
        return 1
    a._reuse_translation = getattr(a, "_last_translation", None)   # 同じ翻訳を使い回す
    return _translate_and_dispatch(a, book, source_book, struct_dump, sheets)


def _make_dsl_step_deps() -> DslStepDeps:
    """★ C7: 呼び出しのたびに毎回新しく組み立てる依存の束（monkeypatch を効かせるため・dsl_step.py 参照）。"""
    return DslStepDeps(
        format_confirmation_line=format_confirmation_line,
        maybe_warn_header_col_mismatch=_maybe_warn_header_col_mismatch,
        maybe_warn_target_overwrite=_maybe_warn_target_overwrite,
        interpretation_summary_line=_interpretation_summary_line, confirm_overwrite_or_gate=_confirm_overwrite_or_gate,
        basrun_apply=basrun_apply, snapshot=snapshot, diff_snapshots=diff_snapshots,
        run_postcondition=run_postcondition, progress_start=progress_start, progress_end=progress_end,
        pivot_caveat=PIVOT_CAVEAT, verify_dsl_args=verify_dsl_args,
        apply_new_column_fallback=_apply_new_column_fallback, build_advisories=build_advisories,
        structural_advisories=_structural_advisories, unrequested_new_sheet_advisory=unrequested_new_sheet_advisory,
        classify_subject_provenance=classify_subject_provenance,   # ★ 単位E
        sheet_conflict_gate=_sheet_conflict_gate)   # ★ 挙動変更#3


def cmd_run_dsl(a: argparse.Namespace, book: Path, source_book: Path, book_meta: dict,
                 op: str, raw_args: dict) -> int:
    """M2b の決定論パイプライン本体。②検証 → ③確認行 → ④codegen → ⑤適用 → ⑥事後条件。
       ★ W3: source_book は cmd_run が翻訳より前に正規化済み・やり直さない。
       ★ C7: ②③⑤⑥は ailine_core.dsl_step の共有エンジン（_run_dsl_plan_step と同じコード・
       「単発 = 1 段の計画」）。--ask・.bas/変更点の印字・_truncation_notice は単発固有のまま。"""
    vocab = load_vocab()
    deps = _make_dsl_step_deps()
    # ★ 挙動変更#2: 対象シートは _cmd_run_dispatch が resolve_target_sheet で一箇所だけ
    #   決めた a._target_sheet を読む（後方互換フォールバックは1枚目・旧挙動と同一）。
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    ground = resolve_dsl_step_args(op, raw_args, a.task, book_meta, vocab, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        print(f"？ {ground.err}")
        return 3
    resolved, inferred = ground.resolved, ground.inferred

    # ★ 挙動変更#2: header_row は「本当の」対象シート(resolved["_target_sheet"])で引き直す。
    #   LOOKUP_FILL は自分の target_sheet slot が最終的な正で、依頼文が参照シート(source_sheet)
    #   の名前も含む場合に一般解決(a._target_sheet・op を知る前の機械的な文字列一致)が
    #   参照シート側へ寄る可能性があるため、op が分かった後のここで必ず読み直す
    #   （他の op は resolved["_target_sheet"] == first_sheet のまま・実質無変化）。
    first_sheet = resolved.get("_target_sheet") or first_sheet
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    use_formula = not getattr(a, "values", False)

    print(render_run_header("DSL 経路", a.model, book.name))
    confirm = print_dsl_confirmation(op, resolved, inferred, a.task, meta=book_meta, warn_book=book, new_cols=None, a=a, deps=deps)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所
        return confirm.gate_exit

    if a.ask:
        try:
            ans = input("続行しますか？ [y/N]: ").strip().lower()
        except EOFError:
            ans = ""
        if ans not in ("y", "yes"):
            print(render_aborted())
            return 1

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = book.with_name(book.stem + ".out" + book.suffix)
    apply_timeout = a.timeout if a.timeout else None   # 0 で無効化（旧挙動 = 無制限）
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    _helper_catalog, helper_files = load_helpers(helpers_dir)

    code = codegen_dsl(op, resolved, book_meta, use_formula=use_formula)
    (workdir / "dsl_attempt.bas").write_text(code, encoding="utf-8")
    # ★ W8a 項目5: 「決定論」はユーザー向け文字列から排除（内部名・関数名は不変）。
    for ln in render_code_block(f"\n─ 生成した .bas（ルール変換・LLM不使用）───────────────", code):
        print(ln)

    # ★ 段1: interpretation/provenance は1箇所（build_interpretation）で組む
    #   （単位C の教訓 ―― 出所を運ぶ場所が2つあると片方だけ更新されて食い違う）。
    #   provenance は resolved["_sources"] をそのまま返す派生ビュー（値・型は不変）。
    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book.name])
    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              "path": "dsl", "command": confirm.line, "postcondition": None,
              "interpretation": interpretation, "provenance": provenance}

    if a.dry:
        # ★ 挙動変更#3: シート衝突②の内部プレビュー（_preview_and_run_on_alternative_sheet）
        #   では「--dry を外して実行」の案内は誤り（この直後に y/N を聞くため）。
        if not getattr(a, "_preview_only", False):
            print("（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        _finish_run(a, book, result, "none")
        return 0

    before = snapshot(source_book)
    shutil.copy2(source_book, out_book)   # 原本は触らず、正規化済みコピーに適用

    # ★ C7: ⑤適用〜⑥事後条件（共有エンジン）。print_changes は単発固有（docstring 参照）。
    apply_result = apply_dsl_step(
        op, resolved, code, apply_target=out_book, before=before, before_charts=before["charts"],
        workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout, header_row=header_row,
        use_formula=use_formula, source_book=source_book, deps=deps,
        apply_progress_label="⏳ LibreOffice で適用中…", print_changes=True)

    if apply_result.runtime_error is not None:
        print(f"× 実行時エラー: {short_error_summary(apply_result.runtime_error)}（詳細は履歴に記録）。")
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["last_error_full"] = apply_result.runtime_error
        _finish_run(a, book, result, "runtime_error", error_detail=apply_result.runtime_error)
        return 1

    after, lines = apply_result.after, apply_result.changes
    # ★ 止血3/C7: 単発は常に呼ぶ（★ C9 で複合計画の DSL 段の穴も塞いだ・dsl_step.py 参照）。
    notice = _truncation_notice(before, after, exhaustive_postcondition=True)
    if notice:
        print(notice)
    # ★ W10b 項目4b(摩擦): 参照専用シートは「変更なし」対象から除外。
    # ★ 単位C(D8): ここにあった `if op == "LOOKUP_FILL"` のハードコードは削除した ──
    #   除外すべきシートは OP_WRITE_TARGET の reads_only 宣言そのものなので、
    #   build_advisories が op/resolved から自分で求める（AGGREGATE/PIVOT の入力シートも同じ）。
    # ★ 誤爆#3: 対象シートを決めた側の判定結果（曖昧なので既定へ後退した）をそのまま運ぶ。
    # ★★ 単位G: 前提の検査を「助言を組む前」に move した（検査の中身は 単位F のまま）。
    #   理由: 中立化（「（既存シート『集計』の更新は意図どおりです）」）は助言の発生源で
    #   決まるので、前提が破れたことを 発生源が知っていなければならない。
    #   ★ 後から該当行を消す方式は採らない ── ailine.py の中立化は W10c/W10d で
    #   「出してから打ち消す」を捨てて発生源へ先取りした経緯がある（この関数の上の注記）。
    #   ★ 印字と関所は 位置を変えていない（後段のまま）。ここでやるのは判定だけ。
    precondition = _maybe_warn_write_precondition(op, before, after, resolved)
    precondition_broken = precondition[0] if precondition else None
    # ★★ 単位H 開示: 関所判定の直後（★ 単位H 導入前は完全な無言だった箇所）。助言ではない
    #   ── 「前回の自分の出力の作り直し」と判定して関所を黙らせた理由そのものを見せる。
    for own_notice in _maybe_own_prior_output_notice(op, before, after, resolved):
        print(own_notice)
    advisories = compose_dsl_step_advisories(   # mode="flat" は単発固有（dsl_step.py 参照）
        "flat", op, resolved, book_meta, a.task, before, after, deps=deps,
        sheet_conflict=getattr(a, "_sheet_conflict", None),
        precondition_broken=precondition_broken) + formula_error_advisory(source_book, out_book, cell_ref=_cell_ref)   # ★ 挙動変更#1(a)
    for adv in advisories:
        print(adv)
    result["changes"] = lines
    result["advisories"] = advisories

    status, reason = apply_result.postcondition_status, apply_result.postcondition_reason
    # ★ 止血1/2: "error"(チェッカー内の予期しない例外)は --json 上 "fail" に丸める。
    result["postcondition"] = "fail" if status == "error" else status
    if status == "error":
        print(f"\n× {reason}")
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_error")
        return 1
    if status == "fail":
        print(f"\n× 適用されたが事後条件を満たさない: {reason}")
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_fail")
        return 1
    if status == "warn":
        # ★ 止血1: 検証対象が少なすぎる場合、「機械検証済み」とは名乗らない。
        print(f"\n⚠ 事後条件を機械検証できなかった（操作:{OP_LABELS.get(op, op)}）: {reason}")
    else:
        # ★ C9: 事後条件が見た中身（例「3 行を検証（降順）」）はここで述べる。✓ とは呼ばない
        #   ―― ✓ は原本(--copy なら .out)が確定した後の1行だけ（_finish_apply）。
        print(f"\n事後条件を確認（操作:{OP_LABELS.get(op, op)}）: {reason}")
    result["ok"] = True

    # ★★ 単位F: 反映の直前（原本はまだ無傷・before/after は両方手元にある）に、宣言した
    #   書き込み領域の前提を確かめる。破れていたら**同じ破壊の関所**へ渡す（新しい関所も
    #   新しい exit code も作らない ―― 上書き警告と同じ引数の位置に、同じ形の1行を渡す）。
    warn_precondition = precondition[1] if precondition else None   # ★ 単位G: 上で 1 度だけ検査済み
    if warn_precondition:
        print(warn_precondition)
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition)
        if gate_exit is not None:
            return gate_exit

    # ★ C9: postcondition が warn（検証対象不足）なら ✓ は名乗らない。scope は「解釈: ...」行から
    #   「解釈: 」を除いた宣言テキスト（＝計画が宣言した対象）。
    # ★★ 単位E: ③（依頼文の語と矛盾する対象がある）なら、事後条件が通っていても ✓ は出さない
    #   ―― 検証したのは「計画どおり」であって、その計画は依頼文の語と食い違っている。
    _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(status != "warn" and not confirm.subject_warnings),
                   scope=confirm.label, scope_note="\n".join(render_scope_notes(list(confirm.unspoken))))

    _finish_run(a, book, result, "none")
    return 0


# ---------------------------------------------------------------------------
# ★ W8a 項目4: 率リテラルの機械スキャン（判断棚から昇格）
#   自由生成(FREEFORM/OUT_OF_VOCAB・単段)された Basic コードに、依頼文にも用語集にも
#   出典の無い『率らしい数値』（消費税8%のつもりの 0.08、税込計算の *1.1 等）が紛れて
#   いないかを機械で見る。実測: 単段の自由生成では「8% 仮定」がノーチェックで
#   「✓できました」を素通りしていた（勝手な税率仮定はモデルの幻覚と見分けがつかない）。
#   ★ 保守的: コメント行(') は対象外・整数(小数点なし)は対象外。発火してもブロックはせず
#   「検算してください」の助言止まり（no-op ガード等と違い、正しさの断定はしない）。
# ---------------------------------------------------------------------------

_BASIC_COMMENT_LINE_RE = re.compile(r"^\s*(?:'|REM\b)", re.I)
_RATE_LITERAL_RE = re.compile(r"(?<![\w.])(\d+\.\d+)(?![\w.])")

# ★ W10b 項目2: 修復ループが行き詰まると『総当たりで色々試す』方向に暴走し、依頼と無関係な
#   操作（ヘルパの全呼び出し等）を追加することがあった実測（operator 第6回査定・3回連続
#   再現。詳細は _known_helper_names/detect_helper_sweep のコメント）。修復メッセージ自体に、
#   依頼の範囲を超えるな、という制約を明示的に加える（no-op/実行時エラーの2種のみ・
#   bad_signature/truncated は構造の修正だけなので対象外）。
_REPAIR_SCOPE_GUARD = "★ 元の依頼に無い操作(無関係なヘルパ呼び出しを含む)を追加してはいけない。"


def _looks_like_rate(value: float) -> bool:
    """0.05〜0.2（例: 消費税8%→0.08）または 1.05〜1.2（例: 税込計算の *1.1）の小数か。"""
    return (0.05 <= value <= 0.2) or (1.05 <= value <= 1.2)


def _rate_explained(value: float, task: str, vocab: dict) -> bool:
    """率らしい数値が依頼文/用語集のどこかで説明されているか（保守的＝広めに許して
       誤検知を避ける）。①コードの生の数字がそのまま依頼文にある ②%/％表記に換算した
       値が依頼文にある ③用語集の登録値が同じ率を指す、のいずれか。"""
    factor = value if value >= 1 else 1 + value
    frac = value if value < 1 else value - 1
    pct = frac * 100
    if any(lit in task for lit in {f"{value:g}", f"{factor:g}", f"{frac:g}"} if lit):
        return True
    if any(f"{p}%" in task or f"{p}％" in task for p in {f"{pct:g}", f"{pct:.0f}"}):
        return True
    for v in vocab.values():
        try:
            fv = float(v)
        except (TypeError, ValueError):
            continue
        if abs(fv - factor) < 1e-6:
            return True
    return False


def scan_rate_literals(code: str, task: str, vocab: dict | None = None) -> list:
    """生成 Basic コード中の『率らしい数値リテラル』のうち、依頼文にも用語集にも出典が
       無いものを、検算を促す助言のリストにする（同じ数値は1回だけ報告）。"""
    vocab = vocab or {}
    seen: set = set()
    out: list = []
    for line in code.splitlines():
        if _BASIC_COMMENT_LINE_RE.match(line):
            continue
        for m in _RATE_LITERAL_RE.finditer(line):
            raw = m.group(1)
            if raw in seen:
                continue
            value = float(raw)
            if not _looks_like_rate(value):
                continue
            if _rate_explained(value, task, vocab):
                continue
            seen.add(raw)
            out.append(f"★ 率らしい数値 ({raw}) が依頼に無いのに使われています — 検算してください")
    return out


# ---------------------------------------------------------------------------
# ★ W10b 項目1/2: 自由生成の関所とヘルパ総なめ検出
#   operator 第6回ブラインド査定（実機）で確定した致命: 自由生成(FREEFORM/OUT_OF_VOCAB)は
#   事後条件チェッカーが無い唯一の経路なのに、成功時は「⚠ 機械保証なし」のソフト警告だけで
#   無条件に適用されていた。修復ループが行き詰まると、依頼と無関係な操作（ヘルパの全呼び出し）
#   を書いて「何か効くかもしれない」を試す暴走（ヘルパ総なめ）が3回連続で再現した。
#   検証できないものは人に確認を返す（忠実度ゲート/破壊の関所と同じ思想）。
# ---------------------------------------------------------------------------

class _FreeformGateAbort(Exception):
    """★ W10b 項目1: run_freeform_plan_step の関所で人が拒否/非対話で確認できなかった時、
       cmd_run_plan まで一気に抜けるための内部シグナル（_confirm_overwrite_or_gate が
       cmd_run_plan の同じフレームで直接 exit code を return しているのと同じ扱いを、
       別関数をまたいでも実現する）。exit_code は呼び出し元がそのまま return すべき値。"""
    def __init__(self, exit_code: int):
        super().__init__(exit_code)
        self.exit_code = exit_code


def _confirm_freeform_apply(a: argparse.Namespace, sweep_warning: str | None = None,
                             step_prefix: str = "") -> int | None:
    """★ W10b 項目1: 自由生成の関所。DSL 経路（②検証→③確認→④codegen→⑤適用→⑥事後条件）
       と違い、自由生成は「変化したか」しか機械確認できず「正しいか」は検証できない唯一の
       経路（事後条件チェッカーが無い）。生成コードを見せた直後・適用の直前に必ず人に確認を
       返す。--allow-freeform が既に立っていれば承知の上として素通し。
       sweep_warning があれば y/N の直前に強調表示する（★ 項目2: ヘルパ総なめの疑い）。
       戻り値: 続行してよければ None、中断すべきなら呼び出し側がそのまま return すべき
       exit code（対話で拒否=1・非対話で確認できない=8）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。単発 FREEFORM では空文字。"""
    if getattr(a, "allow_freeform", False):
        return None
    if sweep_warning:
        print(f"{step_prefix}{sweep_warning}")
    try:
        ans = input(f"{step_prefix}このコードは機械検証できません。適用しますか？ [y/N]: ").strip().lower()
    except EOFError:
        for ln in render_retry_options(step_prefix, [
                ("--allow-freeform", "機械検証できないことを承知の上で適用する")]):
            print(ln)
        return 8
    if ans not in ("y", "yes"):
        print(render_aborted(step_prefix))
        return 1
    return None


_HELPER_SUB_DECL_RE = re.compile(r"^\s*Sub\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", re.M)
_HELPER_CALL_RE = re.compile(r"\bCall\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(")
# ★ 根拠: bench 再現実験（W10b・demo/expense_w10b.xlsx 相当のダミー表で「氏名の列を全部
#   『退職済み』に書き換えて」を実行）で、修復3回目のヘルパ総なめは異なるヘルパ7種
#   （AutoFitColumns/AlignCenter/FormatThousands/VLookupFromTable/PivotSum/SummaryTable/
#   StyleBold）を Call していた一方、正常な単一操作の自由生成は0〜1種しか呼ばない。
#   間の4を閾値に採用（operator の所見「単一操作でCallが4個以上」とも一致）。
_HELPER_SWEEP_THRESHOLD = 4


def _known_helper_names(helper_files: list) -> set:
    """helper_files（load_helpers が返す .bas ファイル一覧）から Sub 名を集める
       （ヘルパ総なめ検出の対象を『実在するヘルパ』だけに絞るため。ユーザー定義の
       Sub まで数えると通常の複数ヘルパ利用と区別できなくなる）。"""
    names: set = set()
    for f in helper_files:
        try:
            names |= set(_HELPER_SUB_DECL_RE.findall(f.read_text(encoding="utf-8")))
        except OSError:
            continue
    return names


def detect_helper_sweep(code: str, helper_names: set) -> str | None:
    """★ W10b 項目2: 自由生成コードが異なるヘルパを閾値(_HELPER_SWEEP_THRESHOLD)以上
       呼んでいたら、依頼と無関係な操作を全部 Call する『ヘルパ総なめ』の疑いを1行で
       返す（無ければ None）。ブロックはしない（誤検知の可能性は残る）— 関所の y/N の
       直前に強調表示し、人の判断を後押しする助言止まり。"""
    if not helper_names:
        return None
    called = set(_HELPER_CALL_RE.findall(code)) & helper_names
    if len(called) < _HELPER_SWEEP_THRESHOLD:
        return None
    names = "、".join(sorted(called))
    return (f"🚨 疑わしい: {len(called)} 種類のヘルパを呼んでいます（{names}）"
            "— 依頼と無関係な操作が混じっていないか確認してください")


def cmd_run_freeform(a: argparse.Namespace, book: Path, source_book: Path,
                      step: dict | None = None) -> int:
    """自由生成経路（従来の cmd_run 本体そのまま。M2a の助言つき）。
       ① 翻訳が CLARIFY にも DSL 語彙にも決まらなかった（FREEFORM・翻訳失敗）ときに使う。
       ★ W3: source_book は cmd_run が翻訳より前に正規化済み（--dry のときは book と同じ・
       正規化していない）。ここでは正規化をやり直さない。
       step は _translate_and_dispatch が振り分けに使った plan[0]（op="FREEFORM"/"OUT_OF_VOCAB"
       と、OUT_OF_VOCAB なら about）。★ K-1: 生成に入る前の通知の理由1行目に使うだけ
       （省略時=FREEFORM 扱い・後方互換）。"""
    refs_dir = Path(a.refs).resolve() if a.refs else DEFAULT_REFS
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    helper_catalog, helper_files = load_helpers(helpers_dir)
    known_helper_names = _known_helper_names(helper_files)   # ★ W10b 項目2: 総なめ検出用
    system = CONTRACT + load_refs(refs_dir) + helper_catalog
    desc = describe_book(book)
    user = f"{desc}\n\nタスク:\n{a.task}\n\n`Sub Run(oDoc As Object)` を1つだけ書け。コードのみ。"
    msgs = [{"role": "system", "content": system}, {"role": "user", "content": user}]

    # ★ W8a 項目5: 「自由生成経路」→「AI が直接作成（機械保証なし）」（operator の
    #   検品リストの語彙翻訳。内部の変数名・関数名・コメントは不変）。
    print(render_run_header("AI が直接作成・機械保証なし", a.model, book.name))
    print(f"■ 参照ライブラリ: {refs_dir}  ({len(list(refs_dir.glob('*.bas'))) if refs_dir.is_dir() else 0} 例)")
    print(f"■ ヘルパ: {helpers_dir}  ({len(helper_files)} 本を同梱・Call で呼ばせる)")
    # ★ K-1: 語彙外に落ちる瞬間の通知。生成が始まる前（★ の1行目より前）に、
    #   理由・費用・次の手を1ブロックで言う。通知だけ・同意の門(y/N)は作らない
    #   （関所は別に _confirm_freeform_apply が適用の直前に持つ・K-2 は意図的に保留）。
    step = step or {}
    reason = freeform_notice_reason(str(step.get("op") or "FREEFORM"), step.get("about") or "")
    print()
    for ln in render_freeform_notice(reason):
        print(ln)

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = book.with_name(book.stem + ".out" + book.suffix)

    apply_timeout = a.timeout if a.timeout else None   # 0 で無効化（旧挙動 = 無制限）

    before = None if a.dry else snapshot(source_book)
    vocab = load_vocab()   # ★ W8a 項目4: 率リテラルスキャンが「用語集に説明があるか」を見る

    result = {"ok": False, "attempts": 0, "task": a.task, "model": a.model,
              "path": "freeform", "command": None, "postcondition": None}
    failure_kind = "none"
    for attempt in range(a.repair + 1):
        result["attempts"] = attempt + 1
        t0 = progress_start(f"⏳ 生成中 ({a.model})…")
        raw = ollama_generate(a.model, msgs, temperature=a.temperature)
        progress_end(t0)
        code = extract_bas(raw)
        (workdir / f"attempt{attempt}.bas").write_text(code, encoding="utf-8")

        for ln in render_code_block(f"\n─ 試行 {attempt+1} ─ 生成した .bas ───────────────", code):
            print(ln)

        if not valid_signature(code):
            print("× 署名が違う（Sub Run(oDoc As Object) が無い）。修復する。")
            failure_kind = "bad_signature"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": "署名が違う。`Sub Run(oDoc As Object)` を1つだけ。コードのみ。"}]
            continue

        # ★ M2a: bad_signature とは別の分類。署名はあるが本体が途中で切れているケース。
        if is_truncated_code(code):
            print("× 生成コードが不完全（途中で切断）。修復する。")
            failure_kind = "truncated"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "コードが途中で切れている（End Sub まで書き切れていない）。"
                      "最初から完全なコードを1つだけ書いて。コードのみ。"}]
            continue

        if a.dry:
            print("\n（--dry: 適用しない。レビュー後に --dry を外して実行）")
            result["ok"] = True
            result["dry"] = True
            failure_kind = "none"
            break

        # ★ W10b 項目1: 自由生成の関所。コードは既に表示済み（上の attempt 表示）・
        #   適用の直前に必ず確認する（原本にはまだ何も触れていない）。
        sweep_warning = detect_helper_sweep(code, known_helper_names)
        gate_exit = _confirm_freeform_apply(a, sweep_warning)
        if gate_exit is not None:
            return gate_exit

        shutil.copy2(source_book, out_book)   # 原本は触らず、正規化済みコピーに適用
        t0 = progress_start("⏳ LibreOffice で適用中…")
        ok, err, rawout = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
        progress_end(t0)
        if not ok:
            # ★ M2a: obasync の生 Python トレースバックを端末にそのまま出さない。
            #   端末は最終行(例外名+メッセージ)だけ。修復ループへは従来どおり err の全文
            #   を渡す（モデルへの情報は減らさない）。全文は履歴 jsonl 側に残す。
            print(f"× 実行時エラー: {short_error_summary(err)}（詳細は履歴に記録）。修復する。")
            failure_kind = "runtime_error"
            result["last_error_full"] = err
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": f"実行時エラー: {err}\nこれを直して。{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        after = snapshot(out_book)
        changed, lines = diff_snapshots(before, after)
        if not changed:
            print("× no-op（実行は成功したが文書に変化が無い）。修復する。")
            failure_kind = "noop"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "実行は成功したが文書に一切変化が無かった（no-op）。"
                      "設定した API が効いていない可能性がある。別の正しい方法で書き直して。"
                      f"{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        # ★ W8a 項目4: 単段の FREEFORM/OUT_OF_VOCAB は、成功しても『機械検証済み』の
        #   ✓ ではなく複数段計画の語彙外段(format_plan_report)と同じ強度の正直な ⚠ 枠で
        #   表示する（実測: 8% 仮定やラベル貼りが「✓できました」で素通りしていた）。
        print("\n⚠ AI が直接作成した処理です（機械保証なし）— 確認してください。変更点:")
        for ln in lines:
            print(ln)
        # ★ 止血3: FREEFORM 経路は no-op ガード/advisories も snapshot() 頼みなので、
        #   検証自体が先頭1000行までしか見ていない（DSL経路より弱い正直さ）。
        notice = _truncation_notice(before, after, exhaustive_postcondition=False)
        if notice:
            print(notice)
        advisories = build_advisories(a.task, before, after,
                                       sheet_conflict=getattr(a, "_sheet_conflict", None))   # ★ 誤爆#3
        # ★ W8a 項目4: 率らしい数値リテラルの機械スキャン。★ 挙動変更#1(a): エラー値増加の網。
        advisories = advisories + scan_rate_literals(code, a.task, vocab) + formula_error_advisory(source_book, out_book, cell_ref=_cell_ref)
        for adv in advisories:
            print(adv)
        result["ok"] = True
        result["changes"] = lines
        result["advisories"] = advisories
        failure_kind = "none"
        # ★ W8b-2 項目1: 自由生成(FREEFORM/OUT_OF_VOCAB)は機械保証が無いので、既定(原本
        #   直接適用)でも trailing メッセージは ⚠「機械保証はありません」側を使う。
        #   ★ C9: それでも最終ファイルの読み戻し（今どうなっているか）は同じように行う。
        _finish_apply(a, book, out_book, workdir, result, machine_verified=False)
        break
    else:
        print(f"\n× {a.repair+1} 回試みたが達成できなかった。")

    _finish_run(a, book, result, failure_kind)
    return 0 if result["ok"] else 1


# ---------------------------------------------------------------------------
# M2c: 複合依頼の計画実行と正直な範囲表示
#   翻訳(①)が返した plan(長さ2以上)を段ごとに実行する。DSL 語彙の段は②〜⑥の決定論
#   パイプライン、語彙外(OUT_OF_VOCAB/FREEFORM)の段は FREEFORM 経路（その段の依頼文だけ）。
#   ★ 黙落ゼロ: 計画に載った段は必ず項目別報告の1行になる。
#   ★ 総合判定は最弱の段に従う。段別の行は evidence だけを述べ、✓ とは呼ばない（C9）。
#   ★ C5: _ITEM_STATUS_MARK / _VERIFY_SCOPE_NOTE(_PLAN) / format_plan_report /
#   overall_verdict は ailine_core/claim.py に移した（『✓』相当の文字列はレンダラ1箇所
#   からしか出さない・冒頭の import で ailine.format_plan_report 等は従来どおり使える）。
#   ★ C9: その ✓ の発生点自体を _finish_apply（原本が確定した後）へ動かした。
# ---------------------------------------------------------------------------

# col系 slot を持つ op → その slot 名（依存つき連鎖の新規列フォールバック対象）。
_COLUMN_ARG_KEYS = {
    "SORT": ("col",), "NUMBER_FORMAT": ("col",), "CHART": ("value_col",),
    "AGGREGATE": ("group_col", "value_col"),
}

# ★ 致命1(W10e): 依頼文に「見出し」の語があるのに BOLD/FILL_COLOR/CENTER_ALIGN の対象が
#   直前までの段が新規作成した列に解決された場合の食い違いを検出する。実測事故の再現形
#   そのもの ―― COMPUTE_COLUMN は target 無指定(新規列作成)の見出しを式そのまま
#   （例:「数量*単価」）で書く（codegen_dsl 参照）。「見出しを太字にして」の段が、
#   複合依頼の翻訳で前段の演算対象名を引きずり、target がその「数量*単価」という
#   *実在する*（が事務上は無関係な）新規列名に解決されてしまうと、check_bold は
#   「対象:col:数量*単価」という*計画どおり*の検証には合格する（が依頼＝見出し行の
#   太字化とは無関係）。★ 保守的: 対象列が『この計画で直前までに新規作成された列』
#   （new_cols）であり、かつ依頼文に「見出し」とある場合だけ発火する（列名を最初から
#   明示したふつうの BOLD/FILL_COLOR/CENTER_ALIGN では絶対に発火しない）。
#   ★ _apply_new_column_fallback 経由(target が未解決語から救済された場合)・素の
#   verify_dsl_args 直解決(target が最初から新規列の実名と一致した場合)のどちらでも
#   同じ理屈で疑わしいため、fallback の発火有無ではなく new_cols 所属の有無で判定する
#   （fallback 有無だけを見ると後者を取りこぼす＝実測の再現形はむしろ後者に近い）。
_HEADER_WORD_RE = re.compile(r"見出し")


def _maybe_warn_header_col_mismatch(op: str, resolved: dict, new_cols: list, task: str) -> str | None:
    """★ 致命1: 上のコメント参照。target がこの計画の直前までに新規作成された列
       （new_cols）を指していて、かつ依頼文に「見出し」とある場合だけ、非ブロッキングの
       助言を返す（M2a と同じ思想＝保守的・ブロックしない・確認を促すだけ）。
       ★ 単位B: 一般則（③）が同じスロットで鳴る時は、この文は単独では出ず、③の ⚠ に
       「（この計画の直前の段で新規作成された列）」として畳み込まれる（dsl_step.py 参照）。
       事実の文言は NEW_COLUMN_ORIGIN 1箇所が持つ ―― 畳み込み側と文面がずれないため。"""
    if op not in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        return None
    target = resolved.get("target", "")
    if not (isinstance(target, str) and target.startswith("col:")):
        return None
    col_name = target[4:]
    if col_name not in (new_cols or []):
        return None
    if not _HEADER_WORD_RE.search(task or ""):
        return None
    return (f"⚠ 対象の列『{col_name}』は{NEW_COLUMN_ORIGIN}です。"
            "依頼に「見出し」とあるため、見出し行（行全体）を意図していないか確認してください")


def _apply_new_column_fallback(op: str, args: dict, headers: list, new_cols: list) -> dict:
    """★ M2c 依存つき連鎖（battery v2 #107 型）: 直前までの段が新規作成した列がちょうど
       1つあり、この段の列参照が現在の実列名のどれとも一致しない場合、その新規列を指して
       いるとみなして args を書き換える（候補が0か2つ以上なら何もしない＝保守的）。
       ★ 書き換えても最終的には verify_dsl_args が実在確認するので、誤った書き換えは
       通常どおりのエラーで止まる（無条件に信じ切らない）。"""
    if len(new_cols) != 1:
        return args
    only = new_cols[0]
    patched = dict(args)
    for k in _COLUMN_ARG_KEYS.get(op, ()):
        v = patched.get(k)
        if isinstance(v, str) and v not in headers and not re.fullmatch(r"\d+", v):
            patched[k] = only
    if op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        t = patched.get("target", "")
        if isinstance(t, str) and t.startswith("col:"):
            name = t[4:]
            if name not in headers and not re.fullmatch(r"\d+", name):
                patched["target"] = f"col:{only}"
    return patched


def run_freeform_plan_step(a: argparse.Namespace, task_text: str, out_book: Path, workdir: Path,
                            refs_dir: Path, helpers_dir: Path, tag: str,
                            apply_timeout: float | None, step_prefix: str = "",
                            vocab: dict | None = None, op: str = "FREEFORM",
                            about: str = "") -> tuple:
    """M2c: 複合計画の語彙外(OUT_OF_VOCAB/FREEFORM)段を FREEFORM 経路で実行する。
       cmd_run_freeform と同じ生成→（★ W10b: 関所→）適用→署名/切断/no-op チェックのループを、
       『その段の依頼文だけ』かつ『out_book の現在の状態』を起点に行う版。
       ★ cmd_run_freeform 本体は変えない（既存の回帰リスクを避けるため意図的に複製する）。
       ★ W10b 項目1: 関所で人が拒否/非対話で確認できなかった場合は _FreeformGateAbort を
       投げて cmd_run_plan まで一気に抜ける（破壊の関所と同じ『計画全体を止める』扱い。
       原本(book)はこの時点でまだ一切触れていない＝out_book はコピーなので安全）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。
       op/about は呼び出し側(cmd_run_plan)の step から渡す。★ K-1: 生成に入る前の通知の
       理由1行目に使うだけ（OUT_OF_VOCAB のときだけ about を添える）。
       ★ W10f 項目2: vocab（率リテラル走査用の用語集）は cmd_run_freeform（単発の自由生成）
       には元から渡っていた(scan_rate_literals)が、複合計画のこの経路には渡っておらず
       A' 原則（LLM に率や値を確定させない）の機械監査が複合計画の自由生成段だけ
       素通りしていた（独立監査が発見・査定5回は偶然当たらなかっただけの穴）。ここで
       同じ走査を通す。
       戻り値: (ok, changes:list[str], advisories:list[str], failure_kind:str, detail:str|None)"""
    helper_catalog, helper_files = load_helpers(helpers_dir)
    known_helper_names = _known_helper_names(helper_files)   # ★ W10b 項目2: 総なめ検出用
    system = CONTRACT + load_refs(refs_dir) + helper_catalog
    desc = describe_book(out_book)
    user = f"{desc}\n\nタスク:\n{task_text}\n\n`Sub Run(oDoc As Object)` を1つだけ書け。コードのみ。"
    msgs = [{"role": "system", "content": system}, {"role": "user", "content": user}]

    # ★ K-1: 語彙外に落ちる瞬間の通知（この段の生成が始まる前・1行に畳んだ版）。
    reason = freeform_notice_reason(op, about)
    print(render_freeform_notice_compact(reason, step_prefix=step_prefix))

    stepsource = workdir / f"{tag}_source{out_book.suffix}"
    shutil.copy2(out_book, stepsource)
    before = snapshot(stepsource)

    failure_kind = "none"
    for attempt in range(a.repair + 1):
        t0 = progress_start(f"⏳ 生成中（語彙外段・{a.model}）…")
        raw = ollama_generate(a.model, msgs, temperature=a.temperature)
        progress_end(t0)
        code = extract_bas(raw)
        (workdir / f"{tag}_attempt{attempt}.bas").write_text(code, encoding="utf-8")

        if not valid_signature(code):
            failure_kind = "bad_signature"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": "署名が違う。`Sub Run(oDoc As Object)` を1つだけ。コードのみ。"}]
            continue
        if is_truncated_code(code):
            failure_kind = "truncated"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "コードが途中で切れている（End Sub まで書き切れていない）。"
                      "最初から完全なコードを1つだけ書いて。コードのみ。"}]
            continue

        # ★ W10b 項目1: 自由生成の関所。単発 cmd_run_freeform と違い、この経路はこれまで
        #   生成コードを一切表示していなかった（黙って確認を求めても判断できない）ので、
        #   ここで初めて表示してから y/N を聞く。
        for ln in render_code_block(f"{step_prefix}─ 生成した .bas（語彙外・AI が直接作成）───────────────",
                                     code, step_prefix=step_prefix):
            print(ln)
        sweep_warning = detect_helper_sweep(code, known_helper_names)
        gate_exit = _confirm_freeform_apply(a, sweep_warning, step_prefix=step_prefix)
        if gate_exit is not None:
            raise _FreeformGateAbort(gate_exit)

        shutil.copy2(stepsource, out_book)
        t0 = progress_start("⏳ LibreOffice で適用中…")
        ok, err, _raw = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
        progress_end(t0)
        if not ok:
            failure_kind = "runtime_error"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": f"実行時エラー: {err}\nこれを直して。{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        after = snapshot(out_book)
        changed, lines = diff_snapshots(before, after)
        if not changed:
            failure_kind = "noop"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "実行は成功したが文書に一切変化が無かった（no-op）。"
                      "設定した API が効いていない可能性がある。別の正しい方法で書き直して。"
                      f"{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        advisories = build_advisories(task_text, before, after,
                                       sheet_conflict=getattr(a, "_sheet_conflict", None))   # ★ 誤爆#3
        # ★ W10f 項目2: 単発 cmd_run_freeform と同じ率リテラルの機械スキャン。この段の
        #   依頼文(task_text)だけを出典として見る（他段の依頼文言に混ざらないよう局所判定）。
        advisories = advisories + scan_rate_literals(code, task_text, vocab) + formula_error_advisory(stepsource, out_book, cell_ref=_cell_ref)   # ★ 挙動変更#1(a)
        # ★ 止血3: 呼び出し元(cmd_run_plan)は lines をそのまま「  {ln}」で表示するだけ
        #   なので、切り詰め注記はここで lines に混ぜて渡す。
        notice = _truncation_notice(before, after, exhaustive_postcondition=False)
        if notice:
            lines = lines + [notice]
        return True, lines, advisories, "none", None

    detail = {
        "bad_signature": "生成コードの署名が不正でした",
        "truncated": "生成コードが途中で切断されました",
        "runtime_error": "実行時エラーが解消しませんでした",
        "noop": "適用しても文書に変化がありませんでした",
    }.get(failure_kind, "原因不明で失敗しました")
    return False, [], [], failure_kind, detail


# ★ C5: format_plan_report / overall_verdict は ailine_core/claim.py に移した
#   （Claim を経由してしか『✓ 機械検証済み』相当の文字列を出さないレンダラ側に統合・
#   冒頭の import で ailine.format_plan_report / ailine.overall_verdict は従来どおり使える）。


# ★ W10d【本命】: 複合計画は段が増えるほど同じ助言（例: 幽霊データ検出）が段の数だけ
#   並びうる。査定で名指しされたオオカミ少年化を避けるため、文言が同じものは1行に畳んで
#   段番号を列挙する（初出順は保つ・文言そのものは変えない＝読み手が信じる根拠を削らない）。
def _group_step_advisories(entries: list) -> list:
    """entries: [(段番号 or None, 助言文言), ...]（None=計画全体に対する助言・段に紐付かない）
       を文言でグルーピングし、初出順を保った [(段番号のリスト, 文言), ...] を返す。
       印字用(_dedup_step_advisories)と --json 用(cmd_run_plan の result["advisories"])が
       同じグルーピングを共有する（二重管理しない）。"""
    order = []
    by_text: dict = {}
    for idx, text in entries:
        if text not in by_text:
            by_text[text] = []
            order.append(text)
        if idx is not None and idx not in by_text[text]:
            by_text[text].append(idx)
    return [(by_text[text], text) for text in order]


def _dedup_step_advisories(entries: list) -> list:
    """複合計画の全段から集めた助言を、印字用の行リストに整形する（同じ文言は1行に畳み、
       該当する段番号を列挙する）。"""
    lines = []
    for idxs, text in _group_step_advisories(entries):
        if idxs:
            prefix = "・".join(f"{i}段目" for i in idxs)
            lines.append(f"  {prefix}: {text}")
        else:
            lines.append(f"  {text}")
    return lines


def _run_dsl_plan_step(i: int, op: str, raw_args: dict, *, task: str, current_meta: dict, original_headers: dict,
                        first_sheet: str | None, out_book: Path, workdir: Path, helper_files, apply_timeout,
                        use_formula: bool, header_rows: dict, before_charts: int, a: argparse.Namespace, vocab: dict,
                        book_name: str, subject_sink: dict | None = None) -> tuple:
    """★ C7: cmd_run_plan の DSL 語彙段の1段分。cmd_run_dsl と同じ ailine_core.dsl_step の共有エンジンを通る
       （非対称は dsl_step.py 参照）。この分離で stage_organs の dsl_plan_step 代表関数はここになる（DoD7）。
       戻り値: (gate_exit, item, plan_json_entry, step_advisories, provenance_entry, mention_exclude_sheets, current_meta)。
       ★ 単位E: subject_sink（呼び出し側が用意する dict）に、この段の対象スロットの出所を積む
       ―― ③ の有無は計画全体の ✓ を左右し、② は ✓ の直後の1文になるので、段の外へ運ぶ必要がある
       （戻り値のタプルはこれ以上広げない ―― 既存の unpack を壊さないための選択）。"""
    step_prefix = f"  {i}段目: "
    deps = _make_dsl_step_deps()
    # 依存つき連鎖: 直前までの段の適用後の実列構成(current_meta)で接地する（新規列フォールバック込み）
    ground = resolve_dsl_step_args(op, raw_args, task, current_meta, vocab,
                                    original_headers=original_headers, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        return (None, (i, f"操作:{OP_LABELS.get(op, op)}", "fail", ground.err),
                {"op": op, "command": None, "status": "fail", "postcondition": None},
                [], None, None, current_meta)
    resolved, inferred = ground.resolved, ground.inferred

    # ★ 致命1(W10e) 要求2: 単発は元々この行を適用前に出す・複合計画は抜けていた（段番号付きで見せる）。
    confirm = print_dsl_confirmation(op, resolved, inferred, task, meta=current_meta, warn_book=out_book,
                                      new_cols=ground.new_cols, a=a, deps=deps, step_prefix=step_prefix)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所（複合計画の段ごと）
        return (confirm.gate_exit, None, None, [], None, None, current_meta)

    if subject_sink is not None:   # ★ 単位E: この段の対象スロットの出所を計画全体へ運ぶ
        subject_sink.setdefault("warnings", []).extend(confirm.subject_warnings)
        for phrase in confirm.unspoken:
            if phrase not in subject_sink.setdefault("unspoken", []):
                subject_sink["unspoken"].append(phrase)
    step_advisories = [confirm.mismatch_warning] if confirm.mismatch_warning else []
    # ★ 段1: interpretation/provenance は1箇所（build_interpretation）で組む（cmd_run_dsl と同じ）。
    #   provenance_entry の中身（キー・値）は今までと完全に同じ（resolved["_sources"] のまま）。
    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book_name])
    provenance_entry = {"step": i, **provenance} if provenance else None

    # ★ 挙動変更#2: cmd_run_dsl と同じ理由（コメント参照）で、この段の「本当の」対象シート
    #   (resolved["_target_sheet"]) を優先する（LOOKUP_FILL 段が計画全体の対象シートと
    #   異なる参照シートを持つ場合の header_row 取り違えを避ける）。
    step_target_sheet = resolved.get("_target_sheet") or first_sheet
    step_header_row = current_meta.get("header_rows", {}).get(step_target_sheet, 1) if step_target_sheet else 1
    code = codegen_dsl(op, resolved, current_meta, use_formula=use_formula)
    (workdir / f"plan_step{i}.bas").write_text(code, encoding="utf-8")

    # ★ W9: INSERT_ROWS/AUTOFIT の事後条件が段ごとの before/after を突き合わせられるようコピー。
    stepsource = workdir / f"plan_step{i}_source{out_book.suffix}"
    shutil.copy2(out_book, stepsource)
    step_before = snapshot(stepsource)   # ★ W10d: 助言計算用（この段の適用直前）

    apply_result = apply_dsl_step(
        op, resolved, code, apply_target=out_book, before=step_before, before_charts=before_charts,
        workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout,
        header_row=step_header_row, use_formula=use_formula, source_book=stepsource, deps=deps,
        apply_progress_label=f"⏳ {i}段目 LibreOffice で適用中…", print_changes=False)

    if apply_result.runtime_error is not None:
        detail = f"実行時エラー: {short_error_summary(apply_result.runtime_error)}"
        return (None, (i, confirm.label, "fail", detail),
                {"op": op, "command": confirm.line, "status": "fail", "postcondition": None,
                 "interpretation": interpretation},
                step_advisories, provenance_entry, None, current_meta)

    # ★ W10d【本命】: mode="structural"（依頼文言との重なり④は呼び出し側が全体で1回評価・dsl_step.py 参照）。
    step_after = apply_result.after
    # ★ C9: 複合計画の DSL 段だけ _truncation_notice が一度も呼ばれていなかった既知の穴
    #   （stage_organs の dsl_plan_step × truncation_notice が None と宣言していた場所）。
    #   ✓ の意味を「読み戻して確かめた」にする以上、「先頭 MAX_ROWS 行しか見ていない」は
    #   ✓ の主張範囲に直接効くので同じ回で塞ぐ。
    step_notice = _truncation_notice(step_before, step_after, exhaustive_postcondition=True)
    if step_notice:
        print(f"{step_prefix}{step_notice}")
    # ★ 単位C(D8): ここも `if op == "LOOKUP_FILL"` のハードコードを宣言読み取りへ置き換えた。
    #   複合計画は④を計画全体で1回だけ評価するので、段ごとの「読むだけのシート」を返して
    #   呼び出し側が全段分を合算する（返り値は集合・空集合なら足すものが無いだけ）。
    mention_exclude_sheets = _declared_reads_only_sheets(op, resolved)
    # ★★ 単位F: 単発(cmd_run_dsl)と同じ位置・同じ関所。段ごとの before/after で見る
    #   （計画全体の before/after で見ると、前の段が既存行を書いたことを次の段の前提破れと
    #   読み違える ―― 宣言は段ごとに違うので、突き合わせも段ごとでなければならない）。
    # ★★ 単位G(複合計画版): 中立化（「（既存シート『集計』の更新は意図どおりです）」）は
    #   助言の発生源で決まるので、単発(5517-5522)と同じく advisories を組む「前」に検査する
    #   ―― ここが単発と非対称のまま(precondition_broken 未配線)だと、破れた宣言でも
    #   肯定文が出てしまう（単位G の保証が複合計画に配線されていなかった不具合）。
    #   ★ 検査は段あたり1度だけ（単位C の教訓）。結果は advisories と下の関所の両方で使い回す。
    _precondition = _maybe_warn_write_precondition(op, step_before, step_after, resolved)
    precondition_broken = _precondition[0] if _precondition else None
    # ★★ 単位H 開示(複合計画版): 単発と同じく関所判定の直後（置き場所は単発と対称）。
    for own_notice in _maybe_own_prior_output_notice(op, step_before, step_after, resolved):
        print(f"{step_prefix}{own_notice}")
    step_advisories.extend(compose_dsl_step_advisories(
        "structural", op, resolved, current_meta, task, step_before, step_after, deps=deps,
        precondition_broken=precondition_broken) + formula_error_advisory(stepsource, out_book, cell_ref=_cell_ref))   # ★ 挙動変更#1(a)

    status, reason = apply_result.postcondition_status, apply_result.postcondition_reason
    # ★ 止血1/2: "error"→fail 扱い。"warn"(検証対象不足)は成功は名乗るが機械検証済みとは言わない。
    if status in ("fail", "error"):
        return (None, (i, confirm.label, "fail", reason),
                {"op": op, "command": confirm.line, "status": "fail", "postcondition": "fail",
                 "interpretation": interpretation},
                step_advisories, provenance_entry, mention_exclude_sheets, current_meta)
    item_status = "warn" if status == "warn" else "ok"
    warn_precondition = _precondition[1] if _precondition else None   # ★ 単位G: 上で 1 度だけ検査済み
    if warn_precondition:
        print(f"{step_prefix}{warn_precondition}")
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition, step_prefix=step_prefix)
        if gate_exit is not None:
            return (gate_exit, None, None, [], None, None, current_meta)
    return (None, (i, confirm.label, item_status, reason),
            {"op": op, "command": confirm.line, "status": item_status, "postcondition": status,
             "interpretation": interpretation},
            step_advisories, provenance_entry, mention_exclude_sheets,
            build_book_meta(out_book, header_rows=header_rows))


def _preview_dsl_plan(a: argparse.Namespace, plan: list, book_meta: dict, vocab: dict, book_name: str) -> list:
    """★ C7: cmd_run_plan の --dry プレビュー（未実行・印字のみ）。--json 用 plan_json を返す。
       分離により cmd_run_plan 自身は verify_dsl_args を直接呼ばなくなる（DoD7 の材料）。"""
    preview_items, plan_json = [], []
    for i, step in enumerate(plan, 1):
        op = step.get("op")
        if op == "CLARIFY":
            q = step.get("question") or "確認が必要です"
            preview_items.append((i, q, "fail", "計画の途中で確認が必要なため対応できません"))
            plan_json.append({"op": "CLARIFY", "command": None, "status": "fail", "postcondition": None})
        elif op not in OP_SCHEMA:
            about = step.get("about") or "内容不明の依頼"
            preview_items.append((i, about, "warn", None))
            plan_json.append({"op": op, "command": about, "status": "warn", "postcondition": None})
        else:
            # ★ 挙動変更#2: --dry プレビューも a._target_sheet を読む（実行と同じ解決）。
            ok_v, resolved, inferred, err = verify_dsl_args(
                op, step.get("args", {}), book_meta, task=a.task, vocab=vocab,
                target_sheet=getattr(a, "_target_sheet", None))
            if ok_v:
                # ★ 挙動変更#3: プレビューの宣言テキストにも対象シートを載せる（実行時の
                #   「解釈:」行と同じ形にする・1枚のブックでは従来どおり付かない）。
                label = format_confirmation_line(op, resolved, inferred, sheets=book_meta.get("sheets"),
                                                 target_sheet=resolved.get("_target_sheet"))[len("解釈: "):]
                preview_items.append((i, label, "ok", None))
                # ★ 段1: プレビューにも interpretation を足す（--dry でも解釈は決まっている）。
                #   単発/実行と同じ組み立て（build_interpretation）を使う ―― 判定は
                #   classify_subject_provenance を1回だけ呼ぶ（単発は print_dsl_confirmation
                #   の内側で済ませているが、プレビューはそれを呼ばないのでここで直接呼ぶ）。
                verdicts = classify_subject_provenance(op, resolved, book_meta, a.task, a)
                interpretation, _provenance = build_interpretation(op, resolved, inferred, verdicts, [book_name])
                plan_json.append({"op": op, "command": label, "status": "ok", "postcondition": None,
                                   "interpretation": interpretation})
            else:
                preview_items.append((i, f"操作:{OP_LABELS.get(op, op)}", "fail", err))
                plan_json.append({"op": op, "command": None, "status": "fail", "postcondition": None})
    # ★ C9: --dry はプレビュー専用レンダラを使う（実行経路と同じ format_plan_report に
    #   status="ok" を流し込んでいたことが「未実行なのに ✓」の直接の原因だった）。
    for ln in format_plan_preview(preview_items):
        print(ln)
    return plan_json


def cmd_run_plan(a: argparse.Namespace, book: Path, source_book: Path, book_meta: dict, plan: list) -> int:
    """M2c: 複合依頼の計画実行本体。段ごとに②検証→③確認→④codegen→⑤適用→⑥事後条件
       （DSL 語彙の段。_run_dsl_plan_step に委譲）または FREEFORM（語彙外の段・依頼文だけを
       渡す）を順に実行し、★ 項目別の honest な報告を出す。総合判定は最弱の段に従う。
       ★ 依存つき連鎖: 各段の接地は直前段を適用した後の列構成(current_meta)で行い、列名
       不一致時は _apply_new_column_fallback が新規列への参照とみなし1回だけ書き換えを試みる。
       ★ W3: header_rows は計画全体を通して不変なので current_meta 再読み込みでも同じ値を渡す。"""
    print(render_run_header(f"複合計画・{len(plan)} 段", a.model, book.name))

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = book.with_name(book.stem + ".out" + book.suffix)
    apply_timeout = a.timeout if a.timeout else None
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    refs_dir = Path(a.refs).resolve() if a.refs else DEFAULT_REFS
    _helper_catalog, helper_files = load_helpers(helpers_dir)
    header_rows = book_meta.get("header_rows", {})
    use_formula, vocab = not getattr(a, "values", False), load_vocab()

    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              "path": "plan", "command": None, "postcondition": None}
    if a.dry:
        print("\n（--dry プレビュー・語彙外の段は実行時に AI が直接作成（機械保証なし）で対応します。未実行）")
        plan_json = _preview_dsl_plan(a, plan, book_meta, vocab, book.name)
        print("\n（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        result["plan"] = plan_json
        _finish_run(a, book, result, "none")
        return 0

    shutil.copy2(source_book, out_book)

    original_headers = {k: list(v) for k, v in book_meta["headers"].items()}
    # ★ 挙動変更#2: 対象シートは _cmd_run_dispatch が resolve_target_sheet で一箇所だけ
    #   決めた a._target_sheet を読む（後方互換フォールバックは1枚目・旧挙動と同一）。
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    before_all = snapshot(out_book)
    before_charts = before_all["charts"]

    current_meta = book_meta
    items: list = []         # (idx, label, status, detail)
    plan_json: list = []     # --json 用（既存キー不変・新規追加）
    plan_provenance: list = []   # ★ A': 段ごとの倍率等の出典（history.jsonl 用）
    step_advisory_entries: list = []   # ★ W10d: [(段番号 or None, 助言文言), ...]
    mention_exclude_sheets: set = set()   # ★ W10d/単位C: 参照専用シート（reads_only 宣言・全段分の合算）
    subject_sink: dict = {"warnings": [], "unspoken": []}   # ★ 単位E: 対象スロットの出所（全段分）

    for i, step in enumerate(plan, 1):
        op = step.get("op")

        if op == "CLARIFY":
            question = step.get("question") or "確認が必要です"
            items.append((i, question, "fail", "計画の途中で確認が必要なため対応できません"))
            plan_json.append({"op": "CLARIFY", "command": None, "status": "fail", "postcondition": None})
            continue

        if op not in OP_SCHEMA:
            about = step.get("about") or "内容不明の依頼"
            # ★ W10b 項目1: 関所で拒否/非対話なら計画全体をここで止める（out_book はコピーなので安全）。
            try:
                okf, changes, advisories, _fkind, detail = run_freeform_plan_step(
                    a, about, out_book, workdir, refs_dir, helpers_dir, f"plan{i}", apply_timeout,
                    step_prefix=f"  {i}段目: ", vocab=vocab, op=op, about=about)
            except _FreeformGateAbort as e:
                return e.exit_code
            if okf:
                items.append((i, about, "warn", None))
                for ln in changes:
                    print(f"  {ln}")
                step_advisory_entries.extend((i, adv) for adv in advisories)   # ★ W10d: ループ後に重複を畳む
            else:
                items.append((i, about, "fail", detail))
            plan_json.append({"op": op, "command": about,
                               "status": "ok" if okf else "fail", "postcondition": None})
            current_meta = build_book_meta(out_book, header_rows=header_rows)
            continue

        # ★ C7: DSL 語彙段。cmd_run_dsl と同じ ailine_core.dsl_step を通る _run_dsl_plan_step に委譲。
        gate_exit, item, plan_json_entry, step_adv, prov_entry, step_reads_only, current_meta = \
            _run_dsl_plan_step(
                i, op, step.get("args", {}), task=a.task, current_meta=current_meta,
                original_headers=original_headers, first_sheet=first_sheet, out_book=out_book,
                workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout,
                use_formula=use_formula, header_rows=header_rows, before_charts=before_charts, a=a, vocab=vocab,
                book_name=book.name, subject_sink=subject_sink)
        if gate_exit is not None:
            return gate_exit
        if item is not None:
            items.append(item)
        if plan_json_entry is not None:
            plan_json.append(plan_json_entry)
        if step_adv:
            step_advisory_entries.extend((i, adv) for adv in step_adv)
        if prov_entry is not None:
            plan_provenance.append(prov_entry)
        if step_reads_only:
            mention_exclude_sheets.update(step_reads_only)

    # ★ W10d【本命】: 依頼文言との重なり④は計画全体に対して1回だけ評価（他段が担当する言及の
    #   誤検知を避ける）。exclude_sheets は全段の「読むだけのシート」(reads_only 宣言)を合算する。
    # ★ 誤爆#3: 対象シートを決めた側が「曖昧なので既定へ後退した」と記録した語も同じ和に足す
    #   （単発の build_advisories が中でやっているのと同じ和・ここは④を直接呼ぶ経路）。
    mention_exclude_sheets |= conflict_excluded_sheets(getattr(a, "_sheet_conflict", None))
    after_all = snapshot(out_book)
    mentions = extract_task_mentions(a.task, before_all["sheets"])
    final_mention_lines = mention_overlap_advisory(
        mentions, before_all, after_all, mention_exclude_sheets or None)
    step_advisory_entries.extend((None, ln) for ln in final_mention_lines)   # None=計画全体

    print()
    for ln in format_plan_report(items):
        print(ln)
    dedup_advisories = _dedup_step_advisories(step_advisory_entries)   # ★ W10d: 重複を畳む
    if dedup_advisories:
        print("\n助言:")
        for ln in dedup_advisories:
            print(ln)
    verdict_line, verdict = overall_verdict(items)
    # ★ C9: 全段 ok のときは verdict_line が None（『✓ すべて機械検証済み』は廃止し、原本が
    #   確定した後の1行＝_finish_apply に移した）。
    # ★ 単位E: 常時の範囲注記はここからも消えた（発火率100%＝情報量ゼロ）。範囲を明示する
    #   役割は、②の run 固有の1文（✓ の直後・render_scope_notes）が引き継ぐ。
    if verdict_line:
        print(f"\n{verdict_line}")

    _changed, difflines = diff_snapshots(before_all, after_all)
    result["plan"] = plan_json
    result["provenance"] = plan_provenance or None
    result["items"] = [{"idx": idx, "label": label, "status": st, "detail": det}
                        for idx, label, st, det in items]
    result["changes"] = difflines
    result["advisories"] = [{"steps": idxs, "text": text}
                             for idxs, text in _group_step_advisories(step_advisory_entries)]

    if verdict == "fail":
        # ★ C9: 査定が名指しした沈黙 ―― 全段破棄で終わるのに原本がどうなったかを言って
        #   いなかった（.out が黙って隣に残ることも無言だった）。
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "plan_step_failed")
        return 1

    result["ok"] = True
    # ★ W8b-2 項目1: 複合計画は総合判定(overall_verdict)に従う。全段機械検証済み(ok)
    #   の時だけ ✓、語彙外/検証不足の段が混じる(warn)なら ⚠「機械保証はありません」側
    #   （自由生成の段が混じっている以上、全体としても機械保証済みとは名乗れない）。
    # ★★ 単位E: ③（依頼文の語と矛盾する対象）を含む段が1つでもあれば ✓ は出さない
    #   ―― 各段の事後条件が「計画どおり」に通っていても、その計画は依頼文と食い違っている。
    _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(verdict == "ok" and not subject_sink["warnings"]),
                   scope="; ".join(label for _idx, label, _st, _det in items),
                   scope_note="\n".join(render_scope_notes(subject_sink["unspoken"])))

    _finish_run(a, book, result, "none")
    return 0


def cmd_stop(a: argparse.Namespace) -> int:
    subprocess.run([sys.executable, str(basrun_path()), "stop"], encoding="utf-8", errors="replace")
    return 0


def cmd_scan(a: argparse.Namespace) -> int:
    """`ailine scan <folder>`: M1読み ── 複数ブックの棚卸し（書き込みゼロ・LO は起動しない）。
       DESIGN-20260821-multifile.md v2 §1(M1読み)・§2(骨)。分類・照合は ailine_core/multifile.py
       に置く。見出し行の推定だけ既存の detect_header_row/_row_char_stats を基準ファイル1冊分
       だけ呼ぶ（このモジュールが持つ元々の機能をそのまま流用・LO 往復は無い）。"""
    folder = Path(a.folder).resolve()
    candidates, excluded = multifile.classify_folder_contents(folder)
    base_path, base_wb = multifile.open_base_workbook(candidates)
    base_headers, base_sheet, header_row, value_col_name = [], None, 1, None
    if base_wb is not None:
        base_sheet = base_wb.sheetnames[0]
        ws = base_wb[base_sheet]
        scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
        rows = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
        row, confident = detect_header_row({"rows": rows})
        header_row = row if confident else 1
        base_headers = multifile.read_row_headers(ws, header_row)
        value_col = multifile.numeric_value_column(ws, header_row, len(base_headers) or MAX_COLS)
        value_col_name = base_headers[value_col - 1] if value_col else None
        base_wb.close()
    files = [multifile.evaluate_file(p, base_headers, base_sheet, header_row, value_col_name)
             for p in candidates]
    result = {"denominator": len(candidates), "base": base_path.name if base_path else None,
              "files": files, "excluded": excluded}
    if a.json:
        print(json.dumps(result, ensure_ascii=False))
        return 0
    for ln in render_scan_report(str(folder), result):
        print(ln)
    return 0


def _peek_headers(path: Path) -> list | None:
    """先頭シートの1行目をヘッダーとして覗き見る（読めなければ None）。
       stack の自己参照除外・関所（署名判定）専用の軽い読み。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    try:
        return multifile.read_row_headers(wb[wb.sheetnames[0]], 1)
    except Exception:
        return None
    finally:
        wb.close()


def _stack_json(result: dict) -> dict:
    """--json 契約（検体で凍結済み）: denominator/stacked_files/rows_written/files/sums/
       mismatches。★ jisaku-review#4: mismatches はテキストの ⚠ と同じ情報を機械可読で
       （ファイルごとの入れ子ではなく、[{file, row, excluded_value, adopted_sum}, ...] の
       平らな形 ── 自動化経路が総なめできるように）。
       ★ P2: sheet_fallbacks（[{name, wanted, used}, ...]）も追加 ── 基準名のシートが無く
       1枚目へ落ちたファイルの機械可読の開示（人間向けは render_stack_report）。"""
    mismatches = [{"file": entry["name"], "row": m["row"],
                   "excluded_value": m["excluded_value"], "adopted_sum": m["adopted_sum"]}
                  for entry in result.get("mismatches", ()) for m in entry["rows"]]
    return {"denominator": result["denominator"], "stacked_files": result["stacked_files"],
            "rows_written": result["rows_written"], "files": result["files"],
            "sums": result["sums"], "mismatches": mismatches,
            "sheet_fallbacks": result.get("sheet_fallbacks", [])}


def _stack_postcondition_fail(label: str, expected, actual) -> int:
    """事後条件①②が破れた時の唯一の出口。★ tmp_out は移さない（out は無傷のまま）。"""
    print(f"⚠ 事後条件が破れた: {label}  元(採用時) {multifile_stack.fmt_num(expected)} / "
          f"出力(書いた直後) {multifile_stack.fmt_num(actual)}")
    return 5


def _run_folder_refuse(op: str, plan_len: int) -> int:
    """M2 の断り（E11）: フォルダに未対応の依頼を**名指し**で断り、次の手を添えて exit 3。
       ★ 黙って1冊目に適用が最悪の形 ── 断る時は原本に一切触れない（この時点で
       書き込みは一度も起きていない）。対応の可否は OP_META の folder 宣言が唯一の出どころ
       （手書きの対応表を別に持たない ── ずれる表を作らない）。"""
    label = OP_META.get(op, {}).get("label") or (op if op else "この依頼")
    if plan_len > 1:
        print(f"？ 複数の操作をまとめた依頼（{plan_len} 段）はフォルダに対応していません。"
              "フォルダに頼めるのは抽出だけです（1 冊を指定すればそのまま頼めます）。")
    else:
        print(f"？ {label}は複数ファイル（フォルダ）に対応していません。"
              f"フォルダに頼めるのは抽出だけです（1 冊を指定すれば『{label}』も頼めます）。")
    print("  （頼める操作の一覧: ailine ops）")
    return 3


def _own_extract_output_status(path: Path, col: str, cmp: str, value) -> tuple:
    """path が①ailine 産か（mark）②M2 抽出の自分の前回出力で、かつ焼いた条件が
       今回と完全一致するか（same_condition）を返す。
       ★ review3#1 critical の直し: 『黙って作り直してよい』の根拠を**印だけ**にしない ──
       印は同じでも条件（列/比較/値）が違う前回出力を、条件を見ずに上書きして消していた
       （実機再現: 長いフォルダ名で切り詰めが起きると別条件が同名に潰れる）。
       同じ判定をここ1箇所に集約し、preflight（40冊読む前）と移す直前の再判定の両方が使う。"""
    headers = _peek_headers(path)
    mark = (multifile_stack.own_output_mark(path, headers) if headers is not None else None)
    if mark != extract_multi.CREATOR_MARK:
        return mark, False
    _creator, description = xml_readback.read_core_properties(path)
    try:
        cond = json.loads(description) if description else None
    except (TypeError, ValueError):
        cond = None
    same_condition = (isinstance(cond, dict) and cond.get("tool") == "ailine"
                       and cond.get("kind") == "extract" and cond.get("column") == col
                       and cond.get("cmp") == cmp and cond.get("value") == value)
    return mark, same_condition


def _refuse_output_conflict(out: Path, mark: str | None) -> int:
    """M2 出力先の関所（exit 7）: 人のファイル、または ailine の別コマンド/別条件の
       出力があって、黙って上書きしてよい根拠が無い時の唯一の出口。"""
    whose = (f"ailine の別のコマンドの出力です（作成: {mark}）" if mark
              else "ailine の印が無い人のファイルです")
    print(f"⚠ 出力先に書けません: {out}")
    print(f"（{out.name} は{whose}。run にはフラグでの上書き許可がありません ── "
          "そのファイルを別の場所へ移すか削除してから、もう一度実行してください）")
    return 7


def cmd_run_folder(a: argparse.Namespace) -> int:
    """`ailine run <フォルダ> "<依頼>"`: M2 ── フォルダの各ブックから条件に一致する行を
       抜き出して1冊に集約する（DESIGN-20260821-multifile.md M2 節・Namakoo 決裁 A 案）。
       配管は M1書き(stack) の再利用（出所列・書き手の印・関所 exit7・自己参照除外・
       workdir→移動）で、この経路固有なのは「翻訳された条件で選ぶ」ことだけ。
       ★ architect 致命4: この経路は check_excel_lock / normalize_book / バックアップ /
       undo 履歴のどれにも触れない（原本は読むだけ・消せるのは出力ブックだけ）。
       フォルダに `open(r+b)` を試みて『Excel で開かれています』と嘘をつく穴は、
       分岐がこの関数へ来ることで構造的に消える。
       ★ 一括検出（Namakoo 決裁 09:22）: ファイル単位の欠陥は最初の1件で止めず全部集めて
       名指しする。★ 憲法2: この関数は ✓ を一切名乗らない（✓ の裏づけを機械の結果
       オブジェクトから出す claim render guard がこの経路にまだ無い ── 迷ったら落とす）。"""
    folder = Path(a.book).resolve()
    as_json = bool(getattr(a, "json", False))
    say = (lambda *args, **kw: None) if as_json else print

    # ① 分母と自己参照除外（V6）── cmd_stack と同じ判定（ailine 産は種類を問わず入力から外す）。
    candidates, _folder_excluded = multifile.classify_folder_contents(folder)
    self_excluded, filtered = [], []
    for p in candidates:
        headers = _peek_headers(p)
        if headers is not None and multifile_stack.is_own_output(p, headers):
            self_excluded.append(p.name)
            continue
        filtered.append(p)
    candidates = filtered
    denominator = len(candidates)

    base_path, base_wb = multifile.open_base_workbook(candidates)
    if base_path is None:
        say(f"■ ailine run（フォルダ抽出）  folder={folder}")
        if self_excluded:
            names = "、".join(f"『{n}』" for n in self_excluded)
            say(f"（自分の出力 {names} を入力から除外しました）")
        say(f"{denominator} ファイル中 0 照合 → 読める .xlsx が無いので抽出できません")
        if as_json:
            print(json.dumps({"out": None, "condition": None, "multifile": {
                "denominator": denominator, "matched_files": 0, "contributing_files": 0,
                "rows_written": 0, "files": [], "skipped": [
                    {"name": p.name, "reason": "旧形式(.xls)または読み込み失敗"} for p in candidates],
                "self_excluded": self_excluded, "sheet_fallbacks": [], "excluded_detail": [],
                "mismatches": [], "rebuilt_own_output": False}}, ensure_ascii=False))
        return 0

    # ② 基準ファイル方式（A1）: 見出し行の推定は既存の器官を1回だけ使い、値として渡す。
    base_sheet = base_wb.sheetnames[0]
    ws = base_wb[base_sheet]
    scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
    rows_stats = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
    row, confident = detect_header_row({"rows": rows_stats})
    header_row = row if confident else 1
    base_headers = multifile.read_row_headers(ws, header_row)
    base_wb.close()

    # ③ 翻訳（7B に渡すのは基準ファイルの見出しだけ ── 40 冊分で prompt を壊さない）。
    book_meta = {"sheets": [base_sheet], "headers": {base_sheet: base_headers},
                 "header_rows": {base_sheet: header_row}}
    translation = translate_task(a.model, a.task, book_meta, temperature=0.1)
    plan = translation.get("plan") if isinstance(translation, dict) else None
    if plan is None and isinstance(translation, dict) and translation.get("op"):
        plan = [translation]          # ★ 後方互換: "plan" で包まない旧形式
    if not isinstance(plan, list) or not plan:
        plan = [{"op": "FREEFORM", "args": {}}]

    step = plan[0] if isinstance(plan[0], dict) else {}
    op = str(step.get("op") or "")
    if len(plan) == 1 and op == "CLARIFY":
        print(f"？ {step.get('question') or '確認が必要です'}")
        print("  （頼める操作の一覧: ailine ops）")
        return 3
    if len(plan) != 1 or not OP_META.get(op, {}).get("folder"):
        return _run_folder_refuse(op, len(plan))

    # ④ 条件の確定（A': 値の数値化は機械が行う ── verify_dsl_args と同じ線）。
    args = step.get("args") or {}
    col = args.get("col") or args.get("column")
    cmp = str(args.get("cmp") or "").strip().lower()
    value = args.get("value")
    if cmp not in _EXTRACT_CMPS:
        print(f"？ 比較『{args.get('cmp')}』は {'/'.join(_EXTRACT_CMPS)} のどれでもありません。"
              "言い方を変えて（例:『金額が40000以上の行を抜き出して』）もう一度お願いします。")
        return 3
    if col not in base_headers:
        print(f"？ 列『{col}』が基準ファイル『{base_path.name}』にありません。"
              f"ある列: {', '.join(base_headers)}")
        return 3
    if value in (None, ""):
        print("？ 抽出する値が依頼文から読み取れません（例:『金額が40000以上の行を抜き出して』）。")
        return 3
    if cmp == "contains":
        value = str(value)
    else:
        try:
            value = float(value)
        except (TypeError, ValueError):
            if cmp != "eq":
                print(f"？ 『{value}』は数値として読めないので {_EXTRACT_CMP_LABELS[cmp]} の"
                      "比較ができません。数値で言い直してください。")
                return 3

    # ⑤ 出力先（Q7: フォルダの親・機械命名）と書き込みの関所（40 冊読む前に判定して印字）。
    #    ★ review3#1/#5: 黙って作り直してよいのは「印」だけでなく「条件も一致」する時だけ。
    #    条件が違う自分の出力と名前が衝突したら、条件のハッシュで別名へ決定論で逃がす
    #    （切り詰めが起きていなくても、表示形式が同じ文字列に丸まる衝突はありうるため
    #    sanitize_filename 単体の直し（③）とは別にここでも見る）。
    cmp_label = _EXTRACT_CMP_LABELS.get(cmp, cmp)
    raw_stem = f"{folder.name}_{col}{_format_extract_value(value)}{cmp_label}"
    out = folder.parent / f"{extract_multi.sanitize_filename(raw_stem)}.xlsx"
    rebuilt_own_output = False
    if out.exists():
        mark, same_condition = _own_extract_output_status(out, col, cmp, value)
        if mark == extract_multi.CREATOR_MARK and same_condition:
            rebuilt_own_output = True      # 同じ条件の前回出力＝黙って作り直してよい
        elif mark == extract_multi.CREATOR_MARK:
            # ★ review3#1②: 条件が違う自分の出力と衝突 ── ハッシュで別名に決定論で逃がす。
            digest = hashlib.sha256(json.dumps(
                {"column": col, "cmp": cmp, "value": value}, sort_keys=True,
                ensure_ascii=False).encode("utf-8")).hexdigest()[:6]
            collided_name = out.name
            out = folder.parent / f"{extract_multi.sanitize_filename(f'{raw_stem}_{digest}')}.xlsx"
            say(f"（同名の前回出力『{collided_name}』は別条件のため、この結果は"
                f"『{out.name}』に保存します）")
            if out.exists():
                mark, same_condition = _own_extract_output_status(out, col, cmp, value)
                if mark == extract_multi.CREATOR_MARK and same_condition:
                    rebuilt_own_output = True
                else:
                    return _refuse_output_conflict(out, mark)
        else:
            return _refuse_output_conflict(out, mark)
    say(f"■ ailine run（フォルダ抽出）  folder={folder}")
    say(f"出力先: {out}")
    say(f"条件: {col} {_format_extract_value(value)} {cmp_label}")

    # ⑥ ファイルごとの評価（★ 一括検出: 欠陥が出ても止めず全部集める）。
    skipped, files_json, excluded_detail, mismatches = [], [], [], []
    sheet_fallbacks, matched_rows_all = [], []
    for p in candidates:
        r = extract_multi.evaluate_and_extract(p, base_headers, base_sheet, header_row,
                                                col, cmp, value)
        if r.sheet_fallback:
            sheet_fallbacks.append({"name": r.name, "wanted": r.sheet_fallback[0],
                                    "used": r.sheet_fallback[1]})
        if r.status == "取れなかった":
            skipped.append({"name": r.name, "reason": r.reason})
            continue
        for values, src_row in r.rows:
            matched_rows_all.append((values, r.name, src_row))
        files_json.append({"name": r.name, "rows_matched": r.rows_matched,
                           "rows_unmatched": r.rows_unmatched,
                           "total_rows_excluded": len(r.excluded), "reordered": r.reordered})
        if r.excluded:
            excluded_detail.append({"name": r.name, "rows": [
                {"row": e.row, "value": e.value, "reason": e.reason} for e in r.excluded]})
        for m in r.mismatches:
            mismatches.append({"name": r.name, "row": m.row,
                               "excluded_value": m.excluded_value, "adopted_sum": m.adopted_sum})

    total_matched = len(matched_rows_all)
    matched_files = len(files_json)
    contributing_files = len({name for _v, name, _r in matched_rows_all})

    # ⑦ 書き出し（workdir→移動）。★ 条件は文書属性に焼く ── verify が出力単体から
    #    条件を復元して同じ検算を再実行できる（信用の条件⑥）。
    out_headers = base_headers + multifile_stack.own_output_headers(base_headers)
    workdir = Path(tempfile.mkdtemp(prefix="ailine_extract_"))
    try:
        tmp_out = workdir / out.name
        wb_out = openpyxl.Workbook()
        wb_out.properties.creator = extract_multi.CREATOR_MARK
        wb_out.properties.description = json.dumps(
            {"tool": "ailine", "kind": "extract", "version": 1,
             "column": col, "cmp": cmp, "value": value, "sheet": base_sheet},
            ensure_ascii=False)
        ws_out = wb_out.active
        ws_out.title = _extract_output_sheet_name(col, cmp, value)
        ws_out.append(out_headers)
        for values, fname, src_row in matched_rows_all:
            ws_out.append(list(values) + [fname, src_row])
        wb_out.save(tmp_out)
        wb_out.close()

        # ⑧ 事後条件: 書いた直後の中身を**独立読み**（xml_readback）で検算する。
        #    元側も候補ファイル全部を読み直す ── 一致0行のファイルは出所列に現れないため、
        #    出所列だけを頼りにすると「1冊まるごと落ちた」が検算をすり抜ける。
        post = multifile_verify.verify_extract(tmp_out, folder, col, cmp, value,
                                                sheet_name=base_sheet, sources=candidates)
        if post.get("mismatch"):
            m = post["mismatch"]
            where = f"Σ{m['column']}" if m["kind"] == "sum" else "採用行数"
            if as_json:
                print(json.dumps({"out": str(out), "postcondition": post,
                                  "written": False}, ensure_ascii=False))
            else:
                print(f"⚠ 事後条件が破れた: {where}  元 {multifile_stack.fmt_num(m['source'])} / "
                      f"出力(書いた直後) {multifile_stack.fmt_num(m['output'])}")
                print(f"（{out.name} は書き込んでいません。元フォルダも変更していません）")
            return 1

        # ⑨ 関所（fail closed）: 移す直前にもう一度見る（前段の判定から時間が経っている）。
        #    ★ review3#1: ここも印だけでなく条件一致まで見る（preflight と同じ判定）。
        if out.exists():
            mark, same_condition = _own_extract_output_status(out, col, cmp, value)
            if not (mark == extract_multi.CREATOR_MARK and same_condition):
                print(f"⚠ 出力先に書けません（実行中に別のファイルが現れました）: {out}")
                return 7
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_out, out)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    # ⑩ 報告（分母つき・行の完全会計・両側の数字）。
    mf = {"denominator": denominator, "matched_files": matched_files,
          "contributing_files": contributing_files, "rows_written": total_matched,
          "files": files_json, "skipped": skipped, "self_excluded": self_excluded,
          "sheet_fallbacks": sheet_fallbacks, "excluded_detail": excluded_detail,
          "mismatches": mismatches, "rebuilt_own_output": rebuilt_own_output}
    if as_json:
        print(json.dumps({"out": str(out), "written": True,
                          "condition": {"column": col, "cmp": cmp, "value": value},
                          "sums": post.get("sums", {}), "multifile": mf}, ensure_ascii=False))
        return 0
    if self_excluded:
        names = "、".join(f"『{n}』" for n in self_excluded)
        say(f"（自分の出力 {names} を入力から除外しました）")
    say(f"{denominator} ファイル中 {matched_files} 照合 → "
        f"{matched_files} 中 {contributing_files} ファイルで計 {total_matched} 行一致")
    # ★ D6 差し戻し（実弾検分・2026-08-21）: 正常なファイルは名指ししない ── 名指しは
    #   異常のあるファイル（取れなかった／閉じる検査の不一致／シート fallback）だけ。
    #   正常分（並べ替え・合計行の除外・行の完全会計）は 1 行の集計に畳む。
    #   全ファイルの内訳は --json に既にそのまま載っている（ここでは減らさない）。
    for f in skipped:
        say(f"  ⚠ {f['name']}: 取れなかった（{f['reason']}）── 中身の検査は未実施")
    reordered_files = [f for f in files_json if f.get("reordered")]
    if reordered_files:
        say(f"  並べ替えて照合: {len(reordered_files)} 冊（内訳は --json）")
    for f in sheet_fallbacks:
        say(f"  {f['name']}: シート『{f['wanted']}』が無いので1枚目『{f['used']}』を使いました")
    if excluded_detail:
        total_excluded_rows = sum(len(entry["rows"]) for entry in excluded_detail)
        say(f"  合計行 {total_excluded_rows} 行を {len(excluded_detail)} 冊で除外"
            "（内訳は --json）")
    for m in mismatches:
        say(f"  ⚠ {m['name']}: 合計行({m['row']}行目) の値 "
            f"{multifile_stack.fmt_num(m['excluded_value'])} ≠ 明細の和 "
            f"{multifile_stack.fmt_num(m['adopted_sum'])}")
    if files_json:   # ★ 憲法⑨ 行の完全会計: どの行もどれかの勘定に入っている（全冊で成立・集計）
        say(f"  行の完全会計: {len(files_json)} 冊すべてで成立"
            "（データ行 = 一致 + 不一致 + 合計行の除外・内訳は --json）")
    say(f"出力データ行数: {total_matched}")
    for col_name, both in post.get("sums", {}).items():
        say(f"Σ{col_name}: 元 {multifile_stack.fmt_num(both['source'])} / "
            f"出力 {multifile_stack.fmt_num(both['output'])}")
    if rebuilt_own_output:
        say(f"（前回の抽出出力『{out.name}』を作り直しました）")
    return 0


def cmd_stack(a: argparse.Namespace) -> int:
    """`ailine stack <folder> --out <path>`: M1書き ── 縦積み（UNION ALL）+ 出所列。
       DESIGN-20260821-multifile.md v2 §1(M1書き)・v2.1(単位L)。列挙・照合・合計行の識別は
       既存部品（ailine_core/multifile.py・total_row.py）を再利用し、この関数は
       積む行の決定・出所列つきの書き出し・関所・事後条件の配線だけを持つ（本体は
       ailine_core/stack.py）。★ workdir は tempfile に作り、最後に out へ移す。"""
    folder = Path(a.folder).resolve()
    out = Path(a.out).resolve()
    candidates, _excluded = multifile.classify_folder_contents(folder)

    # ★ 自己参照除外（V6・architect 致命2 で拡張）: 入力フォルダ内の ailine 産の出力
    #   （out と同じパスに限らず、種類（stack/extract 等）も問わない）は二重計上を防ぐため
    #   入力から除外 + 開示。判定は marks 集合（is_own_output）── 印が違っても ailine 産なら除外。
    self_excluded = []
    filtered = []
    for p in candidates:
        headers = _peek_headers(p)
        if headers is not None and multifile_stack.is_own_output(p, headers):
            self_excluded.append(p.name)
            continue
        filtered.append(p)
    candidates = filtered
    denominator = len(candidates)

    base_path, base_wb = multifile.open_base_workbook(candidates)
    if base_path is None:
        result = {"denominator": denominator, "stacked_files": 0, "rows_written": 0,
                  "files": [], "skipped": [{"name": p.name, "reason": "旧形式(.xls)または読み込み失敗"}
                                            for p in candidates],
                  "sums": {}, "excluded_detail": [], "mismatches": [], "col_a_warnings": [],
                  "sheet_fallbacks": [], "self_excluded": self_excluded, "rebuilt_own_output": False}
        if a.json:
            print(json.dumps(_stack_json(result), ensure_ascii=False))
        else:
            for ln in render_stack_report(str(folder), str(out), result):
                print(ln)
        return 0

    base_sheet = base_wb.sheetnames[0]
    ws = base_wb[base_sheet]
    scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
    rows_stats = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
    row, confident = detect_header_row({"rows": rows_stats})
    header_row = row if confident else 1
    base_headers = multifile.read_row_headers(ws, header_row)
    value_col = multifile.numeric_value_column(ws, header_row, len(base_headers) or MAX_COLS)
    value_col_name = base_headers[value_col - 1] if value_col else None   # 合計行検出の keyed 列（変えない）
    # ★ jisaku-review#3/#6: Σ 照合・報告は全数値列に（value_col_name は合計行検出専用のまま）。
    numeric_cols = multifile_stack.numeric_column_names(ws, header_row, base_headers)
    base_wb.close()

    skipped, files_json, excluded_detail, mismatches, col_a_warnings = [], [], [], [], []
    sheet_fallbacks = []   # ★ P2 開示: 基準名のシートが無く1枚目へ落ちたファイル
    stacked_rows = []   # [(base 列順の値, 元ファイル名, 元行), ...]
    sums_source = {col: 0.0 for col in numeric_cols}

    for p in candidates:
        r = multifile_stack.evaluate_and_stack(p, base_headers, base_sheet, header_row, value_col_name)
        if r.sheet_fallback:
            sheet_fallbacks.append({"name": r.name, "wanted": r.sheet_fallback[0],
                                    "used": r.sheet_fallback[1]})
        if r.status == "積めなかった":
            skipped.append({"name": r.name, "reason": r.reason})
            continue
        for values, src_row in r.rows:
            stacked_rows.append((values, r.name, src_row))
            for col in numeric_cols:
                v = values[base_headers.index(col)]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    sums_source[col] += v
        files_json.append({"name": r.name, "rows_stacked": len(r.rows),
                           "total_rows_excluded": len(r.excluded), "reordered": r.reordered})
        if r.excluded:
            excluded_detail.append({"name": r.name, "rows": [
                {"row": e.row, "value": e.value, "reason": e.reason} for e in r.excluded]})
        if r.mismatches:
            mismatches.append({"name": r.name, "rows": [
                {"row": m.row, "excluded_value": m.excluded_value, "adopted_sum": m.adopted_sum}
                for m in r.mismatches]})
        if r.col_a_mismatch:
            col_a_warnings.append({"name": r.name, "col_a": r.col_a_mismatch[0],
                                   "used_range": r.col_a_mismatch[1]})

    stacked_files = len(files_json)
    prov_headers = multifile_stack.own_output_headers(base_headers)
    out_headers = base_headers + prov_headers
    collision_notice = None
    if tuple(prov_headers) != multifile_stack.PROVENANCE_HEADERS:
        collision_notice = f"列名の衝突: 出所列は {prov_headers[0]} / {prov_headers[1]} として追加"

    workdir = Path(tempfile.mkdtemp(prefix="ailine_stack_"))
    try:
        tmp_out = workdir / out.name
        wb_out = openpyxl.Workbook()
        wb_out.properties.creator = multifile_stack.CREATOR_MARK   # ★ jisaku-review#1: 書き手の印
        ws_out = wb_out.active
        ws_out.title = base_sheet
        ws_out.append(out_headers)
        for values, fname, src_row in stacked_rows:
            ws_out.append(list(values) + [fname, src_row])
        wb_out.save(tmp_out)
        wb_out.close()

        # ★ 事後条件①②: 独立読み実装（xml_readback）で書いた直後の中身を検算する
        #   （openpyxl で書いて openpyxl で読み返すだけでは、同じ道具の同じ盲点を通る）。
        readback = xml_readback.read_grid(tmp_out)
        out_row_nums = xml_readback.data_row_numbers(readback, header_row=1)
        if len(out_row_nums) != len(stacked_rows):
            return _stack_postcondition_fail("採用行数", len(stacked_rows), len(out_row_nums))
        # ★ jisaku-review#3/#6: 事後条件②も全数値列に（最初の不一致で止める）。
        sums_output = {}
        for col in numeric_cols:
            idx = base_headers.index(col) + 1
            total = 0.0
            for rr in out_row_nums:
                v = readback["grid"].get((rr, idx))
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    total += v
            sums_output[col] = total
            if abs(total - sums_source.get(col, 0.0)) > 1e-6:
                return _stack_postcondition_fail(f"Σ{col}", sums_source.get(col, 0.0), total)

        # ★ 関所（writes=new_book）: 移す直前に判定。
        #   ★ architect 致命2: 「作り直してよい」は creator の完全一致（CREATOR_MARK）に限定。
        #   ailine 産だが印が違う出力（例: ailine extract）は「別のコマンドの出力」として
        #   名指しで止める（無警告の作り直しにしない）。
        rebuilt_own_output = False
        if out.exists():
            existing_headers = _peek_headers(out)
            mark = (multifile_stack.own_output_mark(out, existing_headers)
                    if existing_headers is not None else None)
            if mark == multifile_stack.CREATOR_MARK:
                rebuilt_own_output = True
            elif mark is not None:
                if not getattr(a, "overwrite", False):
                    print(f"⚠ 出力先は ailine の別のコマンドの出力です: {out}")
                    print(f"（{out.name}: これは ailine の別のコマンドの出力です（作成: {mark}）。"
                         "承知の上なら --overwrite を付けて実行してください）")
                    return 7
            elif not getattr(a, "overwrite", False):
                # ★ jisaku-review#5: 無言で閉まらない ── 何が邪魔か（名指し）+ 次の手を言う。
                print(f"⚠ 出力先に人のファイルがあります: {out}")
                print(f"（{out.name} は ailine stack の前回出力ではありません。"
                     "承知の上で上書きするなら --overwrite を付けて実行してください）")
                return 7
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_out, out)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    sums = {col: {"source": sums_source[col], "output": sums_output.get(col, sums_source[col])}
            for col in sums_source}
    result = {"denominator": denominator, "stacked_files": stacked_files,
              "rows_written": len(stacked_rows), "files": files_json, "skipped": skipped,
              "sums": sums, "excluded_detail": excluded_detail, "mismatches": mismatches,
              "col_a_warnings": col_a_warnings, "sheet_fallbacks": sheet_fallbacks,
              "self_excluded": self_excluded,
              "rebuilt_own_output": rebuilt_own_output, "collision_notice": collision_notice}
    if a.json:
        print(json.dumps(_stack_json(result), ensure_ascii=False))
        return 0
    for ln in render_stack_report(str(folder), str(out), result):
        print(ln)
    return 0


def cmd_verify(a: argparse.Namespace) -> int:
    """`ailine verify <out.xlsx> <srcfolder>`: 検算の単独再実行（信用の条件⑥）。
       stack の出力ブックと元フォルダだけから、行数照合・数値列ごとの Σ 照合を独立に
       再実行する（読みは ailine_core/xml_readback.py・openpyxl は経由しない）。
       本体（ailine_core/verify.py）が検算そのものを持ち、この関数は配線だけ。
       ★ M2（E13/致命3）: 検算の種類（縦積み/抽出）は出力ブックの印と焼いた条件から
       verify_output が決める。ailine の印が無いブックは合格でも不合格でもなく
       exit 4（「検算できません」）── 0 件照合で空虚な合格を名乗らない。"""
    out = Path(a.out).resolve()
    folder = Path(a.srcfolder).resolve()
    result = multifile_verify.verify_output(out, folder)
    if result.get("unmarked"):
        print(f"× ailine の印がありません。検算できません: {out}")
        return 4
    for ln in render_verify_report(str(out), str(folder), result):
        print(ln)
    return 5 if result.get("mismatch") else 0


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(prog="ailine", description="自然言語 → LibreOffice Basic → 適用 → 検証")
    sub = ap.add_subparsers(dest="cmd", required=True)

    r = sub.add_parser("run", help="タスクを生成・適用・検証する")
    r.add_argument("book", help="対象の文書 (.xlsx / .ods)")
    r.add_argument("task", help="やりたいことを自然言語で")
    r.add_argument("--model", default=DEFAULT_MODEL, help=f"ollama モデル (既定 {DEFAULT_MODEL})")
    r.add_argument("--refs", default=None, help="参照ライブラリのディレクトリ (既定 ./refs)")
    r.add_argument("--helpers", default=None, help="検証済みヘルパのディレクトリ (既定 ./helpers)")
    r.add_argument("--repair", type=int, default=2, help="修復の最大回数 (既定 2)")
    r.add_argument("--temperature", type=float, default=0.2)
    r.add_argument("--dry", action="store_true", help="生成して見せるだけ（適用しない）")
    r.add_argument("--inplace", action="store_true",
                   help="（廃止・後方互換のため受理のみ）既定で原本に直接適用するため不要。"
                        "旧 .out 挙動が欲しければ --copy")
    r.add_argument("--json", action="store_true", help="結果を JSON でも出す")
    r.add_argument("--timeout", type=float, default=DEFAULT_APPLY_TIMEOUT,
                   help=f"basrun apply のタイムアウト秒 (既定 {DEFAULT_APPLY_TIMEOUT:.0f}、"
                        "0 で無効化=旧挙動の無制限)")
    r.add_argument("--ask", action="store_true",
                   help="DSL 経路の確認行の後に y/n で対話する（既定は表示して続行）")
    r.add_argument("--keep-backups", dest="keep_backups", type=int, default=DEFAULT_KEEP_BACKUPS,
                   help=f"原本への反映前のバックアップを book ごとに何世代残すか (既定 {DEFAULT_KEEP_BACKUPS}、"
                        "負数で無制限)")
    r.add_argument("--values", action="store_true",
                   help="COMPUTE_COLUMN を式でなく値ベタ書きにする（既定は式・W3 Part3）")
    r.add_argument("--header-row", dest="header_row", type=int, default=None,
                   help="見出し行を明示指定（1起点。指定時は自動検出をスキップしてこの行を採用）")
    r.add_argument("--sheet", default=None,
                   help="対象シートをシート名で明示指定（省略時は依頼文中のシート名の言及 → "
                        "1枚目。★ 挙動変更#2: 従来は常に1枚目固定だった）")
    r.add_argument("--accept-loss", dest="accept_loss", action="store_true",
                   help="往復忠実度ゲートが検出した喪失を承知の上で原本への反映を続行する"
                        "（ailine undo で戻せる）")
    r.add_argument("--copy", action="store_true",
                   help="原本には触らず <book>.out に結果を作る（既定の原本直接適用をしない・"
                        "旧 --inplace 無指定と同じ挙動。往復忠実度ゲートも走らせない）")
    r.add_argument("--overwrite", action="store_true",
                   help="破壊の関所（既存データを持つ列への上書き確認）を承知の上で"
                        "続行する（ailine undo で戻せる）")
    r.add_argument("--allow-freeform", dest="allow_freeform", action="store_true",
                   help="自由生成の関所（AI が直接作成したコードは機械検証できないという確認）を"
                        "承知の上で続行する（ailine undo で戻せる）")
    r.set_defaults(func=cmd_run)

    s = sub.add_parser("stop", help="起動した LibreOffice を落とす")
    s.set_defaults(func=cmd_stop)

    d = sub.add_parser("doctor", help="セットアップを診断する")
    d.add_argument("--model", default=DEFAULT_MODEL, help=f"確認するモデル (既定 {DEFAULT_MODEL})")
    d.set_defaults(func=cmd_doctor)

    o = sub.add_parser("ops", help="頼める操作の一覧を表示する（何ができるか）")
    o.set_defaults(func=cmd_ops)

    sc = sub.add_parser("scan", help="フォルダ内の複数ブックを棚卸しする（書き込みゼロ）")
    sc.add_argument("folder", help="対象フォルダ（直下の .xlsx / .xls のみ・サブフォルダは見ない）")
    sc.add_argument("--json", action="store_true", help="結果を JSON で出す（stdout は JSON のみ）")
    sc.set_defaults(func=cmd_scan)

    st = sub.add_parser("stack", help="フォルダ内の複数ブックを縦積みする（新ブック + 出所列）")
    st.add_argument("folder", help="対象フォルダ（直下の .xlsx / .xls のみ・サブフォルダは見ない）")
    st.add_argument("--out", required=True, help="出力ブックのパス")
    st.add_argument("--overwrite", action="store_true",
                    help="出力先に人のファイルが既にある時の関所（exit 7）を承知の上で上書きする")
    st.add_argument("--json", action="store_true", help="結果を JSON で出す（stdout は JSON のみ）")
    st.set_defaults(func=cmd_stack)

    vf = sub.add_parser("verify", help="stack の出力を検算だけ独立に再実行する（読むだけ）")
    vf.add_argument("out", help="ailine stack が作った出力ブック")
    vf.add_argument("srcfolder", help="元ファイルがあるフォルダ")
    vf.set_defaults(func=cmd_verify)

    h = sub.add_parser("history", help="実行履歴を表示する")
    h.add_argument("--max", type=int, default=10, help="表示件数（既定 10、新しい順）")
    h.set_defaults(func=cmd_history)

    rs = sub.add_parser("restore", help="原本への反映前のバックアップから復元する（ailine undo と同じ）")
    rs.add_argument("book", help="対象の文書 (.xlsx / .ods)")
    rs.add_argument("--list", action="store_true", help="バックアップ一覧を表示するだけ（復元しない）")
    rs.set_defaults(func=cmd_restore)

    u = sub.add_parser("undo", help="原本への反映前のバックアップから復元する（あと何回戻せるかを表示）")
    u.add_argument("book", help="対象の文書 (.xlsx / .ods)")
    u.add_argument("--list", action="store_true", help="バックアップ一覧を表示するだけ（復元しない）")
    u.set_defaults(func=cmd_undo)

    v = sub.add_parser("vocab", help="用語集（税率等の取り決め値）を編集・表示する")
    vsub = v.add_subparsers(dest="vocab_cmd", required=True)
    va = vsub.add_parser("add", help="語を登録する（例: ailine vocab add 消費税 1.1）")
    va.add_argument("term", help="語（例: 消費税）")
    va.add_argument("value", help="値（倍率。例: 1.1）")
    vl = vsub.add_parser("list", help="登録済みの語を一覧表示する")
    v.set_defaults(func=cmd_vocab)
    return ap


def main(argv=None) -> int:
    # ★ 日本語出力の生存保証（2026-08-21・CI の長期赤で発覚した製品バグ）: 端末やパイプの
    #   符号化が日本語を持たない環境（英語圏 Windows の cp1252 等）では、print が
    #   UnicodeEncodeError で落ちて「黙って失敗するより悪い、途中で死ぬ」になる。
    #   符号化はそのまま・書けない文字だけ置換に倒す（クラッシュしない、が保証の中身）。
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(errors="replace")
        except Exception:
            pass
    a = build_parser().parse_args(argv)
    return a.func(a)


if __name__ == "__main__":
    raise SystemExit(main())
