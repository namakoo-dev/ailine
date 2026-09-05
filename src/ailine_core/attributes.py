"""attributes — 語が実表の**何**を指しているかを、実表の証拠だけから列挙する層。

★★ なぜ要るか（2026-09-05・Namakoo の設計）:
  「もし語彙外で判断できない単語が現れたときに**属性を登録してもらう**のはどうだろう？
    登録する属性は有限と考えられるから列挙でいいが、
    **無関係な属性は表示させないように工夫する必要がある**」

  ★ 「無関係を出さない」の答えは**選択肢を実表から作る**こと ── 語 w について
    実表に証拠が無い属性は、そもそも候補に載らない。紙の上で測った形:

        『平均』  → 列(平均単価) ／ まだ無い操作(平均)   2 つ＝**聞くべき**
        『ボルト』→ 値(品名 列に 3 件)                    1 つ＝聞かない
        『売上』  → シート(売上一覧)                      1 つ＝聞かない
        『原価』  → （実表に手がかりなし）                0  ＝「ありません」

★ ここは**材料を並べるだけ**（cellmap と同じ規律）── 聞くかどうか・覚えるかどうかは
  呼び出し側が決める。★ 覚えた答えも使う時に必ず実表で検算する
  （登録は**解釈の保存**であって**検証の免除ではない** ── Namakoo 決裁 論点 E）。

★ 移植可能性（tests/test_line_budget.py が機械で守る）: ailine を import しない。
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass

KIND_SHEET = "sheet"
KIND_COLUMN = "column"
KIND_VALUE = "value"
KIND_OP_LACKING = "op_lacking"


@dataclass(frozen=True)
class Kind:
    """属性 1 種類の**全部**（名前・証拠の作り方・見せ方）を 1 箇所に持つ。

    ★★ なぜ 1 箇所か（2026-09-05・Namakoo「属性は 4 つで足りるか？」への答え）:
      属性を足すのに **名簿・候補作り・表示文の 3 箇所**を触る形だと、
      片方だけ直る（この repo で何度も起きた「片配線」そのもの）。
      **足りないと分かった時に安く足せるか**が、4 で足りるかより大事だった。

    evidence: (word, ctx) -> [(label, sheet, note, exact), ...]  ── 実表から証拠を作る
              exact=True は**その語ぴったり**の証拠、False は部分一致の弱い証拠。
    describe: (Candidate) -> str                          ── 人に見せる 1 行
    """
    name: str
    evidence: object
    describe: object


def _ev_sheet(word, ctx):
    return [(str(s), str(s), "", _same(word, s))
            for s in ctx.get("sheets") or () if _touches(word, s)]


def _ev_column(word, ctx):
    return [(str(col), str(sheet), "", _same(word, col))
            for sheet, cols in (ctx.get("headers") or {}).items()
            for col in cols or () if _touches(word, col)]


def _ev_value(word, ctx):
    out = []
    for (sheet, col), values in (ctx.get("samples") or {}).items():
        hits = [v for v in (values or ()) if _touches(word, v)]
        if hits:
            out.append((str(col), str(sheet), f"{len(hits)} 件",
                        any(_same(word, v) for v in hits)))
    return out


def _ev_op_lacking(word, ctx):
    return [(str(why or term), "", "", _same(word, term))
            for term, why in (ctx.get("lacks") or {}).items() if _touches(word, term)]


#: 属性の名簿。★ 足すときはここに 1 行 ── 証拠の作り方と見せ方が同じ行に並ぶ。
#: ★ 足していない物とその理由（2026-09-05 の実測 29 語・当たらなかった 12 語の内訳）:
#:    ・扱える操作の名前 / 書式の名前 … **既存の器官が持つ**（別名ストア・op の照合語彙）。
#:      ここに足すと二重化して片配線を生む ── 足さないのが正しい。
#:    ・数・単位・セル番地・日付 … 残差に出ない（数字は除去され、番地は語彙が処理する）。
#:    ・★ 用語集の項目（税率・消費税）… **実在の穴**。5 つ目の第一候補（vocab.json が証拠）。
#:    ・★ 行の名前（小計・合計行）… いまは値として拾えている。取りこぼしを見たら足す。
#:    ・★ 別ファイル（去年のファイル）… multifile 経路で残差が出るかを測ってから。
KIND_TABLE = {
    KIND_SHEET: Kind("シート", _ev_sheet,
                     lambda c: f"シートの名前『{c.label}』"),
    KIND_COLUMN: Kind("列", _ev_column,
                      lambda c: f"列の名前『{c.label}』" + (f"（{c.sheet}）" if c.sheet else "")),
    KIND_VALUE: Kind("値", _ev_value,
                     lambda c: f"{c.label} 列にある値"
                               + (lambda ins: f"（{ins}）" if ins else "")(
                                   "／".join(x for x in (c.sheet, c.note) if x))),
    KIND_OP_LACKING: Kind("まだ無い操作", _ev_op_lacking,
                          lambda c: f"まだ扱えない操作『{c.label}』"),
}

#: 画面の並びが実行のたびに揺れないための順（★ KIND_TABLE の宣言順そのもの）
KINDS = tuple(KIND_TABLE)

#: 値を見に行く深さ（★ 断りの経路で毎回走るので全セルは読まない・Namakoo 決裁 論点 C）
DEFAULT_SAMPLE_ROWS = 200

#: 画面に出す候補の上限（これを超える語は語として広すぎる ── 選ばせても意味がない）
MAX_CANDIDATES = 5


@dataclass(frozen=True)
class Candidate:
    """語がこう読める、という**実表の証拠 1 件**。"""
    kind: str            # KINDS のどれか
    label: str           # 実体（列名・シート名・値・まだ無い操作の名）
    sheet: str = ""      # どのシートの話か（sheet 属性では自分自身）
    note: str = ""       # 「3 件」など、証拠の濃さ
    exact: bool = False  # ★ その語ぴったりか（部分一致なら False ＝ 弱い証拠）

    def describe(self) -> str:
        """人に見せる 1 行（★ 文言は KIND_TABLE の行が持つ ── 聞く側と覚える側でずれない）。"""
        kind = KIND_TABLE.get(self.kind)
        return kind.describe(self) if kind else f"{self.kind}『{self.label}』"


def _same(word: str, name) -> bool:
    """語と名前が**ぴったり**同じか（★ 強い証拠 ── これだけが「そう読みました」と言える）。"""
    return bool(word) and name is not None and word == str(name).strip()


def _touches(word: str, name) -> bool:
    """語と名前が触れているか（一致・どちらかがもう片方を含む）。
       ★ sheet_named_but_missing と同じ判定 ── 片方だけ厳しくすると使う側が混乱する。"""
    if not word or name is None:
        return False
    s = str(name).strip()
    return bool(s) and (word == s or word in s or s in word)


def candidates_for(word: str, *, sheets=(), headers=None, samples=None,
                   lacks=None) -> list:
    """語 word について、実表に証拠のある読み方だけを列挙する。

    sheets:  シート名の並び
    headers: {シート名: [列名, ...]}
    samples: {(シート名, 列名): [値, ...]}   ── sample_columns が作る（無くてもよい）
    lacks:   {語: 説明}  ── op_axes の「まだ無い側」を集めたもの（無くてもよい）

    ★ 返る順は KINDS の順（画面の並びが実行のたびに揺れないため）。
    """
    if not word:
        return []
    ctx = {"sheets": sheets, "headers": headers, "samples": samples, "lacks": lacks}
    out = []
    for name, kind in KIND_TABLE.items():   # ★ 名簿を回すだけ（属性を足しても ここは変わらない）
        for label, sheet, note, exact in kind.evidence(word, ctx) or ():
            out.append(Candidate(name, label, sheet=sheet, note=note, exact=bool(exact)))
    order = {k: i for i, k in enumerate(KINDS)}
    out.sort(key=lambda c: (order.get(c.kind, len(KINDS)), c.sheet, c.label))
    return _collapse_across_sheets(out)


def _collapse_across_sheets(cands: list) -> list:
    """★ 同じ読み方がシート違いで並ぶのを 1 つに畳む（実測で見つけた・紙の予測と 1 件ずれた）。

      『ボルト』→ 『在庫』の品名列にある値 ／ 『売上一覧』の品名列にある値   ← 2 つに見える
      『品名』  → 列の名前『品名』（在庫） ／ 列の名前『品名』（売上一覧）    ← 同上

    ★ 聞いているのは「**その語が何か**」であって「どのシートか」ではない。
      対象シートの決定は resolve_target_sheet の仕事 ── ここで二重に聞かない。
      畳んだ結果 1 つになれば**聞かずに済む**（＝無関係な選択肢を出さない・Namakoo）。
    """
    merged = {}
    for c in cands:
        key = (c.kind, c.label)
        if key not in merged:
            merged[key] = [c, [c.sheet] if c.sheet else [], [c.note] if c.note else []]
            continue
        keep, sheets_, notes_ = merged[key]
        if c.exact and not keep.exact:
            merged[key][0] = keep = c   # ★ 強い証拠が 1 つでもあれば強い側に倒す
        if c.sheet and c.sheet not in sheets_:
            sheets_.append(c.sheet)
        if c.note:
            notes_.append(c.note)   # ★ 件数は足す（下で合算する ── 「1 件・2 件」と並べない）
        del keep
    out = []
    for (kind, label), (first, sheets_, notes_) in merged.items():
        if len(sheets_) <= 1 and len(notes_) <= 1:
            out.append(first)
            continue
        total = 0
        for n in notes_:
            head = "".join(ch for ch in n if ch.isdigit())
            total += int(head) if head else 0
        out.append(Candidate(kind, label, sheet="・".join(sheets_),
                              note=(f"計 {total} 件" if total else first.note),
                              exact=first.exact))
    return out


def sample_columns(path, headers, header_rows=None, max_rows: int = DEFAULT_SAMPLE_ROWS) -> dict:
    """実表の値を**先頭 max_rows 行だけ**見て {(シート名, 列名): [値, ...]} を作る。

    ★ 全セルは読まない（cellmap.read_cellmap は全部読む ── 断りの経路には重すぎる）。
    ★ 読めないブック・壊れたシートでは黙って空を返す（断りの経路を落とさない）。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    rows_of = dict(header_rows or {})
    out = {}
    try:
        for sheet, cols in (headers or {}).items():
            if sheet not in wb.sheetnames or not cols:
                continue
            ws = wb[sheet]
            head = int(rows_of.get(sheet, 1) or 1)
            picked = {}
            for i, row in enumerate(ws.iter_rows(min_row=head + 1,
                                                 max_row=head + max_rows,
                                                 max_col=len(cols), values_only=True), 1):
                for c, v in zip(cols, row):
                    if v is None or (isinstance(v, str) and not v.strip()):
                        continue
                    picked.setdefault(c, []).append(v)
                del i
            for c, vs in picked.items():
                out[(sheet, c)] = vs
    except Exception:
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def book_key(path, sheets=(), headers=None) -> dict:
    """ブックの鍵（★ Namakoo 決裁 論点 B: **パスと列名の署名の両方**を持つ）。

    ★ どちらかが合えば引く ── ファイルを動かしても、列を足しても、片方が生きていれば
      聞き直さずに済む。★ 使う時に必ず実表で検算するので、鍵が外れても**害は無い**
      （合わなければ捨ててまた聞くだけ）── つまり「何回聞かれるか」だけの問題。
    """
    names = "\n".join(f"{s}\t{chr(9).join(str(c) for c in (headers or {}).get(s, ()))}"
                      for s in sorted(sheets or ()))
    return {"path": str(path), "sig": hashlib.sha256(names.encode("utf-8")).hexdigest()[:16]}


def key_matches(saved: dict, now: dict) -> bool:
    """保存した鍵と今の鍵が**どちらか一方でも**合えば同じブックとみなす。"""
    if not isinstance(saved, dict) or not isinstance(now, dict):
        return False
    return bool((saved.get("path") and saved.get("path") == now.get("path"))
                or (saved.get("sig") and saved.get("sig") == now.get("sig")))


def render_question(word: str, cands: list) -> list:
    """候補が 2 つ以上のときに出す前置き（選択肢の行は ask_choice が作る）。"""
    return [f"？ 『{word}』が何を指しているか決められません",
            "  この表には、こう読める手がかりがあります:"]


def render_no_evidence(word: str, sheets=(), headers=None) -> list:
    """候補が 0 のとき ── **実表に何が在るか**を見せて終わる（★ 非対話でも言ってよい）。"""
    cols = []
    for s in sheets or ():
        for c in (headers or {}).get(s, ()) or ():
            if str(c) not in cols:
                cols.append(str(c))
    lines = [f"？ このブックに『{word}』は見当たりません"]
    if sheets:
        lines.append(f"  シート: {'、'.join(str(s) for s in sheets)}")
    if cols:
        lines.append(f"  列: {'、'.join(cols[:12])}" + ("…" if len(cols) > 12 else ""))
    return lines


def render_understood(word: str, cand: Candidate) -> str:
    """候補が 1 つのとき ── **黙って使わず、そう読んだことを必ず言う**
       ★ 呼び出し側は cand.exact のときだけここへ来ること（弱い証拠で断定しない）。
       （★ Namakoo 決裁 論点 A: 印字して進む・ただし覚えない。憲法「参照のズレは
         既定では直さない。補正するなら確認を取ってから」との折り合い）。"""
    return f"  （『{word}』は{cand.describe()}のことと読みました）"


# --- 覚える層（★ 別名ストアの流儀に揃える: 検疫・上限・取り消し）-------------------
#
# ★ 機械が書く層には**取り消し**が要る（別名ストアで決めた 6-4 の規律をそのまま継ぐ）。
# ★ 保存するのは「解釈」だけ ── 使う時に必ず実表で検算する（論点 E）。

#: 実体（列名・値）の長さの上限（★ 語の上限は呼び出し側の語彙規則に合わせる）
MAX_LABEL_LEN = 120


def parse_attributes_json(raw, *, max_entries: int, max_term_len: int) -> list:
    """attributes.json を項目の並びに読み直す。**壊れていても落ちない**
       （形の違う項目は黙って捨てる ── load_vocab / load_aliases と同じ流儀）。"""
    if not isinstance(raw, dict):
        return []
    items = raw.get("entries")
    if not isinstance(items, list):
        return []
    out = []
    for it in items:
        if not isinstance(it, dict):
            continue
        word = it.get("word")
        kind = it.get("kind")
        label = it.get("label")
        book = it.get("book")
        if not (isinstance(word, str) and word.strip() and len(word) <= max_term_len):
            continue
        if kind not in KINDS:
            continue
        if not (isinstance(label, str) and label.strip() and len(label) <= MAX_LABEL_LEN):
            continue
        if not isinstance(book, dict):
            continue
        path, sig = book.get("path"), book.get("sig")
        if not (isinstance(path, str) or isinstance(sig, str)):
            continue
        out.append({"word": word, "kind": kind, "label": label,
                    "sheet": it.get("sheet") if isinstance(it.get("sheet"), str) else "",
                    "book": {"path": path if isinstance(path, str) else "",
                              "sig": sig if isinstance(sig, str) else ""}})
        if len(out) >= max_entries:
            break
    return out


def build_attributes_payload(entries: list) -> dict:
    """保存する形（★ 並び順が取り消しの順序そのもの ── 末尾が最後に登録したもの）。"""
    return {"entries": list(entries or [])}


def find_attribute(entries: list, word: str, now_key: dict) -> dict | None:
    """この語について、**このブックで**覚えた読み方があれば返す（無ければ None）。
       ★ 同じ語が複数あれば**最後に登録したもの**が勝つ（人が言い直した方が新しい）。"""
    if not word:
        return None
    for it in reversed(entries or []):
        if it.get("word") == word and key_matches(it.get("book"), now_key):
            return it
    return None


def entry_still_holds(entry: dict, *, sheets=(), headers=None, samples=None,
                      lacks=None) -> bool:
    """★ 論点 E の要 ── 覚えた読み方が**今の実表でもまだ成り立つか**を確かめる。
       成り立たなければ呼び出し側は捨てて聞き直す（古い登録は害にならない）。"""
    if not isinstance(entry, dict):
        return False
    for c in candidates_for(entry.get("word", ""), sheets=sheets, headers=headers,
                            samples=samples, lacks=lacks):
        if c.kind == entry.get("kind") and c.label == entry.get("label"):
            return True
    return False
