"""stack — M1書き `ailine stack <folder> --out <path>`（縦積み・UNION ALL）の本体。
   DESIGN-20260821-multifile.md v2 §1(M1書き)・v2.1。

   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定（detect_header_row/_row_char_stats）は ailine.py 側（cmd_stack）が既存のものを1回だけ
   呼び、その結果を値としてこの module へ渡す（cmd_scan と同じ配線）。
   ★ 既存部品の再利用: 分母・基準ファイル方式・3判定は multifile.py、合計行の識別と
   閉じる検査は total_row.py（どちらも既存・単位L で完成済み）。この module は
   「積む行の決定」と「出所列つきで積む」という stack 固有の配線だけを持つ。
"""
from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass, field
from xml.etree import ElementTree as ET

import openpyxl

from ailine_core import inspection, multifile, total_row, xml_readback
from ailine_core.filetypes import OPENPYXL_READABLE_SUFFIX
from ailine_core.primitives import fmt_num

PROVENANCE_HEADERS = ("元ファイル", "元行")
# ★ 赤2 の直し（2026-08-21 実機敵対検分）: 署名判定はサフィックス形（元ファイル_2 等）も
#   「自分」と認める。素の名前そのもの、または末尾に _数字 が付いた形のどちらにも当たる。
_PROVENANCE_SIGNATURE_RE = tuple(re.compile(rf"^{re.escape(name)}(_\d+)?$")
                                 for name in PROVENANCE_HEADERS)

# ★ jisaku-review#1 critical の直し: 署名を「列名だけ」から「列名 AND 書き手の印」に。
# stack が出力を書く時に docProps/core.xml の dc:creator へこの印を残す
# （cmd_stack が wb.properties.creator = CREATOR_MARK を設定）。
CREATOR_MARK = "ailine stack"
# ★ architect 致命2 の直し（M2 前置き・2026-08-21）: 「ailine の出力か」は stack 1本だけでなく
# ailine の複数コマンドに広がる。書く側の定数（CREATOR_MARK）は互換のためそのまま残し、
# 読む側の判定はこの集合で行う（stack・extract・そして P 先行 commit で match を追加）。
# ★ CSV 検疫接続（2026-08-22）: `ailine csv` の出力を足す。ailine_core/verify.py の
# _CREATOR_MARKS にも同時に足す（tests/test_stack_e2e.py の同期番人が二重管理のずれを見る）。
CREATOR_MARKS = {"ailine stack", "ailine extract", "ailine match", "ailine csv"}
# ★ M3 P 先行 commit（DESIGN-20260821-multifile.md M3 設計 v2）: match の集約出力
# （1行=1キー）は末尾2列の出所列署名を構造的に持てない。1枚目シート名+固定見出しで判定する。
MATCH_SHEET_NAME = "照合"
MATCH_HEADERS = ("キー", "A側 件数", "A側 合計", "B側 件数", "B側 合計", "差額", "状態")
_CORE_NS = {"cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
           "dc": "http://purl.org/dc/elements/1.1/"}


def col_a_mismatch_is_explained(col_a_count: int, used_range_count: int,
                                 excluded_blank_label_rows: int) -> bool:
    """1列目の非空行数と表の範囲の差が、**合計行として除外した行**で説明できるか。

    ★ なぜ在るか（盲検の査定・2026-08-24）: 小計行のある請求書 3 冊すべてに
      「1列目から数えると 2 行ですが、表の範囲は 3 行あります」が出た。原因は
      **自分が正しく除外した小計行**。日本の請求書は「小計」を金額の隣（右寄せ）に書き、
      1 列目は空にするのが最も普通の形なので、**普通の請求書で必ず鳴る**警告だった。
      オオカミ少年防止を謳う道具が、自分でオオカミ少年になっていた。
    ★ 説明に使えるのは**1列目が空の**除外行だけ ── ラベルが 1 列目に在る合計行は
      col_a_count に既に数えられているので、差の説明にならない（恒真にしない）。
    """
    gap = used_range_count - col_a_count
    if gap <= 0:
        return True
    return gap <= excluded_blank_label_rows

def _is_blank(v) -> bool:
    return total_row._is_blank_cell(v)


def _row_has_any_value(ws, row: int, num_cols: int) -> bool:
    """row の 1..num_cols 列のどこかに値があるか（★ adopted_rows とは別の判定 ──
       金額が空欄でも他の列に値があれば『データ行』とみなす。凍結検体
       test_data_row_with_empty_numeric_cell_is_still_stacked の配線）。"""
    return any(not _is_blank(ws.cell(row=row, column=c).value) for c in range(1, num_cols + 1))


def own_output_headers(headers: list) -> list:
    """利用者の列名と衝突したら機械的サフィックス（元ファイル_2 等）で逃がす。
       戻り値: 実際に使う出所列名2本（衝突が無ければ PROVENANCE_HEADERS そのまま）。"""
    out = []
    for name in PROVENANCE_HEADERS:
        candidate = name
        n = 2
        while candidate in headers or candidate in out:
            candidate = f"{name}_{n}"
            n += 1
        out.append(candidate)
    return out


def is_own_signature(headers: list) -> bool:
    """headers の末尾2列が出所列の見出し（素の名前 または 衝突時のサフィックス形
       元ファイル_N / 元行_N）と一致するか。★ 列名だけの一致であり、これ単独では
       『自分の出力』の証明にならない（jisaku-review#1 実測: たまたま同じ列名の
       人のファイルを誤認する）── 呼び出し側は `is_own_output` を使うこと。"""
    if len(headers) < 2:
        return False
    return all(pat.match(str(h)) for pat, h in zip(_PROVENANCE_SIGNATURE_RE, headers[-2:]))


def _read_creator(path) -> str | None:
    """docProps/core.xml の dc:creator を直読み（zip 直読み・軽い専用の読み）。
       読めなければ None（壊れている/該当なし = 印なし = 他人のファイル扱い ── fail closed）。"""
    try:
        with zipfile.ZipFile(path) as z:
            if "docProps/core.xml" not in z.namelist():
                return None
            root = ET.fromstring(z.read("docProps/core.xml"))
    except Exception:
        return None
    el = root.find("dc:creator", _CORE_NS)
    return el.text if el is not None else None


def _peek_first_sheet_headers(path) -> list | None:
    """1枚目シートの1行目をヘッダーとして覗き見る（読めなければ None）。
       stack/extract kind の列署名判定専用の軽い読み（旧 ailine.py 側 `_peek_headers` を
       この module へ移した ── own_output_mark が path から自分で読む形になったため）。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    try:
        return multifile.read_row_headers(wb.worksheets[0], 1)
    except Exception:
        return None
    finally:
        wb.close()


def _stack_extract_signature(path) -> bool:
    """stack/extract kind の列署名判定: 末尾2列が出所列見出し（サフィックス形含む）と一致するか。
       ★ 既存ロジックの流用（is_own_signature）── 変更なし。"""
    headers = _peek_first_sheet_headers(path)
    return headers is not None and is_own_signature(headers)


def _match_signature(path) -> bool:
    """match kind の列署名判定: 1枚目シート名が MATCH_SHEET_NAME（『照合』）かつ
       見出しが MATCH_HEADERS と一致するか。★ M3 P 先行 commit（M3 設計 v2）: 照合は
       1行=1キーの集約のため、stack/extract のような末尾2列の出所列署名を構造的に
       持てない ── kind ごとに違う判定を KIND_SIGNATURES テーブルへ分ける理由そのもの。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return False
    try:
        ws = wb.worksheets[0]
        if ws.title != MATCH_SHEET_NAME:
            return False
        headers = multifile.read_row_headers(ws, 1)
        return tuple(headers) == MATCH_HEADERS
    except Exception:
        return False
    finally:
        wb.close()


def _csv_signature(path) -> bool:
    """csv kind の列署名判定: CSV 由来の列名は利用者の原本そのまま（任意）のため、
       stack/extract のような末尾2列の出所列署名も、match のような固定シート名/見出しも
       構造的に持てない ── docProps の description が ailine csv の機械可読契約
       （kind:"csv"）と一致するかだけで判定する（署名の代わりに条件を見る）。"""
    _creator, description = xml_readback.read_core_properties(path)
    if not description:
        return False
    try:
        cond = json.loads(description)
    except (TypeError, ValueError):
        return False
    return isinstance(cond, dict) and cond.get("tool") == "ailine" and cond.get("kind") == "csv"


# ★ P 先行 commit（M3 設計 v2）: 署名を「末尾2列」1本槍から kind 別テーブルへ拡張。
# 印（creator）ごとに列署名の判定関数を引く ── 未知の印は握っていない = 他人扱い（fail closed）。
KIND_SIGNATURES = {
    "ailine stack": _stack_extract_signature,
    "ailine extract": _stack_extract_signature,
    "ailine match": _match_signature,
    "ailine csv": _csv_signature,
}


def own_output_mark(path) -> str | None:
    """署名 = kind 別の列署名 **AND** 書き手の印（docProps/core.xml の creator）。
       一致すれば実際の印（例 "ailine stack" / "ailine extract" / "ailine match"）を返す。
       どちらか片方でも欠ければ None（他人のファイル扱い ── fail closed）。
       ★ P 先行 commit（M3 設計 v2）: headers 引数を廃止し path から自分で読む形へ変更 ──
       match の列署名（1枚目シート名+見出し）は headers（列名リストだけ）では判定できない
       （シート名の情報が headers に無い）。呼び出し側は事前に headers を覗いてから渡すのでなく
       この関数へ path だけを渡すこと。
       ★★ jisaku-review#1 critical の直し（実機再現済み）: `is_own_signature`（列名だけ）は
       末尾2列がたまたま『元ファイル』『元行』という名前の人のファイルを前回出力と誤認し、
       --overwrite 無しで無警告上書きしてデータを消した。名前が合っていても印が無ければ
       他人のファイル。
       ★ architect 致命2 の直し: 「自分（ailine の何か）の出力か」（V6 の入力自己参照除外）と
       「作り直してよい前回出力か」（書き込み関所）は問いが違う ── 前者はこの関数の戻り値が
       None でないか、後者は戻り値が CREATOR_MARK と完全一致するか、で呼び出し側が分ける。"""
    creator = _read_creator(path)
    checker = KIND_SIGNATURES.get(creator)
    if checker is None:
        return None
    return creator if checker(path) else None


def is_own_output(path) -> bool:
    """『ailine の何らかのコマンドの出力』か（印は問わない・集合のどれかに当たれば真）。
       ★ V6（入力からの自己参照除外）が使う判定はこれ ── ailine 産は種類を問わず除外する。
       『作り直してよい』（書き込み関所）は別問い ── `own_output_mark` の戻り値を
       CREATOR_MARK と完全一致で見る（呼び出し側 = ailine.py cmd_stack）。"""
    return own_output_mark(path) is not None


# ★ operator 盲検7度目の直し（2026-08-21）: multifile.py へ移した
#   （extract_multi.py とも共有するため）。呼び出し元互換のためこの名前でも参照できる。
numeric_column_names = multifile.numeric_column_names


@dataclass(frozen=True)
class FileStackResult:
    """1ファイルを積んだ（または積めなかった）結果。
       rows: [(base_headers 順の値リスト, 同順の number_format リスト, 元行番号), ...]（積めた時のみ・
       ★ 実視の磨き: number_format は日付セルの時刻の尻尾を消すため元セルから運ぶ）。
       excluded/mismatches: total_row.split_total_rows の戻り値そのまま。
       col_a_mismatch: (col_a_count, used_range_count) 食い違い時のみ（③・可視化専用）。
       findings: M2.5（検分シート）用の inspection.Finding のリスト ── ws を持つこの関数の
       内側で組み立てる（ファイル+シート+セルの3座標が、ここでなら列位置まで正確に引ける）。"""
    name: str
    status: str                      # "積んだ" / "積めなかった"
    reason: str | None = None
    reordered: bool = False
    rows: list = field(default_factory=list)
    excluded: list = field(default_factory=list)
    mismatches: list = field(default_factory=list)
    # ★ 2026-08-24: 値として運べない「中身」（コメント/ハイパーリンク）の件数。
    #   飾り（罫線・塗り）とは分けて数える ── あちらは持ち越さなくて当然だが、
    #   人が打ったメモやリンクは消えたと言わないと気づけない。
    dropped_notes: list = field(default_factory=list)
    col_a_mismatch: tuple | None = None
    sheet_fallback: tuple | None = None   # (wanted, used) ── 基準名のシートが無く1枚目へ落ちた時だけ
    findings: list = field(default_factory=list)   # list[inspection.Finding]（M2.5）


def evaluate_and_stack(path, base_headers: list, base_sheet_name, header_row: int,
                        numeric_col_names: list) -> FileStackResult:
    """1ファイルを基準と照合し、取れていれば『積む行』を確定して値まで読む。
       ★ どんな失敗でも例外を上げず名指し+理由で返す（multifile.evaluate_file と同じ線）。
       ★ operator 盲検7度目の直し（2026-08-21）: 合計行の候補判定を『指定の1本の数値列』
       （旧 value_col_name）から『基準の数値列集合すべて』（numeric_col_names）へ広げる。
       実務標準形（数量・単価つき請求書）は最初の数値列=数量だが、合計の数字は金額列にしか
       無い ── 単一列版は has_number=False で全トリガが沈黙し、Σ が黙って2倍になった。"""
    if path.suffix.lower() != OPENPYXL_READABLE_SUFFIX:
        return FileStackResult(name=path.name, status="積めなかった", reason="旧形式(.xls)")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return FileStackResult(name=path.name, status="積めなかった", reason=f"読み込み失敗: {e}")
    try:
        ws, sheet_fell_back = multifile.find_matching_sheet(wb, base_sheet_name)
        sheet_fallback = (base_sheet_name, ws.title) if sheet_fell_back else None
        other_headers = multifile.read_row_headers(ws, header_row)
        status, detail = multifile.classify_headers(base_headers, other_headers)
        if status == "取れなかった":
            not_taken = [inspection.finding(
                kind=inspection.KIND_NOT_TAKEN, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row), source_value=None, output_value=None,
                next_step=f"見出しが基準と合いません（{detail}）。この冊は積んでいません。")]
            return FileStackResult(name=path.name, status="積めなかった", reason=detail,
                                    sheet_fallback=sheet_fallback, findings=not_taken)
        reordered = bool(detail)

        # 各 base 列 → このファイル自身の列位置（並べ替えファイルで位置がずれる対策）。
        col_for_base = {bh: multifile._column_index(other_headers, bh) for bh in base_headers}
        max_row = ws.max_row or header_row
        all_rows = list(range(header_row + 1, max_row + 1))

        label_col = col_for_base.get(base_headers[0]) if base_headers else None
        value_cols = {name: col_for_base[name] for name in numeric_col_names
                      if col_for_base.get(name)}
        if label_col and value_cols:
            rows_in = []
            for r in all_rows:
                label_val = ws.cell(row=r, column=label_col).value
                vals = {name: ws.cell(row=r, column=idx).value for name, idx in value_cols.items()}
                rows_in.append((r, label_val, vals))
            verdict = total_row.split_total_rows_multi(rows_in)
        else:
            verdict = total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[])

        excluded_rows = {e.row for e in verdict.excluded}
        num_cols = len(other_headers)
        data_rows = [r for r in all_rows if _row_has_any_value(ws, r, num_cols)]
        stack_rows = [r for r in data_rows if r not in excluded_rows]

        rows = []
        dropped_notes = []   # ★ 値として運べない「中身」（コメント/リンク）の記録
        for r in stack_rows:
            cells = [ws.cell(row=r, column=col_for_base[bh]) for bh in base_headers]
            values = [c.value for c in cells]
            # ★ 実視の磨き（2026-08-21）: 元セルの number_format をデータセルへ運ぶ
            #   （日付が『2026-07-09 0:00:00』と時刻付きで出ないように）。決定論
            #   ── 元の書式文字列をそのまま運ぶだけ（乱数・時刻は使わない）。
            formats = [c.number_format for c in cells]
            rows.append((values, formats, r))
            # ★ 2026-08-24: コメントとハイパーリンクは**飾りでなく中身**（「要確認: 入金待ち」
            #   のような、人が打った情報）。縦積みは値と数値書式しか運ばないので黙って
            #   消える ── 消したこと自体は変えずに、消したと言えるように数える。
            for c in cells:
                if c.comment is not None:
                    dropped_notes.append((r, "コメント"))
                if c.hyperlink is not None:
                    dropped_notes.append((r, "ハイパーリンク"))

        col_a_count = sum(1 for r in all_rows if not _is_blank(ws.cell(row=r, column=1).value))
        used_range_count = len(all_rows)
        # ★ 2026-08-24: 差が**合計行として除外した行**で説明できるなら黙る。
        #   1 列目が空の合計行（「小計」を金額の隣に右寄せで書く、日本の請求書で最も普通の形）で
        #   毎回鳴っていた ── オオカミ少年防止を謳う道具が自分でオオカミ少年になっていた。
        excluded_blank_label = sum(
            1 for e in verdict.excluded
            if _is_blank(ws.cell(row=getattr(e, "row", 0) or 0, column=1).value))
        col_a_mismatch = None
        if not col_a_mismatch_is_explained(col_a_count, used_range_count, excluded_blank_label):
            col_a_mismatch = (col_a_count, used_range_count)

        # ★ M2.5: 所見の組み立て（ws がまだ開いているこの関数の内側でだけ、列位置まで
        #   正確な3座標が引ける）。⚠ 相当（閉じる検査の不一致・分母食い違い）とシート
        #   fallback の開示（既存 CLI 報告と同じ基準・inspection.WARN_KINDS）。
        # ★ アンカーは「怪しい数字そのもの」── 単一列版は基準の2列目等へ固定していたが、
        #   複数数値列版では不一致が起きた**その列**（m.column）へ飛ばすのが最も正確
        #   （同じ行が複数列で同時に不一致になりうるため、列ごとに違うセルへ導く）。
        #   着地セルが引けなければ従来どおり label_col へ落ちる（fail closed）。
        # ★ UX 磨き③（Namakoo 実視 2026-08-21 12:01）: 断片（「除外行の値が...」）でなく
        #   1所見1文。状態（両側の数字+動詞）+ 次の手（『クリック』して確認）を1文に言い切る。
        findings = []
        for m in verdict.mismatches:
            m_col = col_for_base.get(m.column) if hasattr(m, "column") else None
            cell = inspection.cell_ref(m_col or label_col or 1, m.row)
            col_label = f"『{m.column}』" if hasattr(m, "column") else ""
            findings.append(inspection.finding(
                kind=inspection.KIND_TOTAL_ROW_MISMATCH, file=path.name, sheet=ws.title,
                cell=cell, source_value=m.excluded_value, output_value=m.adopted_sum,
                next_step=f"合計行({cell}) の{col_label}列の値 {inspection.fmt_num(m.excluded_value)} が"
                          f"明細の和 {inspection.fmt_num(m.adopted_sum)} と合いません。"
                          f"リンクをクリックして {path.name} の {cell} を確認してください"
                          "（除外そのものは維持しています）。"))
        if sheet_fallback:
            findings.append(inspection.finding(
                kind=inspection.KIND_SHEET_FALLBACK, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row),
                source_value=sheet_fallback[0], output_value=sheet_fallback[1],
                next_step=f"基準名のシート『{sheet_fallback[0]}』が見つからないため、"
                          f"1枚目『{sheet_fallback[1]}』を使いました。"
                          "意図したシートか確認してください。"))
        if col_a_mismatch:
            findings.append(inspection.finding(
                kind=inspection.KIND_COL_A_MISMATCH, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row + 1),
                source_value=col_a_mismatch[0], output_value=col_a_mismatch[1],
                next_step=f"A列走査 {col_a_mismatch[0]} 行と used range {col_a_mismatch[1]} 行が"
                          "一致しません。空の書式だけが残った行が無いか確認してください。"))

        return FileStackResult(name=path.name, status="積んだ", reordered=reordered, rows=rows,
                                excluded=verdict.excluded, mismatches=verdict.mismatches,
                                col_a_mismatch=col_a_mismatch, sheet_fallback=sheet_fallback,
                                findings=findings,
                                 dropped_notes=dropped_notes)
    finally:
        wb.close()


def split_own_outputs(candidates):
    """入力候補から **ailine 産の出力**を外し、(残った候補, 外した名前) を返す。

    ★ 2026-08-24（第三波 S1）: この 5 行が cmd_stack と cmd_run_folder に**書き写されて**
    いて、cmd_scan にだけ無かった。実測: 2 冊照合の出力が入力フォルダに残ったまま
    `ailine scan` を掛けると「3 ファイル中 2 照合できた」と分母が汚れ、自分の出力を
    「取れなかった」ファイルとして ⚠ で名指ししていた（stack は正しく除外していた）。
    ★ 三度目の書き写しをせず 1 箇所にする ── 同じ形のバグを何度も直さないために。
    """
    kept, excluded_names = [], []
    for path in candidates:
        if is_own_output(path):
            excluded_names.append(path.name)
        else:
            kept.append(path)
    return kept, excluded_names
