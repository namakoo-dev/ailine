"""inspection — 検分シート（出力ブック内の視覚的誘導）。
   DESIGN-20260821-multifile.md M2.5 節 + 追補（Namakoo 決裁 2026-08-21 10:2x「374行目が
   怪しいです、と言われてもそれを追うのは負担」「視覚的な表示はデモとして効果的」）。

   ★ 原則: 原本には1ピクセルも塗らない。視覚は全部こちら（出力ブック）の内側だけ。
   ★ 住所の3座標（M2.5 追補）: 全所見は ファイル+シート+セル で完全修飾する。
     `Finding` の `sheet` はデフォルト無しの必須引数 ── 行番号だけの所見（シート抜き）を
     組み立てられない、という型の強制（呼び出し側が sheet を省略すると TypeError）。
   ★ ✓ の絶対性（憲法2 の適用拡張）: この module 自身は ✓ を作らない。要約に並べる分母・
     Σ・行の完全会計は、呼び出し側（stack.py/extract_multi.py/ailine.py）が事後条件の
     結果オブジェクトから既に計算した数字を、そのまま並べるだけ。凡例に「色は『怪しい』の
     印であって検証の主張ではない」と明記する。
   ★ 決定論: 時刻・乱数・PC名を一切使わない。セルコメントの author は固定文字列
     （COMMENT_AUTHOR）。同一入力なら検分シートの内容（値・色・コメント）も完全一致する。
   ★ 実装: openpyxl のみ（塗り/コメント/HYPERLINK）。LibreOffice は起動しない。
   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）。
"""
from __future__ import annotations

import datetime
import os
from dataclasses import dataclass

from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from ailine_core.primitives import fmt_num

SHEET_NAME = "検分"
COMMENT_AUTHOR = "ailine"                 # ★ 決定論: 時刻・ユーザー名を使わない固定文字列

# ★ UX 磨き①（Namakoo 実視 2026-08-21 12:01「薄くて気づかない」で淡黄を却下）:
#   Excel の条件付き書式が『悪い値』に使う慣習色（薄赤 FFC7CE）を借りる ── 説明ゼロで
#   意味が読める色。1色のみ・正常行は一切塗らない、という原則は変えない。
_TINT_RGB = "FFFFC7CE"
TINT_FILL = PatternFill(start_color=_TINT_RGB, end_color=_TINT_RGB, fill_type="solid")

HEADER_FONT = Font(bold=True)             # ★ UX 磨き③: 見出し行は太字（自分のブックなので可）

# ★ UX 磨き②: 列幅の機械算出（内容由来・決定論・乱数や現在時刻を使わない）。
#   openpyxl は未設定の列でも既定 13.0 を返す（恒真の罠）── 必ず内容から明示設定する。
_COL_WIDTH_MIN = 8
_COL_WIDTH_MAX = 60
_COL_WIDTH_PADDING = 2
# 全角/CJK 相当（東アジアの文字幅=2・その他は1）とみなす大まかなコードポイント帯。
_WIDE_RANGES = (
    (0x1100, 0x115F), (0x2E80, 0xA4CF), (0xAC00, 0xD7A3), (0xF900, 0xFAFF),
    (0xFF00, 0xFF60), (0xFFE0, 0xFFE6), (0x20000, 0x3FFFD),
)
# ★ jisaku-review4戦目 F2（minor）の直し: date/datetime の表示相当の固定幅。
#   str(datetime.datetime(...)) は number_format が時刻を隠していても
#   『2026-07-09 00:00:00』のように時刻の尾まで数え、列が過大（実測: 幅21）になっていた。
#   セルの実際の見た目（number_format 適用後）は測らず、代表的な日付表示相当の固定幅で数える。
_DATE_DISPLAY_WIDTH = 10


def _char_width(ch: str) -> int:
    cp = ord(ch)
    return 2 if any(lo <= cp <= hi for lo, hi in _WIDE_RANGES) else 1


def _display_width(v) -> int:
    """CJK・全角は2幅、半角は1幅として数える表示幅。
       ★ date/datetime は str() の生の長さでなく固定幅（_DATE_DISPLAY_WIDTH）で数える。"""
    if isinstance(v, datetime.date):   # datetime.datetime は date のサブクラス・両方拾う
        return _DATE_DISPLAY_WIDTH
    return sum(_char_width(ch) for ch in ("" if v is None else str(v)))


def autosize_columns(ws, min_width: int = _COL_WIDTH_MIN, max_width: int = _COL_WIDTH_MAX,
                      padding: int = _COL_WIDTH_PADDING) -> None:
    """列幅を内容（実際に書いたセルの値）だけから機械算出して明示設定する。
       ★ 実害の現物（Namakoo 実視）: 日付列が ### で潰れ、検分のファイル名・所見が
       見切れていた ── 内容の最大表示幅 + 余白で clamp（下限/上限つき）。
       ★ 決定論: 乱数・現在時刻は使わない（同一内容なら同一幅）。"""
    widths: dict = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), _display_width(cell.value))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, w + padding))


def bold_row(ws, row_idx: int, num_cols: int) -> None:
    """見出し行を太字に（UX 磨き③・自分のブックなので書式適用可）。"""
    for c in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=c).font = HEADER_FONT

# 所見の種類（stack.py / extract_multi.py と共有する語彙 ── 手書きの表記ゆれを避ける）。
KIND_NOT_TAKEN = "取れなかった"
KIND_TOTAL_ROW_MISMATCH = "合計行の閉じる検査 不一致"
KIND_COL_A_MISMATCH = "分母食い違い(A列/used range)"
KIND_SHEET_FALLBACK = "シート fallback"
# ★ ⚠ 付きファイル判定（憲法1 相当・淡色+コメントの対象）: 閉じる検査の不一致・分母食い違い
#   だけを「怪しい」とする。取れなかった（行そのものが無い）・シート fallback（開示のみ・
#   render_stack_report でも ⚠ を付けていない）は対象外 ── 既存 CLI 報告の ⚠ 基準と揃える。
WARN_KINDS = frozenset({KIND_TOTAL_ROW_MISMATCH, KIND_COL_A_MISMATCH})

_FINDINGS_HEADER = ["種類", "ファイル", "シート", "セル", "元の値", "採用/出力側の値", "次の手", "リンク"]
_FILEMAP_HEADER = ["ファイル", "使ったシート", "備考"]

_LEGEND = [
    "凡例:",
    "・色は「怪しい」の印であって検証の主張ではありません（検証は Σ・行数などの数字が言います）。",
    "・リンクは Ctrl を押しながらクリックで開きます（LibreOffice の既定。Excel は設定によりそのままクリック）。",
    "・リンクは相対パスです。フォルダごと一緒に移動すれば生きています。ファイル単体だけを移動すると切れます。",
    "・原本には一切印を付けません（塗り・コメントとも出力ブック側だけです）。"
    "リンクの着地セル（クリックすると選択された状態で開きます）が対象のセルです。",
]


def cell_ref(col: int, row: int) -> str:
    """1起点の (列, 行) → Excel セル表記（例 (1,5) → "A5"）。"""
    return f"{get_column_letter(col)}{row}"


@dataclass(frozen=True)
class Finding:
    """1件の所見。★ sheet は必須（デフォルト無し）── M2.5 追補「住所の3座標」を
       型で強制する（file だけ・row だけの所見を組み立てられない）。
       link: ★ M3 gap#1 の直し（2026-08-21）。None（既定）なら従来どおり file/sheet から
       『元ファイルへの外部相対パス』を組み立てる（stack.py/extract_multi.py はこちら・
       挙動は一切変えない）。(target, location) を渡すとそれを**そのまま**使う ──
       target=None ならブック内リンク（例: 明細シートの該当行）になる。M3 は2冊が別
       フォルダでも切れないブック内リンクを使うためにこれを渡す。"""
    kind: str
    file: str
    sheet: str
    cell: str
    source_value: object
    output_value: object
    next_step: str
    link: tuple | None = None


def finding(kind: str, file: str, sheet: str, cell: str, source_value, output_value,
            next_step: str, link: tuple | None = None) -> Finding:
    """Finding を作る唯一の入口。sheet を渡さないと Finding() 自体が TypeError になる
       （キーワード省略も不可 ── デフォルト値を持たない dataclass フィールドの性質）。
       link は省略可（既定 None・従来どおりの外部相対パスリンク）。"""
    return Finding(kind=kind, file=file, sheet=sheet, cell=cell,
                    source_value=source_value, output_value=output_value, next_step=next_step,
                    link=link)


def describe(f: Finding) -> str:
    """所見1件を1行のテキストにする（セルコメント・要約の両方が使う共通の言い方）。
       ★ 両側の数字つき（信用の条件④・感想で終わらせない）。"""
    return (f"{f.kind}: {f.sheet}!{f.cell}  元 {fmt_num(f.source_value)} / "
            f"採用 {fmt_num(f.output_value)}（{f.next_step}）")


def hyperlink_target_path(out_dir, source_path) -> str:
    """出力ブックから元ファイルへの相対パス（OOXML relationship の Target・URI・角括弧なし）。
       ★ v2 直し（2026-08-21・実 XML 検分）: 初版設計の `[相対パス]シート名!セル` は
       Target 1本に丸ごと詰め込む形で、これは HYPERLINK 関数式の構文であって
       relationship の Target ではない ── Excel が解決できない不正な URI になっていた
       （relationship の Target=相対パス・TargetMode=External、hyperlink 要素側に
       location を持つ、というのが OOXML の正式形）。
       ★ 相対パス前提 ── フォルダごと移動すれば生存・単品移動で切れる（凡例に1行）。"""
    rel = os.path.relpath(str(source_path), start=str(out_dir))
    return rel.replace(os.sep, "/")


def hyperlink_location(sheet: str, cell: str) -> str:
    """hyperlink 要素の location（ブック内の『飛び先』・シート名は `'...'` で引用）。
       ★ 日本語シート名の安全側（スペース・記号を含むシート名でも Excel が解釈できる形）。"""
    return f"'{sheet}'!{cell}"


def tint_row(ws, row_idx: int, num_cols: int, comment_col: int, reason_lines: list) -> None:
    """⚠ 付きファイル由来のデータ行を『帯』で塗る（Namakoo 実視2巡目 12:14: 出所セル1個は
       目が拾わない）── 1〜num_cols 列（見出しは含まない・出所列2本も含む）を全部薄赤に。
       理由コメントは従来どおり出所セル1個だけに付ける（重複コメントは作らない）。
       正常行はこの関数を一度も呼ばない（呼び出し側が suspicious なファイルの行だけ選ぶ）。"""
    for c in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=c).fill = TINT_FILL
    ws.cell(row=row_idx, column=comment_col).comment = Comment("\n".join(reason_lines), COMMENT_AUTHOR)


def denominator_lines(denominator: int, processed: int, contributing: int, verb: str) -> list:
    """3段分母: フォルダ内候補 → 照合できた(processed) → 実際に行を持ち込んだ(contributing)。"""
    return [f"{denominator} ファイル中 {processed} 冊 {verb} → "
            f"うち {contributing} 冊が実際に行を持ち込みました"]


def row_accounting_lines(adopted: int, excluded: int, not_taken_files: int,
                          unmatched: int | None = None) -> list:
    """憲法⑨ 行の完全会計: データ行数 = 採用/一致 + (不一致 +) 除外（合計行）。
       取れなかったファイルの中身は『未実施』と開示する（黙る未実施は禁止・憲法2 の一括報告）。

       ★ 実弾検分（2026-08-21 11:4x）で発覚: 抽出（run）経路は『条件に合わなかった行』
       （不一致）の勘定が丸ごと落ちたまま『処理したデータ行 = 採用 + 除外』という、
       式の中では辻褄が合って見える文を出していた（実数はもう1勘定多い）── 憲法2 違反。
       unmatched を渡す経路（抽出）は3勘定・渡さない経路（縦積み ── 不一致という区分自体が
       存在しない）は2勘定のまま。どちらも**渡された勘定の総和だけ**を言う式にする
       （呼び出し側が持っていない勘定をここで作らない・手で再計算もしない）。"""
    if unmatched is None:
        total = adopted + excluded
        lines = [f"行の完全会計: 処理したデータ行 {total} = 採用 {adopted} + 除外(合計行) {excluded}"]
    else:
        total = adopted + unmatched + excluded
        lines = [f"行の完全会計: 処理したデータ行 {total} = "
                 f"一致 {adopted} + 不一致 {unmatched} + 除外(合計行) {excluded}"]
    if not_taken_files:
        lines.append(f"（取れなかった {not_taken_files} 冊は中身の検査が未実施）")
    return lines


def _set_finding_row(ws, row_idx: int, f: Finding, out_dir, source_dir) -> None:
    """所見1行を書く。★ Namakoo 実視2巡目（12:14）: 所見の行そのものも薄赤で塗る
       （種類〜リンクの実セル範囲・_FINDINGS_HEADER の列数ぶん）── 開いた瞬間に目が行く。
       検分シートの所見は定義上すべて「見るべきもの」なので、見出し・要約・凡例の行は
       塗らない一方、所見テーブルの行はどの種類でも塗る（データ面の『正常行は塗らない』
       という原則とは対象が違う ── ここは最初から異常だけが並ぶ表）。"""
    ws.cell(row=row_idx, column=1, value=f.kind)
    ws.cell(row=row_idx, column=2, value=f.file)
    ws.cell(row=row_idx, column=3, value=f.sheet)
    ws.cell(row=row_idx, column=4, value=f.cell)
    ws.cell(row=row_idx, column=5, value=f.source_value)
    ws.cell(row=row_idx, column=6, value=f.output_value)
    ws.cell(row=row_idx, column=7, value=f.next_step)
    if f.link is not None:
        # ★ M3 gap#1: 上書き経路。target=None ならブック内リンク（例: 明細!A5）。
        target, location = f.link
    else:
        target = hyperlink_target_path(out_dir, source_dir / f.file)
        location = hyperlink_location(f.sheet, f.cell)
    link_text = f"{target} > {location}" if target else location
    link_cell = ws.cell(row=row_idx, column=8, value=link_text)
    link_cell.hyperlink = Hyperlink(ref=link_cell.coordinate, target=target, location=location)
    for c in range(1, len(_FINDINGS_HEADER) + 1):
        ws.cell(row=row_idx, column=c).fill = TINT_FILL


def build_sheet(wb, *, findings: list, denominator_lines_: list, accounting_lines: list,
                 sums: dict, file_sheet_map: list, out_dir, source_dir) -> None:
    """出力ブックへ『検分』シートを追加する（既存シートの後ろ・2枚目）。

       findings: list[Finding] ── 所見の表そのもの。
       denominator_lines_ / accounting_lines: 呼び出し側が既に計算した文字列
         （3段分母・行の完全会計）── ここでは手書きの ✓ を作らない（数字を並べるだけ）。
       sums: {列名: {"source": float, "output": float}}（数値列ごとの Σ 両側）。
       file_sheet_map: [(ファイル名, 使ったシート名, 備考), ...]（sheet_fallback の開示込み）。
       out_dir: 出力ブックが最終的に置かれるディレクトリ（HYPERLINK の相対パス起点）。
       source_dir: 元ファイルが置かれているフォルダ（Finding.file から元パスを解決する）。"""
    ws = wb.create_sheet(title=SHEET_NAME)

    row = 1
    ws.cell(row=row, column=1, value="■ 検分")
    row += 2

    ws.cell(row=row, column=1, value="要約")
    row += 1
    for line in list(denominator_lines_) + list(accounting_lines):
        ws.cell(row=row, column=1, value=line)
        row += 1
    for col_name, both in sums.items():
        ws.cell(row=row, column=1,
                value=f"Σ{col_name}: 元 {fmt_num(both['source'])} / 出力 {fmt_num(both['output'])}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="ファイル → 使ったシート")
    row += 1
    for c, h in enumerate(_FILEMAP_HEADER, start=1):
        ws.cell(row=row, column=c, value=h)
    bold_row(ws, row, len(_FILEMAP_HEADER))
    row += 1
    for fname, sheet_used, note in file_sheet_map:
        ws.cell(row=row, column=1, value=fname)
        ws.cell(row=row, column=2, value=sheet_used)
        ws.cell(row=row, column=3, value=note)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="所見")
    row += 1
    for c, h in enumerate(_FINDINGS_HEADER, start=1):
        ws.cell(row=row, column=c, value=h)
    bold_row(ws, row, len(_FINDINGS_HEADER))
    row += 1
    for f in findings:
        _set_finding_row(ws, row, f, out_dir, source_dir)
        row += 1
    row += 1

    for line in _LEGEND:
        ws.cell(row=row, column=1, value=line)
        row += 1

    autosize_columns(ws)   # ★ UX 磨き②: 検分シート自身も内容から列幅を機械算出
