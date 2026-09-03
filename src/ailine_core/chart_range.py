"""chart_range — グラフのデータ範囲から合計行を除く（operator 盲検10度目 ①）。

★ 実測: 部門3行+合計行の表に棒グラフを作ると、データ範囲 `集計!$A$2:$A$5` に合計1750が
第4の柱として混入し、管理者が総額を部門と誤読する形になっていた。stack/extract は
total_row.py の判定で合計行を既に除外しているのに、CHART のデータ範囲だけがその判定に
配線されていなかった（片配線）。

★ ここは total_row.py の既存判定（split_total_rows_multi）をそのまま再利用する ──
新しい合計行判定を書かない（二重実装の禁）。この module がやるのは「実ファイルを読んで
total_row.py が食える形の行データを組み立てる」層だけ。

★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）。openpyxl だけに
依存する（total_row.py 自体は純ロジックのまま・ここが唯一 openpyxl でファイルを読む層）。
"""
from __future__ import annotations

import openpyxl

from ailine_core import total_row
from ailine_core.primitives import is_number as _is_number


def chart_data_last_row(path, sheet_name: str | None = None, header_row: int = 1) -> int:
    """path のシート sheet_name（省略時は先頭シート）を読み、グラフのデータ範囲として
       使ってよい最終行（1起点の Excel 行番号）を返す。

       total_row.split_total_rows_multi が合計行として除外する行（ラベル語/ラベル空白/
       直上空行のいずれか）は最終行の候補から外す。合計行が無ければ、物理データの
       最終行をそのまま返す（★ 誤爆防止: 合計行が無い表では最終行まで使う）。

       ★ ラベル列は先頭列（total_row.py の他の呼び出し箇所と同じ前提）。データが
       1行も無ければ header_row をそのまま返す（呼び出し側が Exit Sub 相当で扱える）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        max_row = ws.max_row or header_row
        max_col = ws.max_column or 1
        rows = []
        for r in range(header_row + 1, max_row + 1):
            label = ws.cell(row=r, column=1).value
            values = {}
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if _is_number(v):
                    values[c] = v
            rows.append((r, label, values))
        verdict = total_row.split_total_rows_multi(rows)
        if verdict.adopted_rows:
            return max(verdict.adopted_rows)
        return max_row if max_row >= header_row + 1 else header_row
    finally:
        wb.close()
