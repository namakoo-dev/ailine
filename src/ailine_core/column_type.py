"""column_type — 列の実体（書換前）から数値/文字列を機械決定する（operator 盲検10度目 ④）。

★ 実測: SET_COLUMN_VALUE は常に setString で書いていたため、数値列（『数量』等）へ
「10」を一括書換すると `'10'`（書式 `@`）の文字列になり、それを参照する SUM が黙って
壊れる。対象列の**書換前**の実体（データ行の非空セルが全部数値）が数値列で、かつ
書き込む値が数値として読めるなら setValue（数値）で書くべき ── それ以外は従来どおり
setString（A' 原則: 型を LLM に決めさせず、対象列の実体+書く値の両方から機械決定する）。

★ ailine を import しない（移植可能性の番人）。openpyxl のみに依存する。
"""
from __future__ import annotations

import openpyxl


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def column_is_all_numeric(path, sheet_name, col_idx: int, header_row: int = 1) -> bool:
    """col_idx は1起点の列番号。header_row の直下から最終データ行まで走査し、
       非空セルが1つ以上あり、かつ全部数値なら True。非空セルが1つも無い列は False
       （実体不明の列を安全側=数値でないほうへ倒す）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        max_row = ws.max_row or header_row
        saw_any = False
        for r in range(header_row + 1, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None or v == "":
                continue
            saw_any = True
            if not _is_number(v):
                return False
        return saw_any
    finally:
        wb.close()


def value_parses_as_number(value) -> float | None:
    """value（str/int/float）が数値として読めれば float を返す。桁区切りのカンマは許容
       （ailine_core.formula_health._parses_as_number と同じ緩さ）。読めなければ None。"""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", ""))
        except ValueError:
            return None
    return None
