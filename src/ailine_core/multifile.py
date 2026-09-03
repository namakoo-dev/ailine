"""multifile — M1読み: `ailine scan <folder>` の棚卸しロジック（書き込みゼロ）。
   DESIGN-20260821-multifile.md v2 §1(M1読み)・§2(骨)・§3(③④分母/原本無変更)。

   ★ LibreOffice は一切起動しない（openpyxl のみで完結・読むだけ）。
   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定（detect_header_row/_row_char_stats）は ailine.py 側（cmd_scan）が既存のものを1回だけ
   呼び、その結果（header_row・base_headers）をこのモジュールの関数へ値として渡す。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from ailine_core.filetypes import CSV_SUFFIX, OPENPYXL_READABLE_SUFFIX, SCAN_CANDIDATE_SUFFIXES
from ailine_core.primitives import column_index as _column_index

_TEMP_PREFIX = "~$"                      # Excel の一時ファイル（開いている間だけ現れる隣接ファイル）
_MAX_HEADER_COLS = 200                   # 見出し行を読む安全上限（ailine.py の MAX_COLS とは独立）


# ★ ailine 自身が置く作業ファイル（分母に数えない ── 人の資料ではない）。
_AILINE_WORKFILES = frozenset({
    "history.jsonl", "run.lock", "vocab.json", "aliases.json",
    "misclass.jsonl", "notice_v2_shown",
})
_AILINE_WORKFILE_SUFFIXES = frozenset({".lock", ".jsonl"})

def classify_folder_contents(folder: Path):
    """folder 直下（サブフォルダの中は見ない）を分類する。
       戻り値: (candidates: 名前順の Path リスト, excluded: {"temp": n, "subdirs": n, "csv": n})。
       ★ 分母そのものが検証対象（V7）── ~$ 一時ファイルとサブフォルダは対象外として数える
       （1件以上あれば呼び出し側が1行ずつ開示する）。
       ★ CSV 検疫接続（2026-08-22・設計 v2「フォルダ実行」節）: .csv は候補にしない点は
       変わらない（1本ずつ `ailine csv` で扱う対象）が、以前は「その他の拡張子は黙って
       無視してよい」に紛れて数えてすらいなかった ── 名指しで断れるよう分母だけ数える
       （挙動の本体は変えない・報告の材料が増えるだけ）。

       ★★ 2026-08-24 の訂正: 旧版はここで「それ以外の拡張子は引き続き黙って無視してよい」と
       **宣言していた**。盲検 2 者が独立に、それが誤りだと実測で示した ── `.xlsm` を混ぜた
       6 冊のフォルダで「3 ファイル中 3 照合できた」「Σ金額 元 4500 / 出力 4500 ✓」と出る。
       **3 冊が無かったことになる。** マクロ入りの請求書テンプレは実際の経理フォルダで
       最も在りうる非 .xlsx で、「黙って無視してよい」対象ではなかった。
       ★ 根は 1 つ ── **分母を「処理できたもの」から作っていた**。分母は
       **フォルダに実際に在るもの**から作る。処理できなかったものは 0 件ではなく
       **名前つきの件数**として分母に残す。"""
    candidates = []
    excluded = {"temp": 0, "subdirs": 0, "csv": 0, "other_format": 0,
                 "other_format_names": []}
    for item in sorted(folder.iterdir(), key=lambda p: p.name):
        if item.is_dir():
            excluded["subdirs"] += 1
            continue
        if not item.is_file():
            continue
        if item.name.startswith(_TEMP_PREFIX):
            excluded["temp"] += 1
            continue
        suffix = item.suffix.lower()
        if suffix in SCAN_CANDIDATE_SUFFIXES:
            candidates.append(item)
        elif suffix == CSV_SUFFIX:
            excluded["csv"] += 1
        elif item.name in _AILINE_WORKFILES or item.suffix.lower() in _AILINE_WORKFILE_SUFFIXES:
            # ★ 2026-08-24: ailine 自身の作業ファイル（history.jsonl / run.lock 等）を
            #   「読めない形式」に数えない。実測で `history.jsonl` が
            #   「.xlsx に保存し直すと扱えます」と案内されていた ── **自分が置いたものを
            #   他人の資料と同じに扱っていた**。分母に入れるのは
            #   「人が置いた、扱えなかったもの」だけ。
            pass
        else:
            # ★ 候補にしなかった拡張子は**名前つきで**分母に残す（.xlsm/.xlsb/.ods 等）。
            #   件数だけでは人は動けないので、どのファイルかを言えるようにする。
            excluded["other_format"] += 1
            excluded["other_format_names"].append(item.name)
    return candidates, excluded


def open_base_workbook(candidates):
    """基準ファイル方式: パス辞書順（呼び出し側で名前順に並べ済み）で最初に読めた .xlsx を
       基準にする。戻り値: (path, workbook) または、読める .xlsx が1つも無ければ (None, None)。
       ★ .xls は openpyxl で開けないため基準になれない（読めたものだけが資格を持つ）。"""
    for path in candidates:
        if path.suffix.lower() != OPENPYXL_READABLE_SUFFIX:
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        return path, wb
    return None, None


def read_row_headers(ws, header_row: int) -> list:
    """header_row（1起点）の見出し名を読む。

    ★★ 2026-08-24 の根治（盲検の実データ耐性レビュー）: 旧版は「先頭列から**連続する**
    非空セル」で切っていた（docstring に「第一波は単純に ── 列の間に空白を挟む見出しは
    扱わない」と当時の割り切りが在った）。結合セルは 2 つ目以降が None なので、
    **`A1:B1` を結合した請求書で見出しがそこで途切れ、その先の列が丸ごと消えていた**。
    実測: `商品情報`(A1:B1 結合) / `数量` / `金額` の表を stack すると
    出力が `('品名','規格','元ファイル','元行')` になり、**数量と金額が列ごと消え、
    しかも exit 0・⚠ なし・Σ金額 の行も出ない**（消えたこと自体が画面に現れない）。
    日本の請求書で結合見出しはほぼ普遍なので、「扱わない」で済む形ではなかった。

    ★ 直し方: **末尾の余白では止まり、途中の空セルでは止まらない。**
      ① 使われている最終列まで見る（`ws.max_column`・上限は _MAX_HEADER_COLS）
      ② 結合セルの中は**左上の値を引き継ぐ**（Excel の見た目と一致させる）
      ③ 引き継ぎ元も無い空セルは、位置を保つために空文字で埋める
         （列位置がずれると、やる側と見る側で別の列を指す事故になる）
      ④ 末尾の空は落とす（無限に伸ばさない）
    """
    merged_value = {}
    for rng in getattr(ws, "merged_cells", None).ranges if getattr(ws, "merged_cells", None) else []:
        if rng.min_row <= header_row <= rng.max_row:
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            if top_left not in (None, ""):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_value[c] = str(top_left)
    last = min(int(getattr(ws, "max_column", 0) or 0), _MAX_HEADER_COLS)
    headers = []
    for c in range(1, last + 1):
        v = ws.cell(row=header_row, column=c).value
        if v in (None, ""):
            v = merged_value.get(c, "")
        headers.append(str(v))
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def find_matching_sheet(wb, base_sheet_name: str | None):
    """他ファイルは基準と同名のシートを探し、無ければ最初のシートで照合する（DESIGN §2骨）。
       戻り値: (worksheet, fallback: bool)。fallback=True は「基準名のシートが無く1枚目へ
       落ちた」こと（★ architect 致命5 前段の開示: 呼び出し側がファイル単位の結果に
       sheet_fallback として載せる）。base_sheet_name が無い（基準ブック自身など）場合は
       比較対象が無いので fallback=False のまま。"""
    if base_sheet_name:
        if base_sheet_name in wb.sheetnames:
            return wb[base_sheet_name], False
        return wb[wb.sheetnames[0]], True
    return wb[wb.sheetnames[0]], False


def classify_headers(base_headers: list, other_headers: list):
    """3判定（列名の完全一致のみが根拠・ゆるい寄せはしない）:
       並びまで一致 → ("取れた", None)
       多重集合が一致・順序だけ違う → ("取れた", "並べ替え")
       それ以外 → ("取れなかった", "欠け/余りの名指し")"""
    if other_headers == base_headers:
        return "取れた", None
    if sorted(other_headers) == sorted(base_headers):
        return "取れた", "並べ替え"
    missing = [h for h in base_headers if h not in other_headers]
    extra = [h for h in other_headers if h not in base_headers]
    parts = []
    if missing:
        parts.append(f"欠け: {', '.join(missing)}")
    if extra:
        parts.append(f"余り: {', '.join(extra)}")
    return "取れなかった", "; ".join(parts) if parts else "列名が一致しません"


def numeric_value_column(ws, header_row: int, num_cols: int) -> int | None:
    """基準シートで、見出し行の下で最初に数値が現れる列（1起点）を返す。無ければ None。
       ★ 単位L の配線: 基準ファイルで1回だけ決める。呼び出し側がこの列の**列名**を
       他ファイルへ渡し、各ファイルはその名前を自分の並びで引き直す（_column_index）
       ── 並べ替えファイルで違う列を数える事故を避ける（implementer 申告・検体化済み）。"""
    max_row = ws.max_row or header_row
    for col in range(1, num_cols + 1):
        for row in range(header_row + 1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return col
    return None


def duplicate_header_names(headers: list) -> list:
    """同じ名前が 2 回以上現れる見出し名（出現順・空文字は数えない）。

    ★ なぜ在るか（盲検 2 者が独立に再現・2026-08-24）: 見出し `品番/備考/金額/備考`
      （社内用メモ と 客先提出用）を stack すると、**客先提出用が社内用に化けた**
      ── 出力 `('0012','社内用メモ',1000,'社内用メモ',…)`・exit 0・Σ 一致 ✓。
      しかも**やる側（stack）と見る側（verify）が同じ「名前→列」の辞書を別々に組む**ので、
      照合も同じように潰れて**恒真**になり、誰も気づけない。
    ★ 根は「列を**名前**で引いている」こと。名前は表示のためのラベルであって、
      同一性の担保ではない（同じ名前の列は実在する）。
    ★ 空文字は数えない ── 結合見出しの直しで位置保持のために入れた埋め草であり、
      「同じ名前の列」ではない。
    """
    seen, dupes = {}, []
    for h in headers or ():
        name = str(h)
        if name == "":
            continue
        seen[name] = seen.get(name, 0) + 1
        if seen[name] == 2:
            dupes.append(name)
    return dupes


def formula_columns_without_cache(path, header_row: int, headers: list,
                                   sheet_name=None, limit: int = 200) -> list:
    """データ行に**数式は在るのに値が無い**列の名前。

    ★ なぜ要るか（2026-08-24）: `=B2*C2` の金額列は data_only=True で None になり、
      数値列と見なされないので **Σ金額 の行がそもそも出ない**。
      「Σ が出ない」＝「検算していない」なのに、出ないことが唯一の信号だった。
    ★ 引数が **ws でなくパス**なのは実測の結果: `data_only=True` で開いた ws では
      数式そのものが見えず（値も式も無い空欄に見える）、検出が常に空になった。
      式を見るには data_only=False で開き直すしかない。
    """
    import openpyxl as _op
    if not headers:
        return []
    try:
        wb_f = _op.load_workbook(path, data_only=False)
    except Exception:
        return []
    try:
        ws = (wb_f[sheet_name] if sheet_name and sheet_name in wb_f.sheetnames
              else wb_f.worksheets[0])
        out = []
        last = min(int(getattr(ws, "max_row", 0) or 0), header_row + limit)
        for c, name in enumerate(headers, start=1):
            if str(name) == "":
                continue
            has_value = has_formula = False
            for r in range(header_row + 1, last + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    has_formula = True
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    has_value = True
            if has_formula and not has_value:
                out.append(str(name))
        return out
    except Exception:
        return []
    finally:
        wb_f.close()


def unverified_numeric_columns(headers: list, verified: list, formula_columns=()) -> list:
    """**検算できなかった列**の名前（見出しの順）。

    ★ 根（今日ずっと出ている家系）: 「無いこと」で伝えようとしていた ──
      分母が消える・列が消える・Σ の行が消える。**出ないものは読めない。**
      検算できなかったなら、できなかったと**書く**。
    ★ 対象は「数値として扱われるはずだったのに扱えなかった列」だけ。
      最初から文字の列（品名など）は検算の対象ではないので言わない（誤爆にしない）。
    """
    verified_set = {str(v) for v in (verified or ())}
    formula_set = {str(f) for f in (formula_columns or ())}
    return [str(h) for h in (headers or ())
            if str(h) != "" and str(h) in formula_set and str(h) not in verified_set]


def numeric_column_names(ws, header_row: int, headers: list) -> list:
    """headers（基準ファイルの列名）のうち、データ行のどこかで数値を持つ列名の一覧。
       ★ jisaku-review#3/#6 の直し: Σ 照合・報告を『最初の数値列』1本だけでなく
       全数値列に広げるための土台。★ operator 盲検7度目の直し: stack.py から
       multifile.py へ移した（extract_multi.py とも共有するため ── 合計行検出を
       全数値列に広げる土台が2箇所必要になった。stack.py 側は再輸出 `numeric_column_names
       = multifile.numeric_column_names` で呼び出し元互換を保つ）。"""
    max_row = ws.max_row or header_row
    out = []
    for i, name in enumerate(headers, start=1):
        for row in range(header_row + 1, max_row + 1):
            v = ws.cell(row=row, column=i).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append(name)
                break
    return out


def total_row_candidate_count(ws, header_row: int, label_col: int, value_col: int) -> int:
    """単位L: ラベル列(label_col)・数値列(value_col) で split_total_rows を走らせ、
       除外（合計行候補）の件数を返す（--json の分布測定の口・DESIGN v2.1）。"""
    from ailine_core.total_row import split_total_rows
    max_row = ws.max_row or header_row
    rows = [(r, ws.cell(row=r, column=label_col).value, ws.cell(row=r, column=value_col).value)
            for r in range(header_row + 1, max_row + 1)]
    return len(split_total_rows(rows).excluded)


def evaluate_file(path: Path, base_headers: list, base_sheet_name: str | None, header_row: int,
                   value_col_name: str | None = None) -> dict:
    """1ファイルを基準と照合する。戻り値: {"name", "status", "reason"(取れなかった時),
       "reordered"(並べ替えで取れた時), "total_row_candidates"(取れた時・列名が引ければ)}。
       ★ どんな失敗でも例外を上げず名指し+理由にして返す
       （$0 条件「黙って失敗する」の裏返し ── 報告が成果物）。
       ★ ラベル列・数値列は基準の**列名**で引き当てる（並べ替えファイルでは位置が違う）。"""
    if path.suffix.lower() != OPENPYXL_READABLE_SUFFIX:
        return {"name": path.name, "status": "取れなかった", "reason": "旧形式(.xls)"}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"name": path.name, "status": "取れなかった", "reason": f"読み込み失敗: {e}"}
    try:
        ws, sheet_fallback = find_matching_sheet(wb, base_sheet_name)
        other_headers = read_row_headers(ws, header_row)
        status, detail = classify_headers(base_headers, other_headers)
        entry = {"name": path.name, "status": status}
        if status == "取れなかった":
            entry["reason"] = detail
        elif detail:   # "並べ替え"
            entry["reordered"] = True
        if sheet_fallback:
            entry["sheet_fallback"] = {"wanted": base_sheet_name, "used": ws.title}
        if status == "取れた" and value_col_name is not None and base_headers:
            label_col = _column_index(other_headers, base_headers[0])
            value_col = _column_index(other_headers, value_col_name)
            if label_col is not None and value_col is not None:
                entry["total_row_candidates"] = total_row_candidate_count(
                    ws, header_row, label_col, value_col)
    finally:
        wb.close()
    return entry
