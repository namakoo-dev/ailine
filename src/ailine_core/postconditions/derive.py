"""事後条件 ── **新しいシートや別ファイルを作る** op。

対象 7 op: 集計 / ピボット / 抽出 / 列抽出 / 帳票作成 / 様式写像 / グラフ

★ この群が確かめること: 作られた側が元と**辻褄が合う**こと（件数・合計・
参照した値）。そして**元のシートが無傷**であること ── 新しい表を作る操作で
元表を書き換えたら、それは気づかれにくい事故になる。

★ CHART だけ事後条件を持たない: openpyxl では図形の中身を取り出せず、
読み戻して形を確かめる手段が無い。**穴ではなく正当な例外**として
`tests/test_postcondition_ledger.py` に理由つきで記録してある。

★ この群は「様式を写す」op を含むが、**何を保存するかを宣言していない**
（値・列の順序・列名・行の同一性・書式・式の参照・合計行の意味）。
2026-09-03 の未着手 ── 地図の投影法が「面積・角度・距離のどれを保存するか選ぶ」
のと同じ形の穴で、日誌 project_basrun_ai_line に条件つきで記録した。
"""
from __future__ import annotations

from pathlib import Path
import zipfile

from openpyxl.utils import get_column_letter

from ailine_core import compare_blocked
from ailine_core import report_group
from ailine_core import total_row
from ailine_core.book_view import BookView
from ailine_core.primitives import is_number as _is_number
from ailine_core.report_per_row import compare_report_cells
from ailine_core.table_scan import _cell_ref, _col_index_by_header, _scan_last_col, _scan_last_row, _used_extent, data_extent

from ailine_core.postconditions._shared import PIVOT_CAVEAT, _ZERO_TARGET_REASON, _cells_for_shift, _extract_predicate, _row_as_shown, compare_moved_rows, note_stringy_numbers

def check_aggregate(path: Path, args: dict, header_row: int = 1, use_formula: bool = False) -> tuple:
    """★ W3: header_row は集計元(対象シート。★ 挙動変更#2 より前は常に1枚目だった)の
       見出し行。出力の「集計」シートは SummaryTable ヘルパが毎回新規作成し常に物理1行目が
       見出し（検出対象外・そのまま）。
       ★ W10f 項目1: use_formula のとき分類列/集計列を data_only(計算後の値)側から読む
       （check_compute_column と同型のバグ。前段が式で作った計算列を group_col/value_col
       にした場合、raw 側は式文字列のままで value_col が『数値でない→0扱い』に落ち、
       偽の不一致 fail になっていた。value_col が0扱いされる止血2 の丸めは残すが、
       それは『本当に非数値/空欄』の行専用にする）。
       AGGREGATE は合計という『全行をまたぐ』検証のため、式にキャッシュ値が無く読めない
       行が1件でもあれば期待値そのものが信頼できない。COMPUTE_COLUMN の行独立検証と違い
       部分採点はせず、fail で理由を示して打ち切る（0cf9218 空虚な検証合格の禁止の趣旨）。"""
    with BookView(path) as bv:
        if "集計" not in bv.sheetnames:
            return "fail", "『集計』シートが無い"
        src = bv.sheet(args.get("_target_sheet"))
        gi = _col_index_by_header(src, args["group_col"], header_row=header_row)
        vi = _col_index_by_header(src, args["value_col"], header_row=header_row)
        if gi is None or vi is None:
            return "fail", "分類列/集計列が見つからない"
        expect: dict = {}
        _stringy_seen: list = []
        uncached = 0
        r = header_row + 1
        while src.cell(row=r, column=1).value not in (None, ""):
            k_raw = src.cell(row=r, column=gi).value
            v_raw = src.cell(row=r, column=vi).value
            if use_formula:
                k = bv.cell_value(r, gi, sheet=args.get("_target_sheet"))
                v = bv.cell_value(r, vi, sheet=args.get("_target_sheet"))
            else:
                k, v = k_raw, v_raw
            k_uncached = use_formula and k is None and isinstance(k_raw, str) and k_raw.startswith("=")
            v_uncached = use_formula and not _is_number(v) and isinstance(v_raw, str) and v_raw.startswith("=")
            if k_uncached or v_uncached:   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
                uncached += 1
                r += 1
                continue
            # ★ 塊③（2026-08-25）: 0 に落とす**前**に、別実装で「数字に見えるか」を見る。
            #   判定は変えない（0 として足す）── 変えるのは言うかどうかだけ。
            _stringy_seen.append(v)
            v = v if _is_number(v) else 0   # ★ 止血2: 非数値/None は0扱い（クラッシュさせない）
            expect[k] = expect.get(k, 0) + v
            r += 1
        if uncached:
            return "fail", (f"分類列/集計列に式はあるがキャッシュ値が無く検証できない行が "
                             f"{uncached} 件あり、集計を検証できません"
                             f"（LibreOffice を通していない可能性）")
        note_stringy_numbers(args, _stringy_seen)
        if not expect:
            return "fail", _ZERO_TARGET_REASON   # ★ 止血1: 集計元データが0件を「合格」にしない
        out = bv.sheet("集計")
        seen = set()
        r = 2
        while True:
            k = out.cell(row=r, column=1).value
            if k in (None, "") or k == "合計":
                break
            v = out.cell(row=r, column=2).value
            v = v if _is_number(v) else 0
            if k not in expect or abs(v - expect[k]) > 1e-6:
                return "fail", f"グループ『{k}』の合計が不一致 (期待 {expect.get(k)} 実際 {v})"
            seen.add(k)
            r += 1
    if seen != set(expect.keys()):
        return "fail", "集計に含まれないグループがある"
    return "pass", f"{len(expect)} グループを検証"

def check_extract_columns(path: Path, args: dict, header_row: int = 1,
                           source_book: Path | None = None) -> tuple:
    """EXTRACT_COLUMNS の事後条件。**選んだ列だけが、値も型も保たれて写った**ことを見る。

    ① 出力シートが在り、見出しが宣言した並びそのもの
    ② 行数が元と同じ（列を選ぶ操作で行は減らない）
    ③ 各行・各列の値が、元の対応するセルと一致（型ごと）
    ④ 元シートは 1 セルも変わっていない（読むだけの操作）
    """
    cols = list(args.get("cols") or [])
    dst = str(args.get("_new_sheet") or "")
    src_sheet = args.get("_target_sheet")
    if not cols or not dst:
        return "warn", "残す列か出力シートが分からないため確かめられていません"
    with BookView(path) as bv:
        if dst not in bv.sheetnames:
            return "fail", f"出力シート『{dst}』がありません"
        out = bv.sheet(dst)
        ws = bv.sheet(src_sheet)
        last, wide = data_extent(ws, header_row)
        hdrs = [str(ws.cell(row=header_row, column=c).value or "") for c in range(1, wide + 1)]
        got_head = [str(out.cell(row=1, column=j + 1).value or "") for j in range(len(cols))]
        if got_head != cols:
            return "fail", (f"出力の見出しが宣言と違う（宣言: {'・'.join(cols)} / "
                             f"実際: {'・'.join(got_head)}）")
        idx = [hdrs.index(c) + 1 for c in cols if c in hdrs]
        if len(idx) != len(cols):
            return "fail", "元の表に無い列が宣言されています"
        rows_src = last - header_row
        rows_out = max(0, data_extent(out, 1)[0] - 1)
        if rows_out != rows_src:
            return "fail", (f"行数が合わない（元 {rows_src} 行 → 出力 {rows_out} 行・"
                             "列を選ぶ操作で行は減らないはず）")
        wrong = []
        for r in range(header_row + 1, last + 1):
            for j, ci in enumerate(idx):
                a_v = bv.cell_value(r - header_row + 1, j + 1, dst)
                b_v = bv.cell_value(r, ci, src_sheet)
                if a_v != b_v:
                    wrong.append(f"{r}行目の『{cols[j]}』")
        if wrong:
            return "fail", (f"写した値が元と違います（{'、'.join(wrong[:5])}）")
    if source_book is not None and Path(source_book).exists():
        with BookView(path) as bv2, BookView(source_book) as bvb:
            after_rows = _cells_for_shift(bv2, src_sheet, header_row, last, wide)
            before_rows = _cells_for_shift(bvb, src_sheet, header_row, last, wide)
        st, info = compare_moved_rows(after_rows, before_rows, "元シート")
        if st == "broken":
            return "fail", f"元シートが変わっています（読むだけの操作のはず）── {info}"
    return "pass", (f"『{'・'.join(cols)}』だけを『{dst}』へ写しました"
                     f"（{rows_src} 行・値も型もそのまま・元シートは無変更）")

def check_pivot(path: Path, args: dict, header_row: int = 1) -> tuple:
    """PIVOT の事後条件。①『ピボット』シートが存在 ②xlsx zip 内に本物の DataPilot
       (xl/pivotTables/) が実在するか、を見る（W8b の忠実度ゲートと同じ zip 直読み技法・
       LO を再起動しない）。★ W9 項目4: DataPilot は開き直すたび再描画で値キャッシュが
       変わりうる既知の癖があるため、集計値そのものの照合（check_aggregate 相当）はせず
       構造の存在確認に留める（PIVOT_CAVEAT として案内も添える）。"""
    with BookView(path, read_only=True) as bv:
        has_sheet = "ピボット" in bv.sheetnames
    if not has_sheet:
        return "fail", "『ピボット』シートが無い"
    try:
        with zipfile.ZipFile(path) as z:
            names = [n.lower() for n in z.namelist()]
    except Exception as e:
        return "fail", f"xlsx を zip として読めない: {e}"
    has_pivot_table = any(n.startswith("xl/pivottables/") for n in names)
    if not has_pivot_table:
        return "fail", "『ピボット』シートはあるが DataPilot(ピボットテーブル)の実体が無い"
    return "pass", f"『ピボット』シートと DataPilot を確認（{PIVOT_CAVEAT}）"

def check_extract(path: Path, args: dict, header_row: int = 1,
                   source_book: Path | None = None) -> tuple:
    """EXTRACT の事後条件（コミット 2edcb08「EXTRACT op」―「番人が入場料を徴収した初の op」）。
       ①行数一致 ②値と型の保存 ③両側の網羅（多い/少ないの両方を落とす）を、1つの位置対応
       比較で同時に見る: 元シートを独立に走査して『条件に一致する行』を上から順に集めた
       ものが expected。出力の各行が expected と完全に同じ順・同じ内容で並んでいれば、
       行数も一致し(①)・値も型も保存され(②。取り違えれば != が拾う ── 数値 59400 と
       文字列 '59400' は Python の != では等しくない＝昨夜の全セル文字列化バグを直接殺す)・
       多く含めても少なく埋めても位置がずれて不一致になる(③)。
       ④ 元シートが無変更（source_book が渡された時だけ突き合わせられる・読むだけの op。
       source_book が無ければ①②③だけを見る＝INSERT_ROWS/AUTOFIT と同じ劣化フォールバック）。"""
    dst_name = args.get("_new_sheet")
    col_name = args.get("col")
    cmp = args.get("cmp")
    # ★ 2026-08-24 の片配線の追補: 日付比較のとき value は**表示用の文字列**で、
    #   閾値の実体は _value_serial（verify_dsl_args が積む）。ここが見ていなかったせいで
    #   正しい抽出が「6行中0行が一致 → 出力は5行」という自己矛盾で失敗していた。
    threshold = args.get("_value_serial", args.get("value"))
    date_mode = bool(args.get("_date_compare"))
    if not dst_name:
        return "fail", "出力シート名が決まっていません（verify_dsl_args を経由していない可能性）"
    match = _extract_predicate(cmp, threshold, date_mode=date_mode)

    with BookView(path) as bv:
        src = bv.sheet(args.get("_target_sheet"))
        src_name = src.title
        if dst_name not in bv.sheetnames:
            return "fail", f"出力シート『{dst_name}』が作られていません"
        col_idx = _col_index_by_header(src, col_name, header_row=header_row)
        if col_idx is None:
            return "fail", f"対象列『{col_name}』が元シート『{src_name}』に見つかりません"
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        total = 0
        expected_rows = []
        unmatched_cells = []
        r = header_row + 1
        # ★★ 2026-08-31: 合計行は抜き出さない ── **分母もそこを外す**。
        #   ★ 片側だけ縮めると「行数が期待と不一致」になる（並べ替えで 2 度踏んだ形）。
        #     宣言（_skip_rows）を**同じ 1 箇所**から生成にも検算にも配る。
        _skip = {int(x) for x in (args.get("_skip_rows") or [])}
        while src.cell(row=r, column=1).value not in (None, ""):
            if r in _skip:
                r += 1
                continue
            total += 1
            cell_v = src.cell(row=r, column=col_idx).value
            if match(cell_v):
                expected_rows.append(_row_as_shown(bv, src_name, r, last_col))
            else:
                unmatched_cells.append(cell_v)
            r += 1
        # ★ 第三波 H3（2026-08-24）: 開示専用 ── 判定は上で終わっており 1 ビットも変えない。
        #   フォルダ経路（extract_multi）と**同じ器官**を呼ぶ（片配線への備え）。
        blocked = compare_blocked.scan_column(unmatched_cells, cmp)
        if total == 0:
            return "fail", _ZERO_TARGET_REASON

        out = bv.sheet(dst_name)
        out_rows = []
        r = 2   # 出力は ExtractRows の仕様どおり常に物理1行目が見出し
        while out.cell(row=r, column=1).value not in (None, ""):
            out_rows.append([out.cell(row=r, column=c).value for c in range(1, last_col + 1)])
            r += 1

    denom = f"{total}行中{len(expected_rows)}行が一致"
    denom += compare_blocked.disclosure_inline(blocked, col_name)
    if len(out_rows) != len(expected_rows):
        return "fail", f"{denom} → 出力は{len(out_rows)}行（行数が期待と不一致）"
    for i, (want, got) in enumerate(zip(expected_rows, out_rows), start=1):
        if want == got:
            continue
        for c, (wv, gv) in enumerate(zip(want, got), start=1):
            if wv != gv:
                letter = get_column_letter(c)
                return "fail", (
                    f"{denom} → 出力{i}行目 {letter}列が元と不一致"
                    f"（元 {wv!r}（{type(wv).__name__}） 出力 {gv!r}（{type(gv).__name__}））"
                )

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            src_before = bv_before.sheet(args.get("_target_sheet"))
            src_after = bv_after.sheet(args.get("_target_sheet"))
            last_row_before = _scan_last_row(src_before, header_row=header_row)
            mismatches = sum(
                1 for r in range(header_row, last_row_before + 1) for c in range(1, last_col + 1)
                if src_before.cell(row=r, column=c).value != src_after.cell(row=r, column=c).value)
        if mismatches:
            return "fail", (f"{denom} を抽出しましたが、元シート『{src_name}』が {mismatches} セル"
                             "変更されています（読むだけのはず）")
        return "pass", f"{denom} → {len(expected_rows)}行を抽出（値・型とも保存・元シート無変更）"

    return "pass", f"{denom} → {len(expected_rows)}行を抽出（値・型とも保存。元シートとの突き合わせ無し）"

def check_report_per_row(path: Path, args: dict, header_row: int = 1,
                          source_book: Path | None = None) -> tuple:
    """REPORT_PER_ROW の事後条件。設計文書の4本柱を機械で確かめる:
       ①枚数の完全会計（データ行 N＝出力 N 枚。合計行は total_row の既存機構で独立に
        再判定する ── verify_dsl_args の勘定を信じ切らず、ここで recount する）
       ②値の3計数（compare_report_cells・印セルだけに絞る・欠落0/不一致0/余剰0）
       ③出所（検分シートの実在）
       ④雛形+データシートの不変（source_book が渡された時だけ突き合わせる）"""
    report_rows = args.get("_report_rows") or []
    placeholders = args.get("_placeholders") or []
    inspection_sheet = args.get("_inspection_sheet")
    template_sheet = args.get("template_sheet")
    src_sheet_name = args.get("_target_sheet")
    if not report_rows:
        return "fail", "出力する行が決まっていません（verify_dsl_args を経由していない可能性）"
    if not placeholders:
        return "fail", "印の一覧が決まっていません（verify_dsl_args を経由していない可能性）"
    if not inspection_sheet:
        return "fail", "検分シート名が決まっていません（verify_dsl_args を経由していない可能性）"

    with BookView(path) as bv:
        if src_sheet_name not in bv.sheetnames:
            return "fail", f"データシート『{src_sheet_name}』がありません"
        src = bv.sheet(src_sheet_name)
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        # ①枚数の完全会計: 合計行の除外を total_row.py で独立に recount する
        #   （verify_dsl_args が決めた _report_rows を鵜呑みにしない）。
        last_row = _scan_last_row(src, header_row=header_row)
        rows_in = []
        for r in range(header_row + 1, last_row + 1):
            label_val = src.cell(row=r, column=1).value
            vals = {c: src.cell(row=r, column=c).value for c in range(2, last_col + 1)}
            rows_in.append((r, label_val, vals))
        verdict = total_row.split_total_rows_multi(rows_in) if rows_in else \
            total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[])
        expected_rows = set(verdict.adopted_rows)
        got_rows = [rr["row"] for rr in report_rows]
        if len(got_rows) != len(set(got_rows)):
            return "fail", "報告シートに同じ元行が重複しています"
        got_row_set = set(got_rows)
        if expected_rows != got_row_set:
            missing = sorted(expected_rows - got_row_set)
            extra = sorted(got_row_set - expected_rows)
            return "fail", (f"データ{len(expected_rows)}行のうち出力は{len(got_row_set)}枚"
                             f"（欠落行 {missing}・余剰行 {extra}）")

        declared_sheet_names = {rr["sheet"] for rr in report_rows}
        missing_sheets = sorted(declared_sheet_names - set(bv.sheetnames))
        if missing_sheets:
            return "fail", f"宣言した報告シートがありません（欠落 {missing_sheets}）"
        # ★★ 2026-08-28（Namakoo が実測・別レイアウトの雛形を 2 枚目として置いた回）:
        #   「孤児シート」の分母を**出力側から**作っていた ── データ/雛形/検分**以外は
        #   全部この操作が作ったはず、という決めつけ。ブックに前から在るシート
        #   （2 つ目の雛形・メモ・参照表）が丸ごと『余剰(孤児シートの疑い)』になり、
        #   正しく 5 枚作れているのに × が出ていた。
        #   ★ 分母は入力側から取る: **前に無くて後に在る**シートだけが、この操作の産物。
        #   ★ 適用前が無い回は断定しない（この柱だけ落として、他の柱は今までどおり見る）。
        if source_book is not None and Path(source_book).exists():
            with BookView(source_book) as bv_b:
                before_sheets = set(bv_b.sheetnames)
            born = set(bv.sheetnames) - before_sheets
            orphans = sorted(born - declared_sheet_names - {inspection_sheet})
            if orphans:
                return "fail", ("この操作が作ったのに宣言していないシートがあります"
                                 f"（孤児シートの疑い {orphans}）")
        if inspection_sheet not in bv.sheetnames:
            return "fail", f"検分シート『{inspection_sheet}』がありません"

        # ②値の3計数: 印セルだけに絞って、各枚を元の行と突き合わせる（型込み等値）。
        for rr in report_rows:
            declared = {}
            for ph in placeholders:
                src_val = src.cell(row=rr["row"], column=ph["col_idx"]).value
                if ph["whole"]:
                    declared[ph["cell"]] = src_val
                else:
                    filler = "" if src_val is None else str(src_val)
                    declared[ph["cell"]] = ph["raw"].replace(
                        "{{" + ph["column_name"] + "}}", filler)
            result = compare_report_cells(path, rr["sheet"], declared)
            if not result.ok:
                return "fail", (
                    f"報告シート『{rr['sheet']}』（元{rr['row']}行目）の印が元の行と"
                    f"一致しません（欠落{len(result.missing)}・不一致{len(result.mismatched)}"
                    f"・余剰{len(result.surplus)}）"
                )

    denom = f"データ{len(report_rows)}行 → 出力{len(report_rows)}枚（印{len(placeholders)}箇所/枚）"

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            for sheet_name in (template_sheet, src_sheet_name):
                sb = bv_before.sheet(sheet_name)
                sa = bv_after.sheet(sheet_name)
                # ★ 2026-08-24: 走査でなく**使用範囲**で測る（A 列が余白の雛形で
                #   比較セル数が 0 になり「無変更」を分母ゼロで宣言していた）。
                br, bc = _used_extent(sb)
                ar, ac = _used_extent(sa)
                lr, lc = max(br, ar), max(bc, ac)
                mismatches = sum(
                    1 for r in range(1, lr + 1) for c in range(1, lc + 1)
                    if sb.cell(row=r, column=c).value != sa.cell(row=r, column=c).value)
                if mismatches:
                    return "fail", (f"{denom} でしたが、シート『{sheet_name}』が {mismatches}"
                                     " セル変更されています（雛形・データシートは読むだけのはず）")
        return "pass", f"{denom}（印の値・型とも保存・雛形/データシート無変更）"

    return "pass", f"{denom}（印の値・型とも保存。雛形/データシートとの突き合わせ無し）"

def check_report_per_group(path: Path, args: dict, header_row: int = 1,
                            source_book: Path | None = None) -> tuple:
    """まとめ版の事後条件。**REPORT_PER_ROW とは別の証明**（同じ関数に混ぜない ──
       混ぜた側の分岐が恒真になっているのが、この repo で一番こわい壊れ方）。

    証明するもの:
      ①グループの完全会計 ── データ行の集合が、グループの行の直和とちょうど一致
        （合計行の除外は total_row.py で**独立に**数え直す）
      ②明細 ── 件数ぶんの行が在り、k 件目が元の k 番目の行と型込みで一致
      ③合計 ── 画面の値が、**別実装（ここで足し直した値）**と一致
        （書いたのは Basic の UNO 側なので、ここは本当に別の実装）
      ④グループ値 ── そのグループ全行で同じ値であることを確かめたうえで一致
      ⑤**印でないセル**が、行を増やしたあとも雛形のまま（ずれで壊れていない）
      ⑥雛形・データシートが無変更（source_book が在る回だけ）
    ★ ⑤が芯: 行の挿入は「増やした所」より「押し下げた所」が静かに壊れる。
    """
    groups = args.get("_groups") or []
    placeholders = args.get("_placeholders") or []
    inspection_sheet = args.get("_inspection_sheet")
    template_sheet = args.get("template_sheet")
    src_sheet_name = args.get("_target_sheet")
    detail_row = args.get("_detail_row")
    if not groups:
        return "fail", "まとめる単位が決まっていません（verify_dsl_args を経由していない可能性）"
    if not inspection_sheet:
        return "fail", "検分シート名が決まっていません（verify_dsl_args を経由していない可能性）"

    with BookView(path) as bv:
        for need, what in ((src_sheet_name, "データシート"), (template_sheet, "雛形"),
                            (inspection_sheet, "検分シート")):
            if need not in bv.sheetnames:
                return "fail", f"{what}『{need}』がありません"
        src = bv.sheet(src_sheet_name)
        tpl = bv.sheet(template_sheet)
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        # ① グループの完全会計（合計行の除外を独立に数え直す）
        last_row = _scan_last_row(src, header_row=header_row)
        rows_in = [(r, src.cell(row=r, column=1).value,
                     {c: src.cell(row=r, column=c).value for c in range(2, last_col + 1)})
                    for r in range(header_row + 1, last_row + 1)]
        verdict = (total_row.split_total_rows_multi(rows_in) if rows_in
                    else total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[]))
        expected = set(verdict.adopted_rows)
        got, seen = set(), []
        for g in groups:
            for r in g["rows"]:
                if r in got:
                    return "fail", f"{r}行目が 2 つのグループに入っています"
                got.add(r)
            seen.append(g["sheet"])
        if expected != got:
            return "fail", (f"データ{len(expected)}行のうち、まとめたのは{len(got)}行"
                             f"（欠落 {sorted(expected - got)}・余剰 {sorted(got - expected)}）")
        if len(set(seen)) != len(seen):
            return "fail", "同じ名前のシートを 2 回宣言しています"

        tpl_last_row, tpl_last_col = _used_extent(tpl)
        mark_cells = {(ph["row"], ph["col"]) for ph in placeholders}
        for g in groups:
            if g["sheet"] not in bv.sheetnames:
                return "fail", f"シート『{g['sheet']}』がありません"
            out = bv.sheet(g["sheet"])
            n = len(g["rows"])
            # ②③④ 印
            for ph in placeholders:
                for out_row in report_group.output_rows_for(ph["row"], detail_row, n):
                    got_v = out.cell(row=out_row, column=ph["col"]).value
                    ref = f"{ph['cell']}→{out_row}行目"
                    if ph["kind"] == "total":
                        want = sum(float(src.cell(row=r, column=ph["col_idx"]).value or 0)
                                    for r in g["rows"])
                        if isinstance(got_v, bool) or not isinstance(got_v, (int, float)):
                            return "fail", (f"『{g['sheet']}』の合計 {ref} が数値ではありません"
                                             f"（{got_v!r}）")
                        if abs(float(got_v) - want) > 1e-9:
                            return "fail", (f"『{g['sheet']}』の合計 {ref} が {got_v!r} ですが、"
                                             f"{list(g['rows'])}行目を足すと {want!r} です")
                        continue
                    if ph["kind"] == "detail":
                        k = report_group.detail_index_for(out_row, detail_row, n)
                        src_row = g["rows"][k]
                    else:
                        src_row = g["rows"][0]
                    want_v = src.cell(row=src_row, column=ph["col_idx"]).value
                    if ph["kind"] == "value":
                        for r in g["rows"]:
                            if src.cell(row=r, column=ph["col_idx"]).value != want_v:
                                return "fail", (f"『{g['sheet']}』の『{ph['column_name']}』が"
                                                 f"元の行で食い違っています（{list(g['rows'])}行目）")
                    if ph["whole"]:
                        if got_v != want_v:
                            return "fail", (f"『{g['sheet']}』の {ref} が {got_v!r} ですが、"
                                             f"元{src_row}行目は {want_v!r} です")
                    else:
                        filler = "" if want_v is None else str(want_v)
                        want_s = ph["raw"].replace("{{" + ph["mark"] + "}}", filler)
                        if str(got_v if got_v is not None else "") != want_s:
                            return "fail", (f"『{g['sheet']}』の {ref} が {got_v!r} ですが、"
                                             f"{want_s!r} のはずです")
            # ⑤ 印でないセルが、押し下げられた先で雛形のまま
            for r in range(1, tpl_last_row + 1):
                for c in range(1, tpl_last_col + 1):
                    if (r, c) in mark_cells:
                        continue
                    want_v = tpl.cell(row=r, column=c).value
                    for out_row in report_group.output_rows_for(r, detail_row, n):
                        got_v = out.cell(row=out_row, column=c).value
                        if got_v != want_v:
                            return "fail", (
                                f"『{g['sheet']}』の {_cell_ref(out_row, c)} が {got_v!r} ですが、"
                                f"雛形の {_cell_ref(r, c)} は {want_v!r} です"
                                "（明細行を増やしたときに、印でないセルが壊れています）")

    denom = (f"取引先{len(groups)}件 → 請求書{len(groups)}枚"
              f"（明細 {sum(len(g['rows']) for g in groups)} 行ぶん・印{len(placeholders)}種）")

    # ⑥ 雛形とデータシートが無変更・この操作が作った宣言外のシートが無い
    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            for sheet_name in (template_sheet, src_sheet_name):
                sb = bv_before.sheet(sheet_name)
                sa = bv_after.sheet(sheet_name)
                br, bc = _used_extent(sb)
                ar, ac = _used_extent(sa)
                lr, lc = max(br, ar), max(bc, ac)
                bad = sum(1 for r in range(1, lr + 1) for c in range(1, lc + 1)
                           if sb.cell(row=r, column=c).value != sa.cell(row=r, column=c).value)
                if bad:
                    return "fail", (f"{denom} でしたが、シート『{sheet_name}』が {bad} セル"
                                     "変更されています（雛形・データシートは読むだけのはず）")
            born = set(bv_after.sheetnames) - set(bv_before.sheetnames)
            orphans = sorted(born - {g["sheet"] for g in groups} - {inspection_sheet})
            if orphans:
                return "fail", ("この操作が作ったのに宣言していないシートがあります"
                                 f"（孤児シートの疑い {orphans}）")
        return "pass", f"{denom}（明細・合計・固定文とも検算済み・雛形/データシート無変更）"

    return "pass", f"{denom}（明細・合計とも検算済み。雛形/データシートとの突き合わせ無し）"

def check_format_map(path: Path, args: dict, header_row: int = 1,
                      source_book: Path | None = None) -> tuple:
    """FORMAT_MAP の事後条件。REPORT_PER_ROW の兄弟（縦の展開）で同じ4本柱を機械で確かめる:
       ①行の完全会計（データ行 N＝出力 N 行。合計行は total_row で独立に recount する）
       ②値の3計数（見出し行＋各データ行の印セルだけ・compare_report_cells・型込み等値）
       ③出所（検分シートの実在）
       ④雛形+データシートの不変（source_book が渡された時だけ突き合わせる）"""
    data_rows = args.get("_data_rows") or []
    placeholders = args.get("_placeholders") or []
    header_texts = args.get("_header_texts") or []
    inspection_sheet = args.get("_inspection_sheet")
    template_sheet = args.get("template_sheet")
    src_sheet_name = args.get("_target_sheet")
    output_sheet = args.get("_output_sheet")
    if not data_rows:
        return "fail", "写す行が決まっていません（verify_dsl_args を経由していない可能性）"
    if not placeholders:
        return "fail", "印の一覧が決まっていません（verify_dsl_args を経由していない可能性）"
    if not inspection_sheet or not output_sheet:
        return "fail", "出力先/検分シート名が決まっていません（verify_dsl_args を経由していない可能性）"

    with BookView(path) as bv:
        if src_sheet_name not in bv.sheetnames:
            return "fail", f"データシート『{src_sheet_name}』がありません"
        if output_sheet not in bv.sheetnames:
            return "fail", f"出力シート『{output_sheet}』がありません"
        if inspection_sheet not in bv.sheetnames:
            return "fail", f"検分シート『{inspection_sheet}』がありません"
        src = bv.sheet(src_sheet_name)
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        # ①行の完全会計: 合計行の除外を total_row.py で独立に recount する
        #   （verify_dsl_args が決めた _data_rows を鵜呑みにしない）。
        last_row = _scan_last_row(src, header_row=header_row)
        rows_in = []
        for r in range(header_row + 1, last_row + 1):
            label_val = src.cell(row=r, column=1).value
            vals = {c: src.cell(row=r, column=c).value for c in range(2, last_col + 1)}
            rows_in.append((r, label_val, vals))
        verdict = total_row.split_total_rows_multi(rows_in) if rows_in else \
            total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[])
        expected_rows = verdict.adopted_rows
        if len(data_rows) != len(set(data_rows)):
            return "fail", "出力行に同じ元行が重複しています"
        if set(expected_rows) != set(data_rows):
            missing = sorted(set(expected_rows) - set(data_rows))
            extra = sorted(set(data_rows) - set(expected_rows))
            return "fail", (f"データ{len(expected_rows)}行のうち出力は{len(data_rows)}行"
                             f"（欠落行 {missing}・余剰行 {extra}）")

        out_ws = bv.sheet(output_sheet)
        out_last_row = _scan_last_row(out_ws, header_row=1)
        if out_last_row != 1 + len(data_rows):
            return "fail", (f"データ{len(data_rows)}行のはずが出力シートは"
                             f"{max(out_last_row - 1, 0)}行です（見出し行を除く）")

        # ②値の3計数: 見出し行＋各データ行（印セルだけに絞る・型込み等値）。
        declared = {}
        for ph, text in zip(placeholders, header_texts):
            declared[f"{get_column_letter(ph['out_col'])}1"] = text
        for out_row, src_row in enumerate(data_rows, start=2):
            for ph in placeholders:
                src_val = src.cell(row=src_row, column=ph["col_idx"]).value
                if ph["whole"]:
                    val = src_val
                else:
                    filler = "" if src_val is None else str(src_val)
                    val = ph["raw"].replace("{{" + ph["column_name"] + "}}", filler)
                declared[f"{get_column_letter(ph['out_col'])}{out_row}"] = val

    result = compare_report_cells(path, output_sheet, declared)
    if not result.ok:
        return "fail", (
            f"出力シート『{output_sheet}』の値が元データと一致しません"
            f"（欠落{len(result.missing)}・不一致{len(result.mismatched)}"
            f"・余剰{len(result.surplus)}）"
        )

    denom = f"データ{len(data_rows)}行 → 出力{len(data_rows)}行（印{len(placeholders)}箇所/行）"

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            for sheet_name in (template_sheet, src_sheet_name):
                sb = bv_before.sheet(sheet_name)
                sa = bv_after.sheet(sheet_name)
                # ★ 2026-08-24: 走査でなく**使用範囲**で測る（A 列が余白の雛形で
                #   比較セル数が 0 になり「無変更」を分母ゼロで宣言していた）。
                br, bc = _used_extent(sb)
                ar, ac = _used_extent(sa)
                lr, lc = max(br, ar), max(bc, ac)
                mismatches = sum(
                    1 for r in range(1, lr + 1) for c in range(1, lc + 1)
                    if sb.cell(row=r, column=c).value != sa.cell(row=r, column=c).value)
                if mismatches:
                    return "fail", (f"{denom} でしたが、シート『{sheet_name}』が {mismatches}"
                                     " セル変更されています（雛形・データシートは読むだけのはず）")
        return "pass", f"{denom}（印の値・型とも保存・雛形/データシート無変更）"

    return "pass", f"{denom}（印の値・型とも保存。雛形/データシートとの突き合わせ無し）"

def _check_report_router(path: Path, args: dict, header_row: int = 1,
                          source_book: Path | None = None) -> tuple:
    """帳票段の事後条件は**2 つある**（1 行 1 枚 / 取引先ごとに 1 枚）。
       どちらを使うかは op 名でなく、**解決済みの宣言**が持っている:
       `_groups` が在る回はまとめ版 ── まとめ版の証明は別関数（混ぜない）。
       ★ 混ぜると、片方の分岐が恒真でも全体は緑に見える（この repo で一番こわい形）。"""
    if args.get("_groups"):
        return check_report_per_group(path, args, header_row, source_book=source_book)
    return check_report_per_row(path, args, header_row, source_book=source_book)
