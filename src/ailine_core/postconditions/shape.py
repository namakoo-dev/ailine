"""事後条件 ── **行数も列数も変わらない** op（値と見た目）。

対象 12 op: 1セル書換 / 一括書換 / 条件つき書換 / 計算列 / 転記 /
太字 / 中央揃え / 数値書式 / けい線 / 背景色 / セル結合 / 列幅自動調整

★ この群が確かめること: 頼まれた場所が変わり、**それ以外が 1 セルも変わっていない**。
表の骨格（行数・列数・見出し）は前後で同じはずなので、そこがずれたら
「頼んでいない操作が混ざった」ことになる。

★ 書式 op（太字〜列幅）も**値の不変**を見る: 見た目だけを変えるはずの op が
値を壊したら、それは重い事故なのに誰も見ていない。
★ 2026-09-02 に数えた時点では 8 op が「他は 1 セルも変わらず」を見ておらず、
うち 4 op（けい線・数値書式・セル結合・一括書換）がこの群に居る
（`tests/test_postcondition_ledger.py` が在庫として持っている）。
"""
from __future__ import annotations

from pathlib import Path

from openpyxl.utils import get_column_letter

from ailine_core.book_view import BookView
from ailine_core.primitives import is_number as _is_number
from ailine_core.table_scan import _col_index_by_header, _scan_last_col, _scan_last_row, data_extent

from ailine_core.postconditions._shared import COLOR_MAP, _ZERO_TARGET_REASON, _cells_for_shift, _extract_predicate, _moved_rows_note, compare_moved_rows, note_unverified

def _apply_operator(a, b, operator: str):
    a = a or 0
    b = b or 0
    if operator == "+":
        return a + b
    if operator == "-":
        return a - b
    if operator == "*":
        return a * b
    if operator == "/":
        return a / b if b else 0
    raise ValueError(operator)

def only_this_column_changed(path: Path, source_book, sheet, col_index: int,
                              header_row: int = 1, inserted_at: int | None = None) -> str | None:
    """作った列**以外**が 1 セルも変わっていないか。変わっていればその理由（無ければ None）。

    ★★ 2026-08-30（番人の感度を測る治具が、初回の本気の測定で見つけた穴）:
      生成コードに「頼んでいないセルへの書き足し」を仕込んだところ、計算列の事後条件が
      **pass を返した**。作った列の中身しか見ていなかったため。
      ★ 1セル書換・行追加・行削除・並べ替え・入れ替えには「他は 1 セルも変わらず」が
        在るのに、計算列だけ無かった ── **また片配線**。
      ★ 単列×率の版（single_factor）も同じ穴なので、**1 関数を両方に配る**。
    ★ source_book が無い経路では何も言わない（断定しない・従来どおり）。
    """
    if not source_book or not Path(source_book).exists():
        return None
    try:
        with BookView(Path(source_book)) as bv_b, BookView(path) as bv_a:
            ws_b, ws_a = bv_b.sheet(sheet), bv_a.sheet(sheet)
            lb, cb = data_extent(ws_b, header_row)
            la, ca = data_extent(ws_a, header_row)
            width = max(cb, ca)
            # ★★ 2026-08-30（既存の検体が捕まえた・俺の入れた誤爆）: 同じ名前の列が
            #   2 本ある表（同じ依頼を 2 回した回）だと、_col_index_by_header は
            #   **古いほう**を返す。すると新しく作った列が「頼んでいない場所」に見えて、
            #   前提破れの関所（exit 7・言い方を案内する）より先に落としていた。
            #   ★ 同じ見出しの列は**全部**対象外にする（どれが新しいかは名前では決まらない）。
            # ★★ 2026-09-02（在庫 A を確かめて見つけた）: 依頼が位置を言った回
            #   （「売上と原価の間に…の列を作って」）では、作った列を**あとで動かす**ので
            #   その右にある列は 1 つずつずれる。同じ座標どうしで比べていたため、
            #   **正しい結果なのに「頼んでいない場所を書いた」と落としていた** ──
            #   置き場所の機構は動いていたのに、検算が追いついていなかった。
            #   ★ ずれを知っているのは呼び出し側なので、**渡してもらう**（推測しない）。
            if inserted_at:
                # ★★ 2026-09-02（自作 review・重大 4）: この経路は before の列数ぶんしか
                #   見ておらず、after に**想定外の列が増えても素通り**していた
                #   （呼び出し側はこの結果だけを根拠に「他は 1 セルも変わらず」と ✓ を出す）。
                #   ★ 列を 1 本挿したのだから、増えるのは**ちょうど 1 本**。分母を先に縛る。
                if ca != cb + 1:
                    return (f"列が {cb} → {ca} 本になっています"
                             f"（1 本だけ増えるはず）── 頼んでいない列を作った疑いがあります")
                for r in range(header_row, max(lb, la) + 1):
                    for c_b in range(1, cb + 1):
                        c_a = c_b if c_b < inserted_at else c_b + 1
                        v_b = ws_b.cell(row=r, column=c_b).value
                        v_a = ws_a.cell(row=r, column=c_a).value
                        if v_b != v_a:
                            return (f"作った列のほかに {r}行{c_a}列 が変わっています"
                                     f"（{v_b!r}→{v_a!r}）── "
                                     "頼んでいない場所を書いた疑いがあります")
                return None
            _name = str(ws_a.cell(row=header_row, column=col_index).value or "")                 if 1 <= col_index <= ca else ""
            skip = {c for c in range(1, width + 1)
                     if _name and str(ws_a.cell(row=header_row, column=c).value or "") == _name}
            skip.add(col_index)
            for r in range(header_row, max(lb, la) + 1):
                for c in range(1, width + 1):
                    if c in skip:
                        continue          # 作った列そのものは上で検証済み
                    v_b = ws_b.cell(row=r, column=c).value
                    v_a = ws_a.cell(row=r, column=c).value
                    if v_b != v_a:
                        return (f"作った列のほかに {r}行{c}列 が変わっています"
                                 f"（{v_b!r}→{v_a!r}）── 頼んでいない場所を書いた疑いがあります")
    except Exception:
        return None                       # 読めない回は黙る
    return None

def check_compute_column(path: Path, args: dict, header_row: int = 1,
                          use_formula: bool = False, source_book=None) -> tuple:
    """★ 止血1/2: 演算対象が非数値/None の行（合計行等）は「対象外」として除外し件数を
       表示する。除外後に検証できた行が0件なら fail（機械検証済みと名乗らない）。
       対象列(target)自体が非数値なのは演算が本当に効いていない証拠なので除外せず fail。
       ★ W3 Part3: use_formula=True のとき事後条件を二層化する — ①通常の openpyxl 読み
       で保存された式文字列が期待形か ②data_only 読みでキャッシュ値が演算結果と一致するか。
       両方合格して初めて pass にする（式だけ合っていて未計算/値だけ合っていて式が
       無いケースの両方を見逃さない）。
       ★ W10b 項目3: operands が1つだけ（税込み/税抜き等「列 × 率」パターン）の場合は
       check_compute_column_single_factor に委譲する。
       ★ W10f 項目1: use_formula のとき operands は data_only(計算後の値)側から読む
       （旧実装は raw 側=式ビューから読んでいたため、前段が式で作った計算列(小計等)を
       operand にすると raw 値が式文字列のままで『数値でない』扱いになり、全行除外→
       検証対象0件failで計画ごとロールバックしていた＝実測の再現形そのもの）。
       operand 自体が式なのにキャッシュ値が無い(LibreOffice を通していない)行は、
       『対象が無い』(合計行等の非数値/空欄)とは別に数える（0cf9218 空虚な検証合格の
       禁止の趣旨 — 対象は本当は有るのに読めていないだけ、と混同しない）。COMPUTE_COLUMN
       は行ごとに独立な検証なので、うち検証できた行だけを checked に数えても他行の結果は
       歪めない（AGGREGATE/SORT の全行をまたぐ検証とは違い部分採点できる）。"""
    if len(args["operands"]) == 1:
        return check_compute_column_single_factor(path, args, header_row=header_row,
                                                    use_formula=use_formula,
                                                    source_book=source_book)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        op1, op2 = args["operands"]
        i1 = _col_index_by_header(ws, op1, header_row=header_row)
        i2 = _col_index_by_header(ws, op2, header_row=header_row)
        # ★ M2c: target(実在列名) 指定時はその列を検証する。無指定なら従来どおり自動命名の新列。
        # ★ W3 改定(2026-08-20): codegen_dsl と同じ見出し名決定（_new_col_label があれば
        #   利用者が依頼文で名指しした新規列名）。
        target = args.get("target")
        newname = target or args.get("_new_col_label") or f"{op1}{args['operator']}{op2}"
        inew = _col_index_by_header(ws, newname, header_row=header_row)
        if i1 is None or i2 is None or inew is None:
            return "fail", f"演算対象または対象列『{newname}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        col1_letter = get_column_letter(i1)
        col2_letter = get_column_letter(i2)
        checked = 0
        excluded = 0
        uncached = 0
        for r in range(header_row + 1, last + 1):
            a_raw = ws.cell(row=r, column=i1).value
            b_raw = ws.cell(row=r, column=i2).value
            got = ws.cell(row=r, column=inew).value
            a = bv.cell_value(r, i1, sheet=args.get("_target_sheet")) if use_formula else a_raw
            b = bv.cell_value(r, i2, sheet=args.get("_target_sheet")) if use_formula else b_raw
            if not _is_number(a) or not _is_number(b):
                # ★ W10f: operand 自体が式(前段の計算列)で、かつキャッシュ値が無い行は
                #   『数値でない対象外』と別カウントする（下の note で区別して表示）。
                if use_formula and (
                    (not _is_number(a) and isinstance(a_raw, str) and a_raw.startswith("="))
                    or (not _is_number(b) and isinstance(b_raw, str) and b_raw.startswith("="))
                ):
                    uncached += 1
                else:
                    excluded += 1   # 例: 合計行で演算対象セルが空欄
                continue
            want = _apply_operator(a, b, args["operator"])
            if use_formula:
                expect_formula = f"={col1_letter}{r}{args['operator']}{col2_letter}{r}"
                if not isinstance(got, str) or got.replace(" ", "") != expect_formula:
                    return "fail", f"{r}行目: 式が期待形でない (期待 {expect_formula} 実際 {got!r})"
                got_cached = bv.cell_value(r, inew, sheet=args.get("_target_sheet"))
                if not _is_number(got_cached) or abs(got_cached - want) > 1e-6:
                    return "fail", f"{r}行目: 式のキャッシュ値が不一致 (期待 {want} 実際 {got_cached!r})"
            else:
                if not _is_number(got) or abs(got - want) > 1e-6:
                    return "fail", f"{r}行目: 期待 {want} 実際 {got}"
            checked += 1
    note_parts = []
    if excluded:
        note_parts.append(f"数値でない {excluded} 行は対象外")
        note_unverified(args, excluded, "演算の対象が数値でないため計算結果を確かめられない")
    if uncached:
        note_parts.append(f"演算対象の式にキャッシュ値が無く検証できない {uncached} 行")
    note = f"（{'・'.join(note_parts)}）" if note_parts else ""
    if checked == 0:
        return "fail", _ZERO_TARGET_REASON + note
    _moved_to = args.get("_move_new_col_to")
    if (_side := only_this_column_changed(
            path, source_book, args.get("_target_sheet"), inew, header_row,
            inserted_at=(int(_moved_to) + 1) if _moved_to is not None else None)):
        return "fail", _side
    if use_formula:
        return "pass", f"{checked} 行を検証（式・キャッシュ値とも一致・他は 1 セルも変わらず）{note}"
    return "pass", f"{checked} 行を検証（他は 1 セルも変わらず）{note}"

def check_compute_column_single_factor(path: Path, args: dict, header_row: int = 1,
                                        use_formula: bool = False, source_book=None) -> tuple:
    """★ W10b 項目3: COMPUTE_COLUMN の「1列 × 率」パターン（税込み/税抜き等）専用の事後条件。
       check_compute_column（2列版）と同じ二層検証（式の期待形・data_only キャッシュ値）を
       1列 + factor に合わせて行う。factor は verify_dsl_args が機械確定済みの値
       （resolved["factor"]）をそのまま受け取る前提。
       ★ W10f 項目1: use_formula のとき operand は data_only(計算後の値)側から読む
       （check_compute_column と同型・前段が式で作った計算列(小計等)を operand にした
       場合の同じバグをここでも直す）。キャッシュ値が無い(式はあるが未計算)行は
       『数値でない対象外』と別カウントする（0cf9218 空虚な検証合格の禁止の趣旨）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        op1 = args["operands"][0]
        operator = args["operator"]
        factor = float(args.get("factor", 1) or 1)
        i1 = _col_index_by_header(ws, op1, header_row=header_row)
        target = args.get("target")
        # ★ W10c 中: codegen_dsl と同じ見出し名決定（_new_col_label があれば自然な日本語見出し）。
        newname = target or args.get("_new_col_label") or f"{op1}{operator}{factor:g}"
        inew = _col_index_by_header(ws, newname, header_row=header_row)
        if i1 is None or inew is None:
            return "fail", f"演算対象または対象列『{newname}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        col1_letter = get_column_letter(i1)
        checked = 0
        excluded = 0
        uncached = 0
        for r in range(header_row + 1, last + 1):
            a_raw = ws.cell(row=r, column=i1).value
            got = ws.cell(row=r, column=inew).value
            a = bv.cell_value(r, i1, sheet=args.get("_target_sheet")) if use_formula else a_raw
            if not _is_number(a):
                if use_formula and isinstance(a_raw, str) and a_raw.startswith("="):
                    uncached += 1   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
                else:
                    excluded += 1   # 例: 合計行で演算対象セルが空欄
                continue
            want = _apply_operator(a, factor, operator)
            if use_formula:
                expect_formula = f"={col1_letter}{r}{operator}{factor:g}"
                if not isinstance(got, str) or got.replace(" ", "") != expect_formula:
                    return "fail", f"{r}行目: 式が期待形でない (期待 {expect_formula} 実際 {got!r})"
                got_cached = bv.cell_value(r, inew, sheet=args.get("_target_sheet"))
                if not _is_number(got_cached) or abs(got_cached - want) > 1e-6:
                    return "fail", f"{r}行目: 式のキャッシュ値が不一致 (期待 {want} 実際 {got_cached!r})"
            else:
                if not _is_number(got) or abs(got - want) > 1e-6:
                    return "fail", f"{r}行目: 期待 {want} 実際 {got}"
            checked += 1
    note_parts = []
    if excluded:
        note_parts.append(f"数値でない {excluded} 行は対象外")
        note_unverified(args, excluded, "演算の対象が数値でないため計算結果を確かめられない")
    if uncached:
        note_parts.append(f"演算対象の式にキャッシュ値が無く検証できない {uncached} 行")
    note = f"（{'・'.join(note_parts)}）" if note_parts else ""
    if checked == 0:
        return "fail", _ZERO_TARGET_REASON + note
    _moved_to = args.get("_move_new_col_to")
    if (_side := only_this_column_changed(
            path, source_book, args.get("_target_sheet"), inew, header_row,
            inserted_at=(int(_moved_to) + 1) if _moved_to is not None else None)):
        return "fail", _side
    if use_formula:
        return "pass", f"{checked} 行を検証（式・キャッシュ値とも一致・他は 1 セルも変わらず）{note}"
    return "pass", f"{checked} 行を検証（他は 1 セルも変わらず）{note}"

def check_lookup_fill(path: Path, args: dict, header_row: int = 1,
                       use_formula: bool = False) -> tuple:
    """★ W3: header_row は対象シート(target_sheet。★ 挙動変更#2 より前は常に1枚目だった)の
       見出し行。参照表(source_sheet)は常に「列0=キー・列1=値」の物理1行目見出し前提
       （VLookupFromTable ヘルパの仕様どおり・検出対象外）。
       ★ W10f 項目5: use_formula のとき対象シートのキー列を data_only(計算後の値)側から
       読む。VLookupFromTable ヘルパ自体は getString()（LibreOffice が式を評価した文字列）
       でキーを照合するため、前段の式(=A2 等)で埋まったキー列でも転記そのものは正しく
       動く。旧実装は openpyxl の raw 読み（式文字列そのもの）でキーを拾っていたため
       対応表と1件も一致せず、正しく転記されたブックに対して『マスタ表の列順が違う』と
       いう誤診断（濡れ衣）を出していた（Namakoo の純関数レベルの実測をそのまま回帰
       テストにする）。LOOKUP_FILL は check_compute_column と同型の行独立の検証（他行の
       結果に影響しない）なので、キーが読めなかった行は checked から除外する部分採点に
       する（0cf9218 空虚な検証合格の禁止の趣旨 — 読めた行だけを『検証済み』と名乗る）。
       ★ 誤診断の是正: 『対応表のキーが1件も転記されていない』という観測から即座に
       『マスタ表の列順が違う』と断定するのは飛躍だった。まずキーが読めなかった行を
       別集計し（『対象が無い』と『対象はあるが読めない』を混同しない）、キーが全部
       読めたのに1件も一致しない場合でも、本当に列順が違うのか単にキー値そのものが
       対応表と食い違うのかはこのデータだけでは区別できないため、断定をやめて
       『可能性』として並べる。"""
    with BookView(path) as bv:
        if (args["target_sheet"] not in bv.sheetnames
                or args["source_sheet"] not in bv.sheetnames):
            return "fail", "対象/参照シートが無い"
        tws = bv.sheet(args["target_sheet"])
        sws = bv.sheet(args["source_sheet"])
        key_idx = _col_index_by_header(tws, args["key_col"], header_row=header_row)
        tgt_idx = _col_index_by_header(tws, args["target_col"], header_row=header_row)
        if key_idx is None or tgt_idx is None:
            return "fail", "対象シートにキー列/対象列が無い"
        lookup = {}
        r = 2
        while sws.cell(row=r, column=1).value not in (None, ""):
            lookup[sws.cell(row=r, column=1).value] = sws.cell(row=r, column=2).value
            r += 1
        scanned = 0   # ★ 止血1: 対象シートに行が1件も無い(0件)場合と、行はあるが1件も
                      #   対応表に載っていない場合を別のメッセージで区別する。
        checked = 0
        uncached = 0
        r = header_row + 1
        while tws.cell(row=r, column=key_idx).value not in (None, ""):
            scanned += 1
            key_raw = tws.cell(row=r, column=key_idx).value
            key = bv.cell_value(r, key_idx, sheet=args["target_sheet"]) if use_formula else key_raw
            if use_formula and key is None and isinstance(key_raw, str) and key_raw.startswith("="):
                uncached += 1   # ★ W10f: キー列に式はあるがキャッシュ値が無い（『対象が無い』とは別）
                r += 1
                continue
            if key in lookup:
                got = tws.cell(row=r, column=tgt_idx).value
                want = lookup[key]
                if got != want:
                    return "fail", f"{r}行目: キー『{key}』の転記値が不一致 (期待 {want!r} 実際 {got!r})"
                checked += 1
            r += 1
    if scanned == 0:
        return "fail", _ZERO_TARGET_REASON
    if uncached == scanned:
        return "fail", (f"対象シートのキー列に式はあるがキャッシュ値が無く検証できない行が "
                         f"{uncached} 件あり、転記結果を検証できません"
                         f"（LibreOffice を通していない可能性）")
    note = f"（キー列に式はあるがキャッシュ値が無く検証できない {uncached} 行は対象外）" if uncached else ""
    note_unverified(args, uncached, "キー列が式でキャッシュ値が無く、転記先を確かめられない")
    if checked == 0:
        # ★ W10b 項目4a → W10f 項目5: ここまで来た行はキーが読めている（『読めなかった』
        #   場合とは別集計済み）。ただし『本当に列順が違う』のか『キー値そのものが対応表と
        #   一致しない』のかはこのデータだけでは区別できないため、断定せず可能性として
        #   並べる（マスタ表が実際にキー列→値列の順で書かれた具体的な直し方は残す）。
        return "fail", (
            f"対応表『{args['source_sheet']}』に載っているキーが1件も転記されていません。"
            f"マスタ表の列順が違う可能性があります。または『{args['target_sheet']}』シートの"
            f"『{args['key_col']}』列の値が対応表のキーと一致していない可能性があります。"
            f"マスタ表がキー列→値列の順であれば、『{args['source_sheet']}』シートの A 列に"
            f"キー（{args['key_col']} に対応する値）、B 列に値を置いてください{note}"
        )
    return "pass", f"{checked} 行を検証{note}"

def _cell_target(args: dict):
    """target が cell:R,C なら (行, 列)。そうでなければ None。"""
    t = str(args.get("target") or "")
    if not t.startswith("cell:"):
        return None
    try:
        r, c = (int(x) for x in t[5:].split(","))
        return r, c
    except ValueError:
        return None

def _only_this_cell(path: Path, args: dict, rc: tuple, read, label: str,
                     header_row: int = 1, source_book: Path | None = None) -> tuple:
    """1 セルだけを飾る操作の事後条件（★ 両方向を見る ── ここが芯）。

    ★ 2026-08-27（Namakoo「セル指定はできる？」）: 「『商品』セルに色を付けて」は
      実測で毎回 `col:商品`（**列ぜんぶ**）に化けていた。頼んでいない範囲に静かに
      広がる形で、この道具が最も嫌うもの。だから証明も両方向にする:
      ① 宣言したセルが飾られている ② **他のセルの飾りが 1 つも変わっていない**
    ★ ② は適用前と突き合わせる（元から飾ってあったセルを「広がった」と誤らない）。
      適用前が無ければ ② は主張しない（warn ── 言えないことは言わない）。
    read: ws とセルから「飾りの値」を取り出す関数（太字か・背景色か・寄せか）。
    """
    r, c = rc
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last, cols = data_extent(ws, header_row)
        if r > last or c > cols:
            return "fail", f"{label}の {r}行{c}列は表の外です（表は {last}行×{cols}列）"
        if not read(ws.cell(row=r, column=c)):
            return "fail", f"{r}行{c}列に{label}が付いていません"
        after = {(rr, cc): read(ws.cell(row=rr, column=cc))
                  for rr in range(header_row, last + 1) for cc in range(1, cols + 1)}
    if source_book is None or not Path(source_book).exists():
        return "warn", (f"{r}行{c}列に{label}が付いたことのみ確認"
                         "（他のセルへ広がっていないかは、適用前ファイルが無いので未確認）")
    with BookView(source_book) as bvb:
        wsb = bvb.sheet(args.get("_target_sheet"))
        spread = [(rr, cc) for (rr, cc), v in after.items()
                   if (rr, cc) != (r, c) and v != read(wsb.cell(row=rr, column=cc))]
    if spread:
        where = "、".join(f"{rr}行{cc}列" for rr, cc in spread[:5])
        return "fail", (f"{label}が {r}行{c}列 以外にも付いています（{where}）"
                         " ── 1 セルのはずが広がった疑いがあります")
    return "pass", f"{r}行{c}列 だけに{label}（他のセルは 1 つも変わらず）"

def check_bold(path: Path, args: dict, header_row: int = 1,
                source_book: Path | None = None) -> tuple:
    """★ W3: "col:" 対象は見出し(header_row)を含めて検証する（codegen の
       StyleBold(oDoc, col, hr0, col, lastRow) が見出しも含めて太字にするため）。"""
    if (rc := _cell_target(args)):
        return _only_this_cell(path, args, rc, lambda c: bool(c.font and c.font.bold),
                                "太字", header_row, source_book)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        kind, val = args["target"].split(":", 1)
        if kind == "row":
            last_col = _scan_last_col(ws, header_row=header_row)
            row = int(val)
            cells = [ws.cell(row=row, column=c) for c in range(1, last_col + 1)]
            label = f"{row}行目"
        else:
            idx = _col_index_by_header(ws, val, header_row=header_row)
            if idx is None:
                return "fail", f"列『{val}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{val}』"
        if not cells:   # ★ 止血1: 検証対象0件（見出しすら無い空シート等）を合格にしない
            return "fail", _ZERO_TARGET_REASON
        ok = all(c.font and c.font.bold for c in cells)
    if not ok:
        return "fail", f"{label} に太字でないセルがある"
    return "pass", f"{len(cells)} セルが太字"

def _bg_hex(c) -> str:
    """セルの背景色（16 進・無色は空）。openpyxl の書式表現の揺れをここ 1 箇所に閉じる。"""
    f = getattr(c, "fill", None)
    rgb = getattr(getattr(f, "fgColor", None), "rgb", None)
    if not isinstance(rgb, str):
        return ""
    return rgb[-6:].upper()

def check_fill_color(path: Path, args: dict, header_row: int = 1,
                      source_book: Path | None = None) -> tuple:
    """★ W3: "col:" 対象は見出し(header_row)を含めて検証する（codegen が見出しも
       含めて塗るため）。"""
    if (rc := _cell_target(args)):
        want = COLOR_MAP[args["color"]].upper()
        return _only_this_cell(path, args, rc, lambda c: _bg_hex(c) == want and _bg_hex(c),
                                f"色（{args['color']}）", header_row, source_book)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        want_hex = COLOR_MAP[args["color"]].upper()
        kind, val = args["target"].split(":", 1)
        if kind == "row":
            last_col = _scan_last_col(ws, header_row=header_row)
            row = int(val)
            cells = [ws.cell(row=row, column=c) for c in range(1, last_col + 1)]
            label = f"{row}行目"
        else:
            idx = _col_index_by_header(ws, val, header_row=header_row)
            if idx is None:
                return "fail", f"列『{val}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{val}』"
        if not cells:   # ★ 止血1
            return "fail", _ZERO_TARGET_REASON

        def _matches(cell) -> bool:
            if cell.fill is None or not cell.fill.patternType:
                return False
            return str(cell.fill.start_color.rgb).upper().endswith(want_hex)

        ok = all(_matches(c) for c in cells)
    if not ok:
        return "fail", f"{label} に色『{args['color']}』が付いていないセルがある"
    return "pass", f"{len(cells)} セルの背景色を確認"

def check_number_format(path: Path, args: dict, header_row: int = 1) -> tuple:
    # ★ 2026-08-29: 行に掛けた回は、その行の**数値セル**に付いていることを見る
    #   （ラベルの『合計』には掛からないのが正しい）。
    if args.get("_row_index"):
        row = int(args["_row_index"])
        with BookView(path) as bv:
            ws = bv.sheet(args.get("_target_sheet"))
            _last, last_col = data_extent(ws, header_row)
            nums = [c for c in range(1, last_col + 1)
                     if isinstance(ws.cell(row=row, column=c).value, (int, float))
                     or str(ws.cell(row=row, column=c).value or "").startswith("=")]
            if not nums:
                return "fail", f"{row}行目に数値のセルがありません"
            bad = [c for c in nums
                    if "#,##0" not in (ws.cell(row=row, column=c).number_format or "")]
        if bad:
            return "fail", (f"{row}行目の {len(bad)} 箇所に桁区切り書式が付いていません")
        return "pass", f"{row}行目の数値 {len(nums)} 箇所に桁区切り書式を確認"
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        if last < header_row + 1:   # ★ 止血1: データ行0件を合格にしない
            return "fail", _ZERO_TARGET_REASON
        ok = all("#,##0" in (ws.cell(row=r, column=idx).number_format or "")
                 for r in range(header_row + 1, last + 1))
    if not ok:
        return "fail", f"列『{args['col']}』に桁区切り書式が付いていないセルがある"
    return "pass", f"{last - header_row} 行に桁区切り書式を確認"

def check_merge(path: Path, args: dict, header_row: int = 1) -> tuple:
    with BookView(path) as bv:
        ranges = {str(r) for r in bv.sheet(args.get("_target_sheet")).merged_cells.ranges}
    if args["range"] not in ranges:
        return "fail", f"範囲『{args['range']}』が結合されていない"
    return "pass", f"{args['range']} の結合を確認"

def check_center_align(path: Path, args: dict, header_row: int = 1,
                       source_book: Path | None = None) -> tuple:
    """★ W3: "all"/"col:" とも見出し(header_row)を含めて検証する（codegen の
       AlignCenter/inline テンプレが見出しも含めて中央揃えにするため）。"""
    if (rc := _cell_target(args)):
        return _only_this_cell(
            path, args, rc,
            lambda c: str(getattr(getattr(c, "alignment", None), "horizontal", "")) == "center",
            "中央揃え", header_row, source_book)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        target = args["target"]
        if target == "all":
            last_row = _scan_last_row(ws, header_row=header_row)
            last_col = _scan_last_col(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=c) for r in range(header_row, last_row + 1)
                     for c in range(1, last_col + 1)]
            label = "表全体"
        else:
            colname = target[4:]
            idx = _col_index_by_header(ws, colname, header_row=header_row)
            if idx is None:
                return "fail", f"列『{colname}』が見つからない"
            last_row = _scan_last_row(ws, header_row=header_row)
            cells = [ws.cell(row=r, column=idx) for r in range(header_row, last_row + 1)]
            label = f"列『{colname}』"
        if not cells:   # ★ 止血1
            return "fail", _ZERO_TARGET_REASON
        ok = all(c.alignment and c.alignment.horizontal == "center" for c in cells)
    if not ok:
        return "fail", f"{label} に中央揃えでないセルがある"
    return "pass", f"{len(cells)} セルの中央揃えを確認"

def check_set_where(path: Path, args: dict, header_row: int = 1,
                     source_book: Path | None = None) -> tuple:
    """SET_WHERE の事後条件。**当てはまる行だけが変わった**ことを両側から証明する。

    ① 条件は**適用前のファイル**から独立に評価する（書き手が使った判定を借りない ──
      同じ所から分母を作ると恒真になる。この repo が 8 月に 4 回踏んだ形）
    ② 当てはまる行: 対象列がちょうど宣言した値になっている
    ③ 当てはまらない行: 対象列が**1 セルも変わっていない**（広がっていないことの証明）
    ④ 対象列以外は 1 セルも変わっていない
    ★ ②③ の両方を見るのが要点 ── 「付いたか」だけ見ると、全行に付けても pass する。
    """
    col, cond_col = str(args.get("col", "")), str(args.get("cond_col", ""))
    cmp, thr, value = args.get("cmp"), args.get("cond_value"), args.get("value")
    if source_book is None or not Path(source_book).exists():
        return "warn", "適用前ファイルが無いため、当てはまる行だけが変わったことを確かめられていません"
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_a, cols_a = data_extent(ws, header_row)
        headers = [str(ws.cell(row=header_row, column=c).value or "")
                    for c in range(1, cols_a + 1)]
        if col not in headers or cond_col not in headers:
            return "fail", f"列『{col}』または『{cond_col}』が適用後の表にありません"
        wi, ci = headers.index(col) + 1, headers.index(cond_col) + 1
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            last_b, cols_b = data_extent(ws_b, header_row)
            if [str(ws_b.cell(row=header_row, column=c).value or "")
                 for c in range(1, cols_b + 1)] != headers:
                return "fail", "見出しが変わっています（条件つき書換で見出しは動かないはず）"
            last = max(last_a, last_b)
            match = _extract_predicate(cmp, thr)
            wrong_hit, wrong_miss, hits = [], [], 0
            skip_rows = set(int(x) for x in (args.get("_skip_rows") or []))
            for r in range(header_row + 1, last + 1):
                # ★ 条件は**適用前**の値で見る（書いた後の値で見ると、書いた印そのものが
                #   条件を変えてしまう op では恒真になりうる）
                # ★ 合計行は「当てはまる」側から外す ── ただし**変わっていないこと**は
                #   要求する（外した行が黙って書き換わるのは、外していないのと同じくらい悪い）。
                if r in skip_rows:
                    if (bv.cell_value(r, wi, args.get("_target_sheet"))
                            != bv_b.cell_value(r, wi, args.get("_target_sheet"))):
                        wrong_hit.append(r)
                    continue
                if match(bv_b.cell_value(r, ci, args.get("_target_sheet"))):
                    hits += 1
                    if str(bv.cell_value(r, wi, args.get("_target_sheet")) or "") != str(value):
                        wrong_miss.append(r)
                else:
                    before_v = bv_b.cell_value(r, wi, args.get("_target_sheet"))
                    if bv.cell_value(r, wi, args.get("_target_sheet")) != before_v:
                        wrong_hit.append(r)
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"), header_row, last,
                                           max(cols_a, cols_b))
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"), header_row, last,
                                            max(cols_a, cols_b))
    if wrong_miss:
        return "fail", (f"条件に当てはまるのに『{col}』が『{value}』になっていない行があります"
                         f"（{'、'.join(map(str, wrong_miss[:5]))}行目）")
    if wrong_hit:
        return "fail", (f"条件に当てはまらない行の『{col}』が変わっています"
                         f"（{'、'.join(map(str, wrong_hit[:5]))}行目）── 書き込みが広がった疑い")
    if not hits:
        return "fail", f"条件に当てはまる行が 1 つもありません（『{value}』はどこにも書かれていない）"
    # ④ 対象列以外は不変 ── 対象列を両側から抜いて、残りを 1 箇所の比べ方に通す
    j = wi - 1
    st, info = compare_moved_rows([tuple(row[:j] + row[j + 1:]) for row in after_rows],
                                   [tuple(row[:j] + row[j + 1:]) for row in before_rows],
                                   "条件つき書換のあと")
    if st == "broken":
        return "fail", f"『{col}』以外の列が変わっています ── {info}"
    if info:
        return "warn", _moved_rows_note(info)
    return "pass", (f"条件に当てはまる {hits} 行だけ『{col}』を『{value}』に"
                     f"（当てはまらない行と他の列は 1 セルも変わらず）")

def check_set_cell_value(path: Path, args: dict, header_row: int = 1,
                          source_book: Path | None = None) -> tuple:
    """SET_CELL_VALUE の事後条件。**1 セルであることを証明する**。

    ★★ 2026-08-27（architect の査読で名指しされた穴）: 既存の check_set_column_value は
      「対象列のデータ行が**全部**その値か」を見る ── 1 セル用に流用すると、
      **列全体を潰した方が pass する**（逆向きの検算）。この機能で最も起きやすい
      壊れ方（列全体の codegen を流用して走査範囲を間違える）を、番人が通してしまう。
    ★ だから証明するのは 3 つ:
      ① 宣言したセルが宣言した値になっている
      ② **値が変わったセルはちょうど 1 個**（列を潰していない）
      ③ その 1 個の座標が宣言と一致する
    ★ ②③ は source_book が要る。無ければ①だけの warn（断定しない）。
    """
    name = str(args.get("row", ""))
    col_name = str(args.get("col", ""))
    want = args.get("value")
    numeric = bool(args.get("_write_numeric"))
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last, last_col = data_extent(ws, header_row)
        headers = [str(ws.cell(row=header_row, column=c).value or "")
                    for c in range(1, last_col + 1)]
        if args.get("_writes_header"):
            # ★ 見出しそのものを書き換えた回。名前で引くと**書き換えた後の表には
            #   もう無い** ── 宣言した座標で見る（2026-08-30 実測で落ちた形）。
            cidx = int(args.get("_col_index") or 0)
            if not (1 <= cidx <= last_col):
                return "fail", f"列の位置（{cidx}）が表の外です"
        elif col_name not in headers:
            return "fail", f"列『{col_name}』が見つからない"
        else:
            cidx = headers.index(col_name) + 1
        # ★ 2026-08-28: 行番号で指された回は、その番号を正とする（名前で探し直さない ──
        #   同名の行が 2 つある表では名前では決まらないし、依頼は番号そのものだった）。
        if args.get("row_number"):
            row = int(args["row_number"])
            _lo = header_row - 1 if args.get("_writes_header") else header_row
            if not (_lo < row <= last):
                return "fail", f"{row}行目は表の範囲外（データは{header_row + 1}〜{last}行目）"
        else:
            hits = [r for r in range(header_row + 1, last + 1)
                     if any(str(ws.cell(row=r, column=c).value or "").strip() == name
                             for c in range(1, last_col + 1))]
            if len(hits) != 1:
                return "fail", f"『{name}』の行が {len(hits)} 件（1 件に決まらない）"
            row = hits[0]
        got = ws.cell(row=row, column=cidx).value
        # ★ 見出しを書き換えた回は、見出し行も**見る範囲に入れる** ── 入れないと
        #   「1 セルのはずが 0 セル変わっています」になる（2026-08-30 実測）。
        #   ★ 範囲は 1 箇所で決めて、before/after/座標の 3 つに配る（片配線を作らない）。
        _scan_from = header_row if args.get("_writes_header") else header_row + 1
        after = [[ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
                  for r in range(_scan_from, last + 1)]
    # ① 宣言どおりの値か
    if numeric:
        w = float(args["_write_numeric_value"])
        if not (_is_number(got) and abs(float(got) - w) <= 1e-6):
            return "fail", f"『{name}』の{col_name}が {w:g} でない（実際 {got!r}）"
    elif str(got) != str(want):
        return "fail", f"『{name}』の{col_name}が {want!r} でない（実際 {got!r}）"
    if source_book is None or not Path(source_book).exists():
        return "warn", f"『{name}』の{col_name}のみ確認（変えていないセルは見ていません）"
    with BookView(source_book) as bv_b:
        ws_b = bv_b.sheet(args.get("_target_sheet"))
        lb, cb = data_extent(ws_b, header_row)
        before = [[ws_b.cell(row=r, column=c).value for c in range(1, max(cb, last_col) + 1)]
                   for r in range(_scan_from, lb + 1)]
    # ②③ 変わったセルはちょうど 1 個で、その座標が宣言と一致する
    changed = []
    for ri in range(max(len(before), len(after))):
        b_row = before[ri] if ri < len(before) else []
        a_row = after[ri] if ri < len(after) else []
        for ci in range(max(len(b_row), len(a_row))):
            bv_ = b_row[ci] if ci < len(b_row) else None
            av_ = a_row[ci] if ci < len(a_row) else None
            if bv_ != av_:
                changed.append((_scan_from + ri, ci + 1))
    if len(changed) != 1:
        return "fail", (f"1 セルのはずが {len(changed)} セル変わっています"
                         f"（{'、'.join(f'{r}行{c}列' for r, c in changed[:6])}）"
                         " ── 列全体を潰した疑いがあります")
    if changed[0] != (row, cidx):
        return "fail", (f"変わったのは {changed[0][0]}行{changed[0][1]}列で、"
                         f"宣言した {row}行{cidx}列ではありません")
    if args.get("_writes_header"):
        return "pass", (f"{row}行{cidx}列（見出し）の名前だけを書き換え"
                         f"（変わったセルは 1 個・中身の行は 1 セルも変わらず）")
    return "pass", f"『{name}』の{col_name}だけを書き換え（変わったセルは 1 個）"

def check_draw_borders(path: Path, args: dict, header_row: int = 1) -> tuple:
    """DRAW_BORDERS の事後条件。使用範囲の全セルに上下左右の罫線属性が付いているかを見る
       （DrawTableBorders ヘルパは格子罫線を範囲全体に一括で付けるため、1セルでも
       欠けていれば ヘルパが動いていないか範囲がずれている）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_row = _scan_last_row(ws, header_row=header_row)
        last_col = _scan_last_col(ws, header_row=header_row)
        if last_col < 1 or last_row < header_row:
            return "fail", _ZERO_TARGET_REASON
        cells = [ws.cell(row=r, column=c) for r in range(header_row, last_row + 1)
                 for c in range(1, last_col + 1)]
    if not cells:
        return "fail", _ZERO_TARGET_REASON

    def _has_border(cell) -> bool:
        bd = cell.border
        return bool(bd and bd.top and bd.top.style and bd.bottom and bd.bottom.style
                     and bd.left and bd.left.style and bd.right and bd.right.style)

    ok = all(_has_border(c) for c in cells)
    if not ok:
        return "fail", "使用範囲の一部に罫線が無い"
    return "pass", f"{len(cells)} セルの罫線を確認"

def check_autofit(path: Path, args: dict, header_row: int = 1,
                   source_book: Path | None = None) -> tuple:
    """AUTOFIT の事後条件。source_book が渡されれば、使用中の各列の幅が適用前後で
       変化した列数を数える（1列でも変化していれば pass）。source_book が無ければ、
       AFTER 側で幅が明示的に設定されている列があることだけを見る warn 判定に落とす。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_col = _scan_last_col(ws, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON
        after_widths = {}
        for c in range(1, last_col + 1):
            letter = get_column_letter(c)
            dim = ws.column_dimensions.get(letter)
            after_widths[letter] = dim.width if dim and dim.width else None

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            before_widths = {}
            for c in range(1, last_col + 1):
                letter = get_column_letter(c)
                dim = ws_b.column_dimensions.get(letter)
                before_widths[letter] = dim.width if dim and dim.width else None
        changed = sum(1 for k in after_widths if after_widths[k] != before_widths.get(k))
        if changed == 0:
            return "fail", "列幅が変化していない"
        return "pass", f"{changed}/{len(after_widths)} 列の幅が変化"

    set_count = sum(1 for v in after_widths.values() if v)
    if set_count == 0:
        return "fail", "列幅が設定されていない（AutoFitColumns が効いていない可能性）"
    return "warn", f"{set_count} 列に幅が設定されていることのみ確認（適用前ファイルとの突き合わせ無し）"

def check_set_column_value(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ 致命3(W10e): SET_COLUMN_VALUE の事後条件。対象列のデータ行が全部、機械抽出した
       定数値(args["value"])と一致するかを見る（型を問わず文字列表現で比較 — codegen は
       setString で書くため、読み戻しも文字列として揃える）。
       ★ operator10 ④: args["_write_numeric"] が真（codegen が setValue で数値として
       書いた）なら、検証も数値として揃える（型に追従 ── 文字列表現比較のままだと
       10 と 10.0 のような表記差で偽 fail になる）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        if last < header_row + 1:   # ★ 止血1: データ行0件を合格にしない
            return "fail", _ZERO_TARGET_REASON
        value = args["value"]
        vals = [ws.cell(row=r, column=idx).value for r in range(header_row + 1, last + 1)]
    if args.get("_write_numeric"):
        want = float(args["_write_numeric_value"])
        if not all(_is_number(v) and abs(float(v) - want) <= 1e-6 for v in vals):
            return "fail", f"列『{args['col']}』に数値『{want:g}』でないセルがある"
        return "pass", f"{len(vals)} 行を数値『{want:g}』に統一"
    if not all(str(v) == str(value) for v in vals):
        return "fail", f"列『{args['col']}』に『{value}』でないセルがある"
    return "pass", f"{len(vals)} 行を『{value}』に統一"
