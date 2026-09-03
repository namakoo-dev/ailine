"""事後条件 ── **行や列の数・位置が変わる** op。

対象 10 op: 並べ替え / 入れ替え / 行追加 / 列追加 / 行挿入 /
行削除 / 列削除 / 重複除去 / 合計追加 / セル分割

★ この群が確かめること: 動いたのが**頼まれた行や列だけ**で、
**中身は動いただけ**（値が書き換わっていない）。位置が変われば式の参照も
追随するので、参照の追随まで含めて見る。

★ 突き合わせは `_shared.compare_moved_rows` を通す: 挿入・削除・並べ替え・入れ替えが
同じ検算を要るのに、片方だけ直すのがこの repo の言う片配線。**1 関数に畳んである**。

★ 合計行はこの群で何度も問題になった: 並べ替えで本体に沈む（2026-08-18）、
列の入れ替えで見出しが動く（08-31）、`#REF!` で壊れる（08-29）。
合計行を「ただのデータ行」として動かすと、表が意味を失う。
"""
from __future__ import annotations

from pathlib import Path
import re

from openpyxl.utils import get_column_letter

from ailine_core import split_cell
from ailine_core.book_view import BookView
from ailine_core.primitives import is_number as _is_number
from ailine_core.sum_identity import rows_matching_sum_above
from ailine_core.table_scan import _cell_ref, _col_index_by_header, _scan_last_col, _scan_last_row, data_extent, extent_gap

from ailine_core.postconditions._shared import _ZERO_TARGET_REASON, _cells_for_shift, _moved_rows_note, _numeric_value, _row_as_shown, compare_moved_rows, note_stringy_numbers, note_unverified

def _dedup_key_display(key_tuple) -> str:
    """check_dedup の名指し用: 正規化キー（(型名, 値) のタプルの列）を人が読める文字列に。"""
    parts = [str(v) for _typ, v in key_tuple]
    return "・".join(parts) if parts else "(空)"

def _fmt_amount(v) -> str:
    """恒等式の説明に出す数値（整数なら小数点を付けない・指数表記にしない）。"""
    return str(int(v)) if float(v).is_integer() else str(v)

def _column_block_values(bv, ws, idx: int, header_row: int, sheet: str | None) -> list:
    """対象列を見出しの次の行から**空欄まで**走査した (行番号, 計算後の値) の並び。

    ★ 行の広がりは式ビュー（そのセルが空欄か）で決め、値は値ビュー（キャッシュ）から取る
      ―― 式で埋まった列を『空』と誤読せず、式文字列を数値と誤読しないための組み合わせ。
    ★ 空欄で切れる＝空行で区切られた2つ目の塊は読まない（別の表を続けて足すと、
      無関係な行が『上の合計』に当たってしまう）。"""
    out: list = []
    r = header_row + 1
    while ws.cell(row=r, column=idx).value not in (None, ""):
        out.append((r, bv.cell_value(r, idx, sheet=sheet)))
        r += 1
    return out

def _nested_total_reason(values: list, sheet_name: str, idx: int) -> str | None:
    """**足し込んだ範囲の最終行**が『自分より上の全部の合計』なら、その1行を名指しする。

    ★ 位置で判定する（2段構え）:
      1. 並びの最後の数値行は、今この op が書いた合計そのもの。範囲から外す
         （自分自身と照合したら恒真式に戻る）。
      2. 残った範囲の中でも、**最終行**で一致したときだけ鳴らす。既にある合計は
         その塊の一番下に在り、二重計上はそういう起き方をする。範囲の**真ん中**で
         一致するのは偶然の側 ―― 実測: demo/sales.xlsx の 部門/金額 =
         100,200,**300**,400,500,250 は 300 が開発部門のただの売上なのに
         100+200 と一致し、README の quickstart が exit 1 で落ちていた。
         手元の実検体 70 数値列で 発火 5 件（うち偽陽性 3）→ 2 件（どちらも本物の
         合計行）に下がった。語も書式も読まない性質は保たれる（算術＋位置だけ）。
    ★ 代償は取り逃がし: 『本体→小計→本体→小計→合計』のように**最終行でない**位置に
      合計がある帳票は、二重に数えていても鳴らない（見えない側に倒した）。"""
    numeric = [(row, value) for row, value in values if _is_number(value)]
    for hit in rows_matching_sum_above(numeric[:-1]):
        if not hit.is_last:
            continue
        span = f"{_cell_ref(hit.term_rows[0], idx)}:{_cell_ref(hit.term_rows[-1], idx)}"
        return (f"{sheet_name}!{_cell_ref(hit.row, idx)} の {_fmt_amount(hit.value)} は "
                f"{span} の合計と一致します"
                f"（この行まで足し込むと既にある合計を二重に数えます）")
    return None

def _total_row_left_the_bottom_reason(path: Path, source_book: Path | None, args: dict,
                                       header_row: int, idx: int, sheet_name: str) -> str | None:
    """並べ替えの前は最下行が『上の全部の合計』だったのに、後ではそうでなくなったら1行返す。

    ★ 判定は「存在」でなく「位置」: 並べ替え前の表にも合計行は在る（それが正常）。
      危険なのは合計と一致する行が**最下行でなくなった**とき ―― 合計行がただのデータとして
      一緒に並べ替えられ、表の途中に混ざった状態。並べ替え前は沈黙し、後にだけ鳴る。
    ★ source_book（適用前のコピー）が無い経路では黙る ―― before が無ければ位置の変化は
      測れない。測れないことを鳴らす側にも合格側にも寄せない（断定しない）。"""
    sheet = args.get("_target_sheet")
    if source_book is None or not Path(source_book).exists():
        return None
    with BookView(source_book) as bv_before:
        ws_before = bv_before.sheet(sheet)
        idx_before = _col_index_by_header(ws_before, args["col"], header_row=header_row)
        if idx_before is None:
            return None
        before_values = _column_block_values(bv_before, ws_before, idx_before, header_row, sheet)
    was_total = next((h for h in rows_matching_sum_above(before_values) if h.is_last), None)
    if was_total is None:
        return None                      # 元から最下行に合計は無い＝この検査の対象外
    with BookView(path) as bv_after:
        ws_after = bv_after.sheet(sheet)
        after_values = _column_block_values(bv_after, ws_after, idx, header_row, sheet)
    if any(h.is_last for h in rows_matching_sum_above(after_values)):
        return None                      # 合計はまだ最下行にある
    moved_to = next((r for r, v in after_values
                     if _is_number(v) and abs(v - was_total.value) <= 1e-6), None)
    where = (f"{sheet_name}!{_cell_ref(moved_to, idx)}" if moved_to
             else "並べ替えた後の表の中に見つかりません")
    return (f"並べ替える前は最下行 {sheet_name}!{_cell_ref(was_total.row, idx_before)} の "
            f"{_fmt_amount(was_total.value)} が上の {len(was_total.term_rows)} 行の合計でしたが、"
            f"並べ替えた後は {where} にあり最下行ではありません"
            f"（合計行がデータとして一緒に並べ替えられています）")

def check_sort(path: Path, args: dict, header_row: int = 1, use_formula: bool = False,
                source_book: Path | None = None) -> tuple:
    """SORT の事後条件。戻り値は (status, reason)。status ∈ {"pass","warn","fail"}。
       ★ 止血1: 検証対象が0件なら fail、1件（順序が定義できない）なら warn とし、
       どちらも「機械検証済み」とは名乗らない。
       ★ 止血2: 合計行等の非数値/None セルは比較から除外し、除外件数を表示する
       （C②: None >= int の生トレースバックの根治）。全部除外なら0件と同じ扱い。
       ★ W3: header_row(1起点、省略時1) が『接地・codegen』と同じ見出し行を指す。
       ★ W10f 項目1: use_formula のとき対象列を data_only(計算後の値)側から読む
       （check_compute_column と同型のバグ — 前段が式で作った計算列(小計等)を SORT の
       対象列にした場合、raw 側は式文字列のままで全行『数値でない』扱いになっていた）。
       SORT は相対順序という『全行をまたぐ』検証のため、式にキャッシュ値が無く読めない
       行が1件でもあれば、その行を除いた残りだけで『順序OK』と判定するのは危険
       （除いた行が実際は順序を崩していても見逃す＝COMPUTE_COLUMN の行独立検証とは違い
       部分採点できない）。0cf9218 空虚な検証合格の禁止の趣旨のまま fail で打ち切る。
       ★ 算術恒等の検算: 並び順が合っていても、合計行がデータとして一緒に並べ替えられて
       表の途中に混ざったら ✓ は出さない（source_book が渡された経路のみ・
       _total_row_left_the_bottom_reason 参照）。"""
    sheet_name = args.get("_target_sheet")
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        sheet_name = sheet_name or ws.title
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        last = _scan_last_row(ws, header_row=header_row)
        # ★★ 2026-08-29（Namakoo が実測）: 合計行を並べ替えの対象から外したら、今度は
        #   **検算側が合計行を数えて**「指定順に並んでいない」と言った（合計は一番大きい）。
        #   ★ 並べ替えなかった行を、並び順の検算に入れない ── 宣言（_sort_end_row）を
        #     そのまま分母にする。宣言と検算が同じ範囲を見る、という当たり前を通す。
        if args.get("_sort_end_row"):
            last = min(last, int(args["_sort_end_row"]))
        # ★ 塊②（2026-08-25）: 分母を**物理の使用範囲**と突き合わせる。
        #   A 列走査は最初の空で止まるので、末尾に A 列が空の行があると
        #   処理からも分母からも消える（実測: 真の分母 5 を「3行中」と言っていた）。
        # ★ 2026-08-25: 食い違いの**記録**は run_postcondition の入口に畳んだ
        #   （note_extent_gap）。ここで gap を取るのは phys_cols を得るためだけ ──
        #   ここでも数えると二重計上になる。
        gap = extent_gap(ws, header_row=header_row)
        # ★ 行の同一性を確かめるための、**物理の列範囲**での行の中身（後段で使う）。
        phys_cols = max(gap["cols_physical"], gap["cols_scanned"])
        after_rows = [tuple(ws.cell(row=r, column=c).value for c in range(1, phys_cols + 1))
                       for r in range(header_row + 1, last + 1)]
        raw_vals = [ws.cell(row=r, column=idx).value for r in range(header_row + 1, last + 1)]
        if use_formula:
            eff_vals = [bv.cell_value(r, idx, sheet=args.get("_target_sheet")) for r in range(header_row + 1, last + 1)]
        else:
            eff_vals = raw_vals
    vals = []
    excluded = 0
    uncached = 0
    for rv, ev in zip(raw_vals, eff_vals):
        # ★ 2026-08-24: 日付/日時もシリアル値として検証対象に入れる（_numeric_value 参照）。
        num = _numeric_value(ev)
        if num is not None:
            vals.append(num)
        elif use_formula and isinstance(rv, str) and rv.startswith("="):
            uncached += 1   # ★ W10f: 式はあるがキャッシュ値が無い（『対象が無い』とは別）
        else:
            excluded += 1
    note = f"（数値でない {excluded} 行は対象外）" if excluded else ""
    # ★ 塊①: 除外を**機械の値**として残す（今までは文章にして捨てていた）。
    note_unverified(args, excluded, "数値でないため並び順を確かめられない")
    if uncached:
        return "fail", (f"並び順の検証対象に式はあるがキャッシュ値が無く検証できない行が "
                         f"{uncached} 件あり、順序を検証できません"
                         f"（LibreOffice を通していない可能性）{note}")
    if len(vals) == 0:
        return "fail", _ZERO_TARGET_REASON + note
    if len(vals) == 1:
        return "warn", f"検証対象が1行のみ（並べ替えの意味がありません）{note}"
    asc = args["order"] == "asc"
    ok = (all(vals[i] <= vals[i + 1] for i in range(len(vals) - 1)) if asc
          else all(vals[i] >= vals[i + 1] for i in range(len(vals) - 1)))
    if not ok:
        return "fail", f"列『{args['col']}』が指定順（{args['order']}）に並んでいない{note}"
    moved = _total_row_left_the_bottom_reason(path, source_book, args, header_row, idx, sheet_name)
    if moved:
        return "fail", moved
    # ★★ 塊②（2026-08-25・中核 op の盲検 致命1）: **行の同一性**を確かめる。
    #   旧版はキー列の単調性しか見ておらず、見出しの無い列が置き去りになって
    #   全行の備考が別商品の物に付け替わっても ✓ を出していた（実測）。
    #   ★ 比べるのは**物理の列範囲**（見出し由来の範囲では、まさに壊れた列が入らない）。
    #   ★ 並べ替えは行の集合を変えない操作なので、行の中身の多重集合が一致すれば同一性が保たれる。
    torn = _sort_rows_lost_their_identity(source_book, after_rows, args, header_row, phys_cols)
    if torn:
        return "fail", torn
    return "pass", f"{len(vals)} 行を検証（{'昇順' if asc else '降順'}）{note}"

def _sort_rows_lost_their_identity(source_book, after_rows: list, args: dict,
                                    header_row: int, phys_cols: int) -> str | None:
    """並べ替えの前後で、**行そのもの**が保たれているか。壊れていれば理由、無事なら None。

    ★ 並べ替えは行を入れ替えるだけで、行の中身は 1 つも変わらないはず。
      前後の行を多重集合として比べれば、順序に依らず同一性が見える。
      ★ 見出し由来の列範囲で比べると、範囲外に落ちて壊れた列がそもそも入らないので
        必ず一致してしまう（＝恒真）。物理の使用範囲で比べる。
    ★ source_book が無い経路では確かめられない ── その時は黙って None（従来どおり）。
    """
    if source_book is None or not Path(source_book).exists():
        return None
    try:
        with BookView(Path(source_book)) as bv:
            ws = bv.sheet(args.get("_target_sheet"))
            last = _scan_last_row(ws, header_row=header_row)
            # ★ 2026-08-29: 並べ替えなかった行（合計行）は**前後どちらの分母からも**外す。
            #   片側だけ縮めると「行数が変わった」と言う ── 分母は 1 箇所で決める。
            if args.get("_sort_end_row"):
                last = min(last, int(args["_sort_end_row"]))
            before_rows = [tuple(ws.cell(row=r, column=c).value for c in range(1, phys_cols + 1))
                            for r in range(header_row + 1, last + 1)]
    except Exception:
        return None
    if len(before_rows) != len(after_rows):
        return (f"並べ替えの前後で行数が変わっています"
                f"（元 {len(before_rows)} 行 / 後 {len(after_rows)} 行）")
    # ★★ 2026-08-25 の回帰の直し（盲検 2 回目 R1）: 相対参照の式は、行が動けば
    #   **変わるのが正しい**（=B2*C2 → =B5*C5）。生の値で比べると「ちぎれた」と
    #   誤判定し、**この製品の看板ユースケース**「金額列を作って金額順に並べる」が
    #   必ず落ちていた（出力は完全に正しいのに検算だけが間違っていた）。
    #   ★ 式のセルは同一性の材料にしない ── 行が動いた事実そのものを表すので、
    #     「変わっていないこと」を要求できない。
    def _identity(row):
        return tuple("" if (isinstance(v, str) and v.startswith("=")) else v for v in row)

    from collections import Counter
    before_rows = [_identity(r) for r in before_rows]
    after_rows = [_identity(r) for r in after_rows]
    lost = Counter(before_rows) - Counter(after_rows)
    if not lost:
        return None
    sample = next(iter(lost))
    shown = "・".join("(空)" if v in (None, "") else str(v) for v in sample[:4])
    return (f"並べ替えで行がちぎれています（{len(list(lost.elements()))} 行が元と一致しません）。"
            f"例: 元にあった行 {shown} が出力にありません"
            " ── 見出しの無い列が一緒に動かなかった可能性があります")

_APPEND_TOTAL_FORMULA_RE = re.compile(
    r"^=SUM\([A-Za-z]{1,3}\d+:INDEX\([A-Za-z]{1,3}:[A-Za-z]{1,3},ROW\(\)-1\)\)(\*[\d.]+)?$",
    re.IGNORECASE)

def check_append_total(path: Path, args: dict, header_row: int = 1) -> tuple:
    """★ W6: APPEND_TOTAL の事後条件（二層・COMPUTE_COLUMN の use_formula 版と同じ考え方）。
       ①式文字列が期待形『=SUM(<列><行>:INDEX(<列>:<列>,ROW()-1))』（★ B: 挿入耐性式・
       factor!=1 のときは末尾に *factor が付く）②data_only 読みのキャッシュ値が
       「列合計×factor」と一致（浮動小数は丸め許容）。対象列が表の最左端でない場合は、
       その左隣セルに期待ラベルが立っていることも確認する（codegen_dsl の配置と対）。
       ★ B: codegen は LO 方言（INDEX の引数区切りがセミコロン）で書くが、保存後は
       カンマ区切りに自動変換される（実測）ので、ここではその保存後のカンマ形と照合する。

       ★ 合計行の位置は『対象列そのものに "=SUM(" で始まる式が最初に現れた行』として直接
       探す（生読み・型判定はしない）。実測で分かった罠が2つあり、どちらもこの探し方だけ
       で同時に避けられる:
       - COMPUTE_COLUMN の式化（W3 Part3・既定）で対象列自体が "=B2*C2" 型の式で埋まって
         いることがある。生セル値の型（数値かどうか）だけで『データ行/合計行』を分けようと
         すると、データ行の式もキャッシュ値未読の合計行の式も同じ『文字列』に見えて区別
         できない（E2E 実測: 小計列を先に式化してから APPEND_TOTAL するケースで発覚）。
       - 最終データ行の判定に列1(左端)を走査する既存の _scan_last_row は使わない。左端が
         分類列で、かつ対象列がその右隣（＝ラベルの置き場所が偶然その分類列と重なる、例:
         2列だけの帳票）だと、書き込んだラベルの文字列自体が『データ行』として誤って
         数えられ off-by-one になる（実測）。
       "=SUM(" という固有の目印を対象列自身の中だけで探すので、他列の中身にも
       COMPUTE_COLUMN の式の形にも影響されない。

       ★★ 算術恒等の検算（二重計上）を**最初に**行う: 下の①②は期待値を「合計式が
       生成したのと同じ範囲」から作るので、既存の合計を足し込んでいても両方通ってしまう
       （検算が被検算と同じ盲点を使う恒真式）。しかも既存の合計が『式』で入っている表では
       "=SUM(" の初出行がその既存行を掴み、①が「合計の式が期待形でない」と**ユーザーが
       自分で書いた行を責める**誤診断を出す。どちらの誤りも、対象列全体を数値として見る
       この検算が先に立てば正しい行を名指しできる（_nested_total_reason 参照）。"""
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        idx = _col_index_by_header(ws, args["col"], header_row=header_row)
        if idx is None:
            return "fail", f"列『{args['col']}』が見つからない"
        nested = _nested_total_reason(
            _column_block_values(bv, ws, idx, header_row, args.get("_target_sheet")),
            args.get("_target_sheet") or ws.title, idx)
        if nested:
            return "fail", nested
        r = header_row + 1
        while True:
            v = ws.cell(row=r, column=idx).value
            if v is None:
                break
            if isinstance(v, str) and v.replace(" ", "").startswith("=SUM("):
                break
            r += 1
        last = r - 1
        if last < header_row + 1:
            return "fail", _ZERO_TARGET_REASON
        total_row = r
        got_formula = ws.cell(row=total_row, column=idx).value
        got_formula_norm = got_formula.replace(" ", "") if isinstance(got_formula, str) else ""
        if not _APPEND_TOTAL_FORMULA_RE.match(got_formula_norm):
            detail = f"{total_row}行目: 合計の式が期待形(挿入耐性 SUM 型)でない (実際 {got_formula!r})"
            return "fail", detail
        label_ok = True
        want_label = str(args.get("label") or "合計")
        got_label = None
        if idx > 1:
            # ★ 2026-08-28: ラベルは**1 列目**（codegen と同じ場所を見る）。
            #   旧は対象列の左隣を見ていた ── その置き方が 1 列目を空にして、
            #   道具自身の走査を止めていた（合計を出すと ✓ が永久に出なかった）。
            got_label = ws.cell(row=total_row, column=1).value
            label_ok = got_label == want_label
        if not label_ok:
            return "fail", f"{total_row}行目: ラベルが期待『{want_label}』と不一致 (実際 {got_label!r})"

        raw_vals = [bv.cell_value(rr, idx, sheet=args.get("_target_sheet")) for rr in range(header_row + 1, last + 1)]
        nums = [v for v in raw_vals if _is_number(v)]
        # ★ 塊③（2026-08-25・致命4）: 画面に 1000・2000・3000 が並んでいるのに合計 4000 に
        #   なるのは、文字列の「2000」が数え上げから落ちているから。判定は変えず（落とす）、
        #   別実装（compare_blocked.looks_numeric）で見て**言う**。
        note_stringy_numbers(args, [v for v in raw_vals if not _is_number(v)])
        got_cached = bv.cell_value(total_row, idx, sheet=args.get("_target_sheet"))
    if not nums:
        return "fail", _ZERO_TARGET_REASON
    factor = float(args.get("factor", 1) or 1)
    want_total = sum(nums) * factor
    if not _is_number(got_cached) or abs(got_cached - want_total) > 1e-6:
        return "fail", f"{total_row}行目: 合計のキャッシュ値が不一致 (期待 {want_total} 実際 {got_cached!r})"
    # ★★ 単位I: 主張の範囲を狭める ── **どの行に合計が立っているか**を pass でも言う。
    #   非対称だった: 失敗時は「5行目: …」と行番号を言うのに、成功時は言っていなかった。
    #   ★ 見せる必要が一番あるのは成功したときの方（盲検査定の致命①: 手入力の合計行がある
    #   ブックで、書かれた先が見えないまま ✓ が出ていた）。
    #   ★ 「何を検証したか」でなく「どこを検証したか」まで述べる。件数だけでは範囲を主張できない。
    return "pass", (f"{total_row}行目の合計を検証"
                    f"（{len(nums)} 行分・式・キャッシュ値・ラベルとも一致）")

def check_add_row(path: Path, args: dict, header_row: int = 1,
                   source_book: Path | None = None) -> tuple:
    """ADD_ROW の事後条件。

    ★ 2026-08-26: Namakoo が GUI を触って見つけた欠け（21 op のどれにも
      「データを 1 行足す」が無かった）への実装。
    ★ 何を証明するか ── **押し下げたこと**（末尾に足しただけではないこと）:
      ① 行数がちょうど 1 増えている
      ② at 行目が、宣言した列に宣言した値を持つ（型込みで等値）
      ③ **at 行目以降の元の行が、1 行ずれてそのまま在る**（上書きしていない）
      ④ at より上の行が 1 セルも変わっていない
    ★ source_book が無ければ②だけを見る warn（断定しない・INSERT_ROWS と同じ劣化）。
    """
    at = int(args["at"])
    values = args.get("values") or {}
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        _last_a, last_col = data_extent(ws, header_row)
        headers = [str(ws.cell(row=header_row, column=c).value or "")
                    for c in range(1, last_col + 1)]
        wrong = []
        for name, want in values.items():
            if str(name) not in headers:
                wrong.append(f"列『{name}』が見つからない")
                continue
            got = ws.cell(row=at, column=headers.index(str(name)) + 1).value
            if got != want:
                wrong.append(f"{name}: {want!r} のはずが {got!r}")
        if wrong:
            return "fail", f"{at}行目の値が宣言と違う（{'／'.join(wrong)}）"
        if source_book is None or not Path(source_book).exists():
            return "warn", f"{at}行目の値のみ確認（適用前ファイルとの突き合わせ無し）"
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            _lb, _cb = data_extent(ws_b, header_row)
            # ★ 前後で列幅を揃えてから読む（片方だけ広いと、ずれが「値の変化」に化ける）
            cols = max(last_col, _cb)
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"),
                                           header_row, _last_a, cols)
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"),
                                            header_row, _lb, cols)
    if len(after_rows) != len(before_rows) + 1:
        return "fail", (f"行数が合わない（適用前 {len(before_rows)} 行 → "
                         f"適用後 {len(after_rows)} 行・1 行増えるはず）")
    k = at - header_row - 1          # 挿入位置（データ行の 0 起点）
    if k < 0 or k > len(before_rows):
        return "fail", f"{at}行目は表の範囲外です（データは {len(before_rows)} 行）"
    # ④ 上は不変 ── ★ 比べ方は 1 箇所（compare_moved_rows）にしか無い
    st_up, info_up = compare_moved_rows(after_rows[:k], before_rows[:k], f"{at}行目より上")
    if st_up == "broken":
        return "fail", f"{at}行目より上の行が変わっている（挿入で既存行を壊した疑い）"
    # ③ 下は 1 行ずれてそのまま
    st_dn, info_dn = compare_moved_rows(after_rows[k + 1:], before_rows[k:], f"{at}行目より下")
    if st_dn == "broken":
        return "fail", (f"{at}行目より下の行が元のままでない ── "
                         "押し下げずに上書きした疑いがあります")
    moved = list(info_up) + list(info_dn)
    if moved:
        return "warn", _moved_rows_note(moved)
    return "pass", f"{at}行目に 1 行追加（上下の行は元のまま・値は宣言どおり）"

def check_delete_rows(path: Path, args: dict, header_row: int = 1,
                       source_book: Path | None = None) -> tuple:
    """DELETE_ROWS の事後条件。**詰めたこと**を証明する。

    ① 行数がちょうど count 減っている
    ② 残った行が、適用前から消した分を抜いた並びと**順序ごと連続で**一致する
       （＝空行が残っていない・別の行を巻き込んでいない）
    ★ 消した中身は呼び出し側が画面に出す（args["_deleted"] に積む）──
      「消えたものは差分に出ない」への処置。
    """
    at = int(args["at"])
    count = int(args.get("count", 1) or 1)
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        _la, last_col = data_extent(ws, header_row)
        if source_book is None or not Path(source_book).exists():
            return "warn", "適用前ファイルが無いため、消えた行を確かめられていません"
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            _lb2, _cb2 = data_extent(ws_b, header_row)
            cols = max(last_col, _cb2)
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"),
                                           header_row, _la, cols)
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"),
                                            header_row, _lb2, cols)
    k = at - header_row - 1
    if k < 0 or k >= len(before_rows):
        return "fail", f"{at}行目は表の範囲外です（データは {len(before_rows)} 行）"
    expected = before_rows[:k] + before_rows[k + count:]
    if len(after_rows) != len(expected):
        return "fail", (f"行数が合わない（適用前 {len(before_rows)} 行から {count} 行消えて "
                         f"{len(expected)} 行のはずが {len(after_rows)} 行）")
    # ★ 削除でも式は上へ追随する ── 挿入と**同じ 1 箇所**で比べる（片配線を作らない）
    st, info = compare_moved_rows(after_rows, expected, "残った行")
    if st == "broken":
        return "fail", "残った行の並びが元と違う ── 詰め方が正しくない疑いがあります"
    note_deleted(args, [tuple(c[1] for c in before_rows[i])
                         for i in range(k, min(k + count, len(before_rows)))])
    if info:
        return "warn", _moved_rows_note(info)
    return "pass", f"{at}行目から {count} 行を削除（残りは順序ごと元のまま）"

def check_add_column(path: Path, args: dict, header_row: int = 1,
                      source_book: Path | None = None) -> tuple:
    """ADD_COLUMN の事後条件。**押し出したこと**を証明する（上書きしていないこと）。

    ① 列がちょうど 1 本増えている
    ② 宣言した位置に、宣言した名前の見出しが在る（名前が空なら見出しも空）
    ③ **他の列が 1 セルも変わっていない** ── ただし式は列が動けば参照が追随する
      （`=B2*C2` → `=B2*D2`）ので、比べ方は compare_moved_rows に任せる（追加・削除・
      入れ替えと**同じ 1 箇所**）。
    ④ 挿した列のデータ行は空（見出し以外を勝手に埋めていない）
    """
    at = int(args.get("_at_col") or 0)
    name = str(args.get("name") or "")
    if at < 1:
        return "warn", "挿す位置が分からないため確かめられていません"
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_a, cols_a = data_extent(ws, header_row)
        headers_after = [str(ws.cell(row=header_row, column=c).value or "")
                          for c in range(1, cols_a + 1)]
        if source_book is None or not Path(source_book).exists():
            return "warn", f"{at}列目に列を挿したことのみ確認（適用前ファイルとの突き合わせ無し）"
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            last_b, cols_b = data_extent(ws_b, header_row)
            headers_before = [str(ws_b.cell(row=header_row, column=c).value or "")
                               for c in range(1, cols_b + 1)]
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"),
                                           header_row, max(last_a, last_b), cols_a)
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"),
                                            header_row, max(last_a, last_b), cols_b)
    if len(headers_after) != len(headers_before) + 1:
        return "fail", (f"列数が合わない（適用前 {len(headers_before)} 列 → "
                         f"適用後 {len(headers_after)} 列・1 本増えるはず）")
    j = at - 1
    if j > len(headers_before):
        return "fail", f"{at}列目は表の外です（適用前は {len(headers_before)} 列）"
    if headers_after[j] != name:
        return "fail", (f"{at}列目の見出しが『{headers_after[j]}』で、"
                         f"宣言した『{name}』と違います")
    want = headers_before[:j] + [name] + headers_before[j:]
    if headers_after != want:
        return "fail", (f"見出しの並びが宣言と違う（適用後: "
                         f"{"、".join(h or "（空）" for h in headers_after)}）")
    # ④ 挿した列のデータ行は空
    filled = [i + header_row + 1 for i, row in enumerate(after_rows)
               if j < len(row) and row[j][1] not in (None, "")]
    if filled:
        return "fail", (f"挿した{at}列目に値が入っています（{'、'.join(map(str, filled[:5]))}行目）"
                         " ── 空の列を作るはずです")
    # ③ 他の列は元のまま（式は追随してよい）
    expected = [tuple(list(row[:j]) + [(False, None, None)] + list(row[j:]))
                 for row in before_rows]
    st, info = compare_moved_rows(after_rows, expected, "挿したあと")
    if st == "broken":
        return "fail", f"{at}列目に列を挿したとき、他の列が変わっています ── {info}"
    if info:
        return "warn", _moved_rows_note(info)
    label = f"『{name}』" if name else "見出しの無い列"
    return "pass", f"{at}列目に{label}を挿入（他の列は 1 セルも変わらず）"

def _check_swap_cells(path: Path, args: dict, header_row: int,
                       source_book) -> tuple:
    """セル 2 つの入れ替えの事後条件。**2 つだけが、互いの値になった**ことを証明する。

    ★★ 2026-08-31（Namakoo が実測）:「丸和物流の単価とみどり建設の単価を入れ替えて」で
      **行を丸ごと入れ替えて ✓ を出していた**（16 セルが動いた・頼んだのは 2 セル）。
      番人は宣言（行の入れ替え）どおりかを見るので、宣言が違えば通ってしまう。
    ★ だからここで見るのは 3 つ:
      ① 宣言した 2 セルが、**互いの値**になっている
      ② 変わったセルは**ちょうど 2 個**（列や行を巻き込んでいない）
      ③ その 2 個の座標が宣言と一致する
    ★ ②③ は適用前が要る。無ければ①だけの warn（断定しない・他の op と同じ劣化）。
    """
    cells = [tuple(c) for c in (args.get("_cells") or [])]
    want = list(args.get("_cell_values") or [])
    if len(cells) != 2 or len(want) != 2:
        return "fail", "入れ替える 2 セルが宣言されていません"
    sheet = args.get("_target_sheet")
    with BookView(path) as bv:
        got = [bv.cell_value(r, c, sheet) if bv.cell_formula(r, c, sheet) is not None
                else bv.sheet(sheet).cell(row=r, column=c).value
                for r, c in cells]
        _last, _wide = data_extent(bv.sheet(sheet), header_row)
        after = {(r, c): bv.sheet(sheet).cell(row=r, column=c).value
                  for r in range(header_row, _last + 1) for c in range(1, _wide + 1)}
    # ① 互いの値になっているか（式セルは計算後の値で見る ── 他の op と同じ線）
    for i, (r, c) in enumerate(cells):
        w = want[1 - i]
        if isinstance(w, str) and w.startswith("="):
            continue                      # 式は②③と読み戻しで見る（文字比較しない）
        if str(got[i]) != str(w):
            return "fail", (f"{r}行{c}列 が {w!r} になっていません（実際 {got[i]!r}）")
    if source_book is None or not Path(source_book).exists():
        return "warn", (f"{cells[0][0]}行{cells[0][1]}列 と {cells[1][0]}行{cells[1][1]}列 "
                         "の入れ替えだけ確認（変えていないセルは見ていません）")
    with BookView(source_book) as bv_b:
        ws_b = bv_b.sheet(sheet)
        lb, cb = data_extent(ws_b, header_row)
        before = {(r, c): ws_b.cell(row=r, column=c).value
                   for r in range(header_row, lb + 1) for c in range(1, cb + 1)}
    # ②③ 変わったのはちょうど 2 個で、その座標が宣言と一致する
    changed = sorted(k for k in set(before) | set(after)
                      if before.get(k) != after.get(k))
    if len(changed) != 2:
        return "fail", (f"2 セルのはずが {len(changed)} セル変わっています"
                         f"（{'、'.join(f'{r}行{c}列' for r, c in changed[:6])}）"
                         " ── 行や列ごと動かした疑いがあります")
    if changed != sorted(cells):
        return "fail", (f"変わったのは {'、'.join(f'{r}行{c}列' for r, c in changed)} で、"
                         f"宣言した {'、'.join(f'{r}行{c}列' for r, c in sorted(cells))} "
                         "ではありません")
    return "pass", (f"{cells[0][0]}行{cells[0][1]}列 と {cells[1][0]}行{cells[1][1]}列 "
                     "の 2 セルだけを入れ替え（他は 1 セルも変わらず）")

def check_swap(path: Path, args: dict, header_row: int = 1,
                source_book: Path | None = None) -> tuple:
    """SWAP の事後条件。**入れ替わったこと**を、適用前との突き合わせで証明する。

    ★★ 実測が設計を決めた（bench/swap_formula_spike_RESULTS.md）: 値を文字として交換すると
      **式が壊れる**（各行の計算結果が他の行の値になる）。見た目は正しく並ぶので人は気づけない。
      ★ だから「並びが入れ替わったか」だけを見てはいけない ── **中身が自分の値のまま
        移ったか**まで見る。式セルは文字ではなく**計算後の値**で突き合わせる。
    ★ 比べ方は 1 箇所（compare_moved_rows）── 追加・削除・入れ替えで同じ関数しか通らない。
    ★ 入れ替えでは、動いていない行の計算結果が変わる理由が無い（挿入と違って行は増減しない）
      ので、開示（warn）ではなく **fail** に倒す。ここが挿入と違う唯一の点。
    """
    axis = args.get("_axis") or "row"
    if axis == "cell":
        return _check_swap_cells(path, args, header_row, source_book)
    a, bname = str(args.get("a", "")), str(args.get("b", ""))
    if source_book is None or not Path(source_book).exists():
        return "warn", "適用前ファイルが無いため、入れ替わったことを確かめられていません"
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        last_a, cols_a = data_extent(ws, header_row)
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            last_b, cols_b = data_extent(ws_b, header_row)
            cols = max(cols_a, cols_b)
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"),
                                           header_row, last_a, cols)
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"),
                                            header_row, last_b, cols)
            headers_after = [str(ws.cell(row=header_row, column=c).value or "")
                              for c in range(1, cols + 1)]
            headers_before = [str(ws_b.cell(row=header_row, column=c).value or "")
                               for c in range(1, cols + 1)]
    if len(after_rows) != len(before_rows):
        return "fail", (f"行数が変わっています（適用前 {len(before_rows)} 行 → "
                         f"適用後 {len(after_rows)} 行・入れ替えで増減はしないはず）")
    if axis == "column":
        try:
            i, j = headers_before.index(a), headers_before.index(bname)
        except ValueError:
            return "fail", f"適用前に『{a}』『{bname}』の列が見つかりません"
        want_headers = list(headers_before)
        want_headers[i], want_headers[j] = want_headers[j], want_headers[i]
        if headers_after != want_headers:
            return "fail", (f"見出しが入れ替わっていません（適用後: "
                             f"{"、".join(h for h in headers_after if h)}）")
        expected = []
        for row in before_rows:
            r = list(row)
            r[i], r[j] = r[j], r[i]
            expected.append(tuple(r))
        where = f"列『{a}』と列『{bname}』"
    else:
        ra, rb = args.get("_a_pos"), args.get("_b_pos")
        if not ra or not rb:
            return "warn", "入れ替える行の位置が分からないため確かめられていません"
        i, j = int(ra) - header_row - 1, int(rb) - header_row - 1
        if not (0 <= i < len(before_rows) and 0 <= j < len(before_rows)):
            return "fail", "入れ替える行が表の範囲外です"
        expected = list(before_rows)
        expected[i], expected[j] = expected[j], expected[i]
        if headers_after != headers_before:
            return "fail", "見出しが変わっています（行の入れ替えで見出しは動かないはず）"
        where = f"『{a}』の行と『{bname}』の行"
    st, info = compare_moved_rows(after_rows, expected, "入れ替え後")
    if st == "broken":
        return "fail", f"{where} が入れ替わっていません ── {info}"
    if info:
        # ★ 入れ替えで計算結果が変わったら、それは実測した「静かに壊れる」形そのもの。
        # ★★ 2026-08-29（Namakoo の実測・デモ材料そのもの）: 「税込み金額と金額を
        #   入れ替えて」で、データ側は正しく追従した（税込み列は `=F2*1.1` になった）のに、
        #   合計式だけ範囲の**片側しか追従せず** `=SUM(E2:INDEX(F:F,ROW()-1))` になり、
        #   E2 から F8 までの**二列**を足していた（両方の合計が 1,000,440）。
        #   ★ 画面には大きな数字が出るだけなので、人はまず気づかない。
        return "fail", (f"{where} を入れ替えたあと、式の計算結果が変わっています ── "
                         + _moved_rows_note(info, why=(
                             "式が入れ替え先の列に付いていきませんでした"
                             "（範囲の片側だけが動くと、二列にまたがる合計になります）")))
    return "pass", f"{where} を入れ替え（中身は自分の値のまま移動・他は 1 セルも変わらず）"

def check_delete_column(path: Path, args: dict, header_row: int = 1,
                         source_book: Path | None = None) -> tuple:
    """DELETE_COLUMN の事後条件。**他の列が 1 セルも変わらない**ことを証明する。"""
    name = str(args["col"])
    with BookView(path) as bv:
        ws = bv.sheet(args.get("_target_sheet"))
        _lc, last_col = data_extent(ws, header_row)
        headers_after = [str(ws.cell(row=header_row, column=c).value or "")
                          for c in range(1, last_col + 1)]
        if name in headers_after:
            return "fail", f"列『{name}』がまだ在ります（削除されていない）"
        if source_book is None or not Path(source_book).exists():
            return "warn", f"列『{name}』が無いことのみ確認（適用前ファイルとの突き合わせ無し）"
        with BookView(source_book) as bv_b:
            ws_b = bv_b.sheet(args.get("_target_sheet"))
            _lrb, lc_b = data_extent(ws_b, header_row)
            headers_before = [str(ws_b.cell(row=header_row, column=c).value or "")
                               for c in range(1, lc_b + 1)]
            if name not in headers_before:
                return "fail", f"適用前にも列『{name}』が無い（消した対象が特定できない）"
            j = headers_before.index(name)
            # ★ 列を消すと、右にあった列は左へ寄る ── 式の列参照も追随する。
            #   行の挿入/削除と**同じ 1 箇所**で比べる（3 経路に同じ判断を書き写さない）。
            after_rows = _cells_for_shift(bv, args.get("_target_sheet"),
                                           header_row, _lc, last_col)
            before_rows = _cells_for_shift(bv_b, args.get("_target_sheet"),
                                            header_row, _lrb, lc_b)
    if headers_after != headers_before[:j] + headers_before[j + 1:]:
        return "fail", "見出しの並びが元と違う（別の列を巻き込んだ疑い）"
    expected = [r[:j] + r[j + 1:] for r in before_rows]
    st, info = compare_moved_rows(after_rows, expected, "残った列")
    if st == "broken":
        return "fail", "残った列の中身が元と違う ── 別の列を巻き込んだ疑いがあります"
    note_deleted(args, [(r[j][1],) for r in before_rows])
    if info:
        return "warn", _moved_rows_note(info)
    return "pass", f"列『{name}』を削除（残りの列は 1 セルも変わらず）"

def note_deleted(args: dict, rows) -> None:
    """消した中身を機械の値として残す（呼び出し側が人に見せる）。

    ★ 「消えたものは差分に出ない」── 削除は**画面の差分に何も出ない**操作なので、
      何を消したかを言わなければ、人は取り返しがつくかを判断できない。
    """
    if not isinstance(args, dict) or not rows:
        return
    args["_deleted"] = [list(r) for r in rows][:20]

def check_insert_rows(path: Path, args: dict, header_row: int = 1,
                       source_book: Path | None = None) -> tuple:
    """INSERT_ROWS の事後条件。source_book(適用前のコピー)が渡されれば、
       ①データ最終行が count 増えている ②挿入位置(at)以降の内容が、適用前の対応行から
       count 行分ずれて現れている（シフト）③挿入された行自体が空欄、を突き合わせる。
       ★ source_book が無い（複合計画の途中段等で before が用意できない経路）場合は、
       挿入位置が空欄であることだけを見る warn 判定に落とす（保守的・断定しない）。"""
    at = int(args["at"])
    count = int(args.get("count", 1) or 1)

    if source_book is None or not Path(source_book).exists():
        with BookView(path) as bv:
            ws = bv.sheet(args.get("_target_sheet"))
            # ★ 2026-08-27: 空行が 1 本あると走査が止まって「検証対象が 0 件」になっていた。
            _lr0, last_col = data_extent(ws, header_row)
            row_cells = [ws.cell(row=at, column=c).value for c in range(1, last_col + 1)]
        if all(v in (None, "") for v in row_cells):
            return "warn", "挿入位置が空欄であることのみ確認（適用前ファイルとの突き合わせ無し）"
        return "fail", f"{at}行目が空欄でない（挿入されていない可能性）"

    with BookView(source_book) as bv_before, BookView(path) as bv_after:
        ws_before = bv_before.sheet(args.get("_target_sheet"))
        last_before, last_col = data_extent(ws_before, header_row)
        if last_col < 1 or last_before < header_row + 1:
            return "fail", _ZERO_TARGET_REASON

        ws_after = bv_after.sheet(args.get("_target_sheet"))

        # ★ 挿入は AFTER 側に意図的な空行（挿入行そのもの）を作るため、連続データを前提にする
        #   _scan_last_row を AFTER 側の「最終行」検出には使わない（挿入行で即座に打ち切られて
        #   しまう）。BEFORE の各データ行が「count 行下」に正しく現れているかを直接照合する。
        mismatches = checked = 0
        for r in range(max(at, header_row + 1), last_before + 1):
            for c in range(1, last_col + 1):
                checked += 1
                if ws_before.cell(row=r, column=c).value != ws_after.cell(row=r + count, column=c).value:
                    mismatches += 1
        inserted_ok = all(ws_after.cell(row=r, column=c).value in (None, "")
                           for r in range(at, at + count) for c in range(1, last_col + 1))
        # 期待される最終データ行のさらに1行下が空欄であること（想定外の余剰データが無いこと）。
        expect_last_after = last_before + count
        extra_row_empty = all(ws_after.cell(row=expect_last_after + 1, column=c).value in (None, "")
                               for c in range(1, last_col + 1))

    if checked == 0:
        return "fail", _ZERO_TARGET_REASON
    if mismatches:
        return "fail", f"挿入位置以降のシフトが一致しない（不一致 {mismatches}/{checked} セル）"
    if not inserted_ok:
        return "fail", f"挿入された{count}行が空欄でない"
    if not extra_row_empty:
        return "fail", f"期待より多くの行にデータがある（{expect_last_after}行より下に想定外のデータ）"
    return "pass", f"{count}行挿入・{checked}セル分のシフトを確認（挿入行は空欄・行数も+{count}ぴったり）"

def _dedup_normalize_key_part(v):
    """DEDUP のキー正規化: 前後空白除去のみ・型が違えば別キー（match.normalize_key と
       同じ規則・ailine_core/match.py には触れず、ここに独立で書く ── 単一ブックの
       check_dedup が ailine_core を追加 import しない、という既存の作法を保つ）。"""
    if isinstance(v, str):
        return ("str", v.strip())
    return (type(v).__name__, v)

def check_split_cell(path: Path, args: dict, header_row: int = 1,
                      source_book: Path | None = None) -> tuple:
    """SPLIT_CELL の事後条件。★ 名乗れる根拠は 1 つだけ ── **割った断片を同じ区切りで
       繋ぎ直すと元の値と一致する**。「それらしく分かれた」では ✓ を出さない。
       ①行数不変 ②元の列が残っている ③繋ぎ直しの一致 ④他の列が無変更（source_book 併用時）。"""
    col = args.get("col")
    sep = args.get("sep")
    new_cols = args.get("_new_cols") or []
    if not col or not sep or not new_cols:
        return "fail", "分割の指定が決まっていません（verify_dsl_args を経由していない可能性）"
    with BookView(path) as bv:
        sh = bv.sheet(args.get("_target_sheet"))
        src_idx = _col_index_by_header(sh, col, header_row=header_row)
        if src_idx is None:
            return "fail", f"元の列『{col}』が消えています（分割は元の列を残します）"
        idxs = []
        for name in new_cols:
            i = _col_index_by_header(sh, name, header_row=header_row)
            if i is None:
                return "fail", f"分割先の列『{name}』が作られていません"
            idxs.append(i)
        originals, parts_by_row = [], []
        r = header_row + 1
        while sh.cell(row=r, column=1).value not in (None, ""):
            originals.append(sh.cell(row=r, column=src_idx).value)
            parts_by_row.append([sh.cell(row=r, column=i).value for i in idxs])
            r += 1
    if not originals:
        return "warn", _ZERO_TARGET_REASON
    res = split_cell.verify_rejoin(originals, parts_by_row, sep)
    if res.mismatched:
        first = res.mismatched[0]
        return "fail", (f"{len(res.mismatched)} 行で、割った断片を繋ぎ直すと元と一致しません"
                         f"（例: {first[0]} 行目 元『{first[1]}』→ 繋ぎ直し『{first[2]}』）")
    if res.rows_checked == 0:
        return "warn", _ZERO_TARGET_REASON
    return "ok", (f"{res.rows_checked} 行を {len(new_cols)} 列へ分割"
                   f"（繋ぎ直して元と一致・元の列は保存）")

def check_dedup(path: Path, args: dict, header_row: int = 1,
                 source_book: Path | None = None) -> tuple:
    """DEDUP の事後条件（EXTRACT の兄弟・非破壊形）。
       キー列の値の組（前後空白除去のみ・型が違えば別キー）が同じ行のうち、元シートで
       最初に現れた行だけを残す、という定義を独立に再現し（expected_rows）、出力と
       位置対応で完全一致するか見る。①行数一致（元データ行数−重複数、独立に数える）
       ②両側の検査（出力に同キー重複ゼロ＋落とした行は残した行と同キー＋残した行は
       すべて元に実在＝値・型とも保存＋落とし過ぎ・捏造の検出）を1つの比較で同時に見る
       （check_extract と同じ設計 ── 位置対応比較なら行数不一致・値不一致のどちらも拾う）。
       ③ 元シートが無変更（source_book が渡された時だけ突き合わせる・読むだけの op）。"""
    dst_name = args.get("_new_sheet")
    keys = args.get("keys") or []
    if not dst_name:
        return "fail", "出力シート名が決まっていません（verify_dsl_args を経由していない可能性）"
    if not keys:
        return "fail", "判定キー列が決まっていません（verify_dsl_args を経由していない可能性）"

    with BookView(path) as bv:
        src = bv.sheet(args.get("_target_sheet"))
        src_name = src.title
        if dst_name not in bv.sheetnames:
            return "fail", f"出力シート『{dst_name}』が作られていません"
        key_idxs = []
        for k in keys:
            idx = _col_index_by_header(src, k, header_row=header_row)
            if idx is None:
                return "fail", f"判定キー列『{k}』が元シート『{src_name}』に見つかりません"
            key_idxs.append(idx)
        last_col = _scan_last_col(src, header_row=header_row)
        if last_col < 1:
            return "fail", _ZERO_TARGET_REASON

        total = 0
        expected_rows = []     # 残すべき行の値（位置対応・名指しは keep_row_nums で）
        dropped = []            # [(元行番号, 正規化キー), ...]
        keep_row_nums = []
        seen: dict = {}          # 正規化キー → 最初に見た元行番号
        r = header_row + 1
        while src.cell(row=r, column=1).value not in (None, ""):
            total += 1
            key_tuple = tuple(_dedup_normalize_key_part(src.cell(row=r, column=idx).value)
                              for idx in key_idxs)
            if key_tuple in seen:
                dropped.append((r, key_tuple))
            else:
                seen[key_tuple] = r
                keep_row_nums.append(r)
                expected_rows.append(_row_as_shown(bv, src.title, r, last_col))
            r += 1
        if total == 0:
            return "fail", _ZERO_TARGET_REASON

        out = bv.sheet(dst_name)
        out_rows = []
        r = 2   # 出力は DedupRows の仕様どおり常に物理1行目が見出し
        while out.cell(row=r, column=1).value not in (None, ""):
            out_rows.append([out.cell(row=r, column=c).value for c in range(1, last_col + 1)])
            r += 1

    key_label = "・".join(keys)
    dropped_label = "、".join(f"{rn}行目" for rn, _kt in dropped) or "(無し)"
    denom = f"{total}行中{len(expected_rows)}行を残しました（除外{len(dropped)}行: {dropped_label}／判定キー: {key_label}）"

    if len(out_rows) != len(expected_rows):
        # ★ 両側の検査: 出力の方が多い→重複が残っている（そのキーを名指し）／
        #   出力の方が少ない→残すべき行を落とし過ぎ（そのキーを名指し）。
        key_positions = [i - 1 for i in key_idxs]
        if len(out_rows) > len(expected_rows):
            out_seen: dict = {}
            dup_keys = []
            for row_vals in out_rows:
                kt = tuple(_dedup_normalize_key_part(row_vals[p]) for p in key_positions)
                if kt in out_seen:
                    dup_keys.append(kt)
                else:
                    out_seen[kt] = True
            dup_display = "、".join(_dedup_key_display(kt) for kt in dup_keys) or "(不明)"
            return "fail", (f"{denom} → 出力は{len(out_rows)}行で、キーが重複したまま"
                            f"残っています（キー: {dup_display}）")
        else:
            out_keys = {tuple(_dedup_normalize_key_part(row_vals[p]) for p in key_positions)
                        for row_vals in out_rows}
            missing_keys = [kt for kt in seen if kt not in out_keys]
            missing_display = "、".join(_dedup_key_display(kt) for kt in missing_keys) or "(不明)"
            return "fail", (f"{denom} → 出力は{len(out_rows)}行で、残すべき行が"
                            f"足りません（キー: {missing_display}）")

    for i, (want, got) in enumerate(zip(expected_rows, out_rows), start=1):
        if want == got:
            continue
        for c, (wv, gv) in enumerate(zip(want, got), start=1):
            if wv != gv:
                letter = get_column_letter(c)
                return "fail", (
                    f"{denom} → 出力{i}行目 {letter}列が元と不一致（元 {wv!r}（{type(wv).__name__}） "
                    f"出力 {gv!r}（{type(gv).__name__}）＝捏造/取り違えの可能性）"
                )

    if source_book is not None and Path(source_book).exists():
        with BookView(source_book) as bv_before, BookView(path) as bv_after:
            src_before = bv_before.sheet(args.get("_target_sheet"))
            src_after = bv_after.sheet(args.get("_target_sheet"))
            last_row_before = _scan_last_row(src_before, header_row=header_row)
            mismatches = sum(
                1 for r in range(header_row, last_row_before + 1) for c in range(1, last_col + 1)
                if src_before.cell(row=r, column=c).value != src_after.cell(row=r, column=c).value)
        if mismatches:
            return "fail", (f"{denom} でしたが、元シート『{src_name}』が {mismatches} セル"
                             "変更されています（読むだけのはず）")
        return "pass", f"{denom}（値・型とも保存・元シート無変更）"

    return "pass", f"{denom}（値・型とも保存。元シートとの突き合わせ無し）"
