"""match — M3 `ailine run <A.xlsx> <B.xlsx> "<依頼>"`（2冊の照合）の本体。
   DESIGN-20260821-multifile.md M3 設計 v2（凍結・2026-08-21 14:5x）。

   芯（買い手の言葉・凍結）: 「候補を並べて差額だけ機械で保証、決めるのは人。
   消し込みまで自動と言われた瞬間、確認に40分戻るので価値0」── この module は判断しない。
   キーごとに両側を数えて差額を計算し、並べて渡すだけ。

   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定（detect_header_row/_row_char_stats）は ailine.py 側（cmd_run_match）が既存のものを
   1回だけ呼び、その結果を値としてこの module へ渡す（stack.py/extract_multi.py と同じ配線）。
   ★ 単位L との違い（設計の明記）: total_row.py の構造トリガ（ラベル空白・直上空行）は
   キー列に使わない ── キー空白行は黙って除外せず『キー不明』の第5区分として必ず表に出す
   （「合計商事」のようなラベル語を含む正当なキーを消さないため）。
   ★ 列署名（MATCH_SHEET_NAME/MATCH_HEADERS）は stack.py（P 先行 commit）が単一の出所 ──
   ここでは import して使い、値を重複定義しない。
"""
from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field

import openpyxl

from ailine_core import inspection, stack as stack_kind

# ★ P 先行 commit（stack.py）が既に凍結した契約 ── 単一の出所から import する
#   （match.py 側で別の値を持つと、署名の判定と実際の出力が食い違う「片配線」になる）。
MATCH_SHEET_NAME = stack_kind.MATCH_SHEET_NAME
MATCH_HEADERS = list(stack_kind.MATCH_HEADERS)
CREATOR_MARK = "ailine match"

DETAIL_SHEET_NAME = "明細"
UNKNOWN_KEY_LABEL = "キー不明"
TOLERANCE = 1e-6


def read_data_rows(ws, header_row: int, headers: list) -> list:
    """header_row より下の全データ行を (行番号, [値...], [number_format...]) で読む
       （headers の列数ぶんだけ）。全列空白の行はスキップする。
       ★ 単位L の構造トリガはここでは一切使わない ── 『何も判断しない読み』だけを提供する
       （合計行かどうかの判断はキー空白の有無だけに委ねる。M3 設計 v2「単位L」節）。
       ★ M2.5 の再発直し（実弾検分）: number_format も運ぶ ── 明細シートの日付セルが
       時刻の尻尾つき（00:00:00）で出ないように（stack.py の実視の磨きと同じ線）。"""
    num_cols = len(headers)
    max_row = ws.max_row or header_row
    rows = []
    for r in range(header_row + 1, max_row + 1):
        cells = [ws.cell(row=r, column=c) for c in range(1, num_cols + 1)]
        values = [c.value for c in cells]
        if any(v not in (None, "") for v in values):
            formats = [c.number_format for c in cells]
            rows.append((r, values, formats))
    return rows


def numeric_columns(headers: list, rows: list) -> set:
    """headers のうち、データ行のどこかに値があり、かつ全ての値が数値(bool除く)である
       列名の集合。★ 値が1つも無い列は数値と認めない（金額列に既定を持たせない、の徹底）。"""
    out = set()
    for i, name in enumerate(headers):
        vals = [values[i] for _r, values, _f in rows if i < len(values)]
        vals = [v for v in vals if v not in (None, "")]
        if vals and all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in vals):
            out.add(name)
    return out


def _is_swallowed_fragment(h: str, task: str, vocabulary: list) -> bool:
    """★ review5#4 の直し（単位B の轍・W3 断片ガードと同じ型）: h の task 中の出現が、
       すべて『h を含む・h より長い、vocabulary 内の別の語』の出現に完全に包含されるか。
       1つでも呑まれない出現があれば False（＝ h 自身が独立した名指しとして生きている）。
       例:『税込金額』を名指ししただけで、部分文字列の『金額』（お預り金額/税込金額の中に
       埋もれているだけ）まで名指し扱いにしない。vocabulary は両冊の全ヘッダー（呑み込む側の
       語が相手側のヘッダーであってもよい ── 呼び出し側が resolve_columns で両側をまとめて渡す）。"""
    positions = [m.start() for m in re.finditer(re.escape(h), task)]
    if not positions:
        return False
    longer = [lh for lh in vocabulary if lh and lh != h and len(lh) > len(h)
              and h in lh and lh in task]
    if not longer:
        return False
    longer_spans = [(m.start(), m.start() + len(lh)) for lh in longer
                     for m in re.finditer(re.escape(lh), task)]
    for p in positions:
        end = p + len(h)
        if not any(span_start <= p and end <= span_end for span_start, span_end in longer_spans):
            return False
    return True


def resolve_role(task: str, headers: list, numeric: set, role: str,
                  vocabulary: list | None = None) -> tuple:
    """機械3段（LLM ゼロ）: ①依頼文に名指しされ、かつ役割に合う型の列がちょうど1本なら採用
       ②それが決まらなければ、役割に合う型の列（依頼文の名指しは問わない）がちょうど1本なら採用
       ③それでも決まらなければ (None, 候補列) を返す（呼び出し側が exit 3 で列挙する）。
       role="key" は非数値列、role="amount" は数値列。★ 金額列に既定を持たせない ──
       候補が2本以上でも1本も無くても None のまま返す（ここで折衷しない）。
       ★ review5#4: ①の名指し判定は部分文字列包含だが、より長い別ヘッダーに完全に呑まれる
       出現は名指しと数えない（_is_swallowed_fragment）。vocabulary 省略時は headers 自身
       （後方互換・単独呼び出しの検体向け）。"""
    def type_ok(h):
        return (h not in numeric) if role == "key" else (h in numeric)

    vocab = vocabulary if vocabulary is not None else headers
    named = [h for h in headers if h and h in task and not _is_swallowed_fragment(h, task, vocab)]
    named_typed = [h for h in named if type_ok(h)]
    if len(named_typed) == 1:
        return named_typed[0], []
    typed = [h for h in headers if type_ok(h)]
    if len(typed) == 1:
        return typed[0], []
    candidates = named_typed if len(named_typed) > 1 else typed
    return None, candidates


@dataclass(frozen=True)
class ColumnResolution:
    """4役割（A側キー/B側キー/A側金額/B側金額）の解決結果。
       ok=False の時は unresolved に (side, role, candidates) のリストが載る
       （★ 一括検出: 最初の1件で止めず、決まらなかった役割を全部集めて返す）。"""
    ok: bool
    key_a: str | None = None
    key_b: str | None = None
    amount_a: str | None = None
    amount_b: str | None = None
    unresolved: list = field(default_factory=list)


def resolve_columns(task: str, headers_a: list, rows_a: list,
                     headers_b: list, rows_b: list) -> ColumnResolution:
    """両冊・両役割ぶん resolve_role をまとめて呼ぶ入口。
       ★ review5#4: 断片ガードの vocabulary は両冊のヘッダー全部の和（順序保持・重複排除）──
       飲み込む側の長い語が相手側の冊のヘッダーであるケース（お預り金額 が 金額 を呑む、等）
       まで正しく拾うため。"""
    vocabulary = list(dict.fromkeys(list(headers_a) + list(headers_b)))
    num_a = numeric_columns(headers_a, rows_a)
    num_b = numeric_columns(headers_b, rows_b)
    key_a, cand_ka = resolve_role(task, headers_a, num_a, "key", vocabulary)
    key_b, cand_kb = resolve_role(task, headers_b, num_b, "key", vocabulary)
    amount_a, cand_aa = resolve_role(task, headers_a, num_a, "amount", vocabulary)
    amount_b, cand_ab = resolve_role(task, headers_b, num_b, "amount", vocabulary)
    unresolved = []
    if key_a is None:
        unresolved.append(("A", "key", cand_ka))
    if key_b is None:
        unresolved.append(("B", "key", cand_kb))
    if amount_a is None:
        unresolved.append(("A", "amount", cand_aa))
    if amount_b is None:
        unresolved.append(("B", "amount", cand_ab))
    return ColumnResolution(ok=not unresolved, key_a=key_a, key_b=key_b,
                             amount_a=amount_a, amount_b=amount_b, unresolved=unresolved)


def normalize_key(v):
    """キーの正規化 = 前後空白除去のみ（Q15: 全角半角・法人格の同一視はしない ── 意味的汚染を
       避ける）。型が違えば別キー（数値の123と文字列の"123"は別扱い ── 黙って型変換しない）。
       None・空文字（空白のみを含む）は None（＝キー不明のバケツへ）。"""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return None if s == "" else ("str", s)
    return (type(v).__name__, v)


def display_key(v):
    """キーの表示値（前後空白は落とすが、それ以外は元の型のまま）。"""
    return v.strip() if isinstance(v, str) else v


def _numeric(v) -> float:
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else 0.0


def _state(a_count: int, b_count: int, diff: float) -> str:
    """状態語は算術のみ（憲法2・design v2「①照合」節）: 『一致』は使わない
       （合算・相殺で嘘になる ── 差額 0 でも件数が違う戊社の形が実例）。"""
    if a_count == 0 and b_count > 0:
        return "B のみ"
    if b_count == 0 and a_count > 0:
        return "A のみ"
    if abs(diff) <= TOLERANCE:
        return "差額 0"
    if diff > 0:
        return f"+{inspection.fmt_num(diff)}"
    return f"−{inspection.fmt_num(abs(diff))}"


def side_pair(g) -> str:
    """『A 186300 / B 0』の 0 が **金額 0** なのか **1 行も無い** のか読めなかった
    （2026-08-24 第三波 S5）。片側が 0 行なら『なし』と書いて区別する
    ── 出ないことは信号でないので、出ていないと書く。★ 実装は 1 つ（2 箇所が呼ぶ）。"""
    def one(label, count, total):
        if count == 0:
            return f"{label} なし（0 行）"
        return f"{label} {inspection.fmt_num(total)}"
    return f"{one('A', g.a_count, g.a_sum)} / {one('B', g.b_count, g.b_sum)}"


def _clean_num(v):
    """整数値は int で書く（650.0 でなく 650）。"""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _new_group(display) -> dict:
    return {"display": display, "a_count": 0, "a_sum": 0.0, "a_rows": [],
            "b_count": 0, "b_sum": 0.0, "b_rows": []}


@dataclass(frozen=True)
class KeyGroup:
    """1キー分（または UNKNOWN_KEY_LABEL の集約行1本）の勘定。"""
    key_display: object
    a_count: int
    a_sum: float
    b_count: int
    b_sum: float
    diff: float
    state: str
    a_rows: list    # [(元行番号, [その側の全列の値], [number_format...]), ...]
    b_rows: list

    @property
    def key(self):
        """key_display の別名。呼び出し側の素朴な `g.key` という想定に応える（読み取り専用・
           frozen dataclass なので通常代入では書き換わらない）。正式なフィールドは
           key_display（キー不明の集約行では UNKNOWN_KEY_LABEL になる）。"""
        return self.key_display


def compute_match(headers_a: list, rows_a: list, key_col_a: str, amount_col_a: str,
                   headers_b: list, rows_b: list, key_col_b: str, amount_col_b: str) -> list:
    """キー勘定の本体。戻り値はソート済み（|差額| 降順 → キー Unicode 昇順 の設計凍結タイブレーク）。
       ★ キー空白行は UNKNOWN_KEY_LABEL の1行に集約（両側とも黙って落とさない・第5区分）。"""
    ka_idx = headers_a.index(key_col_a)
    aa_idx = headers_a.index(amount_col_a)
    kb_idx = headers_b.index(key_col_b)
    ab_idx = headers_b.index(amount_col_b)

    groups: dict = {}
    unknown_a, unknown_b = [], []

    for row_num, values, formats in rows_a:
        raw_key = values[ka_idx] if ka_idx < len(values) else None
        amt = _numeric(values[aa_idx] if aa_idx < len(values) else None)
        nk = normalize_key(raw_key)
        if nk is None:
            unknown_a.append((row_num, values, formats, amt))
            continue
        g = groups.setdefault(nk, _new_group(display_key(raw_key)))
        g["a_count"] += 1
        g["a_sum"] += amt
        g["a_rows"].append((row_num, values, formats))

    for row_num, values, formats in rows_b:
        raw_key = values[kb_idx] if kb_idx < len(values) else None
        amt = _numeric(values[ab_idx] if ab_idx < len(values) else None)
        nk = normalize_key(raw_key)
        if nk is None:
            unknown_b.append((row_num, values, formats, amt))
            continue
        g = groups.setdefault(nk, _new_group(display_key(raw_key)))
        g["b_count"] += 1
        g["b_sum"] += amt
        g["b_rows"].append((row_num, values, formats))

    out = []
    for g in groups.values():
        diff = g["a_sum"] - g["b_sum"]
        out.append(KeyGroup(key_display=g["display"], a_count=g["a_count"], a_sum=g["a_sum"],
                             b_count=g["b_count"], b_sum=g["b_sum"], diff=diff,
                             state=_state(g["a_count"], g["b_count"], diff),
                             a_rows=g["a_rows"], b_rows=g["b_rows"]))

    if unknown_a or unknown_b:
        a_sum = sum(amt for _r, _v, _f, amt in unknown_a)
        b_sum = sum(amt for _r, _v, _f, amt in unknown_b)
        diff = a_sum - b_sum
        out.append(KeyGroup(key_display=UNKNOWN_KEY_LABEL, a_count=len(unknown_a), a_sum=a_sum,
                             b_count=len(unknown_b), b_sum=b_sum, diff=diff, state=UNKNOWN_KEY_LABEL,
                             a_rows=[(r, v, f) for r, v, f, _a in unknown_a],
                             b_rows=[(r, v, f) for r, v, f, _a in unknown_b]))

    out.sort(key=lambda g: (-abs(g.diff), str(g.key_display)))
    return out


@dataclass(frozen=True)
class TotalRowNote:
    """『キーが空 かつ 金額 = 同じ冊の他のデータ行の和』の注記（除外はしない・算術のみ）。
       design v2「単位L」節の合計行対策 ── 実弾検分（2026-08-21）で発覚: 銀行明細の合計行が
       キー不明の袋に無警告で混ざり、本物の名義不明の金額を沈めていた（実害確認済み）。"""
    side: str          # "A" / "B"
    row_num: int
    amount: float
    other_sum: float


def possible_total_row_notes(headers: list, rows: list, key_col: str, amount_col: str,
                              side: str) -> list:
    """キーが空白のデータ行のうち、金額が『同じ冊の他の全データ行の金額の和』と一致する
       ものを注記として返す（除外はしない ── 呼び出し側が両側の数字つきで開示するだけ）。
       ★ 判定は算術のみ（許容誤差 TOLERANCE）。ラベル語（『合計』等）は一切見ない ──
       単位L の構造/ラベルトリガと同じ轍を踏まない（design v2 の明記どおり）。"""
    key_idx = headers.index(key_col)
    amount_idx = headers.index(amount_col)
    total = sum(_numeric(values[amount_idx] if amount_idx < len(values) else None)
                for _r, values, _f in rows)
    notes = []
    for row_num, values, _f in rows:
        raw_key = values[key_idx] if key_idx < len(values) else None
        if normalize_key(raw_key) is not None:
            continue
        amt = _numeric(values[amount_idx] if amount_idx < len(values) else None)
        if amt == 0:
            continue
        other_sum = total - amt
        if abs(amt - other_sum) <= TOLERANCE:
            notes.append(TotalRowNote(side=side, row_num=row_num, amount=amt, other_sum=other_sum))
    return notes


def independent_key_stats(grid: dict, row_numbers: list, headers: list,
                           key_col: str, amount_col: str) -> dict:
    """★ review5#1 critical の直し・M3 verify（_verify_match）と共有する単一の実装
       （片配線の自己点検: 別実装を2つ作らない）。原本から独立に（xml_readback の grid ──
       openpyxl を経由しない別実装）キーごとの件数+金額和を再集計する。compute_match と
       同じキー正規化規則（normalize_key: 前後空白除去のみ・型が違えば別キー）を使う ──
       ここがズレると恒真検査になり偽陽性になる（呼び出し側からの入力は
       xml_readback.read_grid の grid + データ行番号であること）。
       戻り値: {normalize_key の戻り値（None はキー不明）: {"count": int, "sum": float}}。"""
    key_idx = headers.index(key_col) + 1     # grid のキーは1起点の (行, 列)
    amount_idx = headers.index(amount_col) + 1
    stats: dict = {}
    for r in row_numbers:
        raw_key = grid.get((r, key_idx))
        amt = grid.get((r, amount_idx))
        amt = float(amt) if isinstance(amt, (int, float)) and not isinstance(amt, bool) else 0.0
        nk = normalize_key(raw_key)
        entry = stats.setdefault(nk, {"count": 0, "sum": 0.0})
        entry["count"] += 1
        entry["sum"] += amt
    return stats


def independent_key_sums(grid: dict, row_numbers: list, headers: list,
                          key_col: str, amount_col: str) -> dict:
    """independent_key_stats の金額だけを取り出す薄いラッパー（cmd_run_match の
       書き込み時事後条件が使う ── 本体のロジックは independent_key_stats に一本化）。
       戻り値: {normalize_key の戻り値（None はキー不明）: 金額和}。"""
    stats = independent_key_stats(grid, row_numbers, headers, key_col, amount_col)
    return {nk: v["sum"] for nk, v in stats.items()}


def side_totals(groups: list) -> tuple:
    """(A側件数合計, B側件数合計) ── キー不明を含む全キー行の Σ件数（完全会計①の右辺）。"""
    return sum(g.a_count for g in groups), sum(g.b_count for g in groups)


def diff_distribution(groups: list) -> list:
    """差額の値ごとの件数（キー不明を除く）。多い順（Counter.most_common）。検分シート用。"""
    counter = Counter(_clean_num(g.diff) for g in groups if g.key_display != UNKNOWN_KEY_LABEL)
    return counter.most_common()


def fuzzy_candidates(groups: list) -> list:
    """A のみ/B のみ のキーで、空白除去後に部分文字列一致するものを並べるだけ
       （並べ替え・自動採用はしない ── 表記ゆれの『候補』として検分シートに載せるだけ）。"""
    only_a = [g.key_display for g in groups if g.a_count > 0 and g.b_count == 0
              and isinstance(g.key_display, str)]
    only_b = [g.key_display for g in groups if g.b_count > 0 and g.a_count == 0
              and isinstance(g.key_display, str)]
    out = []
    for x in only_a:
        xs = re.sub(r"\s", "", x)
        for y in only_b:
            ys = re.sub(r"\s", "", y)
            if xs and ys and xs != ys and (xs in ys or ys in xs):
                out.append(f"『{x}』(A のみ) と 『{y}』(B のみ) が部分一致")
    return out


def build_workbook(groups: list, headers_a: list, headers_b: list,
                    book_a_name: str, book_b_name: str) -> tuple:
    """出力ブックの1・2枚目（照合・明細）を組み立てる（検分シートは呼び出し側が
       inspection.build_sheet で追加する ── stack.py/extract_multi.py と同じ役割分担）。
       ★ 照合シートの見出しは MATCH_HEADERS と一致させる ── stack.py の署名契約そのもの。
       戻り値: (wb, key_to_detail_row) ── 後者は str(key_display) → 明細シートでその
       キーの最初の行番号（★ gap#1: 検分の所見からブック内リンクで飛ぶ着地点に使う）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = MATCH_SHEET_NAME
    ws.append(list(MATCH_HEADERS))
    for g in groups:
        ws.append([g.key_display, g.a_count, _clean_num(g.a_sum), g.b_count,
                   _clean_num(g.b_sum), _clean_num(g.diff), g.state])
    inspection.bold_row(ws, 1, len(MATCH_HEADERS))
    inspection.autosize_columns(ws)

    ws2 = wb.create_sheet(title=DETAIL_SHEET_NAME)
    detail_headers = ["冊", "元行", "キー"] + list(headers_a) + list(headers_b)
    ws2.append(detail_headers)
    inspection.bold_row(ws2, 1, len(detail_headers))

    entries = []
    for g in groups:
        tinted = abs(g.diff) > TOLERANCE
        for row_num, values, formats in g.a_rows:
            row = [book_a_name, row_num, g.key_display] + list(values) + [None] * len(headers_b)
            fmts = [None, None, None] + list(formats) + [None] * len(headers_b)
            entries.append((str(g.key_display), book_a_name, row_num, row, fmts, tinted))
        for row_num, values, formats in g.b_rows:
            row = [book_b_name, row_num, g.key_display] + [None] * len(headers_a) + list(values)
            fmts = [None, None, None] + [None] * len(headers_a) + list(formats)
            entries.append((str(g.key_display), book_b_name, row_num, row, fmts, tinted))
    entries.sort(key=lambda e: (e[0], e[1], e[2]))

    r = 2
    key_to_detail_row: dict = {}
    for sort_key, _book, _row_num, row, fmts, tinted in entries:
        key_to_detail_row.setdefault(sort_key, r)
        for c, v in enumerate(row, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            # ★ M2.5 の再発直し（実弾検分）: 元セルの number_format を運ぶ
            #   （日付が『2026-07-31 0:00:00』と時刻付きで出ないように・stack.py と同じ線）。
            fmt = fmts[c - 1]
            if fmt:
                cell.number_format = fmt
        if tinted:
            for c in range(1, len(detail_headers) + 1):
                ws2.cell(row=r, column=c).fill = inspection.TINT_FILL
        r += 1
    inspection.autosize_columns(ws2)
    return wb, key_to_detail_row


# ★ gap#1（design v2「誘導の憲法」の適用拡張）: 検分の所見表に載る M3 固有の種類。
KIND_DIFF = "差額あり"
KIND_UNKNOWN_KEY = "キー不明"
KIND_POSSIBLE_TOTAL_ROW = "合計行の可能性"


def build_findings(groups: list, key_to_detail_row: dict, total_notes: list,
                    book_a_name: str, book_b_name: str) -> list:
    """検分の所見表に載せる Finding のリスト（差額あり・A/Bのみ・キー不明・合計行の可能性）。
       ★ gap#1: リンクはブック内（明細シートの該当キーの先頭行）── 2冊が別フォルダでも
       切れない（外部相対パスに依存しない）。次の一手（next_step）は端末報告にもそのまま
       使う（呼び出し側が同じ文言を2箇所に出す・二重管理を避ける）。"""
    findings = []
    for g in groups:
        detail_row = key_to_detail_row.get(str(g.key_display))
        link = ((None, inspection.hyperlink_location(DETAIL_SHEET_NAME,
                                                       inspection.cell_ref(1, detail_row)))
                if detail_row is not None else None)
        cell_ref = inspection.cell_ref(1, detail_row) if detail_row is not None else ""
        if g.key_display == UNKNOWN_KEY_LABEL:
            findings.append(inspection.finding(
                kind=KIND_UNKNOWN_KEY, file=f"{book_a_name} / {book_b_name}",
                sheet=DETAIL_SHEET_NAME, cell=cell_ref,
                source_value=_clean_num(g.a_sum), output_value=_clean_num(g.b_sum),
                next_step=f"キー不明 {g.a_count + g.b_count}行（{side_pair(g)}）。"
                          "名義不明の入金・請求が無いか、明細シートで確認してください。",
                link=link))
        elif abs(g.diff) > TOLERANCE:
            findings.append(inspection.finding(
                kind=KIND_DIFF, file=f"{book_a_name} / {book_b_name}",
                sheet=DETAIL_SHEET_NAME, cell=cell_ref,
                source_value=_clean_num(g.a_sum), output_value=_clean_num(g.b_sum),
                next_step=f"{g.key_display}: {g.state}（{side_pair(g)}）。"
                          "明細シートで内訳を確認してください。",
                link=link))
    unknown_detail_row = key_to_detail_row.get(UNKNOWN_KEY_LABEL)
    unknown_link = ((None, inspection.hyperlink_location(
                        DETAIL_SHEET_NAME, inspection.cell_ref(1, unknown_detail_row)))
                    if unknown_detail_row is not None else None)
    unknown_cell_ref = (inspection.cell_ref(1, unknown_detail_row)
                        if unknown_detail_row is not None else "")
    for note in total_notes:
        book_name = book_a_name if note.side == "A" else book_b_name
        findings.append(inspection.finding(
            kind=KIND_POSSIBLE_TOTAL_ROW, file=book_name, sheet=DETAIL_SHEET_NAME,
            cell=unknown_cell_ref,
            source_value=_clean_num(note.amount), output_value=_clean_num(note.other_sum),
            next_step=f"{book_name} の{note.row_num}行目（金額 {inspection.fmt_num(note.amount)}）は"
                      f"他のデータ行の和（{inspection.fmt_num(note.other_sum)}）と一致します ── "
                      "合計行かもしれません（キーが空欄なので除外はしていません）。",
            link=unknown_link))
    return findings
