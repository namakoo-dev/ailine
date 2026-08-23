"""BookView — openpyxl の式ビュー/値ビューの二重性を1箇所に閉じ込める。

★ C4（再設計 分割の一歩目）: ailine.py の事後条件チェッカー(check_*)は、openpyxl で
開いたブックを「式ビュー」（data_only 未指定・セルに書かれた式文字列そのもの）と
「値ビュー」（data_only=True・LibreOffice/Excel が最後に保存時計算したキャッシュ値）の
二重に読み分ける必要がある（docs/behavior-corpus/nodes/formula-readback-duality.md
参照）。実測: ailine.py の load_workbook 呼び出し約29箇所のうち data_only=True は
4箇所だけという非対称があり、これが直近の5件の欠陥（事後条件が operand を式ビューから
読む）の共通の根だった。個別には直したが、次に増えるチェッカーが同じ間違いをできる
状態が残っていたため、二重性そのものを BookView 1箇所に閉じ込める。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl


class BookView:
    """openpyxl の式ビュー/値ビューの二重性を1箇所に閉じ込める。
       cell_value(r,c) は常に『計算後の値』を返す。cell_formula(r,c) が式を返す。
       『値が無い』と『式はあるがキャッシュ値が無い』を呼び側が区別できること。

       使い方:
           with BookView(path) as bv:
               ws = bv.sheet()                       # 構造走査/スタイル読み用の生 worksheet
               raw = ws.cell(row=r, column=c).value   # 式ビュー（そのまま）
               val = bv.cell_value(r, c)              # 値ビュー（data_only の計算後の値）
               formula = bv.cell_formula(r, c)        # 式なら式文字列・式でなければ None

       値ビュー(data_only=True)の Workbook は cell_value/cell_formula を実際に呼ぶまで
       開かない。use_formula=False で値ビューを一度も読まないチェッカーは、従来どおり
       load_workbook を1回しか呼ばない（挙動もコストも変えない・純リファクタ）。"""

    def __init__(self, path: Path, read_only: bool = False):
        self.path = path
        self._read_only = read_only
        self._wb = None
        self._wb_value = None

    def __enter__(self) -> "BookView":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def close(self) -> None:
        if self._wb is not None:
            self._wb.close()
            self._wb = None
        if self._wb_value is not None:
            self._wb_value.close()
            self._wb_value = None

    @property
    def wb(self):
        """式ビュー（openpyxl の既定・data_only 未指定）の生 Workbook。"""
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.path, read_only=self._read_only)
        return self._wb

    @property
    def sheetnames(self) -> list:
        return self.wb.sheetnames

    def sheet(self, name: str | None = None):
        """式ビューの worksheet（構造走査・スタイル読みはここから）。name 省略時は1枚目。"""
        wb = self.wb
        return wb[name] if name else wb[wb.sheetnames[0]]

    def _value_wb(self):
        if self._wb_value is None:
            self._wb_value = openpyxl.load_workbook(
                self.path, data_only=True, read_only=self._read_only)
        return self._wb_value

    def _value_sheet(self, name: str | None = None):
        wb = self._value_wb()
        return wb[name] if name else wb[wb.sheetnames[0]]

    def cell_value(self, row: int, col: int, sheet: str | None = None):
        """『計算後の値』。式でないセルは式ビューと同じ値、式セルは LibreOffice/Excel が
           保存時に計算したキャッシュ値（未計算なら None）。"""
        return self._value_sheet(sheet).cell(row=row, column=col).value

    def cell_formula(self, row: int, col: int, sheet: str | None = None):
        """式なら式文字列（例 "=A2*B2"）、式でなければ None。"""
        v = self.sheet(sheet).cell(row=row, column=col).value
        return v if isinstance(v, str) and v.startswith("=") else None
