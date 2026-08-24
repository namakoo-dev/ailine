"""report_per_row — 帳票段 REPORT_PER_ROW の純ロジック部品。
   DESIGN-20260823-report-per-row.md（spike の実測で5箇所訂正済み・全文はプラン参照）。

   ★ 憲法の適用: 雛形は人が作る。機械は埋めるだけ ── 機械が触ってよいのは
   雛形の中の印（{{列名}}）が置かれたセルだけ。

   ★ spike の実測による契約（守らないと静かに壊れる。全文は設計文書参照）:
     A. シート名は Excel の規則で機械が自分で切る（禁止文字 \\ / ? * [ ] :・31文字上限）。
        31文字超は LO が無警告で切り詰める（例外も返り値も無い）── 機械が自分で数えて切る。
        ★ 訂正1: extract_multi.sanitize_filename（Windows のファイル名規則）とは別物・流用不可。
     B. copyByName は失敗時に孤児シートを残す ── 呼ぶ前に一意な名前を解決する（再試行禁止）。
        ★ 訂正2: unique_sheet_name が「絶対に失敗しない名前」を呼ぶ前に決める。
     C. ②値の3計数（compare_report_cells）は csv_quarantine.compare_against_quarantine の
        直接転用ではない ── CSV 版は「宣言=シート全体」前提で、雛形由来の非印セルまで
        「余剰」に誤検出する。★ 訂正5: 印セルだけに絞った姉妹関数として新設する
        （_values_agree・xml_readback.read_grid は部品として再利用）。

   ★ ailine を import しない（tests/test_line_budget.py と同じ移植可能性の作法）。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

from ailine_core import xml_readback
from ailine_core.csv_quarantine import _values_agree

# --- A/B: シート名の作法（Excel の規則。Windows のファイル名規則とは別物） -----------

# Excel がシート名に禁じる文字。extract_multi の _FILENAME_FORBIDDEN_RE / ailine.py の
# _EXTRACT_SHEET_NAME_FORBIDDEN_RE と同じ文字集合（同じ理由で3箇所が独立に持つ ──
# ailine_core モジュールは ailine.py を import しない・extract_multi にも依存しない）。
_SHEET_NAME_FORBIDDEN_RE = re.compile(r"[:\\/?*\[\]]")
MAX_SHEET_NAME_LEN = 31   # ★ Excel の上限。LO は無警告で切り詰める（例外も返り値も無い）


def sanitize_sheet_name(raw: str) -> str:
    """Excel のシート名規則に直す（禁止文字→'_'・31文字で機械が自分で切る）。
       空文字列は '_'（Excel はシート名を空にできない）。"""
    s = _SHEET_NAME_FORBIDDEN_RE.sub("_", str(raw))
    if not s:
        s = "_"
    return s[:MAX_SHEET_NAME_LEN]


def unique_sheet_name(raw: str, used) -> str:
    """★ 訂正2: copyByName に失敗しうる名前を渡さない ── 呼ぶ**前**に一意名を解決する。
       sanitize 済みの名前が衝突すれば '_2'・'_3'... を足す（31文字上限は維持したまま
       末尾を削って足す・上限超過での失敗を作らない）。「失敗したら別名で再試行」は禁止
       （孤児シートが積み上がる・設計文書 訂正2）。"""
    used = set(used or ())
    base = sanitize_sheet_name(raw)
    if base not in used:
        return base
    n = 2
    while True:
        suffix = f"_{n}"
        candidate = base[: MAX_SHEET_NAME_LEN - len(suffix)] + suffix
        if candidate not in used:
            return candidate
        n += 1


# --- 印（プレースホルダ）の走査 -------------------------------------------------

_PLACEHOLDER_RE = re.compile(r"\{\{([^{}]+)\}\}")


@dataclass(frozen=True)
class Placeholder:
    """雛形シート上の印1件。
       cell: セル表記（例 "B3"）。row/col: 1起点。column_name: {{...}} の中身（データ列名）。
       whole: セル全体がちょうど "{{列名}}" か（True）／他の文字と同居する部分一致か（False）。
       raw: 印を含むセルの元の文字列（部分一致の復元に使う）。"""
    cell: str
    row: int
    col: int
    column_name: str
    whole: bool
    raw: str


def scan_placeholders(ws, max_row: int, max_col: int) -> list:
    """ws（openpyxl の worksheet）の 1..max_row × 1..max_col を走査し、文字列セルに
       含まれる {{...}} を**すべて**印として集める。★ 2026-08-24 まで「最初の 1 個だけ」
       だったため、2 つ目が生のまま顧客の紙に印字され、しかも事後条件の declared も
       同じ規則で作るので原理的に検出不能だった。
       見つかった順（上から・各行は左から）。"""
    out = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            # ★ 2026-08-24: 初版は search（1 セル 1 個・v1 の割り切り）だった。
            #   `{{取引先}} 様（担当: {{担当者}}）` の 2 つ目が生のまま顧客の紙に印字され、
            #   しかも事後条件の declared も同じ規則で作るので**原理的に検出不能**だった。
            #   全部拾う ── 拾えば埋められるし、埋められなければ列名照合で断れる。
            matches = list(_PLACEHOLDER_RE.finditer(v))
            if not matches:
                continue
            for m in matches:
                name = m.group(1)
                whole = v == "{{" + name + "}}"
                out.append(Placeholder(cell=f"{get_column_letter(c)}{r}", row=r, col=c,
                                        column_name=name, whole=whole, raw=v))
    return out


def cells_with_multiple_placeholders(placeholders) -> list:
    """同じセルに印が 2 つ以上ある箇所 [(セル参照, [列名, ...]), ...]。

    ★ なぜ埋めずに断るのか（2026-08-24）: 1 セルに 2 つあると、埋める側は 1 セルに
      2 回書くことになり**後の値が前を消す**。「それらしく埋まって片方が生で残る」より、
      **雛形を直してくださいと言う方が正しい**（雛形は人が作る、という帳票段の憲法）。
    """
    by_cell = {}
    for ph in placeholders:
        by_cell.setdefault(ph.cell, []).append(ph.column_name)
    return [(cell, names) for cell, names in by_cell.items() if len(names) > 1]


# --- ②値の3計数（印セルだけに絞る・訂正5） --------------------------------------

@dataclass(frozen=True)
class ReportCellsResult:
    """compare_report_cells の戻り値。missing: declared にあるが読み戻しに無いセル参照の列。
       mismatched: [(セル参照, declared値, actual値), ...]。surplus: 常に空
       （この関数は宣言された印セルだけを見る設計そのものが「余剰」を作らない ── 訂正5）。"""
    missing: list = field(default_factory=list)
    mismatched: list = field(default_factory=list)
    surplus: list = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not (self.missing or self.mismatched or self.surplus)


def _parse_cell_ref(ref: str) -> tuple:
    """"B3" → (行, 列) の1起点タプル（openpyxl の座標パーサをそのまま使う）。"""
    col_letters, row = coordinate_from_string(ref)
    return row, column_index_from_string(col_letters)


def compare_report_cells(path, sheet_name: str, declared: dict) -> ReportCellsResult:
    """declared（{セル参照: 期待値}）を宣言分母として、xml_readback で独立に読み戻した
       シートと突き合わせる。★ 訂正5: CSV 版(csv_quarantine.compare_against_quarantine)は
       「宣言=シート全体」前提で、雛形由来の非印セル（見出し・固定文）まで「余剰」に
       誤検出する。ここは**印セルだけ**を見るので surplus は構造的に常に空になる。
       _values_agree（型込み等値）と xml_readback.read_grid は部品として再利用する。"""
    data = xml_readback.read_grid(path, sheet_name=sheet_name)
    grid = data["grid"]
    missing = []
    mismatched = []
    for ref, dval in declared.items():
        key = _parse_cell_ref(ref)
        if key not in grid:
            missing.append(ref)
            continue
        aval = grid[key]
        if not _values_agree(dval, aval):
            mismatched.append((ref, dval, aval))
    return ReportCellsResult(missing=missing, mismatched=mismatched, surplus=[])
