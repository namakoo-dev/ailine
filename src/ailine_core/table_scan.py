"""表の走査 ── シートのどこまでが使われているか、見出しから列を引く、座標を書く。

★ なぜ在るか（2026-09-03・分割の一歩目）: 単一ファイル（17,000 行超）を割る前に、
**事後条件の層が要る荷物**を先に外へ出す。この 7 本は「op ごとの事後条件」から
14〜2 箇所ずつ呼ばれ、かつ**その外からも呼ばれる**ので、事後条件層と一緒に動かすと
`ailine_core → ailine` の循環になる。だから先にここへ置く。

★ 何を置くかの線: **シートを走査して位置や広がりを求めるもの**だけ。
値の意味づけ（合計行か・数値か・空欄か）は置かない ── それは
`total_row` / `primitives` / `column_type` の仕事で、混ぜると
「表のどこを見ているか」と「その値をどう解釈するか」が 1 つのモジュールに同居する。

★ 走査と物理の使用範囲は**別物**（`extent_gap` が在る理由）:
  走査は「見出しから連続して値がある範囲」、物理は「openpyxl が持っている範囲」。
  食い違うのは、離れた場所に値が残っている時 ── その差を数えて開示するために
  両方を持っている。**片方だけ見て『全部見た』と言わない**ための組。
"""
from __future__ import annotations

from openpyxl.utils import get_column_letter

def _cell_ref(row: int, col: int) -> str:
    """**1 起点**の行/列 (r, c) を、人が読める A1 形式（例: C2）にする。

    ★★ この repo には**起点が 2 つある**（2026-09-03 に数えて確かめた）:
      ・**0 起点** … LibreOffice Basic 側。`getCellByPosition(列, 行)`、
        ヘルパの `headerRow` 引数、LLM に見せる「列は 0 起点で 0..N」の説明、
        `used_range`（ailine.py:643 に 0起点→1起点 の変換が在る）
      ・**1 起点** … openpyxl 側と検算側。`_col_index_by_header` の戻り値、
        スナップショットの `r,c`（`for r in range(1, nrow + 1)`）

    ★ この関数は **1 起点の側**にある。docstring は長く「0起点」と書かれていたが、
      呼び元 8 箇所すべてが 1 起点を渡しており、**実装と呼び元は整合していた** ──
      説明だけが Basic 側の世界のまま取り残されていた。
    ★ 0 を渡すと `get_column_letter` が ValueError を投げる（列 0 は存在しない）。
      それでよい ── **黙って `B0` のような嘘の座標を作るより落ちる方が安全**。
      この振る舞いは tests/test_cell_ref_is_one_based.py が凍結している。
    """
    return f"{get_column_letter(col)}{row}"

def _col_index_by_header(ws, name: str, header_row: int = 1):
    """見出し行(既定は物理1行目・header_row で指定可)を左から走査して name に一致する
       列の1起点インデックスを返す。無ければ None。★ W3: header_row 省略時は旧挙動と同一。
       ★ W3: header_row>1 の子見出し行で空欄の列は、book_columns() と同じ規則で
       真上の行を遡って引き継ぐ（多段見出しの先頭列対策。無いと『7月』等キー列より
       手前の空欄列で走査が誤って打ち切られる）。"""
    c = 1
    while True:
        v = ws.cell(row=header_row, column=c).value
        if v in (None, "") and header_row > 1:
            for up in range(header_row - 1, 0, -1):
                uv = ws.cell(row=up, column=c).value
                if uv not in (None, ""):
                    v = uv
                    break
        if v in (None, ""):
            return None
        if str(v) == name:
            return c
        c += 1

def _row_has_any_value(ws, row: int, last_col: int) -> bool:
    """その行に、表の幅の中で 1 つでも値が在るか（2026-09-05）。"""
    for c in range(1, last_col + 1):
        if ws.cell(row=row, column=c).value not in (None, ""):
            return True
    return False


def _scan_last_row(ws, key_col: int = 1, header_row: int = 1) -> int:
    """データの最終行（見出し行を除く）。データが無ければ header_row。

    ★★ 2026-09-05（実物の請求書で実測・段B）: それまでは
      **key_col（既定 1 列目）を上から見て、最初の空で止める**規則だった。
      帳票はこれを満たさない ──

        ・A 列が余白の請求書 → 1 行目から空なので **0 行**と読む（実測: rows_scanned=0）
        ・日付が最初の行にしか無い明細 → 2 行目で止まる
        ・小計や続きの行で 1 列目が空く → そこで止まる

      ★ そして「別のキー列を選び直す」では解けない ── 実物では**どの 1 列を見ても
        明細の全行を覆えない**（日付は先頭行だけ、品名は全行、数量は空の行がある）。

      ★ 直し方: 「**表の幅のどこかに値が在れば行**」に変える。終わりは
        **空行が続いたら**とする（表の下の別の表を巻き込まないため）。

      ★ 呼び出し側 23 箇所には配らない ── 規則はこの 1 本の中で変える
        （同じ判断が N 箇所にあると M 箇所だけ直る、を今日 4 回踏んでいる）。
      ★ key_col は互換のため受け取り続けるが、**空行の判定には使わない**。
        1 列目だけを見る呼び出しが残っていても、規則はここで統一される。
    """
    last_col = _scan_last_col(ws, header_row=header_row)
    if last_col <= 0:
        # ★ 見出しが読めない時だけ、従来どおり key_col を見る（退行させない）
        r = header_row + 1
        while ws.cell(row=r, column=key_col).value not in (None, ""):
            r += 1
        return r - 1
    # ★★ 空行が **1 行**でも在れば、そこが表の終わり。
    #   最初 GAP_ROWS=2 にしたら、既存の番人が捕まえた ──
    #   「空行で区切られた 2 つの塊」を続けて読むと **4600 == 1200+3400** が当たり、
    #   合計行の誤検出が起きる（tests/test_sum_identity.py の F9）。
    #   ★ 空行 1 つは Excel の表では**十分に強い区切り**で、跨いで読む理由が無い。
    #   ★ 「表の途中に空行がある」場合はそこで切れるが、それは**今までと同じ**
    #     （旧走査も 1 列目の空で止まっていた）── 段A がそれを正直に開示する。
    GAP_ROWS = 1
    physical, _ = _used_extent(ws)
    last = header_row
    gap = 0
    r = header_row + 1
    while r <= physical:
        if _row_has_any_value(ws, r, last_col):
            last = r
            gap = 0
        else:
            gap += 1
            if gap >= GAP_ROWS:
                break
        r += 1
    return last


def _used_extent(ws) -> tuple:
    """シートの**使用範囲**（行数, 列数）。中身のあるセルが 1 つも無ければ (0, 0)。

    ★ なぜ在るか（盲検の契約レビュー・2026-08-24）: 「雛形は 1 セルも変わっていない」を
      走査（A 列・1 行目を上/左から見て最初の空で打ち切り）で測っていたため、
      **A 列を余白にした典型的な請求書雛形**（B2 に「請 求 書」）では比較セル数が 0 になり、
      雛形が何セル壊れても pass していた。**分母ゼロの主張は主張ではない。**
    ★ 雛形は「表」ではなく「紙の見た目」なので、走査でなく使用範囲で測るのが正しい。
      openpyxl の max_row/max_column は書式だけのセルも含んで大きく出ることがあるので、
      値のあるセルの最大位置を自分で取る。
    """
    last_r = last_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                if cell.row > last_r:
                    last_r = cell.row
                if cell.column > last_c:
                    last_c = cell.column
    return last_r, last_c

def _scan_last_col(ws, header_row: int = 1) -> int:
    """見出し行の最終列（1 起点）。見出しが 1 つも無ければ 0。

    ★★ 2026-09-05（段B）: それまでは**左から見て最初の空で止める**規則で、
      A 列が余白の請求書では **0 列**を返していた（表が読めない真因の片割れ）。
      ★ book_columns（段C-1）と**同じ規則**にする ── 結合セルは範囲の左上から引き、
        空欄では打ち切らず、空が続いたらそこが右端。
      ★ 規則を 2 箇所に書き写さない: 名前の解決は book_columns、範囲の解決はここ、と
        役目は違うが、**「どこまでが表か」の線は 1 つ**でなければならない。
        だから同じ定数と同じ手順をここに置き、番人が両者の一致を縛る。
    ★ W3: 子見出し行（header_row>1）の空欄列は真上を遡って引き継ぐ（多段見出し）。
    """
    GAP_COLS = 3

    def _effective(c: int):
        v = ws.cell(row=header_row, column=c).value
        if v in (None, ""):
            v = _merged_value(ws, header_row, c)
        if v in (None, "") and header_row > 1:
            for up in range(header_row - 1, 0, -1):
                uv = ws.cell(row=up, column=c).value
                if uv in (None, ""):
                    uv = _merged_value(ws, up, c)
                if uv not in (None, ""):
                    return uv
        return v

    _, physical = _used_extent(ws)
    last = 0
    gap = 0
    for c in range(1, physical + 1):
        if _effective(c) not in (None, ""):
            last = c
            gap = 0
        else:
            gap += 1
            if gap >= GAP_COLS:
                break
    return last


def _merged_value(ws, row: int, col: int):
    """結合セルの中なら範囲の左上の値（2026-09-05・book_columns と同じ規則）。"""
    try:
        for rng in getattr(ws, "merged_cells", None).ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(row=rng.min_row, column=rng.min_col).value
    except Exception:
        return None
    return None


def _rows_beyond_the_table(ws, header_row: int, last_row: int, last_col: int) -> tuple:
    """走査が終わった先に在る行を、「表の続きらしい」と「表の外」に分ける（2026-09-05）。

    ★★ なぜ要るか（実物の請求書で実測）: 走査は 41 行目（税区分の末尾）で正しく
      終わっているのに、物理の使用範囲は 75 行まで在る（備考・振込先・広告文）。
      その差 34 行を「見ていない行」と数えると、**帳票では ✓ が永久に出ない**。
      備考欄や広告文は**処理の対象ではない**ので、見落としではない。

    ★ だが「表の外」と決めつけると危ない ── 表の**途中**に空行があって走査が
      そこで止まった場合、その先はまさに見落としだ（2026-09-05 の朝に直した事故）。
    ★ 見分け方（実測で決めた・Namakoo の判断 A′）: **表と同じ形をしているか**。
      表の行が最大 W 列を埋めているとき、先の行が W の半分より多く埋めていれば
      「表の続きかもしれない」とみなす。
        実測: ⑩途中に空行 → 3/3 で止める（正しい）／実物の備考 → 1/7 で通す（正しい）／
              下に 2 列の別表 → 2/3 で止める（**安全側に倒れる**・嘘はつかない）
    ★ 迷ったら止める。見落としを隠すより、うるさい方がよい。
    """
    def fill(r: int) -> int:
        return sum(1 for c in range(1, last_col + 1)
                   if ws.cell(row=r, column=c).value not in (None, ""))

    body = [fill(r) for r in range(header_row + 1, last_row + 1)]
    width = max(body) if body else 0
    physical, _ = _used_extent(ws)
    like_table, outside = [], []
    for r in range(last_row + 1, physical + 1):
        n = fill(r)
        if n == 0:
            continue
        (like_table if n * 2 > width else outside).append(r)
    return like_table, outside


def extent_gap(ws, header_row: int = 1, key_col: int = 1) -> dict:
    """走査で得た範囲と、**その先に在る表らしい行**の食い違いを数える。

    ★ 2026-08-25（塊②）: 中核 op の盲検が実測した 2 件の根。
      行: `_scan_last_row` が key_col を上から見て最初の空で止まり、末尾の行が
          処理からも分母からも消えた。
      列: 並べ替えの範囲が見出し由来なので、見出しの無い列が範囲外に落ちた。

    ★★ 2026-09-05（段B・A′）: 分母を**物理の使用範囲**から
      「**走査の先に在る、表と同じ形の行**」に狭めた。
      それまでは備考欄・振込先・広告文まで「見ていない行」と数えており、
      帳票では ✓ が原理的に出なかった（実測: 実物で 34 行が誤って計上）。
      ★ 狭めても危険は増えない ── 表と似た形の行は必ず数える（_rows_beyond_the_table）。
      ★ 表の外に在る行は rows_outside として**別に返す**。黙らせるのではなく分けて言う。
    戻り値の rows_* / cols_* はいずれも**見出し行を除いたデータの数**。
    """
    phys_r, phys_c = _used_extent(ws)
    cols_physical = max(0, phys_c)
    last_row = _scan_last_row(ws, key_col=key_col, header_row=header_row)
    rows_scanned = max(0, last_row - header_row)
    cols_scanned = max(0, _scan_last_col(ws, header_row=header_row))
    like_table, outside = _rows_beyond_the_table(ws, header_row, last_row, cols_scanned or phys_c)
    return {
        "rows_scanned": rows_scanned,
        "rows_physical": max(0, phys_r - header_row),
        "rows_missing": len(like_table),
        "rows_outside": len(outside),
        "cols_scanned": cols_scanned, "cols_physical": cols_physical,
        "cols_missing": max(0, cols_physical - cols_scanned),
    }

def data_extent(ws, header_row: int = 1) -> tuple:
    """見出しより下のデータの**物理の**広がり (最終行, 最終列)。

    ★★ 2026-08-27（Namakoo が実測・今日 3 箇所目）: `_scan_last_row` は 1 列目を上から
      見て**最初の空で止まる**。表の途中に空行が 1 本あるだけで、その下が全部消える。
      ・位置の解決が「みかんが見つかりません」になった
      ・事後条件が「検証対象が 0 件」になった
      ・分母が消えた（塊①で直したのと同じ形）
      ★ 3 箇所で同じ穴を開けたので、**器官を 1 つにする**（それぞれに書き写さない）。
    """
    phys_r, phys_c = _used_extent(ws)
    last = max(phys_r, _scan_last_row(ws, header_row=header_row))
    cols = max(phys_c, _scan_last_col(ws, header_row=header_row), 1)
    return last, cols
