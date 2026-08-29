"""cellmap — 表を「座標 → 中身」の表として扱い、操作を**座標の写像**として宣言する層。

★★ なぜ要るか（2026-08-29・Namakoo の設計）:
  「原本から先に全てのセルの位置情報と内容を内部的な表にできないか？
    行の挿入は、今の (x,y) から後ろを (x+1,y) にずらして、(x,y) に入れるだけだよな？」

  ★ この見立ては**宣言の言語としても、検算の物差しとしても正しい**。
    基本操作はどれも「座標の写像 π」と「新しく書く座標の集合 W」の 2 つで書ける:

      行の挿入 k: π(r,c) = (r+1,c) if r>=k else (r,c)      W = {(k, 列): 値}
      行の削除 k: π(r,c) = (r-1,c) if r>k  else (r,c)      W = {}（k 行目は消える）
      行の交換 a,b: π が a↔b を入れ替える                  W = {}
      列の挿入/削除/移動: 同じ形を列でやるだけ
      1 セル書換: π = 恒等                                  W = {(r,c): 値}
      並べ替え: π は行の置換（データが決める）               W = {}

  ★ ただし**実行の手順としては足りない**（実測 2026-08-29）。行を 1 本挿しただけで、
    値以外にこれだけ座標が付け替わった:
        =B3*C3      → =B4*C4        （数式の中の座標）
        =SUM(D2:D4) → =SUM(D2:D5)   （範囲が伸びた ── 単純にずらすと合計から漏れる）
        結合 A8:D8  → A9:D9
        行の高さ 40 が 3 行目 → 4 行目
    値だけを写しても表は壊れる。だから**実行は LibreOffice に任せたまま**にして、
    ここは「何をするかの宣言」と「本当にそうなったかの検算」だけを担う。
    ★ 自分で計算して書き込むと、検算が『自分の計算 vs 自分の書き込み』＝恒真になる。

★ ailine を import しない（ailine_core の作法）。
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field


# --- 座標の写像 -------------------------------------------------------------------

@dataclass(frozen=True)
class Shift:
    """座標の写像。**行と列を同じ形で書く**（Namakoo の指摘した対称性）。

    axis: "row" / "col" / "none"。kind: "identity" / "insert" / "delete" / "swap" / "permute"。
    at/count: 挿入・削除の位置と本数。a/b: 交換する 2 つ。order: 置換（並べ替え）。
    """
    kind: str = "identity"
    axis: str = "none"
    at: int = 0
    count: int = 0
    a: int = 0
    b: int = 0
    order: tuple = ()          # permute: 適用後 i 番目に来る「適用前の番号」

    def moved(self, idx: int) -> int | None:
        """適用前の行(列)番号 → 適用後の番号。消えるなら None。"""
        if self.kind == "identity":
            return idx
        if self.kind == "insert":
            return idx + self.count if idx >= self.at else idx
        if self.kind == "delete":
            if self.at <= idx < self.at + self.count:
                return None
            return idx - self.count if idx >= self.at + self.count else idx
        if self.kind == "swap":
            return self.b if idx == self.a else (self.a if idx == self.b else idx)
        if self.kind == "permute":
            for new_i, old_i in enumerate(self.order, start=1):
                if old_i == idx:
                    return new_i
            return idx
        raise ValueError(f"未知の写像: {self.kind}")

    def map_cell(self, row: int, col: int):
        """(行, 列) → 適用後の (行, 列)。消えるなら None。"""
        if self.axis == "row":
            r = self.moved(row)
            return None if r is None else (r, col)
        if self.axis == "col":
            c = self.moved(col)
            return None if c is None else (row, c)
        return (row, col)

    def describe(self) -> str:
        unit = "行" if self.axis == "row" else ("列" if self.axis == "col" else "")
        if self.kind == "identity":
            return "座標は動かない"
        if self.kind == "insert":
            return f"{self.at}{unit}目に {self.count} 本挿し、以降を後ろへずらす"
        if self.kind == "delete":
            return f"{self.at}{unit}目から {self.count} 本消し、以降を前へ詰める"
        if self.kind == "swap":
            return f"{self.a}{unit}目と {self.b}{unit}目を入れ替える"
        return f"{unit}を並べ替える"


def identity() -> Shift:
    return Shift()


def insert_rows(at: int, count: int = 1) -> Shift:
    return Shift(kind="insert", axis="row", at=at, count=count)


def delete_rows(at: int, count: int = 1) -> Shift:
    return Shift(kind="delete", axis="row", at=at, count=count)


def swap_rows(a: int, b: int) -> Shift:
    return Shift(kind="swap", axis="row", a=a, b=b)


def insert_cols(at: int, count: int = 1) -> Shift:
    return Shift(kind="insert", axis="col", at=at, count=count)


def delete_cols(at: int, count: int = 1) -> Shift:
    return Shift(kind="delete", axis="col", at=at, count=count)


def swap_cols(a: int, b: int) -> Shift:
    return Shift(kind="swap", axis="col", a=a, b=b)


def permute_rows(order) -> Shift:
    return Shift(kind="permute", axis="row", order=tuple(order))


# --- 数式の中の座標も同じ写像で動かす ------------------------------------------------

# ★ 直前がシート修飾（`Sheet2.` / `Sheet2!`）なら**このシートの写像は効かない**。
#   2026-08-29 の検体が捕まえた: `=Sheet2.B3` を `=Sheet2.B4` に動かしていた
#   ── 別のシートに行を挿したわけではないので、これは静かな破壊になる。
_REF = re.compile(r"(?<![.!])(?<![A-Za-z0-9_$])(\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})")


def _col_to_num(s: str) -> int:
    n = 0
    for ch in s:
        n = n * 26 + (ord(ch) - 64)
    return n


def _num_to_col(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def map_formula(formula: str, shift: Shift) -> str | None:
    """数式の中の座標を、**同じ写像で**動かした形にする。消える参照があれば None。

    ★★ 2026-08-29 の実測がこの関数の理由:
      行を 1 本挿すと `=B3*C3` は `=B4*C4` に、`=SUM(D2:D4)` は `=SUM(D2:D5)` になった。
      前者は「参照が 1 つずれる」、後者は「範囲の**終わりだけ**がずれる」── どちらも
      **参照を 1 つずつ写像に通す**という同じ規則で説明できる（範囲は両端を通す）。
    ★ だから検算は「数式の文字列が同じか」ではない。**写像を通したら一致するか**だ
      （文字列で比べると、正しい付け替えを毎回『壊れた』と言う）。
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    out, lost = [], False

    def sub(m):
        nonlocal lost
        cd, cl, rd, rw = m.group(1), m.group(2), m.group(3), m.group(4)
        mapped = shift.map_cell(int(rw), _col_to_num(cl))
        if mapped is None:
            lost = True
            return m.group(0)
        r2, c2 = mapped
        return f"{cd}{_num_to_col(c2)}{rd}{r2}"

    out = _REF.sub(sub, formula)
    return None if lost else out


# --- 座標表 ---------------------------------------------------------------------

@dataclass
class CellMap:
    """1 シートを「座標 → 中身」で持つ。値・数式・結合・行列の寸法まで。

    ★ 実行はしない（読むだけ）。★ 判定もしない（材料を持つだけ）。
    """
    values: dict = field(default_factory=dict)      # (r, c) -> 生の値（数式なら文字列）
    cached: dict = field(default_factory=dict)      # (r, c) -> 計算値（あれば）
    merges: tuple = ()                              # ((r1, c1, r2, c2), ...)
    row_heights: dict = field(default_factory=dict)
    col_widths: dict = field(default_factory=dict)
    max_row: int = 0
    max_col: int = 0

    def nonempty(self):
        return {p for p, v in self.values.items() if v not in (None, "")}

    def at(self, p):
        return self.values.get(p)


def _is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def read_cellmap(path, sheet=None) -> CellMap:
    """xlsx を座標表として読む（openpyxl だけを使う・LO は起動しない）。"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    wsv = wbv[ws.title]
    cm = CellMap(max_row=ws.max_row or 0, max_col=ws.max_column or 0)
    for r in range(1, cm.max_row + 1):
        for c in range(1, cm.max_col + 1):
            v = ws.cell(r, c).value
            if not _is_empty(v):
                cm.values[(r, c)] = v
                cv = wsv.cell(r, c).value
                if cv is not None:
                    cm.cached[(r, c)] = cv
    cm.merges = tuple((m.min_row, m.min_col, m.max_row, m.max_col)
                       for m in ws.merged_cells.ranges)
    cm.row_heights = {r: d.height for r, d in ws.row_dimensions.items() if d.height}
    cm.col_widths = {c: d.width for c, d in ws.column_dimensions.items() if d.width}
    wb.close()
    wbv.close()
    return cm


# --- 法則（これ 1 本で全部の基本操作を検算する）--------------------------------------

@dataclass(frozen=True)
class Verdict:
    ok: bool
    reason: str = ""


def check_shift(before: CellMap, after: CellMap, shift: Shift,
                 writes: dict | None = None, *, check_merges: bool = True) -> Verdict:
    """**1 本の法則**: 適用後の表は、適用前の表を写像で動かしたもの＋宣言した書き込み。

    証明するもの:
      ① 動いた先が元の中身のまま（数式は**写像を通してから**比べる）
      ② 宣言した座標が宣言した値になっている
      ③ **それ以外の座標に何も無い**（頼んでいない所に書いていない ＝ 負の被覆）
      ④ 結合セルも同じ写像で動いている

    ★ op ごとの専用チェッカーを 1 本に畳むための土台。
      「消えたものは diff に出ない」への正面からの答え ── ③が在れば、
      **宣言していない変化は全部**捕まる（op が増えても法則は同じ）。
    """
    writes = {tuple(k): v for k, v in (writes or {}).items()}
    expected = {}
    for p, v in before.values.items():
        q = shift.map_cell(*p)
        if q is None:
            continue                      # 消える座標（削除）
        if isinstance(v, str) and v.startswith("="):
            mv = map_formula(v, shift)
            expected[q] = mv if mv is not None else v
        else:
            expected[q] = v
    expected.update(writes)

    got = {p: v for p, v in after.values.items()}
    # ① ② 宣言どおりか
    for q, want in sorted(expected.items()):
        have = got.get(q)
        if _values_differ(have, want):
            src = [p for p in before.values if shift.map_cell(*p) == q]
            hint = f"（元 {src[0]} から動いたはず）" if src and q not in writes else ""
            return Verdict(False, f"{_ref(q)} が {have!r} ですが、{want!r} のはずです{hint}")
    # ③ 宣言していない所に何も無いか
    extra = sorted(set(got) - set(expected))
    if extra:
        head = "・".join(f"{_ref(q)}={got[q]!r}" for q in extra[:3])
        more = f" ほか {len(extra) - 3} 箇所" if len(extra) > 3 else ""
        return Verdict(False, f"宣言していない所に中身があります: {head}{more}")
    # ④ 結合セル
    if check_merges:
        want_m = set()
        for (r1, c1, r2, c2) in before.merges:
            a, b = shift.map_cell(r1, c1), shift.map_cell(r2, c2)
            if a and b:
                want_m.add((a[0], a[1], b[0], b[1]))
        got_m = set(after.merges)
        if want_m != got_m:
            return Verdict(False, f"結合セルが宣言どおり動いていません: "
                                   f"欠け {sorted(want_m - got_m)}・余り {sorted(got_m - want_m)}")
    return Verdict(True, f"{shift.describe()} ── 宣言した座標だけが宣言どおり変わりました"
                          f"（写した {len(expected) - len(writes)} 箇所・"
                          f"新しく書いた {len(writes)} 箇所・それ以外は 1 箇所も無し）")


def _values_differ(have, want) -> bool:
    if _is_empty(have) and _is_empty(want):
        return False
    if isinstance(have, (int, float)) and isinstance(want, (int, float)):
        return abs(float(have) - float(want)) > 1e-9
    return str(have) != str(want)


def _ref(p) -> str:
    r, c = p
    return f"{_num_to_col(c)}{r}"


# --- 参照のズレ（並べ替え・入れ替えで「指す先の中身」が変わる式）----------------------

_RANGE_RE = re.compile(
    r"(\$?[A-Z]{1,3}\$?[0-9]{1,7})\s*:\s*(\$?[A-Z]{1,3}\$?[0-9]{1,7})")
_QUALIFIED_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z_\u3040-\u30ff\u4e00-\u9fff][\w\u3040-\u30ff\u4e00-\u9fff]*))"
    r"[.!](\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})")
_BARE_RE = re.compile(r"(?<![.!])(?<![A-Za-z0-9_$])(\$?)([A-Z]{1,3})(\$?)([0-9]{1,7})")
# ★ コロンに隣接する参照（相手が関数でも範囲の端）。`B2:INDEX(...)` / `INDEX(...):B9`
_ADJ_COLON_RE = re.compile(
    r"(\$?[A-Z]{1,3}\$?[0-9]{1,7}\s*:)|(:\s*\$?[A-Z]{1,3}\$?[0-9]{1,7})")


def single_cell_refs(formula: str, own_sheet: str) -> list:
    """式の中の**単独セル参照**を [(シート名, 行, 列), ...] で返す（範囲の両端は除く）。

    ★★ 2026-08-29（Namakoo の指摘 → 実測で裏取り）: 並べ替えると、**範囲の外から
      特定の 1 行を指している式**は追従せず、**指す先の中身だけが変わる**。
      実測: `=B3`（ラベルは「ぶどうの金額」）が、並べ替え後に みかん の 200 を指した。
      式は 1 文字も壊れていないので、値でも文字列でも検出できない ── **参照を読むしかない**。
    ★ 範囲（`SUM(B2:B4)`）は除く: そちらは領域を指しているので並べ替えに正しく追従する。
      鳴らすのは「特定の 1 行を指す式」だけ ── ここを分けられるのが、式を読む値打ち。
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    masked = _RANGE_RE.sub(lambda m: "#" * len(m.group(0)), formula)
    # ★★ 2026-08-29（既存の検体が捕まえた誤検出）: `=SUM(B2:INDEX(B:B,ROW()-1))` の
    #   `B2` を「単独参照」と読んでいた ── 相手が関数なので `A1:B2` の形に見えない。
    #   ★ **コロンに隣接する参照は範囲の端**。範囲は領域を指すので鳴らさない。
    masked = _ADJ_COLON_RE.sub(lambda m: "#" * len(m.group(0)), masked)
    out = []
    for m in _QUALIFIED_RE.finditer(masked):
        sheet = m.group(1) or m.group(2)
        out.append((sheet, int(m.group(6)), _col_to_num(m.group(4))))
    # 修飾つきを伏せてから、素の参照を拾う（同じ参照を 2 回数えない）
    rest = _QUALIFIED_RE.sub(lambda m: "#" * len(m.group(0)), masked)
    for m in _BARE_RE.finditer(rest):
        out.append((own_sheet, int(m.group(4)), _col_to_num(m.group(2))))
    return out


def refs_pointing_into(path, target_sheet: str, row_lo: int = 1, row_hi: int = 10 ** 7,
                        col_lo: int = 1, col_hi: int = 10 ** 4) -> list:
    """動かす区画を**外から**指している単独セル参照を集める。

    区画は 行 [row_lo, row_hi] × 列 [col_lo, col_hi]。
      並べ替え → 行で区切る（列は全部）
      列の入れ替え → 列で区切る（行は全部）
    戻り値: [(式の在るシート, 式のセル参照, 式, 指している行, 指している列), ...]。
    ★ 「外から」= 別シート、または同じシートでも区画の外。区画の中の式は
      一緒に動くので正しく追従する（実測で確認済み）。
    """
    import openpyxl
    hits = []
    wb = openpyxl.load_workbook(path)

    def _in_block(sheet, r, c):
        return (sheet == target_sheet and row_lo <= r <= row_hi and col_lo <= c <= col_hi)

    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or not v.startswith("="):
                        continue
                    if _in_block(ws.title, cell.row, cell.column):
                        continue
                    for sh, r, c in single_cell_refs(v, ws.title):
                        if _in_block(sh, r, c):
                            hits.append((ws.title, _ref((cell.row, cell.column)), v, r, c))
                            break
    finally:
        wb.close()
    return sorted(hits, key=lambda t: (t[0], t[3], t[4]))


def reference_drift_note(hits: list) -> str | None:
    """検出した参照のズレを 1 行にする（無ければ None）。

    ★ 直さない ── **直してよいかは人が決める**（Excel も LibreOffice も、範囲の外から
      特定の行を指す式は並べ替えで追従させない＝アドレスに留まるのが既定の意味）。
      「ぶどうの金額 = B3」は行に追従してほしいが、「3行目の値 = B3」は留まってほしい
      ── 機械には区別できない。だから**名指しして人に返す**。
    """
    if not hits:
        return None
    head = "・".join(f"{sh}!{ref}（{f}）" for sh, ref, f, *_ in hits[:3])
    more = f" ほか {len(hits) - 3} 件" if len(hits) > 3 else ""
    return (f"この操作で、**指す先の中身が変わる式**が {len(hits)} 件あります: {head}{more}"
             " ── 式そのものは壊れませんが、指している行が入れ替わります"
             "（直してよいかは人が決めることなので、直していません）")
