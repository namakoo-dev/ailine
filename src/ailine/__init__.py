#!/usr/bin/env python
"""ailine — 自然言語のタスクを、ローカル LLM が LibreOffice Basic に書き起こし、
   basrun で文書に適用し、★ 効果を読み戻して検証する（「走った ≠ できた」）。

    ailine run  <book> "<タスク>"          生成 → 検品ゲート → 原本に反映 → 変化検証 → 修復 → 差分表示
    ailine run  <book> "<タスク>" --dry     生成して見せるだけ（適用しない・レビュー用）
    ailine run  <book> "<タスク>" --copy    原本には触らず <book>.out に結果を作る（旧既定）
    ailine undo <book>                      直前の反映を取り消す（世代バックアップから復元）
    ailine stop                             起動した LibreOffice を落とす（basrun に委譲）

## 設計判断（basrun_spike 2026-08-10 の実証に基づく）

- **モデル非依存**: `--model` で差し替え（既定 qwen2.5-coder:7b）。天井はモデルの大きさでなく
  「参照例の供給＋効果の検証」で上げる（7B が正解例1本で苦手層 0%→67% になった）。
- ★ **検証をループに**: 適用の前後で文書が変化したかを見る **no-op ガード**。
  LibreOffice + LLM は「実行時エラー無しで成功と報告し、実際は何もしない」ことがある
  （もっともらしい UNO の幻覚）。変化ゼロなら失敗として修復に回す。
- ★ **正規化パス（2026-08-14 修正・製品の心臓）**: before スナップショットの前に、
  コピーを LibreOffice で一度（空マクロで）開いて保存する。openpyxl 製ブックは LO の
  初回保存で行高（時に列幅）を実体化する副作用があり、これを先に済ませておかないと
  no-op ガードが偽陽性になる（何もしないマクロでも「変化した」と誤って成功表示していた）。
  コストは LO 往復 1 回（数秒）— 正しさ優先で受け入れる。
- ★ **既定=原本にそのまま反映（W8b-2）**: `<book>` に直接書く。「壊さない」の担保は
  もう「コピーにしか書かない」ことではなく、3重の安全網でまかなう:
  ①往復忠実度ゲート（LO 往復だけで失われる飾りを検品して申告・`--accept-loss`/`--copy` で選ぶ）
  ②世代バックアップ+`ailine undo`（反映前に必ず退避・いつでも戻せる）
  ③ no-op ガード＋事後条件などの機械検証（変化・正しさを見る）。
  従来の「コピーにしか書かない」挙動は `--copy` で選べる（原本は無変更・`<book>.out.xlsx` に生成）。
- **参照ライブラリ**: `refs/*.bas` を few-shot に供給。苦手層（新シート・色）を補う。
  並べ替え・グラフ・ピボットなどの難所は `helpers/*.bas` の検証済みヘルパを `Call` で呼ばせる。
- **レビュー導線**: 生成した .bas と、変わったセルの差分を必ず表示する。

## 正直な限界

- LibreOffice Basic + UNO は学習データが薄く、珍しい操作は外しやすい。参照とヘルパで補う設計
  （太字も当初「環境不可」と誤断したが、実際は `CharWeight`+`CharWeightAsian` の native 書きで解決済み）。
- ローカル LLM(ollama) と LibreOffice(basrun 経由) が要る。**外部送信はしない。**
- no-op ガードが保証するのは「変化したこと」だけ。「**正しいか**」は差分を人が見て判断する。

## 置き場所（上級者向け）

- 既定は `~/.ailine`。環境変数 **AILINE_HOME** を設定すると、履歴/用語集/別名/バックアップ等
  すべての置き場所をそこへ差し替えられる（`resolve_home_dir()` 参照）。
"""
from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
# ★ _date_cls は本体では使わなくなった（2026-09-03 に事後条件の共有部分を
#   ailine_core/postconditions/_shared.py へ移したため）。公開面の凍結
#   (tests/ailine_public_surface.txt) が名前を守っているので、ailine._date_cls として
#   引ける状態は保つ。★ 分割の diff に「名前の削除」を混ぜない。
from datetime import date as _date_cls   # noqa: F401 ── 再輸出（公開面の凍結が守る名前）
from pathlib import Path

# ★ 2026-08-24（盲検の査定で最も痛い指摘・第 1 位）: この 2 つは openpyxl の import
#   ガードより**後ろ**に置かれていて、openpyxl が無い環境では
#   `NameError: name 'exit_environment' is not defined` という生の traceback が出ていた。
#   ── 「足りないものを名指しする」ための ailine doctor すら到達できない。
#   初日・素の環境という、開発機では絶対に踏まない道。定義を利用点の前へ移す。
# ★ 終了コードの分離（2026-08-23 の取り込みで見つかった穴・実測）: 依存の欠落・外部
#   プログラムに繋がらない・入力が無い、といった「実行の前提が満たされていない」状態は、
#   検証の失敗（適用したが事後条件を満たさない = 1）と意味が違う。旧実装は exit_environment("...") が
#   全部 1 に潰れており、CI や自動化から「⚠ が出た」と「道具が壊れた」を区別できなかった。
#   既存の 3(CLARIFY)/4(忠実度)/5(verify)/6(ロック)/7(上書き関所)/8(自由生成の関所) と
#   argparse の 2 を避けて 9 を割り当てる。
EXIT_ENVIRONMENT = 9
# ★ 2026-08-24（初回体験の盲検・致命②）: 5 は golden 表と検体で「書き込めない」に
#   凍結済み。README の表が verify 専用の意味（検算の不一致）しか書いておらず、
#   同じ番号に 2 つの意味が同居していた ── 番号でなく**表**を実態に合わせた。
EXIT_WRITE_BLOCKED = 5
# ★ 2026-08-25（復元の重大6）: 原本への反映そのものが中止された時の番号。
#   「道具が壊れた」ではなく「前提が満たせず反映できなかった」ので 9 と同じ族にする。
EXIT_APPLY_FAILED = 9


def exit_environment(message: str):
    """実行の前提が満たされていない旨を述べて EXIT_ENVIRONMENT で落ちる。
       ★ sys.exit(文字列) は必ず 1 になる ── 意味を持つ番号で落ちる唯一の入口にする。"""
    print(message, file=sys.stderr)
    raise SystemExit(EXIT_ENVIRONMENT)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    exit_environment("openpyxl が要る:  pip install openpyxl")

# ★ C4（再設計 分割の一歩目）: 新しい単位は ailine.py に足さず ailine_core/ に置き、
#   ここからは import するだけにする（tests/ailine_py_line_budget.txt が単調減少を見張る）。
#   ailine_core は ailine.py と同じディレクトリの兄弟パッケージなので、python が script
#   実行時に自動で sys.path[0] に入れる ailine.py の所在ディレクトリからそのまま拾える
#   （別の作業ディレクトリから `python C:\...\ailine.py run ...` と叩いても同じ）。
from ailine_core.book_view import BookView
from ailine_core.claim import (   # ★ C5/C9: Claim 型と『✓』の一元レンダラ（✓ は反映後の1箇所だけ）
    Claim, format_plan_report, format_plan_preview, overall_verdict,
    render_applied_claim, render_applied_unverified, render_applied_unobservable,
    render_scope_notes,   # ★ 単位E: 常時注記を廃止し、その run 固有の②の1文に置き換えた
    count_suspicious_advisories, render_applied_claim_demoted,   # ★ 決裁③: ⚠ による ✓ の降格
)
from ailine_core.dsl_step import (   # ★ C7: 単発 DSL / 複合計画の DSL 段が共有する実行エンジン
    DslStepDeps, resolve_dsl_step_args, print_dsl_confirmation, apply_dsl_step, compose_dsl_step_advisories,
    NEW_COLUMN_ORIGIN,   # ★ 単位B: 「直前の段が作った列」の文言の唯一の出どころ
    is_transient_lo_error, TRANSIENT_LO_RETRY_NOTICE,   # ★ 摩擦⑥: LO 一時不調の凍結マーカーと開示文言の唯一の出どころ
)
from ailine_core.cli_render import (   # ★ C8: 複数経路が同じ形を手書きしていた表示の純関数化
    render_excluded_lines,
    render_code_block, render_retry_options, render_aborted, render_run_header,
    render_backup_list, render_legacy_note,
    render_restore_done, render_vocab_add_result, render_vocab_listing,
    render_alias_listing,   # ★ W10 便A: `ailine alias list`
    render_ops_table,
    freeform_notice_reason, render_freeform_notice_compact,   # ★ K-1（単発向けの旧 render_freeform_notice は廃止）
    render_vocab_miss_refusal,   # ★ freeform 最終決定: 単発の語彙外の断り
    render_scan_report,   # ★ M1読み: `ailine scan`
    render_stack_report, render_verify_report,   # ★ M1書き: `ailine stack` / `ailine verify`
    render_folder_routes,
    render_verify_match_report,   # ★ M3: `ailine verify <出力> <元A> <元B>`（照合出力の検算）
)
from ailine_core.filetypes import (BOOKLIKE_SUFFIXES, CSV_SUFFIX,
                                   OPENPYXL_PROBEABLE_SUFFIXES,
                                   RUN_SUPPORTED_SUFFIXES)   # ★ 拡張子判定の登録簿
from ailine_core import multifile   # ★ M1読み: 多ファイル棚卸し（DESIGN-20260821-multifile.md）
from ailine_core import stack as multifile_stack   # ★ M1書き: 縦積み本体（DESIGN v2 §1 M1書き）
from ailine_core import verify as multifile_verify   # ★ M1書き: `ailine verify` の検算本体
from ailine_core import xml_readback   # ★ 検算の独立読み実装（openpyxl を import しない別実装）
from ailine_core import extract_multi   # ★ M2: `ailine run <フォルダ>`（抽出集約）の本体
from ailine_core import inspection   # ★ M2.5: 検分シート + 視覚的誘導（DESIGN §M2.5）
# ★ 2026-08-24: 一部は**意図した再輸出**（検体が ailine.sanitize_sheet_name の形で
#   見ている）。未使用に見えても消さない ── リンタには noqa で伝える。
from ailine_core.report_per_row import (  # noqa: F401
    cells_with_multiple_placeholders,   # ★ 帳票段: REPORT_PER_ROW の純ロジック部品
    sanitize_sheet_name, unique_sheet_name, scan_placeholders, compare_report_cells,   # noqa: F401 ── 再輸出/在否確認のため残す
)
from ailine_core import report_group   # ★ 帳票段（まとめ版）: 同じ取引先を 1 枚にまとめる純ロジック
from ailine_core import cellmap   # ★ 座標の層: 写像・数式の参照・参照のズレ検出
from ailine_core import match as multifile_match   # ★ M3: `ailine run <A> <B>`（2冊の照合）の本体
from ailine_core import total_row   # ★ operator 盲検7度目: 語のトリップワイヤ（第二の独立検出器）
from ailine_core import csv_quarantine   # ★ CSV 検疫: `ailine csv` / run 暗黙前段の本体
from ailine_core import csv_export   # ★ CSV_EXPORT: `ailine export-csv`（検疫の逆方向）の本体
from ailine_core import date_compare   # ★ EXTRACT の日付範囲比較（台帳 DATE_RANGE_AGG の正体）
from ailine_core import split_cell   # ★ SPLIT_CELL: 1セルの複数値を右の列へ割る（台帳2件）
from ailine_core import pdf_export   # ★ PRINT/EXPORT_DOC: `ailine export-pdf`（台帳4件）
from ailine_core.date_compare import (   # noqa: F401  ← 試験と呼び出し側が ailine. で引く
    parse_date_literal, date_to_serial, classify_date_column,
)
from ailine_core.chart_check import check_chart_series, charts_by_sheet   # ★ グラフ段: 事後条件②（種別+参照の検証）+ operator10 ②（シート別グラフ数）
from ailine_core.chart_range import chart_data_last_row   # ★ operator10 ①: グラフ範囲から合計行を除く
from ailine_core import compare_blocked
from ailine_core.column_type import column_is_all_numeric, value_parses_as_number   # ★ operator10 ④: 型の機械決定
from ailine_core.xml_readback import numeric_cells_became_strings   # ★ operator10 ⑤: 数式セルの偽アラーム防止   # noqa: F401 ── 再輸出/在否確認のため残す
from ailine_core.formula_health import formula_error_advisory, detect_write_target_type_change   # ★ 挙動変更#1(a)(b)
from ailine_core.write_precondition import (   # ★ 単位F/G: 宣言した領域の前提（破れた種類つき）   # noqa: F401 ── 再輸出/在否確認のため残す
    check_write_preconditions_detail,
    own_prior_output_notice_lines,   # ★ 単位H 開示: 関所が黙った理由を1行で見せる
)
from ailine_core.sum_identity import rows_matching_sum_above   # noqa: F401 ── 再輸出   # ★ 算術恒等の検算（二重計上）
from ailine_core import row_identity   # ★ 行内の等式（金額＝件数×単価）が操作で崩れたら言う
from ailine_core.target_sheet import (
    drop_names_covered_by_longer, sheets_named_explicitly,   # ★ 挙動変更#2/#3: 対象シートの決定を一箇所に閉じ込める
    resolve_target_sheet, describe_target_sheet, wrap_basic_for_sheet,
    format_sheet_field, sheet_conflict_choice_lines, conflict_excluded_sheets,
    sheet_names_mentioned_in,   # ★ 単位E: シート名照合の素材（決定側と助言側が共有する）
)
from ailine_core.subject import (   # ★ 単位E: A' 原則を「値」から「対象スロット」へ広げる
    Slot, Consumed as SubjectConsumed, classify_slots,
    COLUMN as SUBJ_COLUMN, REGION as SUBJ_REGION, ROW as SUBJ_ROW,
    SHEET as SUBJ_SHEET, LABEL as SUBJ_LABEL, INPUT as SUBJ_INPUT,
    SHEET_INPUT as SUBJ_SHEET_INPUT,   # ★ operator8 ①: LOOKUP_FILL の source_sheet 消費用
    name_matches_task,   # ★ W3 改定(2026-08-20): 実在しない target が「依頼文の名指し」か
                          #   「翻訳の捏造」かの照合に、単位B の部分文字列規律を再利用する
)
from ailine_core import alias_store   # ★ W10 便A: 別名ストアの検疫/照合/保存形式（純関数）
from ailine_core import suggest as suggest_candidates   # ★ W10 便C2: もしかして提案の候補生成（語としての厳格一致+about）
from ailine_core import residue as suggest_residue   # ★ W10 便C2 S5: もしかして提案の残差検出（純ロジック）
from ailine_core.interpretation import build_interpretation   # ★ 段1: 解釈を機械可読で出す（--json の interpretation/provenance）
from ailine_core.ask_choice import (   # ★ 挙動変更#3: 「選択肢を出して選ばせる」対話部品
    Choice, ask_choice, ask_yes_no, is_interactive,
)
from ailine_core import primitives
from ailine_core.primitives import is_number as _is_number
from ailine_core.table_scan import (   # noqa: F401 ── 再輸出
    # ★ _used_extent / _scan_last_col は本体では使わなくなった（2026-09-03 に
    #   事後条件を postconditions/ へ移したため）。公開面の凍結が名前を守っており、
    #   tests/test_review_wave3.py が ailine._used_extent として直接呼ぶ。
    _cell_ref, _col_index_by_header, _scan_last_row, _used_extent,
    _scan_last_col, extent_gap, data_extent,
)
from ailine_core.postconditions._shared import (   # noqa: F401 ── 再輸出
    COLOR_MAP, PIVOT_CAVEAT, _ZERO_TARGET_REASON,   # ★ 共有する文言・語彙
    # ★ 分割: 事後条件の共有部分。_MOVED_ROWS_WHY は本体では使わないが、
    #   公開面の凍結が名前を守っているので ailine. から引ける状態を保つ。
    _MOVED_ROWS_WHY, _cells_for_shift, _extract_predicate, _moved_rows_note,
    _numeric_value, _row_as_shown, compare_moved_rows, note_stringy_numbers,
    note_unverified,
)
from ailine_core.postconditions.shape import (   # noqa: F401 ── 再輸出（POSTCONDITIONS 辞書と公開面の凍結が名前で引く）
    _apply_operator, _bg_hex, _cell_target, _only_this_cell, check_autofit, check_bold,
    check_center_align, check_compute_column, check_compute_column_single_factor,
    check_draw_borders, check_fill_color, check_lookup_fill, check_merge, check_number_format,
    check_set_cell_value, check_set_column_value, check_set_where, only_this_column_changed,
)
from ailine_core.postconditions.move import (   # noqa: F401 ── 再輸出（POSTCONDITIONS 辞書と公開面の凍結が名前で引く）
    _check_swap_cells, _column_block_values, _dedup_key_display, _dedup_normalize_key_part,
    _fmt_amount, _nested_total_reason, _sort_rows_lost_their_identity,
    _total_row_left_the_bottom_reason, check_add_column, check_add_row, check_append_total,
    check_dedup, check_delete_column, check_delete_rows, check_insert_rows, check_sort,
    check_split_cell, check_swap, note_deleted, _APPEND_TOTAL_FORMULA_RE,
)
from ailine_core.postconditions.derive import (   # noqa: F401 ── 再輸出（POSTCONDITIONS 辞書と公開面の凍結が名前で引く）
    _check_report_router, check_aggregate, check_extract, check_extract_columns,
    check_format_map, check_pivot, check_report_per_group, check_report_per_row,
)
HERE = Path(__file__).resolve().parent
DEFAULT_REFS = HERE / "refs"
DEFAULT_HELPERS = HERE / "helpers"
OLLAMA = os.environ.get("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
DEFAULT_MODEL = os.environ.get("AILINE_MODEL", "qwen2.5-coder:7b")
DEFAULT_APPLY_TIMEOUT = 180.0  # M1: 暴走マクロで無限ハングしないよう既定 ON（--timeout 0 で無効化）

def resolve_home_dir() -> Path:
    """ailine のホームディレクトリ（第二波 ①）。環境変数 AILINE_HOME があればその Path、
       無ければ従来どおり ~/.ailine。★ subprocess 起動のテストにも env 継承で届く構造の
       隔離 ── monkeypatch.setattr は同一プロセスにしか効かないため、`ailine.py` を
       別プロセスとして起動する経路（14 ファイルの subprocess テスト）はこれまで実 home に
       素通りしていた（SEALED-20260823-jisaku-ultra.md 所見⑦）。呼び出しのたび環境変数を
       読む（モジュール import 時点の値に固定しない）ので、テストから直接呼んでも効く。"""
    env = os.environ.get("AILINE_HOME")
    if env:
        return Path(env)
    return Path.home() / ".ailine"


HISTORY_DIR = resolve_home_dir()
HISTORY_FILE = HISTORY_DIR / "history.jsonl"
BACKUP_DIR = HISTORY_DIR / "backups"
DEFAULT_KEEP_BACKUPS = 10   # M2c: book ごとにこの世代数を超えたら古い順に削除する

# ★ W8b 項目6: グローバル run ロック。基盤の LibreOffice(basrun 経由)が単一インスタンス
#   (port 2002)前提のため、ブック単位でなく `ailine run` 全体で1本にする。
RUN_LOCK_FILE = HISTORY_DIR / "run.lock"
# ★ #17: 「同じ依頼で前に通った op」を探して遡る件数。全件読むと履歴が育つほど重くなる。
HISTORY_RECALL_MAX = 500
RUN_LOCK_STALE_SECONDS = 30 * 60   # 30分超のロックは stale とみなして奪取する

# ★ W10a 項目2: 既定変更(W8b-2・原本直接反映)の一度きり告知。marker ファイルの有無だけで
#   判定する（history 等には依存しない・単純に「見せたか」の1ビット）。
NOTICE_V2_FILE = HISTORY_DIR / "notice_v2_shown"
NOTICE_V2_TEXT = (
    "★ このバージョンから、既定で原本に直接反映します（自動バックアップ+undo つき）。"
    "従来のコピー方式は --copy。この告知は一度だけ表示されます"
)

# ★ A': 用語集（税率等の「現場の取り決め値」）。グローバルのみ・ブック別上書きは作らない
#   （YAGNI・受信ファイル注入の予防。サイドカー自動読みは絶対にしない）。
VOCAB_FILE = HISTORY_DIR / "vocab.json"
DEFAULT_VOCAB_MAX_ENTRIES = 200     # 個人利用の上限。無制限にしない
DEFAULT_VOCAB_MAX_TERM_LEN = 40     # 語（キー）の最大長

# ★ W10 便A: 別名ストア（言い回し → op 名）。vocab.json とは別ファイル
#   （load_vocab は float 以外を黙って捨てる設計のため、op 名の文字列は同居できない
#   ── これは vocab 側の設計を歪めない正しい判断・REVIEW-20260822-w10-architect.md 3-2）。
ALIASES_FILE = HISTORY_DIR / "aliases.json"

# ★ 誤分類の実例台帳センサ: vocab_miss と同じ需要センサ方式（記録するだけ・分析/提案/表示は
#   作らない）。破壊の関所で N・undo の2点だけを容疑として拾う（成功 run は何も書かない）。
MISCLASS_FILE = HISTORY_DIR / "misclass.jsonl"




def _find_basrun_path() -> Path | None:
    """basrun.py の場所。環境変数 BASRUN > ailine と並びの checkout の順で探す。
       見つからなければ None（sys.exit しない版。doctor から非致命的に使う）。"""
    env = os.environ.get("BASRUN")
    if env:
        p = Path(env)
        return p if p.exists() else None
    # ★ wheel 化（2026-08-23）で HERE は src/ailine（install 後は site-packages/ailine）に
    #   なった。「並びの checkout」は repo から動かす場合の便宜なので、祖先を数段さかのぼって
    #   探す（src/ailine → src → <repo> → <repo の親> の順で C:\Dev\basrun 等に届く）。
    #   install した利用者にとっての正規の指定は環境変数 BASRUN（上で処理済み）。
    for base in (HERE, *HERE.parents[:3]):
        for name in ("basrun", "nagi-bas"):  # 公開 repo 名 / 作者ローカルの旧ディレクトリ名
            p = base / name / "basrun.py"
            if p.exists():
                return p
    return None


def basrun_path() -> Path:
    """basrun.py の場所。無ければ理由つきで落とす（run から使う致命版）。"""
    p = _find_basrun_path()
    if p is None:
        exit_environment("basrun.py が見つからない: 環境変数 BASRUN にパスを指定するか、"
                 "https://github.com/namakoo-dev/basrun を clone して"
                 "ailine と同じ階層に置く")
    return p


# ---------------------------------------------------------------------------
# 進捗表示（M1: 生成中の完全沈黙を解消。凝った演出はしない — Windows コンソール安全第一）
# ---------------------------------------------------------------------------

def _fmt_elapsed(seconds: float) -> str:
    """経過秒を表示用に整形する（例: 12.34 → '(12.3s)'）。"""
    return f"({seconds:.1f}s)"


def progress_start(label: str) -> float:
    """進捗の開始行を stderr に出す（改行しない＝完了時に経過秒を追記する）。
       --json のときも常に stderr（stdout の機械出力を汚さない）。
       開始時刻(time.monotonic())を返す。"""
    print(label, end="", file=sys.stderr, flush=True)
    return time.monotonic()


def progress_end(start: float) -> None:
    """開始行に経過秒を追記して改行する。"""
    print(" " + _fmt_elapsed(time.monotonic() - start), file=sys.stderr)

CONTRACT = """あなたは LibreOffice Basic を書く。出力は .bas のコードだけ。説明・markdown 柵は禁止。

厳守する契約:
- 先頭に `Option VBASupport 1` と `Option Explicit`。
- 手続きは **ちょうど1つ**、必ず `Sub Run(oDoc As Object)` という名前と署名。
- `ThisComponent` は使わない。文書は引数 oDoc で受け取る。
- シート: `oDoc.Sheets.getByIndex(0)`。セルは 0 起点 `getCellByPosition(列, 行)`。
  値は `.getValue()/.setValue()`、文字は `.getString()/.setString()`。
- 数値書式は `oDoc.getNumberFormats()` の queryKey/addNew を取り `範囲.NumberFormat = nFmt`。
- 列を文字("A")で指す API は使わない（例外で静かに止まる）。数値の列番号だけ。
- ★ 新しいシートを作らず、既存シートの中の最小の変更で達成することを最優先する
  （依頼が明示的にシート新設・ピボットを求めていない限り）。
"""

# セルの状態を snapshot する際の使用範囲の上限（病的に巨大な文書での暴走を防ぐ）
MAX_ROWS = 1000
MAX_COLS = 64


# ---------------------------------------------------------------------------
# ローカル LLM（ollama）
# ---------------------------------------------------------------------------

def ollama_generate(model: str, messages: list, temperature: float = 0.2) -> str:
    body = json.dumps({
        "model": model, "messages": messages, "stream": False,
        "options": {"temperature": temperature, "num_predict": 1600, "num_ctx": 8192},
    }).encode()
    req = urllib.request.Request(f"{OLLAMA}/api/chat", data=body,
                                 headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=300) as r:
            d = json.load(r)
    except urllib.error.HTTPError as e:
        # ★ HTTPError は URLError のサブクラスなので先に拾う。
        #   404 は「繋がっているがモデルが無い」で、接続不能とは原因も対処も別。
        if e.code == 404:
            exit_environment(f"ollama にモデル '{model}' が見つからない (HTTP 404)。\n"
                     f"★ `ollama pull {model}` で取得してから再実行して。")
        exit_environment(f"ollama がエラーを返した ({OLLAMA}): HTTP {e.code} {e.reason}")
    except urllib.error.URLError as e:
        exit_environment(f"ollama に繋がらない ({OLLAMA}): {e}\n"
                 "★ `ollama serve` が動いているか確認。外部送信はしない設計。")
    return d.get("message", {}).get("content", "")


def load_refs(refs_dir: Path) -> str:
    """refs/*.bas を few-shot テキストに連結する。"""
    if not refs_dir.is_dir():
        return ""
    chunks = []
    for f in sorted(refs_dir.glob("*.bas")):
        chunks.append(f"--- 参考例: {f.stem} ---\n{f.read_text(encoding='utf-8').strip()}")
    if not chunks:
        return ""
    return "\n\nこれらは正しい書き方の参考（別タスク）:\n" + "\n".join(chunks) + "\n--- 参考ここまで ---\n"


def load_helpers(helpers_dir: Path) -> tuple:
    """helpers/*.bas を (プロンプト用カタログ, ファイル一覧) にする。
       ★ 難所（ソートの ContainsHeader 等）は人が検証したヘルパに閉じ込め、
          モデルには『Call で呼ぶだけ』させる。arcane 層の確度を上げる要。"""
    files = sorted(helpers_dir.glob("*.bas")) if helpers_dir.is_dir() else []
    if not files:
        return "", []
    srcs = "\n".join(f.read_text(encoding="utf-8").strip() for f in files)
    catalog = (
        "\n\n## 定義済みヘルパ（★ 呼ぶだけ・再定義しない）\n"
        "arcane な操作（並べ替え等）は、自分で書かず次のヘルパを使うこと。\n"
        "★ 呼び方は必ず `Call 名前(引数)` の形（Call を付ける。括弧つきで Call 無しは誤動作する）。\n"
        "★ ヘルパの中身は絶対に書き写すな（SummaryTable 等が長くても）。必ず `Call 名前(...)` の1行だけで呼ぶ。\n"
        "★ W10b: 依頼に無い操作のヘルパは絶対に呼ぶな（『何か効くかもしれない』で総当たりに"
        "色々呼ぶのは禁止。依頼を達成するのに要る分だけを呼ぶ）。\n"
        "★ headerRow 引数は見出し行（0起点）。見出しが物理1行目ならほぼ常に 0。\n"
        "例: 列0〜4の表で金額が列1なら、金額で降順に並べ替え"
        " → `Call SortByColumn(oDoc, 0, 4, 1, False)`（第2引数 lastCol=表の最終列）\n"
        "例: 金額(列1)の棒グラフ（項目名は列0）→ `Call InsertChart(oDoc, 0, 0, 1, \"bar\", 4)`"
        "（第3引数=項目名の列・第4引数=値の列・第5引数=\"bar\"/\"line\"/\"pie\"・"
        "第6引数=データ範囲の最終行(0起点・省略可・省略時は自動検出)）\n"
        "例: A1とB1を結合 → `Call MergeCells(oDoc, 0, 0, 1, 0)`\n"
        "例: 先頭データ行(2行目)の前に1行挿入 → `Call InsertRows(oDoc, 1, 1)`\n"
        "例: 表に罫線を引く → `Call DrawTableBorders(oDoc)`\n"
        "例: 各列の幅を内容に合わせる → `Call AutoFitColumns(oDoc)`\n"
        "例: C列(列2)に、商品名(列0)をキーに『単価表』から値を引く（VLOOKUP相当）"
        " → `Call VLookupFromTable(oDoc, 0, 0, 2, \"単価表\")`（参照表は 列0=キー・列1=値）\n"
        "例: 『ピボット』で部門(列0)ごとに金額(列1)を集計（本物の DataPilot・Excel で操作可）"
        " → `Call PivotSum(oDoc, 0, 1)`\n"
        "例: 『集計表／まとめ』を作る＝部門(列0)ごとの金額(列1)を見栄えのする普通の表に"
        "（罫線・カンマ・太字つき。★『ピボット』と明示されない集計は基本こちら）"
        " → `Call SummaryTable(oDoc, 0, 0, 1)`\n"
        "例: 見出し行(行0, 列0〜4)を太字に → `Call StyleBold(oDoc, 0, 0, 4, 0)`\n"
        f"--- 定義済み（この通り既に存在する。再定義するな）---\n{srcs}\n--- ここまで ---\n")
    return catalog, files


def extract_bas(text: str) -> str:
    m = re.search(r"```(?:\w+)?\s*(.*?)```", text, re.S)
    return (m.group(1) if m else text).strip()


def valid_signature(code: str) -> bool:
    return re.search(r"Sub\s+Run\s*\(\s*oDoc\s+As\s+Object\s*\)", code, re.I) is not None


#  ★ 開始の "Sub 名前(...)" だけを数える（"End Sub" 内の "Sub" を誤って開始として
#     二重カウントしないよう、識別子が続く形に絞る）。
_SUB_OPEN_RE = re.compile(r"\bSub\s+[A-Za-z_]\w*", re.I)
_ENDSUB_RE = re.compile(r"\bEnd\s+Sub\b", re.I)


def is_truncated_code(code: str) -> bool:
    """生成 Basic が途中で切断された兆候を、構造だけを見て検出する（M2a）。
       ★ 正しさは判定しない。CONTRACT は『手続きはちょうど1つ、必ず End Sub で閉じる』
       前提のため、Sub/End Sub の対応と『最後の行がちょうど End Sub か』だけを見れば、
       開き括弧や識別子で途中生成が止まった典型パターン（＝最後の行が End Sub でない）
       を漏れなく拾える。bad_signature（署名自体が無い）とは別の失敗分類。"""
    stripped = code.strip()
    if not stripped:
        return True
    sub_count = len(_SUB_OPEN_RE.findall(stripped))
    endsub_count = len(_ENDSUB_RE.findall(stripped))
    if sub_count == 0 or sub_count != endsub_count:
        return True
    last_line = stripped.splitlines()[-1].strip()
    return not bool(_ENDSUB_RE.fullmatch(last_line))


# ---------------------------------------------------------------------------
# 文書の説明と snapshot（検証の土台）
# ---------------------------------------------------------------------------

def describe_book(path: Path) -> str:
    wb = openpyxl.load_workbook(path, read_only=True)
    lines = [f"シート一覧: {wb.sheetnames}（1枚目 = {wb.sheetnames[0]!r}）"]
    ws = wb[wb.sheetnames[0]]
    nrow, ncol = ws.max_row or 0, ws.max_column or 0
    lines.append(f"1枚目のデータ範囲: 約 {nrow} 行 x {ncol} 列（列は 0 起点で 0..{max(ncol-1,0)}）。")
    headers = []
    for c in range(1, min(ncol, MAX_COLS) + 1):
        v = ws.cell(row=1, column=1 + (c - 1)).value
        if v not in (None, ""):
            headers.append(f"列{c-1}={v}")
    if headers:
        lines.append("行0(見出し): " + ", ".join(headers))
    lines.append("行1以降がデータ。")
    wb.close()
    return "\n".join(lines)


def book_columns(path: Path, header_rows: dict | None = None) -> dict:
    """全シートの見出し行を {シート名: [列名,...]} で返す。
       ★ describe_book は1枚目だけの人間可読版。こちらは M2b の翻訳・検証が使う
       機械可読の接地情報（列は最初の空欄で打ち切る＝連続した見出しだけを列とみなす）。
       ★ W3: header_rows({シート名: 見出し行(1起点)}) を渡すと、そのシートだけは
       物理1行目でなく指定行を見出しとして読む（StructDump の見出し検出結果を反映する）。
       未指定のシートは従来どおり1行目（後方互換・引数省略時は完全に旧挙動）。"""
    header_rows = header_rows or {}
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        hr = header_rows.get(name, 1)
        ncol = min(ws.max_column or 0, MAX_COLS)
        headers = []
        for c in range(1, ncol + 1):
            v = ws.cell(row=hr, column=c).value
            if v in (None, "") and hr > 1:
                # ★ W3: 多段見出し対策。子見出し行(hr>1)で空欄の列は、真上の行を遡って
                #   最初に見つかる非空値を引き継ぐ（D検体: 商品名は1行目だけにあり
                #   2行目(子見出し)の同じ列は空。無いと打ち切り誤検知で列挙が0件になる）。
                for up in range(hr - 1, 0, -1):
                    uv = ws.cell(row=up, column=c).value
                    if uv not in (None, ""):
                        v = uv
                        break
            if v in (None, ""):
                break
            headers.append(str(v))
        out[name] = headers
    wb.close()
    return out


def build_book_meta(path: Path, header_rows: dict | None = None) -> dict:
    """{"sheets": [...], "headers": {シート名: [列名,...]}, "header_rows": {シート名: 行(1起点)},
       "path": path}。M2b 翻訳・検証の接地情報。★ W3: header_rows 省略時は全シート1行目
       （旧挙動と同一）。★ operator10 ①④: "path" は実ファイルを読み直す必要がある機能
       （chart_data_last_row・column_is_all_numeric）専用の追加キー。手組みの book_meta
       （単体テスト・_SAMPLE_META 等）はこのキーを持たないため、それらの経路は
       従来どおり「実ファイルを読まない」判定に自然と後退する（後方互換）。"""
    header_rows = dict(header_rows or {})
    headers = book_columns(path, header_rows)
    resolved_header_rows = {name: header_rows.get(name, 1) for name in headers}
    return {"sheets": list(headers.keys()), "headers": headers,
            "header_rows": resolved_header_rows, "path": path}


# ---------------------------------------------------------------------------
# W3 Part1/2: StructDump（LibreOffice の目で構造を読む）+ 見出し行推定
#
# ★ 正規化パス(normalize_book)の LO 往復に同乗させる（二役）。normalize_book が実行する
#   マクロを「何もしない空マクロ」から「何もしないが構造をテキストへ書き出すマクロ」に
#   差し替えるだけで、追加の LO 起動は発生しない。
# ★ LibreOffice の Cursor.gotoEndOfUsedArea・図形/グラフ/DataPilot 個数は LO でしか
#   正確に取れない（openpyxl の ws.max_row は書式だけ残った幽霊セルで過大評価しうる）。
#   一方、行ごとの書式的特徴(太字数等)・結合範囲は、LO が保存し終えた同じファイルを
#   openpyxl で読むだけで求まる（もう一度 LO を起動しない）。
# ---------------------------------------------------------------------------

STRUCT_HEADER_SCAN_ROWS = 20   # 見出し検出に使う先頭行数（多段見出し・タイトル行を含めても十分な余裕）
STRUCTDUMP_FILENAME = "structdump.txt"   # Basic → Python の受け渡し用（生のタブ区切りテキスト）


def _structdump_macro(out_path: Path) -> str:
    """StructDump 用の Basic マクロ。文書には一切手を触れない（正規化パスの no-op 性質を
       保つ）。シート毎に「実使用範囲(Cursor.gotoEndOfUsedArea)・図形/グラフ/DataPilot 個数」
       だけをタブ区切りテキストで書き出す（Basic に JSON エンコードは無いため、
       機械可読 dict への組み立ては Python 側 build_struct_dump() が担う）。"""
    out_str = str(out_path).replace('"', '""')
    return f'''Option VBASupport 1
Option Explicit

Sub Run(oDoc As Object)
    Dim iFile As Integer, i As Integer, n As Integer
    Dim oSheet As Object, oCursor As Object, oAddr As Object
    iFile = FreeFile
    Open "{out_str}" For Output As #iFile
    n = oDoc.Sheets.Count
    For i = 0 To n - 1
        oSheet = oDoc.Sheets.getByIndex(i)
        oCursor = oSheet.createCursor()
        oCursor.gotoEndOfUsedArea(True)
        oAddr = oCursor.RangeAddress
        Print #iFile, "SHEET" & Chr(9) & oSheet.Name & Chr(9) & oAddr.StartColumn & Chr(9) _
            & oAddr.StartRow & Chr(9) & oAddr.EndColumn & Chr(9) & oAddr.EndRow & Chr(9) _
            & oSheet.DrawPage.Count & Chr(9) & oSheet.Charts.Count & Chr(9) & oSheet.DataPilotTables.Count
    Next i
    Close #iFile
End Sub
'''


def parse_structdump_raw(text: str) -> dict:
    """StructDump の生テキスト（"SHEET"行群・タブ区切り）を
       {シート名: {"used_range": {...0起点...}, "shapes":n, "charts":n, "datapilots":n}} にパースする。
       壊れた/欠けた行は無視する（安全側・落とさない）。"""
    sheets: dict = {}
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) != 9 or parts[0] != "SHEET":
            continue
        try:
            name = parts[1]
            sc, sr, ec, er = int(parts[2]), int(parts[3]), int(parts[4]), int(parts[5])
            shapes, charts, pivots = int(parts[6]), int(parts[7]), int(parts[8])
        except ValueError:
            continue
        sheets[name] = {
            "used_range": {"start_col": sc, "start_row": sr, "end_col": ec, "end_row": er},
            "shapes": shapes, "charts": charts, "datapilots": pivots,
        }
    return sheets


def _row_char_stats(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> dict:
    """StructDump の行ごとの書式的特徴（見出し検出のヒューリスティクスに足る最小限）。
       {行(1起点): {"nonempty": 非空セル数, "str": うち文字列セル数, "bold": 太字セル数}}。"""
    stats = {}
    for r in range(start_row, end_row + 1):
        nonempty = 0
        strcnt = 0
        boldcnt = 0
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v in (None, ""):
                continue
            nonempty += 1
            if isinstance(v, str):
                strcnt += 1
            if cell.font and cell.font.bold:
                boldcnt += 1
        stats[r] = {"nonempty": nonempty, "str": strcnt, "bold": boldcnt}
    return stats


def build_struct_dump(normalized_book: Path, workdir: Path) -> dict:
    """StructDump 本体。normalize_book() が同じ LO 往復で書き出した生ダンプ
       (workdir/STRUCTDUMP_FILENAME) をパースし、正規化済みブックを openpyxl で読んで
       行の書式的特徴・結合範囲を足した機械可読 dict にする（もう一度 LO を起動しない）。
       生ダンプが無い/壊れている場合はそのシートだけ openpyxl の推定で代用する
       （テストでの normalize_book 差し替え等・呼び出し側は既定 header_row=1 に退避できる）。
       戻り値: {"sheets": {シート名: {"used_range":{1起点}, "merges":[...], "shapes":n,
                "charts":n, "datapilots":n, "rows": {行(1起点): {...}}}}}"""
    dump_path = workdir / STRUCTDUMP_FILENAME
    raw_sheets: dict = {}
    if dump_path.exists():
        try:
            raw_sheets = parse_structdump_raw(dump_path.read_text(encoding="utf-8"))
        except Exception:
            raw_sheets = {}
    wb = openpyxl.load_workbook(normalized_book)
    out: dict = {"sheets": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        raw = raw_sheets.get(name)
        if raw is not None:
            ur = raw["used_range"]
            start_row, end_row = ur["start_row"] + 1, ur["end_row"] + 1     # 0起点→1起点
            start_col, end_col = ur["start_col"] + 1, ur["end_col"] + 1
        else:
            start_row, start_col = 1, 1
            end_row = min(ws.max_row or 1, MAX_ROWS)
            end_col = min(ws.max_column or 1, MAX_COLS)
        scan_end = min(end_row, start_row + STRUCT_HEADER_SCAN_ROWS - 1)
        rows = _row_char_stats(ws, start_row, scan_end, start_col, min(end_col, MAX_COLS))
        out["sheets"][name] = {
            "used_range": {"start_row": start_row, "end_row": end_row,
                            "start_col": start_col, "end_col": end_col},
            "merges": sorted(str(r) for r in ws.merged_cells.ranges),
            "shapes": raw["shapes"] if raw else 0,
            "charts": raw["charts"] if raw else 0,
            "datapilots": raw["datapilots"] if raw else 0,
            "rows": rows,
        }
    wb.close()
    return out


def detect_header_row(sheet_struct: dict) -> tuple:
    """(見出し行(1起点) or None, confident: bool)。
       ヒューリスティクス: 「上に結合セルやタイトル行があっても、列方向に複数(2以上)の
       非空文字列セルが並び、その行自体は非空セルが全部文字列（=見出しらしい）で、
       直下の行に文字列でない非空セル(数値等)が混ざる（=データが始まる）」行を強い候補とする。
       ★ D検体（2段見出し）対策: 親見出し行(結合)も str_count>=2 を満たしうるが、その直下も
       まだ見出し（型の混在なし）なので候補から外れる。子見出し行の直下でようやくデータの
       型混在が起き、そこで初めて候補になる（追加ルール不要でこの一般則から自然に解ける）。
       ★ フォールバック: 型混在チェックで候補0件（例: 全列が文字列のみの表）でも、
       「非空文字列セル2以上・行自体は純文字列」を満たす行が候補全体でちょうど1つなら
       それを確信とする（曖昧さが無いケース）。
       強い候補が0個か2個以上（=曖昧）なら (None, False) を返す＝呼び出し側は推測せず CLARIFY。"""
    rows = sheet_struct.get("rows", {})
    if not rows:
        return None, False
    sorted_rows = sorted(rows.keys())
    pure_str_rows = [r for r in sorted_rows
                      if rows[r]["str"] >= 2 and rows[r]["nonempty"] == rows[r]["str"]]

    def _mixture_below(r: int) -> bool:
        nxt = rows.get(r + 1)
        return nxt is not None and nxt["nonempty"] > nxt["str"]

    with_mixture = [r for r in pure_str_rows if _mixture_below(r)]
    # ★★ 2026-08-29（Namakoo が実測・基本操作が丸ごと止まった）: 「丸山工業／PCパーツ」
    #   だけ埋めた行を作ったら、そのシートで**何も**できなくなった
    #   （？ 見出しが何行目か分かりません）。
    #   ★ その行は「非空セルが全部文字列」で、下の行に数字がある ── 見出しの条件を
    #     そのまま満たしてしまう。候補が 2 つになって決められない。
    #   ★ 一般則で切れる: **本物の見出しは表の幅いっぱいに並ぶ**。
    #     途中まで入力した行は 1〜2 セルしか埋まっていない ── 幅が違う。
    #     幅が最大の候補が**ちょうど 1 つ**なら、それが見出し。
    #     同じ幅で並ぶ（見出しが 2 段・表が 2 つ縦に並ぶ等）なら、今までどおり決めない。
    if len(with_mixture) > 1:
        _widest = max(rows[r]["str"] for r in with_mixture)
        _top = [r for r in with_mixture if rows[r]["str"] == _widest]
        if len(_top) == 1:
            with_mixture = _top
    if len(with_mixture) == 1:
        return with_mixture[0], True
    if not with_mixture and len(pure_str_rows) == 1:
        return pure_str_rows[0], True
    # ★ 致命3(W10e) 項目①: 表の全列が文字列型（数値/日付列が一つも無い）の場合、
    #   「型混在」の手がかりが構造的に一度も発生しない（with_mixture は必ず空）ため、
    #   pure_str_rows が『先頭行から続く全データ行』になり、見出し行が一意に決まらず
    #   毎回 CLARIFY に落ちていた（実測: 氏名/部署/備考のようなテキストのみの表）。
    #   スキャン範囲内の非空行が『例外なく全部』文字列のみ（数値/日付の手がかりが
    #   一つも無い）と確認できた場合に限り、最初の候補行（先頭・str>=2）を見出しと
    #   みなす（型混在の手がかりが少しでもある通常のケースはこの分岐に来ない＝
    #   test_detect_header_row_ambiguous_two_equally_valid_candidates_is_not_confident 等の
    #   既存挙動は変えない）。
    if not with_mixture and pure_str_rows:
        no_type_mixture_possible = all(
            info["nonempty"] == info["str"] for info in rows.values() if info["nonempty"] > 0)
        if no_type_mixture_possible:
            return pure_str_rows[0], True
    return None, False


# ★ W8a 項目3: 旧文言は「答えて」と聞くだけで答え方(CLI で何を打てばいいか)が無い行き止まり
#   だった（architect 発見）。--header-row フラグの使い方まで添えて、次のコマンドが打てる形にする。
CLARIFY_HEADER_ROW_QUESTION = ("見出しが何行目か分かりません。"
                                "`--header-row 3` のように指定して再実行してください")


def resolve_header_rows(struct_dump: dict, sheets: list, target_sheet: str | None = None) -> tuple:
    """全シートの見出し行(1起点)を決める。(header_rows: {シート名: 行}, clarify_question|None)。
       ★ 挙動変更#2: 対象シート（target_sheet・省略時は1枚目＝旧挙動と同一）だけ StructDump の
       ヒューリスティクスで推定する（DSL 操作の書き込み対象はこのシート1枚だけのため）。
       他シート（LOOKUP_FILL の参照表等）は物理1行目を既定にする。
       ★ build_struct_dump は元々全シート分の rows(書式的特徴) を計算済み（_structdump_macro が
       `For i = 0 To n - 1` で全シートを走査する）。対象シートが1枚目でなくても StructDump 自体の
       やり直しは不要 — ここで見る辞書のキーを sheets[0] から target_sheet に差し替えるだけでよい。
       StructDump が無い（テストでの normalize_book 差し替え等）場合は全シート1行目のまま
       （旧挙動と同一・CLARIFY は出さない）。自信が持てない場合だけ対象シートについて
       CLARIFY 質問を返す（推測で進まない）。"""
    header_rows = {s: 1 for s in sheets}
    if not sheets:
        return header_rows, None
    sd_sheets = (struct_dump or {}).get("sheets", {})
    target = target_sheet if target_sheet in sheets else sheets[0]
    info = sd_sheets.get(target)
    if info is None:
        return header_rows, None
    row, confident = detect_header_row(info)
    if confident:
        header_rows[target] = row
        return header_rows, None
    return header_rows, CLARIFY_HEADER_ROW_QUESTION


def _struct_dump_info_missing(struct_dump: dict, sheets: list, target_sheet: str | None) -> bool:
    """★ operator8 ③（測定器の修正・2026-08-22）: resolve_header_rows が『StructDump に
       対象シートの情報が無ければ無言で1行目とみなす』フォールバック（同関数 docstring・
       上の `if info is None: return header_rows, None` の分岐）を踏んだかどうかを、
       resolve_header_rows 自身の戻り値（2-tuple・複数の凍結テストが tuple 等価で固定して
       いるため形を変えられない）に触れずに、呼び出し側で独立に判定する。
       ★ 実機で起きる形: LO の一時不調等で build_struct_dump が対象シート分の情報を
       持たずに戻る（operator の実物ファイルで再現した実事故の機構そのもの）。
       判定条件は resolve_header_rows の info 取得と完全に同じにする（ここだけ緩めたり
       厳しくしたりしない）。"""
    if not sheets:
        return False
    target = target_sheet if target_sheet in sheets else sheets[0]
    sd_sheets = (struct_dump or {}).get("sheets", {})
    return sd_sheets.get(target) is None


def _scan_first_rows(path: Path, sheet_name: str, max_rows: int = 10) -> dict:
    """★ operator8 ③: 対象シートの先頭 max_rows 行を実ファイルから読み、
       {行番号(1起点): 非空セル値を文字列化した集合} を返す（空行は含めない）。
       見出し行の検出が外れた/使えなかった場合の敗者復活（列解決が失敗した時だけ参照）
       専用のデータ ―― A' 原則: 実在するセル値だけが材料（LLM は使わない）。
       読めない/シートが無い等はどんな理由でも空 dict（呼び出し側は『従来のまま』に
       フォールバックする）。"""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        out: dict = {}
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows), start=1):
            cells = {str(c.value) for c in row if c.value not in (None, "")}
            if cells:
                out[r] = cells
        return out
    except Exception:
        return {}
    finally:
        wb.close()


def _header_row_hint_for_missing_col(book_meta: dict, sheet_name: str, raw_col) -> str | None:
    """★ operator8 ③: 列解決が失敗したときの敗者復活。book_meta["_row_scan"]
       （_translate_and_dispatch が実ファイルから積む・単体テストで直接 verify_dsl_args を
       呼ぶ場合は無い＝常に None で従来のまま）を見て、要求された列名 raw_col と完全一致する
       セルが**現在の見出し行以外**の行Rに見つかったら、--header-row R への導線を返す。
       見つからなければ None（呼び出し側は従来の「列『X』がありません。ある列:…」のまま）。
       ★ 走査は完全一致のみ（部分一致で誤誘導しない）。"""
    raw = str(raw_col) if raw_col not in (None, "") else ""
    if not raw:
        return None
    scan = (book_meta.get("_row_scan") or {}).get(sheet_name) or {}
    if not scan:
        return None
    current = (book_meta.get("header_rows") or {}).get(sheet_name, 1)
    for r in sorted(scan):
        if r == current:
            continue
        if raw in scan[r]:
            return (f"列『{raw}』は{r}行目に見出しがあるようです。"
                    f"`--header-row {r}` のように指定して再実行してください")
    return None


def _charts_count(path: Path) -> int:
    try:
        with zipfile.ZipFile(path) as z:
            return sum(1 for n in z.namelist()
                       if "chart" in n.lower() and n.lower().endswith(".xml")
                       and "/charts/chart" in n.lower())
    except Exception:
        return 0


def _chart_paths(path: Path) -> frozenset:
    """★ 致命④(2026-08-23レビュー): _charts_count の隣に置く集合版（同じタイミング・
       同じ判定基準で数える）。before との差分で「今回増えた1個」を同定する材料
       （check_chart_series の before_chart_paths に渡す）。"""
    try:
        with zipfile.ZipFile(path) as z:
            return frozenset(n for n in z.namelist()
                             if "chart" in n.lower() and n.lower().endswith(".xml")
                             and "/charts/chart" in n.lower())
    except Exception:
        return frozenset()


def snapshot(path: Path) -> dict:
    """文書の状態を取る。★ 値/数値書式/背景色/太字 に加え、罫線・結合・列幅・行高・
       水平配置も捉える。これで『書式のみ・罫線のみ・列幅のみ・結合のみ・中央揃えのみ』
       の変更も『変化した』と検出でき、no-op 誤検出（＝効いているのに失敗扱い）を防ぐ。"""
    wb = openpyxl.load_workbook(path)
    # ★ 止血3: truncated は「この snapshot が MAX_ROWS/MAX_COLS で切り詰められたか」を
    #   保持する。diff_snapshots 後の表示で「先頭1000行しか見せていない」ことを正直に
    #   注記するために使う（bench/realworld/BASELINE.md の B 検体所見の根治）。
    # ★ operator 指摘②(2026-08-19): true_rows はシート毎の実際の行数（切り詰め前）。
    #   count_reconciliation が「データ 999 行のうち…」と 嘘の分母 を出していた根治に使う。
    #   値は既にこのループで計算していた（捨てていただけ）。
    snap = {"sheets": list(wb.sheetnames), "charts": _charts_count(path),
            # ★ operator10 ②: シート単位のグラフ数（_changed_sheets がセル以外の変化を
            #   このシートの「変わった」に数えるための材料）。
            "chart_counts": charts_by_sheet(path),
            "cells": {}, "merges": {}, "colw": {}, "rowh": {}, "truncated": False,
            "true_rows": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        true_nrow = ws.max_row or 0
        snap["true_rows"][name] = true_nrow
        true_ncol = ws.max_column or 0
        if true_nrow > MAX_ROWS or true_ncol > MAX_COLS:
            snap["truncated"] = True
        nrow = min(true_nrow, MAX_ROWS)
        ncol = min(true_ncol, MAX_COLS)
        for r in range(1, nrow + 1):
            for c in range(1, ncol + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                fill = None
                if cell.fill is not None and cell.fill.patternType:
                    fill = str(cell.fill.start_color.rgb)
                bold = bool(cell.font.bold) if cell.font else False
                numfmt = cell.number_format
                bd = cell.border
                bsig = (bd.left.style, bd.right.style, bd.top.style, bd.bottom.style) if bd else None
                if bsig == (None, None, None, None):
                    bsig = None
                # ★ 水平配置（中央揃え等）。既定は None/'general' 扱い。
                align = cell.alignment.horizontal if cell.alignment else None
                if align == "general":
                    align = None
                if (val in (None, "") and fill is None and not bold
                        and numfmt == "General" and bsig is None and align is None):
                    continue
                snap["cells"][f"{name}!{r},{c}"] = (val, numfmt, fill, bold, bsig, align)
        snap["merges"][name] = sorted(str(rng) for rng in ws.merged_cells.ranges)
        snap["colw"][name] = {k: round(d.width, 2) for k, d in ws.column_dimensions.items() if d.width}
        snap["rowh"][name] = {k: round(d.height, 2) for k, d in ws.row_dimensions.items() if d.height}
    wb.close()
    return snap


# ★ snapshot() の cell tuple の並びと対応する日本語ラベル（M1: 生 tuple 漏れの解消）。
_CELL_FIELD_LABELS = ("値", "数値書式", "背景", "太字", "罫線", "配置")
# snapshot() は完全に既定状態のセルを cells 辞書に入れない（skip 条件参照）。
# 片側が辞書に無い(=完全既定)ときは、この既定 tuple で埋める（全 None で埋めると
# 数値書式が『None→General』という偽の差分を作ってしまう＝実測で見つけた罠）。
_CELL_DEFAULT = (None, "General", None, False, None, None)


def _fmt_cell_value(v) -> str:
    """人が読める値表示（None は '(空)'、文字列はクォート）。"""
    if v is None:
        return "(空)"
    if isinstance(v, str):
        return f"'{v}'"
    return str(v)


def describe_cell_change(before: tuple | None, after: tuple | None) -> str:
    """(値,数値書式,背景,太字,罫線,配置) の tuple 差分を、★ 変わったフィールドだけ
       日本語ラベルで列挙する人間可読の文字列にする（tuple の生表示は出さない）。
       例: 値 'りんご'→'リンゴ', 太字 (空)→True"""
    b = before if before is not None else _CELL_DEFAULT
    a = after if after is not None else _CELL_DEFAULT
    parts = []
    for label, bv, av in zip(_CELL_FIELD_LABELS, b, a):
        if bv != av:
            parts.append(f"{label} {_fmt_cell_value(bv)}→{_fmt_cell_value(av)}")
    return ", ".join(parts) if parts else "(差分なし)"


def diff_snapshots(before: dict, after: dict) -> tuple:
    """(changed: bool, lines: [str])。人が読める変更点も返す。
       セル値/書式/色/太字/罫線・結合・列幅・行高・シート・グラフの変化を見る。"""
    lines = []
    added = [s for s in after["sheets"] if s not in before["sheets"]]
    removed = [s for s in before["sheets"] if s not in after["sheets"]]
    if added:
        lines.append(f"＋シート追加: {added}")
    if removed:
        lines.append(f"－シート削除: {removed}")
    if after["charts"] != before["charts"]:
        lines.append(f"＊グラフ数: {before['charts']} → {after['charts']}")

    # 結合セル
    merge_changes = 0
    for name in after.get("merges", {}):
        b = set(before.get("merges", {}).get(name, []))
        a = set(after["merges"][name])
        for m in sorted(a - b):
            merge_changes += 1; lines.append(f"＋結合 {name}!{m}")
        for m in sorted(b - a):
            merge_changes += 1; lines.append(f"－結合解除 {name}!{m}")

    # 列幅・行高（変わったシート数で示す）
    dim_changes = 0
    for key, label in (("colw", "列幅"), ("rowh", "行高")):
        for name in after.get(key, {}):
            if before.get(key, {}).get(name) != after[key].get(name):
                dim_changes += 1; lines.append(f"＊{label}変更: {name}")

    # セル（値/書式/色/太字/罫線）－ シートごとに自前の見出し→明細（他の種別と揃える）
    keys = set(before["cells"]) | set(after["cells"])
    changed_by_sheet: dict = {}
    cell_changes = 0
    for k in sorted(keys):
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        if b == a:
            continue
        cell_changes += 1
        sheet, rc = k.split("!", 1)
        r_str, c_str = rc.split(",")
        ref = _cell_ref(int(r_str), int(c_str))
        changed_by_sheet.setdefault(sheet, []).append(f"  {ref}: {describe_cell_change(b, a)}")

    shown = 0
    for sheet, clines in changed_by_sheet.items():
        if shown >= 12:  # 全部は出さない。多いときは件数で示す
            break
        lines.append(f"＊セル値変更: {sheet}")
        for cl in clines:
            if shown >= 12:
                break
            lines.append(cl)
            shown += 1
    if cell_changes > shown:
        lines.append(f"  …ほか {cell_changes - shown} セル")

    changed = bool(added or removed or (after["charts"] != before["charts"])
                   or merge_changes or dim_changes or cell_changes)
    return changed, lines


def _truncation_notice(before: dict, after: dict, exhaustive_postcondition: bool) -> str | None:
    """★ 止血3: before/after どちらかの snapshot が MAX_ROWS/MAX_COLS で切り詰められて
       いたら、無言で切らず正直な1行を返す（bench/realworld/BASELINE.md の B 検体所見）。
       ★ 経路で文言を出し分ける:
       - exhaustive_postcondition=True（DSL経路・cmd_run_dsl / cmd_run_plan の DSL 段）:
         事後条件チェッカーは openpyxl で out ファイルを直接開き _scan_last_row で全行を
         走査する（snapshot() の MAX_ROWS とは無関係）。表示だけが切り詰められている。
       - exhaustive_postcondition=False（FREEFORM経路）: no-op ガード・advisories も
         snapshot() 頼みなので、検証自体も先頭1000行までしか見ていない。
       切り詰めが無ければ None。"""
    if not (before.get("truncated") or after.get("truncated")):
        return None
    if exhaustive_postcondition:
        return f"（表示は先頭 {MAX_ROWS} 行の変化のみ。検証・適用は全行に対して実施）"
    return f"（表示は先頭 {MAX_ROWS} 行の変化のみ。検証も先頭 {MAX_ROWS} 行のみ）"


# ---------------------------------------------------------------------------
# 疑わしい変化の機械検出（M2a: 冷間再監査が実測した「✓ の下の失敗」への対抗）
#
# ★ 全部「差分の後の助言」であってブロックしない。誤検知を恐れて条件は保守的に
#   （両条件とも『変更セルの全部がそれに該当』した時だけ発火）。
# ---------------------------------------------------------------------------

def _changed_cells(before: dict, after: dict) -> list:
    """(sheet, row, col) のリスト。値/書式/色/太字/罫線/配置のいずれかが変わったセル。"""
    keys = set(before["cells"]) | set(after["cells"])
    out = []
    for k in keys:
        if before["cells"].get(k) != after["cells"].get(k):
            sheet, rc = k.split("!", 1)
            r_str, c_str = rc.split(",")
            out.append((sheet, int(r_str), int(c_str)))
    return out


def _value_changed_cells(before: dict, after: dict) -> list:
    """(sheet, row, col) のリスト。★ M2c: 値(idx0)が実際に変わったセルだけ
       （書式(罫線・中央揃え等)だけが変わったセルは含めない）。
       幽霊データ/一様埋め検出を『値変更』の部分集合だけで評価するための土台。
       ★ 冷間再監査3回目の実測: 罫線+中央揃え+0埋めが混在すると、書式だけ変わった
       セル（値は不変）が uniform 判定の対象に紛れ込み、一様埋めが見逃されていた。"""
    keys = set(before["cells"]) | set(after["cells"])
    out = []
    for k in keys:
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        b_val = b[0] if b is not None else None
        a_val = a[0] if a is not None else None
        if b_val != a_val:
            sheet, rc = k.split("!", 1)
            r_str, c_str = rc.split(",")
            out.append((sheet, int(r_str), int(c_str)))
    return out


def _used_range(before: dict, sheet: str) -> tuple | None:
    """原本(before snapshot)においてそのシートで実際に値が入っていたセルの矩形
       (min_row, max_row, min_col, max_col)。★ 近似: 値(idx0)が非空のセルだけを対象に
       する（書式のみのセルは『データ』に数えない）。値が1つも無ければ None。"""
    rows, cols = [], []
    prefix = sheet + "!"
    for k, v in before["cells"].items():
        if not k.startswith(prefix):
            continue
        if v[0] in (None, ""):
            continue
        r_str, c_str = k[len(prefix):].split(",")
        rows.append(int(r_str)); cols.append(int(c_str))
    if not rows:
        return None
    return (min(rows), max(rows), min(cols), max(cols))


def detect_ghost_data(before: dict, after: dict, *, new_col_letter: str | None = None,
                       new_row_at_end: bool = False) -> str | None:
    """★ 幽霊データ検出: 変更セルが全部、原本の使用範囲（データが存在した矩形）の
       外に集中している場合だけ疑わしい旨を返す。1セルでも範囲内なら何も言わない
       （保守的。使用範囲が不明なシートが混ざる場合も判定を保留する）。
       ★ M2c: 判定対象は『値変更』の部分集合だけ（書式のみの変更は無視・保守性は部分集合内で維持）。
       ★ W6: 実行中に新規作成されたシート（before["sheets"] に無い）のセルはここでは無視する
       （new_sheet_advisories が別途担当）。以前は『使用範囲が不明なシートが1つでも混ざると
       関数全体が判定を保留する』実装だったため、AGGREGATE/CHART/PivotSum のように新規シート
       を作る操作が絡むたび、他シートの本当のゴーストデータ検出まで丸ごと素通りしていた
       （監査実測。旧シート範囲が不明＝原本に無い＝新規シート、の場合に限って除外することで
       既存シートに対する検出力は変えずに直す）。
       ★ C9: new_col_letter（呼び出し側が op の宣言（OP_WRITE_TARGET）から求めた、今回
       新規に作る列の文字）が与えられ、かつ検出範囲が丸ごとその1列に収まる場合は、
       警告でなく中立表示を返す（旧 _neutralize_new_column_ghost_warning が『出してから
       打ち消す』後処理でやっていたのと同じ判定を、発生源で先取りする）。
       ★ 単位C(D10): new_row_at_end（op の宣言 writes に WRITE_NEW_ROW_AT_END がある＝
       データ最終行の下に行を足す op）が真で、検出セルが全部『原本の最終行より下・かつ
       原本の列範囲の中』に収まる場合も同じく中立表示にする。合計行は定義上ずっと原本の
       使用範囲の外に出るので、この誤警報は毎回・確実に再現していた（APPEND_TOTAL）。
       列の外（右）へ出たセルが混ざる場合は中立化しない＝検出力は保守的に温存する。"""
    changed = _value_changed_cells(before, after)
    if not changed:
        return None
    outside = []
    below_only = True   # ★ 単位C(D10): 範囲外セルが全部「原本の最終行より下・列範囲の中」か
    for sheet, r, c in changed:
        if sheet not in before["sheets"]:
            continue   # ★ W6: 新規作成されたシート＝ここでの判定対象外
        rect = _used_range(before, sheet)
        if rect is None:
            return None  # このシートの原本データ範囲が不明 → 判定を保留
        min_r, max_r, min_c, max_c = rect
        if min_r <= r <= max_r and min_c <= c <= max_c:
            return None  # 1つでも範囲内 → 発火しない
        if not (r > max_r and min_c <= c <= max_c):
            below_only = False
        outside.append((r, c))
    if not outside:
        return None   # 変更が全部、新規シートのセルだけだった
    rows = [r for r, _ in outside]
    cols = [c for _, c in outside]
    min_c, max_c = min(cols), max(cols)
    if new_col_letter is not None:
        # ★ 2026-08-24: 1 列だけでなく**列の集合**を受ける（SPLIT_CELL は N 列作る）。
        #   範囲外セルの列が全部その集合に収まっていれば、宣言どおりの効果＝誤警報。
        raw = new_col_letter if isinstance(new_col_letter, (set, frozenset, list, tuple))             else [new_col_letter]
        declared = set()
        for letter in raw:
            try:
                declared.add(column_index_from_string(letter))
            except (ValueError, AttributeError):
                pass
        if declared and set(cols) <= declared:
            return "（新規列の追加は意図どおりです）"
    if new_row_at_end and below_only:
        return "（表の末尾への追記は意図どおりです）"
    top_left = _cell_ref(min(rows), min_c)
    bot_right = _cell_ref(max(rows), max_c)
    span = top_left if len(outside) == 1 else f"{top_left}:{bot_right}"
    return f"★ 疑わしい: 変更が元データの範囲外です（{span}）"


# ★ W10b 項目4a(摩擦): 「★疑わしい: 変更が元データの範囲外」が、新規列の作成という
#   *意図した*操作でも毎回出ていた実測所見への対応。DSL 経路(COMPUTE_COLUMN・target
#   無指定=新規列作成)に限り、detect_ghost_data が指す範囲が丸ごとその新規列に収まって
#   いれば中立表示に落とす（新規列は原本の使用範囲外に出るのが当然＝誤警報）。
# ★ W10d: COMPUTE_COLUMN 専用の if だったものを OP_WRITE_TARGET の宣言駆動へ一般化した
#   （査定で名指しされたオオカミ少年: LOOKUP_FILL が新規列を作っても同じ誤警報が出ていた）。
#   op ごとの if は増やさない — OP_WRITE_TARGET に既に登録済みの「書き込み先列」宣言を
#   そのまま読み、その列が対象シートの既存見出しに無ければ『新規列を作る効果』とみなす
#   （COMPUTE_COLUMN の target 無指定＝キーが空、LOOKUP_FILL の target_col 有指定だが
#   対象シートにまだ無い列名＝どちらも実行時にその名前で新しい列を作るという同じ意味）。
# ★ C9: 以前は detect_ghost_data がまず警告を出し、_neutralize_new_column_ghost_warning が
#   advisories の該当行を後から中立表示に置き換えていた（『出してから打ち消す』）。W10c/W10d で
#   宣言（OP_WRITE_TARGET）が入力として取れるようになったので、detect_ghost_data 自身が
#   new_col_letter を受け取り、発生源で中立表示を返す形にした（_neutralize_new_column_ghost_warning
#   は削除。判定条件・出力文言は一切変えていない＝ゴールデン差分ゼロで確認済み）。


def _declared_new_column_letter(op: str, resolved: dict, book_meta: dict) -> str | None:
    """op が今回、宣言済みの効果として新規列を作るなら、その列の文字（"C" 等）を返す。
       作らない/対象シートが分からない場合は None。
       ★ W10d 番人の土台: OP_WRITE_TARGET だけを見る。新しい op を足しても
       OP_WRITE_TARGET へ登録さえすれば、ここへの追記なしで正しく判定される
       （test_op_write_target_declares_all_ops が登録漏れ自体を防ぐ）。
       ★ 単位C: 宣言が領域を持つようになったので、まず writes に『新規列』があるかを見る
       （旧形は col_key の有無だけが手掛かりで、新規列を作らない op でも列名が resolved に
       無ければ列文字を返してしまう形だった）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or WRITE_NEW_COLUMN not in write_target.writes:
        return None
    col_key, sheet_key = write_target.col_key, write_target.sheet_key
    if not col_key:
        return None
    if sheet_key:
        sheet = resolved.get(sheet_key)
    else:
        # ★ 挙動変更#2: sheets[0] 決め打ちをやめ、verify_dsl_args が一箇所で決めた
        # resolved["_target_sheet"] を読む（旧値と後方互換のフォールバック付き）。
        sheets = book_meta.get("sheets") or []
        sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    if not sheet:
        return None
    headers = book_meta.get("headers", {}).get(sheet, [])
    col_name = resolved.get(col_key)
    if col_name and col_name in headers:
        return None   # 既存列への書き込み（上書き側の話・新規列ではない）
    new_col_idx = len(headers)   # 0起点・新規列は既存見出しの直後
    return get_column_letter(new_col_idx + 1)


def resolve_new_column_placement(op: str, resolved: dict, book_meta: dict,
                                  task: str, sheet: str | None) -> dict | None:
    """新しい列を作る op に、依頼文の**位置の言い回し**を効かせる（横断層）。

    ★ Namakoo「『〜の右側に』『〜と〜の間に』などは頻出だから全ての操作で有効に」。
      そのとおりで、位置は op の性質でなく**依頼文の性質**。op ごとに if を書くと、
      op が増えるたびに配線が要る（今日 4 回踏んだ形）。宣言を読んで 1 箇所で解く。
    ★ 返り値: {"_move_new_col_to": 0起点の目的地, "_new_col_from": 0起点の作られる場所,
               "_at_basis": 根拠の文}。動かす必要が無ければ None。
    ★ 対象は「1 本の新しい列を右端に作る」と宣言した op だけ:
      - ADD_COLUMN は自分で位置を決める（col_index_key を持つ）ので対象外
      - SPLIT_CELL は 1 回で N 本作る（cols_key）ので対象外 ── 複数本の移動は
        まだ測っていない。**測っていないものを黙って動かさない。**
    """
    wt = OP_WRITE_TARGET.get(op)
    if not wt or WRITE_NEW_COLUMN not in wt.writes:
        return None
    if wt.col_index_key or wt.cols_key:
        return None
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    if not headers:
        return None
    # 既存列への書き込みなら新しい列は生まれない（位置の話は起きない）
    if wt.col_key and str(resolved.get(wt.col_key) or "") in headers:
        return None
    at, note = resolve_col_anchor(task, headers)
    if at is None:
        return None
    src0 = len(headers)          # 0 起点・新規列は既存見出しの直後（右端）
    to0 = int(at) - 1
    if to0 == src0:
        return None              # もともとそこに出来る（動かす必要が無い）
    return {"_move_new_col_to": to0, "_new_col_from": src0, "_at_basis": note}


def _letterlike_header_columns(book_meta: dict | None) -> set:
    """英字だけの**見出し名**（URL / ID / AB …）を、列文字として読んだ時の列番号の集合。

    ★ なぜ要るか（2026-08-24 実測）: 依頼文の「URL列を…」を列文字参照と読んで
      column_index_from_string("URL")=14676 列目とし、「『列URL』は変更されていません」
      という偽の ⚠ を出していた。決裁③でこれが ✓ を △ へ降格させる。op を問わず、
      英字の見出しを持つブック全部で起きる。
    ★ 名前つき列の変更を見る仕組みは別に無いので、ここは「言えないことは言わない」に倒す
      （黙るのであって、正しいと主張するのではない）。"""
    cols = set()
    if not book_meta:
        return cols
    for headers in (book_meta.get("headers") or {}).values():
        for h in headers or ():
            name = str(h).strip()
            if not name or not name.isascii() or not name.isalpha():
                continue
            try:
                cols.add(column_index_from_string(name.upper()))
            except ValueError:
                pass
    return cols


def _declared_kept_subject_cols(op: str, resolved: dict, book_meta: dict | None) -> set:
    """対象列を**意図して変えない**と宣言している op の、その列番号（1起点）の集合。
       ★ 宣言駆動（OP_WRITE_TARGET.keeps_subject）── op ごとの if を増やさない。
       SPLIT_CELL は元の列を残すのが契約なので、「言及された列が変更されていません」は
       誤警報であり、それが ✓ を △ へ降格させていた（実測 2026-08-24）。"""
    wt = OP_WRITE_TARGET.get(op)
    if not wt or not wt.keeps_subject or not resolved or not book_meta:
        return set()
    sheets = book_meta.get("sheets") or []
    sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    headers = book_meta.get("headers", {}).get(sheet, []) if sheet else []
    cols = set()
    for key, _kind in OP_SUBJECT_SLOTS.get(op, ()):  # 対象スロット＝この op の主語の列
        name = resolved.get(key)
        for one in (name if isinstance(name, list) else [name]):
            if one in headers:
                cols.add(headers.index(one) + 1)
    return cols


def _declared_new_column_letters(op: str, resolved: dict, book_meta: dict) -> set:
    """今回の宣言で新しく作られる列の**全部**の列文字。単数宣言(col_key)と
       複数宣言(cols_key)の両方を 1 つの集合にまとめる ── 免除の判定を「1 列だけ」に
       縛らないため（SPLIT_CELL は 1 回で N 列作る）。"""
    letters = set()
    single = _declared_new_column_letter(op, resolved, book_meta)
    if single:
        letters.add(single)
    wt = OP_WRITE_TARGET.get(op)
    # ★ 2026-08-27: 位置を**機械が決めた**op は、右端決め打ちではなくその位置が宣言。
    #   どのキーに入っているかは宣言(col_index_key)が持つ ── op 名の if は増やさない。
    if wt and wt.col_index_key and (resolved or {}).get(wt.col_index_key):
        letters.add(get_column_letter(int(resolved[wt.col_index_key])))
    if wt and wt.cols_key and resolved:
        names = resolved.get(wt.cols_key) or []
        sheets = book_meta.get("sheets") or []
        sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
        headers = book_meta.get("headers", {}).get(sheet, []) if sheet else []
        base = len(headers)   # 0起点・新規列は既存見出しの直後から並ぶ
        for k, name in enumerate(names):
            if name not in headers:
                letters.add(get_column_letter(base + k + 1))
    return letters


def _cell_key_col(key) -> int:
    """snapshot の cells のキー（"シート!行,列"）から列番号（1 起点）を取る。
       読めない形なら 0（＝どの宣言にも当たらない＝疑いを消さない・安全側）。"""
    try:
        return int(str(key).rsplit("!", 1)[1].split(",")[1])
    except (IndexError, ValueError):
        return 0


def detect_uniform_fill(before: dict, after: dict, single_cell: bool = False,
                         new_col_letter=None, proved: bool = False) -> str | None:
    """★ 一様埋め検出: 変更セルの全部で『変化前が空欄』かつ『変化後が全部同一値』
       （特に 0/空文字）の場合だけ疑わしい旨を返す（保守的）。
       ★ M2c: 判定対象は『値変更』の部分集合だけ（罫線・中央揃えなど書式のみが変わった
       セルは対象外にする — 混ざっていると後方の値だけ均一でも見逃していた実測不具合の修正）。"""
    # ★ 2026-08-27: 宣言済みの新しい列に書いた分は疑わない（detect_ghost_data と同じ規則）。
    #   ADD_COLUMN は挿した列の**見出し 1 セル**を書く ── それは「一括書き込み」ではない。
    # ★ 2026-08-27: 事後条件が「どのセルが変わるべきか」を両方向で証明する op では、
    #   この助言は何も足さない（証明の方が強い）── 助言は証明が届かない所にだけ要る。
    if proved:
        return None
    raw = ([new_col_letter] if isinstance(new_col_letter, str)
            else list(new_col_letter or ()))
    declared = set()
    for letter in raw:
        try:
            declared.add(column_index_from_string(letter))
        except (ValueError, AttributeError):
            pass
    keys = set(before["cells"]) | set(after["cells"])
    after_vals = []
    for k in keys:
        if declared and _cell_key_col(k) in declared:
            continue
        b = before["cells"].get(k)
        a = after["cells"].get(k)
        b_val = b[0] if b is not None else None
        a_val = a[0] if a is not None else None
        if b_val == a_val:
            continue  # 値は変わっていない（書式だけの変更は対象外）
        if b_val not in (None, ""):
            return None  # 変化前が空欄でないセルが1つでもある → 発火しない
        after_vals.append(a_val)
    if not after_vals:
        return None
    if len(set(after_vals)) != 1:
        return None
    val = after_vals[0]
    if val in (None, ""):
        return None  # 空欄→空欄は『埋めた』ことにならない
    if single_cell and len(after_vals) == 1:
        # ★ 宣言どおり 1 セルだけ書いた ── 「一括」ではないので疑う理由が無い。
        #   1 セルを超えていたら鳴らす（宣言と実体がずれた時こそ言うべき場面）。
        return None
    return f"★ 疑わしい: 空欄への同一値の一括書き込みです（値 {_fmt_cell_value(val)} × {len(after_vals)} セル）"


def _data_row_count(before: dict, sheet: str, key_col: int) -> int:
    """★ 近似: そのシートの原本使用範囲のうち見出し行(1行目)を除いた行で、
       key_col（左隣接のキー列）が埋まっている行数をデータ行数の目安とする。
       結合セルや空行の扱いまでは厳密化しない、素直な近似。"""
    rect = _used_range(before, sheet)
    if rect is None:
        return 0
    min_r, max_r, _min_c, _max_c = rect
    start = max(min_r + 1, 2)
    prefix = sheet + "!"
    count = 0
    for r in range(start, max_r + 1):
        v = before["cells"].get(f"{prefix}{r},{key_col}")
        if v is not None and v[0] not in (None, ""):
            count += 1
    return count


def count_reconciliation(before: dict, after: dict) -> str | None:
    """変更が単一シート・単一列に集中している場合だけ「データ N 行のうち M 行を変更」
       を添える（りんご欠落型のような『1行だけ抜けている』変更を1秒で見えるように）。
       ★ W8a 項目2: 分子(M)は必ずデータ行のみに数える。_used_range/_data_row_count は
       先頭行(min_r)を見出し行とみなして分母を数えているのに、分子側(changed_rows)は
       全変更行をそのまま数えていたため、見出し行も対象列で変わった場合（例: 新規列を
       作って見出し+データ全行に書く COMPUTE_COLUMN）に『データ5行のうち6行を変更』の
       ような算数が壊れた表示になっていた（実測: e2e_work/w3_e2e3_log.txt）。
       見出し行(min_r)自体の変更・原本の使用範囲より下の新規行の変更は、分子(データ行数)
       には数えず、見出し行の変更だけ別語「＋見出し行」で添える。"""
    changed = _changed_cells(before, after)
    if not changed:
        return None
    sheets = {s for s, _, _ in changed}
    cols = {c for _, _, c in changed}
    if len(sheets) != 1 or len(cols) != 1:
        return None
    sheet = next(iter(sheets))
    col = next(iter(cols))
    rect = _used_range(before, sheet)
    if rect is None:
        return None
    min_r, max_r, min_c, _max_c = rect
    key_col = col - 1 if col > min_c else col
    data_rows = _data_row_count(before, sheet, key_col)
    all_changed_rows = {r for _, r, _ in changed}
    header_changed = min_r in all_changed_rows
    # ★ 分子=データ行のみ: 見出し行(min_r)と、原本の使用範囲より下の新規行(> max_r)は
    #   ここでは「データ行」として数えない（分母(data_rows)も同じ範囲の定義に揃える）。
    data_changed_rows = {r for r in all_changed_rows if min_r < r <= max_r}
    changed_rows = len(data_changed_rows)
    # ★ 2026-08-26（初回体験の盲検 3 回目・CONFUSING 5）: 実測で
    #   「列 D: データ 0 行のうち 3 行を変更（0 行は未変更）」が出た（3+0 ≠ 0）。
    #   根: 分母 data_rows は**隣の列**（key_col = col-1）を数えて作っている。
    #   その列が空（実測は lookup.xlsx の単価列）だと分母が 0 になり、算数が壊れる。
    #   ★ 分母を都合のいい代用品から作らない ── 壊れた時は**物理の使用範囲**から取る。
    #     data_changed_rows は定義上 (min_r, max_r] に収まるので、これで必ず整合する。
    #   ★ 今まで整合していた出力は 1 文字も変わらない（壊れている時だけ差し替える）。
    if changed_rows > data_rows:
        data_rows = max_r - min_r
    unchanged_rows = max(data_rows - changed_rows, 0)
    col_letter = get_column_letter(col)
    # ★ operator 指摘②(2026-08-19): 1500 行のブックで「データ 999 行のうち 999 行を変更」と
    #   出ていた ── snapshot が MAX_ROWS で切れているのに、分母を snapshot から数えていた。
    #   ★ 分母を実データ行数に置き換えることは しない: 1000 行目より下の空/非空は snapshot に
    #   無いので、その数字はでっち上げになる。代わりに 主張を狭める（単位I と同じ型）──
    #   確認した範囲を言い、確認していない残りを 物理行数（true_rows・これは実測）で正直に述べる。
    true_rows = (before.get("true_rows") or {}).get(sheet, 0)
    if before.get("truncated") and true_rows > MAX_ROWS:
        hidden = true_rows - MAX_ROWS
        msg = (f"列 {col_letter}: 確認した先頭 {data_rows} 行のうち {changed_rows} 行を変更"
               f"（{unchanged_rows} 行は未変更）★ {MAX_ROWS} 行目より下の {hidden} 行は確認していない")
    else:
        msg = (f"列 {col_letter}: データ {data_rows} 行のうち {changed_rows} 行を変更"
               f"（{unchanged_rows} 行は未変更）")
    if header_changed:
        msg += "＋見出し行"
    return msg


# --- 依頼文言と変更範囲の重なりチェック（保守的・明示言及がある時だけ）--------

_RE_COL_KANJI_NUM = re.compile(r"列\s*(\d+)")
_RE_COL_LETTER_PREFIX = re.compile(r"列\s*([A-Za-z]{1,3})(?![A-Za-z0-9])")
_RE_COL_LETTER_SUFFIX = re.compile(r"(?<![A-Za-z0-9])([A-Za-z]{1,3})\s*列")
_RE_ROW = re.compile(r"行\s*(\d+)")


def extract_task_mentions(task: str, sheet_names: list) -> dict:
    """タスク文言から明示的な言及だけを正規表現で抜き出す（保守的・誤検知回避優先）。
       戻り値: {"cols": {1起点の列番号(文字表記=曖昧さなし)}, "digit_cols": {数字表記の生の値},
               "rows": {1起点の行番号}, "sheets": {原本に実在するシート名}}
       ★ 列の数字表記(『列2』)は曖昧 — 本ツールの規約(README・ヘルパ)は 0 起点だが、
       人が普段言うのは 1 起点。どちらか一方に断定すると正しい結果へ誤警報を出す
       (2026-08-14 再監査 2 回目で実測:『在庫(列2)』= 0 起点 C 列への正しい変更に
       B 列不在の★が出た)。生の値のまま保持し、照合側で両解釈を許す。"""
    cols = set()
    digit_cols = {int(m.group(1)) for m in _RE_COL_KANJI_NUM.finditer(task)
                  if int(m.group(1)) >= 0}
    for pat in (_RE_COL_LETTER_PREFIX, _RE_COL_LETTER_SUFFIX):
        for m in pat.finditer(task):
            try:
                cols.add(column_index_from_string(m.group(1).upper()))
            except ValueError:
                pass
    rows = {int(m.group(1)) for m in _RE_ROW.finditer(task) if int(m.group(1)) >= 1}
    # ★ 単位E: 「このシート名は依頼文に含まれるか」の照合は、決定側(resolve_target_sheet)と
    #   ここ（助言側）が独立に同じ文字列照合を書いていた。素材を1つに寄せ、決定側は1つに
    #   絞る／助言側は全部使う、という役割の違いだけを残す（判定規則も戻り値も不変）。
    # ★ 片配線の解消（2026-08-24）: 対象シートの決定は部分文字列を畳むのに、言及の
    #   抽出は畳んでいなかった。同じ規則（位置で見る）を ailine_core 側の 1 箇所から使う。
    sheets = set(drop_names_covered_by_longer(
        task, sheet_names_mentioned_in(task, list(sheet_names or ()))))
    return {"cols": cols, "digit_cols": digit_cols, "rows": rows, "sheets": sheets}


def _new_sheets(before: dict, after: dict) -> list:
    """before に無く after に有るシート名（実行中に新規作成されたシート）。順序維持。"""
    return [s for s in after["sheets"] if s not in before["sheets"]]


def _sheet_only_snapshot(snap: dict, sheet: str) -> dict:
    """snap のうち sheet に属する情報だけを持つ縮小 snapshot。
       ★ detect_ghost_data/detect_uniform_fill/count_reconciliation は snapshot() が返す
       dict の形（cells/merges/colw/rowh/sheets）をそのまま前提にした既存ロジックなので、
       新規シート専用に書き直さず、この縮小 snapshot 越しに『再利用』する（new_sheet_advisories
       が使う）。"""
    prefix = sheet + "!"
    return {
        "sheets": [sheet],
        "charts": 0,
        "cells": {k: v for k, v in snap["cells"].items() if k.startswith(prefix)},
        "merges": {sheet: snap.get("merges", {}).get(sheet, [])},
        "colw": {sheet: snap.get("colw", {}).get(sheet, {})},
        "rowh": {sheet: snap.get("rowh", {}).get(sheet, {})},
        "truncated": snap.get("truncated", False),
    }


def new_sheet_advisories(before: dict, after: dict) -> list:
    """★ W6 項目2: 監査所見「新規『集計』シートの全 0 埋めが★素通り」の根治。
       detect_ghost_data / detect_uniform_fill / count_reconciliation は『変更セル全部が
       該当した時だけ発火する』設計のため、新規シートの異常が他シートの正常な変更と混ざると
       『全部該当』が崩れ、丸ごと素通りしていた（例: COMPUTE_COLUMN が本シートの小計に書いた
       正しい値と、AGGREGATE が作った『集計』シートの壊れた全0埋めが同じ diff に同居する）。
       新規シートごとに単独の縮小 before/after（空 before・そのシートだけの after）を作って
       同じ3関数へ通すことで、他シートの変更に影響されず判定する。
       ★ detect_ghost_data は『原本の使用範囲外か』を測る関数だが、新規シートに原本の
       使用範囲は存在しない（何もかもが『新規』であって『範囲外』ではない）。そのため
       常に None を返す（適用はするが構造的に無言＝その関数の既存の誠実さをそのまま踏襲）。"""
    lines = []
    for sheet in _new_sheets(before, after):
        empty_before = {"sheets": [sheet], "charts": 0, "cells": {}, "merges": {sheet: []},
                         "colw": {sheet: {}}, "rowh": {sheet: {}}, "truncated": False}
        sheet_after = _sheet_only_snapshot(after, sheet)
        for fn in (detect_ghost_data, detect_uniform_fill):
            msg = fn(empty_before, sheet_after)
            if msg:
                lines.append(msg.replace("★ 疑わしい: ", f"★ 疑わしい: 新規シート『{sheet}』の", 1))
        recon = count_reconciliation(empty_before, sheet_after)
        if recon:
            lines.append(f"新規シート『{sheet}』 {recon}")
    return lines


_NEW_SHEET_MENTION_RE = re.compile(r"シート|ピボット|別に")


def unrequested_new_sheet_advisory(task: str, before: dict, after: dict, *,
                                    op: str | None = None) -> list:
    """★ W6 項目3（機械側）: 依頼文にシート新設の明示的な言及（『シート』『ピボット』
       『別に』のいずれか）が無いのに新規シートが作られたら申告する。
       ★ 保守的: 言及があれば（AGGREGATE/CHART/PivotSum 等が意図どおり新設したと見なし）沈黙。
       プロンプト側の抑制（CONTRACT の追記）はあくまで誘導であって保証にならないため、
       この機械申告が最終防衛線（feedback_intent_vs_guarantee: 指示は意図、保証は機械）。
       ★ C9: op が『新規シートを作る』と宣言していて（OP_WRITE_TARGET の writes に
       WRITE_NEW_SHEET・AGGREGATE(SummaryTable)/PIVOT(DataPilot)）、かつ今回ちょうど1枚だけ
       新規シートができた場合は、その1枚については警告でなく中立表示を返す
       （旧 _neutralize_declared_new_sheet_warning の後処理を発生源へ先取り。
       2枚以上できた場合は宣言どおりと断定できないので従来どおり全部警告する＝保守的）。
       ★ 単位C: 以前はここ専用の op 名集合 OP_DECLARED_SHEET_EFFECT を別に持っていたが、
       「新規シートを作る」は OP_WRITE_TARGET の宣言そのものなので、宣言を1つに畳んだ
       （宣言が2箇所にあると片方だけ更新されて食い違う）。集合の中身は変わっていない。"""
    new_sheets = _new_sheets(before, after)
    if not new_sheets or _NEW_SHEET_MENTION_RE.search(task):
        return []
    # ★ 帳票段: REPORT_PER_ROW はデータ行数ぶん（N枚・N は依頼のたび変わる）+検分シートを
    #   作るのが定義そのもの ── 「ちょうど1枚」を前提にする他 op の中立化条件をそのまま
    #   当てると、2枚目以降がすべて「依頼にない」と誤爆する。op の宣言(WRITE_NEW_SHEET)が
    #   立っている限り、枚数を問わず中立表示にする。
    # ★ 様式写像段: FORMAT_MAP は「出力シート1枚＋検分シート1枚」の計2枚が定義そのもの
    #   （REPORT_PER_ROW の N枚+検分と同じ理由でここに同居させる）。
    if op in ("REPORT_PER_ROW", "FORMAT_MAP") and _op_writes(op, WRITE_NEW_SHEET):
        return [f"（新規シート {len(new_sheets)} 枚の作成は意図どおりです）"]
    if _op_writes(op, WRITE_NEW_SHEET) and len(new_sheets) == 1:
        return [f"（新規シート『{new_sheets[0]}』の作成は意図どおりです）"]
    return [f"★ 依頼にない新しいシートが作成されました（{s}）" for s in new_sheets]


# ★ 致命2(W10e): 「既存シートの中身が置き換わった」検出。自由生成が依頼と無関係な
#   内容で既存シートを丸ごと上書きした実測事故（「集計」シートが日付別の無関係な内容に
#   すり替わった）への対抗。detect_ghost_data/detect_uniform_fill と同じ保守的な方針
#   （両条件とも『原本にあった非空セルが全部』変わった時だけ発火＝一部だけの更新・
#   再計算は対象外＝誤検知回避優先）。
def existing_sheet_replaced_advisory(before: dict, after: dict, *, op: str | None = None,
                                      precondition_broken: str | None = None) -> list:
    """before・after の両方に実在するシート（新規作成ではない）のうち、原本の使用範囲に
       あった非空セルが【全部】別の値に変わっている場合だけ「中身が置き換わった」を返す。
       一部のセルだけが変わった（値の再計算・部分更新等）場合は対象外（保守的）。
       ★ 空欄への一様書き込み等は detect_uniform_fill が別途担当するので、ここでは
       『置き換え後も何かしら値が残っている』ケースだけを見る（全消去は別の懸念）。
       ★ C9: op が『新規シートを作る』と宣言していて（OP_WRITE_TARGET の writes に
       WRITE_NEW_SHEET）、かつそのシートが OP_DECLARED_SHEET_NAME の宣言どおりの出力先
       （例: AGGREGATE→『集計』）なら、警告でなく中立表示を返す
       （旧 _neutralize_declared_sheet_replace_warning の後処理を発生源へ先取り）。"""
    # ★★ 単位G: 中立化は「前提が成立していた時だけ」。宣言(writes=new_sheet)は警告を黙らせる
    #   権利を持つが、その前提（＝その名前のシートは before に存在しない）が単位F の検査で
    #   破れていたなら、権利を失う ── 破れた宣言に「意図どおりです」と言わせない。
    #   ★ 破れた種類を見る（「何か破れた」では、format_only や reorder の破れで無関係に黙らせる）。
    #   ★ ここで警告に戻すだけで、正常系（前に ailine 自身が作った『集計』の作り直し）を
    #   肯定文に戻すのは 単位H（出所判定）の仕事。G で H をやると、また恒真を踏む。
    declared_sheet = (OP_DECLARED_SHEET_NAME.get(op)
                      if _op_writes(op, WRITE_NEW_SHEET) and precondition_broken != WRITE_NEW_SHEET
                      else None)
    lines = []
    for sheet in before["sheets"]:
        if sheet not in after["sheets"]:
            continue   # シート削除は diff_snapshots が別途拾う
        rect = _used_range(before, sheet)
        if rect is None:
            continue   # 原本にデータが無かった（新規に埋まっただけ）→ここでは判定しない
        min_r, max_r, min_c, max_c = rect
        prefix = sheet + "!"
        total = 0
        changed = 0
        after_has_content = False
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                key = f"{prefix}{r},{c}"
                b = before["cells"].get(key)
                b_val = b[0] if b is not None else None
                if b_val in (None, ""):
                    continue
                total += 1
                a = after["cells"].get(key)
                a_val = a[0] if a is not None else None
                if a_val not in (None, ""):
                    after_has_content = True
                if a_val != b_val:
                    changed += 1
        if total == 0 or changed != total or not after_has_content:
            continue   # 一部だけの変更、または全消去（置き換えではない）→ 発火しない
        if sheet == declared_sheet:
            lines.append(f"（既存シート『{sheet}』の更新は意図どおりです）")
            continue
        lines.append(f"★ 疑わしい: 既存シート『{sheet}』の中身が置き換わりました"
                      f"（元データ {total} セル分が別の内容に変わっています）")
    return lines


# ★ 致命2(W10e): AGGREGATE(集計)/PIVOT(ピボット) が既存の同名シートを宣言どおり
#   再生成する場合（例: 既に「集計」シートがある状態で再度「部門ごとにまとめて」を頼む
#   正常系）は、上の検出と同じ理屈で『中身が置き換わった』が誤って出る。W10c/W10d で
#   確立した「op の宣言済み効果と一致する変化は中立」に乗せて中立化する
#   （helpers/*.bas は出力シート名を固定で決め打つため、ここも固定表で対応づける）。
OP_DECLARED_SHEET_NAME = {"AGGREGATE": "集計", "PIVOT": "ピボット"}


def _changed_sheets(before: dict, after: dict) -> set:
    """何かしら変わったシート名の集合（セル・結合・列幅・行高・グラフ・追加/削除）。
       ★ operator10 ②: グラフだけを挿す run はセル値が一切変わらないため、以前は
       ここが沈黙し「★ 依頼で言及された『集計』は存在しません/変更されていません」の
       誤アラームで ✓ が △ に落ちていた。chart_counts（snapshot() がシート別に持つ）の
       差分もこのシートの「変わった」に数える。"""
    changed = set()
    for name in set(before["sheets"]) | set(after["sheets"]):
        if name not in before["sheets"] or name not in after["sheets"]:
            changed.add(name)
            continue
        if before.get("merges", {}).get(name) != after.get("merges", {}).get(name):
            changed.add(name)
        elif before.get("colw", {}).get(name) != after.get("colw", {}).get(name):
            changed.add(name)
        elif before.get("rowh", {}).get(name) != after.get("rowh", {}).get(name):
            changed.add(name)
    for sheet, _r, _c in _changed_cells(before, after):
        changed.add(sheet)
    before_charts = before.get("chart_counts") or {}
    after_charts = after.get("chart_counts") or {}
    for name in set(before_charts) | set(after_charts):
        if before_charts.get(name, 0) != after_charts.get(name, 0):
            changed.add(name)
    return changed


def mention_overlap_advisory(mentions: dict, before: dict, after: dict,
                              exclude_sheets: set | None = None,
                              exclude_cols: set | None = None) -> list:
    """言及があるのに変更範囲と全く重ならない場合だけ警告する（保守的）。
       数字表記の列は 0 起点/1 起点の両解釈を許し、どちらかが触られていれば沈黙する。
       ★ W10b 項目4b(摩擦): exclude_sheets に載るシート（例: LOOKUP_FILL の参照専用
       source_sheet）は、依頼文に言及があっても『変更されていません』を出さない
       （読み取り専用が正しい操作で、変更が無いのが正常なため誤警報だった）。
       ★ 安全器官の減衰なので保守的に: 抑制は呼び出し側が op から明示的に渡す時だけ・
       既定(None)は従来どおり無抑制。"""
    exclude_sheets = exclude_sheets or set()
    # ★ 2026-08-24: 対象列を**意図して変えない** op（SPLIT_CELL は元の列を残すのが契約）
    #   のために、列にも同じ抑制を用意した（シート側の exclude_sheets と同じ考え方）。
    exclude_cols = exclude_cols or set()
    if not (mentions["cols"] or mentions.get("digit_cols") or mentions["rows"]
            or mentions["sheets"]):
        return []
    changed = _changed_cells(before, after)
    changed_cols = {c for _, _, c in changed}
    changed_rows = {r for _, r, _ in changed}
    changed_sheets = _changed_sheets(before, after)

    lines = []
    for col in sorted(mentions["cols"]):
        if col in exclude_cols:
            continue
        if col not in changed_cols:
            letter = get_column_letter(col)
            lines.append(f"★ 依頼で言及された『列{letter}』は存在しません/変更されていません")
    for n in sorted(mentions.get("digit_cols", set())):
        # 1 起点読み = 列 n / 0 起点読み = 列 n+1 (1 起点換算)。両方外れた時だけ警告し、
        # 警告文は文字に変換せずユーザーの書いた数字のまま返す (推定で上書きしない)
        candidates = {c for c in (n, n + 1) if c >= 1}
        if candidates and not (candidates & changed_cols):
            lines.append(f"★ 依頼で言及された『列{n}』は存在しません/変更されていません")
    for row in sorted(mentions["rows"]):
        if row not in changed_rows:
            lines.append(f"★ 依頼で言及された『行{row}』は存在しません/変更されていません")
    for sheet in sorted(mentions["sheets"]):
        if sheet in exclude_sheets:
            continue
        # ★ 2026-08-28（雛形を 2 種類置いた実測）: 「英文の**雛形**で請求書を作って」で
        #   『雛形』（今回使わなかった方）に誤警報が出て ✓ が △ に落ちた。
        #   人が言った語が、読むだけと宣言済みのシート名の**一部**である回は黙る
        #   （『雛形』は『雛形_英文』の一部 ── 言及は当たっていて、変わらないのが正常）。
        if any(sheet in ex for ex in exclude_sheets):
            continue
        if sheet not in changed_sheets:
            lines.append(f"★ 依頼で言及された『{sheet}』は存在しません/変更されていません")
    return lines


def _structural_advisories(before: dict, after: dict, *, op: str | None = None,
                            resolved: dict | None = None, meta: dict | None = None,
                            precondition_broken: str | None = None,
                            after_path: Path | None = None) -> list:
    """助言のうち『この差分そのものが疑わしいか』を判定する部分だけ
       （①幽霊データ ②一様埋め ③件数の突き合わせ ⑤新規シートの中身・★ W6）。
       依頼文言との重なり(④ mention_overlap_advisory)は含めない。
       ★ W10d: 複合計画(cmd_run_plan)が段ごとの before/after にこの部分だけを
       再利用するために build_advisories から切り出した。④は依頼文全体に対する
       充足を問う質問なので、段ごとの局所的な before/after では判定できない
       （他段が担当する言及まで『この段で変更されていない』と誤検知する）。
       単発 op(build_advisories 経由)ではこれまでどおり④も同じ before/after で
       評価する（そちらは1段しかないため局所=全体で一致し、挙動は不変）。
       ★ C9: op/resolved/meta（呼び出し側が今回の段の宣言済み効果を渡す・省略時は None）は
       detect_ghost_data の new_col_letter 判定と existing_sheet_replaced_advisory の宣言シート判定・
       detect_write_target_type_change（★宣言つき挙動変更#1(b)）にそのまま渡す（旧 _neutralize_* 三兄弟の後処理を発生源へ先取り）。
       ★ operator10 ⑤: after_path（適用後の実ファイル・省略可）は detect_write_target_type_change
       の数式セル偽アラーム判定にのみ使う（無ければ安全側＝数式セルはカウントしない）。"""
    lines = []
    new_col_letter = _declared_new_column_letters(op, resolved, meta) or None \
        if (op and resolved is not None and meta is not None) else None
    new_row_at_end = _op_writes(op, WRITE_NEW_ROW_AT_END)   # ★ 単位C(D10): 合計行は宣言済みの効果
    for fn, kwargs in ((detect_ghost_data, {"new_col_letter": new_col_letter, "new_row_at_end": new_row_at_end}),
                        (detect_uniform_fill,
                          # ★ 2026-08-29: 「既にある合計行に書く」回は**1 セルだけ**の
                          #   書き込み（行を増やさない）── 宣言は op でなく args に在る。
                          {"single_cell": (_op_writes(op, WRITE_SINGLE_CELL)
                                            or bool((resolved or {}).get("_at_row"))),
                           "new_col_letter": new_col_letter,
                           "proved": bool(getattr(OP_WRITE_TARGET.get(op), "proves_which_cells",
                                                    False))})):
        msg = fn(before, after, **kwargs)
        if msg:
            lines.append(msg)
    recon = count_reconciliation(before, after)
    if recon:
        lines.append(recon)
    # ★ 2026-08-24: 1列目の空欄で分母が縮んだ事実を必ず言う（黙って少なく処理しない）。
    #   ★ 付きなので count_suspicious_advisories が拾い、決裁③で ✓→△ に降格する。
    #   ここ（全 op が通る助言の組み立て）に置くので、op ごとの配線漏れが起きない。
    if meta and (path := meta.get("path")):
        sheet = (resolved or {}).get("_target_sheet") or (meta.get("sheets") or [None])[0]
        header_row = (meta.get("header_rows") or {}).get(sheet, 1) if sheet else 1
        try:
            wb_gap = openpyxl.load_workbook(path, read_only=False, data_only=True)
            try:
                if sheet in wb_gap.sheetnames:
                    gap = detect_first_column_gap(wb_gap[sheet], header_row=header_row)
                    if gap:
                        lines.append(gap)
            finally:
                wb_gap.close()
        except Exception:
            pass   # 読めない時は黙る（無関係な入力を巻き添えにしない）
    lines.extend(new_sheet_advisories(before, after))
    lines.extend(existing_sheet_replaced_advisory(before, after, op=op, precondition_broken=precondition_broken) + [m for m in [detect_write_target_type_change(before, after, op=op, resolved=resolved, meta=meta, op_write_target=OP_WRITE_TARGET, is_number=_is_number, after_path=after_path)] if m])   # ★ 致命2(W10e) + 挙動変更#1(b)
    return lines


def build_advisories(task: str, before: dict, after: dict, exclude_sheets: set | None = None, *,
                      op: str | None = None, resolved: dict | None = None,
                      meta: dict | None = None, sheet_conflict=None,
                      precondition_broken: str | None = None,
                      after_path: Path | None = None) -> list:
    """diff の後に表示する助言行を全部集める。
       ①幽霊データ ②一様埋め ③件数の突き合わせ ⑤新規シートの中身（★ W6・
       _structural_advisories が担当） ⑥依頼にないシート新設の申告（★ W6）
       ④依頼文言との重なり。
       ★ C9: op/resolved/meta は _structural_advisories/unrequested_new_sheet_advisory へ
       そのまま横流しする（宣言済み効果の中立化を発生源で先取りするための追加引数・
       省略時は従来どおり無条件で全部発火する）。
       ★ 単位C(D8): 参照専用シート（OP_WRITE_TARGET の reads_only 宣言）は、呼び出し側が
       op ごとの if で渡すのをやめ、ここが宣言から自分で求めて exclude_sheets に足す
       （明示の exclude_sheets は自由生成経路のために残す＝宣言と和を取る）。
       ★ 誤爆#3: sheet_conflict（resolve_target_sheet が「この語は列名とも一致したので
       曖昧＝既定へ後退した」と決めた記録）に載る同名シートも、同じ和に足す ── 助言側で
       「曖昧かどうか」を判定し直さず、決めた側の結果をそのまま運ぶ
       （ailine_core.target_sheet.conflict_excluded_sheets 参照）。
       ★ operator10 ⑤: after_path は _structural_advisories へそのまま横流しするだけ
       （省略可・無ければ従来どおり）。"""
    lines = list(_structural_advisories(before, after, op=op, resolved=resolved, meta=meta,
                                        precondition_broken=precondition_broken, after_path=after_path))
    lines.extend(unrequested_new_sheet_advisory(task, before, after, op=op))
    mentions = extract_task_mentions(task, before["sheets"])
    excluded = (set(exclude_sheets or ()) | _declared_reads_only_sheets(op, resolved)
                | conflict_excluded_sheets(sheet_conflict))
    excluded_cols = (_declared_kept_subject_cols(op, resolved, meta)
                      | _letterlike_header_columns(meta))
    lines.extend(mention_overlap_advisory(mentions, before, after, excluded or None,
                                           excluded_cols or None))
    return lines


# ---------------------------------------------------------------------------
# ★ A': 用語集（vocab）— 税率等の「現場の取り決め値」を LLM でなく辞書に持たせる。
#   ~/.ailine/vocab.json（グローバルのみ）に {"語": 値} の平坦な dict で持つ。
#   設定ファイル経由で唯一「自由入力」がコード生成に届く経路なので、他のどの入力より
#   厳しく検疫する（float() 関門・制御文字禁止・件数/長さ上限。壊れたファイルは
#   クラッシュせず空辞書として扱う）。
# ---------------------------------------------------------------------------

def _sanitize_vocab_term(term) -> str | None:
    """語（キー）が登録可能な形か。空・制御文字・長すぎるものは None（拒否）。"""
    s = str(term).strip()
    if not s or len(s) > DEFAULT_VOCAB_MAX_TERM_LEN:
        return None
    if re.search(r"[\x00-\x1f\x7f]", s):   # 改行等の制御文字禁止（codegen へ渡る経路の防御）
        return None
    return s


def load_vocab(path: Path | None = None) -> dict:
    """~/.ailine/vocab.json を読む。無い/壊れている/形が違う場合は空の辞書を返す
       （★ クラッシュしない・battery や pytest からは path 明示で差し替えて再現性を保つ）。
       エントリごとに語をサニタイズし、値は float() を通ったものだけを採用する。
       件数が上限を超えた分は読み捨てる（先着順・ファイルの並び順に依存）。"""
    p = path or VOCAB_FILE
    if not p.is_file():
        return {}
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(raw, dict):
        return {}
    vocab: dict = {}
    for term, value in raw.items():
        if len(vocab) >= DEFAULT_VOCAB_MAX_ENTRIES:
            break
        clean_term = _sanitize_vocab_term(term)
        if clean_term is None:
            continue
        try:
            vocab[clean_term] = float(value)
        except (TypeError, ValueError):
            continue
    return vocab


def save_vocab(vocab: dict, path: Path | None = None) -> None:
    """vocab を ~/.ailine/vocab.json に上書き保存する。"""
    p = path or VOCAB_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(vocab, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def vocab_add(term: str, value, path: Path | None = None) -> tuple:
    """(ok, message)。term/value を検証してから登録・保存する。
       未登録の新規語で件数が上限に達している場合は拒否する（既存語の値更新は上限に関係なく可）。"""
    clean_term = _sanitize_vocab_term(term)
    if clean_term is None:
        return False, f"語『{term}』は登録できません（空/制御文字/{DEFAULT_VOCAB_MAX_TERM_LEN}文字超）"
    try:
        fval = float(value)
    except (TypeError, ValueError):
        return False, f"値『{value}』が数値ではありません"
    vocab = load_vocab(path)
    if clean_term not in vocab and len(vocab) >= DEFAULT_VOCAB_MAX_ENTRIES:
        return False, f"用語集が上限（{DEFAULT_VOCAB_MAX_ENTRIES}件）に達しています"
    vocab[clean_term] = fval
    save_vocab(vocab, path)
    return True, f"登録: {clean_term} = {fval:g}"


# ---------------------------------------------------------------------------
# ★ W10 便A: 別名ストア（言い回し → op 名）── REVIEW-20260822-w10-architect.md 3-2/3-3/6-4・
#   Namakoo 決裁（二段目翻訳・文字マッチ開始）に基づく前提工事。ここでは登録/照合だけを作る
#   （翻訳経路への接続は便C。lookup_alias は便C2 以降、提案フロー（_maybe_suggest_or_refuse の段0）から呼ばれる
#
#   形式: ~/.ailine/aliases.json に {"aliases": {言い回し: op名}, "order": [登録順]} の
#   平文 JSON（vocab.json とは別ファイル・上記 ALIASES_FILE 参照）。order は undo
#   （直近の登録の取り消し）専用 ── 機械が書く層には取り消しが要る、という6-4の決定。
#
#   検疫は vocab の写経（_sanitize_vocab_term を言い回しにもそのまま使う）+ op 名が
#   OP_META に実在するかの追加チェック + 件数上限は vocab と同じ DEFAULT_VOCAB_MAX_ENTRIES
#   を共有する（★ 別名は「事務の言葉の言い換え」で、税率のような別カテゴリの値ではない
#   ため、上限だけ揃えて枠は分けない判断）。
#
#   照合（lookup_alias）は「語として含む」── 断片ガードは単位B の
#   `_raw_target_not_embedded_in_task` と同じ判定を ailine_core/alias_store.py に写経した
#   ものを使う（3度目の断片問題を踏まないため独立に持つ・設計ノート③）。
# ---------------------------------------------------------------------------

def load_aliases(path: Path | None = None) -> tuple:
    """~/.ailine/aliases.json を読む。無い/壊れている/形が違う場合は空を返す
       （★ クラッシュしない・load_vocab と同じ流儀）。(aliases dict, order list)。"""
    p = path or ALIASES_FILE
    if not p.is_file():
        return {}, []
    try:
        raw = json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}, []
    return alias_store.parse_aliases_json(
        raw, lambda op: op in OP_META, DEFAULT_VOCAB_MAX_ENTRIES, DEFAULT_VOCAB_MAX_TERM_LEN)


def save_aliases(aliases: dict, order: list, path: Path | None = None) -> None:
    """aliases/order を ~/.ailine/aliases.json に上書き保存する。"""
    p = path or ALIASES_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    payload = alias_store.build_aliases_payload(aliases, order)
    p.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def alias_add(phrase: str, op: str, path: Path | None = None) -> tuple:
    """(ok, message)。②検疫: 言い回しの検疫（_sanitize_vocab_term 同等）・op が OP_META に
       実在するか・件数上限（vocab と同じ DEFAULT_VOCAB_MAX_ENTRIES）。
       既存の言い回しへの再登録（op の張り替え）は上限に関係なく可（vocab_add と同じ扱い）。"""
    clean_phrase = _sanitize_vocab_term(phrase)
    if clean_phrase is None:
        return False, f"言い回し『{phrase}』は登録できません（空/制御文字/{DEFAULT_VOCAB_MAX_TERM_LEN}文字超）"
    if op not in OP_META:
        return False, f"op『{op}』は存在しません（実在する操作名のみ登録可。一覧: ailine ops）"
    aliases, order = load_aliases(path)
    if clean_phrase not in aliases and len(aliases) >= DEFAULT_VOCAB_MAX_ENTRIES:
        return False, f"別名ストアが上限（{DEFAULT_VOCAB_MAX_ENTRIES}件）に達しています"
    if clean_phrase in order:
        order.remove(clean_phrase)
    order.append(clean_phrase)   # ★ undo は order の末尾（＝最後に登録/更新された言い回し）を取り消す
    aliases[clean_phrase] = op
    save_aliases(aliases, order, path)
    return True, f"登録: {clean_phrase} → {op}"


def save_alias(phrase: str, op: str, path: Path | None = None) -> bool:
    """★ 凍結済み検体（tests/test_alias_store.py）の契約: 戻り値は bool のみ。
       alias_add の薄いラッパー（CLI 表示用のメッセージを捨てるだけ）。"""
    ok, _msg = alias_add(phrase, op, path)
    return ok


def alias_remove(phrase: str, path: Path | None = None) -> tuple:
    """(ok, message)。④: 登録済みの言い回しを削除する（vocab.json 側は remove を作らない
       という5307の決定はそのまま・機械が書く別名側だけが前提が反転するので remove を持つ）。"""
    aliases, order = load_aliases(path)
    clean_phrase = _sanitize_vocab_term(phrase)
    if clean_phrase is None or clean_phrase not in aliases:
        return False, f"言い回し『{phrase}』は登録されていません"
    del aliases[clean_phrase]
    if clean_phrase in order:
        order.remove(clean_phrase)
    save_aliases(aliases, order, path)
    return True, f"削除: {clean_phrase}"


def alias_undo(path: Path | None = None) -> tuple:
    """(ok, message)。④: 直近の登録（order の末尾）だけを取り消す。"""
    aliases, order = load_aliases(path)
    if not order:
        return False, "取り消せる登録がありません"
    last = order.pop()
    aliases.pop(last, None)
    save_aliases(aliases, order, path)
    return True, f"取り消し: {last}"


def lookup_alias(task: str, path: Path | None = None) -> str | None:
    """⑤: 依頼文 task に登録済みの言い回しが「語として」含まれていれば op 名を返す
       （slot は運ばない・戻り値は op 名の str か None だけ）。
       断片ガード: alias_store.phrase_is_standalone_in_task（単位B と同型の判定）で、
       他の語の断片としてしか出現しない言い回しは当てない。
       複数の言い回しが同時に標準出現でヒットした場合は、最長のものを勝たせる
       （短い言い回しが長い言い回しの一部になっている時、より具体的な方を優先する
       ── ★ このタスクで足した境界検体 tests/test_alias_lookup_boundary.py 参照）。
       ★ このタスクでは翻訳経路からは呼ばれない（接続は便C）。"""
    if not task:
        return None
    aliases, _order = load_aliases(path)
    best_phrase = None
    for phrase in aliases:
        if alias_store.phrase_is_standalone_in_task(phrase, task):
            if best_phrase is None or len(phrase) > len(best_phrase):
                best_phrase = phrase
    return aliases[best_phrase] if best_phrase else None


# --- ★ A': APPEND_TOTAL の倍率(factor)を LLM から切り離し、機械が確定する ------------
#   ①依頼文の明示率を regex で抽出 ②無ければ用語集を引く ③どちらも無く label が税/込を
#   含むなら CLARIFY。LLM が factor を返しても、ここが常に勝つ（verify_dsl_args 側で
#   食い違いを WARN として記録する）。

_RATE_PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[%％]")
_RATE_BAI_RE = re.compile(r"(\d+(?:\.\d+)?)\s*倍")
# ★ W10b 項目3: 「1.1を掛けた/掛ける」「1.1で割った/割る」型（税込み/税抜きの言い換えで
#   「倍」を伴わない場合の実測ギャップ・battery v5 #503 で発覚）。掛けるは n そのまま、
#   割るは 1/n（税抜き＝税込み金額から逆算する倍率）。
_RATE_KAKE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:を|に)?\s*掛け")
_RATE_WARI_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:で|に)?\s*割っ")
_RATE_KEYWORD_RE = re.compile(r"税|倍率")
_RATE_BARE_NUM_RE = re.compile(r"(\d+(?:\.\d+)?)")
# ★ W10c 高: 依頼文に「率らしい語」が一切無いのに COMPUTE_COLUMN の「1列×率」パターン
#   （税込み/税抜き専用）へ誤分類された時に分類そのものを疑うための、より広い信号語。
#   上の各 _RATE_*_RE より緩い（数値を伴わなくても良い）— 「倍率を求める話かどうか」
#   だけを見る。
_RATE_SIGNAL_RE = re.compile(r"[%％]|倍|掛け|割っ|税")
# ★ W10c 中: 新規列の見出しを自然な日本語にするための語（A' 原則: LLM を使わず正規表現の
#   有無判定だけで決める。査定で名指しされた「金額*1.1」という数式風の見出しの対応）。
_TAX_INCLUSIVE_RE = re.compile(r"税込")
_TAX_EXCLUSIVE_RE = re.compile(r"税抜")


def extract_rate_factor(text: str) -> tuple:
    """依頼文から明示の倍率を抽出する。戻り値は (factor, 出典スニペット) か (None, None)。
       ①「10%」「8 ％」型 → 1+n/100 ②「1.1倍」型 → n そのまま ②'「1.1を掛けた」型 → n
       そのまま ②''「1.1で割った」型 → 1/n（税抜き等の逆算） ③「税」「倍率」という語の
       前後8文字だけにある裸の小数（例:「税率0.1」）→ 1未満なら 1+n・1以上ならそのまま
       （無関係な数値の誤爆を避けるため、③だけは税/倍率の語の近傍に絞る）。
       複数の異なる値が見つかった場合は断定しない（None, None・CLARIFY に委ねる）。"""
    if not text:
        return None, None
    candidates: dict = {}   # factor -> 出典スニペット（最初に見つかったもの）
    for m in _RATE_PCT_RE.finditer(text):
        f = round(1 + float(m.group(1)) / 100, 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_BAI_RE.finditer(text):
        f = round(float(m.group(1)), 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_KAKE_RE.finditer(text):
        f = round(float(m.group(1)), 6)
        candidates.setdefault(f, m.group(0))
    for m in _RATE_WARI_RE.finditer(text):
        n = float(m.group(1))
        if n > 0:
            f = round(1 / n, 6)
            candidates.setdefault(f, m.group(0))
    if not candidates:
        for km in _RATE_KEYWORD_RE.finditer(text):
            window = text[max(0, km.start() - 8): km.end() + 8]
            nm = _RATE_BARE_NUM_RE.search(window)
            if nm:
                n = float(nm.group(1))
                f = round((1 + n) if n < 1 else n, 6)
                candidates.setdefault(f, nm.group(0))
    if len(candidates) == 1:
        f, snippet = next(iter(candidates.items()))
        return f, snippet
    return None, None


def lookup_vocab_factor(text: str, vocab: dict) -> tuple:
    """依頼文に用語集の語が部分一致で含まれるかを見る。戻り値は (factor, 用語) か
       (None, None)。複数の異なる語（異なる値）がヒットした場合は断定しない。"""
    if not text or not vocab:
        return None, None
    hits: dict = {}
    for term, value in vocab.items():
        if term and term in text:
            hits.setdefault(value, term)
    if len(hits) == 1:
        value, term = next(iter(hits.items()))
        return value, term
    return None, None


def lookup_vocab_tax_factor(vocab: dict) -> tuple:
    """★ operator8 ②: 恒真式の番人が CLARIFY に倒す**直前**の敗者復活。第一照合
       （lookup_vocab_factor・依頼文の字面部分一致）と依頼文の率抽出（extract_rate_factor）
       の優先順は一切変えない ―― これはその両方が外れた（label が税/込を含むのに倍率が
       確定できない）場合だけ呼ばれる最後の一手。
       label『税込み合計』は語彙 key『消費税』を字面に含まないため第一照合は当たらないが、
       「税込み/税抜き」の依頼で使う倍率は用語集の中でも key に「税」を含む語である
       可能性が高い ―― そこだけ緩めて拾う（A' 原則は維持: 実在する用語集エントリの
       値だけを使い、LLM は使わない）。
       戻り値: (factor, term, candidates)。
         ・相異なる値がちょうど1つ → (その値, 名前, ())。
         ・相異なる値が2つ以上 → (None, None, ((value, term), ...))
           （呼び出し側が候補を名指しした CLARIFY にする ―― 「登録してください」とは言わない。
           登録は既に済んだ手だから）。
         ・0件 → (None, None, ())（呼び出し側は従来どおりの登録案内）。"""
    if not vocab:
        return None, None, ()
    seen: dict = {}   # value -> term（最初に見つかった名前。表示・一意判定の両方に使う）
    for term, value in vocab.items():
        if term and "税" in term:
            seen.setdefault(value, term)
    if len(seen) == 1:
        value, term = next(iter(seen.items()))
        return value, term, ()
    if len(seen) >= 2:
        return None, None, tuple(seen.items())
    return None, None, ()


def _resolve_tax_rescue(context_word: str, context_text: str, vocab: dict | None) -> tuple:
    """★ 致命③(2026-08-23レビュー): lookup_vocab_tax_factor の敗者復活を APPEND_TOTAL・
       COMPUTE_COLUMN の両方から呼ぶ共有実装（レビュー所見: 逐語コピー2箇所のうち
       COMPUTE_COLUMN 側だけ配線が届いていなかった片配線 ── 登録済みの税語彙があるのに
       「登録してください」と嘘をつく）。context_text（APPEND_TOTAL は label、
       COMPUTE_COLUMN は task）に「税」か「込」が無ければ rescue 対象外。
       戻り値: (factor, term, error_message)。
         ・rescue 対象外（税/込を含まない） → (None, None, None)
         ・rescue 成立 → (factor, term, None)
         ・rescue 失敗（候補複数/0件） → (None, None, "エラー文言")。"""
    if not any(k in context_text for k in ("税", "込")):
        return None, None, None
    tax_factor, tax_term, tax_candidates = lookup_vocab_tax_factor(vocab or {})
    if tax_factor is not None:
        return tax_factor, tax_term, None
    if tax_candidates:
        listed = "・".join(f"{term}={value:g}" for value, term in tax_candidates)
        return None, None, (
            f"{context_word}『{context_text}』は税/込を含みますが、用語集に候補が複数あります"
            f"（{listed}）。どちらを使うか依頼文に書いてください（例:「消費税10%」）"
        )
    return None, None, (
        f"{context_word}『{context_text}』は税/込を含みますが倍率が分かりません。"
        "依頼文に税率を書く（例:「消費税10%」）か、用語集に登録してください"
        "（例: ailine vocab add 消費税 1.1）"
    )


# --- ★ A' 原則(致命3・W10e): SET_COLUMN_VALUE が書き込む定数値を LLM から切り離す ------
#   依頼文の引用符（「」『』""''）で囲まれた文字列を機械抽出する。extract_rate_factor と
#   同じ考え方 — ちょうど1つに絞れる時だけ確定・0件/2件以上は CLARIFY に委ねる（None）。
_QUOTE_PATTERNS = (
    re.compile(r"「([^」]+)」"),
    re.compile(r"『([^』]+)』"),
    re.compile(r'"([^"]+)"'),
    re.compile(r"'([^']+)'"),
)


def extract_quoted_literal(text: str) -> str | None:
    """依頼文全体を通して、引用符で囲まれた文字列がちょうど1つだけ見つかった場合に
       その中身を返す。0個/2個以上は曖昧とみなし None（機械確定を諦める＝呼び出し側が
       CLARIFY にする）。"""
    if not text:
        return None
    found = []
    for pat in _QUOTE_PATTERNS:
        found.extend(m.group(1) for m in pat.finditer(text))
    if len(found) == 1:
        return found[0]
    return None


# ---------------------------------------------------------------------------
# M2b: 中間命令言語（DSL）パイプライン
#   ①翻訳(LLM) → ②検証(接地) → ③確認行 → ④決定論 codegen(LLM不使用) → ⑤適用(既存機構)
#   → ⑥op別事後条件(openpyxl で機械検証)
#   ★ 天井は bench/translation_battery.json で実測（凍結合格線: op90%/slot80%/誤断定20%）。
#   CLARIFY・語彙外(FREEFORM)・翻訳失敗は現行の自由生成経路（M2a 助言つき）へ retreat する。
# ---------------------------------------------------------------------------

# ★ W9: op メタデータ（接地フォーム設計の確定分・今回は宣言だけで UI は変えない）。
#   各 op に category(大分類・事務の言葉)・label(操作名・事務の言葉)・synonyms(依頼文で
#   よく使う言い方の例・宣言のみ＝翻訳プロンプトへ機械的に注入はしない) を一箇所で持つ。
#   OP_LABELS は後方互換のためこの dict から導出する（既存コードの OP_LABELS.get(op, op)
#   はそのまま・値は完全に同じ）。
#   ★ W10 便C1: match_phrases を追加（もしかして提案 suggest_ops の照合専用プール）。
#   synonyms とは別物 ── synonyms は `ailine ops` 表示用で3語までに絞る（表が壁にならない
#   ため）。match_phrases は表示しない・件数上限なし。op の**意味**から一般語で書く
#   （凍結セット bench/w10_suggest_frozen_set.json の文言は写さない ── 写すと自己汚染。
#   セットとの重複率は bench/run_w10_suggest_eval.py が測って報告する＝透明化）。
#   label/synonyms も照合プールに含めてよい（suggest_ops 側で合成する）ので、ここでは
#   synonyms と文字列が重複しない言い回しだけを足す。
OP_META = {
    "SORT": {"category": "並べ替える", "label": "並べ替え", "folder": False,
              "synonyms": ["並べ替え", "ソート", "順に並べる"],
              "match_phrases": ["順番", "昇順", "降順", "整列", "並び替える", "並び替え",
                                "順位付け"]},
    "COMPUTE_COLUMN": {"category": "計算する", "label": "計算列", "folder": False,
                         "synonyms": ["計算", "掛け算・割り算", "列同士の演算"],
                         "match_phrases": ["演算", "掛け算", "割り算", "足し算", "引き算",
                                            "式を入れる", "計算式", "算出する", "列を作る"]},
    "LOOKUP_FILL": {"category": "表を編集する", "label": "転記", "folder": False,
                     "synonyms": ["引っ張ってくる", "転記", "VLOOKUP"],
                     "match_phrases": ["参照する", "引き当てる", "突き合わせる", "マスタ参照", "マスタ",
                                        "値を持ってくる", "対応する値を入れる"]},
    "AGGREGATE": {"category": "計算する", "label": "集計", "folder": False,
                   "synonyms": ["集計", "まとめる", "グループごとに小計"],
                   "match_phrases": ["小計", "合算", "グループ集計", "内訳", "サマリー",
                                      "項目ごとにまとめる"]},
    "BOLD": {"category": "見た目を整える", "label": "太字", "folder": False,
              "synonyms": ["太字", "ボールド", "強調"],
              "match_phrases": ["太くする", "強調する", "目立たせる"]},
    "FILL_COLOR": {"category": "見た目を整える", "label": "背景色", "folder": False,
                    "synonyms": ["色を付ける", "塗りつぶす", "ハイライト"],
                    "match_phrases": ["背景に色", "色付け", "着色する", "マーキングする"]},
    "NUMBER_FORMAT": {"category": "見た目を整える", "label": "数値書式", "folder": False,
                        "synonyms": ["桁区切り", "カンマ区切り", "3桁区切り"],
                        "match_phrases": ["数値の表示形式", "3桁ごとに区切る", "通貨表示",
                                           "小数点の桁数", "カンマを入れる"]},
    "MERGE": {"category": "表を編集する", "label": "セル結合", "folder": False,
               "synonyms": ["結合", "セルを繋げる", "セルをまとめる"],
               "match_phrases": ["結合する", "つなげる", "一つのセルにする", "まとめて一体化"]},
    "CHART": {"category": "グラフを作る", "label": "グラフ", "folder": False,
               "synonyms": ["グラフ", "棒グラフ", "チャート"],
               "match_phrases": ["グラフ化", "図にする", "可視化", "折れ線グラフ", "円グラフ",
                                  "チャートを作る"]},
    "CENTER_ALIGN": {"category": "見た目を整える", "label": "中央揃え", "folder": False,
                       "synonyms": ["中央揃え", "センタリング", "真ん中に寄せる"],
                       "match_phrases": ["中央に配置", "センターに合わせる", "均等に中央配置"]},
    "APPEND_TOTAL": {"category": "計算する", "label": "合計追加", "folder": False,
                       "synonyms": ["合計を出す", "税込み合計", "一番下に合計"],
                       "match_phrases": ["合計を追加", "総額を出す", "合計行", "締めの合計"]},
    # ★ W9: 検証済みヘルパ4種の DSL 語彙昇格。
    "INSERT_ROWS": {"category": "表を編集する", "label": "行挿入", "folder": False,
                      "synonyms": ["行を挿入", "行を追加", "行を足す"],
                      "match_phrases": ["行を追加する", "空行を入れる", "行間を空ける"]},
    # ★ 2026-08-26: 表計算の基本操作なのに 21 op のどれにも無かった 3 つ
    #   （Namakoo が GUI を触って実測: 「5行目に商品として梨を追加して」が
    #     行挿入＋一括書換に分解され、4 段とも別々の理由で落ちた）。
    #   ★ 削除はこの道具で最も破壊的な操作 ── 消す前に中身を見せる（下の check_* 参照）。
    "ADD_ROW": {"category": "表を編集する", "label": "行追加", "folder": False,
                 "synonyms": ["行を追加して値を入れる", "データを1行足す", "レコードを追加"],
                 "match_phrases": ["行を追加して", "1行足して", "データを追加"]},
    # ★ 2026-08-27: 1 セルだけ書き換える。**第二段専用**（OPS_DOC に載せない＝語彙コスト 0）。
    #   SET_COLUMN_VALUE と読まれた依頼が「梨の売上を…」のように**行を名指し**していたら、
    #   op を固定して読み直す（ADD_ROW で実証済みの経路をそのまま使う）。
    "SET_CELL_VALUE": {"category": "表を編集する", "label": "1セル書換", "folder": False,
                        "synonyms": ["1つのセルを書き換える", "特定の行の値を変える"],
                        # ★ 番人が空の match_phrases を通さない（作法どおり）。
                        #   第二段専用のつもりでも、一覧（ailine ops）には出るので
                        #   **人が読める語**を持たせる ── 空にするなら免除簿に理由が要る。
                        "match_phrases": ["この行のこの値だけ変える", "1 セルだけ直す"]},
    # ★ 2026-08-27: 入れ替え。**行と列で op を分けない** ── 依頼文（「みかんとぶどうを
    #   入れ替えて」）は、その 2 つが行なのか列なのかを言わない。どちらかは**実表を見た
    #   機械が決める**（verify_dsl_args が _axis を積む）。LLM に軸を当てさせない。
    "SWAP": {"category": "表を編集する", "label": "入れ替え", "folder": False,
              "synonyms": ["入れ替え", "交換", "順番を入れ替える"],
              "match_phrases": ["AとBを入れ替えて", "2つの行を交換", "2つの列を交換"]},
    # ★ 2026-08-27（Namakoo「列の追加はできないの？」）: **削除だけあって追加が無かった。**
    #   行は空行(INSERT_ROWS)と値つき(ADD_ROW)の 2 つがあるのに、列は削除だけ。
    #   実測でも「備考という列を追加して」は語彙外で断られていた（GUI で本人が確認）。
    # ★ 2026-08-27（Namakoo「原価が500以上の項目に◎を付ける」）: 表計算のごく普通の操作
    #   なのに一覧に無かった（SET_COLUMN_VALUE は列を丸ごと同じ値にするだけ）。
    #   実測でも 4/4 で OUT_OF_VOCAB（「条件付き書式」と誤って読まれていた）。
    # ★ 2026-08-27（Namakoo「特定行や特定列の抜き出しができない」）: 行は EXTRACT が
    #   持っていたが、**列を選ぶ手段が 1 つも無かった**。
    "EXTRACT_COLUMNS": {"category": "表を編集する", "label": "列抽出", "folder": False,
                         "synonyms": ["列だけ抜き出す", "必要な列だけ取り出す", "列を選んで別シートへ"],
                         "match_phrases": ["商品と売上の列だけ抜き出して", "必要な列だけ取り出して"]},
    "SET_WHERE": {"category": "表を編集する", "label": "条件つき書換", "folder": False,
                   "synonyms": ["条件に合う行だけ書き換える", "〜以上の行に印を付ける",
                                 "該当する行にだけ値を入れる"],
                   "match_phrases": ["500以上の行に○を付けて", "条件に合う行だけ書き換えて"]},
    "ADD_COLUMN": {"category": "表を編集する", "label": "列追加", "folder": False,
                    "synonyms": ["列を追加", "列を足す", "列を挿入"],
                    "match_phrases": ["列を追加して", "列を足して", "空の列を入れて"]},
    "DELETE_ROWS": {"category": "表を編集する", "label": "行削除", "folder": False,
                     "synonyms": ["行を削除", "行を消す", "行を取り除く"],
                     "match_phrases": ["行を削除して", "行を消して", "不要な行を除く"]},
    "DELETE_COLUMN": {"category": "表を編集する", "label": "列削除", "folder": False,
                       "synonyms": ["列を削除", "列を消す", "列を取り除く"],
                       "match_phrases": ["列を削除して", "列を消して", "この列は要らない"]},
    "DRAW_BORDERS": {"category": "見た目を整える", "label": "けい線", "folder": False,
                       "synonyms": ["けい線を引く", "罫線を引く", "枠線を付ける"],
                       "match_phrases": ["線で囲む", "表に枠を付ける", "けい線"]},
    "AUTOFIT": {"category": "見た目を整える", "label": "列幅自動調整", "folder": False,
                 "synonyms": ["幅を内容に合わせる", "列幅調整", "列を自動調整"],
                 "match_phrases": ["列幅を自動調整", "はみ出しを直す", "列幅を整える"]},
    "PIVOT": {"category": "計算する", "label": "ピボット", "folder": False,
               "synonyms": ["ピボットテーブル", "ピボットで集計", "クロス集計"],
               "match_phrases": ["縦横に組み替える", "行と列を入れ替えて集計", "クロス表にする",
                                 "行と列を入れ替え", "入れ替え"]},
    # ★ 致命3(W10e): 「列を一括で定数に書き換える」の DSL 昇格（査定所見:総務事務が
    #   最も頻繁に行う操作に信頼できる経路が無かった）。
    "SET_COLUMN_VALUE": {"category": "表を編集する", "label": "一括書換", "folder": False,
                           "synonyms": ["全部同じ値にする", "一括で書き換える", "列を統一する"],
                           "match_phrases": ["一斉に置き換える", "同じ内容で埋める", "列をまとめて更新"]},
    # ★ 生まれた時から検証つきの1例目（コミット 2edcb08「EXTRACT op」参照）: 単一条件
    #   （列×比較×値）に一致する行を新シートへ抜き出す。自由生成の実弾2件（全セル文字列化・
    #   空シートで exit 0）を事後条件(check_extract)が直接殺す形で op に昇格させる。
    "EXTRACT": {"category": "表を編集する", "label": "抽出", "folder": True,
                 "synonyms": ["抜き出す", "抽出", "絞り込んでコピー"],
                 "match_phrases": ["抽出する", "条件で絞り込む", "該当行だけ取り出す", "別シートに出す"]},
    # ★ freeform 廃止バンドル前段: DEDUP（EXTRACT の兄弟・非破壊形）。判定キー列の値の
    #   組が同じ行のうち最初の1行だけを新シートへ残す（元シートの行は消さない）。
    #   ★ folder は False（M2 のフォルダ抽出対応は EXTRACT だけ・DEDUP の folder 版は未実装）。
    "SPLIT_CELL": {"category": "表を編集する", "label": "セル分割", "folder": False,
                    "synonyms": ["セルを分ける", "1セルを複数セルに", "項目分割"],
                    "match_phrases": ["別のセルに分ける", "別セルに分ける", "1件ずつ別のセルに",
                                       "区切りで分ける", "セルを分割"]},
    "DEDUP": {"category": "表を編集する", "label": "重複除去", "folder": False,
               "synonyms": ["重複を除く", "重複除去", "重複行を消す", "ユニークにする"],
               "match_phrases": ["ダブりを消す", "重複行を削除"]},
    # ★ 帳票段: 実需 MARKET-20260823-lancers.md 財務書類系7件中5件がこの形
    #   （表の1行を、人が作った定型フォーマットの1枚に転写してN枚出す）。
    #   ★ folder は False（第一波はフォルダ集約に対応しない・単一ブックのみ）。
    "REPORT_PER_ROW": {"category": "表を編集する", "label": "帳票作成", "folder": False,
                         "synonyms": ["請求書を作る", "1件ずつ帳票にする", "雛形に転記"],
                         "match_phrases": ["請求書", "見積書", "領収書", "帳票を作る",
                                            "定型フォーマットに転記", "1行ずつ書類にする",
                                            "行ごとに書類を作る"]},
    # ★ 様式写像段: REPORT_PER_ROW の兄弟（縦の展開）。実需 MARKET-20260823-lancers.md
    #   FORMAT_MAP 5件・別の様式（列の並び・名前・単位が違う）へ機械的に写す需要
    #   （受注CSV→出荷CSV・預金出納帳→弥生会計インポート様式 等）。
    #   ★ folder は False（第一波は単一ブックのみ）。
    "FORMAT_MAP": {"category": "表を編集する", "label": "様式写像", "folder": False,
                    "synonyms": ["別の様式に写す", "決まった形式に変換", "様式に合わせて転記"],
                    "match_phrases": ["様式の形に写して", "様式に合わせて", "決まったフォーマットに変換",
                                       "インポート様式に変換", "様式シートの形に",
                                       # ★ 台帳の実文言（2852253/4139068/4701331）から。
                                       #   「入れ替え」は PIVOT と取り合いになる語なので
                                       #   両方に持たせ、どちらも候補に出す（人が選ぶ）。
                                       "列の並び替え", "入れ替え", "レイアウト変更",
                                       "レイアウトに変更", "フォーマットに変換"]},
}

OP_LABELS = {op: meta["label"] for op, meta in OP_META.items()}

# ★ W10 反復(棄権設計・便④): 「op の意味領域の外側」にある Excel 機能語彙。OP_META の
#   19 op が一切対応しない、よく知られた Excel の機能名を意味から書く（凍結セット
#   bench/w10_suggest_frozen_set.json の文言の写しではない ── 書いた後に
#   bench/run_w10_suggest_eval.py が重複率を透明化する）。ここに挙げた語が依頼文に
#   語として現れたら、pool 側の一致に関わらず候補ゼロで応答する（できないことに候補を
#   出す＝二度目の「できません」の方が重い、という設計の線）。
#   ★ 便C2(2026-08-22 抜き打ち検体で 5/12 誤提示・Namakoo 決裁): この名簿は主対策では
#   なくなった。「非対応機能」は列挙できない開集合であり、増築しても収束しない
#   （現に封印されていた別の12件はこの名簿の外から誤提示した）。主対策は
#   suggest_ops 側の白側の証拠要求の厳格化（語としての一致のみを候補にする）に移した
#   ── この名簿は「効く分には害がない」補助として残すだけで、以後は増築しない。
OUT_OF_SCOPE_TERMS = [
    "印刷", "ページ設定", "印刷範囲", "PDF",
    "パスワード", "シート保護", "保護", "ロック",
    "条件付き書式",
    "入力規則", "ドロップダウン",
    "ふりがな", "フリガナ",
    "ウィンドウ枠の固定", "先頭行固定", "行固定", "列固定", "固定",
    "コメント", "メモ",
    "マクロ", "VBA",
    "画像", "図形", "アイコン",
    "ハイパーリンク", "リンク",
    "共有", "保存", "バックアップ",
    "テーマ", "スタイル", "グラデーション",
]


def _op_match_pool(op: str) -> list:
    """op 1つ分の照合語彙（label+synonyms+match_phrases）。suggest_ops の pool 構築と
       同じ規則を1 op 分だけ切り出したもの（★ W10 便C2 S5: 残差検出(ailine_core/residue.py)
       が「この操作の照合語彙で消費された部分」を判定するのにも同じプールを使う ──
       suggest_ops と別のプールを持つと片方だけ更新されて食い違う事故になるため）。"""
    meta = OP_META.get(op, {})
    return [meta.get("label", op), *meta.get("synonyms", ()), *meta.get("match_phrases", ())]


# ★ operator9 ②: 複合計画の1段について「段の存在自体」に依頼文の根拠があるかを機械照合する。
#   単位E は対象スロットの食い違いしか見ておらず、依頼文に一度も出てこない op の段（捏造段）が
#   湧いても ✓ が出ていた ── ここは cmd_run_plan の DSL 段だけが呼ぶ（単発 cmd_run_dsl は
#   確認段が近いので対象外・自由生成の段は元から ⚠ 経路なので対象外）。
def _phrase_matches_task_allowing_conjugation(phrase: str, task: str) -> bool:
    """alias_store.phrase_is_standalone_in_task と同じ断片ガードに、活用ゆれを1段だけ許す
       ―― pool のフレーズは辞書形（例:「整える」）で書かれているが、依頼文は活用した形
       （「整えて」「整えた」等）で来ることが普通にある。ichidan 動詞（〜える/〜いる）は
       語幹（「る」を落とした形）が活用のあいだ共通なので、辞書形一致が失敗したら語幹
       一致も試す（例:「整える」→「整え」は「整えて」に語として現れる）。
       ★ 既存の suggest_ops/alias_store は厳格一致のまま変えない（誤提示 5/12 の教訓で
       意図的に緩めていない・ここは新設のこの検査専用の緩和）。緩めるのは「一致しにくく
       なる」方向のみで、既存の一致を狭めることは無い（誤爆防止が命の方向と一致）。"""
    if alias_store.phrase_is_standalone_in_task(phrase, task):
        return True
    if len(phrase) > 2 and phrase.endswith("る"):
        return alias_store.phrase_is_standalone_in_task(phrase[:-1], task)
    return False


def _op_has_task_grounding(op: str, resolved: dict, task: str) -> bool:
    """根拠 = (a) その op の照合プール句（_op_match_pool）が依頼文に語として在る
       （_phrase_matches_task_allowing_conjugation・活用ゆれを1段だけ許す）∨
       (b) 解決済み args の値のいずれかが依頼文の語に接地している（断片ガードを流用）。
       ★ 誤爆防止が命（検体の二本立て根拠が凍結形）: 迷ったら「根拠あり」に倒す ──
       task が空/内部キー(_ 始まり)は見ない・1文字の値は偶然一致しすぎるので証拠にしない
       （単位B の _MIN_FRAGMENT=2 と同じ理由）。"""
    if not task:
        return True   # 判定材料が無い ── 誤って★を出さない
    for phrase in _op_match_pool(op):
        if phrase and _phrase_matches_task_allowing_conjugation(phrase, task):
            return True
    for key, value in resolved.items():
        if key.startswith("_"):
            continue
        values = value if isinstance(value, (list, tuple)) else (value,)
        for v in values:
            s = str(v).strip()
            if len(s) >= 2 and alias_store.phrase_is_standalone_in_task(s, task):
                return True
    return False


def suggest_ops(task: str, about: str | None = None, exclude_ops=None) -> list:
    """★ W10 便C2: もしかして提案の候補生成。OP_META の label/synonyms/match_phrases を
       照合プールに組み、ailine_core/suggest.py の純ロジック（語としての厳格一致・
       phrase_is_standalone_in_task と同じ断片ガード）へ薄く委譲する（op 名・一致した
       最長フレーズの文字数降順・最大3・実在 op のみ）。OUT_OF_SCOPE_TERMS を veto
       プールとして渡し、語彙外の強い語が語として現れる依頼は候補ゼロに落とす
       （主対策ではなく補助 ── 主対策は白側の一致要求そのもの）。凍結セットでの測定は
       bench/run_w10_suggest_eval.py、契約は tests/test_suggest_candidates.py。"""
    pool = {op: _op_match_pool(op) for op in OP_META}
    return suggest_candidates.suggest_ops(task, pool, about=about, exclude_ops=exclude_ops,
                                           veto_phrases=OUT_OF_SCOPE_TERMS)


# ---------------------------------------------------------------------------
# ★ W10 便C2 第2段: もしかして提案の判定器（段1の厳格一致(suggest_ops)が空の時だけ呼ばれる）。
#   両面プロンプト（candidates+unsupported）── 2026-08-22 夜の bench 探針（/tmp/judge3.log）で
#   実測済みの形をそのまま製品化する: 陽性対照 18/18・in_vocab recall@1 41/44・
#   true_out_of_vocab 誤提示 0/10・回帰の床 誤提示 3/12。
#   ★ unsupported が非空なら「依頼全体が対応外」という自己申告として candidates を丸ごと
#   捨てる。★ 部分対応（一部は対応・残りは対応外）の自己申告は信用しない ── 実測で 7B の
#   自己棄権（残りは対応外だと自分で申告すること）は 5/6 素通り（並べ替えてから印刷して、
#   のような複合依頼で「印刷」の対応外を自分からは言わない）。残差の検出は別の機械
#   （ailine_core/residue.py・S5）に任せ、ここでは「全く対応できない」の判定だけを信用する。
#   ★ 第5の凍結定数（tests/test_prompt_freeze.py 登録・S6）。W9 の教訓（few-shot/別名を
#   混ぜると壊れる部品）に倣い、op の意味カテゴリの一覧だけを渡す（args のスキーマは不要）。
# ---------------------------------------------------------------------------

SUGGEST_JUDGE_SYSTEM = """あなたは表計算操作の判定係。日本語の依頼が、下の「対応できる操作」の
どれかに言い換えれば一致するかどうかを判定する。あなたの仕事は判定だけで、実際の操作は行わない。

出力形式は必ず {{"candidates": [...], "unsupported": [...]}}。それ以外は書かない。
- candidates: 依頼が言い換えれば一致する操作名（下の一覧の名前そのまま）。一致が無ければ空配列。
  一覧に無い名前を出してはいけない
- unsupported: 依頼**全体**が一覧のどの操作にも一致しない場合だけ、その理由を短い日本語で
  1つ入れる。依頼の一部だけが一覧の操作に一致し、残りが対応外でも、一致した部分がある限り
  unsupported は空配列のままにする（一致した操作は candidates に入れる。残りをどう扱うかは
  あなたの仕事ではない）

★ 自信の無い当てはめより「一致しない」と正直に言う方を選ぶ。JSON のみ出力（説明・markdown
柵は禁止）。

対応できる操作:
{ops}"""


def _judge_ops_catalog() -> str:
    """SUGGEST_JUDGE_SYSTEM の {ops} 差し込み用: OP_META の category/label/synonyms を
       簡潔な一覧にしたもの（args のスキーマは要らない・judge は op を選ぶだけ）。
       ★ OPS_DOC/TRANSLATION_SYSTEM と同じ関係 ── OP_META が伸びても SUGGEST_JUDGE_SYSTEM
       自体のハッシュ（凍結対象）は変わらない。"""
    return "\n".join(
        f"{op}: {meta['category']}・{meta['label']}（{'/'.join(meta['synonyms'])}）"
        for op, meta in OP_META.items())


def build_suggest_judge_messages(task: str, about: str | None) -> list:
    """判定器の messages（system + 実クエリ）を組む。few-shot は持たない（W9 の教訓）。"""
    system = SUGGEST_JUDGE_SYSTEM.format(ops=_judge_ops_catalog())
    about_line = f"\n一次翻訳の要約: 「{about}」" if about else ""
    return [{"role": "system", "content": system},
            {"role": "user", "content": f"依頼: 「{task}」{about_line}"}]


def judge_ops_via_llm(task: str, about: str | None = None) -> list:
    """★ W10 便C2 第2段。厳格一致(suggest_ops)が空の時だけ呼ばれる（+3秒・呼び出し側が
       進捗表示を出す）。戻り値は実在 op のみ・重複除去・最大 suggest_candidates.MAX_CANDIDATES
       件（幻覚 op の構造的封鎖・suggest_ops と同じ考え方）。
       応答が壊れていれば（JSON不正・型不正・ollama 不通）正直に空リストを返す
       （幻覚候補で先へ進まない・呼び出し側は従来の断りへ retreat できる）。"""
    try:
        raw = ollama_generate_json(DEFAULT_MODEL, build_suggest_judge_messages(task, about),
                                    temperature=0.1, num_predict=200)
        data = json.loads(raw)
    except Exception:
        return []
    if not isinstance(data, dict):
        return []
    if data.get("unsupported"):
        return []
    candidates = data.get("candidates")
    if not isinstance(candidates, list):
        return []
    seen = []
    for c in candidates:
        if c in OP_META and c not in seen:
            seen.append(c)
    return seen[:suggest_candidates.MAX_CANDIDATES]


# op → 必須 slot 名のタプル。翻訳直後の slot 欠落チェックと確認行の項目順を兼ねる。
OP_SCHEMA = {
    "SORT": ("col", "order"),
    "COMPUTE_COLUMN": ("operands", "operator"),
    "LOOKUP_FILL": ("target_sheet", "target_col", "source_sheet", "key_col"),
    "AGGREGATE": ("group_col", "value_col"),
    "BOLD": ("target",),
    "FILL_COLOR": ("target", "color"),
    "NUMBER_FORMAT": ("col", "style"),
    "MERGE": ("range",),
    "CHART": ("value_col",),
    "CENTER_ALIGN": ("target",),
    # ★ W6: label/factor は既定値がある任意項目（無指定でも FREEFORM に退避させない）。
    "APPEND_TOTAL": ("col",),
    # ★ W9: INSERT_ROWS は count が既定値ありの任意項目（at だけ必須）。
    #   DRAW_BORDERS/AUTOFIT は引数無し（空タプル＝ any(...) が常に False で FREEFORM に落ちない）。
    "INSERT_ROWS": ("at",),
    # ★ 2026-08-26: 値は列名で受ける ── 実在しない列名を機械が弾ける（幻覚の封鎖）。
    "ADD_ROW": ("at", "values"),
    "DELETE_ROWS": ("at",),
    "DELETE_COLUMN": ("col",),
    # ★ a/b = 入れ替える 2 つの名前。行名か列名かは機械が実表から決める（_axis）。
    "SWAP": ("a", "b"),
    # ★ EXTRACT_COLUMNS: cols(残す列名の並び)。出力シート名は機械が決める（A' 原則）。
    "EXTRACT_COLUMNS": ("cols",),
    # ★ SET_WHERE: col(書き込み先列)・cond_col(条件を見る列)・cmp(比較)。
    #   ★ 書き込む値と閾値は必須 slot に入れない（A' 原則）── 値は依頼文の引用符から、
    #     比較は依頼文の語から、機械が取る（SET_COLUMN_VALUE/EXTRACT と同じ作法）。
    "SET_WHERE": ("col", "cond_col", "cmp"),
    # ★ ADD_COLUMN の必須 slot は無い ── 「原価の右に列を追加して」のように**名前を
    #   言わない依頼が実在する**（Namakoo の実測）。位置も名前も機械が決める/受ける。
    "ADD_COLUMN": (),
    # ★ row=行の名前（中身）・col=列名。値は LLM に決めさせず依頼文から機械が取る（A' 原則）。
    "SET_CELL_VALUE": ("row", "col"),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": ("group_col", "value_col"),
    # ★ 致命3(W10e): value は必須 slot に入れない（LLM に確定させない・A' 原則）。
    #   実際に書き込む値は verify_dsl_args が依頼文の引用符から機械抽出する
    #   （抽出できなければ CLARIFY）。翻訳直後の slot 欠落チェックは col だけを見る。
    "SET_COLUMN_VALUE": ("col",),
    # ★ EXTRACT: col(対象列)・cmp(比較。gte/lte/gt/lt/eq/contains)・value(比較する値)。
    #   出力シート名は verify_dsl_args が機械で決め打ちする（LLM に決めさせない・A' 原則）。
    "EXTRACT": ("col", "cmp", "value"),
    # ★ DEDUP: keys(判定キー列名のリスト)。無ければ CLARIFY（全列一致を既定にしない）。
    #   出力シート名は verify_dsl_args が機械で決め打ちする（EXTRACT と同じ A' 原則）。
    "SPLIT_CELL": ("col", "sep"),
    "DEDUP": ("keys",),
    # ★ 帳票段: REPORT_PER_ROW。template_sheet(人が作った雛形シート)・name_col
    #   (シート名に使う列＝データ行の見出し役)。印({{列名}})の実在検証・出力シート名
    #   （行の値から機械で決め打ち・sanitize_sheet_name/unique_sheet_name）は
    #   verify_dsl_args が行う（EXTRACT/DEDUP と同じ A' 原則 ── LLM に名前を決めさせない）。
    "REPORT_PER_ROW": ("template_sheet", "name_col"),
    # ★ 様式写像段: FORMAT_MAP。template_sheet(人が作った雛形シート)のみ。REPORT_PER_ROW と
    #   違い出力は1枚のシートなので name_col（出力シート名の元列）は要らない。印({{列名}})の
    #   実在検証・出力シート名（雛形名+接尾・sanitize_sheet_name/unique_sheet_name）は
    #   verify_dsl_args が行う（A' 原則 ── LLM に名前を決めさせない）。
    "FORMAT_MAP": ("template_sheet",),
}

# ★ W10c 致命1: 「破壊の関所」（既存列への上書き検知・下の _maybe_warn_target_overwrite）が
#   守る対象を op ごとの if 分岐で持たず宣言駆動にする。旧実装は
#   `if op != "COMPUTE_COLUMN": return None` の1行で、COMPUTE_COLUMN 以外（LOOKUP_FILL 等）
#   は関所が構造的に発火しなかった（監査実測: 存在しない転記先列が無関係な既存列へ
#   解決され、確認なしで上書きされた事故）。
#   ★ 単位C: 宣言の形を「列」から「領域」へ広げた。旧形は (col_key, sheet_key) か None の
#   2択で、「どこに書くか」を列でしか言えなかった。そのため助言の側は宣言を読めず、
#   依頼文の表層語や列の形で代用して毎回誤爆していた（実測2件: APPEND_TOTAL の合計行が
#   必ず「★ 疑わしい: 変更が元データの範囲外です」を出す／AGGREGATE・PIVOT が名指しの
#   入力シートを読むだけなのに「★ …は変更されていません」を出す）。誤爆は条件分岐でなく
#   宣言で消す ── writes（書く領域の種類）と reads_only（参照専用シートの slot 名）を足した。
#   None は廃止し、全 op が WriteTarget を持つ（「安全だから省略した」でなく「対象が無いと
#   確認した」の明示宣言、という旧 None の思想はそのまま writes/col_key=None が担う）。
#   sheet_key が None のときは verify_dsl_args が決めた resolved["_target_sheet"]
#   （後方互換で book_meta の先頭シート）を指す。LOOKUP_FILL だけ target_sheet で明示する。
#   番人: test_op_write_target_declares_all_ops（OP_SCHEMA の全 op に宣言があるか）と
#   test_op_write_target_declarations_are_well_formed（未知の種類・矛盾した組み合わせ）。

# 書く領域の種類の語彙（宣言に書けるのはここに載る種類だけ）。
WRITE_EXISTING_COLUMN = "existing_column"   # 既存列の値を書き換える（＝破壊の関所の対象）
WRITE_NEW_COLUMN = "new_column"             # データの右端に新しい列を作る
WRITE_NEW_ROW_AT_END = "new_row_at_end"     # データ最終行の下に新しい行を足す
WRITE_NEW_SHEET = "new_sheet"               # 新しいシートを作る
WRITE_FORMAT_ONLY = "format_only"           # セルの値は変えない（書式・罫線・列幅・埋め込みグラフ）
WRITE_ROW_SHIFT = "row_shift"               # 行を挿入して既存行を下へずらす（値そのものは残る）
WRITE_REORDER = "reorder"                   # 行を並べ替える（値の集合は保存される）
# ★ 2026-08-26: 「値が消える」を宣言できる種別が **1 つも無かった**（実測で気づいた ──
#   削除を row_shift と宣言したら、番人が「行を動かすだけのはずが値が 3 件消えた」と
#   正しく怒った。宣言の側が嘘をついていた）。消すことを頼まれた op のための種別。
WRITE_REMOVE = "remove"                     # 行/列ごと取り除く（値は減る・詰まる）
# ★ 2026-08-27（Namakoo の手順を自分で通して気づいた）: 「空欄に同じ値を一括で書いた」と
#   いう助言が、**1 セルだけ書く op でも鳴っていた**。空欄の 1 セルを埋めるのは
#   SET_CELL_VALUE の**宣言そのもの**で、疑う理由が無い（事後条件が「変わったのは
#   ちょうど 1 個・座標も宣言どおり」を別に証明している）。鳴らない理由を op 名の if で
#   書かず、**宣言の種別**として持つ ── 新しい op が増えても配線が要らない。
WRITE_SINGLE_CELL = "single_cell"           # 宣言した 1 セルだけを書く
WRITE_KINDS = frozenset({
    WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN, WRITE_NEW_ROW_AT_END, WRITE_NEW_SHEET,
    WRITE_FORMAT_ONLY, WRITE_ROW_SHIFT, WRITE_REORDER, WRITE_REMOVE, WRITE_SINGLE_CELL})


@dataclass(frozen=True)
class WriteTarget:
    """op が「どこに書くか / どこを読むだけか」の宣言。
       writes: 書く領域の種類（WRITE_KINDS の部分集合・空は不可＝必ず何かを宣言する）。
       col_key: 書き込み先列を指す resolved args のキー（既存列を書く op だけが持つ）。
       sheet_key: 対象シート名を指す resolved args のキー（None = resolved["_target_sheet"]）。
       reads_only: 参照専用シートを指す resolved args のキー（そのシートが無変更なのは
                   正常なので、助言側は「変更されていません」を言ってはいけない）。
       cols_key: 書き込み先列**の並び**を指す resolved args のキー（1回で複数の新規列を
                 作る op のため・2026-08-24 SPLIT_CELL）。col_key が単数しか表せず、
                 「範囲外への書き込み」の免除が 1 列にしか効かなかった実測への対応。
       keeps_subject: 対象列を**意図して変えない** op（SPLIT_CELL は元の列を残すのが契約）。
                 助言側が「言及された列が変更されていません」と誤って言うのを止める。
       proves_which_cells: この op の事後条件が「どのセルが変わるべきか」を**両方向**
                 （書かれるべき行は書かれ、それ以外は 1 セルも変わらない）で証明する。
                 True の op では「空欄への同一値の一括書き込み」の助言は何も足さないので
                 出さない ── 助言は**証明が届かない所**にだけ要る。
                 ★ SET_COLUMN_VALUE は False のまま: あちらの事後条件は「全データ行が
                   その値か」しか見ず、元が空欄だったかを問わないので助言が仕事をする。
       col_index_key: 書き込み先列を**位置（1 起点の番号）**で指す resolved args のキー。
                 名前でなく位置で決まる op（ADD_COLUMN・位置は機械が見出しから解決する）
                 のため。col_key/cols_key と同じく、宣言を読むだけで新規列が分かる形に保つ
                 ── op 名の if を増やさない。"""
    writes: tuple = ()
    col_key: str | None = None
    sheet_key: str | None = None
    reads_only: tuple = ()
    cols_key: str | None = None
    keeps_subject: bool = False
    col_index_key: str | None = None
    proves_which_cells: bool = False


OP_WRITE_TARGET = {
    # 並べ替えのみ・値そのものは保存される（書き込み先列という対象は無いと確認した）
    "SORT": WriteTarget(writes=(WRITE_REORDER,)),
    # target 有指定なら既存列・無指定なら新規列（resolved に無い＝関所は素通り）
    "COMPUTE_COLUMN": WriteTarget(writes=(WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN), col_key="target"),
    # target_col が対象シートに実在すれば既存列・無ければその名前で新規列を作る（codegen 参照）
    "LOOKUP_FILL": WriteTarget(writes=(WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN),
                                col_key="target_col", sheet_key="target_sheet",
                                reads_only=("source_sheet",)),
    # 新規シート（SummaryTable）を作るだけ。入力シートは読むだけ＝無変更が正常
    "AGGREGATE": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    "BOLD": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "FILL_COLOR": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "NUMBER_FORMAT": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "MERGE": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # 埋め込みグラフを足すだけ（セルの値も見出しも書かない＝値の書き込み先は無いと確認した）
    "CHART": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "CENTER_ALIGN": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # ★ 単位C(D10): データ末尾の新規行に [ラベル|合計式] を書く（W6・既存列は不可侵）。
    #   合計行は定義上ずっと元データの使用範囲の外に出る＝幽霊データの誤爆源だった。
    #   args の col は「合計する列」であって書き込み先列ではない（col_key に入れない）。
    "APPEND_TOTAL": WriteTarget(writes=(WRITE_NEW_ROW_AT_END,)),
    # 行を挿入するだけ・既存値は下にずれるだけで残る
    "INSERT_ROWS": WriteTarget(writes=(WRITE_ROW_SHIFT,)),
    # ★ 2026-08-26: 追加も削除も既存行を**ずらす**（削除は詰める）。
    # ★★ 2026-08-28（Namakoo が請求書のデモで実測）: 末尾に 1 行足すと毎回 △ に落ちていた。
    #   「変更が元データの範囲外です（A7）」と「空欄への同一値の一括書き込み（× 1 セル）」の
    #   2 件が誤爆する ── どちらも**この op がやると宣言していること**そのものなのに、
    #   宣言に書いていなかったので助言側が知りようがなかった（APPEND_TOTAL・ADD_COLUMN で
    #   同じ形を 2 度直している。3 度目なので宣言を足す側で直す）。
    #   ・末尾に足す回は WRITE_NEW_ROW_AT_END（途中に挿す回は範囲内なので助言は元々出ない）
    #   ・check_add_row は「上は 1 セルも動かず・下は 1 行ずれてそのまま」を両方向で証明する
    #     ＝助言は何も足さない（助言は証明が届かない所にだけ要る）
    "ADD_ROW": WriteTarget(writes=(WRITE_ROW_SHIFT, WRITE_NEW_ROW_AT_END),
                            proves_which_cells=True),
    "DELETE_ROWS": WriteTarget(writes=(WRITE_REMOVE,)),
    "DELETE_COLUMN": WriteTarget(writes=(WRITE_REMOVE,)),
    # ★ 入れ替えは値の多重集合が保存される（reorder）── 前提の番人
    #   (_check_value_multiset) が「動かすだけのはずが値が消えた」を見る。
    "SWAP": WriteTarget(writes=(WRITE_REORDER,)),
    # ★ 途中に挿しても既存の値は消えない（右へずれるだけ）。行の挿入と同じ種別。
    # ★ 2026-08-27（Namakoo が実測）: **新しい列を作る**ことも宣言する。宣言しないと
    #   「変更が元データの範囲外です（D1）」「空欄への一括書き込み」の 2 つが誤爆して
    #   ✓ が △ に落ちる ── 宣言済みの効果を疑わない、という既存の仕組みに乗せる。
    "ADD_COLUMN": WriteTarget(writes=(WRITE_ROW_SHIFT, WRITE_NEW_COLUMN),
                               col_index_key="_at_col", proves_which_cells=True),
    # ★ 新しいシートを作るだけ・入力シートは読むだけ（EXTRACT と同じ）。
    "EXTRACT_COLUMNS": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    # ★ 既存列の一部の行に書く＝破壊の関所の対象（宣言必須）。
    "SET_WHERE": WriteTarget(writes=(WRITE_EXISTING_COLUMN,), col_key="col",
                              proves_which_cells=True),
    # ★ 1 セルは既存列への上書き（前提なし側）。★ ただし「1 セルのはず」は
    #   check_set_cell_value が**変わったセルの数**で証明する（列全体を潰したら落ちる）。
    "SET_CELL_VALUE": WriteTarget(writes=(WRITE_EXISTING_COLUMN, WRITE_SINGLE_CELL),
                                    col_key="col"),
    "DRAW_BORDERS": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    "AUTOFIT": WriteTarget(writes=(WRITE_FORMAT_ONLY,)),
    # 新規シート（DataPilot）を作るだけ。入力シートは読むだけ＝無変更が正常
    "PIVOT": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    # ★ 致命3(W10e): 既存列への一括書き込み＝破壊の関所の対象そのもの（宣言必須）。
    "SET_COLUMN_VALUE": WriteTarget(writes=(WRITE_EXISTING_COLUMN,), col_key="col"),
    # ★ EXTRACT: AGGREGATE/PIVOT と同じ形（新規シートを作るだけ・入力シートは読むだけ）。
    #   出力シート名は動的（col/cmp/value から機械決定）なので OP_DECLARED_SHEET_NAME の
    #   固定表には乗らない ── 単位H(_own_output_headers)側で動的な名前を扱う。
    "EXTRACT": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    # ★ DEDUP: EXTRACT と同じ形（新規シートを作るだけ・入力シートは読むだけ）。
    # ★ 右端に新しい列を「複数」作る（col_key は単数しか表せないので付けない）。
    "SPLIT_CELL": WriteTarget(writes=(WRITE_NEW_COLUMN,), cols_key="_new_cols",
                               keeps_subject=True),
    "DEDUP": WriteTarget(writes=(WRITE_NEW_SHEET,), reads_only=("_target_sheet",)),
    # ★ 帳票段: REPORT_PER_ROW は N 枚の新規シート＋検分シートを作るだけ（新規シートの
    #   前提検査 _check_new_sheet は before に存在しないシートを一切対象にしないので、
    #   1 枚固定を前提にする既存の宣言をそのまま流用できる）。データシート(_target_sheet)と
    #   雛形(template_sheet)はどちらも読むだけ＝助言側の「変更されていません」を抑える。
    "REPORT_PER_ROW": WriteTarget(writes=(WRITE_NEW_SHEET,),
                                    reads_only=("_target_sheet", "template_sheet")),
    # ★ 様式写像段: FORMAT_MAP は1枚の新規シート＋検分シートを作るだけ（REPORT_PER_ROW と
    #   同じ形。データシート(_target_sheet)と雛形(template_sheet)はどちらも読むだけ）。
    "FORMAT_MAP": WriteTarget(writes=(WRITE_NEW_SHEET,),
                               reads_only=("_target_sheet", "template_sheet")),
}


def _op_writes(op: str | None, kind: str) -> bool:
    """op が kind の領域に書くと宣言しているか（op ごとの if を増やさないための問い合わせ口）。"""
    wt = OP_WRITE_TARGET.get(op)
    return bool(wt and kind in wt.writes)


def _declared_reads_only_sheets(op: str | None, resolved: dict | None) -> set:
    """op が「読むだけ」と宣言したシート名の集合（reads_only の slot を resolved で引く）。
       ★ 単位C(D8): 助言側はこれを exclude_sheets として受け取り、読むだけのシートに
       『変更されていません』を言わない。旧実装は `if op == "LOOKUP_FILL"` のハードコード
       2箇所で、AGGREGATE/PIVOT の入力シートは同じ理屈なのに毎回誤爆していた。"""
    wt = OP_WRITE_TARGET.get(op)
    if not wt or not wt.reads_only or resolved is None:
        return set()
    return {resolved[k] for k in wt.reads_only if resolved.get(k)}


# ★★ 単位E: 「利用者が依頼文で名指ししうる対象」の宣言。OP_WRITE_TARGET（どこに**書く**か）
#   とは別の事実なので別の表にする ―― 実証: APPEND_TOTAL は writes=new_row_at_end（書く先は
#   末尾の新規行）だが、利用者が名指しするのは合計する列(col)。同じ事実を2箇所に書くのが
#   禁じ手であって、違う事実を別々に宣言するのは正しい（両者が食い違うことはありえない）。
#   ★ 種別は「解決値の形」で決まる: column=列名 / region=col:列名|row:N|all /
#     row=行番号 / label=書き込むラベル。判定本体は ailine_core/subject.py（純ロジック）。
#   ★ 空タプルは「名指しできる対象が無いと確認した」宣言（DRAW_BORDERS/AUTOFIT は引数無しで
#     表全体・MERGE の range は A1:C1 という機械形式で実在名との照合対象が無い）。
#   ★ 対象シート(_target_sheet)はどの op も持つので、この表でなく共通処理として足す
#     （複数シートのブックだけ・1枚のブックには曖昧さが存在しない＝出力は従来どおり不変）。
#   番人: test_op_subject_slots_declares_all_ops / test_op_subject_slots_are_well_formed。
OP_SUBJECT_SLOTS = {
    "SORT": (("col", SUBJ_COLUMN),),
    # target 無指定（新規列作成）なら resolved に無い＝判定対象そのものが無い。
    # ★ operands は「対象」ではないが計画が実際に使った実在列なので、依頼文の語を消費する
    #   （input 種別・subject.py 参照）。これが無いと「売上から原価を引いた利益列を作って、
    #   利益で降順に」の 2 段目が、誰にも拾われない『売上』『原価』を反証と誤読する。
    "COMPUTE_COLUMN": (("operands", SUBJ_INPUT), ("target", SUBJ_COLUMN)),
    # ★ operator8 ①: source_sheet は「対象」ではないが、依頼文がそのシートだけを名指しした
    #   自然な言い回し（「単価表シートから単価を引っ張ってきて」）で、その言及を
    #   _target_sheet スロットの反証として誤って消費させないよう先に消費する
    #   （INPUT の列版と同じ考え方・sheets 版は SHEET_INPUT）。
    "LOOKUP_FILL": (("target_col", SUBJ_COLUMN), ("key_col", SUBJ_COLUMN),
                     ("source_sheet", SUBJ_SHEET_INPUT)),
    "AGGREGATE": (("group_col", SUBJ_COLUMN), ("value_col", SUBJ_COLUMN)),
    "BOLD": (("target", SUBJ_REGION),),
    "FILL_COLOR": (("target", SUBJ_REGION),),
    "CENTER_ALIGN": (("target", SUBJ_REGION),),
    "NUMBER_FORMAT": (("col", SUBJ_COLUMN),),
    "MERGE": (),
    # ★ グラフ段: category_col は EXTRACT の cmp/value と同じ理由で対象に含めない（多くの
    #   依頼は横軸の列名を明示しない ── 「金額のグラフを作って」は健全系そのもので、
    #   依頼文に無い横軸の既定決定を毎回 ★ で申告すると常時ノイズになる。kind も同様に対象外
    #   （どちらも列名と同種の「利用者が名指しうる対象」ではない）。
    # ★ グラフ段の検分（2026-08-23）: category_col は SUBJ_INPUT ── 依頼文の言及
    #   （「商品ごとの」の商品）を消費するが自身は判定されない。登録しないと、その言及が
    #   value_col への ③（依頼文と矛盾する対象）として誤爆し ✓ が消える
    #   （operator8 ① LOOKUP_FILL の参照シートと同じ形・SHEET_INPUT の列版は既存の SUBJ_INPUT）。
    "CHART": (("value_col", SUBJ_COLUMN), ("category_col", SUBJ_INPUT)),
    # ★ label は「金額の性質の限定（税込み/税抜き）」が依頼文にある時だけ問う（subject.py 参照）。
    "APPEND_TOTAL": (("col", SUBJ_COLUMN), ("label", SUBJ_LABEL)),
    "INSERT_ROWS": (("at", SUBJ_ROW),),
    "ADD_ROW": (("at", SUBJ_ROW),),
    "DELETE_ROWS": (("at", SUBJ_ROW),),
    "DELETE_COLUMN": (("col", SUBJ_COLUMN),),
    # ★ a/b は行名にも列名にもなりうる ── 依頼文が直接名指す「対象」なので
    #   SUBJ_COLUMN 側に置く（EXTRACT の col と同じ扱い・実在照合は verify_dsl_args）。
    "SWAP": (("a", SUBJ_COLUMN), ("b", SUBJ_COLUMN)),
    "SET_WHERE": (("col", SUBJ_COLUMN), ("cond_col", SUBJ_COLUMN)),
    # ★ cols は列名の並び（DEDUP の keys と同じく list を 1 件ずつ展開する仕組みに乗る）。
    "EXTRACT_COLUMNS": (("cols", SUBJ_COLUMN),),
    # ★ name は**これから作る**列なので実在照合の対象にしない（幻覚の封鎖は別口 ──
    #   verify_dsl_args が「同名の列が既に在る」を断る）。位置の基準列は _at_basis に出る。
    "ADD_COLUMN": (),
    "SET_CELL_VALUE": (("col", SUBJ_COLUMN),),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": (("group_col", SUBJ_COLUMN), ("value_col", SUBJ_COLUMN)),
    "SET_COLUMN_VALUE": (("col", SUBJ_COLUMN),),
    # ★ EXTRACT: cmp/value は SET_COLUMN_VALUE の value と同じ理由で対象に含めない
    #   （依頼文が名指しうる「対象」は列だけ・比較の種類や閾値は列名と同種の実在物ではない）。
    "EXTRACT": (("col", SUBJ_COLUMN),),
    # ★ DEDUP: keys は列名のリスト（_subject_slots が list を1件ずつ Slot に展開する・
    #   COMPUTE_COLUMN の operands と同じ仕組み）。SUBJ_COLUMN にする理由: keys は
    #   「計算の入力」ではなく依頼文が直接名指す対象そのもの（EXTRACT の col と同じ扱い）。
    "SPLIT_CELL": (("col", SUBJ_COLUMN),),
    "DEDUP": (("keys", SUBJ_COLUMN),),
    # ★ 帳票段: name_col は依頼文が名指しうる「対象」そのもの（AGGREGATE の group_col と
    #   同じ扱い）。template_sheet は「対象」ではなく、依頼文がその名で言及していれば
    #   消費するだけの入力（LOOKUP_FILL の source_sheet と同じ SHEET_INPUT）。
    "REPORT_PER_ROW": (("name_col", SUBJ_COLUMN), ("template_sheet", SUBJ_SHEET_INPUT)),
    # ★ 様式写像段: FORMAT_MAP は依頼文が名指しうる「対象」を持たない（出力は1枚のシート
    #   そのもので、REPORT_PER_ROW の name_col のような列選びが無い）。template_sheet は
    #   LOOKUP_FILL の source_sheet と同じ SHEET_INPUT（言及していれば消費するだけ）。
    "FORMAT_MAP": (("template_sheet", SUBJ_SHEET_INPUT),),
}


def _subject_slots(op: str, resolved: dict, sheets: list, task: str = "") -> list:
    """宣言(OP_SUBJECT_SLOTS)と resolved から判定対象のスロットを組む。
       ★ 対象シートは全 op 共通で足す ―― ただし**複数シートのブックだけ**
       （1枚しか無いブックに『どのシートか』の曖昧さは存在しない。format_sheet_field/
       describe_target_sheet が沈黙するのと同じ線引き＝単一シート帳票の出力は不変）。"""
    slots = []
    for key, kind in OP_SUBJECT_SLOTS.get(op, ()):
        value = resolved.get(key)
        if value in (None, ""):
            continue
        if isinstance(value, (list, tuple)):   # operands のような複数値の slot は1件ずつ
            slots.extend(Slot(key=key, value=str(v), kind=kind) for v in value if v not in (None, ""))
            continue
        context = ""
        if kind == SUBJ_LABEL:
            # ★ 率が機械確定していれば（＝1.0 でない）、限定は「率:1.1（依頼文: 消費税10%）」
            #   として既に解釈行に出ている ―― 二重に問わない（判定対象そのものにしない）。
            if resolved.get("factor") not in (None, 1.0):
                continue
            # 限定語が対象列の名前に現れていれば、解釈は限定を運んでいる
            # （例: 対象列『税込金額』の合計にラベル『合計』は正しい）。
            context = str(resolved.get("col") or "")
        # ★ 行の位置が「4行目の下」から**引き算で**出たものなら、依頼文と照合する相手は
        #   その導出元（4）── 唯一の産地 row_number_anchor に、決める側と同じ問いを出す。
        _from = None
        if kind == SUBJ_ROW and key == "at":
            _at_a, _n_a, _ = row_number_anchor(task)
            if _at_a is not None and str(_at_a) == str(value):
                _from = _n_a
        slots.append(Slot(key=key, value=str(value), kind=kind, context=context,
                          derived_from=_from))
    target_sheet = resolved.get("_target_sheet")
    # ★ 出どころが「人の明示指定」（--sheet / 画面の選択）なら、依頼文との照合は問わない
    #   ── 人が選んだという事実のほうが、語の一致より強い証拠。
    if resolved.get("_sheet_source") == "cli":
        target_sheet = None
    if target_sheet and len(sheets or []) > 1:
        slots.append(Slot(key="_target_sheet", value=str(target_sheet), kind=SUBJ_SHEET))
    return slots


# ★ 表の行を読む op（連鎖の対象）。書式だけを触る op は元表に掛けたい場合も多いので外す。
PLAN_CHAIN_CONSUMER_OPS = frozenset({
    "AGGREGATE", "PIVOT", "APPEND_TOTAL", "SORT", "COMPUTE_COLUMN", "CHART",
    "EXTRACT", "DEDUP", "REPORT_PER_ROW", "FORMAT_MAP", "SPLIT_CELL", "LOOKUP_FILL",
})


def chain_target_sheet(op: str, task: str, derived_sheets, sheets, headers) -> str | None:
    """複合計画の後段が、前段の作った派生シートを対象にすべきならその名前を返す。

    ★ 3 条件が全部揃った時だけ連鎖する（2026-08-24）:
      ① 直前までの段が派生シート（EXTRACT/DEDUP の出力＝絞り込んだ**同じ表**）を作った
      ② この段が表の行を読む op である
      ③ 依頼文が**シート名を明示していない**（明示があれば人の指定が勝つ）
    ★ 連鎖したことは必ず解釈行に出す ── 黙って対象を変えない。
    ★ なぜ実装したか: ⚠ で止めるだけでは、台帳の最頻の形（期間で絞って集計）が
      一度も完遂しなかった。盲検の 3 人が全員ここで詰まった。
    """
    if not derived_sheets or op not in PLAN_CHAIN_CONSUMER_OPS:
        return None
    # ★ 「はっきりシートとして」名指ししている時だけ人の指定が勝つ。裸の言及
    #   （「売上が60以上」の『売上』＝列名）まで拾うと連鎖が一度も効かない（実測）。
    if sheets_named_explicitly(task or "", list(sheets or [])):
        return None      # ③ 人がシート名を書いている ── その指定が勝つ
    return derived_sheets[-1]["sheet"]


# ★ 道具自身の構造語 ── 依頼者がこれを言っても「列を名指しした」ことにはならない
#   （2026-08-24: 『集計シートで』の『シート』を捏造列と誤断した実測から）。
STRUCTURAL_NOUNS = frozenset({
    "シート", "列", "行", "表", "セル", "ブック", "ファイル", "タブ", "見出し", "項目",
})


def fabricated_subject_refusal(op: str, resolved: dict, book_meta: dict, task: str,
                                sheet: str | None) -> str | None:
    """依頼者が名指しした列がブックに無いのに、機械が**別の実在列で実行しようとして**
       いるなら、適用する前に断る理由を返す（無ければ None）。

    ★ 実測した事故（盲検の査定・2026-08-24）: 日付/取引先/商品/数量/単価 の表に
      「金額で降順に並べ替えて」と頼むと、7B が args を col="数量" に差し替え、
      機械は実在列なので通し、**✓ 機械検証済み**を出して並べ替えた。
      査定者の言葉:「人間なら『金額って列、ありませんけど？』と聞き返す。この道具は
      聞き返さず、原本を書き換え、緑のチェックを付ける。」

    ★ 既存の 3 段階（①照合できた ②無言 ③矛盾）ではこれが ② に落ちていた。② は注記だけで
      ✓ を保つ設計だが、**依頼者は名指ししている**ので無言ではない。無言と捏造を分ける
      材料は既にあった ── 残差（どの引数にも使われなかった依頼文の語）。実測:
        「金額で降順に並べ替えて」(col=数量) → 残差 [金額]  ← 捏造
        「数量で降順に並べ替えて」(col=数量) → 残差 []      ← 正しい
        「降順に並べ替えて」    (col=数量) → 残差 []      ← 無指定（②のまま）
      誤爆しない形で切り分けられる。
    """
    if not task:
        return None
    if resolved.get("_chained_new_column"):
        # ★ 依存つき連鎖の書き換え ── 依頼者の言った名前は、前段が作った列を指している。
        #   捏造ではないので関所の対象外（実測: 「利益で降順に並べ替えて」で前段が
        #   『売上-原価』を自動命名した場合）。
        return None
    slots = [k for k, kind in OP_SUBJECT_SLOTS.get(op, ()) if kind == SUBJ_COLUMN]
    if not slots:
        return None
    headers = list((book_meta.get("headers") or {}).get(sheet, []))
    if not headers:
        return None
    # ① 決まった対象列が、依頼文に語として現れているか
    named = []
    for key in slots:
        value = resolved.get(key)
        for one in (value if isinstance(value, list) else [value]):
            if isinstance(one, str) and one:
                named.append(one)
    if not named or any(n in task for n in named):
        return None      # ① 照合できた ── 関所の対象外
    # ② 依頼文に、どの引数にも使われなかった語が残っているか
    pool = tuple(ph for meta_ in OP_META.values() for ph in meta_.get("match_phrases", ()))
    left = suggest_residue.find_unconsumed_words(task, resolved_args=resolved, pool_phrases=pool) or []
    # ③ その残差のうち、ブックの列名でないもの＝依頼者が言ったのに存在しない列
    # ★ 絞り込み（2026-08-24 の実測で決めた）: 残差が空でないだけでは広すぎた
    #   （既存 1,805 件のうち 24 件が誤爆 ── 「税込み合計を出して」の『税込』『合計』、
    #   「売上高の列を作って」の『売上高』のような**修飾語や新しい列名**まで掴んでいた）。
    #   実際に起きた事故の形だけに絞る: **「〜で」で名指しされた語**。
    #   ★ 正直に残す穴: 「金額を並べ替えて」（を）は掴めない ── 「合計を出して」と
    #   同じ形になり、機械では区別できないため。同じ事故が『を』で再来したら測り直す。
    # ★ 2026-08-24 第三波の実測（実機 CHART が赤くなって発覚）: この関所は**列しか
    #   知らなかった**。「集計シートで部門ごとの棒グラフを作って」の『シート』を
    #   捏造列として掴み、正しい依頼を断っていた（判定に要る三項のうち『実体』が
    #   列だけで、シート名と道具自身の構造語が入っていなかった）。
    #   ★ 実体の第三項を足す ── ① 道具の構造語（列でも捏造でもない）
    #   ② ブックに実在するシート名（『集計シート』のように接尾辞つきで言われる）。
    sheet_names = set(book_meta.get("sheets") or [])
    def _is_real_elsewhere(w: str) -> bool:
        if w in STRUCTURAL_NOUNS:
            return True
        if w in sheet_names:
            return True
        for suffix in ("シート", "表", "タブ"):
            if w.endswith(suffix) and w[: -len(suffix)] in sheet_names:
                return True
        return False
    ghosts = [w for w in left
               if w not in headers and (w + "で") in task and not _is_real_elsewhere(w)]
    if not ghosts:
        return None      # 無指定・修飾語だけ（②のまま・従来どおり通す）
    ghost = "・".join(f"『{w}』" for w in ghosts)
    return (f"依頼文の{ghost}は、シート『{sheet}』の列にありません"
             f"（{named[0]}で実行しようとしていました）。"
             f"ある列: {chr(12289).join(headers)}")


def classify_subject_provenance(op: str, resolved: dict, meta: dict, task: str, a=None) -> list:
    """★ 単位E の入口: この段の対象スロットを①②③に仕分ける（SubjectVerdict のリスト）。
       照合の材料は実在物だけ ―― 対象シートの実在列名・ブックの実在シート名・見出し行。
       ★ 複合計画では meta が『直前までの段を適用した後』の実体なので、前段が作った新規列も
       実在列として材料に入る（症状の検体そのもの: `col:数量*単価` は実在するが、依頼文の
       どの語とも照合できない ―― 断片『数量』『単価』は他の実在列も指しうるので証拠にならない）。
       ★ 「誰も拾わなかった語だけが反証」を段またぎで成立させるため、消費の台帳(Consumed)は
       a（run 全体を通して1つ）に持たせる ―― _sheet_conflict 等と同じ置き場所。a を渡さない
       直接呼び出し（単体テスト）では毎回まっさらな台帳で判定する。"""
    sheets = list((meta or {}).get("sheets") or [])
    target_sheet = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    columns = list(((meta or {}).get("headers") or {}).get(target_sheet, []))
    header_row = ((meta or {}).get("header_rows") or {}).get(target_sheet, 1)
    qualifier = bool(_TAX_INCLUSIVE_RE.search(task or "") or _TAX_EXCLUSIVE_RE.search(task or ""))
    consumed = getattr(a, "_subject_consumed", None) if a is not None else None
    if a is not None and consumed is None:
        consumed = SubjectConsumed()
        a._subject_consumed = consumed
    return classify_slots(_subject_slots(op, resolved, sheets, task or ""), task=task or "",
                           columns=columns, header_row=header_row, sheets=sheets,
                           qualifier_signal=qualifier, consumed=consumed)


# ★ bench/translation_spike.py（実測 v1）と同じ語彙定義（bench 側は比較用に据え置き、
#   本番プロンプトはここが唯一の元）。
OPS_DOC = """SORT: 並べ替え。args: col(列名), order(asc|desc)
COMPUTE_COLUMN: 既存列同士の計算。args: operands(列名2つ), operator(+,-,*,/), target(省略可・実在する列名。
  依頼が「〜に」のように既存列を名指ししたらその列名を入れる。無指定なら新しい列を作る)
  ★ 税込み/税抜き等「1列 × 率」の場合は operands を既存列名1つだけの配列にする。
  operator は * (税込み等、掛ける) か / (税抜き等、割る)。倍率(税率等)の数値はここに入れない
  （数値化はここでは行わない・機械が別途確定する。APPEND_TOTAL の倍率と同じ扱い）
LOOKUP_FILL: 別シートの対応表から値を転記。args: target_sheet, target_col, source_sheet, key_col
  ★ target_col は target_sheet に実在する列名が望ましいが、無ければ依頼文に書かれている
  そのままの列名を入れてよい（実行時にその名前で新しい列を作る）。実在しない列名を
  依頼文に無い別の実在列名に置き換えて誤魔化してはいけない（無関係な列を上書きする事故になる）
AGGREGATE: グループ別に集計表を作る。args: group_col, value_col
BOLD: 太字。args: target("row:行番号" か "col:列名")
FILL_COLOR: 背景色。args: target("row:N"か"col:列名"), color(英語色名)
NUMBER_FORMAT: 数値書式。args: col(列名), style("thousands")
MERGE: セル結合。args: range("A1:C1"形式)
CHART: グラフ。args: value_col(列名),
  kind(省略可・既定"bar"。"bar"=棒, "line"=折れ線/推移, "pie"=円/構成比・割合・内訳。
  依頼文の言い方に合わせて選ぶ),
  category_col(省略可・既定は先頭列。横軸/項目名にする列名)
  ★ kind は依頼文からも機械抽出され、LLM の値と食い違えば機械抽出が優先される
CENTER_ALIGN: 中央揃え。args: target("all" か "col:列名")
APPEND_TOTAL: 列の合計(SUM)を表の最終行の下に追加する（税込み合計等）。args: col(合計する列名),
  label(省略可・既定"合計"。表示ラベル。「税込み合計」等、依頼の言い方をそのまま入れる)
  ★ 合計(SUM)専用。平均・最大・最小など他の統計量は語彙に無い（OUT_OF_VOCAB にする）。
  倍率(税率等)は入れない。数値化はここでは行わない（機械が別途確定する）
INSERT_ROWS: **空行だけ**を挿入する。値を入れる依頼なら ADD_ROW を使う。
  args: at(1起点の行番号。この行の位置に挿入され、既存行は下にずれる), count(省略可・既定1・挿入する行数)
ADD_ROW: 値を入れた行を1本追加する。args: at(1起点の行番号), values(列名→値の対応)
DELETE_ROWS: 行を削除して詰める。args: at(1起点の行番号), count(省略可・既定1)
DELETE_COLUMN: 列を1本削除する。args: col(列名)
DRAW_BORDERS: 依頼文に「けい線/罫線/枠線」という言葉が明示された時だけ使う。表にけい線(格子線)を
  引く。args不要（表全体が対象）★「整えて」「いい感じに」のような具体性の無い依頼には
  絶対に使わない（曖昧なら CLARIFY で確認する）
AUTOFIT: 依頼文に「列幅/幅」という言葉が明示された時だけ使う。列幅を内容に合わせて自動調整する。
  args不要（全列が対象）★ 具体性の無い依頼には使わない（曖昧なら CLARIFY で確認する）
PIVOT: 依頼文に「ピボット」「ピボットテーブル」という言葉が明示された時だけ使う。
  本物のピボットテーブル(DataPilot)を作る。args: group_col(分類する列), value_col(合計する列)
  ★「ピボット」の語が無いグループ別集計（「集計」「まとめる」「小計」「合計がみたい」等、
  ピボットという語を伴わない言い方はすべて）は AGGREGATE を使う（PIVOT は LibreOffice で
  開き直すたび書式が消える癖があるため、書式つきの見栄えが要るなら AGGREGATE の方が適する）
SET_COLUMN_VALUE: 既存列の値を全部、同じ1つの値に書き換える。args: col(書き換える既存列名)
  ★ 実際に書き込む値(value)は依頼文の「」または『』で囲まれた引用を機械が抽出する
  （ここに書いてもよいが、依頼文の引用と食い違えば依頼文側が優先される）。税率等の
  倍率計算(COMPUTE_COLUMN)とは別物 — 依頼が「〜を掛けた」のような計算でなく、
  同じ文字列/値をそのまま代入するだけの依頼はこちらを使う
EXTRACT: 単一条件（列×比較×値）に一致する行だけを新しいシートへ抜き出す。
  args: col(列名), cmp(比較。gte=以上, lte=以下, gt=超, lt=未満, eq=等しい, contains=を含む),
  value(比較する値。数値または文字列。ここでは数値化しなくてよい・機械が確定する)
  ★ 出力シート名は機械が決める（LLM は考えなくてよい）。複数条件(AND/OR)や
  グループごとに分けての抽出は語彙に無い（OUT_OF_VOCAB にする）
DEDUP: 判定キー列（1つ以上）の値の組が同じ行のうち、最初の1行だけを新しいシートへ残す
  （重複除去。元シートの行は消さない・非破壊）。args: keys(判定キー列名のリスト。
  依頼文で名指しされた列だけを入れる)
  ★ keys は依頼文に無い列を推測で入れない。どの列が同じなら重複とみなすか依頼文に
  無ければ CLARIFY で確認する（「全列が一致したら重複」を黙った既定にしない）。
  出力シート名は機械が決める（LLM は考えなくてよい）
REPORT_PER_ROW: 表の1行を雛形シート1枚に写してN枚出す（請求書・見積書・領収書等）。
  args: template_sheet(雛形のシート名), name_col(出力シート名の元になる列)

FORMAT_MAP: 表の各行を雛形シートの様式に写してN行の表を作る（受注CSV→出荷CSV 等）。
SPLIT_CELL: 1つのセルに詰まった複数の値を、区切り(sep)で右の新しい列へ割る。args: col, sep。
  args: template_sheet(雛形のシート名)
"""

# ★ M2c: battery(v1) が実測で取り違えた基本パターンに加え、複合依頼(battery v2)を few-shot で
#   教える。同じ混同/構造を別の言い回しで示す（battery の項目文そのままは使わない＝暗記でなく
#   汎化を確かめる）。
#   ①「引いてくる/転記」は LOOKUP_FILL であって COMPUTE_COLUMN（四則演算）ではない。
#   ②③ 条件付き書式・集計行の追記は語彙に無い操作＝ OUT_OF_VOCAB（曖昧ではないので CLARIFY ではない）。
#   ④ 語彙内(target 付き COMPUTE_COLUMN)＋語彙外の混在。⑤ 語彙内どうしの2連(依存なし)。
TRANSLATION_FEWSHOT = [
    ('対象ブックの構成: {"Sheet": ["商品", "単価"], "商品マスタ": ["商品", "単価"]}\n'
     '依頼: 「商品マスタから商品名を引っ張ってきて明細に入れて」',
     '{"plan": [{"op": "LOOKUP_FILL", "args": {"target_sheet": "Sheet", "target_col": "商品", '
     '"source_sheet": "商品マスタ", "key_col": "商品"}}]}'),
    # ★ W10c 致命2: target_col が対象シートにまだ無いケース（監査実測: これを教えないと
    #   LLM が依頼に無い別の実在列名（この例なら「数量」）を勝手に代入することがあった）。
    ('対象ブックの構成: {"明細": ["商品コード", "数量"], "単価表": ["商品コード", "単価"]}\n'
     '依頼: 「単価表を見て単価を入れて」',
     '{"plan": [{"op": "LOOKUP_FILL", "args": {"target_sheet": "明細", "target_col": "単価", '
     '"source_sheet": "単価表", "key_col": "商品コード"}}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "金額", "在庫"]}\n'
     '依頼: 「金額が1000円未満の行を薄い黄色にして」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "条件付き書式"}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「合計を一番下の行に追加して」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "集計行の追記"}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "数量", "単価", "金額"]}\n'
     '依頼: 「金額に数量×単価を入れて、割引後の金額も出して」',
     '{"plan": ['
     '{"op": "COMPUTE_COLUMN", "args": {"operands": ["数量", "単価"], "operator": "*", "target": "金額"}}, '
     '{"op": "OUT_OF_VOCAB", "about": "割引後の金額"}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額で昇順に並べ替えて、見出し行を太字にして」',
     '{"plan": ['
     '{"op": "SORT", "args": {"col": "金額", "order": "asc"}}, '
     '{"op": "BOLD", "args": {"target": "row:1"}}]}'),
    # ★ W6: APPEND_TOTAL 語彙昇格（監査3回連続失敗の実測による）。
    #   ①明示的な倍率(消費税等)つき ②単純な合計、の両方を教える。
    #   ★ A': 倍率の数値化(1.1等)は LLM に求めない（factor は machine-determined。
    #   verify_dsl_args の extract_rate_factor/lookup_vocab_factor が依頼文から確定する）。
    ('対象ブックの構成: {"Sheet": ["品目", "数量", "単価", "小計"]}\n'
     '依頼: 「税込み合計を一番下に出して（消費税10%）」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": '
     '{"col": "小計", "label": "税込み合計"}}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額の合計を最後に」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額"}}]}'),
    # ★ C9: items_v7 の実測（2026-08-20・round0）で見つかった穴 ── 「税込み」と言っている
    #   のに依頼文に倍率の数字が一切無い依頼（例:「消費税込みでいくらになるか教えて」）を、
    #   モデルが APPEND_TOTAL ではなく OUT_OF_VOCAB に退避させることがあった（既存の
    #   #304/#704 で再現）。倍率が無いなら CLARIFY で聞き返すのが正しい挙動だが、それは
    #   APPEND_TOTAL に翻訳された後でしか起動しない（factor は machine-determined・
    #   verify_dsl_args 側の仕事）。ここでは「倍率の数字が無くても税込み/消費税の言及だけで
    #   op は APPEND_TOTAL」であることを教える（label に税/込を残せば、機械側が倍率不明を
    #   検出して CLARIFY に倒す）。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「消費税込みでいくらになるか教えて」',
     '{"plan": [{"op": "APPEND_TOTAL", "args": {"col": "金額", "label": "消費税込み合計"}}]}'),
    # ★ W9: 検証済みヘルパ4種の語彙昇格。PIVOT/AGGREGATE の分岐（「ピボット」と明示
    #   された時だけ PIVOT・それ以外の集計語は AGGREGATE）と INSERT_ROWS を1例ずつ教える。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門ごとにピボットテーブルで集計して」',
     '{"plan": [{"op": "PIVOT", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門ごとに金額をまとめて」',
     '{"plan": [{"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    ('対象ブックの構成: {"Sheet": ["商品", "金額"]}\n'
     '依頼: 「3行目の前に1行挿入して」',
     '{"plan": [{"op": "INSERT_ROWS", "args": {"at": 3, "count": 1}}]}'),
    # ★ W9 回帰対応（battery v1 再実測で実測）: DRAW_BORDERS/AUTOFIT が「引数不要」なぶん
    #   曖昧な依頼の当てはめ先として誤って選ばれやすかった（「整えて」「いい感じにして」→
    #   誤って DRAW_BORDERS を選ぶ誤断定が発生）。CLARIFY が正しい例を明示する。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「整えて」',
     '{"plan": [{"op": "CLARIFY", "question": "「整える」とは具体的に何をしますか"'
     '"（例: けい線を引く／列幅を合わせる／太字にする）"}]}'),
    # ★ W9 回帰対応: 「ピボット」の語が無いのに漠然とした集計依頼を PIVOT と誤断定した
    #   実測（#14「部署別の売上合計がみたい」→ AGGREGATE が正しい）。同型の言い回しを直接教える。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「部門別の金額合計がみたい」',
     '{"plan": [{"op": "AGGREGATE", "args": {"group_col": "部門", "value_col": "金額"}}]}'),
    # ★ W9 回帰対応: APPEND_TOTAL は合計(SUM)専用。平均等の他統計量は語彙外
    #   （実測: 「平均値を追加」を誤って APPEND_TOTAL にした）。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額の平均値を一番下に追加して」',
     '{"plan": [{"op": "OUT_OF_VOCAB", "about": "平均値の追加"}]}'),
    # ★ W10b 項目3: 税込み/税抜き等「1列 × 率」パターンの語彙昇格（operator 第6回査定の
    #   実測: この言い回しが4種すべて DSL 分類に失敗し自由生成へ退避していた）。
    #   few-shot は最小限（W9 の実測: 足しすぎると別 op の誤断定回帰が出る）。
    #   ★ A': 倍率の数値化は LLM に求めない（factor は machine-determined。verify_dsl_args の
    #   extract_rate_factor/lookup_vocab_factor が依頼文/用語集から機械確定する）。
    #   ★ battery v3 再走で実測した回帰: 「税込み○○を出して」型の言い回しで例を作ると、
    #   APPEND_TOTAL の label 例（「税込み合計を一番下に出して」）と表現が似すぎて、
    #   モデルが APPEND_TOTAL の label フィールドまで省略するようになった(#305 CLARIFY
    #   すべきが確定に化けた)。「〜を追加して」型（新規列作成が明確な言い回し）の例だけ
    #   残し、APPEND_TOTAL の言い回しと重ならないようにする。
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「金額に1.1を掛けた列を追加して」',
     '{"plan": [{"op": "COMPUTE_COLUMN", "args": {"operands": ["金額"], "operator": "*"}}]}'),
    # ★ W10c 高 → 致命3(W10e) で語彙昇格: 「列を全部Xに書き換える」（数値の倍率ではなく
    #   文字列を一律に代入する）は、税込み/税抜き（COMPUTE_COLUMN の1列×率パターン）と
    #   表現が似ているが別物（実測: 「氏名の列を全部『退職済み』に書き換えて」が税率の話と
    #   誤認された事故）。以前は語彙に無く OUT_OF_VOCAB へ退避させていたが、W10e で
    #   SET_COLUMN_VALUE を新設したため、この例も新しい正解へ差し替える（既存 fewshot の
    #   再利用＝件数を増やさない。W9 の教訓＝足しすぎは別 op の誤断定回帰を招く）。
    ('対象ブックの構成: {"Sheet": ["氏名", "部署", "金額"]}\n'
     '依頼: 「氏名の列を全部『退職済み』に書き換えて」',
     '{"plan": [{"op": "SET_COLUMN_VALUE", "args": {"col": "氏名", "value": "退職済み"}}]}'),
    # ★ freeform 廃止バンドル前段: DEDUP の語彙昇格。①判定キー列が依頼文で名指しされて
    #   いれば DEDUP ②名指しが無ければ CLARIFY（全列一致を黙った既定にしない）、の2例で
    #   両方を教える（W9 の教訓＝足しすぎは別 op の誤断定回帰を招くため最小限の2例）。
    ('対象ブックの構成: {"Sheet": ["取引先", "金額"]}\n'
     '依頼: 「取引先が同じ行を重複として除いて」',
     '{"plan": [{"op": "DEDUP", "args": {"keys": ["取引先"]}}]}'),
    ('対象ブックの構成: {"Sheet": ["部門", "金額"]}\n'
     '依頼: 「重複を消して」',
     '{"plan": [{"op": "CLARIFY", "question": '
     '"どの列が同じなら重複とみなしますか（例: 取引先が同じなら重複）"}]}'),
]

TRANSLATION_SYSTEM = """あなたは表計算操作の翻訳係。日本語の依頼を、下の操作語彙を使った「計画」の JSON に翻訳する。
出力形式は必ず {{"plan": [ {{...}}, {{...}}, ... ]}}。それ以外は書かない。
依頼が複数の操作を含むなら、その全部を計画に順番どおり列挙すること。一部を省略してはいけない
（黙って落とすことを禁止・単一の依頼なら要素数1の計画にする）。

計画の各要素は次のどれか一つ:
- {{"op": "<語彙>", "args": {{...}}}}  操作語彙のどれかに当てはまる場合
- {{"op": "OUT_OF_VOCAB", "about": "<何についての依頼か、短く>"}}  操作語彙のどれにも当てはまらない部分
  （条件付き書式・行/シートの削除やコピー・集計行の追記などは語彙外。必ずこの形で計画に残す）
- {{"op": "CLARIFY", "question": "確認文"}}  依頼が曖昧で必須引数を確定できない場合。推測で断定しない

重要な規則:
- 列は必ず「対象ブックの構成」に実在する列名で指定する（番号ではなく）。ただし直前の段が新規作成する
  列を後続の段が参照する場合は、依頼文の言い方のままでよい（実行時に解決する）
- 依頼が既存の列を名指し（「小計に」等）して値を入れる/書き換える場合、COMPUTE_COLUMN の args に
  target(その実在列名) を入れる
- 「整えて」「いい感じにして」のような、何をすればよいか具体的に書いていない依頼は、
  引数が要らない操作(DRAW_BORDERS/AUTOFIT 等)であっても絶対に推測で当てはめない。
  必ず CLARIFY で何をしたいか確認する
- JSON のみ出力（説明・markdown 柵は禁止）

操作語彙:
{ops}"""

TRANSLATION_USER = "対象ブックの構成: {book}\n依頼: 「{text}」"


def build_translation_messages(task: str, book_meta: dict) -> list:
    """翻訳リクエストの messages（system + few-shot + 実クエリ）を組む。"""
    msgs = [{"role": "system", "content": TRANSLATION_SYSTEM.format(ops=OPS_DOC)}]
    for user_ex, assistant_ex in TRANSLATION_FEWSHOT:
        msgs.append({"role": "user", "content": user_ex})
        msgs.append({"role": "assistant", "content": assistant_ex})
    book_desc = json.dumps(book_meta.get("headers", {}), ensure_ascii=False)
    msgs.append({"role": "user", "content": TRANSLATION_USER.format(book=book_desc, text=task)})
    return msgs


def ollama_generate_json(model: str, messages: list, temperature: float = 0.1,
                          num_predict: int = 300) -> str:
    """M2b 翻訳層専用。/api/chat を format="json" 付きで叩く（bench/translation_spike.py
       で実測済みの型。qwen3 系は think:false を足す）。★ 通常生成の ollama_generate とは
       あえて別関数にする — format="json" を全体に効かせると Basic 生成側が壊れるため。
       戻り値は content 文字列（json.loads は呼び出し側 translate_task が行う）。"""
    body = {
        "model": model, "messages": messages, "stream": False, "format": "json",
        "options": {"temperature": temperature, "num_predict": num_predict, "num_ctx": 8192},
    }
    if "qwen3" in model:
        body["think"] = False
    req = urllib.request.Request(f"{OLLAMA}/api/chat", data=json.dumps(body).encode(),
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=300) as r:
        d = json.load(r)
    return d.get("message", {}).get("content", "")


def _normalize_plan_step(data) -> dict:
    """計画の1要素を正規化する（① 翻訳の per-step 版・translate_task の旧 per-op ロジックを継承）。
       ★ 黙って落とさない: 不明な形・語彙外・必須 slot 欠落は必ず何らかの op を持つ dict として
       残す（呼び出し側がその段を FREEFORM(自由生成)/OUT_OF_VOCAB として扱えるように）。
       - モデルが明示した op="OUT_OF_VOCAB" は about を保って素通し。
       - op が語彙外・必須 slot 欠落・非 dict 要素は op="FREEFORM" にする
        （旧 translate_task の『退避は FREEFORM』という分類をそのまま踏襲。単一依頼のときの
        後方互換＝cmd_run の全面自由生成 retreat が変わらないようにするため）。
       ★ W10 前提工事①（architect レビュー致命5-2）: FREEFORM に落ちた**理由**を
       "_fail_reason" として持たせる（"out_of_vocab"=非dict要素/未知op・
       "slot_missing"=必須slot欠落）。呼び出し側(cmd_refuse_vocab_miss)が
       history の failure_kind を下位区分するための材料 ── OUT_OF_VOCAB/CLARIFY の
       dict 形は既存テストの exact-equality を壊さないよう不変のまま。"""
    if not isinstance(data, dict):
        return {"op": "FREEFORM", "args": {}, "_fail_reason": "out_of_vocab"}
    op = str(data.get("op", "")).upper()
    if op == "CLARIFY":
        question = data.get("question")
        return {"op": "CLARIFY", "question": question or "確認が必要です", "args": {}}
    if op == "OUT_OF_VOCAB":
        about = data.get("about")
        return {"op": "OUT_OF_VOCAB", "about": str(about) if about else "内容不明の依頼", "args": {}}
    if op not in OP_SCHEMA:
        return {"op": "FREEFORM", "args": {}, "_fail_reason": "out_of_vocab"}
    args = data.get("args")
    if not isinstance(args, dict):
        # モデルが args で包まず op と slot をフラットに返した場合の救済（寛容に受ける）。
        args = {k: v for k, v in data.items() if k not in ("op", "about", "question")}
    required = OP_SCHEMA[op]
    if any(k not in args or args[k] in (None, "") for k in required):
        return {"op": "FREEFORM", "args": {}, "_fail_reason": "slot_missing"}
    return {"op": op, "args": args}


def translate_task(model: str, task: str, book_meta: dict, temperature: float = 0.1) -> dict:
    """① 翻訳。自然言語タスクを命令言語の「計画」 {"plan": [step, ...]} に翻訳する（M2c）。
       接地: book_meta の実在シート/列名だけを few-shot と一緒に渡す。
       ★ 複合依頼（複数の操作を含む依頼）は plan に全段を列挙する。★ 後方互換: 単一依頼は
       長さ1の計画になる（モデルが "plan" で包まず1操作だけ返した場合もここで長さ1に正規化する）。
       各 step は {"op": <語彙>, "args": {...}} / {"op": "OUT_OF_VOCAB", "about": ...} /
       {"op": "CLARIFY", "question": ...} / {"op": "FREEFORM"}（退避）のいずれか。
       ★ 失敗（API 不通・JSON 不正・空応答）は例外を投げずクラッシュさせない。
       op="FREEFORM" 一段の計画に退避し、呼び出し側が現行の自由生成経路（M2a 助言つき）
       へフォールバックできるようにする。
       ★ W10 前提工事①: この関数レベルの退避（API 不通/JSON 不正/空応答）は
       "_fail_reason": "translate_error" を付ける ── _normalize_plan_step 内の
       out_of_vocab/slot_missing とは別種（翻訳が DSL の形にすら届かなかった）で、
       断りの文言も「照合できませんでした」とは言わない（cmd_refuse_vocab_miss 参照）。"""
    try:
        messages = build_translation_messages(task, book_meta)
        raw = ollama_generate_json(model, messages, temperature=temperature, num_predict=700)
        data = json.loads(raw)
    except Exception:
        return {"plan": [{"op": "FREEFORM", "args": {}, "_fail_reason": "translate_error"}]}
    steps_raw = None
    if isinstance(data, dict) and isinstance(data.get("plan"), list):
        steps_raw = data["plan"]
    elif isinstance(data, dict):
        steps_raw = [data]   # 後方互換: モデルが plan で包まず単一 op を直接返した
    elif isinstance(data, list):
        steps_raw = data
    if not steps_raw:
        return {"plan": [{"op": "FREEFORM", "args": {}, "_fail_reason": "translate_error"}]}
    return {"plan": [_normalize_plan_step(s) for s in steps_raw]}


# ★ W10 便B: 二段目翻訳（op 固定で args だけ埋めさせる）── 頷きの対象（本物の解釈行）と、
#   別名ヒット後の翻訳の両方が使う心臓（REVIEW-20260822-w10-architect.md 1-1 + Namakoo 決裁）。
#   ★ 第4の凍結定数。build_translation_messages/TRANSLATION_SYSTEM/OPS_DOC/TRANSLATION_FEWSHOT
#   は一切触らない（既存3定数の SHA は test_prompt_freeze の番人が別に持つ）。
#   ★ W9 実測（few-shot 1 例で誤断定 27.3%）に倣い、OPS_DOC 全文も別名も見せない ──
#   固定した op 1 つ分のスキーマだけを見せる方が 7B に優しく速い（対象語彙が狭いほど暴走しにくい）。
TRANSLATION_FIXED_OP_SYSTEM = """あなたは表計算操作の翻訳係。操作の種類は既に "{op}" に確定している。
あなたの仕事は、この操作の args（引数）だけを日本語の依頼から埋めることだ。
op を変更してはいけない（別の操作の方が適切だと思っても、必ず "{op}" のまま返す。
op を変えて返しても、呼び出し側は "{op}" を強制するのであなたの判断は無視される）。

出力形式は必ず {{"op": "{op}", "args": {{...}}}}。それ以外は書かない。

この操作の引数: {schema}
列は必ず「対象ブックの構成」に実在する列名で指定する（番号ではなく）。
確定できない引数を推測で埋めてはいけない（分かる範囲だけを埋める）。
JSON のみ出力（説明・markdown 柵は禁止）。"""


# ★ 第二段翻訳でだけ見せる補足（OPS_DOC は 1 文字も増やさない ── 実測で、OPS_DOC に
#   16 行足したら op 一致が 98.1% → 94.2% に落ちた。語彙は増やすほど本流が濁る）。
#   slot 名だけでは意味が伝わらない op にだけ 1 行足す。
_OP_SCHEMA_NOTES = {
    "SWAP": "a と b には、依頼文が『入れ替える』と言っている 2 つの名前を"
             "**そのまま**入れる（行の中身の名前でも、列の見出しでもよい）。"
             "どちらなのかは機械が実際の表を見て決めるので、当てなくてよい。"
             "行番号・列番号・A1 のような座標は入れない。",
    "EXTRACT_COLUMNS": "cols は**残す列の名前の配列**（依頼文に書かれた順）。"
                        "列番号や A1 のような座標は入れない。",
    "SET_WHERE": "col は**印を書き込む列**の名前、cond_col は**条件を見る列**の名前、"
                  "cmp は gte(以上)/lte(以下)/gt(超)/lt(未満)/eq(等しい)/contains(含む) の 1 つ。"
                  "書き込む値と閾値の数字は入れない（機械が依頼文から取る）。",
    "ADD_COLUMN": "args は name（新しい列の見出しに書く名前）だけ。依頼文が名前を"
                   "言っていなければ args を空 {} にする（**作らない**）。"
                   "位置（右/左/末尾）は機械が実表の見出しから決めるので入れない。",
}


def _op_schema_doc(op: str) -> str:
    """固定した op 1 つ分だけのスキーマ文（OPS_DOC 全文でなく OP_SCHEMA[op] から機械生成）。"""
    required = OP_SCHEMA.get(op, ())
    note = _OP_SCHEMA_NOTES.get(op)
    line = (f"{op} は必須の引数なし" if not required
             else f"{op}: args: " + ", ".join(required))
    return line + (chr(10) + note if note else "")


def build_translation_fixed_op_messages(op: str, task: str, book_meta: dict) -> list:
    """② 翻訳の第4の messages ビルダ。build_translation_messages とは別関数
       （op 固定・few-shot 無し・OPS_DOC 全文を見せない ── 混ぜると W9 の 27.3% を繰り返す）。"""
    system = TRANSLATION_FIXED_OP_SYSTEM.format(op=op, schema=_op_schema_doc(op))
    book_desc = json.dumps(book_meta.get("headers", {}), ensure_ascii=False)
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": TRANSLATION_USER.format(book=book_desc, text=task)},
    ]


def translate_task_fixed_op(model: str, op: str, task: str, book_meta: dict,
                             temperature: float = 0.1) -> dict | None:
    """二段目翻訳: op を機械が固定し、その op の args だけを LLM に埋めさせる（W10 便B）。
       頷き（本物の解釈行への確認）と、別名ヒット後の再翻訳の両方が使う心臓。
       ①: LLM が別の op を返しても、返り値の op は固定した op で強制上書きする
       （毒の第一防壁: 頷いた op と違う操作が走る経路を構造的に塞ぐ）。
       ②: 応答が壊れていれば（JSON 不正・op が dict でない・args が dict でない）
       正直に None を返す（幻覚 args で進まない・呼び出し側が CLARIFY に倒せる）。"""
    try:
        messages = build_translation_fixed_op_messages(op, task, book_meta)
        # ★ 検分の差し戻し(2026-08-22): 本流 translate_task と同じ ollama_generate_json
        #   （format=json 強制）を使う ── 素の呼び出しでは 7B が ```json の柵で包み、
        #   正解の応答を parse 失敗で捨てていた（実機 2/5 の真因・片配線の再演）
        raw = ollama_generate_json(model, messages, temperature=temperature, num_predict=300)
        data = json.loads(raw)
    except Exception:
        return None
    if not isinstance(data, dict):
        return None
    args = data.get("args")
    if not isinstance(args, dict):
        return None
    return {"op": op, "args": args}


# --- ② 検証（接地：実在するシート/列名かを機械照合） -------------------------

def _digit_candidates(raw: str, headers: list) -> list:
    """数字表記の列参照を 0 起点/1 起点の両解釈で実在列名に変換した候補（重複除去・順序維持）。
       数字表記でなければ空リスト。"""
    s = str(raw).strip()
    if not re.fullmatch(r"\d+", s):
        return []
    n = int(s)
    cands = []
    for idx in (n, n - 1):        # 0起点読み(n) と 1起点読み(n-1) の両方を試す
        if 0 <= idx < len(headers) and headers[idx] not in cands:
            cands.append(headers[idx])
    return cands


def resolve_col_ref(raw, headers: list) -> tuple:
    """(実在列名 or None, 推定だったか, エラー文 or None)。
       ★ 列名を正とする。実在すればそのまま。数字表記なら 0/1 起点の両候補を試し、
       一意に決まればそれを『推定』として解決、決まらなければ CLARIFY 相当のエラーを返す。"""
    s = str(raw)
    if s in headers:
        return s, False, None
    cands = _digit_candidates(s, headers)
    if len(cands) == 1:
        return cands[0], True, None
    if len(cands) > 1:
        return None, False, f"列『{s}』は複数の解釈が可能で一意に決まりません: {cands}"
    # ★★ 2026-08-28（Namakoo が請求書のデモで実測）: 「F列に…」が断られていた。
    #   人は表計算の座標（A1 方式の列文字）で当たり前に指す。**断るのでなく解ける。**
    #   ★ 見出しに同じ名前が在れば**そちらが勝つ**（上の s in headers が先に返る）ので、
    #     『URL』のような英字の見出しを列文字と誤読する事故は起きない。
    #   ★ 解けたら「推定」として返す（呼び側が解釈行に (推定) を付ける＝黙って決めない）。
    letter = s[:-1] if s.endswith("列") else s
    if letter and re.fullmatch(r"[A-Za-z]{1,2}", letter):
        try:
            idx = column_index_from_string(letter.upper())
        except ValueError:
            idx = 0
        if 1 <= idx <= len(headers):
            return headers[idx - 1], True, None
    known = ", ".join(headers) if headers else "(無し)"
    return None, False, f"列『{s}』がありません。ある列: {known}"


# ★ 単位B 照合の断片ガード（呼び出し側で・単位B 本体＝ailine_core/subject.py は変更しない）。
#   name_matches_task の呼び出し元は「実在するかどうか未確認の raw_target」を渡すことがあり
#   （下の COMPUTE_COLUMN target 経路）、_standalone_occurrence は「他の“実在名”の一部でしか
#   ない出現」しか除外しない ── raw_target が「実在しない・別の複合語」の断片（例:『小計』の
#   『計』）であっても素通りする。ailine_core/subject.py の _MIN_FRAGMENT=2（「1文字の漢字は
#   偶然一致しすぎる」）と同じ理由で、①長さ2未満は最初から証拠にしない、②2文字以上でも
#   依頼文中の全出現が「より長い連続した漢字の内部」（＝別の複合語の内側）でしかないなら
#   証拠にしない。
_CJK_KANJI_RE = re.compile(u"[㐀-䶿一-鿿豈-﫿]")


def _raw_target_not_embedded_in_task(raw_target: str, task: str) -> bool:
    """raw_target の依頼文中の出現のうち、少なくとも1つが「より長い連続した漢字の内部」
       ではない（＝独立した語としての出現がある）なら True。ひらがな/カタカナ/記号は
       日本語の語境界（助詞など）として扱う ―― 漢字が両隣にも続く場合だけ『内部』とみなす。
       出現が無ければ False（そもそも証拠が無い）。"""
    if not raw_target or not task:
        return False
    at = task.find(raw_target)
    if at < 0:
        return False
    n = len(raw_target)
    while at >= 0:
        before_ok = at == 0 or not _CJK_KANJI_RE.match(task[at - 1])
        after_ok = (at + n) >= len(task) or not _CJK_KANJI_RE.match(task[at + n])
        if before_ok and after_ok:
            return True
        at = task.find(raw_target, at + 1)
    return False


def _task_outside_quotes(task: str) -> str:
    """引用符で囲まれた所を空白に潰した依頼文（＝**値でない部分**だけ）。

    ★ 「」の中は「そのセルに書く文字列」── 対象（列・行）の名指しとして読むと、
      値の一部分がたまたま列名と一致しただけで、頼んでいない列が対象になる。
    """
    out = str(task or "")
    for pat in _QUOTE_PATTERNS:
        out = pat.sub(lambda m: " " * len(m.group(0)), out)
    return out


def _task_names_single_real_column(task: str, headers: list) -> str | None:
    """★ operator10 ③ (A' 原則): 依頼文が対象シートの実在列名を独立した語として名指しし、
       候補がちょうど1つに絞れるならその列名を返す（0件/複数件は None ── 曖昧なときは
       機械が断定しない）。断片ガードは COMPUTE_COLUMN の target 判定と同じ
       _raw_target_not_embedded_in_task を再利用する（二重実装しない）。"""
    if not task:
        return None
    # ★★ 2026-08-30（Namakoo が実測）:「A行G列を『税込み金額』に上書き」で、LLM が
    #   実在しない列『税込み金額』を返した。ここは救済のつもりで**別の列『金額』を
    #   採用し**、金額列を丸ごと文字列で潰しにいった（⚠ は出たが止まらない）。
    #   『金額』が依頼文に現れるのは**引用符の中（＝書き込む値）だけ**だった。
    #   ★ 引用符の中は**値**であって、対象の名指しではない ── 証拠に使わない。
    #     （Namakoo の言い方: 「」で囲んだものはセルに入れる値に統一したい）
    outside = _task_outside_quotes(task)
    hits = [h for h in (headers or []) if h and _raw_target_not_embedded_in_task(str(h), outside)]
    return hits[0] if len(hits) == 1 else None


def verify_dsl_args(op: str, args: dict, book_meta: dict, task: str = "", vocab: dict | None = None,
                     target_sheet: str | None = None) -> tuple:
    """② 検証。(ok, resolved_args, inferred_keys, error_message)。
       args のシート/列名が実在するかを機械照合し、実在名に解決する。実在しなければ
       CLARIFY 相当のエラーメッセージを返す（呼び出し側が確認質問として表示する）。
       ★ A': task/vocab は APPEND_TOTAL の倍率(factor)確定専用（他の op は使わない・
       既定値のままで後方互換）。倍率の出典は resolved["_sources"]["factor"] に、
       LLM 由来の値との食い違いは resolved["_warnings"] に積む（戻り値のタプル形は
       変えない＝呼び出し側の unpack を壊さない）。
       ★ 挙動変更#2: target_sheet（省略時は1枚目＝旧挙動と同一）が「対象シート」の
       決定そのもの（resolve_target_sheet が一箇所で決めた値・呼び出し側が渡す）。
       resolved["_target_sheet"] に必ず積む — codegen_dsl・_maybe_warn_target_overwrite・
       _declared_new_column_letter・postcondition の各チェッカーは、みなここを読むだけで
       個別に sheets[0] を仮定しない（LOOKUP_FILL だけは自分自身の target_sheet slot が
       別途あるので、それを resolved["_target_sheet"] にも複製する）。"""
    sheets = book_meta["sheets"]
    headers = book_meta["headers"]
    if not sheets:
        return False, dict(args), set(), "ブックにシートが無い"
    first_sheet = target_sheet if target_sheet in sheets else sheets[0]
    resolved = dict(args)
    resolved["_target_sheet"] = first_sheet
    if book_meta.get("_sheet_source"):
        resolved["_sheet_source"] = book_meta["_sheet_source"]
    inferred: set = set()

    def resolve_in(key: str, sheet_name: str):
        val, was_inferred, err = resolve_col_ref(resolved.get(key), headers.get(sheet_name, []))
        if err:
            # ★ operator10 ③ (A' 原則): LLM の値がこのシートに実在しなくても、依頼文が
            #   対象シートの実在列名を独立した語として一意に名指ししていれば、そちらを
            #   採用する（別シートの同名列に汚染された答えより、依頼文+実在照合が勝つ ──
            #   factor/cmp/kind と同じ「機械抽出が LLM に勝つ」系譜）。依頼文にも手掛かりが
            #   無ければ従来どおり CLARIFY で止める。
            #   ★ 誤爆防止: 既に別のスロット（例: CHART の value_col）が採用した列名は
            #   候補から外す（同じ列名が依頼文に1回出るだけで、値列の言及を横取りして
            #   欠けたスロットへ誤って流用しないため）。
            original = resolved.get(key)
            claimed = {v for k2, v in resolved.items()
                       if k2 != key and isinstance(v, str) and not k2.startswith("_")}
            candidates = [h for h in headers.get(sheet_name, []) if h not in claimed]
            rescued = _task_names_single_real_column(task, candidates)
            if rescued is not None:
                resolved[key] = rescued
                resolved["_warnings"] = resolved.get("_warnings", []) + [
                    f"LLM が返した列『{original}』は対象シートに無いため、依頼文が名指しする"
                    f"列『{rescued}』を採用しました"
                ]
                return None
            # ★ operator8 ③: 列解決の失敗時だけ敗者復活（book_meta に _row_scan が無い
            #   単体テスト等では常に None ＝従来のまま）。
            return _header_row_hint_for_missing_col(book_meta, sheet_name, resolved.get(key)) or err
        resolved[key] = val
        if was_inferred:
            inferred.add(key)
        return None

    def check_sheet(key: str):
        name = resolved.get(key)
        if name not in sheets:
            return f"シート『{name}』がありません。あるシート: {', '.join(sheets)}"
        return None

    if op == "SORT":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        if resolved.get("order") not in ("asc", "desc"):
            return False, resolved, inferred, f"順序『{resolved.get('order')}』は asc/desc のどちらでもありません"
        # ★★ 2026-08-29（Namakoo が実測）: 合計行まで並べ替えの範囲に入れていたので、
        #   降順にすると合計（一番大きい）が**先頭へ飛び**、その式が
        #   `=SUM(#REF!:INDEX(E:E,ROW()-1))` に壊れた。番人は止めたが、人は並べ替えられない。
        #   ★ 合計行は「データ行ではない」── 並べ替えの対象から外し、最下行に残す。
        #     判定は既存の凍結規則を借りる（total_rows_in → row_has_total_word）。
        #   ★ 見つけたら**必ず画面に出す**（黙って行を外さない）。
        _s_sheet = resolved.get("_target_sheet") or first_sheet
        _s_hr = int((book_meta.get("header_rows") or {}).get(_s_sheet, 1) or 1)
        _s_tot = total_rows_in(book_meta, _s_sheet, _s_hr)
        if _s_tot:
            _s_end = min(_s_tot) - 1
            if _s_end >= _s_hr + 1:
                resolved["_sort_end_row"] = _s_end
                # ★ 開示は**解釈行**に出す（警告ではない）。SET_WHERE が合計行を外す時と
                #   同じ口を使う ── 警告にすると決裁③で ✓ が △ に落ち、合計行のある表を
                #   並べ替えるたびに「確かめきれていない」と言うことになる。
                #   ★ 宣言どおりに動いて検算も通っているのだから、それは ✓ でよい。
                resolved["_skip_rows"] = list(_s_tot)
                resolved["_skip_label"] = ("合計行 " + "、".join(
                    f"{r}行目" for r in _s_tot) + "（データ行でないため並べ替えません）")
        # ★ 並べ替えで「指す先の中身が変わる式」を名指しする（★ 付き＝決裁③で ✓→△）。
        #   ここは疑いなので警告でよい ── 合計行の除外（開示）とは性質が違う。
        _s_last = resolved.get("_sort_end_row") or 10 ** 7
        if (_dw := reference_drift_warning(book_meta, _s_sheet,
                                            row_lo=_s_hr + 1, row_hi=_s_last)):
            resolved["_warnings"] = resolved.get("_warnings", []) + [_dw]

    elif op == "COMPUTE_COLUMN":
        operands = resolved.get("operands")
        # ★ W10b 項目3: 税込み/税抜き等「1列 × 率」パターン。operands が列名1つだけの
        #   配列なら『既存列×倍率』とみなす（2列の四則演算とは別モード）。倍率(factor)は
        #   APPEND_TOTAL と同じ A' 原則で LLM から受け取らず機械確定する
        #   （extract_rate_factor/lookup_vocab_factor・regex のみ）。
        single_factor_mode = isinstance(operands, list) and len(operands) == 1
        if not single_factor_mode and not (isinstance(operands, list) and len(operands) == 2):
            return False, resolved, inferred, "演算対象が2つの列名になっていません"

        if single_factor_mode:
            v, was_inferred, err = resolve_col_ref(operands[0], headers.get(first_sheet, []))
            if err:
                return False, resolved, inferred, err
            resolved["operands"] = [v]
            if was_inferred:
                inferred.add("operands")
            if resolved.get("operator") not in ("*", "/"):
                return False, resolved, inferred, (
                    f"演算子『{resolved.get('operator')}』は列1つの計算（税込み/税抜き等）"
                    "では * か / のみ対応です")

            llm_factor_raw = resolved.pop("factor", None)
            text_factor, text_snippet = extract_rate_factor(task)
            vocab_factor, vocab_term = (None, None)
            if text_factor is None:
                vocab_factor, vocab_term = lookup_vocab_factor(task, vocab or {})

            sources: dict = {}
            if text_factor is not None:
                resolved["factor"] = text_factor
                sources["factor"] = f"依頼文: {text_snippet}"
            elif vocab_factor is not None:
                resolved["factor"] = vocab_factor
                sources["factor"] = f"用語集: {vocab_term}"
            else:
                # ★ W10c 高: 依頼文に率らしい語が一切無いのに「1列×率」（税込み/税抜き専用）へ
                #   分類されているのは、分類そのものが誤っている可能性が高い（実測: 「氏名の
                #   列を全部『退職済み』に書き換えて」のような値の一括書き換え依頼が、税率の
                #   話と誤認されて COMPUTE_COLUMN の単列モードに落ちることがあった）。
                #   その場合は「倍率が分からない」でなく、分類自体を疑う文言に変える
                #   （率を要求する op に分類されたのに率の手がかりが無い＝CLARIFY の理由を
                #   正直に言い換える。指示は意図・保証は機械＝プロンプト側だけに頼らない）。
                if not _RATE_SIGNAL_RE.search(task or ""):
                    return False, resolved, inferred, (
                        f"依頼「{task}」は『{v}』列に何らかの倍率（税率等）を掛ける操作として"
                        "解釈しましたが、依頼文に倍率らしき手がかりが見当たりません。"
                        "列の値をそのまま書き換える操作は今のところ対応していません。"
                        "倍率を掛ける処理であれば、依頼文に率を書く（例:「消費税10%」）か、"
                        "用語集に登録してください（例: ailine vocab add 消費税 1.1）"
                    )
                # ★ 致命③(2026-08-23レビュー): 敗者復活（_resolve_tax_rescue・APPEND_TOTAL と
                #   共有）。第一照合（上の text_factor/vocab_factor）の優先順は変えない ──
                #   ここに来るのはその両方が外れ、かつ依頼文に率らしき語（税/込を含む）が
                #   ある場合だけ（片配線の解消: 登録済みの税語彙で「登録してください」と
                #   嘘をつかない）。
                tax_factor, tax_term, tax_err = _resolve_tax_rescue("依頼", task or "", vocab)
                if tax_factor is not None:
                    resolved["factor"] = tax_factor
                    sources["factor"] = f"用語集: {tax_term}（依頼『{task}』の税に適用）"
                elif tax_err:
                    return False, resolved, inferred, tax_err
                else:
                    return False, resolved, inferred, (
                        "倍率（税率等）が分かりません。依頼文に率を書く（例:「消費税10%」）か、"
                        "用語集に登録してください（例: ailine vocab add 消費税 1.1）"
                    )
            if resolved["factor"] <= 0:
                return False, resolved, inferred, f"倍率『{resolved['factor']}』は正の数でなければなりません"
            if sources:
                resolved["_sources"] = sources
            # ★ W10c 中: 新規列の見出しの自然化。旧実装は見出しを f"{op1}{operator}{factor:g}"
            #   （例:「金額*1.1」）という数式風の文字列にしていた（査定で名指し）。target
            #   無指定(新規列作成)かつ依頼文が税込み/税抜きと分かる言い方の場合だけ、
            #   その日本語ラベルを見出しに使う（A' 原則: LLM を使わず正規表現の有無のみで
            #   決める。手がかりが無ければ従来どおりの数式風見出しにフォールバック）。
            if not resolved.get("target"):
                # ★★ 2026-08-30: **依頼文に名前が在るなら、それが名前**（作らない）。
                #   下の「税込〜」は、人が名前を書かなかった時だけの間に合わせ。
                _asked = new_column_name_from_task(task, headers.get(first_sheet, []))
                if _asked:
                    resolved["_new_col_label"] = _asked
                elif _TAX_INCLUSIVE_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税込{v}"
                elif _TAX_EXCLUSIVE_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税抜{v}"
                elif _RATE_KEYWORD_RE.search(task or ""):
                    resolved["_new_col_label"] = f"税込{v}" if resolved["operator"] == "*" else f"税抜{v}"
            if llm_factor_raw not in (None, ""):
                try:
                    llm_factor = float(llm_factor_raw)
                except (TypeError, ValueError):
                    llm_factor = None
                if llm_factor is not None and abs(llm_factor - resolved["factor"]) > 1e-9:
                    mfactor = resolved["factor"]
                    resolved["_warnings"] = [
                        f"LLM が返した倍率({llm_factor:g})と機械抽出の倍率({mfactor:g})が"
                        f"食い違うため機械抽出({mfactor:g})を採用しました"
                    ]
        else:
            new_operands = []
            for o in operands:
                v, was_inferred, err = resolve_col_ref(o, headers.get(first_sheet, []))
                if err:
                    return False, resolved, inferred, err
                new_operands.append(v)
                if was_inferred:
                    inferred.add("operands")
            resolved["operands"] = new_operands
            if resolved.get("operator") not in ("+", "-", "*", "/"):
                return False, resolved, inferred, f"演算子『{resolved.get('operator')}』が不明です"
        # ★ M2c: target(任意) — 依頼が既存列を名指し（「小計に」等）した場合はその列に書く。
        #   無指定なら従来どおり新規列（codegen_dsl 側で分岐）。
        # ★ W3: target が実在しない場合、翻訳が「新しい列の名前」（例:「利益列を作って」の
        #   『利益』）を target と誤って埋めていることが多いと実測された（qwen2.5-coder:7b が
        #   『既存列に書く/新規に作る』の区別を安定して守らない）。実在しない＝一意に決まらない
        #   （複数解釈で曖昧）のとは別の理由なので、その場合だけ target を無指定として扱い
        #   新規列作成にフォールバックする（推測で断定しない CLARIFY の原則は、真に曖昧な
        #   ケース＝digit_candidates の複数一致にだけ残す）。
        # ★ W3 改定(2026-08-20): 上の前提（実在しない target＝ほぼ捏造）が古くなった実測が
        #   出た。「金額を数量×単価で埋めて」（金額列がまだ実在しない構成）で翻訳は正しく
        #   target:"金額" を返す。だが無条件に捨てる旧実装はそれも落とし、新規列が
        #   『数量*単価』に自動命名されていた（利用者が名前を言っているのに無視される・
        #   2026-08-19 のデモ制作で3回踏んだ実害）。★ 依頼文を判定者にする: raw_target が
        #   依頼文に実在する語として機械照合できるなら（単位B の name_matches_task を再利用
        #   ── 素朴な in 判定はしない。「税込金額を…」+target「金額」のような片方向の部分
        #   文字列の穴は単位B が塞いだ形そのものなので、同じ判定を2箇所に書かず呼ぶ）、
        #   捏造ではなく利用者の指名とみなして新規列の名前として使う。依頼文に語が無ければ
        #   従来どおり捏造とみなして捨てる（W3 本来の防御は生きている）。
        if resolved.get("target"):
            raw_target = resolved["target"]
            v, was_inferred, err = resolve_col_ref(raw_target, headers.get(first_sheet, []))
            if err:
                if "一意に決まりません" in err:
                    return False, resolved, inferred, err
                # ★ 単位B 照合の断片ガード（呼び出し側・上の _raw_target_not_embedded_in_task
                #   docstring 参照）: raw_target は「実在するか未確認」の生の文字列なので、
                #   name_matches_task を素通しに使う前に (1) 1文字を弾き (2) 依頼文中の全出現が
                #   「他の複合語（実在列とは限らない）の内部」でしかないなら弾く。
                if (len(raw_target) >= 2
                        and _raw_target_not_embedded_in_task(raw_target, task)
                        and name_matches_task(raw_target, task, others=headers.get(first_sheet, []))):
                    resolved["_new_col_label"] = raw_target
                del resolved["target"]
            elif (task_asks_to_add_a_column(task)
                    and not name_matches_task(v, task,
                                               others=headers.get(first_sheet, []))):
                # ★★ 2026-09-02（130 件の器を広げて初めて見えた・実測）:
                #   「単価の右に、数量と単価をかけた**列を作って**」で、一段目が
                #   target='メモ'（実在するが**空**の列）を返し、道具は新しい列を作らずに
                #   **その列へ書いて ✓ を出していた**。頼んでいない場所に書いている。
                #   ★ W3 は「**実在しない** target ＝ ほぼ捏造」を捨てる。抜けていたのは
                #     「**実在するが、依頼文に無い**」列 ── そこだけ素通りだった。
                #   ★ 判定に語彙の一覧は要らない: 依頼文が「作る」と言っているか
                #     （閉じた文法）と、その名前が依頼文と機械照合できるか（既存の
                #     provenance 層）の 2 つだけ。新しい言い回しが来ても足すものは無い。
                #   ★ 道具は既に気づいていた（★で開示していた）── **止めていなかった**だけ。
                del resolved["target"]          # → 新しい列を作る（自動命名 or 依頼文の名前）
            else:
                resolved["target"] = v
                if was_inferred:
                    inferred.add("target")
                    # ★ W10a 項目3: 数字指定→列名解決の元の表記を残す（解釈要約の表示用・
                    #   例:「列5」→「在庫」列と解決した時、確認行の直後にその経緯を見せる）。
                    resolved["_target_raw"] = raw_target
        # ★★ 2026-09-02: 2 項の演算（売上 − 原価）でも、依頼文の名前を拾う。
        #   名前の抽出は**倍率の枝（税込/税抜）にしかなかった**ので、
        #   「売上から原価を引いた**利益**の列を作って」の見出しが
        #   『売上-原価』（式そのもの）になっていた ── A' 原則が抜けた形。
        #   ★ 人が名前を書いていない時だけ従来どおり式風の見出しに落ちる。
        if not resolved.get("target") and not resolved.get("_new_col_label"):
            # ★ 見出しの一覧を**渡さずに**呼ぶ ── 「既に在る名前」も受け取りたいから。
            _asked_c = new_column_name_from_task(task, [], require_position=False)
            _heads_c = [str(h) for h in (headers.get(first_sheet) or [])]
            if _asked_c and _asked_c in _heads_c:
                # ★★ 2026-09-02（実測で捕まえた実害）: 依頼した名前が**既に在る**時、
                #   「新しい列の名前ではない」として捨てて自動命名に落ちていた。すると
                #   「売上から原価を引いた利益の列を作って」を 2 回実行すると、
                #   1 回目『利益』・2 回目『売上-原価』になり、**見出しが違うので
                #   「見出しも値も同一の列を作りました」の関所が鳴らない** ──
                #   値がそっくり同じ列が静かに 2 本目として増え、✓ まで出た。
                #   （盲検 operator 査定が見つけた事故「不安でもう一回実行」の再来）
                #   ★ 意味で考えても、これは「作る」ではなく「**もう在る**」。
                #     その列を計算し直す依頼と読み、既存の**上書きの関所**に載せる
                #     ── 新しい関所も新しい終了コードも作らない。
                resolved["target"] = _asked_c
            elif _asked_c:
                resolved["_new_col_label"] = _asked_c

    elif op == "LOOKUP_FILL":
        if (err := check_sheet("target_sheet")):
            return False, resolved, inferred, err
        if (err := check_sheet("source_sheet")):
            return False, resolved, inferred, err
        # ★ 挙動変更#2: 旧実装はここで「対象シートは1枚目のみ対応しています」と拒否していた
        #   （散在した『1枚目固定』の一つ・査定の致命そのもの）。LOOKUP_FILL は元々
        #   target_sheet を自分の必須 slot として名前で受け取り check_sheet で実在確認まで
        #   済ませているので、この制限を外すだけで対応できる。resolved["_target_sheet"] は
        #   LOOKUP_FILL 自身の target_sheet を正とする（他 op 用の一般解決 first_sheet より
        #   こちらを優先 — 依頼文に転記先/参照元の2シート名が両方出て一般解決が曖昧に
        #   フォールバックしていても、LOOKUP_FILL のここでの解決は影響を受けない）。
        resolved["_target_sheet"] = resolved["target_sheet"]
        # ★★ 塊③(2/2)・中核 op 致命2（2026-08-24 の盲検）:
        #   書き手（VLookupFromTable）は参照表の**列1（2 番目の列）を値**と決め打ちし、
        #   検算（check_lookup_fill）も同じ決め打ちで期待値を作る ── やる側と見る側が
        #   同じ思い込みを共有しているので必ず一致する（恒真）。
        #   実測: マスタ = 商品/区分/**単価**（3 列目）で「単価を転記して」と頼むと、
        #   **単価の列に「果物」が入って ✓** が出た。数値であるべき列に文字列が入る。
        #   ★ 商品コード/商品名/単価 のような 3 列マスタは実務でごく普通。
        #   → 書く**前**に前提を照合する: 頼まれた列名が参照表の 2 列目でなければ断る。
        #   ★ 見出しが読めない参照表では断らない（根拠が無い時に止めない）。
        #   ★ 断らずに**開示する**理由（実測で 1 度誤爆した）:
        #     事故の形   マスタ=[商品,区分,単価] → 2 列目は「区分」
        #     正しい依頼 明細  =[商品,数量,単価] → 2 列目は「数量」
        #     どちらも「2 列目 ≠ 頼まれた列」で、**列の位置だけでは区別できない**。
        #     断ると正しい依頼まで止める（既存検体で実証）。判定は変えず、
        #     何が書かれるかを名指しして ✓ を △ に降ろす（決裁③の機構に乗せる）。
        _src_headers = list((book_meta.get("headers") or {}).get(resolved["source_sheet"], []))
        _want = resolved.get("target_col")
        if len(_src_headers) > 2 and isinstance(_want, str) and _want in _src_headers:
            _at = _src_headers.index(_want) + 1
            if _at != 2:
                resolved["_warnings"] = resolved.get("_warnings", []) + [
                    f"参照表『{resolved['source_sheet']}』では『{_want}』が {_at} 列目ですが、"
                    f"この転記は 2 列目を値として読む仕組みです（1 列目がキー・2 列目が値）。"
                    f"実際に書き込まれるのは『{_src_headers[1]}』の値です ── "
                    f"意図と違う場合は、キーと『{_want}』だけの表を用意してください"]
        # ★ W10c 致命2: target_col は COMPUTE_COLUMN の target と違い OP_SCHEMA 上は必須
        #   slot なので、LLM は「存在しないなら空にする」を選べない。実測（監査再現）:
        #   対象シートに『単価』列がまだ無いのに転記を頼むと、LLM がそれと無関係な
        #   *実在する*既存列（例:「数量」）の名前を代わりに返すことがある。resolve_col_ref
        #   は実在列名なら無条件で素通しするため、これだけでは見分けられない（そのまま
        #   進めると「数量」が確認なしで上書きされる事故になる）。
        #   ここでは「実在するから信用する」をやめ、根拠を要求する:
        #   ①依頼文にその列名が書かれている ②転記元（source_sheet）の値列
        #   （VLookupFromTable ヘルパの仕様どおり常に列1＝2番目の列）と同じ名前
        #   のどちらかが無いと、実在列であっても信用しない。
        target_headers = headers.get(resolved["target_sheet"], [])
        source_headers = headers.get(resolved["source_sheet"], [])
        value_col_hint = source_headers[1] if len(source_headers) > 1 else None
        raw_target_col = resolved.get("target_col")
        raw_str = str(raw_target_col) if raw_target_col not in (None, "") else ""
        exists = raw_str in target_headers
        mentioned = bool(raw_str) and raw_str in task
        matches_value_col = value_col_hint is not None and raw_str == value_col_hint

        if exists and (mentioned or matches_value_col):
            pass   # 根拠つきで実在列を指名＝そのまま使う（上書き注意は破壊の関所が別途担当）
        elif not exists:
            cands = _digit_candidates(raw_str, target_headers)
            if len(cands) == 1:
                resolved["target_col"] = cands[0]   # 数字表記の推定は従来どおり許容
                inferred.add("target_col")
            elif mentioned:
                # ★ 依頼文にも同じ列名が書かれている＝新規作成が正しい解釈（COMPUTE_COLUMN の
                #   target 無指定＝新規列と同じ考え方）。target_col はそのまま残し、
                #   codegen_dsl 側で新規列として作る。
                pass
            else:
                known = ", ".join(target_headers) if target_headers else "(無し)"
                return False, resolved, inferred, (
                    f"転記先の列『{raw_target_col}』が『{resolved['target_sheet']}』シートに"
                    f"見つかりません。ある列: {known}。新しい列として作る場合は、依頼文に"
                    f"その列名を書いてください（例:「{raw_target_col}という列を作って転記して」）"
                )
        else:
            # ★ 実測の事故そのもの: exists=True だが根拠が無い（依頼文にも書かれておらず、
            #   転記元の値列とも一致しない）＝上書き対象を取り違えている可能性が高い。
            hint = f"（参照表『{resolved['source_sheet']}』の値の列は『{value_col_hint}』です）" \
                if value_col_hint else ""
            return False, resolved, inferred, (
                f"転記先の列『{raw_target_col}』は実在しますが、依頼文にその列名が見当たらず、"
                f"転記元の値とも対応が確認できません{hint}。上書き対象を取り違えている"
                "可能性があるため、意図した列名を依頼文に明記してください"
            )
        if (err := resolve_in("key_col", resolved["target_sheet"])):
            return False, resolved, inferred, err

    elif op == "AGGREGATE":
        if (err := resolve_in("group_col", first_sheet)):
            return False, resolved, inferred, err
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err

    elif op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        target = str(resolved.get("target", ""))
        if target == "all":
            if op != "CENTER_ALIGN":
                return False, resolved, inferred, f"対象『all』は {OP_LABELS[op]} では未対応です"
        elif target.startswith("row:"):
            n = target[4:]
            if not (n.isdigit() and int(n) >= 1):
                return False, resolved, inferred, f"行番号『{n}』が不正です"
            # ★ 単位I: 契約文(1743)・codegen(2588)は CENTER_ALIGN に row: を認めておらず、
            #   ここだけが少数派だった（実測: row:N を通した先で codegen が素の traceback）。
            #   codegen に row: を実装するのは別作業 ―― ここは契約文に合わせて拒否するだけ。
            if op == "CENTER_ALIGN":
                return False, resolved, inferred, (
                    f"対象『{target}』は {OP_LABELS[op]} では未対応です"
                    "（col:列名 か all を使ってください）"
                )
        elif target.startswith("col:"):
            colname = target[4:]
            v, was_inferred, err = resolve_col_ref(colname, headers.get(first_sheet, []))
            if err:
                return False, resolved, inferred, err
            resolved["target"] = f"col:{v}"
            if was_inferred:
                inferred.add("target")
        elif target.startswith("cell:"):
            # ★ cell: は**機械が作る**形（LLM には出させない）。R,C は 1 起点。
            try:
                _r, _c = (int(x) for x in target[5:].split(","))
            except ValueError:
                return False, resolved, inferred, f"セルの指定『{target}』が読めません"
            if _r < 1 or _c < 1:
                return False, resolved, inferred, f"セルの指定『{target}』が範囲外です"
        else:
            return False, resolved, inferred, (
                f"対象『{target}』の形式が不明です（row:N / col:列名 / cell:行,列 / all）")
        if op == "FILL_COLOR":
            color = str(resolved.get("color", "")).lower()
            if color not in COLOR_MAP:
                return False, resolved, inferred, f"色『{color}』は未対応です。使える色: {', '.join(sorted(COLOR_MAP))}"
            resolved["color"] = color

    elif op == "NUMBER_FORMAT":
        # ★ 2026-08-29: 行にも掛けられるようにした（行と列は軸違い）。
        #   行が指定されている回は列を要求しない。
        if resolved.get("row_number"):
            _nf_sheet = resolved.get("_target_sheet") or first_sheet
            _nf_hr = int((book_meta.get("header_rows") or {}).get(_nf_sheet, 1) or 1)
            _nf_row = int(resolved["row_number"])
            if _nf_row <= _nf_hr:
                return False, resolved, inferred, (
                    f"{_nf_row}行目は見出し行（{_nf_hr}行目）またはその上です")
            resolved["_row_index"] = _nf_row
            resolved.pop("col", None)
            resolved["_at_basis"] = f"{_nf_row}行目"
        elif (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        if resolved.get("style") != "thousands":
            return False, resolved, inferred, f"書式『{resolved.get('style')}』は未対応です（対応: thousands）"

    elif op == "MERGE":
        if not re.fullmatch(r"[A-Za-z]{1,3}\d+:[A-Za-z]{1,3}\d+", str(resolved.get("range", ""))):
            return False, resolved, inferred, f"範囲『{resolved.get('range')}』の形式が不正です（例: A1:C1）"

    elif op == "CHART":
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err
        # ★ グラフ段①: kind も A' 原則の中へ（cmp と同じ作法）。依頼文からの機械抽出が
        #   非 None かつ LLM の kind と食い違えば機械が勝つ（EXTRACT の cmp と同じ形）。
        llm_kind_raw = str(resolved.get("kind") or "").strip().lower()
        mechanical_kind = extract_chart_kind_from_task(task)
        if mechanical_kind is not None and mechanical_kind != llm_kind_raw:
            kind = mechanical_kind
            resolved["_warnings"] = resolved.get("_warnings", []) + [
                f"LLM が返した種類({llm_kind_raw or '(空)'})と依頼文の機械抽出({mechanical_kind})が"
                f"食い違うため機械抽出({mechanical_kind})を採用しました"
            ]
        else:
            kind = llm_kind_raw or "bar"
        if kind not in _CHART_KINDS:
            return False, resolved, inferred, (
                f"グラフ種類『{resolved.get('kind')}』は {'/'.join(_CHART_KINDS)} のどれでもありません"
            )
        resolved["kind"] = kind
        # ★ グラフ段②: category_col(省略可・既定は先頭列)。指定があれば実在列検証。
        raw_cat = resolved.get("category_col")
        if raw_cat in (None, ""):
            first_col = (headers.get(first_sheet) or [None])[0]
            if first_col is None:
                return False, resolved, inferred, f"シート『{first_sheet}』に列がありません"
            resolved["category_col"] = first_col
            inferred.add("category_col")
        elif (err := resolve_in("category_col", first_sheet)):
            return False, resolved, inferred, err

    elif op == "APPEND_TOTAL":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        # ★ W6: label は既定値を持つ任意項目。ここで確定させ、codegen/事後条件/
        #   確認行の全部に同じ既定解決を一貫して渡す。
        resolved["label"] = str(resolved.get("label") or "合計")
        label = resolved["label"]

        # ★★ 2026-08-29（Namakoo が実測）: 合計行が**既に在る**表で「単価列の合計行に
        #   単価の合計を書いて」と頼むと、10 行目に『単価合計』という**別の行**が増えた。
        #   ★ 真因: 合計行を「データ行」と数えて、その下に足していた。
        #   ★ 合計行が 1 つに決まり、その列がまだ空なら、**その行に書く**（行は増やさない）。
        #     判定は既存の凍結規則を借りる（total_rows_in → row_has_total_word）──
        #     ここで新しい規則を書かない。同じことを 2 箇所が決めると必ずずれる。
        _tot_sheet = resolved.get("_target_sheet") or first_sheet
        _tot_hr = int((book_meta.get("header_rows") or {}).get(_tot_sheet, 1) or 1)
        _tot_rows = total_rows_in(book_meta, _tot_sheet, _tot_hr)
        if len(_tot_rows) == 1:
            _tr = _tot_rows[0]
            _theads = [str(h) for h in ((book_meta.get("headers") or {}).get(_tot_sheet) or [])]
            _tidx = _theads.index(resolved["col"]) + 1 if resolved["col"] in _theads else 0
            _cur = None
            if _tidx:
                try:
                    with BookView(Path(book_meta["path"])) as _bv:
                        _cur = _bv.sheet(_tot_sheet).cell(row=_tr, column=_tidx).value
                except Exception:
                    _tidx = 0
            if _tidx and (_cur in (None, "") or str(_cur).startswith("=SUM(")):
                resolved["_at_row"] = _tr
                resolved["_at_basis"] = f"既にある合計行＝{_tr}行目（行は増やしません）"
                # ★ ラベルは**その行に既に在る物**が正（LLM の案『単価合計』で検算しない）。
                try:
                    with BookView(Path(book_meta["path"])) as _bv2:
                        _lbl = _bv2.sheet(_tot_sheet).cell(row=_tr, column=1).value
                    if _lbl not in (None, ""):
                        resolved["label"] = str(_lbl)
                        label = resolved["label"]
                except Exception:
                    pass
            # ★★ 2026-08-29: ここで「既に値が入っています」と**断るのはやめた**。
            #   既存の番人（事後条件の算術の検算＝二重計上に ✓ を出さない／単位F の関所）が
            #   同じ事故を既に止めていて、断りを重ねると**その番人の出番が消える**
            #   ── 過去の事故を守っている検体が通らなくなる（実測で 3 本落ちた）。
            #   ★ 埋められる時だけ埋め、それ以外は今までどおり深い番人に任せる。

        # ★ A': factor は LLM から受け取らない。LLM が返した値(あれば)はいったん取り出して
        #   おき、機械抽出/用語集の結果と食い違う場合だけ WARN として記録する（常に機械が勝つ）。
        llm_factor_raw = resolved.pop("factor", None)

        text_factor, text_snippet = extract_rate_factor(task)
        vocab_factor, vocab_term = (None, None)
        if text_factor is None:
            vocab_factor, vocab_term = lookup_vocab_factor(task, vocab or {})

        sources: dict = {}
        if text_factor is not None:
            resolved["factor"] = text_factor
            sources["factor"] = f"依頼文: {text_snippet}"
        elif vocab_factor is not None:
            resolved["factor"] = vocab_factor
            sources["factor"] = f"用語集: {vocab_term}"
        else:
            resolved["factor"] = 1.0

        if resolved["factor"] <= 0:
            return False, resolved, inferred, f"倍率『{resolved['factor']}』は正の数でなければなりません"

        # ★ 恒真式の番人（最優先）: label が「税」/「込」を含むのに倍率が確定できず既定
        #   1.0 のままだと、税抜き金額に「税込み」ラベルが付いた恒真の誤りを事後条件が
        #   pass にしてしまう（args 基準の検証だから）。ここで機械的に CLARIFY へ倒す
        #   （語リストは 税/込 の2語で凍結・むやみに増やさない）。
        # ★ operator8 ②: CLARIFY に倒す直前に敗者復活（lookup_vocab_tax_factor・
        #   docstring 参照）。第一照合（上の text_factor/vocab_factor）の優先順は変えない
        #   ―― ここに来るのはその両方が外れた場合だけ。
        if resolved["factor"] == 1.0 and any(k in label for k in ("税", "込")):
            tax_factor, tax_term, tax_err = _resolve_tax_rescue("ラベル", label, vocab)
            if tax_factor is not None:
                resolved["factor"] = tax_factor
                sources["factor"] = f"用語集: {tax_term}（ラベル『{label}』の税に適用）"
            elif tax_err:
                return False, resolved, inferred, tax_err

        if sources:
            resolved["_sources"] = sources
        if llm_factor_raw not in (None, ""):
            try:
                llm_factor = float(llm_factor_raw)
            except (TypeError, ValueError):
                llm_factor = None
            if llm_factor is not None and abs(llm_factor - resolved["factor"]) > 1e-9:
                mfactor = resolved["factor"]
                resolved["_warnings"] = [
                    f"LLM が返した倍率({llm_factor:g})と機械抽出の倍率({mfactor:g})が"
                    f"食い違うため機械抽出({mfactor:g})を採用しました"
                ]

    # --- ★ 2026-08-26: 表の基本操作 3 種（追加・行削除・列削除）---------------
    elif op in ("ADD_ROW", "DELETE_ROWS"):
        # ★★ 2026-08-27（Namakoo が実測）: 「みかんの下に梨を追加して」が動かなかった。
        #   位置を**行番号**でしか受け取れないのに、人は相対で言う。LLM に数えさせると
        #   外し、空行だけの INSERT_ROWS に落ちていた。
        #   ★ 分担を変える: LLM は「誰の隣か」を言うだけ／**行番号は機械が実表を数えて決める**
        #     （列名の解決を機械 3 段でやっているのと同じ形）。
        #   ★ 機械が決めた位置は LLM の数字より優先する ── 実表を見た側が正しい。
        _sheet0 = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr0 = int((book_meta.get("header_rows") or {}).get(_sheet0, 1) or 1)
        _at_anchor, _anchor_note = resolve_row_anchor(task, book_meta, _sheet0, header_row=_hr0)
        if _anchor_note and _at_anchor is None:
            return False, resolved, inferred, _anchor_note
        if _at_anchor is not None:
            resolved["at"] = _at_anchor
            resolved["_at_basis"] = _anchor_note
        at_raw = str(resolved.get("at", "")).strip()
        if not (at_raw.isdigit() and int(at_raw) >= 1):
            return False, resolved, inferred, f"行番号『{resolved.get('at')}』が不正です（1以上の整数）"
        resolved["at"] = int(at_raw)
        _sheet = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr = int((book_meta.get("header_rows") or {}).get(_sheet, 1) or 1)
        # ★ 見出し行より上を触らせない（表の骨格を壊す操作は受け付けない）。
        if int(at_raw) <= _hr:
            return False, resolved, inferred, (
                f"{at_raw}行目は見出し行（{_hr}行目）またはその上です ── "
                "見出しを壊す操作は受け付けません")
        if op == "DELETE_ROWS":
            c = resolved.get("count")
            if c in (None, ""):
                resolved["count"] = 1
                inferred.add("count")
            else:
                cs = str(c).strip()
                if not (cs.isdigit() and int(cs) >= 1):
                    return False, resolved, inferred, f"削除行数『{c}』が不正です（1以上の整数）"
                resolved["count"] = int(cs)
        else:
            # ★ 値は**列名で**受ける。実在しない列名はここで弾く（幻覚の封鎖）。
            vals = resolved.get("values")
            # ★ 2026-08-27（実測）: LLM は values を**並び**で返すことがある
            #   （['梨', 600, 300]）。列名の対応は**機械が付けられる** ── 左から順に
            #   当てる。多すぎる時だけ断る（推測で余りを捨てない）。
            #   ★ 決めた対応は解釈行に出す（_values_label）── 黙って割り当てない。
            _headers_now = (book_meta.get("headers") or {}).get(
                resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]) or []
            if isinstance(vals, (list, tuple)):
                if len(vals) > len(_headers_now):
                    return False, resolved, inferred, (
                        f"入れる値が {len(vals)} 個ありますが、列は {len(_headers_now)} 本です"
                        f"（ある列: {"、".join(map(str, _headers_now))}）")
                # ★ 2026-08-27（実測・俺が入れた壊し方）: LLM は埋まらない列を None で
                #   返すことがある（['梨', None, None]）。そのまま渡すと codegen が
                #   `str(None)` を書き、セルに**文字列 "None"** が入った。
                #   ★ 指定の無い列には**何も書かない**（空欄のままにする）。
                #   ★ 事後条件はこの壊れ方を捕まえていた（rc=1）── 番人は効いていたが、
                #     壊れた物を作ってから気づく形だったので、入口で落とす。
                vals = {str(h): v for h, v in zip(_headers_now, vals)
                         if v is not None and v != ""}
                resolved["values"] = vals
                inferred.add("values")
            if not isinstance(vals, dict) or not vals:
                return False, resolved, inferred, (
                    "入れる値が読み取れません（列名と値の組で書いてください）")
            headers = (book_meta.get("headers") or {}).get(
                resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]) or []
            unknown = [k for k in vals if str(k) not in [str(h) for h in headers]]
            if unknown:
                return False, resolved, inferred, (
                    f"列『{"、".join(map(str, unknown))}』がこの表にありません"
                    f"（ある列: {"、".join(map(str, headers))}）")
            # ★ 同上: 値が空の列は書かない（"None" という文字列を作らない）。
            resolved["values"] = {str(k): v for k, v in vals.items()
                                   if v is not None and v != ""}
            resolved["_headers"] = [str(h) for h in headers]
            resolved["_values_label"] = "／".join(
                f"{k}={v}" for k, v in resolved["values"].items())
            # ★★ 2026-09-02（README の「既知の問題」に自分で書いていた）:
            #   追加した行に既存の式が引き継がれず、利益列が**空のまま**だった。
            #   宣言した値は正しいので ✓ は正しいが、人の期待とは違う。
            #   ★ 引き継ぐのは「全データ行が式を持つ列」だけ ── 形で決める（列挙しない）。
            #     合計列は E2..E7 が直値なので自然に外れる。
            #   ★ **黙ってやらない。**解釈行に出す（_inherit_label）。
            if op == "ADD_ROW" and resolved.get("at"):
                _ih_sheet = resolved.get("_target_sheet")
                _ih_hr = int(resolved.get("_header_row") or 1)
                _ih_cols, _ih_from = formula_columns_to_inherit(
                    book_meta, _ih_sheet, _ih_hr, int(resolved["at"]),
                    set(resolved["values"].keys()))
                if _ih_cols and _ih_from:
                    resolved["_inherit_cols"] = _ih_cols
                    resolved["_inherit_from"] = _ih_from
                    _hd = resolved["_headers"]
                    resolved["_inherit_label"] = "／".join(
                        (_hd[c] if 0 <= c < len(_hd) else f"{c + 1}列目")
                        for c in _ih_cols) + f"（{_ih_from}行目から）"

    elif op == "SET_CELL_VALUE":
        # ★ 2026-08-27（Namakoo「梨の売上にピンポイントで入れたい」）:
        #   SET_COLUMN_VALUE は**列を丸ごと**同じ値にする op で、1 セルを狙えなかった。
        _sheet_c = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _headers_c = [str(h) for h in
                       ((book_meta.get("headers") or {}).get(_sheet_c) or [])]
        _row_name = str(resolved.get("row", "")).strip()
        _col_name = str(resolved.get("col", "")).strip()
        _row_no = resolved.get("row_number")
        # ★★ 2026-08-30（Namakoo「行と列による一意の指定も出来た方がいい」→ 実測）:
        #   「1行F列を「税込金額(10%)」にして」で、第二段は col に**書き込む値**を入れて
        #   きた（col=『税込金額(10%)』）。列の名前と値が入れ替わっている。
        #   ★ 依頼文が英字で列を名指ししているなら、それが正 ── 機械が実表から決める
        #     （行番号を機械が決めるのと同じ分担・LLM の欄の中身に頼らない）。
        _letters = {m.group(0) for m in _re_a1_col_word.finditer(_task_outside_quotes(task))}
        if len(_letters) == 1:
            _cand = next(iter(_letters)).replace("列", "").strip()
            _v2, _inf2, _err2 = resolve_col_ref(_cand, _headers_c)
            if not _err2:
                _col_name = _v2
        # ★ 2026-08-28: 列は**列文字でも**指せる（「F列に」）── resolve_col_ref が解く。
        if _col_name not in _headers_c:
            _v, _inf, _err = resolve_col_ref(_col_name, _headers_c)
            if _err:
                return False, resolved, inferred, (
                    f"列『{_col_name}』がこの表にありません"
                    f"（ある列: {"、".join(_headers_c)}）")
            if _inf:
                inferred.add("col")
            _col_name = _v
        if not _row_name and not _row_no:
            return False, resolved, inferred, (
                "どの行かが読み取れません（行の名前か行番号で指してください）")
        # ★★ 2026-08-28: 行番号で指された時は**番号を正**にする（人が数えて言っている）。
        #   ★ 名前も同時に在るなら、その行に本当にその名前が在るかを確かめる ──
        #     三項（依頼・宣言・実体）。食い違ったら書かずに断る。
        if _row_no:
            _hr_c = int((book_meta.get("header_rows") or {}).get(_sheet_c, 1) or 1)
            _path_c = book_meta.get("path")
            try:
                with BookView(Path(_path_c)) as _bvc:
                    _wsc = _bvc.sheet(_sheet_c)
                    _lastc, _colsc = data_extent(_wsc, _hr_c)
                    _rowvals = [str(_wsc.cell(row=int(_row_no), column=c).value or "").strip()
                                 for c in range(1, _colsc + 1)] if int(_row_no) <= _lastc else None
            except Exception as e:
                return False, resolved, inferred, f"表を読めませんでした（{type(e).__name__}）"
            # ★★ 2026-08-30（Namakoo「行と列による一意の指定も出来た方がいい。
            #   ピンポイントに操作できるようになる」）: それまで見出し行は一律で断って
            #   いたので、**列の名前を直す手段が 1 つも無かった**（実測: 計算列の見出しが
            #   「金額*1.1」に化けた表を、人が直せない）。
            #   ★ 人が**行番号と列を書いて名指しした**のは、いちばん強い証拠 ──
            #     見出しでも書かせる。ただし黙って書かない（解釈行で必ず言う）。
            #   ★ LLM が推した行・名前から解いた行では、この道は開けない（下の else 側）。
            if _rowvals is None and int(_row_no) > _hr_c:
                return False, resolved, inferred, (
                    f"{_row_no}行目はこの表の範囲外です（見出しは{_hr_c}行目・データは"
                    f"{_hr_c + 1}〜{_lastc}行目）")
            if int(_row_no) < _hr_c:
                return False, resolved, inferred, (
                    f"{_row_no}行目は見出し行（{_hr_c}行目）より上です ── "
                    "表の外には書けません")
            if int(_row_no) == _hr_c:
                if task_names_a_row_number(task) != int(_row_no):
                    return False, resolved, inferred, (
                        f"{_row_no}行目は見出し行です ── 見出しの名前を変えるなら、"
                        f"行番号と列で名指ししてください"
                        f"（例:「{_hr_c}行G列を「新しい名前」にして」）")
                resolved["_writes_header"] = True
            # ★ 実測: 第二段は row に**行番号そのもの**を入れてくることがある（"7"）。
            #   それは名前ではないので、名前としては扱わない（食い違い扱いにしない）。
            if _row_name.isdigit() or _row_name == str(_row_no):
                _row_name = ""
            if _row_name and _row_name not in _rowvals:
                return False, resolved, inferred, (
                    f"{_row_no}行目に『{_row_name}』がありません"
                    f"（その行: {"、".join(v for v in _rowvals if v)}）── "
                    "行番号と名前が食い違っています")
            _hitrow = int(_row_no)
            _note_c = f"{_hitrow}行目（依頼文の行番号）"
            if resolved.get("_writes_header"):
                # ★ 見出しを書き換える回に「対象の行:取引先」と出ると読み手を誤らせる
                #   （実測で出た）── 何をしているのかを、その言葉で言う。
                _row_name = "見出し"
                _note_c = f"{_hitrow}行目（見出し行）── **見出しの名前を変えます**"
            else:
                _row_name = _row_name or (_rowvals[0] if _rowvals and _rowvals[0] else str(_hitrow))
        else:
            # ★ 行が実在し・1 つに決まることを**適用前に**確かめる（推測で別の行に書かない）。
            _hitrow, _note_c = _resolve_named_row(book_meta, _sheet_c, _row_name)
            if _hitrow is None:
                return False, resolved, inferred, _note_c
        resolved["_row_index"] = _hitrow
        # ★ 見出しを書き換えると、**その列は元の名前で引けなくなる** ── 位置を残す
        #   （検算は名前でなく座標で見る）。実測で「列『税込み金額』が見つからない」と
        #   落ちた（書き込み自体は成功していたのに）。
        if _col_name in _headers_c:
            resolved["_col_index"] = _headers_c.index(_col_name) + 1
        resolved["row"] = _row_name
        resolved["col"] = _col_name
        resolved["_headers"] = _headers_c
        resolved["_at_basis"] = _note_c
        # ★ 値は LLM に決めさせず、依頼文から機械が取る（A' 原則・SET_COLUMN_VALUE と同じ線）。
        #   ★ ただし 1 セルなので**裸の数字も受ける** ── 「梨の売上を2000にして」を
        #     引用符の有無で断るのは、道具の都合を人に押し付けている（実測の困りごと）。
        _lit = extract_quoted_literal(task)
        if _lit is None:
            # ★★ 2026-08-29（Namakoo が実測・直した先で出た穴）:
            #   「丸山工業の締め日を**2026/08/31**にして」で **31** が書かれて ✓ が出た。
            #   裸の数字を拾う正規表現が、日付の**末尾だけ**を掴んでいた。
            #   ★ 先に機械の引き算（依頼文から、既に分かっている物を引く）を通す ──
            #     こちらは値を**丸ごと**取るので、途中で切れない。
            _lit = bare_value_from_task(task, _row_name, _col_name, _headers_c)
        if _lit is None:
            _m = _re_bare_number.search(task or "")
            _lit = _m.group(1) if _m else None
        if _lit is None:
            # ★★ 2026-08-29（Namakoo が実測）: 「丸山重工の右にPCパーツ」が
            #   『文字なら「」で囲んで』で断られていた。**引用符は道具の都合**であって、
            #   人の書き方の問題ではない（この repo が何度も自分に言ってきた線）。
            #   ★ A' 原則の芯は「引用符が在ること」ではなく「**依頼文に在る値**であること」。
            #     だから条件をそちらへ置き直す: 依頼文に literal で在り・見出しの語でなく・
            #     行の名前でもない値なら、引用符が無くても受ける。
            #   ★ それでも**画面に出してから書く**（「こう読みました」に値が出る）。
            _cand = str((resolved.get("value") if resolved.get("value") is not None else ""))
            _cand = _cand.strip()
            _bad = {str(h) for h in _headers_c} | {str(_row_name), str(_col_name)}
            if _cand and _cand in (task or "") and _cand not in _bad:
                _lit = _cand
        if _lit is None:
            return False, resolved, inferred, (
                "書き込む値が依頼文から読み取れません"
                "（依頼文に書かれている値をそのまま使います ── "
                "紛らわしいときは「」で囲んでください）")
        resolved["value"] = _lit
        # ★ 2026-08-27（実測）: `_is_number` は**型**で見るので、文字列 "2000" は False。
        #   ここへ来る値は必ず文字列なので、**数字として読めるか**で判定する。
        #   ★ これを外すと `'2000'` が文字列でセルに入り、下流の SUM が静かに壊れる
        #     （この repo が何度も測ってきた形）。
        try:
            resolved["_write_numeric_value"] = float(str(_lit).replace(",", ""))
            resolved["_write_numeric"] = True
        except ValueError:
            pass

    elif op == "DELETE_COLUMN":
        name = str(resolved.get("col", "")).strip()
        headers = (book_meta.get("headers") or {}).get(
            resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]) or []
        if name not in [str(h) for h in headers]:
            return False, resolved, inferred, (
                f"列『{name}』がこの表にありません（ある列: {"、".join(map(str, headers))}）")
        if len([h for h in headers if str(h) != ""]) <= 1:
            return False, resolved, inferred, "列が 1 本しかないので削除できません"
        resolved["col"] = name
        resolved["_headers"] = [str(h) for h in headers]

    # --- ★ 2026-08-27: 列の抽出（残す列を依頼文の実在列から機械が拾う）------------
    elif op == "EXTRACT_COLUMNS":
        _sheet_x = resolved.get("_target_sheet") or first_sheet
        _hdrs_x = [str(h) for h in (headers.get(_sheet_x) or [])]
        # ★ 値は LLM に作らせない: **依頼文に現れる実在の列名**を、出現順に機械が拾う。
        #   LLM の cols は「候補の当たり」としてだけ使い、実在照合を通ったものだけ採る。
        _asked = [c for c in _hdrs_x if c and c in (task or "")]
        _llm_cols = resolved.get("cols")
        if isinstance(_llm_cols, str):
            _llm_cols = [x.strip() for x in _llm_cols.split(",") if x.strip()]
        _llm_cols = [str(c) for c in (_llm_cols or []) if str(c) in _hdrs_x]
        _cols_x = _asked or _llm_cols
        if not _cols_x:
            return False, resolved, inferred, (
                "残す列が依頼文から読み取れません"
                f"（ある列: {'、'.join(_hdrs_x)}）── 列名をそのまま書いてください")
        if len(_cols_x) >= len(_hdrs_x):
            return False, resolved, inferred, (
                "全部の列が指定されています（抜き出す意味がありません）")
        # ★ 依頼文の出現順に並べる（人が書いた順で出す ── 表の順に勝手に直さない）
        _cols_x = sorted(set(_cols_x), key=lambda c: (task or "").find(c))
        resolved["cols"] = _cols_x
        resolved["_cols_label"] = "・".join(_cols_x)
        resolved["_headers"] = _hdrs_x
        resolved["_header_row"] = int((book_meta.get("header_rows") or {}).get(_sheet_x, 1) or 1)
        resolved["_source_headers"] = tuple(_hdrs_x)
        resolved["_new_sheet"] = _EXTRACT_SHEET_NAME_FORBIDDEN_RE.sub(
            "_", "・".join(_cols_x) + "だけ")[:31]

    # --- ★ 2026-08-27: 条件つき書換（値と比較は機械が依頼文から取る）--------------
    elif op == "SET_WHERE":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        # ★ 2026-08-27（Namakoo「置き換えができない」）: 「『A』を『B』に」の形は、
        #   **同じ列の中で A の行だけを B にする**＝条件つき書換の特別な場合
        #   （条件列＝書き込み先列・比較は「等しい」）。新しい op は要らない。
        #   ★ 条件も値も**引用の対**から機械が取る（LLM に決めさせない）。
        _pair = extract_replace_pair(task)
        if _pair:
            resolved.setdefault("cond_col", resolved["col"])
        if (err := resolve_in("cond_col", first_sheet)):
            return False, resolved, inferred, err
        if _pair:
            resolved["cmp"], resolved["cond_value"], resolved["value"] = "eq", _pair[0], _pair[1]
            resolved["_sources"] = {**resolved.get("_sources", {}),
                                     "value": f"依頼文: 「{_pair[0]}」→「{_pair[1]}」"}
            resolved["_headers"] = [str(h) for h in (headers.get(first_sheet) or [])]
            resolved["_header_row"] = int(
                (book_meta.get("header_rows") or {}).get(first_sheet, 1) or 1)
            resolved["_cond_label"] = (f"『{resolved['cond_col']}』が『{_pair[0]}』の行"
                                        f"（→『{_pair[1]}』に）")
            # ★ 合計行は**データ行ではない**ので対象から外す。外したことは必ず画面に出す。
            resolved["_skip_rows"] = total_rows_in(book_meta, first_sheet, resolved["_header_row"])
            if resolved["_skip_rows"]:
                resolved["_skip_label"] = ("合計行 " + "、".join(
                    f"{r}行目" for r in resolved["_skip_rows"]) + "（データ行でないため）")
            _hits_r = _rows_matching(book_meta, first_sheet, resolved["cond_col"], "eq",
                                      _pair[0], resolved["_header_row"])
            if _hits_r is not None:
                if not _hits_r:
                    return False, resolved, inferred, (
                        f"『{resolved['cond_col']}』に『{_pair[0]}』の行がありません"
                        "（ファイルには何も書いていません）")
                resolved["_match_rows"] = _hits_r
                resolved["_match_label"] = (
                    f"{len(_hits_r)} 行（{'、'.join(str(r) for r in _hits_r[:5])}行目"
                    + ("…）" if len(_hits_r) > 5 else "）"))
            return True, resolved, inferred, None
        # ★ 比較は機械が勝つ（EXTRACT と同じ作法・LLM の写し間違いで境界行が混入する）
        _llm_cmp = str(resolved.get("cmp", "")).strip().lower()
        _mech_cmp = extract_cmp_from_task(task)
        if _mech_cmp is not None and _mech_cmp != _llm_cmp:
            resolved["_warnings"] = resolved.get("_warnings", []) + [
                f"LLM が返した比較({_llm_cmp or '(空)'})と依頼文の機械抽出({_mech_cmp})が"
                f"食い違うため機械抽出({_mech_cmp})を採用しました"]
            _cmp = _mech_cmp
        else:
            _cmp = _llm_cmp
        if _cmp not in _EXTRACT_CMPS:
            return False, resolved, inferred, (
                f"比較『{resolved.get('cmp')}』は {'/'.join(_EXTRACT_CMPS)} のどれでもありません")
        resolved["cmp"] = _cmp
        # ★ 閾値は**依頼文の数字**から機械が取る（LLM に確定させない）。
        _nums = _re_threshold_num.findall((task or "").translate(_ZENKAKU_DIGITS))
        if _cmp == "contains":
            _thr = resolved.get("cond_value")
            if _thr in (None, ""):
                return False, resolved, inferred, "条件の値が依頼文から読み取れません"
            resolved["cond_value"] = str(_thr)
        else:
            if len(set(_nums)) != 1:
                return False, resolved, inferred, (
                    "条件の数値が依頼文から一意に読み取れません"
                    f"（見つかった数: {'、'.join(sorted(set(_nums))) or 'なし'}）── "
                    "「原価が500以上の行の…」のように 1 つだけ書いてください")
            resolved["cond_value"] = float(_nums[0])
        # ★ 書き込む値は引用符から（SET_COLUMN_VALUE と同じ関所 ── LLM に作らせない）
        _q = extract_quoted_literal(task)
        if _q is None:
            return False, resolved, inferred, (
                "書き込む値が依頼文から一意に読み取れません。値を「」または『』で囲んで"
                "書いてください（例:「原価が500以上の行のチェック列に『◎』を付けて」）")
        resolved["value"] = _q
        resolved["_sources"] = {**resolved.get("_sources", {}), "value": f"依頼文: 「{_q}」"}
        resolved["_headers"] = [str(h) for h in (headers.get(first_sheet) or [])]
        resolved["_header_row"] = int((book_meta.get("header_rows") or {}).get(first_sheet, 1) or 1)
        _lab = _EXTRACT_CMP_LABELS.get(_cmp, _cmp)
        # ★ 500.0 でなく 500 と出す（人が依頼文に書いた形に近づける・整数なら整数で）
        _cv = resolved["cond_value"]
        _shown = (_cv if _cmp == "contains"
                   else (str(int(_cv)) if float(_cv).is_integer() else _fmt_cell_value(_cv)))
        resolved["_cond_label"] = f"『{resolved['cond_col']}』が {_shown} {_lab}"
        # ★ 当てはまる行を**先に数えて画面に出す**（0 行なら、走らせる前に断る ──
        #   「何も起きなかった」を後から × で知らせるのは、正しくても不親切）。
        # ★ 合計行は**データ行ではない**ので対象から外す。外したことは必ず画面に出す。
        resolved["_skip_rows"] = total_rows_in(book_meta, first_sheet, resolved["_header_row"])
        if resolved["_skip_rows"]:
            resolved["_skip_label"] = ("合計行 " + "、".join(
                f"{r}行目" for r in resolved["_skip_rows"]) + "（データ行でないため）")
        _hits = _rows_matching(book_meta, first_sheet, resolved["cond_col"],
                                _cmp, resolved["cond_value"], resolved["_header_row"])
        if _hits is not None:
            if not _hits:
                return False, resolved, inferred, (
                    f"{resolved['_cond_label']} に当てはまる行がありません"
                    "（条件か値を確かめてください・ファイルには何も書いていません）")
            resolved["_match_label"] = f"{len(_hits)} 行（{'、'.join(str(r) for r in _hits[:5])}行目"
            resolved["_match_label"] += "…）" if len(_hits) > 5 else "）"
            resolved["_match_rows"] = _hits

    # --- ★ 2026-08-27: 列の追加（位置は機械が見出しから決める）--------------------
    elif op == "ADD_COLUMN":
        _sheet_c = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr_c = int((book_meta.get("header_rows") or {}).get(_sheet_c, 1) or 1)
        _headers_c = [str(h) for h in ((book_meta.get("headers") or {}).get(_sheet_c) or [])]
        _name_c = str(resolved.get("name") or "").strip()
        # ★★ 2026-08-27（実測）: 名前を言っていない依頼に対し、LLM が「新しい列」という
        #   **依頼文に無い名前を作って**返す回があった（3 回中 1 回）。A' 原則の違反 ──
        #   値は LLM に確定させない。**依頼文に現れない名前は採らない**（空欄に倒す）。
        #   ★ 空欄は誤った名前より安い: 見出しが空なら △ になり、人が気づける。
        #     もっともらしい名前が付くと、人は「自分がそう言った」と思ってしまう。
        if _name_c and _name_c not in (task or ""):
            resolved["_name_dropped"] = _name_c
            _name_c = ""
        # ★ 同名の列が既に在るなら断る（黙って 2 本目を作らない ── 後で列名の解決が
        #   「2 つあります」で詰まる形を、作る側で防ぐ）。
        if _name_c and _name_c in _headers_c:
            return False, resolved, inferred, (
                f"列『{_name_c}』は既にあります（{_headers_c.index(_name_c) + 1}列目）")
        _at_c, _note_c = resolve_col_anchor(task, _headers_c)
        if _at_c is None and _note_c:
            return False, resolved, inferred, _note_c
        if _at_c is None:
            # 位置の言い回しが無い＝末尾。**黙って決めない**ので根拠を必ず出す。
            _at_c = len(_headers_c) + 1
            _note_c = f"末尾＝{_at_c}列目（依頼文に位置の指定が無いため）"
        # ★★ 2026-08-27（Namakoo が GUI で実測）: 見出しも値も無い列を**末尾**に足すと、
        #   セルは 1 つも増えないので機械には**何も変わって見えない**（物理の使用範囲は
        #   値のあるセルで測るため）。事後条件は正しく「列数が合わない」で × を出すが、
        #   利用者には「動かなかった」としか見えない ── **やる前に断って理由を言う**。
        #   ★ 途中に挿す場合は右の列がずれるので見える（そちらは通す）。
        if not _name_c and int(_at_c) > len(_headers_c):
            return False, resolved, inferred, (
                "見出しも値も無い列を末尾に足しても、ファイルの中身は何も変わりません"
                "（空の列はセルを持たないので機械にも見えません）── "
                "見出しの名前を言ってください（例: 「原価の右にチェックという列を追加して」）")
        resolved["name"] = _name_c
        resolved["_at_col"] = int(_at_c)
        resolved["_at_basis"] = _note_c
        resolved["_headers"] = _headers_c
        resolved["_header_row"] = _hr_c
        # ★ 名前が無いなら「空のまま」と画面に書く（黙って空欄を作らない）。
        # ★ 見出しが空の列を作ると、**その右にある列も走査できなくなる**
        #   （走査は見出し行の最初の空で止まる）── 作る前に言う。判定は △ に落ちる。
        _dropped = resolved.get("_name_dropped")
        resolved["_name_label"] = _name_c or (
            ("（依頼文に無い名前『%s』は採りませんでした・" % _dropped if _dropped
              else "（名前なし・")
            + "見出しは空のまま ── 右にある列も走査できなくなり、判定は △ になります）")

    # --- ★ 2026-08-27: 入れ替え（行か列かは**機械が実表を見て**決める）------------
    elif op == "SWAP":
        # ★★ 2026-08-31: セルの入れ替えは **a/b を要求する前**に見る。
        #   実測: 一段目が OUT_OF_VOCAB を返す言い方（「みどり建設の単価と丸和物流の
        #   単価を入れ替えて」）では a/b が空で、ここで先に落ちていた。
        #   ★ 座標は依頼文と実表だけで解ける ── LLM の返事に依存させない。
        _sheet_s0 = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr_s0 = int((book_meta.get("header_rows") or {}).get(_sheet_s0, 1) or 1)
        _cells0 = swap_targets_are_cells(task, book_meta, _sheet_s0, _hr_s0)
        _a = str(resolved.get("a", "")).strip()
        _b = str(resolved.get("b", "")).strip()
        if not _cells0 and (not _a or not _b):
            return False, resolved, inferred, (
                "入れ替える 2 つを取り出せませんでした"
                "（『みかんとぶどうを入れ替えて』のように 2 つの名前を書いてください）")
        if not _cells0 and _a == _b:
            return False, resolved, inferred, (
                f"『{_a}』と『{_b}』が同じものです（入れ替えになりません）")
        _sheet_s = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr_s = int((book_meta.get("header_rows") or {}).get(_sheet_s, 1) or 1)
        _headers_s = [str(h) for h in ((book_meta.get("headers") or {}).get(_sheet_s) or [])]
        resolved["_headers"] = _headers_s
        resolved["_header_row"] = _hr_s
        resolved["a"], resolved["b"] = _a, _b
        # ★★ 2026-08-31: 行/列を決める**前に**、セルの入れ替えでないかを見る。
        #   実測: 「丸和物流の単価とみどり建設の単価を入れ替えて」で行を丸ごと
        #   入れ替えて ✓ を出していた（頼んだのは 2 セル）。
        if (_cells := _cells0):
            # ★ 中身は**実表から読む**（LLM に値を作らせない・A' 原則）。
            _p_s = book_meta.get("path")
            try:
                with BookView(Path(_p_s)) as _bv_s:
                    _ws_s = _bv_s.sheet(_sheet_s)
                    _vals = [_ws_s.cell(row=r, column=c).value for r, c in _cells]
            except Exception as _e:
                return False, resolved, inferred, f"表を読めませんでした（{type(_e).__name__}）"
            if _vals[0] == _vals[1]:
                return False, resolved, inferred, (
                    "入れ替える 2 つのセルの中身が同じです（入れ替えになりません）")
            resolved["_axis"] = "cell"
            # ★ 2026-08-31: a/b が空の経路（一段目が OUT_OF_VOCAB だった回）だと
            #   解釈行に「入れ替える一方: もう一方:」と**空欄**が出ていた。
            #   嘘の空欄を見せない ── その行を人が呼ぶ名前（1 列目）で埋める。
            if not _a or not _b:
                _n0 = _cell_row_name_for(book_meta, _sheet_s, _cells[0][0], _hr_s)
                _n1 = _cell_row_name_for(book_meta, _sheet_s, _cells[1][0], _hr_s)
                resolved["a"] = _a or (str(_n0) if _n0 else f"{_cells[0][0]}行目")
                resolved["b"] = _b or (str(_n1) if _n1 else f"{_cells[1][0]}行目")
            resolved["_cells"] = [list(c) for c in _cells]
            resolved["_cell_values"] = list(_vals)
            resolved["_a_pos"], resolved["_b_pos"] = _cells[0][0], _cells[1][0]
            resolved["_axis_label"] = (
                f"セル（{_headers_s[_cells[0][1] - 1]} の {_cells[0][0]}行目 と "
                f"{_headers_s[_cells[1][1] - 1]} の {_cells[1][0]}行目）")
            return True, resolved, inferred, None
        as_col = _a in _headers_s and _b in _headers_s
        _ra, _note_a = _resolve_named_row(book_meta, _sheet_s, _a)
        _rb, _note_b = _resolve_named_row(book_meta, _sheet_s, _b)
        as_row = _ra is not None and _rb is not None
        hint = _swap_axis_hint(task)
        # ★ 三項（依頼・宣言・実体）: 依頼文の「行/列」という語と、LLM が挙げた 2 つの名前と、
        #   実際の表。どれか 2 つだけで決めると、欠けた項を代用して恒真になる。
        if as_col and as_row:
            if hint is None:
                return False, resolved, inferred, (
                    f"『{_a}』『{_b}』は列の見出しにも、行の中身にも両方あります ── "
                    "どちらを入れ替えるのか決められません"
                    "（『〜の列を入れ替えて』『〜の行を入れ替えて』と書いてください）")
            as_col, as_row = (hint == "column"), (hint == "row")
        if hint == "row" and not as_row:
            return False, resolved, inferred, f"行として決められません（{_note_a}／{_note_b}）"
        if hint == "column" and not as_col:
            return False, resolved, inferred, (
                f"列として決められません（ある列: {"、".join(_headers_s)}）")
        if not as_col and not as_row:
            return False, resolved, inferred, (
                f"入れ替える対象を決められません（{_note_a}／{_note_b}／"
                f"ある列: {"、".join(_headers_s)}）")
        resolved["_headers"] = _headers_s
        resolved["_header_row"] = _hr_s
        resolved["a"], resolved["b"] = _a, _b
        if as_col:
            resolved["_axis"] = "column"
            resolved["_axis_label"] = "列（見出しで一致）"
            resolved["_a_pos"] = _headers_s.index(_a) + 1
            resolved["_b_pos"] = _headers_s.index(_b) + 1
        else:
            resolved["_axis"] = "row"
            resolved["_axis_label"] = f"行（{_note_a}／{_note_b}）"
            resolved["_a_pos"] = _ra
            resolved["_b_pos"] = _rb
            if min(_ra, _rb) <= _hr_s:
                return False, resolved, inferred, (
                    f"見出し行（{_hr_s}行目）を巻き込む入れ替えは受け付けません")
        # ★★ 2026-08-29: 入れ替えは「表に写像 π を掛ける」ことで、式もその対象。
        #   LibreOffice の自動付け替えに任せず、**π を通した式を自分で書き戻す**。
        _sh = (cellmap.swap_cols(resolved["_a_pos"], resolved["_b_pos"]) if as_col
                else cellmap.swap_rows(resolved["_a_pos"], resolved["_b_pos"]))
        _rw, _rw_why = formula_rewrites_for_shift(
            book_meta, resolved.get("_target_sheet") or first_sheet, _sh)
        if _rw_why and "実行しません" in _rw_why:
            return False, resolved, inferred, _rw_why
        if _rw:
            resolved["_formula_rewrites"] = sorted(
                (r, c, f) for (r, c), f in _rw.items())
            resolved["_formula_rewrites_label"] = (
                f"{len(_rw)} 個の式を、入れ替え後の位置に合わせて書き直します"
                "（操作前と同じ計算結果に戻ることを、適用後に読み戻して確かめます）")
        elif _rw_why:
            resolved["_warnings"] = resolved.get("_warnings", []) + [_rw_why]
        # ★ 入れ替えでも「指す先の中身が変わる式」を名指しする（並べ替えと同じ目）。
        #   ★ 軸で区画が変わるだけ ── 行なら 2 行、列なら 2 列（行と列を同じ形で書く）。
        _sw_sheet = resolved.get("_target_sheet") or first_sheet
        _lo, _hi = min(resolved["_a_pos"], resolved["_b_pos"]), max(resolved["_a_pos"],
                                                                     resolved["_b_pos"])
        _kw = ({"col_lo": _lo, "col_hi": _hi} if as_col
                else {"row_lo": _lo, "row_hi": _hi})
        # ★★ 2026-08-31（Namakoo「この指す中身が変わるとはどういうこと？」→ 実測）:
        #   「金額と単価を入れ替えて」で ⚠ が出るが、**中身が 3 つとも事実に反していた**:
        #     ・「指す先の中身が変わる」→ 変わらない（税込金額は金額を指し続けた）
        #     ・「**行**が入れ替わる」   → 入れ替えたのは列
        #     ・「直していません」       → **直している**（=E2*1.1 → =D2*1.1）
        #   そして嘘の ⚠ のせいで、正しく動いた操作の ✓ が △ に落ちていた。
        #   ★ 片配線の**逆**: 警告は並べ替え（式を直さない）用に作って入れ替えにも配線し、
        #     そのあと入れ替えだけ式を直すようになったのに、警告は昔の前提のまま残った。
        #   ★ 書き直す式は名指しから外す（別シートから指す式は書き直さないので残す）。
        if (_dw := reference_drift_warning(book_meta, _sw_sheet,
                                            rewritten=set(_rw),
                                            unit=("列" if as_col else "行"), **_kw)):
            resolved["_warnings"] = resolved.get("_warnings", []) + [_dw]

    # --- ★ W9: 検証済みヘルパ4種の語彙昇格 -----------------------------------
    elif op == "INSERT_ROWS":
        # ★ 2026-08-27（実測）:「みかんの下に空行を入れて」で LLM が 3 行目と言った
        #   （みかんが 3 行目なので、下は 4 行目）。**位置は op に関係なく位置** ──
        #   同じ機械の解決を通す（片配線を作らない）。
        _sheet_i = resolved.get("_target_sheet") or (book_meta.get("sheets") or [None])[0]
        _hr_i = int((book_meta.get("header_rows") or {}).get(_sheet_i, 1) or 1)
        _at_i, _note_i = resolve_row_anchor(task, book_meta, _sheet_i, header_row=_hr_i)
        # ★ 2026-08-27（実測）: 見つからなかった時に**黙って LLM の行番号へ落ちて**いた。
        #   位置を名指しした依頼で場所が特定できないなら、推測で挿さずに断る
        #   （静かに別の場所へ入るのが一番こわい ── ADD_ROW と同じ線に揃える）。
        if _note_i and _at_i is None:
            return False, resolved, inferred, _note_i
        if _at_i is not None:
            resolved["at"] = _at_i
            resolved["_at_basis"] = _note_i
        at_raw = str(resolved.get("at", "")).strip()
        if not (at_raw.isdigit() and int(at_raw) >= 1):
            return False, resolved, inferred, f"行番号『{resolved.get('at')}』が不正です（1以上の整数）"
        resolved["at"] = int(at_raw)
        count_raw = resolved.get("count")
        if count_raw in (None, ""):
            resolved["count"] = 1
            inferred.add("count")
        else:
            count_str = str(count_raw).strip()
            if not (count_str.isdigit() and int(count_str) >= 1):
                return False, resolved, inferred, f"挿入行数『{count_raw}』が不正です（1以上の整数）"
            resolved["count"] = int(count_str)

    elif op == "DRAW_BORDERS":
        pass   # 引数無し・表全体が対象（検証することが無い）

    elif op == "AUTOFIT":
        pass   # 引数無し・全列が対象（検証することが無い）

    elif op == "PIVOT":
        # ★ AGGREGATE と同じ2 slot（group_col/value_col）。列名の実在確認だけ共有する。
        if (err := resolve_in("group_col", first_sheet)):
            return False, resolved, inferred, err
        if (err := resolve_in("value_col", first_sheet)):
            return False, resolved, inferred, err

    # ★ 致命3(W10e): 一括定数書き換え。値は LLM に確定させない（A' 原則）。
    elif op == "SET_COLUMN_VALUE":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        llm_value_raw = resolved.pop("value", None)
        quoted = extract_quoted_literal(task)
        if quoted is None:
            return False, resolved, inferred, (
                "書き込む値が依頼文から一意に読み取れません。値を「」または『』で囲んで"
                "書いてください（例:「備考列を全部『確認済み』にして」）"
            )
        resolved["value"] = quoted
        resolved["_sources"] = {**resolved.get("_sources", {}), "value": f"依頼文: 「{quoted}」"}
        if llm_value_raw not in (None, "") and str(llm_value_raw) != quoted:
            resolved["_warnings"] = resolved.get("_warnings", []) + [
                f"LLM が返した値('{llm_value_raw}')と依頼文の引用('{quoted}')が食い違うため"
                f"依頼文側('{quoted}')を採用しました"
            ]
        # ★ operator10 ④: 書き込む型を列の実体から機械決定する（A' 原則: LLM に決めさせない）。
        #   book_meta にファイルパスがある（実行時）場合だけ実体を読める ── 手組みの
        #   book_meta（単体テスト等）では従来どおり判定しない（_write_numeric を付けない）。
        book_path = book_meta.get("path")
        if book_path is not None:
            numeric_value = value_parses_as_number(quoted)
            write_numeric = False
            if numeric_value is not None:
                header_row_here = book_meta.get("header_rows", {}).get(first_sheet, 1)
                col_idx1 = headers[first_sheet].index(resolved["col"]) + 1
                try:
                    write_numeric = column_is_all_numeric(book_path, first_sheet, col_idx1,
                                                           header_row_here)
                except Exception:
                    write_numeric = False
            resolved["_write_numeric"] = write_numeric
            if write_numeric:
                resolved["_write_numeric_value"] = numeric_value

    # ★ EXTRACT: 単一条件（col × cmp × value）に一致する行を新シートへ抜き出す
    #   （コミット 2edcb08「EXTRACT op」参照）。col は実在検証、cmp は語彙の6値に限定、value は
    #   gte/lte/gt/lt なら数値必須・eq は数値化できればそのまま数値・できなければ文字列・
    #   contains は常に文字列（A' 原則: 数値化は機械が行う。LLM の言い分をそのまま信じない）。
    elif op == "EXTRACT":
        # ★★ 2026-08-31（Namakoo が実測）:「**5行目の**ヤマノ食品を抜き出して」で
        #   ヤマノ食品が **2 行とも**抜き出されていた（5行目という限定が無視された）。
        #   ★ 三項の番人は鳴っていた（「依頼文が指しているのは: 5行目」）が、
        #     **⚠ を出して進んで**いた ── 何が無視されたかを言っておらず、
        #     利用者からは「効かなかった」に見える。
        #   ★★ そして、そもそも**行番号での抜き出しは語彙に無い**（EXTRACT は
        #     列×比較×値しか持たない）。できないことは ⚠ でなく**断る**。
        #   ★ この分岐には**早い出口が 2 つ**ある（依頼文が実在の値を名指しした道と、
        #     LLM の値をそのまま使う道）。最初は片方にだけ足して**素通り**した
        #     ── 判定は必ず**両方の手前**に置く（今日 3 度目の片配線）。
        if (_x_n := task_names_a_row_number(task)):
            return False, resolved, inferred, (
                f"行番号での抜き出し（{_x_n}行目）には対応していません ── "
                "抜き出しは「どの列がどうなっている行か」で指してください"
                "（例:「取引先がヤマノ食品の行を抜き出して」）")
        # ★★ 2026-08-31（Namakoo の実測から辿って出た別件）:
        #   「金額が60000以上の行を抜き出して」で **合計行（356400）まで抜き出されていた**。
        #   条件には合っているが、**合計はデータ行ではない**。
        #   ★ 並べ替え・条件つき書換では既に外していたのに、**抽出だけ外していなかった**
        #     ── また片配線。判定は凍結済みの規則（total_rows_in）を借りる。
        #   ★ ここも**早い出口より前**に置く（後ろに置いて 1 度素通りさせた）。
        _x_hr = int((book_meta.get("header_rows") or {}).get(first_sheet, 1) or 1)
        if (_x_tot := total_rows_in(book_meta, first_sheet, _x_hr)):
            resolved["_skip_rows"] = list(_x_tot)
            resolved["_skip_label"] = ("合計行 " + "、".join(
                f"{r}行目" for r in _x_tot) + "（データ行でないため抜き出しません）")
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        # ★★ 2026-08-27（Namakoo「みかんの行とりんごの行だけを抽出して」）:
        #   実測で一段目は `contains "リンゴ"`（片仮名の幻覚）や `eq "みかんとりんご"`
        #   （連結）を返していた。どちらも 0 行に当たり、**空の抽出結果が ✓ で出る**。
        #   ★ 比較語（以上/以下/…）が依頼文に**無い**なら、それは条件でなく**名指し**。
        #     表に実在する値のうち依頼文に現れるものを機械が拾い、「どれか」で抽出する。
        #   ★ 比較語が在る時は触らない ── 「原価が500以上」の 500 を名前と読まない。
        _hr_x = int((book_meta.get("header_rows") or {}).get(first_sheet, 1) or 1)
        if extract_cmp_from_task(task) is None:
            _named_vals = task_names_real_values(task, book_meta, first_sheet,
                                                  resolved["col"], _hr_x)
            if _named_vals:
                if str(resolved.get("value") or "") not in _named_vals:
                    resolved["_warnings"] = resolved.get("_warnings", []) + [
                        f"LLM が返した値『{resolved.get('value')}』は列『{resolved['col']}』に"
                        f"無いため、依頼文が名指しする実在の値"
                        f"（{'、'.join(_named_vals)}）を採用しました"]
                # ★★ 2026-09-02（実測で捕まえた・検体が警告していた事故の再演）:
                #   ここは「依頼文が実在の値を名指ししたら機械が勝つ」という正しい規則
                #   だが、**否定を知らなかった**。そのため「味噌汁以外を抜き出して」で
                #   読み直しが cmp=nin を立てた直後にここが `eq` へ上書きし、
                #   **味噌汁だけを抜き出して △ を出していた** ── 逆のことをして合格。
                #   ★ 片配線そのもの: 読み直しに足して、決定の場所に足し忘れた。
                _neg = task_says_except(task)
                if _neg:
                    resolved["cmp"] = "nin"
                    resolved["value"] = list(_named_vals)
                else:
                    resolved["cmp"] = "in" if len(_named_vals) > 1 else "eq"
                    resolved["value"] = (_named_vals if len(_named_vals) > 1
                                          else _named_vals[0])
                resolved["_source_headers"] = tuple(headers.get(first_sheet, []))
                resolved["_new_sheet"] = _extract_output_sheet_name(
                    resolved["col"], resolved["cmp"], resolved["value"])
                return True, resolved, inferred, None
        llm_cmp_raw = str(resolved.get("cmp", "")).strip().lower()
        # ★ operator9 ①: cmp も A' 原則の中へ。依頼文からの機械抽出が非 None かつ LLM の cmp と
        #   食い違えば機械が勝つ（factor/value と同じ作法）。一致 or 機械 None なら現状どおり。
        mechanical_cmp = extract_cmp_from_task(task)
        if mechanical_cmp is not None and mechanical_cmp != llm_cmp_raw:
            cmp = mechanical_cmp
            resolved["_warnings"] = resolved.get("_warnings", []) + [
                f"LLM が返した比較({llm_cmp_raw or '(空)'})と依頼文の機械抽出({mechanical_cmp})が"
                f"食い違うため機械抽出({mechanical_cmp})を採用しました"
            ]
        else:
            cmp = llm_cmp_raw
        if cmp not in _EXTRACT_CMPS:
            return False, resolved, inferred, (
                f"比較『{resolved.get('cmp')}』は {'/'.join(_EXTRACT_CMPS)} のどれでもありません"
            )
        resolved["cmp"] = cmp
        raw_value = resolved.get("value")
        if raw_value in (None, ""):
            return False, resolved, inferred, "抽出する値(value)が依頼文から読み取れません"
        if cmp == "contains":
            resolved["value"] = str(raw_value)
        else:
            try:
                resolved["value"] = float(raw_value)
            except (TypeError, ValueError):
                # ★ 日付の範囲比較（台帳 DATE_RANGE_AGG 2件の正体・2026-08-24）。
                #   「2026/3/26 以降」は数値ではないが、対象列が日付列なら
                #   シリアル値に直せば既存の数値比較でそのまま通る（Basic 側は無改造）。
                #   ★ resolved["value"] は**元の文字列のまま**残す ── 表示（解釈行・出力
                #   シート名）が 46107 になるのを防ぐ。codegen だけが _value_serial を見る。
                parsed_date = parse_date_literal(raw_value)
                # ★ 2026-08-24（盲検の使い勝手レビュー）: 依頼文が「3月26日から4月25日まで」と
                #   年を言っていないのに、LLM が **2023 年**を入れてきた（データは全部 2026 年）。
                #   A' 原則 ── 依頼文に無い年は機械が受け取らない。年は人が決めることで、
                #   LLM が埋めてよい空白ではない。
                if parsed_date is not None and str(parsed_date.year) not in (task or ""):
                    return False, resolved, inferred, (
                        f"依頼文に年が書かれていないため、『{raw_value}』の"
                        f"{parsed_date.year} 年は使えません"
                        f"（何年かを書いて、もう一度お試しください）")
                col_kind, col_has_time = _extract_column_date_kind(
                    book_meta, first_sheet, resolved["col"])
                if parsed_date is not None and col_kind == "date":
                    resolved["value"] = str(raw_value)
                    resolved["_value_serial"] = date_compare.threshold_for(cmp, parsed_date)
                    resolved["_date_compare"] = True
                    if cmp == "eq" and col_has_time:
                        resolved["_warnings"] = resolved.get("_warnings", []) + [
                            f"『{raw_value}』と一致で抽出しますが、列『{resolved['col']}』には"
                            f"時刻が入っています（0時ちょうどの行しか当たりません）"
                        ]
                elif parsed_date is not None and col_kind == "text_date":
                    # ★ 辞書順で黙って比べない ── "2026/3/26" > "2026/12/1" になる。
                    return False, resolved, inferred, (
                        f"列『{resolved['col']}』は日付が**文字列**で入っているため、"
                        f"『{raw_value}』との日付比較ができません"
                        f"（日付の書式に直してから、もう一度お試しください）"
                    )
                elif cmp != "eq":
                    return False, resolved, inferred, (
                        f"比較『{cmp}』には数値の値が必要ですが『{raw_value}』は数値に変換できません"
                    )
                else:
                    resolved["value"] = str(raw_value)
        # ★ 単位H: 出力シートの見出し署名(= 元シートの見出し行そのもの)を _own_output_headers
        #   が組めるよう、決めた材料をここで resolved に積む（他 op の _target_sheet と同じ作法）。
        resolved["_source_headers"] = tuple(headers.get(first_sheet, []))
        resolved["_new_sheet"] = _extract_output_sheet_name(resolved["col"], cmp, resolved["value"])

    # ★ DEDUP（EXTRACT の兄弟）: 判定キー列（1つ以上）の値の組が同じ行のうち最初の1行だけを
    #   新シートへ残す。★ keys が無い/空なら CLARIFY ── 全列一致を黙って既定にしない
    #   （「取引先が同じなら重複」は人の意図であって機械が推測してよい既定ではない）。
    elif op == "SPLIT_CELL":
        if (err := resolve_in("col", first_sheet)):
            return False, resolved, inferred, err
        sep = split_cell.normalize_separator(resolved.get("sep"))
        if not sep:
            return False, resolved, inferred, (
                "区切り(sep)が読み取れません（改行/カンマ/、/スペース などで指定してください）"
            )
        resolved["sep"] = sep
        # ★ 何列必要かは**実データ**が決める（LLM に数えさせない）。
        values = _column_values(book_meta, first_sheet, resolved["col"])
        parts = split_cell.max_parts(values, sep)
        if parts < 2:
            return False, resolved, inferred, (
                f"列『{resolved['col']}』に区切り『{split_cell.describe_separator(sep)}』が"
                f"見つからないため、分けられません"
            )
        resolved["_parts"] = parts
        resolved["_new_cols"] = [f"{resolved['col']}_{k}" for k in range(1, parts + 1)]

    elif op == "DEDUP":
        raw_keys = resolved.get("keys")
        if not isinstance(raw_keys, list) or not raw_keys:
            return False, resolved, inferred, (
                "重複を判定する列が依頼文から読み取れません。どの列が同じなら重複とみなすか、"
                "依頼文に列名を書いてください（例:「取引先が同じ行を重複として除いて」）"
            )
        resolved_keys = []
        for raw_key in raw_keys:
            v, was_inferred, err = resolve_col_ref(raw_key, headers.get(first_sheet, []))
            if err:
                return False, resolved, inferred, err
            resolved_keys.append(v)
            if was_inferred:
                inferred.add("keys")
        resolved["keys"] = resolved_keys
        # ★ 単位H: EXTRACT と同じ作法（出力シートの見出し署名の材料を resolved に積む）。
        resolved["_source_headers"] = tuple(headers.get(first_sheet, []))
        resolved["_new_sheet"] = _dedup_output_sheet_name(resolved_keys)

    # ★ 帳票段: REPORT_PER_ROW（DESIGN-20260823-report-per-row.md）。表の1行を、人が作った
    #   雛形シートの1枚に転写してN枚出す。憲法の適用: 機械が触ってよいのは雛形の中の
    #   印({{列名}})が置かれたセルだけ ── ここで印を実在検証し、印以外は一切触らない前提を
    #   固める（型の出し分け・出力シート名・合計行の除外まで、すべて機械が決め切る）。
    elif op == "REPORT_PER_ROW":
        if (err := check_sheet("template_sheet")):
            return False, resolved, inferred, err
        template_sheet = resolved["template_sheet"]
        if template_sheet == first_sheet:
            return False, resolved, inferred, (
                f"雛形シートとデータシートが同じ『{template_sheet}』です。"
                "雛形は別のシートに用意してください")
        if (err := resolve_in("name_col", first_sheet)):
            return False, resolved, inferred, err

        book_path = book_meta.get("path")
        if book_path is None:
            return False, resolved, inferred, (
                "帳票段はファイルの実体が無いと検証できません（book_meta に path が無い）")
        data_headers = headers.get(first_sheet, [])
        header_row_here = book_meta.get("header_rows", {}).get(first_sheet, 1)

        try:
            wb_tpl = openpyxl.load_workbook(book_path)
        except Exception as e:
            return False, resolved, inferred, f"雛形の読み込みに失敗しました: {e}"
        try:
            tpl_ws = wb_tpl[template_sheet]
            placeholders = scan_placeholders(tpl_ws, tpl_ws.max_row or 1, tpl_ws.max_column or 1)
            # ★ 縦の結合セルは、明細行を増やすと崩れる（値は合うので事後条件は通ってしまう）。
            #   日本の請求書の雛形は結合だらけなので、起きる方に賭けるべき事象（設計査読）。
            tpl_vmerges = [(m.min_row, m.max_row, m.coord)
                            for m in tpl_ws.merged_cells.ranges if m.min_row != m.max_row]
        finally:
            wb_tpl.close()

        if not placeholders:
            return False, resolved, inferred, (
                f"雛形『{template_sheet}』に印（{{{{列名}}}}）が見つかりません。"
                "転記したいセルに {{列名}} の形で印を置いてください")

        # ★ 2026-08-24: 1 セルに印が 2 つ以上あるなら、埋めずに断る。埋めると 1 セルに
        #   2 回書くことになり後の値が前を消す ── 「それらしく埋まって片方が生で残る」
        #   （盲検の査定で名指しされた事故）より、雛形を直してくださいと言う方が正しい。
        if (dupes := cells_with_multiple_placeholders(placeholders)):
            cell, names = dupes[0]
            return False, resolved, inferred, (
                f"雛形『{template_sheet}』の {cell} に印が {len(names)} つあります"
                f"（{chr(12539).join(names)}）。1 つのセルに置ける印は 1 つまでです ── "
                f"別々のセルに分けてください")

        # ★★ 2026-08-28（Namakoo「同名の取引先から複数の発注があるケースでは
        #   請求書を一枚にまとめないといけない」）: 印を 3 種類に仕分ける。
        #   {{列名}} / {{明細:列名}} / {{合計:列名}}。**雛形が形を決める**ので、
        #   依頼文にも一段目の語彙（OPS_DOC）にも 1 文字も足さない。
        mark_layout, layout_err = report_group.classify_placeholders(placeholders)
        if layout_err:
            return False, resolved, inferred, f"雛形『{template_sheet}』: {layout_err}"
        if mark_layout.detail_row is not None:
            crossing = [c for lo, hi, c in tpl_vmerges if lo <= mark_layout.detail_row <= hi]
            if crossing:
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』の明細行（{mark_layout.detail_row}行目）を、"
                    f"縦に結合したセルが横切っています（{'・'.join(crossing[:3])}）。"
                    "明細行は件数ぶん増えるので、この結合は崩れます ── "
                    "結合を解くか、明細行の外へずらしてください")

        resolved_placeholders = []
        for ph in placeholders:
            ph_kind, ph_col = report_group.mark_kind(ph.column_name)
            if ph_col not in data_headers:
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』の印『{{{{{ph.column_name}}}}}』"
                    f"（{ph.cell}）が指す列『{ph_col}』は、データシート"
                    f"『{first_sheet}』に見つかりません。実在する列名を印にしてください"
                )
            col_idx = data_headers.index(ph_col) + 1
            if ph_kind == "total" and not ph.whole:
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』の合計の印『{{{{{ph.column_name}}}}}』"
                    f"（{ph.cell}）は、セル全体を印にしてください（合計は数値です）")
            if not ph.whole:
                # ★ 訂正3: 部分一致の印は原理的に文字列にしかなれない ── 数値列には使わせない
                #   （検体には無いが自分の検体で固定する境界。設計文書の指示どおり）。
                try:
                    is_numeric = column_is_all_numeric(book_path, first_sheet, col_idx,
                                                        header_row_here)
                except Exception:
                    is_numeric = False
                if is_numeric:
                    return False, resolved, inferred, (
                        f"雛形『{template_sheet}』の印『{{{{{ph.column_name}}}}}』"
                        f"（{ph.cell}）はセルの一部分（部分一致）ですが、列『{ph_col}』は"
                        "数値です。数値列には部分一致の印を使えません"
                        "（セル全体を印にしてください: 例 " + "{{" + ph.column_name + "}}）"
                    )
            resolved_placeholders.append({
                "cell": ph.cell, "row": ph.row, "col": ph.col,
                "column_name": ph_col, "kind": ph_kind, "mark": ph.column_name,
                "whole": ph.whole, "raw": ph.raw, "col_idx": col_idx,
            })
        resolved["_placeholders"] = resolved_placeholders

        try:
            wb_data = openpyxl.load_workbook(book_path, data_only=True)
        except Exception as e:
            return False, resolved, inferred, f"データシートの読み込みに失敗しました: {e}"
        try:
            src_ws = wb_data[first_sheet]
            last_row = _scan_last_row(src_ws, header_row=header_row_here)
            rows_in = []
            for r in range(header_row_here + 1, last_row + 1):
                label_val = src_ws.cell(row=r, column=1).value
                vals = {h: src_ws.cell(row=r, column=i + 1).value
                        for i, h in enumerate(data_headers)}
                rows_in.append((r, label_val, vals))
        finally:
            wb_data.close()
        verdict = total_row.split_total_rows_multi(rows_in) if rows_in else total_row.TotalRowVerdict(
            excluded=[], adopted_rows=[], mismatches=[])
        row_values = {r: v for r, _l, v in rows_in}

        # ★★ まとめるか、1 行 1 枚か ── **雛形と実表の両方**が決める（人に選ばせない）:
        #   ・雛形に明細/合計の印が在る → まとめる（1 件でも同じ道を通る）
        #   ・印は無いが同じ名前が 2 行以上ある → **断る**（2 枚に割れた紙は仕事にならない）
        name_col_here = resolved["name_col"]
        name_idx = data_headers.index(name_col_here) + 1
        groups = report_group.build_groups(
            [(r, [row_values[r].get(h) for h in data_headers]) for r in verdict.adopted_rows],
            name_idx)
        grouped = mark_layout.detail_row is not None or bool(mark_layout.total)
        # ★★ 2026-08-28（設計査読で名指しされた・自分で開けかけた穴）:
        #   ここで**断って**はいけない。同名が 2 行あっても正しい帳票がある ──
        #   領収書・納品書は取引ごとに 1 枚だし、締め日違いの月別請求も同じ形
        #   （OPS_DOC 自身が REPORT_PER_ROW の用途に領収書を挙げている）。
        #   既に在る処方は「断ること」ではなく「✓ を出さないこと」だった（2026-08-24）。
        #   ★ 反転させずに、△ の警告文へ**まとめ方への道**を足すだけにする。

        used = set(sheets) | {template_sheet}
        report_rows = []
        if grouped:
            for g in groups:
                sheet_name = unique_sheet_name(str(g.name), used)
                used.add(sheet_name)
                report_rows.append({"row": g.rows[0], "sheet": sheet_name,
                                     "name": g.name, "rows": list(g.rows)})
        else:
            for r in verdict.adopted_rows:
                raw_name = row_values[r].get(name_col_here)
                sheet_name = unique_sheet_name(str(raw_name), used)
                used.add(sheet_name)
                report_rows.append({"row": r, "sheet": sheet_name})
        if not report_rows:
            return False, resolved, inferred, (
                "帳票にするデータ行がありません（表が空か、全行が合計行と判定されました）"
            )
        inspection_sheet = unique_sheet_name(inspection.SHEET_NAME, used)
        used.add(inspection_sheet)

        if grouped:
            # ★ 1 枚に 1 つしか書けない欄が、グループの中で食い違っていないか。
            #   食い違ったら**埋めずに断る** ── 推測で選ぶと、別の担当者の名前が客に届く。
            by_name = {g.name: g for g in groups}
            for rr in report_rows:
                g = by_name[rr["name"]]
                for ph in resolved_placeholders:
                    if ph["kind"] == "value":
                        vals = report_group.value_conflicts(g, row_values, ph["column_name"])
                        if vals:
                            # ★ 「足すなら {{合計:担当}}」は、担当のような文字列の列では
                            #   意味を成さない ── 出せる道だけを名指しする。
                            _way = f"『{{{{明細:{ph['column_name']}}}}}』にしてください"
                            if all(report_group.is_numeric(v) for v in vals):
                                _way = (f"『{{{{明細:{ph['column_name']}}}}}』、"
                                         f"足すなら『{{{{合計:{ph['column_name']}}}}}』"
                                         "にしてください")
                            return False, resolved, inferred, (
                                f"『{g.name}』の {list(g.rows)}行目で"
                                f"『{ph['column_name']}』が食い違っています（{vals}）。"
                                f"1 枚の紙には 1 つしか書けません ── 明細に出すなら{_way}")
                    elif ph["kind"] == "total":
                        _s, serr = report_group.sum_for(g, row_values, ph["column_name"])
                        if serr:
                            return False, resolved, inferred, f"『{g.name}』: {serr}"
            resolved["_groups"] = [{"sheet": rr["sheet"], "name": rr["name"],
                                     "rows": rr["rows"]} for rr in report_rows]
            resolved["_detail_row"] = mark_layout.detail_row
        else:
            # ★ 2026-08-24: 重複を知った瞬間に言う（`_2` を付けたのがその瞬間）。
            #   実測: 3 社の売上表（4 行）で請求書 4 枚・同じ取引先が 2 枚に分かれて ✓ が出た。
            #   ★ 付きなので count_suspicious_advisories が拾い、決裁③で ✓→△ に降格する。
            if (dup := duplicate_name_warning(
                    name_col_here,
                    [row_values[r].get(name_col_here) for r in verdict.adopted_rows])):
                resolved["_warnings"] = resolved.get("_warnings", []) + [dup]
        resolved["_report_rows"] = report_rows
        resolved["_report_sheet_names"] = [rr["sheet"] for rr in report_rows]
        resolved["_inspection_sheet"] = inspection_sheet
        resolved["_source_headers"] = tuple(data_headers)

    # ★ 様式写像段: FORMAT_MAP（DESIGN-20260824-format-map.md）。REPORT_PER_ROW の兄弟
    #   （縦の展開）── 表の1行を、人が作った雛形の1行に転写して N行の新シートを1枚出す。
    #   憲法の適用は同じ: 機械が触ってよいのは雛形の中の印({{列名}})が置かれたセルだけ。
    elif op == "FORMAT_MAP":
        if (err := check_sheet("template_sheet")):
            return False, resolved, inferred, err
        template_sheet = resolved["template_sheet"]
        if template_sheet == first_sheet:
            return False, resolved, inferred, (
                f"雛形シートとデータシートが同じ『{template_sheet}』です。"
                "雛形は別のシートに用意してください")

        book_path = book_meta.get("path")
        if book_path is None:
            return False, resolved, inferred, (
                "様式写像段はファイルの実体が無いと検証できません（book_meta に path が無い）")
        data_headers = headers.get(first_sheet, [])
        header_row_here = book_meta.get("header_rows", {}).get(first_sheet, 1)

        try:
            wb_tpl = openpyxl.load_workbook(book_path)
        except Exception as e:
            return False, resolved, inferred, f"雛形の読み込みに失敗しました: {e}"
        try:
            tpl_ws = wb_tpl[template_sheet]
            placeholders = scan_placeholders(tpl_ws, tpl_ws.max_row or 1, tpl_ws.max_column or 1)
            if not placeholders:
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』に印（{{{{列名}}}}）が見つかりません。"
                    "出力したい列の直下のセルに {{列名}} の形で印を置いてください")

            # ★ 2026-08-24: 1 セルに印が 2 つ以上あるなら、埋めずに断る。埋めると 1 セルに
            #   2 回書くことになり後の値が前を消す ── 「それらしく埋まって片方が生で残る」
            #   （盲検の査定で名指しされた事故）より、雛形を直してくださいと言う方が正しい。
            if (dupes := cells_with_multiple_placeholders(placeholders)):
                cell, names = dupes[0]
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』の {cell} に印が {len(names)} つあります"
                    f"（{chr(12539).join(names)}）。1 つのセルに置ける印は 1 つまでです ── "
                    f"別々のセルに分けてください")

            # ★ 第一波: 印は全部1つの行に置かれている前提（設計文書「見出し行 + 直下1行に印」）。
            #   最初に見つかった印の行を「印行」、その直上を「見出し行」とみなす。
            ph_row = placeholders[0].row
            header_tpl_row = ph_row - 1
            if header_tpl_row < 1:
                return False, resolved, inferred, (
                    f"雛形『{template_sheet}』の印（{ph_row}行目）の上に見出し行がありません。"
                    "印の1つ上の行に出力したい列名を書いてください")

            row_placeholders = sorted(
                (ph for ph in placeholders if ph.row == ph_row), key=lambda p: p.col)
            resolved_placeholders = []
            header_texts = []
            for ph in row_placeholders:
                if ph.column_name not in data_headers:
                    return False, resolved, inferred, (
                        f"雛形『{template_sheet}』の印『{{{{{ph.column_name}}}}}』"
                        f"（{ph.cell}）が指す列『{ph.column_name}』は、データシート"
                        f"『{first_sheet}』に見つかりません。実在する列名を印にしてください"
                    )
                col_idx = data_headers.index(ph.column_name) + 1
                if not ph.whole:
                    # ★ REPORT_PER_ROW と同じ境界: 部分一致の印は原理的に文字列にしかなれない。
                    try:
                        is_numeric = column_is_all_numeric(book_path, first_sheet, col_idx,
                                                            header_row_here)
                    except Exception:
                        is_numeric = False
                    if is_numeric:
                        return False, resolved, inferred, (
                            f"雛形『{template_sheet}』の印『{{{{{ph.column_name}}}}}』"
                            f"（{ph.cell}）はセルの一部分（部分一致）ですが、列『{ph.column_name}』は"
                            "数値です。数値列には部分一致の印を使えません"
                            "（セル全体を印にしてください: 例 " + "{{" + ph.column_name + "}}）"
                        )
                header_texts.append(tpl_ws.cell(row=header_tpl_row, column=ph.col).value)
                resolved_placeholders.append({
                    "cell": ph.cell, "row": ph.row, "col": ph.col,
                    "column_name": ph.column_name, "whole": ph.whole, "raw": ph.raw,
                    "col_idx": col_idx, "out_col": len(resolved_placeholders) + 1,
                })
        finally:
            wb_tpl.close()
        resolved["_placeholders"] = resolved_placeholders
        resolved["_header_texts"] = header_texts
        resolved["_header_tpl_row"] = header_tpl_row
        resolved["_placeholder_tpl_row"] = ph_row

        try:
            wb_data = openpyxl.load_workbook(book_path, data_only=True)
        except Exception as e:
            return False, resolved, inferred, f"データシートの読み込みに失敗しました: {e}"
        try:
            src_ws = wb_data[first_sheet]
            last_row = _scan_last_row(src_ws, header_row=header_row_here)
            rows_in = []
            for r in range(header_row_here + 1, last_row + 1):
                label_val = src_ws.cell(row=r, column=1).value
                vals = {h: src_ws.cell(row=r, column=i + 1).value
                        for i, h in enumerate(data_headers)}
                rows_in.append((r, label_val, vals))
        finally:
            wb_data.close()
        verdict = total_row.split_total_rows_multi(rows_in) if rows_in else total_row.TotalRowVerdict(
            excluded=[], adopted_rows=[], mismatches=[])
        if not verdict.adopted_rows:
            return False, resolved, inferred, (
                "写す行がありません（表が空か、全行が合計行と判定されました）"
            )

        used = set(sheets) | {template_sheet}
        output_sheet = unique_sheet_name(str(template_sheet) + "_出力", used)
        used.add(output_sheet)
        inspection_sheet = unique_sheet_name(inspection.SHEET_NAME, used)
        used.add(inspection_sheet)

        resolved["_data_rows"] = list(verdict.adopted_rows)
        resolved["_output_sheet"] = output_sheet
        resolved["_inspection_sheet"] = inspection_sheet
        resolved["_source_headers"] = tuple(data_headers)

    else:
        return False, resolved, inferred, f"未対応の操作: {op}"

    # ★ 捏造列の関所（2026-08-24・盲検の査定者が値段を決めた一点）。
    #   ここは全経路（単発・複合計画・--dry プレビュー）が通る唯一の場所なので、
    #   ここに置けば片配線にならない。適用より前なので原本には触れない。
    if (fab := fabricated_subject_refusal(op, resolved, book_meta, task, first_sheet)):
        return False, resolved, inferred, fab

    # ★★ 2026-08-27（Namakoo「『〜の右側に』『〜と〜の間に』は頻出だから全ての操作で」）:
    #   位置の言い回しは op の性質ではなく**依頼文の性質**なので、op ごとに配線しない。
    #   ここ（全経路が通る唯一の場所）で 1 回だけ解いて、宣言（WRITE_NEW_COLUMN）を持つ
    #   op すべてに効かせる。codegen 側は wrap() が 1 箇所で MoveColumnTo を足す。
    if (place := resolve_new_column_placement(op, resolved, book_meta, task, first_sheet)):
        resolved.update(place)

    return True, resolved, inferred, None


# --- ③ 確認行（命令言語形式） -------------------------------------------------

_CONFIRM_FIELDS = {
    # ★★ 2026-08-29（Namakoo の通しで実測）: `_skip_label` は**セットされているだけで
    #   どこにも表示されていなかった** ── 「外したことは必ず画面に出す」と書いてある
    #   契約が守られていない（在っても鳴らない、の形）。解釈行に出す口を足す。
    "SORT": (("対象", "col", None), ("順", "order", lambda v: "降順" if v == "desc" else "昇順"),
              ("対象から外した行", "_skip_label", None)),
    # ★ 新しい列の名前は**必ず出す** ── 出ていなかったので、「税込み金額」が
    #   「税込金額」や「金額*1.1」に化けても気づけなかった（2026-08-30 実測）。
    "COMPUTE_COLUMN": (("演算対象", "operands", lambda v: " と ".join(v)), ("演算子", "operator", None),
                        ("対象列", "target", None), ("新しい列の名前", "_new_col_label", None)),
    "LOOKUP_FILL": (("対象シート", "target_sheet", None), ("対象列", "target_col", None),
                     ("参照シート", "source_sheet", None), ("キー列", "key_col", None)),
    "AGGREGATE": (("分類列", "group_col", None), ("集計列", "value_col", None)),
    "BOLD": (("対象", "target", None),),
    "FILL_COLOR": (("対象", "target", None), ("色", "color", None)),
    "NUMBER_FORMAT": (("対象列", "col", None), ("書式", "style", None)),
    "MERGE": (("範囲", "range", None),),
    "CHART": (("値列", "value_col", None),
               ("種類", "kind", lambda v: _CHART_KIND_LABELS.get(v, v)),
               ("横軸列", "category_col", None)),
    "CENTER_ALIGN": (("対象", "target", None),),
    # ★ W8a 項目5: 表示ラベルのみ「倍率」→「率」（税率・掛け率の文脈での事務向け言い換え）。
    #   内部キー("factor")・関数名・コメントは不変。
    "APPEND_TOTAL": (("対象列", "col", None), ("ラベル", "label", None), ("率", "factor", None)),
    # ★ W9: 検証済みヘルパ4種の語彙昇格。
    "INSERT_ROWS": (("挿入位置", "at", None), ("位置の根拠", "_at_basis", None),
                     ("行数", "count", None)),
    "ADD_ROW": (("挿入位置", "at", None), ("位置の根拠", "_at_basis", None),
                 ("入れる値", "_values_label", None),
                 # ★ 2026-09-02: 宣言していないセル（式の列）に書く以上、**先に言う**。
                 #   書いてから知らせるのでは、✓ の意味が広がったことが人に伝わらない。
                 ("式を引き継ぐ列", "_inherit_label", None)),
    "DELETE_ROWS": (("削除位置", "at", None), ("位置の根拠", "_at_basis", None),
                     ("行数", "count", None)),
    "DELETE_COLUMN": (("削除する列", "col", None),),
    # ★ a/b をそのまま出す ── `ailine ops` の「必要な情報」はこの登録簿からラベルを引く
    #   ので、ここに無い slot は生の英字（a・b）のまま人に見えてしまう。
    "SWAP": (("入れ替える一方", "a", None), ("もう一方", "b", None),
              ("何を入れ替えるか", "_axis_label", None)),
    "ADD_COLUMN": (("新しい列の名前", "_name_label", None), ("入れる位置", "_at_basis", None)),
    # ★ cond_col/cmp のラベルもここに要る ── `ailine ops` の「必要な情報」はこの登録簿から
    #   引くので、無いと英字（cond_col・cmp）のまま人に見える（SWAP で 1 度踏んだ）。
    # ★ cols のラベルも要る（`ailine ops` の「必要な情報」はこの登録簿から引く）。
    "EXTRACT_COLUMNS": (("残す列", "cols", lambda v: "・".join(map(str, v or []))),
                         ("出力シート", "_new_sheet", None)),
    "SET_WHERE": (("書き込む列", "col", None), ("書き込む値", "value", None),
                   ("条件を見る列", "cond_col", None),
                   ("比べ方", "cmp", lambda v: _EXTRACT_CMP_LABELS.get(v, v)),
                   ("条件", "_cond_label", None), ("当てはまる行", "_match_label", None),
                   ("対象から外した行", "_skip_label", None)),
    "SET_CELL_VALUE": (("対象の行", "row", None), ("対象列", "col", None),
                        ("書き込む値", "value", None)),
    "DRAW_BORDERS": (),
    "AUTOFIT": (),
    "PIVOT": (("分類列", "group_col", None), ("集計列", "value_col", None)),
    # ★ operator10 ④: 「型」は resolved_args に "_write_numeric" キーがある時だけ表示される
    #   （M2c のフィールド省略・手組みの resolved_args ではキー自体が無いので出ない）。
    "SET_COLUMN_VALUE": (("対象列", "col", None), ("値", "value", None),
                          ("型", "_write_numeric", lambda v: "数値" if v else "文字列"),
                          ("対象から外した行", "_skip_label", None)),
    # ★ 2026-08-31: 合計行を外すようになったので、**外したことを画面に出す**
    #   （08-29 に SORT で同じ形を踏んだ ── 作る側と見せる側の数を検体が縛っている）。
    "EXTRACT": (("対象列", "col", None), ("条件", "cmp", lambda v: _EXTRACT_CMP_LABELS.get(v, v)),
                 ("値", "value", lambda v: _format_extract_value(v)),
                 ("対象から外した行", "_skip_label", None)),
    "SPLIT_CELL": (("対象列", "col", None),
                    ("区切り", "sep", lambda v: split_cell.describe_separator(v))),
    "DEDUP": (("判定キー", "keys", lambda v: "・".join(v)),),
    "REPORT_PER_ROW": (("雛形", "template_sheet", None), ("シート名の元列", "name_col", None)),
    "FORMAT_MAP": (("雛形", "template_sheet", None),),
}

# ★ 複合計画の連鎖の番人（2026-08-24・実測した致命）。
#   「売上が60以上の行だけ現場ごとに集計して」で 1段目 EXTRACT が『売上60以上』を作り、
#   2段目 AGGREGATE が**元の『売上』**を集計して、それでも ✓ が出ていた。
#   各段の事後条件はどちらも真 ── 嘘は段の中でなく**段と段の間**にあった。
#   ★ 自動で連鎖させない理由: 「抽出して、元表に合計を追加して」も正当な計画で、
#   どちらの意図かは機械に決まらない。決まらないものを黙って決めたのが事故そのもの。
#   だから名指しして人に返す（★ 付き助言 → 決裁③ で ✓→△ に降格）。
#   派生シートを作る op（＝出力が「絞り込んだ同じ表」であるもの）だけを対象にする。
PLAN_CHAIN_WARNING_OPS = ("EXTRACT", "DEDUP")


# ★ EXTRACT: 比較の語彙（設計書どおり6種）。gte/lte/gt/lt は数値比較・eq は値の型に応じて
#   数値/文字列どちらでも・contains は常に文字列の部分一致。
# ★ 2026-08-27: "in"（どれか）を足した。値は**一覧**（複数の名前）。
#   意味論は 3 箇所が同時に持つ: ここ / Basic の RowMatches Case 6 /
#   Python の _extract_predicate。凍結した真理値表 tests/test_predicate_truth_table.py
#   が 3 者の一致を縛る ── 変える時は必ず一緒に直すこと。
_EXTRACT_CMPS = ("gte", "lte", "gt", "lt", "eq", "contains", "in")
_EXTRACT_CMP_LABELS = {"gte": "以上", "lte": "以下", "gt": "超", "lt": "未満",
                        "eq": "等しい", "contains": "を含む", "in": "のどれか",
                        "nin": "のどれでもない"}
_EXTRACT_CMP_CODE = {"gte": 0, "lte": 1, "gt": 2, "lt": 3, "eq": 4, "contains": 5, "in": 6, "nin": 7}
_EXTRACT_SHEET_NAME_FORBIDDEN_RE = re.compile(r'[:\\/?*\[\]]')

# ★ operator9 ①: 比較語(cmp)も A' 原則の中に入れる ── value は機械が数値化するのに、cmp の
#   種別(gte/lte/gt/lt/eq/contains)だけは LLM の言い分をそのまま検証していた（「より大きい」→
#   gte・「未満」→lte と写し間違えても素通し・境界値の行が黙って混入する実害）。
#   語の列挙は意味から広め（検体に無い自然な同義語も拾う）。
_EXTRACT_CMP_WORDS = (
    ("gt", ("より大きい", "より大きく", "を超える", "を超えて", "より多い", "より多く",
             "より高い", "より高く")),
    ("lt", ("未満", "より小さい", "より小さく", "より少ない", "より少なく",
             "より安い", "より安く")),
    ("gte", ("以上",)),
    ("lte", ("以下",)),
    ("contains", ("を含む", "を含んで", "が含まれる", "を含める")),
    ("eq", ("と等しい", "に等しい", "と同じ")),
)
# ★ 断片ガード: 「以上」「以下」は文末の定型（「以上です」等）の断片として現れやすいので、
#   直前 _EXTRACT_CMP_NUM_WINDOW 文字以内に数字が無ければ比較語として採用しない
#   （対象を値の近傍の比較語に絞る）。gt/lt/contains/eq の語は文末定型と衝突しないので対象外。
_EXTRACT_CMP_NEEDS_NUM_NEARBY = frozenset({"gte", "lte"})
_EXTRACT_CMP_NUM_RE = re.compile(r'[0-9０-９]')
_EXTRACT_CMP_NUM_WINDOW = 10


def extract_cmp_from_task(task: str) -> str | None:
    """依頼文から比較語を機械抽出する。一致が無ければ None（機械は断定しない）。
       複数の比較語が現れても、依頼文中で最初に出現した(かつ断片ガードを通った)ものを採る。"""
    if not task:
        return None
    best = None   # (出現位置, cmp名)
    for cmp_name, words in _EXTRACT_CMP_WORDS:
        for w in words:
            idx = task.find(w)
            while idx >= 0:
                if cmp_name in _EXTRACT_CMP_NEEDS_NUM_NEARBY:
                    window = task[max(0, idx - _EXTRACT_CMP_NUM_WINDOW):idx]
                    if "。" in window or not _EXTRACT_CMP_NUM_RE.search(window):
                        idx = task.find(w, idx + 1)
                        continue
                if best is None or idx < best[0]:
                    best = (idx, cmp_name)
                break
    return best[1] if best else None


# ★ グラフ段①: kind の機械抽出（cmp と同じ作法・extract_cmp_from_task の兄弟）。
#   折れ線/推移→line・円/構成比/割合/内訳→pie・棒→bar・手掛かりなし→None。
_CHART_KINDS = ("bar", "line", "pie")
_CHART_KIND_LABELS = {"bar": "棒", "line": "折れ線", "pie": "円"}
_CHART_KIND_WORDS = (
    ("line", ("折れ線", "推移")),
    ("pie", ("構成比", "割合", "内訳", "円")),
    # ★ 断片ガード①: 「棒」単独は「相棒」等の複合語と衝突するため、単独の「棒」でなく
    #   「棒グラフ」全体を語にする（凍結検体はどれも「棒グラフ」表記のみで単独「棒」を要らない）。
    ("bar", ("棒グラフ",)),
)
# ★ 断片ガード②: 「円」は単独では通貨表記（「500円」）と衝突するため、直前の文字が
#   数字（半角/全角）なら採用しない（extract_cmp_from_task の gte/lte 数字近傍ガードと同じ考え方）。
_CHART_KIND_YEN_GUARD = frozenset({"円"})
_CHART_KIND_NUM_RE = re.compile(r'[0-9０-９]')


def extract_chart_kind_from_task(task: str) -> str | None:
    """依頼文からグラフ種別を機械抽出する（extract_cmp_from_task と同じ作法）。
       一致が無ければ None（機械は断定しない）。複数の種別語が現れても、依頼文中で
       最初に出現した(かつ断片ガードを通った)ものを採る。"""
    if not task:
        return None
    best = None   # (出現位置, kind名)
    for kind_name, words in _CHART_KIND_WORDS:
        for w in words:
            idx = task.find(w)
            while idx >= 0:
                if (w in _CHART_KIND_YEN_GUARD and idx > 0
                        and _CHART_KIND_NUM_RE.match(task[idx - 1])):
                    idx = task.find(w, idx + 1)
                    continue
                if best is None or idx < best[0]:
                    best = (idx, kind_name)
                break
    return best[1] if best else None


def _format_extract_value(value) -> str:
    """EXTRACT のシート名/確認行に使う値の表示形。整数相当の float は小数点を付けない
       （40000.0 でなく 40000）。
       ★ 2026-08-27: 「どれか」の一覧は中黒で繋ぐ ── 解釈行に ['みかん', 'りんご'] と
         Python の見た目が出ていた（人が読む行に機械の書き方を出さない）。"""
    if isinstance(value, (list, tuple)):
        return "・".join(_format_extract_value(v) for v in value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else f"{value:g}"
    return str(value)


def header_row_drops_columns(ws, chosen_row: int) -> str | None:
    """採用した見出し行より**上**の行に、採用行より多くの見出しが在るなら、
       落ちる列を名指しする 1 文を返す（無ければ None）。

    ★ なぜ在るか（盲検の実データ耐性レビュー・2026-08-24 に実物で再現）:
      `商品情報`(A1:B1 結合) / `数量` / `金額` の請求書で `detect_header_row` は
      **2 行目を confident で返す**（1 行目は結合のせいで非空セルが 1 つに見える）。
      2 行目は `品名`/`規格` の 2 列しかないので**数量と金額が列ごと落ち**、
      それでも exit 0・⚠ なし・**Σ金額 の行も出ない** ── 消えたこと自体が画面に現れない。
      日本の請求書で結合見出しはほぼ普遍なので、黙っていてよい形ではなかった。
    ★ 判定そのものは変えない: 2 行目を選ぶのが正しい表も実在する（多段見出し）。
      ここでやるのは「選ばなかった行に**まだ見出しが在る**なら、それを言う」だけ。
    """
    try:
        from ailine_core.multifile import read_row_headers
        chosen = [h for h in read_row_headers(ws, chosen_row) if h != ""]
        lost = []
        for r in range(1, max(1, chosen_row)):
            above = [h for h in read_row_headers(ws, r) if h != ""]
            # ★ 誤爆の線（実測で決めた）: 上の行が**採用行より広い**時だけ見る。
            #   表題やタイトル（「請 求 書」など 1 セルだけの行）は見出しではないので、
            #   ここで拾うと普通の請求書すべてで鳴る（オオカミ少年）。
            if len(above) <= len(chosen):
                continue
            for h in above:
                if h not in chosen and h not in lost:
                    lost.append(h)
        if not lost:
            return None
        head = "・".join(f"『{h}』" for h in lost[:5])
        more = f" ほか {len(lost) - 5} 列" if len(lost) > 5 else ""
        return (f"★ 疑わしい: 見出しを {chosen_row} 行目と判断したため、"
                f"{head}{more} は取り込まれません"
                f"（結合された見出しがあると起きます ── 見出しを 1 行にまとめると取り込めます）")
    except Exception:
        return None


def _column_values(meta: dict, sheet: str, col: str, limit: int = 500) -> list:
    """対象列の中身を実際に読む（宣言でなく実体を見る側の共通の入口）。
       読めない時は空リスト ── ここで例外を投げると無関係な入力まで巻き添えで壊れる。"""
    try:
        book_path = meta.get("path")
        headers = meta.get("headers", {}).get(sheet) or []
        if not book_path or col not in headers:
            return []
        col_idx1 = headers.index(col) + 1
        header_row = int(meta.get("header_rows", {}).get(sheet, 1))
        wb = openpyxl.load_workbook(book_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet]
            return [row[0] for row in ws.iter_rows(
                min_row=header_row + 1, min_col=col_idx1, max_col=col_idx1,
                max_row=header_row + limit, values_only=True)]
        finally:
            wb.close()
    except Exception:
        return []


def _extract_column_date_kind(meta: dict, sheet: str, col: str) -> tuple:
    """対象列の中身を実際に見て、日付として比較できるかを返す（("date"|"text_date"|
       "other"|"empty", 時刻を含むか)）。

    ★ 宣言でなく実体を見る: 「日付」という見出しでも中身が文字列なら比較してはいけない
      （辞書順で "2026/3/26" > "2026/12/1" になる）。判定は ailine_core.date_compare が
      持ち、ここは**読み出しの配線だけ**（本体を ailine.py に書かない規約）。
    ★ 読めない時は ("other", False) を返して従来の断り文に落とす ── ここで例外を
      投げると、日付と無関係な入力まで巻き添えで壊れる。
    """
    try:
        book_path = meta.get("path")
        if not book_path:
            return "other", False
        headers = meta.get("headers", {}).get(sheet) or []
        if col not in headers:
            return "other", False
        return classify_date_column(_column_values(meta, sheet, col))
    except Exception:
        return "other", False


def _extract_output_sheet_name(col: str, cmp: str, value) -> str:
    """★ A': 出力シート名は機械が決め打ちで組む（LLM に名前を決めさせない・設計書の例そのまま
       ―― 列『金額』・cmp gte・value 40000 → 『金額40000以上』）。
       Excel が禁じる文字(: \\ / ? * [ ])は '_' に置き換え、31文字上限（Excel のシート名
       制限）で切り詰める。"""
    label = _EXTRACT_CMP_LABELS.get(cmp, cmp)
    shown = ("・".join(str(v) for v in value) if isinstance(value, (list, tuple))
              else _format_extract_value(value))
    # ★★ 2026-08-30（Namakoo の実測で画面に出た）: 連結だけだと
    #   『取引先丸和物流・みどり建設のどれか』── 文になっておらず、会社名に見える。
    #   ★ 大小比較（以上・未満）は連結で日本語になる（『金額40000以上』）が、
    #     一致（eq/in）は**助詞が要る** ── そこだけ「が」を挟み、語尾は落とす。
    if cmp in ("eq", "in"):
        name = f"{col}が{shown}"
    elif cmp == "nin":
        # ★ 2026-09-02: 否定は「〜以外」と書く。連結だと『料理味噌汁のどれでもない』
        #   になり日本語にならない（eq/in で助詞を挟んだのと同じ理由）。
        name = f"{col}が{shown}以外"
    else:
        name = f"{col}{shown}{label}"
    return _EXTRACT_SHEET_NAME_FORBIDDEN_RE.sub("_", name)[:31]


def _dedup_output_sheet_name(keys: list) -> str:
    """★ A'（EXTRACT の兄弟）: 出力シート名は機械が決め打ちで組む（LLM に名前を決めさせない）。
       例: keys=["取引先"] → 『取引先の重複除去』。禁止文字の置換・31文字上限は
       _extract_output_sheet_name と同じ規則（_EXTRACT_SHEET_NAME_FORBIDDEN_RE を共有）。"""
    label = "・".join(keys)
    name = f"{label}の重複除去"
    return _EXTRACT_SHEET_NAME_FORBIDDEN_RE.sub("_", name)[:31]


# ★ W9 項目4: PIVOT(DataPilot) の既知の癖（README 記載・再描画で書式が撥ねる）を
#   確認行・結果表示の両方に一言添える。AGGREGATE(SummaryTable) との使い分けを促す。
def format_confirmation_line(op: str, resolved_args: dict, inferred: set,
                              sheets: list | None = None, target_sheet: str | None = None) -> str:
    """命令言語形式の確認行を1行で組む（例: 解釈: 操作:並べ替え 対象:金額 順:降順）。
       推定で埋めた（数字表記から解決した等）引数には (推定) を付ける。
       ★ M2c: キー自体が resolved_args に無い任意項目（COMPUTE_COLUMN の target 等）は
       そのフィールドを丸ごと省略する（必須項目は常に存在するので既存の表示は変わらない）。
       ★ A': resolved_args["_sources"] に該当キーの出典があれば（例: 倍率:1.1）
       末尾に「（用語集: 消費税）」のように出典を添える（verify_dsl_args の APPEND_TOTAL が積む）。
       ★★ 挙動変更#3: sheets/target_sheet を渡すと、複数シートのブックでは先頭に
       `シート:『売上データ』(1枚目)` を載せる（本体は
       ailine_core.target_sheet.format_sheet_field・1枚のブックでは None で従来どおり）。
       その場合 LOOKUP_FILL の「対象シート:」欄は同じ値の二重表示になるので省く
       （転記の「参照シート:」はもう一方のシートなので残す）。"""
    sources = resolved_args.get("_sources") or {}
    sheet_field = format_sheet_field(sheets or [], target_sheet)
    parts = ([sheet_field] if sheet_field else []) + [f"操作:{OP_LABELS.get(op, op)}"]
    for label, key, transform in _CONFIRM_FIELDS.get(op, ()):
        if key not in resolved_args or (sheet_field and key == "target_sheet"):
            continue
        val = resolved_args.get(key)
        shown = transform(val) if transform else val
        tag = "(推定)" if key in inferred else ""
        if key in sources:
            tag += f"（{sources[key]}）"
        parts.append(f"{label}:{shown}{tag}")
    # ★★ 2026-08-27: 位置の言い回しを解いた回は、**どの op でも**根拠を出す。
    #   op ごとの _CONFIRM_FIELDS に足して回ると、足し忘れた op が黙って位置を動かす
    #   （見えない変更が一番こわい）。横断層で解いたものは横断層で見せる。
    if resolved_args.get("_at_basis") and not any(p.startswith("入れる位置:") for p in parts):
        parts.append(f"入れる位置:{resolved_args['_at_basis']}")
    return "解釈: " + " ".join(parts)


# --- ④ 決定論 codegen（op → Basic テンプレ。LLM は一切使わない） --------------

def _wrap_basic(body: str) -> str:
    """CONTRACT と同じ骨格（Option 2行 + Sub Run(oDoc As Object) 1つ）で包む。"""
    return "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\n" + body + "End Sub\n"


def _wrap_basic_for_sheet(body: str, book_meta: dict, target_sheet: str | None) -> str:
    """★ 挙動変更#2: 薄い配線。本体（対象シートの一時的な並べ替えロジック）は
       ailine_core.target_sheet.wrap_basic_for_sheet に置く（移植可能性の番人 —
       tests/ailine_py_line_budget.txt 参照）。_wrap_basic を「対象シートが1枚目のときの
       既定ラップ」としてコールバックで渡す（ailine_core → ailine の逆流を避けるため）。"""
    return wrap_basic_for_sheet(body, _wrap_basic, book_meta.get("sheets") or [], target_sheet)


def _scan_last_row_basic(var: str = "oSheet", key_col: str = "0",
                          start_row: str = "1", min_ok: str | None = None) -> str:
    """走査ループの定型（refs の作法どおり：A列を上から走査して最終データ行を探す）。
       ★ W3: start_row（Basic 0起点の走査開始行＝見出し行の直下）を渡すと、見出しが
       物理1行目でない帳票でも正しい行から走査する（既定 "1" は旧挙動と同一）。
       min_ok（"lastRow がこの値未満なら Exit Sub"の閾値）を渡すと保存境界を調整できる
       （既定は start_row と同じ＝『データ0行なら何もしない』。見出しを含めて範囲を
       スタイリングする操作は min_ok に start_row-1 相当を渡す＝データ0行でも見出しは扱える）。"""
    if min_ok is None:
        min_ok = start_row
    return (f"    lastRow = {start_row}\n"
            f"    Do While {var}.getCellByPosition({key_col}, lastRow).getString() <> \"\"\n"
            f"        lastRow = lastRow + 1\n"
            f"    Loop\n"
            f"    lastRow = lastRow - 1\n"
            f"    If lastRow < {min_ok} Then Exit Sub\n")


def codegen_dsl(op: str, resolved_args: dict, book_meta: dict, use_formula: bool = True) -> str:
    """④ 決定論 codegen。既存ヘルパへの Call を最優先し、無い操作だけテンプレ Basic を書く。
       ★ W3: book_meta["header_rows"] があれば対象シートの見出し行(1起点)をそこから読み
       （★ 挙動変更#2 より前は対象シート＝常に1枚目だった。下記参照）、
       0起点(hr0)に変換して全 op に一貫して渡す（『三層全部が同じ見出し推定を使う』の codegen 側）。
       header_rows が無い/キーが無い book_meta（_SAMPLE_META 等の旧テスト値）は既定1行目
       （hr0=0）＝旧挙動と完全に同一の Basic を生成する。
       use_formula: COMPUTE_COLUMN の既定を式（=B2*C2 等）にする（★ W3 Part3）。False で
       従来の値ベタ書きに戻す（--values）。
       ★ 挙動変更#2: 対象シートは resolved_args["_target_sheet"]（verify_dsl_args が一箇所で
       決めた値）を読む。無ければ従来どおり1枚目（後方互換 — codegen_dsl を直接呼ぶ既存の
       単体テスト・golden はこのキーを持たない args を渡しており、その挙動は変えない）。
       wrap() は _wrap_basic の対象シート対応版（_wrap_basic_for_sheet）に book_meta/
       first_sheet を閉じ込めたショートハンド — 対象シートの決定を codegen 側でも
       一箇所（_wrap_basic_for_sheet）に寄せるための配線。"""
    headers = book_meta["headers"]
    first_sheet = resolved_args.get("_target_sheet") or book_meta["sheets"][0]
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    hr0 = header_row - 1   # Basic 0起点の見出し行

    def wrap(body: str) -> str:
        # ★★ 2026-08-27（Namakoo「位置の言い回しは頻出だから全ての操作で有効に」）:
        #   新しい列を作る op は、既定でデータの**右端**に作る。依頼文が位置を言っていて
        #   機械がそれを解けたなら、作った**あとで動かす**（MoveColumnTo）。
        #   ★ op ごとに codegen を書き換えない ── ここ 1 箇所で、宣言（WRITE_NEW_COLUMN）
        #     を持つ op すべてに同じ手が付く。op が増えても配線が要らない。
        #   ★ 実測（bench/swap_formula_spike_RESULTS.md）: insertByIndex も moveRange も
        #     参照を自動で付け替えるので、動かしても式は壊れない。
        move_to = resolved_args.get("_move_new_col_to")
        if move_to is not None:
            src0 = int(resolved_args.get("_new_col_from", 0))
            body = body + ('    Call MoveColumnTo(oDoc, %d, %d)%s'
                            % (src0, int(move_to), chr(10)))
        return _wrap_basic_for_sheet(body, book_meta, first_sheet)

    if op == "SORT":
        col_idx = headers[first_sheet].index(resolved_args["col"])
        asc = "True" if resolved_args["order"] == "asc" else "False"
        last_col = len(headers[first_sheet]) - 1
        # ★ 合計行が在る回だけ、終わりの行を足して渡す。
        #   ★ 挙動が変わらない回は**生成する文字列も変えない**（凍結した検体・
        #     ゴールデンを理由なく動かさない ── 動かすと差分の意味が薄まる）。
        #   ★ 引数を増やさず**別の腕**を呼ぶ（既存の目録・README・凍結した検体を動かさない）。
        _end = resolved_args.get("_sort_end_row")
        if _end:
            return wrap(f"    Call SortByColumnUpTo(oDoc, {hr0}, {last_col}, {col_idx}, "
                         f"{asc}, {int(_end) - 1})\n")
        return wrap(f"    Call SortByColumn(oDoc, {hr0}, {last_col}, {col_idx}, {asc})\n")

    if op == "LOOKUP_FILL":
        theaders = headers[resolved_args["target_sheet"]]
        key_idx = theaders.index(resolved_args["key_col"])
        src = resolved_args["source_sheet"].replace('"', '""')
        target_col_name = resolved_args["target_col"]
        # ★ W10c 致命2: target_col が対象シートに実在しない場合（verify_dsl_args が依頼文に
        #   同じ列名があると確認済み）は、COMPUTE_COLUMN の新規列作成と同じ考え方で
        #   末尾に新しい列を作ってから転記する（無関係な既存列を上書きしない）。
        if target_col_name in theaders:
            tgt_idx = theaders.index(target_col_name)
            header_write = ""
        else:
            tgt_idx = len(theaders)   # 0起点・次の空き列
            header_name = str(target_col_name).replace('"', '""')
            header_write = (f'    oDoc.Sheets.getByIndex(0).getCellByPosition({tgt_idx}, {hr0})'
                             f'.setString("{header_name}")\n')
        return wrap(header_write +
                            f'    Call VLookupFromTable(oDoc, {hr0}, {key_idx}, {tgt_idx}, "{src}")\n')

    if op == "AGGREGATE":
        g_idx = headers[first_sheet].index(resolved_args["group_col"])
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        return wrap(f"    Call SummaryTable(oDoc, {hr0}, {g_idx}, {v_idx})\n")

    if op == "NUMBER_FORMAT":
        if resolved_args.get("_row_index"):
            _nrow0 = int(resolved_args["_row_index"]) - 1
            _lastc = max(0, len(headers.get(first_sheet) or []) - 1)
            return wrap(f"    Call FormatThousandsRow(oDoc, {_nrow0}, {_lastc})\n")
        col_idx = headers[first_sheet].index(resolved_args["col"])
        return wrap(f"    Call FormatThousands(oDoc, {hr0}, {col_idx})\n")

    if op == "MERGE":
        c1s, r1s, c2s, r2s = re.match(
            r"([A-Za-z]{1,3})(\d+):([A-Za-z]{1,3})(\d+)", resolved_args["range"]).groups()
        col1 = column_index_from_string(c1s.upper()) - 1
        col2 = column_index_from_string(c2s.upper()) - 1
        row1, row2 = int(r1s) - 1, int(r2s) - 1
        return wrap(f"    Call MergeCells(oDoc, {col1}, {row1}, {col2}, {row2})\n")

    if op == "CHART":
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        # ★ グラフ段: category_col/kind は verify_dsl_args が既定を確定させるが、codegen_dsl
        #   を単体で直接呼ぶ(ゴールデン等)呼び出し元もあるため、ここでも既定を持つ
        #   （value_col と同様の作法・省略時は1列目/"bar"）。
        cat_name = resolved_args.get("category_col") or headers[first_sheet][0]
        cat_idx = headers[first_sheet].index(cat_name)
        kind = resolved_args.get("kind") or "bar"
        # ★ operator10 ①: 合計行をグラフ範囲から除く（片配線の解消）。book_meta にファイル
        #   パスがある（実行時）場合だけ実ファイルを読める ── 手組みの book_meta（単体テスト・
        #   ゴールデン等）では従来どおり最終引数を付けない＝InsertChart が自前走査する旧挙動。
        book_path = book_meta.get("path")
        max_row_arg = ""
        if book_path is not None:
            try:
                last_row = chart_data_last_row(book_path, first_sheet, header_row)
                max_row_arg = f", {last_row - 1}"   # Basic 0起点
            except Exception:
                max_row_arg = ""
        return wrap(f'    Call InsertChart(oDoc, {hr0}, {cat_idx}, {v_idx}, "{kind}"{max_row_arg})\n')

    if op == "APPEND_TOTAL":
        # ★ W6: ヘルパは無し（罫線・カンマ等の見栄えまでは踏み込まない・素の SUM 式だけ）。
        #   データ最終行の直下に [ラベル文字列 | 合計式] を書く。
        #   ラベルは対象列の左隣に置く（既存の帳票では『合計』の文字が金額の左に来るのが自然
        #   で、対象列自体を上書きしない＝既存構造を壊さない置き方）。対象列が表の最左端
        #   （col_idx=0）の場合は左隣が無いため、値のみを書きラベルは省略する。
        # ★ B: 挿入耐性式（bench/formula_spike_work2 で実測）。SUM(D2:INDEX(D:D;ROW()-1))
        #   型で書く。INDEX(D:D;ROW()-1) は「この式自身の1行上」を指すので、後でデータ行を
        #   1本挿入しても SUM 範囲が自動で追従する（静的な "D2:D5" 型は追従しない）。
        #   ★ setFormula は LO 方言＝INDEX の2引数区切りはセミコロン(;)。カンマ(,)で
        #   書くと #VALUE!/#NAME? になる（実測: formula_spike_RESULTS.md 追記分）。
        #   保存後は自動でカンマ形 "=SUM(D2:INDEX(D:D,ROW()-1))*1.1" に変換される
        #   （check_append_total はこの保存後カンマ形と照合する）。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        label = str(resolved_args.get("label", "合計")).replace('"', '""')
        factor = float(resolved_args.get("factor", 1) or 1)
        col_letter = get_column_letter(col_idx + 1)
        start_excel_row = hr0 + 2   # データ先頭行（Basic 0起点 hr0+1）の Excel(1起点) 行
        factor_tail = "" if factor == 1 else f"*{factor:g}"
        _fixed_row = resolved_args.get("_at_row")
        if _fixed_row:
            # ★ 既にある合計行に書く（行は増やさない・ラベルはその行に既に在る）。
            body = ("    Dim oSheet As Object, totalRow As Long\n"
                     "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                     "    totalRow = " + str(int(_fixed_row) - 1) + "\n")
        else:
            body = ("    Dim oSheet As Object, lastRow As Long, totalRow As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1))
                    + "    totalRow = lastRow + 1\n")
        if col_idx > 0 and not _fixed_row:
            body += f'    oSheet.getCellByPosition(0, totalRow).setString("{label}")\n'
        body += (f'    oSheet.getCellByPosition({col_idx}, totalRow).setFormula('
                 f'"=SUM(" & "{col_letter}" & {start_excel_row} & ":INDEX(" & "{col_letter}" & '
                 f'":" & "{col_letter}" & ";ROW()-1))" & "{factor_tail}")\n')
        return wrap(body)

    if op == "CENTER_ALIGN":
        if str(resolved_args["target"]).startswith("cell:"):
            _r, _c = (int(x) for x in str(resolved_args["target"])[5:].split(","))
            return wrap('    oDoc.Sheets.getByIndex(0).getCellByPosition(%d, %d)'
                         '.HoriJustify = com.sun.star.table.CellHoriJustify.CENTER%s'
                         % (_c - 1, _r - 1, chr(10)))
        if resolved_args["target"] == "all":
            last_col = len(headers[first_sheet]) - 1
            return wrap(f"    Call AlignCenter(oDoc, {hr0}, {last_col})\n")
        # col:NAME はヘルパ無し → refs の作法（走査して範囲を求め HoriJustify）でテンプレを書く。
        col_idx = headers[first_sheet].index(resolved_args["target"][4:])
        body = ("    Dim oSheet As Object, oRange As Object, lastRow As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                + f"    oRange = oSheet.getCellRangeByPosition({col_idx}, {hr0}, {col_idx}, lastRow)\n"
                "    oRange.HoriJustify = com.sun.star.table.CellHoriJustify.CENTER\n")
        return wrap(body)

    if op == "BOLD":
        target = resolved_args["target"]
        if target.startswith("cell:"):
            _r, _c = (int(x) for x in target[5:].split(","))
            return wrap("    Call StyleBold(oDoc, %d, %d, %d, %d)%s"
                         % (_c - 1, _r - 1, _c - 1, _r - 1, chr(10)))
        if target.startswith("row:"):
            row_idx = int(target[4:]) - 1
            # ★ W3: 列幅は Basic で走査せず、接地済みの見出し列数(headers)から決定論的に
            #   決める（多段見出しで先頭列が空欄のケースで走査が誤って -1 になるのを回避）。
            last_col = len(headers[first_sheet]) - 1
            body = (f"    Call StyleBold(oDoc, 0, {row_idx}, {last_col}, {row_idx})\n")
        else:
            col_idx = headers[first_sheet].index(target[4:])
            body = ("    Dim oSheet As Object, lastRow As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                    + f"    Call StyleBold(oDoc, {col_idx}, {hr0}, {col_idx}, lastRow)\n")
        return wrap(body)

    if op == "FILL_COLOR":
        target = resolved_args["target"]
        hexcolor = COLOR_MAP[resolved_args["color"]]
        if target.startswith("cell:"):
            _r, _c = (int(x) for x in target[5:].split(","))
            return wrap('    oDoc.Sheets.getByIndex(0).getCellByPosition(%d, %d)'
                         '.CellBackColor = &H%s&%s' % (_c - 1, _r - 1, hexcolor, chr(10)))
        if target.startswith("row:"):
            row_idx = int(target[4:]) - 1
            # ★ W3: 列幅は Basic で走査せず、接地済みの見出し列数(headers)から決定論的に
            #   決める（多段見出しで先頭列が空欄のケースで走査が誤って -1 になるのを回避）。
            last_col = len(headers[first_sheet]) - 1
            body = ("    Dim oSheet As Object, c As Integer\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    f"    For c = 0 To {last_col}\n"
                    f"        oSheet.getCellByPosition(c, {row_idx}).CellBackColor = &H{hexcolor}&\n"
                    "    Next c\n")
        else:
            col_idx = headers[first_sheet].index(target[4:])
            body = ("    Dim oSheet As Object, lastRow As Long, r As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1), min_ok=str(hr0))
                    + f"    For r = {hr0} To lastRow\n"
                    f"        oSheet.getCellByPosition({col_idx}, r).CellBackColor = &H{hexcolor}&\n"
                    "    Next r\n")
        return wrap(body)

    if op == "COMPUTE_COLUMN":
        operands = resolved_args["operands"]
        operator = resolved_args["operator"]
        target = resolved_args.get("target")

        if len(operands) == 1:
            # ★ W10b 項目3: 税込み/税抜き等「列 × 率」パターン（1列 + 機械確定した factor）。
            #   codegen は既存の2列版と同じ書き方（式=セル参照、値ベタ書きの選択も同じ）。
            op1 = operands[0]
            i1 = headers[first_sheet].index(op1)
            factor = float(resolved_args["factor"])
            if target:
                new_col = headers[first_sheet].index(target)
                header_write = ""
            else:
                new_col = len(headers[first_sheet])
                # ★ W10c 中: verify_dsl_args が税込み/税抜きと判定できていれば自然な
                #   日本語見出し（例:「税込金額」）を使う。無ければ従来どおりの数式風見出し。
                header_name = str(resolved_args.get("_new_col_label")
                                   or f"{op1}{operator}{factor:g}").replace('"', '""')
                header_write = f"    oSheet.getCellByPosition({new_col}, {hr0}).setString(\"{header_name}\")\n"
            if use_formula:
                col1_letter = get_column_letter(i1 + 1)
                write_line = (f'        oSheet.getCellByPosition({new_col}, i).setFormula('
                              f'"=" & "{col1_letter}" & (i + 1) & "{operator}{factor:g}")\n')
            else:
                write_line = (f"        oSheet.getCellByPosition({new_col}, i).setValue("
                              f"oSheet.getCellByPosition({i1}, i).getValue() {operator} {factor:g})\n")
            body = ("    Dim oSheet As Object, lastRow As Long, i As Long\n"
                    "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                    + _scan_last_row_basic(start_row=str(hr0 + 1))
                    + header_write
                    + f"    For i = {hr0 + 1} To lastRow\n"
                    + write_line
                    + "    Next i\n")
            return wrap(body)

        op1, op2 = operands
        i1 = headers[first_sheet].index(op1)
        i2 = headers[first_sheet].index(op2)
        # ★ M2c: target(実在列名) 指定時はその列に書く（新規列を作らない）。
        #   無指定なら従来どおり次の空き列に新規列を作る。
        if target:
            new_col = headers[first_sheet].index(target)
            header_write = ""   # 既存の見出しはそのまま（上書きしない）
        else:
            new_col = len(headers[first_sheet])   # 0起点で次の空き列
            # ★ W3 改定(2026-08-20): verify_dsl_args が「target は実在しないが依頼文に
            #   実在する語」と判定していれば _new_col_label に利用者の指名が入っている
            #   （単列×率パターンの税込/税抜ラベルと同じ仕組み）。無ければ従来どおりの
            #   数式風見出し。
            header_name = str(resolved_args.get("_new_col_label")
                               or f"{op1}{operator}{op2}").replace('"', '""')
            header_write = f"    oSheet.getCellByPosition({new_col}, {hr0}).setString(\"{header_name}\")\n"
        # ★ W3 Part3: 既定は式（setFormula）。formula_spike の実測どおり、単純な行内の
        #   二項演算（=B2*C2 型）は区切り記号もシート参照も無いため LO 方言(;/.)の
        #   気遣いが不要（多引数関数・シート跨ぎ参照だけが方言の対象＝ここでは無縁）。
        #   行番号だけが Basic 側ループ変数(i)で動くので、その部分だけ実行時に連結する。
        if use_formula:
            col1_letter = get_column_letter(i1 + 1)
            col2_letter = get_column_letter(i2 + 1)
            write_line = (f'        oSheet.getCellByPosition({new_col}, i).setFormula('
                          f'"=" & "{col1_letter}" & (i + 1) & "{operator}" & "{col2_letter}" & (i + 1))\n')
        else:
            write_line = (f"        oSheet.getCellByPosition({new_col}, i).setValue("
                          f"oSheet.getCellByPosition({i1}, i).getValue() {operator} "
                          f"oSheet.getCellByPosition({i2}, i).getValue())\n")
        body = ("    Dim oSheet As Object, lastRow As Long, i As Long\n"
                "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                + _scan_last_row_basic(start_row=str(hr0 + 1))
                + header_write
                + f"    For i = {hr0 + 1} To lastRow\n"
                + write_line
                + "    Next i\n")
        return wrap(body)

    # --- ★ W9: 検証済みヘルパ4種の語彙昇格。いずれも helpers/*.bas のヘルパは headerRow
    #   引数を取らない（物理1行目を前提に自前走査する既存実装・ここでは変更しない）ため、
    #   codegen 側も hr0 を渡さずそのまま Call するだけ。
    if op == "INSERT_ROWS":
        at0 = int(resolved_args["at"]) - 1   # 1起点(Excel行番号) → 0起点(Basic)
        count = int(resolved_args.get("count", 1) or 1)
        return wrap(f"    Call InsertRows(oDoc, {at0}, {count})\n")

    # ★ 2026-08-26: 表の基本操作 3 種（Namakoo が GUI を触って欠けを実測）。
    #   ★ insertByIndex / removeByIndex を使う ── 挿入は**途中でも押し下げ**、
    #     削除は**詰める**（clearContents だと空行が残る）。
    if op == "ADD_ROW":
        at0 = int(resolved_args["at"]) - 1
        values = resolved_args.get("values") or {}
        headers = list(resolved_args.get("_headers") or [])
        idx, vals, kinds = [], [], []
        for name, v in values.items():
            if name not in headers:
                continue
            idx.append(str(headers.index(name)))
            # ★ 型は書く値そのものから決める。数値を文字列で書くと下流の SUM が壊れる
            #   （この repo が何度も測ってきた「静かに壊れる」形）。
            if isinstance(v, bool):
                vals.append("TRUE" if v else "FALSE"); kinds.append("s")
            elif isinstance(v, (int, float)):
                vals.append(repr(v)); kinds.append("n")
            else:
                vals.append(str(v)); kinds.append("s")
        sep = chr(1)
        body = (
            f'    Call AddRowWithValues(oDoc, {at0}, "{",".join(idx)}", '
            f'"{sep.join(vals)}", "{",".join(kinds)}")\n')
        # ★★ 2026-09-02: 既存の行が式で出している列は、新しい行にも式を写す。
        #   ★ 値は作らない（A' 原則）── **隣の行から写すだけ**。
        #     参照の付け替えは LibreOffice にやらせる（自前で式を書き換えない）。
        _inh = list(resolved_args.get("_inherit_cols") or [])
        _src = int(resolved_args.get("_inherit_from") or 0)
        if _inh and _src:
            body += ('    Call FillFormulasFromNeighbour(oDoc, %d, %d, "%s")%s'
                      % (at0, _src - 1, ",".join(str(c) for c in _inh), chr(10)))
        return wrap(body)

    if op == "DELETE_ROWS":
        at0 = int(resolved_args["at"]) - 1
        count = int(resolved_args.get("count", 1) or 1)
        return wrap(f"    Call DeleteRows(oDoc, {at0}, {count})\n")

    if op == "DELETE_COLUMN":
        headers = list(resolved_args.get("_headers") or [])
        col0 = headers.index(resolved_args["col"]) if resolved_args["col"] in headers else 0
        return wrap(f"    Call DeleteColumn(oDoc, {col0})\n")

    if op == "EXTRACT_COLUMNS":
        hdrs = list(resolved_args.get("_headers") or [])
        idx = ",".join(str(hdrs.index(c)) for c in resolved_args["cols"] if c in hdrs)
        hr0x = int(resolved_args.get("_header_row", 1)) - 1
        dst = str(resolved_args["_new_sheet"]).replace(chr(34), chr(34) * 2)
        return wrap('    Call ExtractColumns(oDoc, %d, "%s", "%s")%s'
                     % (hr0x, idx, dst, chr(10)))

    if op == "SET_WHERE":
        headers = list(resolved_args.get("_headers") or [])
        wcol = headers.index(resolved_args["col"]) if resolved_args["col"] in headers else 0
        ccol = (headers.index(resolved_args["cond_col"])
                 if resolved_args["cond_col"] in headers else 0)
        hr0 = int(resolved_args.get("_header_row", 1)) - 1
        code = _EXTRACT_CMP_CODE[resolved_args["cmp"]]
        thr = resolved_args["cond_value"]
        # ★ 2026-08-27（実測・生 traceback を出した）: 閾値が数値かどうかは **cmp ではなく
        #   値そのもの**で決まる。eq は「金額が 100 と等しい」にも「チェックが『◎』と
        #   等しい」にも使う ── cmp で分けると、置き換えの『◎』を float() に渡して落ちる。
        _num = _is_number(thr) and not isinstance(thr, bool)
        thr_lit = (repr(float(thr)) if _num
                    else '"%s"' % str(thr).replace(chr(34), chr(34) * 2))
        val = str(resolved_args["value"]).replace(chr(34), chr(34) * 2)
        # ★ 外す行は**構造の事実**なので Python が渡す（条件の判定は Basic が自分で行う ──
        #   そこを渡すと事後条件が独立した検算でなくなる）。
        skip = ",".join(str(int(r) - 1) for r in (resolved_args.get("_skip_rows") or []))
        return wrap('    Call SetColumnValueWhere(oDoc, %d, %d, %d, %d, %s, "%s", "%s")%s'
                     % (hr0, wcol, ccol, code, thr_lit, val, skip, chr(10)))

    if op == "ADD_COLUMN":
        # ★ 位置は Python が見出しから決めた 1 起点 → Basic は 0 起点。
        #   名前は Basic 側で書く（空なら見出しも空のまま）。
        at0 = int(resolved_args["_at_col"]) - 1
        hr0 = int(resolved_args.get("_header_row", 1)) - 1
        nm = str(resolved_args.get("name") or "").replace(chr(34), chr(34) * 2)
        return wrap('    Call InsertColumnAt(oDoc, %d, "%s", %d)%s' % (at0, nm, hr0, chr(10)))

    if op == "SWAP":
        # ★ 位置は Basic 側にも**名前で**渡す（Python が数えた番号を渡さない）。
        #   Basic は実文書を走査して自分で見つける ── Python の解決と食い違えば、
        #   事後条件が「宣言した位置が入れ替わっていない」と落とす（独立な 2 実装）。
        hr0 = int(resolved_args.get("_header_row", 1)) - 1
        _a = str(resolved_args["a"]).replace(chr(34), chr(34) * 2)
        _b = str(resolved_args["b"]).replace(chr(34), chr(34) * 2)
        if resolved_args.get("_axis") == "cell":
            # ★ 2 セルの入れ替え。中身は verify が実表から読んだもの（LLM 由来ではない）。
            #   ★ 式のセルは setFormula で書く（setString だと文字列になる）。
            _cs = resolved_args["_cells"]
            _vs = resolved_args["_cell_values"]
            _body = ""
            for (r, c), v in zip(_cs, reversed(_vs)):     # ★ 入れ替え＝相手の値を書く
                if isinstance(v, str) and v.startswith("="):
                    _esc = formula_for_basic(v).replace(chr(34), chr(34) * 2)
                    _body += ('    Call SetFormulaAt(oDoc, %d, %d, "%s")%s'
                               % (r - 1, c - 1, _esc, chr(10)))
                elif _is_number(v):
                    _body += ('    Call SetCellAt(oDoc, %d, %d, "%s", "n")%s'
                               % (r - 1, c - 1, repr(float(v)), chr(10)))
                else:
                    _esc = str("" if v is None else v).replace(chr(34), chr(34) * 2)
                    _body += ('    Call SetCellAt(oDoc, %d, %d, "%s", "s")%s'
                               % (r - 1, c - 1, _esc, chr(10)))
            return wrap(_body)
        if resolved_args.get("_axis") == "column":
            _body = ('    Call SwapColumnsByName(oDoc, "%s", "%s", %d)%s'
                      % (_a, _b, hr0, chr(10)))
        else:
            # ★★ 2026-08-31: 行番号で指された回は**座標で**渡す。名前で渡すと Basic は
            #   表から『6行目』という名前を探して見つけられない（実測で別の行が動いた）。
            #   ★ 名前で指された回は今までどおり名前を渡す ── あちらは Basic が自分で
            #     位置を見つけるので、事後条件が独立な検算になる。
            if (_re_row_number_word.fullmatch(_a.strip())
                    and _re_row_number_word.fullmatch(_b.strip())):
                _body = ('    Call SwapRowsAt(oDoc, %d, %d)%s'
                          % (resolved_args["_a_pos"] - 1, resolved_args["_b_pos"] - 1,
                             chr(10)))
            else:
                _body = ('    Call SwapRowsByName(oDoc, "%s", "%s", 0, %d)%s'
                          % (_a, _b, hr0, chr(10)))
        # ★ 入れ替えの**あと**に、写像を通した式を書き戻す（順序が意味を持つ）。
        for _r, _c, _f in resolved_args.get("_formula_rewrites") or ():
            _esc = formula_for_basic(_f).replace(chr(34), chr(34) * 2)
            _body += ('    Call SetFormulaAt(oDoc, %d, %d, "%s")%s'
                       % (_r - 1, _c - 1, _esc, chr(10)))
        return wrap(_body)

    if op == "SET_CELL_VALUE":
        headers = list(resolved_args.get("_headers") or [])
        col0 = headers.index(resolved_args["col"]) if resolved_args["col"] in headers else 0
        v = resolved_args["value"]
        kind = "n" if resolved_args.get("_write_numeric") else "s"
        val = (repr(float(resolved_args["_write_numeric_value"]))
                if kind == "n" else str(v))
        hr0 = int(resolved_args.get("_header_row", 1)) - 1
        # ★ 2026-08-28: 行番号で指された回は**座標で**書く（探し直す相手が無い・同名の
        #   行が 2 つある表でも狙いが定まる）。名前で指された回は今までどおり名前を渡す
        #   ── あちらは Basic が自分で位置を見つけるので、事後条件が独立な検算になる。
        if resolved_args.get("row_number"):
            return wrap('    Call SetCellAt(oDoc, %d, %d, "%s", "%s")%s'
                         % (int(resolved_args["row_number"]) - 1, col0, val, kind, chr(10)))
        return wrap(
            f'    Call SetCellByName(oDoc, "{resolved_args["row"]}", 0, {col0}, '
            f'"{val}", "{kind}", {hr0})\n')

    if op == "DRAW_BORDERS":
        return wrap("    Call DrawTableBorders(oDoc)\n")

    if op == "AUTOFIT":
        return wrap("    Call AutoFitColumns(oDoc)\n")

    if op == "PIVOT":
        g_idx = headers[first_sheet].index(resolved_args["group_col"])
        v_idx = headers[first_sheet].index(resolved_args["value_col"])
        return wrap(f"    Call PivotSum(oDoc, {g_idx}, {v_idx})\n")

    if op == "SET_COLUMN_VALUE":
        # ★ 致命3(W10e): ヘルパ無し・既存列のデータ行全部に同じ値を書く
        #   （CENTER_ALIGN の col: 分岐と同じ「走査してヘッダ直下から最終行まで」の作法）。
        # ★ operator10 ④: verify_dsl_args が列の実体（書換前）から機械決定した型
        #   （resolved_args["_write_numeric"]）に従い、数値列には setValue で数値を書く
        #   （既定/判定できない場合は従来どおり setString）。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        body_head = ("    Dim oSheet As Object, lastRow As Long, r As Long\n"
                     "    oSheet = oDoc.Sheets.getByIndex(0)\n"
                     + _scan_last_row_basic(start_row=str(hr0 + 1)))
        if resolved_args.get("_write_numeric"):
            num = float(resolved_args["_write_numeric_value"])
            num_lit = str(int(num)) if num.is_integer() else repr(num)
            body = (body_head
                    + f"    For r = {hr0 + 1} To lastRow\n"
                    f"        oSheet.getCellByPosition({col_idx}, r).setValue({num_lit})\n"
                    "    Next r\n")
        else:
            value = str(resolved_args["value"]).replace('"', '""')
            body = (body_head
                    + f"    For r = {hr0 + 1} To lastRow\n"
                    f"        oSheet.getCellByPosition({col_idx}, r).setString(\"{value}\")\n"
                    "    Next r\n")
        return wrap(body)

    if op == "EXTRACT":
        # ★ ヘルパへの Call 1行だけ（helpers/AiLineHelpers.bas:ExtractRows）。
        #   型を保つコピー（getValue/setValue・getString/setString の分岐）は helper 側。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        cmp_code = _EXTRACT_CMP_CODE[resolved_args["cmp"]]
        # ★ 日付比較のときだけ、表示用の文字列でなくシリアル値を Basic へ渡す
        #   （resolved["value"] は解釈行と出力シート名のために元の文字列で残してある）。
        value = resolved_args.get("_value_serial", resolved_args["value"])
        if isinstance(value, (list, tuple)):
            # ★「どれか」の一覧は Chr(2) 区切り（値にカンマ・タブが入りうる）。
            #   Basic 側 RowMatches Case 6 が同じ区切りで**丸ごと一致**を見る。
            joined = chr(2).join(str(v) for v in value)
            value_lit = '"' + joined.replace('"', '""').replace(chr(2), '" & Chr(2) & "') + '"'
        elif isinstance(value, str):
            value_lit = '"' + value.replace('"', '""') + '"'
        else:
            value_lit = f"{float(value):g}"
        dst_name = str(resolved_args["_new_sheet"]).replace('"', '""')
        # ★ 外す行は**構造の事実**なので Python が渡す（条件の判定は Basic が自分で行う
        #   ── そこを渡すと事後条件が独立した検算でなくなる）。SET_WHERE と同じ作法。
        # ★ Basic のループは **0 起点**（他のヘルパと同じ）── 1 起点の行番号を渡すと
        #   1 行ずれて効かない（実測: 「8」を渡したのに 8 行目が抜き出された）。
        #   SetColumnValueWhere も同じく -1 して渡している。
        _x_skip = ",".join(str(int(r) - 1) for r in (resolved_args.get("_skip_rows") or []))
        return wrap(f'    Call ExtractRows(oDoc, {hr0}, {col_idx}, {cmp_code}, '
                     f'{value_lit}, "{dst_name}", "{_x_skip}")\n')

    if op == "SPLIT_CELL":
        # ★ ヘルパへの Call 1行だけ（helpers/AiLineHelpers.bas:SplitColumn）。
        #   新しい見出し名は Python 側で決めて渡す（Basic 側で名前を作らない・帳票段と同じ作法）。
        col_idx = headers[first_sheet].index(resolved_args["col"])
        sep_lit = '"' + str(resolved_args["sep"]).replace('"', '""').replace(chr(10), '" & Chr(10) & "') + '"'
        names_csv = ",".join(str(n).replace(",", "_") for n in resolved_args["_new_cols"])
        return wrap(f'    Call SplitColumn(oDoc, {hr0}, {col_idx}, {sep_lit}, "{names_csv}")' + chr(10))

    if op == "DEDUP":
        # ★ ヘルパへの Call 1行だけ（helpers/AiLineHelpers.bas:DedupRows・EXTRACT と同じ作法）。
        #   キー列は複数ありうるため、0起点の列インデックスをカンマ区切りの文字列で渡す
        #   （Basic の Call 引数に配列リテラルを直接書けないため・ヘルパ側で Split する）。
        key_idxs = [headers[first_sheet].index(k) for k in resolved_args["keys"]]
        key_idx_csv = ",".join(str(i) for i in key_idxs)
        dst_name = str(resolved_args["_new_sheet"]).replace('"', '""')
        return wrap(f'    Call DedupRows(oDoc, {hr0}, "{key_idx_csv}", "{dst_name}")\n')

    if op == "REPORT_PER_ROW":
        # ★ ヘルパへの Call を行数ぶん（helpers/AiLineHelpers.bas:FillReportSheet）。
        #   ★ B: 呼ぶ前に一意名を解決済み（verify_dsl_args の unique_sheet_name）── 失敗
        #   しうる名前を copyByName に渡さない。ループは Python 側（Basic 側でシート名を
        #   作らない・設計文書の指示どおり「名前は Python 側が全部決めてから渡す」）。
        template_sheet = str(resolved_args["template_sheet"]).replace('"', '""')
        src_sheet = str(first_sheet).replace('"', '""')
        lines = []
        groups = resolved_args.get("_groups")
        if groups:
            # ★★ まとめ版（2026-08-28）: 1 グループ = 1 枚。明細行が件数ぶん増える。
            #   行を増やすと下がずれる ── ずれの数え方は report_group.output_rows_for に
            #   1 箇所だけ置き、確かめる側もそこを使う（埋める側と数え方が割れない）。
            det0 = resolved_args.get("_detail_row")
            det0 = -1 if det0 is None else int(det0) - 1
            for g in groups:
                gname = str(g["sheet"]).replace('"', '""')
                rows_csv = ",".join(str(int(r) - 1) for r in g["rows"])
                lines.append(
                    f'    Call FillGroupReportSheet(oDoc, "{template_sheet}", "{gname}", '
                    f'"{src_sheet}", "{rows_csv}", {hr0}, {det0})' + chr(10))
        for rr in ([] if groups else resolved_args.get("_report_rows", [])):
            new_name = str(rr["sheet"]).replace('"', '""')
            src_row0 = int(rr["row"]) - 1   # Excel(1起点) → Basic(0起点)
            lines.append(
                f'    Call FillReportSheet(oDoc, "{template_sheet}", "{new_name}", '
                f'"{src_sheet}", {src_row0}, {hr0})\n'
            )
        # ★ 2026-08-24（土台固め）: 検分シートも**同じ Basic の中で**書く。
        #   旧実装は生成後に openpyxl でブックを開き直して足しており、その往復が
        #   xl/drawings の中の図形（描かれた角印・社判）を全部捨てていた（実測）。
        #   LO 側で書けば往復そのものが無くなる ── 追加の LO 起動も要らない。
        insp = resolved_args.get("_inspection_sheet")
        report_rows = resolved_args.get("_report_rows") or []
        if insp and groups:
            # ★ まとめた回は「元の行」が複数 ── 何行をまとめたかまで検分に出す
            #   （どの発注が 1 枚に入ったかを、後から人が追えること）。
            lines.append(inspection_sheet_basic_call(
                insp, ["シート名", "元の行", "まとめた件数"],
                [[g["sheet"], ",".join(str(r) for r in g["rows"]), len(g["rows"])]
                 for g in groups], "ssn"))
        elif insp and report_rows:
            n_ph = len(resolved_args.get("_placeholders") or [])
            lines.append(inspection_sheet_basic_call(
                insp, ["シート名", "元の行", "埋めた印の数"],
                [[rr["sheet"], rr["row"], n_ph] for rr in report_rows], "snn"))
        return wrap("".join(lines))

    if op == "FORMAT_MAP":
        # ★ ヘルパへの Call 1行だけ（helpers/AiLineHelpers.bas:FillFormatMapSheet）。
        #   REPORT_PER_ROW と同じ理由: 出力シート名は呼ぶ前に一意名を解決済み。印の実在は
        #   verify_dsl_args が検証済み・行の一覧(_data_rows)も Python 側で決め切る
        #   （Basic 側でシート名・行の取捨選択をしない）。
        template_sheet = str(resolved_args["template_sheet"]).replace('"', '""')
        src_sheet = str(first_sheet).replace('"', '""')
        dst_sheet = str(resolved_args["_output_sheet"]).replace('"', '""')
        header_tpl_row0 = int(resolved_args["_header_tpl_row"]) - 1
        ph_tpl_row0 = int(resolved_args["_placeholder_tpl_row"]) - 1
        src_rows_csv = ",".join(str(int(r) - 1) for r in resolved_args.get("_data_rows", []))
        # ★ 2026-08-24（土台固め）: 検分シートも同じ Basic の中で書く（帳票段と同じ処置・
        #   同じ関数）。openpyxl の往復が図形を捨てるので、往復自体を無くす。
        fm_calls = (
            f'    Call FillFormatMapSheet(oDoc, "{template_sheet}", "{src_sheet}", "{dst_sheet}", '
            f'{header_tpl_row0}, {ph_tpl_row0}, {hr0}, "{src_rows_csv}")' + chr(10))
        insp = resolved_args.get("_inspection_sheet")
        data_rows = resolved_args.get("_data_rows") or []
        if insp and data_rows:
            n_ph = len(resolved_args.get("_placeholders") or [])
            out_sheet = resolved_args.get("_output_sheet")
            fm_calls += inspection_sheet_basic_call(
                insp, ["出力シート", "出力行", "元の行", "埋めた印の数"],
                [[out_sheet, k + 1, src_row, n_ph]
                 for k, src_row in enumerate(data_rows, start=1)], "snnn")
        return wrap(fm_calls)

    raise ValueError(f"未対応の op: {op}")


# --- ⑥ op 別事後条件（達成の機械検証。openpyxl で out ファイルを読むだけ・LO 不要） ----

def duplicate_name_warning(col: str, values) -> str | None:
    """シート名の元になる列に同じ値が複数あるなら、その事実を名指しする 1 文（無ければ None）。

    ★ 実測（盲検の使い勝手レビュー・2026-08-24）: 取引先 3 社の売上表（4 行）に
      「取引先ごとに請求書を作って」と頼むと**請求書が 4 枚**でき、同じ取引先が
      『あかつき商事』(120,000) と『あかつき商事_2』(64,000) の 2 枚に分かれた。
      それでも「データ4行 → 出力4枚」で ✓ が出る。帳票段は **1 行 1 枚**の op なので
      機械としては正しいが、依頼者の「ごとに」は**まとめて 1 枚**を指している。
    ★ 機械は `_2` を付けたその瞬間に重複を知っている。知っていて黙るのが一番悪い。
      1 枚にまとめるかどうかは人が決めることなので、**名指しして人に返す**
      （★ 付き → 決裁③で ✓→△）。
    """
    seen = {}
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        key = str(v).strip()
        seen[key] = seen.get(key, 0) + 1
    dupes = sorted(((k, n) for k, n in seen.items() if n > 1), key=lambda t: (-t[1], t[0]))
    if not dupes:
        return None
    head = "・".join(f"『{k}』{n} 行" for k, n in dupes[:3])
    more = f" ほか {len(dupes) - 3} 件" if len(dupes) > 3 else ""
    # ★ 印（⚠）は _warnings を印字する側が付ける ── ここで ★ を足すと「⚠ ★」と二重になる。
    # ★ 2026-08-28: 「先に集計してから」は**行き止まり**だった（この道具の中に道が無い）。
    #   まとめ方は在る ── 雛形に明細の印を置けば 1 枚にまとまる。出口を名指しする。
    #   ★ 断らないのは、取引ごとに 1 枚が正しい帳票（領収書・納品書）があるから。
    return (f"列『{col}』に同じ値が複数あります（{head}{more}）。"
             f"1 行につき 1 枚ずつ作るので、同じ相手の書類が別々の枚に分かれます"
             f"（1 枚にまとめるなら、雛形の明細行に『{{{{明細:列名}}}}』の印を、"
             f"合計欄に『{{{{合計:金額}}}}』の印を置いてください。"
             f"取引ごとに 1 枚が正しい書類 ── 領収書・納品書 ── ならこのままで大丈夫です）")


def detect_first_column_gap(ws, header_row: int = 1, look_ahead: int = 200) -> str | None:
    """A 列（1列目）を上から走査して止まった位置より**下にまだ中身がある**なら、その事実を
       名指しする 1 文を返す（無ければ None）。

    ★ なぜ在るか（盲検レビュー・2026-08-24）: データ行の数え方（_scan_last_row）は
      1 列目を上から見て最初の空セルで打ち切る。ところが**計画側も検証側も同じ関数**を
      使うので、「データ N 行 → 出力 N 枚」の完全会計が**恒真**になっていた。
      実測: 5 行の表が 3 行と数えられる。30 社の一覧で 12 行目の顧客名が空（結合セルの
      2 行目・月の区切りの空行は普通の書き方）なら、10 枚しか出ないのに ✓ が出る。
    ★ 直し方の選択: 数え方そのものを変えると、合計行の除外など既存の全機構の前提が動く。
      ここでは**縮んだ事実を必ず言う**に留める（★ 付きなので決裁③で ✓→△ に降格する）。
      「黙って少なく処理する」から「少なく処理したと言う」へ ── これが最小の正直さ。
    """
    stop = header_row
    while ws.cell(row=stop + 1, column=1).value not in (None, ""):
        stop += 1
    counted = stop - header_row
    last_seen = None
    for r in range(stop + 1, stop + 1 + look_ahead):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last_seen = r
    if last_seen is None:
        return None
    return (f"★ 疑わしい: 1列目が {stop + 1} 行目で空いているため、データ行を {counted} 行と数えました"
            f"（{last_seen} 行目にはまだ中身があります）。"
            f"1列目に空欄がある表は、その手前までしか処理されません")


# ★ 止血1/2 共通の文言。事後条件チェッカーは検証対象0件を絶対に「合格」にしない
#   （D検体: no-opを『行数が少なく比較不要』で素通ししていた根治）。
# --- ★ 算術恒等の検算（二重計上・合計行の位置） -------------------------------
#   独立レビューの実測: check_append_total は期待値を「合計式が生成したのと同じ範囲」から
#   作っていた ―― 検算が被検算と同じ盲点を使う恒真式。既存の合計 300 を持つ表に合計を
#   足すと 600 が書かれ「3 行の合計を検証」と言って ✓ が出た。並べ替えにも同型がある
#   （合計行が2行目に来ても「5 行を検証（降順）」で通る）。
#   判定は ailine_core/sum_identity.py（語も書式も読まない純粋な算術）に閉じ込め、
#   ここは「どのセルか」を人が読める文にするだけにする。


def note_extent_gap(out_book: Path, resolved_args: dict, header_row: int = 1) -> None:
    """走査で見た範囲と、**物理の使用範囲**の食い違いを ⚠ に変える ── op に依らず 1 箇所で。

    ★ 2026-08-25（塊①・盲検 2 回目の致命①②）: extent_gap は在ったが、20 ある事後条件の
      うち **check_sort 1 本にしか挿さっていなかった**。同じ検体（末尾行の 1 列目が空）で
      並べ替えだけが △ に落ち、集計・抽出・重複除去・一括書換・太字・桁区切り・グラフは
      すべて ✓ を名乗った ── 見ていない行が在るのは**表の性質**であって op の性質ではない。

    ★ だから各 checker に配って回らない（今日までに片配線を 7 回踏んだ）。
      全 op が必ず通る run_postcondition の入口で、1 度だけ数える。

    ★ 判定（pass/fail）は 1 ビットも変えない。✓ を △ に降ろすのは決裁③の既存の機構に任せる。
    """
    if not isinstance(resolved_args, dict):
        return
    try:
        with BookView(out_book) as bv:
            ws = bv.sheet(resolved_args.get("_target_sheet"))
            gap = extent_gap(ws, header_row=header_row)
    except Exception as e:
        # ★ 出ないことは信号でない: 測れなかったことを黙って飲まず、そう言う。
        note_unverified(resolved_args, 1,
                        f"表の範囲を測れませんでした（{type(e).__name__}）")
        return
    if gap["rows_missing"]:
        # ★★ 2026-08-31（Namakoo が実測・「めちゃくちゃ惜しい」）:
        #   「取引先と項目を入れ替えて」で**合計行の見出し『合計』も一緒に動き**、
        #   1 列目が空になって走査が止まっていた。★ 道具は気づいていた（この ⚠ が出た）
        #   のに、文面は「1 列目が空」としか言わず、**何が起きたか伝わらなかった**。
        #   ★ 挙動は 1 ビットも変えない ── **心当たりを名指しする**だけ。
        #     「合計」は行の見出しであって列の中身ではないので、列を入れ替えると
        #     人の期待とずれる（Excel でも同じことが起きる ── だから言う値打ちがある）。
        _why = "1 列目が空のため走査がそこで止まり、この行を見ていない"
        if str(resolved_args.get("_axis") or "") == "column":
            _why += ("（心当たり: 合計行の見出しが 1 列目に在ると、"
                      "列の入れ替えで一緒に動きます）")
        note_unverified(resolved_args, gap["rows_missing"], _why)
    if gap["cols_missing"]:
        # ★ 2026-08-27 に文言を正した: 「見出しの無い列が N 列」は不正確だった。
        #   走査は見出し行を左から見て**最初の空で止まる**ので、空きの右にある列は
        #   （見出しが在っても）まとめて見えない ── 数えているのはその「見えない列」の数。
        note_unverified(resolved_args, gap["rows_physical"],
                        f"見出し行に空きがあるため、その右の {gap['cols_missing']} 列は"
                        "走査できておらず、何も確かめていない")



def check_chart(path: Path, before_charts: int) -> tuple:
    after = _charts_count(path)
    if after != before_charts + 1:
        return "fail", f"グラフ数が +1 でない（{before_charts} → {after}）"
    return "pass", f"グラフ数 {before_charts} → {after}"


# --- ★ W9: 検証済みヘルパ4種の事後条件 ---------------------------------------
#   ヘルパ本体(helpers/*.bas)が headerRow を取らないのと同じ理由で、これらのチェッカーは
#   openpyxl の生スキャン(_scan_last_row/_scan_last_col)で実データ範囲を都度見つける
#   （header_row を渡しはするが、ヘルパが物理1行目前提で動く以上、通常は 1 のまま使う想定）。

# ★ 2026-08-27: 入れ替えの依頼を見分ける（第二段翻訳へ回すための**証拠**であって、
#   名前をここから取るためのものではない ── 名前は LLM が言い、機械が実表で確かめる）。
_re_swap_ask = re.compile(
    r"(?:入れ?替え|入替|交換|逆に\s*し|前後を\s*入れ)")
# 依頼文が軸（行/列）をはっきり書いている場合だけ、その語を採る。
_re_swap_axis_row = re.compile(r"行\s*(?:同士\s*)?(?:を|の|で)?\s*[^。]{0,6}?(?:入れ?替え|入替|交換)")
_re_swap_axis_col = re.compile(r"列\s*(?:同士\s*)?(?:を|の|で)?\s*[^。]{0,6}?(?:入れ?替え|入替|交換)")


def task_asks_for_a_swap(task: str) -> bool:
    """依頼文が「入れ替え」を求めているか（軸も対象もここでは決めない）。"""
    return bool(_re_swap_ask.search(task or ""))


def _swap_pair_resolves(book_meta: dict, sheet: str | None, a: str, b: str) -> bool:
    """2 つの名前が、実表で**行としても列としても**ちょうど 1 つに解けるか（どちらかでよい）。

    ★ なぜ「読み直す前」に確かめるのか: 一段目が SORT を返す言い方があり（実測:
      「順番を逆にして」）、黙って横取りすると**正当な並べ替えを壊す**。
      読み直してよいのは、機械が対象を解けている時だけ ── ADD_ROW で同じ線を引いた。
    ★ ここでは軸を決めない（決めるのは verify_dsl_args 1 箇所）。解けるかだけを見る。
    """
    if not a or not b or a == b:
        return False
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    if a in headers and b in headers:
        return True
    ra, _ = _resolve_named_row(book_meta, sheet, a)
    rb, _ = _resolve_named_row(book_meta, sheet, b)
    return ra is not None and rb is not None


# 「AとBの〈列〉」── 2 つの行の、同じ 1 列を指す言い方（末尾の「の」まで含めて当てる）。
_re_between_and = re.compile(r"^.*?([^\s、。との]+?)\s*と\s*([^\s、。との]+?)\s*の$")


def swap_targets_are_rows(task: str, book_meta: dict, sheet: str | None,
                           header_row: int = 1) -> list | None:
    """依頼文が **2 つの行**の入れ替えを指しているなら [行番号, 行番号]（でなければ None）。

    ★★ 2026-09-02（入れ替えを効果の検体に載せて初めて見えた）:
      「あかね商事とうえだ物産の行を入れ替えて」で、読み直しの二段目（op を SWAP に
      固定して LLM に聞き直す）が **a='取引先' b='件数'** を返した ── 人が言っていない
      **列名**。しかも実在の列なので `_swap_pair_resolves` は True を返す。
      止まったのは三項の番人が「依頼文の語と照合できない」と気づいたからで、判断は
      正しいが、**利用者の正当な依頼が通らない**。
      ★ 3 表（在庫・名簿・献立）では同じ言い方が 6/6 通っていた ── **LLM の揺れ**。

    ★ 処方は 8/31 にセルでやったものと同じ:
      **LLM に聞く前に、機械だけで 2 つ解けているならそれを使う。**
      依頼文と実表しか見ていないので、LLM の返事より確かで、速い。
    ★ 語彙を数え上げない ── 実表の値が依頼文に literal で現れ、それが**ちょうど 2 行**に
      決まる時だけ。決まらなければ None（推測しない）。
    ★ 見出しの語と、数のように見える値は行の名前にしない
      （`_row_named_anywhere_in_task` と同じ理由 ── 揺れを増幅しない）。
    """
    rows, heads = _table_rows_for_anchor(book_meta, sheet, header_row)
    if not rows:
        return None
    text = _task_outside_quotes(task)
    head_set = {h for h in heads if h}
    hits = {}
    for r, vals in rows.items():
        for v in vals:
            if (not v or v in head_set or len(v) < 2 or v not in text
                    or _is_number_like(v)):
                continue
            hits.setdefault(v, set()).add(r)
    # ★ 1 行に決まる名前だけを採る（同じ値が 2 行に在るなら名前で指せていない）
    named = sorted({next(iter(rs)) for v, rs in hits.items() if len(rs) == 1})
    return named if len(named) == 2 else None


def swap_targets_are_cells(task: str, book_meta: dict, sheet: str | None,
                            header_row: int = 1) -> list | None:
    """依頼文が **2 つのセル**の入れ替えを指しているなら [(行,列), (行,列)]（でなければ None）。

    ★★ 2026-08-31（Namakoo が実測・✓ が出たのに操作が違った）:
      「丸和物流の**単価**とみどり建設の**単価**を入れ替えて」で、
      **行を丸ごと入れ替えて ✓ を出していた**（16 セルが動いた・頼んだのは 2 セル）。
      番人は「宣言どおり行が入れ替わったか」を見るので、**宣言そのものが違えば通る**
      ── 三項（依頼・宣言・実体）の「依頼」が抜けた形。
    ★ 依頼文には証拠が在る: **両側とも「〜の〈列名〉」と列を名指ししている**。
      行の入れ替えなら列は出てこない。
    ★ 言い回しを数え上げない ── 見るのは**実表に在る見出し**と**実表に在る行**だけ。
      「〈何か〉の〈実在する列〉」がちょうど 2 つ在り、その〈何か〉が行として一意に
      決まるときだけ、セルの入れ替えと読む。決まらなければ None（推測しない）。
    """
    text = _task_outside_quotes(task)
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    if not headers:
        return None
    found = []
    for h in sorted(headers, key=len, reverse=True):   # 長い見出しから（部分一致よけ）
        start = 0
        while True:
            i = text.find("の" + h, start)
            if i < 0:
                break
            start = i + 1
            name = text[:i].rsplit("と", 1)[-1].strip(" 　、。")
            for w in ("を", "は", "が", "の"):
                name = name.split(w)[-1] if name.endswith(w) else name
            if name:
                found.append((i, name, h))
    # ★★ 2026-08-31（Namakoo が実測・2 つ目の形）:
    #   「丸和物流**と**近江スチール**の項目**を入れ替えて」は、列名が **1 回しか出ない**。
    #   前の形（〈A〉の〈列〉と〈B〉の〈列〉）だけを見ていたので拾えず、
    #   行を丸ごと入れ替えていた（人は 2 セルのつもり）。
    #   ★ 「〈A〉と〈B〉の〈列〉」＝ **2 つの行の、同じ 1 列**。実表に照らして解ける
    #     ときだけ（列は実在の見出し・A と B は行として一意）── 推測しない。
    if len(found) == 1:
        _i, _name, _h = found[0]
        _pair = _re_between_and.match(text[:_i] + "の")
        if _pair:
            _rows = []
            for _nm in (_pair.group(1).strip(), _pair.group(2).strip()):
                _r, _ = _resolve_named_row(book_meta, sheet, _nm)
                if _r is None:
                    return None
                _rows.append(_r)
            if _rows[0] != _rows[1] and _h in headers:
                _c = headers.index(_h) + 1
                return [(_rows[0], _c), (_rows[1], _c)]
        return None
    if len(found) != 2:
        return None
    cells = []
    for _i, name, h in sorted(found):
        row, note = _resolve_named_row(book_meta, sheet, name)
        if row is None or h not in headers:
            return None
        cells.append((row, headers.index(h) + 1))
    return cells if cells[0] != cells[1] else None


def _swap_axis_hint(task: str) -> str | None:
    """依頼文が『行を』『列を』とはっきり書いているときだけ "row"/"column" を返す。
       ★ 両方書いてある／どちらも無いなら None ── 推測しない（機械が実表で決める側へ回す）。"""
    t = task or ""
    row = bool(_re_swap_axis_row.search(t))
    col = bool(_re_swap_axis_col.search(t))
    if row == col:
        return None
    return "row" if row else "column"


_re_between = re.compile(r"([^\s、。]+?)\s*と\s*([^\s、。]+?)\s*の\s*間")


def _re_anchor(suffix: str):
    return re.compile(r"([^\s、。]+?)" + re.escape(suffix))


_ANCHOR_AFTER = ("の下に", "の下へ", "の後に", "の後ろに", "の次に")
_ANCHOR_BEFORE = ("の上に", "の上へ", "の前に")


_re_row_number_word = re.compile(r"[0-9０-９]+\s*行(?:目)?")
def _row_word_number(word: str) -> int:
    """「4行目」「４行」→ 4（全角も受ける）。"""
    digits = "".join(ch for ch in word if ch.isdigit() or ch in "０１２３４５６７８９")
    return int(digits.translate(_ZENKAKU_DIGITS))


def row_number_anchor(task: str) -> tuple:
    """「4行目の下に」「2行目の前に」── **行番号と向き**だけから位置を出す。

    戻り値: (入れる行, 依頼文が言った行番号, 説明) ── 当たらなければ (None, None, "")。
    ★ 表に訊く必要が無い（純関数）。だから**位置を決める側と、その位置を審査する側の
      両方が同じここを通る**（片配線を作らない ── この repo が何度も踏んだ形）。
    """
    text = (task or "").replace("　", " ")
    # ★★ 2026-08-29（Namakoo）:「4行目と5行目は両方ともヤマノ食品。取引先で指定は
    #   出来ない」── 中身で指せない表では、人は番号でしか言えない。ならば
    #   「4行目と5行目の間に」も同じ引き算で出す（ここも表に訊く必要が無い）。
    m = _re_between.search(text)
    if m:
        a_, b_ = m.group(1).strip(), m.group(2).strip()
        if _re_row_number_word.fullmatch(a_) and _re_row_number_word.fullmatch(b_):
            na, nb = _row_word_number(a_), _row_word_number(b_)
            if nb - na != 1:
                return None, None, ""        # 隣り合っていない ── 決めない
            return nb, na, f"{na}行目と{nb}行目の間＝{nb}行目"
        return None, None, ""                # 片方でも名前なら、表に訊く側の仕事
    for sufs, after in ((_ANCHOR_AFTER, True), (_ANCHOR_BEFORE, False)):
        for suf in sufs:
            m = _re_anchor(suf).search(text)
            if not m:
                continue
            name = m.group(1).strip()
            if not _re_row_number_word.fullmatch(name):
                return None, None, ""      # 名前で指している（表に訊く側の仕事）
            n = _row_word_number(name)
            at = n + 1 if after else n
            return at, n, f"{n}行目の{chr(0x4E0B) if after else chr(0x4E0A)}に入れる＝{at}行目"
    return None, None, ""


# ★★ 2026-08-31（Namakoo の提案した通しを俺が先に走らせて出た・1 幕目が全滅）:
#   「**8行目に**丸山工業の行を作って」で、名前として『8行目に丸山工業』を丸ごと
#   切り出していた（区切りが空白と読点しか無く、**行番号をまたいで飲み込む**）。
#   ★ そのとき task_names_a_row_number は正しく 8 を返していた ── **行番号が
#     分かっているのに、名前の切り出しがそれを無視していた**。
#   ★ 助詞と行番号の語は名前に含まれない ── そこで切る（語彙ではなく文法の線）。
_re_row_of = re.compile(r"([^\s、。をにへはがでとのも]+?)\s*の\s*行")
_re_row_unit = re.compile(r"[0-9０-９]*\s*行\s*(?:を|も)?\s*(?:足|追加|入れ|挿入)")
_re_value_assign = re.compile(r"[^\s、。]+\s*(?:は|を|＝|=)\s*[0-9０-９]")


def row_anchor_names(task: str) -> list:
    """依頼文が**位置の目印**として使っている名前（「丸和物流と近江スチールの間に」の 2 つ）。

    ★ 目印は「置く物」ではない ── ここを分けないと、目印がそのまま新しい行の値になる
      （実測: 「丸和物流と近江スチールの間に北斗精機を作って」で `取引先=丸和物流`）。
    """
    text = _task_outside_quotes(task).replace("　", " ")   # ★ 引用符の中は値（上と同じ線）
    out = []
    m = _re_between.search(text)
    if m:
        out += [m.group(1).strip(), m.group(2).strip()]
    else:
        for suf in list(_ANCHOR_AFTER) + list(_ANCHOR_BEFORE):
            m = _re_anchor(suf).search(text)
            if m:
                out.append(m.group(1).strip())
                break
        else:
            m2 = _re_row_of.search(text)
            if m2:
                out.append(m2.group(1).strip())
    return [s for s in out if s]


def anchor_column_name(book_meta: dict, sheet: str | None, names: list,
                        header_row: int = 1) -> str | None:
    """目印の名前が**実際に入っている列**の見出し（決まらなければ None）。

    ★ 「A と B の間に X」の X は、A・B と**同じ列**の住人（取引先の間には取引先が入る）。
      置き場所を LLM に決めさせると別の列へ入る（実測: `項目=北斗精機`）。
    ★★ 2026-08-29（Namakoo「どうしても中身でさせない場面が出てくる」→ 効果検体で実測）:
      目印が**行番号**（「3行目の上に新品を入れて」）だと、名前が表のどこにも無いので
      列が決まらず、値の行き先を失って**空行**になっていた ── 位置の解決・op の選択に
      続いて、これが同じ非対称の 3 層目。
      ★ ここも表に訊けば決まる: その行が**自分の名前を持っている列**（左から見て最初に
        値の在る列）。「A の隣には A と同じ列の住人が入る」を、行番号で言われた時に
        言い直しただけ ── 語の一覧は増やさない。
      ★ 決めた列は解釈行の「入れる値」にそのまま出る（黙って置かない）。
    """
    path = book_meta.get("path")
    headers = (book_meta.get("headers") or {}).get(sheet) or []
    if not path or not headers or not names:
        return None
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, last_col = data_extent(ws, header_row)
            cols = set()
            width = min(last_col, len(headers))
            for nm in names:
                if _re_row_number_word.fullmatch(nm):
                    rn = _row_word_number(nm)
                    if header_row < rn <= last:
                        for c in range(1, width + 1):
                            if str(ws.cell(row=rn, column=c).value or "").strip():
                                cols.add(c)
                                break
                    continue
                for r in range(header_row + 1, last + 1):
                    for c in range(1, width + 1):
                        if str(ws.cell(row=r, column=c).value or "").strip() == nm:
                            cols.add(c)
    except Exception:
        return None
    return headers[cols.pop() - 1] if len(cols) == 1 else None


def add_row_values_from_request(task: str, book_meta: dict, sheet: str | None,
                                 llm_values, header_row: int = 1) -> dict:
    """新しい行に入れる値を、**依頼文から**決める（LLM の出した値は篩にかけるだけ）。

    ★★ 2026-08-28（Namakoo が実測・3 回とも別々に壊れた）:
      「丸和物流と近江スチールの間に北斗精機を作って」に対し、第二段は
        取引先=丸和物流／項目=北斗精機／件数=1／単価=件／金額=未定／締め日=未定
      を返した。**位置の目印が値になり・置く物が別の列に入り・存在しない値
      （未定・未設定・件）がでっち上げられた**。
    ★ A' 原則をここでも通す ── 値は依頼文に literal で在るものだけ。
      ・目印の名前は落とす（それは位置であって値ではない）
      ・列を人が名指ししている値だけ、その列へ（「売上は600」）
      ・列が名指しされていない値は**目印と同じ列**へ（置く物そのもの）
      ・残りは**空のまま**（埋めない ── 空欄は誤値より安い）
    """
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    anchors = row_anchor_names(task)
    if isinstance(llm_values, dict):
        pairs = [(str(k), v) for k, v in llm_values.items()]
    else:
        pairs = [(headers[i], v) for i, v in enumerate(llm_values or []) if i < len(headers)]
    kept, payload = {}, None
    for col, val in pairs:
        s = str(val).strip()
        if not s or s in anchors or s not in (task or ""):
            continue
        if col in headers and col in (task or ""):
            kept[col] = val
        elif payload is None:
            payload = val
    acol = anchor_column_name(book_meta, sheet, anchors, header_row)
    # ★★ 2026-08-29（84 件の効果検体で最後に残った 1 件）: 「鈴木の上に新品を入れて」で
    #   第二段が値として**目印そのもの**（氏名=鈴木）だけを返した回、篩で全部落ちて
    #   空になり、呼び出し側が「置き換え無し」と見て**悪い値のまま**通していた。
    #   ★ 機械の引き算は、篩が空になった回にも使う（LLM が何も出さない回と同じ扱い）。
    _bare_all = bare_value_from_task(task, anchors, acol, headers)
    if payload is None:
        payload = _bare_all
    if payload is not None and acol:
        # ★★ 2026-08-29（効果検体の第 2 回で出た新しい穴）: 「味噌汁の上に**新品**を入れて」で
        #   値が『新』になった。第二段が『新』を返し、篩は「依頼文に literal で在る」だけを
        #   見ていたので通した ── **短い部分文字列は必ず通ってしまう**。
        #   ★ 機械の引き算（bare_value_from_task）も同じ依頼から値を出せる。
        #     両方とも依頼文由来なら、**長い方**を採る（部分だけ書くのは必ず間違い）。
        if (_bare_all and str(payload) in str(_bare_all)
                and len(str(_bare_all)) > len(str(payload))):
            payload = _bare_all
        kept.setdefault(acol, payload)
    return kept


_RIGHT_WORDS = ("の右", "の隣", "のとなり", "の右隣")
_LEFT_WORDS = ("の左", "の左隣")


_re_row_word = re.compile(r"(?:第)?\s*[0-9０-９]{1,4}\s*行(?:目)?")
_re_a1_col_word = re.compile(r"[A-Za-z]{1,2}\s*列")
# ★ 助詞と語尾は**閉じた文法の集合**（業務語彙の列挙ではない）。落としても意味は減らない。
_TAIL_WORDS = ("にして", "にする", "に変えて", "に変える", "と入れて", "と書いて",
                "を入れて", "を書いて", "を追加して", "を追加", "を記入して", "を記入",
                "にセット", "入れて", "書いて", "変えて", "して", "ください", "です")
_PARTICLES = "をにへはとがのでも、。 　"


# ★ 末尾の「を＋動詞」（を作って／をつくる／を新設して…）。**語でなく形**で書く。
#   ★ 語尾（て・た・る…）は**必須**にする。省略可にすると「担当を佐藤」の
#     『を佐藤』まで食って、値そのものを消してしまった（実測）。
_re_verb_tail = re.compile(
    r"を[^\sをにへはがでとのも、。]{1,6}(?:て|た|る|ます|ください|下さい)\s*$")


def bare_value_from_task(task: str, row_name: str | None, col_name: str | None,
                          headers=None) -> str | None:
    """依頼文から、機械が**引き算で**書き込む値を切り出す。

    ★★ 2026-08-29（Namakoo が実測）: 「丸山重工の右にPCパーツ」で、第二段は
      row/col しか返さず **value を返さなかった**（qwen も gemma4 も）。
      LLM が値を出さないなら、機械が出す ── 機械は既に「誰の行か」「どの列か」を
      知っているので、依頼文からそれらを**引く**だけでいい。
    ★ 引くのは: 行の名前・列の名前・見出しの語・「N行目」「F列」・位置の語・助詞と語尾。
      どれも閉じた集合（業務語彙の列挙ではない）。
    ★ 残りが**依頼文の中に連続した文字列として在る**ことを最後に確かめる
      ── 切れ端を継ぎ足した幽霊の値を作らないため。
    """
    text = (task or "")
    out = text
    # ★ row_name は 1 つとは限らない（「AとBの間に」は目印が 2 つ）── 並びも受ける。
    names = list(row_name) if isinstance(row_name, (list, tuple)) else [row_name]
    for w in names + [col_name] + [str(h) for h in (headers or [])]:
        if w:
            out = out.replace(str(w), " ")
    out = _re_row_word.sub(" ", out)
    out = _re_a1_col_word.sub(" ", out)
    # ★ 位置の語（列の左右だけでなく、**行の上下と間**も落とす）。
    #   ★ 2026-08-29: 「味噌汁**の上に**新品を入れて」で『上に新品』が値になっていた
    #     ── 列の語だけ落として行の語を落としていなかった（また行と列の非対称）。
    for w in ("の右隣", "の左隣", "のとなり", "のあいだ", "の間",
               "の右", "の左", "の隣", "の上", "の下", "の前", "の後ろ", "列"):
        out = out.replace(w, " ")
    for w in _TAIL_WORDS:
        out = out.replace(w, " ")
    # ★★ 2026-08-30（1B の検体で 6 件・7B でも同じ形で落ちていた）:
    #   「ボルトとナットの間に**新品を作って**」で値が取れず、空行の挿入に落ちていた。
    #   「を追加して」は _TAIL_WORDS に在り、「を作って」は無い ── **列挙の穴そのもの**。
    #   ★ 動詞を数え上げると必ず漏れる（この repo が何度も踏んだ形）。
    #     語尾は**閉じた文法**なので、語ではなく**形**で書く:
    #       「を」＋（助詞を含まない短い語）＋（て／た／る／…）が末尾に付いていたら落とす。
    #   ★ セルに書く値は名詞なので、この形が値の一部になることはまずない。
    out = _re_verb_tail.sub(" ", out)
    #   ★ 「1行足して」のように**行そのものを足す**依頼は値ではない（凍結済みの
    #     述語を借りる ── 新しい語を数え上げない）。実測で『足』が値になりかけた。
    if _re_row_unit.search(task or ""):
        return None
    out = out.strip(_PARTICLES).strip()
    while out and out[0] in _PARTICLES:
        out = out[1:]
    while out and out[-1] in _PARTICLES:
        out = out[:-1]
    if not out or " " in out or "　" in out:
        return None                       # 2 つ以上に割れた ── 決めない
    if out not in text:
        return None                       # 連続していない ── 継ぎ足した値は使わない
    # ★★ 2026-08-29: 「スプリング**を作って**」がそのまま値になった。動詞の語尾を
    #   数え上げても必ず漏れる（今日 3 度目）── **文法の線**で弾く:
    #   セルに書く値の中に助詞は入らない。残っていたら、それは文がまだ切れていない証拠。
    if any(ch in out for ch in "をにへはがでとのも"):
        return None
    bad = {str(h) for h in (headers or [])} | {str(row_name or ""), str(col_name or "")}
    if out in bad:
        return None
    return out


def _is_number_like(s) -> bool:
    """数字だけの文字列か（依頼文に出る数と、行の名前を混同しないため）。"""
    try:
        float(str(s).replace(",", ""))
        return True
    except (TypeError, ValueError):
        return False


def _cell_row_name_for(book_meta: dict, sheet, row: int, header_row: int = 1):
    """その行を人が呼ぶときの名前（1 列目の値）。分からなければ None。"""
    rows, _h = _table_rows_for_anchor(book_meta, sheet, header_row)
    vals = rows.get(row) or []
    return vals[0] if vals else None


def resolve_cell_target_from_task(task: str, book_meta: dict, sheet: str | None,
                                   header_row: int = 1):
    """依頼文 → **1 つのセル**（行・列・その根拠）。決まらなければ None。

    ★★ 2026-08-29（Namakoo が実測・両モデルとも外した）:
      「丸山重工の右にPCパーツ」を qwen は SPLIT_CELL（区切り文字を聞く）、
      gemma4 は ADD_ROW at:1（見出し行に挿す）と読んだ。
      「丸山重工の項目をPCパーツにして」は OUT_OF_VOCAB / SET_COLUMN_VALUE（列を全部潰す）。
    ★ どれも「1 セルに書く」だけの依頼で、**機械は既に答えを知っている** ──
      丸山重工は 8 行目、項目は 2 列目。誰も表に訊いていなかった。
    ★ だからモデルを替えても直らない。**行も列も機械が実表から決める**。

    列の決め方は 3 つ（強い順）:
      ① 見出しの名前が依頼文に**ちょうど 1 つ**現れる（「項目を」）
      ② A1 の列名（「F列に」）
      ③ 行の名前が入っているセルからの**相対**（「〜の右」「〜の隣」「〜の左」）
    ★ 2 つ以上の見出しが現れたら決めない（推測で別の列に書かない）。
    """
    rows, heads = _table_rows_for_anchor(book_meta, sheet, header_row)
    if not rows or not heads:
        return None
    # --- 行 ---
    name = None
    row = task_names_a_row_number(task)
    if row is not None and row not in rows:
        return None
    if row is None:
        hit = _row_named_anywhere_in_task(task, rows, heads, require_possessive=True)
        if not hit:
            return None
        row, name = hit
        # ★★ 2026-08-29（既存の検体が捕まえた・俺の横取り）: この読み直しは
        #   「金額の**合計**を一番下に出して」の『合計』や、「数量を**10**に」の『10』を
        #   行の名前と読んで、正当な依頼を 1 セル書換に化けさせた。
        #   ★ 歯止めを 2 つ ── どちらも文法であって業務語彙ではない:
        #     ① 数字は行の名前にしない（依頼文には数字が普通に出る）
        #     ② 人がセルを指すときは「**〜の**」と言う（「丸山重工の項目」「高橋の右」）。
        #        「合計**を**」のように別の助詞が付くなら、それは操作の説明であって
        #        行の名指しではない。
        if _is_number_like(name) or f"{name}の" not in (task or ""):
            return None
    basis_row = f"{row}行目" if name is None else f"『{name}』の行＝{row}行目"

    # --- 列 ---
    text = task or ""
    named = [h for h in heads if h and h in text]
    if len(named) == 1:
        return row, heads.index(named[0]) + 1, f"{basis_row}／列は『{named[0]}』"
    if len(named) > 1:
        return None                      # 見出しが 2 つ以上 ── 決めない

    for raw in re.findall(r"([A-Za-z]{1,2})\s*列", text):
        try:
            idx = column_index_from_string(raw.upper())
        except ValueError:
            continue
        if 1 <= idx <= len(heads):
            # ★ 2026-08-30（実測で画面に出た）: 見出しの無い列だと「列は『』（B列）」と
            #   **空の名前**が出ていた。名前が無いなら英字だけで言う（嘘の空欄を見せない）。
            _hd = str(heads[idx - 1] or "").strip()
            return row, idx, (f"{basis_row}／列は『{_hd}』（{raw.upper()}列）" if _hd
                               else f"{basis_row}／列は {raw.upper()}列（見出しなし）")

    if name:
        try:
            base = rows[row].index(name) + 1
        except ValueError:
            return None
        step = 0
        if any(w in text for w in _RIGHT_WORDS):
            step = 1
        elif any(w in text for w in _LEFT_WORDS):
            step = -1
        if step:
            idx = base + step
            if 1 <= idx <= len(heads):
                where = "右" if step > 0 else "左"
                return row, idx, (f"{basis_row}／列は『{heads[idx - 1]}』"
                                   f"（『{name}』の 1 つ{where}）")
    return None


def value_written_in_task(task: str, llm_value, headers=None) -> str | None:
    """書き込む値は**依頼文に literal で在るもの**だけ（A' 原則）。

    ★ 実測: 第二段は『未定』『未設定』のような、どこにも書かれていない値を返す。
      見出しの語をそのまま値にすることもある ── どちらも入れさせない。
    """
    s = "" if llm_value is None else str(llm_value).strip()
    if not s or s not in (task or ""):
        return None
    if headers and s in {str(h) for h in headers}:
        return None
    return s


# ★ 「取り除く」意味の語。**ここは語の列挙で正しい** ── 判定しているのが
#   「表のどこか」ではなく「人がどの動作を言ったか」だからだ（動作は言葉でしか分からない）。
#   ★ 漏れた時の壊れ方が違うことが大事: 語が無ければ**何も起きない**（今までどおり）。
#     黙って別のことをするのではない。しかも読みは書く前に画面に出る。
_REMOVAL_WORDS = ("削除", "消して", "消す", "除いて", "除く", "取り除", "抜いて", "無くして",
                   "いらない", "要らない", "不要")
# ★ 「〜以外」は**別の意味**（残す側を選ぶ）。今はその操作が無いので、混ぜずに断る。
_EXCEPT_WORDS = ("以外", "を除いた", "を抜いた")


def removal_reading(task: str, book_meta: dict, sheet: str | None, header_row: int = 1):
    """「その行を消す」と読める依頼なら (行番号, 説明) を返す。無ければ None。

    ★★ 2026-08-29（84 件の効果検体で最後まで残った穴・3 表とも同じ形）:
      「味噌汁の行を**除いて**」が、削除にならなかった（抽出に化ける／断られる）。
      一段目は 3 表で EXTRACT / OUT_OF_VOCAB / 条件付き抽出 と返し分けた。
    ★ 「除く」は日本語として 2 通りに読める:
        ① その行を**消す**            → DELETE_ROWS（この道具にある）
        ② それ**以外**を残す/抜き出す → EXTRACT の cmp『〜でない』（**まだ無い**）
      ★ ②の語（以外・を除いた）が在る回は、①に化けさせない ── 断って要望に記録する。
        ここで①を選ぶと、残したかった行を消すという**取り返しのつかない**間違いになる。
    ★ 行そのものが決まらない回も決めない（推測で別の行を消さない）。
    """
    text = task or ""
    if any(w in text for w in _EXCEPT_WORDS):
        return None
    if not any(w in text for w in _REMOVAL_WORDS):
        return None
    if not _re_row_of.search(text) and not _row_named_anywhere_in_task(
            text, *_table_rows_for_anchor(book_meta, sheet, header_row)):
        return None
    at, note = resolve_row_anchor(text, book_meta, sheet, header_row=header_row)
    if at is None or not note:
        return None
    return at, note


def task_says_except(task: str) -> bool:
    """依頼文が「〜以外」を言っているか（語の集合は removal_reading と共有）。"""
    return any(w in (task or "") for w in _EXCEPT_WORDS)


def except_extraction_reading(book_meta: dict, sheet: str | None, task: str,
                               header_row: int = 1) -> tuple:
    """「〜以外を抜き出して」を、抽出（cmp=nin）として読む。決まらなければ (None, None)。

    ★★ 2026-09-02: これまでは**名指しで断って**いた（README「作らなかったこと」）:
      「この述語は Python・Basic・凍結した真理値表の 3 箇所が独立に持つので、
       締切前に触ると 3 つの同期がずれる」── 正しい判断だった。
      ★ 締切は過ぎ、**ずれたら赤くなる番人**を先に置いた
        （tests/test_compare_codes_stay_in_sync.py）ので、触れる状態になった。
    ★ 読みは機械がやる（LLM に「以外」を教えない）── 列も値も
      resolve_named_extraction が実表から解く。同じ道具を使い、否定だけを足す。
    ★ 決まらなければ決めない ── ここで外すと、**残したい行を落とす**。
    """
    if not task_says_except(task):
        return None, None
    return resolve_named_extraction(book_meta, sheet, task, header_row)


CHOICE_PREFIX = "候補: "


def render_refusal(op: str, resolved_or_args, reason: str) -> list:
    """断りを、**利用者の言葉**で 3 行にする。

    ★★ 2026-08-30（Namakoo）:「断りの理由が対象とする非エンジニアにとって理解しがたい。
      『なんでできないの？』という感覚が先に来る」
      実測で出ていた断り:
        ？ 演算子『+』は列1つの計算（税込み/税抜き等）では * か / のみ対応です
        ？ 列『ナット』がこの表にありません（ある列: 品名、棚、数量、備考）
      ★ どちらも**機械の内側の状態**を説明していて、頼んだ人の言葉になっていない。
        「ナットを削除して」と頼んだ人に列の話をしても、何を直せばいいか分からない。
    ★★ 一番効くのは**「こう読みました」の 1 行**だと判断した ── 断りの多くは
      「できない」ではなく「**違うふうに読まれた**」なので、そこを見せれば
      『なんでできないの？』が『ああ、読み違えられたのか』に変わる。
      直し方も自分で分かる（言い直せばよい、と分かる）。
    ★ 理由そのものは消さない ── 事実は落とさない。順番と枠だけ変える。
    """
    lines = ["？ できませんでした"]
    label = OP_LABELS.get(op)
    if label:
        args = resolved_or_args if isinstance(resolved_or_args, dict) else {}
        # ★ 内部の引数名（col=…）ではなく、解釈行と**同じ日本語の欄名**で見せる
        #   ── 画面の他の場所と言葉が揃っていないと、そこでまた迷わせる。
        _labels = {k: lab for lab, k, _fn in _CONFIRM_FIELDS.get(op, ())}
        shown = "、".join(
            f"{_labels.get(k, k)}={v}" for k, v in args.items()
            if not str(k).startswith("_") and v not in (None, "", [], {})
            and not isinstance(v, (dict, list, tuple)))
        head = f"  依頼を『{label}』と読みました"
        lines.append(f"{head}（{shown}）" if shown else head)
    lines.append(f"  止めた理由: {reason}")
    phrases = (OP_META.get(op) or {}).get("match_phrases") or []
    if label and phrases:
        # ★ match_phrases は語の断片のこともある（「演算」「掛け算」）。
        #   「こう言えば通る」と書くと嘘になるので、**その操作を指す言い方**として出す。
        lines.append(f"  『{label}』を指す言い方: "
                      + "／".join(f"「{p}」" for p in phrases[:3]))
    lines.append("  読み方そのものが違うなら、言い直してください（頼める操作の一覧: ailine ops）")
    return lines


def render_choices(choices) -> str:
    """2 通り以上に読める時の**選べる形**（機械可読 1 行 × N）。

    ★★ 2026-08-29: 断りを行き止まりにしない ── 「どちらですか」を返す。
    ★ 形は `候補: <op>	<人が読む説明>`。画面はこれをボタンにし、押した候補を
      `--op` で**固定して**実行する（当て直しが起きない）。
    ★ ポップアップは使わない（モーダルは画面を止める・実測で踏んだ）。
    ★ 説明は op 名でなく**効果**で書く ── 人は op 名を知らない。
    """
    return chr(10).join(f"{CHOICE_PREFIX}{op}{chr(9)}{why}" for op, why in choices)


# ★ 「数値の見せ方」を言う語。**op を当てる**ための語なので列挙で正しい
#   （動作は言葉でしか分からない）。漏れても何も起きない（今までどおり）。
_NUMFMT_WORDS = ("金額表示", "通貨表示", "通貨", "カンマ区切り", "カンマ", "桁区切り",
                  "3桁", "３桁", "円表示", "見やすい数字")


def task_asks_for_number_format(task: str) -> bool:
    """依頼文が桁区切り（金額表示）を求めているか。

    ★★ 2026-08-29（Namakoo が実測）:「合計を金額表示にして」が**合計追加**に読まれ、
      既にある合計をもう一度書いて ✓ が出た（画面は何も変わらない）。
      『合計』は**対象**であって操作ではないのに、そちらに引かれていた。
    ★ 『金額』単体では発火させない ── 「金額列の合計を出して」を横取りしないため。
    """
    return any(w in (task or "") for w in _NUMFMT_WORDS)


def number_format_target(task: str, book_meta: dict, sheet: str | None,
                          header_row: int = 1):
    """桁区切りを掛ける先を機械が決める → ("col", 列名) / ("row", 行番号) / None。

    ★ 列でも行でも掛けられる（行と列は軸違い・今日そこの非対称を 3 回踏んだ）。
    ★ 決まらなければ決めない（推測で別の場所に書式を掛けない）。
    """
    text = task or ""
    # ★★ 2026-08-29: 「合計を**金額表示**にして」の『金額』を列名と読んでいた
    #   （部分文字列の穴 ── この repo で 3 度目）。操作の語を**先に取り除いてから**
    #   列名を探す。「金額列を桁区切りにして」は取り除いても『金額』が残るので当たる。
    stripped = text
    for w in _NUMFMT_WORDS:
        stripped = stripped.replace(w, " ")
    heads = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    named = [h for h in heads if h and h in stripped]
    if len(named) == 1:
        return ("col", named[0])
    if len(named) > 1:
        return None
    rows = total_rows_in(book_meta, sheet, header_row)
    if len(rows) == 1 and any(w in text for w in ("合計", "小計", "総計")):
        return ("row", rows[0])
    return None


def formula_for_basic(formula: str) -> str:
    """xlsx の式（引数の区切りが `,`）を、Basic の `setFormula` が読める形（`;`）にする。

    ★★ 2026-08-29（実測で分かった環境の事実・推測では出ない）:
      同じ式を 3 通り書いて読み戻した:
        =SUM(E2:INDEX(E:E,ROW()-1))  → **#VALUE!**
        =SUM(E2:INDEX(E:E;ROW()-1))  → 通る
        =SUM(E2:E8)                  → 通る（区切りが無いので影響なし）
      ★ `setFormula` の引数区切りは **`;`**。カンマのまま渡すと式は入るが計算できない
        （文字は正しく見えるのに値だけ壊れる ── 一番たちが悪い形）。
    ★ 文字列の中のカンマは触らない（`=IF(A1;"a,b";"c")` の中身を壊さない）。
    """
    out, in_q = [], False
    for ch in str(formula or ""):
        if ch == chr(34):
            in_q = not in_q
        out.append(";" if (ch == "," and not in_q) else ch)
    return "".join(out)


# ★ 式を書き直す本数の上限。これを超えたら書き直さない（黙って諦めない・下で言う）。
FORMULA_REWRITE_LIMIT = 200


def formula_rewrites_for_shift(book_meta: dict, sheet: str | None, shift) -> tuple:
    """操作前の式を写像に通して、**操作後に在るべき式**を並べる。

    戻り値: (書き直し {(行, 列): 式}, 断りの理由 or None)

    ★★ 2026-08-29（Namakoo「合計行ごと参照を変えずに追記したいってことじゃないの？」）:
      「税込み金額と金額を入れ替えて」が × になった。断り自体は正しかった（実測で
      合計式が二列にまたがり、両方 1,000,440 ＝ 金額＋税込み金額 になっていた）。
      ★ だが利用者が欲しいのは**意味を保ったまま位置だけ入れ替わった表**で、
        それは機械が全部言える ── 操作前の式と π が分かっているのだから、
        操作後の式は π(操作前) でしかない。
      ★ だから「後から直す」のではなく **最初から正しく書く**。LibreOffice の
        自動付け替えを当てにしない（範囲の片側だけ動かすことがある・実測）。
      ★ 合っているセルに同じ内容を書き戻すのは無害 ── 見分けるより確実。
    ★ 消える参照が 1 つでもあれば**書き直さずに断る**（壊れた式を書かない）。
    ★ 本数が多すぎる回は書き直さない ── ただし黙らない（呼び出し側が理由を出す）。
    """
    path = book_meta.get("path")
    if not path or not sheet:
        return {}, None
    try:
        cells = {}
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        cells[(cell.row, cell.column)] = v
    except Exception:
        return {}, None                    # 読めない回は何もしない（断定しない）
    if not cells:
        return {}, None
    if len(cells) > FORMULA_REWRITE_LIMIT:
        return {}, (f"式が {len(cells)} 個あるため、参照の書き直しは行いません"
                     f"（上限 {FORMULA_REWRITE_LIMIT} 個）")
    out, lost = cellmap.formulas_after(cells, shift)
    if lost:
        where = "、".join(f"{r}行{c}列" for r, c in lost[:3])
        return {}, f"この操作で参照が消える式があります（{where}）── 実行しません"
    return out, None


def formula_columns_to_inherit(book_meta: dict, sheet: str | None, header_row: int,
                                at: int, declared: set) -> tuple:
    """新しい行に式を引き継ぐ列（0 起点の列番号の一覧）と、**写す元の行**（1 起点）。

    ★★ 2026-09-02（README の「既知の問題」に自分で書いていた）:
      「みかんの下に梨を追加して」の後、梨の行の利益列は**空のまま**だった。
      宣言した値だけを書くので `✓` は正しいが、**人が期待するものとは違う**。
    ★ 式は発明ではない ── **隣の行から写す**（依頼文にも実表にも無い値は作らない＝A' 原則）。
      参照の付け替えは LibreOffice にやらせる（自分で式の文字列を書き換えると、
      それは 2 つ目の参照解決の実装になる ── SwapRowsByName が moveRange を使うのと同じ線）。

    ★ 引き継ぐ条件は **全データ行が式を持っていること**。これで合計列が自然に外れる:
      金額列は E2..E7 が直値・E8 だけ =SUM なので「全部が式」ではない。
      逆に 税込金額 は全行 =E*1.1 なので引き継ぐ。**列挙ではなく形で決める。**
    ★ 合計行は写す元にしない（=SUM を新しい行に配ると壊れる）。判定は既存の
      凍結規則（total_rows_in）を借りる ── 同じことを 2 箇所で決めない。
    ★ 人が値を指定した列は触らない（**人の指定が勝つ**）。

    返り値: (0 起点の列番号のリスト, 写す元の行番号) / 引き継ぐものが無ければ ([], 0)
    """
    path = (book_meta or {}).get("path")
    if not path:
        return [], 0
    totals = set(total_rows_in(book_meta, sheet, header_row))
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, wide = data_extent(ws, header_row)
            rows = [r for r in range(header_row + 1, last + 1) if r not in totals]
            if not rows or wide < 1:
                return [], 0
            heads = [str(ws.cell(row=header_row, column=c).value or "")
                      for c in range(1, wide + 1)]
            cols = []
            for c in range(1, wide + 1):
                if heads[c - 1] in declared:
                    continue                       # ★ 人が指定した列は触らない
                if all(bv.cell_formula(r, c, sheet) is not None for r in rows):
                    cols.append(c - 1)
    except Exception:
        return [], 0                               # 読めない回は黙る（断定しない）
    if not cols:
        return [], 0
    # ★ 写す元は「新しい行のすぐ上のデータ行」。無ければ下から取る。
    above = [r for r in rows if r < at]
    if above:
        return cols, max(above)
    below = [r for r in rows if r >= at]
    if not below:
        return [], 0
    return cols, min(below) + 1                    # ★ 挿入で 1 行ずれた後の位置


def broken_identity_advisory(source_book, out_book, resolved: dict) -> list:
    """操作の前に成り立っていた**列どうしの等式**が崩れていたら、その 1 行（無ければ []）。

    ★★ 2026-08-31（Namakoo）:「金額が入れ替われば付随して関連するセルの内容も
      変えなければいけない。しかもそれが複数の内容に影響する場合はそれらも踏まえて」
      実測: 単価のセルを 2 つ入れ替えると、**直値の 金額（＝件数×単価）が取り残される**。
        件数 12 × 単価 7200 = 86,400 なのに 金額 57,600 のまま。
      それでも「頼まれた 2 セルだけ動いた」は真実なので **✓ が出ていた**。
    ★ 式なら再計算されるので起きない ── **直値で持っている派生列**だけの事故。
    ★ 直さない・**言う**（どう直すかは人が決める ── 参照のズレと同じ線）。
    ★ op を問わず**1 箇所**で見る（入れ替えに限らず、入力を変える操作すべてに効く）。
    """
    sheet = (resolved or {}).get("_target_sheet")
    if not source_book or not Path(source_book).exists() or not sheet:
        return []
    # ★ 見出し行は resolved から取る ── **呼び出し側の変数に頼らない**
    #   （最初は header_row / book_meta を引数で受けたが、5 つの呼び出し場所で
    #     名前が揃っておらず NameError を 2 回出した。渡すものは 1 つに減らす）。
    header_row = int((resolved or {}).get("_header_row") or 1)
    try:
        def _rows(path):
            with BookView(Path(path)) as bv:
                ws = bv.sheet(sheet)
                last, wide = data_extent(ws, header_row)
                heads = [ws.cell(row=header_row, column=c).value
                          for c in range(1, wide + 1)]
                body = [[bv.cell_value(r, c, sheet)
                          if bv.cell_formula(r, c, sheet) is not None
                          else ws.cell(row=r, column=c).value
                          for c in range(1, wide + 1)]
                         for r in range(header_row + 1, last + 1)]
                return heads, body
        heads_b, rows_b = _rows(source_book)
        _heads_a, rows_a = _rows(out_book)
    except Exception:
        return []                      # 読めない回は黙る（断定しない）
    # ★★ 2026-09-01（Namakoo が実演の練習で実測・「成り立っているのに警告が出る」）:
    #   「項目と件数を入れ替えて」で ⚠ が出た。だが 金額＝件数×単価 は**成り立ったまま**
    #   ── 崩れたのではなく、**列が動いた**だけ。等式は列の位置で持っているので、
    #   並びが変われば同じ等式が別の組に見え、「消えた」と誤検出する。
    #   ★ セル 2 つの入れ替え（幕 3）は見出しが動かないので、そちらは今までどおり鳴る。
    # ★★ 2026-09-02（自作 review・致命 2）: 初版は `if heads_b != _heads_a: return []` と
    #   **降りて**いた。だがそれは「列が動いた回」だけでなく、**見出しが変わる操作すべて**
    #   （列追加・列削除・見出しの変更）で検算を丸ごと止める。
    #   実測: 列を 1 本足すついでに直値の派生列を壊しても ⚠ が出なくなっていた ──
    #   **この関数が検出対象にしていた事故クラスを、最も普通の操作で握りつぶしていた。**
    #   ★ しかも同じ関数の docstring に「op を問わず 1 箇所で見る（入れ替えに限らず、
    #     入力を変える操作すべてに効く）」と書いてある ── 自己申告と実装が矛盾していた。
    #   ★ 直しは「降りる」ではなく **名前で対応づける**。位置でなく見出しで並べ直せば、
    #     並べ替えも列追加も列削除も同じ手で扱える（前後の両方に在る列だけを比べる）。
    #     ★ 同じ見出しが 2 本ある表では最初の 1 本を使う（どちらが対応するかは名前で
    #       決まらない ── 決められないものを決めない）。
    def _index_of(heads):
        out = {}
        for i, h in enumerate(heads):
            key = str(h or "")
            if key and key not in out:
                out[key] = i
        return out

    _ib, _ia = _index_of(heads_b), _index_of(_heads_a)
    _common = [h for h in _ib if h in _ia]
    if len(_common) < 3:
        return []                      # 3 列そろわなければ等式は立たない
    heads_b = list(_common)
    rows_b = [[r[_ib[h]] if _ib[h] < len(r) else None for h in _common] for r in rows_b]
    rows_a = [[r[_ia[h]] if _ia[h] < len(r) else None for h in _common] for r in rows_a]
    # ★★ 2026-09-02: 行が増える操作でも見る（増えた行を**含めて**等式が成り立つか）。
    #   ★ 鳴りすぎない: identities() は 3 つとも数が入っている行だけで判定するので、
    #     値を入れなかった新しい行は最初から無視される。鳴るのは
    #     「新しい行に数が揃っていて、しかも等式を満たさない」時だけ。
    #   ★ 呼び分けはここ 1 箇所 ── 4 つある呼び出し側は 1 行も変えない。
    _lost = (row_identity.broken_after_insert(rows_b, rows_a)
              if len(rows_a) > len(rows_b) else row_identity.broken(rows_b, rows_a))
    note = row_identity.describe(_lost, heads_b)
    return [f"⚠ {note}"] if note else []


def reference_drift_warning(book_meta: dict, sheet: str | None, *,
                             row_lo: int = 1, row_hi: int = 10 ** 7,
                             col_lo: int = 1, col_hi: int = 10 ** 4,
                             rewritten=None, unit: str = "行") -> str | None:
    """動かす区画を**外から**指している式を見つけ、1 行にして返す（無ければ None）。

    ★★ 2026-08-29（Namakoo の指摘 → 実測で裏取り）:
      並べ替えると、範囲の外から特定の 1 行を指している式は**追従しない**。
      実測: `=B3`（ラベルは「ぶどうの金額」）が、並べ替え後に みかん の 200 を指した。
      別シートからの `=売上!B2` も同じ（りんご 100 → ぶどう 300）。
      ★ **式は 1 文字も壊れていない**ので、値でも文字列でも捕まらない ── 参照を読むしかない。
      ★ そして ailine は ✓ を出していた（並べ替え自体は宣言どおりだから）。
        「静かに壊れて合格が出る」── この製品が一番嫌う形に、ぴったり当てはまっていた。
    ★ 直さない: Excel も LibreOffice も、範囲の外から特定の行を指す式は並べ替えで
      追従させない（アドレスに留まるのが既定の意味）。
      「ぶどうの金額 = B3」は行に追従してほしいが「3行目の値 = B3」は留まってほしい
      ── 機械には区別できない。**名指しして人に返す**（補正は人が決めてから）。
    ★ 範囲（SUM(B2:B4)）は鳴らさない ── そちらは領域を指すので正しく追従する。
    """
    path = book_meta.get("path")
    if not path or not sheet:
        return None
    try:
        hits = cellmap.refs_pointing_into(Path(path), sheet, row_lo, row_hi, col_lo, col_hi)
    except Exception:
        return None                      # 読めない回は黙る（断定しない）
    # ★ 2026-08-31: **こちらで書き直す式**は、もう「ずれる式」ではないので名指ししない
    #   （別シートから指している式は書き直していないので、そちらは残す）。
    return cellmap.reference_drift_note(
        cellmap.drop_rewritten(hits, rewritten, sheet), unit=unit)


def placements_in_plan(plan) -> dict:
    """計画の中で**位置を作る**段（挿入・追加）を軸ごとに数える。

    ★★ 2026-08-29（Namakoo の設計判断）:
      「行や列を 2 つ以上増やす操作はもともと無いから縛っていい。
        複数行を増やす場合は for 文で順次増やせばいい」
    ★ これは**座標の法則の形とそのまま一致する** ── 1 つの操作 = 1 つの写像 π。
      写像は合成しない、と決めれば「1 回の依頼で 2 本増える」は構造的に起きない。
    ★ 実測（84 件の効果検体・3 表で同じ形）: 「味噌汁の上に新品を入れて」で一段目が
        [INSERT_ROWS at:2（空行）, ADD_ROW at:2（値つき）]
      を返し、両方走って**行が 2 本**増えた。同じ仕事を二重に言っている。
    ★ 数えるのは op 名ではなく**宣言**（新しい op が増えても、宣言さえ書けば数に入る）。
    """
    rows = cols = 0
    for st in plan or []:
        op_ = (st or {}).get("op")
        if not _op_writes(op_, WRITE_ROW_SHIFT):
            continue
        if _op_writes(op_, WRITE_NEW_COLUMN):
            cols += 1
        else:
            rows += 1
    return {"row": rows, "col": cols}


def plan_only_inserts_a_bare_row(plan) -> bool:
    """その計画が「**空行を挿すだけ**」か（値も列も足さない）。

    ★ op 名で数えない ── 宣言（writes）で見る。
      ・行をずらす（WRITE_ROW_SHIFT）
      ・でも新しい行の値は書かない（WRITE_NEW_ROW_AT_END が無い）
      ・列も足さない（WRITE_NEW_COLUMN が無い）
    ★ この形の計画にだけ「値が決まらないなら止まる」を掛ける ── それ以外の計画
      （値を書く・列を足す・行を消す）は、値が無いのが普通なので巻き込まない。
    """
    if not plan or len(plan) != 1:
        return False
    op_ = (plan[0] or {}).get("op")
    return (_op_writes(op_, WRITE_ROW_SHIFT)
            and not _op_writes(op_, WRITE_NEW_ROW_AT_END)
            and not _op_writes(op_, WRITE_NEW_COLUMN))


# ★★ 機械が**依頼文から取り直す**引数（A' 原則で LLM の値を採らないもの）。
#   ★ 2026-08-30（Namakoo「特定条件の行や列の抜き出しができない」）:
#     「丸和物流とみどり建設を抽出して」で、一段目は**値ごとに 1 段ずつ**返した
#       [EXTRACT value:丸和物流, EXTRACT value:みどり建設]
#     機械は各段で値を依頼文から取り直すので、解決後は**同じ抽出が 2 段**になり、
#     2 段目が連鎖の規則で 1 段目の出力を食って落ちた（人は 1 回しか頼んでいない）。
#   ★ だから「同じ仕事か」を見る時は、**機械が取り直す引数を外してから**比べる。
MACHINE_DERIVED_ARGS = {
    "EXTRACT": ("value", "values"),
}


# ★ 一段目が「分かりません」と降りたことを表す op。読み直しが横取りしてよい相手。
GIVING_UP_OPS = ("CLARIFY", "FREEFORM", "OUT_OF_VOCAB")


def plan_is_all_giving_up(plan) -> bool:
    """その計画が**全段とも降りている**か（1 段でも仕事をしているなら False）。

    ★★ 2026-08-30（Namakoo「読み直しって具体的には何をしてるの？」から辿って判明）:
      読み直しは `len(plan) == 1` のときしか発火しなかった。ところが実測では、
      一段目が**2 段の「内容不明」計画**を返すことがある:
          「商品・売上の2列だけ取り出して」
            → 1. 内容不明の依頼 / 2. 内容不明の依頼   ← 読み直しが鳴らない
      ★ 番人は在るのに、**失敗が取る形では発火しない** ──「在っても鳴らない」の再演。
    ★ 横取りしてよいのは**降りている計画だけ**、という元の線はそのまま守る
      （1 段でも実行できる段が在れば触らない）。段数だけを緩める。
    """
    steps = list(plan or [])
    return bool(steps) and all(
        str((st or {}).get("op")) in GIVING_UP_OPS for st in steps)


def fold_identical_steps(plan) -> tuple:
    """**中身がまったく同じ段**を 1 回にまとめる。戻り値: (畳んだ計画, 落とした数)。

    ★★ 2026-08-30（Namakoo「特定条件の行や列の抜き出しができない」）:
      「丸和物流とみどり建設を抽出して」で、一段目が**同じ抽出を 2 段**返した。
      2 段目は連鎖の規則で 1 段目の出力を対象にし、そこから同じ条件で抽出して
      「元シートが 8 セル変更されています」で落ちた ── 人は 1 回しか頼んでいない。
    ★ 同じ op・同じ引数の段は「合成」ではなく**同じ仕事の二重宣言**（行を 2 回
      足そうとする形と同じ・too_many_placements と同じ系譜）。
    ★ 断らずに畳む ── 依頼そのものは曖昧でない。ただし**黙って畳まない**
      （何段落としたかを呼び出し側が言う）。
    """
    out, seen, dropped = [], set(), 0
    for st in plan or []:
        op_ = str((st or {}).get("op"))
        # ★★ 畳むのは**新しいシートを作る段**だけ（宣言で絞る）。
        #   落ちたのはその形だけ ── 2 段目が連鎖の規則で 1 段目の出力を食う。
        #   同じ並べ替えを 2 回のような段は無害なので触らない（実測で、既存の
        #   検体「もう一度金額で降順に並べ替えて」を壊しかけた）。
        if not _op_writes(op_, WRITE_NEW_SHEET):
            out.append(st)
            continue
        args = dict((st or {}).get("args") or {})
        for k in MACHINE_DERIVED_ARGS.get(op_, ()):
            args.pop(k, None)
        key = (op_, repr(sorted(args.items(), key=lambda kv: str(kv[0]))))
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        out.append(st)
    return out, dropped


def too_many_placements(plan) -> str | None:
    """同じ軸に位置を作る段が 2 つ以上あるなら、その理由（無ければ None）。

    ★ 「追加してから並べ替えて」のような**別種の合成**は縛らない ── 縛るのは
      「同じ軸に 2 回place する」形だけ（それは 1 つの仕事の二重宣言）。
    """
    n = placements_in_plan(plan)
    for axis, label in (("row", "行"), ("col", "列")):
        if n[axis] > 1:
            return (f"1 回の依頼で{label}を {n[axis]} 回足そうとしています"
                     f"（{label}を増やすのは 1 回だけです ── "
                     f"複数必要なら、1 本ずつ頼んでください）")
    return None


def insert_rows_should_have_been_add_row(task: str, resolved: dict,
                                          book_meta: dict | None = None,
                                          sheet: str | None = None) -> str | None:
    """`INSERT_ROWS`（空行だけ）に読み取ったが、依頼文が**値も入れろ**と言っている形。

    ★ 2026-08-27（Namakoo が実測）:「みかんとぶどうの間に梨を追加して。売上は600 原価は300」
      が空行 1 本の挿入になった。op の取り違えで、位置の問題ではない。
    ★ 黙って op を書き換えない ── 証拠（依頼文に値の指定が在る）を示して、
      二段目翻訳（op を固定して args だけ埋め直す）へ回す判断材料にする。
    ★ 「言い換えてください」で終わらせない: 利用者の書き方は正しかった。
    """
    text = (task or "")
    # ★ 「空行が欲しい」と明示している依頼には触らない（誤爆させない）。
    if any(w in text for w in ("空行", "空白行", "行を挿入", "行を空け", "行間")):
        return None
    # ★ 2026-08-27（実測・俺の読み直しの誤爆）: 「みかんとぶどうの間に1行足して」を
    #   record の追加と誤解し、`商品=みかんとぶどう` という**値をでっち上げた**。
    #   ★ 足そうとしているのが「**行**」そのものなら、それは空行の挿入 ── 読み直さない。
    if _re_row_unit.search(text):
        return None
    if _re_value_assign.search(text):
        return "依頼文に入れる値の指定があります（行挿入は空行を挿すだけです）"
    # ★ 2026-08-27（2 度目の実測）: 初版は「値の代入」だけを証拠にしていたので、
    #   「みかんとぶどうの間に梨を追加して」（数字が無い）で発火しなかった。
    #   ★ **相対位置が実表で解けること自体**が強い証拠 ── 空行を「みかんとぶどうの
    #     間に」挿してくれ、という依頼は考えにくい。人は record を置く話をしている。
    # ★★ 2026-08-28（Namakoo が実測・3 度目）: 初版は「追加/足し/入れ」という
    #   **動詞の列挙**で発火させていた。「丸和物流と近江スチールの間に北斗精機を
    #   **作って**」が漏れて、空行が挿さった（一段目が CLARIFY を返す回もある）。
    #   ★ 列挙は必ず漏れる ── この repo で何度も踏んだ形。動詞を見るのをやめる。
    #   ★ 証拠は 2 つで足りる: ①相対位置が**実表で解ける** ②置く物の名前が依頼文に在る
    #     （②は呼び出し側が確かめる: 第二段が values を出し、その値が依頼文に literal で
    #      在ること ── 値をでっち上げた回に switch しない）。
    #   ★ 「空行が欲しい」「行を 1 本」は上で既に除いてある（別の意図）。
    at, note = resolve_row_anchor(text, book_meta or {}, sheet)
    if at is not None:
        return f"依頼文が場所を{note}と指しています（行挿入は空行を挿すだけです）"
    return None


_re_bare_number = re.compile(r"([0-9][0-9,]*(?:\.[0-9]+)?)\s*(?:に|へ|と)?\s*(?:し|する|して|に)")


def task_names_real_values(task: str, book_meta: dict, sheet: str | None,
                            col: str, header_row: int = 1) -> list:
    """依頼文が名指ししている、**その列に実在する値**を出現順に返す（重複は畳む）。

    ★ 2026-08-27（Namakoo「みかんの行とりんごの行だけを抽出して」）: 一段目は
      `contains "リンゴ"`（片仮名の幻覚）や `eq "みかんとりんご"`（連結）を返していた。
      どちらも 0 行に当たり、**空の抽出結果が ✓ で出る**ところだった。
    ★ 値は LLM に作らせない ── 表に在る値のうち、依頼文に現れるものだけを拾う。
    ★ 部分文字列で取りこぼさない/拾いすぎないために、**長い値から**当てて位置を潰す
      （「青りんご」と「りんご」が両方在る表で、短い方だけが当たるのを防ぐ）。
    """
    path = book_meta.get("path")
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    if not path or col not in headers:
        return []
    ci = headers.index(col) + 1
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, _c = data_extent(ws, header_row)
            values = []
            for r in range(header_row + 1, last + 1):
                v = ws.cell(row=r, column=ci).value
                s = str(v).strip() if v is not None else ""
                if s and s not in values:
                    values.append(s)
    except Exception:
        return []
    text = task or ""
    found = []
    for v in sorted(values, key=len, reverse=True):
        i = text.find(v)
        if i >= 0:
            found.append((i, v))
            text = text[:i] + (chr(0) * len(v)) + text[i + len(v):]
    return [v for _i, v in sorted(found)]


# ★ 「『商品』セル」「商品の見出しだけ」── **1 セル**を指す言い回し。
#   ★ 実測（2026-08-27）: これらは全部 `col:商品`（列ぜんぶ）に化けていた。
#     頼んでいない範囲に静かに広がる形で、この道具が最も嫌うもの。
#   ★ 2026-08-27（既存の検体が捕まえた・自分で開けた穴）: 「見出しを太字にして」まで
#     1 セルと読んでいた。それは**見出し行ぜんぶ**の意味でもありうる ── 曖昧。
#     ★ 「セル」と書いてあるか、「見出し**だけ**」のように限定が付いた時だけ 1 セルと読む。
#       曖昧なものを勝手に狭めない（広げないのと同じくらい大事）。
_re_one_cell_ask = re.compile(
    r"(?:セル[^。]{0,6}?(?:色|塗|太字|ボールド|強調|中央|センタ)"
    r"|見出し[^。]{0,4}?(?:だけ|のみ)[^。]{0,6}?(?:色|塗|太字|ボールド|強調|中央|センタ))")


def task_asks_for_one_cell(task: str) -> bool:
    return bool(_re_one_cell_ask.search(task or ""))


# ★ 「飾りの種類」だけは語で決まる（色/太字/中央）── これは**言い回しの揺れ**ではなく
#   操作そのものの名前なので、一覧（OP_META の synonyms）と同じ語彙を機械が読む。
_FORMAT_OP_WORDS = (("FILL_COLOR", ("色", "塗り", "塗っ", "ハイライト")),
                     ("BOLD", ("太字", "ボールド", "強調")),
                     ("CENTER_ALIGN", ("中央", "センタリング", "真ん中")))


def format_op_from_task(task: str) -> str | None:
    """依頼文から飾りの op を機械抽出する（見つからなければ None・断定しない）。"""
    best = None
    for op, words in _FORMAT_OP_WORDS:
        for w in words:
            i = (task or "").find(w)
            if i >= 0 and (best is None or i < best[0]):
                best = (i, op)
    return best[1] if best else None


def resolve_named_cell(book_meta: dict, sheet: str | None, name: str,
                        header_row: int = 1) -> tuple:
    """名前 → その値が入っている**1 つのセル**の (行, 列)（1 起点）。
       決められなければ (None, None, 断りの文)。

    ★ 見つからない・複数ある時は**決めない** ── 推測で別のセルを塗るのが一番こわい。
    ★ 探す範囲は**物理の使用範囲**（走査が最初の空で止まる穴を避ける・今週 3 度直した形）。
    ★ 見出し行も対象に含める（「『商品』セル」はふつう見出しを指す）。
    """
    path = book_meta.get("path")
    if not path or not name:
        return None, None, "表を読めないため、どのセルかを決められません"
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, cols = data_extent(ws, header_row)
            hits = [(r, c) for r in range(header_row, last + 1)
                     for c in range(1, cols + 1)
                     if str(ws.cell(row=r, column=c).value or "").strip() == name]
    except Exception as e:
        return None, None, f"表を読めませんでした（{type(e).__name__}）"
    if not hits:
        return None, None, f"『{name}』というセルが見つかりません"
    if len(hits) > 1:
        where = "、".join(f"{r}行{c}列" for r, c in hits[:4])
        return None, None, (f"『{name}』が {len(hits)} 箇所あります（{where}）"
                             " ── どれか決められません")
    return hits[0][0], hits[0][1], f"『{name}』＝{hits[0][0]}行{hits[0][1]}列"


# ★ 「7 行目」「7行」「第7行」── 人は行を**番号**でも指す。
_re_row_number_in_task = re.compile(r"(?:第)?\s*([0-9０-９]{1,4})\s*行(?:目)?")


def task_names_a_row_number(task: str) -> int | None:
    """依頼文が指している行番号（1 起点）。無い/複数あって決まらないなら None。"""
    nums = {int(m.translate(_ZENKAKU_DIGITS))
             for m in _re_row_number_in_task.findall(task or "")}
    return nums.pop() if len(nums) == 1 else None


_re_quoted_value = re.compile(r"[「『\"“]([^」』\"”]{1,40})[」』\"”]")


def task_quotes_a_value(task: str) -> str | None:
    """依頼文が**書き込む値を引用符で名指し**しているか（A' 原則: 値は依頼文から取る）。

    ★ 2026-08-28: 「1 セルのつもりの依頼」を見分ける三項目のうちの 1 つ。
      行を指し・列を指し・**値を引用している**なら、それは 1 セルへの書き込み。
      引用が無い依頼（「利益を計算して」）を横取りしないための項でもある。"""
    m = _re_quoted_value.search(task or "")
    return m.group(1) if m else None


def plan_writes_beyond_one_cell(plan) -> bool:
    """この計画は**1 セルより広く**書くと宣言しているか。

    ★★ op 名を数え上げない ── 今日 3 度目に破れた「除外の列挙」を繰り返さないため、
      各 op が OP_WRITE_TARGET に自分で書いている「書く領域」を読む。
      新しい op が増えても、宣言さえ書けばこの門は自動で効く。"""
    wide = {WRITE_EXISTING_COLUMN, WRITE_NEW_COLUMN, WRITE_NEW_ROW_AT_END, WRITE_ROW_SHIFT}
    for st in plan or []:
        wt = OP_WRITE_TARGET.get((st or {}).get("op"))
        # ★ 「1 セルだけ書く」と自分で宣言している op は、既に落ちている（読み直さない）。
        if wt and (set(wt.writes) & wide) and WRITE_SINGLE_CELL not in wt.writes:
            return True
    return False


def task_points_at_one_row(task: str, book_meta: dict, sheet: str | None) -> str | None:
    """依頼文が**1 行を名指ししている**なら、その根拠の文。していなければ None。

    ★★ 2026-08-28（Namakoo が請求書のデモで実測・今日いちばん悪い形）:
      「7行目の担当を『佐藤』にして」で **担当列が全行『佐藤』になり、✓ が出た**。
      一括書換（列ぜんぶ）の契約としては ✓ は正しい ── だが**依頼は 1 行**だった。
      ★ 判定に三項が要る（依頼・宣言・実体）のに、機械は宣言と実体しか見ていなかった。
    ★ ここは「依頼が行を指しているか」だけを見る（どの行かは別で解く）。
      指しているのに 1 セルへ落とせなかったら、**列全体を書かずに断る**。
    """
    if task_names_a_row_number(task):
        return f"依頼文が『{task_names_a_row_number(task)}行目』と行を指しています"
    if _task_names_a_row(task, book_meta, sheet):
        return f"依頼文が『{_task_names_a_row(task, book_meta, sheet)}』の行を指しています"
    # ★ 名前が**複数行に在る**時も「行を指している」── 決められないだけで、指してはいる。
    #   ここを見落とすと、同名が 2 行あるだけで列全体が潰れる（実測した形）。
    path = book_meta.get("path")
    if not path or not task:
        return None
    hr = int((book_meta.get("header_rows") or {}).get(sheet, 1) or 1)
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, _c = data_extent(ws, hr)
            for r in range(hr + 1, last + 1):
                v = str(ws.cell(row=r, column=1).value or "").strip()
                if v and v in task:
                    return f"依頼文が『{v}』（表に複数あります）を指しています"
    except Exception:
        return None
    return None


def _task_names_a_row(task: str, book_meta: dict, sheet: str | None) -> str | None:
    """依頼文が**表に実在する行の名前**を指しているか。指していればその名前。

    ★ 機械が実表と突き合わせて確かめる ── LLM に「これは 1 セルの話か」を判断させない。
    ★ 1 つに決まらない（同名が 2 行）なら None（推測で書かない）。
    """
    path = book_meta.get("path")
    if not path or not task:
        return None
    hr = int((book_meta.get("header_rows") or {}).get(sheet, 1) or 1)
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, last_col = data_extent(ws, hr)
            names = {}
            for r in range(hr + 1, last + 1):
                v = str(ws.cell(row=r, column=1).value or "").strip()
                if v:
                    names[v] = names.get(v, 0) + 1
    except Exception:
        return None
    hit = [n for n, cnt in names.items() if cnt == 1 and n in task]
    # ★ 一番長い一致を採る（「梨」と「洋梨」が両方在る時に短い方を拾わない）
    return max(hit, key=len) if hit else None


def _resolve_named_row(book_meta: dict, sheet: str | None, name: str) -> tuple:
    """行の名前 → 行番号（1 起点）。決められなければ (None, 断りの文)。

    ★ 2026-08-27: 住所の解決はここ 1 箇所に集める（resolve_row_anchor もこれを使う形へ
      寄せていく）。★ 探す範囲は**物理の使用範囲**（走査が最初の空で止まる穴を避ける）。
    ★ 見つからない・複数ある時は**決めない** ── 推測で別の行に書くのが一番こわい。
    """
    path = book_meta.get("path")
    if not path:
        return None, "表を読めないため、どの行かを決められません"
    hr = int((book_meta.get("header_rows") or {}).get(sheet, 1) or 1)
    # ★★ 2026-08-31（Namakoo が実測・「この基本操作ができない」）:
    #   「6行目と5行目を入れ替えて」が CLARIFY に落ちていた ── **行番号で指すと黙る**、
    #   08-29 に追加・削除で直したのと同じ非対称が、**入れ替えには残っていた**。
    #   ★ ここは住所の解決を集めている 1 箇所なので、ここに足すと全部の op に効く
    #     （入れ替え専用の判定を作らない）。
    if _re_row_number_word.fullmatch(str(name or "").strip()):
        _n = _row_word_number(name)
        if _n > hr:
            return _n, f"{_n}行目（依頼文の行番号）"
        return None, f"{_n}行目は見出し行（{hr}行目）またはその上です"
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, last_col = data_extent(ws, hr)
            hits = [r for r in range(hr + 1, last + 1)
                     if any(str(ws.cell(row=r, column=c).value or "").strip() == name
                             for c in range(1, last_col + 1))]
    except Exception as e:
        return None, f"表を読めませんでした（{type(e).__name__}）"
    if not hits:
        return None, f"『{name}』という行が見つかりません"
    if len(hits) > 1:
        return None, (f"『{name}』が {len(hits)} 行あります"
                       f"（{'、'.join(str(h) for h in hits)}行目）── どれか決められません")
    return hits[0], f"『{name}』の行＝{hits[0]}行目"


def _table_rows_for_anchor(book_meta: dict, sheet, header_row: int) -> tuple:
    """位置解決のために実表を読む（行番号 → 値の並び、と見出しの並び）。読めなければ空。"""
    path = book_meta.get("path")
    if not path:
        return {}, []
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, last_col = data_extent(ws, header_row)
            rows = {r: [str(ws.cell(row=r, column=c).value or "").strip()
                         for c in range(1, last_col + 1)]
                     for r in range(header_row + 1, last + 1)}
            heads = [str(ws.cell(row=header_row, column=c).value or "").strip()
                      for c in range(1, last_col + 1)]
            return rows, heads
    except Exception:
        return {}, []


def _row_named_anywhere_in_task(task: str, rows: dict, headers: list,
                                 require_possessive: bool = False):
    """依頼文に literal で現れる**実在の値**が、ちょうど 1 行にしか無いならその行。

    ★★ 2026-08-28（Namakoo「行の削除もできない」）: 「ナットを削除して」のように、
      人は「〜の行」と言わないことがある。言い回しを足すのではなく**表に訊く**。
    ★ 見出しの語は除く（列名を行の名前と読み違えない）。
    ★ 2 行に当たったら決めない（推測で別の行を消すのが一番こわい）。
    ★★ 2026-08-30（Namakoo「セル指定しているのに値を上書きできない」）:
      「7行B列を『{{合計:税込金額}}』に上書き」で、**引用符の中の『合計』**が表の
      合計行に当たり、そこを狙った操作に読み替えられていた。
      ★ 引用符の中は**値**であって、対象の名指しではない ── ここが 4 つの呼び出しの
        合流点なので、**この 1 行**で全部に効く（呼び出し側に配らない）。
    """
    text = _task_outside_quotes(task)
    heads = {h for h in headers if h}
    best = None
    for r, vals in (rows or {}).items():
        for v in vals:
            # ★★ 2026-08-31（Namakoo「LLM の揺れが一番厄介だ」→ 追ったら半分は機械の責任）:
            #   「金額が**60000**以上の行を抜き出して」で、機械が『60000』を**行の名前**
            #   として解き（金額列に 60000 が在る）、「『60000』の行＝3行目」と確信して
            #   行追加に読み替えていた。★ LLM が揺れた回に、**機械がその揺れを
            #   『確信をもって間違った操作』に育てていた**。
            #   ★ 揺れは消せないが、**増幅しないことはできる** ── 依頼文に出る数は
            #     ほぼ常に閾値や個数で、行の名前ではない。
            #   ★ 判定は既にある `_is_number_like`（「依頼文に出る数と、行の名前を
            #     混同しないため」）を借りる ── 1 箇所でしか使われていなかった。
            if not v or v in heads or len(v) < 2 or v not in text or _is_number_like(v):
                continue
            # ★★ 2026-08-29（Namakoo が実測）: 「丸山工業の担当に『佐藤』を入れて」で
            #   **書き込む値『佐藤』**が別の行の担当欄にも在るため、行の候補が 2 つに
            #   なって「決められない」に落ちていた ── 値を行の名前と読んでいた。
            #   ★ 人が行を指すときは「**〜の**」と言う。セルを指す経路ではそれを要求する
            #     （「ナットを削除して」のように の が無い経路は今までどおり）。
            if require_possessive and f"{v}の" not in text:
                continue
            if best is None:
                best = (r, v)
            elif best[0] != r:
                return None          # 2 行以上に当たる ── 決めない
            elif len(v) > len(best[1]):
                best = (r, v)
    return best


def resolve_row_anchor(task: str, book_meta: dict, sheet: str | None,
                        header_row: int = 1) -> tuple:
    """依頼文の「**みかんの下に**」「**みかんとぶどうの間に**」から行番号を決める。

    ★ 2026-08-27（Namakoo が実測）: ADD_ROW は位置を**行番号**でしか受け取れないのに、
      人は相対で言う。LLM に数えさせると外し、空行だけの INSERT_ROWS に落ちていた。
    ★ 分担を変える: **LLM は「誰の隣か」を言うだけ／行番号は機械が実表を数えて決める**
      （列名の解決を機械 3 段でやっているのと同じ形）。
    ★ 見つからない・複数ある時は**決めない**（推測で行を挿すと、静かに別の場所へ入る）。
    戻り値: (行番号 or None, 説明 or 断りの理由 or None)
    """
    # ★★ 2026-08-30（Namakoo「セル指定しているのに値を上書きできない」）:
    #   「7行B列を『{{合計:税込金額}}』に上書き」で、**引用符の中の『合計』**を位置の
    #   目印として拾い、『合計』の行＝9行目 と解いていた。そのせいで一段目が行の挿入を
    #   返した回に「行追加として読み直しました」が発火し、頼んでいない行が挿さりかけた。
    #   ★ 列では既に塞いだ穴（_task_names_single_real_column）が、行では開いていた
    #     ── **行と列の非対称**、この repo が何度も踏んだ形。
    #   ★ 引用符の中は**値**であって、対象の名指しではない（Namakoo の決めた約束）。
    #     だから位置を探す時は引用符の中を見ない ── 「『みかん』の行を削除して」の
    #     ように名前を引用する書き方は、引用符なしで書いてもらう（列と同じ扱い）。
    text = _task_outside_quotes(task).replace("　", " ")
    want_after, name, second = None, None, None
    m = _re_between.search(text)
    if m:
        want_after, name, second = True, m.group(1).strip(), m.group(2).strip()
    else:
        for suf in _ANCHOR_AFTER:
            m = _re_anchor(suf).search(text)
            if m:
                want_after, name = True, m.group(1).strip()
                break
        if name is None:
            for suf in _ANCHOR_BEFORE:
                m = _re_anchor(suf).search(text)
                if m:
                    want_after, name = False, m.group(1).strip()
                    break
    if not name:
        # ★ 2026-08-27（実測）:「りんごの行を削除して」── 人は行を**中身**で指す。
        #   相対の言い回しが無くても、「<X>の行」なら X を実表で探す。
        m2 = _re_row_of.search(text)
        if m2:
            want_after, name = None, m2.group(1).strip()
    # ★ 2026-08-28: 言い回しが 1 つも当たらない回も、**表に訊いてから**諦める。
    if not name:
        rows_h, heads_h = _table_rows_for_anchor(book_meta, sheet, header_row)
        alt = _row_named_anywhere_in_task(task, rows_h, heads_h)
        if alt:
            return alt[0], f"『{alt[1]}』の行＝{alt[0]}行目"
        return None, None
    # ★ 2026-08-27（自分で入れた誤爆・既存の検体が捕まえた）:
    #   「**2行目の前に**1行挿入して」の「2行目」を中身の名前として探し、
    #   見つからず断っていた。**行番号は名前ではない** ── 表を探しに行かない。
    # ★★ 2026-08-29（Namakoo が実測・「行の追加が出来なくなってる」）:
    #   そのとき「探さない」を「**決めない**」と書いてしまった。結果:
    #     「ヤマノ食品の下に丸山工業の行を作って」→ 5行目・値も入る（✓）
    #     「4行目の下に丸山工業の行を作って」  → 機械が黙る → LLM の 4 がそのまま通り、
    #                                            **上に空行**が挿さった（✗）
    #   ★ 同じ「下に」なのに、**指し方が名前か番号かで結果が変わっていた**。
    #     ここは表に訊く必要すらない ── 番号と向きが揃っているのだから**引き算で出る**。
    #   ★ 位置が出れば、`insert_rows_should_have_been_add_row` の証拠①も立つので、
    #     値つきの行（ADD_ROW）へ回る ── 空行に落ちる道が同時に塞がる。
    if _re_row_number_word.fullmatch(name):
        at, _n, note = row_number_anchor(task)
        return (at, note) if at is not None else (None, None)
    path = book_meta.get("path")
    if not path:
        return None, None
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            # ★★ 2026-08-27（Namakoo が実測・俺が新しい所で開けた同じ穴）:
            #   `_scan_last_row` は 1 列目を上から見て**最初の空で止まる**。
            #   下書きに空行が 1 本あると、その下の「みかん」を探せず、位置解決が黙って
            #   失敗して LLM の行番号がそのまま通っていた。
            #   ★ **探す範囲は物理の使用範囲から取る**（今週この repo が 3 度直した形）。
            last, last_col = data_extent(ws, header_row)
            ws_rows = {r: [str(ws.cell(row=r, column=c).value or "").strip()
                            for c in range(1, last_col + 1)]
                        for r in range(header_row + 1, last + 1)}
            headers_here = [str(ws.cell(row=header_row, column=c).value or "").strip()
                             for c in range(1, last_col + 1)]
            hits = [r for r, vals in ws_rows.items() if name in vals]
    except Exception:
        return None, None
    if not hits:
        # ★★ 2026-08-28（Namakoo「行の削除もできない」）: 「ナット**を**削除して」が
        #   『1行目は見出し行です』で断られていた。人は「〜の行」と言わないこともある。
        #   ★ 言い回しを足すのではなく、**表に訊く**: 依頼文に literal で現れる値が
        #     この表のちょうど 1 行にしか無いなら、それがその行。
        #     （列名も見出しも除く ── 「数量が100未満の行」のような条件文は当たらない）
        alt = _row_named_anywhere_in_task(task, ws_rows, headers_here)
        if alt:
            return alt[0], f"『{alt[1]}』の行＝{alt[0]}行目"
        # ★★ 2026-08-31（通しの 1 幕目で全滅した形）:「8行目に丸山工業の行を作って」
        #   ── **これから置く**行なので、名前が表に無いのは当たり前。
        #   ★ 依頼文が行番号を名指ししているなら、それが場所（表に無いことは断りの
        #     理由にならない）。実測では task_names_a_row_number が 8 を返せていたのに、
        #     名前が見つからないほうで先に断っていた。
        _n_here = task_names_a_row_number(task)
        if _n_here and _n_here > header_row:
            return _n_here, f"{_n_here}行目（依頼文の行番号）"
        return None, (f"『{name}』という行が見つかりません"
                       "（この表に在る値で指してください・行番号でも指せます）")
    if len(hits) > 1:
        # ★★ 2026-08-29（Namakoo）:「どうしても中身でさせない場面が出てくる。例えば
        #   4行目と5行目は両方ともヤマノ食品。取引先で指定は出来ない」── そのとおりで、
        #   ここは**断って終わる場所ではなく、行番号の道へ渡す場所**。
        #   ★ 候補の行番号は機械がもう知っている ── そのまま言う（人に数え直させない）。
        _rows = "、".join(str(h) for h in hits)
        _ex = f"{hits[0]}行目"
        return None, (f"『{name}』が {len(hits)} 行あります（{_rows}行目） ── どれか決められません。"
                       f"行番号で指してください（例:「{_ex}の下に…」「{_ex}を削除して」）")
    row = hits[0]
    if want_after is None:          # 「<X>の行」＝ その行そのもの
        return row, f"『{name}』の行＝{row}行目"
    at = row + 1 if want_after else row
    where = "下" if want_after else "上"
    note = f"『{name}』（{row}行目）の{where}＝{at}行目"
    if second:
        note += f"（『{second}』との間）"
    return at, note


# ★ 2026-08-27: 列の相対位置。行（_ANCHOR_AFTER/_BEFORE）と**同じ形**で持つ ──
#   「位置は op に依らず位置」なので、片方だけ賢くしない。
_COL_AFTER = ("の右に", "の右へ", "の右側に", "の後ろに", "のうしろに", "の次に")
_COL_BEFORE = ("の左に", "の左へ", "の左側に", "の前に", "の手前に")
# 「原価と売上の右側に」＝ 2 つのうち右の方の隣（Namakoo が挙げた実例）。
_re_col_pair = re.compile(r"([^\s、。]+?)\s*と\s*([^\s、。]+?)\s*の\s*(右|左)")
# 依頼文が「列を追加/足す/挿入」と言っているか（第二段へ回すための証拠）。
# ★ 2026-08-27（自分で開けた穴・実機の検体が捕まえた）: 「列を**入れ替え**て」が
#   「入れ」に当たって列追加として横取りされ、入れ替えが動かなくなった。
#   ★ 語の一部が別の語の一部でありうる ── 部分文字列の穴は、この repo で 2 度目。
# ★ 2026-08-27（Namakoo「◎を入れて では動作しない」）: 「列**に**『◎』を入れて」まで
#   列追加として拾っていた。★ 列の直後の助詞で分かれる ── 「列**を**追加/入れる」は
#   列そのものが対象、「列**に**…を入れる」は列が**行き先**。助詞は意味を運んでいる。
_re_add_col_ask = re.compile(r"列\s*(?:を|の)\s*[^。]{0,4}?(?:追加|足し|足す|挿入|入れ(?!替)|作)")


# ★ 「列を追加して」という依頼に対して、一段目の答えを**そのまま残してよい** op。
#   中身のある列を作る op だけ ── それ以外は軸か操作が違う（実測で INSERT_ROWS・
#   SPLIT_CELL が返ってきた）。除外を数え上げると、返しうる op が増えるたび穴が開く。
KEEP_FOR_COLUMN_REQUEST = ("COMPUTE_COLUMN", "LOOKUP_FILL")


# ★ 条件つきの書き込みを求めているか（比較語があり、かつ書き込む値が引用されている）。
#   ★ 2 つとも要る: 比較語だけなら EXTRACT（抜き出す）かもしれないし、引用だけなら
#     SET_COLUMN_VALUE（列を丸ごと）。**両方揃った時だけ**この op を疑う。
# ★ 「列だけ抜き出す」── 行の抽出（EXTRACT）と区別するのは**「列」という語**だけ。
#   「〜の列だけ」「必要な列だけ」。抽出の動詞（抜き出す/抽出/取り出す）は EXTRACT と共通。
_re_extract_cols_ask = re.compile(r"列[^。]{0,8}?(?:だけ|のみ)[^。]{0,8}?(?:抜き出|抽出|取り出|残)")


_re_extract_ask = re.compile(r"(?:抜き出|抽出|取り出)")


def resolve_named_extraction(book_meta: dict, sheet: str | None, task: str,
                              header_row: int = 1) -> tuple:
    """依頼文が「どの列の、どの値の行を」抜き出したいのかを実表から決める。
       決まらなければ (None, None)。

    ★ 2026-08-27（実測）: 同じ依頼文で、一段目が EXTRACT を返す回と OUT_OF_VOCAB に
      落ちる回があった（2/3 と 1/3）。落ちた回は「もしかして: 抽出？」の確認に回り、
      **聞かれる回と聞かれない回が偶然で決まる**形になっていた。
      ★ 機械が列も値も解けているなら、迷う理由が無い（実表を見た側が確かなことを知って
        いる）── 他の読み直しと同じ線を引く。
    ★ 列は「名指しされた値が一番多く在る列」。同数で並んだら**決めない**（推測しない）。
    """
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    best = []
    for col in headers:
        vals = task_names_real_values(task, book_meta, sheet, col, header_row)
        if vals:
            best.append((len(vals), col, vals))
    if not best:
        return None, None
    best.sort(key=lambda x: -x[0])
    if len(best) > 1 and best[0][0] == best[1][0]:
        return None, None
    return best[0][1], best[0][2]


def task_asks_to_extract_columns(task: str) -> bool:
    return bool(_re_extract_cols_ask.search(task or ""))


# ★ 置き換え「『A』を（全て）『B』に」の形。★ 動詞（書き換え/置換/直し…）は見ない ──
#   並べ始めると並べ忘れた言い方が黙って落ちる。**助詞が意味を運ぶ**（A を … B に）。
_re_replace_pair = re.compile(
    r"[「『]([^」』]+)[」』]\s*を\s*[^。]{0,8}?[「『]([^」』]+)[」』]\s*に")


def extract_replace_pair(task: str):
    """依頼文から「『A』を … 『B』に」の A, B を取る。無ければ None。
       ★ 引用が 2 つあると extract_quoted_literal は「一意に読み取れない」と断る（正しい）。
         断るだけで終えず、**2 つある時の意味**（置き換え）をここで読む。"""
    m = _re_replace_pair.search((task or "").replace("　", " "))
    if not m:
        return None
    src, dst = m.group(1).strip(), m.group(2).strip()
    return (src, dst) if src and dst and src != dst else None


def task_asks_for_a_replace(task: str) -> bool:
    return extract_replace_pair(task) is not None


def task_asks_for_a_conditional_write(task: str) -> bool:
    return bool(task and extract_cmp_from_task(task)
                 and extract_quoted_literal(task) is not None)


def task_asks_to_add_a_column(task: str) -> bool:
    """依頼文が「列を追加」を求めているか（位置も名前もここでは決めない）。"""
    return bool(_re_add_col_ask.search(task or ""))


def _header_index(headers: list, name: str) -> tuple:
    """列名 → (1 起点の位置, 実際の見出し名)。『原価列』のように「列」が付いた言い方も受ける。
       決まらなければ (None, name)（推測しない）。
       ★ 実際の見出し名も返すのは、解釈行に**表に在る名前**を出すため
         （『原価列』と書かれても『原価』と表示する ── 人が突き合わせられる形にする）。"""
    names = [str(h) for h in headers]
    if name in names:
        return names.index(name) + 1, name
    if name.endswith("列") and name[:-1] in names:
        return names.index(name[:-1]) + 1, name[:-1]
    return None, name


# ★ 条件の閾値を依頼文から取るための素材（EXTRACT の value を LLM に任せていたのと違い、
#   こちらは**依頼文の数字**しか受け取らない ── 印を付ける操作は範囲を外すと静かに広がる）。
_ZENKAKU_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")
_re_threshold_num = re.compile(r"\d+(?:\.\d+)?")


def total_rows_in(book_meta: dict, sheet: str | None, header_row: int = 1) -> list:
    """データ行ではない「合計行」の行番号（1 起点）。

    ★★ 2026-08-28（Namakoo が請求書のデモで実測）: 「金額が10万以上の行に印を付けて」が
      **合計行にも印を付けた**。条件としては真だが、合計行は請求の行ではない ── 意味が違う。
    ★ 判定は既存の凍結規則を借りる（ailine_core.total_row.row_has_total_word:
      合計/小計/総計は部分一致・『計』は完全一致・『設計部』等は誤爆しない断片ガードつき）。
      ここで新しい規則を書かない ── 同じことを 2 箇所が別々に決めると必ずずれる。
    ★ 見つけたら**必ず画面に出す**（黙って行を外さない）。
    """
    path = book_meta.get("path")
    if not path:
        return []
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, cols = data_extent(ws, header_row)
            out = []
            for r in range(header_row + 1, last + 1):
                vals = [ws.cell(row=r, column=c).value for c in range(1, cols + 1)]
                if total_row.row_has_total_word(vals):
                    out.append(r)
    except Exception:
        return []
    return out


def _rows_matching(book_meta: dict, sheet: str | None, cond_col: str, cmp: str,
                    threshold, header_row: int = 1):
    """条件に当てはまるデータ行（1 起点の行番号）を、実表を読んで数える。
       読めなければ None（分からないことを 0 と言わない）。

    ★ なぜ**適用前に**数えるか: 0 行なら書き込みは何も起こさず、事後条件は
      「変化なし」で落ちる。正しいが、利用者には「動かなかった」としか見えない
      （列追加で同じ形を実測した）。★ 走らせる前に、当てはまる行数を画面に出す。
    ★ 走査は**物理の使用範囲**から（1 列目の空で止まる罠を避ける）。
    ★ 合計行は**データ行ではない**ので外す（外したことは呼び側が画面に出す）。
    """
    path = book_meta.get("path")
    if not path:
        return None
    skip = set(total_rows_in(book_meta, sheet, header_row))
    headers = [str(h) for h in ((book_meta.get("headers") or {}).get(sheet) or [])]
    if cond_col not in headers:
        return None
    ci = headers.index(cond_col) + 1
    match = _extract_predicate(cmp, threshold)
    try:
        with BookView(Path(path)) as bv:
            ws = bv.sheet(sheet)
            last, _cols = data_extent(ws, header_row)
            return [r for r in range(header_row + 1, last + 1)
                     if r not in skip and match(bv.cell_value(r, ci, sheet))]
    except Exception:
        return None


# 「〜の右に」等の位置の言い回し（列版・_COL_AFTER/_COL_BEFORE と同じ語彙）。
_re_after_position = re.compile("(?:" + "|".join(
    re.escape(w) for w in (*_COL_AFTER, *_COL_BEFORE)) + ")")
# 名前のうしろに付く「列を追加して」等。★ 語尾は閉じた文法の集合（業務語彙ではない）。
# 修飾節の終わり ── 助詞、または活用語尾（引い**た** / 掛け**て** / 足し**た**）。
# ★ 「の」は入れない（「粗利の列」の の は名前の側）。★ 業務語彙ではなく閉じた文法。
_re_clause_end = re.compile(r"[をにへはがでとも]|[ぁ-ん](?:た|て|だ)")
# ★ 位置語が無い回の入口 ── 「**作る**」と言っている語尾だけ（裸の「列」は入れない）。
_NEW_COL_MAKE_TAILS = ("という列を追加して", "という列を作って", "という列を追加",
                        "という列を作る", "の列を追加して", "の列を作って",
                        "の列を追加", "の列を作る", "列を追加して", "列を作って",
                        "列を追加", "列を作る")
_NEW_COL_TAILS = ("という列を追加して", "という列を作って", "という列を追加", "という列を作る",
                   "という列", "の列を追加して", "の列を作って", "の列を追加", "の列を作る",
                   "列を追加して", "列を作って", "列を追加", "列を作る", "列",
                   "を追加して", "を作って", "を追加", "を作る", "を入れて", "を足して")


def new_column_name_from_task(task: str, headers=None, *,
                               require_position: bool = True) -> str | None:
    """依頼文が名指ししている、**新しい列の名前**（決まらなければ None）。

    ★★ 2026-08-30（Namakoo が実測・下書きに 2 本できた）:
      「金額の右に税込み金額を追加」を 2 回頼んで、見出しが
        1 回目「税込金額」（**「み」が落ちた**）／2 回目「金額*1.1」（**式が名前になった**）
      になった。前者は道具が `f"税込{列名}"` と**作った**名前、後者は式そのもの。
      ★ どちらも A' 原則（値も名前も依頼文から取る）が抜けていた ── **人が書いた
        名前がそこに在るのに、機械が別の名前を発明していた**。
      ★ しかも解釈行に名前が出ていなかったので、間違いに気づく手がかりが無かった。
    ★ 引き算で切り出す: 位置の言い回し（「〜の右に」）の**うしろ**から、語尾を落とす。
      全体を置換しない ── 「金額の右に税込み金額」で『金額』を全部消すと『税込み』になる。
    ★ 実在する見出しと同じ名前なら None（それは新しい列ではない）。
    """
    text = _task_outside_quotes(task).replace(chr(12288), " ")
    m = None
    for m2 in _re_after_position.finditer(text):
        m = m2                       # 最後の位置語のうしろを見る（「AとBの右に X」）
    if m:
        name = text[m.end():]
    elif require_position:
        # ★★ 2026-09-02: 既定は**位置語が在る時だけ**（従来どおり）。
        #   実測で分かったこと: 位置語なしを無条件に許すと、税の枝（W10c で設計）まで
        #   書き換わる ──「税込みの列を追加して」で見出しが『税込み』になった。
        #   『税込み』は名前ではなく**修飾語**で、機械が組む『税込金額』のほうが良い。
        #   ★ 測っていない所まで直しを広げない（断る範囲を広げるのと同じ失敗）。
        #     欠けていたのは**2 項の演算**の枝だけなので、そこだけ明示的に呼ぶ。
        return None
    else:
        # ★★ 2026-09-02（A の確認中に見つけた）: 位置を言わない依頼では、ここに
        #   入る前に空文字になっていた ── 「売上から原価を引いた**利益**の列を作って」で
        #   見出しが『売上-原価』（式そのもの）になっていた。A' 原則が抜けた形。
        #   ★ 位置語が無い時は、**語尾の手前まで**を候補にして、その中の
        #     **修飾節の終わり**から始める（引き算は位置語の時と同じ考え方）。
        #     節の終わり = 助詞（を に へ は が で と も）か、活用語尾（〜た/て/だ）。
        #     ★ 「の」は**入れない** ── 「粗利の列」の の は名前側に属する。
        #   ★★ 初版は語尾を先に切らずに走査したので、「作**って**」自身が節の終わりに
        #     当たり、名前ごと飲み込んで空になっていた（実測で捕まえた）。
        # ★★ 実測で捕まえた誤爆（既存の検体が赤くした）: 語尾に **裸の「列」** を
        #   許すと、「A行G列を『税込み金額』に上書き」で『A行G』を新しい列の名前として
        #   拾った。★ 位置語が無い回は、**作る**と言っている語尾だけを入口にする
        #   （裸の「列」は「〜の右に 利益列」のような位置語つきの回のためのもの）。
        _end = min((text.find(w) for w in _NEW_COL_MAKE_TAILS if text.find(w) > 0),
                    default=-1)
        if _end < 0:
            return None                  # 「〜の列を作って」の形 が無い＝名指しでない
        _head = text[:_end]
        _cut = 0
        for _mb in _re_clause_end.finditer(_head):
            _cut = _mb.end()
        name = _head[_cut:]
    if not name.strip():
        return None
    for w in _NEW_COL_TAILS:         # 長い語尾から落とす（並びが長さ順）
        i = name.find(w)
        if i > 0:
            name = name[:i]
            break
    name = name.strip().strip("、。 ")
    while name and name[0] in "をにへはがでとのも 　":
        name = name[1:]
    if len(name) < 2 or " " in name:
        return None
    if any(ch in name for ch in "をにへはがでとも"):
        return None                  # 文がまだ切れていない（助詞が残っている）
    if name in {str(h) for h in (headers or [])}:
        return None                  # 既にある列 ── 新しい名前ではない
    return name if name in text else None


def resolve_col_anchor(task: str, headers: list) -> tuple:
    """依頼文の「**原価の右に**」「**原価と売上の右側に**」から、新しい列が入る位置
       （1 起点）を決める。

    ★ 分担は行と同じ: **LLM は「誰の隣か」を言うだけ／位置は機械が実表の見出しから決める。**
    ★ 見つからない・決められない時は**決めない**（黙って末尾に付けない ── 静かに
      違う場所へ入るのが一番こわい、を列でも同じに扱う）。
    戻り値: (位置 or None, 説明 or 断りの理由 or None)。
            (None, None) = 位置の言い回しが**そもそも無い**（呼び側が末尾を選べる）
    """
    text = (task or "").replace("　", " ")
    names = [str(h) for h in headers]
    # ★★ 2026-08-29（84 件の効果検体で 3 表とも同じ形で落ちた）:
    #   「料理と主材料の**間に**区分の列を追加して」が解けず、黙って末尾に付いていた。
    #   行は `_re_between`（「AとBの間」）を持っているのに、列は「右／左」しか
    #   見ていなかった ── **行と列の非対称**。Namakoo が名指しした所そのもの。
    #   ★ 同じ正規表現を列にも通す（軸が違うだけで、位置の言い回しは同じ）。
    mb = _re_between.search(text)
    if mb:
        a, c = mb.group(1).strip(), mb.group(2).strip()
        ia, a = _header_index(names, a)
        ic, c = _header_index(names, c)
        if ia is not None and ic is not None:
            hi = max(ia, ic)
            return hi, f"『{a}』と『{c}』の間＝{hi}列目"
        # ★ 見出しに無いなら、それは列の話ではない（行の「間」かもしれない）── 触らない。
    m = _re_col_pair.search(text)
    if m:
        a, c, side = m.group(1).strip(), m.group(2).strip(), m.group(3)
        ia, a = _header_index(names, a)
        ic, c = _header_index(names, c)
        if ia is None or ic is None:
            missing = [x for x, i in ((a, ia), (c, ic)) if i is None]
            return None, (f"『{"』『".join(missing)}』という列がありません"
                           f"（ある列: {"、".join(names)}）")
        lo, hi = min(ia, ic), max(ia, ic)
        at = hi + 1 if side == "右" else lo
        return at, f"『{a}』と『{c}』の{side}＝{at}列目"
    for suf in _COL_AFTER + _COL_BEFORE:
        m = _re_anchor(suf).search(text)
        if not m:
            continue
        idx, name = _header_index(names, m.group(1).strip())
        if idx is None:
            return None, (f"『{name}』という列がありません"
                           f"（ある列: {"、".join(names)}）")
        after = suf in _COL_AFTER
        at = idx + 1 if after else idx
        return at, f"『{name}』（{idx}列目）の{"右" if after else "左"}＝{at}列目"
    return None, None


def inspection_sheet_basic_call(sheet_name: str, header: list, rows: list,
                                 types: str) -> str:
    """検分シートを LibreOffice 側で書く Basic の 1 行を組む。

    ★ 2026-08-24（土台固め）: 旧実装は openpyxl でブックを開き直して検分シートを足して
    いた。openpyxl の往復は xl/drawings の**中身の図形**（描かれた角印・社判・
    テキストボックス）を捨てる ── ファイル名は残るので、忠実度ゲートの
    ファイル名比較にも掛からなかった。実測: 雛形に角印のある請求書で、LO が正しく
    N 枚へ複製した角印を最後の openpyxl 往復が全部消し、✓ が出ていた。
    ★ LO 経路は図形を保つと実測済みなので、書き手を LO へ寄せて往復ごと無くす。
    ★ 実装は 1 つ（帳票段と様式写像段が同じ関数を呼ぶ ── 書き写さない）。

    types は列ごとの型（"s"=文字列 / "n"=数値）。Excel のシート名・セル値は制御文字を
    含めないので、レコード Chr(30) / フィールド Chr(31) の区切りは安全。
    """
    # ★ 区切りは **Basic の Chr(30)/Chr(31) 式**として書く ── 生の制御文字を .bas に
    #   埋めない（この repo は生成物に混ざった制御文字で一度事故を起こしている）。
    def q(v):
        return '"' + ("" if v is None else str(v)).replace('"', '""') + '"'

    def rec(values):
        return " & Chr(31) & ".join(q(v) for v in values)

    parts = [rec(header)] + [rec(r) for r in rows]
    payload_expr = " & Chr(30) & ".join(parts) if parts else '""'
    name = str(sheet_name).replace('"', '""')
    return (f'    Call WriteInspectionSheet(oDoc, "{name}", '
            f'{payload_expr}, "{types}")' + chr(10))


POSTCONDITIONS = {
    "SORT": check_sort, "COMPUTE_COLUMN": check_compute_column,
    "LOOKUP_FILL": check_lookup_fill, "AGGREGATE": check_aggregate,
    "BOLD": check_bold, "FILL_COLOR": check_fill_color,
    "NUMBER_FORMAT": check_number_format, "MERGE": check_merge,
    "CENTER_ALIGN": check_center_align, "APPEND_TOTAL": check_append_total,
    # ★ W9: 検証済みヘルパ4種。
    "INSERT_ROWS": check_insert_rows, "DRAW_BORDERS": check_draw_borders,
    # ★ 2026-08-26: 表の基本操作 3 種
    "ADD_ROW": check_add_row, "DELETE_ROWS": check_delete_rows,
    "DELETE_COLUMN": check_delete_column, "SET_CELL_VALUE": check_set_cell_value,
    "SWAP": check_swap, "ADD_COLUMN": check_add_column, "SET_WHERE": check_set_where,
    "EXTRACT_COLUMNS": check_extract_columns,
    "AUTOFIT": check_autofit, "PIVOT": check_pivot,
    # ★ 致命3(W10e):
    "SET_COLUMN_VALUE": check_set_column_value,
    "EXTRACT": check_extract,
    "SPLIT_CELL": check_split_cell,
    "DEDUP": check_dedup,
    # ★ 帳票段:
    "REPORT_PER_ROW": _check_report_router,
    # ★ 様式写像段:
    "FORMAT_MAP": check_format_map,
}


def run_postcondition(op: str, out_book: Path, resolved_args: dict, before_charts: int = 0,
                       header_row: int = 1, use_formula: bool = False,
                       source_book: Path | None = None,
                       before_chart_paths: frozenset | None = None) -> tuple:
    """⑥ op 別事後条件。(status, reason)。status ∈ {"pass","warn","fail","error"}。
       CHART だけ before_charts と比較する専用の形。
       ★ W3: header_row（1起点、省略時1）を全チェッカーに一貫して渡す（『三層全部が
       同じ見出し推定を使う』の事後条件側）。
       ★ W10f 項目1: use_formula は元は COMPUTE_COLUMN 専用（W3 Part3）だったが、
       operand を式ビューから読む同型バグが SORT/AGGREGATE にもあったため、その2つにも
       渡すよう広げた（同じ理由で check_compute_column_single_factor は
       check_compute_column からの委譲で既に受け取っている）。
       ★ W10f 項目5: LOOKUP_FILL のキー列にも同型バグ(対象シートのキー列を式ビューから
       読む)があったため渡すよう広げた。
       ★ W9: source_book（適用前のコピー・無ければ None）は INSERT_ROWS/AUTOFIT だけが使う
       （before/after の突き合わせが要る2op・他は無視される安全なキーワード引数）。
       ★ 止血2: チェッカー内で予期しない例外が起きても生の Python トレースバックを
       出さない。ここで必ず捕まえて "error" ステータス + 要約1行に変換する
       （C②の教訓: 事後条件チェッカー自身のクラッシュがユーザーに未捕捉のまま漏れていた）。
       ★ 致命④(2026-08-23レビュー): before_chart_paths（snapshot()["chart_paths"]・
       apply 前の chart XML パス集合）を CHART の check_chart_series へそのまま渡す。
       None なら check_chart_series 側が従来どおり先頭一致で見る（後方互換）。"""
    try:
        # ★ 全 op の合流点。見ていない行/列が在るのは表の性質なので、op を問わず先に数える。
        note_extent_gap(out_book, resolved_args, header_row)
        if op == "CHART":
            # ★ グラフ段: 事後条件を二層にする。①グラフ数+1（旧 check_chart・恒真殺しの
            #   手前）②その1個が意図した種別/値列を指しているか（check_chart_series・
            #   恒真殺し本体）。①が fail/error ならそこで止める（②は「グラフが在る」前提）。
            status, reason = check_chart(out_book, before_charts)
            if status != "pass":
                return status, reason
            with BookView(out_book) as bv:
                ws = bv.sheet(resolved_args.get("_target_sheet"))
                val_idx = _col_index_by_header(ws, resolved_args["value_col"], header_row=header_row)
                cat_name = resolved_args.get("category_col")
                cat_idx = (_col_index_by_header(ws, cat_name, header_row=header_row)
                           if cat_name else None)
            if val_idx is None:
                return "fail", f"値列『{resolved_args['value_col']}』が見つからない"
            return check_chart_series(
                out_book, kind=resolved_args.get("kind") or "bar",
                value_col_letter=get_column_letter(val_idx),
                category_col_letter=get_column_letter(cat_idx) if cat_idx else None,
                before_chart_paths=before_chart_paths)
        fn = POSTCONDITIONS.get(op)
        if fn is None:
            return "fail", f"未対応の op: {op}"
        if op == "COMPUTE_COLUMN":
            # ★ 2026-08-30: 「作った列以外は 1 セルも変わらず」を見るため before が要る
            #   （番人の感度を測る治具が、ここが素通りするのを見つけた）。
            return fn(out_book, resolved_args, header_row, use_formula,
                       source_book=source_book)
        if op == "SORT":
            # ★ 算術恒等の検算: 合計行が最下行から動いたかは before が無いと測れない。
            return fn(out_book, resolved_args, header_row, use_formula=use_formula,
                       source_book=source_book)
        if op in ("AGGREGATE", "LOOKUP_FILL"):
            return fn(out_book, resolved_args, header_row, use_formula=use_formula)
        if op in ("INSERT_ROWS", "AUTOFIT", "EXTRACT", "DEDUP", "REPORT_PER_ROW", "FORMAT_MAP",
                   "ADD_ROW", "DELETE_ROWS", "DELETE_COLUMN", "SET_CELL_VALUE", "SWAP",
                   "ADD_COLUMN", "SET_WHERE",
                   "EXTRACT_COLUMNS", "BOLD", "FILL_COLOR", "CENTER_ALIGN"):
            return fn(out_book, resolved_args, header_row, source_book=source_book)
        return fn(out_book, resolved_args, header_row)
    except Exception as e:
        return "error", f"事後条件の検証に失敗: {type(e).__name__}: {e}"


# ---------------------------------------------------------------------------
# basrun 経由の適用
# ---------------------------------------------------------------------------

def _timeout_error_message(timeout: float) -> str:
    return f"実行時エラー: マクロが {timeout:.0f} 秒で終了しない（無限ループの可能性）"


def short_error_summary(err: str) -> str:
    """★ M2a: obasync 由来の生 Python トレースバックを端末にそのまま出さないための
       整形。最終行（例外名+メッセージ）だけを取り出す。全文は履歴 jsonl 側に残す
       （修復ループへは従来どおり err の全文を渡す＝モデルへの情報は減らさない）。"""
    lines = [ln for ln in (err or "").splitlines() if ln.strip()]
    return lines[-1].strip() if lines else "(詳細不明)"


def _kill_process_tree(pid: int) -> None:
    """PID 指定でプロセスツリーを kill する。
       ★ taskkill /IM（名前一括）は使わない — 無関係な他プロセスも巻き込む
       （2026-08 の Senior MCP 事故の教訓）。必ず特定した PID だけを狙う。"""
    if os.name == "nt":
        subprocess.run(["taskkill", "/PID", str(pid), "/T", "/F"],
                        capture_output=True)
    else:
        import signal
        try:
            os.killpg(os.getpgid(pid), signal.SIGKILL)
        except Exception:
            pass


def basrun_apply(book: Path, code: str, workdir: Path, helper_files=(),
                  timeout: float | None = DEFAULT_APPLY_TIMEOUT) -> tuple:
    """生成コードを book に適用する。(ok, error_or_None, raw_output)。
       helper_files があれば同じ src に置く＝同じライブラリに同期され、Gen から呼べる。

    ★ M1: timeout（既定 180 秒、None/0 で無効）を basrun 側の `apply --timeout` に
       転送する。basrun は自分の内部タイムアウトで、ハングした接続先の LibreOffice
       だけを stop_office() で終了させる（taskkill 一括はしない、既存機構）。
       ここではさらに外側の安全網として、basrun 自身が固まった場合に備えて
       ゆとり(+30秒)を持たせた outer timeout を持ち、それも超えたら basrun.py の
       プロセスを PID 指定で kill する。どちらの経路で落ちても、修復ループには
       同じ「実行時エラー: マクロが N 秒で終了しない」で渡す（正直な分類）。"""
    src = workdir / "src"
    if src.exists():
        shutil.rmtree(src)
    src.mkdir(parents=True)
    for hf in helper_files:
        shutil.copy2(hf, src / hf.name)
    (src / "Gen.bas").write_text(code, encoding="utf-8")

    cmd = [sys.executable, str(basrun_path()), "apply", str(book), str(src), "AiLine", "Gen.Run"]
    if timeout:
        cmd += ["--timeout", str(timeout)]
    outer_timeout = (timeout + 30) if timeout else None

    proc = subprocess.Popen(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        text=True, encoding="utf-8", errors="replace")  # ★ cp932 事故を避ける
    try:
        out, err_out = proc.communicate(timeout=outer_timeout)
    except subprocess.TimeoutExpired:
        # ★ basrun 自身の内部タイムアウト＋stop_office() が働かなかった場合の安全網。
        _kill_process_tree(proc.pid)
        try:
            out, err_out = proc.communicate(timeout=5)
        except Exception:
            out, err_out = "", ""
        raw = (out or "") + "\n" + (err_out or "")
        return False, _timeout_error_message(timeout), raw

    raw = (out or "") + "\n" + (err_out or "")
    if proc.returncode != 0:
        # basrun 自身の内部タイムアウトで落ちた場合も、同じ「実行時エラー」分類に正規化する。
        if timeout and "秒応答しなかった" in raw:
            return False, _timeout_error_message(timeout), raw
        return False, raw.strip()[-800:], raw
    return True, None, raw


# ★ 正規化パス専用の空マクロ。何もしない（Call すら書かない）が、basrun_apply の
#   `doc.store()` を一度通すことで LibreOffice 側の初回保存の実体化を先に済ませる。
NOOP_MACRO = "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\nEnd Sub\n"
#  ★ W3: 正規化パスは NOOP_MACRO でなく _structdump_macro() を使う（同じ LO 往復で
#     構造読み取りを二役させる）。文書へは何も書かない点は NOOP_MACRO と同じ。


def _stop_office() -> None:
    """basrun.py stop を呼んで LibreOffice を落とす（M2c: normalize_book の自動リトライ用）。
       ★ ここでの失敗は無視する（次の apply がどのみち再起動を試みる。taskkill 一括はしない
       既存機構の外側なので、ここでは basrun 自身の stop に委譲するだけ）。"""
    try:
        subprocess.run([sys.executable, str(basrun_path()), "stop"],
                        capture_output=True, encoding="utf-8", errors="replace")
    except Exception:
        pass


def normalize_book(book: Path, workdir: Path,
                    timeout: float | None = DEFAULT_APPLY_TIMEOUT) -> Path:
    """コピーを LibreOffice で一度（空マクロで）開いて保存する ＝ P0 の正規化パス。

    LibreOffice は openpyxl 製（＝ LO で保存されたことがない）ブックを初回保存する際、
    行高（時に列幅）を実体化する。before スナップショットをこの実体化の**前**に取ると、
    その副作用が「マクロが変化させた」と誤検出される（no-op ガードの偽陽性・製品の心臓）。
    先にこの正規化を一度済ませておけば、以降の before/after 比較はマクロの実際の効果
    だけを見る。コストは LO 往復 1 回（数秒）— 正しさ優先で受け入れる。
    参考: ailine-ts の tests/e/_harness.ts normalizeThroughLibreOffice が同じ手当てを
    テスト側で先に実装していた（挙動の参考。製品経路に入れるのはこちらが初）。

    ★ M2c: 監査2回で2回再現した既知の摩擦（RuntimeException: Could not create system
    bitmap! 等、LibreOffice 側の一時的な描画/接続不調）への低リスク対処。1回失敗しても
    即座に落とさず、stop（LibreOffice を落とす）→ 再起動を挟んで1回だけ自動リトライする。
    それでも失敗したら現行どおりのエラーで落ちる（無限リトライはしない）。"""
    normalized = workdir / ("normalized" + book.suffix)
    shutil.copy2(book, normalized)
    dump_path = workdir / STRUCTDUMP_FILENAME
    struct_code = _structdump_macro(dump_path)
    ok, err, _ = basrun_apply(normalized, struct_code, workdir, timeout=timeout)
    if not ok:
        _stop_office()
        shutil.copy2(book, normalized)   # 中途半端な保存状態を残さず作り直す
        ok, err, _ = basrun_apply(normalized, struct_code, workdir, timeout=timeout)
    if not ok:
        exit_environment(f"正規化パスに失敗した（LibreOffice で開けなかった）: {err}")
    return normalized


def success_message(result: dict) -> str | None:
    """★ の注意書きは『変化を検出して適用が成功した』ときだけ出す。
       失敗(exit 1)や --dry（何も適用していない）で出すのは不誠実（P1）。"""
    if result.get("ok") and not result.get("dry"):
        return "★ 変化は検出したが『正しいか』は上の差分を見て判断してください（no-op ガードは正しさを保証しない）。"
    return None


# ---------------------------------------------------------------------------
# --inplace バックアップ + restore（M2a）
#
# ★ --inplace は原本を上書きするので、上書き前に必ず ~/.ailine/backups/ へコピーする。
#   バックアップに失敗したら --inplace 自体を中止する（安全側。原本は無変更のまま）。
#   restore は復元前の現状も退避してから上書きする＝復元自体も取り消せる。
# ---------------------------------------------------------------------------

# ★ W8b: 秒精度(旧・6桁)・マイクロ秒精度(新・12桁)・衝突回避の "-N" 連番つき、の
#   いずれも受け付ける（後方互換）。
_BACKUP_TS_RE = re.compile(r"^\d{8}T\d{6}(?:\d{6})?Z(?:-\d+)?$")
_BACKUP_TS_SEQ_RE = re.compile(r"-(\d+)$")

# ★ W11: undo が「復元前の現状」を退避する棚。名前空間ディレクトリの**さらに下**に置く
#   （BACKUP_DIR/<ns>/undo/）。list_backups は名前空間ディレクトリ直下のファイルしか
#   見ないので、ここに置いたものは自動的に「遡りの履歴」から外れる。
#   なぜ分けるか（実測した不具合）: 退避を実編集の世代と同じ棚に積むと、undo のたびに
#   履歴が伸び、しかも最も古い状態に着いた後は「現在の中身と同じ退避」が最新世代として
#   並ぶため、次の undo が『その1つ内側』＝直前に打ち消したはずの新しい状態を釣り上げて
#   いた（N 回編集して N+2 回 undo すると v1 が復活する）。退避を消すのでなく、
#   **数える場所と遡りの参照点から外す**ことで直す（安全網は減らさない）。
UNDO_SHELF_DIRNAME = "undo"


class BrokenBackupError(Exception):
    """復元元が Excel として開けない ── 原本に被せる前に止める（復元の致命②）。"""


class NoOlderBackupError(Exception):
    """★ W11: 最も古い世代に着いていて、これ以上遡れない（＝undo の端）。
       「バックアップが1つも無い」(FileNotFoundError)とは別物なので型で分ける。"""


def _utc_ts() -> str:
    """ファイル名に使える UTC タイムスタンプ（例: 20260814T120000123456Z）。
       ★ W8b: マイクロ秒まで含める（実測: それでも Windows の壁時計分解能は粗く
       20万回の連続呼び出しで56通りしか値が変わらない。衝突自体は make_backup 側の
       "-N" 連番フォールバックで最終的に防ぐ・ここは表示上の精度向上に留める）。"""
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%f") + "Z"


def _ts_sort_key(ts: str):
    """ts 文字列（秒/マイクロ秒精度・衝突回避の "-N" 連番つき）を新しい順ソート用の
       比較可能な値 (datetime, 連番) にする。壊れた形式は最古扱い。"""
    body = ts
    seq = 0
    m = _BACKUP_TS_SEQ_RE.search(body)
    if m:
        seq = int(m.group(1))
        body = body[:m.start()]
    if body.endswith("Z"):
        body = body[:-1]
    for fmt in ("%Y%m%dT%H%M%S%f", "%Y%m%dT%H%M%S"):
        try:
            return (datetime.strptime(body, fmt), seq)
        except ValueError:
            continue
    return (datetime.min, seq)


def _backup_namespace(book: Path) -> str:
    """book が置かれているフォルダごとのバックアップ名前空間（sha1 の先頭8桁）。
       ★ W8b 項目3: 同名ファイルが別フォルダにある場合（A\\見積.xlsx / B\\見積.xlsx）、
       旧・フラットな BACKUP_DIR では stem+suffix だけで一致させていたため取り違え
       （undo 混線）が起きえた。フォルダの絶対パスでハッシュ化した名前空間ディレクトリへ
       分離する（book が相対パスで渡されても resolve() してから使う）。"""
    return hashlib.sha1(str(Path(book).resolve().parent).encode("utf-8")).hexdigest()[:8]


def undo_shelf_dir(book: Path) -> Path:
    """★ W11: undo が取る「復元前の現状」の退避先（遡りの履歴には数えない棚）。"""
    return BACKUP_DIR / _backup_namespace(book) / UNDO_SHELF_DIRNAME


def backup_path_for(book: Path, ts: str | None = None, shelf: bool = False) -> Path:
    ts = ts or _utc_ts()
    base = undo_shelf_dir(book) if shelf else BACKUP_DIR / _backup_namespace(book)
    return base / f"{book.stem}.{ts}{book.suffix}"


def keep_backups_arg(value: str) -> int:
    """`--keep-backups` の値。★ 0 は受け付けない。

    ★ 2026-08-25（復元の致命3・盲検）: 0 は `backups[0:]`＝**いま作った分ごと全部消す**
      という意味になり、安全網ゼロで原本を書き換えていた。それでいて
      「（もとに戻す: ailine undo）」と表示していた ── 嘘。
      ★ `--help` は「負数で無制限」と書くが、**0 を無制限と読む利用者は普通に居る**
        （多くの CLI がそう）。その 1 文字で原本が戻らなくなるので、ここで止める。
    """
    import argparse as _ap
    try:
        n = int(value)
    except ValueError:
        raise _ap.ArgumentTypeError(f"整数で指定してください: {value!r}")
    if n == 0:
        raise _ap.ArgumentTypeError(
            "0 は「バックアップを作った直後に全部消す」という意味になり、"
            "原本を戻せなくします。無制限にしたいなら -1 を、"
            "世代を残すなら 1 以上を指定してください")
    return n


def _pruned_marker_path(book: Path) -> Path:
    """「この本は剪定で古い世代を捨てたことがある」の印。★ 停止メッセージを正直にするため。"""
    return BACKUP_DIR / _backup_namespace(book) / f".pruned-{book.stem}{book.suffix}"


def pruned_generations_note(book: Path, keep: int) -> str | None:
    """剪定でこれから捨てる世代を人へ告げる 1 行（捨てるものが無ければ None）。

    ★ 2026-08-25（復元の致命4）: 剪定は完全に無言だった。上限を超えて古い世代を
      黙って捨てたあと、undo が「最も古い状態です」と言う ── 実際は
      「**まだ残っている中で**一番古い」でしかなく、原本は既に消してある。
    """
    if keep < 0:
        return None
    stale = list_backups(book)[keep:]
    if not stale:
        return None
    try:
        marker = _pruned_marker_path(book)
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text("1", encoding="utf-8")
    except OSError:
        pass
    return (f"（古い世代 {len(stale)} 件を捨てました ── 上限 {keep} 世代。"
            "これより前へは戻せなくなります）")


def prune_backups(book: Path, keep: int = DEFAULT_KEEP_BACKUPS, shelf: bool = False) -> list:
    """★ M2c: book の世代のうち keep 件を超える古いもの（list_backups は新しい順）を削除する。
       戻り値は削除したパスのリスト。keep < 0 は「無制限（削除しない）」扱い。
       ★ W11: shelf=True のときは undo の退避棚を対象にする（棚も同じ上限で剪定する）。"""
    if keep < 0:
        return []
    backups = list_undo_shelf(book) if shelf else list_backups(book)
    stale = backups[keep:]
    deleted = []
    for p in stale:
        try:
            p.unlink()
            deleted.append(p)
        except OSError:
            pass
    return deleted


def make_backup(book: Path, keep: int = DEFAULT_KEEP_BACKUPS, shelf: bool = False) -> Path:
    """book のバックアップを ~/.ailine/backups/<名前空間>/ に作る。戻り値はバックアップ先。
       ★ 失敗したら例外を投げる（呼び出し側が --inplace 中止の判断に使う）。
       ★ M2c: 新しいバックアップを作った後、keep 世代を超えた古いものを剪定する
       （既定 DEFAULT_KEEP_BACKUPS=10。無制限にすると個人開発機のディスクを静かに食う）。
       ★ W8b 項目3: 新規のバックアップは必ず名前空間ディレクトリへ書く（フラット領域には
       もう書かない・読み取り専用の後方互換は list_backups 側で担う）。
       ★ W8b: Windows の壁時計分解能は実測で粗く（20万回の連続呼び出しで56通りしか
       値が変わらない）、restore_backup が「復元前の現状」を退避する高速な連続呼び出し
       等でファイル名が衝突しうる。衝突したら "-N" 連番を足して必ず別ファイルにする
       （既存の世代を上書きで消さない・回帰テストで自己顕在化した実際の不具合の修正）。
       ★ W11: shelf=True は undo が「復元前の現状」を退避するときの置き場（BACKUP_DIR/
       <ns>/undo/）。遡りの履歴には数えないが、ファイルとしては同じ作法で残す
       （undo を可逆にする性質＝退避そのものは減らさない）。"""
    # ★ 2026-08-25（復元の致命1の後半）: undo の直後に run すると、「いま復元したばかりの
    #   内容」がもう一度世代に積まれる。歩みは（同一性で判定するようになったので）壊れないが、
    #   **中身の同じ世代を 2 回通る**ので、使う側には「undo が効いていない」ように見える。
    #   最新世代と 1 バイトも違わないなら、積まない ── 世代列は「変化の履歴」であって
    #   「実行の履歴」ではない。★ 退避棚（shelf）は別: undo 自体を可逆にするための記録なので
    #   同じ中身でも残す。
    if not shelf:
        existing = list_backups(book)
        if existing:
            try:
                if existing[0].read_bytes() == book.read_bytes():
                    prune_backups(book, keep=keep, shelf=shelf)
                    return existing[0]
            except OSError:
                pass
    ts = _utc_ts()
    dst = backup_path_for(book, ts=ts, shelf=shelf)
    dst.parent.mkdir(parents=True, exist_ok=True)
    n = 2
    while dst.exists():
        dst = backup_path_for(book, ts=f"{ts}-{n}", shelf=shelf)
        n += 1
    shutil.copy2(book, dst)
    if not shelf:
        # ★ 2026-08-25（復元の致命4）: 剪定は完全に無言だった。捨てる**前**に言う。
        note = pruned_generations_note(book, keep)
        if note:
            print(note)
        # ★ 2026-08-25: 実編集の世代を積んだ ── これから書き換わる中身は、どの世代とも
        #   同じでない（＝新しい編集の直後）。undo の退避（shelf）では動かさない。
        _write_undo_pointer(book, None)
    prune_backups(book, keep=keep, shelf=shelf)
    return dst


def _parse_backup_name(name: str, stem: str, suffix: str) -> str | None:
    """バックアップのファイル名が `<stem>.<ts><suffix>` の形かを見て、ts を返す
       （形が違えば None）。"""
    prefix = stem + "."
    if not (name.startswith(prefix) and name.endswith(suffix)):
        return None
    ts = name[len(prefix):len(name) - len(suffix)]
    return ts if _BACKUP_TS_RE.match(ts) else None


def _gather_backups(directory: Path, stem: str, suffix: str) -> list:
    """directory 直下（再帰しない）の `<stem>.<ts><suffix>` 形のファイルを (ts, Path) で集める。
       iterdir で拾うのはファイルだけ＝サブディレクトリ（名前空間・undo の退避棚）自体を
       誤ってバックアップと数えない。並べ替えは呼び出し側の責任。"""
    found = []
    if directory.is_dir():
        for p in directory.iterdir():
            if not p.is_file():
                continue
            ts = _parse_backup_name(p.name, stem, suffix)
            if ts is not None:
                found.append((ts, p))
    return found


def _sorted_newest_first(found: list) -> list:
    # ★ W8b: 秒精度(旧)とマイクロ秒精度(新)が混在しうるため、生文字列の辞書順ではなく
    #   _ts_sort_key() でパースした実時刻順に並べる（桁数違いの文字列比較は時刻順にならない）。
    found.sort(key=lambda pair: _ts_sort_key(pair[0]), reverse=True)
    return [p for _ts, p in found]


def list_backups(book: Path) -> list:
    """book に対応するバックアップ（＝遡れる実編集の世代）を新しい順で返す。
       ★ W8b 項目3: 名前空間ディレクトリ BACKUP_DIR/<ns>/ を主として見る。
       旧フラット領域（BACKUP_DIR 直下・名前空間分離前の名残）も読み取り専用互換で
       あわせて見る（新規はもう書かない）。
       ★ W11: undo の退避棚 BACKUP_DIR/<ns>/undo/ は**含めない**（直下しか見ないので
       自動的に外れる）。棚を見たいときは list_undo_shelf() を使う。"""
    stem, suffix = book.stem, book.suffix
    return _sorted_newest_first(
        _gather_backups(BACKUP_DIR / _backup_namespace(book), stem, suffix))


def _renamed_or_moved_note(book: Path) -> str:
    """同じフォルダの世代置き場に**別の名前の**世代が在れば、そう言う。

    ★ 2026-08-26（復元の盲検 3 回目・重大5）: ファイル名を変えたり移動したりすると
      `× <名前> のバックアップが無い` とだけ言って終わっていた。世代はそのまま在るのに、
      利用者には「消えた」としか見えない ── 命綱が「無い」と言うのと
      「別の名前で在る」は別物。
    ★ 勝手に結び付けない（同名でない物を当てるのは、致命①で塞いだばかりの事故の形）。
      在り処と、名前で結び付いていることだけを言う。
    """
    ns_dir = BACKUP_DIR / _backup_namespace(book)
    if not ns_dir.is_dir():
        return ""
    others = [q for q in ns_dir.iterdir()
               if q.is_file() and not q.name.startswith(".")]
    if not others:
        return ""
    return (f"（このフォルダの世代置き場には別の名前の世代が {len(others)} 件あります: "
            f"{ns_dir} ── 世代は「フォルダ＋ファイル名」で結び付けているので、"
            "名前を変えたり移したりすると辿れなくなります）")


def list_legacy_backups(book: Path) -> list:
    """旧フラット領域（BACKUP_DIR 直下・名前空間分離前の名残）にある**同名**の世代。

    ★ 2026-08-25（復元の致命①・盲検 2 回目）: ここは以前 list_backups が黙って
      混ぜていた。直下のファイルは**フォルダの情報を持たない**ので、同名なら
      誰のものでも遡り履歴に入る ── 実際に再現した:

          A 社の古い世代が直下に残っている状態で B 社の同名ブックを編集
          → B の遡り履歴に A 社の中身が入る

      ★ この機械の実 ~/.ailine/backups/ 直下にも 4 件在った。

    ★ 番人が 2 本、互いに反対を守っていた（「混線しない」と「旧領域へ遡れる」）。
      どちらも単独では緑で、**組み合わせが誰も試されていなかった**。
      両方を同時に真にする道は 1 つ ── 出所が分からないものを**勝手に使わない**、
      しかし**在ることは言う**（断れない時は開示する）。読めるし、手で戻せる。
    """
    return _sorted_newest_first(_gather_backups(BACKUP_DIR, book.stem, book.suffix))


def list_undo_shelf(book: Path) -> list:
    """★ W11: undo が取った「復元前の現状」の退避を新しい順で返す（遡りには数えない）。"""
    return _sorted_newest_first(_gather_backups(undo_shelf_dir(book), book.stem, book.suffix))


def _undo_pointer_path(book: Path) -> Path:
    """「いま、どの世代の上に立っているか」を記録する印の置き場。

    ★ 2026-08-25（復元の致命1・盲検の最重）: 現在地を**内容の等値**で決めていたため、
    同じ内容の世代が 2 つ並んだ瞬間に位置が確定できず、undo が振動して原本へ永久に
    到達できなくなっていた（トリガは `run → undo → run` ── README が勧める使い方）。
    ★ 内容は世代の一意キーではない。**どの世代から復元したか**を覚えておく。
    """
    return BACKUP_DIR / _backup_namespace(book) / f".at-{book.stem}{book.suffix}"


def _read_undo_pointer(book: Path) -> str | None:
    try:
        name = _undo_pointer_path(book).read_text(encoding="utf-8").strip()
    except OSError:
        return None
    return name or None


def _write_undo_pointer(book: Path, backup_name: str | None) -> None:
    """backup_name=None は「新しい編集の直後（どの世代の上でもない）」の意味。"""
    path = _undo_pointer_path(book)
    try:
        if backup_name is None:
            path.unlink(missing_ok=True)
            return
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(backup_name, encoding="utf-8")
    except OSError:
        pass


def _undo_position(book: Path, backups: list) -> int | None:
    """book の現在地が世代列(新しい順)のどこかの添字。どこでもなければ None
       （＝実編集の直後で、まだ 1 段も遡っていない）。

    ★ 判定は「どの世代から復元したか」の記録（同一性）で行う。内容は
      **その記録がまだ有効かの確認**にだけ使う ── 人が Excel で直接編集していたら
      指し先は当てにならないので、新しい編集の直後として扱う。
    ★ 記録が無い置き場（この仕組みより前に作られたバックアップ）は、従来どおり
      内容の一致で探す（後方互換・そこでは重複の事故が残るが、壊れはしない）。
    """
    try:
        current = book.read_bytes()
    except OSError:
        return None

    pointed = _read_undo_pointer(book)
    if pointed:
        for i, p in enumerate(backups):
            if p.name != pointed:
                continue
            try:
                if p.read_bytes() == current:
                    return i          # 記録どおりの世代の上に立っている
            except OSError:
                pass
            return None               # 人が直接編集した ── 指し先は当てにならない
        return None                   # 指し先の世代が剪定で消えた

    for i, p in enumerate(backups):   # 記録が無い置き場（後方互換）
        try:
            if p.read_bytes() == current:
                return i
        except OSError:
            continue
    return None


def undo_steps_left(book: Path, backups: list | None = None) -> int:
    """★ W11: 今の位置から**あと何回 undo できるか**。バックアップの総数ではない
       （総数を数えていたので、undo が退避を積むたびに『あと N 回』が増えていた）。
       現在地が世代列の i 番目なら、残りは i より古い世代の数 = len-1-i。
       現在地がどこにも無い（＝実編集の直後）なら全世代を遡れるので len。"""
    backups = list_backups(book) if backups is None else backups
    if not book.exists():
        return len(backups)
    i = _undo_position(book, backups)
    return len(backups) if i is None else len(backups) - 1 - i


def _shelve_bytes(book: Path, data: bytes) -> None:
    """復元**前**の中身を undo の棚へ退避する（restore 自体を可逆にする材料）。

    ★ 2026-08-25（復元の重大9）: 旧実装は `make_backup(book, shelf=True)` を
      書き込みの前に呼んでいたので、①書き込みが失敗しても積まれ ②何も変わらない
      no-op でも積まれた。実測では棚 10 件すべてが同一内容で埋まり、本物が押し出された。
      ★ ここは「今のファイル」ではなく「**さっきまでの中身**」を積む必要があるので、
      make_backup（ファイルを読む）ではなくバイト列を直接書く。
    """
    try:
        dst = backup_path_for(book, ts=_utc_ts(), shelf=True)
        dst.parent.mkdir(parents=True, exist_ok=True)
        n = 2
        while dst.exists():
            dst = backup_path_for(book, ts=f"{_utc_ts()}-{n}", shelf=True)
            n += 1
        dst.write_bytes(data)
        prune_backups(book, keep=DEFAULT_KEEP_BACKUPS, shelf=True)
    except OSError:
        pass


def restore_backup(book: Path) -> Path:
    """book を「1つ前の世代」から復元する。戻り値は使ったバックアップの Path。
       バックアップが1つも無ければ例外を投げる。
       ★ W8b-2: 連続 undo が実編集履歴を1段ずつ正しく遡れるようにする（B2 事故バッテリ
       ①で実測: 素朴に『常に最新のバックアップを使う』実装だと、undo が作る『復元前の
       現状の退避』が次回の undo で『最新』として選ばれてしまい、2回目の undo が
       1段先(=1回目の undo を打ち消す)に進んでしまっていた＝多段 undo が成立しない）。
       book の現在の中身と一致するバックアップがあれば、そのすぐ内側(より古い方)を
       復元先にする（＝現在地がバックアップ履歴のどこかに『既にいる』とみなし、そこから
       もう1段遡る）。一致するものが無ければ（＝直前の実編集の直後・通常の最初の undo）
       最新のバックアップを復元する。復元前の現状は必ず退避する（undo 自体も可逆）。
       ★ W11: 最も古い世代に着いていたら NoOlderBackupError を投げて**止まる**
       （旧実装は同じものを復元して『✓ 復元した』と名乗っていた＝何もしていないのに成功）。
       ★ W11: 退避先は undo の棚（undo_shelf_dir）で、遡りの履歴には混ぜない
       （混ぜていたので、端に着いた後の undo が退避を最新世代として釣り上げていた）。"""
    backups = list_backups(book)   # 新しい順
    # ★ 2026-08-25（復元の致命①）: 旧領域の同名世代は**使わない**が、
    #   「無い」と言い切ると嘘になる場面がある ── 在ることと、なぜ使わないかを言う。
    legacy_note = "".join(render_legacy_note(list_legacy_backups(book)))
    if not backups:
        raise FileNotFoundError(f"{book.name} のバックアップが無い"
                                 f"{_renamed_or_moved_note(book)}{legacy_note}")

    target = backups[0]
    if book.exists():
        i = _undo_position(book, backups)
        if i is not None:
            if i + 1 >= len(backups):
                # ★ 2026-08-25（復元の致命4）: 「最も古い状態です」は「今のこれが原本だ」と
                #   読める断定文。剪定で原本を捨てている時は嘘になる。
                pruned = _pruned_marker_path(book).exists()
                tail = ("残っている中で一番古い状態です ── 上限を超えた古い世代は"
                        "剪定で捨てられており、原本には戻せません"
                        if pruned else "最も古い状態です")
                raise NoOlderBackupError(
                    f"{book.name} をこれ以上は戻せません（{tail}）{legacy_note}")
            target = backups[i + 1]
    # ★ 2026-08-25（復元の重大9・盲検）: 退避は**書き込みが成功してから**積む。
    #   旧版は書き込みの**前**に呼んでいたので、読み取り専用で 3 回失敗させたら
    #   棚が 2→5 件に増えた。しかも致命1 のループでは棚 10 件すべてが同一内容の
    #   原本コピーで埋まり、**本物の run1/run2 の結果が押し出されて全滅**した。
    #   ★ 棚は「undo をやり直す材料」── 何も起きなかった回に積むと、材料の方が消える。
    # ★ 2026-08-25（復元の致命②・片配線の 4 度目）: 「壊れた結果を原本に被せない」検査は
    #   **今朝、反映側にだけ**入れた（_why_output_is_unusable）。命綱の側は素の copy2 で、
    #   開けもしないファイルを原本に上書きして「✓ 復元した」と名乗っていた。
    #   ★ 別の世代へ勝手にずらさない ── どれを使うかは人が決める（黙って代用しない）。
    # ★ 基準は「今の原本が満たしているもの」── 三項目で見る。
    #   原本が開ける xlsx なのに復元元が開けないなら、被せた瞬間に確実に壊れる → 断る。
    #   原本が既に開けないなら、それは**救出**の最中で、同じ物差しを当てる根拠が無い
    #   （ここで断ると、壊れたファイルを直したい人の命綱を塞ぐ）。
    if book.exists() and _why_output_is_unusable(book) is None:
        broken = _why_output_is_unusable(target)
        if broken:
            raise BrokenBackupError(
                f"バックアップ {target.name} が開けません（{broken}）。"
                f"原本は変更していません ── 世代は {target.parent} に在ります")
    _prev = book.read_bytes() if book.exists() else None
    shutil.copy2(target, book)
    if _prev is not None and _prev != book.read_bytes():
        # 中身が実際に変わった時だけ退避する（同じ中身なら、やり直す材料にならない）
        _shelve_bytes(book, _prev)
    # ★ 2026-08-25: 「どの世代の上に立ったか」を記録する。内容の等値で当てないための第三項。
    _write_undo_pointer(book, target.name)
    return target


class NothingToRedoError(Exception):
    """やり直せるものが無い（undo をしていない・既に全部やり直した）。"""


def redo_last_undo(book: Path) -> Path:
    """直前の undo を**やり直す**。戻り値は使った退避の Path。

    ★★ 2026-09-02（README の「既知の問題」に自分で書いていた）: undo に redo が無く、
      戻しすぎると進めなかった。
    ★ 材料は既に在る ── undo は毎回「戻す前の中身」を棚へ退避している
      （`_shelve_bytes`・「undo 自体も可逆」）。**取り出す口を作るだけ**で、
      新しい保管場所も新しい概念も作らない。

    ★ 棚は積み下ろしの山（新しい順）。取り出したものは**棚から外す** ──
      外さないと、同じ状態を何度もやり直せてしまい「あと何回進めるか」が嘘になる。
    ★ やり直した後の現在地は「新しい編集の直後」と同じ（どの世代の上でもない）ので、
      印を消しておく。
      ★★ 正直に書く: **この 1 行は現状では効いていない**（2026-09-02 に実測）。
        `_undo_position` は「印が指す世代と今の中身が違えば無効」と見るので、
        redo の後は必ず印が無効になり、消しても消さなくても同じ場所へ行く。
        変異試験でも差が出せず、**効くことを示す検体を作れなかった**。
      ★ それでも残すのは、この repo が「**内容は世代の一意キーではない**」という
        事故（同じ内容の世代が 2 つ並んで undo が振動した）を実際に踏んでいるから
        ── 内容の一致だけに頼らない形にしておく。番人は付いていない、と分かるように書く。
    ★ 壊れた退避を原本に被せない ── undo と**同じ物差し**を当てる
      （原本が開けるのに退避が開けないなら断る。原本が既に開けないなら救出の最中なので
        同じ物差しを当てる根拠が無い）。
    """
    shelved = list_undo_shelf(book)
    if not shelved:
        raise NothingToRedoError(
            f"{book.name} はやり直せません（直前の undo がありません）")
    target = shelved[0]
    if book.exists() and _why_output_is_unusable(book) is None:
        broken = _why_output_is_unusable(target)
        if broken:
            raise BrokenBackupError(
                f"退避 {target.name} が開けません（{broken}）。"
                f"原本は変更していません ── 退避は {target.parent} に在ります")
    # ★★ 2026-09-02（自作 review・致命 1）: ここは**上書き前に今の中身を退避していなかった**。
    #   実測: undo のあとに別の編集を挟んでから redo すると、その編集内容が
    #   **警告なしに完全消失**した（backups にも棚にも残らず、rglob で全探索して不在を確認）。
    #   しかも画面には「✓ やり直しました」しか出ない。
    #   ★ 姉妹関数 restore_backup は上書き前に `_prev` を取って退避している ── **片配線**。
    #   ★ 直しは「棚に積む」ではなく **世代に積む**（make_backup）。
    #     棚に積むと redo が自分の直前状態を拾い直して往復が壊れる。
    #     世代に積めば `ailine undo` で普通に取り戻せる ── 新しい概念を作らない。
    #   ★ make_backup は最新世代と 1 バイトも違わなければ積まない（世代列は変化の履歴）。
    try:
        make_backup(book)
    except OSError as e:
        raise BrokenBackupError(
            f"やり直す前に今の内容を退避できませんでした（{e}）。原本は変更していません")
    try:
        shutil.copy2(target, book)
    except OSError as e:
        # ★ 2026-09-02（自作 review・致命 3 の後半）: ここが素だと、他プロセスが
        #   排他オープン中に**生の traceback** が出る（この repo が「重大7」で踏んだ形）。
        raise BrokenBackupError(
            f"やり直せませんでした（{e}）── 原本を開いているアプリを閉じてから、"
            "もう一度お試しください")
    # ★ 使った退避は外す（山を 1 つ下ろす）。消せなくても進めたことは事実なので黙って続ける。
    try:
        target.unlink()
    except OSError:
        pass
    _write_undo_pointer(book, None)
    return target


def cmd_redo(a: argparse.Namespace) -> int:
    """`ailine redo` ── 直前の undo をやり直す。

    ★ undo と同じ実行ロックを取る（原本を書き換えるため）。
    ★ フォルダには redo が無い（undo と同じ理由 ── 戻す対象が構造的に無い）。
    """
    book = Path(a.book).resolve()
    if book.is_dir():
        print("× フォルダに対する redo はありません"
               f"（原本は読んでいません）: {book}")
        return 1

    def _body() -> int:
        # ★★ 2026-09-02（自作 review・致命 3）: redo は**ロックの関所を一度も通っていなかった**。
        #   この repo は「run は Excel ロックで止まるのに undo は素通り」（復元の致命5）を
        #   既に踏み、番人を「1 本で 4 経路を縛る」形にしていた。
        #   ★ 俺が**5 本目の経路を作って配線しなかった** ── 在っても鳴らない、そのもの。
        blocked = refuse_if_locked(book)
        if blocked is not None:
            return blocked
        try:
            used = redo_last_undo(book)
        except NothingToRedoError as e:
            print(f"× {e}")
            print("  → 戻すなら ailine undo、世代の一覧は ailine undo --list")
            return 1
        except BrokenBackupError as e:
            print(f"× {e}")
            return 1
        left = len(list_undo_shelf(book))
        print(f"✓ {book.name} をやり直しました（{used.name} から）")
        print(f"  あと {left} 回やり直せます" if left
               else "  これ以上やり直せるものはありません")
        return 0

    return under_run_lock(_body)


def cmd_restore(a: argparse.Namespace) -> int:
    """`ailine restore` は `ailine undo` と**同じ仕事**をする（undo は restore の昇格版）。

    ★ 2026-08-25（復元の中#13・盲検）: 同じ `restore_backup` を呼ぶのに、restore 側だけが
      劣化版になっていた ── フォルダガードが無く「× w10 のバックアップが無い」と的外れな
      理由を言い、Excel ロックの関所も、例外を言葉にする処理も、残り回数の表示も無かった。
      undo で直したものが restore に届かない（片配線）。
    ★ 根治: 分岐を持たず**委譲する**。2 つ書かなければ、ずれようがない。
    """
    return cmd_undo(a)


def cmd_undo(a: argparse.Namespace) -> int:
    """★ W8b 項目5: `restore` の昇格。真実の源はバックアップファイル自体
       （history.jsonl には依存しない＝history が壊れていても undo できる）。
       名前空間対応(item3)は list_backups/restore_backup 経由でそのまま効く。
       復元後、まだ戻せる回数（★ W11: バックアップの総数ではなく**あと何段遡れるか**）を
       添える。端（最も古い状態）に着いたら復元せずに非零で止まる。
       ★ M2: フォルダ抽出（run <フォルダ>）は原本を読むだけでバックアップも履歴も作らない
       ── 戻す対象が構造的に存在しないので、フォルダには「無い」と正直に言って止まる。"""
    book = Path(a.book).resolve()
    if book.is_dir():
        print(f"× フォルダに対する undo はありません"
              f"（原本は読んでいません。消せるのは出力ブックだけです）: {book}")
        return 1
    if a.list:
        backups = list_backups(book)
        for ln in render_backup_list(book.name, backups,
                                      shelved=len(list_undo_shelf(book)),
                                      shelf_dir=undo_shelf_dir(book),
                                      legacy=list_legacy_backups(book)):
            print(ln)
        return 0
    # ★ 2026-08-25（復元の重大8）: undo も原本を書き換える ── run と同じ実行ロックを取る。
    #   （--list は読むだけなので上で先に返している）
    return under_run_lock(lambda: _cmd_undo_body(a, book))


def _cmd_undo_body(a: argparse.Namespace, book: Path) -> int:
    # ★ 2026-08-25（復元の致命5）: run は Excel ロックで止まるのに、undo は素通りしていた。
    #   「Excel で結果を見て、気に入らないから戻す」は undo の**最も自然な使い方**で、
    #   そこだけ関所が無かった。同じ検出器・同じ文言を通す。
    blocked = refuse_if_locked(book)
    if blocked is not None:
        return blocked
    try:
        used = restore_backup(book)
    except (FileNotFoundError, NoOlderBackupError, BrokenBackupError) as e:
        print(f"× {e}")
        return 1
    except PermissionError as e:
        # ★ 重大7（同じ盲検）: 読み取り専用の原本で**生の traceback** が出ていた。
        #   命綱がスタックトレースで死ぬのが最悪なので、必ず言葉にする。
        print(f"× {book.name} に書き込めませんでした（{e}）。")
        print("  → ファイルが読み取り専用か、他のアプリが開いていないか確認してください")
        return EXIT_WRITE_BLOCKED
    except OSError as e:
        print(f"× 復元に失敗しました（{e}）。原本は変更していません。")
        return 1
    print(render_restore_done(book.name, used.name, remaining=undo_steps_left(book)))
    # ★ 誤分類の実例台帳センサ②: undo が成功した＝その run の判断を人がひっくり返した
    #   容疑。task は history の該当ブックの直近エントリから引く（無ければ落とさず空文字）。
    _record_misclass_suspect("undo", _last_task_for_book(book), None, book)
    return 0


# ---------------------------------------------------------------------------
# ★ W8b: 安全器官（既定の反転は次コミット。今回は原本を直接書く危険を減らす下ごしらえ）
# ---------------------------------------------------------------------------

def refuse_if_run_cannot_handle(book: Path) -> int | None:
    """`ailine run` が扱えない形式なら、**触る前に**人の言葉で断る。

    ★★ 実測（README「未実装」に自分で書いていた）: `.ods` を渡すと**生の traceback**が
      出ていた。`--help` は 3 箇所で「.xlsx / .ods」と約束しているのに、
      `build_book_meta` は openpyxl なので読めない ── **約束だけが先行していた。**
      ★ 直しは 2 つで 1 組: 約束のほうを実体に合わせ（help から .ods を外す）、
        それでも来た時に**説明して**止まる。

    ★★ ここを `refuse_if_locked` に相乗りさせない（＝ undo も通る場所に置かない）。
      この repo は 2026-08 の盲検で出た「`.ods` の拒否を全形式に広げよう」を
      **却下している**: 断る範囲を広げると、**命綱（undo・バックアップ）に届く前に
      止まる経路**ができるため。だから関所は **run の入口 1 箇所だけ**に置く。
      undo は「壊れた形式でも戻せる」ままにする。

    ★ 断りは行き止まりにしない ── **直し方**を言う（この repo の作法）。
    """
    suffix = book.suffix.lower()
    if suffix in RUN_SUPPORTED_SUFFIXES:
        return None
    print(f"× {book.name} は、この道具が操作できる形式ではありません"
           f"（{suffix or '拡張子なし'}）。")
    print("  → 操作できるのは: " + "／".join(sorted(RUN_SUPPORTED_SUFFIXES)))
    if suffix in (".ods", ".ots"):
        print("  → LibreOffice で開いて「Excel 2007-365 (.xlsx)」として"
               "保存し直すと、そのまま扱えます")
    elif suffix in (".xls", ".xlsb", ".xlt"):
        print("  → 旧形式です。Excel か LibreOffice で .xlsx として保存し直してください")
    print("  → 中身を見るだけなら、この道具は使わずにそのまま開けます（原本は触っていません）")
    return EXIT_ENVIRONMENT


def refuse_if_locked(book: Path) -> int | None:
    """書き込みが塞がれていれば、理由を人の言葉で言って EXIT_WRITE_BLOCKED を返す。

    ★ 2026-08-26（初回体験の盲検 3 回目・CONFUSING 3）: 同じ判断が **4 箇所**に
      書き写されていて、3 通りに散らばっていた ──
        run（単一ブック）: 断定しない・心当たり 2 行（08-24 に直した正しい形）
        undo:             断定しない・心当たり 1 行（最後の 1 行が抜けている）
        run（2 冊照合）:   **「Excel で開かれています」と断定**・しかも
                          `{lock_a}` で **タプルをそのまま印字**していた
      ★ 直しは「3 つとも直す」ではなく **1 つに畳んで呼び出し側に持たせない**
        （今日までに片配線を 8 回踏んでいる）。番人も 1 本で 4 経路を縛る。

    ★ 重大7（復元の盲検 3 回目）の一部: Excel が異常終了すると `~$` が残り、
      undo まで恒久的に塞がれる。**回避フラグは足さない**（設計判断は別途）が、
      残骸である可能性と、消せば直ることは言う ── 断れない時は開示する。
    """
    reason = check_excel_lock(book)
    if not reason:
        return None
    kind, detail = reason
    print(f"× {detail}。")
    if kind == "excel":
        print("  → Excel で開いています。閉じてから実行してください")
        print("  → Excel を開いていないのに出る場合は、同じフォルダの"
              f"「~${book.name}」が前回の異常終了の残骸です（消せば進めます）")
    elif kind == "libreoffice":
        # ★ 「書けないから止める」ではなく「**あとで消えるから止める**」。理由を言う。
        print("  → LibreOffice で開いています。閉じてから実行してください")
        print("  → 開いたまま書き換えると、そのあと LibreOffice 側で保存したときに"
              "**開いた時点の内容で上書き**され、変更が消えます")
        print("  → 開いていないのに出る場合は、同じフォルダの"
              f"「.~lock.{book.name}#」が前回の異常終了の残骸です（消せば進めます）")
    else:
        # ★ 原因を断定しない ── 見たのは「書けない」ことだけ。心当たりを並べる。
        print("  → 心当たり: Excel などで開いている / 読み取り専用 / "
              "書き込み権限が無い / 同期中（OneDrive 等）")
        print("  → 開いていないのに出る場合は、ファイルのプロパティで"
              "「読み取り専用」と権限を確認してください")
    return EXIT_WRITE_BLOCKED


def check_excel_lock(book: Path) -> tuple | None:
    """book が Excel 等で開かれている兆候を機械的に見る。開かれていそうなら理由の
       文字列（人間可読）、そうでなければ None。
       ★ W8b 項目2: ①同フォルダの Excel ロックファイル(~$<name>) の存在
       ②open(book, 'r+b') を試みて PermissionError になるか、の2つを見る（保守的
       ＝どちらかに該当したら『開かれている可能性』として止める。誤検知より、
       書き込み中の文書を壊さない方を優先する）。run の最初（LO 起動・翻訳より前）
       に呼ぶ（--copy 時も含め常に同じ判定にする＝整合性の観点で経路を分けない）。"""
    # ★ 2026-08-24（初回体験の盲検・致命②）: 旧版はどちらの兆候でも
    #   「Excel で開かれています」と**原因を断定**していた。実測: Windows の ACL で
    #   書き込みを拒否しただけのファイル（Excel は 1 つも動いていない）に対して
    #   「Excel で開かれています。閉じてから実行してください」と言った。
    #   使う側はタスクマネージャを開いて Excel を探し、見つからず、次の手を失う。
    #   ★ 見たものと、その解釈を分ける ── 断定できるのはロックファイルが在る時だけ。
    lock_file = book.parent / f"~${book.name}"
    if lock_file.exists():
        return ("excel", f"Excel のロックファイル {lock_file.name} が在ります")
    # ★★ 2026-08-30（Namakoo「LO を開いた状態で表の更新はできないの？」→ 実測）:
    #   Excel のロック（~$名前）は見ていたのに、**LibreOffice のロック（.~lock.名前#）は
    #   見ていなかった** ── ailine 自身が LibreOffice を使う道具なのに、片方だけ。
    #   ★ 実測: ロックファイルを置いても素通りして書き込めた。
    #   ★ 危ないのは書けることではなく、この順序:
    #       ① 人が LO で開く → ② ailine が書く（成功） → ③ 人が LO 側で保存する
    #       → **開いた時点の古い中身で上書き**され、ailine の変更が黙って消える。
    #     この道具が一番嫌う「静かに失われる」形なので、書く前に止める。
    lo_lock = book.parent / f".~lock.{book.name}#"
    if lo_lock.exists():
        return ("libreoffice", f"LibreOffice のロックファイル {lo_lock.name} が在ります")
    try:
        with open(book, "r+b"):
            pass
    except PermissionError:
        return ("unwritable", "このファイルに書き込めません")
    except OSError:
        return None   # その他の I/O エラーはここでは判定しない（保守的・誤検知回避）
    return None


# --- ★ W8b 項目1: 往復忠実度ゲート -------------------------------------------
#   normalize_book の LO 往復『だけ』で失われる飾り（原本にはあり、正規化後には無い）を
#   検出する。マクロの効果とは無関係（normalize は何もしない空マクロで一度 LO を通す
#   だけの工程）。喪失ゼロなら無言（体験は不変）。喪失があれば --inplace の直前で
#   申告し、--accept-loss / --copy のどちらかを選ばせる（exit 4）。

_FIDELITY_RELS_RE = re.compile(r"^xl/worksheets/_rels/.*\.rels$", re.I)


def _fidelity_zip_members(path: Path) -> set:
    try:
        with zipfile.ZipFile(path) as z:
            return {n.lower() for n in z.namelist()}
    except Exception:
        return set()


def _classify_fidelity_member(name: str) -> str | None:
    """xlsx zip 内の1エントリが『LO 往復で失われがちな飾り』のどのカテゴリか。
       対象外（本文と無関係な構造ファイル等）なら None。"""
    if name.startswith("xl/pivotcache") or name.startswith("xl/pivottables/"):
        return "ピボットテーブル"
    if name.startswith("xl/drawings/"):
        return "図形/描画"
    if name.startswith("xl/media/"):
        return "画像"
    if name == "xl/vbaproject.bin":
        return "VBA マクロ"
    if _FIDELITY_RELS_RE.match(name):
        return "リンク情報(_rels)"
    return None


def count_drawing_objects(path) -> dict:
    """xlsx の中の描画オブジェクトを**中身まで数える**。

    ★ 2026-08-24（土台固め）の実測で開いた穴: 忠実度ゲートは zip の**ファイル名**だけを
    比べていた。openpyxl の往復は xl/drawings/drawing1.xml を**残したまま中の図形だけ**
    捨てるので、ファイル名の集合は 1 つも変わらず、喪失 0 件と報告された（実測: 帳票段が
    雛形の角印を消して ✓ を出した）。★ ファイルが在ることは、中身が在ることではない。

    戻り値: {カテゴリ: 個数}。sp(図形/テキストボックス) と pic(画像) を分けて数える。
    """
    import re as _r
    sp_re = _r.compile("<(?:[a-zA-Z]+:)?sp[ >]")
    pic_re = _r.compile("<(?:[a-zA-Z]+:)?pic[ >]")
    counts = {}
    try:
        with zipfile.ZipFile(path) as z:
            for name in z.namelist():
                if not (name.startswith("xl/drawings/") and name.endswith(".xml")):
                    continue
                text = z.read(name).decode("utf-8", errors="replace")
                n_sp = len(sp_re.findall(text))
                n_pic = len(pic_re.findall(text))
                if n_sp:
                    counts["図形/描画"] = counts.get("図形/描画", 0) + n_sp
                if n_pic:
                    counts["画像"] = counts.get("画像", 0) + n_pic
    except Exception:
        return {}
    return counts


def check_drawing_content_loss(original, produced) -> list:
    """描画オブジェクトが**減った**分をカテゴリ別に返す（増えた分は数えない ──
       帳票段が雛形を N 枚に複製して図形が増えるのは正常）。
       戻り値: [(カテゴリ, 減った数, 元の数, 今の数), ...]。"""
    before = count_drawing_objects(original)
    if not before:
        return []
    after = count_drawing_objects(produced)
    out = []
    for cat, n_before in sorted(before.items()):
        n_after = after.get(cat, 0)
        if n_after < n_before:
            out.append((cat, n_before - n_after, n_before, n_after))
    return out


def check_zip_fidelity_loss(original: Path, normalized: Path) -> list:
    """(a) original の zip 構成要素のうち normalized で消えたものをカテゴリ別に集計する。
       戻り値: [(カテゴリ, 消えた件数), ...]（件数0のカテゴリは含めない・カテゴリ名順）。
       ★ zip として読めない（xlsx でない・壊れている）場合は空リスト（保守的・誤検知しない）。"""
    before = _fidelity_zip_members(original)
    if not before:
        return []
    after = _fidelity_zip_members(normalized)
    lost = before - after
    counts: dict = {}
    for name in lost:
        cat = _classify_fidelity_member(name)
        if cat:
            counts[cat] = counts.get(cat, 0) + 1
    return sorted(counts.items())


def _count_cf_and_dv(wb) -> tuple:
    """ブック全体の(条件付き書式の件数, 入力規則の件数)を合計する。"""
    cf_total = 0
    dv_total = 0
    for name in wb.sheetnames:
        ws = wb[name]
        cf = getattr(ws, "conditional_formatting", None)
        if cf is not None:
            cf_total += sum(1 for _ in cf)
        dv = getattr(ws, "data_validations", None)
        if dv is not None:
            dv_total += len(dv.dataValidation)
    return cf_total, dv_total


def check_openpyxl_fidelity_loss(original: Path, normalized: Path) -> list:
    """(b) openpyxl で条件付き書式/入力規則の件数が正規化後に減っていないかを見る。
       戻り値: [(カテゴリ, before件数, after件数), ...]（減っていないカテゴリは含めない）。
       ★ どちらか一方でも開けない場合は空リスト（保守的・誤検知しない）。"""
    try:
        wb_b = openpyxl.load_workbook(original)
        cf_b, dv_b = _count_cf_and_dv(wb_b)
        wb_b.close()
        wb_a = openpyxl.load_workbook(normalized)
        cf_a, dv_a = _count_cf_and_dv(wb_a)
        wb_a.close()
    except Exception:
        return []
    out = []
    if cf_a < cf_b:
        out.append(("条件付き書式", cf_b, cf_a))
    if dv_a < dv_b:
        out.append(("入力規則", dv_b, dv_a))
    return out


def check_round_trip_fidelity(original: Path, normalized: Path) -> dict:
    """往復忠実度ゲート本体。{"lost": bool, "items": [{"label":str,"count":int}, ...]}。
       ★ history.jsonl の fidelity フィールドにそのまま記録する形（機械可読）。"""
    items = []
    seen_labels = set()
    for cat, n in check_zip_fidelity_loss(original, normalized):
        items.append({"label": cat, "count": n})
        seen_labels.add(cat)
    # ★ 2026-08-24: ファイル名の集合だけでは、**部品が残ったまま中身だけ抜ける**形を
    #   取り逃がす（実測: 帳票段が drawing1.xml を残して中の角印だけ捨て、喪失 0 件で
    #   ✓ が出た）。中の個数も数える。ファイル名側で既に数えたカテゴリは重複させない。
    for cat, n, b, a in check_drawing_content_loss(original, normalized):
        if cat in seen_labels:
            continue
        items.append({"label": cat, "count": n, "before": b, "after": a})
    for cat, b, a in check_openpyxl_fidelity_loss(original, normalized):
        items.append({"label": cat, "count": b - a, "before": b, "after": a})
    return {"lost": bool(items), "items": items}


def format_output_fidelity_warning(fidelity: dict, name: str) -> str:
    """★ 2026-08-24: 「これから失う」予告（format_fidelity_warning）ではなく、
       **もう失った**の報告。人に渡す最終ファイルに対して回す。
       ★ ⚠ で始めるので、決裁③の count_suspicious_advisories が拾って ✓ を △ に降ろす。"""
    parts = "・".join(f"{it['label']} {it['count']} 件"
                      for it in fidelity.get("items", []))
    return (f"⚠ {name} では、元のファイルにあった飾りが失われています（{parts}）"
            f"{chr(10)}  → この処理では保てません。飾りが要る書類は、元のファイルを"
            " LibreOffice や Excel で直接編集してください")


def format_fidelity_warning(fidelity: dict) -> str:
    """人間可読の申告文（例:「⚠ このファイルには、処理すると失われる飾りがあります
       （条件付き書式 3 件・図形/描画 1 件）」）。"""
    parts = "・".join(f"{it['label']} {it['count']} 件" for it in fidelity.get("items", []))
    return f"⚠ このファイルには、処理すると失われる飾りがあります（{parts}）"


# --- ★ W8b 項目4: アトミック置換（--inplace の torn-write 窓の根治） --------------

def _is_our_scratch_output(book: Path, out_book: Path) -> bool:
    """out_book を「今回の run の作業ファイル」として消してよいか。

    ★ 2026-08-26（復元の致命1）: 原本反映の run は `<stem>.out.xlsx` を作業ファイルに
      使うが、**`--copy` の成果物と同じ名前**なので、消すと人の成果物が消える。
    ★ 判定には三項が要る（依頼／宣言／実体）。ここで使えるのは実体だけ ──
      **この run が始まる前から在ったか**を見る。前から在ったなら、それは前回の
      `--copy` の成果物か人の物で、今回の作業ファイルではない。
    ★ 印は使えない（単一ブック経路の .out には印が付かない ── 実測済み）。
    """
    stamp = getattr(_is_our_scratch_output, "_pre_existing", None)
    if stamp is None:
        return True          # 記録が無い（旧経路）── 従来どおり
    return str(out_book.resolve()) not in stamp


def note_pre_existing_output(out_book: Path) -> None:
    """run の**開始時**に、出力先が既に在ったかを控える（消してよいかの実体の項）。"""
    seen = getattr(_is_our_scratch_output, "_pre_existing", None)
    if seen is None:
        seen = set()
        _is_our_scratch_output._pre_existing = seen
    if Path(out_book).exists():
        seen.add(str(Path(out_book).resolve()))


def atomic_replace_inplace(book: Path, out_book: Path, workdir: Path,
                            keep_backups: int = DEFAULT_KEEP_BACKUPS) -> tuple:
    """--inplace の実体。(ok: bool, error_message: str|None)。ok=False の場合、
       out_book はそのまま残り、原本(book)は無変更（呼び出し側が『--inplace は中止』
       として表示する）。
       ★ 手順: ①バックアップ（失敗したら即中止・原本は触らない）
       ②原本と同じボリューム上の staging(workdir/staged<suffix>)へコピー
       ③os.replace(staging, book) で原子的に置換（POSIX rename(2)/Windows
       MoveFileEx の原子性保証＝torn write の窓が無い。同一ボリュームである
       ことは staging を book と同じ親フォルダ配下(workdir)に置くことで保証する）。
       os.replace が失敗したら（バックアップは既に確保済みの上で）shutil.copy2
       へフォールバックし、その旨を1行表示する。成功時は out_book を削除する
       （原本に反映済みの内容と同じものを残しておく理由が無い＝旧 shutil.move
       と同じ最終状態）。"""
    try:
        make_backup(book, keep=keep_backups)
    except Exception as e:
        return False, f"バックアップに失敗したため --inplace を中止した（原本は無変更）: {e}"

    def _discard(p: Path) -> None:
        try:
            if p.exists():
                p.unlink()
        except OSError:
            pass

    staging = workdir / f"staged{book.suffix}"
    try:
        shutil.copy2(out_book, staging)
        os.replace(staging, book)
    except OSError as e:
        # ★★ 2026-08-26（復元の盲検 3 回目・致命3）: ここには
        #   `shutil.copy2(out_book, book)` ── **原本へ直接書くフォールバック**が在った。
        #   copy2 は dst を開いた瞬間に切り詰めるので、途中で失敗すると原本が 0 バイトになる。
        #   それで False を返すと、呼び出し側は必ず「原本は変更していません」と印字する
        #   ── **壊した上で無変更を名乗る**。障害注入で再現済み（ENOSPC / size 0 / BadZipFile）。
        #   ★ 根治は「フォールバックを直す」ではなく **原本へ直接書く経路を持たないこと**。
        #     置換は必ず rename で行う（rename は成功か失敗かしかなく、半端な原本を作らない）。
        #   ★ 元のフォールバックの建前はクロスデバイスだった。それは staging の置き場を
        #     **原本と同じフォルダ**に取り直せば消える ── だからもう一度 rename する。
        _discard(staging)
        near = book.parent / f".ailine_staged_{book.stem}{book.suffix}"
        try:
            shutil.copy2(out_book, near)
            os.replace(near, book)
        except OSError as e2:
            _discard(near)
            return False, (f"置換に失敗した（原本は無変更・バックアップは確保済み）: {e2}")
        print("⚠ 作業フォルダからの置換に失敗したため、原本と同じフォルダで置換した"
              f"（原本は壊していません・バックアップは確保済み）: {e}")
    finally:
        _discard(staging)

    # ★★ 2026-08-26（復元の盲検 3 回目・致命1）: ここは無条件に `<stem>.out.xlsx` を消す。
    #   原本反映 run も作業ファイルに**同じ名前**を使うので、`--copy` で作った成果物と
    #   同じ場所を掴んで消していた（undo は book しか守らないので戻せない）。
    #   ★ 8/25 に入れた関所は run の**入口**だけを見ており、この**出口**を見ていなかった
    #     ── 同じ .out を巡る 2 経路のうち片方しか塞いでいない（片配線）。
    #   ★ 入口と同じ判断をここでも書き写さず、**同じ器官に問い合わせる**:
    #     「これは今回の run が作った物か」だけを消す根拠にする。
    if _is_our_scratch_output(book, out_book):
        try:
            if out_book.exists() and out_book != book:
                out_book.unlink()
        except OSError:
            pass
    elif out_book.exists():
        # ★ 消さなかったことを黙らない（出ないことは信号でない）。
        print(f"（{out_book.name} は今回の run より前から在ったので消していません）")
    return True, None


# --- ★ W8b 項目6: グローバル run ロック --------------------------------------

def own_image_name() -> str:
    """自分の実行ファイル名（小文字）。ロックに焼いて、後で同一性の照合に使う。"""
    try:
        return Path(sys.executable).name.lower()
    except Exception:
        return ""


def _pid_alive(pid: int, expect_image: str | None = None) -> bool:
    """PID が生きているか（確実な保証は無いが十分・追加の依存(psutil 等)は増やさない）。
       判定できない場合は「生きている」扱い（安全側＝奪取しない）。

       ★ 2026-08-24: 旧実装は tasklist の出力に **PID の数字が含まれるか**の部分文字列
       判定だった。前の ailine が終わった後にその PID を**無関係な別プロセスが取り直す**と
       「まだ生きている」と誤判定し、ロックが居座って直後の run が exit 6 になる。
       ★ 根: 判定に要る三項（誰のロックか／今その PID は誰か／同じか）のうち、
       **誰のロックか**を持っていなかった。取得時に実行ファイル名を焼いておけば、
       ここで推測（「python っぽい名前か」）をしなくて済む。
       expect_image が None（古い形式のロック）なら従来どおりの緩い判定に落ちる。"""
    if os.name == "nt":
        try:
            out = subprocess.run(["tasklist", "/FI", f"PID eq {pid}", "/FO", "CSV", "/NH"],
                                  capture_output=True, text=True, encoding="utf-8", errors="replace")
            text = (out.stdout or "").strip()
            if not text or "INFO:" in text:
                return False           # 一致するタスクが無い＝死んでいる
            first = text.splitlines()[0]
            if expect_image:
                image = first.split('","')[0].lstrip('"').lower()
                return image == expect_image
            return str(pid) in first
        except Exception:
            return True
    else:
        try:
            os.kill(pid, 0)
            return True
        except ProcessLookupError:
            return False
        except OSError:
            return True


def _read_lock_info(path: Path) -> dict | None:
    """ロックファイルの**説明**（誰が持っているか）を読む。判定には使わない。

    ★ 2026-08-24: ファイル全体を読むと、末尾に掛けた OS の範囲ロックを跨いで
    PermissionError になる（Windows は自分からも読めなくする）。実測では、
    盲検の画面に `pid=?・?` と出て**説明が説明になっていなかった**。
    説明が書かれている範囲だけを読む。
    """
    try:
        return json.loads(_lock_info_path(path).read_text(encoding="utf-8")) or None
    except Exception:
        return None


def _lock_is_stale(info: dict | None) -> bool:
    """既存ロックが奪取してよい(stale)か。①壊れた/読めないロック ②pid が自分自身
       （同一プロセス内の前回呼び出しが解放し損ねた・テスト等で頻出）
       ③pid が既に存在しない ④30分超、のいずれか。"""
    if info is None:
        return True
    pid = info.get("pid")
    if not isinstance(pid, int):
        return True
    if pid == os.getpid():
        return True
    if not _pid_alive(pid, expect_image=info.get("image")):
        return True
    try:
        ts = datetime.fromisoformat(info.get("ts", ""))
        age = (datetime.now(timezone.utc) - ts).total_seconds()
        if age > RUN_LOCK_STALE_SECONDS:
            return True
    except Exception:
        pass
    return False


# ★ 2026-08-24（初回体験の盲検・致命③）: ロックの持ち主判定を **OS に持たせる**。
#
# 経緯: 旧版は「PID が生きているか」で判定していた。同じ日の夕方に
#   「誰のロックか（実行ファイル名）」を焼く三項化をしたが、**まだ足りなかった** ──
#   記録したのは `python.exe` で、それはこの機械で走る**あらゆる python** に一致する。
#   テストで python を大量に起動するので PID の使い回しが日常的に起き、
#   死んだ ailine のロックが「生きている」と誤判定されて exit 6 で人を止めた
#   （盲検の使い手が 2 回踏んだ・どちらも PID は実在しなかった）。
#
# ★ 根治: 生死を**推測しない**。ロックファイルに OS の排他ロックを掛け、
#   プロセスが死んだら OS が必ず解放する。掛けられれば空き、掛けられなければ誰かが居る。
#   pid/ts/image は今も書くが、それは**人へ見せる説明**専用で、判定には一切使わない。
_RUN_LOCK_HANDLE = None   # プロセスの生存期間ずっと開いておく（閉じると解放される）
# ★ 鍵と説明書きを**別ファイル**にする。同じファイルでロックと読み書きを両立させようと
#   すると、Windows の msvcrt.locking が掛けた範囲を自分からも読めなくするため、
#   バイト位置の細工が要って複雑になった（真夜中に一度そこへ迷い込んだ）。
#   run.lock = OS の鍵だけ / run.lock.info = 人が読む説明だけ。どちらも素直になる。
def _lock_info_path(lock_path) -> Path:
    return Path(str(lock_path) + ".info")


def _try_os_lock(fd) -> bool:
    """fd に排他ロックを掛けられたか。掛けられなければ誰かが持っている。"""
    try:
        if os.name == "nt":
            import msvcrt
            os.lseek(fd, 0, os.SEEK_SET)
            msvcrt.locking(fd, msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        return True
    except OSError:
        return False


def _release_os_lock(fd) -> None:
    try:
        if os.name == "nt":
            import msvcrt
            os.lseek(fd, 0, os.SEEK_SET)
            msvcrt.locking(fd, msvcrt.LK_UNLCK, 1)
        else:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_UN)
    except OSError:
        pass


def acquire_run_lock(path: Path | None = None) -> tuple:
    """(acquired: bool, message: str|None)。ailine run 全体で 1 本の実行ロック。

    ★ 基盤の LibreOffice は単一インスタンス(port 2002)前提なので、ブック単位でなく
    プロセス全体で 1 本にする。

    ★ 2026-08-24: 判定を **OS の排他ロック**に置き換えた（PID の生死を推測しない）。
    プロセスが死ねば OS が必ず解放するので、居座りが原理的に起きない。
    ファイルの中身（pid/ts）は**人へ見せる説明**にだけ使う ── 判定には使わない。
    """
    global _RUN_LOCK_HANDLE
    p = path or RUN_LOCK_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    if _RUN_LOCK_HANDLE is not None:
        # ★ 2026-08-24 の自分の穴: 「非 None なら持っている」とだけ見ていたので、
        #   **別のパスのロックを持っていても通していた**（検体を並べて走らせて発覚）。
        #   同一プロセスの二度取りを許すのは、同じ鍵に対してだけ。
        if Path(_RUN_LOCK_HANDLE[1]) == Path(p):
            return True, None
        return False, f"このプロセスは既に別の実行ロックを持っています（{_RUN_LOCK_HANDLE[1]}）"

    try:
        fd = os.open(str(p), os.O_CREAT | os.O_RDWR)
    except OSError as e:
        return False, f"実行ロックを作れません（{p}）: {e}"

    # 空ファイルには範囲ロックを掛けられないので 1 バイト用意する
    try:
        if os.fstat(fd).st_size == 0:
            os.write(fd, b" ")
    except OSError:
        pass

    if not _try_os_lock(fd):
        holder = _read_lock_info(p) or {}
        os.close(fd)
        pid = holder.get("pid", "?")
        ts = holder.get("ts", "?")
        return False, (f"別の ailine が実行中です（pid={pid}・{ts}）。"
                       "終わるのを待ってから、もう一度実行してください")

    # 取れた ── 誰の物かを書き直す（説明用）
    try:
        info = {"pid": os.getpid(),
                "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "image": own_image_name()}
        _lock_info_path(p).write_text(json.dumps(info, ensure_ascii=False), encoding="utf-8")
    except OSError:
        pass
    _RUN_LOCK_HANDLE = (fd, p)
    return True, None


def under_run_lock(fn):
    """原本を書き換える処理を、グローバル実行ロックで挟む。

    ★ 2026-08-25（復元の重大8・盲検）: ロックを取っていたのは `run` だけで、
      **undo / restore は原本を書くのに取っていなかった**。実測:
          run  → exit=6  × 別の ailine が実行中です
          undo → exit=0  ✓ … から復元した        ← 素通り
      run 実行中の undo は、run 末尾の atomic_replace_inplace に上書きされるうえ、
      その run が「復元したばかりの内容」を世代として積む ── 致命1 の引き金を自分で引く。
    ★ 同じ形を書き写さない: run も undo も restore も、この 1 つを通る。
    """
    acquired, msg = acquire_run_lock()
    if not acquired:
        print(f"× {msg}")
        return 6
    try:
        return fn()
    finally:
        release_run_lock()


def release_run_lock(path: Path | None = None) -> None:
    """ロックを解放してファイルを消す。★ 解放し忘れてもプロセスの終了で OS が外す。"""
    global _RUN_LOCK_HANDLE
    if _RUN_LOCK_HANDLE is None:
        return
    fd, p = _RUN_LOCK_HANDLE
    _RUN_LOCK_HANDLE = None
    try:
        os.lseek(fd, 0, os.SEEK_SET)
    except OSError:
        pass
    _release_os_lock(fd)
    try:
        os.close(fd)
    except OSError:
        pass
    for target in (p, _lock_info_path(p)):
        try:
            target.unlink()
        except OSError:
            pass



def maybe_show_notice_v2(path: Path | None = None) -> bool:
    """★ W10a 項目2: 既定変更(原本直接反映)の一度きり告知。marker ファイルが無ければ
       表示して作る（以後は無言）。戻り値は「今回表示したか」（テスト用）。
       marker の読み書き失敗（権限等）で run 本体を落とさないよう、書き込み失敗は無視する
       （その場合は次回また表示されうる＝安全側に倒す）。
       ★ path 省略時は HISTORY_FILE.parent（＝実運用では NOTICE_V2_FILE と同じ
       ~/.ailine/）から求める（固定の NOTICE_V2_FILE を直接使わない）。実測: この関数は
       cmd_run のたび無条件に呼ぶため、HISTORY_FILE を monkeypatch 済みの大量のテストが
       それに便乗して安全になる（このファイルは backups と違って『一度書いたら消えない』
       ため、実ユーザーの ~/.ailine/notice_v2_shown をテストが汚すと『初回表示』が二度と
       出せなくなる＝汚染の被害が他の一時ファイルより重い）。"""
    p = path or (HISTORY_FILE.parent / "notice_v2_shown")
    if p.exists():
        return False
    print(NOTICE_V2_TEXT)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(datetime.now(timezone.utc).isoformat(timespec="seconds"), encoding="utf-8")
    except OSError:
        pass
    return True


# ---------------------------------------------------------------------------
# doctor（M1: セットアップ診断）
# ---------------------------------------------------------------------------

def _load_module_from_path(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _check_python_version() -> tuple:
    ok = sys.version_info >= (3, 10)
    detail = "" if ok else f"現在 {sys.version.split()[0]}。3.10 以上へ更新して"
    return ok, detail


def _check_openpyxl() -> tuple:
    try:
        import openpyxl as _op  # noqa: F401 — 到達確認のみ
        return True, ""
    except ImportError:   # noqa: F401 ── 再輸出/在否確認のため残す
        return False, "pip install openpyxl"


def _check_ollama_reachable(timeout: float = 3.0) -> tuple:
    try:
        with urllib.request.urlopen(f"{OLLAMA}/api/tags", timeout=timeout) as r:
            json.load(r)
        return True, ""
    except Exception as e:
        return False, f"`ollama serve` を起動して（{e}）"


def _check_model_available(model: str, timeout: float = 3.0) -> tuple:
    try:
        with urllib.request.urlopen(f"{OLLAMA}/api/tags", timeout=timeout) as r:
            d = json.load(r)
    except Exception:
        return False, "ollama に繋がらないため確認できない（先に ollama 到達を直して）"
    names = {m.get("name") for m in d.get("models", [])}
    if model in names:
        return True, ""
    return False, f"`ollama pull {model}` で取得して"


def _check_libreoffice() -> tuple:
    p = _find_basrun_path()
    if p is None:
        return False, "basrun.py が無いため確認できない"
    try:
        mod = _load_module_from_path(p, "_ailine_basrun_probe")
        d = mod.office_dir()
        return True, str(d)
    except SystemExit as e:
        return False, str(e)
    except Exception as e:
        return False, f"検出に失敗: {e}"


def _check_basrun() -> tuple:
    p = _find_basrun_path()
    if p is None:
        return False, ("環境変数 BASRUN に basrun.py のパスを指定するか、"
                        "https://github.com/namakoo-dev/basrun を clone して"
                        "ailine と同じ階層に置いて")
    return True, str(p)


def _check_demo_dir() -> tuple:
    d = HERE / "demo"
    ok = d.is_dir() and any(d.glob("*.xlsx"))
    detail = "" if ok else f"{d} に .xlsx サンプルが無い"
    return ok, detail


def doctor_checks(model: str = DEFAULT_MODEL) -> list:
    """(項目名, ok, 詳細/直し方) のリスト。判定ロジックだけを持ち、副作用(print)は
       cmd_doctor 側に置く（テストしやすくするため分離）。"""
    return [
        ("python 3.10+", *_check_python_version()),
        ("openpyxl", *_check_openpyxl()),
        (f"ollama 到達 ({OLLAMA})", *_check_ollama_reachable()),
        (f"モデル '{model}'", *_check_model_available(model)),
        ("LibreOffice", *_check_libreoffice()),
        ("basrun.py", *_check_basrun()),
        ("demo/", *_check_demo_dir()),
        # ★ W10a 項目2: 動作可否のチェックではなく設定の告知（常に ok=True）。
        #   既定が原本直接反映に変わったこと(W8b-2〜)を doctor でも確認できるようにする。
        ("既定動作", True, "原本直接（v2〜）。--copy で従来のコピー方式"),
    ]


# ★ W8a 項目5: doctor の内部名（"ollama 到達 (URL)" 等）は operator（事務職）には
#   意味が伝わらない。内部名は残したまま(技術者が追える)、事務向けの一行説明を併記する。
#   prefix 一致（name には URL/モデル名など動的な値が混ざるため完全一致にしない）。
#   ★ ダミー名（テストの "a"/"b" 等）はどれにも一致しないため従来どおり内部名のみ表示する。
_DOCTOR_BUSINESS_NOTES = (
    ("python 3.10+", "実行に必要なプログラム言語が使えます"),
    ("openpyxl", "Excel ファイルを読み書きする部品が使えます"),
    ("ollama 到達", "AI エンジン (ollama) に接続できています"),
    ("モデル", "使う AI モデルの準備ができています"),
    ("LibreOffice", "文書を開いて処理する土台があります"),
    ("basrun.py", "文書に処理を適用する仕組みがあります"),
    ("demo/", "動作確認用のサンプル文書があります"),
)


def _doctor_business_note(name: str) -> str | None:
    for prefix, note in _DOCTOR_BUSINESS_NOTES:
        if name.startswith(prefix):
            return note
    return None


def format_doctor_report(results: list) -> tuple:
    """(表示テキスト, all_ok)。"""
    lines = []
    all_ok = True
    for name, ok, detail in results:
        mark = "✓" if ok else "×"
        note = _doctor_business_note(name)
        shown = f"{name}（{note}）" if note else name
        if ok:
            line = f"{mark} {shown}" + (f" ({detail})" if detail else "")
        else:
            all_ok = False
            line = f"{mark} {shown}" + (f" — {detail}" if detail else "")
        lines.append(line)
    return "\n".join(lines), all_ok


def cmd_doctor(a: argparse.Namespace) -> int:
    text, all_ok = format_doctor_report(doctor_checks(a.model))
    print(text)
    return 0 if all_ok else 1


def cmd_ops(a: argparse.Namespace) -> int:
    """★ `ailine ops`: 頼める操作の一覧。盲検査定 2 本が独立に MISSING の筆頭へ置いた
       「こう頼めばこれができる」の対応表。中身は登録簿から生成する（手書きしない）。
       ★ 2026-08-27（Namakoo「登録はドロップダウンで」）: --json を足した。
         画面が一覧を**持たない**という線を守るため ── 人が読む表を画面側で
         parse させると、そこが 2 つ目の一覧になる（形が変わった日に静かにずれる）。"""
    if getattr(a, "json", False):
        print(json.dumps({"ops": [{"op": op, "label": m["label"],
                                    "category": m["category"],
                                    "says": list(m.get("synonyms") or [])}
                                   for op, m in OP_META.items()]}, ensure_ascii=False))
        return 0
    for line in render_ops_table(OP_META, OP_SCHEMA, _CONFIRM_FIELDS):
        print(line)
    # ★ 第三波 S6: 複数ファイルの入口も見せる（argparse の登録簿から生成・手書きしない）。
    sub_actions = [ac for ac in build_parser()._actions
                    if isinstance(ac, argparse._SubParsersAction)]
    #   引数の形（位置引数の並び）も argparse 本体から取る ── 雛形で書くとずれる。
    def _positional_shape(name: str) -> str:
        sp = sub_actions[0].choices.get(name) if sub_actions else None
        if sp is None:
            return ""
        return " ".join(f"<{ac.metavar or ac.dest}>" for ac in sp._actions
                        if not ac.option_strings and ac.dest != "help")
    pairs = [(ch.dest, ch.help or "", _positional_shape(ch.dest))
              for ac in sub_actions for ch in ac._choices_actions]
    for line in render_folder_routes(pairs):
        print(line)
    return 0


# ---------------------------------------------------------------------------
# 実行履歴（M1: 最小版・将来の需要センサー）
# ---------------------------------------------------------------------------

def build_history_entry(result: dict, book: Path, task: str, model: str, failure_kind: str,
                         error_detail: str | None = None) -> dict:
    """1 run の結果を history.jsonl の 1 行分の dict にする（純ロジック・テスト用に分離）。
       ★ M2a: error_detail は runtime_error 時の生エラー全文（端末には最終行だけ出す
       ため、詳細を追いたい時の唯一の入り口になる）。
       ★ W8a 項目1: "dry" は --dry（見せるだけ・未適用）で走ったかどうか。result["dry"] を
       そのまま bool 化する（result に無ければ False＝実適用と同じ扱い）。旧 history.jsonl
       の行（このキーが無い）は read_history/format_history_table 側で dict.get("dry", False)
       により実適用扱いのまま読める（後方互換・新旧を混在させても壊れない）。
       ★ W8b 項目1: "fidelity" は往復忠実度ゲートの検出結果（--inplace が要求され、
       かつ --dry でない run でのみ実際に計算される）。ゲートを走らせなかった run は
       None のまま（既存キーと同じ形＝無ければ None）。"""
    return {
        "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "book": str(book),
        "task": task,
        "model": model,
        "ok": bool(result.get("ok")),
        "dry": bool(result.get("dry")),
        "attempts": result.get("attempts", 0),
        "failure_kind": failure_kind,
        "error_detail": error_detail,
        # ★ #17: 何で通ったか。過去の成功を思い出す材料（提案までで、実行はしない）。
        "op": result.get("op"),
        "ops": result.get("ops"),
        "changes": (result.get("changes") or [])[:3],
        "out": result.get("out"),
        # ★★ 2026-08-26（復元の盲検 3 回目・致命2）: `out_sha` は _finish_run が確かに
        #   作っていたのに、**ここがキーを固定列挙していて写していなかった**。
        #   結果 history 全行で欠落 → 関所は `stamped is None` で常に素通り →
        #   `_refuse_edited_output` は**到達不能**だった（＝あの直しは一度も発火していない）。
        #   ★ 番人が通した理由の方が重い: 検体が history の行を**手で書いて**いて、
        #     本番の書き手（この関数）を一度も通っていなかった ── 継ぎ目を跨いでいない。
        "out_sha": result.get("out_sha"),
        # ★ M2b: DSL 経路(path="dsl")では命令言語の確認文(command)と事後条件の合否を残す。
        #   自由生成経路(path="freeform")では両方 None のまま（既存キーは不変）。
        "path": result.get("path", "freeform"),
        "command": result.get("command"),
        "postcondition": result.get("postcondition"),
        # ★ A': 用語集/依頼文から機械確定した値（APPEND_TOTAL の倍率等）の出典。無ければ None。
        "provenance": result.get("provenance"),
        "fidelity": result.get("fidelity"),
    }


def _file_digest(path: Path) -> str | None:
    """ファイルの指紋（sha256 先頭 16 桁）。読めなければ None ── 無いことは信号でない。"""
    try:
        return hashlib.sha256(Path(path).read_bytes()).hexdigest()[:16]
    except OSError:
        return None


def append_history(entry: dict, path: Path | None = None) -> None:
    """history.jsonl に 1 行 append する。★ 失敗したら例外を投げる（run 本体を落とさ
       ないための try は呼び出し側(cmd_run)が持つ。ここでは書き込みロジックだけ）。"""
    p = path or HISTORY_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def op_that_worked_before(task: str, entries) -> tuple:
    """同じ依頼文で**過去に成功した** op と、その日付を返す（無ければ (None, None)）。

    ★ 2026-08-24（#17）: 同じブックに一字一句同じ依頼を 2 回投げると、1 回目は通り
    2 回目は「頼める操作の一覧に照合できませんでした」で落ちた（LLM のサンプリングの揺れ）。
    ★ 直す対象は揺れそのものではない（temperature を 0 にしても別の入力で揺れる）。
    直すのは「揺れたときに人が困る」方 ── 道具が**自分の過去を思い出す**。

    誤爆を作らない条件（実装前に決めて凍結した）:
      - 一致は **依頼文の完全一致**のみ（部分一致だと別の依頼に前回の op を当てる）
      - **成功した run のみ**を材料にする（失敗した run の op を勧めない）
      - 材料が無ければ黙る（初回・別 PC で挙動が変わらない）
      - ★ 黙って実行しない ── 呼び出し側は候補として並べ、根拠（日付）を必ず言う
    """
    want = (task or "").strip()
    if not want:
        return None, None
    # ★ 渡された並び順に依存しない（read_history は新しい順・ファイルは古い順で、
    #   初版は reversed() を掛けて**最古の成功**を拾っていた ── 自分で用意した
    #   古い順の検体で試したので気づけなかった）。ts で選ぶ。
    hits = []
    for entry in (entries or []):
        if not isinstance(entry, dict) or not entry.get("ok"):
            continue
        if (entry.get("task") or "").strip() != want:
            continue
        op = entry.get("op")
        if not op or op == "PLAN":     # PLAN は 1 op に決まらないので勧めない
            continue
        hits.append((str(entry.get("ts") or ""), op))
    if not hits:
        return None, None
    ts, op = max(hits)
    return op, ts[:10]


def read_history(path: Path | None = None, max_n: int = 10) -> list:
    """新しい順に最大 max_n 件を返す。壊れた行は読み飛ばす。"""
    p = path or HISTORY_FILE
    if not p.exists():
        return []
    entries = []
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            entries.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return list(reversed(entries))[:max_n]


def append_misclass(entry: dict, path: Path | None = None) -> None:
    """misclass.jsonl に 1 行 append する（誤分類の実例台帳センサ本体）。★ history と
       同じ流儀 ── 失敗したら例外を投げる（run を壊さないための try は呼び出し側が持つ）。"""
    p = path or MISCLASS_FILE
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def _record_misclass_suspect(signal: str, task: str, plan, book) -> None:
    """誤分類容疑を1行記録する。①②の記録点が共通で呼ぶ薄い配線。書き込み失敗は run を
       壊さない（history の WARN と同じ作法・失敗しても標準エラーに1行出すだけ）。"""
    try:
        append_misclass({
            "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "signal": signal,
            "task": task,
            "plan": plan,
            "book": str(book) if book is not None else None,
        })
    except Exception as e:
        print(f"WARN: 誤分類容疑の記録に失敗した: {e}", file=sys.stderr)


def _last_task_for_book(book: Path) -> str:
    """②undo が使う: book に一致する history.jsonl の直近エントリの task。
       無ければ空文字列（記録を落とさないための穏やかなフォールバック）。"""
    for e in read_history(max_n=1000):
        if e.get("book") == str(book):
            return str(e.get("task", ""))
    return ""


def format_history_table(entries: list) -> str:
    """人が読める表形式。履歴が無ければ「履歴はまだ無い」を返す。
       ★ W8a 項目1: dry(下見・未適用)の行は末尾に「(下見)」を付けて実適用と区別する
       （「dry-run」は事務の言葉ではないため表示には出さない）。dict にキーが無い旧行は
       dict.get("dry", False) で実適用扱いのまま読める（後方互換）。"""
    if not entries:
        return "履歴はまだ無い"
    header = f"{'日時':<20} {'結果':<4} {'試行':<4} {'モデル':<20} {'文書':<20} タスク"
    lines = [header]
    for e in entries:
        mark = "✓" if e.get("ok") else "×"
        ts = str(e.get("ts", ""))
        attempts = str(e.get("attempts", ""))
        model = str(e.get("model", ""))
        book = Path(str(e.get("book", ""))).name
        task = str(e.get("task", ""))
        line = f"{ts:<20} {mark:<4} {attempts:<4} {model:<20} {book:<20} {task}"
        if e.get("dry", False):
            line += "  (下見)"
        kind = e.get("failure_kind")
        if not e.get("ok") and kind not in (None, "none"):
            # ★ W10 前提工事①: failure_kind は記録側で「語彙外/out_of_vocab」等に
            #   下位区分したが、表示は従来どおり上位ラベル「語彙外」に畳む
            #   （表示互換のため・区分は history.jsonl の生データ側で見る）。
            shown = kind.split("/", 1)[0] if isinstance(kind, str) else kind
            line += f"  [{shown}]"
        lines.append(line)
    return "\n".join(lines)


def cmd_history(a: argparse.Namespace) -> int:
    entries = read_history(max_n=a.max)
    print(format_history_table(entries))
    return 0


# --- ★ A': 用語集(vocab) コマンド --------------------------------------------

def cmd_vocab(a: argparse.Namespace) -> int:
    """`ailine vocab add <語> <値>` / `ailine vocab list`。
       ★ remove は作らない（今の実需では add/list の2本で足りる。壊れたら vocab.json を
       直接編集すればよい）。"""
    if a.vocab_cmd == "add":
        ok, msg = vocab_add(a.term, a.value)
        print(render_vocab_add_result(ok, msg))
        return 0 if ok else 1
    # list
    vocab = load_vocab()
    for ln in render_vocab_listing(vocab, VOCAB_FILE):
        print(ln)
    return 0


# --- ★ W10 便A: 別名ストア(alias) コマンド ------------------------------------

def cmd_alias(a: argparse.Namespace) -> int:
    """`ailine alias add <言い回し> <OP>` / `alias list` / `alias remove <言い回し>` /
       `alias undo`。cmd_vocab の様式の写経 ── ただし vocab.json 側の remove 不在の決定
       （5307行付近のコメント）はここでは触らない: 別名は機械が書く層なので、こちらだけ
       remove/undo（直近の登録の取り消し）を持つ（設計ノート④）。"""
    if a.alias_cmd == "add":
        ok, msg = alias_add(a.phrase, a.op)
        print(render_vocab_add_result(ok, msg))
        return 0 if ok else 1
    if a.alias_cmd == "remove":
        ok, msg = alias_remove(a.phrase)
        print(render_vocab_add_result(ok, msg))
        return 0 if ok else 1
    if a.alias_cmd == "undo":
        ok, msg = alias_undo()
        print(render_vocab_add_result(ok, msg))
        return 0 if ok else 1
    # list
    aliases, order = load_aliases()
    for ln in render_alias_listing(aliases, order, ALIASES_FILE):
        print(ln)
    return 0


# ---------------------------------------------------------------------------
# run コマンド本体
# ---------------------------------------------------------------------------

def _finish_run(a: argparse.Namespace, book: Path, result: dict, failure_kind: str,
                 error_detail: str | None = None) -> None:
    """--json 出力・成功時の注意書き・履歴の記録。cmd_refuse_vocab_miss / cmd_run_dsl /
       cmd_run_plan の共通末尾。
       ★ DSL 経路(path="dsl")・複合計画経路(path="plan")は達成/総合判定の行を既に自分で
       出しているので、success_message() の『正しいかは差分を見て判断』（自由生成向けの
       注意書き）はここでは出さない。"""
    # ★ W8b 項目1: 往復忠実度ゲートの検出結果（cmd_run 側で計算済みなら a._fidelity に
    #   乗っている）を history に写す。ゲートを走らせなかった run は None のまま。
    if "fidelity" not in result:
        result["fidelity"] = getattr(a, "_fidelity", None)
    # ★ C9: --json の新キー。既存キーの意味は変えず、「何と照合し・どのファイルを読み戻して
    #   ✓ と言ったか」を足すだけ。✓ を出さなかった run（--dry・失敗・機械保証なし）は空リスト
    #   ＝『主張していない』が機械可読に残る。
    result.setdefault("claims", [])
    # ★ 適用まで行かなかった run（--dry・失敗・断り）は**何も主張していない**。
    #   キーは必ず在って、値でそれが分かる形にする（無いことで伝えない）。
    result.setdefault("verdict", "not_applied")
    result.setdefault("warning_count", 0)
    # ★ 2026-08-25（復元の致命③）: 出力先の関所が「この道具が過去にそこへ書いたか」
    #   だけで通していたので、**利用者がその後どれだけ手を入れても素通り**した。
    #   判定には三項が要る（依頼/宣言/実体）── 実体の項として、書いた物の指紋を残す。
    if result.get("out"):
        result["out_sha"] = _file_digest(Path(result["out"]))
    if a.json:
        print("\n" + json.dumps(result, ensure_ascii=False))
    if result.get("path") not in ("dsl", "plan"):
        msg = success_message(result)
        if msg:
            print("\n" + msg)
    try:
        detail = error_detail if error_detail is not None else (
            result.get("last_error_full") if failure_kind == "runtime_error" else None)
        append_history(build_history_entry(result, book, a.task, a.model, failure_kind,
                                            error_detail=detail))
    except Exception as e:
        print(f"WARN: 履歴の記録に失敗した: {e}", file=sys.stderr)


# ★ C9: ✓ の唯一の根拠。適用が全部終わった「最終ファイル」を openpyxl で開き直し、
#   機械が読み取れる事実だけで「今このファイルはこうなっている」を述べる。
#   ★★ ここで事後条件を再実行しない（明示的に否定された設計）: APPEND_TOTAL の後に SORT が
#   来れば合計行の位置は正当に動き（check_append_total は "=SUM(" の初出行で合計行を探す）、
#   再実行すると正しい run が偽 fail になる。読み戻すのは「反映が成功したこと」だけでよい。
def observe_book_state(path: Path) -> tuple:
    """(evidence: str|None, error: str|None)。全シート・全行を走査する（切り詰めない＝
       Claim.observation_complete=True の根拠）。evidence は最終ファイルだけから独立に
       再導出できる事実（シート名・行数・列数・値のあるセル数）に限る。"""
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return None, short_error_summary(str(e))
    try:
        parts = []
        for ws in wb.worksheets:
            filled = sum(1 for row in ws.iter_rows() for c in row if c.value not in (None, ""))
            parts.append(f"{ws.title}: {ws.max_row}行×{ws.max_column}列・値のあるセル {filled}")
        return "・".join(parts), None
    finally:
        wb.close()


# ★ W8b-2 項目1: 既定=原本直接適用の終端メッセージを一箇所に集約する
#   （cmd_run_dsl/cmd_run_freeform/cmd_run_plan の3箇所が同じ形だったのを統合）。
#   pending/confirm の中間状態は作らない — undo 一本（architect 判定）。
# ★★ C9: 『✓』の発生点をここ1箇所へ動かした。原本(--copy なら .out)が確定した後に
#   読み戻し、その結果だけを Claim にして描く。段別 ✓・--dry の ✓・反映前の ✓ は廃止。
def _finish_failed_apply(a: argparse.Namespace, book: Path, result: dict) -> int:
    """反映に失敗して終わる時の、唯一の出口。

    ★★ 2026-08-26（Namakoo が実測・盲検の中9 と同じ根）: ここは
      `return EXIT_APPLY_FAILED` で **_finish_run を通らずに抜けて**いた。
      その結果:
        ・「作業結果は <名前>.out.xlsx に残っています」と**言っておきながら**、
          その物を作ったことを**履歴に記録していなかった**
        ・次の run は出力先の関所で「この道具が書いた記録がありません」と塞がれる
        ・人が手でファイルを消すまで、その本には**二度と実行できない**（行き止まり）
      ★ 自分の言葉と自分の記録が食い違うと、自分の関所が自分を締め出す。
      ★ 4 箇所に書き写されていたので、1 つの器官に畳んだ（また片配線を作らない）。
    """
    _finish_run(a, book, result, "apply_failed")
    return EXIT_APPLY_FAILED


def _finish_apply(a: argparse.Namespace, book: Path, out_book: Path, workdir: Path,
                   result: dict, machine_verified: bool, scope: str = "",
                   scope_note: str = "", warning_count: int = 0) -> bool:
    """--copy（a.inplace が False）なら .out のまま（原本は無変更）。既定(a.inplace)なら
       backup+原子的置換(atomic_replace_inplace)で原本へ反映する。そのうえで**最終ファイルを
       読み戻し**、machine_verified=True なら ✓（★ 決裁③: warning_count>0 なら△に降格）の
       1行を、False（自由生成・検証対象不足の段を含む計画）なら ⚠ の1行を出す。
       読み戻せなかったら ✓/△ どちらも出さない。
       scope は照合した宣言（Claim.scope・machine_verified=True のとき必須）、
       scope_note は経路別の範囲注記（単発/複合計画）。
       ★ 決裁③(2026-08-22): warning_count は呼び出し側が数えた「疑わしい系の ⚠」の総数
       （count_suspicious_advisories・単一 op/複合計画それぞれの advisories + 単位F/G の
       前提破れメッセージから求める。呼び出し側が数えるのは、この関数自身は advisories の
       出所（単発 build_advisories 系 / 複合計画の段別集計）を知らないため ── 判定基準
       （★ 付きだけ数える）は claim.py 側に1箇所で持つ）。1件でもあれば ✓ を出さず△にする
       （machine_verified=False の ⚠ 経路には影響しない ── そちらは元から ✓ を名乗らない）。
       戻り値: 置換が成功した(または --copy で置換不要だった)か。"""
    # ★ 忠実度は**置換より前**に測る（book がまだ原本・out_book が成果物）。
    #   --copy でも --inplace でも成果物は out_book なので、1 本の測定で両経路を覆う。
    _output_fidelity = check_round_trip_fidelity(book, out_book)
    # ★★ 2026-08-25（復元の中10・盲検）: 原本に被せる**前**に「そもそも開けるか」を見る。
    #   旧版は反映の関門に「開ける xlsx か」の検査が無く、壊れた成果物（zip として
    #   読めない等）をそのまま原本へ被せてから「読み戻して確認できませんでした」と
    #   言っていた ── **確認は原本を潰した後**だった。報告は正直だが順序が逆。
    #   ★ 原本はまだ無傷なこの位置でしか止められない。
    if a.inplace:
        broken = _why_output_is_unusable(out_book)
        if broken:
            print(f"× 作った結果が壊れているため、原本には反映しませんでした（{broken}）")
            print(_untouched_original_line(book, out_book))
            result["out"] = str(out_book)
            return False
    if a.inplace:
        ok_ip, err_ip = atomic_replace_inplace(
            book, out_book, workdir, keep_backups=getattr(a, "keep_backups", DEFAULT_KEEP_BACKUPS))
        if not ok_ip:
            print(f"× {err_ip}")
            print(_untouched_original_line(book, out_book))
            result["out"] = str(out_book)
            return False
        final, trailer = book, "（もとに戻す: ailine undo）"
        result["out"] = str(book)
    else:
        final, trailer = out_book, f"（原本 {book.name} は変更していません）"
        result["out"] = str(out_book)

    # ★ 2026-08-24（土台固め）: 忠実度ゲートを **人に渡す最終ファイル**にも回す。
    #   旧: 正規化の直後（原本 vs LO を通した直後）だけを見ていた。しかも --inplace の
    #   時だけ。実測した穴は 3 つとも同じ根だった:
    #     ① 帳票段が検分シートを openpyxl で足す往復で、雛形の角印を全枚から消す
    #     ② 様式写像段が同じ形（コード同一）
    #     ③ --copy はゲートを素通りし、出力から VBA が消えたまま ✓ を出す
    #        ── しかも --copy は**ゲート自身が案内する逃げ道**だった
    #   ★ ここは ✓ を出す唯一の choke point なので、ここで見れば経路を問わず塞がる
    #     （新しい op を足しても自動で守られる ── 個別に配線しない）。
    #   ★ 見るのは「消えたもの」だけ（増えた分は数えない）── 帳票段が N 枚足すのは正常。
    # ★ 恒真を切る: --inplace では atomic_replace_inplace の後、book は**結果そのもの**に
    #   なる（自分と自分を比べれば喪失は永遠に 0）。だから測定はこの関数の入口
    #   （置換前）で済ませてある ── ここでは結果を読むだけ。
    lost_in_output = _output_fidelity
    if lost_in_output.get("lost"):
        result["output_fidelity"] = lost_in_output
        # ★ --accept-loss は「失ってよい」と**利用者が先に選んだ**場合。事実は必ず言うが、
        #   選んだ人を ⚠ で責め直して ✓ を降ろすのは筋が違う ── 中立の報告にする。
        #   選んでいない経路（--copy 等・ゲートが走らないまま失う）は従来どおり ⚠。
        if getattr(a, "accept_loss", False):
            parts = "・".join(f"{it['label']} {it['count']} 件"
                               for it in lost_in_output.get("items", []))
            print(f"（承知のうえで飾りを失いました: {parts}）")
        else:
            msg = format_output_fidelity_warning(lost_in_output, final.name)
            print(msg)
            warning_count += count_suspicious_advisories([msg])

    evidence, err = observe_book_state(final)
    # ★★ 2026-08-26: 判定（✓ / △ / ⚠）を**決めているこの場所で**機械可読にも出す。
    #   ★ なぜ: 画面に出る印は今まで**文字としてしか存在しなかった**ので、別の入口
    #     （GUI・自動化）から使うには印を読み取り直すか、条件を書き写すしかなかった。
    #     書き写せばそれは 2 つ目の実装で、今週ずっと潰してきた欠陥をこちらで新造する。
    #   ★ 決めるのは 1 箇所・映すのは何箇所でも、という形にする。
    #   verdict の意味:
    #     "verified"   … 機械検証済み（✓）
    #     "warned"     … 検証はしたが疑わしい ⚠ が在る（△・決裁③の降格）
    #     "unverified" … 機械保証なし（自由生成・検証対象不足）
    #     "unobservable" … 適用したが読み戻せなかった（何も保証しない）
    result["verdict"] = ("unobservable" if err is not None else
                          ("warned" if warning_count > 0 else "verified") if machine_verified
                          else "unverified")
    result["warning_count"] = warning_count
    if err is not None:
        for ln in render_applied_unobservable(final.name, err):
            print(ln)
    elif machine_verified:
        claim = Claim(verified=True, basis="declaration", scope=scope, evidence=evidence,
                       observation_complete=True, observed_on=str(final), observed_after_apply=True)
        # ★ 決裁③(2026-08-22): 疑わしい ⚠ が1件でも出た run は ✓ を名乗らない（△ に降格）。
        render_fn = render_applied_claim_demoted if warning_count > 0 else render_applied_claim
        lines = render_fn(claim, final.name, warning_count) if warning_count > 0 else render_fn(claim, final.name)
        for ln in lines:
            print(ln)
        if scope_note:
            print(scope_note)
        # ★ --json: 既存キーの意味は変えず、claims を足すだけ（何と照合し・どのファイルを
        #   読み戻して言っているのかを機械可読にする）。
        result["claims"] = [{"basis": claim.basis, "compared_with": claim.scope,
                              "observed_on": claim.observed_on}]
    else:
        for ln in render_applied_unverified(final.name, evidence):
            print(ln)
    print(trailer)
    return True


def likely_cause_of_no_change(book_path, sheet_name=None) -> list:
    """事後条件が破れたとき、**なぜ効かなかったか**の心当たりを述べる行を返す。

    ★ 2026-08-24（土台固め）: 飾りの生存表を作っている最中に、対照実験で確定した ──
    表の範囲に**結合セル**があると LibreOffice の並べ替えが黙って何もしない
    （同じブックから結合セルだけ外すと ✓ になる）。ailine は嘘の ✓ を出さずに
    exit 1 で落ちるので「壊さない」は守れているが、**理由を言わない**ので使う側は
    そこで詰まる ── 今日ずっと直してきた「判定は正しいが理由を言う口が無い」の形。

    ★ 断定しない: 「効かなかった原因はこれ」ではなく「心当たり」として出す
    （結合セルが在っても効く操作はある）。
    """
    lines = []
    try:
        with BookView(book_path) as bv:
            ws = bv.sheet(sheet_name)
            merged = list(ws.merged_cells.ranges)
    except Exception:
        return []
    if merged:
        shown = "・".join(str(m) for m in merged[:3])
        more = f"（ほか {len(merged) - 3} 件）" if len(merged) > 3 else ""
        lines.append(f"  心当たり: シート『{sheet_name or ""}』に結合セルが "
                      f"{len(merged)} 件あります（{shown}{more}）")
        lines.append("  → 表の中に結合セルがあると、並べ替えなどの操作が"
                     "何もせずに終わることがあります（実測）。"
                     "結合を解除してからお試しください")
    return lines


def _why_output_is_unusable(path: Path) -> str | None:
    """成果物が Excel として開けない理由（開けるなら None）。

    ★ 2026-08-25（復元の中10）: 原本へ被せる前の最後の確認。ここで止めれば原本は無傷。
    ★ 中身の正しさは見ない ── それは事後条件の仕事。ここが見るのは「開けるか」だけ。
    ★ 2026-08-25: openpyxl が読めない拡張子（.ods 等）では**調べようがない**。
      調べられないことを「壊れている」と言えば、命綱を丸ごと塞いでしまう
      ── 出ないことは信号でない（逆向きにも同じ）。None を返して黙って通す。
      （今の版では .ods は build_book_meta の時点で落ちるので、ここは将来への保険。）
    """
    if Path(path).suffix.lower() not in OPENPYXL_PROBEABLE_SUFFIXES:
        return None
    try:
        with BookView(path) as bv:
            if not bv.sheetnames:
                return "シートが 1 枚もありません"
    except Exception as e:
        return f"{type(e).__name__}: {e}"
    return None


def _untouched_original_line(book: Path, out_book: Path) -> str:
    """★ C9: 失敗して終わるときに必ず出す1行。査定2本が「原本がどうなったか分からない」と
       書いた沈黙の穴 ―― 途中で止まった run は原本無変更を名乗らず、`.out` が黙って隣に
       残ることも告げていなかった。"""
    return f"（原本 {book.name} は変更していません。作業結果は {out_book.name} に残っています）"


def cmd_run(a: argparse.Namespace) -> int:
    """run コマンドの入口。★ W8b 項目6: 実処理(_cmd_run_body)の前後をグローバル run
       ロックで挟む（基盤の LibreOffice が単一インスタンス前提のため、ブック単位でなく
       ailine run 全体で1本）。取得できなければ即 exit 6（LO 起動・翻訳より前）。
       finally で確実に解放する（sys.exit も SystemExit 例外なので finally は通る）。"""
    return under_run_lock(lambda: _cmd_run_body(a))


def _cmd_run_body(a: argparse.Namespace) -> int:
    """run コマンドの本体。★ W3: 正規化パス(＋StructDump による見出し行推定)を
       翻訳より前に一度だけ行う（『三層全部が同じ見出し推定を使う』ための土台。
       translate_task 自身の接地(book_meta)も検出した見出し行を使う）。
       --dry は従来どおり LibreOffice に触れない（見出しは物理1行目のまま・E2E 対象外）。
       ★ W8a 項目3: --header-row N（1起点）が指定されていれば StructDump 検出を丸ごと
       スキップしてその行を採用する（CLARIFY の行き止まりから抜ける唯一の入り口）。
       ★ W8b 項目2: Excel ロック検出は LO 起動・翻訳より前（一番最初）に行う。
       ★ W8b 項目1: --inplace が要求され --dry でない場合だけ、正規化直後に往復忠実度
       ゲートを見る（原本にまだ一切触れていない段階＝『原本に触る前に申告と選択』）。
       ★ W8b 項目4: 自分の workdir(.ailine_<stem>/) は run の終了時に必ず掃除する
       （成功・失敗とも・GC の作り込みはしない＝自分の後始末だけ）。
       ① 見出し行推定（StructDump） → 自信不足なら CLARIFY して exit 3
       ② 翻訳（計画）→
       - 計画が空/1段で CLARIFY → 質問して exit 3
       - 計画が空/1段で DSL 語彙 → ③〜⑥の決定論パイプライン(cmd_run_dsl)
       - 計画が空/1段でそれ以外(FREEFORM・翻訳失敗) → ★ freeform 最終決定: 即座の断り
         （cmd_refuse_vocab_miss・生成には入らない）
       - 計画が2段以上(複合依頼) → 段ごとに honest な項目別実行(cmd_run_plan)（M2c）
       ★ 後方互換: translate_task が "plan" で包まない旧形式（bare {"op":...}）を返した場合
       （テストの monkeypatch を含む）も、その dict をそのまま単一段として扱う。"""
    # ★ M3 P: run の位置引数の arity 判定を一番最初に一度だけ（フォルダ分岐の隣・
    #   翻訳より前）。task は nargs="+" で受けている ── 先頭が実在ファイルなら2冊目の
    #   パス、残りが依頼文。★ LLM には一切渡さない判定（M8: prompt 3 ハッシュ不変の根拠）
    #   ── 実在ファイルかどうかの機械の事実だけで分岐する。2冊目のつもりの引数が
    #   実在しなければ、黙って依頼文の一部として単一ブック経路へ流さず名指しで聞く
    #   （凍結検体 test_nonexistent_second_path_is_asked_not_swallowed_as_task）。
    task_tokens = a.task if isinstance(a.task, list) else [a.task]
    book1 = Path(a.book)
    if not book1.is_dir() and book1.is_file() and len(task_tokens) >= 2:
        second_arg = task_tokens[0]
        second_path = Path(second_arg)
        # ★ review5#3 の直し（実機再現: 依頼文の1語目が拡張子なしの実在ファイル名と
        #   偶然一致し、単一ブックの依頼が2冊照合に誤読された）: 2冊目分岐の発火は
        #   「実在」だけでなく「表計算らしい形（拡張子 or パス区切り）」も要る AND 条件。
        #   拡張子なしの偶然一致は表計算らしくない＝発火しない＝従来どおり依頼文として扱う。
        looks_like_path = _looks_like_second_book_path(second_arg)
        if second_path.is_file() and looks_like_path:
            return cmd_run_match(a, book1.resolve(), second_path.resolve(),
                                  " ".join(task_tokens[1:]).strip())
        if looks_like_path and not second_path.is_file():
            print(f"？ 2冊目として指定した『{second_arg}』が見つかりません。"
                  "パスを確認してから、もう一度実行してください"
                  "（1冊だけの依頼なら、依頼文だけを1つの引数として渡してください）。")
            return 3
    a.task = " ".join(task_tokens)   # ★ 以降の全経路は従来どおり a.task を str として読む

    # ★ M2（architect 致命4）: book の位置がディレクトリなら多ファイル分岐へ ── **一番最初**に
    #   分ける。ここから下は1冊のブック前提の器官（ロック検出・正規化・バックアップ・undo）で、
    #   フォルダを渡すと check_excel_lock の open(r+b) が PermissionError になり
    #   「Excel で開かれています」という嘘の診断を返していた（凍結検体あり）。
    if Path(a.book).is_dir():
        return cmd_run_folder(a)
    # ★ CSV 検疫接続（DESIGN-20260821-multifile.md「CSV 検疫 設計 v2」B・暗黙前段）:
    #   .csv は既存の1冊機械（正規化・LO 適用・normalize_book/basrun_apply）に絶対に
    #   渡さない ── フォルダ分岐と同格の、ここが一番最初の分かれ目。
    if Path(a.book).is_file() and Path(a.book).suffix.lower() == CSV_SUFFIX:
        return _cmd_run_csv_prestage(a)
    maybe_show_notice_v2()   # ★ W10a 項目2: 既定変更の一度きり告知（run の一番最初）

    book = Path(a.book).resolve()
    if not book.exists():
        exit_environment(f"文書が無い: {book}")

    # ★ 形式の関所は**ロックより先**（扱えない形式に「Excel で開いています」と
    #   言うと、心当たりが的外れになる）。★ run の入口だけ ── undo には掛けない。
    unusable = refuse_if_run_cannot_handle(book)
    if unusable is not None:
        return unusable

    blocked = refuse_if_locked(book)
    if blocked is not None:
        return blocked

    # ★ 2026-08-25（復元の致命2）: 出力先に**人のファイル**が在れば、触る前に止める。
    #   LO 起動・翻訳より前 ── ロック検出と同じ位置に置く（壊してから気づかない）。
    conflict = refuse_if_output_is_someone_elses(book)
    if conflict is not None:
        return conflict

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    try:
        return _cmd_run_dispatch(a, book, workdir)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _cmd_run_dispatch(a: argparse.Namespace, book: Path, workdir: Path) -> int:
    # ★ W8b-2 項目2: 既定 = 原本直接適用（旧 --inplace 相当）。--copy で旧 .out 挙動を
    #   温存する。--inplace は廃止した旧フラグだが、互換のため受理はして移行メッセージ
    #   だけ出す（指定してもしなくても挙動は同じ＝既定が原本適用）。
    if getattr(a, "inplace", False):
        print("★ --inplace は廃止されました。既定で原本に直接適用します"
              "（従来の .out 挙動が欲しい場合は --copy を使ってください）。")
    a.inplace = not getattr(a, "copy", False)

    apply_timeout = a.timeout if a.timeout else None   # 0 で無効化（旧挙動 = 無制限）

    source_book = book
    struct_dump: dict = {}
    if not a.dry:
        t0 = progress_start("⏳ 初回準備（文書の正規化+構造読み取り）…")
        source_book = normalize_book(book, workdir, timeout=apply_timeout)
        progress_end(t0)
        struct_dump = build_struct_dump(source_book, workdir)

        # ★ W8b 項目1: 原本に実際に適用する時だけ、まだ触れていないこの時点で往復忠実度
        #   を見る（喪失ゼロなら無言・体験は不変）。★ W8b-2 項目4: --copy 時は原本に
        #   一切触れないためゲート自体を走らせない（無駄な zip/openpyxl 比較も省く）。
        if a.inplace:
            fidelity = check_round_trip_fidelity(book, source_book)
            a._fidelity = fidelity
            if fidelity["lost"]:
                print(format_fidelity_warning(fidelity))
                if getattr(a, "accept_loss", False):
                    print("→ --accept-loss 指定のため続行します（失われても ailine undo で元に戻せます）")
                else:
                    for ln in render_retry_options("", [
                            ("--accept-loss", "失われてよい（バックアップから ailine undo で復元可能）"),
                            ("--copy", "原本には触らず .out に結果を作る（原本は無変更）")]):
                        print(ln)
                    return 4

    # ★ 挙動変更#3: 衝突検出のため headers も使う。build_book_meta は元から1回呼んでいて
    #   sheets しか読んでいなかった ── 同じ戻り値の headers を渡すだけ（ブックは開き直さない）。
    probe_meta = build_book_meta(source_book)
    sheets = probe_meta.get("sheets", [])

    # ★ 挙動変更#2: 対象シートの決定はここ1箇所（resolve_target_sheet）。それより後の
    #   全処理（見出し行検出・翻訳→検証→codegen→事後条件）は a._target_sheet を読むだけで
    #   個別に「1枚目」を仮定しない。--header-row/翻訳より前＝原本の実処理が始まる前に
    #   決め、実処理の前に必ず宣言する（査定所見:「これがあれば事故は防げた」）。
    target_sheet, sheet_source, sheet_err, sheet_conflict = resolve_target_sheet(
        a.task, sheets, getattr(a, "sheet", None), headers=probe_meta.get("headers"))
    if sheet_err:
        print(f"？ {sheet_err}")
        return 3
    a._target_sheet = target_sheet
    # ★★ 2026-08-27（Namakoo が GUI で実測）: 画面のシート選択から --sheet が来た回に
    #   「対象シートは依頼文の語と機械照合できません」が立って ✓ が落ちていた。
    #   ★ **人が選択で名指ししたのは、文字列照合より強い証拠**だ。出どころを運んで、
    #     その回だけ問わない（黙らせるのではなく、根拠が別に在ると認める）。
    a._sheet_source = sheet_source
    a._sheet_conflict = sheet_conflict      # ★ 挙動変更#3: 3択の関門(_sheet_conflict_gate)が読む
    a._rerun_ctx = (book, source_book, struct_dump, sheets)
    # ★ operator8 ①: sheet_source=="task" の推測（依頼文中の裸/マーカー付き言及）は、
    #   op が LOOKUP_FILL だと分かった時に「実は参照シート(source_sheet)だった」と
    #   判明することがある（このブックは2シート・依頼文が言及したのは参照シートだけ、
    #   という自然な言い回し）。誤った「操作するシート」を印字してしまう前に、この源だけ
    #   接地（op が分かるところ）まで印字を遅らせる。"cli"/"default" 由来は常に正しい
    #   （--sheet 明示 or 唯一の既定）ので従来どおり即時に印字する。
    if sheet_source == "task" and len(sheets) > 1:
        a._pending_sheet_announce = (sheets, target_sheet, sheet_source)
    else:
        announce = describe_target_sheet(sheets, target_sheet, sheet_source)
        if announce:
            print(announce)
    return _translate_and_dispatch(a, book, source_book, struct_dump, sheets)


def _flush_pending_sheet_announce(a: argparse.Namespace) -> None:
    """★ operator8 ①: 遅延した事前行をそのまま（推測どおり）印字する
       （op が LOOKUP_FILL 以外だった・または LOOKUP_FILL でも推測が正しかった場合）。"""
    pending = getattr(a, "_pending_sheet_announce", None)
    a._pending_sheet_announce = None
    if not pending:
        return
    sheets, target_sheet, sheet_source = pending
    announce = describe_target_sheet(sheets, target_sheet, sheet_source)
    if announce:
        print(announce)


def _announce_lookup_fill_target_sheet(a: argparse.Namespace, sheets: list, args: dict) -> None:
    """★ operator8 ①: LOOKUP_FILL と分かった時点で、遅延させた事前行を接地する。
       翻訳が返した target_sheet が推測（task 由来）と食い違うなら、依頼文が名指ししたのは
       参照シート(source_sheet)の方だった＝推測は誤りだったということ。誤った主張
       （「操作するシート: N枚目『（実は参照シート）』」）は印字せず、本当の target_sheet を
       source="default"（依頼文からの判断ではなく、参照シート以外の消去法で決まった）で
       出し直す。a._target_sheet もここで訂正する（以降 codegen/事後条件が読む値と揃える・
       verify_dsl_args が resolved["_target_sheet"] で二重に決め直すのと結果は同じだが、
       事前行と実際の対象が食い違って見えないようにする）。
       食い違いが無ければ（推測どおり・または target_sheet が実在しない）、遅延した事前行を
       そのまま出す。"""
    pending = getattr(a, "_pending_sheet_announce", None)
    real_target = args.get("target_sheet") if isinstance(args, dict) else None
    if pending and isinstance(real_target, str) and real_target in sheets and real_target != a._target_sheet:
        a._pending_sheet_announce = None
        a._target_sheet = real_target
        announce = describe_target_sheet(sheets, real_target, "default")
        if announce:
            print(announce)
    else:
        _flush_pending_sheet_announce(a)


def _announce_report_per_row_target_sheet(a: argparse.Namespace, sheets: list, args: dict) -> None:
    """★ 帳票段: operator8①（LOOKUP_FILL の source_sheet 混同）と同型のバグへの対処。
       実機で再現: 依頼文が『雛形を使って請求書を作って』のように雛形シートを名指しすると、
       op 判明**前**の一般解決(resolve_target_sheet)が「依頼文が言及した唯一のシート」＝
       雛形をデータシートと誤認する（雛形は必ずデータシートと別、という事実を一般解決は
       知らない）。REPORT_PER_ROW の raw args には LOOKUP_FILL の target_sheet に相当する
       『データシート名』の直接申告が無い（template_sheet だけが分かる）ため、
       _announce_lookup_fill_target_sheet と同じ食い違い判定はできない ── 代わりに
       「推測が template_sheet と一致し、かつ残りのシートがちょうど1枚」という
       曖昧さの無い場合に限って、その1枚へ訂正する（3枚以上のブックは無理に当てず、
       verify_dsl_args の『雛形とデータシートが同じ』エラーで正直に止める）。"""
    template = args.get("template_sheet") if isinstance(args, dict) else None
    if isinstance(template, str) and template == a._target_sheet:
        others = [s for s in sheets if s != template]
        if len(others) == 1:
            a._pending_sheet_announce = None
            a._target_sheet = others[0]
            announce = describe_target_sheet(sheets, others[0], "default")
            if announce:
                print(announce)
            return
    _flush_pending_sheet_announce(a)


def _announce_format_map_target_sheet(a: argparse.Namespace, sheets: list, args: dict) -> None:
    """★ 様式写像段: _announce_report_per_row_target_sheet と全く同じ理由・同じ形
       （template_sheet しか申告されないため LOOKUP_FILL 型の食い違い判定はできない ──
       推測が template_sheet と一致し、残りのシートがちょうど1枚の時だけ訂正する）。"""
    template = args.get("template_sheet") if isinstance(args, dict) else None
    if isinstance(template, str) and template == a._target_sheet:
        others = [s for s in sheets if s != template]
        if len(others) == 1:
            a._pending_sheet_announce = None
            a._target_sheet = others[0]
            announce = describe_target_sheet(sheets, others[0], "default")
            if announce:
                print(announce)
            return
    _flush_pending_sheet_announce(a)


def _sheet_looks_like_template(source_book: Path, sheet_name: str) -> bool:
    """★ 様式写像段/帳票段: sheet_name に {{列名}} の印が1つでもあれば雛形候補とみなす
       （scan_placeholders を流用）。開けない/読めない場合は False（保守的 ── 誤って
       target_sheet を切り替えない側に倒す）。"""
    try:
        wb = openpyxl.load_workbook(source_book)
    except Exception:
        return False
    try:
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        return bool(scan_placeholders(ws, ws.max_row or 1, ws.max_column or 1))
    except Exception:
        return False
    finally:
        wb.close()


def _translate_and_dispatch(a: argparse.Namespace, book: Path, source_book: Path,
                             struct_dump: dict, sheets: list) -> int:
    """対象シートが決まった後の残り（見出し行 → 翻訳 → 計画の振り分け）。
       ★ 挙動変更#3 のために _cmd_run_dispatch から切り出した: シート名の衝突の3択で
       「もう一方のシートを見てみる」が選ばれたとき、a._target_sheet を差し替えて
       **翻訳からやり直す**ため（見出し行の検出も翻訳結果も対象シートに依存するので、
       翻訳を使い回すのでは「やり直し」にならない）。
       ★ a._reuse_translation が積まれている場合だけ翻訳を省く ── 3択の②で
       プレビュー用に一度翻訳したものを、y の後の本番実行で使い回すため（同じ対象シート・
       同じ依頼文に対して ollama を2回叩かない）。"""
    target_sheet = getattr(a, "_target_sheet", None)
    forced_header_row = getattr(a, "header_row", None)
    if forced_header_row:
        # ★ W8a 項目3: --header-row 指定時は検出(StructDump ヒューリスティクス)を丸ごと
        #   スキップし、その行を対象シートの見出しとして採用する（他シートは既定1行目のまま
        #   ＝ resolve_header_rows の既定と同じ扱い）。CLARIFY には絶対に落ちない。
        header_rows = {s: 1 for s in sheets}
        if target_sheet:
            header_rows[target_sheet] = forced_header_row
        clarify_q = None
    else:
        header_rows, clarify_q = resolve_header_rows(struct_dump, sheets, target_sheet=target_sheet)
        # ★ 様式写像段/帳票段: 依頼文が雛形シートを名指しすると、op判明**前**の一般解決
        #   (resolve_target_sheet)が「依頼文が言及した唯一のシート」＝雛形をデータシートと
        #   誤認する（_announce_report_per_row_target_sheet/_announce_format_map_target_sheet
        #   と同型のバグ）。雛形は小さな構造（見出し+印だけの数行）のことが多く、見出し行
        #   検出が確信を持てず CLARIFY に落ちて翻訳（op判明）にすら届かないことがある実機の
        #   再現形 ── DESIGN-20260824-format-map.md「template_sheet の扱いを必ず決める」。
        #   雛形候補（{{列名}}の印を持つシート）を対象にしていて、残りのシートがちょうど
        #   1枚なら、翻訳より前にそちらへ切り替えて検出をやり直す（誤りなら
        #   verify_dsl_args の『雛形とデータシートが同じ』等で正直に止める・3枚以上は無理に
        #   当てない ── announce 側の訂正と同じ保守的な線引き）。
        if clarify_q and target_sheet and _sheet_looks_like_template(source_book, target_sheet):
            others = [s for s in sheets if s != target_sheet]
            if len(others) == 1:
                a._target_sheet = others[0]
                target_sheet = others[0]
                header_rows, clarify_q = resolve_header_rows(struct_dump, sheets, target_sheet=target_sheet)
        # ★ operator8 ③: resolve_header_rows が「StructDump に対象シートが無い→無言で
        #   1行目」フォールバックを踏んだら、その無言の仮定を開示する（LO の一時不調等で
        #   起きうる・実機の再現形）。clarify_q が立つ分岐とは互いに排他（docstring 参照）。
        #   ★ --dry は元から struct_dump を作らない（意図的なスキップ・_cmd_run_dispatch
        #   参照）ため対象外 ―― LO 不調と区別が付かない誤発火を避ける。
        if (clarify_q is None and not getattr(a, "dry", False)
                and _struct_dump_info_missing(struct_dump, sheets, target_sheet)):
            print(f"（見出し行の自動検出が使えなかったため、"
                  f"『{target_sheet or (sheets[0] if sheets else '')}』シートの1行目を"
                  "見出しとみなしています。違う場合は --header-row で指定してください）")
    if clarify_q:
        # ★ operator8 ①: 翻訳まで届かずに止まる分岐 ―― op が分からないので LOOKUP_FILL 訂正の
        #   出番は無い。遅延させていた推測どおりの事前行をここで出す（従来と同じタイミング差）。
        _flush_pending_sheet_announce(a)
        print(f"？ {clarify_q}")
        return 3

    book_meta = build_book_meta(source_book, header_rows=header_rows)
    # ★ operator8 ③: 列解決が失敗した時の敗者復活（_header_row_hint_for_missing_col）用の
    #   材料。実ファイルからここで一度だけ読む（各シート先頭~10行・軽量）。
    book_meta["_row_scan"] = {s: _scan_first_rows(source_book, s) for s in sheets}
    # ★ 対象シートの**出どころ**を運ぶ（cli=人が --sheet や画面で明示指定した）。
    #   ここに載せれば verify_dsl_args の署名を変えずに全経路へ届く。
    book_meta["_sheet_source"] = getattr(a, "_sheet_source", None)
    translation = getattr(a, "_reuse_translation", None)
    a._reuse_translation = None
    # ★★ 2026-08-28（表記ゆれの treadmill を降りるための入口）: --op で操作を**人が固定**
    #   できるようにした。一段目（言い回しから op を当てる段）を飛ばして、第二段に args
    #   だけを埋め直させる ── 既に在る器官（translate_task_fixed_op）へ配線するだけ。
    #   ★ 揺れを 1 つずつ矯正しても、別の言い方でまた外れる。人が「これだ」と言える道を
    #     常設するのが構造的な答えで、画面の「こう読みました→選び直す」もここを通る。
    forced_op = getattr(a, "op", None)
    if forced_op:
        if forced_op not in OP_SCHEMA:
            print(f"？ そんな操作はありません: 『{forced_op}』（一覧: ailine ops）")
            return 3
        t0 = progress_start(f"⏳ 翻訳中（操作は『{OP_LABELS[forced_op]}』に固定）…")
        fixed = translate_task_fixed_op(a.model, forced_op, a.task, book_meta)
        progress_end(t0)
        if not fixed:
            print(f"？ 『{OP_LABELS[forced_op]}』として読み取れませんでした"
                   "（依頼文に、対象の列や値が書かれているか確かめてください）")
            return 3
        _fargs = dict(fixed.get("args") or {})
        # ★ 人が固定した回でも、**依頼文から機械が読める事実**は足す（A' 原則）。
        #   行番号は「row を名前で探す」実装では拾えないので、ここで渡す。
        _frow = task_names_a_row_number(a.task)
        if _frow and "row" in OP_SCHEMA[forced_op]:
            _fargs["row_number"] = _frow
        # ★★ 人の選択は尊重する ── 別の op へ**黙って読み直さない**。
        #   ただし「1 行を指す依頼」を列ぜんぶ書き換える op で実行するのは、
        #   画面に出した「こう読みました」と結果が食い違う ── 断って選び直させる。
        if plan_writes_beyond_one_cell([{"op": forced_op}]) and task_quotes_a_value(a.task):
            _pts = task_points_at_one_row(a.task, book_meta, target_sheet)
            if _pts:
                print(f"？ {_pts}が、『{OP_LABELS[forced_op]}』は"
                       "その列のデータ行を**全部**書き換えます。"
                       "1 か所だけ直すなら『1セル書換』を選んでください")
                return 3
        translation = {"plan": [{"op": forced_op, "args": _fargs}]}
        a._forced_op = forced_op
    elif translation is None:
        t0 = progress_start(f"⏳ 翻訳中 ({a.model})…")
        translation = translate_task(a.model, a.task, book_meta, temperature=0.1)
        progress_end(t0)
    a._last_translation = translation

    plan = translation.get("plan") if isinstance(translation, dict) else None
    if not isinstance(plan, list) or not plan:
        if isinstance(translation, dict) and translation.get("op"):
            plan = [translation]
        else:
            plan = [{"op": "FREEFORM", "args": {}}]

    # ★★ 2026-08-27（Namakoo が実測）:「みかんとぶどうの間に梨を追加して。売上は600」が
    #   **空行 1 本の挿入**になった。位置ではなく op の取り違えで、しかも LLM は
    #   複数段の計画（行挿入＋一括書換）を返す ── だから計画の**長さを見る前に**判じる。
    #   ★ 黙って op を書き換えない ── 依頼文に**値の指定が在る**という証拠を確かめ、
    #     二段目翻訳（op を固定して args だけ埋め直す・既にある機構）へ回す。
    #     そして**根拠を必ず画面に出す**（勝手に別の操作へすり替えない）。
    #   ★ 「言い換えてください」で終わらせない: 利用者の書き方は正しかった。
    # ★★ 2026-08-27（Namakoo「揺れ無しで追加するにはどうしたらいい？」）:
    #   同じ依頼文で、通る回と聞き返す回があった（7B のサンプリングの揺れ）。
    #   ★ 聞き返しの機構は弱めない ── **機械が場所も値も解けている時だけ**読み直す。
    #     その時に迷う理由が無い（実表を見た側が、モデルより確かなことを知っている）。
    #   ★ 読み直したことは必ず画面に出す（黙って聞き返しを握り潰さない）。
    # ★★ 2026-08-27（Namakoo「梨の売上にピンポイントで入れたい」）:
    #   「梨の売上を2000にして」は SET_COLUMN_VALUE（列を丸ごと）と読まれ、
    #   「値を『』で囲め」と断られていた。だが依頼文は**行を名指し**している。
    #   ★ 行の名前が実表で 1 つに決まるなら、狙いは 1 セル ── 第二段で読み直す。
    #     （OPS_DOC は 1 文字も増やさない ── 第二段は op を固定してスキーマだけ見せる）
    # ★ 読み直しの塊はすべて**対象シート**で解く（1 枚目を仮定しない・上の _sheet_hint と同じ）。
    _sheet_h = (getattr(a, "_target_sheet", None)
                 or (book_meta.get("sheets") or [None])[0])
    # ★★ 2026-08-27（Namakoo「◎を入れて では動作しない」の**構造側**の真因）:
    #   読み直しの塊が 5 つ並んでいて、**後の塊が前の塊の結果を上書きしていた**。
    #   実測: 「チェック列に『◎』を入れて」は正しく条件つき書換に読み直された直後、
    #   列追加の塊がそれを奪って「見出しの無い列を末尾に」に化けていた。
    #   ★ 読み直しは**1 回だけ**。先に決まったものが勝つ（下へ行くほど証拠が弱い順に並べる）。
    #   ★ 個々の塊の条件をいくら賢くしても、この形の事故は消えない ── 塊が増えるたび
    #     「まだ上書きされない」ことを人が確かめる羽目になる。機械で 1 回に縛る。
    #   ★★ 人が op を固定した回は、読み直しを**一切しない**（画面に出した
    #     「こう読みました」と、実際に走る操作が食い違わないため）。
    _reread_done = bool(getattr(a, "_forced_op", None))

    # ★★ 2026-08-29（自分で開けた片配線・実測で捕まえた）: 「〜以外」の断りを
    #   読み直しの門の**内側**に置いていたので、先に別の読み直しが印を立てた回に
    #   素通りし、「味噌汁**以外**を抜き出して」が味噌汁**だけ**を抜き出して ✓ になった。
    #   ★ 断りは読み直しではない ── 門の外に、独立した関所として置く。
    # ★ 2026-09-02: 「〜以外」は読めるようになった（cmp=nin）。ここの関所は外した。
    #   ★ ただし**読めなかった回**は今までどおり断る ── 下の読み直しで
    #     列も値も決まらなければ、黙って別のことをしない。

    def _already_places_a_row(st):
        # ★★ 2026-08-29（Namakoo の通しで実測）: 「件数の合計も合計行に入れて」が
        #   **行追加**に化けた。一段目は 3/3 とも正しく APPEND_TOTAL を返していたのに、
        #   読み直しの門が「行を**ずらす**」op だけを『もう置けている』と数えていて、
        #   合計行のように**ずらさずに末尾へ置く** op が素通りしていた。
        #   ★ 見るべきは「新しい行に中身を置く」と宣言しているか、の 1 点だけ。
        #     ADD_ROW も APPEND_TOTAL も真。INSERT_ROWS（空行を挿すだけ）は偽のまま。
        return _op_writes((st or {}).get("op"), WRITE_NEW_ROW_AT_END)

    def _already_writes_one_cell(st):
        return _op_writes((st or {}).get("op"), WRITE_SINGLE_CELL)

    def _is_a_different_job(st):
        op_ = (st or {}).get("op")
        return any(_op_writes(op_, k) for k in (WRITE_FORMAT_ONLY, WRITE_REMOVE, WRITE_REORDER))

    # ★★ 2026-08-29（Namakoo が実測・qwen も gemma4 も外した）:
    #   「丸山重工の右にPCパーツ」→ qwen は SPLIT_CELL（区切り文字を聞き返す）、
    #   gemma4 は ADD_ROW at:1（見出し行に挿す）。
    #   「丸山重工の項目をPCパーツにして」→ OUT_OF_VOCAB / SET_COLUMN_VALUE（列を全部潰す）。
    #   ★ どれも「1 セルに書く」だけの依頼で、機械は既に答えを持っている
    #     （丸山重工は 8 行目・項目は 2 列目）。**モデルを替えても直らない** ──
    #     誰も表に訊いていなかっただけ。行も列も機械が実表から決める。
    #   ★ 値だけは第二段に出させて、依頼文に literal で在るものだけ通す（A' 原則）。
    #   ★ 計画の**長さ**で門を閉じない（実測: 同じ依頼で 1 段と 2 段が返り分かれ、
    #     2 段の回だけ素通りしていた ── 長さは依頼の性質ではなくモデルの気分）。
    #     見るのは「どの段も 1 セル書換でなく、どの段も別の仕事でない」こと。
    # ★★ 2026-08-31（Namakoo が実測・✓ が出たのに依頼と違う操作）:
    #   「丸和物流と近江スチールの**項目を入れ替えて**」が **1 セル書換**に読み直され、
    #   近江スチールの項目に**「丸和物流」と書いて** ✓ を出していた（鋼材加工が消えた）。
    #   ★ 番人は「1 セルだけ書き換えた」を正しく検算している ── **宣言が依頼と違う**
    #     （三項の「依頼」が抜けた形・今日 2 度目）。
    #   ★ 依頼文が**入れ替え**と言っているなら、1 セル書換に読み替えてはいけない。
    #     入れ替えは 2 か所が動く操作で、1 セル書換は 1 か所しか動かない ── 別の仕事。
    if (not _reread_done and plan
            and not task_asks_for_a_swap(a.task)
            and not any(_already_writes_one_cell(st) for st in plan)
            and not any(_is_a_different_job(st) for st in plan)):
        _cell = resolve_cell_target_from_task(a.task, book_meta, _sheet_h)
        if _cell:
            _crow, _ccol, _cnote = _cell
            _cheads = [str(h) for h in ((book_meta.get("headers") or {}).get(_sheet_h) or [])]
            _cfix = translate_task_fixed_op(a.model, "SET_CELL_VALUE", a.task, book_meta)
            _cval = value_written_in_task(a.task, (_cfix or {}).get("args", {}).get("value"),
                                           _cheads)
            if _cval is None and 1 <= _ccol <= len(_cheads):
                # ★ 第二段が value を返さない回がある（実測: qwen も gemma4 も row/col だけ）。
                #   LLM が出さないなら機械が出す ── 依頼文から、既に分かっている物を引く。
                _cval = bare_value_from_task(
                    a.task, _cell_row_name_for(book_meta, _sheet_h, _crow),
                    _cheads[_ccol - 1], _cheads)
            if _cval is not None and 1 <= _ccol <= len(_cheads):
                print(f"（『1セル書換』として読み直しました ── {_cnote}）")
                plan = [{"op": "SET_CELL_VALUE",
                          "args": {"row_number": _crow, "col": _cheads[_ccol - 1],
                                    "value": _cval}}]
                _reread_done = True
    #   ★★ 2026-08-28（Namakoo が実測・今日いちばん悪い形）: 「7行目の担当を『佐藤』に」で
    #     **担当列が全行『佐藤』になり ✓ が出た**。一括書換の契約としては ✓ は正しいが、
    #     依頼は 1 行だった ── 三項（依頼・宣言・実体）のうち**依頼を見ていなかった**。
    #   ★ 人は行を**番号でも**指す（「7行目」）。名前だけを見ていたので拾えず、しかも
    #     同名が 2 行あると名前でも拾えなかった（実測: ヤマノ食品が 2 行）。
    #   ★ 依頼が行を指しているのに 1 セルへ落とせないなら、**列全体を書かずに断る**。
    #   ★★ 2026-08-28（第二波・実測）: 「7 行F列に『佐藤』を追加」は一段目が 3/3 で
    #     **ADD_ROW**（行の追加）を返していた。op 名で門を作っていたので素通り ──
    #     op 名の数え上げは今日 3 度目に破れた形なので、ここは**宣言**で門を作る。
    #     三項（行を指す・列を指す・値を引用する）が揃った時だけ 1 セルへ落とす。
    if not _reread_done and plan_writes_beyond_one_cell(plan):
        _row_no = task_names_a_row_number(a.task)
        _named = _task_names_a_row(a.task, book_meta, _sheet_h)
        _wide = any((st or {}).get("op") == "SET_COLUMN_VALUE" for st in plan)
        _one_cell = False
        _points = task_points_at_one_row(a.task, book_meta, _sheet_h) if _wide else None
        # ★ 比較語のある依頼は条件つき書換であって 1 セルではない（500 を名前と読まない）。
        if (_row_no or _named) and task_quotes_a_value(a.task) and extract_cmp_from_task(a.task) is None:
            _fx = translate_task_fixed_op(a.model, "SET_CELL_VALUE", a.task, book_meta)
            if _fx and (_fx.get("args") or {}).get("col"):
                _args = dict(_fx["args"])
                if _row_no:
                    _args["row_number"] = _row_no
                    _why = f"依頼文が『{_row_no}行目』と行を指しています"
                else:
                    # ★ 2026-08-29（実測）: 第二段は row に**シート名**を返すことがある
                    #   （'8月請求'）。機械が実表で解いた名前で**上書きする**
                    #   ── setdefault だと LLM の嘘が勝つ。
                    _args["row"] = _named
                    _why = f"依頼文が『{_named}』の行を名指ししています"
                print(f"（『一括書換』でなく『1セル書換』として読み直しました ── {_why}）")
                plan = [{"op": "SET_CELL_VALUE", "args": _args}]
                _one_cell = _reread_done = True
        # ★ ここは読み直しでなく**断り**（印は立てない・立った印を見るのでもない）。
        #   1 セルへ落とせた回だけ黙る、という 1 つの局所変数で決める。
        if _points and not _one_cell:
            # ★ 落とせなかった。列全体を書けば「宣言どおり」で ✓ が出てしまう ── 断る。
            # ★★ 2026-08-29: ここは**本物の 2 択**が残っている唯一の場所
            #   （1 セルか、列ぜんぶか）。行き止まりにせず、**選べる形**で返す。
            #   ★ 候補は機械可読の 1 行で出す（画面がボタンにする・CLI は --op で選べる）。
            #     ポップアップは使わない ── モーダルは画面を止める（実測で踏んだ）。
            print(f"？ {_points}が、どのセルかを決められませんでした。"
                   "列全体は勝手に書き換えません ── どちらか選んでください")
            _col_hint = str((plan[0] or {}).get("args", {}).get("col") or "")
            print(render_choices([
                ("SET_CELL_VALUE",
                 f"その 1 セルだけを書き換える（行番号で言い直してください"
                 f"・例:「7行目の{_col_hint or '担当'}を『佐藤』に」）"),
                ("SET_COLUMN_VALUE",
                 f"『{_col_hint}』列のデータ行を**全部**書き換える"
                 if _col_hint else "その列のデータ行を**全部**書き換える"),
            ]))
            return 3

    # ★★ 2026-08-27（Namakoo「原価が500以上の項目に◎を付ける」）:
    #   実測 4/4 で OUT_OF_VOCAB（しかも「条件付き書式」と誤って読まれていた ──
    #   人が欲しいのは**値**であって書式ではない）。断る側なので横取りして悪くならない。
    # ★★ 2026-08-27（Namakoo「セル指定はできる？『商品』セルに色を付けて など」）:
    #   実測 3/3 で `col:商品`（**列ぜんぶ**）に化けていた。頼んでいない範囲へ静かに
    #   広がる形で、この道具が最も嫌うもの ── 「動かない」より悪い。
    #   ★ 依頼文が「セル」「見出しだけ」と言い、機械がその値を**1 箇所に**特定できるなら、
    #     それは列ではなく 1 セル。位置は機械が実表から決める（LLM に座標を出させない）。
    #   ★ 一段目が断りに落ちる回もある（実測: 「みかんのセルを黄色にして」で OUT_OF_VOCAB）。
    #     その時は飾りの種類を語から決めて、第二段に args だけ埋め直させる。
    _fmt_op = (plan[0] or {}).get("op") if len(plan) == 1 else None
    if (not _reread_done and task_asks_for_one_cell(a.task) and len(plan) == 1
            and _fmt_op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN",
                             "CLARIFY", "FREEFORM", "OUT_OF_VOCAB")):
        if _fmt_op in ("CLARIFY", "FREEFORM", "OUT_OF_VOCAB"):
            _guess = format_op_from_task(a.task)
            _fmt2 = (translate_task_fixed_op(a.model, _guess, a.task, book_meta)
                      if _guess else None)
            _fmt_op = _guess if _fmt2 else _fmt_op
            _fmt_args = dict((_fmt2 or {}).get("args") or {})
        else:
            _fmt_args = dict((plan[0] or {}).get("args") or {})
        _tgt = str(_fmt_args.get("target") or "")
        _nm = _tgt[4:] if _tgt.startswith("col:") else ""
        if not _nm and _fmt_op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
            # ★ 第二段が target を出さない回もある ── 依頼文が名指しする**実在の値**を探す。
            _cands = [v for v in (task_names_real_values(a.task, book_meta, _sheet_h, col)
                                   for col in ((book_meta.get("headers") or {}).get(_sheet_h) or []))
                       for v in v]
            _hdrs_f = [str(h) for h in ((book_meta.get("headers") or {}).get(_sheet_h) or [])]
            _cands += [h for h in _hdrs_f if h and h in (a.task or "")]
            _nm = _cands[0] if len(set(_cands)) == 1 else ""
        if _nm and _fmt_op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
            _cr, _cc, _cnote = resolve_named_cell(book_meta, _sheet_h, _nm)
            if _cr:
                print(f"（『{_nm}』の**1 セル**として読み直しました ── {_cnote}。"
                       "列ぜんぶには広げません）")
                _fmt_args["target"] = f"cell:{_cr},{_cc}"
                plan = [{"op": _fmt_op, "args": _fmt_args}]
                _reread_done = True

    # ★★ 2026-08-27（実測・同じ依頼で聞かれたり聞かれなかったり）: 名指しの行の抽出。
    #   一段目は 2/3 で EXTRACT、1/3 で OUT_OF_VOCAB（→「もしかして」の確認）だった。
    #   ★ 機械が列も値も解けているなら迷う理由が無い ── 他の読み直しと同じ線。
    #   ★★ 2026-08-27（GUI で実測・除外の列挙が 3 度目に破れた）: 一段目は同じ依頼文で
    #     EXTRACT / OUT_OF_VOCAB / CLARIFY を返し分ける。**どれを数え上げても抜ける** ──
    #     今日 3 度目なので、ここも許可の列挙へ裏返す。
    #     依頼文が「抜き出す」と言い、機械が列も値も解けているなら、他の op は当たっていない
    #     （列の抽出だけは別の op が担当するので、そこだけ残す）。
    #     ★ ただし比較語（以上/以下…）が在る依頼には触らない ── それは条件であって
    #       名指しではない（「原価が500以上の行を抜き出して」の 500 を名前と読まない）。
    if (not _reread_done and _re_extract_ask.search(a.task or "")
            and extract_cmp_from_task(a.task) is None
            and not task_asks_to_extract_columns(a.task) and len(plan) == 1
            and (plan[0] or {}).get("op") != "EXTRACT_COLUMNS"):
        # ★★ 2026-09-02: 「〜以外」は**否定**として読む（同じ読み直しの中で分ける）。
        #   ★ 別の門を作らない ── 門が増えるほど「どちらが先に立つか」で事故が起きる
        #     （2026-08-27 に読み直しが 5 つ並んで上書きし合った形）。
        if task_says_except(a.task):
            _ncol, _nvals = except_extraction_reading(book_meta, _sheet_h, a.task)
            if _ncol and _nvals:
                print(f"（『抽出』として読み直しました ── 『{_ncol}』が"
                       f"{'・'.join(_nvals)} **以外**の行を抜き出します）")
                plan = [{"op": "EXTRACT", "args": {"col": _ncol, "cmp": "nin",
                                                    "value": list(_nvals)}}]
                _reread_done = True
        if not _reread_done:
            _xcol, _xvals = resolve_named_extraction(book_meta, _sheet_h, a.task)
            if _xcol and _xvals:
                print(f"（『抽出』として読み直しました ── 『{_xcol}』が"
                       f"{'・'.join(_xvals)} の行を抜き出します）")
                plan = [{"op": "EXTRACT", "args": {"col": _xcol, "cmp": "eq",
                                                    "value": _xvals[0]}}]
                _reread_done = True

    # ★★ 2026-08-27（Namakoo「特定行や特定列の抜き出しができない」）: 列の抽出。
    #   実測で一段目は OUT_OF_VOCAB（「複数条件の抽出」と誤読）を返していた。
    #   ★ 2026-08-30: 段数の縛りを外した（全段が降りている計画も拾う）── 実測で
    #     「商品・売上の2列だけ取り出して」が 2 段の内容不明になり、素通りしていた。
    if (not _reread_done and task_asks_to_extract_columns(a.task)
            and (plan_is_all_giving_up(plan)
                  or (len(plan) == 1 and (plan[0] or {}).get("op") == "EXTRACT"))):
        print("（『列抽出』として読み直しました ── 依頼文が「列だけ」を指しています）")
        plan = [{"op": "EXTRACT_COLUMNS", "args": {}}]
        _reread_done = True

    # ★★ 2026-08-27（Namakoo「置き換えができない」）: 「チェック列の『◎』を全て『合格』に」は
    #   一段目が SET_COLUMN_VALUE（列を丸ごと『合格』に）を返していた ── 空欄の行まで
    #   潰す。機械は引用が 2 つあるので「値が一意に読み取れない」と正しく断っていたが、
    #   **断って終わり**だった。★ 引用が 2 つある時の意味（置き換え）を読む。
    #   ★ col は一段目が既に当てている（実測 3/3）ので、そのまま使って LLM を呼び直さない。
    if (not _reread_done and task_asks_for_a_replace(a.task) and len(plan) == 1
            and (plan[0] or {}).get("op") in ("SET_COLUMN_VALUE", "CLARIFY", "FREEFORM",
                                                "OUT_OF_VOCAB", "SET_WHERE")):
        _rp_args = dict((plan[0] or {}).get("args") or {})
        if not _rp_args.get("col"):
            _rp = translate_task_fixed_op(a.model, "SET_WHERE", a.task, book_meta)
            _rp_args = dict((_rp or {}).get("args") or {})
        if _rp_args.get("col"):
            _src, _dst = extract_replace_pair(a.task)
            print(f"（『置き換え』として読み直しました ── 『{_src}』の行だけを"
                   f"『{_dst}』にします。列を丸ごとは書き換えません）")
            plan = [{"op": "SET_WHERE", "args": {"col": _rp_args["col"]}}]
            _reread_done = True

    #   ★ 2026-08-30: 同上。実測で「売上が1000以上の行の担当を『佐藤』にして」が
    #     『抽出＋一括書換』の 2 段になり、**別シートを作って担当列を丸ごと潰す**
    #     ところだった。★ ここは「降りている計画」ではないので、**抽出＋書換の 2 段**
    #     という形そのものを名指しで拾う（条件つき書換を分解した形）。
    _cw_split = (len(plan) == 2
                  and str((plan[0] or {}).get("op")) == "EXTRACT"
                  and str((plan[1] or {}).get("op")) in ("SET_COLUMN_VALUE", "SET_WHERE"))
    if (not _reread_done and task_asks_for_a_conditional_write(a.task)
            and (plan_is_all_giving_up(plan) or _cw_split
                  or (len(plan) == 1
                       and (plan[0] or {}).get("op") == "SET_COLUMN_VALUE"))):
        _sw2 = translate_task_fixed_op(a.model, "SET_WHERE", a.task, book_meta)
        _sw2_args = dict((_sw2 or {}).get("args") or {})
        if _sw2_args.get("col") and _sw2_args.get("cond_col"):
            print("（『条件つき書換』として読み直しました ── "
                   "依頼文が「条件に当てはまる行だけ」を指しています）")
            plan = [{"op": "SET_WHERE", "args": _sw2_args}]
            _reread_done = True

    # ★★ 2026-08-27（Namakoo「列の追加はできないの？」）:
    #   実測: 「備考という列を追加して」も「原価の右に列を追加して」も語彙外で断られていた
    #   （削除だけあって追加が無い ── 行の時と同じ形の欠け）。
    #   ★ 横取りするのは**断る側**だけ（CLARIFY/FREEFORM/OUT_OF_VOCAB）。
    #     「利益の列を追加して」は一段目が COMPUTE_COLUMN を返す ── そちらの方が
    #     良い答えなので触らない（計算できるなら計算列にする）。
    #   ★★ 2026-08-27（自分で実測）: 一段目は**同じ依頼文で回ごとに違う op** を返す。
    #     「原価列の右に列を追加して」で INSERT_ROWS（行挿入）が返り、そのまま走れば
    #     **列を頼まれて行を挿す**。断るより悪い（軸そのものが違う）。
    #     ★ 依頼文が「列」と言っているなら、行の op は軸を間違えている ── 横取りする。
    #   ★ 計画が 1 段の時だけ差し替える（複合依頼の他の段を巻き添えにしない）。
    _headers_add = (book_meta.get("headers") or {}).get(_sheet_h) or []
    #   ★★ 2026-08-27（Namakoo が GUI で実測・2 件目）: 一段目が **ADD_COLUMN を自分で
    #     返す**回がある（実測 6 回中 5 回）。その場合ここを素通りするので、
    #     一段目が名前を入れてこなかった依頼（「チェックという列を追加して」）で
    #     **名前が空のまま**走り、見出しの無い列ができていた。
    #     ★ 読み直しの対象に ADD_COLUMN 自身も入れる ── ただし**名前が空の時だけ**
    #       第二段に聞き直す（実測で第二段は 6/6 で『チェック』を返す）。
    #   ★★ 2026-08-27（3 件目・実測）: 除外する op を数え上げる書き方だと、一段目が
    #     返しうる op が増えるたびに穴が開く（実測で SPLIT_CELL まで返ってきた ──
    #     「列を追加して」に対してセルの分割を走らせるところだった）。
    #     ★ **残してよい op を挙げる**形へ裏返す: 列を追加する依頼に対して正当なのは
    #       「中身のある列を作る」op（計算列・転記）だけ。それ以外は軸か操作が違う。
    if (not _reread_done and task_asks_to_add_a_column(a.task) and len(plan) == 1
            and (plan[0] or {}).get("op") not in KEEP_FOR_COLUMN_REQUEST):
        _was_add_col = (plan[0] or {}).get("op") == "ADD_COLUMN"
        _ac_args = dict((plan[0] or {}).get("args") or {}) if _was_add_col else {}
        if not str(_ac_args.get("name") or "").strip():
            _ac = translate_task_fixed_op(a.model, "ADD_COLUMN", a.task, book_meta)
            _ac_args = dict((_ac or {}).get("args") or {})
        if not _was_add_col:
            _at_add, _note_add = resolve_col_anchor(a.task, _headers_add)
            print(f"（『列追加』として読み直しました ── 依頼文が列の追加を指しています: "
                   f"{_note_add or "末尾"}）")
        plan = [{"op": "ADD_COLUMN", "args": _ac_args}]
        _reread_done = True

    # ★★ 2026-08-27（Namakoo「行や列の入れ替えを実装できるか」）:
    #   実測すると「みかんとぶどうを入れ替えて」は CLARIFY（「どちらの列ですか」）、
    #   「売上と原価を入れ替えて」は OUT_OF_VOCAB に落ちていた ── **断る側**なので、
    #   ここで読み直しても悪くなりようがない（成功していた経路を横取りしない）。
    #   ★ 軸（行か列か）は LLM に当てさせない: 依頼文はたいてい言わないし、
    #     **実表を見た機械のほうが確かなことを知っている**（verify_dsl_args が決める）。
    #   ★ SORT も読み直しの対象に入れる（実測: 「みかんとぶどうの順番を逆にして」で
    #     一段目が SORT を返した ── そのまま走れば**表全体を並べ替えて ✓ を出す**。
    #     間違った操作を自信をもって実行する形で、断るより悪い）。
    #   ★ ただし SORT を横取りする以上、条件を厳しくする: **2 つの名前が実表で
    #     ちょうど 1 つに解ける時だけ**読み直す（解けないなら元の計画のまま進める）。
    #   ★★ 2026-08-29（Namakoo が実測・今日 4 度目の「列挙は漏れる」）:
    #     「税込み金額列と金額列を入れ替えて」が **COMPUTE_COLUMN**（計算列）に読まれ、
    #     金額列を掛け算で潰しかけた（関所が止めた）。門が op 名の列挙だったので素通り。
    #   ★ 門は証拠で作る: 依頼文が入れ替えを言い、**機械が対象を 2 つとも解けている**なら
    #     読み直す。既に SWAP で読めている回だけ触らない（op 名の列挙をやめる）。
    if (not _reread_done and plan and task_asks_for_a_swap(a.task)
            and not any((st or {}).get("op") == "SWAP" for st in plan)):
        # ★★ 2026-08-31（Namakoo が実測・「単価の入れ替えに対応していないようにみえる」）:
        #   「みどり建設の単価と丸和物流の単価を入れ替えて」が **OUT_OF_VOCAB** で
        #   終わっていた。セルの入れ替えは前日に実装したが、それは SWAP の**検証段**に
        #   置いたので、**op が SWAP にならなければ一度も呼ばれない**。
        #   ★ また「番人は在るが、失敗が取る形では鳴らない」── この repo で何度も踏む形。
        #   ★ ここで LLM に聞く前に、**機械だけで 2 つのセルが解けているなら**それを使う。
        #     依頼文と実表しか見ていないので、LLM の返事より確かで、速い。
        if swap_targets_are_cells(a.task, book_meta, _sheet_h):
            print("（『入れ替え』として読み直しました ── "
                   "依頼文が 2 つのセルの入れ替えを指しています）")
            plan = [{"op": "SWAP", "args": {}}]
            _reread_done = True
    if (not _reread_done and plan and task_asks_for_a_swap(a.task)
            and _swap_axis_hint(a.task) != "column"
            and not any((st or {}).get("op") == "SWAP" for st in plan)):
        # ★★ 2026-09-02（入れ替えを効果の検体に載せて初めて見えた）:
        #   「あかね商事とうえだ物産の行を入れ替えて」で、下の二段目（op を SWAP に
        #   固定して LLM に聞き直す）が **a='取引先' b='件数'** を返した ──
        #   人が言っていない**列名**。実在の列なので _swap_pair_resolves は True。
        #   止めたのは三項の番人（依頼文の語と照合できない）で判断は正しいが、
        #   **利用者の正当な依頼が通らない**。3 表では同じ言い方が 6/6 通っていた
        #   ── つまり **LLM の揺れ**。
        #   ★ セルの時と同じ処方: **LLM に聞く前に、機械だけで解けているならそれを使う。**
        #     依頼文と実表しか見ていないので、LLM の返事より確かで、速い。
        #   ★ 行番号で渡す ── 既に 4 表 4/4 で通っている道に載せる（新しい道を作らない）。
        if (_rows2 := swap_targets_are_rows(a.task, book_meta, _sheet_h)):
            print("（『入れ替え』として読み直しました ── "
                   f"依頼文が {_rows2[0]}行目 と {_rows2[1]}行目 を指しています）")
            plan = [{"op": "SWAP", "args": {"a": f"{_rows2[0]}行目",
                                             "b": f"{_rows2[1]}行目"}}]
            _reread_done = True
    if (not _reread_done and plan and task_asks_for_a_swap(a.task)
            and not any((st or {}).get("op") == "SWAP" for st in plan)):
        _sw = translate_task_fixed_op(a.model, "SWAP", a.task, book_meta)
        _sw_args = (_sw or {}).get("args") or {}
        _sa, _sb = str(_sw_args.get("a", "")).strip(), str(_sw_args.get("b", "")).strip()
        # ★★ 2026-08-29: 門を広げた途端、元の狭い門が守っていた物が壊れた ──
        #   「税込み金額の**順番を逆にして**」で、第二段が**相手をでっち上げて**
        #   入れ替えに化けた（正当な並べ替えを壊す）。
        #   ★ A' 原則をここにも通す: **入れ替える 2 つは、どちらも依頼文に在ること**。
        #     片方しか書かれていない依頼は、入れ替えの依頼ではない。
        if (_sa and _sb and _sa in (a.task or "") and _sb in (a.task or "")
                and _swap_pair_resolves(book_meta, _sheet_h, _sa, _sb)):
            print(f"（『入れ替え』として読み直しました ── 依頼文が"
                   f"『{_sw_args["a"]}』と『{_sw_args["b"]}』の入れ替えを指しています）")
            plan = [{"op": "SWAP", "args": dict(_sw_args)}]
            _reread_done = True

    #   ★★ 2026-08-28（Namakoo が実測・3 度目の「列挙は漏れる」）: 一段目は同じ依頼文で
    #     INSERT_ROWS / CLARIFY / EXTRACT を返し分ける。op 名を数え上げても必ず漏れる。
    #   ★ 宣言で門を作る: 「行をずらして**値も書く**」と宣言している op（＝ADD_ROW）で
    #     既に読めているなら触らない。逆に、**明らかに別の仕事**を宣言している op
    #     （見た目だけ／削除／並べ替え）にも触らない。それ以外は読み直しの候補。
    #   ★★ 2026-08-29（84 件の効果検体・3 表で同じ形）: 「味噌汁の**上に**新品を入れて」で
    #     **行が 2 本**増えた。一段目が
    #       [INSERT_ROWS at:2（空行）, ADD_ROW at:2（値つき）]
    #     という 2 段の計画を返し、両方走っていた。
    #   ★ 1 つの依頼に**配置は 1 回**。行をずらすと宣言した段が 2 つ以上あるなら、
    #     それは同じ仕事を二重に言っている ── 1 本に畳む（宣言で数える・op 名で数えない）。
    # ★★ 2026-08-29（84 件の効果検体で最後まで残った穴）: 「〜の行を除いて」が
    #   削除にならなかった（3 表で EXTRACT / OUT_OF_VOCAB / 条件付き抽出 に返り分かれた）。
    #   ★ 「除く」は 2 通りに読める。**できる読みだけ**を提案し、できない読み（〜以外を
    #     残す）は名指しで断る ── 黙って逆のことをすると、残したい行を消す事故になる。
    if (not _reread_done and plan
            and not any(_op_writes((st or {}).get("op"), WRITE_REMOVE) for st in plan)):
        _rm = removal_reading(a.task, book_meta, _sheet_h)
        if _rm:
            _rat, _rnote = _rm
            print(f"（『行削除』として読み直しました ── {_rnote}）")
            plan = [{"op": "DELETE_ROWS", "args": {"at": _rat, "count": 1}}]
            _reread_done = True

    # ★★ 2026-08-29（84 件の効果検体で最後に残った 1 件・また片配線）:
    #   「鈴木**の上に**新品を入れて」で **氏名=鈴木**（＝位置の目印そのもの）が
    #   新しい行に書かれた。値の篩は「読み直した経路」にだけ入れていて、
    #   一段目が最初から ADD_ROW を返した回は素通りしていた。
    #   ★ 処方は「両方に入れる」ではなく「**必ず同じ関数を通す**」── 経路が増えても
    #     篩が外れない形にする（この repo が 3 度直してきた形）。
    for _st in plan:
        if (_st or {}).get("op") != "ADD_ROW":
            continue
        _raw_vals = (_st.get("args") or {}).get("values")
        _sieved = add_row_values_from_request(
            a.task, book_meta, (book_meta.get("sheets") or [None])[0], _raw_vals)
        if _sieved and _sieved != _raw_vals:
            _st.setdefault("args", {})["values"] = _sieved

    # ★★ 2026-08-29（Namakoo が実測）:「合計を金額表示にして」が**合計追加**に読まれ、
    #   既にある合計をもう一度書いて ✓ が出た（画面は何も変わらない）。
    #   『合計』は**対象**であって操作ではない ── 依頼文が「数値の見せ方」を言っていて、
    #   掛ける先が機械で 1 つに決まるなら、書式として読み直す。
    if (not _reread_done and plan and task_asks_for_number_format(a.task)
            and not any((st or {}).get("op") == "NUMBER_FORMAT" for st in plan)):
        _nf = number_format_target(a.task, book_meta, _sheet_h)
        if _nf:
            _kind, _what = _nf
            _nargs = {"style": "thousands"}
            if _kind == "col":
                _nargs["col"] = _what
                _nnote = f"列『{_what}』"
            else:
                _nargs["row_number"] = _what
                _nnote = f"{_what}行目"
            print(f"（『数値書式』として読み直しました ── 依頼文が桁区切りを指しています: "
                   f"{_nnote}）")
            plan = [{"op": "NUMBER_FORMAT", "args": _nargs}]
            _reread_done = True

    _row_placing = sum(1 for st in plan if _op_writes((st or {}).get("op"), WRITE_ROW_SHIFT))
    if (not _reread_done and plan and not any(_is_a_different_job(st) for st in plan)
            and (_row_placing > 1
                  or (len(plan) == 1 and not _already_places_a_row(plan[0])))):
        # ★★ 2026-08-30（同じ実測）: ここが **1 枚目**を見ていた。画面で『雛形』を
        #   選んでいるのに『8月請求』の 9 行目で位置を解き、別のシートの行を根拠に
        #   op を乗り換えていた。★ 対象シートは既に 1 箇所で決まっている ── それを使う。
        _sheet_hint = (getattr(a, "_target_sheet", None)
                        or (book_meta.get("sheets") or [None])[0])
        _why = insert_rows_should_have_been_add_row(a.task, {}, book_meta, _sheet_hint)
        if _why:
            _fixed = translate_task_fixed_op(a.model, "ADD_ROW", a.task, book_meta)
            _vals = (_fixed or {}).get("args", {}).get("values") or {}
            # ★ A' 原則: 置く物の名前は**依頼文から**取れていること。
            #   でっち上げた値（『商品=みかんとぶどう』の実測）で op を乗り換えない。
            # ★ 実測: 第二段は values を dict でも list でも返す（形を決めつけない）。
            # ★★ 値は機械が依頼文から決める（LLM の出した値は篩にかけるだけ）。
            #   実測で「位置の目印が値になり・置く物が別の列に入り・未定/未設定が
            #   でっち上げられる」の 3 つが同時に起きた。
            _clean = add_row_values_from_request(a.task, book_meta, _sheet_hint, _vals)
            if _fixed and _clean:
                print(f"（『行挿入』でなく『行追加』として読み直しました ── {_why}）")
                _aargs = dict(_fixed["args"])
                _aargs["values"] = _clean
                plan = [{"op": "ADD_ROW", "args": _aargs}]
                _reread_done = True
            elif plan_only_inserts_a_bare_row(plan):
                # ★★ 2026-08-29（効果検体で実測・「3行目と4行目の間に新品を作って」）:
                #   ここまで来た時点で、機械は「**値を入れる行が欲しい依頼だ**」と
                #   分かっている（それが _why）。なのに値が決まらなかった回は、
                #   そのまま**空行を挿して ✓ を出して**いた ── 宣言（空行を挿す）と
                #   実体は一致するので検算は通る。だが依頼とは違う。
                #   ★ 三項のうち「依頼」を捨てた形の再演。この製品の芯は
                #     「機械で確かめられないものに ✓ を出さない」── **壊す前に止まる**。
                #   ★ 断りっぱなしにしない: 何が決まらなかったかと、通る言い方を言う。
                print(f"？ 入れる値を依頼文から決められません ── {_why}")
                print("  （値と列を言ってください: 例「3行目の下に新品を追加して」"
                       "「みかんの下に梨を追加して。売上は600」）")
                print("  （空の行が欲しいなら: 例「3行目の下に1行挿入して」）")
                return 3

    # ★★ 2026-08-30: 中身がまったく同じ段は 1 回にまとめる（連鎖で 2 段目が 1 段目の
    #   出力を食う前に畳む ── 順番が意味を持つ）。黙って畳まず、落とした数を言う。
    if plan and len(plan) > 1:
        plan, _folded = fold_identical_steps(plan)
        if _folded:
            print(f"（同じ操作が {_folded + 1} 回書かれていたので 1 回にまとめました）")

    # ★★ 関所（2026-08-29・Namakoo の設計判断）: 同じ軸に位置を作る段が 2 つ以上ある
    #   計画は実行しない。上の読み直しで 1 本に畳めていればここは通る ── 畳めなかった
    #   回だけ、**壊す前に**止まる（当て物でなく関所にするのが芯）。
    if (_dup := too_many_placements(plan)):
        print(f"？ {_dup}")
        return 3

    if len(plan) == 1:
        step = plan[0]
        op = step.get("op")
        # ★ operator8 ①: op が分かった ―― LOOKUP_FILL なら遅延した事前行を接地して
        #   訂正の要否を判断する。それ以外は推測どおりに出す（従来と同じ文言・タイミング差のみ）。
        if op == "LOOKUP_FILL":
            _announce_lookup_fill_target_sheet(a, sheets, step.get("args") or {})
        elif op == "REPORT_PER_ROW":
            _announce_report_per_row_target_sheet(a, sheets, step.get("args") or {})
        elif op == "FORMAT_MAP":
            _announce_format_map_target_sheet(a, sheets, step.get("args") or {})
        else:
            _flush_pending_sheet_announce(a)
        if op == "CLARIFY":
            question = step.get("question") or "確認が必要です"
            print(f"？ {question}")
            # ★ 行き止まりに出口を置く（盲検査定 A の実測: 語彙外の依頼を 4 回言い直して
            #   4 回とも質問返しになり「普通の購入検討者ならここで評価を終える」）。
            #   聞き返しは「言い方が悪い」場合と「そもそも対応していない」場合を
            #   区別できない ── 区別する手段を毎回そえる。
            print("  （頼める操作の一覧: ailine ops）")
            return 3
        # ★ 帳票段: REPORT_PER_ROW は「データ行 N＝出力 N 枚」という op の形が
        #   cmd_run_dsl/dsl_step.py の共有エンジン（1シートへの書き込み前提）と構造的に
        #   違う（検分シートを LO 適用の直後・事後条件チェックの直前に足す必要があり、
        #   apply_dsl_step は basrun_apply→run_postcondition を1呼び出しの中で行うため
        #   割り込めない）── ②検証・③確認は共有するが④〜⑥は専用の cmd_run_report_per_row
        #   を通す（cmd_run_folder 等、構造が違う op が独自のトップレベル関数を持つのと同じ
        #   作法）。OP_SCHEMA には登録したままにする（他の宣言駆動の番人がそのまま効くため）。
        if op == "REPORT_PER_ROW":
            return cmd_run_report_per_row(a, book, source_book, book_meta, step.get("args", {}))
        # ★ 様式写像段: FORMAT_MAP も REPORT_PER_ROW と同じ理由（検分シートを LO 適用の
        #   直後・事後条件チェックの直前に足す必要がある）で専用の実行経路を通す。
        if op == "FORMAT_MAP":
            return cmd_run_format_map(a, book, source_book, book_meta, step.get("args", {}))
        if op in OP_SCHEMA:
            return cmd_run_dsl(a, book, source_book, book_meta, op, step.get("args", {}))
        return _maybe_suggest_or_refuse(a, book, source_book, book_meta, sheets, step)

    # ★ operator8 ①: 複合計画は対象範囲外（このブリーフの検体は単発 LOOKUP_FILL のみ）。
    #   遅延させた推測をそのまま出す ―― 複合計画の段別 LOOKUP_FILL 訂正は今回やらない。
    _flush_pending_sheet_announce(a)
    return cmd_run_plan(a, book, source_book, book_meta, plan)


def _column_existing_value_count(book_path: Path, sheet_name: str, col_name: str,
                                  header_row: int = 1) -> int:
    """★ M2c / W10a: target(既存列指定)列に、見出し行を除いて値が入っているセルの件数。
       上書き検知の明示（確認行の注意書き）と W10a の破壊の関所（確認メッセージの件数）が
       共有する。読めない/列やシートが見つからない場合は 0（保守的に『無い』扱い＝誤って
       止めない）。★ W3: header_row(1起点)で見出しの実位置を受け取る
       （省略時は物理1行目・旧挙動と同一）。"""
    try:
        wb = openpyxl.load_workbook(book_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return 0
        ws = wb[sheet_name]
        idx = _col_index_by_header(ws, col_name, header_row=header_row)
        if idx is None:
            wb.close()
            return 0
        last = _scan_last_row(ws, header_row=header_row)
        count = sum(1 for r in range(header_row + 1, last + 1)
                    if ws.cell(row=r, column=idx).value not in (None, ""))
        wb.close()
        return count
    except Exception:
        return 0


def _maybe_warn_target_overwrite(op: str, resolved: dict, book_meta: dict, book_path: Path) -> str | None:
    """★ M2c 項目2 / W10c 致命1: OP_WRITE_TARGET が宣言する書き込み先列に既存値がある場合、
       上書きになる旨の1行を返す（無ければ None・確認行に明示するため）。
       ★ W10c: 対象を COMPUTE_COLUMN 専用の if から OP_WRITE_TARGET の宣言読み取りへ
       一般化した（監査実測: LOOKUP_FILL がこの関所を素通りしていた事故の再発防止。
       OP_WRITE_TARGET のコメント参照）。
       ★ W10a 項目1: この検出（と件数）を「破壊の関所」（原本適用時に確認を挟む・
       cmd_run_dsl/cmd_run_plan 側）がそのまま流用する（検出ロジックを二重管理しない）。
       ★ 単位C: 宣言が領域を持つようになったので、『既存列を書く』と宣言した op だけを
       対象にする（col_key を持つのはその op だけ＝番人テストが両者の一致を検査する）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or not write_target.col_key:
        return None
    col_key, sheet_key = write_target.col_key, write_target.sheet_key
    col_name = resolved.get(col_key)
    if not col_name:
        return None
    if sheet_key:
        sheet_name = resolved.get(sheet_key)
    else:
        # ★ 挙動変更#2: sheets[0] 決め打ちをやめ、verify_dsl_args が一箇所で決めた
        # resolved["_target_sheet"] を読む（旧値と後方互換のフォールバック付き）。
        sheets = book_meta.get("sheets") or []
        sheet_name = resolved.get("_target_sheet") or (sheets[0] if sheets else None)
    if not sheet_name:
        return None
    header_row = book_meta.get("header_rows", {}).get(sheet_name, 1)
    # ★★ 2026-08-27（1 セル書換を足して分かった）: この関所は col_key を持つ op に対して
    #   **列全体**の既存値を数える。1 セルだけを書く op に当てると、触りもしない
    #   他の行の値まで「上書きします」と言って止める（実測: 空のセルに書くのに 3 件と言われた）。
    #   ★ 宣言が「どの範囲を書くか」を持っていないのが根 ── **1 セル用は 1 セルだけ見る**。
    if op == "SET_CELL_VALUE":
        row_name = resolved.get("row")
        if not row_name:
            return None
        hit, _why = _resolve_named_row(book_meta, sheet_name, str(row_name))
        if hit is None:
            return None
        try:
            with BookView(Path(book_path)) as bv:
                ws = bv.sheet(sheet_name)
                _last, last_col = data_extent(ws, header_row)
                headers = [str(ws.cell(row=header_row, column=c).value or "")
                            for c in range(1, last_col + 1)]
                if col_name not in headers:
                    return None
                cur = ws.cell(row=hit, column=headers.index(col_name) + 1).value
        except Exception:
            return None
        if cur in (None, ""):
            return None      # 空のセルに書くだけ ── 壊すものが無い
        return (f"★ 『{row_name}』の{col_name}には既に {cur!r} が入っています"
                 "（この 1 セルだけを上書きします）")
    count = _column_existing_value_count(book_path, sheet_name, col_name, header_row=header_row)
    if count > 0:
        return f"★ 対象列『{col_name}』には既存の値が {count} 件あります（上書きします）"
    return None


def _own_output_headers(op: str, resolved: dict | None):
    """★★ 単位H: 「宣言した出力先に在るのは、自分の前回の出力か」を見分ける見出し署名。

    ★ 署名は実装から取る（想像しない）: `helpers/AiLineHelpers.bas` の SummaryTable は
      A1 に分類列の見出し名を、B1 に "合計 - " & 集計列名 を書く（同ファイル 374-375 行）。
    ★ PIVOT は対象外。PivotSum は本物の DataPilot を LibreOffice に描かせるので、
      出力の見出しがコード上に無く、静的に決められない。
      → 発火条件つきで残す: DataPilot の実出力の見出しを 1 度実機で観測できたら、ここに足す。
        それまで PIVOT の 2 回目は関所が鳴る（＝安全側で、うるさい側に倒れている）。
    ★★ EXTRACT（単位H の2例目・コミット 2edcb08「EXTRACT op」）: ExtractRows ヘルパは
      出力の1行目に元シートの見出し行をそのままコピーする（helpers/AiLineHelpers.bas 参照）。
      出力シート名は固定でなく col/cmp/value から動的に決まる（_extract_output_sheet_name。
      verify_dsl_args が resolved["_new_sheet"] に積む）ので OP_DECLARED_SHEET_NAME の
      固定表には乗らない ── 署名の材料(見出し全体)は resolved["_source_headers"]（同じく
      verify_dsl_args が積む）を読む。
    """
    if not resolved:
        return None
    if op == "AGGREGATE":
        sheet = OP_DECLARED_SHEET_NAME.get(op)
        group, value = resolved.get("group_col"), resolved.get("value_col")
        if not (sheet and group and value):
            return None
        return {sheet: (group, f"合計 - {value}")}
    if op == "EXTRACT":
        sheet = resolved.get("_new_sheet")
        source_headers = resolved.get("_source_headers")
        if not (sheet and source_headers):
            return None
        return {sheet: tuple(source_headers)}
    # ★ DEDUP（単位H の3例目）: DedupRows も出力の1行目に元シートの見出し行をそのまま
    #   コピーする（helpers/AiLineHelpers.bas 参照）── EXTRACT と全く同じ署名の作り方。
    if op == "DEDUP":
        sheet = resolved.get("_new_sheet")
        source_headers = resolved.get("_source_headers")
        if not (sheet and source_headers):
            return None
        return {sheet: tuple(source_headers)}
    return None


def _maybe_warn_write_precondition(op: str, before: dict, after: dict, resolved: dict | None = None):
    """★★ 単位F: op が宣言した書き込み領域(OP_WRITE_TARGET.writes)の**前提**が、適用後の
       実測で破れていれば1行を返す（無ければ None）。判定本体は
       ailine_core/write_precondition.py（純ロジック・前提の一覧はそこの docstring）。
       ★ 上の _maybe_warn_target_overwrite（適用**前**に列の既存値を数える）と、この関数
       （適用**後**に before/after を突き合わせる）は、同じ破壊の関所の2つの入力にすぎない
       ―― 返すのは同じ形の1行で、止める判断は両方とも _confirm_overwrite_or_gate が行う。
       col_key を持つ 3 op しか守れていなかった関所に、行・シート・書式・並べ替えの腕を
       足すのがこの関数（関所そのものは増やさない）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target:
        return None
    _writes = write_target.writes
    # ★★ 2026-08-29（Namakoo が実測）: 「既にある合計行に書く」回は**末尾に足していない**
    #   ── 宣言（new_row_at_end）のままだと「末尾に足すはずが既存の行を書き換えた」と
    #   誤警報する。その回だけ**1 セルの書き込み**として扱う（宣言でなく引数から分かる事実・
    #   位置がずれる回に位置ベースの前提を外すのと同じ形）。
    if (resolved or {}).get("_at_row"):
        _writes = tuple(k for k in _writes if k != WRITE_NEW_ROW_AT_END) + (WRITE_SINGLE_CELL,)
    return check_write_preconditions_detail(
        _writes, before, after,
        cell_ref=_cell_ref, fmt_value=_fmt_cell_value,
        own_output_headers=_own_output_headers(op, resolved),
        # ★ この回、新しい列を依頼文の位置へ動かしたなら、右側の列は 1 つずつずれる
        #   ── 位置で比べる前提はその回だけ使えない（宣言でなく引数から分かる事実）。
        positions_shifted=bool((resolved or {}).get("_move_new_col_to") is not None))


def _maybe_own_prior_output_notice(op: str, before: dict, after: dict, resolved: dict | None = None) -> list:
    """★★ 単位H 開示: _looks_like_own_prior_output が真になって new_sheet の前提検査を
       スキップした（＝関所が黙った）ときの理由を、1行ずつ開示する（無ければ空リスト）。
       助言ではない ── 前提検査（_maybe_warn_write_precondition）が黙って通した理由の
       開示そのもの。呼び出し側は前提検査の直後で呼ぶ（単発・複合計画の両経路）。
       ★ new_sheet を writes に宣言していない op は最初から対象外（own 判定自体が起きない）。"""
    write_target = OP_WRITE_TARGET.get(op)
    if not write_target or "new_sheet" not in write_target.writes:
        return []
    return own_prior_output_notice_lines(before, after, own_output_headers=_own_output_headers(op, resolved))


def _interpretation_summary_line(resolved: dict, inferred: set) -> str | None:
    """★ W10a 項目3: 実行前の解釈要約。target が数字表記から列名へ推定解決され、かつ
       （呼び出し側が別途 _maybe_warn_target_overwrite で）既存データありと分かっている
       場合だけ、その経緯を1文で見せる（「監査要望3」＝数字指定→列名解決の可視化）。
       それ以外（target が最初から実在列名で指定された等）は None（何も語ることが無い）。"""
    if "target" not in inferred or resolved.get("_target_raw") is None:
        return None
    return f"→『{resolved['_target_raw']}』は既存の『{resolved['target']}』列と解釈しました（既存データあり）"


def _confirm_overwrite_or_gate(a: argparse.Namespace, warn_overwrite: str | None,
                                step_prefix: str = "", subject_mismatch: bool = False) -> int | None:
    """★ W10a 項目1: 破壊の関所。既定(原本へ直接反映)で、既存データへの上書きが起きる
       操作（_maybe_warn_target_overwrite が検出）は、--ask 無指定でも確認を挟む
       （監査実測: target が誤って既存列に解決され、確認なしで実データが上書きされた
       事故の再発防止）。--copy/--dry 時は原本に触れない/何もしないため素通し。
       --overwrite が既に立っていれば承知の上として素通し。--ask が既に立っている場合は
       cmd_run_dsl 側の汎用確認で兼ねる（二重に聞かない）。
       戻り値: 続行してよければ None、中断すべきなら呼び出し側がそのまま return すべき
       exit code（対話で拒否=1・非対話で確認できない=7）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。単発 DSL では空文字。
       ★★ 単位E: subject_mismatch（依頼文の語と解決値が矛盾する対象がある＝③）も**この関所を
       再利用**して確認する。新しい関所も新しい exit code も作らない ―― 止める条件
       （原本へ直接反映する時だけ・--dry/--copy/--ask/--overwrite では素通し）も、拒否 1 /
       非対話 7 も、上書きの関所とまったく同じ。違うのは聞く文と、非対話時に示す逃げ道だけ。
       ★ 対象シートの3択（_sheet_conflict_gate）を既に出した run では重ねて聞かない
       （同じ対象について2度尋ねることになる。✓ の抑制と ⚠ の印字はそのまま行う）。"""
    reason = warn_overwrite or (subject_mismatch and not getattr(a, "_sheet_conflict_asked", False))
    if not (reason and getattr(a, "inplace", False) and not getattr(a, "dry", False)
            and not getattr(a, "ask", False) and not getattr(a, "overwrite", False)):
        return None
    prompt = "上書きしますか？" if warn_overwrite else "この対象で実行しますか？"
    try:
        ans = input(f"{step_prefix}{prompt} [y/N]: ").strip().lower()
    except EOFError:
        options = [("--overwrite", "上書きを承知して続行する（バックアップから ailine undo で戻せる）"),
                   ("--copy", "原本には触らず .out に結果を作る（原本は無変更）")] if warn_overwrite \
            else [("--copy", "原本には触らず .out に結果を作る（原本は無変更）"),
                  ("--sheet / 列名を依頼文に明記", "対象を依頼文で名指しして、もう一度実行する")]
        for ln in render_retry_options(step_prefix, options):
            print(ln)
        return 7
    if ans not in ("y", "yes"):
        print(render_aborted(step_prefix))
        # ★ 誤分類の実例台帳センサ①: 「上書きしますか？」（=破壊の関所そのもの）で N を
        #   選んだときだけ拾う。シート衝突3択や freeform 関所（subject_mismatch のみの
        #   reason）は対象外 ── 第一波は overwrite 関所のみ（広げるときは検体から）。
        if warn_overwrite:
            _record_misclass_suspect("gate_decline", getattr(a, "task", ""),
                                      getattr(a, "_last_translation", None) or {},
                                      getattr(a, "book", None))
        return 1
    return None


def _stdin_isatty() -> bool:
    """標準入力が端末か（対話してよいか）。★ ここを関数にしてあるのはテストから
       差し替えるため（stdin そのものを触らずに TTY/非 TTY を作り分ける）。"""
    try:
        return bool(sys.stdin) and sys.stdin.isatty()
    except Exception:
        return False


def _sheet_conflict_gate(a: argparse.Namespace, op: str) -> int | None:
    """★ 挙動変更#3: シート名の衝突で既定(1枚目)へ後退した時だけ、「解釈:」行の直後に
       3択を出す薄い配線（選択肢の文言は ailine_core.target_sheet、対話の仕組みは
       ailine_core.ask_choice が本体）。
       戻り値: そのまま続行してよければ None、呼び出し側がそのまま return すべき
       exit code なら int（②の実行結果もここに含まれる）。
       ★ 聞くのは衝突で後退した時だけ・TTY の時だけ（--json/パイプ/CI/--dry では
       絶対に入力を待たない ── 止めると動いていたスクリプトが黙って壊れる）。"""
    conflict = getattr(a, "_sheet_conflict", None)
    if not conflict or getattr(a, "_sheet_conflict_asked", False):
        return None
    # ★ 誤爆#3: ここで a._sheet_conflict は消さない ── 「曖昧なので既定へ後退した」という
    #   判定結果は実行の最後（助言）まで運ぶ必要がある。消してよいのは『もう聞いた』という
    #   対話側の状態だけなので、別のフラグに分けた（②で対象シートを差し替えたときだけは
    #   後退そのものが取り消されるので、_preview_and_run_on_alternative_sheet が記録を落とす）。
    a._sheet_conflict_asked = True   # 二重に聞かない（②のやり直しでも再発火させない）
    lines, choices = sheet_conflict_choice_lines(conflict, op, OP_LABELS.get(op, op))
    result = ask_choice(lines, [Choice(key=k, text=t) for k, t in choices],
                        interactive=is_interactive(stdin_isatty=_stdin_isatty(),
                                                   json_mode=getattr(a, "json", False),
                                                   dry=getattr(a, "dry", False)))
    if result.key in (None, "1"):
        return None            # 聞かなかった/①＝上の解釈のとおり（既定＝今までの挙動）
    if result.key == "3":
        print(render_aborted())
        return 1               # 既存の中止系（対話で拒否）と同じ exit code
    return _preview_and_run_on_alternative_sheet(a, conflict)


def _preview_and_run_on_alternative_sheet(a: argparse.Namespace, conflict) -> int:
    """★ 挙動変更#3 の②: もう一方のシートを対象に**翻訳からやり直し**、その解釈行と
       --dry 相当のプレビューを見せた上で y/N を聞く。y ならそちらで実行、N ならやめる
       （原本は無変更）。この時点で原本にはまだ一切触れていない。"""
    book, source_book, struct_dump, sheets = a._rerun_ctx
    a._target_sheet = conflict.alternative
    # ★ 誤爆#3: ②を選んだ時点で「既定へ後退した」という判定は取り消された（そのシート
    #   自身が対象になる）。記録を落として、助言側のシート言及の抑制も効かせない。
    a._sheet_conflict = None
    original_dry, a.dry, a._preview_only = getattr(a, "dry", False), True, True
    rc = _translate_and_dispatch(a, book, source_book, struct_dump, sheets)
    a.dry, a._preview_only = original_dry, False
    if rc != 0:
        return rc
    if not ask_yes_no("この内容で実行しますか？ [y/N]: ",
                      interactive=is_interactive(stdin_isatty=_stdin_isatty(),
                                                 json_mode=getattr(a, "json", False),
                                                 dry=original_dry)):
        print(render_aborted())
        return 1
    a._reuse_translation = getattr(a, "_last_translation", None)   # 同じ翻訳を使い回す
    return _translate_and_dispatch(a, book, source_book, struct_dump, sheets)


def _make_dsl_step_deps() -> DslStepDeps:
    """★ C7: 呼び出しのたびに毎回新しく組み立てる依存の束（monkeypatch を効かせるため・dsl_step.py 参照）。"""
    return DslStepDeps(
        format_confirmation_line=format_confirmation_line,
        maybe_warn_header_col_mismatch=_maybe_warn_header_col_mismatch,
        maybe_warn_target_overwrite=_maybe_warn_target_overwrite,
        interpretation_summary_line=_interpretation_summary_line, confirm_overwrite_or_gate=_confirm_overwrite_or_gate,
        basrun_apply=basrun_apply, stop_office=_stop_office,   # ★ 摩擦⑥: LO 一時不調の再試行が使う
        snapshot=snapshot,
        # ★ 2026-08-25（復元の中10）: 適用直後の関門（原本に被せる前の最前線）。
        why_output_is_unusable=_why_output_is_unusable, diff_snapshots=diff_snapshots,
        run_postcondition=run_postcondition, progress_start=progress_start, progress_end=progress_end,
        pivot_caveat=PIVOT_CAVEAT, verify_dsl_args=verify_dsl_args,
        apply_new_column_fallback=_apply_new_column_fallback, build_advisories=build_advisories,
        structural_advisories=_structural_advisories, unrequested_new_sheet_advisory=unrequested_new_sheet_advisory,
        classify_subject_provenance=classify_subject_provenance,   # ★ 単位E
        sheet_conflict_gate=_sheet_conflict_gate)   # ★ 挙動変更#3


def cmd_run_dsl(a: argparse.Namespace, book: Path, source_book: Path, book_meta: dict,
                 op: str, raw_args: dict) -> int:
    """M2b の決定論パイプライン本体。②検証 → ③確認行 → ④codegen → ⑤適用 → ⑥事後条件。
       ★ W3: source_book は cmd_run が翻訳より前に正規化済み・やり直さない。
       ★ C7: ②③⑤⑥は ailine_core.dsl_step の共有エンジン（_run_dsl_plan_step と同じコード・
       「単発 = 1 段の計画」）。--ask・.bas/変更点の印字・_truncation_notice は単発固有のまま。"""
    vocab = load_vocab()
    deps = _make_dsl_step_deps()
    # ★ 挙動変更#2: 対象シートは _cmd_run_dispatch が resolve_target_sheet で一箇所だけ
    #   決めた a._target_sheet を読む（後方互換フォールバックは1枚目・旧挙動と同一）。
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    ground = resolve_dsl_step_args(op, raw_args, a.task, book_meta, vocab, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        for _ln in render_refusal(op, raw_args, ground.err):
            print(_ln)
        return 3
    resolved, inferred = ground.resolved, ground.inferred

    # ★ 挙動変更#2: header_row は「本当の」対象シート(resolved["_target_sheet"])で引き直す。
    #   LOOKUP_FILL は自分の target_sheet slot が最終的な正で、依頼文が参照シート(source_sheet)
    #   の名前も含む場合に一般解決(a._target_sheet・op を知る前の機械的な文字列一致)が
    #   参照シート側へ寄る可能性があるため、op が分かった後のここで必ず読み直す
    #   （他の op は resolved["_target_sheet"] == first_sheet のまま・実質無変化）。
    first_sheet = resolved.get("_target_sheet") or first_sheet
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    use_formula = not getattr(a, "values", False)

    print(render_run_header("DSL 経路", a.model, book.name))
    confirm = print_dsl_confirmation(op, resolved, inferred, a.task, meta=book_meta, warn_book=book, new_cols=None, a=a, deps=deps)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所
        return confirm.gate_exit

    if a.ask:
        try:
            ans = input("続行しますか？ [y/N]: ").strip().lower()
        except EOFError:
            ans = ""
        if ans not in ("y", "yes"):
            print(render_aborted())
            return 1

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = out_book_path(book)
    apply_timeout = a.timeout if a.timeout else None   # 0 で無効化（旧挙動 = 無制限）
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    _helper_catalog, helper_files = load_helpers(helpers_dir)

    code = codegen_dsl(op, resolved, book_meta, use_formula=use_formula)
    (workdir / "dsl_attempt.bas").write_text(code, encoding="utf-8")
    # ★ W8a 項目5: 「決定論」はユーザー向け文字列から排除（内部名・関数名は不変）。
    for ln in render_code_block("\n─ 生成した .bas（ルール変換・LLM不使用）───────────────", code):
        print(ln)

    # ★ 段1: interpretation/provenance は1箇所（build_interpretation）で組む
    #   （単位C の教訓 ―― 出所を運ぶ場所が2つあると片方だけ更新されて食い違う）。
    #   provenance は resolved["_sources"] をそのまま返す派生ビュー（値・型は不変）。
    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book.name])
    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              # ★ 2026-08-24（#17）: 実行した op を**機械の値として**持たせる。
              #   表示テキスト（解釈行）から復元してはいけない ── 文言を変えた瞬間に壊れる。
              "op": op,
              "path": "dsl", "command": confirm.line, "postcondition": None,
              "interpretation": interpretation, "provenance": provenance}

    if a.dry:
        # ★ 挙動変更#3: シート衝突②の内部プレビュー（_preview_and_run_on_alternative_sheet）
        #   では「--dry を外して実行」の案内は誤り（この直後に y/N を聞くため）。
        if not getattr(a, "_preview_only", False):
            print("（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        _finish_run(a, book, result, "none")
        return 0

    before = snapshot(source_book)
    # ★ 致命④: chart_paths は snapshot() の戻り値（golden 検体が丸ごと JSON へ落とす辞書）
    #   には入れない ── before_charts の数え上げと同じタイミングで、_charts_count の隣の
    #   集合版関数を直接呼ぶだけにする（同じファイル・まだ何も適用していない時点）。
    before_chart_paths = _chart_paths(source_book)
    shutil.copy2(source_book, out_book)   # 原本は触らず、正規化済みコピーに適用

    # ★ C7: ⑤適用〜⑥事後条件（共有エンジン）。print_changes は単発固有（docstring 参照）。
    apply_result = apply_dsl_step(
        op, resolved, code, apply_target=out_book, before=before, before_charts=before["charts"],
        before_chart_paths=before_chart_paths,
        workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout, header_row=header_row,
        use_formula=use_formula, source_book=source_book, deps=deps,
        apply_progress_label="⏳ LibreOffice で適用中…", print_changes=True)

    if apply_result.runtime_error is not None:
        print(f"× 実行時エラー: {short_error_summary(apply_result.runtime_error)}（詳細は履歴に記録）。")
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["last_error_full"] = apply_result.runtime_error
        _finish_run(a, book, result, "runtime_error", error_detail=apply_result.runtime_error)
        return 1

    after, lines = apply_result.after, apply_result.changes
    # ★ 止血3/C7: 単発は常に呼ぶ（★ C9 で複合計画の DSL 段の穴も塞いだ・dsl_step.py 参照）。
    notice = _truncation_notice(before, after, exhaustive_postcondition=True)
    if notice:
        print(notice)
    # ★ W10b 項目4b(摩擦): 参照専用シートは「変更なし」対象から除外。
    # ★ 単位C(D8): ここにあった `if op == "LOOKUP_FILL"` のハードコードは削除した ──
    #   除外すべきシートは OP_WRITE_TARGET の reads_only 宣言そのものなので、
    #   build_advisories が op/resolved から自分で求める（AGGREGATE/PIVOT の入力シートも同じ）。
    # ★ 誤爆#3: 対象シートを決めた側の判定結果（曖昧なので既定へ後退した）をそのまま運ぶ。
    # ★★ 単位G: 前提の検査を「助言を組む前」に move した（検査の中身は 単位F のまま）。
    #   理由: 中立化（「（既存シート『集計』の更新は意図どおりです）」）は助言の発生源で
    #   決まるので、前提が破れたことを 発生源が知っていなければならない。
    #   ★ 後から該当行を消す方式は採らない ── ailine.py の中立化は W10c/W10d で
    #   「出してから打ち消す」を捨てて発生源へ先取りした経緯がある（この関数の上の注記）。
    #   ★ 印字と関所は 位置を変えていない（後段のまま）。ここでやるのは判定だけ。
    precondition = _maybe_warn_write_precondition(op, before, after, resolved)
    precondition_broken = precondition[0] if precondition else None
    # ★★ 単位H 開示: 関所判定の直後（★ 単位H 導入前は完全な無言だった箇所）。助言ではない
    #   ── 「前回の自分の出力の作り直し」と判定して関所を黙らせた理由そのものを見せる。
    for own_notice in _maybe_own_prior_output_notice(op, before, after, resolved):
        print(own_notice)
    advisories = compose_dsl_step_advisories(   # mode="flat" は単発固有（dsl_step.py 参照）
        "flat", op, resolved, book_meta, a.task, before, after, deps=deps,
        sheet_conflict=getattr(a, "_sheet_conflict", None),
        precondition_broken=precondition_broken, after_path=out_book) + formula_error_advisory(source_book, out_book, cell_ref=_cell_ref) + broken_identity_advisory(source_book, out_book, resolved if isinstance(resolved, dict) else {})   # ★ 挙動変更#1(a)
    for adv in advisories:
        print(adv)
    result["changes"] = lines
    result["advisories"] = advisories

    status, reason = apply_result.postcondition_status, apply_result.postcondition_reason
    # ★ 止血1/2: "error"(チェッカー内の予期しない例外)は --json 上 "fail" に丸める。
    result["postcondition"] = "fail" if status == "error" else status
    if status == "error":
        print(f"\n× {reason}")
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_error")
        return 1
    if status == "fail":
        print(f"\n× 適用されたが事後条件を満たさない: {reason}")
        # ★ 2026-08-24: 「効かなかった」だけでなく心当たりも言う（1 実装・全経路）。
        for _ln in likely_cause_of_no_change(
                out_book, resolved.get("_target_sheet") if isinstance(resolved, dict) else None):
            print(_ln)
        print(_untouched_original_line(book, out_book))   # ★ C9: 失敗の沈黙を塞ぐ
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_fail")
        return 1
    if status == "warn":
        # ★ 止血1: 検証対象が少なすぎる場合、「機械検証済み」とは名乗らない。
        print(f"\n⚠ 事後条件を機械検証できなかった（操作:{OP_LABELS.get(op, op)}）: {reason}")
    else:
        # ★ C9: 事後条件が見た中身（例「3 行を検証（降順）」）はここで述べる。✓ とは呼ばない
        #   ―― ✓ は原本(--copy なら .out)が確定した後の1行だけ（_finish_apply）。
        print(f"\n事後条件を確認（操作:{OP_LABELS.get(op, op)}）: {reason}")
    result["ok"] = True

    # ★★ 単位F: 反映の直前（原本はまだ無傷・before/after は両方手元にある）に、宣言した
    #   書き込み領域の前提を確かめる。破れていたら**同じ破壊の関所**へ渡す（新しい関所も
    #   新しい exit code も作らない ―― 上書き警告と同じ引数の位置に、同じ形の1行を渡す）。
    warn_precondition = precondition[1] if precondition else None   # ★ 単位G: 上で 1 度だけ検査済み
    if warn_precondition:
        print(warn_precondition)
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition)
        if gate_exit is not None:
            return gate_exit

    # ★ C9: postcondition が warn（検証対象不足）なら ✓ は名乗らない。scope は「解釈: ...」行から
    #   「解釈: 」を除いた宣言テキスト（＝計画が宣言した対象）。
    # ★★ 単位E: ③（依頼文の語と矛盾する対象がある）なら、事後条件が通っていても ✓ は出さない
    #   ―― 検証したのは「計画どおり」であって、その計画は依頼文の語と食い違っている。
    # ★ 決裁③(2026-08-22): 疑わしい ⚠ の総数 = advisories 中の ★/⚠ 付き件数 + 単位F/G の
    #   前提破れ（あれば必ず ★ 付きの1件）+ 確認段で「⚠ 」前置で印字済みの
    #   resolved["_warnings"]（LLM 由来の値と機械抽出の食い違い ── 片配線の追補 2026-08-22:
    #   advisories にも前提破れにも入らないため、ここで明示的に数える。素の文字列は印を
    #   持たないので len で数える）。0 なら従来どおり ✓。
    warning_count = (count_suspicious_advisories(advisories) + (1 if warn_precondition else 0)
                      + len(resolved.get("_warnings", [])))
    # ★ 2026-08-25（復元の重大6・盲検）: _finish_apply は置換の成否を返すのに、
    #   呼び出し元 4 箇所が**全部戻り値を捨てていた**。バックアップに失敗して原本反映を
    #   中止しても exit 0 ── スクリプトから回す利用者は「反映されなかった」を検出できない。
    #   ★ 表示は正直だった（「原本は無変更」と言っていた）。嘘だったのは**終了コード**。
    _applied = _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(status != "warn" and not confirm.subject_warnings),
                   scope=confirm.label, scope_note="\n".join(render_scope_notes(list(confirm.unspoken))),
                   warning_count=warning_count)
    if not _applied:
        return _finish_failed_apply(a, book, result)

    _finish_run(a, book, result, "none")
    return 0


# ---------------------------------------------------------------------------
# ★ 帳票段: REPORT_PER_ROW 専用の実行経路（DESIGN-20260823-report-per-row.md）。
#   ②検証(resolve_dsl_step_args)・③確認(print_dsl_confirmation) は cmd_run_dsl と共有する
#   が、④codegen〜⑥事後条件は専用に組む ── 検分シート（③出所の置き場）を basrun_apply の
#   直後・run_postcondition の直前に Python 側で足す必要があり、共有エンジン
#   (dsl_step.apply_dsl_step) は basrun_apply→run_postcondition を1呼び出しの中で行うため
#   割り込めない。
# ---------------------------------------------------------------------------



def cmd_run_report_per_row(a: argparse.Namespace, book: Path, source_book: Path,
                            book_meta: dict, raw_args: dict) -> int:
    """REPORT_PER_ROW 専用の実行経路。②③は cmd_run_dsl と同じ器官を呼ぶ（DslStepDeps 経由）。
       ④codegen〜⑥事後条件は cmd_run_dsl と同じ手順を踏むが、apply_dsl_step は使わず
       basrun_apply→検分シート追加→事後条件、の順で自分で呼ぶ（検分シートを間に挟むため）。
       ★ ASSUMED: 摩擦⑥（LO 一時的な不調の1回リトライ）は本経路には持ち込まない
       （第一波・他 op と違い据え置き。発火したら通常の実行時エラーとして正直に失敗する）。"""
    op = "REPORT_PER_ROW"
    vocab = load_vocab()
    deps = _make_dsl_step_deps()
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    ground = resolve_dsl_step_args(op, raw_args, a.task, book_meta, vocab, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        for _ln in render_refusal(op, raw_args, ground.err):
            print(_ln)
        return 3
    resolved, inferred = ground.resolved, ground.inferred

    first_sheet = resolved.get("_target_sheet") or first_sheet
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    use_formula = not getattr(a, "values", False)

    print(render_run_header("DSL 経路", a.model, book.name))
    confirm = print_dsl_confirmation(op, resolved, inferred, a.task, meta=book_meta, warn_book=book,
                                      new_cols=None, a=a, deps=deps)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所
        return confirm.gate_exit

    if a.ask:
        try:
            ans = input("続行しますか？ [y/N]: ").strip().lower()
        except EOFError:
            ans = ""
        if ans not in ("y", "yes"):
            print(render_aborted())
            return 1

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = out_book_path(book)
    apply_timeout = a.timeout if a.timeout else None
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    _helper_catalog, helper_files = load_helpers(helpers_dir)

    code = codegen_dsl(op, resolved, book_meta, use_formula=use_formula)
    (workdir / "dsl_attempt.bas").write_text(code, encoding="utf-8")
    for ln in render_code_block("\n─ 生成した .bas（ルール変換・LLM不使用）───────────────", code):
        print(ln)

    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book.name])
    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              # ★ 2026-08-24（#17）: 実行した op を**機械の値として**持たせる。
              #   表示テキスト（解釈行）から復元してはいけない ── 文言を変えた瞬間に壊れる。
              "op": op,
              "path": "dsl", "command": confirm.line, "postcondition": None,
              "interpretation": interpretation, "provenance": provenance}

    if a.dry:
        if not getattr(a, "_preview_only", False):
            print("（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        _finish_run(a, book, result, "none")
        return 0

    before = snapshot(source_book)
    before_chart_paths = _chart_paths(source_book)
    shutil.copy2(source_book, out_book)

    t0 = progress_start("⏳ LibreOffice で適用中…")
    okrun, err_apply, _raw = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
    progress_end(t0)
    if not okrun:
        print(f"× 実行時エラー: {short_error_summary(err_apply)}（詳細は履歴に記録）。")
        print(_untouched_original_line(book, out_book))
        result["last_error_full"] = err_apply
        _finish_run(a, book, result, "runtime_error", error_detail=err_apply)
        return 1

    # ★ 2026-08-24（土台固め）: 検分シートは **Basic の中で**書くようになった
    #   （_dsl_to_basic の REPORT_PER_ROW 節）。ここで openpyxl で開き直していた旧経路は
    #   xl/drawings の中の図形を捨てるので撤去した ── 二重に書かないためでもある。

    after = snapshot(out_book)
    changed, lines = diff_snapshots(before, after)
    print("\n変更点:" if changed else "\n（文書に変化は検出されなかった）")
    for ln in lines:
        print(ln)

    notice = _truncation_notice(before, after, exhaustive_postcondition=True)
    if notice:
        print(notice)

    precondition = _maybe_warn_write_precondition(op, before, after, resolved)
    precondition_broken = precondition[0] if precondition else None
    for own_notice in _maybe_own_prior_output_notice(op, before, after, resolved):
        print(own_notice)
    advisories = compose_dsl_step_advisories(
        "flat", op, resolved, book_meta, a.task, before, after, deps=deps,
        sheet_conflict=getattr(a, "_sheet_conflict", None),
        precondition_broken=precondition_broken, after_path=out_book) + formula_error_advisory(
            source_book, out_book, cell_ref=_cell_ref) + broken_identity_advisory(
            source_book, out_book,
            resolved if isinstance(resolved, dict) else {})
    for adv in advisories:
        print(adv)
    result["changes"] = lines
    result["advisories"] = advisories

    status, reason = run_postcondition(op, out_book, resolved, before_charts=before["charts"],
                                        header_row=header_row, use_formula=use_formula,
                                        source_book=source_book, before_chart_paths=before_chart_paths)
    result["postcondition"] = "fail" if status == "error" else status
    if status == "error":
        print(f"\n× {reason}")
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_error")
        return 1
    if status == "fail":
        print(f"\n× 適用されたが事後条件を満たさない: {reason}")
        # ★ 2026-08-24: 「効かなかった」だけでなく心当たりも言う（1 実装・全経路）。
        for _ln in likely_cause_of_no_change(
                out_book, resolved.get("_target_sheet") if isinstance(resolved, dict) else None):
            print(_ln)
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_fail")
        return 1
    if status == "warn":
        print(f"\n⚠ 事後条件を機械検証できなかった（操作:{OP_LABELS.get(op, op)}）: {reason}")
    else:
        print(f"\n事後条件を確認（操作:{OP_LABELS.get(op, op)}）: {reason}")
    result["ok"] = True

    warn_precondition = precondition[1] if precondition else None
    if warn_precondition:
        print(warn_precondition)
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition)
        if gate_exit is not None:
            return gate_exit

    warning_count = (count_suspicious_advisories(advisories) + (1 if warn_precondition else 0)
                      + len(resolved.get("_warnings", [])))
    # ★ 2026-08-25（復元の重大6・盲検）: _finish_apply は置換の成否を返すのに、
    #   呼び出し元 4 箇所が**全部戻り値を捨てていた**。バックアップに失敗して原本反映を
    #   中止しても exit 0 ── スクリプトから回す利用者は「反映されなかった」を検出できない。
    #   ★ 表示は正直だった（「原本は無変更」と言っていた）。嘘だったのは**終了コード**。
    _applied = _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(status != "warn" and not confirm.subject_warnings),
                   scope=confirm.label, scope_note="\n".join(render_scope_notes(list(confirm.unspoken))),
                   warning_count=warning_count)
    if not _applied:
        return _finish_failed_apply(a, book, result)

    _finish_run(a, book, result, "none")
    return 0


# ---------------------------------------------------------------------------
# ★ 様式写像段: FORMAT_MAP 専用の実行経路（DESIGN-20260824-format-map.md）。
#   REPORT_PER_ROW の兄弟（縦の展開）── ②③は同じ器官(DslStepDeps)を共有し、
#   ④codegen〜⑥事後条件は cmd_run_report_per_row と全く同じ形（検分シートを
#   basrun_apply の直後・run_postcondition の直前に Python 側で足す）で組む。
# ---------------------------------------------------------------------------



def cmd_run_format_map(a: argparse.Namespace, book: Path, source_book: Path,
                        book_meta: dict, raw_args: dict) -> int:
    """FORMAT_MAP 専用の実行経路。cmd_run_report_per_row と同じ構造（②③は共有・
       ④codegen〜⑥事後条件は自分で basrun_apply→検分シート追加→事後条件、の順で呼ぶ）。
       ★ ASSUMED: REPORT_PER_ROW と同じく摩擦⑥（LO 一時不調の1回リトライ）は本経路に
       持ち込まない（第一波・発火したら通常の実行時エラーとして正直に失敗する）。"""
    op = "FORMAT_MAP"
    vocab = load_vocab()
    deps = _make_dsl_step_deps()
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    ground = resolve_dsl_step_args(op, raw_args, a.task, book_meta, vocab, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        for _ln in render_refusal(op, raw_args, ground.err):
            print(_ln)
        return 3
    resolved, inferred = ground.resolved, ground.inferred

    first_sheet = resolved.get("_target_sheet") or first_sheet
    header_row = book_meta.get("header_rows", {}).get(first_sheet, 1)
    use_formula = not getattr(a, "values", False)

    print(render_run_header("DSL 経路", a.model, book.name))
    confirm = print_dsl_confirmation(op, resolved, inferred, a.task, meta=book_meta, warn_book=book,
                                      new_cols=None, a=a, deps=deps)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所
        return confirm.gate_exit

    if a.ask:
        try:
            ans = input("続行しますか？ [y/N]: ").strip().lower()
        except EOFError:
            ans = ""
        if ans not in ("y", "yes"):
            print(render_aborted())
            return 1

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = out_book_path(book)
    apply_timeout = a.timeout if a.timeout else None
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    _helper_catalog, helper_files = load_helpers(helpers_dir)

    code = codegen_dsl(op, resolved, book_meta, use_formula=use_formula)
    (workdir / "dsl_attempt.bas").write_text(code, encoding="utf-8")
    for ln in render_code_block("\n─ 生成した .bas（ルール変換・LLM不使用）───────────────", code):
        print(ln)

    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book.name])
    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              # ★ 2026-08-24（#17）: 実行した op を**機械の値として**持たせる。
              #   表示テキスト（解釈行）から復元してはいけない ── 文言を変えた瞬間に壊れる。
              "op": op,
              "path": "dsl", "command": confirm.line, "postcondition": None,
              "interpretation": interpretation, "provenance": provenance}

    if a.dry:
        if not getattr(a, "_preview_only", False):
            print("（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        _finish_run(a, book, result, "none")
        return 0

    before = snapshot(source_book)
    before_chart_paths = _chart_paths(source_book)
    shutil.copy2(source_book, out_book)

    t0 = progress_start("⏳ LibreOffice で適用中…")
    okrun, err_apply, _raw = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
    progress_end(t0)
    if not okrun:
        print(f"× 実行時エラー: {short_error_summary(err_apply)}（詳細は履歴に記録）。")
        print(_untouched_original_line(book, out_book))
        result["last_error_full"] = err_apply
        _finish_run(a, book, result, "runtime_error", error_detail=err_apply)
        return 1

    # ★ 2026-08-24（土台固め）: 検分シートは Basic の中で書くようになった
    #   （_dsl_to_basic の FORMAT_MAP 節）。openpyxl で開き直す旧経路は図形を捨てるので撤去。

    after = snapshot(out_book)
    changed, lines = diff_snapshots(before, after)
    print("\n変更点:" if changed else "\n（文書に変化は検出されなかった）")
    for ln in lines:
        print(ln)

    notice = _truncation_notice(before, after, exhaustive_postcondition=True)
    if notice:
        print(notice)

    precondition = _maybe_warn_write_precondition(op, before, after, resolved)
    precondition_broken = precondition[0] if precondition else None
    for own_notice in _maybe_own_prior_output_notice(op, before, after, resolved):
        print(own_notice)
    advisories = compose_dsl_step_advisories(
        "flat", op, resolved, book_meta, a.task, before, after, deps=deps,
        sheet_conflict=getattr(a, "_sheet_conflict", None),
        precondition_broken=precondition_broken, after_path=out_book) + formula_error_advisory(
            source_book, out_book, cell_ref=_cell_ref) + broken_identity_advisory(
            source_book, out_book,
            resolved if isinstance(resolved, dict) else {})
    for adv in advisories:
        print(adv)
    result["changes"] = lines
    result["advisories"] = advisories

    status, reason = run_postcondition(op, out_book, resolved, before_charts=before["charts"],
                                        header_row=header_row, use_formula=use_formula,
                                        source_book=source_book, before_chart_paths=before_chart_paths)
    result["postcondition"] = "fail" if status == "error" else status
    if status == "error":
        print(f"\n× {reason}")
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_error")
        return 1
    if status == "fail":
        print(f"\n× 適用されたが事後条件を満たさない: {reason}")
        # ★ 2026-08-24: 「効かなかった」だけでなく心当たりも言う（1 実装・全経路）。
        for _ln in likely_cause_of_no_change(
                out_book, resolved.get("_target_sheet") if isinstance(resolved, dict) else None):
            print(_ln)
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "postcondition_fail")
        return 1
    if status == "warn":
        print(f"\n⚠ 事後条件を機械検証できなかった（操作:{OP_LABELS.get(op, op)}）: {reason}")
    else:
        print(f"\n事後条件を確認（操作:{OP_LABELS.get(op, op)}）: {reason}")
    result["ok"] = True

    warn_precondition = precondition[1] if precondition else None
    if warn_precondition:
        print(warn_precondition)
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition)
        if gate_exit is not None:
            return gate_exit

    warning_count = (count_suspicious_advisories(advisories) + (1 if warn_precondition else 0)
                      + len(resolved.get("_warnings", [])))
    # ★ 2026-08-25（復元の重大6・盲検）: _finish_apply は置換の成否を返すのに、
    #   呼び出し元 4 箇所が**全部戻り値を捨てていた**。バックアップに失敗して原本反映を
    #   中止しても exit 0 ── スクリプトから回す利用者は「反映されなかった」を検出できない。
    #   ★ 表示は正直だった（「原本は無変更」と言っていた）。嘘だったのは**終了コード**。
    _applied = _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(status != "warn" and not confirm.subject_warnings),
                   scope=confirm.label, scope_note="\n".join(render_scope_notes(list(confirm.unspoken))),
                   warning_count=warning_count)
    if not _applied:
        return _finish_failed_apply(a, book, result)

    _finish_run(a, book, result, "none")
    return 0


# ---------------------------------------------------------------------------
# ★ W8a 項目4: 率リテラルの機械スキャン（判断棚から昇格）
#   自由生成(FREEFORM/OUT_OF_VOCAB・単段)された Basic コードに、依頼文にも用語集にも
#   出典の無い『率らしい数値』（消費税8%のつもりの 0.08、税込計算の *1.1 等）が紛れて
#   いないかを機械で見る。実測: 単段の自由生成では「8% 仮定」がノーチェックで
#   「✓できました」を素通りしていた（勝手な税率仮定はモデルの幻覚と見分けがつかない）。
#   ★ 保守的: コメント行(') は対象外・整数(小数点なし)は対象外。発火してもブロックはせず
#   「検算してください」の助言止まり（no-op ガード等と違い、正しさの断定はしない）。
# ---------------------------------------------------------------------------

_BASIC_COMMENT_LINE_RE = re.compile(r"^\s*(?:'|REM\b)", re.I)
_RATE_LITERAL_RE = re.compile(r"(?<![\w.])(\d+\.\d+)(?![\w.])")

# ★ W10b 項目2: 修復ループが行き詰まると『総当たりで色々試す』方向に暴走し、依頼と無関係な
#   操作（ヘルパの全呼び出し等）を追加することがあった実測（operator 第6回査定・3回連続
#   再現。詳細は _known_helper_names/detect_helper_sweep のコメント）。修復メッセージ自体に、
#   依頼の範囲を超えるな、という制約を明示的に加える（no-op/実行時エラーの2種のみ・
#   bad_signature/truncated は構造の修正だけなので対象外）。
_REPAIR_SCOPE_GUARD = "★ 元の依頼に無い操作(無関係なヘルパ呼び出しを含む)を追加してはいけない。"


def _looks_like_rate(value: float) -> bool:
    """0.05〜0.2（例: 消費税8%→0.08）または 1.05〜1.2（例: 税込計算の *1.1）の小数か。"""
    return (0.05 <= value <= 0.2) or (1.05 <= value <= 1.2)


def _rate_explained(value: float, task: str, vocab: dict) -> bool:
    """率らしい数値が依頼文/用語集のどこかで説明されているか（保守的＝広めに許して
       誤検知を避ける）。①コードの生の数字がそのまま依頼文にある ②%/％表記に換算した
       値が依頼文にある ③用語集の登録値が同じ率を指す、のいずれか。"""
    factor = value if value >= 1 else 1 + value
    frac = value if value < 1 else value - 1
    pct = frac * 100
    if any(lit in task for lit in {f"{value:g}", f"{factor:g}", f"{frac:g}"} if lit):
        return True
    if any(f"{p}%" in task or f"{p}％" in task for p in {f"{pct:g}", f"{pct:.0f}"}):
        return True
    for v in vocab.values():
        try:
            fv = float(v)
        except (TypeError, ValueError):
            continue
        if abs(fv - factor) < 1e-6:
            return True
    return False


def scan_rate_literals(code: str, task: str, vocab: dict | None = None) -> list:
    """生成 Basic コード中の『率らしい数値リテラル』のうち、依頼文にも用語集にも出典が
       無いものを、検算を促す助言のリストにする（同じ数値は1回だけ報告）。"""
    vocab = vocab or {}
    seen: set = set()
    out: list = []
    for line in code.splitlines():
        if _BASIC_COMMENT_LINE_RE.match(line):
            continue
        for m in _RATE_LITERAL_RE.finditer(line):
            raw = m.group(1)
            if raw in seen:
                continue
            value = float(raw)
            if not _looks_like_rate(value):
                continue
            if _rate_explained(value, task, vocab):
                continue
            seen.add(raw)
            out.append(f"★ 率らしい数値 ({raw}) が依頼に無いのに使われています — 検算してください")
    return out


# ---------------------------------------------------------------------------
# ★ W10b 項目1/2: 自由生成の関所とヘルパ総なめ検出
#   operator 第6回ブラインド査定（実機）で確定した致命: 自由生成(FREEFORM/OUT_OF_VOCAB)は
#   事後条件チェッカーが無い唯一の経路なのに、成功時は「⚠ 機械保証なし」のソフト警告だけで
#   無条件に適用されていた。修復ループが行き詰まると、依頼と無関係な操作（ヘルパの全呼び出し）
#   を書いて「何か効くかもしれない」を試す暴走（ヘルパ総なめ）が3回連続で再現した。
#   検証できないものは人に確認を返す（忠実度ゲート/破壊の関所と同じ思想）。
# ---------------------------------------------------------------------------

class _FreeformGateAbort(Exception):
    """★ W10b 項目1: run_freeform_plan_step の関所で人が拒否/非対話で確認できなかった時、
       cmd_run_plan まで一気に抜けるための内部シグナル（_confirm_overwrite_or_gate が
       cmd_run_plan の同じフレームで直接 exit code を return しているのと同じ扱いを、
       別関数をまたいでも実現する）。exit_code は呼び出し元がそのまま return すべき値。"""
    def __init__(self, exit_code: int):
        super().__init__(exit_code)
        self.exit_code = exit_code


def _confirm_freeform_apply(a: argparse.Namespace, sweep_warning: str | None = None,
                             step_prefix: str = "") -> int | None:
    """★ W10b 項目1: 自由生成の関所。DSL 経路（②検証→③確認→④codegen→⑤適用→⑥事後条件）
       と違い、自由生成は「変化したか」しか機械確認できず「正しいか」は検証できない唯一の
       経路（事後条件チェッカーが無い）。生成コードを見せた直後・適用の直前に必ず人に確認を
       返す。--allow-freeform が既に立っていれば承知の上として素通し。
       sweep_warning があれば y/N の直前に強調表示する（★ 項目2: ヘルパ総なめの疑い）。
       戻り値: 続行してよければ None、中断すべきなら呼び出し側がそのまま return すべき
       exit code（対話で拒否=1・非対話で確認できない=8）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。単発 FREEFORM では空文字。"""
    if getattr(a, "allow_freeform", False):
        return None
    if sweep_warning:
        print(f"{step_prefix}{sweep_warning}")
    try:
        ans = input(f"{step_prefix}このコードは機械検証できません。適用しますか？ [y/N]: ").strip().lower()
    except EOFError:
        for ln in render_retry_options(step_prefix, [
                ("--allow-freeform", "機械検証できないことを承知の上で適用する")]):
            print(ln)
        return 8
    if ans not in ("y", "yes"):
        print(render_aborted(step_prefix))
        return 1
    return None


_HELPER_SUB_DECL_RE = re.compile(r"^\s*Sub\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", re.M)
_HELPER_CALL_RE = re.compile(r"\bCall\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(")
# ★ 根拠: bench 再現実験（W10b・demo/expense_w10b.xlsx 相当のダミー表で「氏名の列を全部
#   『退職済み』に書き換えて」を実行）で、修復3回目のヘルパ総なめは異なるヘルパ7種
#   （AutoFitColumns/AlignCenter/FormatThousands/VLookupFromTable/PivotSum/SummaryTable/
#   StyleBold）を Call していた一方、正常な単一操作の自由生成は0〜1種しか呼ばない。
#   間の4を閾値に採用（operator の所見「単一操作でCallが4個以上」とも一致）。
_HELPER_SWEEP_THRESHOLD = 4


def _known_helper_names(helper_files: list) -> set:
    """helper_files（load_helpers が返す .bas ファイル一覧）から Sub 名を集める
       （ヘルパ総なめ検出の対象を『実在するヘルパ』だけに絞るため。ユーザー定義の
       Sub まで数えると通常の複数ヘルパ利用と区別できなくなる）。"""
    names: set = set()
    for f in helper_files:
        try:
            names |= set(_HELPER_SUB_DECL_RE.findall(f.read_text(encoding="utf-8")))
        except OSError:
            continue
    return names


def detect_helper_sweep(code: str, helper_names: set) -> str | None:
    """★ W10b 項目2: 自由生成コードが異なるヘルパを閾値(_HELPER_SWEEP_THRESHOLD)以上
       呼んでいたら、依頼と無関係な操作を全部 Call する『ヘルパ総なめ』の疑いを1行で
       返す（無ければ None）。ブロックはしない（誤検知の可能性は残る）— 関所の y/N の
       直前に強調表示し、人の判断を後押しする助言止まり。"""
    if not helper_names:
        return None
    called = set(_HELPER_CALL_RE.findall(code)) & helper_names
    if len(called) < _HELPER_SWEEP_THRESHOLD:
        return None
    names = "、".join(sorted(called))
    return (f"🚨 疑わしい: {len(called)} 種類のヘルパを呼んでいます（{names}）"
            "— 依頼と無関係な操作が混じっていないか確認してください")


# ★ W10 前提工事①（architect レビュー致命5-2）: 単発の語彙外(FREEFORM/OUT_OF_VOCAB)が
#   落ちた理由を history の failure_kind に持ち込むための対応表。上位ラベル「語彙外」は
#   接頭辞として残す（format_history_table の表示互換のため・表示自体は従来どおり
#   「語彙外」に畳んで良いが、記録側は区分を保持するというレビュー指示）。
_VOCAB_MISS_KIND_PREFIX = "語彙外"
_VOCAB_MISS_REASONS = ("out_of_vocab", "slot_missing", "translate_error")


def _vocab_miss_reason(step: dict) -> str:
    """単発の語彙外 step から失敗理由を1つに決める。
       - op="OUT_OF_VOCAB"（モデルが明示的に「照合したが無い」と答えた）→ out_of_vocab
       - op="FREEFORM" は _normalize_plan_step / translate_task が付けた
         "_fail_reason"（out_of_vocab/slot_missing/translate_error）をそのまま使う。
       - _fail_reason が無い（例: テストが translate_task 自体を monkeypatch して
         生の {"op": "FREEFORM"} を返す等、正規化を経ていない経路）は out_of_vocab を
         既定にする ── 「未知の理由」を「翻訳失敗」と偽って接続エラー文言を出すよりは
         安全側（照合できなかった、が最も広く成立する既定）。"""
    if step.get("op") == "OUT_OF_VOCAB":
        return "out_of_vocab"
    reason = step.get("_fail_reason")
    return reason if reason in _VOCAB_MISS_REASONS else "out_of_vocab"


def cmd_refuse_vocab_miss(a: argparse.Namespace, book: Path, step: dict | None = None) -> int:
    """★ freeform 最終決定（DESIGN-20260821-multifile.md「freeform 最終決定」節・
       Namakoo 2026-08-21 19:37「廃止しよう」で確定）: 単発の語彙外（翻訳が
       FREEFORM/OUT_OF_VOCAB を返した経路）は生成に入らず即座に断る。
       旧 cmd_run_freeform（AI 直接生成→適用ループ）はここで廃止した ── 理由は設計書
       参照（不完全な機能が見えたままだと信頼感が失われる・警告は査定から守らない
       （実測）・可逆性の非対称（廃止は git から復活可・悪印象は不可逆）・恩恵層が
       実測上不在）。将来の復活は発火条件つきで設計書側に凍結（ここには実装しない）。

       ★ vocab_miss の記録: history.jsonl に failure_kind="語彙外/<理由>" で残す
       （依頼文・book・ts は build_history_entry が普段どおり詰める）。★ W10 前提工事①
       （architect レビュー致命5-2）: 従来は理由を問わず一色で「語彙外」だった
       （未知op/必須slot欠落/translate_task自体の例外・空応答が全部合流していた）。
       ここで _vocab_miss_reason() が out_of_vocab/slot_missing/translate_error に
       仕分ける（上位ラベル「語彙外」は接頭辞として残す）。生成も適用も一切していない
       ので result は ok=False・attempts=0・changes=[] のまま ―― _finish_run/
       build_history_entry を素通しで再利用する（頻度×原始性の二軸で開発キューへ、
       という設計書の使い道に machine-readable な形で残す）。

       ★ 断りの文言は理由で変えない ── ただし translate_error（ollama 不通/JSON不正/
       空応答で翻訳がそもそも DSL の形にならなかった経路）だけは「頼める操作の一覧に
       照合できませんでした」と言うと嘘になる（照合を試みてすらいない）ため、
       render_vocab_miss_refusal に translate_error=True を渡して理由行を差し替える。

       ★ --allow-freeform は受理する（後方互換のため flag 自体は argparse に残す）が、
       廃止告知を1行足すだけで断り自体は変えない（自由生成そのものへは戻らない）。

       ★ 複合計画の語彙外段（run_freeform_plan_step）はここを経由しない ── このブリーフの
       対象は単発経路だけ（compound-plan 側は意図的に変えていない）。"""
    step = step or {}
    about = str(step.get("about") or "").strip()
    reason = _vocab_miss_reason(step)
    for ln in render_vocab_miss_refusal(about, sunset_notice=bool(getattr(a, "allow_freeform", False)),
                                         translate_error=(reason == "translate_error")):
        print(ln)
    result = {"ok": False, "attempts": 0, "task": a.task, "model": a.model,
              "path": "vocab_miss", "command": None, "postcondition": None,
              "changes": [], "out": str(book)}
    _finish_run(a, book, result, failure_kind=f"{_VOCAB_MISS_KIND_PREFIX}/{reason}")
    # ★ 2026-08-24（初回体験の盲検・致命①）: translate_error（ollama 不通・モデル未取得・
    #   空応答）は**語彙の問題ではなく環境の問題**なのに、語彙外と同じ exit 3 を返していた。
    #   README の終了コード表は「9 = 実行の前提が無い（ollama 不通・モデル未取得）」と
    #   書いており、道具と文書が食い違っていた。
    #   ★ 実害: 使う側に「言い換えてください」と案内する ── 直し方が言い回しでは無いのに。
    return EXIT_ENVIRONMENT if reason == "translate_error" else 3


# ---------------------------------------------------------------------------
# ★ W10 便C2: もしかして提案の UX 配線。out_of_vocab（_vocab_miss_reason 参照）の断りの
#   直前に挟む。slot_missing/translate_error はここに来ず従来の断りへ直行する
#   （REVIEW-20260822-w10-architect.md「フローの設計」・Namakoo 決裁）。
# ---------------------------------------------------------------------------

def _ground_via_fixed_op(a: argparse.Namespace, task: str, book_meta: dict, op: str):
    """op を固定した二段目翻訳(translate_task_fixed_op)→接地(verify_dsl_args)を1回で行う。
       戻り値: 成功すれば (raw_args, resolved, inferred)。二段目翻訳の応答が壊れている
       （None・op不一致・args が dict でない）か、接地に失敗すれば (None, None, None)。
       段0(別名直行)と段1/2(もしかして提案)の両方がこの同じ手順を使う ── 二段目翻訳の
       スキーマ（op 固定）と接地の規則を2箇所に書くと片方だけ直されて食い違う事故になる。"""
    fixed = translate_task_fixed_op(a.model, op, task, book_meta, temperature=0.1)
    if not fixed or fixed.get("op") != op or not isinstance(fixed.get("args"), dict):
        return None, None, None
    ok, resolved, inferred, _err = verify_dsl_args(
        op, fixed["args"], book_meta, task=task, vocab=load_vocab(),
        target_sheet=getattr(a, "_target_sheet", None))
    if not ok:
        return None, None, None
    return fixed["args"], resolved, inferred


def _maybe_suggest_or_refuse(a: argparse.Namespace, book: Path, source_book: Path,
                              book_meta: dict, sheets: list, step: dict) -> int:
    """★ W10 便C2: もしかして提案の入口。out_of_vocab でだけ発火する（それ以外の理由は
       従来の断りへ直行）。
       段0: lookup_alias がヒットしたら、もしかしてを経ずに二段目翻訳→接地→適用まで直行
       する（登録済みの言い回しは、既に一度 y で頷いた実績がある信頼できる経路・S1後半）。
       段1: suggest_ops（厳格一致・0秒）。空なら段2: judge_ops_via_llm（+3秒・進捗表示）
       ── ★ about は段2にだけ渡す（段1に渡すと一次翻訳の要約自体が語彙のラベルを含む
       ことがあり、厳格一致が常に先取りしてしまい判定器の出番が無くなる。凍結検体
       test_judge_tier_fires_only_when_strict_empty の実測で判明）。
       候補の先頭だけを試す（外れたら2番目は試さず正直に断る）。
       二段目翻訳→接地まで通ったら、もしかしてブロック（解釈行+残差行）を見せて y/N を聞く
       （非対話は ask_yes_no が既定 False を返すだけ・表示は既に済んでいる＝S4）。
       y: 通常の run フロー(cmd_run_dsl)で適用し、成功後に暗黙登録+開示（S1）。
       N: 従来の断り + misclass 第3信号 signal="suggest_decline"（対話で明示的に断った時
       だけ記録する・S3）。
       候補なし/二段目翻訳の応答が壊れている/接地失敗: 従来の断りのまま。"""
    if _vocab_miss_reason(step) != "out_of_vocab":
        return cmd_refuse_vocab_miss(a, book, step)
    task = a.task
    about = str(step.get("about") or "").strip() or None

    alias_op = lookup_alias(task)
    if alias_op:
        args, _resolved, _inferred = _ground_via_fixed_op(a, task, book_meta, alias_op)
        if args is not None:
            return cmd_run_dsl(a, book, source_book, book_meta, alias_op, args)
        return cmd_refuse_vocab_miss(a, book, step)

    candidates = suggest_ops(task)
    if not candidates:
        t0 = progress_start("⏳ 似た操作を確認中…")
        candidates = judge_ops_via_llm(task, about)
        progress_end(t0)
    # ★ #17: それでも候補が無いなら、**同じ依頼で前に通った op** を思い出す。
    #   既存の候補は押しのけない（候補が出ている回はこの経路に入らない）。
    remembered_op, remembered_on = (None, None)
    if not candidates:
        remembered_op, remembered_on = op_that_worked_before(
            task, read_history(max_n=HISTORY_RECALL_MAX))
        if remembered_op:
            candidates = [remembered_op]
    if not candidates:
        return cmd_refuse_vocab_miss(a, book, step)

    op = candidates[0]
    args, resolved, inferred = _ground_via_fixed_op(a, task, book_meta, op)
    if args is None:
        return cmd_refuse_vocab_miss(a, book, step)

    if remembered_op and op == remembered_op:
        # ★ #17: 出す時は必ず根拠を言う（黙って前回の op を実行しない）。
        print(f"（前回この依頼は『{OP_LABELS.get(op, op)}』で通っています"
              f"{'・' + remembered_on if remembered_on else ''}）")
    print(f"もしかして: {OP_LABELS.get(op, op)}？")
    print(format_confirmation_line(op, resolved, inferred, sheets=sheets,
                                    target_sheet=resolved.get("_target_sheet")))
    note = suggest_residue.render_residue_note(
        suggest_residue.find_unconsumed_words(task, resolved, _op_match_pool(op)))
    if note:
        print(note)

    interactive = is_interactive(stdin_isatty=_stdin_isatty(), json_mode=getattr(a, "json", False),
                                  dry=getattr(a, "dry", False))
    if not ask_yes_no("この操作を適用しますか？ [y/N]: ", interactive=interactive, default=False):
        if interactive:   # ★ 非対話の素通りは「断った」のではない ── 第3信号は N の時だけ
            _record_misclass_suspect("suggest_decline", task,
                                      getattr(a, "_last_translation", None) or {}, book)
        return cmd_refuse_vocab_miss(a, book, step)

    rc = cmd_run_dsl(a, book, source_book, book_meta, op, args)
    if rc == 0:
        # ★ 致命⑥(2026-08-23レビュー): save_alias の bool を見ずに常に「登録しました」と
        #   断定していた（実測: 40字超の言い回しで登録に失敗しても成功を騙る）。
        #   alias_add の (ok, msg) 形で理由を受け取り、失敗時は正直に開示する。
        ok, msg = alias_add(task, op)
        if ok:
            print("この言い回しを登録しました（alias undo で取り消せます）")
        else:
            print(f"{msg}（この言い回しは登録されませんでした。"
                  "次回もまた「もしかして」で確認します）")
    return rc


# ---------------------------------------------------------------------------
# M2c: 複合依頼の計画実行と正直な範囲表示
#   翻訳(①)が返した plan(長さ2以上)を段ごとに実行する。DSL 語彙の段は②〜⑥の決定論
#   パイプライン、語彙外(OUT_OF_VOCAB/FREEFORM)の段は FREEFORM 経路（その段の依頼文だけ）。
#   ★ 黙落ゼロ: 計画に載った段は必ず項目別報告の1行になる。
#   ★ 総合判定は最弱の段に従う。段別の行は evidence だけを述べ、✓ とは呼ばない（C9）。
#   ★ C5: _ITEM_STATUS_MARK / _VERIFY_SCOPE_NOTE(_PLAN) / format_plan_report /
#   overall_verdict は ailine_core/claim.py に移した（『✓』相当の文字列はレンダラ1箇所
#   からしか出さない・冒頭の import で ailine.format_plan_report 等は従来どおり使える）。
#   ★ C9: その ✓ の発生点自体を _finish_apply（原本が確定した後）へ動かした。
# ---------------------------------------------------------------------------

# col系 slot を持つ op → その slot 名（依存つき連鎖の新規列フォールバック対象）。
_COLUMN_ARG_KEYS = {
    "SORT": ("col",), "NUMBER_FORMAT": ("col",), "CHART": ("value_col", "category_col"),
    "AGGREGATE": ("group_col", "value_col"),
}

# ★ 致命1(W10e): 依頼文に「見出し」の語があるのに BOLD/FILL_COLOR/CENTER_ALIGN の対象が
#   直前までの段が新規作成した列に解決された場合の食い違いを検出する。実測事故の再現形
#   そのもの ―― COMPUTE_COLUMN は target 無指定(新規列作成)の見出しを式そのまま
#   （例:「数量*単価」）で書く（codegen_dsl 参照）。「見出しを太字にして」の段が、
#   複合依頼の翻訳で前段の演算対象名を引きずり、target がその「数量*単価」という
#   *実在する*（が事務上は無関係な）新規列名に解決されてしまうと、check_bold は
#   「対象:col:数量*単価」という*計画どおり*の検証には合格する（が依頼＝見出し行の
#   太字化とは無関係）。★ 保守的: 対象列が『この計画で直前までに新規作成された列』
#   （new_cols）であり、かつ依頼文に「見出し」とある場合だけ発火する（列名を最初から
#   明示したふつうの BOLD/FILL_COLOR/CENTER_ALIGN では絶対に発火しない）。
#   ★ _apply_new_column_fallback 経由(target が未解決語から救済された場合)・素の
#   verify_dsl_args 直解決(target が最初から新規列の実名と一致した場合)のどちらでも
#   同じ理屈で疑わしいため、fallback の発火有無ではなく new_cols 所属の有無で判定する
#   （fallback 有無だけを見ると後者を取りこぼす＝実測の再現形はむしろ後者に近い）。
_HEADER_WORD_RE = re.compile(r"見出し")


def _maybe_warn_header_col_mismatch(op: str, resolved: dict, new_cols: list, task: str) -> str | None:
    """★ 致命1: 上のコメント参照。target がこの計画の直前までに新規作成された列
       （new_cols）を指していて、かつ依頼文に「見出し」とある場合だけ、非ブロッキングの
       助言を返す（M2a と同じ思想＝保守的・ブロックしない・確認を促すだけ）。
       ★ 単位B: 一般則（③）が同じスロットで鳴る時は、この文は単独では出ず、③の ⚠ に
       「（この計画の直前の段で新規作成された列）」として畳み込まれる（dsl_step.py 参照）。
       事実の文言は NEW_COLUMN_ORIGIN 1箇所が持つ ―― 畳み込み側と文面がずれないため。"""
    if op not in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        return None
    target = resolved.get("target", "")
    if not (isinstance(target, str) and target.startswith("col:")):
        return None
    col_name = target[4:]
    if col_name not in (new_cols or []):
        return None
    if not _HEADER_WORD_RE.search(task or ""):
        return None
    return (f"⚠ 対象の列『{col_name}』は{NEW_COLUMN_ORIGIN}です。"
            "依頼に「見出し」とあるため、見出し行（行全体）を意図していないか確認してください")


def _apply_new_column_fallback(op: str, args: dict, headers: list, new_cols: list) -> dict:
    """★ M2c 依存つき連鎖（battery v2 #107 型）: 直前までの段が新規作成した列がちょうど
       1つあり、この段の列参照が現在の実列名のどれとも一致しない場合、その新規列を指して
       いるとみなして args を書き換える（候補が0か2つ以上なら何もしない＝保守的）。
       ★ 書き換えても最終的には verify_dsl_args が実在確認するので、誤った書き換えは
       通常どおりのエラーで止まる（無条件に信じ切らない）。"""
    if len(new_cols) != 1:
        return args
    only = new_cols[0]
    patched = dict(args)
    for k in _COLUMN_ARG_KEYS.get(op, ()):
        v = patched.get(k)
        if isinstance(v, str) and v not in headers and not re.fullmatch(r"\d+", v):
            patched[k] = only
    if op in ("BOLD", "FILL_COLOR", "CENTER_ALIGN"):
        t = patched.get("target", "")
        if isinstance(t, str) and t.startswith("col:"):
            name = t[4:]
            if name not in headers and not re.fullmatch(r"\d+", name):
                patched["target"] = f"col:{only}"
    if patched != args:
        # ★ 2026-08-24: 捏造列の関所（fabricated_subject_refusal）に、この書き換えは
        #   「依頼者の言った名前が前段の作った列を指している」正当な連鎖だと知らせる。
        #   これが無いと「利益で並べ替えて」（前段が『売上-原価』を作った）まで断ってしまう。
        patched["_chained_new_column"] = only
    return patched


def run_freeform_plan_step(a: argparse.Namespace, task_text: str, out_book: Path, workdir: Path,
                            refs_dir: Path, helpers_dir: Path, tag: str,
                            apply_timeout: float | None, step_prefix: str = "",
                            vocab: dict | None = None, op: str = "FREEFORM",
                            about: str = "") -> tuple:
    """M2c: 複合計画の語彙外(OUT_OF_VOCAB/FREEFORM)段を FREEFORM 経路で実行する。
       旧・単発 cmd_run_freeform と同じ生成→（★ W10b: 関所→）適用→署名/切断/no-op
       チェックのループを、『その段の依頼文だけ』かつ『out_book の現在の状態』を起点に
       行う版。★ freeform 最終決定（2026-08-21）: 単発側は cmd_refuse_vocab_miss に
       置き換わって cmd_run_freeform 自体は無くなったが、この複合計画側は意図的に
       変えていない（このブリーフの対象は単発経路だけ）── 生成が残る唯一の経路。
       ★ W10b 項目1: 関所で人が拒否/非対話で確認できなかった場合は _FreeformGateAbort を
       投げて cmd_run_plan まで一気に抜ける（破壊の関所と同じ『計画全体を止める』扱い。
       原本(book)はこの時点でまだ一切触れていない＝out_book はコピーなので安全）。
       step_prefix は複合計画の段番号表示用（例: "  2段目: "）。
       op/about は呼び出し側(cmd_run_plan)の step から渡す。★ K-1: 生成に入る前の通知の
       理由1行目に使うだけ（OUT_OF_VOCAB のときだけ about を添える）。
       ★ W10f 項目2: vocab（率リテラル走査用の用語集）は cmd_run_freeform（単発の自由生成）
       には元から渡っていた(scan_rate_literals)が、複合計画のこの経路には渡っておらず
       A' 原則（LLM に率や値を確定させない）の機械監査が複合計画の自由生成段だけ
       素通りしていた（独立監査が発見・査定5回は偶然当たらなかっただけの穴）。ここで
       同じ走査を通す。
       戻り値: (ok, changes:list[str], advisories:list[str], failure_kind:str, detail:str|None)"""
    helper_catalog, helper_files = load_helpers(helpers_dir)
    known_helper_names = _known_helper_names(helper_files)   # ★ W10b 項目2: 総なめ検出用
    system = CONTRACT + load_refs(refs_dir) + helper_catalog
    desc = describe_book(out_book)
    user = f"{desc}\n\nタスク:\n{task_text}\n\n`Sub Run(oDoc As Object)` を1つだけ書け。コードのみ。"
    msgs = [{"role": "system", "content": system}, {"role": "user", "content": user}]

    # ★ K-1: 語彙外に落ちる瞬間の通知（この段の生成が始まる前・1行に畳んだ版）。
    reason = freeform_notice_reason(op, about)
    print(render_freeform_notice_compact(reason, step_prefix=step_prefix))

    stepsource = workdir / f"{tag}_source{out_book.suffix}"
    shutil.copy2(out_book, stepsource)
    before = snapshot(stepsource)

    failure_kind = "none"
    for attempt in range(a.repair + 1):
        t0 = progress_start(f"⏳ 生成中（語彙外段・{a.model}）…")
        raw = ollama_generate(a.model, msgs, temperature=a.temperature)
        progress_end(t0)
        code = extract_bas(raw)
        (workdir / f"{tag}_attempt{attempt}.bas").write_text(code, encoding="utf-8")

        if not valid_signature(code):
            failure_kind = "bad_signature"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": "署名が違う。`Sub Run(oDoc As Object)` を1つだけ。コードのみ。"}]
            continue
        if is_truncated_code(code):
            failure_kind = "truncated"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "コードが途中で切れている（End Sub まで書き切れていない）。"
                      "最初から完全なコードを1つだけ書いて。コードのみ。"}]
            continue

        # ★ W10b 項目1: 自由生成の関所。旧・単発 cmd_run_freeform と違い、この経路はこれまで
        #   生成コードを一切表示していなかった（黙って確認を求めても判断できない）ので、
        #   ここで初めて表示してから y/N を聞く。
        for ln in render_code_block(f"{step_prefix}─ 生成した .bas（語彙外・AI が直接作成）───────────────",
                                     code, step_prefix=step_prefix):
            print(ln)
        sweep_warning = detect_helper_sweep(code, known_helper_names)
        gate_exit = _confirm_freeform_apply(a, sweep_warning, step_prefix=step_prefix)
        if gate_exit is not None:
            raise _FreeformGateAbort(gate_exit)

        shutil.copy2(stepsource, out_book)
        t0 = progress_start("⏳ LibreOffice で適用中…")
        ok, err, _raw = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
        progress_end(t0)
        if not ok and is_transient_lo_error(err):
            # ★ 摩擦⑥: LO の一時不調 ── コード自体の問題ではないので、LLM への「直して」
            #   依頼(msgs)は消費しない。stop → 無垢の stepsource から作り直し → 1回だけ
            #   再試行する（dsl_step.apply_dsl_step と同型・normalize_book/M2c 由来）。
            _stop_office()
            shutil.copy2(stepsource, out_book)
            print(f"{step_prefix}{TRANSIENT_LO_RETRY_NOTICE}")
            t0 = progress_start("⏳ LibreOffice で適用中…")
            ok, err, _raw = basrun_apply(out_book, code, workdir, helper_files, timeout=apply_timeout)
            progress_end(t0)
        if not ok:
            failure_kind = "runtime_error"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content": f"実行時エラー: {err}\nこれを直して。{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        after = snapshot(out_book)
        changed, lines = diff_snapshots(before, after)
        if not changed:
            failure_kind = "noop"
            msgs += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "実行は成功したが文書に一切変化が無かった（no-op）。"
                      "設定した API が効いていない可能性がある。別の正しい方法で書き直して。"
                      f"{_REPAIR_SCOPE_GUARD}コードのみ。"}]
            continue

        advisories = build_advisories(task_text, before, after,
                                       sheet_conflict=getattr(a, "_sheet_conflict", None),
                                       after_path=out_book)   # ★ 誤爆#3
        # ★ W10f 項目2: 旧・単発 cmd_run_freeform と同じ率リテラルの機械スキャン。この段の
        #   依頼文(task_text)だけを出典として見る（他段の依頼文言に混ざらないよう局所判定）。
        advisories = advisories + scan_rate_literals(code, task_text, vocab) + formula_error_advisory(stepsource, out_book, cell_ref=_cell_ref)   # ★ 挙動変更#1(a)
        # ★ 止血3: 呼び出し元(cmd_run_plan)は lines をそのまま「  {ln}」で表示するだけ
        #   なので、切り詰め注記はここで lines に混ぜて渡す。
        notice = _truncation_notice(before, after, exhaustive_postcondition=False)
        if notice:
            lines = lines + [notice]
        return True, lines, advisories, "none", None

    detail = {
        "bad_signature": "生成コードの署名が不正でした",
        "truncated": "生成コードが途中で切断されました",
        "runtime_error": "実行時エラーが解消しませんでした",
        "noop": "適用しても文書に変化がありませんでした",
    }.get(failure_kind, "原因不明で失敗しました")
    return False, [], [], failure_kind, detail


# ★ C5: format_plan_report / overall_verdict は ailine_core/claim.py に移した
#   （Claim を経由してしか『✓ 機械検証済み』相当の文字列を出さないレンダラ側に統合・
#   冒頭の import で ailine.format_plan_report / ailine.overall_verdict は従来どおり使える）。


# ★ W10d【本命】: 複合計画は段が増えるほど同じ助言（例: 幽霊データ検出）が段の数だけ
#   並びうる。査定で名指しされたオオカミ少年化を避けるため、文言が同じものは1行に畳んで
#   段番号を列挙する（初出順は保つ・文言そのものは変えない＝読み手が信じる根拠を削らない）。
def _group_step_advisories(entries: list) -> list:
    """entries: [(段番号 or None, 助言文言), ...]（None=計画全体に対する助言・段に紐付かない）
       を文言でグルーピングし、初出順を保った [(段番号のリスト, 文言), ...] を返す。
       印字用(_dedup_step_advisories)と --json 用(cmd_run_plan の result["advisories"])が
       同じグルーピングを共有する（二重管理しない）。"""
    order = []
    by_text: dict = {}
    for idx, text in entries:
        if text not in by_text:
            by_text[text] = []
            order.append(text)
        if idx is not None and idx not in by_text[text]:
            by_text[text].append(idx)
    return [(by_text[text], text) for text in order]


def _dedup_step_advisories(entries: list) -> list:
    """複合計画の全段から集めた助言を、印字用の行リストに整形する（同じ文言は1行に畳み、
       該当する段番号を列挙する）。"""
    lines = []
    for idxs, text in _group_step_advisories(entries):
        if idxs:
            prefix = "・".join(f"{i}段目" for i in idxs)
            lines.append(f"  {prefix}: {text}")
        else:
            lines.append(f"  {text}")
    return lines


def _run_dsl_plan_step(i: int, op: str, raw_args: dict, *, task: str, current_meta: dict, original_headers: dict,
                        first_sheet: str | None, out_book: Path, workdir: Path, helper_files, apply_timeout,
                        use_formula: bool, header_rows: dict, before_charts: int, a: argparse.Namespace, vocab: dict,
                        book_name: str, subject_sink: dict | None = None,
                        suspicion_sink: list | None = None,
                        derived_sheets: list | None = None,
                        before_chart_paths: frozenset | None = None) -> tuple:
    """★ C7: cmd_run_plan の DSL 語彙段の1段分。cmd_run_dsl と同じ ailine_core.dsl_step の共有エンジンを通る
       （非対称は dsl_step.py 参照）。この分離で stage_organs の dsl_plan_step 代表関数はここになる（DoD7）。
       戻り値: (gate_exit, item, plan_json_entry, step_advisories, provenance_entry, mention_exclude_sheets, current_meta)。
       ★ 単位E: subject_sink（呼び出し側が用意する dict）に、この段の対象スロットの出所を積む
       ―― ③ の有無は計画全体の ✓ を左右し、② は ✓ の直後の1文になるので、段の外へ運ぶ必要がある
       （戻り値のタプルはこれ以上広げない ―― 既存の unpack を壊さないための選択）。
       ★ 決裁③(2026-08-22): suspicion_sink（呼び出し側が用意する list）に、この段で
       「印字済みだが advisories に入らない疑わしい ⚠」を積む ── subject_sink と同じ
       side-channel の作法（戻り値のタプルは広げない）。advisories に混ぜると「助言:」の
       集約表示で二重表示になるため、こちらの棚で運ぶ。積むのは 2 家系（片配線の追補
       2026-08-22 で 1 → 2）: ①単位F/G の前提破れメッセージ（★ 付き・関所の直前で印字済み）
       ②resolved["_warnings"]（LLM 由来の値と機械抽出の食い違い・確認段で「⚠ 」前置で
       印字済み）。✓→△ 降格の判定材料としてだけ、呼び出し側 cmd_run_plan が集計する。"""
    step_prefix = f"  {i}段目: "
    deps = _make_dsl_step_deps()
    # ★ 連鎖: 前段の派生シートを対象にすべきならここで差し替える（決めるのは 1 箇所）。
    chained = chain_target_sheet(op, task, derived_sheets,
                                  current_meta.get("sheets") or [],
                                  current_meta.get("headers") or {})
    if chained:
        first_sheet = chained
    # 依存つき連鎖: 直前までの段の適用後の実列構成(current_meta)で接地する（新規列フォールバック込み）
    ground = resolve_dsl_step_args(op, raw_args, task, current_meta, vocab,
                                    original_headers=original_headers, first_sheet=first_sheet, deps=deps)
    if not ground.ok:
        return (None, (i, f"操作:{OP_LABELS.get(op, op)}", "fail", ground.err),
                {"op": op, "command": None, "status": "fail", "postcondition": None},
                [], None, None, current_meta)
    resolved, inferred = ground.resolved, ground.inferred

    # ★ 致命1(W10e) 要求2: 単発は元々この行を適用前に出す・複合計画は抜けていた（段番号付きで見せる）。
    confirm = print_dsl_confirmation(op, resolved, inferred, task, meta=current_meta, warn_book=out_book,
                                      new_cols=ground.new_cols, a=a, deps=deps, step_prefix=step_prefix)
    if confirm.gate_exit is not None:   # ★ W10a 項目1: 破壊の関所（複合計画の段ごと）
        return (confirm.gate_exit, None, None, [], None, None, current_meta)

    if subject_sink is not None:   # ★ 単位E: この段の対象スロットの出所を計画全体へ運ぶ
        subject_sink.setdefault("warnings", []).extend(confirm.subject_warnings)
        for phrase in confirm.unspoken:
            if phrase not in subject_sink.setdefault("unspoken", []):
                subject_sink["unspoken"].append(phrase)
    if suspicion_sink is not None:   # ★ 決裁③ 片配線の追補: 確認段で「⚠ 」前置印字済みの
        #   LLM/機械抽出の食い違い（print_dsl_confirmation 内で印字・advisories に入らない）
        #   を ✓→△ 降格の判定材料へ運ぶ（印字と同じ「⚠ 」前置の形で積む）
        suspicion_sink.extend(f"⚠ {w}" for w in resolved.get("_warnings", []))
    step_advisories = [confirm.mismatch_warning] if confirm.mismatch_warning else []
    # ★ operator9 ②: 段の存在自体に依頼文の根拠が無ければ名指しする（count_suspicious_advisories
    #   が ★ 付きを拾い、決裁③の ✓→△ 降格に自動で乗る）。"N段目:" の付与は呼び出し側
    #   （cmd_run_plan・_dedup_step_advisories）に任せる ── 他の ★ 付き助言と同じ作法。
    if not _op_has_task_grounding(op, resolved, task):
        step_advisories.append(
            f"★ （{OP_LABELS.get(op, op)}）は依頼文に根拠が見つかりません "
            "── 意図しない操作の可能性があります")
    # ★ 段1: interpretation/provenance は1箇所（build_interpretation）で組む（cmd_run_dsl と同じ）。
    #   provenance_entry の中身（キー・値）は今までと完全に同じ（resolved["_sources"] のまま）。
    interpretation, provenance = build_interpretation(op, resolved, inferred, confirm.verdicts, [book_name])
    provenance_entry = {"step": i, **provenance} if provenance else None

    # ★ 挙動変更#2: cmd_run_dsl と同じ理由（コメント参照）で、この段の「本当の」対象シート
    #   (resolved["_target_sheet"]) を優先する（LOOKUP_FILL 段が計画全体の対象シートと
    #   異なる参照シートを持つ場合の header_row 取り違えを避ける）。
    step_target_sheet = resolved.get("_target_sheet") or first_sheet
    # ★ 連鎖の番人: 前段が派生シートを作ったのに、この段が別のシート（＝元表）を見ていたら
    #   名指しする。★ 付きなので count_suspicious_advisories が拾い、決裁③の降格に乗る。
    if chained and step_target_sheet == chained:
        src_step = next((d["step"] for d in derived_sheets if d["sheet"] == chained), None)
        step_advisories.append(
            f"（{src_step}段目の出力『{chained}』を対象にしました）")
    elif derived_sheets:
        derived_names = [d["sheet"] for d in derived_sheets]
        if step_target_sheet not in derived_names:
            made = "・".join(f"{d['step']}段目の『{d['sheet']}』" for d in derived_sheets)
            step_advisories.append(
                f"★ 前段が作った{made}ではなく『{step_target_sheet}』"
                f"（元の表）を見ています ── 絞り込んだ結果に対する操作を意図していたなら、"
                f"シート名を依頼文に書いて実行し直してください"
            )
    step_header_row = current_meta.get("header_rows", {}).get(step_target_sheet, 1) if step_target_sheet else 1
    code = codegen_dsl(op, resolved, current_meta, use_formula=use_formula)
    (workdir / f"plan_step{i}.bas").write_text(code, encoding="utf-8")

    # ★ W9: INSERT_ROWS/AUTOFIT の事後条件が段ごとの before/after を突き合わせられるようコピー。
    stepsource = workdir / f"plan_step{i}_source{out_book.suffix}"
    shutil.copy2(out_book, stepsource)
    step_before = snapshot(stepsource)   # ★ W10d: 助言計算用（この段の適用直前）

    apply_result = apply_dsl_step(
        op, resolved, code, apply_target=out_book, before=step_before, before_charts=before_charts,
        before_chart_paths=before_chart_paths,
        workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout,
        header_row=step_header_row, use_formula=use_formula, source_book=stepsource, deps=deps,
        apply_progress_label=f"⏳ {i}段目 LibreOffice で適用中…", print_changes=False,
        step_prefix=step_prefix)   # ★ 摩擦⑥: 再試行の開示行も段番号つきで揃える

    if apply_result.runtime_error is not None:
        detail = f"実行時エラー: {short_error_summary(apply_result.runtime_error)}"
        return (None, (i, confirm.label, "fail", detail),
                {"op": op, "command": confirm.line, "status": "fail", "postcondition": None,
                 "interpretation": interpretation},
                step_advisories, provenance_entry, None, current_meta)

    # ★ W10d【本命】: mode="structural"（依頼文言との重なり④は呼び出し側が全体で1回評価・dsl_step.py 参照）。
    step_after = apply_result.after
    # ★ C9: 複合計画の DSL 段だけ _truncation_notice が一度も呼ばれていなかった既知の穴
    #   （stage_organs の dsl_plan_step × truncation_notice が None と宣言していた場所）。
    #   ✓ の意味を「読み戻して確かめた」にする以上、「先頭 MAX_ROWS 行しか見ていない」は
    #   ✓ の主張範囲に直接効くので同じ回で塞ぐ。
    step_notice = _truncation_notice(step_before, step_after, exhaustive_postcondition=True)
    if step_notice:
        print(f"{step_prefix}{step_notice}")
    # ★ 単位C(D8): ここも `if op == "LOOKUP_FILL"` のハードコードを宣言読み取りへ置き換えた。
    #   複合計画は④を計画全体で1回だけ評価するので、段ごとの「読むだけのシート」を返して
    #   呼び出し側が全段分を合算する（返り値は集合・空集合なら足すものが無いだけ）。
    mention_exclude_sheets = _declared_reads_only_sheets(op, resolved)
    # ★★ 単位F: 単発(cmd_run_dsl)と同じ位置・同じ関所。段ごとの before/after で見る
    #   （計画全体の before/after で見ると、前の段が既存行を書いたことを次の段の前提破れと
    #   読み違える ―― 宣言は段ごとに違うので、突き合わせも段ごとでなければならない）。
    # ★★ 単位G(複合計画版): 中立化（「（既存シート『集計』の更新は意図どおりです）」）は
    #   助言の発生源で決まるので、単発(5517-5522)と同じく advisories を組む「前」に検査する
    #   ―― ここが単発と非対称のまま(precondition_broken 未配線)だと、破れた宣言でも
    #   肯定文が出てしまう（単位G の保証が複合計画に配線されていなかった不具合）。
    #   ★ 検査は段あたり1度だけ（単位C の教訓）。結果は advisories と下の関所の両方で使い回す。
    _precondition = _maybe_warn_write_precondition(op, step_before, step_after, resolved)
    precondition_broken = _precondition[0] if _precondition else None
    # ★★ 単位H 開示(複合計画版): 単発と同じく関所判定の直後（置き場所は単発と対称）。
    for own_notice in _maybe_own_prior_output_notice(op, step_before, step_after, resolved):
        print(f"{step_prefix}{own_notice}")
    step_advisories.extend(compose_dsl_step_advisories(
        "structural", op, resolved, current_meta, task, step_before, step_after, deps=deps,
        precondition_broken=precondition_broken, after_path=out_book) + formula_error_advisory(stepsource, out_book, cell_ref=_cell_ref) + broken_identity_advisory(stepsource, out_book, resolved if isinstance(resolved, dict) else {}))   # ★ 挙動変更#1(a)

    status, reason = apply_result.postcondition_status, apply_result.postcondition_reason
    # ★ 止血1/2: "error"→fail 扱い。"warn"(検証対象不足)は成功は名乗るが機械検証済みとは言わない。
    if status in ("fail", "error"):
        return (None, (i, confirm.label, "fail", reason),
                {"op": op, "command": confirm.line, "status": "fail", "postcondition": "fail",
                 "interpretation": interpretation},
                step_advisories, provenance_entry, mention_exclude_sheets, current_meta)
    item_status = "warn" if status == "warn" else "ok"
    # ★ 連鎖の番人: この段が派生シートを作ったことを、後段のために記録する
    #   （失敗した段は記録しない ── 作られていないシートを後段に突き付けない）。
    if derived_sheets is not None and op in PLAN_CHAIN_WARNING_OPS and resolved.get("_new_sheet"):
        derived_sheets.append({"step": i, "op": op, "sheet": resolved["_new_sheet"]})
    warn_precondition = _precondition[1] if _precondition else None   # ★ 単位G: 上で 1 度だけ検査済み
    if warn_precondition:
        print(f"{step_prefix}{warn_precondition}")
        if suspicion_sink is not None:   # ★ 決裁③: ✓→△ 降格の判定材料へ運ぶ
            suspicion_sink.append(warn_precondition)
        gate_exit = _confirm_overwrite_or_gate(a, warn_precondition, step_prefix=step_prefix)
        if gate_exit is not None:
            return (gate_exit, None, None, [], None, None, current_meta)
    return (None, (i, confirm.label, item_status, reason),
            {"op": op, "command": confirm.line, "status": item_status, "postcondition": status,
             "interpretation": interpretation},
            step_advisories, provenance_entry, mention_exclude_sheets,
            build_book_meta(out_book, header_rows=header_rows))


def _preview_dsl_plan(a: argparse.Namespace, plan: list, book_meta: dict, vocab: dict, book_name: str) -> list:
    """★ C7: cmd_run_plan の --dry プレビュー（未実行・印字のみ）。--json 用 plan_json を返す。
       分離により cmd_run_plan 自身は verify_dsl_args を直接呼ばなくなる（DoD7 の材料）。"""
    preview_items, plan_json = [], []
    for i, step in enumerate(plan, 1):
        op = step.get("op")
        if op == "CLARIFY":
            q = step.get("question") or "確認が必要です"
            preview_items.append((i, q, "fail", "計画の途中で確認が必要なため対応できません"))
            plan_json.append({"op": "CLARIFY", "command": None, "status": "fail", "postcondition": None})
        elif op not in OP_SCHEMA:
            about = step.get("about") or "内容不明の依頼"
            preview_items.append((i, about, "warn", None))
            plan_json.append({"op": op, "command": about, "status": "warn", "postcondition": None})
        else:
            # ★ 挙動変更#2: --dry プレビューも a._target_sheet を読む（実行と同じ解決）。
            ok_v, resolved, inferred, err = verify_dsl_args(
                op, step.get("args", {}), book_meta, task=a.task, vocab=vocab,
                target_sheet=getattr(a, "_target_sheet", None))
            if ok_v:
                # ★ 挙動変更#3: プレビューの宣言テキストにも対象シートを載せる（実行時の
                #   「解釈:」行と同じ形にする・1枚のブックでは従来どおり付かない）。
                label = format_confirmation_line(op, resolved, inferred, sheets=book_meta.get("sheets"),
                                                 target_sheet=resolved.get("_target_sheet"))[len("解釈: "):]
                preview_items.append((i, label, "ok", None))
                # ★ 段1: プレビューにも interpretation を足す（--dry でも解釈は決まっている）。
                #   単発/実行と同じ組み立て（build_interpretation）を使う ―― 判定は
                #   classify_subject_provenance を1回だけ呼ぶ（単発は print_dsl_confirmation
                #   の内側で済ませているが、プレビューはそれを呼ばないのでここで直接呼ぶ）。
                verdicts = classify_subject_provenance(op, resolved, book_meta, a.task, a)
                interpretation, _provenance = build_interpretation(op, resolved, inferred, verdicts, [book_name])
                plan_json.append({"op": op, "command": label, "status": "ok", "postcondition": None,
                                   "interpretation": interpretation})
            else:
                preview_items.append((i, f"操作:{OP_LABELS.get(op, op)}", "fail", err))
                plan_json.append({"op": op, "command": None, "status": "fail", "postcondition": None})
    # ★ C9: --dry はプレビュー専用レンダラを使う（実行経路と同じ format_plan_report に
    #   status="ok" を流し込んでいたことが「未実行なのに ✓」の直接の原因だった）。
    for ln in format_plan_preview(preview_items):
        print(ln)
    return plan_json


def cmd_run_plan(a: argparse.Namespace, book: Path, source_book: Path, book_meta: dict, plan: list) -> int:
    """M2c: 複合依頼の計画実行本体。段ごとに②検証→③確認→④codegen→⑤適用→⑥事後条件
       （DSL 語彙の段。_run_dsl_plan_step に委譲）または FREEFORM（語彙外の段・依頼文だけを
       渡す）を順に実行し、★ 項目別の honest な報告を出す。総合判定は最弱の段に従う。
       ★ 依存つき連鎖: 各段の接地は直前段を適用した後の列構成(current_meta)で行い、列名
       不一致時は _apply_new_column_fallback が新規列への参照とみなし1回だけ書き換えを試みる。
       ★ W3: header_rows は計画全体を通して不変なので current_meta 再読み込みでも同じ値を渡す。"""
    print(render_run_header(f"複合計画・{len(plan)} 段", a.model, book.name))

    workdir = book.parent / f".ailine_{book.stem}"
    workdir.mkdir(exist_ok=True)
    out_book = out_book_path(book)
    apply_timeout = a.timeout if a.timeout else None
    helpers_dir = Path(a.helpers).resolve() if a.helpers else DEFAULT_HELPERS
    refs_dir = Path(a.refs).resolve() if a.refs else DEFAULT_REFS
    _helper_catalog, helper_files = load_helpers(helpers_dir)
    header_rows = book_meta.get("header_rows", {})
    use_formula, vocab = not getattr(a, "values", False), load_vocab()

    result = {"ok": False, "attempts": 1, "task": a.task, "model": a.model,
              # ★ #17: 段ごとの op を全部残す（PLAN は 1 依頼に複数 op）。
              "op": "PLAN",
              "ops": [st.get("op") for st in plan if isinstance(st, dict) and st.get("op")],
              "path": "plan", "command": None, "postcondition": None}
    if a.dry:
        print("\n（--dry プレビュー・語彙外の段は実行時に AI が直接作成（機械保証なし）で対応します。未実行）")
        plan_json = _preview_dsl_plan(a, plan, book_meta, vocab, book.name)
        print("\n（--dry: 適用しない。レビュー後に --dry を外して実行）")
        result["ok"] = True
        result["dry"] = True
        result["plan"] = plan_json
        _finish_run(a, book, result, "none")
        return 0

    shutil.copy2(source_book, out_book)

    original_headers = {k: list(v) for k, v in book_meta["headers"].items()}
    # ★ 挙動変更#2: 対象シートは _cmd_run_dispatch が resolve_target_sheet で一箇所だけ
    #   決めた a._target_sheet を読む（後方互換フォールバックは1枚目・旧挙動と同一）。
    first_sheet = getattr(a, "_target_sheet", None) or (book_meta["sheets"][0] if book_meta.get("sheets") else None)
    before_all = snapshot(out_book)
    before_charts = before_all["charts"]
    # ★ 致命④: 「今回増えた1個」の同定に使う（snapshot() の辞書には入れない・上のコメント参照）。
    before_chart_paths = _chart_paths(out_book)

    current_meta = book_meta
    items: list = []         # (idx, label, status, detail)
    plan_json: list = []     # --json 用（既存キー不変・新規追加）
    plan_provenance: list = []   # ★ A': 段ごとの倍率等の出典（history.jsonl 用）
    step_advisory_entries: list = []   # ★ W10d: [(段番号 or None, 助言文言), ...]
    mention_exclude_sheets: set = set()   # ★ W10d/単位C: 参照専用シート（reads_only 宣言・全段分の合算）
    subject_sink: dict = {"warnings": [], "unspoken": []}   # ★ 単位E: 対象スロットの出所（全段分）
    suspicion_sink: list = []   # ★ 決裁③: 印字済みだが advisories に入らない疑わしい ⚠
    derived_sheets: list = []   # ★ 連鎖の番人: 前段が作った派生シート（EXTRACT/DEDUP の出力）
    #   （単位F/G の前提破れ + LLM/機械抽出の食い違い・全段分・✓→△ 降格の材料）

    for i, step in enumerate(plan, 1):
        op = step.get("op")

        if op == "CLARIFY":
            question = step.get("question") or "確認が必要です"
            items.append((i, question, "fail", "計画の途中で確認が必要なため対応できません"))
            plan_json.append({"op": "CLARIFY", "command": None, "status": "fail", "postcondition": None})
            continue

        if op not in OP_SCHEMA:
            about = step.get("about") or "内容不明の依頼"
            # ★ W10b 項目1: 関所で拒否/非対話なら計画全体をここで止める（out_book はコピーなので安全）。
            try:
                okf, changes, advisories, _fkind, detail = run_freeform_plan_step(
                    a, about, out_book, workdir, refs_dir, helpers_dir, f"plan{i}", apply_timeout,
                    step_prefix=f"  {i}段目: ", vocab=vocab, op=op, about=about)
            except _FreeformGateAbort as e:
                return e.exit_code
            if okf:
                items.append((i, about, "warn", None))
                for ln in changes:
                    print(f"  {ln}")
                step_advisory_entries.extend((i, adv) for adv in advisories)   # ★ W10d: ループ後に重複を畳む
            else:
                items.append((i, about, "fail", detail))
            plan_json.append({"op": op, "command": about,
                               "status": "ok" if okf else "fail", "postcondition": None})
            current_meta = build_book_meta(out_book, header_rows=header_rows)
            continue

        # ★ C7: DSL 語彙段。cmd_run_dsl と同じ ailine_core.dsl_step を通る _run_dsl_plan_step に委譲。
        gate_exit, item, plan_json_entry, step_adv, prov_entry, step_reads_only, current_meta = \
            _run_dsl_plan_step(
                i, op, step.get("args", {}), task=a.task, current_meta=current_meta,
                original_headers=original_headers, first_sheet=first_sheet, out_book=out_book,
                workdir=workdir, helper_files=helper_files, apply_timeout=apply_timeout,
                use_formula=use_formula, header_rows=header_rows, before_charts=before_charts, a=a, vocab=vocab,
                book_name=book.name, subject_sink=subject_sink, suspicion_sink=suspicion_sink,
                derived_sheets=derived_sheets, before_chart_paths=before_chart_paths)
        if gate_exit is not None:
            return gate_exit
        if item is not None:
            items.append(item)
        if plan_json_entry is not None:
            plan_json.append(plan_json_entry)
        if step_adv:
            step_advisory_entries.extend((i, adv) for adv in step_adv)
        if prov_entry is not None:
            plan_provenance.append(prov_entry)
        if step_reads_only:
            mention_exclude_sheets.update(step_reads_only)

    # ★ W10d【本命】: 依頼文言との重なり④は計画全体に対して1回だけ評価（他段が担当する言及の
    #   誤検知を避ける）。exclude_sheets は全段の「読むだけのシート」(reads_only 宣言)を合算する。
    # ★ 誤爆#3: 対象シートを決めた側が「曖昧なので既定へ後退した」と記録した語も同じ和に足す
    #   （単発の build_advisories が中でやっているのと同じ和・ここは④を直接呼ぶ経路）。
    mention_exclude_sheets |= conflict_excluded_sheets(getattr(a, "_sheet_conflict", None))
    after_all = snapshot(out_book)
    mentions = extract_task_mentions(a.task, before_all["sheets"])
    final_mention_lines = mention_overlap_advisory(
        mentions, before_all, after_all, mention_exclude_sheets or None)
    step_advisory_entries.extend((None, ln) for ln in final_mention_lines)   # None=計画全体

    print()
    for ln in format_plan_report(items):
        print(ln)
    dedup_advisories = _dedup_step_advisories(step_advisory_entries)   # ★ W10d: 重複を畳む
    if dedup_advisories:
        print("\n助言:")
        for ln in dedup_advisories:
            print(ln)
    verdict_line, verdict = overall_verdict(items)
    # ★ C9: 全段 ok のときは verdict_line が None（『✓ すべて機械検証済み』は廃止し、原本が
    #   確定した後の1行＝_finish_apply に移した）。
    # ★ 単位E: 常時の範囲注記はここからも消えた（発火率100%＝情報量ゼロ）。範囲を明示する
    #   役割は、②の run 固有の1文（✓ の直後・render_scope_notes）が引き継ぐ。
    if verdict_line:
        print(f"\n{verdict_line}")

    _changed, difflines = diff_snapshots(before_all, after_all)
    result["plan"] = plan_json
    result["provenance"] = plan_provenance or None
    result["items"] = [{"idx": idx, "label": label, "status": st, "detail": det}
                        for idx, label, st, det in items]
    result["changes"] = difflines
    result["advisories"] = [{"steps": idxs, "text": text}
                             for idxs, text in _group_step_advisories(step_advisory_entries)]

    if verdict == "fail":
        # ★ C9: 査定が名指しした沈黙 ―― 全段破棄で終わるのに原本がどうなったかを言って
        #   いなかった（.out が黙って隣に残ることも無言だった）。
        print(_untouched_original_line(book, out_book))
        result["out"] = str(out_book)
        _finish_run(a, book, result, "plan_step_failed")
        return 1

    result["ok"] = True
    # ★ W8b-2 項目1: 複合計画は総合判定(overall_verdict)に従う。全段機械検証済み(ok)
    #   の時だけ ✓、語彙外/検証不足の段が混じる(warn)なら ⚠「機械保証はありません」側
    #   （自由生成の段が混じっている以上、全体としても機械保証済みとは名乗れない）。
    # ★★ 単位E: ③（依頼文の語と矛盾する対象）を含む段が1つでもあれば ✓ は出さない
    #   ―― 各段の事後条件が「計画どおり」に通っていても、その計画は依頼文と食い違っている。
    # ★ 決裁③(2026-08-22): 疑わしい ⚠ の総数 = 全段の advisories 中の ★ 付き件数
    #   （step_advisory_entries・重複畳み前の生数を数える ── dedup_advisories は表示用に
    #   畳んだ後の行数で、同じ ⚠ が複数段に跨って1行に畳まれることがあるため件数がずれる）
    #   + suspicion_sink（単位F/G 前提破れ + LLM/機械抽出の食い違い ── 印字済みだが
    #   advisories に含めていない別集計・_run_dsl_plan_step の docstring 参照）。
    warning_count = (count_suspicious_advisories(text for _idx, text in step_advisory_entries)
                      + len(suspicion_sink))
    # ★ 2026-08-25（復元の重大6・盲検）: _finish_apply は置換の成否を返すのに、
    #   呼び出し元 4 箇所が**全部戻り値を捨てていた**。バックアップに失敗して原本反映を
    #   中止しても exit 0 ── スクリプトから回す利用者は「反映されなかった」を検出できない。
    #   ★ 表示は正直だった（「原本は無変更」と言っていた）。嘘だったのは**終了コード**。
    _applied = _finish_apply(a, book, out_book, workdir, result,
                   machine_verified=(verdict == "ok" and not subject_sink["warnings"]),
                   scope="; ".join(label for _idx, label, _st, _det in items),
                   scope_note="\n".join(render_scope_notes(subject_sink["unspoken"])),
                   warning_count=warning_count)
    if not _applied:
        return _finish_failed_apply(a, book, result)

    _finish_run(a, book, result, "none")
    return 0


def cmd_stop(a: argparse.Namespace) -> int:
    subprocess.run([sys.executable, str(basrun_path()), "stop"], encoding="utf-8", errors="replace")
    return 0


def cmd_scan(a: argparse.Namespace) -> int:
    """`ailine scan <folder>`: M1読み ── 複数ブックの棚卸し（書き込みゼロ・LO は起動しない）。
       DESIGN-20260821-multifile.md v2 §1(M1読み)・§2(骨)。分類・照合は ailine_core/multifile.py
       に置く。見出し行の推定だけ既存の detect_header_row/_row_char_stats を基準ファイル1冊分
       だけ呼ぶ（このモジュールが持つ元々の機能をそのまま流用・LO 往復は無い）。"""
    folder = Path(a.folder).resolve()
    candidates, excluded = multifile.classify_folder_contents(folder)
    # ★ 2026-08-24 第三波 S1: scan にだけこの配線が無く、自分の出力（2冊照合の結果等）を
    #   棚卸しの分母に数えて「取れなかった」と ⚠ で名指ししていた（stack/run には在った）。
    candidates, self_excluded = multifile_stack.split_own_outputs(candidates)
    base_path, base_wb = multifile.open_base_workbook(candidates)
    base_headers, base_sheet, header_row, value_col_name = [], None, 1, None
    if base_wb is not None:
        base_sheet = base_wb.sheetnames[0]
        ws = base_wb[base_sheet]
        scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
        rows = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
        row, confident = detect_header_row({"rows": rows})
        header_row = row if confident else 1
        base_headers = multifile.read_row_headers(ws, header_row)
        value_col = multifile.numeric_value_column(ws, header_row, len(base_headers) or MAX_COLS)
        value_col_name = base_headers[value_col - 1] if value_col else None
        base_wb.close()
    files = [multifile.evaluate_file(p, base_headers, base_sheet, header_row, value_col_name)
             for p in candidates]
    result = {"denominator": len(candidates), "base": base_path.name if base_path else None,
              "files": files, "excluded": excluded, "self_excluded": self_excluded}
    if a.json:
        print(json.dumps(result, ensure_ascii=False))
        return 0
    for ln in render_scan_report(str(folder), result):
        print(ln)
    return 0


def _stack_json(result: dict) -> dict:
    """--json 契約（検体で凍結済み）: denominator/stacked_files/rows_written/files/sums/
       mismatches。★ jisaku-review#4: mismatches はテキストの ⚠ と同じ情報を機械可読で
       （ファイルごとの入れ子ではなく、[{file, row, excluded_value, adopted_sum}, ...] の
       平らな形 ── 自動化経路が総なめできるように）。
       ★ P2: sheet_fallbacks（[{name, wanted, used}, ...]）も追加 ── 基準名のシートが無く
       1枚目へ落ちたファイルの機械可読の開示（人間向けは render_stack_report）。"""
    mismatches = [{"file": entry["name"], "row": m["row"],
                   "excluded_value": m["excluded_value"], "adopted_sum": m["adopted_sum"]}
                  for entry in result.get("mismatches", ()) for m in entry["rows"]]
    return {"denominator": result["denominator"], "stacked_files": result["stacked_files"],
            "rows_written": result["rows_written"], "files": result["files"],
            "sums": result["sums"], "mismatches": mismatches,
            "sheet_fallbacks": result.get("sheet_fallbacks", []),
            "total_word_warnings": result.get("total_word_warnings", [])}


def _stack_postcondition_fail(label: str, expected, actual) -> int:
    """事後条件①②が破れた時の唯一の出口。★ tmp_out は移さない（out は無傷のまま）。"""
    print(f"⚠ 事後条件が破れた: {label}  元(採用時) {primitives.fmt_num(expected)} / "
          f"出力(書いた直後) {primitives.fmt_num(actual)}")
    return 5


def _stack_attribution_fail(mismatch: dict) -> int:
    """★ jisaku-review4戦目 F1 major の直し: cmd_run_folder は verify_extract を再利用して
       帰属検算（review3#3）を無償継承していたが、cmd_stack の書き込み時経路は自前の
       行数+Σ だけで帰属を見ていなかった ── Σ 保存のまま値だけ入れ替える変異が exit 0 で
       素通りする実機再現が根拠。行数/Σ と同じ『移す前の tmp_out』段で
       verify.verify_output（帰属検算まで含む独立読み）を呼び、ここで拾う。"""
    print(f"⚠ 事後条件が破れた: 帰属  {mismatch['file']} の {mismatch['src_row']}行目 "
          f"列『{mismatch['column']}』 元(採用時) {primitives.fmt_num(mismatch['source'])} / "
          f"出力(書いた直後) {primitives.fmt_num(mismatch['output'])}")
    return 5


# ★ review5#2 の直し（実機再現: .xlsm の打ち間違いが黙って依頼文に落ちた）: 個別拡張子の
#   列挙だけでは次の未知の拡張子（.numbers 等）でまた同じ穴になる。一般則を併用する ──
#   末尾が「. + 英数字2〜5文字」ならファイル名の拡張子らしいと見る（依頼文の1語目が
#   たまたまこの形になることはまず無い ── 自然文の単語はピリオドで終わらない）。
_GENERIC_EXT_RE = re.compile(r"\.[A-Za-z0-9]{2,5}$")


def _looks_like_second_book_path(token: str) -> bool:
    """依頼文の一部でなく『2冊目のつもりのパス』らしいか（M3 arity 判定の後半）。
       ★ 実在確認は呼び出し側が既にやっている（この関数は『実在しないが2冊目らしい』を
       見分けるためだけ）── 表計算らしい拡張子・拡張子らしい語尾・パス区切りのいずれかで真。
       依頼文の1トークン目がたまたまこれに当てはまることはまず無い（自然文はこの形にならない）。"""
    p = Path(token)
    if p.suffix.lower() in BOOKLIKE_SUFFIXES:
        return True
    if "/" in token or "\\" in token:
        return True
    return bool(_GENERIC_EXT_RE.search(token))


def _run_folder_refuse(op: str, plan_len: int) -> int:
    """M2 の断り（E11）: フォルダに未対応の依頼を**名指し**で断り、次の手を添えて exit 3。
       ★ 黙って1冊目に適用が最悪の形 ── 断る時は原本に一切触れない（この時点で
       書き込みは一度も起きていない）。対応の可否は OP_META の folder 宣言が唯一の出どころ
       （手書きの対応表を別に持たない ── ずれる表を作らない）。"""
    label = OP_META.get(op, {}).get("label") or (op if op else "この依頼")
    if plan_len > 1:
        print(f"？ 複数の操作をまとめた依頼（{plan_len} 段）はフォルダに対応していません。"
              "フォルダに頼めるのは抽出だけです（1 冊を指定すればそのまま頼めます）。")
    else:
        print(f"？ {label}は複数ファイル（フォルダ）に対応していません。"
              f"フォルダに頼めるのは抽出だけです（1 冊を指定すれば『{label}』も頼めます）。")
    print("  （頼める操作の一覧: ailine ops）")
    return 3


def _own_extract_output_status(path: Path, col: str, cmp: str, value) -> tuple:
    """path が①ailine 産か（mark）②M2 抽出の自分の前回出力で、かつ焼いた条件が
       今回と完全一致するか（same_condition）を返す。
       ★ review3#1 critical の直し: 『黙って作り直してよい』の根拠を**印だけ**にしない ──
       印は同じでも条件（列/比較/値）が違う前回出力を、条件を見ずに上書きして消していた
       （実機再現: 長いフォルダ名で切り詰めが起きると別条件が同名に潰れる）。
       同じ判定をここ1箇所に集約し、preflight（40冊読む前）と移す直前の再判定の両方が使う。"""
    mark = multifile_stack.own_output_mark(path)
    if mark != extract_multi.CREATOR_MARK:
        return mark, False
    _creator, description = xml_readback.read_core_properties(path)
    try:
        cond = json.loads(description) if description else None
    except (TypeError, ValueError):
        cond = None
    same_condition = (isinstance(cond, dict) and cond.get("tool") == "ailine"
                       and cond.get("kind") == "extract" and cond.get("column") == col
                       and cond.get("cmp") == cmp and cond.get("value") == value)
    return mark, same_condition


def out_book_path(book: Path) -> Path:
    """`--copy` と作業用に使う `<名前>.out.<拡張子>` の場所。★ 実装は 1 つ。

    ★ 2026-08-25（復元の致命2・盲検）: この式が 4 箇所に書き写されていて、
      どこにも**存在確認が無かった**。利用者が自分で作った `売上.out.xlsx` は
      一言も無く上書きされ、原本反映が成功すると unlink で消えていた。
      フォルダ経路には同じ危険への関所が既に在るのに、単一ブック経路に無かった（片配線）。
    """
    return book.with_name(book.stem + ".out" + book.suffix)


def refuse_if_output_is_someone_elses(book: Path) -> int | None:
    """出力先に**人のファイル**が在れば exit 7 で止める。無ければ None。

    ★ ailine 産（前回の .out）は従来どおり黙って作り直す ── 印で見分ける
      （フォルダ経路の関所と同じ規則・同じ出口）。
    """
    out = out_book_path(book)
    # ★ 2026-08-26（復元の致命1）: 入口と出口で同じことを 2 度判断させない。
    #   ここは run の一番最初に必ず通る ── 「始まる前から在ったか」をここで控える。
    #   出口（atomic_replace_inplace の後始末）はこの控えだけを見て消すかを決める。
    note_pre_existing_output(out)
    if not out.exists():
        return None
    # ★ 根拠は「この道具が過去にそこへ書いたか」── 履歴が out を記録している。
    #   ★ 印（stack/extract の CREATOR_MARK）は使えない: 単一ブック経路の .out には
    #     印が付かない（実測で確認 ── 印は複数ファイル経路だけの仕組みだった）。
    #   ★ 履歴が無い/読めない時は**止める側**に倒す（人のファイルを壊すより、
    #     一度断って人に確認してもらう方が安い）。
    try:
        target = str(out.resolve())
        for entry in read_history(max_n=HISTORY_RECALL_MAX):
            if not isinstance(entry, dict):
                continue
            recorded = entry.get("out")
            if not (recorded and str(Path(recorded).resolve()) == target):
                continue
            # ★ 2026-08-25（復元の致命③・**今朝入れたこの関所そのものの穴**）:
            #   「自分が書いた場所」だけでは足りない。README が慎重な人に勧める
            #   `--copy` の成果物は、書いた**後**に人が手を入れて育てる物で、
            #   それが原本反映 1 回で警告なしに消えていた。
            #   ★ 三項目（実体）で見る: いま在る物が、俺が置いた物のままか。
            stamped = entry.get("out_sha")
            if stamped is None:
                return None      # 指紋を残す前の古い記録 ── 従来どおり（判定材料が無い）
            if stamped == _file_digest(out):
                return None      # 俺が置いたまま ── 作り直してよい
            return _refuse_edited_output(out)
    except Exception:
        pass
    return _refuse_output_conflict(out, None)


def _refuse_edited_output(out: Path) -> int:
    """自分が書いた物だが、その後**人が手を入れている** ── 消さずに断る（exit 7）。"""
    print(f"⚠ 出力先に書けません: {out}")
    print(f"（{out.name} は ailine が作った物ですが、そのあと変更されています。"
          "作業内容が消えるので上書きしません ── 別の場所へ移すか削除してから、"
          "もう一度実行してください）")
    return 7


def _refuse_output_conflict(out: Path, mark: str | None) -> int:
    """M2 出力先の関所（exit 7）: 人のファイル、または ailine の別コマンド/別条件の
       出力があって、黙って上書きしてよい根拠が無い時の唯一の出口。"""
    # ★★ 2026-08-26（複数ファイルの盲検 中9 + Namakoo が実際に踏んだ）:
    #   「ailine の印が無い**人のファイル**です」と**断定**していた。しかし実測では、
    #   それは ailine が数分前に作り、「作業結果はここに残っています」と自分で案内した
    #   ファイルだった（失敗した run は history に out を残さないので、印が見つからない）。
    #   ★ 見たものと、その解釈を分ける ── 分かるのは「この道具が書いた記録が無い」ことだけ。
    #     誰の物かは断定しない（安全側に倒すのは正しいが、説明が嘘なのは別の問題）。
    if mark:
        whose = f"ailine の別のコマンドの出力です（作成: {mark}）"
    else:
        whose = ("この道具が書いた記録がありません（人が置いたファイルか、"
                 "途中で失敗した run が残した作業結果のどちらかです）")
    print(f"⚠ 出力先に書けません: {out}")
    print(f"（{out.name} は{whose}。run にはフラグでの上書き許可がありません ── "
          "そのファイルを別の場所へ移すか削除してから、もう一度実行してください）")
    return 7


def cmd_run_folder(a: argparse.Namespace) -> int:
    """`ailine run <フォルダ> "<依頼>"`: M2 ── フォルダの各ブックから条件に一致する行を
       抜き出して1冊に集約する（DESIGN-20260821-multifile.md M2 節・Namakoo 決裁 A 案）。
       配管は M1書き(stack) の再利用（出所列・書き手の印・関所 exit7・自己参照除外・
       workdir→移動）で、この経路固有なのは「翻訳された条件で選ぶ」ことだけ。
       ★ architect 致命4: この経路は check_excel_lock / normalize_book / バックアップ /
       undo 履歴のどれにも触れない（原本は読むだけ・消せるのは出力ブックだけ）。
       フォルダに `open(r+b)` を試みて『Excel で開かれています』と嘘をつく穴は、
       分岐がこの関数へ来ることで構造的に消える。
       ★ 一括検出（Namakoo 決裁 09:22）: ファイル単位の欠陥は最初の1件で止めず全部集めて
       名指しする。★ 憲法2: この関数は ✓ を一切名乗らない（✓ の裏づけを機械の結果
       オブジェクトから出す claim render guard がこの経路にまだ無い ── 迷ったら落とす）。"""
    folder = Path(a.book).resolve()
    as_json = bool(getattr(a, "json", False))
    say = (lambda *args, **kw: None) if as_json else print

    # ① 分母と自己参照除外（V6）── cmd_stack と同じ判定（ailine 産は種類を問わず入力から外す）。
    candidates, folder_excluded = multifile.classify_folder_contents(folder)
    candidates, self_excluded = multifile_stack.split_own_outputs(candidates)
    denominator = len(candidates)

    base_path, base_wb = multifile.open_base_workbook(candidates)
    if base_path is None:
        say(f"■ ailine run（フォルダ抽出）  folder={folder}")
        if self_excluded:
            names = "、".join(f"『{n}』" for n in self_excluded)
            say(f"（自分の出力 {names} を入力から除外しました）")
        for line in render_excluded_lines(folder_excluded):
            say(line)
        say(f"{denominator} ファイル中 0 照合 → 読める .xlsx が無いので抽出できません")
        if as_json:
            print(json.dumps({"out": None, "condition": None, "multifile": {
                "denominator": denominator, "matched_files": 0, "contributing_files": 0,
                "rows_written": 0, "files": [], "skipped": [
                    {"name": p.name, "reason": "旧形式(.xls)または読み込み失敗"} for p in candidates],
                "self_excluded": self_excluded, "sheet_fallbacks": [], "excluded_detail": [],
                "mismatches": [], "rebuilt_own_output": False}}, ensure_ascii=False))
        return 0

    # ② 基準ファイル方式（A1）: 見出し行の推定は既存の器官を1回だけ使い、値として渡す。
    base_sheet = base_wb.sheetnames[0]
    ws = base_wb[base_sheet]
    scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
    rows_stats = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
    row, confident = detect_header_row({"rows": rows_stats})
    header_row = row if confident else 1
    base_headers = multifile.read_row_headers(ws, header_row)
    # ★ operator 盲検7度目の直し（2026-08-21）: extract 経路も合計行検出を基準の
    #   数値列すべてに揃える（旧仕様は条件列だけを見ていた ── 偶然 条件列=金額 で
    #   助かっていただけ）。base_wb を閉じる前に ws から読む（stack と同じ配線）。
    numeric_cols = multifile.numeric_column_names(ws, header_row, base_headers)
    base_wb.close()

    # ③ 翻訳（7B に渡すのは基準ファイルの見出しだけ ── 40 冊分で prompt を壊さない）。
    book_meta = {"sheets": [base_sheet], "headers": {base_sheet: base_headers},
                 "header_rows": {base_sheet: header_row}}
    translation = translate_task(a.model, a.task, book_meta, temperature=0.1)
    plan = translation.get("plan") if isinstance(translation, dict) else None
    if plan is None and isinstance(translation, dict) and translation.get("op"):
        plan = [translation]          # ★ 後方互換: "plan" で包まない旧形式
    if not isinstance(plan, list) or not plan:
        plan = [{"op": "FREEFORM", "args": {}}]

    step = plan[0] if isinstance(plan[0], dict) else {}
    op = str(step.get("op") or "")
    if len(plan) == 1 and op == "CLARIFY":
        print(f"？ {step.get('question') or '確認が必要です'}")
        print("  （頼める操作の一覧: ailine ops）")
        return 3
    if len(plan) != 1 or not OP_META.get(op, {}).get("folder"):
        return _run_folder_refuse(op, len(plan))

    # ④ 条件の確定（A': 値の数値化は機械が行う ── verify_dsl_args と同じ線）。
    args = step.get("args") or {}
    col = args.get("col") or args.get("column")
    llm_cmp_raw = str(args.get("cmp") or "").strip().lower()
    # ★ operator9 ①（片配線の追補）: このフォルダ抽出経路は verify_dsl_args を通らず cmp を
    #   独自に確定するため、同じ機械抽出+開示を個別に配線する（verify_dsl_args の EXTRACT 分岐
    #   参照・同じ食い違いが起きうる別経路）。
    mechanical_cmp = extract_cmp_from_task(a.task)
    cmp_mismatch_warning = None
    if mechanical_cmp is not None and mechanical_cmp != llm_cmp_raw:
        cmp = mechanical_cmp
        cmp_mismatch_warning = (
            f"LLM が返した比較({llm_cmp_raw or '(空)'})と依頼文の機械抽出({mechanical_cmp})が"
            f"食い違うため機械抽出({mechanical_cmp})を採用しました")
    else:
        cmp = llm_cmp_raw
    value = args.get("value")
    if cmp not in _EXTRACT_CMPS:
        print(f"？ 比較『{args.get('cmp')}』は {'/'.join(_EXTRACT_CMPS)} のどれでもありません。"
              "言い方を変えて（例:『金額が40000以上の行を抜き出して』）もう一度お願いします。")
        return 3
    if col not in base_headers:
        print(f"？ 列『{col}』が基準ファイル『{base_path.name}』にありません。"
              f"ある列: {', '.join(base_headers)}")
        return 3
    if value in (None, ""):
        print("？ 抽出する値が依頼文から読み取れません（例:『金額が40000以上の行を抜き出して』）。")
        return 3
    if cmp == "contains":
        value = str(value)
    else:
        try:
            value = float(value)
        except (TypeError, ValueError):
            if cmp != "eq":
                print(f"？ 『{value}』は数値として読めないので {_EXTRACT_CMP_LABELS[cmp]} の"
                      "比較ができません。数値で言い直してください。")
                return 3

    # ⑤ 出力先（Q7: フォルダの親・機械命名）と書き込みの関所（40 冊読む前に判定して印字）。
    #    ★ review3#1/#5: 黙って作り直してよいのは「印」だけでなく「条件も一致」する時だけ。
    #    条件が違う自分の出力と名前が衝突したら、条件のハッシュで別名へ決定論で逃がす
    #    （切り詰めが起きていなくても、表示形式が同じ文字列に丸まる衝突はありうるため
    #    sanitize_filename 単体の直し（③）とは別にここでも見る）。
    cmp_label = _EXTRACT_CMP_LABELS.get(cmp, cmp)
    raw_stem = f"{folder.name}_{col}{_format_extract_value(value)}{cmp_label}"
    out = folder.parent / f"{extract_multi.sanitize_filename(raw_stem)}.xlsx"
    rebuilt_own_output = False
    if out.exists():
        mark, same_condition = _own_extract_output_status(out, col, cmp, value)
        if mark == extract_multi.CREATOR_MARK and same_condition:
            rebuilt_own_output = True      # 同じ条件の前回出力＝黙って作り直してよい
        elif mark == extract_multi.CREATOR_MARK:
            # ★ review3#1②: 条件が違う自分の出力と衝突 ── ハッシュで別名に決定論で逃がす。
            digest = hashlib.sha256(json.dumps(
                {"column": col, "cmp": cmp, "value": value}, sort_keys=True,
                ensure_ascii=False).encode("utf-8")).hexdigest()[:6]
            collided_name = out.name
            out = folder.parent / f"{extract_multi.sanitize_filename(f'{raw_stem}_{digest}')}.xlsx"
            say(f"（同名の前回出力『{collided_name}』は別条件のため、この結果は"
                f"『{out.name}』に保存します）")
            if out.exists():
                mark, same_condition = _own_extract_output_status(out, col, cmp, value)
                if mark == extract_multi.CREATOR_MARK and same_condition:
                    rebuilt_own_output = True
                else:
                    return _refuse_output_conflict(out, mark)
        else:
            return _refuse_output_conflict(out, mark)
    say(f"■ ailine run（フォルダ抽出）  folder={folder}")
    say(f"出力先: {out}")
    say(f"条件: {col} {_format_extract_value(value)} {cmp_label}")
    if cmp_mismatch_warning:
        say(f"⚠ {cmp_mismatch_warning}")

    # ⑥ ファイルごとの評価（★ 一括検出: 欠陥が出ても止めず全部集める）。
    skipped, files_json, excluded_detail, mismatches = [], [], [], []
    sheet_fallbacks, matched_rows_all = [], []
    blocked_total, blocked_samples = 0, []   # ★ 第三波 H3: 数字に見える文字列（開示専用）
    uncached_total = 0                        # ★ 致命③: 条件列の「数式だが値が無い」セル
    extract_dropped = {}   # ★ 2026-08-24: 値として運べない中身（コメント/リンク）
    all_findings = []   # ★ M2.5: 検分シートの所見（inspection.Finding）
    file_sheet_map = []   # ★ M2.5: [(ファイル名, 使ったシート, 備考), ...]
    for p in candidates:
        r = extract_multi.evaluate_and_extract(p, base_headers, base_sheet, header_row,
                                                numeric_cols, col, cmp, value)
        all_findings.extend(r.findings)
        sheet_used = r.sheet_fallback[1] if r.sheet_fallback else base_sheet
        if r.sheet_fallback:
            sheet_fallbacks.append({"name": r.name, "wanted": r.sheet_fallback[0],
                                    "used": r.sheet_fallback[1]})
        if r.status == "取れなかった":
            skipped.append({"name": r.name, "reason": r.reason})
            file_sheet_map.append((r.name, sheet_used, f"取れなかった（{r.reason}）"))
            continue
        file_sheet_map.append((r.name, sheet_used, "並べ替えて照合" if r.reordered else ""))
        for _row, _kind in getattr(r, "dropped_notes", ()) or ():
            extract_dropped[_kind] = extract_dropped.get(_kind, 0) + 1
        uncached_total += getattr(r, "uncached_in_column", 0) or 0
        if r.blocked:
            blocked_total += r.blocked["count"]
            for sm in r.blocked["samples"]:
                if sm not in blocked_samples:
                    blocked_samples.append(sm)
        for values, formats, src_row in r.rows:
            matched_rows_all.append((values, formats, r.name, src_row))
        files_json.append({"name": r.name, "rows_matched": r.rows_matched,
                           "rows_unmatched": r.rows_unmatched,
                           "total_rows_excluded": len(r.excluded), "reordered": r.reordered})
        if r.excluded:
            excluded_detail.append({"name": r.name, "rows": [
                {"row": e.row, "value": e.value, "reason": e.reason} for e in r.excluded]})
        for m in r.mismatches:
            mismatches.append({"name": r.name, "row": m.row,
                               "excluded_value": m.excluded_value, "adopted_sum": m.adopted_sum})

    total_matched = len(matched_rows_all)
    matched_files = len(files_json)
    contributing_files = len({name for _v, _fmt, name, _r in matched_rows_all})
    suspicious_files = {f.file for f in all_findings if f.kind in inspection.WARN_KINDS}
    # ★ 第二の独立検出器（operator 盲検7度目 修正2・恒真切り）: stack と同じ線。
    #   列解決に依存せず、抽出結果として積んだ行の全セル値を走査して合計語を名指しする
    #   （除外はしない・書き込みは止めない ── 下の post 由来の abort ガードとは別経路）。
    trip_rows = [(fname, src_row, values) for values, _fmt, fname, src_row in matched_rows_all]
    total_word_warnings = [{"file": fname, "row": row_num, "word": word}
                           for fname, row_num, word in total_row.total_word_trip_findings(trip_rows)]

    # ⑦ 書き出し（workdir→移動）。★ 条件は文書属性に焼く ── verify が出力単体から
    #    条件を復元して同じ検算を再実行できる（信用の条件⑥）。
    out_headers = base_headers + multifile_stack.own_output_headers(base_headers)
    workdir = Path(tempfile.mkdtemp(prefix="ailine_extract_"))
    try:
        tmp_out = workdir / out.name
        wb_out = openpyxl.Workbook()
        wb_out.properties.creator = extract_multi.CREATOR_MARK
        wb_out.properties.description = json.dumps(
            {"tool": "ailine", "kind": "extract", "version": 1,
             "column": col, "cmp": cmp, "value": value, "sheet": base_sheet},
            ensure_ascii=False)
        ws_out = wb_out.active
        ws_out.title = _extract_output_sheet_name(col, cmp, value)
        ws_out.append(out_headers)
        prov_col_idx = len(base_headers) + 1   # ★ M2.5: 「元ファイル」列（出所列の1本目）
        for i, (values, formats, fname, src_row) in enumerate(matched_rows_all, start=2):
            ws_out.append(list(values) + [fname, src_row])
            for c, fmt in enumerate(formats, start=1):   # ★ 実視の磨き: 元の number_format を運ぶ
                ws_out.cell(row=i, column=c).number_format = fmt
            if fname in suspicious_files:
                reason_lines = [inspection.describe(f) for f in all_findings
                                if f.file == fname and f.kind in inspection.WARN_KINDS]
                inspection.tint_row(ws_out, i, len(out_headers), prov_col_idx, reason_lines)
        inspection.bold_row(ws_out, 1, len(out_headers))   # ★ UX 磨き③: 見出し行
        inspection.autosize_columns(ws_out)   # ★ UX 磨き②: 列幅を内容から機械算出
        wb_out.save(tmp_out)

        # ⑧ 事後条件: 書いた直後の中身を**独立読み**（xml_readback）で検算する。
        #    元側も候補ファイル全部を読み直す ── 一致0行のファイルは出所列に現れないため、
        #    出所列だけを頼りにすると「1冊まるごと落ちた」が検算をすり抜ける。
        post = multifile_verify.verify_extract(tmp_out, folder, col, cmp, value,
                                                sheet_name=base_sheet, sources=candidates)
        # ★ operator 盲検7度目 修正2: verify_extract の mismatch（単数）は語のトリップワイヤ
        #   （kind="total_word"）も含みうるが、これは除外しない設計の検出専用 ── 書き込みを
        #   止める理由にはしない（止めるのは row_count/sum/attribution だけ・従来どおり）。
        blocking_mismatch = post.get("mismatch")
        if blocking_mismatch and blocking_mismatch.get("kind") == "total_word":
            blocking_mismatch = None
        if blocking_mismatch:
            m = blocking_mismatch
            where = f"Σ{m['column']}" if m["kind"] == "sum" else "採用行数"
            wb_out.close()
            if as_json:
                print(json.dumps({"out": str(out), "postcondition": post,
                                  "written": False}, ensure_ascii=False))
            else:
                # ★ 2026-08-24（第三波 H1）: 数字の差だけを出して**どのファイルが原因か
                #   一言も言わない**のが盲検の致命だった（scan/stack は名指しするのに）。
                #   名前と理由は skipped に既に在る ── 在るのに使っていなかった。
                #   ★ 理由を**結論より先に**出す（読む順序が原因→結果になるように）。
                for sk in skipped[:5]:
                    print(f"  ⚠ {sk['name']}: 照合できませんでした（{sk['reason']}）")
                if len(skipped) > 5:
                    print(f"  … 他 {len(skipped) - 5} 件")
                if skipped:
                    print("  → この冊の行が『元』に数えられ、出力には現れないため差が出ます")
                    print("     見出しの行や列名を揃えるか、この冊を別フォルダへ移してお試しください")
                print(f"⚠ 事後条件が破れた: {where}  元 {primitives.fmt_num(m['source'])} / "
                      f"出力(書いた直後) {primitives.fmt_num(m['output'])}")
                print(f"（{out.name} は書き込んでいません。元フォルダも変更していません）")
            return 1

        # ★ M2.5①: 検分シート（出力2枚目）── 事後条件(post)が通った直後の数字だけを並べる
        #   （✓ の絶対性の適用拡張・手書きの ✓ を作らない・Σ両側は post['sums'] をそのまま使う）。
        inspection.build_sheet(
            wb_out, findings=all_findings,
            denominator_lines_=inspection.denominator_lines(
                denominator, matched_files, contributing_files, "照合"),
            accounting_lines=inspection.row_accounting_lines(
                adopted=total_matched,
                excluded=sum(len(entry["rows"]) for entry in excluded_detail),
                not_taken_files=len(skipped),
                # ★ 実弾検分の直し: 不一致（条件に合わなかった行）を3勘定目に足す。
                #   files_json の rows_unmatched は --json と同じ結果オブジェクト由来
                #   （手で再計算しない）。
                unmatched=sum(f["rows_unmatched"] for f in files_json)),
            sums=post.get("sums", {}), file_sheet_map=file_sheet_map,
            out_dir=out.parent, source_dir=folder)
        wb_out.save(tmp_out)
        wb_out.close()

        # ⑨ 関所（fail closed）: 移す直前にもう一度見る（前段の判定から時間が経っている）。
        #    ★ review3#1: ここも印だけでなく条件一致まで見る（preflight と同じ判定）。
        if out.exists():
            mark, same_condition = _own_extract_output_status(out, col, cmp, value)
            if not (mark == extract_multi.CREATOR_MARK and same_condition):
                print(f"⚠ 出力先に書けません（実行中に別のファイルが現れました）: {out}")
                return 7
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_out, out)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    # ⑩ 報告（分母つき・行の完全会計・両側の数字）。
    mf = {"denominator": denominator, "matched_files": matched_files,
          "contributing_files": contributing_files, "rows_written": total_matched,
          "files": files_json, "skipped": skipped, "self_excluded": self_excluded,
          "sheet_fallbacks": sheet_fallbacks, "excluded_detail": excluded_detail,
          "mismatches": mismatches, "total_word_warnings": total_word_warnings,
          "rebuilt_own_output": rebuilt_own_output,
          "blocked_stringy": ({"count": blocked_total, "samples": blocked_samples[:3]}
                               if blocked_total else None),
          "uncached_in_column": uncached_total}
    if as_json:
        print(json.dumps({"out": str(out), "written": True,
                          "condition": {"column": col, "cmp": cmp, "value": value},
                          "sums": post.get("sums", {}), "multifile": mf}, ensure_ascii=False))
        return 0
    if self_excluded:
        names = "、".join(f"『{n}』" for n in self_excluded)
        say(f"（自分の出力 {names} を入力から除外しました）")
    for line in render_excluded_lines(folder_excluded):
        say(line)
    say(f"{denominator} ファイル中 {matched_files} 照合 → "
        f"{matched_files} 中 {contributing_files} ファイルで計 {total_matched} 行一致")
    # ★ 第三波 H3（盲検）: 金額が文字列（"1,000" / △1,500 / 全角）の実物で『計 0 行一致』と
    #   だけ言って終わっていた。判定は変えない ── 理由を言う口が無かっただけ。
    # ★ 2026-08-24: コメント/リンクは飾りでなく中身 ── 消えたと言う（stack と同じ線）。
    if extract_dropped:
        _parts = "・".join(f"{k} {n} 件" for k, n in sorted(extract_dropped.items()))
        say(f"  ⚠ 元のセルにあった{_parts}は運んでいません"
            "（値でないため）。元のブックで確認してください")
    for line in compare_blocked.disclosure_lines(
            {"count": blocked_total, "samples": blocked_samples[:3]} if blocked_total else None,
            col, total_matched):
        say(line)
    # ★★ 2026-08-26（複数ファイルの盲検・致命③）: stack は「検算していません」と言うのに、
    #   この経路には警告そのものが無かった（片配線）── 金額が全部数式の請求書に対して
    #   『計 0 行一致』＋『行の完全会計: 成立』＋ exit 0 という**嘘の成功報告**が出ていた。
    if uncached_total:
        say(f"  ⚠ 『{col}』の {uncached_total} 件は数式で、計算結果がファイルに"
            "入っていないため**条件に合うか確かめられませんでした**"
            "（合わなかったのではありません）")
        say("  → Excel か LibreOffice で一度開いて保存し直すと、計算結果が入ります")
    # ★ D6 差し戻し（実弾検分・2026-08-21）: 正常なファイルは名指ししない ── 名指しは
    #   異常のあるファイル（取れなかった／閉じる検査の不一致／シート fallback）だけ。
    #   正常分（並べ替え・合計行の除外・行の完全会計）は 1 行の集計に畳む。
    #   全ファイルの内訳は --json に既にそのまま載っている（ここでは減らさない）。
    for f in skipped:
        say(f"  ⚠ {f['name']}: 取れなかった（{f['reason']}）── 中身の検査は未実施")
    reordered_files = [f for f in files_json if f.get("reordered")]
    if reordered_files:
        say(f"  並べ替えて照合: {len(reordered_files)} 冊（内訳は --json）")
    for f in sheet_fallbacks:
        say(f"  {f['name']}: シート『{f['wanted']}』が無いので1枚目『{f['used']}』を使いました")
    if excluded_detail:
        total_excluded_rows = sum(len(entry["rows"]) for entry in excluded_detail)
        say(f"  合計行 {total_excluded_rows} 行を {len(excluded_detail)} 冊で除外"
            "（内訳は --json）")
    for m in mismatches:
        say(f"  ⚠ {m['name']}: 合計行({m['row']}行目) の値 "
            f"{primitives.fmt_num(m['excluded_value'])} ≠ 明細の和 "
            f"{primitives.fmt_num(m['adopted_sum'])}")
    for w in total_word_warnings:
        say(f"  ⚠ {w['file']} の{w['row']}行目に合計語『{w['word']}』を含む行が"
            "積まれています（除外していません・確認してください）")
    if files_json:   # ★ 憲法⑨ 行の完全会計: どの行もどれかの勘定に入っている（全冊で成立・集計）
        say(f"  行の完全会計: {len(files_json)} 冊すべてで成立"
            "（データ行 = 一致 + 不一致 + 合計行の除外・内訳は --json）")
    say(f"出力データ行数: {total_matched}")
    for col_name, both in post.get("sums", {}).items():
        say(f"Σ{col_name}: 元 {primitives.fmt_num(both['source'])} / "
            f"出力 {primitives.fmt_num(both['output'])}")
    if rebuilt_own_output:
        say(f"（前回の抽出出力『{out.name}』を作り直しました）")
    return 0


def _unreadable_book_for_match_message(path: Path) -> str:
    """M3 で1冊が読めなかった時の断り文言。★ 検体④の直し（2026-08-22）: .csv は
       openpyxl で読めないので必ずここに来るが、旧文言「.xlsx 形式か確認してください」は
       .csv に対する誤誘導（.csv は形式が壊れているのではなく、そもそも別の入口が要る）。
       csv の扱い（`ailine csv`）へ名指しで誘導する。"""
    if path.suffix.lower() == CSV_SUFFIX:
        return (f"？ {path.name} は csv 形式なので、このままでは照合できません。"
                f"先に `ailine csv {path.name}` で xlsx に変換してから、"
                "その xlsx 同士（または xlsx と csv 変換後のファイル）で照合してください。")
    return f"？ {path.name} を読めませんでした。壊れていないか、.xlsx 形式か確認してください。"


def _peek_match_book(path: Path):
    """M3 専用の軽い読み: 見出し行推定（既存の StructDump ヒューリスティクスを1回だけ）+
       全データ行の読み取り。戻り値 (header_row, headers, rows) ── 読めなければ
       (None, None, None)（呼び出し側が名指しで exit 3 にする）。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None, None, None
    try:
        ws = wb.worksheets[0]
        scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
        rows_stats = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
        row, confident = detect_header_row({"rows": rows_stats})
        header_row = row if confident else 1
        headers = multifile.read_row_headers(ws, header_row)
        rows = multifile_match.read_data_rows(ws, header_row, headers)
        return header_row, headers, rows
    except Exception:
        return None, None, None
    finally:
        wb.close()


def _match_condition(key_a: str, key_b: str, amount_a: str, amount_b: str, book_b_name: str,
                     headers_a: list, headers_b: list) -> dict:
    """出力ブックの docProps/description へ焼く条件（機械可読・verify の入口が読む）。
       ★ 絶対パスを焼かない（検分ごと発注者に渡るため）── book_b はファイル名のみ。
       ★ _verify_match 用に両冊のヘッダー全部も焼く（verify.py の _find_header_row が
       xml_readback だけで見出し行を引き当てるのに必要 ── 列名だけでは足りない・
       openpyxl の StructDump ヒューリスティクスを verify 側に持ち込まないための代替）。"""
    return {"tool": "ailine", "kind": "match", "key_a": key_a, "key_b": key_b,
            "amount_a": amount_a, "amount_b": amount_b, "b": book_b_name,
            "headers_a": list(headers_a), "headers_b": list(headers_b)}


def _own_match_output_status(path: Path, cond: dict) -> tuple:
    """path が①ailine 産か（mark）②M3 照合の自分の前回出力で、かつ焼いた条件が
       今回と完全一致するか（same_condition）を返す（_own_extract_output_status と同じ線）。"""
    mark = multifile_stack.own_output_mark(path)
    if mark != multifile_match.CREATOR_MARK:
        return mark, False
    _creator, description = xml_readback.read_core_properties(path)
    try:
        existing = json.loads(description) if description else None
    except (TypeError, ValueError):
        existing = None
    same_condition = existing == cond
    return mark, same_condition


def _match_postcondition_fail(label: str, source, output) -> int:
    """M3 事後条件の破れ（両側の数字つき・design v2: 破れは exit 1）。"""
    print(f"⚠ 事後条件が破れた: {label}  元(算出) {inspection.fmt_num(source)} / "
          f"出力(書いた直後) {inspection.fmt_num(output)}")
    return 1


def cmd_run_match(a: argparse.Namespace, book_a: Path, book_b: Path, task: str) -> int:
    """`ailine run <A.xlsx> <B.xlsx> "<依頼>"`: M3 ── 2冊の照合（突き合わせ）。
       DESIGN-20260821-multifile.md M3 設計 v2（凍結）。芯: 候補を並べて差額だけ機械で保証、
       決めるのは人 ── この経路は判断しない（消し込み・照合完了は名乗らない・憲法2）。
       ★ 翻訳（translate_task）はこの経路で呼ばない（M8: LLM には1語も足さない）── 列対応は
       機械3段（依頼文の名指し→型→曖昧なら exit 3）で決める。原本2冊は読むだけ。"""
    as_json = bool(getattr(a, "json", False))
    say = (lambda *args, **kw: None) if as_json else print

    for _b in (book_a, book_b):
        blocked = refuse_if_locked(_b)
        if blocked is not None:
            return blocked

    sha_a = hashlib.sha256(book_a.read_bytes()).hexdigest()
    sha_b = hashlib.sha256(book_b.read_bytes()).hexdigest()
    if sha_a == sha_b:
        say(f"（{book_a.name} と {book_b.name} は同一の内容です（sha256 一致）。"
            "同じデータとして続行します）")

    header_row_a, headers_a, rows_a = _peek_match_book(book_a)
    if headers_a is None:
        print(_unreadable_book_for_match_message(book_a))
        return 3
    header_row_b, headers_b, rows_b = _peek_match_book(book_b)
    if headers_b is None:
        print(_unreadable_book_for_match_message(book_b))
        return 3

    # ★ 列対応（機械3段・LLM ゼロ）: 依頼文の名指し → 型で絞る → 曖昧なら exit 3（候補つき）。
    #   ★ 一括検出: 決まらなかった役割を全部集めてから報告する（1件目で止めない）。
    resolution = multifile_match.resolve_columns(task, headers_a, rows_a, headers_b, rows_b)
    if not resolution.ok:
        say(f"■ ailine run（2冊の照合）  A={book_a}  B={book_b}")
        for side, role, candidates in resolution.unresolved:
            label = "キー" if role == "key" else "金額"
            book_label = book_a.name if side == "A" else book_b.name
            if candidates:
                cand_txt = "、".join(str(c) for c in candidates)
                say(f"？ {book_label} の{label}列が依頼文から決まりません。候補: {cand_txt}。"
                    f"依頼文に列名を含めて（例:『{candidates[0]}を{label}に』）もう一度実行してください。")
            else:
                say(f"？ {book_label} に{label}に使える列が見つかりません。")
        return 3
    key_a, key_b, amount_a, amount_b = (resolution.key_a, resolution.key_b,
                                         resolution.amount_a, resolution.amount_b)

    groups = multifile_match.compute_match(headers_a, rows_a, key_a, amount_a,
                                            headers_b, rows_b, key_b, amount_b)
    a_total_keyed, b_total_keyed = multifile_match.side_totals(groups)
    # ★ 実弾検分の差し戻し#1: 合計行対策（design v2「単位L」節）。キーが空 かつ 金額 =
    #   同じ冊の他のデータ行の和、を注記として集める（除外はしない・算術のみ）。
    total_notes = (multifile_match.possible_total_row_notes(headers_a, rows_a, key_a, amount_a, "A")
                   + multifile_match.possible_total_row_notes(headers_b, rows_b, key_b, amount_b, "B"))

    # ★ 出力先（A の親 + 機械命名・sanitize+条件ハッシュ・「照合」を名前に含める）と
    #   書き込みの関所（40行読む前に判定して印字 ── stack/extract と同じ配役）。
    cond = _match_condition(key_a, key_b, amount_a, amount_b, book_b.name, headers_a, headers_b)
    digest = hashlib.sha256(json.dumps(cond, sort_keys=True, ensure_ascii=False)
                            .encode("utf-8")).hexdigest()[:6]
    raw_stem = f"{book_a.stem}_照合_{digest}"
    out = book_a.parent / f"{extract_multi.sanitize_filename(raw_stem)}.xlsx"
    rebuilt_own_output = False
    if out.exists():
        mark, same_condition = _own_match_output_status(out, cond)
        if mark == multifile_match.CREATOR_MARK and same_condition:
            rebuilt_own_output = True
        else:
            return _refuse_output_conflict(out, mark)

    say(f"■ ailine run（2冊の照合）  A={book_a}  B={book_b}")
    say(f"出力先: {out}")
    say(f"キー: {key_a}(A) / {key_b}(B)　金額: {amount_a}(A) / {amount_b}(B)")

    workdir = Path(tempfile.mkdtemp(prefix="ailine_match_"))
    try:
        tmp_out = workdir / out.name
        wb_out, key_to_detail_row = multifile_match.build_workbook(
            groups, headers_a, headers_b, book_a.name, book_b.name)
        # ★ 実弾検分の差し戻し#3（gap#1 の解消）: 差額あり・A/Bのみ・キー不明・合計行の可能性を
        #   検分の所見表に載せ、ブック内リンク（明細シートの該当キーの先頭行）で誘導する。
        findings = multifile_match.build_findings(groups, key_to_detail_row, total_notes,
                                                   book_a.name, book_b.name)
        wb_out.properties.creator = multifile_match.CREATOR_MARK
        wb_out.properties.description = json.dumps(cond, ensure_ascii=False)
        denom_lines = [f"A: {len(rows_a)} 行 / B: {len(rows_b)} 行 → {len(groups)} キーに整理"
                       "（キー不明を含む）"]
        acct_lines = [f"行の完全会計 A: {len(rows_a)} = キー行 {a_total_keyed}",
                      f"行の完全会計 B: {len(rows_b)} = キー行 {b_total_keyed}"]
        inspection.build_sheet(
            wb_out, findings=findings, denominator_lines_=denom_lines, accounting_lines=acct_lines,
            sums={}, file_sheet_map=[(book_a.name, "1枚目", ""), (book_b.name, "1枚目", "")],
            out_dir=out.parent, source_dir=book_a.parent)
        kb_ws = wb_out[inspection.SHEET_NAME]
        extra_lines = ["", "差額分布（Counter・キー不明を除く）:"]
        dist = multifile_match.diff_distribution(groups)
        if dist:
            extra_lines += [f"・差額 {inspection.fmt_num(v)} が {n} 件" for v, n in dist]
        else:
            extra_lines.append("（無し）")
        fuzzy = multifile_match.fuzzy_candidates(groups)
        extra_lines += ["", "表記ゆれ候補（空白除去後の部分文字列一致・並べるだけ・自動採用しない）:"]
        extra_lines += fuzzy if fuzzy else ["（無し）"]
        extra_lines += ["", "未実施: どの行がどの行に対応するかは検査していません（キー単位の集計のみ）。"]
        r = kb_ws.max_row + 1
        for line in extra_lines:
            kb_ws.cell(row=r, column=1, value=line)
            r += 1
        inspection.autosize_columns(kb_ws)
        wb_out.save(tmp_out)

        # ★ 事後条件（design v2「事後条件」節）: xml_readback の独立読みで書いた直後の
        #   中身を検算する。破れは両側の数字つき exit 1（stack/extract は exit 5 ── M3 は
        #   design 側の指示どおり exit 1 を使う）。
        readback = xml_readback.read_grid(tmp_out, sheet_name=multifile_match.MATCH_SHEET_NAME)
        out_rows = xml_readback.data_row_numbers(readback, header_row=1)
        if len(out_rows) != len(groups):
            wb_out.close()
            return _match_postcondition_fail("行数(照合表)", len(groups), len(out_rows))
        grid = readback["grid"]
        col_a_count, col_b_count, col_diff, col_a_sum, col_b_sum = 2, 4, 6, 3, 5
        a_count_out = sum(grid.get((r, col_a_count), 0) or 0 for r in out_rows)
        b_count_out = sum(grid.get((r, col_b_count), 0) or 0 for r in out_rows)
        data_a = xml_readback.read_grid(book_a)
        a_headers_x = xml_readback.header_names(data_a, header_row=header_row_a)
        a_rows_x = [r for r in xml_readback.data_row_numbers(data_a, header_row_a)
                    if xml_readback.row_has_any_value(data_a, r, len(a_headers_x))]
        data_b = xml_readback.read_grid(book_b)
        b_headers_x = xml_readback.header_names(data_b, header_row=header_row_b)
        b_rows_x = [r for r in xml_readback.data_row_numbers(data_b, header_row_b)
                    if xml_readback.row_has_any_value(data_b, r, len(b_headers_x))]
        if a_count_out != len(a_rows_x):
            wb_out.close()
            return _match_postcondition_fail("完全会計(A)", len(a_rows_x), a_count_out)
        if b_count_out != len(b_rows_x):
            wb_out.close()
            return _match_postcondition_fail("完全会計(B)", len(b_rows_x), b_count_out)
        # ★ review5#1 critical の直し: 件数だけでなく金額（a_sum/b_sum）も原本から独立に
        #   （xml_readback の grid ── openpyxl を経由しない別実装で）再集計し、書いた
        #   セルと全キーで突き合わせる。内部整合（diff=a-b）だけでは compute_match の
        #   金額そのものの改竄を見逃す（片配線だった箇所）。
        a_sums_x = multifile_match.independent_key_sums(data_a["grid"], a_rows_x, a_headers_x,
                                                          key_a, amount_a)
        b_sums_x = multifile_match.independent_key_sums(data_b["grid"], b_rows_x, b_headers_x,
                                                          key_b, amount_b)
        for r in out_rows:
            key_val = grid.get((r, 1))
            a_sum_v = grid.get((r, col_a_sum), 0) or 0
            b_sum_v = grid.get((r, col_b_sum), 0) or 0
            diff_v = grid.get((r, col_diff), 0) or 0
            nk = None if key_val == multifile_match.UNKNOWN_KEY_LABEL                 else multifile_match.normalize_key(key_val)
            expect_a = a_sums_x.get(nk, 0.0)
            expect_b = b_sums_x.get(nk, 0.0)
            if abs(a_sum_v - expect_a) > multifile_match.TOLERANCE:
                wb_out.close()
                return _match_postcondition_fail(f"A側合計の独立再集計({key_val})", expect_a, a_sum_v)
            if abs(b_sum_v - expect_b) > multifile_match.TOLERANCE:
                wb_out.close()
                return _match_postcondition_fail(f"B側合計の独立再集計({key_val})", expect_b, b_sum_v)
            if abs(diff_v - (a_sum_v - b_sum_v)) > multifile_match.TOLERANCE:
                wb_out.close()
                return _match_postcondition_fail(f"差額の算術({r}行目)", a_sum_v - b_sum_v, diff_v)
        wb_out.close()

        # ★ 原本2冊 sha 無変更（読むだけの経路だが、書き込み経路のどこかで誤って開いて
        #   保存していないかの最終防衛）。
        if hashlib.sha256(book_a.read_bytes()).hexdigest() != sha_a:
            return _match_postcondition_fail("原本無変更(A)", sha_a, "変化あり")
        if hashlib.sha256(book_b.read_bytes()).hexdigest() != sha_b:
            return _match_postcondition_fail("原本無変更(B)", sha_b, "変化あり")

        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_out, out)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    # ★ D6: 正常キーの洪水を作らない ── 差額ありキーだけ名指し、正常分は1行に畳む。
    # ★ 実弾検分の差し戻し#1/#2/#3: 端末の1行は所見（findings）の next_step をそのまま
    #   使う ── 検分シートと文言が二重管理にならない。キー不明の重複文言（『キー不明: キー不明』）
    #   は next_step 側の書式変更（build_findings）で解消済み。
    mismatched = [g for g in groups if abs(g.diff) > multifile_match.TOLERANCE]
    ok_count = len(groups) - len(mismatched)
    # ★ 2026-08-24（第三波 S5）: 「12 キー中 0 キーが差額 0」は、二重の否定めいて読めず
    #   何件ずれているのかが一目で分からなかった。両側の数字を素直に並べる。
    key_label = key_a if key_a == key_b else f"{key_a}/{key_b}"
    say(f"『{key_label}』{len(groups)} 件: 差額なし {ok_count} 件 / "
        f"差額あり {len(mismatched)} 件")
    for f in findings:
        say(f"  ⚠ {f.next_step}")
    if rebuilt_own_output:
        say(f"（前回の照合出力『{out.name}』を作り直しました）")
    if as_json:
        print(json.dumps({
            "out": str(out), "key_a": key_a, "key_b": key_b, "amount_a": amount_a,
            "amount_b": amount_b, "a_rows": len(rows_a), "b_rows": len(rows_b),
            "keys": len(groups), "mismatched": len(mismatched),
        }, ensure_ascii=False))
    return 0


def _own_csv_output_status(path: Path, source_sha256: str) -> tuple:
    """path が①ailine 産か（mark）②CSV 検疫の自分の前回出力で、かつ元 CSV の sha256 が
       今回と完全一致するか（same_source）を返す（_own_extract_output_status と同じ線・
       CSV は「条件」の代わりに「どの原本から作ったか」を同一性の根拠にする）。"""
    mark = multifile_stack.own_output_mark(path)
    if mark != csv_quarantine.CREATOR_MARK:
        return mark, False
    _creator, description = xml_readback.read_core_properties(path)
    try:
        cond = json.loads(description) if description else None
    except (TypeError, ValueError):
        cond = None
    same_source = (isinstance(cond, dict) and cond.get("tool") == "ailine"
                   and cond.get("kind") == "csv" and cond.get("source_sha256") == source_sha256)
    return mark, same_source


@dataclass
class _CsvEvaluation:
    """CSV 検疫の評価結果（書き込みより前・出力パスに依存しない）。
       error が None でなければ硬い拒否（行数上限超過・文字コード判定不能）── 呼び出し側は
       書き込みに進まず名指しで断る。"""
    error: str = None
    sha256: str = None
    encoding: object = None
    parsed: object = None
    classifications: list = None
    warnings: list = None


def _csv_record_label(parsed, raw_idx: int, header_offset: int) -> str:
    """レコード番号 + 物理行範囲の名指し（検分の名指し用）。raw_idx はヘッダを含む
       物理レコード列全体に対する0起点（csv_quarantine.ParseResult の契約どおり）。"""
    if header_offset and raw_idx == 0:
        return "見出し行"
    data_idx = raw_idx - header_offset
    if parsed.records and 0 <= data_idx < len(parsed.records):
        rec = parsed.records[data_idx]
        return f"レコード{raw_idx + 1}（物理 {rec.first_line}〜{rec.last_line} 行）"
    return f"レコード{raw_idx + 1}"


def _format_offending_records(parsed, header_offset: int, rec_indices: list, col_idx: int) -> str:
    """検分の差し戻し（2026-08-22）: undecidable 列の ⚠ を列名だけで終わらせず、
       原因セルのレコード（物理行）と値まで名指しする（憲法①「最悪でも修正箇所に
       誘導」）。rec_indices は csv_quarantine.undecidable_offending_indices が返す
       列内 0 起点 index（= parsed.records の index と同じ空間）。
       先頭3件+『ほかN件』で列挙を打ち切る（rec_indices が空なら空文字を返す）。"""
    if not rec_indices:
        return ""
    shown = rec_indices[:3]
    parts = []
    for rec_idx in shown:
        raw_idx = rec_idx + header_offset
        label = _csv_record_label(parsed, raw_idx, header_offset)
        rec = parsed.records[rec_idx]
        value = rec.cells[col_idx] if col_idx < len(rec.cells) else ""
        parts.append(f"{label}: {value}")
    text = "、".join(parts)
    remaining = len(rec_indices) - len(shown)
    if remaining > 0:
        text += f"、ほか{remaining}件"
    return f"（該当: {text}）"


def _evaluate_csv(csv_path: Path) -> _CsvEvaluation:
    """`ailine csv` / run 暗黙前段が共有する検疫パイプライン（書き込みより前の段）。
       ★ 行数上限は decode より前に数える（csv_quarantine.MAX_ROWS・設計 v2 バー）。
       ★ ここは ailine_core/csv_quarantine.py の決定論の関数を呼ぶだけ ── LLM は
       一切呼ばない（憲法: LLM はデータに1バイトも触らない）。"""
    raw = csv_path.read_bytes()
    line_count = raw.count(b"\n") + 1
    if line_count > csv_quarantine.MAX_ROWS:
        return _CsvEvaluation(
            error=f"行数が多すぎます（目安 {line_count} 行 > 上限 {csv_quarantine.MAX_ROWS} 行）。"
                  "ファイルを分割してから実行してください。")
    sha = csv_quarantine.sha256_bytes(raw)
    try:
        enc = csv_quarantine.detect_encoding(raw)
    except csv_quarantine.UndecidableEncodingError as e:
        return _CsvEvaluation(error=str(e), sha256=sha)
    text = raw.decode(enc.encoding)
    parsed = csv_quarantine.parse_csv(text, has_header=True, source_sha256=sha)
    columns = csv_quarantine.build_columns(parsed.header, parsed.records)
    classifications = [csv_quarantine.classify_column(col) for col in columns]

    warnings = []
    # ★ 第二波 ③ 追補: 文字コードの曖昧さは全ての報告関数が冒頭行
    #   「文字コード: {encoding} で読みました（…両方成立…）」で必ず開示する（enc.ambiguous
    #   参照）。ここでも ⚠ として二重に積むと、実在する内容語彙（例: 日本語見出し）を持つ
    #   きれいな CSV が「制御文字1件」のような単一の実欠陥に対しても「⚠ 2 件」と数える
    #   （実欠陥1件+既出の曖昧さ再掲1件）── 制御文字の二重報告と同根の重複。utf-8 は
    #   確定的に選ばれ実際に使われているので、これは是正の必要な ⚠ ではなく開示で足りる。
    for rename in parsed.header_renames:
        warnings.append(f"見出し『{rename.original}』が重複していたため"
                        f"『{rename.renamed}』に機械リネームしました")
    header_offset = 1 if parsed.header else 0
    for raw_idx, expected, actual in parsed.column_count_mismatches:
        warnings.append(f"{_csv_record_label(parsed, raw_idx, header_offset)}: "
                        f"列数が {actual} 列（期待 {expected} 列）")
    for raw_idx in parsed.unterminated_quote_records:
        warnings.append(f"{_csv_record_label(parsed, raw_idx, header_offset)}: "
                        "引用符が閉じていない疑いがあります")
    for i, cls in enumerate(classifications):
        if cls.warn:
            name = parsed.header[i] if i < len(parsed.header) else f"列{i + 1}"
            reason_txt = "・".join(_CSV_REASON_LABELS.get(r, r) for r in cls.reasons)
            offending = csv_quarantine.undecidable_offending_indices(columns[i], cls.reasons)
            detail = _format_offending_records(parsed, header_offset, offending, i)
            warnings.append(f"列『{name}』: 判定不能（{reason_txt}）── "
                            f"文字列として保持し、この列に判定を出しません{detail}")
    for rec_idx, rec in enumerate(parsed.records):
        raw_idx = rec_idx + header_offset
        # ★ 第二波 ③（本家 bug_001）: 制御文字の検出は writer 側
        #   write_result.removed_control_chars に一本化する（re.sub 全件で厳密・行/列/
        #   コードの名指しはそちらで足りる）。ここで重ねて拾うと同じ1件が2行の ⚠ になり、
        #   warn_count（「⚠ N 件」の N）も二重計上される（1件を2件と数える）。
        for col_idx, length in csv_quarantine.overlong_cells(rec.cells):
            warnings.append(f"{_csv_record_label(parsed, raw_idx, header_offset)} "
                            f"{col_idx + 1}列目: {length} 文字（Excel 上限 32,767 超）")
    warnings.extend(csv_quarantine.detect_excel_damage(columns, classifications, header=parsed.header))

    return _CsvEvaluation(error=None, sha256=sha, encoding=enc, parsed=parsed,
                           classifications=classifications, warnings=warnings)


def _write_csv_output(evaluation: _CsvEvaluation, out_path: Path) -> tuple:
    """検疫結果を xlsx へ書き、事後条件（compare_against_quarantine）を検算する。
       own 印（csv_quarantine.CREATOR_MARK）と機械可読の条件（kind:"csv"・
       source_sha256）を docProps へ焼く（own_output_mark 経路が使う・stack/extract と
       同じ配線）。"""
    description = json.dumps({"tool": "ailine", "kind": "csv", "version": 1,
                              "source_sha256": evaluation.sha256}, ensure_ascii=False)
    write_result = csv_quarantine.write_quarantined_xlsx(
        evaluation.parsed, evaluation.classifications, out_path,
        has_header=True, creator=csv_quarantine.CREATOR_MARK, description=description)
    compare_result = csv_quarantine.compare_against_quarantine(write_result.declared, out_path)
    return write_result, compare_result


# ★ 列判定の開示語彙（tests/test_csv_truth_table.py の reason コード → 人の言葉）。
#   凍結対象ではない（人間向けの言い回しであって規則そのものではない）── 語彙を増やす時は
#   ここへ足すだけでよい（未知のコードはそのままの文字列を出す・fail-open で開示は続ける）。
_CSV_REASON_LABELS = {
    "leading_zero": "先頭ゼロ",
    "formula_head": "先頭が = か @（数式のような書式）",
    "digit_overflow": "16桁以上（Excel の精度限界を超える）",
    "surrounding_space": "前後に空白",
    "fullwidth_digit": "全角数字を含む",
    "accounting_negative": "会計負数の書式（△ か括弧）の可能性",
    "empty_column": "全セル空欄",
    "excel_error_token": "Excel エラー値のような文字列",
    "wareki_out_of_scope": "和暦（対象外）",
    "comma_grouped": "カンマ区切りの桁区切り",
    "eight_digit_maybe_date": "8桁の数字（日付の可能性あり・数値のまま保持）",
    "mixed_confident": "日付と数値が同居",
    "calendar_invalid": "暦として成立しない日付形式",
    "comma_inconsistent": "カンマ区切りの有無が列内で不揃い",
}


def _csv_kind_label(cls) -> str:
    reason_txt = "・".join(_CSV_REASON_LABELS.get(r, r) for r in cls.reasons)
    if cls.kind == "number":
        return "数値として読み取り" + (f"（{reason_txt}）" if reason_txt else "")
    if cls.kind == "date":
        return "日付として読み取り"
    if cls.kind == "string":
        return "文字列として保持" + (f"（{reason_txt}）" if reason_txt else "")
    return f"判定不能・文字列として保持（{reason_txt}）"


def _render_csv_report(csv_path: Path, out_path: Path, evaluation: _CsvEvaluation,
                       write_result, compare_result) -> list:
    """`ailine csv` の人間向け報告（REVIEW-20260822-csv-architect.md §『✓ 文例』の凍結形）。
       ★ 決裁③(2026-08-22): ⚠ が1件でもあれば ✓ でなく △ を名乗る。csv の主張は Claim
       機構と別物なので、Claim を無理に通さず文言だけ △ 系に揃える。
       ★「正しく読み込みました」とは言わない ── 言えるのは転送段の宣言（欠落/不一致/余剰の
       3計数）+ 列ごとの判定開示 + 文字列保持列は Σ 検算していない開示 + 原本 sha 一致だけ。"""
    lines = [f"■ ailine csv  file={csv_path}"]
    enc = evaluation.encoding
    enc_note = "（utf-8/cp932 どちらでも読めたため utf-8 で続行）" if enc.ambiguous else ""
    lines.append(f"文字コード: {enc.encoding} で読みました{enc_note}")
    lines.append(f"出力先: {out_path}")
    warn_count = len(evaluation.warnings) + len(write_result.removed_control_chars)
    claim = (f"読み取った {write_result.rows_written} 行×{write_result.cols_written} 列を"
             f" 1 セルも変えずに書いた（欠落{len(compare_result.missing)}・"
             f"不一致{len(compare_result.mismatched)}・余剰{len(compare_result.surplus)}）")
    if warn_count:
        lines.append(f"△ {claim} ── ただし ⚠ {warn_count} 件を先に確認してください")
    else:
        lines.append(f"✓ {claim}")
    header = evaluation.parsed.header
    string_kept = []
    for i, cls in enumerate(evaluation.classifications):
        name = header[i] if i < len(header) else f"列{i + 1}"
        lines.append(f"  {name}: {_csv_kind_label(cls)}")
        if cls.kind in ("string", "undecidable"):
            string_kept.append(name)
    if string_kept:
        lines.append(f"（{'、'.join(string_kept)} は文字列として保持したため、"
                     "Σ で検算していません）")
    for w in evaluation.warnings:
        lines.append(f"  ⚠ {w}")
    for row, col, code in write_result.removed_control_chars:
        lines.append(f"  ⚠ {row}行目{col}列目: 制御文字 {code} を除去して書きました")
    lines.append(f"原本 CSV: 変更なし（sha256 {evaluation.sha256} 一致）")
    return lines


def _render_csv_transfer_failure(csv_path: Path, out_path: Path, evaluation: _CsvEvaluation,
                                 write_result, compare_result) -> list:
    """★ 致命①(2026-08-23レビュー): compare_result（欠落/不一致/余剰）が非 ok の時に
       `ailine csv` が呼ぶ報告。✓ でも △ でもない ── 「1 セルも変えずに書いた」という
       転送の主張自体が成立していないので、× 側（該当セルを名指しして exit 非 0）に倒す。
       暗黙前段 _cmd_run_csv_prestage は元から `or not compare_result.ok` を見ていた
       （片配線の解消・明示コマンド側にも同じ判定を通す）。"""
    lines = [f"■ ailine csv  file={csv_path}"]
    enc = evaluation.encoding
    enc_note = "（utf-8/cp932 どちらでも読めたため utf-8 で続行）" if enc.ambiguous else ""
    lines.append(f"文字コード: {enc.encoding} で読みました{enc_note}")
    lines.append(f"出力先: {out_path}")
    lines.append(f"× 読み取った {write_result.rows_written} 行×{write_result.cols_written} 列を"
                 f" 1 セルも変えずに書けませんでした"
                 f"（欠落{len(compare_result.missing)}・不一致{len(compare_result.mismatched)}・"
                 f"余剰{len(compare_result.surplus)}）")
    for w in evaluation.warnings:
        lines.append(f"  ⚠ {w}")
    for row, col, code in write_result.removed_control_chars:
        lines.append(f"  ⚠ {row}行目{col}列目: 制御文字 {code} を除去して書きました")
    for row, col in compare_result.missing:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で欠落しました")
    for row, col, dval, aval in compare_result.mismatched:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で値が変わりました（{dval!r} → {aval!r}）")
    for row, col, aval in compare_result.surplus:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で余剰として現れました（{aval!r}）")
    lines.append(f"原本 CSV: 変更なし（sha256 {evaluation.sha256} 一致）")
    return lines


def cmd_run_csv(a: argparse.Namespace) -> int:
    """`ailine csv <file>`: CSV 検疫の明示入口（設計 v2「入口2つ」の②・7B ゼロ・0秒）。
       DESIGN-20260821-multifile.md「CSV 検疫 設計 v2」・REVIEW-20260822-csv-architect.md。
       ★ normalize_book/basrun_apply のどちらも呼ばない（LO の CSV インポートが 0 落ちの
       発生源・実測 0123→123 ── この経路は LO に一切触れない・構造の番人）。"""
    csv_path = Path(a.file).resolve()
    if not csv_path.exists():
        print(f"文書が無い: {csv_path}")
        return 1
    evaluation = _evaluate_csv(csv_path)
    if evaluation.error:
        print(f"× {evaluation.error}")
        return 3
    out = csv_path.with_suffix(".xlsx")
    if out.exists():
        mark, same_source = _own_csv_output_status(out, evaluation.sha256)
        if not (mark == csv_quarantine.CREATOR_MARK and same_source):
            return _refuse_output_conflict(out, mark)
    write_result, compare_result = _write_csv_output(evaluation, out)
    if not compare_result.ok:
        for ln in _render_csv_transfer_failure(csv_path, out, evaluation, write_result, compare_result):
            print(ln)
        _record_csv_conversion_history(csv_path, out, ok=False)
        return 3
    for ln in _render_csv_report(csv_path, out, evaluation, write_result, compare_result):
        print(ln)
    _record_csv_conversion_history(csv_path, out, ok=True)
    return 0


def _record_csv_conversion_history(csv_path: Path, out_path: Path, ok: bool) -> None:
    """★ 第二波 ①(AILINE_HOME): `ailine csv` の変換も history.jsonl に残す（run と同じ
       HISTORY_FILE・AILINE_HOME 配下）。build_history_entry は DSL の run 向けの形
       （task/model/command/postcondition 等）なので流用せず、csv 変換に要る最小限の
       形で直接 append する。書き込みに失敗しても csv 変換自体の結果（rc・出力ファイル）
       は変えない（履歴は付帯情報）。"""
    try:
        append_history({
            "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "book": str(csv_path),
            "task": f"csv: {csv_path.name} → {out_path.name}",
            "model": None,
            "ok": ok,
            "dry": False,
            "attempts": 1,
            "failure_kind": None if ok else "csv_transfer_failed",
            "error_detail": None,
            "changes": [],
            "out": str(out_path),
            "path": "csv",
            "command": None,
            "postcondition": None,
            "provenance": None,
            "fidelity": None,
        })
    except OSError:
        pass


def _cmd_run_csv_prestage(a: argparse.Namespace) -> int:
    """`ailine run <file.csv> "タスク"`: 設計 v2「入口2つ」の①（暗黙前段）── 検疫して
       <csv_stem>.xlsx を作り、以後は既存の xlsx 機械（_cmd_run_body）へ a.book を
       差し替えて渡す。
       ★ .csv がこの先（normalize_book/basrun_apply）へ絶対に渡らない構造的保証: ここで
       作るのは実在する普通の .xlsx ファイルであり、以後の経路はもともと .xlsx しか
       受け取らない ── 「.csv」という概念自体がここで完全に消える。
       ★ 決裁: ⚠ が1件でも出たら続行しない（undecidable を含むデータに op を適用するのは
       第一波の外・--force 系の逃げ道は作らない）。"""
    csv_path = Path(a.book).resolve()
    evaluation = _evaluate_csv(csv_path)
    if evaluation.error:
        print(f"× {evaluation.error}")
        return 3
    out = csv_path.with_suffix(".xlsx")
    if out.exists():
        mark, same_source = _own_csv_output_status(out, evaluation.sha256)
        if not (mark == csv_quarantine.CREATOR_MARK and same_source):
            return _refuse_output_conflict(out, mark)
    write_result, compare_result = _write_csv_output(evaluation, out)
    warn_count = len(evaluation.warnings) + len(write_result.removed_control_chars)
    if warn_count or not compare_result.ok:
        print(f"■ ailine run（CSV 検疫）  file={csv_path}")
        print(f"⚠ {csv_path.name} の読み取りに確認事項があるため、続行しません"
              f"（原本は無変更・検疫結果は {out.name} として書きました）。")
        for w in evaluation.warnings:
            print(f"  ⚠ {w}")
        for row, col, code in write_result.removed_control_chars:
            print(f"  ⚠ {row}行目{col}列目: 制御文字 {code} を除去して書きました")
        if not compare_result.ok:
            print(f"  ⚠ 転送の検算で 欠落{len(compare_result.missing)}・"
                  f"不一致{len(compare_result.mismatched)}・余剰{len(compare_result.surplus)} を検出しました")
        print(f"（内容を確認し、『ailine run {out.name} <依頼>』のように"
              "xlsx を直接指定してやり直してください）")
        return 3
    print(f"（{csv_path.name} を検疫し、{out.name} として書きました。以後はこのファイルに対して実行します）")
    a.book = str(out)
    return _cmd_run_body(a)


def _record_csv_export_history(book_path: Path, out_path: Path, sheet: str, ok: bool) -> None:
    """★ `ailine csv`（_record_csv_conversion_history）と対で history.jsonl に残す。
       書き込みに失敗しても export-csv 自体の結果は変えない（履歴は付帯情報）。"""
    try:
        append_history({
            "ts": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "book": str(book_path),
            "task": f"export-csv: {book_path.name}[{sheet}] → {out_path.name}",
            "model": None,
            "ok": ok,
            "dry": False,
            "attempts": 1,
            "failure_kind": None if ok else "csv_export_roundtrip_mismatch",
            "error_detail": None,
            "changes": [],
            "out": str(out_path),
            "path": "export-csv",
            "command": None,
            "postcondition": None,
            "provenance": None,
            "fidelity": None,
        })
    except OSError:
        pass


def _export_out_path(a: argparse.Namespace, default_path: Path) -> tuple:
    """書き出し系の出力先と、上書きの関所。戻り値 (出力先, 断る理由 or None)。

    ★ 2026-08-26（データの出入口の盲検・高7）: この関所は export-csv にしか無く、
      `export-pdf --out keep.pdf` は既存の PDF を**黙って上書き**していた（exit 0）。
      しかも下の docstring は「export-pdf には --out が在ったので非対称でもあった」と
      **自覚を書きながら**、直したのは --out が無かった側だけだった。
      ★ 片配線の自覚つきの片配線 ── 2 箇所に書き写さず、1 つの器官にして両方が通る。
    """
    raw = getattr(a, "out", None)
    out_path = Path(raw).resolve() if raw else default_path
    if out_path.exists() and not getattr(a, "overwrite", False):
        return out_path, (f"× 出力先 {out_path} が既にあります。"
                           f"別名にするなら --out、上書きしてよければ --overwrite を付けてください")
    return out_path, None


def _export_csv_out_path(a: argparse.Namespace, book_path: Path) -> tuple:
    """`export-csv` の出力先と、上書きの関所（_export_out_path への委譲）。

    ★ なぜ在るか（盲検レビュー・2026-08-24）: 出力先は book.with_suffix(".csv") 固定で
      --out が無く、出納帳の 4 月分を出した後に 5 月分を出すと**同じ名前で黙って上書き**
      していた。元 CSV から作った xlsx なら、人が置いた元ファイルを潰す。
      export-pdf には --out が在ったので、同じ道具の中で非対称でもあった。
    ★ 関所の形と exit 7 は既存の出力先の関所に合わせる（同じ意味の旗を 2 通り作らない）。
    """
    return _export_out_path(a, book_path.with_suffix(".csv"))


def cmd_export_csv(a: argparse.Namespace) -> int:
    """`ailine export-csv <book> --sheet <name> [--encoding utf-8|cp932]`: CSV_EXPORT の
       明示入口（DESIGN-20260824-format-map.md「CSV_EXPORT の憲法」）。★ csv_quarantine
       （CSV→xlsx の検疫）の逆方向 ── 書いた CSV を読み戻して元シートと突き合わせ、
       1セルも変えずに書いたことを主張する前に必ず検算する。LLM は使わない（0秒起動）。"""
    book_path = Path(a.book).resolve()
    if not book_path.exists():
        print(f"文書が無い: {book_path}")
        return 1
    enc = csv_export.resolve_encoding(a.encoding)
    if enc is None:
        print(f"× 未対応の文字コード: {a.encoding}（utf-8 / cp932 のみ対応）")
        return 3

    try:
        grid = csv_export.read_source(book_path, a.sheet)
    except Exception as e:
        print(f"× 読み込みに失敗しました: {e}")
        return 1
    if grid.sheet_fallback:
        print(f"× シート『{a.sheet}』がありません")
        return 1

    out_path, refuse = _export_csv_out_path(a, book_path)
    if refuse:
        print(f"■ ailine export-csv  file={book_path}  sheet={a.sheet}")
        print(refuse)
        return 7
    try:
        write_result = csv_export.build_csv(grid, enc)
    except csv_export.EncodingWriteError as e:
        print(f"× {e}")
        return 3

    try:
        with open(out_path, "wb") as f:
            f.write(write_result.raw_bytes)
    except OSError as e:
        print(f"■ ailine export-csv  file={book_path}  sheet={a.sheet}")
        print(f"× 書き出しに失敗しました: {e}"
              "（別のアプリが開いている可能性があります）")
        return 1

    # ★ 2026-08-24（盲検の契約レビュー）: 読み戻すのは**ディスク上のファイル**。
    #   初版はメモリ上の raw_bytes を読み直しており、「1 セルも変えずに書いた」は
    #   バイト列についての主張で、`売上.csv` 自体は一度も読んでいなかった
    #   ── 書き込みが途中で切れても、別のアプリが握っていても、✓ が出る形だった。
    try:
        written = out_path.read_bytes()
    except OSError as e:
        print(f"■ ailine export-csv  file={book_path}  sheet={a.sheet}")
        print(f"× 書き出したファイルを読み戻せませんでした: {e}")
        return 1
    if written != write_result.raw_bytes:
        print(f"■ ailine export-csv  file={book_path}  sheet={a.sheet}")
        print(f"× 書き出したファイルの中身が、書いたはずの内容と違います"
              f"（{len(write_result.raw_bytes)} バイトのつもりが {len(written)} バイト）")
        return 3
    roundtrip = csv_export.verify_roundtrip(write_result.declared, written, enc)

    bom_note = "（BOM 付き）" if enc.bom else "（BOM 無し）"
    lines = [f"■ ailine export-csv  file={book_path}  sheet={a.sheet}",
             f"出力先: {out_path}",
             f"文字コード: {enc.label}{bom_note} で書きました",
             csv_export.QUOTING_DISCLOSURE,
             "改行: CRLF"]
    body = (f"{write_result.rows_written} 行×{write_result.cols_written} 列を 1 セルも"
            f"変えずに書いた（欠落{len(roundtrip.missing)}・不一致{len(roundtrip.mismatched)}"
            f"・余剰{len(roundtrip.surplus)}）")
    # ★★ 2026-08-26（データの出入口の盲検・致命3）: 数式だがキャッシュ値が無いセルは
    #   grid から消えるので**分母からも消え**、空欄で書き出しても「欠落 0」が成立した
    #   ── 金額列が全部数式の見積書は、金額が全部空の CSV が ✓ で出る。
    #   同じブック・同じコマンドで、LO の再計算の有無だけで結果が変わり、どちらも ✓ だった。
    #   ★ 「検算していません」ではなく **「値が空になりました」** と実害の形で言う。
    uncached = list(getattr(grid, "uncached_formulas", ()) or ())
    for r, c in uncached[:5]:
        lines.append(f"⚠ {r}行目{c}列目は数式で、計算結果がファイルに入っていないため"
                     "**空欄で書き出しました**（元の値は分かりません）")
    if len(uncached) > 5:
        lines.append(f"⚠ 同じ形のセルがほかに {len(uncached) - 5} 個あります")
    if uncached:
        lines.append("  → Excel か LibreOffice で一度開いて保存し直すと、計算結果が入ります")
    if roundtrip.ok and not uncached:
        lines.append(f"✓ シートの {body}")
        for ln in lines:
            print(ln)
        _record_csv_export_history(book_path, out_path, a.sheet, ok=True)
        return 0
    if roundtrip.ok:
        # ★ 書けた分は本当に書けている ── ✓ は名乗らないが × でもない（△）。
        lines.append(f"△ シートの {body} ── ただし上の ⚠ は確かめられていません")
        for ln in lines:
            print(ln)
        _record_csv_export_history(book_path, out_path, a.sheet, ok=True)
        return 0

    # ★ 恒真殺し: 読み戻しで食い違ったら ✓ を名乗らない（csv_quarantine と同じ規律）。
    lines.append(f"× シートの {body.replace('1 セルも変えずに書いた', '1 セルも変えずに書けなかった')}")
    for row, col in roundtrip.missing:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で欠落しました")
    for row, col, dval, aval in roundtrip.mismatched:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で値が変わりました（{dval!r} → {aval!r}）")
    for row, col, aval in roundtrip.surplus:
        lines.append(f"  ⚠ {row}行目{col}列目: 転送で余剰として現れました（{aval!r}）")
    for ln in lines:
        print(ln)
    _record_csv_export_history(book_path, out_path, a.sheet, ok=False)
    return 3


def _soffice_to_pdf(book_path, out_path, sheet=None, orientation=None, fit_to_width=False) -> tuple:
    """LibreOffice を headless で呼んで PDF を作る（実 LO を呼ぶ唯一の門＝窒息点）。
       戻り値 (ok, 理由)。★ basrun と同じ office_dir() で soffice を探す ── 探し方を
       2 通り持つと片方だけ直る（片配線）ので、既にある入口を使う。

       ★ sheet_index（2026-08-24）: soffice の `--convert-to pdf` にシート指定は無い。
         初版は sheet 引数を受け取りながら**一切使っておらず、ブック全体が PDF になっていた**
         ── 帳票段で 50 社分のシートを作った後に 1 社を指定すると、全 50 社分と元データが
         1 つの PDF に入るのに ✓ が出た（**他人の売上を同封して顧客に送る事故**）。
         呼び出し側が「指定シートだけを残した一時コピー」を作って渡す形にした。"""
    basrun = _find_basrun_path()
    if basrun is None:
        return False, "LibreOffice の場所が分からない（basrun.py が見つかりません）"
    try:
        mod = _load_module_from_path(basrun, "_ailine_basrun_pdf")
        office = Path(mod.office_dir())
    except Exception as e:
        return False, f"LibreOffice を見つけられません: {e}"
    soffice = office / ("soffice.exe" if os.name == "nt" else "soffice")
    if not soffice.exists():
        return False, f"soffice が無い: {soffice}"
    # ★★ 2026-08-26（データの出入口の盲検・致命4）: ここは `--outdir <出力先の親>` を
    #   渡していたので、soffice は**必ず一度 `<ブック名>.pdf` を出力先フォルダに作る**。
    #   `--out` で別名を指定していても、そこに同名の PDF が在れば予告なく消える
    #   （実測: 顧客へ送った確定版 `請求書.pdf` が exit 0 のまま消滅）。
    #   ★ 根治: **人のフォルダを soffice に触らせない**。専用の一時フォルダへ出してから
    #     宣言した名前へ移す ── 出力先に現れるファイルは `out_path` ただ 1 つになる。
    _pdf_tmp = tempfile.TemporaryDirectory(prefix="ailine_pdfout_")
    out_dir = Path(_pdf_tmp.name)
    cmd = [str(soffice), "--headless", "--norestore", "--convert-to", "pdf",
            "--outdir", str(out_dir), str(book_path)]
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=180,
                               encoding="utf-8", errors="replace")
    except Exception as e:
        return False, f"変換に失敗しました: {e}"
    made = out_dir / (Path(book_path).stem + ".pdf")
    if not made.exists():
        _pdf_tmp.cleanup()
        return False, f"PDF が作られませんでした（{(proc.stderr or proc.stdout or "").strip()[:200]}）"
    try:
        # 一時フォルダは別ボリュームでありうるので move（replace はクロスデバイスで落ちる）
        shutil.move(str(made), str(out_path))
    except OSError as e:
        return False, f"PDF を {out_path} へ置けませんでした: {e}"
    finally:
        _pdf_tmp.cleanup()
    return True, ""


def _prepare_book_for_print(src: Path, dst: Path, sheet: str,
                             orientation=None, fit_to_width: bool = False,
                             keep_only_sheet: bool = False) -> None:
    """印刷用の調整をした**コピー**を作る（原本には一切触らない）。
       ★ 列幅は中身の長さから決める ── 全角を 2 文字ぶんとして数える
       （日本語の帳票で「あかつき商事」が切れて PDF から消えた実測への対応）。"""
    wb = openpyxl.load_workbook(src)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    if keep_only_sheet:
        # ★ 指定シート以外を落とす ── これが「--sheet が効く」の実体。
        for name in [n for n in wb.sheetnames if n != ws.title]:
            del wb[name]
    if fit_to_width:
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                width = sum(2 if ord(ch) > 127 else 1 for ch in text)
                letter = cell.column_letter
                if width > widths.get(letter, 0):
                    widths[letter] = width
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = min(width + 2, 120)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    if orientation:
        ws.page_setup.orientation = orientation
    wb.save(dst)
    wb.close()


def cmd_export_pdf(a: argparse.Namespace) -> int:
    """`ailine export-pdf <book> [--sheet S] [--out PATH]`: 表を紙の形で外へ出す
       （台帳 PRINT 2 件 + EXPORT_DOC 2 件）。LLM は使わない（0 秒起動）。

       ★ 主張の形は export-csv と同じ ── 出した後に**読み戻して**確かめてから ✓ を言う。
         読み戻しの道具（pdfplumber）が居ない環境では PDF は作るが ✓ を名乗らない。"""
    book_path = Path(a.book).resolve()
    if not book_path.exists():
        print(f"文書が無い: {book_path}")
        return 1
    try:
        wb = openpyxl.load_workbook(book_path, data_only=True)
    except Exception as e:
        print(f"× 読み込みに失敗しました: {e}")
        return 1
    sheet = a.sheet or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        print(f"× シート『{sheet}』がありません（ある: {chr(12289).join(wb.sheetnames)}）")
        wb.close()
        return 1
    values = [c for row in wb[sheet].iter_rows(values_only=True) for c in row if c is not None]
    wb.close()

    out_path, _refuse_pdf = _export_out_path(a, book_path.with_suffix(".pdf"))
    if _refuse_pdf:
        print(f"■ ailine export-pdf  file={book_path}  sheet={sheet}")
        print(_refuse_pdf)
        return 7
    # ★ 実測（2026-08-24）: 既定のまま出すと**列幅で文字が切れて PDF から消える**
    #   （「あかつき商事」が落ちた）。Excel 印刷の古典的な事故で、読み戻しが掴んだ。
    #   --fit-to-width / --orientation は原本を触らず**一時コピー**に効かせる。
    # ★ 2026-08-24: シート指定は**必ず**一時コピーで効かせる（soffice にシート指定が無い）。
    #   初版は --sheet を受け取って無視しており、他社の請求書を同封する事故になっていた。
    # ★ 2026-08-24（第三波 #9・実測）: この一時コピーは openpyxl の往復なので、
    #   **画像でない図形（描かれた角印・社判・テキストボックス）が落ちる**
    #   （画像として貼られた印は残る ── repo 内で両方を測って確かめた）。
    #   出来上がった PDF は完成品に見えるので、消える前に名指しする
    #   （「消えたものは差分に出ない」への処置）。
    vanishing = pdf_export.vanishing_shapes(book_path)
    source = book_path
    tmp_holder = tempfile.TemporaryDirectory(prefix="ailine_pdf_")
    source = Path(tmp_holder.name) / book_path.name
    try:
        _prepare_book_for_print(book_path, source, sheet,
                                 orientation=a.orientation, fit_to_width=a.fit_to_width,
                                 keep_only_sheet=True)
    except Exception as e:
        print(f"× 印刷用の調整に失敗しました: {e}")
        tmp_holder.cleanup()
        return 1
    ok, why = _soffice_to_pdf(source, out_path, sheet=sheet,
                               orientation=a.orientation, fit_to_width=a.fit_to_width)
    if tmp_holder is not None:
        tmp_holder.cleanup()
    lines = [f"■ ailine export-pdf  file={book_path}  sheet={sheet}"]
    lines.extend(pdf_export.vanishing_shapes_warning(vanishing))
    if not ok:
        lines.append(f"× {why}")
        for ln in lines:
            print(ln)
        return EXIT_ENVIRONMENT
    lines.append(f"出力先: {out_path}")

    check = pdf_export.verify_values_in_pdf(out_path, values)
    if not check.available:
        lines.append("⚠ PDF は作りましたが、機械保証はありません"
                      "（テキスト層の読み戻しに pdfplumber が要ります: pip install pdfplumber）")
        for ln in lines:
            print(ln)
        return 0
    if check.missing:
        lines.append(f"× シートの {check.checked} 個の値のうち {len(check.missing)} 個が"
                      f"PDF の中に見つかりません（読み戻しで確認）")
        for v in check.missing[:10]:
            lines.append(f"  ⚠ 『{v}』が PDF に見当たりません")
        if not a.fit_to_width:
            lines.append("  → 列幅で文字が切れている可能性があります。"
                          "`--fit-to-width` を付けて出し直してください")
        if len(check.missing) > 10:
            lines.append(f"  … 他 {len(check.missing) - 10} 個")
        lines.append("  （数値の書式や列幅で表示が変わっている可能性があります）")
        for ln in lines:
            print(ln)
        return 3
    # ★ 決裁③（✓ の絶対性）: 疑わしい ⚠ が 1 件でも在れば ✓ を名乗らない。
    #   値は載っていても、消えた角印は「宣言どおり」ではない。
    if vanishing:
        lines.append(f"△ シートの {check.checked} 個の値が PDF に載っていることは"
                     "確認しました ── ただし上の図形は消えています")
        for ln in lines:
            print(ln)
        return 0
    lines.append(f"✓ シートの {check.checked} 個の値が PDF に載っていることを"
                  f"読み戻して確認しました（欠落 0）")
    for ln in lines:
        print(ln)
    return 0

def bundled_demo_dir() -> Path:
    """同梱サンプルの置き場所（パッケージの中）。

    ★ repo からでも install 後でも同じ場所を指す ── `__file__` の隣を見るので、
      site-packages に入っても壊れない（refs/helpers と同じ作法）。
    """
    return Path(__file__).resolve().parent / "demo"


def cmd_demo(a: argparse.Namespace) -> int:
    """`ailine demo`: 同梱サンプルを手元に出して、**次に打つ行**まで見せる。

    ★ なぜ在るか（盲検の査定・2026-08-24）: README の最初のコマンドが落ちていた
      （`demo/sample.xlsx` は repo に無く、実体は `src/ailine/demo/`）。
      しかも **install した人はどちらのパスも持っていない**ので、README を直すだけでは
      install 経路が救われない。査定者の言葉:
      「値段を止めているのは製品の能力ではなく、**能力に到達するまでの距離**だ」。
    ★ だから「置き場所を教える」ではなく「手元に出して次の一手を見せる」にした。
    """
    src_dir = bundled_demo_dir()
    if not src_dir.exists():
        print(f"× 同梱サンプルが見つかりません（{src_dir}）")
        return EXIT_ENVIRONMENT
    dest_dir = Path(a.out).resolve() if getattr(a, "out", None) else Path.cwd()
    wanted = ["sample.xlsx", "lookup.xlsx", "sales.xlsx"]
    available = [n for n in wanted if (src_dir / n).exists()]
    if not available:
        print(f"× 同梱サンプルが空です（{src_dir}）")
        return EXIT_ENVIRONMENT
    clashes = [n for n in available if (dest_dir / n).exists()]
    if clashes and not getattr(a, "overwrite", False):
        print(f"■ ailine demo  出力先={dest_dir}")
        print(f"× {chr(12289).join(clashes)} が既にあります。"
              f"別の場所に出すなら --out、上書きしてよければ --overwrite を付けてください")
        return 7
    copied = []
    try:
        dest_dir.mkdir(parents=True, exist_ok=True)
        for name in available:
            shutil.copy2(src_dir / name, dest_dir / name)
            copied.append(name)
    except OSError as e:
        print(f"× サンプルを置けませんでした: {e}")
        return EXIT_ENVIRONMENT
    print(f"■ ailine demo  出力先={dest_dir}")
    for name in copied:
        print(f"  置いた: {name}")
    print()
    # ★ 2026-08-24 の実測: サンプルを置いて「次にこれを打って」と勧めたら、
    #   install した人の環境では basrun.py が無くてその一手が落ちた。
    #   **置いただけでは距離は縮まらない。** 前提が欠けているなら、勧める前に言う。
    try:
        missing = [(name, hint) for name, ok, hint in doctor_checks(getattr(a, "model", None)
                                                                     or "qwen2.5-coder:7b")
                    if not ok]
    except Exception:
        missing = []
    if missing:
        print("先に足りないものがあります:")
        for name, hint in missing:
            print(f"  × {name}" + (f" ── {hint}" if hint else ""))
        print()
        print("揃ったら `ailine doctor` で全部 ○ になることを確かめてから、もう一度ここへ。")
        return 0
    print("次にこれを打ってみてください:")
    print(f'  ailine run {copied[0]} "売上から原価を引いた利益の列を作って"')
    print()
    print("★ 原本は自動でバックアップされ、`ailine undo` で戻せます。")
    print("★ うまく動かないときは `ailine doctor` が在否を名指しで教えます。")
    return 0


def cmd_stack(a: argparse.Namespace) -> int:
    """`ailine stack <folder> --out <path>`: M1書き ── 縦積み（UNION ALL）+ 出所列。
       DESIGN-20260821-multifile.md v2 §1(M1書き)・v2.1(単位L)。列挙・照合・合計行の識別は
       既存部品（ailine_core/multifile.py・total_row.py）を再利用し、この関数は
       積む行の決定・出所列つきの書き出し・関所・事後条件の配線だけを持つ（本体は
       ailine_core/stack.py）。★ workdir は tempfile に作り、最後に out へ移す。"""
    folder = Path(a.folder).resolve()
    out = Path(a.out).resolve()
    candidates, excluded = multifile.classify_folder_contents(folder)

    # ★ 自己参照除外（V6・architect 致命2 で拡張）: 入力フォルダ内の ailine 産の出力
    #   （out と同じパスに限らず、種類（stack/extract 等）も問わない）は二重計上を防ぐため
    #   入力から除外 + 開示。判定は marks 集合（is_own_output）── 印が違っても ailine 産なら除外。
    candidates, self_excluded = multifile_stack.split_own_outputs(candidates)
    denominator = len(candidates)

    base_path, base_wb = multifile.open_base_workbook(candidates)
    if base_path is None:
        result = {"denominator": denominator, "stacked_files": 0, "rows_written": 0,
                  "files": [], "skipped": [{"name": p.name, "reason": "旧形式(.xls)または読み込み失敗"}
                                            for p in candidates],
                  "sums": {}, "excluded_detail": [], "mismatches": [], "col_a_warnings": [],
                  "sheet_fallbacks": [], "self_excluded": self_excluded, "rebuilt_own_output": False,
                  "excluded": excluded, "file_written": False}
        if a.json:
            print(json.dumps(_stack_json(result), ensure_ascii=False))
        else:
            for ln in render_stack_report(str(folder), str(out), result):
                print(ln)
        return 0

    base_sheet = base_wb.sheetnames[0]
    ws = base_wb[base_sheet]
    scan_end = min(ws.max_row or 1, MAX_ROWS, STRUCT_HEADER_SCAN_ROWS)
    rows_stats = _row_char_stats(ws, 1, scan_end, 1, min(ws.max_column or 1, MAX_COLS))
    row, confident = detect_header_row({"rows": rows_stats})
    header_row = row if confident else 1
    base_headers = multifile.read_row_headers(ws, header_row)
    # ★ 2026-08-24: 見出し行の選び方で**列が丸ごと落ちる**なら、それを言う。
    #   結合見出しの請求書で数量・金額が消えても exit 0・⚠ なし・Σ 行も出ない、という
    #   実測（盲検の実データ耐性レビュー）への処置。判定そのものは変えていない。
    header_drop_warning = header_row_drops_columns(ws, header_row)
    # ★ 2026-08-24（第二波 M3）: 同名の列がある表は**積まずに断る**。
    #   実測: 見出し `品番/備考/金額/備考`（社内用 と 客先提出用）で客先提出用が
    #   社内用に化け、exit 0・Σ 一致 ✓ で通っていた。しかもやる側と見る側が同じ
    #   「名前→列」の辞書を別々に組むので**照合も同じように潰れて恒真**だった。
    #   ★ 値が黙って化けるより、断られる方がはるかに安い。
    dupe_headers = multifile.duplicate_header_names(base_headers)
    if dupe_headers:
        names = "・".join(f"『{h}』" for h in dupe_headers)
        print(f"■ ailine stack  folder={folder}  out={out}")
        print(f"× 基準ファイル {base_path.name} に同じ名前の列があります（{names}）。"
              f"名前で列を引くため、どちらの値かを取り違えます"
              f"── 見出しを別々の名前にしてから、もう一度お試しください")
        return 3
    # ★ operator 盲検7度目の直し（2026-08-21）: 合計行検出は基準の数値列すべてを見る
    #   （旧 value_col_name の『最初の数値列』1本だけを keyed 列にする形は廃止 ── 実務標準形
    #   （数量・単価つき請求書）で合計の数字が金額列にしか無く、全トリガが沈黙していた）。
    numeric_cols = multifile_stack.numeric_column_names(ws, header_row, base_headers)
    # ★ 2026-08-24（第二波 M6）: 検算できなかった列を**黙って落とさない**。
    #   `=B2*C2` の金額列は data_only では None になり数値列と見なされないので、
    #   実測では **Σ金額 の行がそもそも出なかった** ── 「Σ が出ない」＝「検算していない」
    #   なのに、出ないことが唯一の信号だった。**出ないものは読めない。**
    formula_only = multifile.formula_columns_without_cache(base_path, header_row, base_headers,
                                                          sheet_name=base_sheet)
    unverified_cols = multifile.unverified_numeric_columns(
        base_headers, numeric_cols, formula_columns=formula_only)
    base_wb.close()

    skipped, files_json, excluded_detail, mismatches, col_a_warnings = [], [], [], [], []
    sheet_fallbacks = []   # ★ P2 開示: 基準名のシートが無く1枚目へ落ちたファイル
    stacked_rows = []   # [(base 列順の値, 元ファイル名, 元行), ...]
    sums_source = {col: 0.0 for col in numeric_cols}
    all_findings = []   # ★ M2.5: 検分シートの所見（inspection.Finding・ファイルごとに ws 側で組立済み）
    file_sheet_map = []   # ★ M2.5: [(ファイル名, 使ったシート, 備考), ...]（fallback 開示込み）

    dropped_notes_total = {}   # ★ 2026-08-24: 値として運べない「中身」の集計
    for p in candidates:
        r = multifile_stack.evaluate_and_stack(p, base_headers, base_sheet, header_row, numeric_cols)
        all_findings.extend(r.findings)
        for _row, kind in getattr(r, "dropped_notes", ()) or ():
            dropped_notes_total[kind] = dropped_notes_total.get(kind, 0) + 1
        sheet_used = r.sheet_fallback[1] if r.sheet_fallback else base_sheet
        if r.sheet_fallback:
            sheet_fallbacks.append({"name": r.name, "wanted": r.sheet_fallback[0],
                                    "used": r.sheet_fallback[1]})
        if r.status == "積めなかった":
            skipped.append({"name": r.name, "reason": r.reason})
            file_sheet_map.append((r.name, sheet_used, f"取れなかった（{r.reason}）"))
            continue
        file_sheet_map.append((r.name, sheet_used, "並べ替えて照合" if r.reordered else ""))
        for values, formats, src_row in r.rows:
            stacked_rows.append((values, formats, r.name, src_row))
            for col in numeric_cols:
                v = values[base_headers.index(col)]
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    sums_source[col] += v
        files_json.append({"name": r.name, "rows_stacked": len(r.rows),
                           "total_rows_excluded": len(r.excluded), "reordered": r.reordered})
        if r.excluded:
            excluded_detail.append({"name": r.name, "rows": [
                {"row": e.row, "value": e.value, "reason": e.reason} for e in r.excluded]})
        if r.mismatches:
            mismatches.append({"name": r.name, "rows": [
                {"row": m.row, "excluded_value": m.excluded_value, "adopted_sum": m.adopted_sum}
                for m in r.mismatches]})
        if r.col_a_mismatch:
            col_a_warnings.append({"name": r.name, "col_a": r.col_a_mismatch[0],
                                   "used_range": r.col_a_mismatch[1]})

    stacked_files = len(files_json)
    contributing_files = sum(1 for f in files_json if f["rows_stacked"] > 0)
    suspicious_files = {f.file for f in all_findings if f.kind in inspection.WARN_KINDS}
    # ★ 第二の独立検出器（operator 盲検7度目 修正2・恒真切り）: 列解決に一切依存せず、
    #   積んだ行の全セル値を走査して合計語（合計/小計/総計/計）を持つ行を名指しする
    #   （除外はしない ── 検出器1が沈黙しても黙って倍額にはならない、が保証の中身。
    #   誤爆（摘要の『7月合計分』等）は ⚠ 1個の確認コストで受ける）。
    trip_rows = [(fname, src_row, values) for values, _fmts, fname, src_row in stacked_rows]
    total_word_warnings = [{"file": fname, "row": row_num, "word": word}
                           for fname, row_num, word in total_row.total_word_trip_findings(trip_rows)]
    prov_headers = multifile_stack.own_output_headers(base_headers)
    out_headers = base_headers + prov_headers
    collision_notice = None
    if tuple(prov_headers) != multifile_stack.PROVENANCE_HEADERS:
        collision_notice = f"列名の衝突: 出所列は {prov_headers[0]} / {prov_headers[1]} として追加"

    workdir = Path(tempfile.mkdtemp(prefix="ailine_stack_"))
    try:
        tmp_out = workdir / out.name
        wb_out = openpyxl.Workbook()
        wb_out.properties.creator = multifile_stack.CREATOR_MARK   # ★ jisaku-review#1: 書き手の印
        ws_out = wb_out.active
        ws_out.title = base_sheet
        ws_out.append(out_headers)
        prov_col_idx = len(base_headers) + 1   # ★ M2.5: 「元ファイル」列（出所列の1本目）
        for i, (values, formats, fname, src_row) in enumerate(stacked_rows, start=2):
            ws_out.append(list(values) + [fname, src_row])
            for c, fmt in enumerate(formats, start=1):   # ★ 実視の磨き: 元の number_format を運ぶ
                ws_out.cell(row=i, column=c).number_format = fmt
            if fname in suspicious_files:
                # ★ M2.5②: ⚠ 付きファイル由来のデータ行だけ淡色 + コメント（正常行は塗らない）。
                reason_lines = [inspection.describe(f) for f in all_findings
                                if f.file == fname and f.kind in inspection.WARN_KINDS]
                inspection.tint_row(ws_out, i, len(out_headers), prov_col_idx, reason_lines)
        inspection.bold_row(ws_out, 1, len(out_headers))   # ★ UX 磨き③: 見出し行
        inspection.autosize_columns(ws_out)   # ★ UX 磨き②: 列幅を内容から機械算出
        wb_out.save(tmp_out)

        # ★ 事後条件①②: 独立読み実装（xml_readback）で書いた直後の中身を検算する
        #   （openpyxl で書いて openpyxl で読み返すだけでは、同じ道具の同じ盲点を通る）。
        readback = xml_readback.read_grid(tmp_out)
        out_row_nums = xml_readback.data_row_numbers(readback, header_row=1)
        if len(out_row_nums) != len(stacked_rows):
            wb_out.close()
            return _stack_postcondition_fail("採用行数", len(stacked_rows), len(out_row_nums))
        # ★ jisaku-review#3/#6: 事後条件②も全数値列に（最初の不一致で止める）。
        sums_output = {}
        for col in numeric_cols:
            idx = base_headers.index(col) + 1
            total = 0.0
            for rr in out_row_nums:
                v = readback["grid"].get((rr, idx))
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    total += v
            sums_output[col] = total
            if abs(total - sums_source.get(col, 0.0)) > 1e-6:
                wb_out.close()
                return _stack_postcondition_fail(f"Σ{col}", sums_source.get(col, 0.0), total)

        # ★ jisaku-review4戦目 F1 major: 行数/Σ が合っても帰属（どの行がどのファイルの
        #   何行目か）が嘘かもしれない（review3#3 と同型・cmd_run_folder は verify_extract
        #   経由で既に持っていたが cmd_stack は非対称に持っていなかった）。verify.py の
        #   独立読み（openpyxl を経由しない別実装）で移す前の tmp_out を検算する。
        attribution_check = multifile_verify.verify_output(tmp_out, folder)
        attribution_mismatch = attribution_check.get("mismatch")
        if attribution_mismatch and attribution_mismatch.get("kind") == "attribution":
            wb_out.close()
            return _stack_attribution_fail(attribution_mismatch)

        # ★ M2.5①: 検分シート（出力2枚目）── 事後条件が通った直後の数字だけを並べる
        #   （✓ の絶対性の適用拡張・手書きの ✓ を作らない）。1枚目（データ）はもう独立読みで
        #   検算済みなので、ここで検分シートを足して再保存しても事後条件の対象はぶれない
        #   （xml_readback は1枚目のシート順しか見ない ── verify も同じ前提）。
        sums = {col: {"source": sums_source[col], "output": sums_output.get(col, sums_source[col])}
                for col in sums_source}
        inspection.build_sheet(
            wb_out, findings=all_findings,
            denominator_lines_=inspection.denominator_lines(
                denominator, stacked_files, contributing_files, "積んだ"),
            accounting_lines=inspection.row_accounting_lines(
                adopted=len(stacked_rows),
                excluded=sum(len(entry["rows"]) for entry in excluded_detail),
                not_taken_files=len(skipped)),
            sums=sums, file_sheet_map=file_sheet_map, out_dir=out.parent, source_dir=folder)
        wb_out.save(tmp_out)
        wb_out.close()

        # ★ 関所（writes=new_book）: 移す直前に判定。
        #   ★ architect 致命2: 「作り直してよい」は creator の完全一致（CREATOR_MARK）に限定。
        #   ailine 産だが印が違う出力（例: ailine extract）は「別のコマンドの出力」として
        #   名指しで止める（無警告の作り直しにしない）。
        rebuilt_own_output = False
        if out.exists():
            mark = multifile_stack.own_output_mark(out)
            if mark == multifile_stack.CREATOR_MARK:
                rebuilt_own_output = True
            elif mark is not None:
                if not getattr(a, "overwrite", False):
                    print(f"⚠ 出力先は ailine の別のコマンドの出力です: {out}")
                    print(f"（{out.name}: これは ailine の別のコマンドの出力です（作成: {mark}）。"
                         "承知の上なら --overwrite を付けて実行してください）")
                    return 7
            elif not getattr(a, "overwrite", False):
                # ★ jisaku-review#5: 無言で閉まらない ── 何が邪魔か（名指し）+ 次の手を言う。
                print(f"⚠ 出力先に人のファイルがあります: {out}")
                print(f"（{out.name} は ailine stack の前回出力ではありません。"
                     "承知の上で上書きするなら --overwrite を付けて実行してください）")
                return 7
        out.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(tmp_out, out)
    finally:
        shutil.rmtree(workdir, ignore_errors=True)

    result = {"denominator": denominator, "stacked_files": stacked_files,
              "rows_written": len(stacked_rows), "files": files_json, "skipped": skipped,
              "dropped_notes": dropped_notes_total,
              "sums": sums, "excluded_detail": excluded_detail, "mismatches": mismatches,
              "col_a_warnings": col_a_warnings, "header_drop_warning": header_drop_warning, "unverified_cols": unverified_cols, "sheet_fallbacks": sheet_fallbacks,
              "self_excluded": self_excluded, "total_word_warnings": total_word_warnings,
              "rebuilt_own_output": rebuilt_own_output, "collision_notice": collision_notice,
              "excluded": excluded}
    # ★ 2026-08-24（盲検の使い勝手レビュー）: 積めなかった冊が在るのに exit 0 だった。
    #   しかも Σ の「元」は積めた冊だけの和なので**必ず一致する（恒真）**。
    #   得意先 1 社 283,500 円が消えても、スクリプトからは「成功・Σ 一致」に見えていた。
    #   ★ 出力は作る（人が続きを決められるように）。だが **0 では終わらない**。
    left_behind = len(skipped) + int((excluded or {}).get("other_format", 0) or 0)
    stack_exit = 0 if left_behind == 0 else 5
    if a.json:
        print(json.dumps(_stack_json(result), ensure_ascii=False))
        return stack_exit
    for ln in render_stack_report(str(folder), str(out), result):
        print(ln)
    if left_behind:
        print(f"× 取り込めなかったものが {left_behind} 件あります"
              f"（出力は作りました ── 上の名指しを確認してください）")
    return stack_exit


def cmd_verify(a: argparse.Namespace) -> int:
    """`ailine verify <out.xlsx> <srcfolder>` または `ailine verify <out.xlsx> <元A> <元B>`:
       検算の単独再実行（信用の条件⑥）。stack/extract は出力ブック+元フォルダから、
       match（照合）は出力ブック+元2冊から、それぞれ独立に検算する（読みは
       ailine_core/xml_readback.py・openpyxl は経由しない）。本体（ailine_core/verify.py）が
       検算そのものを持ち、この関数は配線+分岐だけ。
       ★ M3 設計 v2「verify」節: sources の個数で分岐（1個=従来の元フォルダ形 / 2個=
       照合の元2冊形）。それ以外の個数は使い方の誤りとして名指しで止める。
       ★ M2（E13/致命3）: 検算の種類は出力ブックの印と焼いた条件から決まる。ailine の印が
       無い/条件が読めないブックは合格でも不合格でもなく exit 4（「検算できません」）──
       0 件照合で空虚な合格を名乗らない。"""
    out = Path(a.out).resolve()
    sources = a.sources
    # ★ 実弾検分（2026-08-21）: 存在しないパスを「印がありません」と誤診していた ──
    #   誤診は次の手を間違わせる（印の問題だと思って原本を疑い始める）。無いなら無いと言う。
    #   従来形（sources 1 個）の元はフォルダ・照合形（2 個）の元はファイル。
    if not out.is_file():
        print(f"× ファイルが見つかりません: {out}")
        return 4
    if len(sources) == 2:
        for s in sources:
            if not Path(s).is_file():
                print(f"× ファイルが見つかりません: {Path(s).resolve()}")
                return 4
    elif len(sources) == 1 and not Path(sources[0]).is_dir():
        print(f"× フォルダが見つかりません: {Path(sources[0]).resolve()}")
        return 4
    if len(sources) == 1:
        folder = Path(sources[0]).resolve()
        result = multifile_verify.verify_output(out, folder)
        if result.get("unmarked"):
            print(f"× ailine の印がありません。検算できません: {out}")
            return 4
        if result.get("unsupported"):
            print(f"× {result['unsupported']}")
            return 4
        for ln in render_verify_report(str(out), str(folder), result):
            print(ln)
        return 5 if result.get("mismatch") else 0
    if len(sources) == 2:
        book_a = Path(sources[0]).resolve()
        book_b = Path(sources[1]).resolve()
        result = multifile_verify.verify_match_output(out, book_a, book_b)
        if result.get("unmarked"):
            print(f"× ailine の印がありません。検算できません: {out}")
            return 4
        if result.get("unsupported"):
            print(f"× {result['unsupported']}")
            return 4
        for ln in render_verify_match_report(str(out), str(book_a), str(book_b), result):
            print(ln)
        return 0 if result.get("ok") else 5
    print("× verify の引数は「元フォルダ1個」または「元A 元B の2冊」のどちらかです"
          f"（{len(sources)} 個渡されました）。")
    return 1   # ★ f6_exit_codes.md: 2 は argparse 予約・ailine 自身は使わない（汎用失敗の1を使う）


def build_parser() -> argparse.ArgumentParser:
    ap = argparse.ArgumentParser(prog="ailine", description="自然言語 → LibreOffice Basic → 適用 → 検証")
    sub = ap.add_subparsers(dest="cmd", required=True)

    r = sub.add_parser("run", help="タスクを生成・適用・検証する")
    r.add_argument("book", help="対象の文書 (.xlsx) またはフォルダ")
    # ★ M3 P: task は nargs="+" で受ける（2冊照合 `ailine run A.xlsx B.xlsx "依頼"` の
    #   2冊目パス+依頼文を同じ位置引数列で拾うため）。1冊経路では従来どおり要素数1の
    #   リストになり、_cmd_run_body の冒頭で通常の文字列へ畳み戻す（下流は全部 str のまま）。
    r.add_argument("task", nargs="+", help="やりたいことを自然言語で（2冊照合は 2冊目のパスに続けて）")
    r.add_argument("--model", default=DEFAULT_MODEL, help=f"ollama モデル (既定 {DEFAULT_MODEL})")
    r.add_argument("--refs", default=None, help="参照ライブラリのディレクトリ (既定 ./refs)")
    r.add_argument("--helpers", default=None, help="検証済みヘルパのディレクトリ (既定 ./helpers)")
    r.add_argument("--repair", type=int, default=2, help="修復の最大回数 (既定 2)")
    r.add_argument("--temperature", type=float, default=0.2)
    r.add_argument("--dry", action="store_true", help="生成して見せるだけ（適用しない）")
    r.add_argument("--op", default=None,
                    help="操作を人が固定する（言い回しから当てる段を飛ばす・一覧: ailine ops）")
    r.add_argument("--inplace", action="store_true",
                   help="（廃止・後方互換のため受理のみ）既定で原本に直接適用するため不要。"
                        "旧 .out 挙動が欲しければ --copy")
    r.add_argument("--json", action="store_true", help="結果を JSON でも出す")
    r.add_argument("--timeout", type=float, default=DEFAULT_APPLY_TIMEOUT,
                   help=f"basrun apply のタイムアウト秒 (既定 {DEFAULT_APPLY_TIMEOUT:.0f}、"
                        "0 で無効化=旧挙動の無制限)")
    r.add_argument("--ask", action="store_true",
                   help="DSL 経路の確認行の後に y/n で対話する（既定は表示して続行）")
    r.add_argument("--keep-backups", dest="keep_backups", type=keep_backups_arg,
                    default=DEFAULT_KEEP_BACKUPS,
                   help=f"原本への反映前のバックアップを book ごとに何世代残すか (既定 {DEFAULT_KEEP_BACKUPS}、"
                        "負数で無制限)")
    r.add_argument("--values", action="store_true",
                   help="COMPUTE_COLUMN を式でなく値ベタ書きにする（既定は式・W3 Part3）")
    r.add_argument("--header-row", dest="header_row", type=int, default=None,
                   help="見出し行を明示指定（1起点。指定時は自動検出をスキップしてこの行を採用）")
    r.add_argument("--sheet", default=None,
                   help="対象シートをシート名で明示指定（省略時は依頼文中のシート名の言及 → "
                        "1枚目。★ 挙動変更#2: 従来は常に1枚目固定だった）")
    r.add_argument("--accept-loss", dest="accept_loss", action="store_true",
                   help="往復忠実度ゲートが検出した喪失を承知の上で原本への反映を続行する"
                        "（ailine undo で戻せる）")
    r.add_argument("--copy", action="store_true",
                   help="原本には触らず <book>.out に結果を作る（既定の原本直接適用をしない・"
                        "旧 --inplace 無指定と同じ挙動。往復忠実度ゲートも走らせない）")
    r.add_argument("--overwrite", action="store_true",
                   help="破壊の関所（既存データを持つ列への上書き確認）を承知の上で"
                        "続行する（ailine undo で戻せる）")
    r.add_argument("--allow-freeform", dest="allow_freeform", action="store_true",
                   help="（廃止・後方互換のため受理のみ）自由生成は提供していません。指定しても"
                        "語彙外の断り + 廃止告知が出るだけです（機械検証できない操作は行わない方針）")
    r.set_defaults(func=cmd_run)

    s = sub.add_parser("stop", help="起動した LibreOffice を落とす")
    s.set_defaults(func=cmd_stop)

    d = sub.add_parser("doctor", help="セットアップを診断する")
    d.add_argument("--model", default=DEFAULT_MODEL, help=f"確認するモデル (既定 {DEFAULT_MODEL})")
    d.set_defaults(func=cmd_doctor)

    o = sub.add_parser("ops", help="頼める操作の一覧を表示する（何ができるか）")
    o.add_argument("--json", action="store_true",
                    help="機械可読で出す（画面や道具が一覧を自前で持たないため）")
    o.set_defaults(func=cmd_ops)

    cv = sub.add_parser("csv", help="CSV を検疫して隣に xlsx を作る（0落ち等を守って壊さず開ける形に）")
    cv.add_argument("file", help="対象の .csv ファイル")
    cv.set_defaults(func=cmd_run_csv)

    ec = sub.add_parser("export-csv", help="xlsx のシートを CSV へ書き出す（検疫の逆方向・0落ちを作らない）")
    ec.add_argument("book", help="対象の .xlsx ファイル")
    ec.add_argument("--sheet", required=True, help="書き出すシート名")
    ec.add_argument("--encoding", default=None,
                    help="出力の文字コード（既定 utf-8・BOM付き。会計ソフト向けに cp932 も選べる）")
    ec.add_argument("--out", default=None, help="出力先の .csv（既定 同名 .csv）")
    ec.add_argument("--overwrite", action="store_true", help="出力先が既にあっても上書きする（関所 exit 7）")
    ec.set_defaults(func=cmd_export_csv)

    dm = sub.add_parser("demo", help="同梱のサンプルを手元に出す（最初の 1 回はこれ）")
    dm.add_argument("--out", default=None, help="出力先のフォルダ（既定 今いる場所）")
    dm.add_argument("--overwrite", action="store_true", help="同名のファイルがあっても上書きする")
    dm.set_defaults(func=cmd_demo)

    ep = sub.add_parser("export-pdf", help="xlsx を PDF へ書き出す（出した PDF を読み戻して確かめる）")
    ep.add_argument("book", help="対象の .xlsx ファイル")
    ep.add_argument("--sheet", default=None, help="書き出すシート名（既定 1枚目）")
    ep.add_argument("--out", default=None, help="出力先の .pdf（既定 同名 .pdf）")
    ep.add_argument("--orientation", default=None, choices=["portrait", "landscape"],
                     help="用紙の向き")
    ep.add_argument("--fit-to-width", action="store_true", help="横幅を1ページに収める")
    ep.add_argument("--overwrite", action="store_true",
                     help="出力先が既にあっても上書きする（関所 exit 7）")
    ep.set_defaults(func=cmd_export_pdf)

    sc = sub.add_parser("scan", help="フォルダ内の複数ブックを棚卸しする（書き込みゼロ）")
    sc.add_argument("folder", help="対象フォルダ（直下の .xlsx を処理・.xls/.csv は数えて名指しで断る・サブフォルダは見ない）")
    sc.add_argument("--json", action="store_true", help="結果を JSON で出す（stdout は JSON のみ）")
    sc.set_defaults(func=cmd_scan)

    st = sub.add_parser("stack", help="フォルダ内の複数ブックを縦積みする（新ブック + 出所列）")
    st.add_argument("folder", help="対象フォルダ（直下の .xlsx を処理・.xls/.csv は数えて名指しで断る・サブフォルダは見ない）")
    st.add_argument("--out", required=True, help="出力ブックのパス")
    st.add_argument("--overwrite", action="store_true",
                    help="出力先に人のファイルが既にある時の関所（exit 7）を承知の上で上書きする")
    st.add_argument("--json", action="store_true", help="結果を JSON で出す（stdout は JSON のみ）")
    st.set_defaults(func=cmd_stack)

    vf = sub.add_parser("verify", help="stack/extract/match の出力を検算だけ独立に再実行する（読むだけ）")
    vf.add_argument("out", help="ailine が作った出力ブック")
    # ★ M3 設計 v2「verify」節: 位置引数を nargs 化し「2個=従来形(出力+元フォルダ) /
    #   3個=照合形(出力+元A+元B)」で分岐する（既存 stack/extract の2引数形は不変）。
    vf.add_argument("sources", nargs="+",
                    help="元フォルダ（1個・stack/extract）または 元A 元B（2個・照合出力）")
    vf.set_defaults(func=cmd_verify)

    h = sub.add_parser("history", help="実行履歴を表示する")
    h.add_argument("--max", type=int, default=10, help="表示件数（既定 10、新しい順）")
    h.set_defaults(func=cmd_history)

    rs = sub.add_parser("restore", help="原本への反映前のバックアップから復元する（ailine undo と同じ）")
    rs.add_argument("book", help="対象の文書 (.xlsx)")
    rs.add_argument("--list", action="store_true", help="バックアップ一覧を表示するだけ（復元しない）")
    rs.set_defaults(func=cmd_restore)

    u = sub.add_parser("undo", help="原本への反映前のバックアップから復元する（あと何回戻せるかを表示）")
    u.add_argument("book", help="対象の文書 (.xlsx)")
    u.add_argument("--list", action="store_true", help="バックアップ一覧を表示するだけ（復元しない）")
    u.set_defaults(func=cmd_undo)

    rd = sub.add_parser("redo", help="直前の undo をやり直す（あと何回やり直せるかを表示）")
    rd.add_argument("book", help="対象の文書 (.xlsx)")
    rd.set_defaults(func=cmd_redo)

    v = sub.add_parser("vocab", help="用語集（税率等の取り決め値）を編集・表示する")
    vsub = v.add_subparsers(dest="vocab_cmd", required=True)
    va = vsub.add_parser("add", help="語を登録する（例: ailine vocab add 消費税 1.1）")
    va.add_argument("term", help="語（例: 消費税）")
    va.add_argument("value", help="値（倍率。例: 1.1）")
    vsub.add_parser("list", help="登録済みの語を一覧表示する")
    v.set_defaults(func=cmd_vocab)

    al = sub.add_parser("alias", help="別名（言い回し → 操作名）を編集・表示する")
    alsub = al.add_subparsers(dest="alias_cmd", required=True)
    ala = alsub.add_parser("add", help="言い回しを登録する（例: ailine alias add 大きい順にして SORT）")
    ala.add_argument("phrase", help="言い回し（例: 大きい順にして）")
    ala.add_argument("op", help="op 名（例: SORT。実在する操作名のみ・一覧: ailine ops）")
    alsub.add_parser("list", help="登録済みの別名を一覧表示する")
    alr = alsub.add_parser("remove", help="言い回しを削除する")
    alr.add_argument("phrase", help="削除する言い回し")
    alsub.add_parser("undo", help="直近の登録を取り消す")
    al.set_defaults(func=cmd_alias)
    return ap


def main(argv=None) -> int:
    # ★ 日本語出力の生存保証（2026-08-21・CI の長期赤で発覚した製品バグ）: 端末やパイプの
    #   符号化が日本語を持たない環境（英語圏 Windows の cp1252 等）では、print が
    #   UnicodeEncodeError で落ちて「黙って失敗するより悪い、途中で死ぬ」になる。
    #   符号化はそのまま・書けない文字だけ置換に倒す（クラッシュしない、が保証の中身）。
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(errors="replace")
        except Exception:
            pass
    a = build_parser().parse_args(argv)
    return a.func(a)


if __name__ == "__main__":
    raise SystemExit(main())


def _console_entry() -> int:
    """`ailine` コマンド（wheel の console_scripts）の入口。
       ★ main(argv=None) は sys.argv[1:] を読む既存の形をそのまま使う ── 引数の解釈を
       二重化しない。戻り値の終了コードは console_scripts が sys.exit に渡す。"""
    return main()
