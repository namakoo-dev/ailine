# ★ 変更耐性テスト（opt-in・bench 層専用。cmd_run/cmd_run_plan の既定パイプラインには入れない）
#   設計 B: APPEND_TOTAL の合計式が「後からデータ行を1本挿入しても正しく追従するか」を、
#   実際に LibreOffice で挿入 → calculateAll() で再計算 → 読み戻して測る。
#
#   ★ なぜ bench 層に置き、cmd_run の既定パイプラインに混ぜないか:
#   run_postcondition（⑥事後条件）は「openpyxl で out ファイルを直接開くだけ・LO 不要」を
#   不変条件にしている（軽量・決定論的であることの土台）。ここに LO の再起動を挟むと毎回の
#   ユーザー時間が伸び、その不変条件を壊す。挿入耐性は「今すぐ壊れていないか」ではなく
#   「明日も壊れないか」を測る別種の検証なので、ailine.py run とは別の opt-in スクリプトに
#   分離する（設計レビューで指摘された不変条件違反への対応）。
#
#   手順: ① openpyxl でテスト表を作る ② ailine.codegen_dsl（本番と同じ関数）で
#   APPEND_TOTAL の挿入耐性式を basrun_apply（本番と同じ適用経路）で書き込む
#   ③ 合計行の直前に1行挿入・値を書く・oDoc.calculateAll()（★ 呼ばないとキャッシュが
#   古いまま偽合格になる） ④ openpyxl(data_only) で合計セルの再計算後キャッシュ値が
#   「新しい行を含めた合計×factor」と一致するかを見る。
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
WORK = HERE / "resilience_check_work"
WORK.mkdir(exist_ok=True)
AILINE_DIR = HERE.parent
sys.path.insert(0, str(AILINE_DIR))
import ailine  # noqa: E402


def make_book(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in [["品目", "数量", "単価", "小計"], ["りんご", 3, 120, 360],
                ["みかん", 5, 80, 400], ["ぶどう", 2, 150, 300]]:
        ws.append(row)
    wb.save(path)


def main() -> int:
    book = WORK / "resilience.xlsx"
    make_book(book)
    workdir = WORK / ".ailine_resilience"
    workdir.mkdir(exist_ok=True)
    _catalog, helper_files = ailine.load_helpers(AILINE_DIR / "helpers")

    book_meta = {"sheets": ["Sheet"], "headers": {"Sheet": ["品目", "数量", "単価", "小計"]}}
    resolved = {"col": "小計", "label": "税込み合計", "factor": 1.1}

    # ① APPEND_TOTAL を codegen_dsl（本番と同じ関数）で適用する
    code1 = ailine.codegen_dsl("APPEND_TOTAL", resolved, book_meta, use_formula=True)
    ok, err, raw = ailine.basrun_apply(book, code1, workdir, helper_files)
    if not ok:
        print("× ①APPEND_TOTAL 適用に失敗:", err, raw[-500:])
        return 1

    wb_v = openpyxl.load_workbook(book, data_only=True)
    before_total = wb_v.active["D5"].value   # データ3行(2-4行目)+合計は5行目・対象列は小計(D)
    wb_v.close()
    expect_before = round((360 + 400 + 300) * 1.1, 6)
    print(f"① 適用直後の合計(D5): {before_total!r}（期待 {expect_before:g}）")

    # ② 合計行(0起点 row=4)の直前に1行挿入し、値を書いて calculateAll する。
    #   InsertRows は helpers/AiLineHelpers.bas（basrun_apply が同じ src に同梱するので
    #   Call できる）。calculateAll は明示的に呼ぶ（呼ばないとキャッシュが古いまま）。
    insert_code = (
        "Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\n"
        "    Dim oSheet As Object\n"
        "    oSheet = oDoc.Sheets.getByIndex(0)\n"
        "    Call InsertRows(oDoc, 4, 1)\n"          # 0起点 row=4 (旧合計行=Excel5行目)の前に挿入
        '    oSheet.getCellByPosition(0, 4).setString("メロン")\n'
        "    oSheet.getCellByPosition(1, 4).setValue(1)\n"
        "    oSheet.getCellByPosition(2, 4).setValue(1000)\n"
        "    oSheet.getCellByPosition(3, 4).setValue(1000)\n"
        "    oDoc.calculateAll()\n"
        "End Sub\n"
    )
    ok, err, raw = ailine.basrun_apply(book, insert_code, workdir, helper_files)
    if not ok:
        print("× ②行挿入+再計算に失敗:", err, raw[-500:])
        return 1

    wb_v2 = openpyxl.load_workbook(book, data_only=True)
    after_total = wb_v2.active["D6"].value   # 挿入で合計行が1行下がる（旧D5→D6）
    wb_v2.close()
    expected = round((360 + 400 + 300 + 1000) * 1.1, 6)
    is_ok = ailine._is_number(after_total) and abs(after_total - expected) < 1e-6
    print(f"② 行挿入+再計算後の合計(D6): {after_total!r}（期待 {expected:g}）"
          f" → {'✓ 挿入耐性あり' if is_ok else '× 挿入耐性なし'}")
    return 0 if is_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
