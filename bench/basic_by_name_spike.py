"""位置の解決を **Basic 側**でやれるか実機で測る（2026-08-27・Namakoo の問い）。

★ 問い: Python が実表を読んで行番号を決める代わりに、生成する Basic に
  「みかんの下」と**名前で**渡して、Basic が走査すれば済まないか。
  済むなら architect の見積もり（住所リゾルバ 180 行 + 一般化）が大きく縮む。

測ること:
  A. 名前で行を探して、その下に値つきの行を足せるか
  B. 名前で 2 行を見つけて moveRange で入れ替えられるか（式が追随するか）
  C. 名前で列を見つけて、その右に列を挿せるか
  D. 見つからない/複数ある時に、Basic が**それと分かる形で**止まれるか
"""
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402

HELPERS = '''
' 名前で行を探す。0 起点の行 index を返す。見つからなければ -1、複数なら -2。
Function FindRowByName(oSheet As Object, sName As String, nCol As Integer) As Integer
    Dim oCur As Object, lastRow As Long, r As Long, hit As Integer, n As Integer
    oCur = oSheet.createCursor()
    oCur.gotoEndOfUsedArea(False)
    lastRow = oCur.RangeAddress.EndRow
    hit = -1 : n = 0
    For r = 1 To lastRow
        If Trim(oSheet.getCellByPosition(nCol, r).getString()) = sName Then
            hit = r : n = n + 1
        End If
    Next r
    If n = 0 Then
        FindRowByName = -1
    ElseIf n > 1 Then
        FindRowByName = -2
    Else
        FindRowByName = hit
    End If
End Function

' 名前で列を探す（見出し行 0）。0 起点の列 index / -1 / -2。
Function FindColByName(oSheet As Object, sName As String) As Integer
    Dim oCur As Object, lastCol As Integer, c As Integer, hit As Integer, n As Integer
    oCur = oSheet.createCursor()
    oCur.gotoEndOfUsedArea(False)
    lastCol = oCur.RangeAddress.EndColumn
    hit = -1 : n = 0
    For c = 0 To lastCol
        If Trim(oSheet.getCellByPosition(c, 0).getString()) = sName Then
            hit = c : n = n + 1
        End If
    Next c
    If n = 0 Then
        FindColByName = -1
    ElseIf n > 1 Then
        FindColByName = -2
    Else
        FindColByName = hit
    End If
End Function
'''


def _book(p: Path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["商品", "数量", "単価", "金額"])
    for i, (n, q, u) in enumerate([("りんご", 2, 500), ("みかん", 3, 800),
                                    ("ぶどう", 4, 900)], start=2):
        ws.cell(i, 1, n); ws.cell(i, 2, q); ws.cell(i, 3, u)
        ws.cell(i, 4, f"=B{i}*C{i}")
    wb.save(p); return p


def _run(book: Path, body: str):
    code = ("Option VBASupport 1\nOption Explicit\n" + HELPERS
            + "\nSub Run(oDoc As Object)\n"
              "    Dim oSheet As Object\n    oSheet = oDoc.Sheets.getByIndex(0)\n"
            + body + "End Sub\n")
    with tempfile.TemporaryDirectory() as wd:
        return ailine.basrun_apply(book, code, Path(wd), timeout=120)


def _show(p: Path, tag: str):
    raw = openpyxl.load_workbook(p)["売上"]
    val = openpyxl.load_workbook(p, data_only=True)["売上"]
    print(f"  {tag}:")
    for r in range(1, raw.max_row + 1):
        print("   ", [(raw.cell(r, c).value, val.cell(r, c).value)
                       for c in range(1, raw.max_column + 1)])


def main() -> int:
    tmp = Path(tempfile.mkdtemp())

    pa = _book(tmp / "a.xlsx")
    ok, err, _ = _run(pa, '''    Dim n As Integer
    n = FindRowByName(oSheet, "みかん", 0)
    If n < 0 Then Exit Sub
    oSheet.Rows.insertByIndex(n + 1, 1)
    oSheet.getCellByPosition(0, n + 1).setString("梨")
    oSheet.getCellByPosition(1, n + 1).setValue(5)
''')
    print(f"A 名前で探して下に足す: ok={ok} {err or ''}"); _show(pa, "結果")

    pb = _book(tmp / "b.xlsx")
    ok, err, _ = _run(pb, """    Dim r1 As Integer, r2 As Integer, tmpRow As Integer
    Dim oCur As Object
    Dim oR As New com.sun.star.table.CellRangeAddress
    Dim oD As New com.sun.star.table.CellAddress
    r1 = FindRowByName(oSheet, "みかん", 0)
    r2 = FindRowByName(oSheet, "ぶどう", 0)
    If r1 < 0 Or r2 < 0 Then Exit Sub
    oCur = oSheet.createCursor()
    oCur.gotoEndOfUsedArea(False)
    tmpRow = oCur.RangeAddress.EndRow + 2
    oR.Sheet = 0 : oR.StartColumn = 0 : oR.EndColumn = 3
    oD.Sheet = 0 : oD.Column = 0
    ' 1) みかんの行を空き地へ退避
    oR.StartRow = r1 : oR.EndRow = r1
    oD.Row = tmpRow
    oSheet.moveRange(oD, oR)
    ' 2) ぶどうの行を みかんの跡地へ
    oR.StartRow = r2 : oR.EndRow = r2
    oD.Row = r1
    oSheet.moveRange(oD, oR)
    ' 3) 退避したみかんを ぶどうの跡地へ
    oR.StartRow = tmpRow : oR.EndRow = tmpRow
    oD.Row = r2
    oSheet.moveRange(oD, oR)
""")
    print("")
    print("B 名前で 2 行を入れ替え（空き地経由の moveRange）: ok=" + str(ok) + " " + str(err or ""))
    _show(pb, "結果")

    pc = _book(tmp / "c.xlsx")
    ok, err, _ = _run(pc, '''    Dim c As Integer
    c = FindColByName(oSheet, "数量")
    If c < 0 Then Exit Sub
    oSheet.Columns.insertByIndex(c + 1, 1)
    oSheet.getCellByPosition(c + 1, 0).setString("利益")
''')
    print(f"\nC 名前で列を探して右に挿す: ok={ok} {err or ''}"); _show(pc, "結果")

    pd = _book(tmp / "d.xlsx")
    ok, err, out = _run(pd, '''    Dim n As Integer
    n = FindRowByName(oSheet, "すいか", 0)
    If n = -1 Then
        oSheet.getCellByPosition(9, 0).setString("NOT_FOUND")
    ElseIf n = -2 Then
        oSheet.getCellByPosition(9, 0).setString("AMBIGUOUS")
    End If
''')
    mark = openpyxl.load_workbook(pd)["売上"].cell(1, 10).value
    print(f"\nD 見つからない時に Basic が知らせられるか: ok={ok} 印={mark!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
