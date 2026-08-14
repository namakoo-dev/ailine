# 式化スパイク: setFormula の主要パターン実測と =VLOOKUP #VALUE! の真因特定
# 目的: 計算系 op の既定を「値ベタ書き」から「式」に反転できるかの判定材料
# (事務職の信頼は式が上: 編集に追随・セルで根拠が見える・Excel の仕事の見た目)
#
# 測るもの (各パターン × 書き方の行列):
#   P1 行ごとの積:   =B2*C2 (行複製)
#   P2 範囲合計:     =SUM(D2:D4)
#   P3 シート跨ぎ参照 (VLOOKUP): 区切り(, vs ;) × シート参照(. vs !) の 4 変種
#   P4 ROUND-TRIP:   保存後に openpyxl で 式文字列 と data_only キャッシュ値 の両方が読めるか
#
# 手順: openpyxl でテストブック生成 → basrun apply で式書き込みマクロを実行
#      → openpyxl (通常+data_only) で照合 → RESULTS.md に行列を出す
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
WORK = HERE / "formula_spike_work"
WORK.mkdir(exist_ok=True)
AILINE = HERE.parent
BASRUN = Path("C:/Dev/basrun/basrun.py")


def make_book(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active  # "Sheet"
    for row in [["品目", "数量", "単価", "小計"], ["りんご", 3, 120, None],
                ["みかん", 5, 80, None], ["ぶどう", 2, 150, None]]:
        ws.append(row)
    ref = wb.create_sheet("単価表")
    for row in [["商品", "単価"], ["りんご", 120], ["みかん", 80], ["ぶどう", 150]]:
        ref.append(row)
    wb.save(path)


CASES = [
    ("P1_product", 'oCell = oSheet.getCellByPosition(3, {r})\n    oCell.setFormula("=B{r1}*C{r1}")', None),
    ("P2_sum", 'oCell = oSheet.getCellByPosition(4, 1)\n    oCell.setFormula("=SUM(D2:D4)")', None),
    ("P3_vlookup_comma_bang", 'oCell = oSheet.getCellByPosition(5, 1)\n    oCell.setFormula("=VLOOKUP(A2,単価表!A2:B4,2,0)")', None),
    ("P3_vlookup_semi_bang", 'oCell = oSheet.getCellByPosition(5, 2)\n    oCell.setFormula("=VLOOKUP(A3;単価表!A2:B4;2;0)")', None),
    ("P3_vlookup_comma_dot", 'oCell = oSheet.getCellByPosition(6, 1)\n    oCell.setFormula("=VLOOKUP(A2,単価表.A2:B4,2,0)")', None),
    ("P3_vlookup_semi_dot", 'oCell = oSheet.getCellByPosition(6, 2)\n    oCell.setFormula("=VLOOKUP(A3;単価表.A2:B4;2;0)")', None),
]


def build_macro() -> str:
    lines = ["Option VBASupport 1", "Option Explicit", "Sub Run(oDoc As Object)",
             "    Dim oSheet As Object, oCell As Object",
             "    oSheet = oDoc.Sheets.getByIndex(0)"]
    for name, tmpl, _ in CASES:
        if name == "P1_product":
            for r in (1, 2, 3):
                lines.append("    " + tmpl.format(r=r, r1=r + 1))
        else:
            lines.append("    " + tmpl)
    lines.append("End Sub")
    return "\n".join(lines) + "\n"


def main():
    book = WORK / "spike.xlsx"
    make_book(book)
    src = WORK / "src"
    src.mkdir(exist_ok=True)
    (src / "Gen.bas").write_text(build_macro(), encoding="utf-8")

    r = subprocess.run([sys.executable, str(BASRUN), "apply", str(book), str(src),
                        "FSpike", "Gen.Run"], capture_output=True, text=True,
                       encoding="utf-8", timeout=300)
    print("basrun exit:", r.returncode)
    if r.returncode != 0:
        print(r.stdout[-800:], r.stderr[-800:])

    wb_f = openpyxl.load_workbook(book)     # 式文字列
    wb_v = openpyxl.load_workbook(book, data_only=True)  # キャッシュ値
    ws_f, ws_v = wb_f.active, wb_v.active
    report = []
    checks = [
        ("P1 D2 (=B2*C2)", "D2", 360), ("P1 D3", "D3", 400), ("P1 D4", "D4", 300),
        ("P2 E2 (=SUM)", "E2", 1060),
        ("P3 F2 (comma+bang)", "F2", 120), ("P3 F3 (semi+bang)", "F3", 80),
        ("P3 G2 (comma+dot)", "G2", 120), ("P3 G3 (semi+dot)", "G3", 80),
    ]
    for label, addr, expect in checks:
        formula = ws_f[addr].value
        value = ws_v[addr].value
        ok = value == expect
        report.append(f"| {label} | `{formula}` | {value} | {'✓' if ok else '×'} (期待 {expect}) |")
        print(f"{'✓' if ok else '×'} {label}: 式={formula!r} 値={value!r} 期待={expect}")

    out = ["# 式化スパイク結果", "", "| ケース | 保存された式 | キャッシュ値 | 判定 |",
           "|---|---|---|---|"] + report
    (HERE / "formula_spike_RESULTS.md").write_text("\n".join(out) + "\n", encoding="utf-8")
    print("\n→ bench/formula_spike_RESULTS.md")


if __name__ == "__main__":
    main()
