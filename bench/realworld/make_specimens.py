"""実務帳票 battery の検体生成スクリプト（測定タスク用・決定論的）。

ailine.py / helpers / refs は一切変更しない。ここで作るのは bench/realworld/*.xlsx の
入力データだけ。再実行可能（既存ファイルを上書きする）。

検体:
  A title_rows.xlsx    見出しが1行目でなく3行目にある帳票（結合タイトル行あり）
  B large.xlsx          3,000行のデータ（MAX_ROWS=1000 の死角を実測するため）
  C formulas.xlsx        数式セル(D列=B*C, 7行目=SUM, F1=TODAY()の揮発性セル)を持つ帳票
  D merged_head.xlsx     2段見出し（結合セルで上期/下期、2行目に月名。7月を含む）
  E cf.xlsx              条件付き書式つき（★ W8b: 往復忠実度ベースライン用）
  F datavalidation.xlsx  入力規則(プルダウン)つき（★ W8b）
  G shapes.xlsx          図形(オートシェイプ)つき（★ W8b・openpyxl に高水準 API が無いため
                          zip 後処理で最小限の DrawingML を直接注入。実際に LibreOffice で
                          開けるかは bench/fidelity_baseline.py 実行時に正直に記録する）
  H rows1500.xlsx        1,500行のデータ（★ W8b・MAX_ROWS=1000 超の忠実度確認用）
"""
import shutil
import zipfile
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

HERE = Path(__file__).resolve().parent


def make_title_rows() -> Path:
    """A: 1行目=結合タイトル、2行目=作成日、3行目=見出し、4行目以降=データ10行。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    ws["A1"] = "◯◯株式会社 売上台帳"
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = "作成日: 2026/08/01"

    headers = ["商品", "金額", "在庫", "売上", "原価"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)

    products = ["りんご", "みかん", "ぶどう", "もも", "なし",
                "いちご", "メロン", "バナナ", "キウイ", "さくらんぼ"]
    # 決定論的な値（行番号から導出）
    for i, name in enumerate(products):
        row = 4 + i
        kingaku = 1000 + i * 137          # 金額
        zaiko = 50 - i * 3                # 在庫
        uriage = kingaku * (10 - i)       # 売上
        genka = int(kingaku * 0.6)        # 原価
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=kingaku)
        ws.cell(row=row, column=3, value=zaiko)
        ws.cell(row=row, column=4, value=uriage)
        ws.cell(row=row, column=5, value=genka)

    out = HERE / "title_rows.xlsx"
    wb.save(out)
    return out


def make_large() -> Path:
    """B: 1行目見出し（商品コード/数量/単価）、2〜3001行目データ3,000行。決定論的な値。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    ws["A1"] = "商品コード"
    ws["B1"] = "数量"
    ws["C1"] = "単価"

    N = 3000
    for i in range(1, N + 1):
        row = 1 + i
        code = f"P{i:05d}"
        suryo = (i % 50) + 1                    # 数量: 1..50 の周期
        tanka = 100 + (i * 37) % 9000            # 単価: 100..9099 の規則的な分布
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=suryo)
        ws.cell(row=row, column=3, value=tanka)

    out = HERE / "large.xlsx"
    wb.save(out)
    return out


def make_formulas() -> Path:
    """C: 1行目見出し（商品/数量/単価/金額/備考）、2〜6行目データ5行。
       D列は式 =B{row}*C{row}、7行目は =SUM(D2:D6)、F1 は =TODAY()（揮発性）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    headers = ["商品", "数量", "単価", "金額", "備考"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    products = ["ノート", "鉛筆", "消しゴム", "定規", "ハサミ"]
    for i, name in enumerate(products):
        row = 2 + i
        suryo = 10 + i * 5
        tanka = 100 + i * 50
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=suryo)
        ws.cell(row=row, column=3, value=tanka)
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
        ws.cell(row=row, column=5, value="")

    # 最下行に SUM 式
    ws.cell(row=7, column=1, value="合計")
    ws.cell(row=7, column=4, value="=SUM(D2:D6)")

    # 揮発性セル（TODAY()）
    ws["F1"] = "=TODAY()"

    out = HERE / "formulas.xlsx"
    wb.save(out)
    return out


def make_merged_head() -> Path:
    """D: 2段見出し。1行目は「上期」(B1:C1結合)「下期」(D1:E1結合)、2行目に月名
       （4月/5月/7月/8月 — 7月を必ず含める）。3行目以降にデータ数行。A列は商品名。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"

    ws["A1"] = "商品"
    ws["B1"] = "上期"
    ws.merge_cells("B1:C1")
    ws["D1"] = "下期"
    ws.merge_cells("D1:E1")

    months = ["4月", "5月", "7月", "8月"]
    for c, m in enumerate(months, start=2):
        ws.cell(row=2, column=c, value=m)
    ws.cell(row=2, column=1, value="")  # A2 は空（A列見出しは商品のみ1行目）

    products = ["りんご", "みかん", "ぶどう", "もも", "なし"]
    for i, name in enumerate(products):
        row = 3 + i
        ws.cell(row=row, column=1, value=name)
        for j, m in enumerate(months):
            col = 2 + j
            val = 100 + i * 17 + j * 23
            ws.cell(row=row, column=col, value=val)

    out = HERE / "merged_head.xlsx"
    wb.save(out)
    return out


def make_cf() -> Path:
    """E: 見出し(1行目)+データ10行。金額列(B)に「50,000円超は赤字」条件付き書式を
       つける（★ W8b: 往復忠実度ベースラインの正例）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    headers = ["商品", "金額"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    products = ["りんご", "みかん", "ぶどう", "もも", "なし",
                "いちご", "メロン", "バナナ", "キウイ", "さくらんぼ"]
    for i, name in enumerate(products):
        row = 2 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=1000 + i * 5137)   # 一部が50,000円を超える
    ws.conditional_formatting.add(
        "B2:B11", CellIsRule(operator="greaterThan", formula=["50000"],
                              stopIfTrue=False))
    out = HERE / "cf.xlsx"
    wb.save(out)
    return out


def make_datavalidation() -> Path:
    """F: 見出し(1行目)+データ10行。備考列(C)に入力規則(プルダウン: 済/未/保留)を
       つける（★ W8b: 往復忠実度ベースラインの正例）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    headers = ["商品", "金額", "備考"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    products = ["りんご", "みかん", "ぶどう", "もも", "なし",
                "いちご", "メロン", "バナナ", "キウイ", "さくらんぼ"]
    dv = DataValidation(type="list", formula1='"済,未,保留"', allow_blank=True)
    ws.add_data_validation(dv)
    for i, name in enumerate(products):
        row = 2 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=1000 + i * 137)
        dv.add(ws.cell(row=row, column=3))
    out = HERE / "datavalidation.xlsx"
    wb.save(out)
    return out


# --- G: 図形(オートシェイプ) ---------------------------------------------------
# ★ openpyxl には worksheet レベルの汎用図形(オートシェイプ)を追加する高水準 API が
#   無い（画像(add_image)・チャートはあるが、矩形/矢印等の DrawingML shape は未対応）。
#   実務帳票にはよくある要素なので、保存済み xlsx の zip を後処理し、最小限の
#   DrawingML(<xdr:sp>) を直接注入する。ここで作った zip 構造が本物として妥当かは
#   bench/fidelity_baseline.py 実行時に実際に LibreOffice で開けるかで検証する
#   （開けなければ「作れない」と正直に記録する・この関数自体は失敗しても例外を投げない）。

_DRAWING_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:from><xdr:col>4</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
    <xdr:to><xdr:col>6</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>4</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
    <xdr:sp macro="" textlink="">
      <xdr:nvSpPr>
        <xdr:cNvPr id="2" name="Rectangle 1"/>
        <xdr:cNvSpPr/>
      </xdr:nvSpPr>
      <xdr:spPr>
        <a:xfrm><a:off x="0" y="0"/><a:ext cx="1000000" cy="500000"/></a:xfrm>
        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
        <a:solidFill><a:srgbClr val="4472C4"/></a:solidFill>
      </xdr:spPr>
      <xdr:txBody>
        <a:bodyPr/>
        <a:lstStyle/>
        <a:p><a:r><a:t>Shape</a:t></a:r></a:p>
      </xdr:txBody>
    </xdr:sp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>
"""

_DRAWING_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>
"""

_SHEET_RELS_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>
</Relationships>
"""


def _inject_autoshape(src: Path, dst: Path) -> None:
    """src(openpyxl 保存済み・単一シート) の zip を読み、図形1個を追加した dst を書く。"""
    with zipfile.ZipFile(src) as zin:
        names = zin.namelist()
        items = {n: zin.read(n) for n in names}

    sheet_name = "xl/worksheets/sheet1.xml"
    sheet_xml = items[sheet_name].decode("utf-8")
    if "</worksheet>" not in sheet_xml:
        raise ValueError("sheet1.xml に </worksheet> が無い（openpyxl の出力形式が変わった？）")
    # ★ openpyxl が書く <worksheet> ルートは既定で xmlns:r（relationships 名前空間）を
    #   宣言していない。<drawing r:id="..."/> は "r" prefix を使うため、無いと
    #   unbound prefix で壊れた xml になる。無ければ自分で足す。
    if "xmlns:r=" not in sheet_xml.split(">", 1)[0]:
        sheet_xml = sheet_xml.replace(
            "<worksheet ",
            '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
            1)
    sheet_xml = sheet_xml.replace("</worksheet>", '<drawing r:id="rId1"/></worksheet>')
    items[sheet_name] = sheet_xml.encode("utf-8")

    items["xl/drawings/drawing1.xml"] = _DRAWING_XML.encode("utf-8")
    items["xl/drawings/_rels/drawing1.xml.rels"] = _DRAWING_RELS_XML.encode("utf-8")
    items["xl/worksheets/_rels/sheet1.xml.rels"] = _SHEET_RELS_XML.encode("utf-8")

    ct_name = "[Content_Types].xml"
    ct_xml = items[ct_name].decode("utf-8")
    override = ('<Override PartName="/xl/drawings/drawing1.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
    if override not in ct_xml:
        ct_xml = ct_xml.replace("</Types>", override + "</Types>")
    items[ct_name] = ct_xml.encode("utf-8")

    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)


def make_shapes() -> Path:
    """G: 見出し(1行目)+データ5行 に加え、矩形のオートシェイプを1個持つ。
       ★ openpyxl 単体では作れないため zip 後処理（_inject_autoshape）で足す。
       失敗したら例外を投げず None を返す（呼び出し側が「作れなかった」と正直に記録する）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    headers = ["商品", "金額"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    products = ["りんご", "みかん", "ぶどう", "もも", "なし"]
    for i, name in enumerate(products):
        row = 2 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=1000 + i * 137)

    tmp = HERE / "_shapes_base.xlsx"
    out = HERE / "shapes.xlsx"
    wb.save(tmp)
    try:
        _inject_autoshape(tmp, out)
    except Exception as e:
        print(f"× shapes.xlsx は作れなかった（zip 後処理に失敗）: {e}")
        tmp.unlink(missing_ok=True)
        return None
    tmp.unlink(missing_ok=True)
    return out


def make_rows1500() -> Path:
    """H: 1行目見出し（商品コード/数量/単価）、2〜1501行目データ1,500行。
       ★ W8b: MAX_ROWS=1000 超での忠実度チェックの実測用（B の large.xlsx=3000行より
       小さく、往復忠実度に絞った軽い検体として別立てにする）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "商品コード"
    ws["B1"] = "数量"
    ws["C1"] = "単価"
    N = 1500
    for i in range(1, N + 1):
        row = 1 + i
        ws.cell(row=row, column=1, value=f"P{i:05d}")
        ws.cell(row=row, column=2, value=(i % 50) + 1)
        ws.cell(row=row, column=3, value=100 + (i * 37) % 9000)
    out = HERE / "rows1500.xlsx"
    wb.save(out)
    return out


def main():
    paths = [make_title_rows(), make_large(), make_formulas(), make_merged_head(),
              make_cf(), make_datavalidation(), make_shapes(), make_rows1500()]
    for p in paths:
        if p is None:
            print("生成: （shapes.xlsx は作れなかった・上のメッセージ参照）")
        else:
            print(f"生成: {p}")


if __name__ == "__main__":
    main()
