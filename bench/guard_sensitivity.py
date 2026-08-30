#!/usr/bin/env python
"""番人の感度を測る ── 生成コードをわざと壊し、事後条件が何件捕まえるかを数える。

★★ 2026-08-30（Namakoo）:「保証の網をすり抜けるケースがないともいいきれない」
  そのとおりで、**完全性は証明できない**（穴が無いことは示せない）。
  ★ だが「穴があるかもしれません」で終わらせずに済む ── **感度は測れる**。
    わざと N 通り壊して、番人が何件鳴ったかを数える。証明ではなく測定値。

やり方（★ 本体は 1 行も触らない）:
  ① 既知の正しい (op, 引数) から Basic を生成し、適用して事後条件を走らせる
     → **陽性対照**: ここが pass しなければ検体か治具が壊れている（測定器を先に疑う）
  ② 同じ Basic を変異させて適用し、同じ事後条件を走らせる
  ③ 変異後のファイルが正解と**同じ**なら、その変異は効いていない（不活性）
     → 分母から外し、**外した数を必ず表示する**（黙って切り詰めない）
  ④ 残りについて、fail/error を返したら「捕まえた」、pass を返したら「見逃した」

★ LLM は使わない ── 翻訳の精度ではなく**番人の感度**だけを測る。
"""

from __future__ import annotations

import re
import shutil
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import ailine  # noqa: E402


# --- 検体（表と、既知の正しい操作）-------------------------------------------------------

def _sales(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "売上", "原価"])
    for r in (["りんご", 1200, 700], ["みかん", 800, 300], ["ぶどう", 1500, 900]):
        ws.append(r)
    wb.save(path)


HEADERS = ["商品", "売上", "原価"]
META = {"sheets": ["売上"], "headers": {"売上": HEADERS}, "header_rows": {"売上": 1}}

# (名前, op, 引数) ── 引数は verify_dsl_args を通した形（_target_sheet 等を含む）
CASES = [
    ("1セル書換", "SET_CELL_VALUE",
     {"row": "みかん", "col": "売上", "value": "2000", "_row_index": 3,
      "_headers": HEADERS, "_header_row": 1, "_write_numeric": True,
      "_write_numeric_value": 2000.0}),
    ("行追加", "ADD_ROW",
     {"at": 3, "values": {"商品": "梨"}, "_headers": HEADERS, "_header_row": 1}),
    ("行削除", "DELETE_ROWS",
     {"at": 3, "count": 1, "_headers": HEADERS, "_header_row": 1}),
    ("列削除", "DELETE_COLUMN",
     {"col": "原価", "_headers": HEADERS, "_header_row": 1}),
    ("並べ替え", "SORT",
     {"col": "売上", "order": "desc", "_headers": HEADERS, "_header_row": 1}),
    ("入れ替え(列)", "SWAP",
     {"a": "売上", "b": "原価", "_axis": "column", "_a_pos": 2, "_b_pos": 3,
      "_headers": HEADERS, "_header_row": 1}),
    ("計算列", "COMPUTE_COLUMN",
     {"operands": ["売上", "原価"], "operator": "-", "_headers": HEADERS,
      "_header_row": 1, "_new_col_label": "利益"}),
]


# --- 変異（生成した Basic を壊す）---------------------------------------------------------
#
# ★ どれも **Basic として正しいまま**にする（構文エラーは番人の手前で落ちるので
#   感度を測ったことにならない）。

_POS = re.compile(r"getCellByPosition\((\s*[0-9]+\s*),(\s*[^,)]+?\s*)\)")


def m_nothing(code: str) -> str:
    """何もしない ── いちばん基本の恒真殺し。番人が鳴らなければ、その番人は無意味。"""
    return re.sub(r"(Sub Run\(oDoc As Object\)\n).*?(\nEnd Sub)",
                   r"\1    Dim dummy As Integer\2", code, flags=re.S)


def m_shift_col(code: str) -> str:
    """1 列ずれて書く。"""
    done = [False]

    def sub(m):
        if done[0]:
            return m.group(0)
        done[0] = True
        return f"getCellByPosition({int(m.group(1)) + 1},{m.group(2)})"
    return _POS.sub(sub, code)


def m_shift_row(code: str) -> str:
    """1 行ずれて書く（行が数字で書かれている呼び出しだけ）。"""
    def sub(m):
        try:
            return f"getCellByPosition({m.group(1)},{int(m.group(2)) + 1})"
        except ValueError:
            return m.group(0)
    return _POS.sub(sub, code, count=1)


def m_extra_write(code: str) -> str:
    """頼んでいないセルに 1 つ書き足す（『他は 1 セルも変わらず』を試す）。"""
    extra = ('    oDoc.Sheets.getByIndex(0).getCellByPosition(0, 3)'
              '.setString("よけいな値")\n')
    return code.replace("End Sub", extra + "End Sub", 1)


def m_drop_a_call(code: str) -> str:
    """ヘルパ呼び出しを 1 本落とす（書いたつもりで書いていない形）。"""
    lines = code.split("\n")
    for i, ln in enumerate(lines):
        if ln.strip().startswith("Call "):
            return "\n".join(lines[:i] + lines[i + 1:])
    return code


def m_off_by_one_arg(code: str) -> str:
    """ヘルパ呼び出しの**最後の数値引数**を 1 ずらす（位置や本数の取り違え）。"""
    def sub(m):
        return f"{m.group(1)}{int(m.group(2)) + 1}{m.group(3)}"
    return re.sub(r"(Call \w+\([^)]*?)(\d+)(\s*\)\s*$)", sub, code,
                   count=1, flags=re.M)


# ★★ 2026-08-30（初回の測定を自分で疑って追加）: 最初の変異は getCellByPosition を
#   前提にしていたので、**ヘルパ Call 形式の op にまったく当たっていなかった**
#   （不活性 15 件・当たらない変異で 100% を出しても、測っていないのと同じ）。
#   ★ Call の引数そのものを壊す変異を足す ── ここが実際の書き間違いの形。

_CALL = re.compile(r"Call (\w+)\(([^)]*)\)")


def _edit_call_args(code: str, fn):
    """最初の Call の引数列を fn で書き換える（変えられなければ元のまま）。"""
    def sub(m):
        parts = [p.strip() for p in m.group(2).split(",")]
        out = fn(parts)
        return m.group(0) if out is None else f"Call {m.group(1)}({', '.join(out)})"
    return _CALL.sub(sub, code, count=1)


def m_call_first_number(code: str) -> str:
    """Call の**最初の数値引数**を 1 ずらす（位置の取り違え）。"""
    def fn(parts):
        for i, p in enumerate(parts):
            if p.lstrip("-").isdigit():
                parts[i] = str(int(p) + 1)
                return parts
        return None
    return _edit_call_args(code, fn)


def m_call_swap_numbers(code: str) -> str:
    """Call の数値引数を 2 つ入れ替える（行と列を取り違える形）。"""
    def fn(parts):
        idx = [i for i, p in enumerate(parts) if p.lstrip("-").isdigit()]
        if len(idx) < 2 or parts[idx[0]] == parts[idx[1]]:
            return None
        parts[idx[0]], parts[idx[1]] = parts[idx[1]], parts[idx[0]]
        return parts
    return _edit_call_args(code, fn)


def m_call_wrong_name(code: str) -> str:
    """Call の文字列引数を、**実在する別の列名**に差し替える（狙う先を間違える形）。"""
    def fn(parts):
        for i, p in enumerate(parts):
            if p.startswith(chr(34)) and p.endswith(chr(34)) and len(p) > 2:
                cur = p[1:-1]
                for h in HEADERS:
                    if h != cur:
                        parts[i] = chr(34) + h + chr(34)
                        return parts
                return None
        return None
    return _edit_call_args(code, fn)


def m_call_flip_bool(code: str) -> str:
    """Call の True/False を反転する（昇順と降順の取り違え）。"""
    def fn(parts):
        for i, p in enumerate(parts):
            if p in ("True", "False"):
                parts[i] = "False" if p == "True" else "True"
                return parts
        return None
    return _edit_call_args(code, fn)


MUTATIONS = [
    ("何もしない", m_nothing),
    ("Callの位置を1ずらす", m_call_first_number),
    ("Callの数値2つを入れ替える", m_call_swap_numbers),
    ("Callの対象名を別の列に", m_call_wrong_name),
    ("Callの真偽を反転", m_call_flip_bool),
    ("1列ずらす", m_shift_col),
    ("1行ずらす", m_shift_row),
    ("余計に書く", m_extra_write),
    ("呼び出しを1本落とす", m_drop_a_call),
    ("引数を1ずらす", m_off_by_one_arg),
]


# --- 実行 ---------------------------------------------------------------------------------

def _cells(path: Path) -> list:
    wb = openpyxl.load_workbook(path)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        out.append((name, tuple(tuple(r) for r in ws.iter_rows(values_only=True))))
    return out


def _apply(src: Path, code: str, workdir: Path, tag: str):
    """src のコピーに code を適用して、その結果のパスを返す（失敗なら None）。"""
    out = workdir / f"{tag}.xlsx"
    shutil.copy2(src, out)
    helpers = ROOT / "src" / "ailine" / "helpers"
    ok, err, _raw = ailine.basrun_apply(
        out, code, workdir, helper_files=tuple(sorted(helpers.glob("*.bas"))),
        timeout=90)
    return out if ok else None


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="ailine_guard_") as tmp:
        work = Path(tmp)
        src = work / "src.xlsx"
        _sales(src)

        caught = missed = inert = na = 0
        broken_control = []
        misses = []

        for name, op, args in CASES:
            a = dict(args)
            a["_target_sheet"] = "売上"
            code = ailine.codegen_dsl(op, a, META)

            # ① 陽性対照 ── 正しいコードは pass するはず
            base = _apply(src, code, work, f"{op}_base")
            if base is None:
                broken_control.append(f"{name}: 適用に失敗")
                continue
            status, why = ailine.run_postcondition(op, base, a, header_row=1,
                                                    use_formula=True, source_book=src)
            if status not in ("pass", "warn"):
                broken_control.append(f"{name}: 正しいコードが {status}（{why[:60]}）")
                continue
            good = _cells(base)

            for mname, fn in MUTATIONS:
                code2 = fn(code)
                if code2 == code:
                    # ★ そもそも当たらなかった変異（この op の書き方に該当箇所が無い）
                    na += 1
                    continue
                out2 = _apply(src, code2, work, f"{op}_{mname}")
                if out2 is None:
                    # 適用そのものが落ちた＝壊れたことは分かる（捕まえた側に数える）
                    caught += 1
                    print(f"  ○ {name} / {mname}  ── 適用が失敗（壊れたと分かる）")
                    continue
                if _cells(out2) == good:
                    inert += 1
                    print(f"  ─ {name} / {mname}  ── 結果が変わらない（不活性・分母から除外）")
                    continue
                st, w = ailine.run_postcondition(op, out2, a, header_row=1,
                                                  use_formula=True, source_book=src)
                if st in ("fail", "error"):
                    caught += 1
                    print(f"  ○ {name} / {mname}  ── 捕まえた: {w[:70]}")
                else:
                    missed += 1
                    misses.append(f"{name} / {mname}（{st}: {w[:60]}）")
                    print(f"  ✗ {name} / {mname}  ── **見逃した**（{st}）")

        total = caught + missed
        print()
        if broken_control:
            print("★ 陽性対照が通らなかった検体（治具側の問題・分母に入れていない）:")
            for b in broken_control:
                print(f"    - {b}")
        print(f"効いた変異 {total} 件: 捕まえた {caught}／見逃した {missed}")
        print(f"  分母から外した内訳: 当たらなかった {na} 件"
               f"（その op の書き方に該当箇所が無い）／"
               f"当てたが結果が変わらなかった {inert} 件")
        if total:
            print(f"  番人の感度 {caught / total * 100:.1f}%")
        if misses:
            print("★ 見逃した変異（ここが網の穴）:")
            for m in misses:
                print(f"    - {m}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
