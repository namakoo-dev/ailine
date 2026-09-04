"""基本操作 × 言い回しの揺れ × 表の形 ── **効果**で測る検体。

★★ なぜ要るか（2026-08-29・Namakoo「2 日前から同様の事例を繰り返している」）:
  この 2 日、壊れは全部**人が手で触って**見つかった。俺は分母を持たずに 1 件ずつ
  直していた ── 直った数は分かるが、**残りがどれだけあるか**を誰も知らない。
  ★ 既に在る凍結検体（translation_battery）は **op 名の一致**しか見ない。
    op が合っていても「担当列が全行書き換わる」事故は起きるし、op が違っても
    結果が正しいことはある（読み直しの層が拾う）。**測るべきは効果**。

測り方: 表を作る → 依頼文を投げる → **出来上がったファイルを読んで**、
        意図した効果になっているかだけを見る（op 名も内部の経路も見ない）。

  合格 = 期待した変化がちょうど起きている（かつ他が壊れていない）
  断り = ？ で止まった（**壊していない**ので、間違いより軽い ── 別枠で数える）
  失敗 = 何か違うことをした（一番重い）

使い方:
    python bench/basic_ops_matrix.py                 # 全部
    python bench/basic_ops_matrix.py --table 在庫    # 表を絞る
    python bench/basic_ops_matrix.py --op cell       # 操作を絞る
    python bench/basic_ops_matrix.py --out result.json

★ 実物の ollama + LibreOffice が要る（遅い）。CI では走らせない。
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from pathlib import Path

import inspect
import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(ROOT / "src"))

# --- 表（語彙も列名も形も違う 3 つ）-----------------------------------------------------

TABLES = {
    "在庫": {
        "sheet": "在庫",
        "headers": ["品名", "棚", "数量", "備考"],
        "rows": [["ボルト", "A-1", 120, ""], ["ナット", "A-2", 80, ""],
                  ["ワッシャー", "B-1", 300, ""]],
    },
    "名簿": {
        "sheet": "名簿",
        "headers": ["氏名", "所属", "内線", "メモ"],   # ★ 4 列目はわざと見慣れない語
        # ★ 2026-08-29 の 1 回目で分かった検体の欠陥: 内線を 101/202/303 にすると
        #   303 = 101+202 になり、**合計行の番人が最下行を合計と誤検出**して並べ替えが
        #   × になった（番人は正しく疑っている ── 悪いのは検体）。
        #   ★ 検体の数字が偶然の関係を持たないようにする。
        "rows": [["山田", "営業", 101, ""], ["鈴木", "経理", 202, ""],
                  ["高橋", "総務", 305, ""]],
    },
    # ★★ 2026-09-02: **計算列（COMPUTE_COLUMN）を 1 件も測れていなかった**。
    #   理由は表の側にあった ── 既存の 3 表はどれも数値の列が 1 本しかない
    #   （在庫=数量／名簿=内線／献立=分量）ので、2 列の演算を頼めなかった。
    #   ★ 実演の幕 1 で使う op なのに、効果の分母に入っていなかった。表を 1 つ足す。
    "見積": {
        "sheet": "見積",
        "headers": ["品名", "数量", "単価", "メモ"],
        # ★ 偶然の等式を作らない（合計行の番人が最下行を誤検出しないため）。
        "rows": [["机", 3, 12000, ""], ["椅子", 8, 4500, ""], ["棚", 2, 30000, ""]],
    },
    # ★★ 2026-09-02: **入れ替え（SWAP）を 1 件も測れていなかった** ── 8/31 に見つけた
    #   13 件の欠陥のうち **5 件が入れ替え**だったのに、効果の分母に乗っていなかった。
    #   ★ 入れ替えの本題は「並びが変わったか」ではなく「**式が自分の行を指し続けるか**」。
    #     実測（bench/swap_formula_spike_RESULTS.md）で、値を交換する実装だと各行の
    #     金額が他の行の値になり、**並びは正しく見えるので人が気づけない**。
    #     それを測るには式のある表が要る ── 既存の表の期待値を動かさないよう別に立てる。
    "請求": {
        "sheet": "請求",
        "headers": ["取引先", "件数", "単価", "金額"],
        "rows": [["あかね商事", 3, 12000, "=B2*C2"],
                  ["いろは工業", 8, 4500, "=B3*C3"],
                  ["うえだ物産", 2, 30000, "=B4*C4"]],
    },
    "献立": {
        "sheet": "献立",
        "headers": ["料理", "主材料", "分量", "備考"],
        "rows": [["カレー", "牛肉", 4, ""], ["味噌汁", "豆腐", 2, ""],
                  ["サラダ", "レタス", 3, ""]],
    },
}


def _build(table_key: str, path: Path) -> Path:
    t = TABLES[table_key]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t["sheet"]
    ws.append(t["headers"])
    for r in t["rows"]:
        ws.append(list(r))
    wb.save(path)
    return path


# --- 期待する効果（op 名でなく、出来上がったファイルで見る）----------------------------

def _grid(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
             for r in range(1, ws.max_row + 1)]


def cell_becomes(row: int, col: int, value):
    """その 1 セルだけが value になり、**他は 1 セルも変わっていない**こと。"""
    def check(before, after):
        if len(before) != len(after):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        diff = [(r, c) for r in range(len(after)) for c in range(len(after[r]))
                 if before[r][c] != after[r][c]]
        if diff != [(row - 1, col - 1)]:
            return False, f"変わったセルが {diff}（{row}行{col}列だけのはず）"
        got = after[row - 1][col - 1]
        if str(got) != str(value):
            return False, f"値が {got!r}（{value!r} のはず）"
        return True, ""
    return check


def row_added_at(at: int, first_value):
    """at 行目に 1 行増え、1 列目がその値。**他の行は順序ごとそのまま**。"""
    def check(before, after):
        if len(after) != len(before) + 1:
            return False, f"行数が {len(before)}→{len(after)}（1 増えるはず）"
        got = after[at - 1][0]
        if str(got) != str(first_value):
            return False, f"{at}行目の 1 列目が {got!r}（{first_value!r} のはず）"
        rest = after[:at - 1] + after[at:]
        if rest != before:
            return False, "他の行が変わっている"
        return True, ""
    return check


def row_deleted(name):
    """その名前の行が消え、**他は順序ごとそのまま**。"""
    def check(before, after):
        want = [r for r in before if str(r[0]) != str(name)]
        if len(after) != len(before) - 1:
            return False, f"行数が {len(before)}→{len(after)}（1 減るはず）"
        if after != want:
            return False, "消えた行が違う（または他の行が動いた）"
        return True, ""
    return check


def column_added_named(name, at_index=None):
    """見出しに name の列が増える（位置指定があればその位置）。値は増やさない。"""
    def check(before, after):
        if len(after) != len(before):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        heads = [str(h or "") for h in after[0]]
        if name not in heads:
            return False, f"見出しに『{name}』が無い（{heads}）"
        if at_index is not None and heads.index(name) + 1 != at_index:
            return False, f"『{name}』が {heads.index(name) + 1} 列目（{at_index} 列目のはず）"
        return True, ""
    return check


def computed_column(name, at_index=None):
    """計算列が、**依頼文の名前**で、指定の位置に、**式**として入ること。

    ★★ 2026-09-02: ここまで効果で見る ── 見出しだけ合っていても、値が直値なら
      翌月ずれる（この repo が何度も踏んだ「静かに壊れる」形）。
    ★ 名前は依頼文から取る（A' 原則）。名前を言っていない回は数式風の見出しでよいので、
      その回は name=None を渡す（位置と式だけ見る）。
    """
    def check(before, after):
        if len(after) != len(before):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        heads = [str(h or "") for h in after[0]]
        if len(heads) != len(before[0]) + 1:
            return False, f"列が 1 本増えていない（{heads}）"
        if name is not None:
            if name not in heads:
                return False, f"見出しに『{name}』が無い（{heads}）"
            idx = heads.index(name)
        else:
            idx = next((i for i, h in enumerate(heads)
                         if i >= len(before[0]) or h != str(before[0][i] or "")), -1)
            if idx < 0:
                return False, f"新しい列が見つからない（{heads}）"
        if at_index is not None and idx + 1 != at_index:
            return False, f"新しい列が {idx + 1} 列目（{at_index} 列目のはず）"
        vals = [after[r][idx] for r in range(1, len(after))]
        if not vals or not all(isinstance(v, str) and v.startswith("=") for v in vals):
            return False, f"式になっていない（{vals}）"
        return True, ""
    return check



class _Books:
    """検算に渡す「操作の前と後のブック」。★ 値の格子では見えないものを読むための口。

    ★ なぜ要るか（2026-09-04）: 太字・けい線・数値書式・セル結合・列幅は
      **値の格子に 1 文字も出ない**。だから「書式を付けたつもりが値を壊した」を
      効果の側から見られなかった（事後条件の台帳でも DRAW_BORDERS / NUMBER_FORMAT /
      MERGE が『値の不変を見ていない』として在庫に載っている）。
    """

    def __init__(self, src, out, sheet):
        self.src, self.out, self.sheet = src, out, sheet

    def cells(self, which):
        """(前 or 後) のシートを開いて、セルの二次元配列を返す。"""
        wb = openpyxl.load_workbook(self.src if which == 'before' else self.out)
        ws = wb[self.sheet]
        return ws, wb


def _fmt_signature(ws):
    """そのシートの**書式だけ**を拾う（値は見ない）。"""
    out = {}
    for r in range(1, (ws.max_row or 0) + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            cell = ws.cell(r, c)
            f = cell.font
            b = cell.border
            out[(r, c)] = (
                bool(f and f.bold),
                (cell.alignment.horizontal if cell.alignment else None),
                cell.number_format,
                (str(cell.fill.start_color.rgb) if cell.fill and
                 cell.fill.patternType else None),
                (b.left.style, b.right.style, b.top.style, b.bottom.style) if b else None,
            )
    return out


def format_applied(kind: str, at=None):
    """★ 書式 op の検算 ── 「値が 1 つも変わらず」かつ「書式が実際に付いた」。

    kind: bold / align / numfmt / fill / border / merge / width
    at:   (行, 列) を指定すると**そのセルに**付いたことまで見る（None なら『どこかに』）

    ★ 2 つを対で見るのが肝 ── 値の不変だけだと「何もしなかった」が通り、
      書式が付いたことだけだと「ついでに値を壊した」が通る。
    """
    idx = {'bold': 0, 'align': 1, 'numfmt': 2, 'fill': 3, 'border': 4}

    def check(before, after, books):
        # ★ セル結合だけは値が変わる ── Excel の仕様で**左上以外の値は破棄**される
        #   （2026-09-04 に実測: A1:B1 を結合すると B1 の『棚』が消える）。
        #   ★ 製品はこれを検知して ✓ でなく △ に落とし、走査できない旨を開示している。
        #     だから「値の不変」を求めるのは**検体の側の誤り**だった。
        #   ★ ただし『何が消えたか』は言っていない ── 事後条件の台帳が持つ
        #     「MERGE: 結合で消える値を見ていない」は別の課題として生きている。
        if kind != 'merge' and before != after:
            diff = [(r, c) for r in range(min(len(before), len(after)))
                    for c in range(min(len(before[r]), len(after[r])))
                    if before[r][c] != after[r][c]]
            return False, f'値が変わった {diff[:3]}（書式だけを変えるはず）'
        ws_b, wb_b = books.cells('before')
        ws_a, wb_a = books.cells('after')
        try:
            if kind == 'merge':
                got = {str(m) for m in ws_a.merged_cells.ranges}
                had = {str(m) for m in ws_b.merged_cells.ranges}
                if not (got - had):
                    return False, 'セルが結合されていない'
                return True, ''
            if kind == 'width':
                wa = {k: d.width for k, d in ws_a.column_dimensions.items() if d.width}
                wb_ = {k: d.width for k, d in ws_b.column_dimensions.items() if d.width}
                if wa == wb_:
                    return False, '列幅が 1 つも変わっていない'
                return True, ''
            i = idx[kind]
            sb, sa = _fmt_signature(ws_b), _fmt_signature(ws_a)
            changed = [k for k in sa if sb.get(k, (None,) * 5)[i] != sa[k][i]]
            if not changed:
                return False, f'{kind} が 1 セルも変わっていない'
            if at is not None and at not in changed:
                return False, f'{kind} が {at} に付いていない（付いたのは {changed[:3]}）'
            return True, ''
        finally:
            wb_b.close(); wb_a.close()
    return check
def _both(*checks):
    """複数の期待を**すべて**満たすこと（片方だけ見て通すと、見た目だけ正しい回を拾う）。"""
    def check(before, after):
        for c in checks:
            ok, why = c(before, after)
            if not ok:
                return False, why
        return True, ""
    return check


def rows_swapped(name_a, name_b):
    """2 行が入れ替わり、**中身は行ごと付いていく**。他の行は動かない。"""
    def check(before, after):
        if len(after) != len(before):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        def _find(grid, name):
            return next((i for i, r in enumerate(grid) if str(r[0]) == name), None)
        ib, jb = _find(before, name_a), _find(before, name_b)
        ia, ja = _find(after, name_a), _find(after, name_b)
        if None in (ib, jb, ia, ja):
            return False, f"行が見つからない（{[r[0] for r in after]}）"
        if (ia, ja) != (jb, ib):
            return False, f"入れ替わっていない（{name_a}: {ib}→{ia} / {name_b}: {jb}→{ja}）"
        # ★★ 2026-09-02 に検体を直した（**製品は正しく、俺の期待が間違っていた**）:
        #   式のある表では、行が動くと式の**文字が変わる**のが正しい
        #   （=B2*C2 が 2 行目に残る ＝ 自分の行を指し続ける）。
        #   丸ごと一致で見ると、その正しい振る舞いを「壊れた」と読む。
        #   ★ しかも間違いの向きが悪い ── **値だけ交換する実装なら式は動かない**ので、
        #     丸ごと一致は**その事故のほうを通してしまう**。
        #   ★ ここは「式でないセルが行ごと付いてくる」で見る。
        #     式が正しい行を指しているかは formulas_point_at_their_own_row が別に見る。
        def _plain(row):
            return [None if (isinstance(v, str) and v.startswith("=")) else v for v in row]
        if _plain(after[ia]) != _plain(before[ib]) or _plain(after[ja]) != _plain(before[jb]):
            return False, "行の中身が付いてきていない（値だけ交換した疑い）"
        rest_b = [_plain(r) for i, r in enumerate(before) if i not in (ib, jb)]
        rest_a = [_plain(r) for i, r in enumerate(after) if i not in (ia, ja)]
        if rest_b != rest_a:
            return False, "他の行が変わっている"
        return True, ""
    return check


def columns_swapped(name_a, name_b):
    """2 列が入れ替わり、**列の中身は見出しに付いていく**。行数は変わらない。"""
    def check(before, after):
        if len(after) != len(before):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        hb = [str(x or "") for x in before[0]]
        ha = [str(x or "") for x in after[0]]
        if sorted(hb) != sorted(ha):
            return False, f"見出しの顔ぶれが変わった {hb}→{ha}"
        if name_a not in hb or name_b not in hb:
            return False, f"対象の見出しが無い（{hb}）"
        ib, jb = hb.index(name_a), hb.index(name_b)
        ia, ja = ha.index(name_a), ha.index(name_b)
        if (ia, ja) != (jb, ib):
            return False, f"入れ替わっていない（{name_a}: {ib}→{ia} / {name_b}: {jb}→{ja}）"
        for r in range(1, len(after)):
            if after[r][ia] != before[r][ib] or after[r][ja] != before[r][jb]:
                return False, f"{r + 1}行目の中身が見出しに付いてきていない"
        return True, ""
    return check


def cells_swapped(r1, c1, r2, c2):
    """**その 2 セルだけ**が互いの値になる（他は 1 セルも変わらない）。"""
    def check(before, after):
        if len(before) != len(after):
            return False, f"行数が変わった {len(before)}→{len(after)}"
        diff = sorted((r, c) for r in range(len(after))
                       for c in range(len(after[r])) if before[r][c] != after[r][c])
        want = sorted([(r1 - 1, c1 - 1), (r2 - 1, c2 - 1)])
        if diff != want:
            return False, f"変わったセルが {diff}（{want} だけのはず）"
        if (after[r1 - 1][c1 - 1] != before[r2 - 1][c2 - 1]
                or after[r2 - 1][c2 - 1] != before[r1 - 1][c1 - 1]):
            return False, "互いの値になっていない"
        return True, ""
    return check


def formulas_point_at_their_own_row(col_index):
    """★ 入れ替えの本題: 動いた行の式が、**自分の行**を指し続けていること。

    ★ 見た目（並び）が正しくても、値を交換する実装だと各行の計算結果が他の行の値になる
      ── 人の目では気づけない「静かに壊れる」形（実測で設計が決まった）。
    ★ ここでは式の**文字**を見る: N 行目の式が N 行目のセルを参照していること。
    """
    def check(before, after):
        import re
        for r in range(1, len(after)):
            v = after[r][col_index - 1]
            if not (isinstance(v, str) and v.startswith("=")):
                return False, f"{r + 1}行目が式でない（{v!r}）"
            rows = {int(m) for m in re.findall(r"[A-Z]+(\d+)", v)}
            if rows and rows != {r + 1}:
                return False, f"{r + 1}行目の式が他の行を指している（{v!r}）"
        return True, ""
    return check


def column_deleted(name):
    def check(before, after):
        heads_b = [str(h or "") for h in before[0]]
        heads_a = [str(h or "") for h in after[0]]
        if name in heads_a:
            return False, f"『{name}』がまだ在る"
        if len(heads_a) != len(heads_b) - 1:
            return False, f"列数が {len(heads_b)}→{len(heads_a)}（1 減るはず）"
        return True, ""
    return check


def rows_sorted_by(col_index: int, desc: bool):
    def check(before, after):
        if len(after) != len(before):
            return False, "行数が変わった"
        vals = [r[col_index - 1] for r in after[1:]]
        want = sorted(vals, reverse=desc)
        if vals != want:
            return False, f"並びが {vals}（{want} のはず）"
        return True, ""
    return check


# --- 検体（基本操作 × 言い回し）---------------------------------------------------------
#
# ★ 言い回しは「思いつく限り」ではなく、**人が実際に書きそうな形**を並べる。
#   ここが痩せていると「高精度」の主張が空になる。

def _cases_for(key: str):
    t = TABLES[key]
    h = t["headers"]
    r1, r2, r3 = (row[0] for row in t["rows"])
    c1 = h[0]           # 1 列目の見出し
    c2 = h[1]           # 2 列目
    c3 = h[2]           # 3 列目（数値）
    out = []

    # ★ 「請求」は**式のある表**。入れ替えで式が壊れないかだけを測る専用の表なので、
    #   汎用の 6 op は回さない（既存の期待値を式のぶん書き換えずに済ませる）。
    if key == "請求":
        # ⑦ 入れ替え ── 本題は「**式が自分の行を指し続けるか**」。
        for task in (f"{r1}と{r3}の行を入れ替えて", f"{r1}の行と{r3}の行を交換して"):
            out.append(("swap", task, _both(rows_swapped(r1, r3),
                                             formulas_point_at_their_own_row(4))))
        out.append(("swap", "2行目と4行目を入れ替えて",
                     _both(rows_swapped(r1, r3), formulas_point_at_their_own_row(4))))
        # 列の入れ替え（式は列の移動に追随する）
        out.append(("swap", "件数と単価の列を入れ替えて",
                     _both(columns_swapped("件数", "単価"),
                            formulas_point_at_their_own_row(4))))
        # セル 2 つ ── 頼んだ 2 つ**だけ**が動くこと
        out.append(("swap", f"{r1}の単価と{r3}の単価を入れ替えて",
                     cells_swapped(2, 3, 4, 3)))
        out.append(("swap", f"{r1}と{r3}の単価を入れ替えて",
                     cells_swapped(2, 3, 4, 3)))
        return out

    # ① セルに値を入れる（1 セルだけ）
    for task in (f"{r2}の{c2}を「東棟」にして",
                  f"{r2}の{c2}に東棟と入れて",
                  f"{r2}の{c2}を東棟に変えて",
                  f"3行目の{c2}を「東棟」にして",
                  f"{r2}の右に東棟",
                  f"{r2}の隣に東棟"):
        out.append(("cell", task, cell_becomes(3, 2, "東棟")))
    out.append(("cell", f"{r3}の{c3}を「999」にして", cell_becomes(4, 3, 999)))
    out.append(("cell", f"4行目の{c3}を999にして", cell_becomes(4, 3, 999)))

    # ② 行を足す（名前つき・位置指定）
    for task in (f"{r1}と{r2}の間に新品を作って",
                  f"{r1}と{r2}の間に新品を追加して",
                  f"{r1}と{r2}の間に新品を入れて",
                  f"{r1}の下に新品を追加して"):
        out.append(("row_add", task, row_added_at(3, "新品")))
    out.append(("row_add", f"{r3}の下に新品を追加して", row_added_at(5, "新品")))
    out.append(("row_add", f"{r2}の上に新品を入れて", row_added_at(3, "新品")))
    # ★★ 2026-08-29（Namakoo）:「どうしても中身でさせない場面が出てくる。例えば
    #   4行目と5行目は両方ともヤマノ食品。取引先で指定は出来ない」
    #   ★ 同じ値が 2 行あれば中身では指せない ── **番号でしか言えない場面がある**。
    #   実測で、番号で言うと 4 行目に空行が挿さっていた（下ではなく上・値も入らない）。
    for task, at in (("3行目の下に新品を追加して", 4), ("3行目の上に新品を入れて", 3),
                      ("3行目と4行目の間に新品を作って", 4)):
        out.append(("row_add", task, row_added_at(at, "新品")))

    # ③ 行を消す
    for task in (f"{r2}の行を削除して", f"{r2}を削除して", f"{r2}の行を消して",
                  f"{r2}の行を除いて", "3行目を削除して"):
        out.append(("row_del", task, row_deleted(r2)))

    # ④ 列を足す
    for task in (f"{c2}の右に区分という列を追加して",
                  f"{c2}の右に区分の列を作って",
                  f"{c1}と{c2}の間に区分の列を追加して"):
        idx = 3 if "右" in task else 2
        out.append(("col_add", task, column_added_named("区分", idx)))
    out.append(("col_add", "区分という列を追加して", column_added_named("区分")))

    # ⑤ 列を消す
    for task in (f"{c2}の列を削除して", f"{c2}列を消して"):
        out.append(("col_del", task, column_deleted(c2)))

    # ⑥ 並べ替え
    out.append(("sort", f"{c3}で降順に並べ替えて", rows_sorted_by(3, True)))
    out.append(("sort", f"{c3}の大きい順にして", rows_sorted_by(3, True)))
    out.append(("sort", f"{c3}の少ない順に並べて", rows_sorted_by(3, False)))

    # ⑦ 入れ替え（★ 2026-09-02 に足した分母 ── 8/31 の欠陥 13 件中 5 件がここだった）
    for task in (f"{r1}と{r3}の行を入れ替えて", f"{r1}の行と{r3}の行を交換して",
                  f"{r1}と{r3}の順番を入れ替えて"):
        out.append(("swap", task, rows_swapped(r1, r3)))
    out.append(("swap", "2行目と4行目を入れ替えて", rows_swapped(r1, r3)))
    out.append(("swap", f"{c2}と{c3}の列を入れ替えて", columns_swapped(c2, c3)))
    out.append(("swap", f"{r1}の{c3}と{r3}の{c3}を入れ替えて", cells_swapped(2, 3, 4, 3)))

    # ⑧ 計算列 ── 数値が 2 本ある表でだけ測れる（★ 2026-09-02 に足した分母）。
    #   ここまで測って初めて「名前・位置・式」の 3 つが揃ったと言える。
    if key == "見積":
        # 名前を言った回 ── **依頼文の名前**が見出しになること（A' 原則）
        out.append(("compute", "数量と単価をかけた金額の列を作って",
                     computed_column("金額")))
        out.append(("compute", "数量に単価をかけた小計の列を追加して",
                     computed_column("小計")))
        # 名前 ＋ 位置
        out.append(("compute", "単価の右に、数量と単価をかけた金額の列を作って",
                     computed_column("金額", 4)))
        out.append(("compute", "数量と単価の間に、数量と単価をかけた金額の列を作って",
                     computed_column("金額", 3)))
        # ★ 名前を言っていない回 ── 発明しない（数式風の見出しでよい）。位置と式だけ見る。
        out.append(("compute", "数量と単価をかけた列を作って", computed_column(None)))
        out.append(("compute", "単価の右に、数量と単価をかけた列を作って",
                     computed_column(None, 4)))
    # ⑨ 書式 ── ★ 値の格子には出ない op（2026-09-04 に足した分母）。
    #   ★ ここが無防備だった理由: 太字やけい線は**値を壊しても格子に差が出ない**ので、
    #     効果の検体では 1 件も測れていなかった。事後条件の台帳でも DRAW_BORDERS /
    #     NUMBER_FORMAT / MERGE が「値の不変を見ていない」として在庫に載っている。
    #   ★ 検算は必ず 2 つを対で見る ── 「値が 1 つも変わらない」かつ「書式が実際に付いた」。
    #     片方だけだと『何もしなかった』か『ついでに値を壊した』が通る。
    out.append(("fmt_bold", "見出しを太字にして", format_applied("bold")))
    out.append(("fmt_bold", "1行目を太字にして", format_applied("bold")))
    out.append(("fmt_align", f"{c2}の列を中央揃えにして", format_applied("align")))
    out.append(("fmt_border", "表にけい線を引いて", format_applied("border")))
    out.append(("fmt_fill", "見出しに背景色を付けて", format_applied("fill")))
    out.append(("fmt_width", "列幅を自動調整して", format_applied("width")))
    # ★ 数値書式は「数値の列がある表」でだけ意味を持つ
    if key in ("見積", "請求", "在庫"):
        out.append(("fmt_numfmt", f"{c3}に桁区切りを付けて", format_applied("numfmt")))
    # ★ セル結合は「1 行目を横につなぐ」が最も素直な依頼
    out.append(("fmt_merge", "1行目のA列とB列を結合して", format_applied("merge")))
    return out


def run_one(table_key: str, task: str, check, workdir: Path, timeout: float):
    src = _build(table_key, workdir / "in.xlsx")
    sheet = TABLES[table_key]["sheet"]
    before = _grid(src, sheet)
    out = workdir / "in.out.xlsx"
    if out.exists():
        out.unlink()
    env = {**__import__("os").environ, "PYTHONPATH": str(ROOT / "src")}
    # ★ 2026-08-30: 入口を差し替えられるようにした（AILINE_ENTRY）。
    #   既定は製品そのもの。段階的に聞く実験（bench/staged_translate.py）を
    #   **同じ検体で**測るためだけの口で、製品の挙動は 1 ビットも変わらない。
    _entry = __import__("os").environ.get("AILINE_ENTRY")
    _cmd = ([sys.executable, _entry] if _entry else [sys.executable, "-m", "ailine"])
    p = subprocess.run(
        _cmd + ["run", str(src), task, "--copy",
                 "--sheet", sheet, "--timeout", "90"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=timeout, cwd=str(ROOT), env=env)
    stdout = p.stdout or ""
    if p.returncode == 3 or "？" in stdout.split("\n")[-6:] and p.returncode != 0:
        return "refused", _first_line(stdout, "？")
    if p.returncode != 0:
        return "failed", _first_line(stdout, "×") or f"exit {p.returncode}"
    if not out.exists():
        return "failed", "出力が無い"
    # ★ 2026-09-04: 書式の op（太字・けい線・数値書式…）は**値の格子に出ない**ので、
    #   検算に「ファイルを読む口」を渡せるようにした。★ 既存の 2 引数の検算は無傷 ──
    #   3 つ目を受け取る検算にだけ渡す（形を変えると 160 件を全部書き直すことになる）。
    after = _grid(out, sheet)
    if len(inspect.signature(check).parameters) >= 3:
        ok, why = check(before, after, _Books(src, out, sheet))
    else:
        ok, why = check(before, after)
    return ("pass", "") if ok else ("failed", why)


def _first_line(text: str, mark: str) -> str:
    for ln in (text or "").split("\n"):
        s = ln.strip()
        if s.startswith(mark):
            return s[:160]
    return ""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--table", default=None, choices=sorted(TABLES))
    ap.add_argument("--op", default=None)
    ap.add_argument("--out", default=None)
    ap.add_argument("--timeout", type=float, default=420)
    ns = ap.parse_args()

    keys = [ns.table] if ns.table else sorted(TABLES)
    results, tally = [], {"pass": 0, "refused": 0, "failed": 0}
    with tempfile.TemporaryDirectory(prefix="ailine_matrix_") as tmp:
        work = Path(tmp)
        for key in keys:
            for op, task, check in _cases_for(key):
                if ns.op and op != ns.op:
                    continue
                try:
                    verdict, why = run_one(key, task, check, work, ns.timeout)
                except subprocess.TimeoutExpired:
                    verdict, why = "failed", "時間切れ"
                tally[verdict] += 1
                results.append({"table": key, "op": op, "task": task,
                                 "verdict": verdict, "why": why})
                mark = {"pass": "✓", "refused": "？", "failed": "×"}[verdict]
                print(f"{mark} [{key}/{op}] {task}" + (f"  ── {why}" if why else ""))

    n = sum(tally.values())
    print()
    print(f"合計 {n} 件: ✓ {tally['pass']}  ？断り {tally['refused']}  × 失敗 {tally['failed']}")
    if n:
        print(f"  意図どおり {tally['pass'] / n * 100:.1f}%  "
               f"／ 壊していない（✓+断り） {(tally['pass'] + tally['refused']) / n * 100:.1f}%")
    if ns.out:
        Path(ns.out).write_text(json.dumps(
            {"tally": tally, "results": results}, ensure_ascii=False, indent=2),
            encoding="utf-8")
        print(f"  → {ns.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
