"""行/列を動かした時、式の相対参照がどう追随するかを**実機で測る**。

★ なぜ在るか（2026-08-27・architect の査読）: 入れ替え（SWAP）と列の途中挿入を
  設計で決めることはできない ── LibreOffice が `=B2*C2` の参照をどう扱うかは
  実機に聞くしかない。★ ここが赤なら SWAP の設計そのものが変わる。
★ この repo の作法: 測る前に「何が分かれば設計が決まるか」を書いておく。

問い:
  Q1 行を入れ替えたら、各行の式は**自分の行**を指し続けるか（＝計算結果が付いてくるか）
  Q2 列を途中に挿したら、右側の式は**ずれた参照**に追随するか
  Q3 入れ替えに使う API（moveRange / 値の入れ替え）で結果が変わるか
"""
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "src"))
import ailine  # noqa: E402


def _book(p: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["商品", "数量", "単価", "金額"])
    for i, (name, q, u) in enumerate([("りんご", 2, 500), ("みかん", 3, 800),
                                       ("ぶどう", 4, 900)], start=2):
        ws.cell(i, 1, name); ws.cell(i, 2, q); ws.cell(i, 3, u)
        ws.cell(i, 4, f"=B{i}*C{i}")
    wb.save(p)
    return p


def _run_basic(book: Path, body: str) -> tuple:
    code = ("Option VBASupport 1\nOption Explicit\n\nSub Run(oDoc As Object)\n"
            "    Dim oSheet As Object\n    oSheet = oDoc.Sheets.getByIndex(0)\n"
            + body + "End Sub\n")
    with tempfile.TemporaryDirectory() as wd:
        return ailine.basrun_apply(book, code, Path(wd), timeout=120)


def _read(p: Path):
    raw = openpyxl.load_workbook(p)["売上"]
    val = openpyxl.load_workbook(p, data_only=True)["売上"]
    return [[(raw.cell(r, c).value, val.cell(r, c).value) for c in range(1, 6)]
             for r in range(1, 5)]


def main() -> int:
    tmp = Path(tempfile.mkdtemp())
    print("★ 式の追随を実機で測る（LibreOffice 経由）")

    # Q1: 2 行を入れ替える（moveRange 相当を Basic で）
    p1 = _book(tmp / "swap_rows.xlsx")
    ok, err, _ = _run_basic(p1, """
    Dim oCell As Object, i As Integer
    Dim tmpV As Variant, tmpF As String
    ' 2 行目と 3 行目を、セルごと入れ替える（式は式のまま運ぶ）
    For i = 0 To 3
        tmpF = oSheet.getCellByPosition(i, 1).getFormula()
        oSheet.getCellByPosition(i, 1).setFormula(oSheet.getCellByPosition(i, 2).getFormula())
        oSheet.getCellByPosition(i, 2).setFormula(tmpF)
    Next i
""")
    print(f"\nQ1 行の入れ替え（式を式のまま運ぶ）: ok={ok} {err or ''}")
    for row in _read(p1):
        print("   ", row)

    # Q2: 列を途中に挿す（B の右＝index 2 に 1 列）
    p2 = _book(tmp / "insert_col.xlsx")
    ok2, err2, _ = _run_basic(p2, "    oSheet.Columns.insertByIndex(2, 1)\n")
    print(f"\nQ2 列を途中に挿す（数量の右に 1 列）: ok={ok2} {err2 or ''}")
    for row in _read(p2):
        print("   ", row)

    # Q3: LibreOffice 自身の行移動（moveRange）で参照が追随するか
    p3 = _book(tmp / "move_rows.xlsx")
    body3 = (
        "    Dim oRange As New com.sun.star.table.CellRangeAddress" + chr(10)
        + "    Dim oDest As New com.sun.star.table.CellAddress" + chr(10)
        + "    oSheet.Rows.insertByIndex(1, 1)" + chr(10)
        + "    oRange.Sheet = 0 : oRange.StartRow = 3 : oRange.EndRow = 3" + chr(10)
        + "    oRange.StartColumn = 0 : oRange.EndColumn = 3" + chr(10)
        + "    oDest.Sheet = 0 : oDest.Row = 1 : oDest.Column = 0" + chr(10)
        + "    oSheet.moveRange(oDest, oRange)" + chr(10)
        + "    oSheet.Rows.removeByIndex(3, 1)" + chr(10))
    ok3, err3, _ = _run_basic(p3, body3)
    print("")
    print("Q3 LibreOffice の moveRange で みかんの行を上へ: ok=" + str(ok3) + " " + str(err3 or ""))
    for row in _read(p3):
        print("   ", row)

    print("\n★ 読み方: 1 つ目が式・2 つ目が計算結果。")
    print("   Q1 は金額が『自分の行の数量×単価』になっているか")
    print("   Q2 は金額の式が =B*C から =B*D へずれているか")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
