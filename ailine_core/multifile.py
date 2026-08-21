"""multifile — M1読み: `ailine scan <folder>` の棚卸しロジック（書き込みゼロ）。
   DESIGN-20260821-multifile.md v2 §1(M1読み)・§2(骨)・§3(③④分母/原本無変更)。

   ★ LibreOffice は一切起動しない（openpyxl のみで完結・読むだけ）。
   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定（detect_header_row/_row_char_stats）は ailine.py 側（cmd_scan）が既存のものを1回だけ
   呼び、その結果（header_row・base_headers）をこのモジュールの関数へ値として渡す。
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

_TEMP_PREFIX = "~$"                      # Excel の一時ファイル（開いている間だけ現れる隣接ファイル）
_EXCEL_EXTS = {".xlsx", ".xls"}
_MAX_HEADER_COLS = 200                   # 見出し行を読む安全上限（ailine.py の MAX_COLS とは独立）


def classify_folder_contents(folder: Path):
    """folder 直下（サブフォルダの中は見ない）を分類する。
       戻り値: (candidates: 名前順の Path リスト, excluded: {"temp": n, "subdirs": n})。
       ★ 分母そのものが検証対象（V7）── ~$ 一時ファイルとサブフォルダは対象外として数える
       （1件以上あれば呼び出し側が1行ずつ開示する）。その他の拡張子は黙って無視してよい。"""
    candidates = []
    excluded = {"temp": 0, "subdirs": 0}
    for item in sorted(folder.iterdir(), key=lambda p: p.name):
        if item.is_dir():
            excluded["subdirs"] += 1
            continue
        if not item.is_file():
            continue
        if item.name.startswith(_TEMP_PREFIX):
            excluded["temp"] += 1
            continue
        if item.suffix.lower() in _EXCEL_EXTS:
            candidates.append(item)
    return candidates, excluded


def open_base_workbook(candidates):
    """基準ファイル方式: パス辞書順（呼び出し側で名前順に並べ済み）で最初に読めた .xlsx を
       基準にする。戻り値: (path, workbook) または、読める .xlsx が1つも無ければ (None, None)。
       ★ .xls は openpyxl で開けないため基準になれない（読めたものだけが資格を持つ）。"""
    for path in candidates:
        if path.suffix.lower() != ".xlsx":
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception:
            continue
        return path, wb
    return None, None


def read_row_headers(ws, header_row: int) -> list:
    """header_row（1起点）の先頭列から連続する非空セルを見出し名として読む
       （第一波は単純に ── 列の間に空白を挟む見出しは扱わない・DESIGN §2骨）。"""
    headers = []
    for c in range(1, _MAX_HEADER_COLS + 1):
        v = ws.cell(row=header_row, column=c).value
        if v in (None, ""):
            break
        headers.append(str(v))
    return headers


def find_matching_sheet(wb, base_sheet_name: str | None):
    """他ファイルは基準と同名のシートを探し、無ければ最初のシートで照合する（DESIGN §2骨）。
       戻り値: (worksheet, fallback: bool)。fallback=True は「基準名のシートが無く1枚目へ
       落ちた」こと（★ architect 致命5 前段の開示: 呼び出し側がファイル単位の結果に
       sheet_fallback として載せる）。base_sheet_name が無い（基準ブック自身など）場合は
       比較対象が無いので fallback=False のまま。"""
    if base_sheet_name:
        if base_sheet_name in wb.sheetnames:
            return wb[base_sheet_name], False
        return wb[wb.sheetnames[0]], True
    return wb[wb.sheetnames[0]], False


def classify_headers(base_headers: list, other_headers: list):
    """3判定（列名の完全一致のみが根拠・ゆるい寄せはしない）:
       並びまで一致 → ("取れた", None)
       多重集合が一致・順序だけ違う → ("取れた", "並べ替え")
       それ以外 → ("取れなかった", "欠け/余りの名指し")"""
    if other_headers == base_headers:
        return "取れた", None
    if sorted(other_headers) == sorted(base_headers):
        return "取れた", "並べ替え"
    missing = [h for h in base_headers if h not in other_headers]
    extra = [h for h in other_headers if h not in base_headers]
    parts = []
    if missing:
        parts.append(f"欠け: {', '.join(missing)}")
    if extra:
        parts.append(f"余り: {', '.join(extra)}")
    return "取れなかった", "; ".join(parts) if parts else "列名が一致しません"


def numeric_value_column(ws, header_row: int, num_cols: int) -> int | None:
    """基準シートで、見出し行の下で最初に数値が現れる列（1起点）を返す。無ければ None。
       ★ 単位L の配線: 基準ファイルで1回だけ決める。呼び出し側がこの列の**列名**を
       他ファイルへ渡し、各ファイルはその名前を自分の並びで引き直す（_column_index）
       ── 並べ替えファイルで違う列を数える事故を避ける（implementer 申告・検体化済み）。"""
    max_row = ws.max_row or header_row
    for col in range(1, num_cols + 1):
        for row in range(header_row + 1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return col
    return None


def _column_index(headers: list, name: str) -> int | None:
    """headers（そのファイル自身の並び）の中で name と同名の列位置（1起点）。無ければ None。"""
    try:
        return headers.index(name) + 1
    except ValueError:
        return None


def total_row_candidate_count(ws, header_row: int, label_col: int, value_col: int) -> int:
    """単位L: ラベル列(label_col)・数値列(value_col) で split_total_rows を走らせ、
       除外（合計行候補）の件数を返す（--json の分布測定の口・DESIGN v2.1）。"""
    from ailine_core.total_row import split_total_rows
    max_row = ws.max_row or header_row
    rows = [(r, ws.cell(row=r, column=label_col).value, ws.cell(row=r, column=value_col).value)
            for r in range(header_row + 1, max_row + 1)]
    return len(split_total_rows(rows).excluded)


def evaluate_file(path: Path, base_headers: list, base_sheet_name: str | None, header_row: int,
                   value_col_name: str | None = None) -> dict:
    """1ファイルを基準と照合する。戻り値: {"name", "status", "reason"(取れなかった時),
       "reordered"(並べ替えで取れた時), "total_row_candidates"(取れた時・列名が引ければ)}。
       ★ どんな失敗でも例外を上げず名指し+理由にして返す
       （$0 条件「黙って失敗する」の裏返し ── 報告が成果物）。
       ★ ラベル列・数値列は基準の**列名**で引き当てる（並べ替えファイルでは位置が違う）。"""
    if path.suffix.lower() != ".xlsx":
        return {"name": path.name, "status": "取れなかった", "reason": "旧形式(.xls)"}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {"name": path.name, "status": "取れなかった", "reason": f"読み込み失敗: {e}"}
    try:
        ws, sheet_fallback = find_matching_sheet(wb, base_sheet_name)
        other_headers = read_row_headers(ws, header_row)
        status, detail = classify_headers(base_headers, other_headers)
        entry = {"name": path.name, "status": status}
        if status == "取れなかった":
            entry["reason"] = detail
        elif detail:   # "並べ替え"
            entry["reordered"] = True
        if sheet_fallback:
            entry["sheet_fallback"] = {"wanted": base_sheet_name, "used": ws.title}
        if status == "取れた" and value_col_name is not None and base_headers:
            label_col = _column_index(other_headers, base_headers[0])
            value_col = _column_index(other_headers, value_col_name)
            if label_col is not None and value_col is not None:
                entry["total_row_candidates"] = total_row_candidate_count(
                    ws, header_row, label_col, value_col)
    finally:
        wb.close()
    return entry
