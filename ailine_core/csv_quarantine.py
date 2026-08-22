"""csv_quarantine — CSV→xlsx の型検疫。DESIGN-20260821-multifile.md
「CSV 検疫 設計 v2」・REVIEW-20260822-csv-architect.md・tests/test_csv_truth_table.py
（真理値表・凍結済み）の実装。

★ 分類規則の正本は tests/test_csv_truth_table.py の冒頭コメント。ここに書く規則は
  その実装であって、食い違えば直すのはこの module（真理値表は手で凍結・変更禁止）。

★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）。依存は
  stdlib + openpyxl のみ（xml_readback は ailine_core の下位モジュールとして import
  可 ── 検算に使う。禁止されているのは ailine.py への逆流だけ）。

## 憲法: LLM はデータに 1 バイトも触らない
  ここに書く判定はすべて決定論の Python。7B の仕事は意図分類と列の名指しだけ
  （このモジュールの外の話）。

## 型検疫の核心（0 落ち根絶）
  全セルを文字列として読み、列単位で 4 分類（number/date/string/undecidable）を
  多数決なしで決める。一票拒否権 a〜g のどれか 1 セルでも該当すれば、その列は
  number/date になれず string へ落ちる（規則の全文は真理値表冒頭コメント）。

## 実測済みの罠（★ ここを外すと壊れる）
  - `=SUM(1,2)` を openpyxl セルへ素で代入すると `cell.data_type` が自動で 'f'
    （数式）になり、xml_readback からは <f>..</f><v></v> で値が脱落する。
    対策: value 代入の**後**に `cell.data_type = "s"` へ上書きする（実測: 上書きは
    save まで保持され、readback 側は文字列として正しく読める）。
  - C0 制御文字を含む文字列を openpyxl に代入すると `IllegalCharacterError` で
    save 前から落ちる（実測）。書く前に除去し、除去した事実を計上する。
  - 日付は openpyxl の日付書式が「日付/時刻の書式」と xml_readback に認識される
    書式コード（y/m/d/h のトークンを含む）が要る。`number_format = "yyyy-mm-dd"`
    を明示すれば、xml_readback._is_date_numfmt がカスタム numFmt からでも認識する
    （実機検証済み）。
"""
from __future__ import annotations

import csv
import datetime
import hashlib
import io
import re
from dataclasses import dataclass

import openpyxl

from ailine_core import xml_readback

# =====================================================================
# 文字コード検出（BOM 最優先 → UTF-8 → cp932 の順・実測根拠は真理値表 ENC_TRUTH）
# =====================================================================

_BOM_UTF8 = b"\xef\xbb\xbf"
_BOM_UTF16 = (b"\xff\xfe", b"\xfe\xff")   # ★ 対象外宣言（設計 v2）── 黙って誤爆させない


class UndecidableEncodingError(ValueError):
    """utf-8 でも cp932 でも復号できないバイト列（EUC/JIS/UTF-16 は設計 v2 で対象外）。"""


@dataclass(frozen=True)
class EncodingResult:
    """encoding: "utf-8-sig" | "utf-8" | "cp932"。
       ambiguous: 両方の符号化で復号可能だった（内容は違う）── UTF-8 続行+開示の対象。"""
    encoding: str
    ambiguous: bool


def detect_encoding(raw: bytes) -> EncodingResult:
    """BOM 最優先 → UTF-8 → cp932 の順で判定する（実測: 短文 UTF-8 の cp932 黙読
    36〜41%・逆方向 0% ── だから UTF-8 を先に試す）。

    ★ 純 ASCII バイト列は UTF-8/cp932 のどちらで読んでも同一内容になるため、
      cp932 を試すまでもなく非曖昧（ambiguous=False）と即決する。
    """
    if raw.startswith(_BOM_UTF8):
        return EncodingResult(encoding="utf-8-sig", ambiguous=False)
    if raw.startswith(_BOM_UTF16):
        raise UndecidableEncodingError("UTF-16 は対象外（設計 v2・名指しで宣言）")
    if all(b < 0x80 for b in raw):
        return EncodingResult(encoding="utf-8", ambiguous=False)

    try:
        raw.decode("utf-8")
    except UnicodeDecodeError:
        utf8_ok = False
    else:
        utf8_ok = True

    if utf8_ok:
        try:
            raw.decode("cp932")
        except UnicodeDecodeError:
            return EncodingResult(encoding="utf-8", ambiguous=False)
        return EncodingResult(encoding="utf-8", ambiguous=True)

    try:
        raw.decode("cp932")
    except UnicodeDecodeError as e:
        raise UndecidableEncodingError(
            "utf-8 でも cp932 でも復号できない（EUC/JIS 等は対象外・設計 v2）"
        ) from e
    return EncodingResult(encoding="cp932", ambiguous=False)


# =====================================================================
# セル単位の隔離（列 kind とは独立）
# =====================================================================

# openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE と同じ範囲（実測: これらを含む文字列は
# save 前から IllegalCharacterError で落ちる）。openpyxl 内部を import せず自前で
# 同じ判定を持つ ── この module 自身の契約として独立に持つ。
_CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

# Excel の 1 セル文字数上限は 32,767。32,768 文字以上は Excel が壊す領域
# （openpyxl 自体は無警告で保存する ── 実測）。
_OVERLONG_THRESHOLD = 32768


def control_char_cells(cells) -> list:
    """C0 制御文字を含むセルを名指しする。戻り値: [(idx, "U+XXXX"), ...]
       （1 セルにつき最初に見つかった 1 文字だけを報告する）。"""
    hits = []
    for idx, cell in enumerate(cells):
        m = _CONTROL_CHAR_RE.search(cell)
        if m:
            hits.append((idx, f"U+{ord(m.group()):04X}"))
    return hits


def overlong_cells(cells) -> list:
    """32,768 文字以上のセルを名指しする。戻り値: [(idx, 長さ), ...]。"""
    return [(i, len(c)) for i, c in enumerate(cells) if len(c) >= _OVERLONG_THRESHOLD]


# =====================================================================
# 列単位の型分類（真理値表が正・tests/test_csv_truth_table.py 冒頭コメント）
# =====================================================================

_NUMBER_RE = re.compile(r"^-?\d{1,15}(\.\d+)?$")
_LEADING_ZERO_RE = re.compile(r"^-?0\d")
_DIGIT_RUN_RE = re.compile(r"^-?(\d+)(?:\.\d+)?$")          # 桁溢れ判定用（15 桁上限なし）
_COMMA_GROUP_RE = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$")
_FULLWIDTH_DIGITS = "０１２３４５６７８９"
_SLASH_DATE_RE = re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})$")
_DASH_DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_EIGHT_DIGIT_RE = re.compile(r"^\d{8}$")
_WAREKI_KANJI_RE = re.compile(r"^(明治|大正|昭和|平成|令和)\d{1,2}年\d{1,2}月\d{1,2}日$")
_WAREKI_ABBR_RE = re.compile(r"^[MTSHR]\d{1,2}\.\d{1,2}\.\d{1,2}$")
_EXCEL_ERROR_TOKENS = frozenset(
    {"#N/A", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#NULL!"}
)

_DIGIT_OVERFLOW_LEN = 16   # Excel の 15 桁精度で壊れる境界（整数部）


@dataclass(frozen=True)
class ColumnClassification:
    """kind: "number" | "date" | "string" | "undecidable"。
       reasons: 開示の語（真理値表の語彙）。warn: kind=="undecidable" の時だけ True。"""
    kind: str
    reasons: list
    warn: bool


def _digit_run_length(cell: str):
    """符号を除いた整数部の桁数（小数部・カンマは対象外）。数字の並びでなければ None。"""
    m = _DIGIT_RUN_RE.match(cell)
    return len(m.group(1)) if m else None


def _is_wareki(cell: str) -> bool:
    return bool(_WAREKI_KANJI_RE.match(cell) or _WAREKI_ABBR_RE.match(cell))


def _date_shape_match(cell: str):
    return _SLASH_DATE_RE.match(cell) or _DASH_DATE_RE.match(cell)


def _valid_date_from_match(m) -> bool:
    y, mo, d = (int(x) for x in m.groups())
    try:
        datetime.date(y, mo, d)
    except ValueError:
        return False
    return True


def classify_column(cells) -> ColumnClassification:
    """列（セル文字列の並び）を 4 分類する。空セルは投票しない（規則の全文は
    tests/test_csv_truth_table.py 冒頭コメント。ここはその実装）。"""
    non_empty = [c for c in cells if c != ""]
    if not non_empty:
        return ColumnClassification(kind="string", reasons=["empty_column"], warn=False)

    # --- 一票拒否権 a〜g（1 セルでも該当すれば number/date になれない）---
    veto_reasons = set()
    for c in non_empty:
        if _LEADING_ZERO_RE.match(c):
            veto_reasons.add("leading_zero")
        if c[:1] in ("=", "@"):
            veto_reasons.add("formula_head")
        n = _digit_run_length(c)
        if n is not None and n >= _DIGIT_OVERFLOW_LEN:
            veto_reasons.add("digit_overflow")
        if c != c.strip():
            veto_reasons.add("surrounding_space")
        if any(ch in _FULLWIDTH_DIGITS for ch in c):
            veto_reasons.add("fullwidth_digit")
        if c.startswith("△") or (len(c) >= 2 and c.startswith("(") and c.endswith(")")):
            veto_reasons.add("accounting_negative")

    if veto_reasons:
        return ColumnClassification(kind="string", reasons=sorted(veto_reasons), warn=False)

    # --- string 側の開示（number/date になれないことが分かってから付ける訳ではなく、
    #     最終的に string に落ちた時だけ意味を持つ ── 先に集めておく）---
    disclose = set()
    for c in non_empty:
        if c in _EXCEL_ERROR_TOKENS:
            disclose.add("excel_error_token")
        if _is_wareki(c):
            disclose.add("wareki_out_of_scope")

    date_matches = [_date_shape_match(c) for c in non_empty]
    is_num_flags = [bool(_NUMBER_RE.match(c)) for c in non_empty]
    date_valid_flags = [m is not None and _valid_date_from_match(m) for m in date_matches]

    # --- 確信クラスの衝突（日付と数値が同居）---
    if any(date_valid_flags) and any(is_num_flags):
        return ColumnClassification(kind="undecidable", reasons=["mixed_confident"], warn=True)

    # --- 日付（列一貫のみ・2 形式混在は許す）---
    if all(m is not None for m in date_matches):
        if all(date_valid_flags):
            return ColumnClassification(kind="date", reasons=[], warn=False)
        return ColumnClassification(kind="undecidable", reasons=["calendar_invalid"], warn=True)

    # --- 数値（カンマ列は列一貫時のみ+開示）---
    has_comma = any("," in c for c in non_empty)
    is_candidate = [is_num_flags[i] or ("," in non_empty[i]) for i in range(len(non_empty))]
    if all(is_candidate):
        consistent = True
        for c in non_empty:
            if "," in c:
                if not _COMMA_GROUP_RE.match(c):
                    consistent = False
                    break
            elif has_comma:
                n = _digit_run_length(c) or 0
                if n > 3:
                    consistent = False
                    break
        if not consistent:
            return ColumnClassification(kind="undecidable", reasons=["comma_inconsistent"], warn=True)
        reasons = []
        if has_comma:
            reasons.append("comma_grouped")
        if any(_EIGHT_DIGIT_RE.match(c) for c in non_empty):
            reasons.append("eight_digit_maybe_date")
        return ColumnClassification(kind="number", reasons=reasons, warn=False)

    # --- 確信をもって入らないものはバイト忠実に文字列保持（疑いではなく開示）---
    return ColumnClassification(kind="string", reasons=sorted(disclose), warn=False)


# =====================================================================
# CSV パース（csv.reader ベース・物理行範囲つき・重複ヘッダ機械リネーム）
# =====================================================================

@dataclass(frozen=True)
class ParsedRecord:
    """1 レコード分の生セル列（文字列のまま）+ 物理行範囲（1 起点・引用内改行対応）。"""
    cells: list
    first_line: int
    last_line: int


@dataclass(frozen=True)
class HeaderRename:
    """重複ヘッダの機械リネーム 1 件分の開示。"""
    index: int
    original: str
    renamed: str


@dataclass(frozen=True)
class ParseResult:
    """★ column_count_mismatches・unterminated_quote_records は「物理レコード列全体
    （ヘッダを含む・has_header=True なら index 0 がヘッダ）」に対する 0 起点 index を指す
    ── records（データ行のみ）の index とは空間が違うことに注意。"""
    header: list
    header_renames: list
    records: list
    column_count_mismatches: list   # (raw_index, expected_cols, actual_cols)
    unterminated_quote_records: list   # raw_index の列
    source_sha256: str


def sha256_bytes(raw: bytes) -> str:
    """原本 CSV バイト列の sha256（無変更確認の材料）。"""
    return hashlib.sha256(raw).hexdigest()


def _dedup_headers(header: list):
    """重複ヘッダを機械リネームする。1 回目の出現はそのまま、2 回目以降は
    `<元名>__dupN` へ（N は出現回数）。衝突が衝突を生まないよう、生成した名前も
    既出名として扱う。"""
    seen: dict = {}
    out = []
    renames = []
    for i, name in enumerate(header):
        if name not in seen:
            seen[name] = 1
            out.append(name)
            continue
        seen[name] += 1
        candidate = f"{name}__dup{seen[name]}"
        while candidate in seen:
            seen[name] += 1
            candidate = f"{name}__dup{seen[name]}"
        seen[candidate] = 1
        out.append(candidate)
        renames.append(HeaderRename(index=i, original=name, renamed=candidate))
    return out, renames


def parse_csv(text: str, *, has_header: bool = True, source_sha256: str = None) -> ParseResult:
    """csv.reader ベースでレコードを切り出す（型判定は一切しない ── classify_column
    の仕事）。

    ★ 物理行範囲: `csv.reader` の `line_num` は引用内改行を跨いだ物理行を正しく
      数える（Python csv モジュールの実測済み挙動）── これを使って
      「レコード N は物理 M〜L 行」を機械で作る（検分シートの名指し用）。
    ★ 閉じない引用の乖離: そのレコードが占める物理行区間に現れる `"` の総数が
      奇数なら、引用が閉じきっていない疑いとして ⚠ 計上する（csv モジュール自体は
      例外を上げずに黙って読み切ってしまうため、独自に検出する）。
    """
    reader = csv.reader(io.StringIO(text))
    lines = text.splitlines()
    raw_records: list = []
    unterminated: list = []
    prev_line = 0
    for row in reader:
        start = prev_line + 1
        end = reader.line_num
        raw_records.append(ParsedRecord(cells=row, first_line=start, last_line=end))
        segment = lines[start - 1:end]
        if sum(line.count('"') for line in segment) % 2 != 0:
            unterminated.append(len(raw_records) - 1)
        prev_line = end

    header: list = []
    header_renames: list = []
    if has_header and raw_records:
        header, header_renames = _dedup_headers(raw_records[0].cells)
        data_records = raw_records[1:]
    else:
        data_records = raw_records

    expected_cols = len(raw_records[0].cells) if raw_records else 0
    mismatches = [
        (i, expected_cols, len(rec.cells))
        for i, rec in enumerate(raw_records)
        if len(rec.cells) != expected_cols
    ]

    return ParseResult(
        header=header,
        header_renames=header_renames,
        records=data_records,
        column_count_mismatches=mismatches,
        unterminated_quote_records=unterminated,
        source_sha256=source_sha256,
    )


def build_columns(header: list, records: list) -> list:
    """classify_column に渡す列単位のセル列を組み立てる小さな糊。ヘッダ長と
    データ行の最大列数の大きい方を列数とし、短い行は無いセルとして扱う
    （投票しない＝空文字列と同じ効果）。"""
    num_cols = len(header)
    for rec in records:
        num_cols = max(num_cols, len(rec.cells))
    columns = []
    for col_idx in range(num_cols):
        columns.append(
            [rec.cells[col_idx] if col_idx < len(rec.cells) else "" for rec in records]
        )
    return columns


# =====================================================================
# 書き出し（openpyxl・分類に従って型を守る）
# =====================================================================


def _strip_control_chars(text: str):
    """C0 制御文字を除去し、(除去後の文字列, 除去した "U+XXXX" のリスト) を返す。"""
    removed = []

    def _sub(m):
        removed.append(f"U+{ord(m.group()):04X}")
        return ""

    cleaned = _CONTROL_CHAR_RE.sub(_sub, text)
    return cleaned, removed


def _parse_number_value(cell: str):
    body = cell.replace(",", "")
    return float(body) if "." in body else int(body)


def _parse_date_value(cell: str):
    m = _date_shape_match(cell)
    if not m:
        return None
    y, mo, d = (int(x) for x in m.groups())
    try:
        return datetime.date(y, mo, d)
    except ValueError:
        return None


@dataclass(frozen=True)
class WriteResult:
    """declared: {(row, col) 1 起点: 値} ── この xlsx に書いたと主張する値
    （compare_against_quarantine の宣言分母）。空セルは declared に入れない
    （xml_readback.read_grid も空セルを保持しない ── 対称性）。
    removed_control_chars: (row, col, "U+XXXX") の列 ── 隔離して除去したセル。"""
    declared: dict
    removed_control_chars: list
    rows_written: int
    cols_written: int


def write_quarantined_xlsx(parsed: ParseResult, classifications: list, out_path,
                            *, has_header: bool = True) -> WriteResult:
    """parsed（parse_csv の結果）と classifications（列 0 起点対応の
    ColumnClassification の並び）に従って openpyxl で xlsx を書く。

    ★ 数式化け対策（実測の罠）: string 分類のセルは value 代入の後で
      `cell.data_type = "s"` に上書きする。これをしないと "=SUM(1,2)" のような
      値が openpyxl に数式(t='f')と誤認され、xml_readback からは値が脱落する。
    ★ 制御文字は書く前に除去する（実測: 素で渡すと IllegalCharacterError で
      save が落ちる）。除去した事実は removed_control_chars に名指しで残す。
    ★ undecidable 分類の列も string と同じく文字列保持で書く（⚠ は列の判定に
      付くのであって、セルの書き方を変える理由ではない ── 「壊さず持つ」）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    declared: dict = {}
    removed: list = []

    def _write_string_cell(row, col, text):
        cleaned, rm = _strip_control_chars(text)
        for cp in rm:
            removed.append((row, col, cp))
        cell = ws.cell(row=row, column=col)
        cell.value = cleaned
        cell.data_type = "s"
        if cleaned != "":
            declared[(row, col)] = cleaned

    row_no = 1
    max_cols = len(parsed.header)
    if has_header and parsed.header:
        for col_idx, name in enumerate(parsed.header):
            _write_string_cell(row_no, col_idx + 1, name)
        row_no += 1

    for rec in parsed.records:
        max_cols = max(max_cols, len(rec.cells))
        for col_idx, raw_cell in enumerate(rec.cells):
            col = col_idx + 1
            if raw_cell == "":
                continue
            classification = classifications[col_idx] if col_idx < len(classifications) else None
            kind = classification.kind if classification else "string"

            if kind == "number":
                try:
                    value = _parse_number_value(raw_cell)
                except ValueError:
                    _write_string_cell(row_no, col, raw_cell)
                else:
                    ws.cell(row=row_no, column=col, value=value)
                    declared[(row_no, col)] = value
            elif kind == "date":
                value = _parse_date_value(raw_cell)
                if value is None:
                    _write_string_cell(row_no, col, raw_cell)
                else:
                    cell = ws.cell(row=row_no, column=col)
                    cell.value = value
                    cell.number_format = "yyyy-mm-dd"
                    declared[(row_no, col)] = value
            else:
                _write_string_cell(row_no, col, raw_cell)
        row_no += 1

    wb.save(out_path)
    return WriteResult(
        declared=declared,
        removed_control_chars=removed,
        rows_written=row_no - 1,
        cols_written=max_cols,
    )


# =====================================================================
# 検算（宣言分母 vs xml_readback の読み戻し・dict 同士の == 比較はしない）
# =====================================================================


@dataclass(frozen=True)
class CompareResult:
    """missing: declared にあるが読み戻しに無い (row,col)。
       mismatched: 両方にあるが型込みで値が違う (row,col,declared値,actual値)。
       surplus: 読み戻しにあるが declared に無い (row,col,actual値)。"""
    missing: list
    mismatched: list
    surplus: list

    @property
    def ok(self) -> bool:
        return not (self.missing or self.mismatched or self.surplus)


def _values_agree(declared, actual) -> bool:
    """型込み等値・TOLERANCE 不使用。数値どうしは int/float の表現差（5 と 5.0）を
    同一視するが、種別（数値/文字列/日付）を跨いだ一致は認めない。"""
    if isinstance(declared, bool) or isinstance(actual, bool):
        return declared is actual
    if isinstance(declared, (int, float)) and isinstance(actual, (int, float)):
        return declared == actual
    if isinstance(declared, datetime.date) and isinstance(actual, datetime.date):
        return declared == actual
    if isinstance(declared, str) and isinstance(actual, str):
        return declared == actual
    return False


def compare_against_quarantine(declared: dict, out_path, *, sheet_name: str = None) -> CompareResult:
    """declared（write_quarantined_xlsx が返した {(row,col):値}）を宣言分母として、
    書いた xlsx を xml_readback で読み戻し、欠落/不一致/余剰を数える。
    ★ dict 同士の == 比較はしない（3 計数を別々に出すことで、何が違うかを言える）。"""
    data = xml_readback.read_grid(out_path, sheet_name=sheet_name)
    grid = data["grid"]

    missing = []
    mismatched = []
    for key, dval in declared.items():
        if key not in grid:
            missing.append(key)
            continue
        aval = grid[key]
        if not _values_agree(dval, aval):
            mismatched.append((key[0], key[1], dval, aval))

    surplus = [(r, c, v) for (r, c), v in grid.items() if (r, c) not in declared]

    return CompareResult(missing=missing, mismatched=mismatched, surplus=surplus)
