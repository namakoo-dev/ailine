"""stack — M1書き `ailine stack <folder> --out <path>`（縦積み・UNION ALL）の本体。
   DESIGN-20260821-multifile.md v2 §1(M1書き)・v2.1。

   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定（detect_header_row/_row_char_stats）は ailine.py 側（cmd_stack）が既存のものを1回だけ
   呼び、その結果を値としてこの module へ渡す（cmd_scan と同じ配線）。
   ★ 既存部品の再利用: 分母・基準ファイル方式・3判定は multifile.py、合計行の識別と
   閉じる検査は total_row.py（どちらも既存・単位L で完成済み）。この module は
   「積む行の決定」と「出所列つきで積む」という stack 固有の配線だけを持つ。
"""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from xml.etree import ElementTree as ET

import openpyxl

from ailine_core import inspection, multifile, total_row

PROVENANCE_HEADERS = ("元ファイル", "元行")
# ★ 赤2 の直し（2026-08-21 実機敵対検分）: 署名判定はサフィックス形（元ファイル_2 等）も
#   「自分」と認める。素の名前そのもの、または末尾に _数字 が付いた形のどちらにも当たる。
_PROVENANCE_SIGNATURE_RE = tuple(re.compile(rf"^{re.escape(name)}(_\d+)?$")
                                 for name in PROVENANCE_HEADERS)

# ★ jisaku-review#1 critical の直し: 署名を「列名だけ」から「列名 AND 書き手の印」に。
# stack が出力を書く時に docProps/core.xml の dc:creator へこの印を残す
# （cmd_stack が wb.properties.creator = CREATOR_MARK を設定）。
CREATOR_MARK = "ailine stack"
# ★ architect 致命2 の直し（M2 前置き・2026-08-21）: 「ailine の出力か」は stack 1本だけでなく
# ailine の複数コマンドに広がる。書く側の定数（CREATOR_MARK）は互換のためそのまま残し、
# 読む側の判定はこの集合で行う（当面 stack と、将来の extract を先取りして凍結）。
CREATOR_MARKS = {"ailine stack", "ailine extract"}
_CORE_NS = {"cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
           "dc": "http://purl.org/dc/elements/1.1/"}


def _is_blank(v) -> bool:
    return total_row._is_blank_cell(v)


def _row_has_any_value(ws, row: int, num_cols: int) -> bool:
    """row の 1..num_cols 列のどこかに値があるか（★ adopted_rows とは別の判定 ──
       金額が空欄でも他の列に値があれば『データ行』とみなす。凍結検体
       test_data_row_with_empty_numeric_cell_is_still_stacked の配線）。"""
    return any(not _is_blank(ws.cell(row=row, column=c).value) for c in range(1, num_cols + 1))


def own_output_headers(headers: list) -> list:
    """利用者の列名と衝突したら機械的サフィックス（元ファイル_2 等）で逃がす。
       戻り値: 実際に使う出所列名2本（衝突が無ければ PROVENANCE_HEADERS そのまま）。"""
    out = []
    for name in PROVENANCE_HEADERS:
        candidate = name
        n = 2
        while candidate in headers or candidate in out:
            candidate = f"{name}_{n}"
            n += 1
        out.append(candidate)
    return out


def is_own_signature(headers: list) -> bool:
    """headers の末尾2列が出所列の見出し（素の名前 または 衝突時のサフィックス形
       元ファイル_N / 元行_N）と一致するか。★ 列名だけの一致であり、これ単独では
       『自分の出力』の証明にならない（jisaku-review#1 実測: たまたま同じ列名の
       人のファイルを誤認する）── 呼び出し側は `is_own_output` を使うこと。"""
    if len(headers) < 2:
        return False
    return all(pat.match(str(h)) for pat, h in zip(_PROVENANCE_SIGNATURE_RE, headers[-2:]))


def _read_creator(path) -> str | None:
    """docProps/core.xml の dc:creator を直読み（zip 直読み・軽い専用の読み）。
       読めなければ None（壊れている/該当なし = 印なし = 他人のファイル扱い ── fail closed）。"""
    try:
        with zipfile.ZipFile(path) as z:
            if "docProps/core.xml" not in z.namelist():
                return None
            root = ET.fromstring(z.read("docProps/core.xml"))
    except Exception:
        return None
    el = root.find("dc:creator", _CORE_NS)
    return el.text if el is not None else None


def own_output_mark(path, headers: list) -> str | None:
    """署名 = 列名の一致 **AND** 書き手の印（docProps/core.xml の creator）。
       一致すれば実際の印（例 "ailine stack" / "ailine extract"）を返す。どちらか片方でも
       欠ければ None（他人のファイル扱い ── fail closed）。
       ★★ jisaku-review#1 critical の直し（実機再現済み）: `is_own_signature`（列名だけ）は
       末尾2列がたまたま『元ファイル』『元行』という名前の人のファイルを前回出力と誤認し、
       --overwrite 無しで無警告上書きしてデータを消した。名前が合っていても印が無ければ
       他人のファイル。
       ★ architect 致命2 の直し: 「自分（ailine の何か）の出力か」（V6 の入力自己参照除外）と
       「作り直してよい前回出力か」（書き込み関所）は問いが違う ── 前者はこの関数の戻り値が
       None でないか、後者は戻り値が CREATOR_MARK と完全一致するか、で呼び出し側が分ける。"""
    if not is_own_signature(headers):
        return None
    creator = _read_creator(path)
    return creator if creator in CREATOR_MARKS else None


def is_own_output(path, headers: list) -> bool:
    """『ailine の何らかのコマンドの出力』か（印は問わない・集合のどれかに当たれば真）。
       ★ V6（入力からの自己参照除外）が使う判定はこれ ── ailine 産は種類を問わず除外する。
       『作り直してよい』（書き込み関所）は別問い ── `own_output_mark` の戻り値を
       CREATOR_MARK と完全一致で見る（呼び出し側 = ailine.py cmd_stack）。"""
    return own_output_mark(path, headers) is not None


def numeric_column_names(ws, header_row: int, headers: list) -> list:
    """headers（基準ファイルの列名）のうち、データ行のどこかで数値を持つ列名の一覧。
       ★ jisaku-review#3/#6 の直し: Σ 照合・報告を『最初の数値列』1本だけでなく
       全数値列に広げるための土台（合計行検出の keyed 列＝ multifile.numeric_value_column
       の1本はここでは変えない・呼び出し側で従来どおり別に決める）。"""
    max_row = ws.max_row or header_row
    out = []
    for i, name in enumerate(headers, start=1):
        for row in range(header_row + 1, max_row + 1):
            v = ws.cell(row=row, column=i).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                out.append(name)
                break
    return out


@dataclass(frozen=True)
class FileStackResult:
    """1ファイルを積んだ（または積めなかった）結果。
       rows: [(base_headers 順の値リスト, 同順の number_format リスト, 元行番号), ...]（積めた時のみ・
       ★ 実視の磨き: number_format は日付セルの時刻の尻尾を消すため元セルから運ぶ）。
       excluded/mismatches: total_row.split_total_rows の戻り値そのまま。
       col_a_mismatch: (col_a_count, used_range_count) 食い違い時のみ（③・可視化専用）。
       findings: M2.5（検分シート）用の inspection.Finding のリスト ── ws を持つこの関数の
       内側で組み立てる（ファイル+シート+セルの3座標が、ここでなら列位置まで正確に引ける）。"""
    name: str
    status: str                      # "積んだ" / "積めなかった"
    reason: str | None = None
    reordered: bool = False
    rows: list = field(default_factory=list)
    excluded: list = field(default_factory=list)
    mismatches: list = field(default_factory=list)
    col_a_mismatch: tuple | None = None
    sheet_fallback: tuple | None = None   # (wanted, used) ── 基準名のシートが無く1枚目へ落ちた時だけ
    findings: list = field(default_factory=list)   # list[inspection.Finding]（M2.5）


def evaluate_and_stack(path, base_headers: list, base_sheet_name, header_row: int,
                        value_col_name: str | None) -> FileStackResult:
    """1ファイルを基準と照合し、取れていれば『積む行』を確定して値まで読む。
       ★ どんな失敗でも例外を上げず名指し+理由で返す（multifile.evaluate_file と同じ線）。"""
    if path.suffix.lower() != ".xlsx":
        return FileStackResult(name=path.name, status="積めなかった", reason="旧形式(.xls)")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return FileStackResult(name=path.name, status="積めなかった", reason=f"読み込み失敗: {e}")
    try:
        ws, sheet_fell_back = multifile.find_matching_sheet(wb, base_sheet_name)
        sheet_fallback = (base_sheet_name, ws.title) if sheet_fell_back else None
        other_headers = multifile.read_row_headers(ws, header_row)
        status, detail = multifile.classify_headers(base_headers, other_headers)
        if status == "取れなかった":
            not_taken = [inspection.finding(
                kind=inspection.KIND_NOT_TAKEN, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row), source_value=None, output_value=None,
                next_step=f"見出しが基準と合いません（{detail}）。この冊は積んでいません。")]
            return FileStackResult(name=path.name, status="積めなかった", reason=detail,
                                    sheet_fallback=sheet_fallback, findings=not_taken)
        reordered = bool(detail)

        # 各 base 列 → このファイル自身の列位置（並べ替えファイルで位置がずれる対策）。
        col_for_base = {bh: multifile._column_index(other_headers, bh) for bh in base_headers}
        max_row = ws.max_row or header_row
        all_rows = list(range(header_row + 1, max_row + 1))

        label_col = col_for_base.get(base_headers[0]) if base_headers else None
        value_col = col_for_base.get(value_col_name) if value_col_name else None
        if label_col and value_col:
            triples = [(r, ws.cell(row=r, column=label_col).value,
                        ws.cell(row=r, column=value_col).value) for r in all_rows]
            verdict = total_row.split_total_rows(triples)
        else:
            verdict = total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[])

        excluded_rows = {e.row for e in verdict.excluded}
        num_cols = len(other_headers)
        data_rows = [r for r in all_rows if _row_has_any_value(ws, r, num_cols)]
        stack_rows = [r for r in data_rows if r not in excluded_rows]

        rows = []
        for r in stack_rows:
            cells = [ws.cell(row=r, column=col_for_base[bh]) for bh in base_headers]
            values = [c.value for c in cells]
            # ★ 実視の磨き（2026-08-21）: 元セルの number_format をデータセルへ運ぶ
            #   （日付が『2026-07-09 0:00:00』と時刻付きで出ないように）。決定論
            #   ── 元の書式文字列をそのまま運ぶだけ（乱数・時刻は使わない）。
            formats = [c.number_format for c in cells]
            rows.append((values, formats, r))

        col_a_count = sum(1 for r in all_rows if not _is_blank(ws.cell(row=r, column=1).value))
        used_range_count = len(all_rows)
        col_a_mismatch = None
        if col_a_count != used_range_count:
            col_a_mismatch = (col_a_count, used_range_count)

        # ★ M2.5: 所見の組み立て（ws がまだ開いているこの関数の内側でだけ、列位置まで
        #   正確な3座標が引ける）。⚠ 相当（閉じる検査の不一致・分母食い違い）とシート
        #   fallback の開示（既存 CLI 報告と同じ基準・inspection.WARN_KINDS）。
        # ★ 合計行の閉じる検査 不一致の HYPERLINK は「ラベル(ID)列」でも「値列」でもなく
        #   基準の2列目（例: 取引先/摘要 ── ID の次に来る記述列）へ飛ばす。両側の数字は
        #   所見の行自体（元の値/採用側の値の列）に既に載っているので、リンク先まで同じ
        #   数値セルへ飛ばすのは冗長 ── 人が行を「誰の・何の行か」で認識できる列を選ぶ
        #   （ID→摘要/取引先→金額、という帳票の一般的な並びに沿う）。
        # ★ アンカーは「怪しい数字そのもの」（閉じなかった合計の値セル）。クリックの
        #   着地点に迷いを作らない（検分シートの本文が両側の数字を持つのは別の役割）。
        anchor_col = value_col or label_col or 1
        # ★ UX 磨き③（Namakoo 実視 2026-08-21 12:01）: 断片（「除外行の値が...」）でなく
        #   1所見1文。状態（両側の数字+動詞）+ 次の手（『クリック』して確認）を1文に言い切る。
        findings = []
        for m in verdict.mismatches:
            cell = inspection.cell_ref(anchor_col, m.row)
            findings.append(inspection.finding(
                kind=inspection.KIND_TOTAL_ROW_MISMATCH, file=path.name, sheet=ws.title,
                cell=cell, source_value=m.excluded_value, output_value=m.adopted_sum,
                next_step=f"合計行({cell}) の値 {inspection.fmt_num(m.excluded_value)} が"
                          f"明細の和 {inspection.fmt_num(m.adopted_sum)} と合いません。"
                          f"リンクをクリックして {path.name} の {cell} を確認してください"
                          "（除外そのものは維持しています）。"))
        if sheet_fallback:
            findings.append(inspection.finding(
                kind=inspection.KIND_SHEET_FALLBACK, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row),
                source_value=sheet_fallback[0], output_value=sheet_fallback[1],
                next_step=f"基準名のシート『{sheet_fallback[0]}』が見つからないため、"
                          f"1枚目『{sheet_fallback[1]}』を使いました。"
                          "意図したシートか確認してください。"))
        if col_a_mismatch:
            findings.append(inspection.finding(
                kind=inspection.KIND_COL_A_MISMATCH, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row + 1),
                source_value=col_a_mismatch[0], output_value=col_a_mismatch[1],
                next_step=f"A列走査 {col_a_mismatch[0]} 行と used range {col_a_mismatch[1]} 行が"
                          "一致しません。空の書式だけが残った行が無いか確認してください。"))

        return FileStackResult(name=path.name, status="積んだ", reordered=reordered, rows=rows,
                                excluded=verdict.excluded, mismatches=verdict.mismatches,
                                col_a_mismatch=col_a_mismatch, sheet_fallback=sheet_fallback,
                                findings=findings)
    finally:
        wb.close()


def fmt_num(v) -> str:
    """数値表示: 整数値は小数点なしで（650.0 でなく 650）。"""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    return str(int(f)) if f.is_integer() else str(f)
