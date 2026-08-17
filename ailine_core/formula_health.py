"""formula_health — 宣言つき挙動変更 #1: 型破壊の安全網。

★ ブラインド査定の実測: `原価` 列（数値）を SET_COLUMN_VALUE で文字列 `"0円"` に一括書換
すると、その列を参照する数式（例: `=D2-C2`）が `#VALUE!` に壊れる。なのに事後条件
チェッカー(check_set_column_value)は「対象列が指定文字列 N 件になったか」だけを見ており
「✓ 達成を機械検証済み」が出ていた（波及した数式破壊を誰も見ていない）。

(a) 一般の網: 適用の前後でエラー値セル(#VALUE!/#REF!/#DIV/0!/#N/A 等)を比較し、新たに
    エラーになったセルがあれば助言を返す。型破壊に限らずあらゆる波及被害を対象にする
    （査定者の言う「一番怖い種類の事故」への対抗・こちらが本命）。
(b) 狙い撃ちの関所……ではなく助言: OP_WRITE_TARGET が宣言する書き込み先列で、数値だった
    セルが数値に見えない文字列に変わった場合に予防的な確認行を返す。★ 関所（対話確認/
    非対話は停止）にはしていない ── 型が変わること自体は「税抜き列を税込みへ書き換える」
    のような正常な用途でも起き得るし、"0円" のように非数値文字列を意図的に書きたい
    場面もあるため、機械が意図まで断定できない（誤検知が出やすい・査定の+$30要望その2への
    予防止まりの応答という設計判断）。実害（数式が実際に壊れたか）は (a) が別途拾う。

★ ailine.py の行数は tests/ailine_py_line_budget.txt で凍結済み（増える方向の更新は禁止・
tests/test_line_budget.py 参照）。新しいコードは C4〜C7 の他モジュールと同じ理由でここに
置く。OP_WRITE_TARGET/`_is_number` 等 ailine.py 側の宣言・判定は、モジュール読み込み時点
ではなく呼び出し時点で引数として受け取る（dsl_step.DslStepDeps と同じ理由 ── テストの
monkeypatch より先に関数参照を固定しない）。
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable

import openpyxl

# OOXML のエラー値。openpyxl は data_only=True で開くと、LibreOffice/Excel が保存時に
# 計算・保存したキャッシュ値をそのまま文字列として返す（エラー値も同じ経路で読める）。
_ERROR_VALUES = frozenset({
    "#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!", "#CALC!",
})


def _error_cells(path: Path) -> dict:
    """path をキャッシュ値(data_only)で開き、エラー値を持つセルを
       {(sheet, row, col): エラー文字列, ...} で返す。式そのものは openpyxl が計算しない
       ため、ここで見えるのは basrun_apply（LibreOffice 経由で保存）を通した後のファイル
       だけ。ファイルが開けない/壊れている場合は空 dict（保守的＝誤検知回避）。"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return {}
    out: dict = {}
    try:
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and v in _ERROR_VALUES:
                        out[(name, cell.row, cell.column)] = v
    finally:
        wb.close()
    return out


def formula_error_advisory(before_path: Path, after_path: Path, *, cell_ref: Callable) -> list:
    """(a) 適用前後でエラー値セルを比較し、新たにエラーになったセル（前はエラーでなかった・
       今はエラー）が1件でもあれば助言を返す。
       ★ 単純な総数比較ではなくセル単位の新規発生で見る ── 総数が同じでも「あるセルが
       壊れ・別のセルが直った」場合は実質は別の問題であり、総数増減だけの比較では見逃す
       （ブリーフの「個数を数え、増えていたら報告」の趣旨をセル単位に強めた設計判断）。
       cell_ref: (row, col) -> "B2" のような表示用文字列に変換する関数（ailine.py の
       `_cell_ref` を渡す想定・表示フォーマットを二重管理しない）。"""
    before_errors = _error_cells(before_path)
    after_errors = _error_cells(after_path)
    new_breaks = {k: v for k, v in after_errors.items() if k not in before_errors}
    if not new_breaks:
        return []
    items = sorted(new_breaks.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2]))
    shown = items[:5]
    refs = "、".join(f"{sheet}!{cell_ref(r, c)}={val}" for (sheet, r, c), val in shown)
    more = f"、ほか{len(items) - 5}件" if len(items) > 5 else ""
    return [f"★ 疑わしい: 適用後にエラー値のセルが増えました（計{len(items)}件）: {refs}{more}"]


def _parses_as_number(s: str) -> bool:
    """文字列が数値そのものの見た目か（桁区切りのカンマは許容）。"105" は数値扱い・
       "0円" は非数値扱い ── SET_COLUMN_VALUE は常に setString で書くため、書いた中身が
       数字そのものか単位付き等の非数値文字列かで型変化の実害の芽の有無を分ける。"""
    try:
        float(s.replace(",", ""))
        return True
    except (ValueError, TypeError, AttributeError):
        return False


def detect_write_target_type_change(before: dict, after: dict, *, op: str | None, resolved: dict | None,
                                     meta: dict | None, op_write_target: dict, is_number: Callable) -> str | None:
    """(b) OP_WRITE_TARGET が宣言する書き込み先列（狙い撃ち）で、変化前が数値・変化後が
       数値に見えない文字列のセルが1件でもあれば予防の確認行を返す（関所にはしない・
       理由はモジュール docstring 参照）。新規列（元の型という概念が無い）は対象外。
       ★ 単位C: op_write_target の値は `.col_key` / `.sheet_key` を持つ宣言オブジェクト
       （呼び出し側が渡す・ここは属性を読むだけで ailine を import しない）。"""
    if not (op and resolved is not None and meta is not None):
        return None
    write_target = op_write_target.get(op)
    if not write_target or not write_target.col_key:
        return None
    col_key, sheet_key = write_target.col_key, write_target.sheet_key
    col_name = resolved.get(col_key)
    if not col_name:
        return None
    sheet = resolved.get(sheet_key) if sheet_key else next(iter(meta.get("sheets") or []), None)
    if not sheet or sheet not in before.get("sheets", []):
        return None
    headers = meta.get("headers", {}).get(sheet, [])
    if col_name not in headers:
        return None
    col_idx = headers.index(col_name) + 1
    prefix = sheet + "!"
    flips = 0
    for k, b in before["cells"].items():
        if not k.startswith(prefix):
            continue
        _, rc = k.split("!", 1)
        _, c_str = rc.split(",")
        if int(c_str) != col_idx:
            continue
        if not is_number(b[0] if b else None):
            continue
        a = after["cells"].get(k)
        a_val = a[0] if a else None
        if isinstance(a_val, str) and not _parses_as_number(a_val):
            flips += 1
    if flips == 0:
        return None
    return (f"（確認）列『{col_name}』は元は数値でしたが、{flips} 件のセルが数値に見えない"
            f"文字列に変わりました。この列を参照する数式があれば壊れていないか確認してください")
