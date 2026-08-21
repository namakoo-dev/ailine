"""extract_multi — M2 `ailine run <フォルダ> "<依頼>"`（抽出集約）の本体。
   DESIGN-20260821-multifile.md M2 節（Namakoo 決裁 2026-08-21 09:1x・A 案）。

   ★ ailine を import しない（tests/test_line_budget.py の移植可能性番人）── 見出し行の
   推定・翻訳・出力シート名の決め打ちは ailine.py 側（cmd_run_folder）が行い、その結果を
   値として渡す（cmd_scan / cmd_stack と同じ配線）。
   ★ 既存部品の再利用: 分母・基準ファイル方式・3判定は multifile.py、合計行の識別と
   閉じる検査は total_row.py。この module は「条件で選ぶ行の決定」と「出力ファイル名の
   sanitize」という M2 固有の配線だけを持つ。
   ★ predicate は ailine._extract_predicate の**独立再実装**（import しない）。同じ勘定を
   2箇所が違う実装で書き、tests/test_predicate_truth_table.py の手書きの表を校正原器に
   両方を突き合わせる（xml_readback.py の docstring と同じ作法）。
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field

import openpyxl

from ailine_core import inspection, multifile, total_row

# ★ 書き手の印。stack.py の CREATOR_MARKS が「将来の extract を先取りして凍結」した
#   文字列と完全一致させる（own_output_mark / is_own_output が M2 の出力を
#   『ailine 産』と正しく認めるための唯一の接点）。
CREATOR_MARK = "ailine extract"

# eq の許容誤差（tests/test_predicate_truth_table.py が凍結した意味論）。
TOLERANCE = 1e-6

# --- ファイル名の sanitize（★ Excel の**シート名**規則とは別物 ---------------
# シート名は ailine._extract_output_sheet_name（: \ / ? * [ ] と 31 文字上限）。
# こちらは Windows の**ファイル名**規則（< > : " / \ | ? * と制御文字・末尾の . と空白・
# 予約デバイス名）。同じ「名前を安全にする」でも禁止集合が違うので寄せない。
_FILENAME_FORBIDDEN_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_RESERVED_NAMES = ({"CON", "PRN", "AUX", "NUL"}
                   | {f"COM{i}" for i in range(1, 10)}
                   | {f"LPT{i}" for i in range(1, 10)})
_MAX_STEM = 100          # 拡張子と親パスの余裕を見た保守的な上限（MAX_PATH 対策）
_HASH_LEN = 6            # 切り詰め衝突時に足す sha256 の桁数


def _is_number(v) -> bool:
    """bool は int のサブクラスだが数値としては扱わない（total_row._is_number と同じ線）。"""
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def predicate(cmp: str, threshold):
    """EXTRACT の判定（★ ailine._extract_predicate の独立再実装・仕様は
       tests/test_predicate_truth_table.py の凍結した表）。

       - gte/lte/gt/lt: 本物の数値（int/float・bool 除く）にだけ効く。文字列数値・日付・
         None・bool は不一致（黙って型変換しない）
       - eq: 条件値が数値なら許容誤差 TOLERANCE で数値比較。数値でなければ文字列の完全一致
       - contains: 文字列セルのみ（数値セルを文字列化しない）
    """
    threshold_is_number = _is_number(threshold)

    def _match(cell_value) -> bool:
        if cmp == "contains":
            if threshold is None or not isinstance(cell_value, str):
                return False
            return str(threshold) in cell_value
        if cmp == "eq":
            if threshold_is_number:
                if not _is_number(cell_value):
                    return False
                return abs(float(cell_value) - float(threshold)) <= TOLERANCE
            return str(cell_value) == str(threshold)
        if not _is_number(cell_value) or not threshold_is_number:
            return False
        v, t = float(cell_value), float(threshold)
        if cmp == "gte":
            return v >= t
        if cmp == "lte":
            return v <= t
        if cmp == "gt":
            return v > t
        if cmp == "lt":
            return v < t
        return False
    return _match


def sanitize_filename(stem: str) -> str:
    """出力ファイル名の幹（拡張子なし）を Windows で作れる形に直す。
       - 禁止文字（< > : " / \\ | ? * と制御文字）は '_' に置換
       - 末尾の '.' と空白を剥がす（エクスプローラが作れない形）
       - 予約デバイス名（CON/PRN/AUX/NUL/COM1-9/LPT1-9・大小問わず）は '_' を足して回避
       - 長すぎる名前は _MAX_STEM で切り詰める

       ★ review3#1/#5 の直し（実機再現: 長いフォルダ名で別条件の出力が同名に潰れ、
       1回目の結果が無警告消去された）: 切り詰めが起きた時は**常に**元の（切り詰め前の）
       名前 全体 の sha256 先頭 6 桁を付ける ── 呼び出し側が taken（既に使った名前）の
       集合を渡してくれることに頼らない（配線されない対策コードを持たない）。フォルダ名が
       枠を使い切る形（切り詰め前の共通部分が _MAX_STEM を超える）でも、条件の違いは
       ハッシュの違いとして必ず残るので構造的に潰れない。
       ★ 切り詰めが起きなければ従来どおり素の名前（既存呼び出し元の期待を変えない）。"""
    original = str(stem)
    s = _FILENAME_FORBIDDEN_RE.sub("_", original)
    s = s.rstrip(". 　")
    if not s:
        s = "ailine_extract"
    if s.upper() in _RESERVED_NAMES or s.upper().split(".")[0] in _RESERVED_NAMES:
        s = f"{s}_"
    if len(s) > _MAX_STEM:
        digest = hashlib.sha256(original.encode("utf-8")).hexdigest()[:_HASH_LEN]
        head = s[:max(1, _MAX_STEM - _HASH_LEN - 1)].rstrip(". 　") or "ailine_extract"
        s = f"{head}_{digest}"
    return s


def _row_has_any_value(ws, row: int, num_cols: int) -> bool:
    """row の 1..num_cols 列のどこかに値があるか（stack._row_has_any_value と同じ線 ──
       条件列が空欄でも他の列に値があれば『データ行』＝会計の分母に数える）。"""
    return any(not total_row._is_blank_cell(ws.cell(row=row, column=c).value)
               for c in range(1, num_cols + 1))


@dataclass(frozen=True)
class FileExtractResult:
    """1ファイルから抽出した（または抽出できなかった）結果。
       rows: [(base_headers 順の値リスト, 同順の number_format リスト, 元行番号), ...]（★ 一致した
       行だけ・number_format は日付セルの時刻の尻尾を消すため元セルから運ぶ）。
       rows_matched / rows_unmatched: 合計行の除外を引いた後の候補行の内訳
       （★ 憲法⑨ 行の完全会計: データ行数 = 一致 + 不一致 + 除外）。
       excluded/mismatches: total_row.split_total_rows の戻り値そのまま。"""
    name: str
    status: str                      # "取れた" / "取れなかった"
    reason: str | None = None
    reordered: bool = False
    rows: list = field(default_factory=list)
    rows_matched: int = 0
    rows_unmatched: int = 0
    excluded: list = field(default_factory=list)
    mismatches: list = field(default_factory=list)
    sheet_fallback: tuple | None = None   # (wanted, used) ── 基準名のシートが無く1枚目へ落ちた時だけ
    findings: list = field(default_factory=list)   # list[inspection.Finding]（M2.5・stack と同じ線）


def evaluate_and_extract(path, base_headers: list, base_sheet_name, header_row: int,
                          cond_col_name: str, cmp: str, value) -> FileExtractResult:
    """1ファイルを基準と照合し、取れていれば条件に一致する行を確定して値まで読む。
       ★ どんな失敗でも例外を上げず名指し+理由で返す（multifile.evaluate_file と同じ線）。
       ★ E3（M2 検体の筆頭）: 合計行の除外（単位L）は**条件の適用より前**に回す ──
       『金額 40000 以上』のような抽出では合計行が必ず条件を満たすため、順序を誤ると
       合計行が一致行として混ざり、そのまま二重計上になる。
       ★ 合計行検出の数値列は multifile.numeric_value_column の『最初の数値列』ではなく
       **条件列そのもの**を使う（M2 の選択）── 検算側（verify.verify_extract）と同じ列で
       同じ除外を再現できないと、書いた側と検算側が食い違う。"""
    if path.suffix.lower() != ".xlsx":
        return FileExtractResult(name=path.name, status="取れなかった", reason="旧形式(.xls)")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return FileExtractResult(name=path.name, status="取れなかった", reason=f"読み込み失敗: {e}")
    try:
        ws, sheet_fell_back = multifile.find_matching_sheet(wb, base_sheet_name)
        sheet_fallback = (base_sheet_name, ws.title) if sheet_fell_back else None
        other_headers = multifile.read_row_headers(ws, header_row)
        status, detail = multifile.classify_headers(base_headers, other_headers)
        if status == "取れなかった":
            not_taken = [inspection.finding(
                kind=inspection.KIND_NOT_TAKEN, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row), source_value=None, output_value=None,
                next_step=f"見出しが基準と合いません（{detail}）。この冊は対象外です"
                          "（抽出していません）。")]
            return FileExtractResult(name=path.name, status="取れなかった", reason=detail,
                                      sheet_fallback=sheet_fallback, findings=not_taken)

        col_for_base = {bh: multifile._column_index(other_headers, bh) for bh in base_headers}
        max_row = ws.max_row or header_row
        all_rows = list(range(header_row + 1, max_row + 1))

        label_col = col_for_base.get(base_headers[0]) if base_headers else None
        value_col = col_for_base.get(cond_col_name)
        if label_col and value_col:
            triples = [(r, ws.cell(row=r, column=label_col).value,
                        ws.cell(row=r, column=value_col).value) for r in all_rows]
            verdict = total_row.split_total_rows(triples)
        else:
            verdict = total_row.TotalRowVerdict(excluded=[], adopted_rows=[], mismatches=[])
        excluded_rows = {e.row for e in verdict.excluded}

        num_cols = len(other_headers)
        data_rows = [r for r in all_rows if _row_has_any_value(ws, r, num_cols)]
        candidate_rows = [r for r in data_rows if r not in excluded_rows]

        match_fn = predicate(cmp, value)
        matched_rows, unmatched_rows = [], []
        for r in candidate_rows:
            cell = ws.cell(row=r, column=value_col).value if value_col else None
            (matched_rows if match_fn(cell) else unmatched_rows).append(r)

        rows = []
        for r in matched_rows:
            cells = [ws.cell(row=r, column=col_for_base[bh]) for bh in base_headers]
            values = [c.value for c in cells]
            # ★ 実視の磨き（2026-08-21）: 元セルの number_format をデータセルへ運ぶ
            #   （日付が『2026-07-09 0:00:00』と時刻付きで出ないように）。決定論
            #   ── 元の書式文字列をそのまま運ぶだけ（乱数・時刻は使わない）。
            formats = [c.number_format for c in cells]
            rows.append((values, formats, r))

        # ★ M2.5: 所見の組み立て（stack.evaluate_and_stack と同じ線 ── ws がまだ開いている
        #   この関数の内側でだけ列位置まで正確な3座標が引ける）。
        # ★ アンカーは「怪しい数字そのもの」（閉じなかった合計の値セル）。クリックの
        #   着地点に迷いを作らない（検分シートの本文が両側の数字を持つのは別の役割）。
        # ★ UX 磨き③（Namakoo 実視 2026-08-21 12:01）: 断片でなく1所見1文
        #   （stack.evaluate_and_stack と同じ文形・inspection.fmt_num で両側の数字を言う）。
        anchor_col = value_col or label_col or 1
        findings = []
        for m in verdict.mismatches:
            cell = inspection.cell_ref(anchor_col, m.row)
            findings.append(inspection.finding(
                kind=inspection.KIND_TOTAL_ROW_MISMATCH, file=path.name, sheet=ws.title,
                cell=cell, source_value=m.excluded_value, output_value=m.adopted_sum,
                next_step=f"合計行({cell}) の値 {inspection.fmt_num(m.excluded_value)} が"
                          f"明細の和 {inspection.fmt_num(m.adopted_sum)} と合いません。"
                          f"リンクをクリックして {path.name} の {cell} を確認してください"
                          "（除外そのものは維持しています）。"))
        if sheet_fallback:
            findings.append(inspection.finding(
                kind=inspection.KIND_SHEET_FALLBACK, file=path.name, sheet=ws.title,
                cell=inspection.cell_ref(1, header_row),
                source_value=sheet_fallback[0], output_value=sheet_fallback[1],
                next_step=f"基準名のシート『{sheet_fallback[0]}』が見つからないため、"
                          f"1枚目『{sheet_fallback[1]}』を使いました。"
                          "意図したシートか確認してください。"))

        return FileExtractResult(name=path.name, status="取れた", reordered=bool(detail),
                                  rows=rows, rows_matched=len(matched_rows),
                                  rows_unmatched=len(unmatched_rows),
                                  excluded=verdict.excluded, mismatches=verdict.mismatches,
                                  sheet_fallback=sheet_fallback, findings=findings)
    finally:
        wb.close()
