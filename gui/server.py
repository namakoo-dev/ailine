"""ailine の GUI ── **薄い殻**。判定は 1 ビットも作らない。

★ 設計の縛り（これが守れなくなったら GUI ごと捨てる）:
  画面に出る ✓ / △ / ⚠ は、`ailine run --json` が返す `verdict` を**そのまま映す**。
  ここで advisories を数え直したり、postcondition から印を導いたりしない ──
  それは 2 つ目の実装で、この repo が 2026-08 の盲検で 4 回踏んだ欠陥
  （「検算の分母が、疑うべき対象と同じ所から作られる」）を自分で新造することになる。
  ★ 決めるのは 1 箇所（ailine 本体）・映すのは何箇所でも。

★ 依存を足さない: Python 標準ライブラリだけ（scripts/ci_parity.py の番人が守る）。
★ localhost のみに bind する（外に開かない）。
"""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
import threading
import webbrowser
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
PAGE = HERE / "index.html"
RUN_TIMEOUT = 300


def _ailine(args: list) -> tuple:
    """ailine を子プロセスで動かす。戻り値 (returncode, stdout, json or None)。

    ★ import して呼ばない: 画面の都合で本体の内部状態を触らないため
      （殻はあくまで外から使う人と同じ入口を叩く）。
    """
    cmd = [sys.executable, "-m", "ailine"] + args
    # ★★ 実測（2026-08-26）: これを付けないと、子プロセスは **site-packages に入っている
    #   古い版**を import する（cwd に `ailine/` が無いので `-m` が拾えない）。
    #   盲検 2 回目で検分者が古いタグを測ってしまったのと同じ形の事故を、
    #   今度は自分の画面が起こすところだった ── **殻は必ずこの repo の本体を叩く**。
    import os as _os
    env = dict(_os.environ)
    src = str(REPO / "src")
    env["PYTHONPATH"] = src + _os.pathsep + env.get("PYTHONPATH", "")
    env.setdefault("PYTHONUTF8", "1")
    proc = subprocess.run(cmd, cwd=str(REPO), capture_output=True, text=True,
                           encoding="utf-8", errors="replace", timeout=RUN_TIMEOUT, env=env)
    out = (proc.stdout or "") + (proc.stderr or "")
    payload = None
    for line in (proc.stdout or "").splitlines():
        s = line.strip()
        if s.startswith("{") and s.endswith("}"):
            try:
                payload = json.loads(s)
            except json.JSONDecodeError:
                pass
    return proc.returncode, out, payload


# ★ 下書きの続き（本 → 作業中の下書きファイル）。段取りだけを持つ・判定は持たない。
_DRAFTS: dict = {}
_LAST_MULTI: dict = {}
_DRAFT_LAST_TASK: dict = {}
DRAFT_SUFFIX = "（下書き）"


def _draft_path(book: Path) -> Path:
    """下書きの置き場。★ 名前を積み重ねない（.out.out.out… にしない）。"""
    if book.stem.endswith(DRAFT_SUFFIX):
        return book
    return book.with_name(book.stem + DRAFT_SUFFIX + book.suffix)

def _native_dialog(kind: str, start: str = "", name: str = "") -> str:
    """本物のエクスプローラを開く。**サーバと画面が同じ機械**だから成り立つ。

    ★ なぜサーバ側で開くか: ブラウザの `<input type=file>` は**完全なパスを返さない**
      （セキュリティ上そうなっている）。この道具は原本の場所を知らないと何もできない
      ので、パスが要る。localhost に閉じた作りだから、素直に OS のダイアログを使う。
    ★ tkinter は標準ライブラリ（新しい依存は増えない ── 番人が確認する）。
    ★ 前面に出す小細工が要る（ブラウザの裏に出ると「固まった」に見える）。
    """
    import tkinter
    from tkinter import filedialog
    root = tkinter.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()
    try:
        if kind == "folder":
            got = filedialog.askdirectory(title="フォルダを選ぶ", initialdir=start or None)
        elif kind == "save":
            got = filedialog.asksaveasfilename(
                title="名前を付けて保存", initialdir=start or None,
                initialfile=name or None, defaultextension=".xlsx",
                filetypes=[("Excel ブック", "*.xlsx"), ("すべて", "*.*")])
        else:
            got = filedialog.askopenfilename(
                title="ファイルを選ぶ", initialdir=start or None,
                filetypes=[("Excel ブック", "*.xlsx"), ("すべて", "*.*")])
    finally:
        root.destroy()
    return got or ""


MAX_ROWS = 200
MAX_COLS = 40


def _read_sheet(path: Path, sheet: str | None) -> dict:
    """表を画面に出すために読む。**判定はしない** ── 見せるだけ。

    ★ 読み手は ailine 自身の独立読み（ailine_core.xml_readback）を使う。
      ここで新しい読み実装を書けば、画面が「本体とは違うもの」を見せうる
      ── 2026-08-26 に、この repo の 2 つの読み実装が食い違っていて
      `9:00` が `1899-12-30T09:00:00` になっていたのを実測している。
      両者が一致することは tests/test_readers_agree.py が機械で守っている。
    ★ 大きい表は途中で切るが、**切ったことは必ず言う**（出ないことは信号でない）。
    """
    try:
        sys.path.insert(0, str(REPO / "src"))
        from ailine_core import xml_readback
        import openpyxl
    except Exception as e:
        return {"error": f"読み手を用意できません: {e}"}
    try:
        # ★★ 2026-08-27（Namakoo が実測・俺が連れてきたバグ）:
        #   read_only のブックは **close しないとファイルハンドルが残る**。
        #   実行の直前に「前」を撮るようにしたら、その掴んだままのハンドルで
        #   下書きを置換できなくなり、**毎回 exit 9 で失敗**していた
        #   （画面には「できていない」としか見えない）。
        #   ★ 見るために開いたものは、見終わったら必ず閉じる。
        _wb0 = openpyxl.load_workbook(path, read_only=True)
        names = list(_wb0.sheetnames)
        _wb0.close()
    except Exception as e:
        return {"error": f"開けません: {type(e).__name__}: {e}"}
    try:
        data = xml_readback.read_grid(path, sheet_name=sheet)
    except Exception as e:
        return {"error": f"読めません: {type(e).__name__}: {e}", "sheets": names}
    grid = data["grid"]
    rows_n, cols_n = data["max_row"], data["max_col"]
    shown_r, shown_c = min(rows_n, MAX_ROWS), min(cols_n, MAX_COLS)
    # ★ 2026-08-26（Namakoo の指摘）: 値だけ出すと**中央揃え・太字・背景色・罫線が
    #   何も見えない**。24 op のうち 5 つが見た目の操作なのに、確かめる手段が無かった。
    #   → 書式も一緒に運ぶ（openpyxl は既に製品の依存・新しい依存は増やさない）。
    #   ★ 判定はしない。見せるだけ。
    style = _read_styles(path, data["sheet_name"], shown_r, shown_c)
    style_error = style.pop(("error",), None)
    cells = [[{"t": _cell_text(grid.get((r, c))), **style.get((r, c), {})}
               for c in range(1, shown_c + 1)]
              for r in range(1, shown_r + 1)]
    note = ""
    if rows_n > shown_r or cols_n > shown_c:
        note = f"表が大きいので {shown_r} 行 × {shown_c} 列だけ出しています（実際は {rows_n} 行 × {cols_n} 列）"
    return {"sheets": names, "sheet": data["sheet_name"], "cells": cells,
            "rows": rows_n, "cols": cols_n, "note": note,
            "style_error": style_error,
            "shapes": _count_shapes(path, data["sheet_name"]),
            "uncached": len(data.get("uncached_formulas") or ())}


def _count_shapes(path: Path, sheet: str) -> dict:
    """グラフ・画像の数。★ 表には**絶対に映らない**ので、数だけでも言う。

    ★ 「消えたものは差分に出ない」の家系: グラフが増えた/消えたことは値の表に一切出ない。
      画面が黙れば、人は「何も起きなかった」と読む。
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        n = {"charts": len(getattr(ws, "_charts", []) or []),
             "images": len(getattr(ws, "_images", []) or [])}
        wb.close()
        return n
    except Exception:
        return {}


def _read_styles(path: Path, sheet: str, rows: int, cols: int) -> dict:
    """セルの見た目（太字・寄せ・背景色・数値書式・罫線）を読む。**判定はしない**。"""
    out: dict = {}
    try:
        # ★ 実測で踏んだ: ここで import し忘れ、NameError が下の except に握り潰されて
        #   **書式が黙って出なくなっていた**（ファイルには太字も #,##0 も入っていた）。
        #   ★ 握り潰す except は「出ないことを信号にする」── 気づけない形だった。
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    except Exception as e:
        out[("error",)] = str(e)
        return out
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            info = {}
            if cell.font is not None and cell.font.bold:
                info["b"] = 1
            al = getattr(cell.alignment, "horizontal", None)
            if al:
                info["a"] = al
            fill = getattr(cell.fill, "start_color", None)
            rgb = getattr(fill, "rgb", None) if fill else None
            if isinstance(rgb, str) and rgb not in ("00000000",) and rgb.endswith(("FFFFFF",)) is False:
                info["g"] = rgb[-6:]
            nf = cell.number_format
            if nf and nf != "General":
                info["n"] = nf
            b = cell.border
            if b and any(getattr(getattr(b, side), "style", None)
                          for side in ("left", "right", "top", "bottom")):
                info["k"] = 1
            if info:
                out[(r, c)] = info
    try:
        wb.close()
    except Exception:
        pass
    return out


def _cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


def _default_dir() -> Path:
    """最初に開くフォルダ。同梱のサンプルがあればそこ、無ければ repo 直下。"""
    for cand in (REPO / "demo_gui", REPO / "demo"):
        if cand.is_dir():
            return cand
    return REPO


class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a):        # 端末を汚さない
        pass

    def _send(self, code: int, body: bytes, ctype: str) -> None:
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _json(self, code: int, obj: dict) -> None:
        self._send(code, json.dumps(obj, ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8")

    def _url(self):
        """★ http.server は要求行を latin-1 で str にするので、日本語のパスを
           クエリに乗せると壊れる（実測: `1_売上.xlsx` が読めなかった）。
           バイト列へ戻してから UTF-8 として読み直す。"""
        raw = self.path
        try:
            raw = raw.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
        return urlparse(raw)

    def do_GET(self):
        u = self._url()
        if u.path in ("/", "/index.html"):
            self._send(200, PAGE.read_bytes(), "text/html; charset=utf-8")
            return
        if u.path == "/api/ops":
            # ★ 「何が頼めるか」を画面に出す。CLI の `ailine ops` と同じものを、
            #   同じ本体から取る（画面が一覧を持たない ── 増えた op が勝手に落ちない）。
            rc, out, _payload = _ailine(["ops"])
            self._json(200, {"rc": rc, "text": out})
            return
        if u.path == "/api/aliases":
            rc, out, _payload = _ailine(["alias", "list"])
            self._json(200, {"rc": rc, "text": out})
            return
        if u.path == "/api/pick":
            q = parse_qs(u.query)
            kind = (q.get("kind") or ["file"])[0]
            start = (q.get("start") or [""])[0]
            name = (q.get("name") or [""])[0]
            try:
                got = _native_dialog(kind, start, name)
            except Exception as e:
                self._json(200, {"path": "", "error": f"ダイアログを開けません: {e}"})
                return
            self._json(200, {"path": got})
            return
        if u.path == "/api/sheet":
            q = parse_qs(u.query)
            path = Path((q.get("path") or [""])[0])
            want = (q.get("sheet") or [None])[0]
            self._json(200, _read_sheet(path, want))
            return
        if u.path == "/api/files":
            q = parse_qs(u.query)
            folder = Path((q.get("dir") or [str(_default_dir())])[0]).expanduser()
            try:
                folder = folder.resolve()
                items = sorted((p for p in folder.iterdir()
                                 if p.is_file() and p.suffix.lower() == ".xlsx"
                                 and not p.name.startswith("~$")),
                                key=lambda q: q.name)
            except OSError as e:
                self._json(200, {"dir": str(folder), "files": [], "error": str(e)})
                return
            # ★ 画面側でパスを**組み立てさせない**（区切り文字の二重化を実測で踏んだ）。
            #   完全なパスはここで作って、そのまま返す ── 組み立てが 1 箇所なら、ずれない。
            # ★ ailine 自身が作った下書き（.out.xlsx）は**隠さず、そう見せる**
            #   （一覧から消すと「無い」と読まれる ── 出ないことは信号でない）。
            self._json(200, {"dir": str(folder),
                              "files": [{"name": p.name, "path": str(p),
                                          "draft": p.name.endswith(".out.xlsx")}
                                         for p in items]})
            return
        self._json(404, {"error": "not found"})

    def do_POST(self):
        u = urlparse(self.path)
        try:
            length = int(self.headers.get("Content-Length") or 0)
            req = json.loads(self.rfile.read(length) or b"{}")
        except Exception as e:
            self._json(400, {"error": f"読めない要求: {e}"})
            return
        book = str(Path(req.get("book") or "").expanduser())
        if not book:
            self._json(400, {"error": "ファイルが選ばれていません"})
            return
        try:
            if u.path == "/api/run":
                # ★★ 2026-08-26（Namakoo が実測）: 「梨を追加して」→「梨の売上を2000に」と
                #   続けると、**1 つ目の操作が消えた**。
                #   根: 下書きは毎回 `run <原本> --copy` なので、**毎回原本から作り直す**。
                #   人は「1 つやって、次をやる」と積み上げるのに、道具は積み上げない。
                #   ★ 直し: 下書きは**作業用のファイルを 1 つ持ち続ける**。
                #     1 回目は原本から作り（--copy）、2 回目からはその下書き自身に反映する。
                #     下書きにもバックアップと undo が効く（普通のブックと同じ扱い）。
                #   ★ ここは段取りであって判定ではない（verdict は本体の値をそのまま運ぶ）。
                task = req.get("task") or ""
                if not req.get("copy"):
                    _before = _read_sheet(Path(book), None)
                    rc, out, payload = _ailine(["run", book, task, "--json"])
                    _DRAFTS.pop(book, None)     # 原本に反映したら下書きの役目は終わり
                    self._json(200, {"rc": rc, "text": out, "json": payload,
                                      "before": _before, "target": book})
                    return
                else:
                    # ★★ 2026-08-26（Namakoo が実測・俺の直しが連れてきたバグ）:
                    #   初版は `--copy` の成果物（<名前>.out.xlsx）に反映し続けた。すると
                    #   作業ファイルが <名前>.out.out.xlsx になり、失敗した回の残骸が残って
                    #   **次から関所に塞がれ、行き止まりになった**。
                    #   ★ 名前を積み重ねる設計が間違い。下書きは**人が読める名前のファイル
                    #     1 つ**にして、そこへ普通に反映する（backup も undo も普通に効く）。
                    # ★★ 2026-08-27（Namakoo が 2 度実測）: 昨夜「積み上げ」にしたのは
                    #   「梨を追加」→「梨の売上を2000に」で 1 つ目が消えるのを直すため
                    #   だった。だがそれで **試すたびに増える**という別の困りごとを作った。
                    #   ★ 分けるべきだったのは「試す」と「積む」── 既定は**原本から**。
                    #     積むのは人が明示した時だけ（continue=true）。
                    #     VS Code 系の作法（編集をその場で確定させず人が受け入れる）と同じ線。
                    draft = _draft_path(Path(book))
                    # ★★ 2026-08-27（3 度目の設計変更・Namakoo が正した）:
                    #   08-26 夜 積み上げ → 08-27 昼 原本から → 今ここ。
                    #   「原本から」にしたら、**続きの依頼で前の結果が消えた**
                    #   （「梨を追加」→「梨の売上を2000に」で梨の行ごと消える）── 08-26 夜と
                    #   同じ困りごとを俺が作り直した。
                    #   ★ 正しい形: **逐次保存（積み上げ）が既定**。繰り返して増えた分は
                    #     **undo で 1 つ戻す** ── 直し方は「やり直さない」ではなく「戻せる」。
                    #     やり直したい時だけ「原本からやり直す」を押す。
                    _cont = not bool(req.get("restart"))
                    _note = ""
                    if not _cont or not draft.exists():
                        if draft.exists() and draft.read_bytes() != Path(book).read_bytes():
                            # ★ 前の下書きを黙って捨てない ── 1 つだけ退避して残す。
                            keep = draft.with_name(draft.stem + "（前回）" + draft.suffix)
                            try:
                                if keep.exists():
                                    keep.unlink()
                                draft.replace(keep)
                                _note = f"（下書きを原本から作り直しました。前の下書きは {keep.name} に残しています）"
                            except OSError:
                                _note = ""
                        shutil.copy2(book, draft)
                    _DRAFTS[book] = str(draft)
                    # ★★ 2026-08-27（Namakoo が実測・俺の画面の壊れ方）:
                    #   「操作する前」に**原本**を、「操作したあと」に**下書き**を出していた。
                    #   別のファイル同士を並べていたので、差分も色も意味を成さない
                    #   ── 正しく動いた操作が「できていない」に見えた。
                    #   ★ 前後は**同じファイル**で撮る。実行の直前にここで読む（画面が
                    #     後から取りに行くと、もう変わっている）。
                    _before = _read_sheet(draft, None)
                    # ★★ 2026-08-27（Namakoo「一つ追加すればいいのに複数追加してしまう」）:
                    #   実測すると 1 回の実行は 1 行しか足していなかった ── 押した回数だけ
                    #   増えていた（下書きは積み上げる設計にしたので、それ自体は正しい）。
                    #   ★ 壊れていたのは**動作でなく、積み上げが見えないこと**。
                    #     同じ依頼を繰り返したら、そう言う（黙って 2 つ目を作らない）。
                    _DRAFT_LAST_TASK[str(draft)] = task
                    rc, out, payload = _ailine(["run", str(draft), task, "--json"])
                    if _note:
                        out = _note + chr(10) + out
                    if _cont:
                        out = ("（下書きに続けて保存しました ── 1 つ戻すなら「もとに戻す」）"
                                + chr(10) + out)
                    self._json(200, {"rc": rc, "text": out, "json": payload,
                                      "before": _before, "target": str(draft)})
                    return
            elif u.path == "/api/folder":
                # ★ 複数ファイルの経路（scan / stack / run <フォルダ>）。
                #   ★ 需要はここに寄っている（実測の需要地図: 上位 5 件中 4 件が複数ファイル）。
                #   ★ ここでも判定は作らない ── 本体の出力をそのまま運ぶ。
                folder = req.get("folder") or ""
                kind = req.get("kind") or "scan"
                if kind == "scan":
                    rc, out, payload = _ailine(["scan", folder])
                elif kind == "stack":
                    dst = req.get("dst") or str(Path(folder) / "縦積み.xlsx")
                    rc, out, payload = _ailine(["stack", folder, "--out", dst])
                    if rc == 0:
                        _LAST_MULTI["out"] = dst
                        _LAST_MULTI["src"] = folder
                elif kind == "extract":
                    rc, out, payload = _ailine(["run", folder, req.get("task") or ""])
                elif kind == "verify":
                    made = _LAST_MULTI.get("out")
                    if not made:
                        # ★ 判定は作らない（json は空のまま）。画面には言葉だけ返す。
                        self._json(200, {"rc": 1, "json": None,
                                          "text": "先に縦積みを作ってください（検算はその結果に対して行います）"})
                        return
                    rc, out, payload = _ailine(["verify", made, _LAST_MULTI.get("src") or folder])
                else:
                    self._json(400, {"error": "不明な操作"})
                    return
                self._json(200, {"rc": rc, "text": out, "json": payload,
                                  "made": _LAST_MULTI.get("out") if kind == "stack" else None})
                return
            elif u.path == "/api/save_as":
                # ★ 「保存」は**複製**であって移動ではない（元の物は残す）。
                #   下書きを育てたまま、清書だけ別名で外へ出せる形にする。
                src = Path(req.get("path") or book)
                dst = Path(req.get("dst") or "")
                if not src.exists():
                    self._json(200, {"rc": 1, "json": {"verdict": "not_applied"},
                                      "text": f"保存するものがありません: {src}"})
                    return
                try:
                    shutil.copy2(src, dst)
                    msg = f"{dst} に保存しました（{src.name} はそのまま残っています）"
                    code = 0
                except OSError as e:
                    msg, code = f"保存できませんでした: {e}", 1
                self._json(200, {"rc": code, "json": {"verdict": "not_applied"}, "text": msg})
                return
            elif u.path == "/api/clear_leftover":
                # ★★ 2026-08-26（Namakoo が 2 度実測）: 前の run が残した作業結果が
                #   出力先を塞ぎ、**人が手でファイルを消すまで先へ進めない**状態になった。
                #   ★ 製品は断ったままにする（人が置いたファイルを黙って消さない設計は正しい）。
                #     画面は、断り文が言っている「別の場所へ移すか削除してから」を
                #     **人が 1 手でできる形**にする ── 決めるのは人、実行だけを楽にする。
                #   ★ 中身を見せてから消す（何を捨てるか分からないまま押させない）。
                target = Path(req.get("path") or "")
                if not target.exists():
                    self._json(200, {"rc": 0, "json": {"verdict": "not_applied"},
                                      "text": "片づけるものはありません。"})
                    return
                trash = target.with_name(target.stem + "（捨てた）" + target.suffix)
                try:
                    if trash.exists():
                        trash.unlink()
                    target.rename(trash)      # ★ 消さずに改名する（取り返しを残す）
                    msg = (f"{target.name} を {trash.name} に改名しました"
                            "（消していません）。もう一度実行できます。")
                except OSError as e:
                    msg = f"片づけられませんでした: {e}"
                self._json(200, {"rc": 0, "json": {"verdict": "not_applied"}, "text": msg})
                return
            elif u.path == "/api/alias_add":
                # ★ 2026-08-27（Namakoo）: 「できない操作も登録まで誘導されないし、
                #   登録画面も実装されていない」── CLI には alias add が在るのに、
                #   画面からは一生辿り着けなかった。**在るのに届かない**は無いのと同じ。
                phrase, op = (req.get("phrase") or "").strip(), (req.get("op") or "").strip()
                if not phrase or not op:
                    self._json(200, {"rc": 1, "json": None,
                                      "text": "言い回しと操作の両方を指定してください"})
                    return
                rc, out, payload = _ailine(["alias", "add", phrase, op])
                self._json(200, {"rc": rc, "text": out, "json": None})
                return
            elif u.path == "/api/open":
                # ★ 表に映らないもの（グラフ・図形・印刷の見え方）は、本物のアプリで見るのが早い。
                #   画面で再現しようとすると**画面が 2 つ目の実装**になる ── 見せる相手は実物。
                target = req.get("path") or book
                try:
                    import os as _o
                    _o.startfile(str(Path(target)))     # Windows の既定アプリで開く
                    msg = f"{Path(target).name} を開きました（表に映らないものはこちらで）"
                except Exception as e:
                    msg = f"開けませんでした: {e}"
                self._json(200, {"rc": 0, "json": {"verdict": "not_applied"}, "text": msg})
                return
            elif u.path == "/api/draft_reset":
                # 下書きを捨てて、次は原本からやり直す（消しはしない ── 残す）
                dropped = _DRAFTS.pop(book, None)
                _DRAFT_LAST_TASK.pop(str(_draft_path(Path(book))), None)
                self._json(200, {"rc": 0, "json": {"verdict": "not_applied"},
                                  "text": (f"下書きを手放しました（{Path(dropped).name} は"
                                            "残っています）。次は原本から作り直します。"
                                            if dropped else "手放す下書きはありません。")})
                return
            elif u.path == "/api/undo":
                # ★★ 2026-08-27（Namakoo「1つ前の操作まで undo で戻れるように」）:
                #   画面の「もとに戻す」は**原本**を対象にしていた。原本には
                #   バックアップが無い（触っていないので当然）ので、何も戻せなかった。
                #   ★ 戻す相手は**いま触っている物** ── 下書きがあれば下書き。
                _target_undo = _DRAFTS.get(book) or book
                rc, out, payload = _ailine(["undo", _target_undo])
                out = f"（戻す相手: {Path(_target_undo).name}）" + chr(10) + out
            elif u.path == "/api/history":
                rc, out, payload = _ailine(["undo", book, "--list"])
            else:
                self._json(404, {"error": "not found"})
                return
        except subprocess.TimeoutExpired:
            self._json(200, {"rc": None, "text": "時間内に終わりませんでした", "json": None})
            return
        # ★ verdict は payload をそのまま渡す（ここで作らない・書き換えない）
        self._json(200, {"rc": rc, "text": out, "json": payload})


def main() -> int:
    port = 8760
    srv = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    url = f"http://127.0.0.1:{port}/"
    print(f"ailine GUI: {url}  （終了は Ctrl+C）")
    threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        srv.serve_forever()
    except KeyboardInterrupt:
        print("\n終了します")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
