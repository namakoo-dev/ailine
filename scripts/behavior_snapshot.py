# -*- coding: utf-8 -*-
"""挙動の写真 ── 直す前と後で、**画面に出る文字が何処で変わったか**を全部並べる。

★ なぜ在るか（Namakoo の懸念・2026-08-24）:
「上がってきたバグを修正する際に、新たなバグを連れてこないかを慎重に検討する必要がある」
── これは理論でなく、この日 3 回起きた:
  ① 帳票段を足したら、その検分シートが雛形の角印を全枚から消した
  ② 「全滅は仕様」と直した時、製品の表示文に**嘘**を書いた（数値書式は運んでいる）
  ③ リンタの指摘を一括置換で直して verify.py を壊した（タプル代入の片側を抜いた）

テストは「正しさ」を見る。この器官は **「動いたかどうか」を見ない ── 何が変わったかだけ**
を見る。正しさを知らなくても『意図していない所が動いた』は分かる。
★ 「消えたものは差分に出ない」への対策でもある: 出力が**減った**ことも差分に出る。

使い方:
    python scripts/behavior_snapshot.py --save before   # 直す前
    ...修正...
    python scripts/behavior_snapshot.py --diff before   # 何処が動いたかを全部出す

★ LLM を呼ぶ経路は入れない（サンプリングで毎回変わるので差分が意味を持たない）。
   決定論の経路だけを撮る ── それでも CLI の表面のかなりを覆う。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
STORE = ROOT / ".ailine_snapshots"


def _build_corpus(work: Path) -> dict:
    """決まった中身のブック/CSV を作る（毎回まったく同じもの）。"""
    import openpyxl
    from openpyxl.comments import Comment
    made = {}

    p = work / "plain.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    for r in [["取引先", "金額", "数量"], ["あかつき商事", 12000, 3],
               ["みどり工業", 8500, 2], ["合計", 20500, 5]]:
        ws.append(r)
    wb.save(p); made["plain"] = p

    p2 = work / "merged.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "売上"
    ws.append(["取引先", "金額"]); ws.append(["a", 100])
    ws.append(["   ", None]); ws.append(["b", 200])
    ws.merge_cells("A5:B5")
    wb.save(p2); made["merged"] = p2

    p3 = work / "notes.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["取引先", "金額", "備考"]); ws.append(["a", 1000, "x"])
    ws["C2"].comment = Comment("要確認", "経理")
    ws["B2"].number_format = "#,##0"
    wb.save(p3); made["notes"] = p3

    folder = work / "books"; folder.mkdir()
    for name, rows in (("a.xlsx", [["商品", "金額"], ["x", 80000]]),
                        ("b.xlsx", [["商品", "金額"], ["y", 90000]]),
                        ("c.xlsx", [["株式会社ほげ 御中"], [], ["商品", "金額"], ["z", 70000]])):
        wb = openpyxl.Workbook(); ws = wb.active
        for r in rows: ws.append(r)
        wb.save(folder / name)
    made["folder"] = folder

    strfolder = work / "strbooks"; strfolder.mkdir()
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["商品", "金額"]); ws.append(["p", "1,000"]); ws.append(["q", "80,000"])
    wb.save(strfolder / "a.xlsx")
    made["strfolder"] = strfolder

    csvp = work / "in.csv"
    csvp.write_bytes("品番,金額\n0123,1000\n".encode("cp932"))
    made["csv"] = csvp

    empty = work / "empty"; empty.mkdir(); made["empty"] = empty
    return made


def _scenarios(c: dict) -> list:
    """(名前, argv) の並び。★ LLM を呼ぶ経路は入れない（毎回変わるため）。"""
    return [
        ("help", ["--help"]),
        ("ops", ["ops"]),
        ("help_run", ["run", "--help"]),
        ("help_stack", ["stack", "--help"]),
        ("help_verify", ["verify", "--help"]),
        ("scan_folder", ["scan", str(c["folder"])]),
        ("scan_str", ["scan", str(c["strfolder"])]),
        ("scan_empty", ["scan", str(c["empty"])]),
        ("stack_folder", ["stack", str(c["folder"]), "--out", str(c["folder"].parent / "s1.xlsx")]),
        ("stack_merged_dir", ["stack", str(c["strfolder"]), "--out", str(c["folder"].parent / "s2.xlsx")]),
        ("stack_empty", ["stack", str(c["empty"]), "--out", str(c["folder"].parent / "s3.xlsx")]),
        ("csv", ["csv", str(c["csv"])]),
        ("vocab_list", ["vocab", "list"]),
        ("alias_list", ["alias", "list"]),
        ("history", ["history"]),
        ("missing_file", ["run", str(c["folder"].parent / "nope.xlsx"), "並べ替えて", "--copy"]),
        ("bad_subcommand", ["nosuchcmd"]),
    ]


def _run_one(argv, cwd, home):
    env = dict(os.environ)
    env["AILINE_HOME"] = str(home)
    env["PYTHONPATH"] = str(ROOT / "src")
    env["PYTEST_DISABLE_PLUGIN_AUTOLOAD"] = "1"
    r = subprocess.run([sys.executable, "-m", "ailine", *argv], cwd=str(cwd), env=env,
                        capture_output=True, text=True, encoding="utf-8",
                        errors="replace", timeout=300)
    return r


def _normalize(text: str, work: Path) -> str:
    """毎回変わる部分（一時パス・所要秒・日時）を伏せる ── そこは差分にしない。"""
    import re
    t = (text or "").replace(str(work), "<WORK>").replace(str(work).replace("\\", "/"), "<WORK>")
    t = re.sub(r"\(\d+\.\d+s\)", "(<T>s)", t)
    t = re.sub(r"\d{4}-\d{2}-\d{2}T[\d:+\-]+", "<TS>", t)
    t = re.sub(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}", "<TS>", t)
    t = re.sub(r"_[0-9a-f]{6}\.xlsx", "_<HASH>.xlsx", t)
    t = re.sub(r"pytest-of-\w+[\/]pytest-\d+", "<PYTEST>", t)
    return t


def capture() -> dict:
    work = Path(tempfile.mkdtemp(prefix="ailine_snap_"))
    home = work / "home"; home.mkdir()
    corpus = _build_corpus(work)
    out = {}
    for name, argv in _scenarios(corpus):
        r = _run_one(argv, work, home)
        out[name] = {
            "exit": r.returncode,
            "stdout": _normalize(r.stdout, work),
            "stderr": _normalize(r.stderr, work),
        }
    shutil.rmtree(work, ignore_errors=True)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="挙動の写真を撮る／前の写真と比べる")
    ap.add_argument("--save", metavar="NAME", help="この名前で写真を保存する")
    ap.add_argument("--diff", metavar="NAME", help="保存した写真と今を比べる")
    a = ap.parse_args()
    STORE.mkdir(exist_ok=True)

    if a.save:
        snap = capture()
        (STORE / f"{a.save}.json").write_text(
            json.dumps(snap, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"✓ {len(snap)} 場面を撮った → {STORE / (a.save + '.json')}")
        return 0

    if a.diff:
        path = STORE / f"{a.diff}.json"
        if not path.exists():
            print(f"× 写真が無い: {path}", file=sys.stderr)
            return 2
        before = json.loads(path.read_text(encoding="utf-8"))
        after = capture()
        changed = []
        for name in sorted(set(before) | set(after)):
            b, af = before.get(name), after.get(name)
            if b is None:
                changed.append((name, "場面が増えた", "", "")); continue
            if af is None:
                changed.append((name, "場面が消えた", "", "")); continue
            if b["exit"] != af["exit"]:
                changed.append((name, "exit", str(b["exit"]), str(af["exit"])))
            for stream in ("stdout", "stderr"):
                if b[stream] != af[stream]:
                    changed.append((name, stream, b[stream], af[stream]))
        if not changed:
            print(f"✓ {len(after)} 場面すべて、一字も変わっていません")
            return 0
        print(f"⚠ {len(changed)} 箇所が変わりました（{len(after)} 場面中）\n")
        import difflib
        for name, kind, b, af in changed:
            print(f"── {name} / {kind} " + "─" * 40)
            if kind == "exit":
                print(f"   {b} → {af}")
                continue
            for line in difflib.unified_diff(b.splitlines(), af.splitlines(),
                                              lineterm="", n=1):
                if line.startswith(("---", "+++", "@@")):
                    continue
                print("   " + line)
        print("\n★ 意図した変更か、1 つずつ確かめてください。"
              "意図していないものが 1 つでもあれば、それが連れてきたバグです。")
        return 1

    ap.print_help()
    return 2


if __name__ == "__main__":
    sys.exit(main())
