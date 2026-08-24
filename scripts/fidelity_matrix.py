# -*- coding: utf-8 -*-
"""飾りの生存表 ── どの経路が Excel の何を保つかを、機械が列挙して測る。

★ なぜ在るか（2026-08-24）: 「黙って壊れる」欠陥をこの日 3 件見つけた（帳票段が角印を
消す / 様式写像段が同じ / --copy が VBA を落とす）。どれも ✓ が出ていた。共通するのは
**人が気づけない**こと ── 出力は完成品に見える。俺の勘で探しても、今日は 3 件とも
「たぶん大丈夫」と思っていた側から出た。

だから列挙は機械にやらせる。ブックに Excel の飾りを一通り仕込み、各経路へ通し、
何が残ったかを数える。忠実度ゲートが知っているカテゴリ（図形/画像/VBA/ピボット/
条件付き書式/入力規則/_rels）**以外**にも当てるのが目的 ── ゲートが知らないものは
今この瞬間も黙って消えている可能性がある。

使い方: python scripts/fidelity_matrix.py [--json]
"""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import openpyxl  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Border, PatternFill, Side  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402


# --- 検出器（1 特徴 = 1 関数・戻り値は個数）--------------------------------------
# ★ openpyxl に聞けないもの（VBA・図形）は zip を直接読む（別実装で測る）。

def _zip_text(path, prefix):
    try:
        with zipfile.ZipFile(path) as z:
            return "".join(z.read(n).decode("utf-8", "replace")
                            for n in z.namelist()
                            if n.startswith(prefix) and n.endswith(".xml"))
    except Exception:
        return ""


def _zip_has(path, name):
    try:
        with zipfile.ZipFile(path) as z:
            return any(n.lower() == name for n in z.namelist())
    except Exception:
        return False


def f_shapes(path):
    import re
    return len(re.findall("<(?:[a-zA-Z]+:)?sp[ >]", _zip_text(path, "xl/drawings/")))


def f_images(path):
    import re
    return len(re.findall("<(?:[a-zA-Z]+:)?pic[ >]", _zip_text(path, "xl/drawings/")))


def f_vba(path):
    return 1 if _zip_has(path, "xl/vbaproject.bin") else 0


def _wb(path, **kw):
    return openpyxl.load_workbook(path, **kw)


def f_merged(path):
    wb = _wb(path)
    try:
        return sum(len(wb[s].merged_cells.ranges) for s in wb.sheetnames)
    finally:
        wb.close()


def f_formulas(path):
    wb = _wb(path)
    try:
        n = 0
        for s in wb.sheetnames:
            for row in wb[s].iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        n += 1
        return n
    finally:
        wb.close()


def f_formula_cache(path):
    """数式の**キャッシュ値**（開かずに読める計算結果）。消えると他ツールが 0 を読む。"""
    wb = _wb(path, data_only=True)
    try:
        wb2 = _wb(path)
        n = 0
        for s in wb2.sheetnames:
            for row in wb2[s].iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        if wb[s][c.coordinate].value is not None:
                            n += 1
        wb2.close()
        return n
    finally:
        wb.close()


def f_conditional(path):
    wb = _wb(path)
    try:
        return sum(sum(1 for _ in wb[s].conditional_formatting) for s in wb.sheetnames)
    finally:
        wb.close()


def f_validation(path):
    wb = _wb(path)
    try:
        return sum(len(wb[s].data_validations.dataValidation) for s in wb.sheetnames)
    finally:
        wb.close()


def f_comments(path):
    wb = _wb(path)
    try:
        return sum(1 for s in wb.sheetnames for row in wb[s].iter_rows()
                    for c in row if c.comment is not None)
    finally:
        wb.close()


def f_hyperlinks(path):
    wb = _wb(path)
    try:
        # ★ ws._hyperlinks は保存往復で空になることがある ── セル側から数える
        #   （初版はこちらを見ていて、仕込んだのに 0 と出た＝測定器の故障）。
        return sum(1 for s in wb.sheetnames for row in wb[s].iter_rows()
                    for c in row if c.hyperlink is not None)
    finally:
        wb.close()


def f_defined_names(path):
    wb = _wb(path)
    try:
        return len(list(wb.defined_names))
    finally:
        wb.close()


def f_autofilter(path):
    wb = _wb(path)
    try:
        return sum(1 for s in wb.sheetnames if wb[s].auto_filter.ref)
    finally:
        wb.close()


def f_print_area(path):
    wb = _wb(path)
    try:
        return sum(1 for s in wb.sheetnames if wb[s].print_area)
    finally:
        wb.close()


def f_borders(path):
    wb = _wb(path)
    try:
        n = 0
        for s in wb.sheetnames:
            for row in wb[s].iter_rows():
                for c in row:
                    b = c.border
                    if b and any(getattr(b, side).style for side in
                                  ("left", "right", "top", "bottom")):
                        n += 1
        return n
    finally:
        wb.close()


def f_fills(path):
    wb = _wb(path)
    try:
        n = 0
        for s in wb.sheetnames:
            for row in wb[s].iter_rows():
                for c in row:
                    if c.fill and c.fill.fill_type == "solid":
                        n += 1
        return n
    finally:
        wb.close()


def f_frozen(path):
    wb = _wb(path)
    try:
        return sum(1 for s in wb.sheetnames if wb[s].freeze_panes)
    finally:
        wb.close()


def f_number_format(path):
    """★ 2026-08-24 の見落としの入口: この列が無かったせいで「stack は書式を全滅させる」と
    誤って一般化し、その嘘を製品の表示文に書いた（実際は #,##0 も yyyy-mm-dd も運んでいる）。
    測っていない列は、無いのではなく**見えていない**。"""
    wb = _wb(path)
    try:
        return sum(1 for s_ in wb.sheetnames for row in wb[s_].iter_rows()
                    for c in row if c.number_format not in (None, "General"))
    finally:
        wb.close()


FEATURES = [
    ("図形(角印)", f_shapes), ("画像", f_images), ("VBA", f_vba),
    ("結合セル", f_merged), ("数式", f_formulas), ("数式のキャッシュ値", f_formula_cache),
    ("条件付き書式", f_conditional), ("入力規則", f_validation),
    ("コメント", f_comments), ("ハイパーリンク", f_hyperlinks),
    ("名前定義", f_defined_names), ("オートフィルタ", f_autofilter),
    ("印刷範囲", f_print_area), ("罫線", f_borders), ("塗りつぶし", f_fills), ("数値書式", f_number_format),
    ("ウィンドウ枠固定", f_frozen),
]


def measure(path) -> dict:
    out = {}
    for label, fn in FEATURES:
        try:
            out[label] = fn(path)
        except Exception as e:
            out[label] = f"測定不能({type(e).__name__})"
    return out


def _populate_formula_cache(path: Path) -> None:
    """soffice に一度通して数式のキャッシュ値を作る（飾りの注入より**前**に呼ぶ）。"""
    # ★ soffice は PATH に無い（ailine 本体も basrun 経由で場所を引いている）──
    #   which で探して None を返し、キャッシュ値の列が永久に 0 のままになっていた。
    try:
        import ailine as _al
        mod = _al._load_module_from_path(_al.basrun_path(), "_ailine_basrun_fm")
        office = Path(mod.office_dir())
        exe = office / ("soffice.exe" if os.name == "nt" else "soffice")
        if not exe.exists():
            return
        exe = str(exe)
    except Exception:
        exe = shutil.which("soffice") or shutil.which("soffice.exe")
        if not exe:
            return
    with tempfile.TemporaryDirectory() as td:
        try:
            subprocess.run([exe, "--headless", "--convert-to", "xlsx", "--outdir", td,
                            str(path)], capture_output=True, timeout=180)
        except Exception:
            return
        made = Path(td) / path.name
        if made.exists():
            shutil.copy2(made, path)


def _inject_drawing(path, body, png_bytes, rel_ns, xdr_ns, sheet_file, drawing_no):
    """まだ drawing を持たないシートに、新しい drawing 一式を足す。

    ★ 実測（2026-08-24）: 既存の drawing1.xml（openpyxl がコメントのために作る）へ
    「並べて」書いたら zip にエントリが重複し、LibreOffice が開けない壊れた xlsx に
    なった（exit 9 が出て ailine のバグに見えた）。「混ぜる」も名前空間が違って
    黙って落ちた。★ 衝突しない番号を使うのが正解 ── 実際の請求書も、角印は雛形シートに
    付いていて他の drawing とは別物になる。
    """
    dname = f"xl/drawings/drawing{drawing_no}.xml"
    drels_name = f"xl/drawings/_rels/drawing{drawing_no}.xml.rels"
    sheet_rels = f"xl/worksheets/_rels/{Path(sheet_file).name}.rels"
    img = f"image{drawing_no}.png"
    drawing = f'<wsDr xmlns="{xdr_ns}">{body}</wsDr>'
    drels = (f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
             f'<Relationship Id="rId1" Type="{rel_ns}/image" Target="../media/{img}"/>'
             f'</Relationships>')
    add_rel = (f'<Relationship Id="rIdDraw{drawing_no}" Type="{rel_ns}/drawing" '
               f'Target="../drawings/drawing{drawing_no}.xml"/>')
    tmp = str(path) + ".d"
    with zipfile.ZipFile(path) as zi:
        names = set(zi.namelist())
        assert dname not in names, f"衝突する drawing 番号: {dname}"
        with zipfile.ZipFile(tmp, "w") as zo:
            for item in zi.infolist():
                data = zi.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    if 'Extension="png"' not in text:
                        text = text.replace("</Types>",
                                            '<Default Extension="png" ContentType="image/png"/></Types>')
                    text = text.replace("</Types>",
                                        f'<Override PartName="/{dname}" ContentType='
                                        '"application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                                        "</Types>")
                    data = text.encode("utf-8")
                elif item.filename == sheet_file:
                    data = data.decode("utf-8").replace(
                        "</worksheet>",
                        f'<drawing xmlns:r="{rel_ns}" r:id="rIdDraw{drawing_no}"/></worksheet>'
                    ).encode("utf-8")
                elif item.filename == sheet_rels:
                    data = data.decode("utf-8").replace(
                        "</Relationships>", add_rel + "</Relationships>").encode("utf-8")
                zo.writestr(item, data)
            if sheet_rels not in names:
                zo.writestr(sheet_rels,
                            '<Relationships xmlns="http://schemas.openxmlformats.org/'
                            f'package/2006/relationships">{add_rel}</Relationships>')
            zo.writestr(f"xl/media/{img}", png_bytes)
            zo.writestr(dname, drawing)
            zo.writestr(drels_name, drels)
    os.replace(tmp, path)


# --- 検体（飾りを全部盛ったブック）--------------------------------------------------

def build_specimen(path: Path) -> Path:
    """Excel の飾りを一通り仕込んだブック。★ 仕込めたかを最後に自分で確かめる
       （治具が壊れていると『消えた』でなく『最初から無い』を測ってしまう）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売上"
    ws.append(["取引先", "金額", "数量", "小計", "備考"])
    for name, amount, qty in (("あかつき商事", 12000, 3), ("みどり工業", 8500, 2),
                               ("そら建設", 30000, 5)):
        ws.append([name, amount, qty, f"={openpyxl.utils.get_column_letter(2)}"
                                       f"{ws.max_row + 1}*"
                                       f"{openpyxl.utils.get_column_letter(3)}"
                                       f"{ws.max_row + 1}", ""])
    # 数値書式（桁区切りと日付）── ★ 初版はこれを仕込み忘れ、列が空のまま
    #   「stack は書式を全滅させる」と誤って一般化した。
    import datetime
    for r in range(2, 5):
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=5, value=datetime.date(2026, 7, 31)).number_format = "yyyy-mm-dd"
    ws.merge_cells("A5:B5")
    ws["A5"] = "合計欄（結合）"
    ws["E2"].comment = Comment("要確認", "ailine")
    ws["E3"].hyperlink = "https://example.invalid/"
    ws["E3"].value = "リンク"
    side = Side(style="thin")
    ws["A1"].border = Border(left=side, right=side, top=side, bottom=side)
    ws["B1"].fill = PatternFill("solid", fgColor="FFFF00")
    ws.conditional_formatting.add("B2:B4",
                                   CellIsRule(operator="greaterThan", formula=["10000"],
                                              fill=PatternFill("solid", fgColor="FFC7CE")))
    dv = DataValidation(type="list", formula1='"可,否"')
    ws.add_data_validation(dv)
    dv.add("E4")
    ws.auto_filter.ref = "A1:E4"
    ws.print_area = "A1:E5"
    ws.freeze_panes = "A2"
    wb.defined_names.add(openpyxl.workbook.defined_name.DefinedName(
        "取引先一覧", attr_text="売上!$A$2:$A$4"))
    tpl = wb.create_sheet("雛形")
    tpl["A1"] = "請求書"
    tpl["A3"] = "宛先"
    tpl["B3"] = "{{取引先}}"
    tpl["A5"] = "ご請求金額"
    tpl["B5"] = "{{金額}}"
    wb.save(path)

    # ★ 数式のキャッシュ値は openpyxl では作れない（計算しないので）。LO に一度通して
    #   計算させる ── これをやらないと「キャッシュ値の喪失」を永久に測れない
    #   （0 から 0 へ減っても差が出ない＝恒真）。
    _populate_formula_cache(path)

    # 図形（角印）と画像は openpyxl では入れられない形なので zip で足す。
    # ★ tests の _write_drawing は「シート rels が無い」前提で書いてあり、ここでは
    #   ハイパーリンクで既に rels が在るため zip にエントリが**重複**して openpyxl が
    #   読めなくなった（実測）。既存 rels に**混ぜる**版をここに持つ。
    sys.path.insert(0, str(ROOT / "tests"))
    from test_vanishing_shapes import _PIC, _SHAPE, _png, _REL, _XDR   # noqa: E402
    # 角印は**雛形シート**（2 枚目）に付ける ── 実際の請求書と同じ形で、
    # コメントが作る drawing1 とも衝突しない。
    _inject_drawing(path, _PIC + _SHAPE, _png(), _REL, _XDR,
                     "xl/worksheets/sheet2.xml", 9)

    # VBA も足す（マクロ有効ブックの代用・在るか消えるかだけを見る）
    tmp = str(path) + ".t"
    with zipfile.ZipFile(path) as zi, zipfile.ZipFile(tmp, "w") as zo:
        for item in zi.infolist():
            data = zi.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = data.decode("utf-8").replace(
                    "</Types>", '<Override PartName="/xl/vbaProject.bin" ContentType='
                                '"application/vnd.ms-office.vbaProject"/></Types>').encode("utf-8")
            zo.writestr(item, data)
        zo.writestr("xl/vbaProject.bin", b"FAKE_VBA_PROJECT_MARKER")
    os.replace(tmp, path)


    # ★ 治具の自己検査: 仕込めなかった特徴が在れば、その列は測っても意味が無い
    got = measure(path)
    # ★ 「0 件」だけでなく「測定不能」も治具の失敗として拾う ── 初版は 0 しか見ておらず、
    #   検体が壊れて全部 KeyError になっているのに『仕込めなかった: []』と報告した。
    missing = [k for k, v in got.items() if not isinstance(v, int) or v == 0]
    return path, got, missing


# --- 経路（ailine の各コマンドへ通す）------------------------------------------------

def _run(argv, cwd, env_home):
    env = dict(os.environ)
    env["AILINE_HOME"] = str(env_home)
    env["PYTHONPATH"] = str(ROOT / "src")
    r = subprocess.run([sys.executable, "-m", "ailine", *argv], cwd=str(cwd), env=env,
                        capture_output=True, text=True, encoding="utf-8",
                        errors="replace", timeout=900)
    return r


def paths_to_measure(spec: Path, work: Path, home: Path):
    """[(経路名, 出力パス or None, 備考), ...] を返す。★ 出力が出来なかった経路も
       None で必ず返す ── 黙って表から消えると『測った』ように見える。"""
    out = []

    # ① 並べ替え（--copy）── 最も普通の単発 op
    a = work / "sort.xlsx"
    shutil.copy2(spec, a)
    r = _run(["run", str(a), "金額で降順に並べ替えて", "--copy", "--accept-loss"], work, home)
    produced = a.with_name(a.stem + ".out.xlsx")
    out.append(("run（並べ替え）", produced if produced.exists() else None,
                 f"exit={r.returncode}"))

    # ② 帳票作成（雛形 → N 枚）
    b = work / "report.xlsx"
    shutil.copy2(spec, b)
    r = _run(["run", str(b), "売上表から取引先ごとに請求書を作って", "--copy", "--accept-loss"],
              work, home)
    produced = b.with_name(b.stem + ".out.xlsx")
    out.append(("run（帳票作成）", produced if produced.exists() else None,
                 f"exit={r.returncode}"))

    # ③ 縦積み（フォルダ）── 新しいブックを作る経路
    folder = work / "stackdir"
    folder.mkdir(exist_ok=True)
    shutil.copy2(spec, folder / "a.xlsx")
    stacked = work / "stacked.xlsx"
    r = _run(["stack", str(folder), "--out", str(stacked)], work, home)
    out.append(("stack（縦積み）", stacked if stacked.exists() else None,
                 f"exit={r.returncode}"))

    return out


def main() -> int:
    as_json = "--json" in sys.argv
    work = Path(tempfile.mkdtemp(prefix="ailine_fidelity_"))
    home = work / "home"
    home.mkdir()
    spec, planted, missing = build_specimen(work / "specimen.xlsx")
    if missing:
        print(f"⚠ 検体に仕込めなかった特徴: {'・'.join(missing)}")
        print("  → その行は『消えた』でなく『最初から無い』ので、下の表では - と書く")

    rows = paths_to_measure(spec, work, home)
    table = {}
    for label, produced, note in rows:
        table[label] = (measure(produced) if produced else None, note)

    if as_json:
        print(json.dumps({"planted": planted,
                          "missing": missing,
                          "paths": {k: {"note": v[1], "measured": v[0]}
                                     for k, v in table.items()}},
                          ensure_ascii=False, indent=2))
        return 0

    names = list(table)
    width = max(len(f) for f, _ in FEATURES) + 2
    print()
    print("飾りの生存表（元 → 各経路の出力）")
    print("=" * (width + 6 + sum(len(n) + 3 for n in names)))
    print("特徴".ljust(width) + "元".rjust(4) + "".join(n.rjust(len(n) + 3) for n in names))
    lost_any = []
    for feat, _fn in FEATURES:
        before = planted.get(feat)
        cells = []
        for n in names:
            measured, _note = table[n]
            if measured is None:
                cells.append("×")
                continue
            after = measured.get(feat)
            if not isinstance(before, int) or before == 0:
                cells.append("-")
            elif not isinstance(after, int):
                cells.append("?")
            elif after >= before:
                cells.append("○")
            elif after == 0:
                cells.append("全滅")
                lost_any.append((feat, n, before, after))
            else:
                cells.append(f"{after}/{before}")
                lost_any.append((feat, n, before, after))
        print(feat.ljust(width) + str(before).rjust(4)
              + "".join(c.rjust(len(n) + 3) for c, n in zip(cells, names)))
    print()
    for n in names:
        print(f"  {n}: {table[n][1]}")
    print()
    if lost_any:
        print("失われたもの:")
        for feat, path_name, b, a in lost_any:
            print(f"  ⚠ {path_name} で『{feat}』が {b} → {a}")
    else:
        print("この検体の範囲では、失われたものはありません")
    print(f"\n作業ディレクトリ: {work}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
